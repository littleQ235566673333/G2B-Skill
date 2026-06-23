from openpyxl import load_workbook

# Load the workbook and select the active worksheet
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun1/eval_51289_tc1/input.xlsx')
ws = wb.active

# Read the first (A1:H1) and second (A2:H2) rows
labels = [ws.cell(row=1, column=c).value for c in range(1, 9)]
numbers = [ws.cell(row=2, column=c).value for c in range(1, 9)]
data = list(zip(labels, numbers, range(8)))

# Rank cats and dogs
cat = sorted([(num, idx) for lbl, num, idx in data if lbl == 'cat'], reverse=True)
dog = sorted([(num, idx) for lbl, num, idx in data if lbl == 'dog'], reverse=True)

# Prepare output
output = [''] * 8
if len(dog) > 0:
    output[dog[0][1]] = 'dog1'
if len(dog) > 1:
    output[dog[1][1]] = 'dog2'
if len(cat) > 0:
    output[cat[0][1]] = 'cat1'
if len(cat) > 1:
    output[cat[1][1]] = 'cat2'

for i, val in enumerate(output):
    ws.cell(row=4, column=i + 1).value = val

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun1/eval_51289_tc1/output.xlsx')
