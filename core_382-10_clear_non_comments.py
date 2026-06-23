from openpyxl import load_workbook

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_4/regression_gate/after_fix/core_382-10/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_4/regression_gate/after_fix/core_382-10/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Only keep cells with value 'Comments' in column A, clear all others (except header)
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=1)
    if cell.value != 'Comments':
        cell.value = None

wb.save(output_path)