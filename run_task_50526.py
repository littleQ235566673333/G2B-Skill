import openpyxl

# Load workbook
wbin = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_7/regression_gate/before_pass/core_50526/input.xlsx'
wout = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_7/regression_gate/before_pass/core_50526/output.xlsx'
wb = openpyxl.load_workbook(wbin)
ws = wb.active

# Read headers (row 1) and lookup values (col 1 from row 2 onwards)
headers = [cell.value for cell in ws[1]][1:]  # Exclude A1, assume B1:... are color names
lookup_rows = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row+1) if ws.cell(row=r, column=1).value}  # Map lookup key to row

# Get the lookup cell value (B6, i.e., ws['B6'])
lookup_value = ws['B6'].value
row = lookup_rows.get(lookup_value)
results = []
if row:
    for i, col_header in enumerate(headers, 2):  # start from column 2 (B)
        cell_value = ws.cell(row=row, column=i).value
        if cell_value is not None and cell_value > 0:
            results.append(col_header)
# Place results (or blanks if not enough results)
for i in range(2): # Only B9 and B10
    ws.cell(row=9+i, column=2).value = results[i] if i < len(results) else ''

wb.save(wout)
