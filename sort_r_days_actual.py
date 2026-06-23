import openpyxl
from openpyxl.utils import get_column_letter
import shutil

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/group_208-20/r2/evolve_208-20/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/group_208-20/r2/evolve_208-20/output.xlsx'

# Copy file to preserve all styles and non-targeted sheets
shutil.copyfile(input_path, output_path)

wb = openpyxl.load_workbook(output_path)
ws = wb['Q']

# Extract headers and data (A2:F2, A3:F37)
header_row = 2
start_row = 3
end_row = 37

header = [ws.cell(row=header_row, column=col).value for col in range(1, 7)]  # A-F

data_rows = []
for r in range(start_row, end_row + 1):
    data_rows.append([
        ws.cell(row=r, column=col).value for col in range(1, 7)
    ])

# Sort by R DAYS (column E, index 4), descending; None/empty treated as lowest

def sort_key(row):
    val = row[4]
    if val is None:
        return float('-inf')  # place Nones at end
    return val

sorted_rows = sorted(data_rows, key=sort_key, reverse=True)

# Write back header and sorted rows to A2:F37 (keep A1:F1 as original)
for c in range(1, 7):
    ws.cell(row=header_row, column=c).value = header[c-1]
for row_idx, data in enumerate(sorted_rows, start=start_row):
    for c in range(1, 7):
        ws.cell(row=row_idx, column=c).value = data[c-1]

wb.save(output_path)
