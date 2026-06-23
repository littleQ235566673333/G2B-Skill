import datetime
from openpyxl import load_workbook, styles
from copy import copy

def main():
    in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_2/regression_gate/before_pass/core_32337/input.xlsx'
    out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_2/regression_gate/before_pass/core_32337/output.xlsx'
    wb = load_workbook(in_path)
    ws = wb['Sheet1']

    header = [cell.value for cell in ws[2]]
    col_e = header.index('Expected Result') + 1
    col_dob = header.index('DATE OF BIRTH') + 1
    col_year_age = header.index('YEAR (age)') + 1
    col_category = header.index('CATEGORY') + 1
    col_age_year = header.index('age (year)') + 1
    # Insert a new column after age (year)
    ws.insert_cols(col_age_year + 1)
    ws.cell(row=2, column=col_age_year + 1).value = 'Actual Age'
    # Copy fill from 'age (year)' - safely
    refcell = ws.cell(row=3, column=col_age_year)
    refstyle = refcell.fill
    fill = styles.PatternFill(
        fill_type=refstyle.fill_type,
        fgColor=copy(refstyle.fgColor),
        bgColor=copy(refstyle.bgColor))
    # Format header
    header_cell = ws.cell(row=2, column=col_age_year + 1)
    header_cell.font = styles.Font(bold=True)
    header_cell.alignment = styles.Alignment(horizontal='center', vertical='top')
    header_cell.fill = fill

    for r in range(3, 16):
        agecell = ws.cell(row=r, column=col_age_year + 1)
        dob_ref = ws.cell(row=r, column=col_dob).coordinate
        report_ref = '$B$1'
        agecell.value = f'=INT(YEARFRAC({dob_ref}, {report_ref}))'
        agecell.font = styles.Font(bold=True)
        agecell.alignment = styles.Alignment(horizontal='center', vertical='center')
        agecell.fill = fill
        agecell.number_format = '0'
    h_col = col_year_age
    cat_col = col_category
    age_col = col_age_year
    for r in range(3, 16):
        ws.cell(row=r, column=col_e).value = f'=INDEX(${chr(64+cat_col)}$3:${chr(64+cat_col)}$15, MATCH({ws.cell(row=r, column=age_col).coordinate}, ${chr(64+h_col)}$3:${chr(64+h_col)}$15, 0))'
    wb.save(out_path)

if __name__ == '__main__':
    main()
