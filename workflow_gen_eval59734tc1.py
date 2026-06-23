import pandas as pd
from openpyxl import load_workbook

inp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/eval_seed42_rerun2/eval_59734_tc1/input.xlsx'
outp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/eval_seed42_rerun2/eval_59734_tc1/output.xlsx'

wb = load_workbook(inp)
ws = wb.active

df = pd.read_excel(inp)

for i in range(1, 16):
    numval = df.loc[i, 'A'] if 'A' in df.columns else df.iloc[i, 0]
    o_val = df.loc[i, 'F'] if 'F' in df.columns else df.iloc[i, 5]
    p_val = df.loc[i, 'G'] if 'G' in df.columns else df.iloc[i, 6]
    val = ''
    if all(col in df.columns for col in ['A','F','G','V']):
        matches = df[(df['A']==numval) & (df['F']==o_val) & (df['G']==p_val)]['V']
        val = matches.values[0] if len(matches) > 0 else ''
    ws.cell(row=i+1, column=2).value = val

wb.save(outp)
