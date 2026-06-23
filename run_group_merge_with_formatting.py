import openpyxl
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_8/regression_gate/after_fix/core_177-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_8/regression_gate/after_fix/core_177-6/output.xlsx'

wbin = openpyxl.load_workbook(input_path)
wsin = wbin.active

rows = list(wsin.iter_rows(min_row=1, max_col=18, max_row=wsin.max_row, values_only=False))
header = [cell.value for cell in rows[0]]
data = [row for row in rows[1:] if any(cell.value is not None for cell in row)]

groups = defaultdict(list)
for row in data:
    key = row[7].value  # Column H (0-indexed 7)
    groups[key].append(row)

def copy_cell_format_safe(src_cell, tgt_cell):
    for attr in ['font', 'fill', 'border', 'number_format', 'alignment', 'protection']:
        try:
            setattr(tgt_cell, attr, getattr(src_cell, attr))
        except Exception:
            pass

if 'combined' in wbin.sheetnames:
    del wbin['combined']
wsout = wbin.create_sheet('combined')

# Write header
for c in range(18):
    src = rows[0][c]
    cell = wsout.cell(row=1, column=c+1, value=src.value)
    copy_cell_format_safe(src, cell)

# Output groups
for i, (group_key, group_rows) in enumerate(groups.items()):
    row_out_idx = i+2
    src_row = group_rows[0]
    for c in range(18):
        src_cell = src_row[c]
        out_cell = wsout.cell(row=row_out_idx, column=c+1)
        if c < 8:  # A-H
            out_cell.value = src_cell.value
            copy_cell_format_safe(src_cell, out_cell)
        else:  # I-R, sum, format
            vals = [r[c].value for r in group_rows if r[c].value not in (None, '')]
            s = Decimal('0.00')
            for v in vals:
                try:
                    s += Decimal(str(v).replace(',',''))
                except Exception:
                    pass
            formatted = '' if s == 0 else '{:.2f}'.format(s.quantize(Decimal('.01'), rounding=ROUND_HALF_UP))
            out_cell.value = formatted
            copy_cell_format_safe(src_cell, out_cell)

wbin.save(output_path)
wbin.close()
