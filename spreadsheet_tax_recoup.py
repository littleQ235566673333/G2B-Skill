import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_36097_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_36097_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Let's assume columns:
# Original Cost (A), ITV (B), Sale Price (C) are given, column H is result, column I is for formatting
# If columns differ please clarify

for row in range(3, 7):
    original_cost = ws[f'A{row}'].value
    itv = ws[f'B{row}'].value
    sale_price = ws[f'C{row}'].value
    # Loss: sold for less than cost
    loss = max(0, original_cost - sale_price)
    profit = sale_price - original_cost

    if sale_price < original_cost:
        # Loss case
        result = itv + loss
    elif profit < original_cost:
        # Profit but less than cost
        result = profit
    else:
        # Profit >= cost
        result = original_cost - itv

    ws[f'H{row}'].value = result

    # Copy formatting from column I
    ws[f'H{row}'].font = ws[f'I{row}'].font
    ws[f'H{row}'].fill = ws[f'I{row}'].fill
    ws[f'H{row}'].border = ws[f'I{row}'].border
    ws[f'H{row}'].alignment = ws[f'I{row}'].alignment
    ws[f'H{row}'].number_format = ws[f'I{row}'].number_format
    ws[f'H{row}'].protection = ws[f'I{row}'].protection

wb.save(output_path)
