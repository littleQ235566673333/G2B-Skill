import openpyxl
from openpyxl.cell.cell import MergedCell

# Load the workbook and sheet
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_55060_tc1/input.xlsx')
ws = wb.active

# Read the value from I12
value = ws['I12'].value

# J23:N23 is columns 10-14, row 23
for col in range(10, 15):
    cell = ws.cell(row=23, column=col)
    if isinstance(cell, MergedCell):
        # Unmerge the merged cell range containing this cell
        for merged_range in ws.merged_cells.ranges:
            if (23, col) >= (merged_range.min_row, merged_range.min_col) and (23, col) <= (merged_range.max_row, merged_range.max_col):
                ws.unmerge_cells(str(merged_range))
                break
        cell = ws.cell(row=23, column=col)
    if value in [None, '']:
        cell.value = None
    else:
        cell.value = value

# Save to output file
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_55060_tc1/output.xlsx')
