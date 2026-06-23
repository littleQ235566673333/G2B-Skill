import openpyxl

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/group_59129/r0/evolve_59129/input.xlsx')
ws = wb['Sheet1']

# Read hire & term dates
hire_dates = []
term_dates = []
for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
    hire_dates.append(row[0])
    term_dates.append(row[1])

# Read analysis months from E1:P1
months = [cell.value for cell in ws[1][4:16]]
results = []
for m in months:
    count = 0
    for h, t in zip(hire_dates, term_dates):
        if h is not None and h <= m and (t is None or t > m):
            count += 1
    results.append(count)

# Write results to E2:P2
for col, res in enumerate(results, start=5):
    ws.cell(row=2, column=col, value=res)

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/group_59129/r0/evolve_59129/output.xlsx')
