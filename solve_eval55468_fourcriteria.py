import openpyxl

def main():
    in_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_55468_tc1/input.xlsx'
    out_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_55468_tc1/output.xlsx'
    wb = openpyxl.load_workbook(in_path)
    ws = wb['Sheet1']

    # Fetch criteria values from header cells
    AE4 = ws['AE4'].value  # Vertical crit 1
    AD4 = ws['AD4'].value  # Vertical crit 2
    AC4 = ws['AC4'].value  # Horizontal crit 1
    AB4 = ws['AB4'].value  # Horizontal crit 2

    # Data matrix: C5:Z10
    data_start_row = 5
    data_end_row = 10
    data_start_col = 3  # C
    data_end_col = 26   # Z

    # Horizontal header rows
    header_row_1 = 3    # C3:Z3 (AB4 matches)
    header_row_2 = 4    # C4:Z4 (AC4 matches)
    # Vertical header columns
    header_col_1 = 1    # A5:A10 (AE4 matches)
    header_col_2 = 2    # B5:B10 (AD4 matches)

    # Find matching column (horizontal, two criteria)
    candidate_cols = []
    for col in range(data_start_col, data_end_col+1):
        val1 = ws.cell(row=header_row_1, column=col).value
        val2 = ws.cell(row=header_row_2, column=col).value
        if val1 == AB4 and val2 == AC4:
            candidate_cols.append(col)
    if not candidate_cols:
        ws['AE5'] = 'No match col'
        wb.save(out_path)
        return
    col_match = candidate_cols[0]  # Take first matched
    
    # Find matching row (vertical, two criteria)
    candidate_rows = []
    for row in range(data_start_row, data_end_row+1):
        val1 = ws.cell(row=row, column=header_col_1).value
        val2 = ws.cell(row=row, column=header_col_2).value
        if val1 == AE4 and val2 == AD4:
            candidate_rows.append(row)
    if not candidate_rows:
        ws['AE5'] = 'No match row'
        wb.save(out_path)
        return
    row_match = candidate_rows[0]

    result = ws.cell(row=row_match, column=col_match).value
    ws['AE5'] = result
    wb.save(out_path)

if __name__ == '__main__':
    main()
