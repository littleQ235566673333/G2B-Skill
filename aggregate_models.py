import openpyxl

# File paths
input_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_49945_tc1/input.xlsx"
output_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_49945_tc1/output.xlsx"

# Load the workbook and select the active worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read model and quantity columns (assume user's ranges: C4:C11 for model, D4:D11 for qty)
models = [ws[f"C{row}"].value for row in range(4, 12)]
quantities = [ws[f"D{row}"].value for row in range(4, 12)]

# Find unique models in the order they appear (since user wants results in G4:G6)
unique_models = []
for m in models:
    if m not in unique_models:
        unique_models.append(m)

# Calculate the sum for each unique model
for idx, model in enumerate(unique_models):
    sum_qty = sum(q for m, q in zip(models, quantities) if m == model)
    # Write to G4, G5, G6 ...
    ws[f"G{4+idx}"].value = sum_qty

# Save the results
wb.save(output_path)
