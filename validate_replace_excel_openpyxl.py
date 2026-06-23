import openpyxl
from openpyxl.styles import PatternFill

# File paths
data_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_120-24_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_120-24_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(data_path)
ws = wb['Sheet1']

# Columns (Excel is 1-based)
col_BL = 64  # BL
col_BG = 59  # BG
col_AY = 51  # AY
col_BN = 67  # BN

# Fill color
fill_lag_populated = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

# The range specified is BN2:BN45
for row in range(2, 46):
    val_BL = ws.cell(row=row, column=col_BL).value
    val_BG = ws.cell(row=row, column=col_BG).value
    val_AY = ws.cell(row=row, column=col_AY).value
    
    # Check BL == 'LAG' and BG in set, then handle AY
    if val_BL == 'LAG' and val_BG in ['OCOGS - Spares - Transfer Price Overhead', 'OFSS - ORCL Consulting Prod Cost I/C']:
        if val_AY in ['BOA_033E', 'BOA_011G']:
            ws.cell(row=row, column=col_BN).value = 'OCOGS - Spares - Transfer Price Overhead'
        else:
            ws.cell(row=row, column=col_BN).value = 'OFSS - ORCL Consulting Prod Cost I/C'
        # If BL == 'LAG' and BN is now populated, fill color
        if ws.cell(row=row, column=col_BN).value:
            ws.cell(row=row, column=col_BL).fill = fill_lag_populated

wb.save(output_path)
