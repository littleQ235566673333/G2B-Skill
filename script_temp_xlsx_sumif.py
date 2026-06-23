from openpyxl import load_workbook
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/before_pass/core_42181/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/before_pass/core_42181/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def contains_no_digits(s):
    return not bool(re.search(r'[0-9]', str(s)))

# Range parameters
start_row, end_row = 4, 10
ref_col = 'I'
sum_col = 'B'
lookup_cell = 'A13'
output_cell = 'B13'
lookup_value = str(ws[lookup_cell].value)

total = 0
for r in range(start_row, end_row + 1):
    ref_val = ws[f'{ref_col}{r}'].value
    sum_val = ws[f'{sum_col}{r}'].value
    if ref_val is not None and sum_val is not None:
        ref_val_str = str(ref_val)
        if lookup_value in ref_val_str and contains_no_digits(ref_val_str):
            try:
                total += float(sum_val)
            except ValueError:
                pass
ws[output_cell] = total
wb.save(output_path)