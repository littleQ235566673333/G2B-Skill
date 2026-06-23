import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_40892_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_40892_tc1/output.xlsx'

# Load the workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read list of colors from D2:D17
color_cells = ws['D'][1:16]  # openpyxl is 0-based, but this is correct
colors = [str(cell.value).strip() for cell in color_cells if cell.value]

# For each row 2-17, extract mentioned color from col 1, write to col 2
for row in range(2, 18):
    text = str(ws.cell(row=row, column=1).value or '').lower()
    matched_color = ''
    for color in colors:
        # Use lowercase for matching
        if color.lower() in text:
            matched_color = color
            break
    ws.cell(row=row, column=2).value = matched_color

wb.save(output_path)
