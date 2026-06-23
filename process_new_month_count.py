from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/group_45707/r3/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/group_45707/r3/evolve_45707/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# D2:D69
start_row, end_row = 2, 69
A = [ws[f'A{row}'].value for row in range(1, end_row + 2)]  # Includes header and one extra
C = [ws[f'C{row}'].value for row in range(1, end_row + 2)]

for i in range(start_row, end_row + 1):
    date_next = A[i] if i < len(A) else None
    cell = ws.cell(row=i, column=4)  # D column
    if isinstance(date_next, datetime) and date_next.day == 1:
        year, month = date_next.year, date_next.month
        count = 0
        for j in range(1, end_row + 1):
            date_j = A[j-1]
            if isinstance(date_j, datetime) and date_j.year == year and date_j.month == month:
                if C[j-1] == 1:
                    count += 1
        cell.value = count
    else:
        cell.value = None

wb.save(output_path)
