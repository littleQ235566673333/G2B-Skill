import openpyxl
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/regression_gate/before_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/regression_gate/before_pass/core_55421/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

lookup = defaultdict(list)
# Build lookup for all unique numbers in Column A (rows 2-20)
for row in ws.iter_rows(min_row=2, min_col=1, max_col=5, max_row=20):
    a = row[0].value
    d = row[3].value
    e = row[4].value
    lookup[a].append({'row': row[0].row, 'status': d, 'date': e})

result = {}
for a, entries in lookup.items():
    statuses = set(e['status'] for e in entries if e['status'])
    has_no_show = 'NO SHOW' in statuses
    has_sch = 'SCH' in statuses
    if statuses == {'SCH'}:
        out = 'FUTURE'
        for e in entries:
            result[e['row']] = out
    elif has_no_show and has_sch:
        out = 'NS/SCHED'
        for e in entries:
            result[e['row']] = out
    elif has_no_show:
        for e in entries:
            if e['status'] == 'NO SHOW':
                if e['date']:
                    result[e['row']] = 'NO ACTION NEEDED'
                else:
                    result[e['row']] = 'CALL PT'
            else:
                result[e['row']] = ''
    else:
        for e in entries:
            result[e['row']] = ''

for row in range(2, 21):
    ws.cell(row=row, column=6, value=result.get(row, ''))
wb.save(output_path)
