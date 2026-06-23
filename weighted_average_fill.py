from openpyxl import load_workbook

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_183-8_tc1/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_183-8_tc1/output.xlsx'

wb = load_workbook(input_fp)
ws = wb['Sheet1']

# For columns J(10), K(11), L(12), rows 3-6, write weighted average formula
# Weight column: I (9) (Sum of Gen.)

for col, letter in zip([10, 11, 12], ['J', 'K', 'L']):
    for row in range(3, 7):
        cell = f'{letter}{row}'
        # Weighted avg: SUMPRODUCT($I$3:$I$6, J3:J6) / SUM($I$3:$I$6)
        ws[cell] = f"=SUMPRODUCT($I$3:$I$6,{letter}$3:{letter}$6)/SUM($I$3:$I$6)"

wb.save(output_fp)
