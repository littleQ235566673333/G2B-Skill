import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

infile = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_1/group_44017/r2/evolve_44017/input.xlsx"
outfile = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_1/group_44017/r2/evolve_44017/output.xlsx"
wb = openpyxl.load_workbook(infile)
ws = wb["Data"]

# Col mappings
col_base = 'W'      # base rate
col_eff = 'L'       # effective date
col_freq = 'J'      # frequency
incs = ['M','N','O','P']  # up to 4 increases
start_col = 30 # AD
end_col = 41 # AO
row_data_start = 14
row_data_end = 42

# Prepare formulas by col
for row in range(row_data_start, row_data_end+1):
    for idx, col in enumerate(range(start_col, end_col+1)):
        out_cell = f"{get_column_letter(col)}{row}"
        date_cell = f"{get_column_letter(col)}$9"
        base_cell = f"${col_base}{row}"
        freq_cell = f"${col_freq}{row}"
        effdate_cell = f"${col_eff}{row}"
        inc_cells = [f"${let}{row}" for let in incs]
        # Generate cascading IF with INT date math. Use DATEDIF in Excel ("M") for full months
        # _EXCEL_ DATEDIF(start,end,"M") counts months. (If not supported, fallback to ((YEAR-END)*12 + (MONTH-END)) - ((YEAR-START)*12 + (MONTH-START)).)
        
        # Main formula logic, as described
        formula = (
            f'=IF({date_cell} < {effdate_cell}, "", '
            f'IFERROR('
            f'{base_cell}'
            f'* (1+IF(({date_cell}>={effdate_cell}), IF({inc_cells[0]}<>"", {inc_cells[0]}, 0), 0))'
            f'* (1+IF(({date_cell}>=EDATE({effdate_cell},1*{freq_cell})), IF({inc_cells[1]}<>"", {inc_cells[1]}, 0), 0))'
            f'* (1+IF(({date_cell}>=EDATE({effdate_cell},2*{freq_cell})), IF({inc_cells[2]}<>"", {inc_cells[2]}, 0), 0))'
            f'* (1+IF(({date_cell}>=EDATE({effdate_cell},3*{freq_cell})), IF({inc_cells[3]}<>"", {inc_cells[3]}, 0), 0),"")'
            f'))'
        )
        ws[out_cell] = formula
        # Remove any fill for clean cells
        ws[out_cell].fill = openpyxl.styles.PatternFill(fill_type=None)

wb.save(outfile)
