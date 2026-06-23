import openpyxl

# File paths
i_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_49945_tc1/input.xlsx'
o_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_49945_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(i_path)
ws = wb.active

# Read the target models from G4:G6
models = [ws.cell(row=row, column=7).value for row in range(4, 7)]

# Read models and quantities (C4:C11, D4:D11)
model_range = [ws.cell(row=row, column=3).value for row in range(4, 12)]
qty_range = [ws.cell(row=row, column=4).value for row in range(4, 12)]

# Sum the quantities per model and write to G4:G6
for i, model in enumerate(models):
    total = sum(qty for m, qty in zip(model_range, qty_range) if m == model)
    ws.cell(row=4 + i, column=7).value = total

wb.save(o_path)
