from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_3/regression_gate/after_pass/core_32337/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_3/regression_gate/after_pass/core_32337/output.xlsx'
wb = load_workbook(input_fp)
ws = wb['Sheet1']

# Obtain the fill of column O header for style consistency
age_header_fill_src = ws.cell(row=2, column=15).fill
fill = PatternFill(
    fill_type=age_header_fill_src.fill_type,
    fgColor=age_header_fill_src.fgColor.rgb,
    bgColor=age_header_fill_src.bgColor.rgb
)

actual_age_col = 16  # column P
header_cell = ws.cell(row=2, column=actual_age_col)
header_cell.value = 'Actual Age'
header_cell.font = Font(bold=True)
header_cell.alignment = Alignment(horizontal='center', vertical='top')
header_cell.fill = fill

# From row 3 to 15, write formula for Actual Age, and do formatting
for i in range(3, 16):
    cell = ws.cell(row=i, column=actual_age_col)
    cell.value = f'=ROUNDDOWN((($B$1-C{i})/365),0)'
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = fill

# Write the formula in E3:E15 for CATEGORY lookup
for i in range(3, 16):
    formula = f'=IFERROR(XLOOKUP(P{i},H$3:H$15,I$3:I$15,"") ,"")'
    ws.cell(row=i, column=5).value = formula

wb.save(output_fp)
