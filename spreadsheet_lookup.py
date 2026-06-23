import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_1/regression_gate/after_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_1/regression_gate/after_pass/core_18935/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read the Data Table into a lookup structure
data_table = []
for row in ws.iter_rows(min_row=5, max_row=10, values_only=True):
    if row[0] is not None and row[1] is not None:
        work = row[0]
        material = row[1]
        # Categories are from columns C, D, E (i.e., index 2,3,4)
        categories = row[2:5]
        data_table.append({'Work': work, 'Material': material, 'Categories': categories})
        
# Mapping from Category Type name ("Category 1/2/3") to index
category_type_to_index = {f'Category {i+1}': i for i in range(3)}

# Fill Report Format D17:D22 based on report rows 17-22
for report_row in range(17, 23):
    work = ws[f'A{report_row}'].value
    category_type = ws[f'B{report_row}'].value
    material = ws[f'C{report_row}'].value
    
    # Lookup in data table
    match = None
    for entry in data_table:
        if entry['Work'] == work and entry['Material'] == material:
            idx = category_type_to_index.get(category_type, None)
            if idx is not None:
                match = entry['Categories'][idx]
            break
    ws[f'D{report_row}'] = match

wb.save(output_path)
