import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

def parse_frequency(val):
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        try:
            return int(float(val))
        except Exception:
            return None
    return None

def parse_increase(val):
    try:
        return float(val)
    except Exception:
        return 0.0

def increment_months(dt, n):
    # Advances the date 'dt' by n months
    y, m = dt.year, dt.month + n
    while m > 12:
        y += 1
        m -= 12
    while m < 1:
        y -= 1
        m += 12
    return dt.replace(year=y, month=m)

wb = openpyxl.load_workbook('results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_8/task_44017/r2/evolve_44017/input.xlsx')
ws = wb.active

col_start = 30  # AD
col_end = 42    # AO (exclusive)
target_rows = range(14, 43)
date_row = 9

date_cols = list(range(col_start, col_end))

# Search all likely columns for a real datetime in row 9 (Z to AO)
date_seed = None
for c in range(26, 42):
    v = ws.cell(row=date_row, column=c).value
    if isinstance(v, datetime):
        date_seed = v
        break
if not date_seed:
    # If still not found, just use July 1, 2022 as a fallback
    date_seed = datetime(2022, 7, 1)

actual_dates = [date_seed]
while len(actual_dates) < len(date_cols):
    prev = actual_dates[-1]
    year, month, day = prev.year, prev.month, 1
    month += 1
    if month > 12:
        year += 1
        month = 1
    actual_dates.append(datetime(year, month, day))

for row in target_rows:
    freq = ws.cell(row=row, column=10).value
    freq_n = parse_frequency(freq)
    eff_date = ws.cell(row=row, column=12).value
    base_rate = ws.cell(row=row, column=23).value
    incrs_raw = [ws.cell(row=row, column=c).value for c in range(13,17)]
    increases = [parse_increase(x) for x in incrs_raw]
    if not (isinstance(eff_date, datetime) and freq_n and base_rate):
        for col in date_cols:
            ws.cell(row=row, column=col, value=None)
        continue
    phases = [eff_date]
    for i in range(1, 4):
        phases.append(increment_months(eff_date, freq_n * i))
    for col_idx, (col, phase_dt) in enumerate(zip(date_cols, actual_dates)):
        if phase_dt < eff_date:
            ws.cell(row=row, column=col, value=None)
            continue
        # Figure out how many increases have phased in
        n_incr = 0
        for i, start in enumerate(phases):
            if phase_dt >= start and increases[i] != 0:
                n_incr += 1
        value = base_rate
        for i in range(n_incr):
            value *= (1 + increases[i])
        ws.cell(row=row, column=col, value=round(value,2))

wb.save('results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_8/task_44017/r2/evolve_44017/output.xlsx')
