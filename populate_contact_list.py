from openpyxl import load_workbook
from datetime import datetime

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/regression_gate/after_pass/core_41589/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/regression_gate/after_pass/core_41589/output.xlsx"

wb = load_workbook(input_path)
ws = wb["Contact List"]

now = datetime.now()

for row in range(4, ws.max_row + 1):  # starting from row 4
    last_contact = ws[f"H{row}"].value
    yn_value = ws[f"I{row}"].value
    result = "NO ACTION"

    date_ok = False
    if isinstance(last_contact, datetime):
        date_ok = True
    elif last_contact is not None:
        try:
            last_contact = datetime.strptime(str(last_contact), "%Y-%m-%d")  # try ISO format
            date_ok = True
        except:
            try:
                last_contact = datetime.strptime(str(last_contact), "%m/%d/%Y")  # US format
                date_ok = True
            except:
                date_ok = False

    yn_value_clean = str(yn_value).strip().upper() if yn_value is not None else None

    if yn_value_clean == "YES" and date_ok:
        if (now - last_contact).days <= 30:
            result = "HOLD"
        else:
            result = "TOUCH BASE"
    else:
        result = "NO ACTION"
    ws[f"J{row}"] = result

wb.save(output_path)
