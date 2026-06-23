import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/regression_gate/after_pass/core_84-40/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/regression_gate/after_pass/core_84-40/output.xlsx'
wb = openpyxl.load_workbook(input_path)
list_ws_name = None
for ws in wb.worksheets:
    if ws.title.strip().lower() == 'list':
        list_ws_name = ws.title
        break
if list_ws_name is None:
    raise ValueError('No sheet named list or LIST found')
list_ws = wb[list_ws_name]
# Clear data in C2:E1000
for row in list_ws.iter_rows(min_row=2, max_row=1000, min_col=3, max_col=5):
    for cell in row:
        cell.value = None
# Collect data from each sheet (excluding 'LIST')
sheet_summaries = []
for ws in wb.worksheets:
    if ws.title == list_ws_name:
        continue
    sum_c = 0
    sum_d = 0
    for row in ws.iter_rows(min_row=1):
        try:
            c = row[2].value if len(row) > 2 else None
            d = row[3].value if len(row) > 3 else None
            if isinstance(c, (int, float)):
                sum_c += c
            if isinstance(d, (int, float)):
                sum_d += d
        except Exception:
            continue
    sheet_summaries.append((ws.title, sum_c, sum_d))
sheet_summaries.sort()
# Write results to LIST
total_c = 0
total_d = 0
for idx, (sheet, sum_c, sum_d) in enumerate(sheet_summaries, start=2):
    list_ws[f'B{idx}'] = sheet
    list_ws[f'C{idx}'] = sum_c
    list_ws[f'D{idx}'] = sum_d
    list_ws[f'E{idx}'] = sum_c - sum_d
    total_c += sum_c
    total_d += sum_d
# Total row
total_row = len(sheet_summaries) + 2
list_ws[f'B{total_row}'] = 'Total'
list_ws[f'C{total_row}'] = total_c
list_ws[f'D{total_row}'] = total_d
list_ws[f'E{total_row}'] = total_c - total_d
wb.save(output_path)
