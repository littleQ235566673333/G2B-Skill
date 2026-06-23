import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_2/regression_gate/after_pass/core_50796/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_2/regression_gate/after_pass/core_50796/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Collect preorder sums based on column B and values in column C for rows 12-22
preorders = {}
for row in range(12, 23):  # inclusive of row 22
    key = ws.cell(row, 2).value  # Column B
    val = ws.cell(row, 3).value  # Column C
    if key is not None and isinstance(val, (int, float)):
        preorders[key] = preorders.get(key, 0) + val

# For output (rows 2-5), column B needs the total for each descriptor in col A
for out_row in range(2, 6):  # 2,3,4,5
    descriptor = ws.cell(out_row, 1).value  # Column A
    ws.cell(out_row, 2).value = preorders.get(descriptor, 0)

wb.save(output_path)
