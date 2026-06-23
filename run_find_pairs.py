import re
from openpyxl import load_workbook

def is_vowel(c):
    return c.upper() in 'AEIOU'

def find_pairs(words):
    seen = set()
    pairs = []
    vowels = 'AEIOU'
    words_set = set(words)
    for w in words:
        s_indices = [i for i, ch in enumerate(w) if ch.upper() == 'S']
        if not s_indices:
            continue
        for v in vowels:
            candidate = ''.join(
                v if i in s_indices else ch for i, ch in enumerate(w))
            if candidate != w and candidate in words_set and candidate.upper() != w.upper():
                # Check only S/vowel diffs
                valid = True
                for i in range(len(w)):
                    if w[i].upper() == 'S':
                        if candidate[i].upper() != v:
                            valid = False
                            break
                    else:
                        if candidate[i] != w[i]:
                            valid = False
                            break
                if valid:
                    p = tuple(sorted([w, candidate], key=lambda x: x.upper()))
                    if p not in seen:
                        seen.add(p)
                        pairs.append(p)
    return pairs

# Load workbook and detect dictionary column
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/regression_gate/before_fix/core_146-49/input.xlsx')
ws = wb['Sheet1']
columns = ws.max_column
rows = ws.max_row
dict_col = None
for col in range(1, columns+1):
    col_vals = [ws.cell(row=r, column=col).value for r in range(1, min(21, rows+1))]
    vals = [v for v in col_vals if isinstance(v, str) and v.strip() != '']
    if len(vals) > 0 and all(re.fullmatch(r'[A-Za-z]+', v) for v in vals):
        dict_col = col
        break
if not dict_col:
    raise Exception('No dictionary column found')
    
words = []
for r in range(1, rows+1):
    val = ws.cell(row=r, column=dict_col).value
    if isinstance(val, str) and val.strip():
        words.append(val.strip())

pairs = find_pairs(words)
pairs = sorted(pairs, key=lambda pair: (pair[0].upper(), pair[1].upper()))

# Write to G/H, start at G2/H2, but final result sorted to G1/H1
start_row = 2
output_col1 = 7 # G
output_col2 = 8 # H
for i, (a, b) in enumerate(pairs):
    ws.cell(row=start_row + i, column=output_col1).value = a
    ws.cell(row=start_row + i, column=output_col2).value = b

# Sort output pairs and move to G1/H1
from operator import itemgetter
total_pairs = len(pairs)
output_vals = [(ws.cell(row=start_row + i, column=output_col1).value,
                ws.cell(row=start_row + i, column=output_col2).value) for i in range(total_pairs)]
output_vals.sort(key=lambda pair: (pair[0].upper(), pair[1].upper()))
for i, (v1, v2) in enumerate(output_vals):
    ws.cell(row=1 + i, column=output_col1).value = v1
    ws.cell(row=1 + i, column=output_col2).value = v2

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/regression_gate/before_fix/core_146-49/output.xlsx')
