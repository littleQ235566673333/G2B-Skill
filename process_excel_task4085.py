import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_6/regression_gate/after_pass/core_408-5/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_6/regression_gate/after_pass/core_408-5/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
rows = list(ws.iter_rows(values_only=False))
header_row = [cell.value for cell in rows[0]]

# Get column indexes (0-based)
def get_idx(col_name):
    try:
        return header_row.index(col_name)
    except ValueError:
        return None

col_c_idx = get_idx('C') or 2
col_d_idx = get_idx('D') or 3
col_e_idx = get_idx('E') or 4

# Find sum of D where E is 0 (excluding header)
sum_d_e0 = 0
for r in rows[1:]:
    e_val = r[col_e_idx].value
    d_val = r[col_d_idx].value
    if e_val == 0 and d_val is not None:
        try:
            sum_d_e0 += float(d_val)
        except Exception:
            pass

# Find rows to keep and BR1 Sales row index
data_rows = []
br1_row_idx = None
for i, r in enumerate(rows[1:], start=2):
    c_val = r[col_c_idx].value if col_c_idx is not None else None
    e_val = r[col_e_idx].value
    # Check for blank row
    row_is_blank = all((cell.value is None or (isinstance(cell.value, str) and not cell.value.strip())) for cell in r)
    if row_is_blank:
        continue
    # Keep BR1 Sales row and save its row index for writing the sum
    if c_val is not None and 'BR1 Sales' in str(c_val):
        br1_row_idx = i
        data_rows.append((i, r))
    elif e_val != 0:
        data_rows.append((i, r))

# Overwrite D value next to BR1 Sales with the sum
if br1_row_idx is not None:
    ws.cell(row=br1_row_idx, column=col_d_idx+1).value = sum_d_e0

# Remove all non-header rows
for i in range(ws.max_row, 1, -1):
    ws.delete_rows(i, 1)

# Re-add only kept data rows
insert_idx = 2
for _, r in data_rows:
    for ci, cell in enumerate(r, 1):
        ws.cell(row=insert_idx, column=ci).value = cell.value
    insert_idx += 1

wb.save(output_path)
