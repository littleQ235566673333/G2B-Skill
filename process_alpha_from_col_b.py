import re
from openpyxl import load_workbook

# Define input/output paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_8/regression_gate/after_pass/core_290-27/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_8/regression_gate/after_pass/core_290-27/output.xlsx"

wb = load_workbook(input_path)
ws = wb.active

# Helper function to check if cell value is a date
from openpyxl.utils.datetime import from_excel, CALENDAR_MAC_1904
from datetime import datetime

def is_date(cell):
    if cell.is_date:
        return True
    # Also check for datetime type in value (in case cell.is_date isn't reliable)
    return isinstance(cell.value, datetime)

# Apply to A14:I137 (Excel uses 1-based indices)
for row in ws.iter_rows(min_row=14, max_row=137, min_col=2, max_col=2):
    cell = row[0]
    if cell.value is None:
        continue
    if is_date(cell):
        continue
    val = str(cell.value)
    # Pattern: 2 or 3 uppercase letters, then optional space, then digits (rest)
    # Remove those leading letters and space.
    # Keep values like '1', '2018-07-12', '12/31/2021' as-is.
    # Eg: 'PID1'->'1', 'GG 1'->'1', 'PID2'->'2'
    new_val = re.sub(r'^[A-Z]{2,3}\s?(?=\d)', '', val)
    cell.value = new_val

wb.save(output_path)
