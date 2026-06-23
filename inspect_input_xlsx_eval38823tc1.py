from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun1/eval_38823_tc1/input.xlsx'
wb = load_workbook(input_path)
ws = wb.active
header = [cell.value for cell in ws[1]]
print('HEADER:', header)
# Print first 10 rows in first 10 columns for visual check
for r in range(1, 11):
    print([ws.cell(row=r, column=c).value for c in range(1, 11)])
