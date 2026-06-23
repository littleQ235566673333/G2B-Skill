import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/eval_seed42_rerun2/eval_58723_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/eval_seed42_rerun2/eval_58723_tc1/output.xlsx'

df = pd.read_excel(input_path)

# Standardize Entry Time to string if it's datetime
if pd.api.types.is_datetime64_any_dtype(df['Entry Time']):
    df['Entry Time'] = df['Entry Time'].dt.time.astype(str)

# Combine Entry Date and Entry Time to get full Timestamp
full_entry = pd.to_datetime(df['Entry Date'].astype(str) + ' ' + df['Entry Time'].astype(str), errors='coerce')
df['FullEntry'] = full_entry
labels = [''] * len(df)

for name in df['Name'].unique():
    records = df[df['Name'] == name]
    # Only consider rows with a valid FullEntry
    valid_entries = records['FullEntry'].notna()
    if valid_entries.any():
        max_time = records.loc[valid_entries, 'FullEntry'].max()
        idx_latest = records[records['FullEntry'] == max_time].index
        for ix in records.index:
            if ix in idx_latest:
                labels[ix] = 'Latest'
            else:
                labels[ix] = 'Not Latest'
    else:
        for ix in records.index:
            labels[ix] = ''

# Write to column 'M' (13th column, cells M2:M41)
wb = load_workbook(input_path)
ws = wb.active
for r, value in enumerate(labels, start=2):
    ws[f'M{r}'] = value
wb.save(output_path)
