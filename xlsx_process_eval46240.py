import openpyxl
from datetime import datetime

# Paths
input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_46240_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_46240_tc1/output.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 5):  # Rows 2 to 4
    h_val = ws[f'H{row}'].value
    if h_val is None or (isinstance(h_val, str) and h_val.strip() == ''):
        result = 'No'
    elif isinstance(h_val, str):
        if h_val.strip().lower() == 'n/a':
            result = 'N/A'
        elif h_val.strip().lower() == 'yes':
            result = 'Yes'
        else:
            # Try to parse as date
            try:
                dt = datetime.strptime(h_val.strip(), '%Y-%m-%d')
                result = 'Yes'
            except Exception:
                result = 'No'
    elif isinstance(h_val, datetime):
        result = 'Yes'
    else:
        result = 'Yes'  # For Excel date type or other detected dates
    ws[f'J{row}'] = result

wb.save(output_path)
