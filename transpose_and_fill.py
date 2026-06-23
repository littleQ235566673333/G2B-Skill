import openpyxl
import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/group_290-1/r0/evolve_290-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/group_290-1/r0/evolve_290-1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Find out how columns K:U (11-21) are filled for rows 2, 3. Existing ones show a pattern.
def fill_transpose(start_row, src_idx, dst_col_offset, length):
    # Copies value from start_row, columns 8/9 (H, I) downward to K:U (11-21)
    for i in range(length):
        src_row = start_row + i
        # The letter and number alternate in destination columns
        number = ws.cell(row=src_row, column=9).value  # I
        letters = ws.cell(row=src_row, column=8).value  # H
        ws.cell(row=start_row, column=11 + dst_col_offset + 2 * i, value=number)
        ws.cell(row=start_row, column=12 + dst_col_offset + 2 * i, value=letters)

# Process all rows 2-10; the first two are already given, fill as in the pattern for following rows
fill_transpose(4, 8, 0, 3)
fill_transpose(5, 8, 0, 2)
fill_transpose(6, 8, 0, 1)
fill_transpose(7, 8, 0, 1)
fill_transpose(8, 8, 0, 1)
fill_transpose(9, 8, 0, 1)

wb.save(output_path)
print('Transposed and saved to output.')
