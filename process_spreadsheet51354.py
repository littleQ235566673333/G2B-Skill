import openpyxl
from openpyxl.styles import PatternFill
import calendar
import re

def parse_month_year(text):
    # Find 'Mon YY' or 'MonYY' at end
    match = re.search(r'([A-Za-z]{3})\s?([0-9]{2})$', text)
    if match:
        mon, year = match.group(1), match.group(2)
        return mon, year
    return None, None

def increment_month(mon, year):
    try:
        month_number = list(calendar.month_abbr).index(mon.title())
        year_num = int(year)
        if month_number == 0:
            return '', year
        month_number += 1
        if month_number > 12:
            month_number = 1
            year_num += 1
        return calendar.month_abbr[month_number], f'{year_num % 100:02}'
    except:
        return '', year

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_51354_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_51354_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

# Process rows 2 to 6
for row in range(2, 7):
    value = ws[f'A{row}'].value or ''
    mon, year = parse_month_year(value)

    # Column D: just the year, apply color
    ws[f'D{row}'] = year
    ws[f'D{row}'].fill = fill

    # Column E: increment month and format as 'MMM YY'
    imon, iyear = increment_month(mon, year)
    ws[f'E{row}'] = f'{imon} {iyear}'.strip() if imon else ''

wb.save(output_path)
