import openpyxl
from datetime import time
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_4/regression_gate/before_fix/core_52305/input.xlsx')
ws = wb['Sheet1']

# Map headers
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
header_to_col = {h: i+1 for i,h in enumerate(header_row)}

# Read Name headers J5:N5
name_headers = [ws.cell(row=5, column=col).value for col in range(10, 15)]
# Read time intervals H6:H24, I6:I24
time_starts = [ws.cell(row=row, column=8).value for row in range(6, 25)]
time_ends = [ws.cell(row=row, column=9).value for row in range(6, 25)]

# Read main data rows (A2:B1000+)
data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[2] is not None and row[1] is not None:  # Name and Time
        data.append({'Type': row[0], 'Time': row[1], 'Name': row[2], 'Destination': row[3]})

def time_only(dt):
    '''Extract python time from datetime/time, or None.'''
    if isinstance(dt, time):
        return dt
    elif hasattr(dt, 'time'):
        return dt.time()
    return None

# Compute results for J6:N24
result_matrix = []
for t_start, t_end in zip(time_starts, time_ends):
    row_list = []
    for name in name_headers:
        count = 0
        for rec in data:
            entry_time = time_only(rec['Time'])
            if entry_time is None or name is None:
                continue
            cond_name = (rec['Name'] == name)
            cond_time = False
            # Only count if time_start/end are set
            if t_start is not None and t_end is not None and entry_time is not None:
                cond_time = (t_start <= entry_time <= t_end)
            if cond_name and cond_time:
                count += 1
        row_list.append(count)
    result_matrix.append(row_list)

# Write to J6:N24
for i, row_vals in enumerate(result_matrix):
    for j, val in enumerate(row_vals):
        ws.cell(row=6 + i, column=10 + j, value=val)
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_4/regression_gate/before_fix/core_52305/output.xlsx')
print('Wrote heat-map table to J6:N24!')
