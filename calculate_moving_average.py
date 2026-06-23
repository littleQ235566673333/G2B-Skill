import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_3/group_4714/r0/evolve_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_3/group_4714/r0/evolve_4714/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet2']

data = []
for row in ws.iter_rows(min_row=2, max_row=25, min_col=1, max_col=4, values_only=False):
    emp = row[0].value
    month = row[2].value
    hours = row[3].value
    if emp is not None and month is not None and hours is not None:
        data.append({'row': row[0].row, 'emp': emp, 'month': month, 'hours': hours})

emp_map = defaultdict(list)
for entry in data:
    emp_map[entry['emp']].append(entry)
for emp in emp_map:
    emp_map[emp].sort(key=lambda x: x['month'])

for entry in data:
    emp = entry['emp']
    cur_month = entry['month']
    row_idx = entry['row']
    rows_for_emp = emp_map[emp]
    # Get only up to the current row's month
    window = [e for e in rows_for_emp if e['month'] <= cur_month]
    # Keep only last 4
    window = window[-4:]
    if len(window) < 4:
        ws[f'E{row_idx}'] = 'n/a'
    else:
        avg = sum(e['hours'] for e in window) / 4
        ws[f'E{row_idx}'] = int(avg) if avg == int(avg) else round(avg, 1)

wb.save(output_path)
print('Done')
