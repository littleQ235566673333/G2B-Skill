import openpyxl
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Input/output paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_4/group_44017/r0/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_4/group_44017/r0/evolve_44017/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Constants for columns
ROW_START, ROW_END = 14, 42               # Output rows
COL_START, COL_END = 30, 41               # AD (30) to AO (41)
DATE_ROW = 9                              # Dates in row 9
BASE_RATE_COL = 23                        # W
EFF_DATE_COL = 12                         # L
FREQ_COL = 10                             # J
INCREASE_COLS = [13, 14, 15, 16]          # M, N, O, P

def to_datetime(value):
    if isinstance(value, str):
        try:
            return pd.to_datetime(value)
        except:
            return None
    if isinstance(value, (datetime, pd.Timestamp)):
        return pd.to_datetime(value)
    return None

# Read column dates (row 9, AD:AO)
dates = [to_datetime(ws.cell(row=DATE_ROW, column=col).value) for col in range(COL_START, COL_END+1)]

for row in range(ROW_START, ROW_END+1):
    base_rate = ws.cell(row=row, column=BASE_RATE_COL).value
    eff_date = to_datetime(ws.cell(row=row, column=EFF_DATE_COL).value)
    freq = ws.cell(row=row, column=FREQ_COL).value
    inc_pcts = [ws.cell(row=row, column=col).value for col in INCREASE_COLS]
    for i in range(len(inc_pcts)):
        if inc_pcts[i] is not None and inc_pcts[i] > 1:
            inc_pcts[i] = inc_pcts[i] / 100.0
        elif inc_pcts[i] is None:
            inc_pcts[i] = 0.0
    for cidx, col in enumerate(range(COL_START, COL_END+1)):
        cell_date = dates[cidx]
        if not cell_date or not eff_date or cell_date < eff_date:
            ws.cell(row=row, column=col).value = None
            continue
        # How many increases are in effect
        n_eff = 0
        for nth in range(4):
            if freq is not None:
                increase_effective_date = eff_date + pd.DateOffset(months=freq * nth)
            else:
                increase_effective_date = eff_date
            if cell_date >= increase_effective_date and inc_pcts[nth]:
                n_eff += 1
            else:
                break
        # Calculate cumulative rate
        cum_factor = 1.0
        for nth in range(n_eff):
            cum_factor *= (1 + inc_pcts[nth])
        new_rate = base_rate * cum_factor if base_rate is not None else None
        ws.cell(row=row, column=col).value = round(new_rate, 2) if new_rate is not None else None
        # Remove yellow fill if present
        cell = ws.cell(row=row, column=col)
        cell.fill = openpyxl.styles.PatternFill(fill_type=None)

wb.save(output_path)
