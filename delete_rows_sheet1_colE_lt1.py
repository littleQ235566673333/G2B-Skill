import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_374-18_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_374-18_tc1/output.xlsx'

# Load workbook and target sheet (Sheet1 is the working sheet, not Imported Data)
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Identify rows to delete (where col E/outstanding < 1, keep header)
rows_to_delete = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=5).value
    try:
        if val is not None and float(val) < 1:
            rows_to_delete.append(row)
    except (TypeError, ValueError):
        continue

# Delete rows in reverse order to maintain row indices
for row in reversed(rows_to_delete):
    ws.delete_rows(row)

# Save after deletion
wb.save(output_path)

# Copy A2:G6 from cleaned Sheet1 to same positions in Sheet1
wb2 = openpyxl.load_workbook(output_path)
ws2 = wb2['Sheet1']

rows_to_copy = list(ws2.iter_rows(min_row=2, max_row=6, max_col=7, values_only=True))
for i, row in enumerate(rows_to_copy, start=2):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=val)
wb2.save(output_path)
