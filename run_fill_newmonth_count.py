import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_4/group_45707/r3/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_4/group_45707/r3/evolve_45707/output.xlsx'
df = pd.read_excel(input_path, engine='openpyxl')
wb = load_workbook(input_path)
ws = wb.active
dates = pd.to_datetime(df.iloc[:,0], errors='coerce')
col_C = df.iloc[:,2]
for row in range(2, 70):  # D2:D69 => row 2 to 69 (inclusive)
    idx_next = row  # Next row index (0-based)
    if idx_next >= len(dates):
        ws.cell(row=row, column=4).value = None
        continue
    next_day = dates.iloc[idx_next]
    if pd.isnull(next_day):
        ws.cell(row=row, column=4).value = None
        continue
    if next_day.day == 1:
        month_mask = (dates.dt.month == next_day.month) & (dates.dt.year == next_day.year)
        count_ones = (col_C[month_mask] == 1).sum()
        ws.cell(row=row, column=4).value = int(count_ones)
    else:
        ws.cell(row=row, column=4).value = None
wb.save(output_path)
