import openpyxl
from openpyxl.styles import PatternFill

# Input and output file paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_6/regression_gate/after_pass/core_51249/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_6/regression_gate/after_pass/core_51249/output.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Green fill for cells, RGB: 226-239-218 = E2EFDA
fill = PatternFill(fill_type="solid", fgColor="E2EFDA")

# Specify the target cells for output
result_cells = ['D1', 'D5', 'D9']

for cell in result_cells:
    # Determine corresponding B row
    row = ws[cell].row
    b1_cell = f'B{row}'
    b2_cell = f'B{row+1}'

    b1 = ws[b1_cell].value or ""
    b2 = ws[b2_cell].value or ""

    # Logic as per user spec
    if b1 == 'Description A' and (b2 == "" or b2 is None):
        ws[cell] = 'Single A'
    elif b1 == 'Description B' and (b2 == "" or b2 is None):
        ws[cell] = 'Single B'
    elif b1 == 'Description A' and b2 == 'Description B':
        ws[cell] = 'Multiple'
    # If you want to clear otherwise, do this, else skip setting value
    # else:
    #    ws[cell] = ''

    # Apply fill if cell now has content
    if ws[cell].value not in (None, ""):
        ws[cell].fill = fill

wb.save(output_path)
