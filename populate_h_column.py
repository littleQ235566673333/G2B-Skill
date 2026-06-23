import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta

# File paths
input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_50631_tc1/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_50631_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_fp)
ws = wb['Sheet1']

# Styles
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
calibri_font = Font(name='Calibri')

# Helper to read holidays in column J
holiday_dates = []
for row in range(7, ws.max_row+1):
    cell = ws[f'J{row}']
    if cell.value:
        try:
            holiday_dates.append(cell.value.date() if isinstance(cell.value, datetime) else datetime.strptime(str(cell.value), '%m/%d/%Y').date())
        except Exception:
            continue

# Get start and end dates from B3 and F3
start_date = ws['B3'].value
end_date = ws['F3'].value
if not isinstance(start_date, datetime):
    start_date = datetime.strptime(str(start_date), '%m/%d/%Y')
if not isinstance(end_date, datetime):
    end_date = datetime.strptime(str(end_date), '%m/%d/%Y')

start_dt = start_date.date()
end_dt = end_date.date()

# Helper: is_workday
def is_workday(date):
    return date.weekday() < 5 and date not in holiday_dates

# Generate workdays between start and end
workdays = []
date = start_dt
while date <= end_dt:
    if is_workday(date):
        workdays.append(date)
    date += timedelta(days=1)

# Fill H7:H38
h_col = 'H'
h_start = 7
h_end = 38
work_idx = 0
for row in range(h_start, h_end+1):
    h_cell = ws[f'{h_col}{row}']
    h_cell.font = calibri_font
    # Check if cell is filled yellow
    if h_cell.fill.start_color.rgb == 'FFFF00' or h_cell.fill.start_color.index == 'FFFF00':
        # Assign next workday if available
        if work_idx < len(workdays):
            h_cell.value = workdays[work_idx]
            h_cell.fill = yellow_fill
            work_idx += 1
        else:
            h_cell.value = None
    else:
        h_cell.value = None

# Hide gridlines
ws.sheet_view.showGridLines = False

# Save to output
wb.save(output_fp)
