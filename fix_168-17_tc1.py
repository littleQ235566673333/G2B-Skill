from openpyxl import load_workbook
input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed0/eval_168-17_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed0/eval_168-17_tc1/output.xlsx'
wb = load_workbook(input_path)
sheet_name = 'Statement'
if sheet_name not in wb.sheetnames:
    raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
ws = wb[sheet_name]
row_found = None
for i, row in enumerate(ws.iter_rows(min_row=1, max_col=1), start=1):
    cell_val = row[0].value
    if cell_val == 'Invoice No.':
        row_found = i
        break
if row_found is None:
    raise ValueError("'Invoice No.' not found in column A of 'Statement' sheet.")
if row_found > 1:
    ws.delete_rows(1, row_found - 1)
wb.save(output_path)
