import openpyxl

# Load the workbook and select the first worksheet
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_57743_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_57743_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Build lookup dictionary from column C (model) and D (price)
model_to_price = {}
for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
    model = row[0].value
    price = row[1].value
    if model is not None:
        model_to_price[model] = price

# For each model in A2:A19, lookup and write price to B2:B19
for i in range(2, 20):  # Rows 2 to 19 inclusive
    model_num = ws[f'A{i}'].value
    price = model_to_price.get(model_num, "")  # Blank if not found
    ws[f'B{i}'] = price

wb.save(output_path)
