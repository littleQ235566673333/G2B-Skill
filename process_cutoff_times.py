import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, time, timedelta

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_35739_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_35739_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

def get_excel_time(dt):
    '''Convert a time or datetime to Excel float time.'''
    if isinstance(dt, datetime):
        dt = dt.time()
    return (dt.hour * 3600 + dt.minute * 60 + dt.second) / 86400

for row in range(2, 101):
    a_val = ws[f'A{row}'].value
    b_val = ws[f'B{row}'].value
    cell = ws[f'C{row}']
    # If A or B is blank, result should be blank
    if not a_val or not b_val:
        cell.value = None
        continue
    # Try to parse B value as time
    b_time = None
    if isinstance(b_val, float) and 0 <= b_val < 1:
        # Excel float time
        total_seconds = int(round(b_val * 86400))
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        b_time = time(hours, minutes, seconds)
    elif isinstance(b_val, datetime):
        b_time = b_val.time()
    elif isinstance(b_val, time):
        b_time = b_val
    if b_time is None:
        cell.value = None
        continue
    # Calculate cut off (handle crossing midnight)
    cut_time_dt = datetime.combine(datetime(2000,1,1), b_time) - timedelta(minutes=30)
    cutoff_time = cut_time_dt.time()
    cell.value = get_excel_time(cutoff_time)
    cell.number_format = 'h:mm AM/PM'

wb.save(output_path)
