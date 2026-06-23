import sys
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun2/eval_55708_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun2/eval_55708_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Find header row and columns
header_row = None
for row in ws.iter_rows(min_row=1, max_row=10):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and 'department' in cell.value.lower():
            header_row = cell.row
            break
    if header_row:
        break
assert header_row is not None
header_cells = list(ws[header_row])
headers = {cell.value: idx for idx, cell in enumerate(header_cells)}
# Map header names (case-insensitive, typo-tolerant for 'turnaroud')
header_map = {str(k).lower(): k for k in headers}
dept_key   = header_map[[k for k in header_map if 'department' in k][0]]
stat_key   = header_map[[k for k in header_map if 'status' in k][0]]
tat_key    = header_map[[k for k in header_map if ('turnaroud' in k and 'time' in k)] [0]]
idx_dept, idx_stat, idx_tat = headers[dept_key], headers[stat_key], headers[tat_key]
data = []
for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, values_only=True):
    if all(x is None for x in row): continue
    data.append(row)
depts = [ws.cell(row=11+i, column=1).value for i in range(3)]
results = []
for dept in depts:
    if not dept:
        results.append('')
        continue
    vals = [r[idx_tat] for r in data if r[idx_dept] == dept and r[idx_stat] in ('In Progress','In Review') and r[idx_tat] is not None and r[idx_tat] >= 6]
    if vals:
        results.append(sum(vals)/len(vals))
    else:
        results.append('')
for i, val in enumerate(results):
    ws.cell(row=11+i, column=2, value=val if val != '' else None)
    ws.cell(row=11+i, column=2).number_format = 'General'
wb.save(output_path)
