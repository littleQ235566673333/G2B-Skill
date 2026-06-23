from openpyxl import load_workbook

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_4/group_290-1/r0/evolve_290-1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_4/group_290-1/r0/evolve_290-1/output.xlsx"

wb = load_workbook(input_path)
ws = wb.active

# Columns K-U are 11..21
src_cols = list("KLMNOPQRSTU")
row_start, row_end = 2, 10
col_start, col_end = 11, 21  # K=11, U=21

# Preserve answer examples already shown in K2:U10
pre_filled = {}
for r in range(row_start, row_end+1):
    for c in range(col_start, col_end+1):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            pre_filled[(r, c)] = v

# Identify the source column containing the data to transpose
all_data = []
for col in range(1, ws.max_column+1):
    if ws.cell(row=1, column=col).value is None:
        continue
    col_letter = ws.cell(row=1, column=col).column_letter
    if col_letter in src_cols:
        continue
    values = []
    for row in range(row_start, ws.max_row+1):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            values.append(v)
    if len(values) >= (row_end - row_start + 1)*(col_end - col_start + 1):
        all_data = values[:(row_end - row_start + 1)*(col_end - col_start + 1)]
        break

if all_data:
    num_cols = col_end - col_start + 1
    num_rows = row_end - row_start + 1
    idx = 0
    for r in range(row_start, row_end+1):
        for c in range(col_start, col_end+1):
            if (r, c) not in pre_filled:
                ws.cell(row=r, column=c).value = all_data[idx]
            idx += 1

wb.save(output_path)
