import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_30930_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_30930_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

start_row = 2
end_row = 66

# We'll be interested in column A (1), column B (2), column C (3)
count = 0
last_1_row = start_row - 1  # The row before start_row, for correct range

for row in range(start_row, end_row + 1):
    b_val = ws.cell(row=row, column=2).value

    if b_val == 1:
        # Count values in column A from (last_1_row + 1) to (row - 1)
        positive_count = 0
        for r in range(last_1_row + 1, row):
            a_val = ws.cell(r, column=1).value
            if a_val is not None and isinstance(a_val, (int, float)) and a_val > 0:
                positive_count += 1
        # Write result to Column C of this row
        ws.cell(row=row, column=3, value=positive_count)
        last_1_row = row
    else:
        ws.cell(row=row, column=3, value=None)

wb.save(output_path)