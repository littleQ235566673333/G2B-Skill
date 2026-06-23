import openpyxl
from openpyxl.styles import PatternFill, Font

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_48620_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_48620_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# First, read all products and properties
data = list(ws.iter_rows(min_row=2, max_col=2, values_only=True))
products = [(row[0], row[1]) for row in data if row[0] is not None]

# Read the search values in D2:D7
search_terms = [ws.cell(row=row, column=4).value for row in range(2, 8)]

# For each search, find all matches in same row and join with ", ", or blank if no match
results = []
for term in search_terms:
    matches = [p[1] for p in products if p[0] == term and p[1] is not None]
    results.append(', '.join(matches) if matches else '')

# Write results to E2:E7, display 0 as blank
for i, val in enumerate(results, start=2):
    if val == '0' or val == 0:
        ws.cell(row=i, column=5, value=None)
    else:
        ws.cell(row=i, column=5, value=val)

# Formatting: #FCE4D6 and Calibri for E2:E7
fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
font = Font(name='Calibri')
for i in range(2, 8):
    cell = ws.cell(row=i, column=5)
    cell.fill = fill
    cell.font = font
    if cell.value == '0' or cell.value == 0:
        cell.value = None  # Show 0s as blanks

wb.save(output_path)
