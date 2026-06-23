import openpyxl

# Load workbooks
in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_8/regression_gate/after_fix/core_48975/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_8/regression_gate/after_fix/core_48975/output.xlsx'
wb = openpyxl.load_workbook(in_path)
ws_in = wb['Input']
ws_out = wb['Output']

out_start_row = 11
out_col = 2  # Column B
written = 0

for row in ws_in.iter_rows(min_row=2):  # Skip header
    a = row[0].value  # Column A
    b = row[1].value  # Column B
    todo = row[4].value if len(row) > 4 else None  # Column E ('To Do')
    # Check for 'yes', no blanks in A or B
    if todo and str(todo).strip().lower() == 'yes' and a not in [None, ''] and b not in [None, '']:
        ws_out.cell(row=out_start_row+written, column=out_col, value=str(a))
        ws_out.cell(row=out_start_row+written, column=out_col+1, value=str(b))
        written += 1

wb.save(out_path)
