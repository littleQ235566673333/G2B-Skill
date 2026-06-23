import openpyxl

# Load the workbook and select active sheet
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/regression_gate/after_fix/core_48969/input.xlsx')
ws = wb.active

# Determine left and right table bounds by checking headers around B3 and F3
# Left Table assumed at B4:C5 (2 rows x 2 cols below B3:C3 headers)
left = [[ws.cell(row=i, column=j).value for j in range(2, 4)] for i in range(4, 6)]
# Right Table assumed at F4:G5 (2 rows x 2 cols below F3:G3 headers)
right = [[ws.cell(row=i, column=j).value for j in range(6, 8)] for i in range(4, 6)]

# Generate cell-by-cell boolean comparison matrix
rows = len(left)
cols = len(left[0])
comparison = [[left[r][c] == right[r][c] for c in range(cols)] for r in range(rows)]

# Write the result into cells J3:K4 (so that it appears as J3:K4, with header row in J3/K3 usually)
for r in range(rows):
    for c in range(cols):
        ws.cell(row=3 + r, column=10 + c, value=comparison[r][c])

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/regression_gate/after_fix/core_48969/output.xlsx')
