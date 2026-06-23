import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font

# Input and output paths
input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_36097_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_36097_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Map headers to column indexes
headers = {cell.value: idx for idx, cell in enumerate(ws[2], start=1)}

# Try to infer column positions for relevant fields
itv_col = headers.get('ITV')
orig_cost_col = headers.get('Original Cost')
sale_price_col = headers.get('Sale Price')

if not itv_col or not orig_cost_col or not sale_price_col:
    # Fallback: just pick columns by index as an assumption (A=1, B=2, ...) for columns D, E, F
    itv_col, orig_cost_col, sale_price_col = 4, 5, 6

start_row, end_row = 3, 6

for row in range(start_row, end_row+1):
    itv = ws.cell(row=row, column=itv_col).value or 0
    orig_cost = ws.cell(row=row, column=orig_cost_col).value or 0
    sale_price = ws.cell(row=row, column=sale_price_col).value or 0
    
    profit = sale_price - itv
    loss = sale_price < itv
    # Logic: If asset sold at a loss --> recoup ITV adjusted by the loss
    if loss:
        result = itv + (sale_price - itv)  # Actually sale_price (sale at loss), but keeping logic explicit
    # If profit is less than original cost: recoup the full profit
    elif profit < orig_cost:
        result = profit
    # Otherwise: recoup the cost basis reduced by ITV
    else:
        result = orig_cost - itv
    # Copy formatting from I
    src_cell = ws.cell(row=row, column=9)
    dest_cell = ws.cell(row=row, column=8)
    dest_cell.value = result
    if src_cell.has_style:
        dest_cell.font = Font(name=src_cell.font.name, size=src_cell.font.size, bold=src_cell.font.bold,
                             italic=src_cell.font.italic, vertAlign=src_cell.font.vertAlign,
                             underline=src_cell.font.underline, strike=src_cell.font.strike, color=src_cell.font.color)
        dest_cell.fill = PatternFill(fill_type=src_cell.fill.fill_type, fgColor=src_cell.fill.fgColor, bgColor=src_cell.fill.bgColor)
        dest_cell.border = Border(
            left=src_cell.border.left, right=src_cell.border.right,
            top=src_cell.border.top, bottom=src_cell.border.bottom
        )
        dest_cell.alignment = Alignment(
            horizontal=src_cell.alignment.horizontal, vertical=src_cell.alignment.vertical
        )
        dest_cell.number_format = src_cell.number_format

wb.save(output_file)
