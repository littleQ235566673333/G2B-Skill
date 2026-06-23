import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_5/regression_gate/before_fix/core_59129/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_5/regression_gate/before_fix/core_59129/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Assuming columns: A = Hire date, B = Termination date, header in row 1, data in rows 2-22
# E1:P1 are months for which headcount is to be calculated

hire_dates = [ws[f'A{row}'].value for row in range(2, 23)]
term_dates = [ws[f'B{row}'].value for row in range(2, 23)]

months = [ws[f'{get_column_letter(col)}1'].value for col in range(5, 17)] # E1:P1

def parse_month(val):
    # Accept datetime or parse common month formats
    if isinstance(val, datetime):
        return val.replace(day=1)
    strval = str(val)
    for fmt in ('%b %Y', '%B %Y', '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
        try:
            dt = datetime.strptime(strval, fmt)
            return dt.replace(day=1)
        except:
            continue
    return None
months = [parse_month(m) for m in months]

counts = []
for month in months:
    if not month:
        counts.append('')
        continue
    # month_end: last day of the month
    month_end = (month.replace(day=28) + pd.Timedelta(days=4)).replace(day=1) - pd.Timedelta(days=1)
    count = 0
    for hire, term in zip(hire_dates, term_dates):
        if not hire:
            continue
        # Must be hired on/before the first of the month
        if hire <= month:
            # Not terminated, or terminated after the end of the month
            if not term or term > month_end:
                count += 1
    counts.append(count)

# Write results to E2:P2
for i, val in enumerate(counts):
    ws.cell(row=2, column=5+i).value = val

wb.save(output_path)
