import openpyxl
import pandas as pd

# Load workbook and sheets
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_7/regression_gate/before_fix/core_45896/input.xlsx')
ws = wb['Volym P5_P6_2023']

# Load all data from ZORD sheet via pandas for ease
zord_df = pd.read_excel('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_7/regression_gate/before_fix/core_45896/input.xlsx', sheet_name='ZORD')

# Map for A (lookup col) and C (date col; 0-based)
zord_lookup = {}
for ix, row in zord_df.iterrows():
    key = row[0]
    date = row[2]
    if pd.notnull(date):
        zord_lookup.setdefault(key, []).append(date)

def format_date(dt):
    # Accept both pandas.Timestamp and datetime/date
    if pd.isnull(dt): return None
    if hasattr(dt, 'strftime'):
        return dt.strftime('%d/%m/%Y')
    try:
        return pd.to_datetime(dt).strftime('%d/%m/%Y')
    except:
        return str(dt)

# For each row 2 to 10 in Volym, get matches from ZORD and join dates
for row in range(2, 11):
    val = ws[f'A{row}'].value
    if val is None:
        ws[f'I{row}'] = None
        continue
    if val in zord_lookup:
        unique_dates = sorted(set(format_date(d) for d in zord_lookup[val] if d and format_date(d)))
        ws[f'I{row}'] = ','.join(unique_dates)
    else:
        ws[f'I{row}'] = ''

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_7/regression_gate/before_fix/core_45896/output.xlsx')
