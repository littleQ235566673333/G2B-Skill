import openpyxl
from datetime import datetime

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed2/eval_46240_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed2/eval_46240_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 5):
    val = ws[f'H{row}'].value
    # 1. If cell H contains 'N/A', J= 'N/A'
    if val == 'N/A':
        result = 'N/A'
    # 2. If H contains 'Yes' or a date:
    elif val == 'Yes':
        result = 'Yes'
    elif isinstance(val, datetime):
        result = 'Yes'
    elif isinstance(val, str):
        # Try to parse as date
        try:
            datetime.strptime(val, '%Y-%m-%d')  # try ISO, else fail
            result = 'Yes'
        except Exception:
            # Check for slash-separated date (common in Excel e.g., mm/dd/yyyy)
            try:
                datetime.strptime(val, '%m/%d/%Y')
                result = 'Yes'
            except Exception:
                # Not a date
                if val.strip() == '':
                    result = 'No'
                else:
                    result = 'No'
    elif val is None or (isinstance(val, str) and val.strip() == ''):
        result = 'No'
    else:
        result = 'No'
    ws[f'J{row}'] = result

wb.save(output_path)
