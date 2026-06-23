import openpyxl
from collections import defaultdict

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_250-20_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_250-20_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['RNM']

rows = list(ws.iter_rows(values_only=False))
header = [cell.value for cell in rows[0]]

# Find index of columns B, C, J
col_B = 1  # COLUMN B
col_C = 2  # COLUMN C
col_J = 9  # COLUMN J

# Dictionary to group rows by (B, C)
groups = defaultdict(list)
for row in rows[1:]:
    # Skip rows that are entirely empty
    if all(cell.value is None for cell in row):
        continue
    key = (row[col_B].value, row[col_C].value)
    groups[key].append(row)

new_rows = [header]
for key, group_cells in groups.items():
    # Sum column J (MATCHED_QTY) values for this group
    total = sum(cell[col_J].value or 0 for cell in group_cells)
    # Take the first row's values (to keep representative info)
    new_row_vals = [cell.value for cell in group_cells[0]]
    new_row_vals[col_J] = total  # Update MATCHED_QTY to sum
    new_rows.append(new_row_vals)

# Clear and write back:
for i, row_vals in enumerate(new_rows[:20]):
    for j, val in enumerate(row_vals):
        ws.cell(row=i+1, column=j+1, value=val)

# Remove rows beyond row 20 if present
total_rows = ws.max_row
if total_rows > 20:
    for i in range(total_rows, 20, -1):
        ws.delete_rows(i)

wb.save(output_path)
