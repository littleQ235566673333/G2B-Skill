from openpyxl import load_workbook

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_1/regression_gate/after_pass/core_302-1/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_1/regression_gate/after_pass/core_302-1/output.xlsx'

wb = load_workbook(in_path)
items_ws = wb['ITEMS']

# Read items column B from ITEMS sheet (rows 2-8)
items = [items_ws.cell(row=i, column=2).value for i in range(2, 9)]

for sheet_name in ['SHEET1', 'SHEET2']:
    ws = wb[sheet_name]
    for i, val in enumerate(items, start=2):
        ws.cell(row=i, column=2).value = val

wb.save(out_path)
