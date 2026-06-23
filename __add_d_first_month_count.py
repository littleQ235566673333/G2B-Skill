import pandas as pd
from openpyxl import load_workbook
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_1/group_45707/r3/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_1/group_45707/r3/evolve_45707/output.xlsx'

# Load with pandas to process logic
# Assume (A: dates, B: ?, C: values, D: insert, E: header present) - keep structure

df = pd.read_excel(input_path)
dates = pd.to_datetime(df.iloc[:,0], errors='coerce')
col_c = df.iloc[:,2]
results = [''] * len(df)

for i in range(1, len(df)):
    prev_day = dates.iloc[i-1]
    this_day = dates.iloc[i]
    if pd.notnull(this_day) and this_day.day == 1:
        # this is a first of the month, so mark in previous row
        # Count how many 1s in col_c for this month & year
        mask = (dates.dt.year == this_day.year) & (dates.dt.month == this_day.month)
        count_ones = ((col_c == 1) & mask).sum()
        results[i-1] = count_ones

# Now use openpyxl to write results in D2:D69 (1-based; D2 is row=2, col=4)
wb = load_workbook(input_path)
ws = wb.active
for idx, val in enumerate(results[:68], start=2):
    ws.cell(row=idx, column=4).value = val if val != '' else None
wb.save(output_path)
