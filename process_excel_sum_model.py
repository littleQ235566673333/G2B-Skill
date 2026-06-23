import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun2/eval_49945_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun2/eval_49945_tc1/output.xlsx'

# Read the data
wb = load_workbook(input_path)
ws = wb.active

target_cells = ['G4', 'G5', 'G6']
target_models = [ws[cell].value for cell in target_cells]

# Use pandas to get model and quantity data, assuming Model in col 2 ("C"), Qty in col 3 ("D")
df = pd.read_excel(input_path, header=None)
models = df.iloc[3:11, 2]  # C4:C11
quantities = df.iloc[3:11, 3]  # D4:D11
data = pd.DataFrame({'Model': models, 'Quantity': quantities})

for i, cell in enumerate(target_cells):
    model = target_models[i]
    if pd.isna(model) or model == None:
        ws[cell] = ''
    else:
        sum_qty = data.loc[data['Model'] == model, 'Quantity'].sum()
        ws[cell] = sum_qty if not pd.isna(sum_qty) else ''

wb.save(output_path)
