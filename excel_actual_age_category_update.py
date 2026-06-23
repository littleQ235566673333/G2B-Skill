from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_1/regression_gate/before_pass/core_32337/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_1/regression_gate/before_pass/core_32337/output.xlsx'

wb = load_workbook(in_path)
ws = wb.active

# Insert Actual Age column after age (year), which is col 15
target_col = 16
ws.insert_cols(target_col)

# Header and fill copy
ws.cell(row=2, column=target_col).value = 'Actual Age'
orig_fill = ws.cell(row=2, column=15).fill
# Create a new PatternFill with same attributes, if fill is present
if isinstance(orig_fill, PatternFill) and orig_fill.patternType:
    fill = PatternFill(patternType=orig_fill.patternType, fgColor=orig_fill.fgColor, bgColor=orig_fill.bgColor)
else:
    fill = None

for r in range(3, 16):
    ws.cell(row=r, column=target_col).value = '=DATEDIF($C{0},$B$1,"Y")'.format(r)
    ws.cell(row=r, column=target_col).font = Font(bold=True)
    ws.cell(row=r, column=target_col).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=r, column=target_col).number_format = '0'
    if fill:
        ws.cell(row=r, column=target_col).fill = fill

ws.cell(row=2, column=target_col).font = Font(bold=True)
ws.cell(row=2, column=target_col).alignment = Alignment(horizontal='center', vertical='top')
if fill:
    ws.cell(row=2, column=target_col).fill = fill

# Expected Result formulas (E3:E15)
for r in range(3, 16):
    ws.cell(row=r, column=5).value = '=INDEX($I$3:$I$15, MATCH($O{0}, $H$3:$H$15, 0))'.format(r)

wb.save(out_path)
