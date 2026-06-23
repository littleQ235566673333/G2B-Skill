import openpyxl
import operator

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_22-47_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_22-47_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

def find_data_start(ws, col):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col), 1):
        cell = row[0].value
        if cell and isinstance(cell, str):
            return i + 1
    return 2 # fallback

data_start = min(find_data_start(ws, 2), find_data_start(ws, 3))

rows = []
for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, min_col=2, max_col=10):
    b, c, d, e, f, g, h, i, j = [cell.value for cell in row]
    if (b is not None or c is not None):
        rows.append({'B': b, 'C': c, 'J': j, 'Row': row, 'orig_row_idx': row[0].row})

filtered = []
seen = set()
for r in rows:
    b, c = r['B'], r['C']
    # skip empty, repeated header, or duplicate B+C
    if not b or not c or (b, c) in seen or (isinstance(b, str) and b.lower() == 'name'):
        continue
    seen.add((b, c))
    filtered.append(r)

helper_j = []
for i in range(data_start, ws.max_row+1):
    val = ws.cell(row=i, column=10).value
    if val and str(val).strip():
        if val not in helper_j: helper_j.append(val)

j_ordered = []
not_in_j = []
if helper_j:
    for name in helper_j:
        for r in filtered:
            if r['B'] == name:
                j_ordered.append(r)
    for r in filtered:
        if r['B'] not in helper_j:
            not_in_j.append(r)
else:
    not_in_j = sorted(filtered, key=lambda x: (str(x['B']).lower() if x['B'] is not None else ''))
combined = j_ordered + not_in_j

# Write output to F2:H10, up to 9 rows
for idx in range(9):
    ws['F' + str(2 + idx)].value = None
    ws['G' + str(2 + idx)].value = None
    ws['H' + str(2 + idx)].value = None
for idx, r in enumerate(combined[:9]):
    ws['F' + str(2 + idx)].value = r['B']
    ws['G' + str(2 + idx)].value = r['C']
    ws['H' + str(2 + idx)].value = r['J']

# Sort only G and H (columns G&H, output in G2:H10) by H (lowest first)
values_gh = []
for i in range(2, 11):
    g = ws['G'+str(i)].value
    h = ws['H'+str(i)].value
    values_gh.append((g,h))
non_empty_gh = [gh for gh in values_gh if not (gh[0] is None and gh[1] is None)]
# replace None with '' for sorting
sort_safe_gh = [(g if g is not None else '', h if h is not None else '') for g, h in non_empty_gh]
sorted_gh = sorted(sort_safe_gh, key=operator.itemgetter(1, 0))
for i, (g, h) in enumerate(sorted_gh):
    ws['G'+str(2+i)].value = g if g != '' else None
    ws['H'+str(2+i)].value = h if h != '' else None
for j in range(len(sorted_gh), 9):
    ws['G'+str(2+j)].value = None
    ws['H'+str(2+j)].value = None

wb.save(output_path)
