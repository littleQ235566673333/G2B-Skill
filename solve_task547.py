from openpyxl import load_workbook

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_547-43_s1/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_547-43_s1/output.xlsx'

wb = load_workbook(INPUT)
emp_ws = wb['Emp']
lookup_ws = wb['Lookup']
expected_ws = wb['Expected']

# Read headers from Emp table
emp_headers = [emp_ws.cell(1, c).value for c in range(1, emp_ws.max_column + 1)]
fixed_headers = emp_headers[:4]  # seqno, Empno, ename, Jobid
variable_headers = emp_headers[4:]  # columns to unpivot/transpose to rows

# Build lookup mapping
lookup = {}
for r in range(2, lookup_ws.max_row + 1):
    var = lookup_ws.cell(r, 1).value
    cat = lookup_ws.cell(r, 2).value
    subcat = lookup_ws.cell(r, 3).value
    if var is not None:
        lookup[var] = (cat, subcat)

# Clear destination region A1:H19 as requested area
for r in range(1, 20):
    for c in range(1, 9):
        expected_ws.cell(r, c).value = None

# Write header
out_headers = fixed_headers + ['Variable', 'Category', 'Subcategory', 'Value']
for c, header in enumerate(out_headers, start=1):
    expected_ws.cell(1, c).value = header

# Unpivot Emp rows into Expected
out_row = 2
for r in range(2, emp_ws.max_row + 1):
    fixed_vals = [emp_ws.cell(r, c).value for c in range(1, 5)]
    if all(v is None for v in fixed_vals):
        continue
    for idx, var_name in enumerate(variable_headers, start=5):
        value = emp_ws.cell(r, idx).value
        category, subcategory = lookup.get(var_name, (None, None))
        row_vals = fixed_vals + [var_name, category, subcategory, value]
        for c, v in enumerate(row_vals, start=1):
            expected_ws.cell(out_row, c).value = v
        out_row += 1

wb.save(OUTPUT)

# verification print
wb2 = load_workbook(OUTPUT)
ws2 = wb2['Expected']
for r in range(1, 20):
    print([ws2.cell(r, c).value for c in range(1, 9)])
