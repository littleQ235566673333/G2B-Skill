import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_250-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_250-20_tc1/output.xlsx'

# Load workbook and 'RNM' sheet.
wb = openpyxl.load_workbook(input_path)
sheet = wb['RNM']

# Extract the first 20 rows and 10 columns (A1:J20)
rows = list(sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10, values_only=True))
header = rows[0]
data = rows[1:]

# Convert to DataFrame
columns = list(header)
df = pd.DataFrame(data, columns=columns)

# Rename for clarity
group_cols = ['NUMBER', 'LINE NO']
sum_col = 'MATCHED_QTY'

if not df.empty:
    grp = df.groupby(group_cols, dropna=False, as_index=False)
    df_sum = grp.agg({sum_col: 'sum'})
    df_first = grp.nth(0).reset_index(drop=True)
    df_first[sum_col] = df_sum[sum_col]
    df_out = df_first[columns]
else:
    df_out = df.copy()

# Pad to 19 rows (excluding header)
while len(df_out) < 19:
    df_out.loc[len(df_out)] = [None]*len(columns)

output_rows = [header] + df_out.values.tolist()

wb.remove(sheet)
ws = wb.create_sheet('RNM')
for row in output_rows:
    ws.append(row)

wb.save(output_path)
