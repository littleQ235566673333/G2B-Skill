from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_8/regression_gate/before_pass/core_493-18/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_8/regression_gate/before_pass/core_493-18/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Collect reference values from column F (excluding header, rows 2-7)
ref_values = set()
for row in ws.iter_rows(min_row=2, max_row=7, min_col=6, max_col=6):
    v = row[0].value
    if v is not None and v != "":
        ref_values.add(v)

# Store matching data from columns A-C for shifting up
retained_rows = []
for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=3):
    a = row[0].value
    if a in ref_values:
        retained_rows.append([cell.value for cell in row])
    # else will clear later

# Clear columns A-C for all rows 2-7
for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=3):
    for cell in row:
        cell.value = None

# Write back retained rows, shifted up
for i, values in enumerate(retained_rows):
    for j, value in enumerate(values):
        ws.cell(row=2+i, column=1+j, value=value)

# Reapply autofilter to columns A1:C7
ws.auto_filter.ref = 'A1:C7'

wb.save(output_path)
