from openpyxl import load_workbook
from datetime import datetime

# Input and output paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/group_45707/r2/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/group_45707/r2/evolve_45707/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Cache all dates and occurrences for quick lookup
dates = []
occurrences = []
for r in range(2, 70):  # D2:D69
    v_date = ws.cell(row=r, column=1).value
    if isinstance(v_date, str) and v_date.startswith('='):
        # Evaluate date formula if possible (sheet seems to use consecutive formulas)
        try:
            prev = dates[-1]
            # Add one day
            v_date = prev + timedelta(days=1)
        except:
            pass
    dates.append(v_date)
    occ = ws.cell(row=r, column=3).value
    occurrences.append(occ)

# For rows 2 to 69 (so indices 0..67), set D only if A of next row is first of month
for idx in range(68):
    next_date = dates[idx+1] if idx+1 < len(dates) else None
    d_cell = ws.cell(row=idx+2, column=4)  # D column
    if isinstance(next_date, datetime) and next_date.day == 1:
        # Count how many occurence == 1 for that month and year
        count = sum(
            (dates[i].year == next_date.year) and (dates[i].month == next_date.month) and (occurrences[i] == 1)
            for i in range(len(dates)) if isinstance(dates[i], datetime)
        )
        d_cell.value = count
    else:
        d_cell.value = None  # Completely empty

# Save to output
wb.save(output_path)
