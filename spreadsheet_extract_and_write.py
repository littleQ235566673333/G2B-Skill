import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_178-22_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_178-22_tc1/output.xlsx'

# Load workbook and source sheet
wb = openpyxl.load_workbook(input_path)
sheet1 = wb['Sheet1']

# Extract headers
headers = [cell.value for cell in sheet1[1][:3]]

# Prepare the filtered data
filtered_rows = []
for row in sheet1.iter_rows(min_row=2, values_only=True):
    col_b = row[1]
    col_c = row[2]
    if (
        col_b == 'TELIVISION' or
        col_c == 'CLASS III' or
        col_c == 'CLASS IV'
    ):
        filtered_rows.append(row[:3])

# Create Sheet2
if 'Sheet2' not in wb:
    sheet2 = wb.create_sheet('Sheet2')
else:
    sheet2 = wb['Sheet2']

# Write headers at row 2
for j, hdr in enumerate(headers, start=1):
    sheet2.cell(row=2, column=j, value=hdr)

# Write filtered rows starting at row 3
for i, row_data in enumerate(filtered_rows, start=3):
    for j, val in enumerate(row_data, start=1):
        sheet2.cell(row=i, column=j, value=val)

# Save output
wb.save(output_path)
