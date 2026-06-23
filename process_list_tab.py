from openpyxl import load_workbook

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_2/regression_gate/after_pass/core_84-40/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_2/regression_gate/after_pass/core_84-40/output.xlsx'

wb = load_workbook(input_file)

sheet_summaries = []
list_sheet_name = None
for name in wb.sheetnames:
    if name.lower() == 'list':
        list_sheet_name = name
        break

target_sheets = [name for name in wb.sheetnames if name.lower() != 'list']
for sheet_name in target_sheets:
    ws = wb[sheet_name]
    # Helper to sum numbers in column (skip header)
    def col_sum(column_letter):
        values = [cell.value for cell in ws[column_letter][1:] if isinstance(cell.value, (int, float))]
        return sum(values)
    sum_c = col_sum('C')
    sum_d = col_sum('D')
    sheet_summaries.append((sheet_name, sum_c, sum_d))

if list_sheet_name is None:
    ws_list = wb.create_sheet('LIST')
else:
    ws_list = wb[list_sheet_name]

# Clear previous data before new fill
for row in ws_list.iter_rows(min_row=2, max_row=ws_list.max_row, min_col=2, max_col=5):
    for cell in row:
        cell.value = None

start_row = 2
for i, (name, total_c, total_d) in enumerate(sheet_summaries):
    ws_list.cell(row=start_row + i, column=2, value=name)  # B
    ws_list.cell(row=start_row + i, column=3, value=total_c)  # C
    ws_list.cell(row=start_row + i, column=4, value=total_d)  # D
    ws_list.cell(row=start_row + i, column=5, value=total_c - total_d)  # E

total_row = start_row + len(sheet_summaries)
ws_list.cell(row=total_row, column=2, value='TOTAL')
# Use standard .value assignment, not f-string expressions in-line
ws_list.cell(row=total_row, column=3).value = "=SUM(C{}:C{})".format(start_row, total_row-1)
ws_list.cell(row=total_row, column=4).value = "=SUM(D{}:D{})".format(start_row, total_row-1)
ws_list.cell(row=total_row, column=5).value = "=SUM(E{}:E{})".format(start_row, total_row-1)

wb.save(output_file)
