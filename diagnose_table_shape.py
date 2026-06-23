import openpyxl

# Input and output file paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_48983_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_48983_tc1/output.xlsx"

tb = openpyxl.load_workbook(input_path)
ws = tb.active

target_start_row = 6
target_categories_rows = range(target_start_row, target_start_row+6)  # 6-11
# L = 12, M = 13, ... S = 19 (13..19 included)
target_brand_cols = range(13, 20)

target_brands = [ws.cell(row=target_start_row, column=col).value for col in target_brand_cols]
categories = [ws.cell(row=row, column=12).value for row in target_categories_rows]

# Debug print categories and brands to find mismatches
print('Target categories:', categories)
print('Target brands:', target_brands)

# Let's scan for possible source tables in A:K and rows 1:20, print the found text for user diagnosis
for r in range(1, 21):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 12)]
    print(f'Row {r}:', row_vals)
