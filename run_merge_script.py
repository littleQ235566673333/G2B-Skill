import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_267-21_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_267-21_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sh_merging = wb['merging']
sh_sh1 = wb['SH1']
sh_rp1 = wb['RP1']

def make_id_qty_dict(sheet):
    return {str(sheet.cell(row=i, column=1).value): sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)}

sh1_dict = make_id_qty_dict(sh_sh1)
rp1_dict = make_id_qty_dict(sh_rp1)

for row in range(2, 12):
    id_val = sh_merging.cell(row=row, column=2).value
    id_val_str = str(id_val) if id_val is not None else ""
    sh1_qty = sh1_dict.get(id_val_str, "-")
    sh_merging.cell(row=row, column=3, value=sh1_qty if sh1_qty is not None else "-")
    rp1_qty = rp1_dict.get(id_val_str, "-")
    sh_merging.cell(row=row, column=4, value=rp1_qty if rp1_qty is not None else "-")

wb.save(output_path)
