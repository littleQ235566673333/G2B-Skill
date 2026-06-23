from openpyxl import load_workbook
from datetime import time, datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_8/regression_gate/before_fix/core_49667/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_8/regression_gate/before_fix/core_49667/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active
# Columns F (6) to AT (46) are meeting blocks
time_cols = list(range(6, 47))
output_cols = [2, 3, 4, 5]  # B, C, D, E

# Helper: Convert time column header value to float hour
# Could be float, or datetime.time, or string

def time_to_float(t):
    if isinstance(t, (float, int)):
        return float(t)
    elif isinstance(t, time):
        return t.hour + t.minute/60.0 + t.second/3600.0
    elif isinstance(t, datetime):
        return t.hour + t.minute/60.0 + t.second/3600.0
    elif isinstance(t, str):
        try:
            dt = datetime.strptime(t, '%H:%M')
            return dt.hour + dt.minute/60.0
        except:
            return None
    return None

# Read the header times from row 1
header_times = [ws.cell(row=1, column=col).value for col in time_cols]
float_times = [time_to_float(t) for t in header_times]

for i in range(2, 17):  # Rows 2 to 16
    row_blocks = [ws.cell(row=i, column=col).value for col in time_cols]
    blocks = []
    block_start = None
    for idx, val in enumerate(row_blocks):
        if val == 'm':
            if block_start is None:
                block_start = idx
        else:
            if block_start is not None:
                blocks.append((block_start, idx - 1))
                block_start = None
    if block_start is not None:
        # Block goes to end
        blocks.append((block_start, len(row_blocks)-1))
    # Write results for up to 2 meeting blocks
    for n in range(2):
        c_start = output_cols[2*n]
        c_end = output_cols[2*n+1]
        if n < len(blocks):
            sidx, eidx = blocks[n]
            start_time = float_times[sidx]
            finish_time = float_times[eidx+1] if (eidx+1)<len(float_times) else float_times[-1]
            ws.cell(row=i, column=c_start).value = start_time
            ws.cell(row=i, column=c_end).value = finish_time
        else:
            ws.cell(row=i, column=c_start).value = None
            ws.cell(row=i, column=c_end).value = None

wb.save(output_path)
