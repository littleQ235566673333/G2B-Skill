import pandas as pd
from openpyxl import load_workbook

# Input/output file paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_6/group_42930/r1/evolve_42930/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_6/group_42930/r1/evolve_42930/output.xlsx'

# Read the data with pandas (to manipulate, assuming default sheet)
df = pd.read_excel(input_path)

# Guess columns (since user did not specify; try to auto-detect)
colnames = df.columns.str.lower()
subdiv_col = [c for c in df.columns if 'subdiv' in c.lower()][0]
year_col = [c for c in df.columns if 'year' in c.lower() or 'built' in c.lower()][0]

# 1. Find earliest house built in each subdivision
earliest = df.groupby(subdiv_col)[year_col].min().reset_index()
early_sorted = earliest.sort_values(year_col).reset_index(drop=True)
early_sorted['SubdivisionNumber'] = range(1, len(early_sorted)+1)

# 2. Map these numbers back to main df, add a column for sorting
mapping = dict(zip(early_sorted[subdiv_col], early_sorted['SubdivisionNumber']))
df['SubdivisionNumber'] = df[subdiv_col].map(mapping)

# 3. Sort df as required (subdivision number, year built, subdivision name for stability)
df_sorted = df.sort_values(['SubdivisionNumber', year_col, subdiv_col]).reset_index(drop=True)

# 4. Write the subdivision number for C2:C22 using openpyxl
wb = load_workbook(input_path)
ws = wb.active
for i in range(2, 23):
    house_idx = i-2  # 0-based indexing
    if house_idx < len(df_sorted):
        ws.cell(row=i, column=3).value = int(df_sorted.iloc[house_idx]['SubdivisionNumber'])

wb.save(output_path)
