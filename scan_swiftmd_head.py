import pandas as pd
from openpyxl import load_workbook
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_1/group_91-34/r2/evolve_91-34/input.xlsx'
sheet_name = 'SwiftMD'
wb = load_workbook(input_path, data_only=True)
ws = wb[sheet_name]
# Print cells B1:O20 (to match user cell refs)
for row in ws.iter_rows(min_row=1, max_row=20, min_col=2, max_col=15):
    print([cell.value for cell in row])
