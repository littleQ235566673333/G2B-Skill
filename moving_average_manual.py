import pandas as pd
from openpyxl import load_workbook

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_3/regression_gate/before_fix/core_4714/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_3/regression_gate/before_fix/core_4714/output.xlsx'

# Read all columns as default (likely: A-Employee, B, C-Month, D-Hours)
df = pd.read_excel(input_file)

results = []
for idx, row in df.iterrows():
    employee = row[0]
    curr_month = row[2]
    employee_rows = df[df.iloc[:,0] == employee]
    eligible_months = employee_rows[employee_rows.iloc[:,2] <= curr_month]
    # Sort months descending, pick top 4
    eligible_months = eligible_months.sort_values(by=eligible_months.columns[2], ascending=False)
    if eligible_months.shape[0] < 4:
        results.append("n/a")
    else:
        window = eligible_months.head(4)
        if window.iloc[:,3].isnull().any():
            results.append("n/a")
        else:
            avg = window.iloc[:,3].mean()
            if avg % 1 == 0:
                results.append(str(int(avg)))
            else:
                results.append(f'{avg:.1f}')

# Write results to E2:E25 using openpyxl
wb = load_workbook(input_file)
ws = wb.active
for i in range(24):  # E2:E25 = rows 2..25
    if i < len(results):
        ws[f'E{i+2}'] = results[i]
wb.save(output_file)
