import openpyxl
from openpyxl.utils import get_column_letter

# Load workbook and sheet
input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_59224_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_59224_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get Project Start Date
project_start = ws['B2'].value

# Range for dates and output
start_row = 4
end_row = 14
col_c = 'C'
col_d = 'D'
col_e = 'E'

select_period_row = None

# First, evaluate where "Select Period" should appear
for row in range(start_row, end_row+1):
    date_c = ws[f'{col_c}{row}'].value
    date_d = ws[f'{col_d}{row}'].value
    if date_c is not None and date_d is not None:
        # Check if project_start is between date_c and date_d
        if project_start > date_c and project_start < date_d:
            select_period_row = row
            break

# Fill E column according to user's instructions
for row in range(start_row, end_row+1):
    if select_period_row is not None and row <= select_period_row:
        ws[f'{col_e}{row}'] = 'Select Period'
    else:
        ws[f'{col_e}{row}'] = 'Do Not Select'

# Save to output
wb.save(output_path)
