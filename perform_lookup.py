import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_7/regression_gate/after_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_7/regression_gate/after_pass/core_18935/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active
# Data table starts at row 4, headers in row 4 (A4:E4)
data_table_col_headers = [ws.cell(row=4, column=col).value for col in range(1,6)] # Work, Material, Category 1, Category 2, Category 3
data = []
row = 5
while ws.cell(row=row, column=1).value is not None:
    work = ws.cell(row=row, column=1).value
    material = ws.cell(row=row, column=2).value
    category_1 = ws.cell(row=row, column=3).value
    category_2 = ws.cell(row=row, column=4).value
    category_3 = ws.cell(row=row, column=5).value
    data.append({'Work': work, 'Material': material, 'Category 1': category_1, 'Category 2': category_2, 'Category 3': category_3})
    row += 1
# For each report row: rows 17-22: ('Work', 'Category Type', 'Material', 'Category Value')
for out_row in range(17, 23):
    work = ws.cell(row=out_row, column=1).value
    category_type = ws.cell(row=out_row, column=2).value
    material = ws.cell(row=out_row, column=3).value
    match = next((row for row in data if row['Work'] == work and row['Material'] == material), None)
    value = match[category_type] if match and category_type in match else None
    ws.cell(row=out_row, column=4, value=value)
wb.save(output_path)
