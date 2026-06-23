from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42/eval_46240_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42/eval_46240_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

for row in range(2, 5):
    h_val = ws[f'H{row}'].value
    # Rule: 'Yes' if date or 'Yes', 'N/A' if N/A, 'No' if blank
    if h_val is None or (isinstance(h_val, str) and h_val.strip() == ''):
        result = 'No'
    elif isinstance(h_val, str) and h_val.strip().upper() == 'N/A':
        result = 'N/A'
    elif (isinstance(h_val, str) and h_val.strip().upper() == 'YES') or hasattr(h_val, 'year'):
        result = 'Yes'
    else:
        result = 'No'
    ws[f'J{row}'] = result

wb.save(output_path)
