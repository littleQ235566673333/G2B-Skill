import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_61-4_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_61-4_tc1/output.xlsx'

# Load workbook and input sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['input']
data = list(ws.values)
headers = data[0][:8]
rows = [r[:8] for r in data[1:] if r[0] is not None]
df = pd.DataFrame(rows, columns=headers)

output_rows = []

for stock in df['Stock Name'].unique():
    stock_df = df[df['Stock Name'] == stock].reset_index(drop=True)
    neg_mask = stock_df['Change'] < 0
    # Find consecutive runs
    run_id = (neg_mask != neg_mask.shift()).cumsum()
    grouped = stock_df[neg_mask].groupby(run_id[neg_mask])
    for grp_idx, group in grouped:
        # We only process runs of length > 0 (consecutive negatives)
        if len(group) > 0:
            first_row = group.iloc[0]
            last_row = group.iloc[-1]
            out = [
                first_row['DATE '],
                stock,
                first_row['OPENP* '],
                group['HIGH '].max(),
                group['LOW '].min(),
                last_row['CLOSEP* '],
                group['VOLUME'].sum()
            ]
            output_rows.append(out)

# Prepare output sheet
def write_output(wb, data, sheetname, startcell):
    if sheetname in wb.sheetnames:
        ws_out = wb[sheetname]
    else:
        ws_out = wb.create_sheet(sheetname)
    
    # Write headers
    ws_out[startcell].value = 'DATE'
    ws_out[startcell[0]+str(int(startcell[1])+0)].offset(column=1).value = 'Stock Name'
    ws_out[startcell[0]+str(int(startcell[1])+0)].offset(column=2).value = 'OPENP'
    ws_out[startcell[0]+str(int(startcell[1])+0)].offset(column=3).value = 'HIGH'
    ws_out[startcell[0]+str(int(startcell[1])+0)].offset(column=4).value = 'LOW'
    ws_out[startcell[0]+str(int(startcell[1])+0)].offset(column=5).value = 'CLOSEP'
    ws_out[startcell[0]+str(int(startcell[1])+0)].offset(column=6).value = 'VOLUME'

    # Write data
    for idx, row in enumerate(data):
        for jdx, val in enumerate(row):
            ws_out.cell(row=2+idx, column=1+jdx).value = val

write_output(wb, output_rows, 'output', 'A2')
wb.save(output_path)
