from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun2/eval_183-8_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun2/eval_183-8_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

def fill_weighted_formulas(start_row, end_row, gen_col, val_cols, out_cols):
    # weighted average: SUMPRODUCT(value, weight) / SUM(weight)
    for i, ocol in enumerate(out_cols):
        formula = f"=SUMPRODUCT({val_cols[i]}{start_row}:{val_cols[i]}{end_row},{gen_col}{start_row}:{gen_col}{end_row})/SUM({gen_col}{start_row}:{gen_col}{end_row})"
        ws[f'{ocol}{end_row + 1}'] = formula

# Cols: C=Gen (3), D=%PLF (4), E=M/C (5), F=Grid (6)
# Output: J, K, L  (10, 11, 12)
# Groups start: 3, 9, 15, 21, 27, 33, 39 (a-g based on CSV positions)
groups = [3, 9, 15, 21, 27, 33, 39]
group_size = 4
for start in groups:
    end = start + group_size - 1
    fill_weighted_formulas(start, end, 'C', ['D', 'E', 'F'], ['J', 'K', 'L'])

wb.save(output_path)
