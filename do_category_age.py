from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_2/regression_gate/before_pass/core_32337/input.xlsx')
ws = wb.active

# Use header row 2
def find_col(header):
    for cell in ws[2]:
        if cell.value and str(cell.value).strip().lower() == header.lower():
            return cell.column
    raise Exception(f'Header not found: {header}')

col_e = find_col('Expected Result')
age_col = find_col('age (year)')
cat_col = find_col('CATEGORY')
dob_col = find_col('DATE OF BIRTH')

# Prepare column letters
age_letter = get_column_letter(age_col)
cat_letter = get_column_letter(cat_col)
dob_letter = get_column_letter(dob_col)

# For the formula, report date is in $A$1
# 1. Set formula in E3:E15 to lookup CATEGORY by age
for row in range(3, 16):
    formula = f'=INDEX(${cat_letter}$3:${cat_letter}$15,MATCH({age_letter}{row},${age_letter}$3:${age_letter}$15,0))'
    ws.cell(row=row, column=col_e).value = formula

# 2. Insert new column after 'age (year)' for Actual Age
insert_col = age_col + 1
ws.insert_cols(insert_col)
ws.cell(row=2, column=insert_col, value='Actual Age')

# Copy fill from prior column
fill_ref = ws.cell(row=3, column=age_col).fill
for row in range(3, ws.max_row+1):
    # Use $A$1 for report date
    formula = f'=INT(DATEDIF({dob_letter}{row},$A$1,"Y"))'
    c = ws.cell(row=row, column=insert_col)
    c.value = formula
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')
    if fill_ref.patternType:
        c.fill = PatternFill(patternType=fill_ref.patternType, fgColor=fill_ref.fgColor, bgColor=fill_ref.bgColor)
# Header formatting: bold, center, top align, fill
h = ws.cell(row=2, column=insert_col)
h.font = Font(bold=True)
h.alignment = Alignment(horizontal='center', vertical='top')
if fill_ref.patternType:
    h.fill = PatternFill(patternType=fill_ref.patternType, fgColor=fill_ref.fgColor, bgColor=fill_ref.bgColor)

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_2/regression_gate/before_pass/core_32337/output.xlsx')
