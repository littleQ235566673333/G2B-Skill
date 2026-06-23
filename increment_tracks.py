import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_16511_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_16511_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

albums = [ws[f'B{i}'].value for i in range(2, 11)]
track_numbers = []
album_count = {}

for album in albums:
    if album not in album_count:
        album_count[album] = 1
    else:
        album_count[album] += 1
    track_numbers.append(album_count[album])

for idx, num in enumerate(track_numbers, start=2):
    ws[f'F{idx}'] = num

wb.save(output_path)
