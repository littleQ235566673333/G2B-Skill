import openpyxl
import pandas as pd
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_2/regression_gate/after_fix/core_91-34/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_2/regression_gate/after_fix/core_91-34/output.xlsx'
sheet_name = 'SwiftMD'
wb = openpyxl.load_workbook(input_path)
ws = wb[sheet_name]
header_row = 2
first_col_data = 3
max_col = ws.max_column
headers = [h for h in ws.iter_rows(min_row=header_row, max_row=header_row, min_col=first_col_data, max_col=max_col, values_only=True)][0]
def header_clean(h):
    return str(h).replace("'", "").replace('"', "").strip() if h else None
headers = [header_clean(h) for h in headers]
data = [tuple(ws.iter_rows(min_row=r, max_row=r, min_col=first_col_data, max_col=max_col, values_only=True))[0] for r in range(header_row+1, ws.max_row+1) if any(ws.iter_rows(min_row=r, max_row=r, min_col=first_col_data, max_col=max_col, values_only=True))]
df = pd.DataFrame(data, columns=headers)
def col_by_keyword(keyword):
    for h in headers:
        if h and keyword in h.lower():
            return h
    raise ValueError(f'header for {keyword} not found: {headers}')
lname_col = col_by_keyword('last name')
fname_col = col_by_keyword('first name')
dob_col   = col_by_keyword('date of birth')
dup_col   = col_by_keyword('duplicate')
rel_col   = col_by_keyword('relationship')
df_yes = df[df[dup_col].astype(str).replace({"'": "", '"': ""}, regex=True).str.strip().str.lower() == 'yes']
groups = df_yes.groupby([lname_col, fname_col, dob_col])
rows_to_remove = set()
for key, group in groups:
    group_df = group
    rels = group_df[rel_col].astype(str).replace({"'": "", '"': ""}, regex=True).str.strip().str.lower().tolist()
    orig_idx = group_df.index.tolist()
    employees = [rel for rel in rels if rel == 'employee']
    if not employees and len(group_df) > 1:
        rows_to_remove.update(orig_idx[1:])
filtered_df = df.drop(rows_to_remove)
wb_out = openpyxl.load_workbook(input_path)
ws_out = wb_out[sheet_name]
start_row, start_col = 2, 2
for col_offset, value in enumerate(headers, start_col):
    ws_out.cell(row=start_row, column=col_offset, value=value)
for r_index, row in enumerate(filtered_df.itertuples(index=False), start_row+1):
    for c_index, cell_value in enumerate(row, start_col):
        ws_out.cell(row=r_index, column=c_index, value=cell_value)
wb_out.save(output_path)
print('DONE')
