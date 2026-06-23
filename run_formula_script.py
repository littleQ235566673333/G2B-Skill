import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/group_45707/r2/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/group_45707/r2/evolve_45707/output.xlsx'

# Read data with pandas
df = pd.read_excel(input_path)

# Prepare to update with openpyxl
wb = load_workbook(input_path)
ws = wb.active

def is_first_of_month(date):
    return date.day == 1

for i in range(1, 69):  # for D2:D69, i is 1-based for D2
    row_idx = i + 1
    if i >= len(df):
        ws[f'D{row_idx}'] = None
        continue
    try:
        this_date = pd.to_datetime(df.iloc[i, 0])
    except Exception:
        ws[f'D{row_idx}'] = None
        continue
    # Check if "the next date" (i+1th row) is the 1st
    if is_first_of_month(this_date):
        month, year = this_date.month, this_date.year
        # Count values in col C with value 1, in same month/year
        mask = (pd.to_datetime(df.iloc[:,0]).dt.year == year) & (pd.to_datetime(df.iloc[:,0]).dt.month == month)
        count_1s = (df.loc[mask, df.columns[2]] == 1).sum()
        ws[f'D{row_idx}'] = count_1s
    else:
        ws[f'D{row_idx}'] = None

wb.save(output_path)
