from openpyxl import load_workbook, styles
from copy import copy

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_1/regression_gate/after_pass/core_32337/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_1/regression_gate/after_pass/core_32337/output.xlsx"

wb = load_workbook(input_path)
ws = wb["Sheet1"]

header_row = 2
first_data_row, last_data_row = 3, 15
col_report_date = 2 # $B$1
col_name, col_gender, col_dob = 1, 2, 3
col_expected_result = 5
col_category = 9 # I
col_year_age = 15 # O in 1-based openpyxl (due to initial extra columns)

# Fill Expected Result (E3-E15) with formula referencing CATEGORY based on age (from col O)
for row in range(first_data_row, last_data_row + 1):
    # Formula: simply reference the CATEGORY column of this row, which points to I (col_category)
    ws.cell(row=row, column=col_expected_result).value = f"=I{row}"

# Insert Actual Age after age (year) (new P column after O)
col_age_actual = col_year_age + 1
ws.insert_cols(col_age_actual)
actual_age_header_cell = ws.cell(row=header_row, column=col_age_actual)
actual_age_header_cell.value = 'Actual Age'

# Copy fill from the left
header_fill = ws.cell(row=header_row, column=col_age_actual-1).fill
actual_age_header_cell.fill = copy(header_fill)

# Fill and format Actual Age values for each data row
for row in range(first_data_row, last_data_row + 1):
    # ROUNDDOWN(($B$1 - Cx)/365, 0), where $B$1 is report date, Cx is DOB
    dob_cell = ws.cell(row=row, column=col_dob).coordinate
    formula = f"=ROUNDDOWN((($B$1-{dob_cell})/365),0)"
    cell = ws.cell(row=row, column=col_age_actual)
    cell.value = formula
    cell.font = styles.Font(bold=True)
    cell.alignment = styles.Alignment(horizontal='center', vertical='center')
    cell.number_format = '0'

# Header formatting
actual_age_header_cell.alignment = styles.Alignment(horizontal='center', vertical='top', wrap_text=True)
actual_age_header_cell.font = styles.Font(bold=True)

wb.save(output_path)
print('Done')
