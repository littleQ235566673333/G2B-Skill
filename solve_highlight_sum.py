import openpyxl
from openpyxl.styles import PatternFill
import itertools

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_254-34_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_254-34_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)

# Remove Sheet3 if exists
if 'Sheet3' in wb.sheetnames:
    wb.remove(wb['Sheet3'])

ws = wb['Before']

# Fill definition for #00FF00
fill_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type="solid")

# Extract all numerical cell coordinates and their values
cell_values = []
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell_values.append((cell.coordinate, cell.value))

# Find a combination of cells whose sum is between 994108 and 994112
min_sum = 994108
max_sum = 994112
solution = None

# Try combinations (efficient for small sets)
for r in range(2, len(cell_values)+1):
    for combo in itertools.combinations(cell_values, r):
        value_sum = sum([v for _, v in combo])
        if min_sum <= value_sum <= max_sum:
            solution = combo
            break
    if solution:
        break

if solution:
    # Highlight the cells in green
    for coord, _ in solution:
        ws[coord].fill = fill_green
    # Place the sum found in cell C2
    ws['C2'].value = sum(v for _, v in solution)
else:
    ws['C2'].value = "No valid combination"

wb.save(output_path)
