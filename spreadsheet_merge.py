import openpyxl
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/after_fix/core_60-7/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/after_fix/core_60-7/output.xlsx'

wb = openpyxl.load_workbook(input_path)

def read_sheet(sheet_name):
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    return data

# Read sheets
existing_data = read_sheet('Existing Task')
additions_data = read_sheet('Additions')
retired_data = read_sheet('Retired')

# Assume headers exist in all, and are the same/compatible
def rows_to_dicts(rows):
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:] if any(x is not None for x in row)]

existing_dicts = rows_to_dicts(existing_data)
additions_dicts = rows_to_dicts(additions_data)
retired_dicts = rows_to_dicts(retired_data)

def serialize_row(row, headers):
    return tuple(row.get(h, None) for h in headers)

# Use all column names in order from Existing Task for output
output_headers = existing_data[0][:5]

# 1. Combine Existing + Additions
all_rows = existing_dicts + additions_dicts
# 2. Remove rows present in Retired
retired_set = set(serialize_row(r, output_headers) for r in retired_dicts)

net_rows = []
seen = set()
for row in all_rows:
    srow = serialize_row(row, output_headers)
    if srow not in retired_set and srow not in seen:
        net_rows.append(row)
        seen.add(srow)

# Output to 'Consolidated Tracker' from A3
ws = wb['Consolidated Tracker']
start_row = 3
start_col = 1
for c, header in enumerate(output_headers, start_col):
    ws.cell(row=start_row, column=c, value=header)
for r, row in enumerate(net_rows, start=1):
    for c, header in enumerate(output_headers, start_col):
        ws.cell(row=start_row + r, column=c, value=row.get(header, None))

wb.save(output_path)
