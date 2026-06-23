import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_3/regression_gate/before_pass/core_408-5/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_3/regression_gate/before_pass/core_408-5/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
data = rows[1:]

# Step 1: Sum Column D where Column E == 0
sum_d_col_e_zero = sum((row[3] for row in data if row[4] == 0 and row[3] is not None), 0)

# Step 2: Find row with 'BR1 Sales' in Column C and its worksheet index
br1_index = None
for i, row in enumerate(data):
    if row[2] and 'BR1 Sales' in str(row[2]):
        br1_index = i + 2  # +2 for 1-based + header
        break
# Step 3: Insert the sum
if br1_index is not None:
    ws.cell(row=br1_index, column=4).value = sum_d_col_e_zero

# Step 4: Build data to keep (removing specified rows and blank lines)
cleaned_rows = [headers]
for row in data:
    col_c = row[2] if len(row) > 2 else None
    col_e = row[4] if len(row) > 4 else None
    is_br1 = col_c and 'BR1 Sales' in str(col_c)
    is_col_e_zero = (col_e == 0)
    is_blank = all((v is None or (isinstance(v, str) and v.strip() == '')) for v in row)
    if not is_blank and (not is_col_e_zero or is_br1):
        cleaned_rows.append(row)

# Step 5: Overwrite the sheet
ws.delete_rows(2, ws.max_row)
for i, row in enumerate(cleaned_rows[1:], start=2):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

wb.save(output_path)
