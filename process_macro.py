import openpyxl
from datetime import datetime
from collections import defaultdict

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_567-21_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_567-21_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ds = wb['Sheet1']

header = [cell.value for cell in ds[1]]
rows = list(ds.iter_rows(min_row=2, values_only=True))

rows_by_key = defaultdict(list)
def parse_date(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == '':
            return None
        try:
            return datetime.strptime(val, '%Y/%m')
        except Exception:
            return None
    return val
for row in rows:
    acct = row[0]
    assoc = row[1]
    dateval = row[4]
    dt = parse_date(dateval)
    if dt:
        rows_by_key[(acct, assoc)].append((dt, row))
filtered_rows = []
for key, record_list in rows_by_key.items():
    if not record_list:
        continue
    # get max date
    max_dt = max(dt for dt, _ in record_list)
    # collect all rows with max date
    for dt, r in record_list:
        if dt == max_dt:
            filtered_rows.append(r)
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Sheet1'
for idx, val in enumerate(header, 1):
    ws_out.cell(row=1, column=idx).value = val
for i, row in enumerate(filtered_rows, start=2):
    for j, val in enumerate(row, start=1):
        ws_out.cell(row=i, column=j, value=val)
wb_out.save(output_path)
