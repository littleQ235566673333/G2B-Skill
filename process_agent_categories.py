import openpyxl

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_53161_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_53161_tc1/output.xlsx"
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Agents and categories (columns E:M, rows 4-9)
agents = []
categories = ['Category 1','Category 2','Category 3','Category 4','Category 5','Category 6','Category 7','Category 8','Category 9']
cat_data = {}
for row in range(4, 10):
    agent = ws.cell(row=row, column=5).value
    if not agent:
        continue
    agents.append(agent)
    cat_data[agent] = {}
    for ci, col in enumerate(range(6, 15)):
        val = ws.cell(row=row, column=col).value
        cat = categories[ci]
        cat_data[agent][cat] = 1 if val else 0

# Schedule info (columns F:AE (=6:31), rows 13-18)
intervals = [ws.cell(row=12, column=col).value for col in range(6,32)]
time_col_map = {t: col for col, t in zip(range(6,32), intervals) if t}
sched_data = {}
for row in range(13, 19):
    agent = ws.cell(row=row, column=5).value
    if not agent:
        continue
    sched_data[agent] = set()
    for t, col in time_col_map.items():
        val = ws.cell(row=row, column=col).value
        if val == 1:
            sched_data[agent].add(t)

# Write output sums F22:AE29
for cat_idx, cat in enumerate(categories):
    row = 22 + cat_idx
    for int_idx, t in enumerate(intervals):
        col = 6 + int_idx
        if t is None:
            continue
        count = 0
        for agent in agents:
            if t in sched_data.get(agent, set()) and cat_data.get(agent, {}).get(cat, 0):
                count += 1
        ws.cell(row=row, column=col).value = count
wb.save(output_path)
