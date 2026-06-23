import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_59224_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_59224_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

b2 = ws['B2'].value
select_period_row = None

# Find the first row where the formula would return 'Select Period'
for row in range(4, 15):
    c = ws[f'C{row}'].value
    d = ws[f'D{row}'].value
    if (b2 is not None and c is not None and d is not None and b2 > c and b2 < d):
        select_period_row = row
        break

for row in range(4, 15):
    if select_period_row is not None and row <= select_period_row:
        ws[f'E{row}'].value = 'Select Period'
    else:
        ws[f'E{row}'].value = 'Do Not Select'

wb.save(output_path)
