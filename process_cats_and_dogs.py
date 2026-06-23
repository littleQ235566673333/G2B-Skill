import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_51289_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_51289_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read row 1 and 2
labels = [ws[f'{chr(65+i)}1'].value for i in range(8)]
numbers = [ws[f'{chr(65+i)}2'].value for i in range(8)]

from collections import defaultdict
data = defaultdict(list)
for idx, (label, num) in enumerate(zip(labels, numbers)):
    if label in ['cat', 'dog'] and isinstance(num, (int, float)):
        data[label].append((num, idx))

rankings = ['']*8
for animal in ['cat', 'dog']:
    # Sort for each by value descending, keeping track of original idx
    vals = sorted(data[animal], reverse=True)
    for rank, (_, idx) in enumerate(vals[:2]):
        rankings[idx] = f'{animal}{rank+1}'

# Write to A4:H4
for i, value in enumerate(rankings):
    ws[f'{chr(65+i)}4'].value = value

wb.save(output_path)
