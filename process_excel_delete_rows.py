import openpyxl

# Define input and output paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_23-24_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_23-24_tc1/output.xlsx"

# Load workbook and active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get all words from column I (assume header at row 1)
words_to_delete = set()
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=9).value  # Column I = 9
    if val is not None:
        words_to_delete.add(str(val).strip())

# Collect rows to keep
rows_to_keep = []
for row in range(2, ws.max_row + 1):  # assume header stays
    word = ws.cell(row=row, column=1).value  # Column A = 1
    if word is None or str(word).strip() not in words_to_delete:
        rows_to_keep.append([ws.cell(row=row, column=i).value for i in range(1, 6)])

# Clear out A2:E in sheet, rewrite header and kept rows
for row in range(2, ws.max_row + 1):
    for col in range(1, 6):
        ws.cell(row=row, column=col).value = None

# Write back kept rows
for idx, vals in enumerate(rows_to_keep, start=2):
    for col, val in enumerate(vals, start=1):
        ws.cell(row=idx, column=col).value = val

# Save result
wb.save(output_path)
