from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_2/regression_gate/after_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_2/regression_gate/after_pass/core_41601/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Students']

for row in range(2, 8):
    # The formula will look like: =INDIRECT("'"&A2&"'!C2")
    formula = "=INDIRECT(""'""&A{}&""'!C2"")".format(row)
    ws[f'E{row}'] = formula

wb.save(output_path)
