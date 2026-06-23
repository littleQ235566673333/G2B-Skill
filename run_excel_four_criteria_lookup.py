import openpyxl

# Load the input workbook and select active sheet
wb = openpyxl.load_workbook('results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_55468_tc1/input.xlsx')
ws = wb.active

# Read lookup values from relevant headers
criteria_X1 = ws['AC4'].value  # e.g., X1 header
criteria_X2 = ws['AB4'].value  # e.g., X2 header
criteria_Y1 = ws['AE4'].value  # e.g., Y1 header
criteria_Y2 = ws['AD4'].value  # e.g., Y2 header

# Find the range boundaries (static based on given formulas)
data_range = ws['C5':'Z10']

# Read horizontal headers (X) in rows 3 and 4 for columns C:Z
x1_headers = [ws.cell(row=4, column=col).value for col in range(3, 27)]  # C4:Z4
x2_headers = [ws.cell(row=3, column=col).value for col in range(3, 27)]  # C3:Z3

# Read vertical headers (Y) in columns A and B for rows 5:10
y1_headers = [ws.cell(row=row, column=31).value for row in range(5, 11)] # AE for readability, unused
y2_headers = [ws.cell(row=row, column=30).value for row in range(5, 11)] # AD for readability, unused

# But actual used:
vert1_headers = [ws.cell(row=row, column=1).value for row in range(5, 11)] # A5:A10
vert2_headers = [ws.cell(row=row, column=2).value for row in range(5, 11)] # B5:B10

# Find matching coordinates
row_idx = None
for i, (v1, v2) in enumerate(zip(vert1_headers, vert2_headers)):
    if v1 == criteria_Y1 and v2 == criteria_Y2:
        row_idx = i
        break
col_idx = None
for j, (h1, h2) in enumerate(zip(x1_headers, x2_headers)):
    if h1 == criteria_X1 and h2 == criteria_X2:
        col_idx = j
        break

# If match found, fetch the value; else, set error
output = '#N/A'
if row_idx is not None and col_idx is not None:
    output = data_range[row_idx][col_idx].value

# Write result to AE5
ws['AE5'] = output

# Save to the specified output file
wb.save('results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_55468_tc1/output.xlsx')
