from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_1/group_44017/r3/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_1/group_44017/r3/evolve_44017/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

no_fill = PatternFill(fill_type=None)

def get_numeric(val, default=1):
    try:
        if val is None or val == '' or (isinstance(val, str) and not val.replace('.', '', 1).isdigit()):
            return default
        return float(val)
    except Exception:
        return default

def parse_date(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d-%b-%y', '%Y/%m/%d'):
            try:
                return datetime.strptime(val, fmt)
            except Exception:
                continue
    return None

for data_row in range(14, 43):
    base_rate = ws.cell(row=data_row, column=23).value # W
    eff_date_raw = ws.cell(row=data_row, column=12).value # L
    eff_date = parse_date(eff_date_raw)
    freq_raw = ws.cell(row=data_row, column=10).value # J
    freq = get_numeric(freq_raw, 1)
    increases = [
        ws.cell(row=data_row, column=13).value or 0, # M
        ws.cell(row=data_row, column=14).value or 0, # N
        ws.cell(row=data_row, column=15).value or 0, # O
        ws.cell(row=data_row, column=16).value or 0, # P
    ]
    increases = [get_numeric(inc, 0) for inc in increases]
    for i, col in enumerate(range(30, 44)):
        date_cell_raw = ws.cell(row=9, column=col).value
        date_cell = parse_date(date_cell_raw)
        cell = ws.cell(row=data_row, column=col)
        if not (date_cell and eff_date and date_cell >= eff_date):
            cell.value = None
            cell.fill = no_fill  # Remove any fill
            continue
        months_since_effect = (date_cell.year - eff_date.year) * 12 + (date_cell.month - eff_date.month)
        wave = 0
        while wave < 4 and months_since_effect >= wave * freq:
            wave += 1
        multiplier = 1
        for w in range(wave):
            multiplier *= 1 + (increases[w] / 100)
        if base_rate is not None:
            cell.value = round(base_rate * multiplier, 2)
        else:
            cell.value = None
        cell.fill = no_fill

wb.save(output_path)
