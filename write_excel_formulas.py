import openpyxl
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_2/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_2/regression_gate/after_pass/core_50526/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Identify columns/rows
header_start_col = 2 # B
header_end_col = ws.max_column
header_range = f"$B$1:${get_column_letter(header_end_col)}$1"
data_start_row = 2  # after header
max_row = ws.max_row
data_range = f"$B$2:${get_column_letter(header_end_col)}${max_row}"
lookup_range = f"$A$2:$A${max_row}"

for i in range(2): # B9, B10
    cell = ws.cell(row=9+i, column=2)
    formula = (
        f"=IFERROR(INDEX({header_range}, "
        f"SMALL(IF(INDEX({data_range}, MATCH($B$6, {lookup_range}, 0), COLUMN({header_range})-COLUMN($B$1)+1)>0, "
        f"COLUMN({header_range})-COLUMN($B$1)+1), ROW(A1)" + (f"+{i}" if i>0 else "") + ")), \"\")"
    )
    cell.value = formula

wb.save(output_path)
print('Done')
