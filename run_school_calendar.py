import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_3/group_50916/r0/evolve_50916/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_3/group_50916/r0/evolve_50916/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Find cycle days list and class schedule table
cycle_days = []
schedule = []
for i in range(2,9):  # A2:A8
    cycle_days.append(ws[f'A{i}'].value)
    # For each day, get columns C-H (3-8)
    schedule.append([ws.cell(row=i, column=col).value for col in range(3,9)])

def lookup_classes(cycle_value):
    '''Given a cycle_value, return the classes row for that cycle day, or blanks if not match'''
    if cycle_value is None:
        return ['']*6
    for i, val in enumerate(cycle_days):
        if val == cycle_value:
            return schedule[i]
    return ['']*6

# Rows 12-14, Cols C-H
for row in range(12,15):
    cycle_entry = ws[f'A{row}'].value
    classes = lookup_classes(cycle_entry)
    for jj, col in enumerate(range(3,9)):
        ws.cell(row=row, column=col).value = classes[jj]

wb.save(output_path)
