import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_5/regression_gate/after_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_5/regression_gate/after_pass/core_55421/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

def find_rows_with_number(search_number):
    result_rows = []
    for r in range(2, 21):
        a_val = ws[f'A{r}'].value
        if a_val == search_number:
            result_rows.append(r)
    return result_rows

def get_statuses_dates_for_number(search_number):
    rows = find_rows_with_number(search_number)
    statuses = set()
    date_with_noshow = False
    empty_date_with_noshow = False
    for row in rows:
        status = (ws[f'D{row}'].value or '').strip()
        date = ws[f'E{row}'].value
        statuses.add(status)
        if status == 'NO SHOW':
            if date:
                date_with_noshow = True
            else:
                empty_date_with_noshow = True
    return statuses, date_with_noshow, empty_date_with_noshow

def classify(a_val):
    if a_val is None:
        return ''
    statuses, date_with_noshow, empty_date_with_noshow = get_statuses_dates_for_number(a_val)
    if statuses == {'SCH'}:
        return 'FUTURE'
    if 'SCH' in statuses and 'NO SHOW' in statuses:
        return 'NS/SCHED'
    if statuses == {'NO SHOW'}:
        if date_with_noshow:
            return 'NO ACTION NEEDED'
        if empty_date_with_noshow:
            return 'CALL PT'
    if 'NO SHOW' in statuses and date_with_noshow:
        return 'NO ACTION NEEDED'
    if 'NO SHOW' in statuses and empty_date_with_noshow:
        return 'CALL PT'
    return ''

for r in range(2, 21):
    a_val = ws[f'A{r}'].value
    res = classify(a_val)
    ws[f'F{r}'] = res

wb.save(output_path)
