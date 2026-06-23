import itertools
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

inp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun2/eval_254-34_tc1/input.xlsx'
outp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun2/eval_254-34_tc1/output.xlsx'

# Load numbers and workbook
wb = load_workbook(inp)
ws = wb['Before']
df = pd.read_excel(inp, sheet_name='Before')
values = df.iloc[:,0].values

found = False
for lo in range(994108, 994113):
    for r in range(2, min(10, len(values)+1)):
        for combo in itertools.combinations(enumerate(values), r):
            idx, nums = zip(*combo)
            s = sum(nums)
            if s == lo:
                fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                for i in idx:
                    ws[f'A{i+1}'].fill = fill
                ws['C2'] = s
                found = True
                break
        if found:
            break
    if found:
        break
if 'Sheet3' in wb.sheetnames:
    del wb['Sheet3']
wb.save(outp)
if found:
    print(f"Done, sum: {s} at rows: {[i+1 for i in idx]}")
else:
    print('No valid combination found')
