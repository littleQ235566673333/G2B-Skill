import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_1/group_516-46/r0/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_1/group_516-46/r0/evolve_516-46/output.xlsx'

# Read the sheet 'ورقة1'
df = pd.read_excel(input_path, sheet_name='ورقة1', header=0)

# Restrict to columns A-E, skip the header row
cols = df.columns[:5]
data = df.loc[:, cols].iloc[1:].copy()

# Convert dates to datetime (assuming B col is the date column)
data[cols[1]] = pd.to_datetime(data[cols[1]], errors='coerce')

# Drop rows without a brand or date
data = data.dropna(subset=[cols[0], cols[1]])

# Group by brand, select rows with the last date per brand, and combine qty
result_rows = []
for brand, rows in data.groupby(cols[0]):
    last_date = rows[cols[1]].max()
    latest = rows[rows[cols[1]] == last_date]
    sum_qty = latest[cols[3]].astype(float).sum()
    first_row = latest.iloc[0].copy()
    first_row[cols[3]] = sum_qty
    # If more than one entry for the latest date/brand, flag as 'Modified'
    merged_flag = 'Modified' if len(latest) > 1 else ''
    result_rows.append(list(first_row.values) + [merged_flag])

# Compose output DataFrame
output_cols = list(cols) + ['Merged']
output = pd.DataFrame(result_rows, columns=output_cols)

# Write to the output Excel starting at H2 in 'ورقة1'
wb = load_workbook(input_path)
ws = wb['ورقة1']
start_row, start_col = 2, 8 # H2: row=2, col=8

# Write headers
for j, val in enumerate(output_cols):
    ws.cell(row=start_row, column=start_col + j, value=val)
# Write data
for i, row in enumerate(output.values.tolist()):
    for j, val in enumerate(row):
        ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)

wb.save(output_path)
