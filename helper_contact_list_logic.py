import datetime
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_2/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_2/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']
today = datetime.datetime.today()
row = 4
h_cell = ws[f'H{row}'].value
i_cell = ws[f'I{row}'].value

value = 'NO ACTION'  # Default
try:
    if isinstance(h_cell, datetime.datetime):
        days_diff = (today - h_cell).days
    elif isinstance(h_cell, str):
        # Attempt to parse date from string
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d'):  # common formats
            try:
                h_date = datetime.datetime.strptime(h_cell, fmt)
                days_diff = (today - h_date).days
                break
            except Exception:
                continue
        else:
            days_diff = None
    else:
        days_diff = None

    if days_diff is not None:
        if days_diff <= 30 and str(i_cell).strip().upper() == 'YES':
            value = 'HOLD'
        elif days_diff > 30 and str(i_cell).strip().upper() == 'YES':
            value = 'TOUCH BASE'
        else:
            value = 'NO ACTION'
    else:
        value = 'NO ACTION'
except Exception as e:
    value = 'NO ACTION'

ws[f'J{row}'] = value
wb.save(output_path)
