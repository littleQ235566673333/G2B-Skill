import openpyxl
from openpyxl.utils import column_index_from_string
from datetime import datetime, timedelta

# Paths
infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/group_44017/r3/evolve_44017/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/group_44017/r3/evolve_44017/output.xlsx'

wb = openpyxl.load_workbook(infile)
ws = wb.active

output_start_col = column_index_from_string('AD') # 30
output_end_col = column_index_from_string('AO') # 41
output_cols = range(output_start_col, output_end_col+1)
output_start_row = 14
output_end_row = 42
data_rows = range(output_start_row, output_end_row+1)

# Map month columns to their dates from row 9
date_by_col = {}
for col in output_cols:
    cell_val = ws.cell(row=9, column=col).value
    if isinstance(cell_val, datetime):
        date_by_col[col] = cell_val
    else:
        # Try to parse if needed
        try:
            date_by_col[col] = datetime.strptime(str(cell_val), '%Y-%m-%d')
        except Exception:
            date_by_col[col] = None

# For each row, extract relevant fields and fill output
for row in data_rows:
    base_rate = ws.cell(row=row, column=23).value # W (base rate)
    eff_date = ws.cell(row=row, column=12).value # L
    freq = ws.cell(row=row, column=10).value # J
    increases = [ws.cell(row=row, column=13+i).value for i in range(4)] # M:P
    # Clean increases values (as percent/100)
    increases = [(x/100.0 if isinstance(x, (int, float)) else 0.0) for x in increases]
    # Parse freq
    try:
        freq = int(freq)
    except Exception:
        freq = 12 # default to annual if blank
    for idx, col in enumerate(output_cols):
        cell = ws.cell(row=row, column=col)
        month_date = date_by_col.get(col, None)
        # Compare output date to effective date
        if not (base_rate and eff_date and month_date):
            cell.value = ''
            continue
        # Only fill if month_date >= eff_date
        if month_date < eff_date:
            cell.value = ''
            continue
        # Calculate how many increases have been applied
        increases_applied = 0
        for i in range(4):
            # +0 for the first wave, +freq for each
            try:
                wave_month = eff_date.month + freq * i
                wave_year = eff_date.year + (wave_month-1)//12
                wave_month = ((wave_month-1)%12)+1
                wave_date = eff_date.replace(year=wave_year, month=wave_month)
            except Exception:
                # Fallback to timedelta if month math fails
                wave_date = eff_date + timedelta(days=30 * freq * i)
            if month_date >= wave_date:
                increases_applied += 1
        # Calc cumulative multiplier
        multiplier = 1.0
        for i in range(increases_applied):
            multiplier *= 1 + increases[i]
        # Write calculated value
        if base_rate is not None:
            cell.value = round(base_rate * multiplier, 5)
        else:
            cell.value = ''

wb.save(outfile)
