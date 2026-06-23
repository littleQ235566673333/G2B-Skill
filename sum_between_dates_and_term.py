import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def main():
    input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_38823_tc1/input.xlsx'
    output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_38823_tc1/output.xlsx'

    # Load full data and search terms
    df = pd.read_excel(input_path, sheet_name=0)

    # Open workbook for writing
    wb = load_workbook(input_path)
    ws = wb.active

    # Assume columns:
    # Date (col X, type date), Value (col Y, to sum), Category/Text (col Z, where to search),
    # H4:H7: search terms, I4:I7: output cells
    # Find header row and column indices
    header = {v: i for i, v in enumerate(df.columns)}
    date_col = [c for c in df.columns if 'date' in str(c).lower()][0]
    value_col = [c for c in df.columns if 'value' in str(c).lower()][0]
    search_col = [c for c in df.columns if c not in [date_col, value_col]][0]

    # Read start/end dates
    start_date = ws['F2'].value
    end_date = ws['F3'].value
    if isinstance(start_date, str):
        start_date = pd.to_datetime(start_date)
    if isinstance(end_date, str):
        end_date = pd.to_datetime(end_date)

    for i, row in enumerate(range(4, 8), start=0):
        search_term = ws[f'H{row}'].value
        if not search_term:
            ws[f'I{row}'] = ''
            continue
        cond = (
            (df[date_col] >= start_date) &
            (df[date_col] <= end_date) &
            (df[search_col].astype(str).str.contains(str(search_term), na=False, case=False))
        )
        sum_val = df.loc[cond, value_col].sum()
        ws[f'I{row}'] = float(sum_val) if pd.notnull(sum_val) else 0
    wb.save(output_path)

if __name__ == '__main__':
    main()
