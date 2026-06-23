import pandas as pd
from openpyxl import load_workbook
from datetime import timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_2/group_59595/r0/evolve_59595/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_2/group_59595/r0/evolve_59595/output.xlsx'

# Adjust columns: Data starts at row 3, columns 'Unnamed: 0' (dates), 'Unnamed: 1' (points)
df = pd.read_excel(input_path, skiprows=2, usecols=[0,1], nrows=16, names=['Date', 'Points'])
df['Date'] = pd.to_datetime(df['Date'])

sums = []
for idx, this_date in enumerate(df['Date']):
    # Past 7 days includes today and 6 days back
    start_date = this_date - timedelta(days=6)
    mask = (df['Date'] >= start_date) & (df['Date'] <= this_date)
    total = df.loc[mask, 'Points'].sum()
    sums.append(total)

wb = load_workbook(input_path)
ws = wb.active
for i, s in enumerate(sums, start=4): # write from C4 to C19
    ws[f'C{i}'] = s
wb.save(output_path)
