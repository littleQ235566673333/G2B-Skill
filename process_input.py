import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_13284_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_13284_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
base_sheet = wb['Base']
streets_sheet = wb['Streets']

# Read all valid streets rows into a list of dicts
streets_data = []
for row in streets_sheet.iter_rows(min_row=2, values_only=True):
    street_B = row[1]  # column B
    min_val = row[2]   # column C
    max_val = row[3]   # column D
    street_F = row[5]  # column F
    if min_val is not None and max_val is not None:
        streets_data.append({'B': street_B, 'C': min_val, 'D': max_val, 'F': street_F})

# Fill E2:E26 in Base
for base_row in range(2, 27):
    value_C = base_sheet.cell(row=base_row, column=3).value
    value_D = base_sheet.cell(row=base_row, column=4).value
    output = ''
    if value_C is not None and value_D is not None:
        try:
            value_D_num = float(value_D)
        except Exception:
            value_D_num = None
        if value_D_num is not None:
            for s in streets_data:
                if s['B'] == value_C:
                    try:
                        min_num = float(s['C'])
                        max_num = float(s['D'])
                    except Exception:
                        continue
                    if min_num <= value_D_num <= max_num:
                        output = s['F']
                        break
    base_sheet.cell(row=base_row, column=5).value = output

wb.save(output_path)
