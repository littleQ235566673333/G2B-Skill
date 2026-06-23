import openpyxl
from openpyxl.styles import PatternFill

# File paths
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_120-24_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_120-24_tc1/output.xlsx'

# Targets
bg_targets = [
    'OCOGS - Spares - Transfer Price Overhead',
    'OFSS - ORCL Consulting Prod Cost I/C'
]
ay_targets = ['BOA_033E', 'BOA_011G']
set_ocogs = 'OCOGS - Spares - Transfer Price Overhead'
set_ofss = 'OFSS - ORCL Consulting Prod Cost I/C'

# Column indices (Excel is 1-based)
COL_BL = 64   # BL
COL_BG = 59   # BG
COL_AY = 51   # AY
COL_BN = 66   # BN

# Fill color for #0070C0
fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

for row in range(2, 46):  # From 2 to 45 (inclusive)
    bl = ws.cell(row=row, column=COL_BL).value
    bg = ws.cell(row=row, column=COL_BG).value
    ay = ws.cell(row=row, column=COL_AY).value

    if bl == 'LAG' and bg in bg_targets:
        if ay in ay_targets:
            ws.cell(row=row, column=COL_BN).value = set_ocogs
        else:
            ws.cell(row=row, column=COL_BN).value = set_ofss

    # Set fill in BL if BL == 'LAG' and BN (column 66) is now populated
    bn_val = ws.cell(row=row, column=COL_BN).value
    if bl == 'LAG' and bn_val:
        ws.cell(row=row, column=COL_BL).fill = fill

wb.save(output_path)
