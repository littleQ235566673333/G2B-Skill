import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_38969_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_38969_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 21):  # Rows 2 to 20
    S = ws[f'S{row}'].value
    U = ws[f'U{row}'].value
    result = None
    # Check if S is error
    if S == 'error' or (isinstance(S, str) and S.lower() == 'error'):
        result = 'Upload'
    elif S != 'error' and S is not None:
        # We only evaluate U if S is not error
        try:
            U_val = float(U)
            if -1 <= U_val <= 1:
                result = 'Do not Upload'
            else:
                result = 'to Check'
        except (TypeError, ValueError):
            # U is not a valid number, fallback
            result = 'to Check'
    else:
        result = 'to Check'
    ws[f'R{row}'] = result

wb.save(output_path)
