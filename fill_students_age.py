from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_4/regression_gate/after_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_4/regression_gate/after_pass/core_41601/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Students']

# Write the correct formula in E2:E7: =INDIRECT("'"&A2&"'!C2")
for i in range(2, 8):
    ws[f'E{i}'] = f"=INDIRECT(\"'\"&A{i}&\"'!C2\")"

wb.save(output_path)
