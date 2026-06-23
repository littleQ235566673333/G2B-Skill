import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_9448_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_9448_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheet = wb['Data']

# Define columns I (9) to T (20), rows 9-18, header row is 7
start_col = 9  # I
end_col = 20   # T
rows = range(9, 19)  # U9:U18

for row in rows:
    last_sale_col = None
    # Iterate columns from T -> I
    for col in range(end_col, start_col - 1, -1):
        cell = sheet.cell(row=row, column=col)
        # Check for number (sale quantity > 0)
        if isinstance(cell.value, (int, float)) and cell.value > 0:
            last_sale_col = col
            break
    if last_sale_col:
        header = sheet.cell(row=7, column=last_sale_col).value
        # Try to parse header as a month/year, e.g., 'Apr 2023', 'April', etc.
        date_val = None
        # Examples: 'Apr 2023', 'April 2023', '2023-04', '04/2023', 'April'
        try:
            for fmt in ["%b %Y", "%B %Y", "%Y-%m", "%m/%Y", "%B", "%b"]:
                try:
                    dt = datetime.strptime(str(header), fmt)
                    # Always set date as 1st of the month
                    date_val = dt.replace(day=1)
                    break
                except Exception:
                    continue
            if date_val is None:
                # If parsing fails, set text
                sheet.cell(row=row, column=21).value = header
            else:
                sheet.cell(row=row, column=21).value = date_val
                sheet.cell(row=row, column=21).number_format = 'mm/dd/yyyy'
        except Exception:
            sheet.cell(row=row, column=21).value = header
    else:
        sheet.cell(row=row, column=21).value = None

wb.save(output_path)
