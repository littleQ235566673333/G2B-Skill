import openpyxl
from collections import defaultdict
from datetime import datetime

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_1/group_91-34/r0/evolve_91-34/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_1/group_91-34/r0/evolve_91-34/output.xlsx'

wb = openpyxl.load_workbook(in_path)
ws = wb['SwiftMD']
rows = list(ws.iter_rows(min_row=2, max_row=42, max_col=15))
headers = [c.value for c in rows[0]]
data = []
for r in rows[1:]:
    data.append([c.value for c in r])

# Build index for groups
groups = defaultdict(list)
for idx, row in enumerate(data):
    # Last Name C: idx=1, First Name D=2, DOB F=4, Duplicate? G=5, Relationship H=6
    key = (row[1], row[2], row[4])
    groups[key].append((idx, row))

# Figure out rows to remove (just 1 per duplicate non-employee group)
to_remove = set()
for key, recs in groups.items():
    # Find those where Duplicate? == 'Yes' and Relationship != 'Employee'
    non_emp_yes = [(i, r) for i, r in recs if r[5] == 'Yes' and (r[6] or '').strip() != 'Employee']
    has_employee = any((r[6] or '').strip() == 'Employee' for _, r in recs)
    # If any is an employee, skip
    if has_employee:
        continue
    # If 2+ non-employee dups, mark one for removal
    if len(non_emp_yes) > 1:
        # Remove the one with highest index (stable)
        to_remove.add(non_emp_yes[0][0])

# Now, rebuild the data array
cleaned_data = [data[i] for i in range(len(data)) if i not in to_remove]

# Write back headers and cleaned data to B2:O42 (match original layout)
# Write headers (already in row 2)
for j, val in enumerate(headers):
    ws.cell(row=2, column=j+1, value=val)
# Write data
for i, row in enumerate(cleaned_data):
    for j, val in enumerate(row):
        ws.cell(row=i+3, column=j+1, value=val)
# Blank the rest
for i in range(len(cleaned_data)+3, 43):
    for j in range(1, 16):
        ws.cell(row=i, column=j, value=None)

wb.save(out_path)
