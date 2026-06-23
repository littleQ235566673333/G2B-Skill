from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42/eval_16511_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42/eval_16511_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

for row in range(2, 11):
    ws[f'F{row}'] = f'=IF(D{row}=D{row-1},F{row-1}+1,1)'

wb.save(output_path)
