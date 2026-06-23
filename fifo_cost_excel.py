from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def fifo_per_unit_formula_for_row(row):
    # Define inventory columns and their costs in sequential order
    inv_cols = [1, 3, 5, 7, 9]  # B (beg inv), D (PO-1), F (PO-2), H (PO-3), J (PO-4)
    cost_cols = [2, 4, 6, 8, 10] # C (beg cost), E (cost1), G (cost2), I (cost3), K (cost4)
    # Excel columns are 0-based here, but openpyxl is 1-based, adjust below

    # Build formula string in Excel syntax for FIFO cost
    base_col = ord('A')
    # Inventory on hand reference in column L
    cur_inv_ref = f"L{row}"
    
    parts = []
    qty_refs = []
    cost_refs = []
    for i in range(len(inv_cols)):
        q_ref = f"{get_column_letter(inv_cols[i]+1)}{row}"
        c_ref = f"{get_column_letter(cost_cols[i]+1)}{row}"
        qty_refs.append(q_ref)
        cost_refs.append(c_ref)

    # Build piecewise sumproduct
    # We'll walk through the inventory as FIFO, accumulating from beginning and each PO
    # Running sum of available units, and for partial usage, effective cost

    # Use only Excel formula logic here, not Python calculation
    # We'll simulate the running total required to match Current Inv (column L)

    # FIFO sumproduct logic via piecewise SUM
    # We break into chunks: for each layer, if still more required, use the full layer, otherwise, only take the remainder
    formula = ""
    used = "0"
    cost_sum = []
    qty_sum = []
    remaining = f"$L{row}"  # Current inventory to fill
    for i in range(len(qty_refs)):
        # units to take from this layer: min(remaining, this_qty)
        layer_amt = f"MIN({remaining},{qty_refs[i]})"
        # cost this layer: min(remaining, qty) * cost
        cost_sum.append(f"{layer_amt}*{cost_refs[i]}")
        qty_sum.append(layer_amt)
        remaining = f"MAX(0,{remaining}-{qty_refs[i]})"

    sum_cost = "+".join(cost_sum)
    sum_qty = f"$L{row}"
    formula = f"=IF({sum_qty}=0,0,({sum_cost})/{sum_qty})"
    return formula

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/group_39432/r0/evolve_39432/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/group_39432/r0/evolve_39432/output.xlsx"

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

for excel_row in range(2, 6):  # Rows 2-5
    formula = fifo_per_unit_formula_for_row(excel_row)
    ws[f"M{excel_row}"] = formula

wb.save(output_path)
