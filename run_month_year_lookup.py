import openpyxl

# Open the workbook and select the active sheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_3/group_39515/r3/evolve_39515/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_3/group_39515/r3/evolve_39515/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get date headers (C1:N1)
headers = [ws.cell(row=1, column=j).value for j in range(3, 15)]  # C1:N1

import re
from calendar import month_abbr

def normalize_month(text):
    """
    Normalize a month string (e.g., 'Jan', 'January', '01', etc.) to a 3-letter abbreviation lowercase
    """
    text = str(text).strip()
    # Try to extract month from numeric (e.g. '1' or '01')
    if text.isdigit():
        idx = int(text)
        if 1 <= idx <= 12:
            return month_abbr[idx].lower()
    # Try matching using calendar month abbreviations
    for idx in range(1, 13):
        abbr = month_abbr[idx].lower()
        full = abbr.capitalize()
        if text.lower().startswith(abbr) or text.lower().startswith(full.lower()) or text.lower().startswith(full.lower()[:-1]):
            return abbr
    # Return the first 3 letters, lowercase (fallback)
    return text[:3].lower()

def normalize_year(text):
    y_search = re.search(r'\d{4}', str(text))
    if y_search:
        return y_search.group()
    return str(text)

def find_col_index_by_month_year(month_name, year_value):
    target_month = normalize_month(month_name)
    target_year = normalize_year(year_value)
    for idx, val in enumerate(headers):
        if not val:
            continue
        header = str(val)
        # Try to extract year and month from header
        # Accept formats like 'Jan 2022', '2022 Jan', 'January 2022', '2022-01', '01/2022', etc.
        y_in_header = normalize_year(header)
        all_parts = re.split(r'[\s/-]', header)
        month_in_header = None
        for part in all_parts:
            normalized_part = normalize_month(part)
            if normalized_part in [m.lower() for m in month_abbr if m]:
                month_in_header = normalized_part
                break
        if month_in_header == target_month and y_in_header == target_year:
            return idx
    return None

for i in range(2, 14):  # O2:O13 is rows 2 to 13
    month_cell = ws.cell(row=i, column=1).value  # A
    year_cell = ws.cell(row=i, column=2).value   # B
    if month_cell is None or year_cell is None:
        ws.cell(row=i, column=15).value = None  # O
        continue
    col_idx = find_col_index_by_month_year(month_cell, year_cell)
    if col_idx is None:
        ws.cell(row=i, column=15).value = None
        continue
    val = ws.cell(row=i, column=col_idx+3).value  # C is column 3 offset
    ws.cell(row=i, column=15).value = val

wb.save(output_path)
