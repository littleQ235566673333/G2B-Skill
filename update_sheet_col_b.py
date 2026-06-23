import openpyxl

# Paths
input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/regression_gate/before_pass/core_302-1/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/regression_gate/before_pass/core_302-1/output.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_fp)
items_ws = wb['ITEMS']

# Read all reference values in column B (B2 ...)
items_col_b = [items_ws[f'B{i}'].value for i in range(2, items_ws.max_row+1)]

def update_col_b(sheet_name, num_rows):
    ws = wb[sheet_name]
    for idx in range(num_rows):
        # Sheet's cell
        row = idx + 2  # Starts at B2
        # Only update if row exists in both ITEMS and this sheet
        if idx < len(items_col_b):
            if ws[f'B{row}'].value != items_col_b[idx]:
                ws[f'B{row}'].value = items_col_b[idx]

# Update Sheet1 B2:B8 (rows 2-8, 7 rows)
update_col_b('SHEET1', 7)
# Update Sheet2 B2:B7 (rows 2-7, 6 rows)
update_col_b('SHEET2', 6)

wb.save(output_fp)
