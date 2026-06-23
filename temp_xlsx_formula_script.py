from openpyxl import load_workbook
import string

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_1/group_36277/r2/evolve_36277/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_1/group_36277/r2/evolve_36277/output.xlsx'

wb = load_workbook(in_path)
ws = wb.active  # Use the first sheet

header_row = 1
first_col = 1
header_cols = ws.max_column

# Excel column naming (A, B, ..., Z, AA, AB, ...)
def col_letter(col):
    result = ''
    while col:
        col, rem = divmod(col-1, 26)
        result = chr(65 + rem) + result
    return result

max_row = ws.max_row

data_rows = range(2, 6)  # I2:I5

top_left = col_letter(first_col) + str(header_row+1)
bottom_right = col_letter(header_cols) + str(max_row)
header_left = col_letter(first_col) + str(header_row)
header_right = col_letter(header_cols) + str(header_row)

for r in data_rows:
    # INDEX(data_range, relative row, matched col)
    # ROW()-1 to get 1 for row 2, etc.
    formula = f'=INDEX(${col_letter(first_col)}${header_row+1}:${col_letter(header_cols)}${max_row},ROW()-1,MATCH($I$1,${col_letter(first_col)}${header_row}:${col_letter(header_cols)}${header_row},0))'
    ws[f'I{r}'] = formula

wb.save(out_path)
