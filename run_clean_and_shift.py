from openpyxl import load_workbook

# File paths
IN_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_5/regression_gate/after_pass/core_493-18/input.xlsx'
OUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_5/regression_gate/after_pass/core_493-18/output.xlsx'

wb = load_workbook(IN_PATH)
ws = wb.active

# 1. Collect filter set from F
vals_f = set()
for row in ws.iter_rows(min_row=2, max_col=6, min_col=6, values_only=True):
    if row[0] is not None:
        vals_f.add(row[0])

# 2. Step 1: Clear A, B, C for rows where A is not in F
data_abc = []
rows = ws.max_row
for i in range(2, rows+1):
    val_a = ws.cell(i, 1).value
    if val_a in vals_f:
        data_abc.append([ws.cell(i, 1).value, ws.cell(i, 2).value, ws.cell(i, 3).value])
    else:
        ws.cell(i, 1).value = None
        ws.cell(i, 2).value = None
        ws.cell(i, 3).value = None

# 3. Step 2: Shift remaining (non-None) ABC cells upwards
# (remove blank rows in ABC, fill from row 2)
for i, vals in enumerate(data_abc, start=2):
    ws.cell(i, 1).value = vals[0]
    ws.cell(i, 2).value = vals[1]
    ws.cell(i, 3).value = vals[2]

# After shifting, blank out anything below the last valid in ABC
for i in range(2+len(data_abc), rows+1):
    ws.cell(i, 1).value = None
    ws.cell(i, 2).value = None
    ws.cell(i, 3).value = None

# 4. Set autofilter on A1:C1
ws.auto_filter.ref = 'A1:C1'

wb.save(OUT_PATH)
print('Done')
