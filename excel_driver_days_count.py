import openpyxl
import numpy as np

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_57989_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_57989_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read relevant data (Assume data is within rows 1-22, cols A-U)
data = []
for row in ws.iter_rows(min_row=1, max_row=22, min_col=1, max_col=21, values_only=True):
    data.append(list(row))

# Identify day-of-week columns
header = data[0]
day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
weekday_columns = {day: [] for day in day_names}

for col_idx, col_name in enumerate(header):
    for d in day_names:
        if d.lower() in str(col_name).lower():
            weekday_columns[d].append(col_idx)

# Find driver names (assume they're in the first column, starting from row 2)
driver_names = []
driver_rows = {}
for i in range(1, len(data)):
    driver = data[i][0]
    if driver and driver not in driver_names:
        driver_names.append(driver)
        driver_rows[driver] = i

# Synthesis table: Count non-empty for each (driver, weekday)
summary = np.zeros((len(driver_names), len(day_names)), dtype=int)
for d_idx, driver in enumerate(driver_names):
    r = driver_rows[driver]
    for wd_idx, wd in enumerate(day_names):
        cnt = 0
        for col in weekday_columns[wd]:
            val = data[r][col]
            if val is not None and str(val).strip() != '':
                cnt += 1
        summary[d_idx, wd_idx] = cnt

# Write results starting at B25
start_row = 25
start_col = 2 # Column B

# Write weekday headers
for j, wd in enumerate(day_names):
    ws.cell(start_row, start_col + j, wd)
# Write driver names and counts
for i, driver in enumerate(driver_names):
    ws.cell(start_row + 1 + i, start_col - 1, driver)
    for j in range(len(day_names)):
        ws.cell(start_row + 1 + i, start_col + j, summary[i, j])

wb.save(output_path)
print('Done.')
