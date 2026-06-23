import pandas as pd
from openpyxl import load_workbook

i_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_2/regression_gate/before_pass/core_9726/input.xlsx'
o_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_2/regression_gate/before_pass/core_9726/output.xlsx'

# Read the relevant sheet
full = pd.read_excel(i_path, sheet_name='Sheet1')

# Table 1: A-C, Table 2: H-J. But columns are 'Student ID', 'Visit number', 'Visit date',...,'Student ID.1','Submission date', 'Visit date.1'
t1 = full[['Student ID', 'Visit date']].dropna(subset=['Student ID', 'Visit date'])
t2 = full[['Student ID.1', 'Submission date']].copy().dropna(subset=['Student ID.1', 'Submission date'])

t1['Visit date'] = pd.to_datetime(t1['Visit date'], errors='coerce')
t2['Submission date'] = pd.to_datetime(t2['Submission date'], errors='coerce')
# Some IDs may be floats, ensure as str

def get_prev_visit(student, sub_date):
    filt = (t1['Student ID'].astype(str) == str(student)) & (t1['Visit date'] < sub_date)
    visits = t1.loc[filt, 'Visit date']
    return visits.max() if not visits.empty else pd.NaT

# Compute results
visit_dates = [get_prev_visit(sid, subd) for sid, subd in zip(t2['Student ID.1'], t2['Submission date'])]

# Write results to Sheet1, col J (index 10), starting from row 2
wb = load_workbook(i_path)
ws = wb['Sheet1']
row_lookup = full[['Student ID.1', 'Submission date']].dropna(subset=['Student ID.1', 'Submission date'])

for i, visit_date in enumerate(visit_dates, start=2):
    ws[f'J{i}'] = visit_date.strftime('%d-%b-%y') if pd.notnull(visit_date) else ''

wb.save(o_path)
