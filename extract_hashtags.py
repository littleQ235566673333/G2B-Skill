from openpyxl import load_workbook
import re

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_5/group_57113/r3/evolve_57113/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_5/group_57113/r3/evolve_57113/output.xlsx'

# Load the workbook and worksheet
wb = load_workbook(input_path)
ws = wb.active

def extract_hashtags(text):
    '''Extract hashtags (without #) from a tweet text'''
    if not text:
        return []
    return [tag[1:] for tag in re.findall(r'#\w+', str(text))]

# Fill B2:C3 per instruction (adjacent cells, so up to 2 hashtags for this example, per tweet)
for row in range(2, 4):  # B2:C3 means rows 2 and 3
    tweet = ws.cell(row=row, column=1).value
    tags = extract_hashtags(tweet)
    # Write up to 2 hashtags, as output is B and C columns only
    for i in range(2):
        ws.cell(row=row, column=2 + i).value = tags[i] if i < len(tags) else None

wb.save(output_path)
