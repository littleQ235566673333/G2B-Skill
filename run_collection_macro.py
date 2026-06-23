import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def safe_float(x):
    if pd.isna(x) or x == '-' or x == '' or x is None:
        return 0
    try:
        return float(x)
    except Exception:
        return 0

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/group_398-14/r2/evolve_398-14/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/group_398-14/r2/evolve_398-14/output.xlsx'

def load_data():
    all_sheets = pd.read_excel(INPUT, sheet_name=None)
    return all_sheets

def aggregate_rows(all_sheets, header):
    rows = []
    for sname, sheet in all_sheets.items():
        if sname.lower() == 'collection':
            continue
        if set(header) <= set(sheet.columns):
            rows += sheet[list(header)].to_dict('records')
    from collections import defaultdict
    grouped = defaultdict(list)
    for row in rows:
        key = (row['TY'], row['OR'])
        grouped[key].append(row)
    compacted = []
    for key, group in grouped.items():
        base = {k: v for k, v in group[0].items() if k not in ['SALE','RET']}
        sale_sum = sum([safe_float(g['SALE']) for g in group])
        ret_sum  = sum([safe_float(g['RET'])  for g in group])
        base['SALE'] = sale_sum
        base['RET']  = ret_sum
        compacted.append(base)
    return compacted

def prep_collection_table(all_sheets):
    ws_c = all_sheets.get('COLLECTION', None)
    if ws_c is None:
        raise ValueError('COLLECTION sheet missing!')
    existing = ws_c.copy()
    header = list(existing.columns)
    header = header[:7]
    agg_rows = aggregate_rows(all_sheets, header[:-1])
    index_map = {(row['TY'], row['OR']): i for i, row in existing.iterrows()}
    collected = []
    used_keys = set()
    for i, erow in existing.iterrows():
        key = (erow['TY'], erow['OR'])
        match = next((r for r in agg_rows if (r['TY'], r['OR']) == key), None)
        d = {c: erow[c] for c in header if c in erow}
        if match is not None:
            d['SALE'] = match['SALE']
            d['RET'] = match['RET']
        collected.append(d)
        used_keys.add(key)
    for r in agg_rows:
        key = (r['TY'], r['OR'])
        if key not in used_keys:
            d = dict(r)
            for h in header:
                if h not in d:
                    d[h] = None
            collected.append(d)
    for d in collected:
        sale = d.get('SALE', None)
        ret = d.get('RET', None)
        bal = None
        if sale is not None and ret is not None and sale != '-' and ret != '-':
            try:
                bal = float(sale)-float(ret)
            except:
                bal = None
        d['BALANCE'] = bal
    out_rows = collected[:8]
    final_data = [header]
    for d in out_rows:
        row = []
        for h in header:
            v = d.get(h, None)
            if h == 'BALANCE':
                if v is None or pd.isna(v):
                    row.append('-')
                else:
                    if v == 0:
                        row.append('-')
                    else:
                        row.append(round(v,2))
            else:
                if v is None or pd.isna(v) or v == '':
                    row.append('-')
                else:
                    row.append(v)
        final_data.append(row)
    return final_data

def write_collection(final_data):
    wb = load_workbook(INPUT)
    if 'COLLECTION' not in wb.sheetnames:
        ws = wb.create_sheet('COLLECTION')
    else:
        ws = wb['COLLECTION']
    for i in range(2,10):
        for j in range(1,8):
            ws.cell(row=i, column=j).value = None
    for r, row in enumerate(final_data[1:], start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 2:
                cell.font = Font(name='Calibri', size=11, bold=False)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                col = c
                align_right = [1,5,6,7]
                align_left = [2,3,4]
                if col in align_right:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                if col in align_left:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.font = Font(name='Calibri', size=11, bold=False)
                if col == 7 and isinstance(val, (int,float)) and val < 0:
                    cell.font = Font(name='Calibri', size=11, color='FF0000', bold=False)
    for row in ws.iter_rows(min_row=2, max_row=9, min_col=1, max_col=7):
        for cell in row:
            cell.font = Font(name='Calibri', size=11, bold=False)
    wb.save(OUTPUT)

sheets = load_data()
final_data = prep_collection_table(sheets)
write_collection(final_data)
