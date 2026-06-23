import openpyxl

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_1/group_387-16/r0/evolve_387-16/input.xlsx')
ws = wb['Sheet1']

# Read Value/Binaries columns
value_col = [ws[f'A{i}'].value for i in range(3, 19)]
bin_col = [ws[f'B{i}'].value for i in range(3, 19)]
# Read Result values
result_vals = [ws[f'D{i}'].value for i in range(5, 19) if ws[f'D{i}'].value is not None]

# For each result value, remove one instance from value_col/bin_col
def remove_first(rows, v):
    for i, val in enumerate(rows):
        if val == v:
            return rows[:i] + rows[i+1:], i
    return rows, -1

values, binaries = value_col[:], bin_col[:]
for rv in result_vals:
    values, idx = remove_first(values, rv)
    if idx != -1:
        binaries = binaries[:idx] + binaries[idx+1:]

# Fill up with blanks to keep 16 rows
while len(values) < 16:
    values.append("")
    binaries.append("")

# Write back A3:B18
for i in range(16):
    ws[f'A{i+3}'] = values[i]
    ws[f'B{i+3}'] = binaries[i]

# Recalculate solver and difference
solver_result = sum(v * b for v, b in zip(values, binaries) if isinstance(v, (int, float)) and isinstance(b, (int, float)))
target = ws['E2'].value
ws['H2'] = solver_result
ws['J2'] = solver_result - target

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_1/group_387-16/r0/evolve_387-16/output.xlsx')
