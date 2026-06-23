import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import collections

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_2/regression_gate/after_fix/core_177-6/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_2/regression_gate/after_fix/core_177-6/output.xlsx'

wb = load_workbook(in_path)
ws = wb.active
headers = [c.value for c in ws[1][:18]]  # A-R
rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=18, values_only=True):
    if any(row):
        rows.append(list(row))
merged = collections.OrderedDict()
for row in rows:
    key = row[7]
    if key not in merged:
        merged[key] = [row]
    else:
        merged[key].append(row)
def to_num(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0
def merge_rows(group):
    out = list(group[0][:8])
    for col in range(8, 18):
        val = sum(to_num(r[col]) for r in group)
        out.append(val)
    return out
merged_rows = [merge_rows(group) for group in merged.values()]
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = 'combined'
ws_out.append(headers)
for row in merged_rows[:7]:
    ws_out.append(row)
#--- Fix: Copy styles as new objects ---#
for idx, cell in enumerate(ws_out[1], 1):
    scell = ws.cell(row=1, column=idx)
    if scell.font:
        cell.font = Font(name=scell.font.name, bold=scell.font.bold, italic=scell.font.italic, vertAlign=scell.font.vertAlign,
                         underline=scell.font.underline, strike=scell.font.strike, color=scell.font.color, sz=scell.font.sz)
    if scell.alignment:
        cell.alignment = Alignment(horizontal=scell.alignment.horizontal, vertical=scell.alignment.vertical)
    ws_out.row_dimensions[1].height = ws.row_dimensions[1].height
for row_idx, row in enumerate(ws_out.iter_rows(min_row=2, max_row=8, min_col=9, max_col=18), start=2):
    for col_idx, cell in enumerate(row, start=9):
        val = cell.value
        cell.number_format = '0.00'
        if isinstance(val, (int, float)) and abs(val) < 1e-8:
            cell.value = None
def autofit(ws, from_col, to_col):
    for col_idx in range(from_col, to_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
autofit(ws_out, 1, 18)
wb_out.save(out_path)
