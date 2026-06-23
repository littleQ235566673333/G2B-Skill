from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the workbook and worksheet
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun2/eval_56419_tc1/input.xlsx')
ws = wb.active

# Find column indices for 'Type' and 'Quantity'
header_row = 1
quantity_col = None
type_col = None
for col in range(1, ws.max_column+1):
    value = ws.cell(row=header_row, column=col).value
    if value is not None:
        if str(value).strip().lower() == 'quantity':
            quantity_col = col
        if str(value).strip().lower() == 'type':
            type_col = col

# Get list of Types with non-zero Quantity
non_zero_types = []
for row in range(2, ws.max_row+1):
    qty = ws.cell(row=row, column=quantity_col).value
    typ = ws.cell(row=row, column=type_col).value
    if qty is not None and typ is not None and qty != 0:
        non_zero_types.append(typ)

# Prepare for output
output_start_row = 2
output_end_row = 27
output_col = 8  # H
fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

for i in range(output_start_row, output_end_row + 1):
    cell = ws.cell(row=i, column=output_col)
    if (i - output_start_row) < len(non_zero_types):
        cell.value = non_zero_types[i - output_start_row]
    else:
        cell.value = "N/A"
    cell.fill = fill

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun2/eval_56419_tc1/output.xlsx')
