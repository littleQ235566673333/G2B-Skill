import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_3/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_3/regression_gate/after_pass/core_50526/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get the lookup value
lookup_value = ws['B6'].value

# Get column headers (assume row 1 is header, columns start from B)
col_labels = [ws.cell(row=1, column=col).value for col in range(2, ws.max_column+1)]

# Find the row where the first column matches the lookup value
found_row = None
for r in range(2, ws.max_row+1):
    if ws.cell(row=r, column=1).value == lookup_value:
        found_row = r
        break

results = []
if found_row:
    for idx, col in enumerate(range(2, ws.max_column+1)):
        v = ws.cell(row=found_row, column=col).value
        if isinstance(v, (int, float)) and v > 0:
            results.append(col_labels[idx])

# Write results to B9, B10 (etc)
for i in range(2):
    ws.cell(row=9+i, column=2).value = results[i] if i < len(results) else None

wb.save(output_path)
