import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/regression_gate/after_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/regression_gate/after_pass/core_9726/output.xlsx'

# Load all tables from all sheets
xls = pd.ExcelFile(input_path)
all_sheets = {name: xls.parse(name) for name in xls.sheet_names}

def find_table_by_cols(sheets, required_cols):
    for name, df in sheets.items():
        dfcols = [str(c).strip().lower() for c in df.columns]
        if all(rc.lower() in dfcols for rc in required_cols):
            return name, df
    return None, None

# Table 1: has a 'Visit date', by student
t1_name, t1 = find_table_by_cols(all_sheets, ['student', 'visit date'])
# Table 2: must have 'Submission date', 'student' columns
t2_name, t2 = find_table_by_cols(all_sheets, ['student', 'submission date'])

if t1 is None or t2 is None:
    raise Exception('Could not find tables by expected columns')

def excel_to_datetime(value):
    if pd.isna(value): return None
    if isinstance(value, datetime): return value
    try:
        return pd.to_datetime(value, dayfirst=True, errors='coerce')
    except Exception:
        return None

t1['Visit date'] = t1['Visit date'].apply(excel_to_datetime)
t2['Submission date'] = t2['Submission date'].apply(excel_to_datetime)

visit_map = {}
for student, group in t1.groupby('Student'):
    vsorted = sorted([d for d in group['Visit date'] if pd.notna(d)])
    visit_map[student] = vsorted

last_visits = []
for idx, row in t2.iterrows():
    student = row['Student']
    subm_date = row['Submission date']
    vdates = visit_map.get(student, [])
    prior_visits = [v for v in vdates if pd.notna(v) and v < subm_date]
    if prior_visits:
        last_visit = max(prior_visits)
        last_visits.append(last_visit)
    else:
        last_visits.append(None)

t2['Visit date'] = last_visits

wb = load_workbook(input_path)
ws = wb[t2_name]

# Find start row for Table 2 from headers
for row in ws.iter_rows(min_row=1, max_row=20):
    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in row]
    if 'student' in headers and 'submission date' in headers:
        start_row = row[0].row
        break
else:
    raise Exception('Could not find Table 2 header row')

header_row = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[start_row]]
if 'visit date' in header_row:
    visitdate_col = header_row.index('visit date') + 1
else:
    raise Exception('Could not find Visit date column in Table 2')

for i, value in enumerate(last_visits):
    cell = ws.cell(row=start_row + 1 + i, column=visitdate_col)
    if pd.isna(value) or value is None:
        cell.value = None
    else:
        cell.value = value.strftime('%d-%b-%y')

wb.save(output_path)
