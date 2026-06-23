import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_8/group_91-34/r3/evolve_91-34/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_8/group_91-34/r3/evolve_91-34/output.xlsx'
sheetname = 'SwiftMD'

# Read the data, headers on row 1

df = pd.read_excel(input_path, sheet_name=sheetname, header=1)

key_cols = ['Last Name', 'First Name', 'Date Of Birth']

if not all(k in df.columns for k in key_cols):
    print("Key columns not found. Found columns:", df.columns)
    raise ValueError("Key columns missing!")

# Function to determine indices to delete
indices_to_drop = []
for _, group in df.groupby(key_cols):
    # Only consider if all are marked 'Yes' for Duplicate
    if group[key_cols[0]].isna().all() or group[key_cols[1]].isna().all() or group[key_cols[2]].isna().all():
        continue
    dup_yes = group['Duplicate?'].fillna('').str.upper() == 'YES'
    if not dup_yes.all():
        continue
    rel_is_employee = group['Relationship'].fillna('').str.upper() == 'EMPLOYEE'
    # If none are 'Employee' and >1 member, delete one
    if not rel_is_employee.any() and len(group) > 1:
        indices_to_drop.append(group.index[0])

df_out = df.drop(index=indices_to_drop)

# Write back to Excel at B2:O42
wb = load_workbook(input_path)
ws = wb[sheetname]

# Clear B2:O42
for row in ws.iter_rows(min_row=2, max_row=42, min_col=2, max_col=15):
    for cell in row:
        cell.value = None

# Write output dataframe (max 41 rows) to B2:O...
for i, row in enumerate(df_out.itertuples(index=False), 2):
    if i > 42:
        break
    for j, value in enumerate(row, 2):
        ws.cell(row=i, column=j, value=value)

wb.save(output_path)
