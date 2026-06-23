"""
XLSX cell-level comparison helper for SpreadSheetBench evaluation.

Matches the official SpreadSheetBench evaluation methodology:
  - Recalculates formulas via soffice before comparison (open_spreadsheet.py)
  - Loads workbooks with data_only=True to read cached computed values
  - Uses SpreadSheetBench's exact value normalization (transform_value)
  - Handles datetime, float rounding, None/empty equivalence

Reference:
  - Evaluation: https://github.com/RUCKBReasoning/SpreadsheetBench/blob/main/evaluation/evaluation.py
  - Recalc: https://github.com/RUCKBReasoning/SpreadsheetBench/blob/main/evaluation/open_spreadsheet.py
"""

import datetime
import os
import shutil
import subprocess
import tempfile
from contextlib import contextmanager
from pathlib import Path

import openpyxl

try:
    import fcntl
except ImportError:  # pragma: no cover - Windows fallback
    fcntl = None


# ═══════════════════════════════════════════════════════════════════════════
# Formula recalculation (matches official open_spreadsheet.py)
# ═══════════════════════════════════════════════════════════════════════════

@contextmanager
def _file_lock(lock_path: str | Path | None):
    """Cross-process advisory lock used to serialize LibreOffice calls."""
    if not lock_path or fcntl is None:
        yield
        return

    lock_path = Path(lock_path)
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    with open(lock_path, "w") as lock_file:
        fcntl.flock(lock_file, fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(lock_file, fcntl.LOCK_UN)


def recalculate_xlsx(
    xlsx_path: str | Path,
    timeout: int = 120,
    lock_path: str | Path | None = None,
) -> bool:
    """Recalculate formulas in an xlsx file using LibreOffice headless.

    Matches the official SpreadSheetBench evaluation step
    (open_spreadsheet.py): the file is opened by soffice which
    evaluates all formulas, then saved back so openpyxl can read
    cached values.

    Uses an isolated user profile per call so multiple concurrent
    soffice instances don't conflict on the profile lock.

    Returns True if recalculation succeeded, False otherwise.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return False

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            # Isolated user profile prevents soffice lock conflicts
            # when multiple instances run concurrently
            profile_dir = Path(tmpdir) / "profile"
            profile_dir.mkdir()
            outdir = Path(tmpdir) / "out"
            outdir.mkdir()

            soffice_bin = os.environ.get("SE_PIPELINE_REAL_SOFFICE", "soffice")
            env_lock = os.environ.get("SE_PIPELINE_SOFFICE_LOCK")
            effective_lock = lock_path or env_lock

            with _file_lock(effective_lock):
                result = subprocess.run(
                    [
                        soffice_bin,
                        f"-env:UserInstallation=file://{profile_dir}",
                        "--headless",
                        "--calc",
                        "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                        "--outdir", str(outdir),
                        str(xlsx_path),
                    ],
                    timeout=timeout,
                    capture_output=True,
                )
            if result.returncode != 0:
                return False

            # Move recalculated file back
            recalc_file = outdir / xlsx_path.name
            if recalc_file.exists():
                shutil.move(str(recalc_file), str(xlsx_path))
                return True
            return False
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError):
        return False


def parse_answer_position(answer_position: str) -> list[tuple[str | None, str]]:
    """Parse answer_position into list of (sheet_name, cell_range) tuples.

    Based on the official SpreadsheetBench evaluation (evaluation.py,
    compare_workbooks) with one fix: commas inside single-quoted sheet
    names are preserved. The official code has a known bug where
    answer_position.split(',') breaks on sheet names like
    "'b2b, sez, de'!A5:V10". This fix handles that correctly.

    All other behavior matches the official implementation:
      - sheet_name stripped of surrounding single quotes
      - cell_range stripped of surrounding single quotes
      - split('!') to separate sheet from range
    """
    ranges = []
    # Split on commas NOT inside single quotes (fixes official bug)
    parts = []
    current = []
    in_quotes = False
    for ch in answer_position:
        if ch == "'":
            in_quotes = not in_quotes
            current.append(ch)
        elif ch == "," and not in_quotes:
            parts.append("".join(current))
            current = []
        else:
            current.append(ch)
    if current:
        parts.append("".join(current))

    for part in parts:
        part = part.strip()
        if "!" in part:
            sheet, cell_range = part.split("!", 1)
            sheet = sheet.lstrip("'").rstrip("'")
            cell_range = cell_range.lstrip("'").rstrip("'")
            ranges.append((sheet, cell_range))
        else:
            ranges.append((None, part))
    return ranges


# ═══════════════════════════════════════════════════════════════════════════
# SpreadSheetBench-compatible value normalization
# (from evaluation.py: transform_value + compare_cell_value)
# ═══════════════════════════════════════════════════════════════════════════

def _datetime_to_float(dt: datetime.datetime) -> float:
    """Convert datetime to Excel serial number (matches SpreadSheetBench)."""
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def _transform_value(v):
    """Normalize a cell value for comparison (matches SpreadSheetBench exactly).

    - int/float (incl. bool) → round to 2 decimals
    - datetime.time → "HH:MM" string
    - datetime.datetime → Excel serial number (rounded to 0 decimals)
    - str that looks numeric → round to 2 decimals as float

    NOTE: No bool special case — matches official evaluation.py exactly.
    Python's bool is a subclass of int, so True→1.0, False→0.0.
    """
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, datetime.time):
        return str(v)[:-3]
    if isinstance(v, datetime.datetime):
        return round(_datetime_to_float(v), 0)
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            pass
    return v


def _values_match(v1, v2) -> bool:
    """Compare two cell values using SpreadSheetBench's methodology.

    Values are first normalized via _transform_value(), then compared
    with type-aware equality and None/empty equivalence.
    """
    v1 = _transform_value(v1)
    v2 = _transform_value(v2)

    # None/empty equivalence
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True

    # Type-aware comparison (after normalization)
    if type(v1) != type(v2):
        return False

    return v1 == v2


# ═══════════════════════════════════════════════════════════════════════════
# Cell extraction
# ═══════════════════════════════════════════════════════════════════════════

def _format_cell_value(val) -> str:
    """Format a raw cell value as a human-readable string for display."""
    if val is None:
        return "<empty>"
    if isinstance(val, float):
        # Avoid ugly float repr (e.g., 1.0000000000000002)
        if val == int(val):
            return str(int(val))
        return f"{val:.6g}"
    return str(val)


def extract_cells(
    xlsx_path: str | Path,
    answer_position: str,
) -> dict[str, object]:
    """Extract cell values at answer_position from an XLSX file.

    Loads with data_only=True to read cached computed values (not formula
    strings), matching SpreadSheetBench's evaluation methodology.

    Returns dict mapping "SheetName!CellRef" to raw Python values
    (int, float, str, datetime, None, etc.).
    """
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    cells = {}

    for sheet_name, cell_range in parse_answer_position(answer_position):
        # Match official: default to first sheet name (not active sheet)
        target_name = sheet_name if sheet_name else wb.sheetnames[0]
        prefix = f"{sheet_name}!" if sheet_name else ""

        # Gracefully handle missing worksheets (matches official behavior:
        # cell_level_compare returns False when sheet_name not in wb_proc)
        if target_name not in wb.sheetnames:
            cells[f"{prefix}{cell_range}"] = None
            continue

        ws = wb[target_name]

        try:
            region = ws[cell_range]
        except Exception:
            cells[f"{prefix}{cell_range}"] = None
            continue

        # ws[range] returns a Cell for single cell, or tuple of tuples for range
        if not isinstance(region, tuple):
            region = ((region,),)
        elif region and not isinstance(region[0], tuple):
            region = (region,)

        for row in region:
            for cell in row:
                cells[f"{prefix}{cell.coordinate}"] = cell.value

    wb.close()
    return cells


def cells_to_text(cells: dict[str, object], max_cells: int = 50) -> str:
    """Format extracted cells as readable text. Truncates if too many."""
    items = list(cells.items())
    if len(items) > max_cells:
        shown = items[:max_cells]
        lines = [f"  {ref}: {_format_cell_value(val)}" for ref, val in shown]
        lines.append(f"  ... ({len(items) - max_cells} more cells)")
    else:
        lines = [f"  {ref}: {_format_cell_value(val)}" for ref, val in items]
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════
# Comparison functions
# ═══════════════════════════════════════════════════════════════════════════

def compare_xlsx(
    output_path: str | Path,
    answer_path: str | Path,
    answer_position: str,
) -> tuple[str, str]:
    """Compare output XLSX against answer XLSX at specified cells.

    Returns (actual_text, expected_text) for the UserSimulator evaluator.
    Both texts show computed values (not formula strings).
    """
    expected_cells = extract_cells(answer_path, answer_position)
    expected_text = cells_to_text(expected_cells)

    output_path = Path(output_path)
    if not output_path.exists():
        actual_text = "<agent did not produce an output file>"
        return actual_text, expected_text

    actual_cells = extract_cells(output_path, answer_position)
    actual_text = cells_to_text(actual_cells)

    return actual_text, expected_text


def compute_accuracy(
    output_path: str | Path,
    answer_path: str | Path,
    answer_position: str,
    recalculate: bool = True,
    recalc_timeout: int = 120,
    lock_path: str | Path | None = None,
) -> dict:
    """Compute cell-level accuracy between output and answer XLSX.

    Uses SpreadSheetBench's official evaluation methodology:
      - Recalculates formulas via soffice before comparison (if enabled)
      - Loads with data_only=True (cached computed values)
      - Normalizes values via _transform_value()
      - Type-aware comparison with None/empty equivalence

    Returns dict with match_count, total_count, accuracy, and mismatches.
    """
    output_path = Path(output_path)
    if not output_path.exists():
        expected = extract_cells(answer_path, answer_position)
        return {
            "match_count": 0,
            "total_count": len(expected),
            "accuracy": 0.0,
            "mismatches": {
                ref: {"expected": _format_cell_value(val), "actual": "<no file>"}
                for ref, val in expected.items()
            },
        }

    # Recalculate formulas before comparison (official SpreadSheetBench step)
    if recalculate:
        recalculate_xlsx(output_path, timeout=recalc_timeout, lock_path=lock_path)

    expected = extract_cells(answer_path, answer_position)
    actual = extract_cells(output_path, answer_position)

    matches = 0
    mismatches = {}
    for ref, exp_val in expected.items():
        act_val = actual.get(ref)
        if _values_match(act_val, exp_val):
            matches += 1
        else:
            mismatches[ref] = {
                "expected": _format_cell_value(exp_val),
                "actual": _format_cell_value(act_val),
            }

    total = len(expected)
    return {
        "match_count": matches,
        "total_count": total,
        "accuracy": matches / total if total > 0 else 1.0,
        "mismatches": mismatches,
    }


def compute_accuracy_on_copy(
    output_path: str | Path,
    answer_path: str | Path,
    answer_position: str,
    recalc_timeout: int = 120,
    lock_path: str | Path | None = None,
) -> dict:
    """Compute accuracy after recalculating a temporary copy of output.xlsx.

    The official SpreadsheetBench evaluation opens and saves output files to
    populate cached formula values. For live monitoring we want that same
    semantics without mutating executor artifacts, so this helper copies the
    workbook to a temporary directory, recalculates the copy, and compares it.
    """
    output_path = Path(output_path)
    if not output_path.exists():
        return compute_accuracy(
            output_path, answer_path, answer_position, recalculate=False,
        )

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_output = Path(tmpdir) / output_path.name
        shutil.copy2(output_path, tmp_output)
        return compute_accuracy(
            tmp_output,
            answer_path,
            answer_position,
            recalculate=True,
            recalc_timeout=recalc_timeout,
            lock_path=lock_path,
        )


def format_comparison(
    output_path: str | Path,
    answer_path: str | Path,
    answer_position: str,
    recalculate: bool = True,
    recalc_timeout: int = 120,
    lock_path: str | Path | None = None,
) -> str:
    """Produce human-readable comparison text for the UserSimulator evaluator.

    Returns text showing expected vs actual computed cell values with a summary.
    The evaluator sees real values (not formula strings), enabling accurate
    feedback on whether the agent's output is correct.
    """
    # Recalculate once, then extract and compare without re-recalculating
    if recalculate:
        output_p = Path(output_path)
        if output_p.exists():
            recalculate_xlsx(output_p, timeout=recalc_timeout, lock_path=lock_path)
    actual_text, expected_text = compare_xlsx(output_path, answer_path, answer_position)
    acc = compute_accuracy(output_path, answer_path, answer_position, recalculate=False)

    lines = [
        "## Expected Cell Values",
        expected_text,
        "",
        "## Actual Cell Values",
        actual_text,
        "",
        f"## Summary",
        f"{acc['match_count']}/{acc['total_count']} cells match "
        f"({acc['accuracy']:.1%} accuracy)",
    ]

    if acc["mismatches"]:
        mismatch_details = []
        for ref, vals in list(acc["mismatches"].items())[:20]:
            mismatch_details.append(
                f"  {ref}: expected \"{vals['expected']}\", got \"{vals['actual']}\""
            )
        lines.append("Mismatches:")
        lines.extend(mismatch_details)
        if len(acc["mismatches"]) > 20:
            lines.append(f"  ... ({len(acc['mismatches']) - 20} more)")

    return "\n".join(lines)
