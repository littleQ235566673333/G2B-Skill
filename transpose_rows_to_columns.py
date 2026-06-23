import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_42902_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_42902_tc1/output.xlsx'

# Load the workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read all values from the first column
list_values = []
for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
    value = row[0]
    if value is not None:
        list_values.append(value)

# Group every 3 rows as a record and write to D, E, F
row_counter = 1
for i in range(0, len(list_values), 3):
    group = list_values[i:i+3]
    # Each group should have exactly 3 rows, pad if necessary
    while len(group) < 3:
        group.append('')
    ws[f'D{row_counter}'] = group[0]
    ws[f'E{row_counter}'] = group[1]
    ws[f'F{row_counter}'] = group[2]
    row_counter += 1

wb.save(output_path)
