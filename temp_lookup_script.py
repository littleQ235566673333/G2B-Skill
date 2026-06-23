import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_1/regression_gate/before_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_1/regression_gate/before_pass/core_50526/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

lookup_cell = 'B6'
lookup_value = ws[lookup_cell].value

# Try to find lookup value in first column, then second column if not found.
row_idx = None
for i in range(2, ws.max_row + 1):
    # In case lookup values are in column 1 (A) or column 2 (B)
    if ws.cell(row=i, column=1).value == lookup_value or ws.cell(row=i, column=2).value == lookup_value:
        row_idx = i
        break
if row_idx is None:
    raise ValueError(f'Lookup value {lookup_value} not found in column A or B')

# Identify headers and matching >0 values
results = []
for col in range(2, ws.max_column + 1):  # Skip col=1 if that's the index/label column
    value = ws.cell(row=row_idx, column=col).value
    header = ws.cell(row=1, column=col).value
    if isinstance(value, (int, float)) and value > 0 and isinstance(header, str):
        results.append(header)

# Write results in B9, B10
for i in range(2):
    cell = ws.cell(row=9 + i, column=2)
    if i < len(results):
        cell.value = results[i]
    else:
        cell.value = None

wb.save(output_path)
