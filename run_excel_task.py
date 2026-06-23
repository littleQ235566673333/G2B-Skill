import openpyxl

# Input and output paths
input_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_333-29_tc1/input.xlsx"
output_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_333-29_tc1/output.xlsx"

def main():
    wb = openpyxl.load_workbook(input_path)
    sheet_a = wb["A"]
    sheet_b = wb["B"]

    # 1. Find first 'Yes' or 'NA' in col P (16) in Sheet 'A'
    target_row_p = None
    for row in range(2, sheet_a.max_row + 1):
        value = sheet_a.cell(row=row, column=16).value
        if str(value).strip().lower() in ["yes", "na"]:
            target_row_p = row
            break

    if target_row_p is not None:
        # 2. Search col L (12) above found row for first 100, get its row
        row_with_100 = None
        for row in range(target_row_p-1, 1, -1):
            val = sheet_a.cell(row=row, column=12).value
            if (val == 100):
                row_with_100 = row
                break
        # 3. Get date from col F (6) of found row, or fallback if not found
        if row_with_100 is not None:
            target_date = sheet_a.cell(row=row_with_100, column=6).value
        else:
            target_date = None
    else:
        target_date = sheet_a.cell(row=3, column=6).value

    # 4. Match F1 text (sheet A) to col C (3) in Sheet B, paste target_date to col F of found row
    # Otherwise, paste it into the same row for F4 in Sheet B
    text_f1 = sheet_a['F1'].value
    match_row_b = None
    for row in range(2, sheet_b.max_row + 1):
        if sheet_b.cell(row=row, column=3).value == text_f1:
            match_row_b = row
            break
    # Always paste identified result to B'!F4
    sheet_b['F4'].value = target_date

    wb.save(output_path)

if __name__ == "__main__":
    main()
