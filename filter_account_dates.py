import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42/eval_567-21_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42/eval_567-21_tc1/output.xlsx'

# Read into DataFrame (for efficiency)
df = pd.read_excel(input_path, sheet_name='Sheet1')

# Drop rows with no date in column E (index 4)
df = df[df.iloc[:, 4].notna()]

# Standardize date col (E, 0-based col 4) as string for max comparison (since 'YYYY/MM' is string comparable)
df['date_str'] = df.iloc[:, 4].astype(str)

groups = df.groupby([df.columns[0], df.columns[1]])
most_recent = groups['date_str'].transform('max')
# Keep only those with the most recent date in E, but allow duplicates for ties
df_keep = df[df['date_str'] == most_recent]

# Drop the helper column
df_keep = df_keep.drop(columns=['date_str'])

# Write back only columns A-G, up to 25 results (since output is 'Sheet1'!A3:G28)
wb = load_workbook(input_path)
ws = wb['Sheet1']

# Erase area A3:G28 first
for row in ws.iter_rows(min_row=3, max_row=28, min_col=1, max_col=7):
    for cell in row:
        cell.value = None

# Write header at row 3
for j, col in enumerate(df_keep.columns[:7], 1):
    ws.cell(row=3, column=j, value=col)

# Write up to 25 rows from row 4
df_out = df_keep.iloc[:25, :7]
for i, row in enumerate(df_out.itertuples(index=False), 4):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)

wb.save(output_path)
