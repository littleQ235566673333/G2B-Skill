import openpyxl

def check_status(column_a_list, column_b_list):
    # Priority order: Potato->Worst, Tomato->Ignore, Pickle->Bad
    for status, name in [('Worst', 'Potato'), ('Ignore', 'Tomato'), ('Bad', 'Pickle')]:
        for a, b in zip(column_a_list, column_b_list):
            if a == name and b is False:
                return status
    return 'Good'

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_42198_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_42198_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 8):  # 2 to 7 inclusive
    values_a = []
    values_b = []
    for r in range(2, row+1):
        values_a.append(ws[f'A{r}'].value)
        val_b = ws[f'B{r}'].value
        # Openpyxl may store boolean as Python bool or 0/1 or string
        if type(val_b) is bool:
            bool_val = val_b
        elif isinstance(val_b, (int, float)):
            bool_val = bool(val_b)
        elif isinstance(val_b, str):
            if val_b.upper() == 'FALSE':
                bool_val = False
            elif val_b.upper() == 'TRUE':
                bool_val = True
            else:
                bool_val = None
        else:
            bool_val = None
        values_b.append(bool_val)
    status = check_status(values_a, values_b)
    ws[f'C{row}'] = status

wb.save(output_path)
