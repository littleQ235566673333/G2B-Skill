import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/group_49667/r2/evolve_49667/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/group_49667/r2/evolve_49667/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

time_labels = [ws[f'{gcl(c)}1'].value for c in range(6, 47)]
def time_from_label(lbl):
    return datetime.datetime.strptime(lbl, '%H:%M').time() if isinstance(lbl, str) else lbl
times = [time_from_label(lbl) for lbl in time_labels]

def extract_meeting_blocks(row_vals):
    m_blocks = []
    in_block = False
    start_idx = None
    for ti, val in enumerate(row_vals):
        if val == 'm' and not in_block:
            in_block = True
            start_idx = ti
        if (val != 'm' and in_block) or (in_block and ti == len(row_vals)-1):
            in_block = False
            end_idx = ti if val != 'm' else ti+1
            m_blocks.append((start_idx, end_idx))
        if len(m_blocks) == 2:
            break
    return m_blocks

for r in range(2, 17):
    row_vals = [ws[f'{gcl(c)}{r}'].value for c in range(6,47)]
    meeting_blocks = extract_meeting_blocks(row_vals)
    for blocki in range(2):
        c_start, c_end = 2 + blocki*2, 3 + blocki*2
        if blocki < len(meeting_blocks):
            start, end = meeting_blocks[blocki]
            ws.cell(row=r, column=c_start).value = times[start]
            ws.cell(row=r, column=c_end).value = times[end]
        else:
            ws.cell(row=r, column=c_start).value = None
            ws.cell(row=r, column=c_end).value = None

wb.save(output_path)
