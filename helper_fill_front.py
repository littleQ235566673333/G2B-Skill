import openpyxl
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42/eval_9391_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42/eval_9391_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_front = wb['Front']
ws_data = wb['Data']

# Collect agent names from Front!A2:A12
agent_rows = list(range(2,13))
agent_letters = [ws_front[f'A{row}'].value for row in agent_rows]

# Dates from Front!B1, C1
date_cells = ['B1','C1']
front_dates = [ws_front[c].value for c in date_cells]

# Data sheet: gather all rows
rows = list(ws_data.iter_rows(values_only=True))

# Find all the row indices in Data where a date starts a block
from collections import OrderedDict
data_date_rows = OrderedDict()
for idx, row in enumerate(rows):
    if isinstance(row[0], (datetime.datetime, datetime.date)):
        d = row[0].date() if isinstance(row[0], datetime.datetime) else row[0]
        data_date_rows[d] = idx

def get_agent_f_value_for_date(agent_letter, date):
    if not date:
        return ''
    d = date.date() if isinstance(date, datetime.datetime) else date
    if d not in data_date_rows:
        return ''
    idx_start = data_date_rows[d]
    # First agent data is two rows after the date (skip date + header)
    first_agent = idx_start + 2
    # Build mapping: agent letter (A) => that row
    agent_map = {}
    for i in range(first_agent, len(rows)):
        r = rows[i]
        cell = r[0]
        if isinstance(cell, (datetime.datetime, datetime.date)) or cell is None:
            break
        agent_map[str(cell).strip()] = r
    r = agent_map.get(str(agent_letter).strip(), None)
    if r is None:
        return ''
    val = r[5] if len(r) > 5 else ''
    return val if val not in (None, 0) else ''

for col_idx, date in enumerate(front_dates, start=2): # B=2, C=3
    for row_offset, agent in enumerate(agent_letters):
        value = get_agent_f_value_for_date(agent, date)
        ws_front.cell(row=agent_rows[row_offset], column=col_idx).value = value

wb.save(output_path)
print('Wrote results to output.')
