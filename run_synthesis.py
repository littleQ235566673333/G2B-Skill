import openpyxl

# Load workbook and sheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_57989_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_57989_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Identify where drivers and days are
# Assume: first row is days, first column is drivers, data is in A2:U21
# Weekdays in header row (row 1, from B1 to H1)
# Drivers in column A (from A2 downward)

day_names = [ws.cell(row=1, column=col).value for col in range(2, 9)] # B1:H1

drivers = []
row = 2
while ws.cell(row=row, column=1).value:
    drivers.append(ws.cell(row=row, column=1).value)
    row += 1

driver_row_map = {driver: idx+2 for idx, driver in enumerate(drivers)}  # +2 due to header
weekday_col_map = {day: idx+2 for idx, day in enumerate(day_names)}     # +2 due to header

# Create a synthesis matrix (drivers x days)
synthesis = []
for driver in drivers:
    row_id = driver_row_map[driver]
    row_counts = []
    for day in day_names:
        col_id = weekday_col_map[day]
        # Count non-empty values for this driver and this weekday (all columns for this driver and all rows for this day)
        count = 0
        cell_value = ws.cell(row=row_id, column=col_id).value
        if cell_value not in [None, '']:
            count += 1
        row_counts.append(count)
    synthesis.append(row_counts)

# Write results to B25:H43
for i, driver_counts in enumerate(synthesis):
    for j, value in enumerate(driver_counts):
        ws.cell(row=25+i, column=2+j, value=value)

wb.save(output_path)
