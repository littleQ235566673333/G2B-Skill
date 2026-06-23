from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42/eval_120-24_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42/eval_120-24_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Column indices (1-based)
col_ay = column_index_from_string('AY')  # FAH Parent Desc (condition 3)
col_bf = column_index_from_string('BF')  # Expected Result in BG (condition 2)
col_bl = column_index_from_string('BL')  # COA_Child (condition 1)
col_bn = column_index_from_string('BN')  # Target output column (empty header)

# Constants
strings_bg = ['OCOGS - Spares - Transfer Price Overhead', 'OFSS - ORCL Consulting Prod Cost I/C']
codes_ay = ['BOA_033E', 'BOA_011G']
fill_color = PatternFill(fill_type='solid', fgColor='0070C0')

for row in range(2, 46):  # BN2 to BN45 requested
    val_bl = ws.cell(row=row, column=col_bl).value
    val_bf = ws.cell(row=row, column=col_bf).value
    val_ay = ws.cell(row=row, column=col_ay).value

    if val_bl == 'LAG' and val_bf in strings_bg:
        if val_ay in codes_ay:
            target_val = 'OCOGS - Spares - Transfer Price Overhead'
        else:
            target_val = 'OFSS - ORCL Consulting Prod Cost I/C'
        ws.cell(row=row, column=col_bn).value = target_val
        # Only fill BL when both BL=='LAG' and BN is written (i.e., this block executes)
        ws.cell(row=row, column=col_bl).fill = fill_color
    # Otherwise: do not touch BN or BL fill

wb.save(output_path)
print('Done.')
