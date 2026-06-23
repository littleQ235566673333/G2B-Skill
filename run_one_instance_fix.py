import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/group_387-16/r0/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/group_387-16/r0/evolve_387-16/output.xlsx'

# Read data

df = pd.read_excel(input_path, sheet_name='Sheet1')
# Get the correct column names
value_col = 'Value' if 'Value' in df.columns else df.columns[1]
bin_col = 'Binaries' if 'Binaries' in df.columns else df.columns[2]
result_col = 'Result values' if 'Result values' in df.columns else df.columns[3]

value_list = df[value_col].tolist()
bin_list = df[bin_col].tolist()
result_values = df[result_col].dropna().tolist()

# Remove only one instance in 'Value' and paired 'Binaries' for each result_value
for rv in result_values:
    if rv in value_list:
        idx = value_list.index(rv)
        value_list[idx] = None
        bin_list[idx] = None

# Compact (remove Nones, shifting all remaining values up)
filtered = [(v, b) for v, b in zip(value_list, bin_list) if v is not None]

# Now update the df, filling blanks after the compacted values
for i, (v, b) in enumerate(filtered):
    df.at[i, value_col] = v
    df.at[i, bin_col] = b
for j in range(len(filtered), len(df)):
    df.at[j, value_col] = None
    df.at[j, bin_col] = None

# Compute Solver result (sum of column A, ignoring NaN and non-numeric)
A = df.iloc[:,0]
A_numeric = pd.to_numeric(A, errors='coerce')
solver_result = A_numeric.dropna().sum()

# Compute difference vs target if present
if 'Target value' in df.columns:
    target_val = df['Target value'].iloc[0]
else:
    target_val = None
difference = solver_result - target_val if target_val is not None else None

# Write result & difference in extra columns for writing
# But only for output positioning. Actual outputs to col 5/6 in openpyxl section.
df['Solver result'] = None
df['Difference'] = None
if len(df) > 0:
    df.at[0, 'Solver result'] = solver_result
    df.at[0, 'Difference'] = difference

# Write back to output file (A2:D18: results, E2: solver, F2: diff)
wb = load_workbook(input_path)
ws = wb['Sheet1']

N = 17  # rows A2:A18
for i in range(N):
    for j in range(4):  # Columns A,B,C,D
        val = df.iloc[i, j] if i < len(df) else None
        ws.cell(row=2+i, column=1+j, value=val)
# Write solver result and difference in E2, F2
ws.cell(row=2, column=5, value=solver_result)
ws.cell(row=2, column=6, value=difference)
wb.save(output_path)
