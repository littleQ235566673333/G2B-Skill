import openpyxl
import re

def extract_number(cell_value):
    """Extract the first occurrence of a number (integer or float) from a cell value that may contain both numbers and text."""
    if cell_value is None:
        return None
    match = re.search(r"[\d,.]+", str(cell_value))
    if match:
        num_str = match.group(0).replace(",", "")
        try:
            value = float(num_str)
        except ValueError:
            value = None
        return value
    return None

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_49036_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_49036_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)
ws = wb["Dashboard"]

b5_val = ws["B5"].value
b6_val = ws["B6"].value

num_b5 = extract_number(b5_val)
num_b6 = extract_number(b6_val)

if num_b5 is not None and num_b6 is not None and num_b5 != 0:
    win_rate = num_b6 / num_b5
    pct_str = f"{win_rate*100:.2f}% WIN RATE"  # Two decimal percent
else:
    pct_str = "N/A WIN RATE"

ws["B8"] = pct_str

wb.save(output_path)
