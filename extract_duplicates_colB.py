import openpyxl
from collections import defaultdict, Counter

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_50971_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_50971_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read column A and B (assuming first row is header)
col_a = []
col_b = []
for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
    a, b = row
    col_a.append(a.value)
    col_b.append(b.value)

# Count occurrences in column A
counter_a = Counter(col_a)

duplicate_values = {value for value, count in counter_a.items() if count > 1}

# Collect column B values associated with duplicates in column A
result = []
for a, b in zip(col_a, col_b):
    if a in duplicate_values:
        result.append(b)

# Write the results to column G starting from G3
start_row = 3
for idx, value in enumerate(result[:11]):  # G3:G13 is 11 cells
    ws.cell(row=start_row + idx, column=7, value=value)

wb.save(output_path)
