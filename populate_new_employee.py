import openpyxl

# File paths
target_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_32093_tc1/output.xlsx"
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_32093_tc1/input.xlsx"

# Load workbook and select active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Process rows 2 to 15
for row in range(2, 16):
    current_emp = ws[f'B{row}'].value
    new_emp1 = ws[f'C{row}'].value
    new_emp2 = ws[f'D{row}'].value
    new_emp3 = ws[f'E{row}'].value

    # Find first non-empty new employee; otherwise use current employee
    result = next((emp for emp in [new_emp1, new_emp2, new_emp3] if emp not in [None, ""]), current_emp)
    ws[f'F{row}'] = result

# Save workbook
wb.save(target_path)
