from openpyxl import load_workbook

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_9/task_39432/r0/evolve_39432/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_9/task_39432/r0/evolve_39432/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet1']

header_row_idx, header_row = 1, [
    'item', 'beginning inventory', 'beginnng cost',
    'po-1', 'cost', 'po-2', 'cost', 'po-3', 'cost', 'po-4', 'cost',
    'current inv', 'current cost'
]
col_map = {h: idx+1 for idx, h in enumerate(header_row)}
def norm_col(name):
    for k in col_map:
        if name in k:
            return col_map[k]
    raise Exception(f"Col {name} not found")

num_rows = 4
results = []
for i in range(header_row_idx+1, header_row_idx+1+num_rows):
    row_vals = [ws.cell(row=i, column=j+1).value for j in range(len(header_row))]
    try:
        begin_inv = float(row_vals[norm_col('beginning inventory')-1])
    except Exception:
        begin_inv = 0
    try:
        begin_cost = float(row_vals[norm_col('beginnng cost')-1])
    except Exception:
        begin_cost = 0
    po_blocks = []
    for po_num in range(1, 5):
        qty_idx = norm_col(f'po-{po_num}') - 1
        cost_idx = qty_idx + 1
        try:
            qty = float(row_vals[qty_idx])
            cost = float(row_vals[cost_idx])
            po_blocks.append((qty, cost))
        except Exception:
            continue
    fifo_stream = []
    if begin_inv > 0 and begin_cost > 0:
        fifo_stream.append((begin_inv, begin_cost))
    fifo_stream.extend(po_blocks)
    try:
        current_inv = float(row_vals[norm_col('current inv')-1])
    except Exception:
        current_inv = 0
    needed = current_inv
    fifo_picks = []
    for qty, cost in fifo_stream:
        if needed <= 0:
            break
        take = min(qty, needed)
        fifo_picks.append((take, cost))
        needed -= take
    if not fifo_picks or sum(q for q,_ in fifo_picks) == 0:
        per_unit = ''
    elif len(fifo_picks) == 1:
        per_unit = fifo_picks[0][1]
    else:
        tc = sum(q*cost for q,cost in fifo_picks)
        tq = sum(q for q,_ in fifo_picks)
        per_unit = round(tc/tq,2) if tq else ''
    results.append(per_unit)
# Write results to column "current cost" (M)
per_unit_col = norm_col('current cost')
for idx, val in enumerate(results):
    ws.cell(row=header_row_idx+1+idx, column=per_unit_col, value=val)
wb.save(output_path)
