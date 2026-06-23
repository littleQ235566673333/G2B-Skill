import openpyxl
from collections import defaultdict

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_41978_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_41978_tc1/output.xlsx'

# Load workbook and select active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read years and status from specified ranges
years = [ws[f'G{row}'].value for row in range(14, 186)]
open_status = [ws[f'J{row}'].value for row in range(14, 186)]

# Count open cases per year (exclude None and blank years)
year_open_count = defaultdict(int)
for year, status in zip(years, open_status):
    if year is not None and str(year).strip() != '' and str(status).strip().lower() == 'open':
        year_open_count[year] += 1

# Sort years for output
years_sorted = sorted(year_open_count.keys())

# Output in cells I2 to I11
start_row = 2
for idx, year in enumerate(years_sorted[:10]):
    cell = f'I{start_row + idx}'
    ws[cell].value = year_open_count[year]

# Save to output path
wb.save(output_path)
