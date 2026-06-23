import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/regression_gate/after_pass/core_302-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/regression_gate/after_pass/core_302-1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
items_ws = wb['ITEMS']

# For each row, get the value from ITEMS column B
items_col_b = {}
for row in range(2, items_ws.max_row + 1):
    items_col_b[row] = items_ws.cell(row=row, column=2).value

# Apply to SHEET1 and SHEET2
def replace_with_items(sheet_name, start_row, end_row):
    ws = wb[sheet_name]
    for row in range(start_row, end_row + 1):
        item_b_value = items_col_b.get(row)
        ws_b_value = ws.cell(row=row, column=2).value
        if ws_b_value != item_b_value:
            ws.cell(row=row, column=2, value=item_b_value)

replace_with_items('SHEET1', 2, 8)
replace_with_items('SHEET2', 2, 7)

wb.save(output_path)
