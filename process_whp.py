from openpyxl import load_workbook

# Paths
INPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_7/regression_gate/after_fix/core_54474/input.xlsx'
OUTPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_7/regression_gate/after_fix/core_54474/output.xlsx'

wb = load_workbook(INPUT_PATH)
ws_whp = wb['WHP']
ws_data = wb['WHP DATA']   # Fixed capitalization

# Try detecting site names in E7:E8 of WHP tab
sites = []
for row in ws_whp.iter_rows(min_row=7, max_row=8, min_col=5, max_col=5):
    for cell in row:
        if cell.value:
            sites.append(cell.value)

# Read WHP data header and records
header_row = next(ws_data.iter_rows(min_row=1, max_row=1, values_only=True))
data_rows = list(ws_data.iter_rows(min_row=2, max_col=len(header_row), values_only=True))
data_dicts = [dict(zip(header_row, data)) for data in data_rows]

# Write info for each site into E7:G8
for idx, site in enumerate(sites):
    found = next((row for row in data_dicts if row[header_row[0]] == site), None)
    ws_whp.cell(row=7+idx, column=5, value=site)
    if found:
        for col in range(1, 3):  # F and G (columns 6, 7)
            ws_whp.cell(row=7+idx, column=5+col, value=found.get(header_row[col], ''))
    else:
        for col in range(1, 3):
            ws_whp.cell(row=7+idx, column=5+col, value='')

wb.save(OUTPUT_PATH)
