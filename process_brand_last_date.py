import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_8/regression_gate/after_fix/core_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_8/regression_gate/after_fix/core_516-46/output.xlsx'

sheet_name = 'ورقة1'
df = pd.read_excel(input_path, sheet_name=sheet_name)

# Focus on data from A2:E (skip header)
data = df.iloc[1:, :5]
data = data.dropna(how='all')  # remove fully empty rows

# Detect likely columns
col_lower = [str(c).lower() for c in data.columns]
date_col = [i for i,c in enumerate(col_lower) if 'date' in c or 'تاريخ' in c]
brand_col = [i for i,c in enumerate(col_lower) if 'brand' in c or 'علامة' in c or 'صنف' in c]
qty_col = [i for i,c in enumerate(col_lower) if 'qty' in c or 'كمية' in c or 'عدد' in c]
date_idx = date_col[0] if date_col else 0
brand_idx = brand_col[0] if brand_col else 1
qty_idx = qty_col[0] if qty_col else 4

# Rename to canonical
canonical = {data.columns[date_idx]: 'date', data.columns[brand_idx]: 'brand', data.columns[qty_idx]: 'qty'}
data = data.rename(columns=canonical)

# Clean dates, keep only valid rows
data['date'] = pd.to_datetime(data['date'], errors='coerce')
data = data.dropna(subset=['date','brand','qty'])

# Find LAST date per brand
last_dates = data.groupby('brand')['date'].transform('max')
latest = data[data['date'] == last_dates]

# Aggregate qty for duplicated entries on last date per brand
other_cols = [c for c in data.columns if c not in ['brand','date','qty']]
agg_dict = {k: 'first' for k in other_cols}
agg_dict.update({'qty': 'sum'})
agg = latest.groupby(['brand','date'], as_index=False).agg(agg_dict)

# Annotate modifications if there are duplicates
agg['modifications'] = ''
for idx, group in latest.groupby(['brand','date']):
    if len(group) > 1:
        agg.loc[(agg['brand']==idx[0]) & (agg['date']==idx[1]), 'modifications'] = 'Combined entries: ' + ', '.join(str(x) for x in group['qty'].values)

# Output columns for H2:L4: brand, date, ColC, ColD, qty, modifications
colC = other_cols[0] if len(other_cols)>0 else ''
colD = other_cols[1] if len(other_cols)>1 else ''
output = agg[['brand','date',colC,colD,'qty','modifications']].copy()
output.rename(columns={'brand':'Brand','date':'Date',colC:'ColC',colD:'ColD','qty':'Qty','modifications':'Modifications'}, inplace=True)
output = output.iloc[:3] # only put results in H2:L4 region

wb = load_workbook(input_path)
ws = wb[sheet_name]
header = ['Brand','Date','ColC','ColD','Qty','Modifications']
for j,col in enumerate(header,8):
    ws.cell(row=1, column=j, value=col)
for i, row in enumerate(output.itertuples(index=False),2):
    for j, v in enumerate(row,8):
        ws.cell(row=i, column=j, value=v)
wb.save(output_path)
