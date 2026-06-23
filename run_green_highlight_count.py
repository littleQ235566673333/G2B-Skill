import openpyxl

def is_green(cell):
    # Checks if cell is filled with green
    fill = cell.fill
    if fill.patternType is None:
        return False
    # Common green RGBs, include both pure and typical office shades
    greens = {'FF00FF00','FF92D050','FF00B050','FF9BBB59'}
    return fill.fgColor is not None and str(fill.fgColor.rgb).upper() in greens

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_43657_tc1/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_43657_tc1/output.xlsx'

# Load the workbook and the sheet
wb = openpyxl.load_workbook(input_fp)
ws = wb.active

# The area C3:G7 contains the names
name_range = ws['C3':'G7']
# The lookup names (L2:L8)
name_list = [ws[f'L{i}'].value for i in range(2,9)]

results = []
for target in name_list:
    count = 0
    for row in name_range:
        for cell in row:
            if cell.value and target:
                # Match name in cell and check green
                if (str(target).strip().lower() == str(cell.value).strip().lower()) and is_green(cell):
                    count += 1
    results.append(count)

# Write results to K2:K8
for idx, count in enumerate(results):
    ws[f'K{idx+2}'] = count

wb.save(output_fp)
