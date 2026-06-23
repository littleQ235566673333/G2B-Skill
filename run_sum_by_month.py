import openpyxl
from datetime import datetime
import calendar
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun1/eval_57590_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun1/eval_57590_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

A26 = ws['A26'].value

def month_from_string(s):
    for i, name in enumerate(calendar.month_name):
        if name and name.lower() in s.lower():
            return i
    return None

target_year = None
target_month = None
if isinstance(A26, str):
    match = re.search(r'(January|February|March|April|May|June|July|August|September|October|November|December)', A26, re.I)
    if match:
        target_month = month_from_string(match.group(0))
    year_match = re.search(r'(\d{4})', A26)
    if year_match:
        target_year = int(year_match.group(1))
else:
    try:
        target_year = A26.year
        target_month = A26.month
    except Exception:
        pass
if not target_month:
    raise ValueError('Could not identify target month from A26')

sum_val = 0
for row in ws.iter_rows(min_row=2, max_col=9, values_only=False):
    date_cell = row[2] # Column C
    val_cell = row[8] # Column I
    date_val = date_cell.value
    if not date_val:
        continue
    if isinstance(date_val, str):
        try:
            date_val = datetime.strptime(date_val[:10], '%Y-%m-%d')
        except Exception:
            continue
    if not isinstance(date_val, datetime):
        continue
    if date_val.month == target_month and (target_year is None or date_val.year == target_year):
        val = val_cell.value
        if isinstance(val, (int, float)):
            sum_val += val
ws['B26'] = sum_val
wb.save(output_path)
