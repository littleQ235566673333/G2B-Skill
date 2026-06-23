import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/regression_gate/before_fix/core_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/regression_gate/before_fix/core_57033/output.xlsx'

# Load both sheets with pandas
df_s4 = pd.read_excel(input_path, sheet_name='Sheet4')
df_md = pd.read_excel(input_path, sheet_name='CBtrans')

# Define the relevant columns (note case difference!)
md_company_col = 'company'
s4_company_col = 'Company'
account_col = 'account'
xchar_col = 'xchar'

start_row, end_row = 2, 7  # K2:K7 inclusive
match_results = []

for idx in range(start_row-2, end_row):    # 0-based index; K2 = row 2 = idx 0
    s_row = df_s4.iloc[idx]
    is_match = (
        (df_md[md_company_col] == s_row[s4_company_col]) &
        (df_md[account_col] == s_row[account_col]) &
        (df_md[xchar_col] == s_row[xchar_col])
    ).any()
    match_results.append('Match' if is_match else '-')

# Open workbook with openpyxl
wb = load_workbook(input_path)
ws4 = wb['Sheet4']

# Prepare cell shading
fill = PatternFill("solid", fgColor="FF66CC")
for i, value in enumerate(match_results, start=start_row):
    cell = ws4[f'K{i}']
    cell.value = value.capitalize()  # regular (Sentence) casing
    cell.fill = fill

wb.save(output_path)
