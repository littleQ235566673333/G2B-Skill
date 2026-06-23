import openpyxl

# Load the workbook and select active sheet
wb = openpyxl.load_workbook('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_1563_tc1/input.xlsx')
ws = wb.active

prev_val = None
for row in range(2, 31):
    val = ws[f'A{row}'].value
    if val is not None and val != '':
        prev_val = val
    ws[f'B{row}'].value = prev_val

wb.save('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_1563_tc1/output.xlsx')
