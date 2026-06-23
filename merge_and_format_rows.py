import openpyxl
from copy import copy
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/before_fix/core_177-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/before_fix/core_177-6/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_in = wb['DATA']
ws_out = wb['combined']

# Map header names to indices
headers = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
field_to_col = {h: idx for idx, h in enumerate(headers)}

col_h_idx = field_to_col['ComboKey']  # merge key column
agg_start, agg_end = 8, 17  # columns I:R are index 8-17 (0-based), inclusive

# Gather all source rows
src_rows = list(ws_in.iter_rows(min_row=2, max_row=ws_in.max_row, values_only=False))
grouped = defaultdict(list)

# Group rows by reference column
for row in src_rows:
    val = row[col_h_idx].value
    # Only consider non-empty key (skip empty 'ComboKey')
    if val is not None and str(val).strip():
        grouped[val].append(row)

# Write header row, also copy formatting from input
for col, cell in enumerate(ws_in[1], 1):
    tgt = ws_out.cell(row=1, column=col, value=cell.value)
    tgt.font = copy(cell.font)
    tgt.fill = copy(cell.fill)
    tgt.border = copy(cell.border)
    tgt.alignment = copy(cell.alignment)
    tgt.number_format = copy(cell.number_format)

# For each group, construct merged output
for out_row_idx, (the_key, group_rows) in enumerate(grouped.items(), 2):
    # Copy non-agg columns (A:H, 0-based 0-7)
    for col in range(0, agg_start):
        src_cell = group_rows[0][col]
        tgt_cell = ws_out.cell(row=out_row_idx, column=col+1, value=src_cell.value)
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.alignment = copy(src_cell.alignment)
        tgt_cell.number_format = copy(src_cell.number_format)
    # Merge Agg: I:R (col 8-17), sum values
    for col in range(agg_start, agg_end+1):
        agg_vals = []
        for row in group_rows:
            v = row[col].value
            # Accept blank (None, '', '\xa0') as skip/0
            if isinstance(v, (int, float)):
                if v != 0 and v is not None:
                    agg_vals.append(float(v))
            elif isinstance(v, str) and v.strip() and v.strip() not in ['\xa0', '']:
                try:
                    fv = float(v.replace(',', ''))
                    if fv != 0:
                        agg_vals.append(fv)
                except:
                    continue
        total = sum(agg_vals) if agg_vals else 0
        # Formatting: 2 decimals, no commas, blank if zero
        out_val = '' if total == 0 else f"{total:.2f}"
        tgt_cell = ws_out.cell(row=out_row_idx, column=col+1, value=out_val)
        # Copy formatting from first in-group
        src_cell = group_rows[0][col]
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.alignment = copy(src_cell.alignment)
        tgt_cell.number_format = '0.00'  # enforce for these cols

wb.save(output_path)
