import openpyxl

# Paths
input_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_42198_tc1/input.xlsx"
output_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_42198_tc1/output.xlsx"

# Load workbook and select active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Function to check condition in cumulative range
# Priority: Potato/FALSE -> "Worst", Tomato/FALSE -> "Ignore", Pickle/FALSE -> "Bad", else "Good"
def cumulative_eval(to_row):
    # Python index is 0-based, Excel rows start from 1
    for i in range(2, to_row+1):
        val_a = ws[f'A{i}'].value
        val_b = ws[f'B{i}'].value
        if val_a == "Potato" and val_b is False:
            return "Worst"
    for i in range(2, to_row+1):
        val_a = ws[f'A{i}'].value
        val_b = ws[f'B{i}'].value
        if val_a == "Tomato" and val_b is False:
            return "Ignore"
    for i in range(2, to_row+1):
        val_a = ws[f'A{i}'].value
        val_b = ws[f'B{i}'].value
        if val_a == "Pickle" and val_b is False:
            return "Bad"
    return "Good"

# Apply for rows 2 to 7 (inclusive)
for row in range(2, 8):
    result = cumulative_eval(row)
    ws[f'C{row}'] = result

# Save the workbook
wb.save(output_path)
