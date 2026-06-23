from openpyxl import load_workbook
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_12864_tc1/input.xlsx')
ws = wb['Sheet2']
for row in range(2, 13):
    formula = '=INDEX(Sheet1!B:B, MATCH(A{0}, Sheet1!D:D, 0)) & ", " & INDEX(Sheet1!C:C, MATCH(A{0}, Sheet1!D:D, 0))'.format(row)
    ws[f'B{row}'] = formula
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_12864_tc1/output.xlsx')
