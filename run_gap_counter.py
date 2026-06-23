import openpyxl
import sys

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/regression_gate/after_fix/core_50521/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/regression_gate/after_fix/core_50521/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

def count_blanks_between_first_and_second_numeric(row_data):
    # row_data includes: row[1] ... row[12] (months columns, 12 months)
    num_indices = [i for i, val in enumerate(row_data) if isinstance(val, (int, float)) and val is not None]
    if not num_indices:
        return None # No numeric found
    first_idx = num_indices[0]
    if row_data[first_idx] > 1:
        return 1
    if len(num_indices) < 2:
        return None # Only one or zero numerical
    second_idx = num_indices[1]
    # Count blanks between (exclusive of start, inclusive of end)
    return (second_idx - first_idx)

# The relevant data rows are 4, 5, 6 (1-based)
for excel_row, output_cell in zip([4,5,6], ['N4','N5','N6']):
    row = list(ws.iter_rows(min_row=excel_row, max_row=excel_row, values_only=True))[0]
    # row[1:13] corresponds to months (B~M)
    result = count_blanks_between_first_and_second_numeric(row[1:13])
    ws[output_cell] = result

wb.save(output_path)
