from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_5/regression_gate/after_fix/core_50916/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_5/regression_gate/after_fix/core_50916/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Read 7-day cycle mapping: A2:A8 are cycle days, C2:H8 are classes for each period
table_start_row = 2
table_end_row = 8
class_start_col = 3 # 'C'
class_end_col = 8 # 'H'

# Build mapping from cycle number to list of classes
cycle_to_classes = {}
for r in range(table_start_row, table_end_row + 1):
    cycle_day = ws.cell(row=r, column=1).value  # Column 1 = 'A'
    periods = [ws.cell(row=r, column=c).value for c in range(class_start_col, class_end_col + 1)]
    cycle_to_classes[cycle_day] = periods

# User's cycle day numbers are in A12:A14
input_cycle_start_row = 12
input_cycle_end_row = 14
input_cycle_col = 1 # 'A'
output_start_col = 3 # 'C'
output_end_col = 8 # 'H'

for row in range(input_cycle_start_row, input_cycle_end_row + 1):
    cycle_day_num = ws.cell(row=row, column=input_cycle_col).value
    classes = cycle_to_classes.get(cycle_day_num, [None]*(output_end_col - output_start_col + 1))
    for col in range(output_start_col, output_end_col + 1):
        val = classes[col - output_start_col] if classes else None
        ws.cell(row=row, column=col).value = val

wb.save(output_path)
