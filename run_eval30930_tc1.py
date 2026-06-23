import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_30930_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_30930_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read values in columns A and B (rows 2 to 66, zero-based: 0-64)
a_vals = [ws[f'A{row}'].value for row in range(2, 67)]
b_vals = [ws[f'B{row}'].value for row in range(2, 67)]

# Prepare output for column C (same range)
results = [''] * 65

start = 0
while start < 65:
    try:
        idx = b_vals.index(1, start)
    except ValueError:
        break
    next_start = idx + 1
    try:
        next_idx = b_vals.index(1, next_start)
    except ValueError:
        next_idx = 65
    # Count values > 0 in a_vals[idx+1:next_idx]
    count = sum(1 for x in a_vals[idx+1:next_idx] if x is not None and isinstance(x, (int, float)) and x > 0)
    results[idx] = count
    start = next_idx

# Write results to column C (rows 2-66)
for i, val in enumerate(results):
    ws.cell(row=i+2, column=3, value=val if val != '' else None)

wb.save(output_path)
