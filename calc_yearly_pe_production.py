import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_2/group_47766/r2/evolve_47766/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_2/group_47766/r2/evolve_47766/output.xlsx'

wb = load_workbook(input_file)
ws = wb.active

# Read all rows via pandas (since we only need ranges, can concat as before)
df = pd.read_excel(input_file)

data_rows = pd.concat([
    df.iloc[7:37],
    df.iloc[40:58],
    df.iloc[61:74]
])

def excel_serial_to_datetime(serial):
    try:
        # Excel's epoch is 1899-12-30
        return datetime(1899, 12, 30) + timedelta(days=float(serial))
    except Exception:
        return pd.NaT

# Date filter
year = 2023
start_date = datetime(year, 1, 1)
end_date = datetime(year, 12, 31)

def is_date_in_range(val):
    dt = excel_serial_to_datetime(val)
    if pd.isna(dt):
        return False
    return start_date <= dt <= end_date

pe_rows = data_rows[data_rows['Unnamed: 7'] == 'PE']
pe_rows = pe_rows[pe_rows['Unnamed: 5'].apply(is_date_in_range)]
total = pe_rows['Unnamed: 2'].apply(pd.to_numeric, errors='coerce').sum()

ws.cell(row=40, column=11).value = total
wb.save(output_file)
print('Written yearly PE total to K40:', total)
