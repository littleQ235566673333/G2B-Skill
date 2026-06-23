import pandas as pd
from openpyxl import load_workbook

def process_excel(input_path, output_path):
    # Read input data, no headers
    df = pd.read_excel(input_path, sheet_name='Sheet1', header=None)
    # Remove first two rows if they are empty or not part of data
    df = df.iloc[2:].reset_index(drop=True)
    # Drop rows with missing column 4 (date in E)
    df = df.dropna(subset=[4])
    # Parse dates in col 4 ('yyyy/mm') - coerce errors
    df['E_date'] = pd.to_datetime(df[4], format='%Y/%m', errors='coerce')
    df = df.dropna(subset=['E_date'])
    # Group by (col 0, col 1), keep all rows with the latest date
    def keep_latest(group):
        mx = group['E_date'].max()
        return group[group['E_date'] == mx]
    df_out = df.groupby([0, 1], group_keys=False).apply(keep_latest)
    df_out = df_out.drop(columns=['E_date'])

    # Write results into output region 'Sheet1'!A3:G28
    wb = load_workbook(input_path)
    ws = wb['Sheet1']
    # Clear output area first
    for r in range(3, 29):
        for c in range(1, 8):
            ws.cell(row=r, column=c).value = None
    # Write rows (max 26)
    for i, row in enumerate(df_out.iloc[:26].itertuples(index=False), start=3):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j).value = v
    wb.save(output_path)

process_excel(
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r1/eval_567-21_tc1/input.xlsx',
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r1/eval_567-21_tc1/output.xlsx'
)
