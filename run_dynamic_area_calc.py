import openpyxl
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_6/group_263-1/r1/evolve_263-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_6/group_263-1/r1/evolve_263-1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']
from collections import defaultdict
areas = defaultdict(float)
for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
    mtrl, w, h = row
    if mtrl is not None and isinstance(w, (int, float)) and isinstance(h, (int, float)):
        areas[mtrl] += w * h
# These are the materials shown as example in the sheet, order matters for answer positions
material_order = ['glass', 'metal', 'PVC']
for idx, mtrl in enumerate(material_order, start=2):
    area_val = areas.get(mtrl, None)
    cell = ws.cell(row=idx, column=8)  # H column
    cell.value = area_val if area_val is not None else None
wb.save(output_path)
