from openpyxl import load_workbook
from datetime import datetime

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_8/regression_gate/after_pass/core_41589/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_8/regression_gate/after_pass/core_41589/output.xlsx'
wb = load_workbook(in_path)
ws = wb['Contact List']
today = datetime.today()

def action_logic(last_contact, yesno):
    if not last_contact or not yesno or yesno.strip().upper() != 'YES':
        return 'NO ACTION'
    if isinstance(last_contact, datetime):
        delta = (today - last_contact).days
    else:
        try:
            dt = datetime.strptime(str(last_contact), '%Y-%m-%d')
            delta = (today - dt).days
        except Exception:
            return 'NO ACTION'
    if delta <= 30:
        return 'HOLD'
    if delta > 30:
        return 'TOUCH BASE'
    return 'NO ACTION'

for row in range(4, ws.max_row + 1):
    h_val = ws[f'H{row}'].value
    i_val = ws[f'I{row}'].value
    ws[f'J{row}'] = action_logic(h_val, i_val)

wb.save(out_path)
