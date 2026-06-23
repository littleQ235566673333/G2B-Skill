from openpyxl import load_workbook

# Input and output paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_3/regression_gate/after_fix/core_54274/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_3/regression_gate/after_fix/core_54274/output.xlsx"

# Open workbook and access first sheet
wb = load_workbook(input_path)
ws = wb.active

# Columns mapping
col = {'Target': 'B', 'Stretch': 'C', 'Over Stretch': 'D', 'Total': 'F', 'Points': 'G'}

target_cell = f"{col['Target']}4"
stretch_cell = f"{col['Stretch']}4"
overstretch_cell = f"{col['Over Stretch']}4"
total_cell = f"{col['Total']}4"
points_cell = f"{col['Points']}4"

# Build the nested IF formula with caps
# Formula logic:
# - If total <= target: total/target
# - If total > target and < stretch: MIN(total/target, 1.09)
# - If total = stretch: MIN(total/target, 1.1)
# - If total > stretch and < overstretch: MIN(total/target, 1.19)
# - If total >= overstretch: MIN(total/target, 1.2)

formula = (
    f"=IF({total_cell}<={target_cell},"
    f"{total_cell}/{target_cell},"
    f"IF({total_cell}<{stretch_cell},"
        f"MIN({total_cell}/{target_cell},1.09),"
        f"IF({total_cell}={stretch_cell},"
            f"MIN({total_cell}/{target_cell},1.1),"
            f"IF({total_cell}<{overstretch_cell},"
                f"MIN({total_cell}/{target_cell},1.19),"
                f"MIN({total_cell}/{target_cell},1.2)"
            f")"
        f")"
    f")"
")"
)

# Write the formula to points cell
ds = ws[points_cell]
ds.value = formula

# Save the output
wb.save(output_path)
