import pandas as pd
from openpyxl import load_workbook
from datetime import time

def extract_time(dt):
    if pd.isnull(dt): return None
    if hasattr(dt, 'time'):
        return dt.time()
    try:
        return pd.to_datetime(dt).time()
    except Exception:
        return None

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_4/group_52305/r3/evolve_52305/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_4/group_52305/r3/evolve_52305/output.xlsx'

# Load all sheets and find main data sheet
xls = pd.ExcelFile(input_path)
sheet_names = xls.sheet_names
data_df = None
data_sheet = None
for s in sheet_names:
    df = pd.read_excel(input_path, sheet_name=s)
    if len(df.columns) > 4:
        data_df = df
        data_sheet = s
        break
if data_df is None:
    raise Exception('No suitable data sheet found')

wb = load_workbook(input_path)
ws = wb[data_sheet]

start_row, end_row, start_col, end_col = 6, 24, 10, 14
rows = list(range(start_row, end_row+1))
cols = list(range(start_col, end_col+1))
# Get row labels (names) from column I (9), rows 6 to 24
name_labels = [ws.cell(row=r, column=start_col-1).value for r in rows]
# Get time ranges from row 5 (top of grid), columns J-N (10-14)
time_labels = [ws.cell(row=start_row-1, column=c).value for c in cols]
# If headers are not present, just skip (do not fill output)
if not any(name_labels) or not any(time_labels):
    wb.save(output_path)
    exit()

# Detect columns
possible_time_cols = [c for c in data_df.columns if 'time' in str(c).lower()]
name_col = data_df.columns[0]
time_col = possible_time_cols[0] if possible_time_cols else data_df.columns[1]

for i, name in enumerate(name_labels):
    if name is None or name == '':
        continue
    subset = data_df[data_df[name_col] == name]
    for j, t_lbl in enumerate(time_labels):
        if not t_lbl or not isinstance(t_lbl, str) or '-' not in t_lbl:
            continue
        t1, t2 = t_lbl.split('-')
        try:
            lower = pd.to_datetime(t1.strip()).time()
            upper = pd.to_datetime(t2.strip()).time()
        except Exception:
            continue
        count = subset[subset[time_col].apply(lambda x: lower <= extract_time(x) <= upper if extract_time(x) else False)].shape[0]
        ws.cell(row=start_row + i, column=start_col + j).value = count

wb.save(output_path)
