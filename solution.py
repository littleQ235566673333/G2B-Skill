import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_5/regression_gate/before_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_5/regression_gate/before_pass/core_50526/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Get lookup value from B6
lookup_value = ws['B6'].value

# Find the row matching the lookup value
lookup_row = None
for row in range(2, ws.max_row + 1):  # Start from 2 to skip header
    if ws.cell(row=row, column=1).value == lookup_value:
        lookup_row = row
        break

labels = []
if lookup_row is not None:
    # Iterate across columns from column 2 onwards (labels)
    for col in range(2, ws.max_column + 1):
        cell_value = ws.cell(row=lookup_row, column=col).value
        if isinstance(cell_value, (int, float)) and cell_value > 0:
            labels.append(ws.cell(row=1, column=col).value)

# Write the results to B9, B10, blank if not enough
for i in range(2):
    ws.cell(row=9 + i, column=2).value = labels[i] if i < len(labels) else ''

wb.save(output_path)
