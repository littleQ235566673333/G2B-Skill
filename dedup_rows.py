from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/group_227-40/r1/evolve_227-40/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/group_227-40/r1/evolve_227-40/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Read the source data from A2:E5
source_rows = list(ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=5, values_only=True))

def dedup_row(row):
    seen = set()
    out = []
    for v in row:
        key = str(v).strip() if v is not None else None
        if key in seen:
            out.append("")
        else:
            out.append(v if v is not None else "")
            seen.add(key)
    while len(out) < 5:
        out.append("")
    return out[:5]

results = [dedup_row(list(row)) for row in source_rows]

# Write the results to Sheet1!A14:E17
start_row = 14
for i, row in enumerate(results):
    for j, value in enumerate(row):
        ws.cell(row=start_row + i, column=1 + j, value=value)

wb.save(output_path)
