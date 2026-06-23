import openpyxl

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun2/eval_42198_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun2/eval_42198_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_file)
ws = wb.active

def get_value(a_list, b_list):
    # Priority: Potato+FALSE: Worst; Tomato+FALSE: Ignore; Pickle+FALSE: Bad; else Good
    for veg, out in [('Potato', 'Worst'), ('Tomato', 'Ignore'), ('Pickle', 'Bad')]:
        for a, b in zip(a_list, b_list):
            if a == veg and b is False:
                return out
    return 'Good'

for i in range(2, 8):  # C2:C7
    # Cumulative range: rows 2 to i
    a_slice = [ws[f'A{r}'].value for r in range(2, i+1)]
    b_slice = [ws[f'B{r}'].value for r in range(2, i+1)]
    result = get_value(a_slice, b_slice)
    ws[f'C{i}'] = result

wb.save(output_file)
