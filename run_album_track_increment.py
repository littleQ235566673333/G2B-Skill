from openpyxl import load_workbook
import string

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun1/eval_16511_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun1/eval_16511_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Inspect headers to find Album Name column
headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column+1)]

try:
    album_col_idx = headers.index('Album Name') + 1 # Excel columns are 1-based
except ValueError:
    raise Exception('No Album Name column found')

album_col_letter = string.ascii_uppercase[album_col_idx-1]
output_col = 6  # F
start_row = 2
end_row = 10

for row in range(start_row, end_row+1):
    formula = f'=COUNTIF(${album_col_letter}${start_row}:${album_col_letter}{row},{album_col_letter}{row})'
    ws.cell(row=row, column=output_col).value = formula

wb.save(output_path)
