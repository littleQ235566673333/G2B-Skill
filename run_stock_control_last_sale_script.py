import openpyxl
from datetime import datetime

# Load the workbook and the relevant sheet
infile = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_9448_tc1/input.xlsx'
outfile = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_9448_tc1/output.xlsx'
wb = openpyxl.load_workbook(infile)
ws = wb['Data']

first_data_row = 9
last_data_row = 18
start_col = 9   # I
end_col = 20    # T
target_col = 21 # U
header_row = 7

def parse_month_year(header):
    # Try to parse header as different possible month formats
    if isinstance(header, datetime):
        return datetime(header.year, header.month, 1)
    if not isinstance(header, str):
        return None
    
    header = header.strip()
    short_months = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 
                    'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
    long_months = {'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6, 'July': 7,
                   'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12}
    # Try regex parse
    import re
    # Patterns for Mar 2024, March 2024, 03/2024, 2024-03, etc.
    month_year = None
    
    # Try common formats
    for fmt in ('%B-%Y', '%b-%Y', '%d-%m-%Y', '%Y-%m', '%b %Y', '%B %Y', '%d/%m/%Y', '%m/%Y', '%B, %Y', '%b, %Y'):
        try:
            d = datetime.strptime(header, fmt)
            return datetime(d.year, d.month, 1)
        except Exception:
            continue
    
    # Try for "Mar 2024" or "March 2024"
    match = re.match(r'(\w+)[ ,/-]+(\d{4})', header)
    if match:
        month_part = match.group(1)
        year_part = match.group(2)
        if month_part in short_months:
            return datetime(int(year_part), short_months[month_part], 1)
        if month_part in long_months:
            return datetime(int(year_part), long_months[month_part], 1)
    # Try for numeric month/year (e.g. 3/2024)
    match = re.match(r'([01]?[0-9])[ /-](\d{4})', header)
    if match:
        return datetime(int(match.group(2)), int(match.group(1)), 1)
    return None

for row in range(first_data_row, last_data_row + 1):
    last_col = None
    for col in range(end_col, start_col - 1, -1): # From T to I
        cell = ws.cell(row=row, column=col)
        if cell.value not in (None, '', 0):
            last_col = col
            break
    output_cell = ws.cell(row=row, column=target_col)
    if last_col:
        header = ws.cell(row=header_row, column=last_col).value
        dt = parse_month_year(header)
        if dt:
            output_cell.value = dt
            output_cell.number_format = 'dd/mm/yyyy'
        else:
            output_cell.value = header # fallback: raw header text
    else:
        output_cell.value = None

wb.save(outfile)
