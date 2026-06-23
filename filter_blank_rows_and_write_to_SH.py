from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_3/regression_gate/after_pass/core_160-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_3/regression_gate/after_pass/core_160-6/output.xlsx'

wb = load_workbook(input_path)
ws = wb['SH']

header = [cell.value for cell in ws[1]][:12]

data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=12, values_only=True):
    if any(cell is not None and str(cell).strip() != '' for cell in row):
        data.append(list(row[:12]))

while len(data) < 6:
    data.append([None]*12)
# Only keep the first 6 data rows if more
out_data = [header] + data[:6]

for i, row in enumerate(out_data, start=6):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j).value = val

wb.save(output_path)
