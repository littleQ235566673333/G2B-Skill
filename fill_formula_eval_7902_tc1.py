import openpyxl
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42/eval_7902_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42/eval_7902_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Formula Required']

for row in range(3, 7):  # rows 3-6
    for col in range(4, 11):  # D=4 to J=10
        col_letter = get_column_letter(col)
        prev_col_letter = get_column_letter(col-1)
        main_ref = f'{col_letter}{row}'
        comp_ref = f'{prev_col_letter}{row}'
        # Formula logic: If cell value equals the previous cell in the same row, return itself.
        # Otherwise, if cell empty, return empty. Else VLOOKUP.
        formula = f'=IF({main_ref}={comp_ref},{main_ref},IF({main_ref}="","",VLOOKUP({main_ref},Grouping!$A$2:$I$7,3,FALSE)))'
        ws[f'{col_letter}{row}'] = formula

wb.save(output_path)
