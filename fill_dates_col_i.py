import openpyxl
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_3/regression_gate/before_fix/core_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_3/regression_gate/before_fix/core_45896/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Volym P5_P6_2023']
zord = wb['ZORD'] if 'ZORD' in wb.sheetnames else None

def get_dates_for_a(a_value, zord_ws):
    if not zord_ws:
        return []
    dates = []
    for row in zord_ws.iter_rows(min_row=2, values_only=True):
        if row[0] == a_value:
            val = row[2]  # ZORD!C:C
            if isinstance(val, (datetime.datetime, datetime.date)):
                dates.append(val)
    return sorted(dates)

for row in range(2, 11):  # I2:I10
    a_val = ws[f'A{row}'].value
    date_objs = get_dates_for_a(a_val, zord)
    if date_objs:
        formatted = [d.strftime('%d/%m/%Y') for d in date_objs]
        ws[f'I{row}'] = ','.join(formatted)
    else:
        ws[f'I{row}'] = ''

wb.save(output_path)
