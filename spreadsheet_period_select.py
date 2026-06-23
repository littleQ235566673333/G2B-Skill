import openpyxl

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_59224_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_59224_tc1/output.xlsx'

# Load workbook and select active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read the Project Start Date from B2
project_start = ws['B2'].value
# Target rows for the date ranges and output
start_row = 4
end_row = 14
rows = list(range(start_row, end_row + 1))

# Build lists of start and end dates
C_dates = [ws[f'C{row}'].value for row in rows]
D_dates = [ws[f'D{row}'].value for row in rows]

select_idx = None
for i, (cdate, ddate) in enumerate(zip(C_dates, D_dates)):
    if cdate is None or ddate is None:
        continue
    if cdate < project_start < ddate:
        select_idx = i
        break

for i, row in enumerate(rows):
    cell = ws[f'E{row}']
    if select_idx is not None and i <= select_idx:
        cell.value = 'Select Period'
    else:
        cell.value = 'Do Not Select'

# Save to output file
wb.save(output_path)
