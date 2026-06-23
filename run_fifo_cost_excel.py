import openpyxl

# Load the workbook and select the active sheet
input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_9/task_39432/r3/evolve_39432/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_9/task_39432/r3/evolve_39432/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Assuming (from description):
# A: Item, B: Beginning Inventory Level, C: Cost per Item from B
# D,E,... (pairs): units received and cost per units from multiple purchase orders
# L: Inventory on hand (to consume FIFO)
# M: Where to write resulting FIFO cost

for row in range(2, 6):  # for rows 2 to 5 (M2:M5)
    qty_needed = ws[f'L{row}'].value
    fifo_batches = []
    # 1. Beginning Inventory
    beg_qty = ws[f'B{row}'].value
    beg_cost = ws[f'C{row}'].value
    if beg_qty and beg_qty > 0:
        fifo_batches.append({'qty': beg_qty, 'cost': beg_cost})
    # 2. Process PO columns
    col = 4
    max_col = ws.max_column
    while col + 1 <= max_col and ws.cell(row=1, column=col).value and ws.cell(row=1, column=col+1).value:
        u_col = col
        c_col = col + 1
        po_qty = ws.cell(row=row, column=u_col).value
        po_cost = ws.cell(row=row, column=c_col).value
        if po_qty and po_qty > 0:
            fifo_batches.append({'qty': po_qty, 'cost': po_cost})
        col += 2
    # 3. Calculate FIFO cost per unit
    units_left = qty_needed
    fifo_costs = []
    for batch in fifo_batches:
        take = min(batch['qty'], units_left)
        if take > 0:
            fifo_costs.extend([batch['cost']] * take)
            units_left -= take
        if units_left <= 0:
            break
    # Defensive: if not enough inventory, use what is available
    if len(fifo_costs) == 0:
        per_unit_cost = 0
    else:
        per_unit_cost = round(sum(fifo_costs) / len(fifo_costs), 2)
    ws[f'M{row}'] = per_unit_cost

wb.save(output_path)
