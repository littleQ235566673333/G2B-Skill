import openpyxl
input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r3/eval_370-43_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r3/eval_370-43_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
sheetnames = wb.sheetnames
if 'Before Insert Row' not in sheetnames:
    raise Exception(f'Worksheet not found: Before Insert Row in {sheetnames}')
ws = wb['Before Insert Row']
rows = list(ws.iter_rows(values_only=True))
# Insert rows above every line with "X" in column A (A7:A1000), working backwards for row safety
start_idx, end_idx = 6, 999  # zero-based index for row 7 to 1000
mod_rows = rows.copy()
for idx in range(end_idx, start_idx - 1, -1):
    if idx < len(mod_rows) and str(mod_rows[idx][0]).strip().upper() == 'X':
        mod_rows.insert(idx, tuple([None]*len(mod_rows[idx])))
# Remove old sheet and add new one
wb.remove(ws)
ws_new = wb.create_sheet('Before Insert Row')
for row in mod_rows:
    ws_new.append(row)
wb.save(output_path)
