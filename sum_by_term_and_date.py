import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Input and output file paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_38823_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_38823_tc1/output.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Locate date range (assume in F2 and G2)
start_date = ws['F2'].value
end_date = ws['G2'].value
if isinstance(start_date, str):
    start_date = datetime.strptime(start_date, '%Y-%m-%d')
if isinstance(end_date, str):
    end_date = datetime.strptime(end_date, '%Y-%m-%d')

# Columns:
# - Dates in column C
# - Search-in column (for search term) in column D
# - Values to sum in column E

# Data window - assuming headers in row 3, data starts row 4
start_row = 4
# Find the last data row
last_data_row = ws.max_row

# Search terms in H4:H7
search_terms = [ws[f'H{row}'].value for row in range(4,8)]

# For each term, sum qualifying values
results = []
for term in search_terms:
    term_sum = 0
    for row in range(start_row, last_data_row+1):
        date_cell = ws[f'C{row}'].value
        # Convert date if text
        if isinstance(date_cell, str):
            try:
                date_val = datetime.strptime(date_cell, '%Y-%m-%d')
            except:
                continue
        elif isinstance(date_cell, datetime):
            date_val = date_cell
        else:
            continue
        if date_val < start_date or date_val > end_date:
            continue
        search_col_val = ws[f'D{row}'].value
        if search_col_val and term in str(search_col_val):
            value_cell = ws[f'E{row}'].value
            if isinstance(value_cell, (int, float)):
                term_sum += value_cell
    results.append(term_sum)

# Write results to I4:I7
for i, value in enumerate(results):
    ws[f'I{4+i}'] = value

wb.save(output_path)
