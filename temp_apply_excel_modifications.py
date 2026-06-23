import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_4/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_4/regression_gate/after_pass/core_32337/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']
# 1. Set formula for E3:E15: lookup CATEGORY from I, given "age (year)" in O matches H.
for row in range(3, 16):
    ws['E'+str(row)] = '=INDEX($I$3:$I$15,MATCH(O{0},$H$3:$H$15,0))'.format(row)
# 2. Insert column after O (O is col 15, so new col 16)
ws.insert_cols(16)
ws.cell(row=2, column=16, value='Actual Age')
# Style for the header, match fill from previous
col15_fill = ws.cell(row=2, column=15).fill
fill = PatternFill(fill_type=col15_fill.fill_type, fgColor=col15_fill.fgColor, bgColor=col15_fill.bgColor) if col15_fill.fill_type else PatternFill()
hdr = ws.cell(row=2, column=16)
hdr.alignment = Alignment(horizontal='center', vertical='top')
hdr.font = Font(bold=True)
hdr.fill = fill
# Values and formatting for each row
for row in range(3, 16):
    c = ws.cell(row=row, column=16)
    c.value = '=DATEDIF(C{0},$B$1,"Y")'.format(row)
    c.number_format = '0'
    c.alignment = Alignment(horizontal='center')
    c.font = Font(bold=True)
    c.fill = fill
wb.save(output_path)
