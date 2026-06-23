import openpyxl

# Load the workbook and select the active sheet
input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_32093_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_32093_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# For rows 2 to 15, read B, C, D, E; set F.
for row in range(2, 16):
    current = ws[f'B{row}'].value
    new1 = ws[f'C{row}'].value
    new2 = ws[f'D{row}'].value
    new3 = ws[f'E{row}'].value
    
    # Find first non-empty new employee name
    result = new1 or new2 or new3 or current
    ws[f'F{row}'] = result

wb.save(output_path)
