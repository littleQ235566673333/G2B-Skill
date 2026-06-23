import openpyxl

# Paths
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_49945_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_49945_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# For G4:G6, sum the values in D4:D11 for model in C4:C11 matching F4:F6
for output_row, lookup_row in zip([4,5,6], [4,5,6]):
    model_name = ws[f'F{lookup_row}'].value
    total = 0
    for row in range(4,12):
        if ws[f'C{row}'].value == model_name:
            qty = ws[f'D{row}'].value
            if qty is not None:
                total += qty
    ws[f'G{output_row}'] = total

wb.save(output_path)
