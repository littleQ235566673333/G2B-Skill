import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

# Input/Output paths
file_in = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_8/regression_gate/before_fix/core_177-6/input.xlsx'
file_out = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_8/regression_gate/before_fix/core_177-6/output.xlsx'

# Load workbook and sheets
wb = load_workbook(file_in)
ws = wb['DATA']
ws_out = wb['combined']

# Extract headers for region A1:R1
headers = [ws.cell(row=1, column=i+1).value for i in range(18)]

# Prepare data: collect all rows
all_rows = []
for row in ws.iter_rows(min_row=2, max_col=18, values_only=False):
    cell_dict = {}
    for i, cell in enumerate(row):
        cell_dict[headers[i]] = cell.value
        cell_dict[f'_cell_{i}'] = cell
    all_rows.append(cell_dict)

# Group by column H (index 7)
group_key = headers[7]
grouped = {}
for row in all_rows:
    key = row[group_key]
    if key not in grouped:
        grouped[key] = []
    grouped[key].append(row)

# Write header to output
for col_idx in range(18):
    src = ws.cell(row=1, column=col_idx+1)
    tgt = ws_out.cell(row=1, column=col_idx+1)
    tgt.value, tgt.data_type = src.value, src.data_type
    tgt._style = copy(src._style)

# Write up to 7 grouped merged rows
row_idx = 2
for group_rows in list(grouped.values())[:7]:
    # For A:H, copy from first occurrence
    for col_idx in range(18):
        src = group_rows[0][f'_cell_{col_idx}']
        tgt = ws_out.cell(row=row_idx, column=col_idx+1)
        if col_idx < 8:
            # Copy content and formatting from first row in group
            tgt.value = src.value
            tgt.data_type = src.data_type
            tgt._style = copy(src._style)
        else:
            # Columns I:R = 8:17, consolidate: sum numeric, otherwise ignore
            vals = [r[headers[col_idx]] for r in group_rows]
            # Filter: accept int/float, or str that is non-empty and numeric
            nm = []
            for v in vals:
                if isinstance(v, (float, int)):
                    nm.append(float(v))
                elif isinstance(v, str) and v not in ('', '\xa0', None) and v.strip() != '':
                    try:
                        nm.append(float(v))
                    except:
                        pass
            val = sum(nm) if nm else 0
            tgt.value = ('%.2f' % val) if val != 0 else ''
            tgt.data_type = 'n' if val != 0 else 's'
            tgt.number_format = '0.00'
            # Copy style from first row in group (could be refined to most common)
            tgt._style = copy(src._style)
    row_idx += 1
    if row_idx > 8:
        break

wb.save(file_out)
print('Done')
