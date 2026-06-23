import openpyxl
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_2/group_45896/r1/evolve_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_2/group_45896/r1/evolve_45896/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Volym P5_P6_2023']
ws_zord = wb['ZORD']

# Gather all dates per material in ZORD
material_dates = {}
for row in ws_zord.iter_rows(min_row=2, values_only=True):
    material = row[0]
    date = row[1]
    if material and isinstance(date, datetime):
        material_dates.setdefault(material, []).append(date)

for mat in material_dates:
    material_dates[mat].sort()

for i in range(2, 11):  # rows 2-10 inclusive
    material = ws.cell(row=i, column=1).value
    dates = material_dates.get(material, [])
    # Format all dates as DD/MM/YYYY
    dates_fmt = ','.join([d.strftime('%d/%m/%Y') for d in dates]) if dates else ''
    ws.cell(row=i, column=9, value=dates_fmt)  # I column = 9

wb.save(output_path)
