from openpyxl import load_workbook
import re

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v3/train/iter_2/regression_gate/before_pass/core_290-27/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v3/train/iter_2/regression_gate/before_pass/core_290-27/output.xlsx"

wb = load_workbook(input_path)
ws = wb.active

start_row, end_row = 14, 137

for row in ws.iter_rows(min_row=start_row, max_row=end_row):
    cell_b = row[1]  # Column B is index 1
    val = cell_b.value
    if isinstance(val, str):
        # Remove leading 2-3 uppercase letters and optional space before numbers
        # Ex: PID1 -> 1, GG 1 -> 1
        new_val = re.sub(r"^[A-Z]{2,3}\s*", "", val)
        # Only update if something was actually removed and what's left is numeric (allowing spaces)
        if new_val != val and new_val.strip() != "" and new_val.replace(" ", "").isdigit():
            cell_b.value = new_val.strip()

wb.save(output_path)
