import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_8/regression_gate/before_fix/core_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_8/regression_gate/before_fix/core_4714/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

rows = list(ws.iter_rows(min_row=2, max_row=25, min_col=1, max_col=4, values_only=True))
# Each row: (A, B, C, D) but we need only (A: employee, C: month, D: hours)
results = []
for idx, row in enumerate(rows):
    employee = row[0]
    month = row[2]
    hours = row[3]
    # Gather all records for this employee up to this row (for windowing)
    history = []
    for jdx in range(len(rows)):
        if rows[jdx][0] == employee:
            if rows[jdx][2] is not None and jdx <= idx:
                history.append((rows[jdx][2], rows[jdx][3]))
    history.sort(key=lambda x: x[0])
    cur_idx = None
    for j, hisrow in enumerate(history):
        if hisrow[0] == month and hisrow[1] == hours:
            cur_idx = j
    if cur_idx is None or cur_idx < 3:
        results.append('n/a')
    else:
        win = history[cur_idx-3:cur_idx+1]
        win_hours = [h[1] for h in win if h[1] is not None]
        if len(win_hours) < 4:
            results.append('n/a')
        else:
            avg = sum(win_hours) / 4
            if avg == int(avg):
                avg_out = str(int(avg))
            else:
                avg_out = f'{avg:.2f}'.rstrip('0').rstrip('.')
            results.append(avg_out)
# Write to E2:E25
for i, val in enumerate(results):
    ws[f'E{i+2}'] = val
wb.save(output_path)
print('Done')
