from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_3/regression_gate/before_fix/core_11276/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_3/regression_gate/before_fix/core_11276/output.xlsx"

wb = load_workbook(input_path)
ws = wb.active  # Assuming edit is on first sheet

# Fill F3:AJ3 (columns 6 to 36) with weekday formula referencing F4:AJ4
for col in range(6, 37):  # F=6, AJ=36
    date_cell_label = f'{get_column_letter(col)}4'
    formula = f'=TEXT({date_cell_label}, "DDD")'
    ws.cell(row=3, column=col).value = formula

wb.save(output_path)
