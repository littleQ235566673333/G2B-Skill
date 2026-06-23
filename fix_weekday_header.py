from openpyxl import load_workbook

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_5/regression_gate/after_fix/core_11276/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_5/regression_gate/after_fix/core_11276/output.xlsx"

wb = load_workbook(input_path)
ws = wb.active  # Use first sheet unless specified otherwise

# cols F (6) to AJ (36), fill formulas in row 3, referencing row 4
def col_range(start, end):
    return range(start, end+1)

for col in col_range(6, 36):  # F=6, AJ=36
    date_ref = ws.cell(row=4, column=col).coordinate
    ws.cell(row=3, column=col).value = f'=TEXT({date_ref}, "DDD")'

wb.save(output_path)
