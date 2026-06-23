from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import datetime
from copy import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_2/regression_gate/before_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_2/regression_gate/before_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

header_row = 2
start_row, end_row = 3, 15

# Map headers correctly
headers = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row])}
dob_col = headers.get('DATE OF BIRTH')
age_col = headers.get('age (year)')
category_col = headers.get('CATEGORY')
expected_result_col = headers.get('Expected Result’') or headers.get('Expected Result') or 5

# Actual Age column after age(year), add if missing
actual_age_col = age_col + 1

# Fill/color/align from age(year) col
from copy import copy
age_header = ws.cell(header_row, age_col)
age_fill = copy(age_header.fill)
center_align = Alignment(horizontal='center', vertical='center')
top_align = Alignment(horizontal='center', vertical='top')
bold_font = Font(bold=True)

# Get report date from B1 (cell B1)
report_date = ws['B1'].value

for row in range(start_row, end_row + 1):
    dob = ws.cell(row, dob_col).value
    act_age = None
    if isinstance(dob, datetime.datetime) and isinstance(report_date, datetime.datetime):
        act_age = int((report_date - dob).days // 365.25)
    cell = ws.cell(row, actual_age_col)
    cell.value = act_age
    cell.alignment = center_align
    cell.font = bold_font
    cell.fill = age_fill
    cell.number_format = '0'

header = ws.cell(header_row, actual_age_col)
header.value = 'Actual Age'
header.alignment = top_align
header.font = bold_font
header.fill = age_fill

# Formula setup for Expected Result (column E)
cat_letter = get_column_letter(category_col)
age_letter = get_column_letter(age_col)
actual_letter = get_column_letter(actual_age_col)
for row in range(start_row, end_row + 1):
    formula = f"=INDEX(${cat_letter}$3:${cat_letter}$15, MATCH({actual_letter}{row}, ${age_letter}$3:${age_letter}$15, 0))"
    ws.cell(row, expected_result_col).value = formula

wb.save(output_path)
print('done')
