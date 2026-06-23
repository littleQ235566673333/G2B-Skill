import openpyxl
from collections import defaultdict

INPUT = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_57989_tc1/input.xlsx"
OUTPUT = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_57989_tc1/output.xlsx"
SHEET = "Sheet1"

def main():
    wb = openpyxl.load_workbook(INPUT)
    ws = wb[SHEET]

    # --- Read synthesis table drivers and weekdays headings (output space) ---
    out_drivers = []
    for row in ws.iter_rows(min_row=25, max_row=43, min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v and str(v).lower().startswith('driver'):
            out_drivers.append(v)
    out_weekdays = []
    for col in ws.iter_cols(min_row=24, max_row=24, min_col=2, max_col=8, values_only=True):
        v = col[0]
        if v and v.strip():
            out_weekdays.append(v.strip())

    # --- Locate input trip data area ---
    # Input driver names in A3:A21 (? scanned until first blank cell or title row)
    drivers = []
    for row in ws.iter_rows(min_row=3, max_row=21, min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v and str(v).lower().startswith('driver'):
            drivers.append(v)
    # Input weekdays: B1:U1 (col 2–21)
    weekdays = [cell.value for cell in ws[1][1:21]]

    # List all (col_idx, weekday_value) for input area
    weekday_col_indices = [(i+2, wd) for i, wd in enumerate(weekdays)]
    # Build lookup for each driver/weekday to cell-count
    driver_weekday_counts = defaultdict(lambda: defaultdict(int))
    for driver_idx, driver in enumerate(drivers, start=3):
        for col_idx, weekday in weekday_col_indices:
            val = ws.cell(row=driver_idx, column=col_idx).value
            if val is not None and str(val).strip() != '':
                driver_weekday_counts[driver][weekday] += 1

    # --- Aggregate synthesis, write to B25:H43 ---
    for out_rowi, out_driver in enumerate(out_drivers, start=25):
        for out_colj, out_weekday in enumerate(out_weekdays, start=2):
            # Synthesis table only covers one week (Friday to Thursday), need all those weekdays, any week
            # So sum for input, for all columns where weekday matches out_weekday
            count = driver_weekday_counts[out_driver][out_weekday] if out_weekday in driver_weekday_counts[out_driver] else 0
            ws.cell(row=out_rowi, column=out_colj, value=count)

    wb.save(OUTPUT)

if __name__ == "__main__":
    main()
