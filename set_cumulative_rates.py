import openpyxl
from openpyxl.styles import PatternFill

def colnum_to_letter(n):
    result = ''
    while n:
        n, r = divmod(n-1, 26)
        result = chr(65+r) + result
    return result

# Load workbook and sheet
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/group_44017/r1/evolve_44017/input.xlsx')
ws = wb['Data']

# Column indexes (1-based)
date_row = 9
start_row, end_row = 14, 42
start_col, end_col = 30, 41  # AD:AO

# We'll set formulas for each cell in AD14:AO42
for row in range(start_row, end_row+1):
    for col in range(start_col, end_col+1):
        excel_col = colnum_to_letter(col)
        cell = ws.cell(row=row, column=col)
        months_offset = col - start_col
        # References
        base = f'W{row}'
        effdate = f'L{row}'
        freq = f'J{row}'
        inc1 = f'M{row}'
        inc2 = f'N{row}'
        inc3 = f'O{row}'
        inc4 = f'P{row}'
        # Date cell for this column
        date_cell = f'{excel_col}{date_row}'
        # Formula: Only calculate if the date >= effdate
        formula = (
            f'=IF({date_cell}>={effdate},'
            f'{base}*'
            f'(1+IF((FLOOR((({date_cell}-{effdate})/30)/{freq},1)>=1),{inc1},0))'
            f'* (1+IF((FLOOR((({date_cell}-{effdate})/30)/{freq},1)>=2),{inc2},0))'
            f'* (1+IF((FLOOR((({date_cell}-{effdate})/30)/{freq},1)>=3),{inc3},0))'
            f'* (1+IF((FLOOR((({date_cell}-{effdate})/30)/{freq},1)>=4),{inc4},0)),"")'
            f')'
        )
        cell.value = formula
        # Remove yellow fill if present
        if hasattr(cell, 'fill'):
            cell.fill = PatternFill(fill_type=None)

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/group_44017/r1/evolve_44017/output.xlsx')
print('done')
