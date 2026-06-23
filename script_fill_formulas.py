from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/before_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/before_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

header_row = 2
headers = [cell.value for cell in ws[header_row]]
field_to_col = {name: idx+1 for idx, name in enumerate(headers) if name}

cat_col = field_to_col.get('CATEGORY', 9) # col I
expected_col = field_to_col.get('Expected Result', 5) # col E
age_year_col = field_to_col.get('age (year)', 15) # col O
name_col = field_to_col.get('NAME', 1) # col A
dob_col = field_to_col.get('DATE OF BIRTH', 3) # col C

# Fill E3:E15 with formula referencing I3:I15
for r in range(3, 16):
    ws.cell(row=r, column=expected_col).value = f"=INDEX($I$3:$I$15,ROW()-2)"

# Insert new col after O = 15, so new actual age at P = 16
ws.insert_cols(age_year_col + 1)
actual_age_col = age_year_col + 1

# Convert fill for actual copying
src_fill = ws.cell(row=header_row, column=age_year_col).fill
if isinstance(src_fill, PatternFill) and src_fill.patternType is not None:
    new_fill = PatternFill(patternType=src_fill.patternType,
                          fgColor=src_fill.fgColor.rgb,
                          bgColor=src_fill.bgColor.rgb)
else:
    new_fill = PatternFill()

# Write header for Actual Age, with formatting
header_cell = ws.cell(row=header_row, column=actual_age_col)
header_cell.value = 'Actual Age'
header_cell.font = Font(bold=True)
header_cell.alignment = Alignment(horizontal='center', vertical='top')
header_cell.fill = new_fill

# Value fill for data cells (match from O3 - if none, leave blank)
data_fill = ws.cell(row=3, column=age_year_col).fill
if isinstance(data_fill, PatternFill) and data_fill.patternType is not None:
    cell_data_fill = PatternFill(patternType=data_fill.patternType, fgColor=data_fill.fgColor.rgb,
                                 bgColor=data_fill.bgColor.rgb)
else:
    cell_data_fill = PatternFill()

for r in range(3, 16):
    cell = ws.cell(row=r, column=actual_age_col)
    cell.value = f"=INT(($B$1-C{r})/365.25)"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = cell_data_fill

wb.save(output_path)
print('DONE')
