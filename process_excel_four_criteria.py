import openpyxl

# Paths
input_path = "results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_55468_tc1/input.xlsx"
output_path = "results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_55468_tc1/output.xlsx"

def find_four_criteria_value(ws):
    # Assume the structure based on the formula and problem description
    # - Horizontal headers: row 3 (Z3:...)
    # - Second horizontal headers: row 4
    # - Vertical headers: column A (A5:A10)
    # - Second vertical headers: column B (B5:B10)
    # - Data: C5:Z10 (as in your formulas)
    #
    # The lookup values are:
    # AC4 = ws['AC4'] (horizontal)
    # AB4 = ws['AB4'] (horizontal)
    # AE4 = ws['AE4'] (vertical)
    # AD4 = ws['AD4'] (vertical)
    #
    # We want the value in the data table where headers match all four conditions

    # Read in header and lookup values
    ac4 = ws['AC4'].value
    ab4 = ws['AB4'].value
    ae4 = ws['AE4'].value
    ad4 = ws['AD4'].value

    # Identify relevant rows (vertical criteria)
    v_headers_1 = [ws[f'A{row}'].value for row in range(5, 11)]
    v_headers_2 = [ws[f'B{row}'].value for row in range(5, 11)]

    # Find row index matching both vertical headers
    row_index = None
    for i, (vh1, vh2) in enumerate(zip(v_headers_1, v_headers_2)):
        if vh1 == ae4 and vh2 == ad4:
            row_index = i
            break

    if row_index is None:
        return '#N/A (Row not found)'

    # Identify relevant columns (horizontal criteria)
    # Row 3: ws['C3':'Z3'], Row 4: ws['C4':'Z4']
    col_headers_1 = [ws.cell(row=3, column=col).value for col in range(3, 27)]
    col_headers_2 = [ws.cell(row=4, column=col).value for col in range(3, 27)]

    col_index = None
    for j, (ch1, ch2) in enumerate(zip(col_headers_1, col_headers_2)):
        if ch1 == ab4 and ch2 == ac4:
            col_index = j
            break

    if col_index is None:
        return '#N/A (Col not found)'

    # Data block starts at (row 5, col 3)
    data_cell = ws.cell(row=5 + row_index, column=3 + col_index)
    return data_cell.value

# Load input
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Find the value
result = find_four_criteria_value(ws)

# Write to AE5
ws['AE5'].value = result

# Save
wb.save(output_path)
