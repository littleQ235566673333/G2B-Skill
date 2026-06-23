from openpyxl import load_workbook

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_5/regression_gate/before_fix/core_280-17/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_5/regression_gate/before_fix/core_280-17/output.xlsx'

wb = load_workbook(in_path)
ws = wb.active
rows = list(ws.iter_rows(min_row=1, values_only=True))

value_to_last_idx = {}
for idx, row in enumerate(rows, start=1):
    val = row[1]  # Second column (num)
    value_to_last_idx[val] = idx
rows_to_keep = set(value_to_last_idx.values())

# Delete non-kept rows from bottom to top
for del_idx in reversed(range(1, len(rows)+1)):
    if del_idx not in rows_to_keep:
        ws.delete_rows(del_idx, 1)

wb.save(out_path)
