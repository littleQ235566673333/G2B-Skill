from openpyxl import load_workbook

# Load input file
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/before_fix/core_54274/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/before_fix/core_54274/output.xlsx'
wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# Mapping columns manually from printout
# Target: B4, Stretch: C4, Over Stretch: D4, Total: F4, Points: G4
# Write nested IFs with caps as per reference:
ws['G4'] = (
    '=IF(F4<=B4, F4/B4,'
    'IF(F4<B4, F4/B4,'                    # never happens: F4<=B4 checked first
    'IF(AND(F4>B4,F4<C4), MIN(F4/B4,1.09),'
    'IF(F4=C4, MIN(F4/B4,1.1),'
    'IF(AND(F4>C4,F4<D4), MIN(F4/B4,1.19),'
    'IF(F4>=D4, MIN(F4/B4,1.2), "")))))))'
)

# Simplify the redundant 2nd condition
ws['G4'] = (
    '=IF(F4<=B4, F4/B4,'
    'IF(AND(F4>B4,F4<C4), MIN(F4/B4,1.09),'
    'IF(F4=C4, MIN(F4/B4,1.1),'
    'IF(AND(F4>C4,F4<D4), MIN(F4/B4,1.19),'
    'IF(F4>=D4, MIN(F4/B4,1.2), ""))))))'
)

# Save result
wb.save(output_path)
