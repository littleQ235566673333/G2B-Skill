import openpyxl
from openpyxl.styles import PatternFill, Font

# Input/output paths
input_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_50631_tc1/input.xlsx"
output_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_50631_tc1/output.xlsx"

# Load workbook and the relevant sheet ('Sheet1')
wb = openpyxl.load_workbook(input_path)
ws = wb["Sheet1"]

# Use Calibri font throughout
def apply_calibri(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Calibri")

apply_calibri(ws)

# Hide gridlines (for openpyxl, set sheet property)
ws.sheet_view.showGridLines = False

# Identify holiday range (column J, starting at J3 and down)
holiday_cells = []
for row in range(3, ws.max_row+1):
    holiday_cell = ws[f"J{row}"]
    if holiday_cell.value:
        holiday_cells.append(f"J{row}")
holiday_range = f"J3:J{row}" if holiday_cells else ""

# Fill color #FFFF00
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Write the formula in H7:H38
target_start = 7
h_start = f"H{target_start}"
h_end = 38
for r in range(target_start, h_end+1):
    # Compose formula
    formula = f"=IF($B$3+ROWS($B$7:$B{r})-1<=$F$3,WORKDAY($B$3+ROWS($B$7:$B{r})-1,0,$J$3:$J${ws.max_row}),"")"
    ws[f"H{r}"].value = formula
    ws[f"H{r}"].fill = yellow_fill
    ws[f"H{r}"].font = Font(name="Calibri")

wb.save(output_path)
