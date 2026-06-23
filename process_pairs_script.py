import openpyxl
from openpyxl import load_workbook
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/group_146-49/r3/evolve_146-49/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/group_146-49/r3/evolve_146-49/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet1']
words = []
max_col = ws.max_column
max_row = ws.max_row

# Guess data column (most likely in A, or first non-empty column)
first_col = 1
for col in range(1, max_col+1):
    # Check for at least one non-empty cell except maybe headers
    if any(ws.cell(row=r, column=col).value for r in range(1, min(20,max_row+1))):
        first_col = col
        break
for row in range(2, max_row+1): # skip header
    val = ws.cell(row=row, column=first_col).value
    if isinstance(val, str) and val.strip():
        words.append(val.strip().upper())

vowels = 'AEIOU'
pairs = set()
length_buckets = defaultdict(list)
for w in words:
    length_buckets[len(w)].append(w)
for wordlist in length_buckets.values():
    n = len(wordlist)
    for i in range(n):
        w1 = wordlist[i]
        for j in range(i+1, n):
            w2 = wordlist[j]
            if w1 == w2: continue
            # Compare pattern
            if len(w1) != len(w2): continue
            # 1. Is w2 a vowelized version of w1 for exactly one vowel?
            match = True
            v_found = None
            for ch1, ch2 in zip(w1, w2):
                if ch1 == 'S':
                    if ch2 not in vowels:
                        match = False
                        break
                    if v_found is None:
                        v_found = ch2
                    elif v_found != ch2:
                        match = False
                        break
                else:
                    if ch1 != ch2:
                        match = False
                        break
            if match and v_found:
                pairs.add((w1, w2))
            # 2. Is w1 a vowelized version of w2?
            match = True
            v_found = None
            for ch1, ch2 in zip(w2, w1):
                if ch1 == 'S':
                    if ch2 not in vowels:
                        match = False
                        break
                    if v_found is None:
                        v_found = ch2
                    elif v_found != ch2:
                        match = False
                        break
                else:
                    if ch1 != ch2:
                        match = False
                        break
            if match and v_found:
                pairs.add((w2, w1))
# Output pairs sorted case-insensitively, start at G1:H1
pairs = sorted(pairs, key=lambda x: (x[0].lower(), x[1].lower()))
for idx, (w1, w2) in enumerate(pairs):
    ws.cell(row=idx+1, column=7).value = w1
    ws.cell(row=idx+1, column=8).value = w2
wb.save(output_path)
