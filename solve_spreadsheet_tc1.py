import openpyxl
from openpyxl.utils import get_column_letter

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_53161_tc1/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_53161_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_fp)
ws = wb['Sheet1']

# 1. Parse Agent Categories table (rows 3-8, cols E:M)
categories_header_row = 3
categories_data_rows = range(4, 10)  # 4,5,6,7,8,9
categories_cols = range(5, 14)  # E(5):M(14)

agent_category_map = {}
category_names = []

for row in categories_data_rows:
    agent = ws.cell(row=row, column=5).value
    agent_category_map[agent] = []
    for col in categories_cols:
        cat_value = ws.cell(row=row, column=col).value
        category = ws.cell(row=categories_header_row, column=col).value
        if row == categories_data_rows[0] and category:
            category_names.append(category)
        if cat_value:
            agent_category_map[agent].append(category)

# 2. Parse Schedule table (rows 12-17, cols F:AE)
schedule_header_row = 12
schedule_data_rows = range(13, 19)  # 13,14,15,16,17,18
schedule_agent_col = 5  # E
schedule_cols = range(6, 32)  # F(6):AE(31)

interval_times = []
schedule_map = {}
for col in schedule_cols:
    value = ws.cell(row=schedule_header_row, column=col).value
    interval_times.append(value)
for row in schedule_data_rows:
    agent = ws.cell(row=row, column=schedule_agent_col).value
    schedule_map[agent] = []
    for idx, col in enumerate(schedule_cols):
        if ws.cell(row=row, column=col).value:
            schedule_map[agent].append(interval_times[idx])

# 3. For each interval, sum by category
total_table_start_row = 22
category_rows = range(total_table_start_row+1, total_table_start_row+1+len(category_names))

totals = {cat: [0 for _ in interval_times] for cat in category_names}

for interval_idx, interval in enumerate(interval_times):
    # find agents scheduled for this interval
    scheduled_agents = [agent for agent, intervals in schedule_map.items() if interval in intervals]
    # for each category
    for cat in category_names:
        count = sum(1 for agent in scheduled_agents if cat in agent_category_map.get(agent, []))
        totals[cat][interval_idx] = count

# 4. Write results to F22:AE29
for i, cat in enumerate(category_names):
    row = total_table_start_row + 1 + i
    ws.cell(row=row, column=5).value = cat
    for j, count in enumerate(totals[cat]):
        ws.cell(row=row, column=6 + j).value = count

wb.save(output_fp)
