import pandas as pd
from openpyxl import load_workbook

file_in = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun2/eval_267-21_tc1/input.xlsx'
file_out = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun2/eval_267-21_tc1/output.xlsx'

wb = load_workbook(file_in)
ws_merge = wb['merging']
ws_sh1 = wb['SH1']
ws_rp1 = wb['RP1']

# Read IDs from Column B2:B11 in 'merging'
ids = [ws_merge.cell(row=r, column=2).value for r in range(2, 12)]

# Helper to build a dict for SH1, RP1
def id_to_qty(sheet):
    d = {}
    maxcol = sheet.max_column
    hdrs = [sheet.cell(row=1, column=c).value for c in range(1, maxcol+1)]
    try:
        id_col = hdrs.index('ID') + 1
    except:
        return {}
    try:
        qty_col = hdrs.index('QTY') + 1
    except:
        return {}
    for row in range(2, sheet.max_row+1):
        idv = sheet.cell(row=row, column=id_col).value
        qtyv = sheet.cell(row=row, column=qty_col).value
        d[idv] = qtyv
    return d

data_sh1 = id_to_qty(ws_sh1)
data_rp1 = id_to_qty(ws_rp1)

for i, idv in enumerate(ids):
    cval = data_sh1.get(idv, '-') if idv is not None else '-'
    dval = data_rp1.get(idv, '-') if idv is not None else '-'
    ws_merge.cell(row=2+i, column=3).value = cval
    ws_merge.cell(row=2+i, column=4).value = dval

wb.save(file_out)
