from openpyxl import load_workbook

def remove_row_duplicates(values):
    seen = set()
    result = []
    for v in values:
        if v in seen and v is not None and v != "":
            result.append("")
        else:
            seen.add(v)
            result.append(v)
    return result

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/regression_gate/before_fix/core_227-40/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/regression_gate/before_fix/core_227-40/output.xlsx"
wb = load_workbook(input_path)
ws = wb["Sheet1"]

# Based on user sample sizes, let's assume the input to process is in A2:E5 (4 rows, 5 columns)
start_in_row = 2
end_in_row = 5
start_out_row = 14

for i, data_row in enumerate(range(start_in_row, end_in_row + 1)):
    vals = [ws.cell(row=data_row, column=c).value for c in range(1, 6)]
    cleaned = remove_row_duplicates(vals)
    for col_idx, val in enumerate(cleaned, 1):
        ws.cell(row=start_out_row + i, column=col_idx).value = val

wb.save(output_path)
