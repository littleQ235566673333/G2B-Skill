from openpyxl import load_workbook

# Path setup
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_7/group_290-1/r1/evolve_290-1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_7/group_290-1/r1/evolve_290-1/output.xlsx"

# Load workbook and select sheet
wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# For each row 2 to 10
for r in range(2, 11):
    col_k = 11  # K
    col_u = 21  # U
    # Gather values for columns K to U by even-odd chunks (pattern matches pairs of columns after I/H)
    values = []
    for i in range(col_k, col_u+1, 2):
        idx = i
        val1 = ws.cell(row=r, column=idx).value
        val2 = ws.cell(row=r, column=idx+1).value
        values.extend([val1, val2])
    # Write the values back (should match K-U with pos correspondence)
    for offset, value in enumerate(values):
        target_col = col_k + offset
        ws.cell(row=r, column=target_col).value = value

# Save output
wb.save(output_path)
