import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_1/group_91-34/r3/evolve_91-34/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_1/group_91-34/r3/evolve_91-34/output.xlsx'
sheet = 'SwiftMD'

# Use header row 1, so data starts at row 2 (index 1)
df = pd.read_excel(input_path, sheet_name=sheet, header=1)

# Mark for deletion
col_last = 'Last Name'
col_first = 'First Name'
col_dob = 'Date Of Birth'
col_dup = 'Duplicate?'
col_rel = 'Relationship'

df['__to_delete'] = False

def safe_str(x):
    return str(x).strip().lower() if pd.notnull(x) else ''

for key, subdf in df[df[col_dup]=='Yes'].groupby([col_last, col_first, col_dob]):
    non_employee = subdf[subdf[col_rel].map(safe_str) != 'employee']
    if len(non_employee) > 1:
        idx = non_employee.index[0]
        df.loc[idx, '__to_delete'] = True

# Remove marked rows
df_result = df[~df['__to_delete']].drop(columns='__to_delete')

# Write back to B2:O42 on the sheet (header+40 rows of data)
wb = load_workbook(input_path)
ws = wb[sheet]

# Clear target range
for r in ws.iter_rows(min_row=2, max_row=42, min_col=2, max_col=15):
    for cell in r:
        cell.value = None

# Write headers
for col_idx, col_name in enumerate(df_result.columns, start=2):
    ws.cell(row=2, column=col_idx, value=col_name)

# Write new data (header row is at row 2, data starts row 3)
for row_idx, row in enumerate(df_result.head(40).itertuples(index=False), start=3):
    for col_idx, val in enumerate(row, start=2):
        ws.cell(row=row_idx, column=col_idx, value=val)

wb.save(output_path)
