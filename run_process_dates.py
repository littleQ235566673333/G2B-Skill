import openpyxl
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_8/regression_gate/after_fix/core_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_8/regression_gate/after_fix/core_45896/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws_main = wb['Volym P5_P6_2023']
ws_zord = wb['ZORD']
def format_date(v):
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, (int, float)) and v > 0:
        try:
            return openpyxl.utils.datetime.from_excel(v).strftime('%d/%m/%Y')
        except Exception:
            return str(v)
    return str(v)
for row in range(2, 11):
    key = ws_main.cell(row=row, column=1).value
    dates = []
    seen = set()
    for zord_row in ws_zord.iter_rows(min_row=2, values_only=True):
        if zord_row[0] == key:
            val = zord_row[2]
            fmt_val = format_date(val)
            if fmt_val not in seen and fmt_val.strip() != '':
                seen.add(fmt_val)
                dates.append(fmt_val)
    ws_main.cell(row=row, column=9).value = ','.join(dates)
wb.save(output_path)
