import pandas as pd
from openpyxl import load_workbook

def parse_month(x):
    try:
        return pd.to_datetime(x)
    except:
        return pd.NaT

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_4/regression_gate/before_fix/core_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_4/regression_gate/before_fix/core_4714/output.xlsx'

df = pd.read_excel(input_path)
df['Month_parsed'] = df.iloc[:,2].apply(parse_month)

rows = df.shape[0]
res = [''] * rows
for i in range(rows):
    emp = df.iloc[i, 0]
    this_month = df.iloc[i]['Month_parsed']
    emp_rows = df[df.iloc[:,0] == emp]
    window = emp_rows[(emp_rows['Month_parsed'] <= this_month)].sort_values('Month_parsed', ascending=False).head(4)
    if window.shape[0] < 4:
        res[i] = 'n/a'
    else:
        avg = window.iloc[:,3].mean()
        res[i] = str(int(avg)) if avg == int(avg) else '{:.2f}'.format(avg)

wb = load_workbook(input_path)
ws = wb.active
for ix, v in enumerate(res[:24], 2):
    ws.cell(row=ix, column=5, value=v)
wb.save(output_path)
