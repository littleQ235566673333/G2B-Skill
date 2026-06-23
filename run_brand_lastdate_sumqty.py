import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_516-46/r1/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_516-46/r1/evolve_516-46/output.xlsx'
sheet_name = 'ورقة1'

# Load dataframe
df = pd.read_excel(input_path, sheet_name=sheet_name, header=None)
df = df.iloc[1:]  # Skip the first row, as A2 is the data start
cols = ['Brand', 'Qty', 'Other1', 'Other2', 'Date']
df = df.iloc[:, 0:5]
df.columns = cols
df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
df = df.dropna(subset=['Brand', 'Qty', 'Date'], how='any')

# Find last date per brand
last_dates = df.groupby('Brand')['Date'].transform('max')
df_last = df[df['Date'] == last_dates]

# Combine for same brand and date (summing Qty)
grouped = df_last.groupby(['Brand','Date'], as_index=False).agg({
    'Qty': 'sum',
    'Other1': 'first', # Take the first of Other1, Other2 (if they match within brand-date)
    'Other2': 'first'
})

# Prepare writing results to Excel
wb = load_workbook(input_path)
ws = wb[sheet_name]

# Write headers (optional, if wanted): H1:L1 -> cols
for j, col in enumerate(['Brand', 'Qty', 'Other1', 'Other2', 'Date'], start=8):
    ws.cell(row=1, column=j, value=col)

# Write 3 rows (up to L4; adjust if needed)
for i, row in enumerate(grouped.values[:3], start=2):
    for j, value in enumerate(row, start=8):
        ws.cell(row=i, column=j, value=value)

wb.save(output_path)
