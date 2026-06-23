from openpyxl import load_workbook

infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_7/regression_gate/before_pass/core_41601/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_7/regression_gate/before_pass/core_41601/output.xlsx'
wb = load_workbook(infile)
ws = wb['Students']
for row in range(2, 8):
    name = ws.cell(row=row, column=1).value
    if name:
        # Correct Excel INDIRECT formula for sheet name in A{row}
        formula = f"=INDIRECT('"'{name}'"!C2')"
        ws.cell(row=row, column=5, value=formula)
    else:
        ws.cell(row=row, column=5, value=None)
wb.save(outfile)
