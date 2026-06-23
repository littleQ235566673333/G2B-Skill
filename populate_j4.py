from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_3/regression_gate/after_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_3/regression_gate/after_pass/core_41589/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Contact List']

today = datetime.now().date()
cell_h = ws['H4'].value
cell_i = ws['I4'].value
result = 'NO ACTION'

if cell_h and cell_i:
    if isinstance(cell_h, datetime):
        contact_date = cell_h.date()
    elif isinstance(cell_h, str):
        # Try a couple date formats
        try:
            contact_date = datetime.strptime(cell_h, '%Y-%m-%d').date()
        except ValueError:
            try:
                contact_date = datetime.strptime(cell_h, '%m/%d/%Y').date()
            except ValueError:
                contact_date = None
    else:
        contact_date = None

    if contact_date and isinstance(cell_i, str) and cell_i.strip().upper() == 'YES':
        delta = (today - contact_date).days
        if delta <= 30:
            result = 'HOLD'
        else:
            result = 'TOUCH BASE'

ws['J4'] = result
wb.save(output_path)
