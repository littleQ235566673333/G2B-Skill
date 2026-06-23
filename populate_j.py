from openpyxl import load_workbook
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']
today = datetime.today()
row = 4
while True:
    date_cell = ws[f'H{row}'].value
    yesno_cell = ws[f'I{row}'].value
    # Stop if entire row is empty (assuming both columns H and I being empty signals end)
    if date_cell is None and yesno_cell is None:
        break
    response = 'NO ACTION'
    if yesno_cell and str(yesno_cell).strip().upper() == 'YES':
        # Handle date value
        days_diff = None
        if isinstance(date_cell, datetime):
            days_diff = (today - date_cell).days
        elif date_cell is not None:
            # Try parsing common string formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']:
                try:
                    date_obj = datetime.strptime(str(date_cell), fmt)
                    days_diff = (today - date_obj).days
                    break
                except Exception:
                    continue
        # Now apply logic
        if days_diff is not None:
            if days_diff <= 30:
                response = 'HOLD'
            else:
                response = 'TOUCH BASE'
        else:
            response = 'NO ACTION'
    ws[f'J{row}'] = response
    row += 1
wb.save(output_path)
