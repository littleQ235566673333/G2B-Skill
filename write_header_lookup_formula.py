from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load workbook and sheet
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_2/group_36277/r3/evolve_36277/input.xlsx')
ws = wb.active

# Sheet dimensions
header_row = 1
max_col = ws.max_column
max_row = ws.max_row

# Excel column letter ranges
header_range = f'$A${header_row}:${get_column_letter(max_col)}${header_row}'
data_range = f'$A$2:${get_column_letter(max_col)}${max_row}'

# Insert formula for I2:I5 (assuming lookup key in H2:H5)
for r in range(2, 6):
    formula = f'=INDEX({data_range}, ROW()-1, MATCH(H{r}, {header_range}, 0))'
    ws[f'I{r}'] = formula

# Save results
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_2/group_36277/r3/evolve_36277/output.xlsx')
