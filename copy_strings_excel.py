from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_230-16_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_230-16_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Before']

for row in range(1, 13):
    val = ws[f'A{row}'].value
    ws[f'B{row}'].value = val
    ws[f'A{row}'].value = ''

wb.save(output_path)
