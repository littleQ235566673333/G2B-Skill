from openpyxl import load_workbook

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_7/regression_gate/before_pass/core_41601/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_7/regression_gate/before_pass/core_41601/output.xlsx'

wb = load_workbook(input_fp)
ws_students = wb['Students']

# Write age-reference formulas for each student in E2:E7
for idx in range(2, 8):  # Excel rows 2 to 7
    name = ws_students.cell(row=idx, column=1).value
    out_cell = ws_students.cell(row=idx, column=5)  # Column E
    if not name:
        out_cell.value = ''
        continue
    # Sheet names with spaces or special chars must be quoted
    special = any(c in name for c in " []*?'")
    safe_name = f"'{name}'" if special else name
    out_cell.value = f"={safe_name}!C2"

wb.save(output_fp)
