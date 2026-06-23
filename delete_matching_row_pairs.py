import openpyxl
import os

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r2/eval_493-5_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r2/eval_493-5_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)

# Verify sheet presence
sheetnames = wb.sheetnames
if 'Imported Data' not in sheetnames:
    raise ValueError(f'Sheet "Imported Data" not found. Available: {sheetnames}')
ws = wb['Imported Data']

# ---- Localize header row and map columns ----
header_row_num = None
headers = None
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    norm = [str(cell).strip().lower() if cell else '' for cell in row]
    if 'reference' in norm and 'narrative' in norm:
        headers = norm
        header_row_num = i
        break
if header_row_num is None:
    raise RuntimeError('Sheet missing the required headers (reference, narrative) in the first 10 rows.')

def find_col(header_pattern):
    for idx, h in enumerate(headers):
        if header_pattern in h:
            return idx
    raise ValueError(f'Header containing pattern "{header_pattern}" not found in headers: {headers}')

col_ref = find_col('reference')
col_c   = find_col('debit')   # Column C → Debit
col_d   = find_col('credit')  # Column D → Credit
col_f   = find_col('narrative')

# ---- Read all rows below header ----
data = []
for row in ws.iter_rows(min_row=header_row_num+1, max_row=ws.max_row, values_only=True):
    if not any(x is not None and str(x).strip() != '' for x in row):
        continue
    data.append(list(row))

# ---- Find pairs to delete ----
delete_idxs = set()
for i, row1 in enumerate(data):
    for j, row2 in enumerate(data):
        if i >= j:
            continue
        # criteria: reference, narrative must match; col_c=col_d/col_d=col_c
        ref_match = row1[col_ref] == row2[col_ref]
        narrative_match = row1[col_f] == row2[col_f]
        # One row's C matches other's D, and vice versa
        cross_1 = row1[col_c] == row2[col_d]
        cross_2 = row1[col_d] == row2[col_c]
        if ref_match and narrative_match and (cross_1 or cross_2):
            delete_idxs.update([i, j])

# ---- Create cleaned data ----
cleaned = [row for idx, row in enumerate(data) if idx not in delete_idxs]

# ---- Output: A1:F10 (row 1 = header, then up to 9 data rows) ----
out_ws = ws
first_6_headers = [ws.cell(row=header_row_num, column=j+1).value for j in range(6)]

for ridx in range(10):
    if ridx == 0:
        outrow = first_6_headers
    else:
        # Only up to available rows
        outrow = cleaned[ridx-1][:6] if (ridx-1) < len(cleaned) else ['']*6
    for cidx, val in enumerate(outrow, 1):
        out_ws.cell(row=ridx+1, column=cidx, value=val)

wb.save(output_path)
