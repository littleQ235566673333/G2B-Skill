import openpyxl

# Load workbook and sheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_42198_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_42198_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Store values for processing
data = []
for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=2, values_only=True):
    data.append(row)

results = []
for end in range(1, len(data)+1):
    # Cumulative range: rows 0 to end-1 in data
    sub = data[:end]
    found = False
    # Priority 1: Potato & FALSE
    for a, b in sub:
        if a == 'Potato' and b is False:
            results.append('Worst')
            found = True
            break
    if found:
        continue
    # Priority 2: Tomato & FALSE
    for a, b in sub:
        if a == 'Tomato' and b is False:
            results.append('Ignore')
            found = True
            break
    if found:
        continue
    # Priority 3: Pickle & FALSE
    for a, b in sub:
        if a == 'Pickle' and b is False:
            results.append('Bad')
            found = True
            break
    if found:
        continue
    # Otherwise
    results.append('Good')

# Write to column C (3), rows 2-7
for i, value in enumerate(results, start=2):
    ws.cell(row=i, column=3, value=value)

# Save
wb.save(output_path)
