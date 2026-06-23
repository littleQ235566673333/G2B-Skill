from openpyxl import load_workbook
from datetime import datetime

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_7/regression_gate/after_pass/core_41589/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_7/regression_gate/after_pass/core_41589/output.xlsx'

wb = load_workbook(in_path)
ws = wb['Contact List']
today = datetime.today()

for row in ws.iter_rows(min_row=4, min_col=8, max_col=10):
    h = row[0].value  # Column H
    i = row[1].value  # Column I
    j_cell = row[2]  # Column J
    result = 'NO ACTION'
    if h and i and str(i).strip().upper() == 'YES':
        if isinstance(h, datetime):
            delta = (today - h).days
            if delta <= 30:
                result = 'HOLD'
            else:
                result = 'TOUCH BASE'
    j_cell.value = result

wb.save(out_path)
