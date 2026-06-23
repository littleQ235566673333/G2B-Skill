import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_58032_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_58032_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Create a new worksheet with the same data and formats as Sheet1
wb.copy_worksheet(ws)
ws_new = wb.worksheets[-1]
ws_new.title = 'Sheet1_TMP'

# Arial font for all cells
def set_arial_font(cell, bold=None, italic=None):
    cell.font = Font(name='Arial', size=cell.font.size if cell.font else 11,
                     bold=bold if bold is not None else (cell.font.bold if cell.font else False),
                     italic=italic if italic is not None else (cell.font.italic if cell.font else False))

for row in ws_new.iter_rows():
    for cell in row:
        if cell.value is not None:
            set_arial_font(cell)

# Row 1 titles: bold, italic, bordered
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
for cell in ws_new[1]:
    set_arial_font(cell, bold=True, italic=True)
    cell.border = border

# Column D: fill from D2 down with #CCCCCC
gray_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
for row in ws_new.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        cell.fill = gray_fill

# Formula for A2:A35, fill color #FCD5B4
peach_fill = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')
for i in range(2, 36):
    formula = '=IFERROR(INDEX($H$2:$H$100, MATCH(1, ($I$2:$I$100=B{row}) * ($J$2:$J$100=C{row}), 0)), "")'.format(row=i)
    cell = ws_new[f'A{i}']
    cell.value = formula
    cell.fill = peach_fill
    set_arial_font(cell)

# Hide gridlines
ws_new.sheet_view.showGridLines = False

# Remove original Sheet1, rename TMP
orig = wb['Sheet1']
del wb['Sheet1']
ws_new.title = 'Sheet1'

wb.save(output_path)
print('Done')
