from openpyxl import load_workbook
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_7/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_7/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']

last_contact = ws['H4'].value
regular_contact = ws['I4'].value

today = datetime.today()
action = 'NO ACTION'
if isinstance(last_contact, datetime) and regular_contact and regular_contact.strip().upper() == 'YES':
    if (today - last_contact).days <= 30:
        action = 'HOLD'
    else:
        action = 'TOUCH BASE'

ws['J4'] = action
wb.save(output_path)
