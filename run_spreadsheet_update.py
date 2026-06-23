from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/after_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Map headers and columns
headers = [cell.value for cell in ws[2]]
header_to_col = {name: idx + 1 for idx, name in enumerate(headers)}

CATEGORY_col = header_to_col['CATEGORY']
DOB_col = header_to_col['DATE OF BIRTH']
age_year_col = header_to_col['age (year)']

# Prepare 'Expected Result' (col E)
for row in range(3, 16):
    ws.cell(row=row, column=5).value = f'={get_column_letter(CATEGORY_col)}{row}'

# Insert 'Actual Age' right after 'age (year)'
actual_age_col = age_year_col + 1
ws.cell(row=2, column=actual_age_col, value='Actual Age')

# Get the report date reference (cell B1)
report_date_cell = ws['B1'].coordinate

for row in range(3, 16):
    age_cell = ws.cell(row=row, column=actual_age_col)
    age_cell.value = f'=ROUNDDOWN(($B$1-C{row})/365,0)'
    age_cell.number_format = '0'
    age_cell.alignment = Alignment(horizontal='center', vertical='center')
    age_cell.font = Font(bold=True)

# Safely match fill (color & style) from left col
left_cell = ws.cell(row=2, column=age_year_col)
left_fill = left_cell.fill
if isinstance(left_fill, PatternFill):
    fill_to_apply = PatternFill(
        fill_type=left_fill.fill_type,
        fgColor=left_fill.fgColor.rgb,
        bgColor=left_fill.bgColor.rgb)
else:
    fill_to_apply = PatternFill(fill_type=None)
for row in range(2, 16):
    ws.cell(row=row, column=actual_age_col).fill = fill_to_apply

# Top-align, center, bold header
hdr_cell = ws.cell(row=2, column=actual_age_col)
hdr_cell.alignment = Alignment(horizontal='center', vertical='top')
hdr_cell.font = Font(bold=True)

wb.save(output_path)
