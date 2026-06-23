import openpyxl
from datetime import datetime, timedelta, time

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_35739_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_35739_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

for row in range(2, 101):
    a = ws[f'A{row}'].value
    b = ws[f'B{row}'].value
    if a is None and b is None:
        ws[f'C{row}'].value = None
    elif b is not None:
        # Parse time
        std_time = None
        if isinstance(b, datetime):
            std_time = b.time()
        elif isinstance(b, time):
            std_time = b
        else:
            try:
                std_time = datetime.strptime(str(b), '%H:%M:%S').time()
            except Exception:
                ws[f'C{row}'].value = None
                continue
        # Calculate cutoff, cross midnight if needed
        base_date = datetime(2000,1,2) if std_time < time(0,30,0) else datetime(2000,1,1)
        full_dt = datetime.combine(base_date, std_time)
        cutoff_dt = full_dt - timedelta(minutes=30)
        ws[f'C{row}'].value = cutoff_dt.time()
    else:
        ws[f'C{row}'].value = None

wb.save(output_path)
