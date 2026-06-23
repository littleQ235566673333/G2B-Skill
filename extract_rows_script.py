import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_178-22_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_178-22_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws1 = wb['Sheet1']
rows = list(ws1.iter_rows(values_only=True))
header = rows[0]
filtered = []

for row in rows[1:]:
    b = row[1]
    c = row[2]
    if b == 'TELIVISION' or c == 'CLASS III' or c == 'CLASS IV':
        filtered.append(row)

# Remove Sheet2 if it exists
if 'Sheet2' in wb.sheetnames:
    wb.remove(wb['Sheet2'])
ws2 = wb.create_sheet('Sheet2')

# Insert blank row at the top so headers go in A2
ws2.insert_rows(1)

# Write headers to A2
for col, value in enumerate(header, 1):
    ws2.cell(row=2, column=col, value=value)

# Write filtered data starting from A3
for r_idx, row in enumerate(filtered, start=3):
    for c_idx, value in enumerate(row, start=1):
        ws2.cell(row=r_idx, column=c_idx, value=value)

wb.save(output_path)
