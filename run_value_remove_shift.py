import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_1/group_387-16/r0/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_1/group_387-16/r0/evolve_387-16/output.xlsx'

# Load the data into DataFrame
df = pd.read_excel(input_path, sheet_name='Sheet1', header=None)

# Find columns by header name
header_row = df.index[df.apply(lambda row: row.astype(str).str.contains('Value|Result values').any(), axis=1)][0]
headers = df.iloc[header_row].values.tolist()
df_data = pd.read_excel(input_path, sheet_name='Sheet1', header=header_row)

# Get correct column names
value_col = [c for c in df_data.columns if 'Value' == str(c).strip()][0]
binaries_col = [c for c in df_data.columns if 'Binaries' in str(c)][0]
result_values_col = [c for c in df_data.columns if 'Result values' in str(c)][0]
target_col = [c for c in df_data.columns if 'Target value' in str(c)][0]

# Process the list to remove just one instance for each match
values = df_data[value_col].tolist()
binaries = df_data[binaries_col].tolist()
result_values = df_data[result_values_col].tolist()

# Remove just one occurrence for each value in result_values from values and synchronize binaries
for key in result_values:
    if key in values:
        idx = values.index(key)
        values.pop(idx)
        binaries.pop(idx)

# Pad with empty strings to retain dataframe size
new_length = len(df_data)
values_new = values + [''] * (new_length - len(values))
binaries_new = binaries + [''] * (new_length - len(binaries))

df_data[value_col] = values_new
df_data[binaries_col] = binaries_new

# Remove blank cells and shift up (just non-empty entries only, then fill rest with blanks), for both columns
def shift_up(col_values):
    entries = [x for x in col_values if pd.notna(x) and x != '']
    return entries + [''] * (new_length - len(entries))

df_data[value_col] = shift_up(df_data[value_col])
df_data[binaries_col] = shift_up(df_data[binaries_col])

# Solver result is sum of col A
col_A = df_data[df_data.columns[0]].apply(pd.to_numeric, errors='coerce').fillna(0)
solver_result = col_A.sum()
target = df_data[target_col].iloc[0] if pd.notna(df_data[target_col].iloc[0]) else 0
diff = solver_result - target

df_data['Solver result'] = ''
df_data['Difference'] = ''
if len(df_data) > 0:
    df_data.at[0, 'Solver result'] = solver_result
    df_data.at[0, 'Difference'] = diff

# Load workbook for writing the results
wb = load_workbook(input_path)
ws = wb['Sheet1']
# Write new data back to sheet (A2:D18)
rows = df_data.iloc[:17, :4].values
for i, row in enumerate(rows, start=2):
    for j, v in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=v)
# Write Solver result and Difference
ws.cell(row=2, column=4, value=solver_result)
ws.cell(row=3, column=4, value=diff)

wb.save(output_path)
