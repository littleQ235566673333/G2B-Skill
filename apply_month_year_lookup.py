import openpyxl
import calendar

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_3/group_39515/r2/evolve_39515/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_3/group_39515/r2/evolve_39515/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Build header mapping: (month_abbr, year) -> column index
header_row = 1
col_map = {}
for col in range(3, 15):  # Columns C (3) to N (14)
    v = ws.cell(row=header_row, column=col).value
    if v is not None:
        try:
            m, y = v.strip().split()
            # Normalize month to 3-letter abbreviation, capitalized
            m = m[:3].capitalize()
            col_map[(m, y)] = col
        except Exception:
            pass

# For translating month names to abbr
full_to_abbr = {calendar.month_name[i].lower(): calendar.month_abbr[i] for i in range(1, 13)}
abbr3 = {calendar.month_abbr[i].lower(): calendar.month_abbr[i] for i in range(1, 13)}

def normalize_month(mstr):
    m = str(mstr).strip().lower()
    # Try direct match
    if m in full_to_abbr:
        return full_to_abbr[m]
    # Try if already abbr
    m3 = m[:3]
    if m3 in abbr3:
        return abbr3[m3]
    # Try capitalize first 3 letters
    return m3.capitalize()

# Fill values for O2:O13
for row in range(2, 14):
    mval = ws.cell(row=row, column=1).value
    yval = ws.cell(row=row, column=2).value
    if mval is not None and yval is not None:
        m3 = normalize_month(mval)
        col = col_map.get((m3, str(yval)))
        if col:
            ws.cell(row=row, column=15, value=ws.cell(row=row, column=col).value)
        else:
            ws.cell(row=row, column=15, value=None)  # No match found
    else:
        ws.cell(row=row, column=15, value=None)

wb.save(output_path)
