import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_191-40_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_191-40_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']
max_col = 8
max_row = 85

def is_empty_row(row):
    return all((ws.cell(row=row, column=c).value in (None, '')) for c in range(1, max_col+1))

def is_formula(val):
    return isinstance(val, str) and val.startswith('=')

# Convert common text-formula appearances to formulas
def text_to_formula(val):
    if not isinstance(val, str):
        return None
    m = re.match(r'^=\"?(.+?)\"?$', val)
    if m:
        return f'={m.group(1)}'
    if val.startswith('='):
        return val
    return None

blocks = []
r = 1
while r <= max_row:
    while r <= max_row and is_empty_row(r):
        r += 1
    block_start = r
    while r <= max_row and not is_empty_row(r):
        r += 1
    block_end = r - 1
    if block_start <= block_end:
        blocks.append((block_start, block_end))
    r += 1

col_widths = {i: ws.column_dimensions[get_column_letter(i)].width for i in range(1, max_col+1)}

data_wb = openpyxl.load_workbook(input_path, data_only=True)
data_ws = data_wb['Sheet1']

rows_output = []
for block_start, block_end in blocks:
    block_data = []
    for real_row_idx, r in enumerate(range(block_start, block_end+1)):
        row_cells = []
        for c in range(1, max_col+1):
            val = ws.cell(row=r, column=c).value
            if c in (4, 5, 6):
                if isinstance(val, str) and not is_formula(val):
                    fval = text_to_formula(val)
                    row_cells.append(fval if fval is not None else val)
                else:
                    row_cells.append(val)
            else:
                row_cells.append(val)
        block_data.append((r, row_cells))  # Store worksheet row for lookup
    # Sort by evaluated value in D (col index 3)
    def d_value(row_info):
        ws_row, _ = row_info
        val = data_ws.cell(row=ws_row, column=4).value
        try:
            return float(val) if val is not None else float('-inf')
        except Exception:
            return float('-inf')
    block_data.sort(key=d_value, reverse=True)
    # Only keep row data (strip out real worksheet row info)
    for real_row, row_vals in block_data:
        rows_output.append(row_vals)
    if (block_start, block_end) != blocks[-1]:
        rows_output.append(['']*max_col)

for r in range(1, max_row+1):
    for c in range(1, max_col+1):
        ws.cell(row=r, column=c).value = None

for i, row in enumerate(rows_output):
    if i >= max_row:
        break
    for j, val in enumerate(row):
        cell = ws.cell(row=i+1, column=j+1)
        if j+1 in (4,5,6) and isinstance(val, str) and is_formula(val):
            cell.value = val
        else:
            cell.value = val

# Number format and style for D–F
for r in range(1, max_row+1):
    for c in range(1, max_col+1):
        cell = ws.cell(row=r, column=c)
        if c in (4,5,6):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            value = cell.value
            if ws.cell(row=r, column=c).value is not None:
                if isinstance(value, (int, float)):
                    if abs(value - round(value)) < 1e-7:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.0'
        # Other columns, leave formatting untouched
for i, w in col_widths.items():
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(output_path)
