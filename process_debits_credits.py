import openpyxl

# File paths
i_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_469-9_tc1/input.xlsx'
o_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_469-9_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(i_file)
ws = wb.active

# Add headers
ws['H1'] = 'Debits'
ws['I1'] = 'Credits'

# Process rows 2-10 (since H1:I10 required for answer)
for row in range(2, 11):  # Rows 2 to 10
    value = ws[f'C{row}'].value
    debit, credit = '', ''
    if value is not None:
        try:
            val = float(value)
            if val < 0:
                debit = abs(val)
            elif val > 0:
                credit = abs(val)
        except (ValueError, TypeError):
            pass  # Skip non-numeric values
    ws[f'H{row}'] = debit
    ws[f'I{row}'] = credit

# Save workbook
wb.save(o_file)
