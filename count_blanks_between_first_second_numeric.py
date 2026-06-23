from openpyxl import load_workbook

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/group_50521/r1/evolve_50521/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/group_50521/r1/evolve_50521/output.xlsx'
wb = load_workbook(in_path)
ws = wb.active

for out_row in range(4, 7):
    values = [ws.cell(row=out_row, column=col).value for col in range(2, 14)]
    num_indices = [idx for idx, val in enumerate(values) if isinstance(val, (int, float))]
    if not num_indices or len(num_indices) < 2:
        result = None
    else:
        if values[num_indices[0]] > 1:
            result = 1
        else:
            start = num_indices[0]
            end = num_indices[1]
            blank_count = sum(1 for v in values[start+1:end] if v is None or v == "")
            result = blank_count
    ws.cell(row=out_row, column=14, value=result)

wb.save(out_path)
