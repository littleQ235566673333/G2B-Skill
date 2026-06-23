import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_50971_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_50971_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

df = pd.read_excel(input_path)

duplicated_a = df[df.duplicated('COLLM', keep=False)].copy()
duplicated_a_sorted = duplicated_a.sort_values('COLLM')
duplicated_groups = duplicated_a_sorted.groupby('COLLM')['Value'].apply(list)

unique_as = [ws[f'F{row}'].value for row in range(3, 14)]  # F3:F13
output_values = []
for a in unique_as:
    if a in duplicated_groups:
        output_values.extend(duplicated_groups[a])

for idx in range(11):
    cell = f'G{3+idx}'
    ws[cell] = output_values[idx] if idx < len(output_values) else None
wb.save(output_path)
