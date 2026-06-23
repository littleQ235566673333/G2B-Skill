import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_1/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_1/regression_gate/after_pass/core_50526/output.xlsx'

# Load workbook & active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read lookup value from B6
lookup_value = ws['B6'].value

# Get headers (skip col A)
headers = [ws.cell(row=1, column=col).value for col in range(2, ws.max_column+1)]

# Find row with A==lookup_value (skip header row)
rownum = None
for row in range(2, ws.max_row+1):
    if ws.cell(row=row, column=1).value == lookup_value:
        rownum = row
        break

result = []
if rownum:
    # In found row, for each col B+, if value>0, collect header
    for col in range(2, ws.max_column+1):
        val = ws.cell(row=rownum, column=col).value
        if isinstance(val, (int, float)) and val > 0:
            result.append(ws.cell(row=1, column=col).value)

# Write outputs in B9, B10
for idx in range(2):
    ws.cell(row=9+idx, column=2, value=result[idx] if idx < len(result) else None)

wb.save(output_path)
