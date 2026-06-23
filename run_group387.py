import openpyxl
import numpy as np

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_2/group_387-16/r1/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_2/group_387-16/r1/evolve_387-16/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# 'Value': column A starting at row 3, up to first empty
# 'Binaries': column B starting at row 3, up to first empty
# 'Target value': cell E2
values = []
binaries = []
row = 3
while True:
    v = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    if v is None and b is None:
        break
    values.append(v)
    binaries.append(b)
    row += 1

n = len(values)
# Result values start at D5 downward (label is D4)
result_values = []
row = 5
while True:
    rv = ws.cell(row=row, column=4).value
    if rv is None:
        break
    result_values.append(rv)
    row += 1

# Target value: E2
try:
    target_value = float(ws['E2'].value)
except:
    target_value = None

# Remove only one instance for each match
values_mask = [True] * n
binaries_mask = [True] * n
for rv in result_values:
    # find first occurrence
    for idx, v in enumerate(values):
        if values_mask[idx] and v == rv:
            values_mask[idx] = False  # mark for removal
            binaries_mask[idx] = False
            break

# Build compacted lists
values_kept = [v for v, keep in zip(values, values_mask) if keep]
binaries_kept = [b for b, keep in zip(binaries, binaries_mask) if keep]

# Pad to original length
while len(values_kept) < n:
    values_kept.append(None)
while len(binaries_kept) < n:
    binaries_kept.append(None)

# Write back A3 and B3 down
for i in range(n):
    ws.cell(row=3+i, column=1).value = values_kept[i]
    ws.cell(row=3+i, column=2).value = binaries_kept[i]

# Solver result: sum of non-None values in A3:A(3+n-1)
solver_result = sum([v for v in values_kept if isinstance(v, (int, float)) and v is not None])
if target_value is not None:
    difference = solver_result - target_value
else:
    difference = None

# Write results: 'Solver result' to H2, difference to K2
ws['H2'] = solver_result
ws['K2'] = difference

wb.save(output_path)
