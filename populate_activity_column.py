import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_33157_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_33157_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Activity columns and their headers
activity_cols = ['B', 'D', 'F', 'H']
activity_headers = [ws[f'{col}1'].value for col in activity_cols]

for row in range(2, 7):  # for rows 2 to 6 as required
    date = ws[f'J{row}'].value
    found = False
    for idx, col in enumerate(activity_cols):
        if ws[f'{col}{row}'].value == date:
            ws[f'K{row}'] = activity_headers[idx]
            found = True
            break
    if not found:
        ws[f'K{row}'] = None  # If no match is found, set as None

wb.save(output_path)
