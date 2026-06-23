from openpyxl import load_workbook
from datetime import datetime, timedelta
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_4/group_45707/r2/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_4/group_45707/r2/evolve_45707/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

dates = []
c_values = []
for row in range(2, 70):
    date_val = ws[f'A{row}'].value
    c_val = ws[f'C{row}'].value
    dates.append(date_val)
    c_values.append(c_val)

ones_month_year = defaultdict(int)
for d, c in zip(dates, c_values):
    if isinstance(d, datetime):
        key = (d.year, d.month)
        if c == 1:
            ones_month_year[key] += 1

for i in range(len(dates)):
    if not isinstance(dates[i], datetime):
        ws[f'D{i+2}'].value = None
        continue
    next_day = dates[i] + timedelta(days=1)
    if next_day.day == 1:
        key = (next_day.year, next_day.month)
        count_ones = ones_month_year.get(key, 0)
        ws[f'D{i+2}'].value = count_ones
    else:
        ws[f'D{i+2}'].value = None

wb.save(output_path)
