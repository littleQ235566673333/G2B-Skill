from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_8/regression_gate/after_fix/core_39432/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_8/regression_gate/after_fix/core_39432/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Helper: Get (units, cost) for all batches for a given row
# (Beginning, then each PO if present)
def get_batches(row):
    batches = []
    # Beginning Inventory
    begin_units = row[1]
    begin_cost = row[2]
    if begin_units is not None and begin_cost is not None:
        batches.append((begin_units, begin_cost))
    # PO Columns: pairs start at col 4 (index 3)
    for i in range(3, 11, 2):
        units = row[i]
        cost = row[i+1]
        if units is not None and cost is not None:
            batches.append((units, cost))
    return batches

def fifo_current_cost(batches, current_inventory):
    '''
    batches: List of (units, cost), oldest-to-newest
    current_inventory: Number of units on hand (int)
    Returns: weighted average cost of last `current_inventory` units using FIFO
    '''
    # Traverse batches from last to first (LIFO) to accumulate the last-in inventory
    total_units = 0
    total_cost = 0.0
    needed = current_inventory
    # Go backward
    for units, cost in reversed(batches):
        take = min(units, needed)
        total_units += take
        total_cost += take * cost
        needed -= take
        if needed == 0:
            break
    if total_units == 0:
        return None
    return round(total_cost/total_units, 4)

# Process rows 2 to 5 (1-based, so 2~5)
for i in range(2, 6):
    row = [ws.cell(row=i, column=j).value for j in range(1, 13)]
    batches = get_batches(row)
    current_inventory = ws.cell(row=i, column=12).value # Col L
    if current_inventory is None or not batches:
        ws.cell(row=i, column=13).value = None
        continue
    cost = fifo_current_cost(batches, int(current_inventory))
    ws.cell(row=i, column=13).value = cost

wb.save(output_path)
print('FIFO per unit cost calculated and written to', output_path)