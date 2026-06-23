import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_38969_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_38969_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 21):
    s_val = ws[f'S{row}'].value
    u_val = ws[f'U{row}'].value
    result = None

    # If S2 = error (literal string "error")
    if isinstance(s_val, str) and s_val.strip().lower() == 'error':
        result = 'Upload'
    else:
        try:
            u_val_num = float(u_val)
            if -1 <= u_val_num <= 1:
                result = 'Do not Upload'
            elif u_val_num > 1:
                result = 'to Check'
            elif u_val_num < -1:
                result = 'to Check'
        except (TypeError, ValueError):
            # u_val is not numeric so no action
            result = None

    ws[f'R{row}'] = result

wb.save(output_path)
