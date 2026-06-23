from openpyxl import load_workbook

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_3/group_118-50/r3/evolve_118-50/input.xlsx')
ws = wb['Sheet1']
with open('pairs_sorted.txt') as f:
    lines = [l.strip().split('\t') for l in f if l.strip()]
for idx, line in enumerate(lines):
    ws.cell(row=idx+2, column=3).value = line[0]
    ws.cell(row=idx+2, column=4).value = line[1]
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_3/group_118-50/r3/evolve_118-50/output.xlsx')
