import openpyxl

def count_instances_by_header(input_path, output_path, header_title, value_to_count):
    # Load workbook and sheet
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Find the header row (assume it's the first row)
    header_row = 1
    headers = [cell.value for cell in ws[header_row]]
    
    # Identify column index matching header_title
    try:
        col_idx = headers.index(header_title) + 1  # openpyxl columns are 1-based
    except ValueError:
        raise Exception(f"Header title '{header_title}' not found.")

    # Count instances in the column (excluding header)
    count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col_idx).value
        if cell_value == value_to_count:
            count += 1

    # Write the result to cell I3
    ws['I3'] = count

    # Save the workbook to desired output_path
    wb.save(output_path)

# Set parameters
def main():
    input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_46897_tc1/input.xlsx'
    output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_46897_tc1/output.xlsx'
    header_title = 'YourHeader'      # <-- Set your header title here
    value_to_count = 'YourValue'    # <-- Set the text or number you want to count
    
    count_instances_by_header(input_path, output_path, header_title, value_to_count)

if __name__ == '__main__':
    main()
