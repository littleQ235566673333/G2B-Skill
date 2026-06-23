import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_45372_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_45372_tc1/output.xlsx'

# Open the workbook and the first worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Setup fill color for E2:E15
fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

# Target time for comparison
switch_time = datetime.strptime('09:45', '%H:%M').time()

for row in range(2, 16):  # E2:E15, corresponds to rows 2-15, inclusive
    time_cell = ws[f'D{row}']
    if time_cell.value is None:
        ws[f'E{row}'].value = None
        ws[f'E{row}'].fill = fill
        continue

    # Handle time in column D, which might be string or datetime.time
    time_val = time_cell.value
    if isinstance(time_val, str):
        try:
            time_val = datetime.strptime(time_val.strip(), '%H:%M').time()
        except Exception:
            ws[f'E{row}'].value = None
            ws[f'E{row}'].fill = fill
            continue
    
    if time_val < switch_time:
        val = ws[f'B{row}'].value
    else:
        val = ws[f'C{row}'].value

    # Show blank if 0 or None
    ws[f'E{row}'].value = '' if (val == 0 or val is None) else val
    # Apply fill
    ws[f'E{row}'].fill = fill

# Save the modified workbook
wb.save(output_path)
