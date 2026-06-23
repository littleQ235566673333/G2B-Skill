from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from datetime import datetime, timedelta

infile = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed0/eval_35739_tc1/input.xlsx'
outfile = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed0/eval_35739_tc1/output.xlsx'
wb = load_workbook(infile)
ws = wb['Sheet1']
# Use header mapping
header_row = 1
headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[header_row]]
def col(name):
    name = name.strip().lower()
    return headers.index(name) + 1 if name in headers else None
col_date = col('date')
col_std = col('std')
col_cutoff = col('cut off')

for i in range(2, 101):
    date_val = ws.cell(row=i, column=col_date).value
    std_val = ws.cell(row=i, column=col_std).value
    cut_cell = ws.cell(row=i, column=col_cutoff)
    if not date_val or not std_val:
        cut_cell.value = None
        continue
    
    bv = std_val
    time_ok = False
    # Accept 'HH:MM' as string, or Excel time as float
    if isinstance(bv, str):
        try:
            bv = datetime.strptime(bv.strip(), '%H:%M').time()
            time_ok = True
        except: 
            cut_cell.value = None
            continue
    elif isinstance(bv, (float, int)):
        try:
            bv = from_excel(bv)
            if isinstance(bv, datetime):
                bv = bv.time()
                time_ok = True
        except:
            cut_cell.value = None
            continue
    if not time_ok:
        cut_cell.value = None
        continue
    # Calculate cutoff, handle midnight
    dt = datetime.combine(datetime(2000,1,1), bv)
    cutoff = dt - timedelta(minutes=30)
    if cutoff.day < dt.day:
        cutoff_time = (cutoff + timedelta(days=1)).time()
    else:
        cutoff_time = cutoff.time()
    cut_cell.value = cutoff_time.strftime('%H:%M')

wb.save(outfile)
