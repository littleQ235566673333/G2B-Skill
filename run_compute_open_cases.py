import openpyxl

INPUT = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed42_rerun1_s0/eval_41978_tc1/input.xlsx'
OUTPUT = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed42_rerun1_s0/eval_41978_tc1/output.xlsx'

wb = openpyxl.load_workbook(INPUT)
sheet = wb.active

data = []
years = set()
first_data_row = 14
last_data_row = 185
for row in range(first_data_row, last_data_row+1):
    year = sheet.cell(row=row, column=7).value  # Column G
    status = sheet.cell(row=row, column=10).value  # Column J
    if year is not None:
        years.add(year)
    data.append({'row': row, 'year': year, 'status': status})

# Get sorted unique years (as strings for consistency)
years = sorted([y for y in years if y is not None], key=lambda x: str(x))

# Count 'open' for each year
open_counts = {}
for y in years:
    open_counts[y] = sum(
        (str(r['status']).strip().lower() == 'open' and str(r['year']) == str(y))
        for r in data
    )

# Write results to I2:I11 (col 9)
for i, y in enumerate(years[:10]):
    target_cell = sheet.cell(row=2+i, column=9)
    target_cell.value = open_counts[y]

wb.save(OUTPUT)
