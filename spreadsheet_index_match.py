import openpyxl

input_file = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_12864_tc1/input.xlsx'
output_file = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_12864_tc1/output.xlsx'

# Load workbook and sheets
wb = openpyxl.load_workbook(input_file)
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']

# Build lookup dict from Sheet2: Column B and C (assuming header in row 1)
data_lookup = {}
for row in sheet2.iter_rows(min_row=2, max_col=3):
    key = row[0].value  # Column B
    value_c = row[1].value  # Column C
    data_lookup[key] = (row[0].value, value_c)  # (date, associated data)

# Get the values to match from Sheet1 Column D (assuming header in row 1)
match_values = [sheet1.cell(row=row, column=4).value for row in range(2, 13)]

results = []
for val in match_values:
    # Find corresponding date (col B) and associated data (col C) in Sheet2
    result = data_lookup.get(val, (None, None))
    results.append(result)

# Write results in Sheet2 B2:B12
for idx, (date, data) in enumerate(results, start=2):
    sheet2.cell(row=idx, column=2).value = date
    sheet2.cell(row=idx, column=3).value = data

wb.save(output_file)
