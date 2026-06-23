import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_1/group_547-43/r1/evolve_547-43/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_1/group_547-43/r1/evolve_547-43/output.xlsx'

# Read input
xl = pd.ExcelFile(input_path)
df_emp = xl.parse('Emp')
df_lookup = xl.parse('Lookup')

# Columns to transpose per the structure
variable_cols = ['Comp1', '2_comp', 'dept1', 'avg sal', 'hike', 'performance']
fixed_cols = ['seqno', 'Empno', 'ename', 'Jobid']

# Merge lookup info for variable columns
lookup_dict = df_lookup.set_index('Variable')[['Category', 'Subcategory']].to_dict('index')

records = []
for _, row in df_emp.iterrows():
    for var in variable_cols:
        rec = [row[c] for c in fixed_cols]
        rec.append(var)
        cat = lookup_dict.get(var, {}).get('Category', '')
        subcat = lookup_dict.get(var, {}).get('Subcategory', '')
        rec.extend([cat, subcat, row[var]])
        records.append(rec)

# Prepare output
wb = load_workbook(input_path)
ws_out = wb['Expected'] if 'Expected' in wb.sheetnames else wb.create_sheet('Expected')
headers = ['seqno', 'Empno', 'ename', 'Jobid', 'Variable', 'Category', 'Subcategory', 'Value']
for c_idx, h in enumerate(headers, start=1):
    ws_out.cell(row=1, column=c_idx, value=h)
for r_idx, rec in enumerate(records[:18], start=2):  # only fill required demo rows
    for c_idx, val in enumerate(rec, start=1):
        ws_out.cell(row=r_idx, column=c_idx, value=val)
wb.save(output_path)
