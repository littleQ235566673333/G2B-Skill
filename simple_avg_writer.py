from openpyxl import load_workbook
import re
from openpyxl.utils import get_column_letter

def normalize_header(header):
    return re.sub(r'\s+', '', str(header)).lower() if header else ''

in_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed2/eval_183-8_tc1/input.xlsx'
out_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed2/eval_183-8_tc1/output.xlsx'
wb = load_workbook(in_path)
ws = wb['Sheet1']

header_row = 1
row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
headers = [normalize_header(cell) for cell in row]

# Find first three populated (non-blank header) data columns
col_targets = []
raw_names = []
for idx, h in enumerate(headers):
    if h != '':
        col_targets.append(idx+1)  # openpyxl col index is 1-based
        raw_names.append(row[idx])
    if len(col_targets) == 3:
        break
# Output goes in J, K, L (cols 10, 11, 12)
output_cols = [10, 11, 12]
data_start = header_row + 1
# Scan down to find last data row (column C must be filled, i.e., col_targets[0]).
data_end = data_start
while True:
    v = ws.cell(row=data_end, column=col_targets[0]).value
    if v is None or (isinstance(v, str) and v.strip()==''):
        break
    data_end += 1
data_end -= 1
# Write the simple (unweighted) average formula for these columns.
for out_idx, tgt_idx in enumerate(col_targets):
    col_letter = get_column_letter(tgt_idx)
    rng = f"${col_letter}${data_start}:${col_letter}${data_end}"
    formula = f"=AVERAGE({rng})"
    for r in range(3, 7):  # J3:J6, K3:K6, L3:L6
        ws.cell(row=r, column=output_cols[out_idx]).value = formula
wb.save(out_path)
