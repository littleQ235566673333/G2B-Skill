import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_35739_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_35739_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

def correct_cutoff_time(std_time):
    # std_time is a time or datetime.time instance
    # Subtract 30 mins, handle midnight crossover
    dummy_date = datetime(2000, 1, 1, std_time.hour, std_time.minute, std_time.second)
    cutoff_dt = dummy_date - timedelta(minutes=30)
    return cutoff_dt.time()

for row in range(2, 101):
    a_val = ws[f'A{row}'].value
    b_val = ws[f'B{row}'].value
    
    if a_val is None or b_val is None:
        ws[f'C{row}'].value = None
    else:
        # If datetime, extract .time()
        if isinstance(b_val, datetime):
            std_time = b_val.time()
        else:
            std_time = b_val if hasattr(b_val, 'hour') else None
        if std_time:
            cutoff_time = correct_cutoff_time(std_time)
            ws[f'C{row}'].value = cutoff_time
        else:
            ws[f'C{row}'].value = None

wb.save(output_path)
