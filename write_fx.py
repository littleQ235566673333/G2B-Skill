from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2-PRUNED/eval_seed42/eval_42515_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2-PRUNED/eval_seed42/eval_42515_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

start_output_row = 5
end_output_row = 19

# Place formulas in F5:F19
# Each one references B5 and B8:B22. For F5, use B8; F6 -> B9; ..., F19 -> B22
for row in range(start_output_row, end_output_row+1):
    # INDEX($B$8:$B$22, row-4) gives B8 for row=5, B9 for row=6, ..., B22 for row=19
    formula = f"=SUM(1/((1/B{row})*INDEX($B$8:$B$22,{row-4})))"
    ws[f'F{row}'] = formula

wb.save(output_path)
