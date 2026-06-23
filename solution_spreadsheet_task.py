import openpyxl
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_7/regression_gate/after_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_7/regression_gate/after_pass/core_55421/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
COL_A = 1
COL_D = 4
COL_E = 5
COL_F = 6
status_data = defaultdict(list)
row_range = range(2, 21)
# Collect all statuses and date info for each number in Column A
for row in row_range:
    a_val = ws.cell(row=row, column=COL_A).value
    d_val = ws.cell(row=row, column=COL_D).value
    e_val = ws.cell(row=row, column=COL_E).value
    status_data[a_val].append((row, d_val, e_val))
# Now assign the correct output to F2:F20
for row in row_range:
    a_val = ws.cell(row=row, column=COL_A).value
    d_val = ws.cell(row=row, column=COL_D).value
    e_val = ws.cell(row=row, column=COL_E).value
    all_status = [item[1] for item in status_data[a_val]]
    sched_only = all(x=='SCH' for x in all_status)
    has_sch = 'SCH' in all_status
    has_ns = 'NO SHOW' in all_status
    if sched_only and has_sch:
        ws.cell(row=row, column=COL_F).value = 'FUTURE'
    elif has_sch and has_ns:
        ws.cell(row=row, column=COL_F).value = 'NS/SCHED'
    elif d_val == 'NO SHOW':
        if e_val:
            ws.cell(row=row, column=COL_F).value = 'NO ACTION NEEDED'
        else:
            ws.cell(row=row, column=COL_F).value = 'CALL PT'
wb.save(output_path)
