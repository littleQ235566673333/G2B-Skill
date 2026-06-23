import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_3/regression_gate/after_pass/core_84-40/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_3/regression_gate/after_pass/core_84-40/output.xlsx'

wb = openpyxl.load_workbook(input_path)
list_sheet_name = None
for s in wb.sheetnames:
    if s.strip().lower() == 'list':
        list_sheet_name = s
        break
if not list_sheet_name:
    list_sheet = wb.create_sheet('LIST')
else:
    list_sheet = wb[list_sheet_name]
    # Clear existing data in C2:E (all rows except header)
    for row in list_sheet['C2:E'+str(list_sheet.max_row)]:
        for cell in row:
            cell.value = None
# Identify sheets before the 'LIST' sheet
list_index = wb.sheetnames.index(list_sheet.title)
data_sheets = wb.sheetnames[:list_index]
results = []
for sname in data_sheets:
    ws = wb[sname]
    sum_c = 0
    sum_d = 0
    for r in range(2, ws.max_row + 1):
        v_c = ws.cell(row=r, column=3).value
        v_d = ws.cell(row=r, column=4).value
        if isinstance(v_c, (int, float)):
            sum_c += v_c
        if isinstance(v_d, (int, float)):
            sum_d += v_d
    results.append((sname, sum_c, sum_d, sum_c - sum_d))
# Write results to 'LIST' sheet, starting at C2, D2, E2
start_row = 2
for idx, (sname, sum_c, sum_d, diff) in enumerate(results):
    list_sheet.cell(row=start_row + idx, column=2, value=sname)
    list_sheet.cell(row=start_row + idx, column=3, value=sum_c)
    list_sheet.cell(row=start_row + idx, column=4, value=sum_d)
    list_sheet.cell(row=start_row + idx, column=5, value=diff)
# Add totals row
if results:
    total_c = sum(r[1] for r in results)
    total_d = sum(r[2] for r in results)
    total_diff = sum(r[3] for r in results)
    total_row = start_row + len(results)
    list_sheet.cell(row=total_row, column=2, value='Total')
    list_sheet.cell(row=total_row, column=3, value=total_c)
    list_sheet.cell(row=total_row, column=4, value=total_d)
    list_sheet.cell(row=total_row, column=5, value=total_diff)
wb.save(output_path)
