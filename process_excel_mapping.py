import openpyxl

# Load the workbook and the first sheet
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_472-15_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_472-15_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read the value in A1
a1_value = str(ws["A1"].value)

# Map the values to numbers
mapping = {
    '4Ozark': 1,
    '3Tall': 2,
    '1Jasper': 3,
    '2GWood': 4,
    '5Dawson': 5,
    '8CPark': 6
}

# Default to blank unless match found
result = mapping.get(a1_value, "")
ws["B2"] = result

# Save to output file
wb.save(output_path)
