import openpyxl
from collections import defaultdict

# Input and output paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_250-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_250-20_tc1/output.xlsx'

def process_rnm_sheet(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb['RNM']
    rows = list(ws.iter_rows(min_row=1, max_col=10, values_only=True))

    header = rows[0]
    data_rows = rows[1:]

    # Dictionary to group by (B, C) key and sum J
    grouped = defaultdict(lambda: [None]*10)
    for row in data_rows:
        key = (row[1], row[2])  # Columns B & C
        if grouped[key][0] is None:
            grouped[key] = list(row)
        else:
            grouped[key][9] += row[9]  # Sum column J (index 9)

    # Prepare result: header + one row per unique (B, C)
    result_rows = [header] + list(grouped.values())

    # Write back to the first 20 rows (A1:J20)
    out_ws = wb.copy_worksheet(ws)
    out_ws.title = 'RNM'
    for i in range(20):
        for j in range(10):
            val = result_rows[i][j] if i < len(result_rows) else None
            out_ws.cell(row=i+1, column=j+1, value=val)
    wb.save(output_path)

process_rnm_sheet(input_path, output_path)
