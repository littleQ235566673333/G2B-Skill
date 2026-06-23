import pandas as pd
from openpyxl import load_workbook
import string

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_7/regression_gate/after_fix/core_56274/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_7/regression_gate/after_fix/core_56274/output.xlsx'

sheets = pd.read_excel(input_path, sheet_name=None)
# Find sheet and columns
lookup_sheet = None
fiscal_col = None
for sheet_name, df in sheets.items():
    for col in df.columns:
        if 'Fiscal Month' in str(col):
            lookup_sheet = df
            fiscal_col = col
            break
    if lookup_sheet is not None:
        break
if lookup_sheet is None:
    raise Exception('No Fiscal Month column found!')

open_col = debits_col = credits_col = close_col = None
for col in lookup_sheet.columns:
    col_l = col.lower()
    if 'opening' in col_l:
        open_col = col
    if 'debit' in col_l:
        debits_col = col
    if 'credit' in col_l:
        credits_col = col
    if 'closing' in col_l:
        close_col = col
if not all([open_col, debits_col, credits_col, close_col]):
    raise Exception('Some columns are missing: found', open_col, debits_col, credits_col, close_col)
cols = list(lookup_sheet.columns)

def idx_to_excel(idx):
    # idx is 1-based
    if idx <= 26:
        return string.ascii_uppercase[idx-1]
    else:
        idx -= 1
        return string.ascii_uppercase[idx // 26 - 1] + string.ascii_uppercase[idx % 26]

sheetname = [k for k,v in sheets.items() if v is lookup_sheet][0]
if ' ' in sheetname or '-' in sheetname:
    sheet_disp = f"'{sheetname}'"
else:
    sheet_disp = sheetname
# Find 1-based column indices
fiscal_idx = cols.index(fiscal_col)+1
open_idx = cols.index(open_col)+1
debits_idx = cols.index(debits_col)+1
credits_idx = cols.index(credits_col)+1
close_idx = cols.index(close_col)+1

# Target formulas
formulas = [
    f"=XLOOKUP(D7,{sheet_disp}!{idx_to_excel(fiscal_idx)}:{idx_to_excel(fiscal_idx)},{sheet_disp}!{idx_to_excel(open_idx)}:{idx_to_excel(open_idx)},\"\",0)",
    f"=XLOOKUP(D7,{sheet_disp}!{idx_to_excel(fiscal_idx)}:{idx_to_excel(fiscal_idx)},{sheet_disp}!{idx_to_excel(debits_idx)}:{idx_to_excel(debits_idx)},\"\",0)",
    f"=XLOOKUP(D7,{sheet_disp}!{idx_to_excel(fiscal_idx)}:{idx_to_excel(fiscal_idx)},{sheet_disp}!{idx_to_excel(credits_idx)}:{idx_to_excel(credits_idx)},\"\",0)",
    f"=XLOOKUP(D7,{sheet_disp}!{idx_to_excel(fiscal_idx)}:{idx_to_excel(fiscal_idx)},{sheet_disp}!{idx_to_excel(close_idx)}:{idx_to_excel(close_idx)},\"\",0)",
]

wb = load_workbook(input_path)
ws = wb.active
ws['D9'] = formulas[0]
ws['D10'] = formulas[1]
ws['D11'] = formulas[2]
ws['D12'] = formulas[3]
wb.save(output_path)
