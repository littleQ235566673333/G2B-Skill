import openpyxl

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_52541_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_52541_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Process rows J6:J10
for row in range(6, 11):
    amt_outstanding = ws[f'H{row}'].value
    overdue_days = ws[f'I{row}'].value
    # Convert amount outstanding to float if possible
    try:
        amt_out = float(amt_outstanding)
    except (TypeError, ValueError):
        amt_out = None
    # Convert overdue days to integer if possible
    try:
        od_days = int(overdue_days)
    except (TypeError, ValueError):
        od_days = None

    if amt_out is not None and amt_out < 0:
        remark = 'Prepaid'
    elif od_days is None:
        remark = ''
    elif od_days < 90:
        remark = 'Call Customer'
    else:
        remark = 'Bad Debts'
    ws[f'J{row}'] = remark

wb.save(output_path)
