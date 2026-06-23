import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_50768_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_50768_tc1/output.xlsx'

# Header is on row 2 (index 2); first three columns are relevant.
df = pd.read_excel(input_path, header=2, usecols=[0, 1, 2])
rid_col = 'ID'
impact_col = 'Impact'
likelihood_col = 'Liklihood'  # typo matches sheet

# Drop fully blank rows and any risk rows missing any key

data = df.dropna(subset=[rid_col, impact_col, likelihood_col])

# Step 2: Create 3x3 matrix (Impact 1-3 y, Likelihood 1-3 x)
matrix = [[[] for _ in range(3)] for _ in range(3)]
for _, row in data.iterrows():
    try:
        impact = int(row[impact_col])
        likelihood = int(row[likelihood_col])
        risk_id = str(row[rid_col])
    except Exception as e:
        continue
    if 1 <= impact <= 3 and 1 <= likelihood <= 3:
        matrix[3 - impact][likelihood - 1].append(risk_id)  # Invert impact for row position

# Step 3: Write output to output.xlsx, cells F12:H14
wb = load_workbook(input_path)
ws = wb.active
row_offset = 12
col_offset = 6 # F=6, G=7, H=8
for i in range(3):
    for j in range(3):
        val = ', '.join(matrix[i][j]) if matrix[i][j] else ''
        ws.cell(row=row_offset + i, column=col_offset + j, value=val)

wb.save(output_path)
