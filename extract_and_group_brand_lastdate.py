import pandas as pd
from openpyxl import load_workbook

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_3/regression_gate/before_fix/core_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_3/regression_gate/before_fix/core_516-46/output.xlsx'
sheet_name = 'ورقة1'

# Load data starting at A2 (A2:E*)
df = pd.read_excel(input_path, sheet_name=sheet_name, header=0, usecols='A:E')
df = df.dropna(how='all', subset=None)

# Find last date
latest_date = df.iloc[:,0].max()

# Subset for last date only
df_last = df[df.iloc[:,0] == latest_date]

# Group by brand, summing quantity (col E) on last date
brand_col = df.columns[1]
qty_col = df.columns[-1]
group_cols = [brand_col]
df_grp = df_last.groupby(group_cols, as_index=False)[qty_col].sum()
df_grp.insert(0, df.columns[0], latest_date)

# Track duplications (modifications for duplicated last dates)
modifications = {}
for brand in df_last[brand_col].unique():
    entries = df_last[df_last[brand_col] == brand]
    if len(entries) > 1:
        modifications[brand] = entries[qty_col].tolist()

# Write output into input file location H2:L4 (columns H-L, rows 2/3/4)
wb = load_workbook(input_path)
ws = wb[sheet_name]
# Headers in H1:L1, data in H2:L4
for idx, h in enumerate(df.columns):
    ws.cell(row=1, column=8+idx).value = h
for i, row in enumerate(df_grp.values):
    for j, v in enumerate(row):
        ws.cell(row=2+i, column=8+j).value = v
# Write modification notes in next cell/row after values
row_offset = 2 + len(df_grp)
col_note = 8+len(df_grp.columns)
if modifications:
    for k, (brand, vals) in enumerate(modifications.items()):
        ws.cell(row=row_offset+k, column=col_note, value=f'Modified for {brand}: ' + ','.join(map(str, vals)))
wb.save(output_path)
