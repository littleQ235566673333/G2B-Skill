from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r1/eval_57743_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r1/eval_57743_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Build a dict of model number (C) to price (D)
model_to_price = {}
for row in range(2, ws.max_row + 1):
    model = ws.cell(row=row, column=3).value  # Column C
    price = ws.cell(row=row, column=4).value  # Column D
    if model is not None:
        model_to_price[str(model)] = price

# For each model in A2:A19, lookup price from C:D, write to B2:B19
for row in range(2, 20):
    model = ws.cell(row=row, column=1).value  # Column A
    price = model_to_price.get(str(model), None) if model is not None else None
    ws.cell(row=row, column=2).value = price if price is not None else ''

wb.save(output_path)
