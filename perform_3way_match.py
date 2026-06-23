import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/group_57033/r0/evolve_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/group_57033/r0/evolve_57033/output.xlsx'

# Load workbook and sheets
wb = openpyxl.load_workbook(input_path)
sheet4 = wb['Sheet4']
cbtrans_sheet = wb['CBtrans']

# Get headers for reference
header4 = [cell.value for cell in sheet4[1]]
header_cb = [cell.value for cell in cbtrans_sheet[1]]

# Find column indices for matching (headers can be case-sensitive)
def find_col(header, name):
    for idx, value in enumerate(header, 1):
        if value and value.strip().lower() == name.strip().lower():
            return idx
    raise Exception(f"Column '{name}' not found.")

col_company_4 = find_col(header4, 'company')
col_account_4 = find_col(header4, 'account')
col_xchar_4   = find_col(header4, 'xchar')

col_company_cb = find_col(header_cb, 'company')
col_account_cb = find_col(header_cb, 'account')
col_xchar_cb   = find_col(header_cb, 'xchar')

# Color fill
fill = PatternFill(start_color='FF66CC', end_color='FF66CC', fill_type='solid')

for row_idx in range(2, 8):  # Rows 2 to 7 inclusive
    company = sheet4.cell(row=row_idx, column=col_company_4).value
    account = sheet4.cell(row=row_idx, column=col_account_4).value
    xchar   = sheet4.cell(row=row_idx, column=col_xchar_4).value
    
    match_found = False
    for cb_row in cbtrans_sheet.iter_rows(min_row=2, max_row=cbtrans_sheet.max_row,
                                          min_col=1, max_col=len(header_cb)):
        cb_company = cb_row[col_company_cb-1].value
        cb_account = cb_row[col_account_cb-1].value
        cb_xchar   = cb_row[col_xchar_cb-1].value
        if (company==cb_company and account==cb_account and xchar==cb_xchar):
            match_found = True
            break
    # Write the result
    result_cell = sheet4.cell(row=row_idx, column=11)  # Column K is 11
    result_cell.value = 'Match' if match_found else '-'
    result_cell.fill = fill
    if result_cell.value:
        result_cell.value = result_cell.value.capitalize()

wb.save(output_path)
