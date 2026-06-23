import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/group_516-46/r0/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/group_516-46/r0/evolve_516-46/output.xlsx'
sheet_name = 'ورقة1'

# Read the data (A2:E...)
df = pd.read_excel(input_path, sheet_name=sheet_name, usecols='A:E', skiprows=1)
cols = df.columns
brand_col = [c for c in cols if 'brand' in str(c).lower() or 'اسم' in str(c).lower()]
date_col = [c for c in cols if 'date' in str(c).lower() or 'تاريخ' in str(c).lower()]
qty_col = [c for c in cols if 'qty' in str(c).lower() or 'كمية' in str(c).lower()]

# Fallback if not detect by keywords
brand_col = brand_col[0] if brand_col else cols[1]
date_col = date_col[0] if date_col else cols[0]
qty_col = qty_col[0] if qty_col else cols[2]

# Coerce date
df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
df = df.dropna(subset=[brand_col, date_col])

# 1. Get most recent date for each brand
last_dates = df.groupby(brand_col)[date_col].max().reset_index()
# 2. For each brand, filter only those rows that match their last date
merged = pd.merge(df, last_dates, left_on=[brand_col, date_col], right_on=[brand_col, date_col])
# 3. Sum qty by brand and date
merged2 = merged.groupby([brand_col, date_col], as_index=False)[qty_col].sum()
# 4. Mark if brand has duplicates (multiple combined on that date)
mod_info = merged.groupby(brand_col).size().reset_index(name='count')
mod_brands = set(mod_info[mod_info['count']>1][brand_col])
merged2['modification'] = merged2[brand_col].apply(lambda b: 'Combined' if b in mod_brands else '')

# 5. Write the results to H2:L (columns H-L, i.e., columns 8-12)
wb = load_workbook(input_path)
ws = wb[sheet_name]
data = [merged2.columns.tolist()] + merged2.values.tolist()
for i, row in enumerate(data):
    for j, val in enumerate(row):
        ws.cell(row=i+2, column=8+j, value=val)
wb.save(output_path)
