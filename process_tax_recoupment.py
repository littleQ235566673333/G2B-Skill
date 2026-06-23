import openpyxl

# Load the workbook and select the active worksheet
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_36097_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_36097_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(3, 7):
    cost = ws[f'B{row}'].value
    itv = ws[f'C{row}'].value
    sale = ws[f'D{row}'].value

    # Compute profit and loss
    profit = sale - itv
    loss = itv - sale if sale < itv else 0

    # Apply recoup/scrapping logic
    if sale < itv:  # sold at a loss
        value = itv - loss
    elif profit < cost:
        value = profit
    else:
        value = cost - itv

    # Preserve formatting from column I
    source_cell = ws[f'I{row}']
    target_cell = ws[f'H{row}']
    target_cell.value = value
    if source_cell.has_style:
        target_cell._style = source_cell._style

wb.save(output_path)
