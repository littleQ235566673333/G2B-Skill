from openpyxl import load_workbook

replacements = {
    'UNATIDE STATES AMERICAN': 'USA',
    'FRANCE': 'FRA',
    'KOREA': 'KOR',
    'GERMANY': 'GR',
    'ITALY': 'IT'
}

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun2/eval_438-18_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun2/eval_438-18_tc1/output.xlsx'

wb = load_workbook(input_path)
input_ws = wb['INPUT']

if 'OUTPUT' in wb.sheetnames:
    output_ws = wb['OUTPUT']
    for row in range(1, 7):
        for col in range(1, 5):
            output_ws.cell(row=row, column=col).value = None
else:
    output_ws = wb.create_sheet('OUTPUT')

for row in range(1, 7):
    for col in range(1, 5):
        value = input_ws.cell(row=row, column=col).value
        output_ws.cell(row=row, column=col).value = value
    country_val = input_ws.cell(row=row, column=6).value
    if country_val in replacements:
        output_ws.cell(row=row, column=4).value = replacements[country_val]
    else:
        output_ws.cell(row=row, column=4).value = country_val

wb.save(output_path)
