import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/before_fix/core_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/before_fix/core_4714/output.xlsx'

df = pd.read_excel(input_path)

def parse_month(val):
    # Try to auto-parse or treat already-string values
    if isinstance(val, str):
        return pd.to_datetime(val, errors='coerce')
    return val

df['C_parsed'] = df.iloc[:,2].apply(parse_month)
# sort for windowing
df.sort_values([df.columns[0], 'C_parsed'], inplace=True)
results = ['n/a'] * len(df)
for idx, row in df.iterrows():
    employee = row[0]
    date = row['C_parsed']
    # get all records for this employee before and including 'date'
    rows_for_emp = df[(df[df.columns[0]]==employee) & (df['C_parsed']<=date)]
    # restrict to last 4 months including this
    # Only keep 4 latest
    rows_for_emp = rows_for_emp.sort_values('C_parsed').tail(4)
    if len(rows_for_emp)==4:
        avg = rows_for_emp[df.columns[3]].mean()
        # Output as int when whole, else as 2 decimals string
        res = str(int(avg)) if avg==int(avg) else '{:.2f}'.format(avg)
        results[idx]=res
    else:
        results[idx]='n/a'
# Write to correct sheet and E2:E25
wb = load_workbook(input_path)
ws = wb.active
for i, val in enumerate(results[:24], start=2):
    ws[f'E{i}']=val
wb.save(output_path)
