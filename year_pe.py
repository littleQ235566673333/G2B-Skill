import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_4/group_47766/r2/evolve_47766/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_4/group_47766/r2/evolve_47766/output.xlsx'

df = pd.read_excel(input_path, sheet_name='Total (2)', header=None)

def serial_to_date(serial):
    if pd.isnull(serial): return None
    try:
        origin = datetime(1899, 12, 30)
        if serial < 61:
            return origin + timedelta(days=serial-1)
        else:
            return origin + timedelta(days=serial)
    except:
        return None

def safe_float(x):
    try: return float(eval(str(x)))
    except Exception:
        return None

# Data blocks
comm_col, date_col, agent_col = 2, 5, 7
rows1 = range(8,38)
data1 = [(df.iloc[r-1,comm_col], df.iloc[r-1,date_col], df.iloc[r-1,agent_col]) for r in rows1]
rows2 = range(41,59)
data2 = [(df.iloc[r-1,comm_col], df.iloc[r-1,date_col], df.iloc[r-1,agent_col]) for r in rows2]

commission1 = [(safe_float(c), serial_to_date(d)) for (c,d,a) in data1 if isinstance(a, str) and 'PE' in a]
commission2 = [(safe_float(c), serial_to_date(d)) for (c,d,a) in data2 if isinstance(a, str) and 'PE' in a]

YEAR = 2023
total = 0.0
for (amt, date) in commission1 + commission2:
    if amt is not None and date is not None and date.year == YEAR:
        total += amt

# Write result
wb = load_workbook(input_path)
ws = wb['Total (2)']
ws['K40'] = total
wb.save(output_path)
print('Yearly PE 2023:', total)
