import openpyxl

# File paths
input_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_45937_tc1/input.xlsx"
output_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_45937_tc1/output.xlsx"

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read kms ranges and their corresponding categories/percentages
def read_legend(ws, start_row, end_row, col_cat, col_min, col_max):
    legend = []
    for row in range(start_row, end_row + 1):
        cat = ws[f'{col_cat}{row}'].value
        min_km = ws[f'{col_min}{row}'].value
        max_km = ws[f'{col_max}{row}'].value
        legend.append((cat, min_km, max_km))
    return legend

# Example legend location (H20:I22 for kms range, assume N for max if exists or I is max)
legend = read_legend(ws, 20, 22, 'H', 'I', 'I')

# For each value in D7:D9, find the matching category
for row in range(7, 10):
    kms = ws[f'D{row}'].value
    category = None
    for cat, min_km, max_km in legend:
        try:
            if min_km is not None and max_km is not None:
                if min_km <= kms <= max_km:
                    category = cat
                    break
            elif min_km is not None:
                if kms >= min_km:
                    category = cat
                    break
        except TypeError:
            continue
    ws[f'E{row}'] = category

# Save to output
wb.save(output_path)
print("Completed column E with category based on kms range.")
