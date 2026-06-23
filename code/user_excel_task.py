from openpyxl import load_workbook

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42/eval_333-29_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42/eval_333-29_tc1/output.xlsx"

wb = load_workbook(input_path)
ws_A = wb['A']
ws_B = wb['B']

# Step 1: Find first P with 'Yes' or 'NA'
first_p_row = None
for row in range(2, ws_A.max_row + 1):  # Skip header
    val = ws_A[f'P{row}'].value
    if val == 'Yes' or val == 'NA':
        first_p_row = row
        break

# Step 2 & fallback
if first_p_row:
    # Go upward in L column from (first_p_row-1) to 2 (header is 1)
    target_l_row = None
    for crow in range(first_p_row - 1, 1, -1):
        if ws_A[f'L{crow}'].value == 100:
            target_l_row = crow
            break
    # The corresponding date
    found_date = ws_A[f'F{target_l_row}'].value if target_l_row else None
else:
    found_date = ws_A['F3'].value  # fallback from instructions

# Step 3: Check text in A!F1 vs B!C*
a_f1 = ws_A['F1'].value
match_row = None
for row in range(2, ws_B.max_row + 1):
    if ws_B[f'C{row}'].value == a_f1:
        match_row = row
        break
if match_row and found_date is not None:
    ws_B[f'F{match_row}'] = found_date

# Step 4: Write found date to B!F4 as final answer
ws_B['F4'] = found_date

# Save and set 'A' as active
wb.active = wb.sheetnames.index('A')
wb.save(output_path)
