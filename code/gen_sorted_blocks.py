import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

input_file = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY/eval_191-40_tc1/input.xlsx'
output_file = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY/eval_191-40_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_file)
wb_data = openpyxl.load_workbook(input_file, data_only=True)
ws = wb['Sheet1']
ws_data = wb_data['Sheet1']

# Preserve original column widths
col_widths = {get_column_letter(col): ws.column_dimensions[get_column_letter(col)].width for col in range(1, ws.max_column+1)}

def is_blank_row(row):
    # Considers a row blank if all displayed cells (in used columns) are blank/None
    return all((cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == '')) for cell in row)

def get_block_indices(ws, min_row=1, max_row=85, max_col=8):
    blocks = []
    block = []
    for i, row in enumerate(ws.iter_rows(min_row=min_row, max_col=max_col, max_row=max_row), min_row):
        if is_blank_row(row):
            if block:
                blocks.append(block)
                block = []
        else:
            block.append(i)
    if block:
        blocks.append(block)
    return blocks

def restore_formula(val):
    # Converts "'=B5+C5" to "=B5+C5" if needed
    if isinstance(val, str):
        v = val.strip()
        if v.startswith("'="):
            return v[1:]
        if v.startswith('='):
            return v
    return val

def get_formula(row_idx, col):
    # Recreate original formulas dynamically given the new row (for D=4, E=5, F=6)
    if col == 4:
        return f"=B{row_idx}+C{row_idx}"
    elif col == 5:
        return f"=D{row_idx}+E{row_idx}"
    elif col == 6:
        return f"=B{row_idx}-C{row_idx}"
    return None

def format_number(cell, value):
    # Format as 0 if int, else 0.0
    try:
        v = float(value)
        if abs(v - int(round(v))) < 1e-8:
            cell.number_format = '0'
        else:
            cell.number_format = '0.0'
    except:
        cell.number_format = '0.0'

def style_bold_center(cell):
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)

def sorted_block_rows(ws, ws_data, block_rows):
    rows = []
    for r in block_rows:
        rec = []
        for c in range(1, 9):
            val = ws.cell(row=r, column=c).value
            rec.append(val)
        # Evaluated D
        d_val = ws_data.cell(row=r, column=4).value
        rec.append(d_val)
        rec.append(r) # store original row
        rows.append(rec)
    # Sort descending by D
    rows.sort(key=lambda x: (float(x[8]) if x[8] is not None else float('-inf')), reverse=True)
    return rows

def apply_block(ws, block_rows_struct, dest_start_row):
    # block_rows_struct: output of sorted_block_rows (each entry is 0-7 for cols A-H, 8: evaluated D, 9: orig row)
    # dest_start_row: where to start writing
    for offset, rec in enumerate(block_rows_struct):
        rownum = dest_start_row + offset
        for c in range(1,9):
            cell = ws.cell(row=rownum, column=c)
            val = rec[c-1]
            # For D,E,F: always ensure formula restored for this row
            if c in (4,5,6):
                if c==4:
                    cell.value = f"=B{rownum}+C{rownum}"
                elif c==5:
                    cell.value = f"=D{rownum}+E{rownum}"
                elif c==6:
                    cell.value = f"=B{rownum}-C{rownum}"
                style_bold_center(cell)
                # Show value via data_only for number_format
                data_val = ws_data.cell(row=rec[9], column=c).value
                format_number(cell, data_val)
            else:
                cell.value = val
    return dest_start_row + len(block_rows_struct)

# Main processing
blocks = get_block_indices(ws, min_row=1, max_row=85, max_col=8)
# Blank area first
for r in range(1, 86):
    for c in range(1, 9):
        ws.cell(row=r, column=c).value = None
write_row = 1
for idx, block in enumerate(blocks):
    sorted_rows = sorted_block_rows(ws, ws_data, block)
    write_row = apply_block(ws, sorted_rows, write_row)
    # After a block, leave one blank row unless it's the last
    if idx != len(blocks)-1:
        write_row +=1
# Restore column widths
for col, width in col_widths.items():
    if width:
        ws.column_dimensions[col].width = width
wb.save(output_file)
