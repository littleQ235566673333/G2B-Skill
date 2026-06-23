import pandas as pd
from openpyxl import load_workbook
import math

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_3/regression_gate/before_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_3/regression_gate/before_pass/core_18935/output.xlsx'

wb = load_workbook(input_path)
sheetnames = wb.sheetnames
df_dict = pd.read_excel(input_path, sheet_name=None)

def get_lookup_table_sheetname():
    for name, df in df_dict.items():
        first_row = df.iloc[0].astype(str).tolist()
        if any('Data Table' in x for x in first_row):
            return name
        if df.shape[1] > 6:
            return name
    return sheetnames[0]

def get_report_sheetname():
    for name, df in df_dict.items():
        if any('Work Criteria' in str(x) for x in df.iloc[:,0]):
            return name
    return sheetnames[0]

lookup_table_sheet = get_lookup_table_sheetname()
report_sheet = get_report_sheetname()
lookup_df = df_dict[lookup_table_sheet]
report_df = df_dict[report_sheet]
ws = wb[report_sheet]

start_row = 16  # Excel 1-based, row 17 is idx 16
results = []
for i in range(6):
    row = start_row + i
    work_criteria = ws[f'B{row+1}'].value
    cat_type = ws[f'C{row+1}'].value
    material = ws[f'D{row+1}'].value
    results.append({'row': row+1, 'work_criteria': work_criteria, 'cat_type': cat_type, 'material': material})

header_row = 0
for idx, row in lookup_df.iterrows():
    if ('Criteria' in str(row[0])) and ('Category' in str(row[1])) and ('Material' in str(row[2])):
        header_row = idx
        break
lookup_header = lookup_df.iloc[header_row,:].values.tolist()

def safe_str(x):
    if isinstance(x, str):
        return x.lower().strip()
    if isinstance(x, float) and (math.isnan(x) or x is None):
        return ''
    return str(x).lower().strip()

colmap = {safe_str(k): i for i, k in enumerate(lookup_header) if safe_str(k) != ''}

def find_category_value(criteria, cat_type, material):
    for idx, row in lookup_df.iterrows():
        if idx <= header_row:
            continue
        vals = row.values.tolist()
        def val_or_empty(colname):
            ci = colmap.get(colname, None)
            if ci is None or ci >= len(vals):
                return ''
            return vals[ci]
        if (str(val_or_empty('criteria')).strip() == str(criteria).strip() and
            str(val_or_empty('category type')).strip() == str(cat_type).strip() and
            str(val_or_empty('material')).strip() == str(material).strip()):
            for key in colmap:
                if 'value' in key:
                    vi = colmap[key]
                    if vi < len(vals):
                        return vals[vi]
            return vals[-1]
    return None

for idx, res in enumerate(results):
    val = find_category_value(res['work_criteria'], res['cat_type'], res['material'])
    cell = f'D{17+idx}'
    ws[cell] = val

wb.save(output_path)
