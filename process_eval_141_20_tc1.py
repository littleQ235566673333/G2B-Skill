import openpyxl
from openpyxl.styles import PatternFill

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r3/eval_141-20_tc1/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r3/eval_141-20_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_fp)
ws_pl = wb['PL Recon Items']
ws_stmt = wb['Statement Recon Items']

# Helper to detect yellow fill
YELLOWS = {'FFFFFF00', 'FFFF00'}
def is_yellow(cell):
    fill = cell.fill
    if fill is None or fill.fill_type != 'solid':
        return False
    fg = getattr(fill, 'fgColor', None)
    if fg is None or fg.type != 'rgb':
        return False
    color = fg.rgb.upper()
    if color.startswith('FF') and len(color) == 8:
        return color in YELLOWS
    if len(color) == 6:
        return 'FF'+color in YELLOWS
    return False

# Get all data rows (excluding header)
pl_rows = list(ws_pl.iter_rows(min_row=2, values_only=False))
st_rows = list(ws_stmt.iter_rows(min_row=2, values_only=False))

# Find yellow-highlighted rows
pl_yellow_idx = [i for i, row in enumerate(pl_rows) if any(is_yellow(cell) for cell in row)]
st_yellow_idx = [i for i, row in enumerate(st_rows) if any(is_yellow(cell) for cell in row)]

# (invoice, value) in PL, (reference, value) in Statement
pl_keys = set((row[2].value, row[3].value) for i, row in enumerate(pl_rows) if i in pl_yellow_idx)
st_keys = set((row[5].value, row[8].value) for i, row in enumerate(st_rows) if i in st_yellow_idx)

# Intersection = match in both yellow sets
matching_keys = pl_keys & st_keys

pl_del_idx = [i for i, row in enumerate(pl_rows) if (row[2].value, row[3].value) in matching_keys]
st_del_idx = [i for i, row in enumerate(st_rows) if (row[5].value, row[8].value) in matching_keys]

for idx in sorted(pl_del_idx, reverse=True):
    ws_pl.delete_rows(idx+2)
for idx in sorted(st_del_idx, reverse=True):
    ws_stmt.delete_rows(idx+2)

wb.save(output_fp)
