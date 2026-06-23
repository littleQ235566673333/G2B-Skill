import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_7/regression_gate/before_pass/core_3413/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_7/regression_gate/before_pass/core_3413/output.xlsx'

# Load Excel to pandas and openpyxl
# Let pandas read just first sheet (to match openpyxl default)
df = pd.read_excel(input_path, engine='openpyxl')
wb = load_workbook(input_path)
ws = wb.active

# Determine exact column names for safety
cols = list(df.columns)
# Assume first three columns: Department, RU, Value (or similar)
# If user uses different column names, correct as needed.
if len(cols) >= 3:
    department_col = cols[0]
    ru_col = cols[1]
    value_col = cols[2]
else:
    raise Exception('Not enough columns in input!')

# Loop over G3:G6 (corresponds to E3:F6 for criteria)
for out_row in range(3, 7):  # Excel rows 3..6
    dep = ws[f'E{out_row}'].value
    ru = ws[f'F{out_row}'].value
    result = None
    # Require department, can be strict on RU
    if dep is not None:
        # Try to match both Department and RU
        df_match = df[(df[department_col] == dep) & (df[ru_col] == ru)]
        if len(df_match) > 0:
            # Combination exists
            result = df_match[value_col].sum()
        else:
            # If not, sum all for that department (ignore RU)
            df_dept = df[df[department_col] == dep]
            result = df_dept[value_col].sum()
    ws[f'G{out_row}'] = result if result is not None else ''

wb.save(output_path)
