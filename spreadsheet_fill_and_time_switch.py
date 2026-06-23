import openpyxl
from openpyxl.styles import PatternFill

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_45372_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_45372_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)
ws = wb.active

fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

for row in range(2, 16):
    time_value = ws[f'D{row}'].value
    b_value = ws[f'B{row}'].value
    c_value = ws[f'C{row}'].value

    # Handle both string and datetime.time
    if isinstance(time_value, str):
        check_time = time_value
    elif hasattr(time_value, 'strftime'):
        check_time = time_value.strftime("%H:%M")
    else:
        check_time = str(time_value)

    if check_time < "09:45":
        result = b_value
    else:
        result = c_value

    final_val = "" if result in [0, None, "0", 0.0] else result
    ws[f'E{row}'] = final_val
    ws[f'E{row}'].fill = fill

wb.save(output_path)
