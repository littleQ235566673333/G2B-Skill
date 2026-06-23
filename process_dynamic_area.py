import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/group_263-1/r2/evolve_263-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/group_263-1/r2/evolve_263-1/output.xlsx'

# Read input
df = pd.read_excel(input_path)

# Only use the first three relevant columns, by hard index as names are clear and fixed
material_col, width_col, height_col = 'Mtrl', 'Width', 'Height'
df = df[[material_col, width_col, height_col]].dropna()

# Ensure numeric types for area calculation
df[width_col] = pd.to_numeric(df[width_col], errors='coerce')
df[height_col] = pd.to_numeric(df[height_col], errors='coerce')
df = df.dropna()

# Calculate area
df['Area'] = df[width_col] * df[height_col]

# Group by material and sum areas
result = df.groupby(material_col)['Area'].sum().reset_index()

# Load the output workbook
wb = load_workbook(input_path)
ws = wb.active

# Write results to H2:H4
for i in range(3):
    cell = f'H{i+2}'
    if i < len(result):
        value = f"{result.iloc[i,0]}: {result.iloc[i,1]:.2f}"
    else:
        value = None
    ws[cell] = value

wb.save(output_path)
