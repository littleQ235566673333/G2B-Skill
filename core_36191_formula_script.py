from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_4/regression_gate/before_pass/core_36191/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_4/regression_gate/before_pass/core_36191/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Formula to extract everything before the last space (first & surname):
formula_c2 = '=LEFT(B2,LOOKUP(2,1/(MID(B2,ROW($1:$99),1)=" "),ROW($1:$99))-1)'
formula_c3 = '=LEFT(B3,LOOKUP(2,1/(MID(B3,ROW($1:$99),1)=" "),ROW($1:$99))-1)'

ws['C2'] = formula_c2
ws['C3'] = formula_c3

wb.save(output_path)
print('Done')
