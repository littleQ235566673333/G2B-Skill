import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_120-24_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_120-24_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

blue_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')

for row in range(2, 46):  # From 2 to 45
    BL = ws[f'BL{row}'].value
    BG = ws[f'BG{row}'].value
    AY = ws[f'AY{row}'].value

    if BL == 'LAG' and BG in [
        'OCOGS - Spares - Transfer Price Overhead',
        'OFSS - ORCL Consulting Prod Cost I/C'
    ]:
        if AY in ['BOA_033E', 'BOA_011G']:
            ws[f'BN{row}'].value = 'OCOGS - Spares - Transfer Price Overhead'
        else:
            ws[f'BN{row}'].value = 'OFSS - ORCL Consulting Prod Cost I/C'

    # Fill color if BL == 'LAG' and BN is populated
    BN = ws[f'BN{row}'].value
    if BL == 'LAG' and BN:
        ws[f'BL{row}'].fill = blue_fill

wb.save(output_path)
