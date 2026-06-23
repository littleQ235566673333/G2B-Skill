from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_38823_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_38823_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

for row in range(4, 8):
    formula = (
        '=SUMIFS($C$3:$C$10, $A$3:$A$10, ">="&$E$4, $A$3:$A$10, "<="&$F$4, $B$3:$B$10, "*"&H{row}&"*")'
    ).replace('{row}', str(row))
    ws[f'I{row}'] = formula

wb.save(output_path)
