from openpyxl import load_workbook
from datetime import timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_2/group_59595/r2/evolve_59595/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_2/group_59595/r2/evolve_59595/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

data = []
for row in ws.iter_rows(min_row=4, max_row=19, min_col=1, max_col=2, values_only=True):
    data.append(row)
dates = [row[0] for row in data]
points = [row[1] for row in data]

for i in range(len(dates)):
    date_i = dates[i]
    if date_i is None:
        continue
    seven_days_prior = date_i - timedelta(days=6)
    sum_points = 0
    for j in range(len(dates)):
        if dates[j] is not None and seven_days_prior <= dates[j] <= date_i:
            if points[j] is not None:
                sum_points += points[j]
    ws.cell(row=4+i, column=3, value=sum_points)

wb.save(output_path)
