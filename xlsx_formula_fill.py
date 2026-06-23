from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_4/group_36277/r2/evolve_36277/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_4/group_36277/r2/evolve_36277/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active  # Assume first sheet
headers = [cell.value for cell in ws[1]]

# Use 2nd header as example if possible
target_header = headers[1] if len(headers) > 1 else headers[0]
last_col = ws.max_column
last_row = ws.max_row

# Data range: A2:<last_col><last_row>
data_range = f'${get_column_letter(1)}$2:${get_column_letter(last_col)}${last_row}'
header_row_range = f'${get_column_letter(1)}$1:${get_column_letter(last_col)}$1'

for output_row in range(2, 6):
    formula = f'=INDEX({data_range},ROW()-1,MATCH("{target_header}",{header_row_range},0))'
    ws[f'I{output_row}'] = formula

wb.save(output_path)
