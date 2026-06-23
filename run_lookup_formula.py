import openpyxl

# Load input workbook
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_2/regression_gate/before_pass/core_50526/input.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get lookup value in B6
lookup_value = ws['B6'].value

# Get headers (B1, C1, ...) values only (skip A1, index 0)
header = [cell.value for cell in ws[1]][1:]
results = []

# Find the row number for the lookup value in column A
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True):
    if row[0] == lookup_value:
        for idx, v in enumerate(row[1:]): # skip col A which is row[0]
            if isinstance(v, (int, float)) and v > 0:
                results.append(header[idx])
        break

# Write results to B9, B10
for i in range(2):
    ws.cell(row=9 + i, column=2, value=results[i] if i < len(results) else None)

# Save to output
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_2/regression_gate/before_pass/core_50526/output.xlsx'
wb.save(output_path)
