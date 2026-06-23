import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/regression_gate/after_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/regression_gate/after_pass/core_18935/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Locate Data Table bounds (row 4 header, data till row 10)
data_start = 5
data_end = 10
category_cols = {'Category 1': 3, 'Category 2': 4, 'Category 3': 5} # 0-based index

# Build Data Table dict for fast lookup
# (work, material) -> row
data_table = {}
for r in range(data_start, data_end+1):
    work = ws.cell(row=r, column=1).value
    material = ws.cell(row=r, column=2).value
    if work and material:
        data_table[(work, material)] = [ws.cell(row=r, column=i).value for i in range(3, 6)]

# For each row in report (rows 17-22)
for idx, r in enumerate(range(17, 23), start=0):
    work = ws.cell(row=r, column=1).value
    cat_type = ws.cell(row=r, column=2).value
    material = ws.cell(row=r, column=3).value
    col_idx = category_cols.get(cat_type)
    val = None
    if (work, material) in data_table and col_idx is not None:
        val = data_table[(work, material)][col_idx-3]  # adjust for zero index
    ws.cell(row=r, column=4).value = val

wb.save(output_path)
print('Category values have been populated in D17:D22.')