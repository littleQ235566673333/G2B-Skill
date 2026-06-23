import openpyxl
from datetime import datetime

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_58723_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_58723_tc1/output.xlsx'

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Extract names and entry times
names = []
times = []
rows = []
for row in range(2, 42):  # M2:M41 => rows 2-41
    name = ws[f'C{row}'].value
    entry_time = ws[f'I{row}'].value
    if isinstance(entry_time, str):
        try:
            entry_time = datetime.strptime(entry_time, '%Y-%m-%d %H:%M:%S')
        except Exception:
            entry_time = None
    names.append(name)
    times.append(entry_time)
    rows.append(row)

# Build mapping from name to latest entry time
from collections import defaultdict
name_times = defaultdict(list)
for name, entry_time in zip(names, times):
    if name is not None and entry_time is not None:
        name_times[name].append(entry_time)

latest_time_per_name = {name: max(times) for name, times in name_times.items() if times}

# Write 'Latest' or 'Not Latest' to column M
for idx, (name, entry_time) in enumerate(zip(names, times)):
    label = ''
    if name is not None and entry_time is not None:
        if entry_time == latest_time_per_name.get(name):
            label = 'Latest'
        else:
            label = 'Not Latest'
    ws[f'M{rows[idx]}'] = label

wb.save(output_path)
