import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/regression_gate/before_fix/core_60-7/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/regression_gate/before_fix/core_60-7/output.xlsx'

# Load all relevant sheets with pandas for easy manipulation
ds = pd.read_excel(input_path, sheet_name=None)
existing = ds['Existing Task']
additions = ds['Additions']
retired = ds['Retired']

# Use all columns for safe row comparison
key_cols = existing.columns.tolist()

# Combine Existing + Additions, removing duplicates
combo = pd.concat([existing, additions], ignore_index=True)
combo_nodup = combo.drop_duplicates(subset=key_cols, keep='last')

# Remove (subtract) any row that matches 'retired' (full row subtraction)
if not retired.empty:
    merged = combo_nodup.merge(retired, how='left', indicator=True, on=key_cols)
    final = merged[merged['_merge'] == 'left_only'].drop(columns=['_merge'])
else:
    final = combo_nodup.copy()

# Prepare to write at A3: get 'Consolidated Tracker' columns as reference
wb = load_workbook(input_path)
ws = wb['Consolidated Tracker']
header = [cell.value for cell in ws[2] if cell.value is not None]
# align order if possible
if set(header) == set(final.columns):
    final = final[header]

# Write values to Excel starting from A3 (row index 3, col 1)
row_start = 3
col_start = 1
for row_offset, vals in enumerate(final.itertuples(index=False), start=0):
    for col_offset, v in enumerate(vals):
        ws.cell(row=row_start + row_offset, column=col_start + col_offset, value=v)

to_clear = ws.max_row
# Optional: clear below output in case old data lingers
for r in range(row_start + len(final), to_clear + 1):
    for c in range(col_start, col_start + len(header)):
        ws.cell(row=r, column=c, value=None)

wb.save(output_path)
