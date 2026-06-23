import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/group_516-46/r1/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/group_516-46/r1/evolve_516-46/output.xlsx'
sheet_name = 'ورقة1'
# Read with pandas

df = pd.read_excel(input_path, sheet_name=sheet_name)
cols = df.columns.tolist()

# We only want actual data rows, lets drop all-na rows
main_data = df.loc[:, cols[:5]].dropna(how='all')

# Try to infer important columns
brand_col = [c for c in main_data.columns if 'brand' in c.lower()]
qty_col = [c for c in main_data.columns if 'qty' in c.lower()]
date_col = [c for c in main_data.columns if 'date' in c.lower()]

if not brand_col:
    brand_col = [cols[1]]
if not qty_col:
    qty_col = [cols[4]]
if not date_col:
    date_col = [cols[0]]

bcol = brand_col[0]
qcol = qty_col[0]
dcol = date_col[0]

# Convert date to datetime
main_data[dcol] = pd.to_datetime(main_data[dcol], errors='coerce')

# Remove rows with no date or no brand
main_data = main_data.dropna(subset=[bcol, dcol])

# Find last date per brand
grouped = main_data.groupby(bcol)
last_date = grouped[dcol].transform('max')
last_rows = main_data[main_data[dcol] == last_date]

# Now group by brand + date to sum qty and check if there are multiple entries
res = last_rows.groupby([bcol, dcol])[qcol].agg(['sum', 'count']).reset_index()

# Prepare modifications
res['Modification'] = res['count'].apply(lambda c: 'Combined' if c>1 else '')

# For user's convenience, rename for nice header
res_out = res.copy()
res_out.columns = ["Brand", "Date", "Total Qty", "Entry Count", "Modification"]

# Now write to output excel in columns H2:L (header in H1:L1)
wb = load_workbook(input_path)
ws = wb[sheet_name]

header = ["Brand", "Date", "Total Qty", "Modification"]
for idx, val in enumerate(header, 8):
    ws.cell(row=1, column=idx, value=val)

for i, row in res_out.iterrows():
    ws.cell(row=2+i, column=8, value=row["Brand"])
    ws.cell(row=2+i, column=9, value=row["Date"])
    ws.cell(row=2+i, column=10, value=row["Total Qty"])
    ws.cell(row=2+i, column=11, value=row["Modification"])

wb.save(output_path)
