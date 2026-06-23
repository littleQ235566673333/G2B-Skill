import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_5835_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_5835_tc1/output.xlsx'

# Load workbook and select active worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get LOG VALUEs from column I (I3:I19)
for row in range(3, 20):
    log_value = ws[f'I{row}'].value
    ws[f'C{row}'].value = log_value

# Save output workbook
wb.save(output_path)
