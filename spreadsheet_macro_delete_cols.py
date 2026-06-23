from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_534-26_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_534-26_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Get current year and month
today = datetime.today()
this_year = today.year
this_month = today.month

# Find which columns to delete
columns_to_delete = []
for col in range(1, ws.max_column + 1):
    date_val = ws.cell(row=1, column=col).value
    if isinstance(date_val, datetime):
        year = date_val.year
        month = date_val.month
    elif isinstance(date_val, str):
        try:
            # Try parsing ISO format first
            parsed = datetime.strptime(date_val, '%Y-%m-%d')
            year, month = parsed.year, parsed.month
        except Exception:
            try:
                # Try MM/DD/YYYY
                parsed = datetime.strptime(date_val, '%m/%d/%Y')
                year, month = parsed.year, parsed.month
            except Exception:
                continue
    else:
        continue  # Not a date, skip
    # Check if older than current month/year
    if (year < this_year) or (year == this_year and month < this_month):
        columns_to_delete.append(col)

# Delete columns in reverse order to avoid index issues
for idx in sorted(columns_to_delete, reverse=True):
    ws.delete_cols(idx)

wb.save(output_path)
