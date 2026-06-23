from openpyxl import load_workbook
import json

# Dump full headers for inspection
INPUT = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r2/eval_120-24_tc1/input.xlsx'
wb = load_workbook(INPUT)
ws = wb['Sheet1']
rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
print(json.dumps(rows[0]), flush=True)
