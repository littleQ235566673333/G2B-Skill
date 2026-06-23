from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']
today = datetime.today()

# Start from row 4 (per user's example), go down for each non-empty H
row = 4
while True:
    h_val = ws[f'H{row}'].value
    i_val = ws[f'I{row}'].value
    if h_val is None and (i_val is None or str(i_val).strip() == ''):
        break  # stop if both are empty (end of data)
    yesno = str(i_val).strip().upper() if i_val is not None else ''
    result = 'NO ACTION'
    if isinstance(h_val, datetime):
        days_ago = (today - h_val).days
        if yesno == 'YES':
            if days_ago <= 30:
                result = 'HOLD'
            else:
                result = 'TOUCH BASE'
    elif yesno == 'YES' and not h_val:
        result = 'NO ACTION'
    ws[f'J{row}'] = result
    row += 1

wb.save(output_path)
