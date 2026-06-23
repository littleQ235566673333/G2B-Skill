from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_4/regression_gate/after_pass/core_32337/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_4/regression_gate/after_pass/core_32337/output.xlsx'

wb = load_workbook(input_file)
ws = wb['Sheet1']

# 1. Direct CATEGORY lookup formula for E3:E15
for row in range(3, 16):
    ws[f'E{row}'] = '=INDEX($I$3:$I$15, MATCH(O{0}, $O$3:$O$15, 0))'.format(row)

# 2. Insert Actual Age column after "age (year)" (O, col 15), so after col 15 -> col 16 (P)
ws.insert_cols(16)
ws['P2'] = 'Actual Age'

# Get fill color/style from O2
source_fill = ws['O2'].fill
if isinstance(source_fill, PatternFill):
    match_fill = PatternFill(
        fill_type=source_fill.fill_type,
        fgColor=source_fill.fgColor.rgb,
        bgColor=source_fill.bgColor.rgb
    )
else:
    match_fill = PatternFill(fill_type=None)
# Apply formatting to header
header = ws['P2']
header.font = Font(bold=True)
header.alignment = Alignment(horizontal='center', vertical='top')
header.fill = match_fill

for row in range(3, 16):
    ws[f'P{row}'] = '=ROUNDDOWN((($B$1-C{0})/365),0)'.format(row)
    cell = ws[f'P{row}']
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = match_fill

wb.save(output_file)
