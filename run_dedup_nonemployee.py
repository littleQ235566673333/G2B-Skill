import pandas as pd
from openpyxl import load_workbook

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_6/regression_gate/before_fix/core_91-34/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_6/regression_gate/before_fix/core_91-34/output.xlsx"
sheet_name = "SwiftMD"
start_col, end_col = 2, 15  # B:O
start_row, end_row = 2, 42

# Read with correct header row (row index 1 in Excel == header=1)
df = pd.read_excel(input_path, sheet_name=sheet_name, header=1, dtype=str)
headers = df.columns.tolist()
for_canon = ['Last Name', 'First Name', 'Date Of Birth']

# Some DOB might be named 'Date Of Birth', some 'DOB'; try to standardize
if 'Date Of Birth' in df.columns:
    dob_col = 'Date Of Birth'
elif 'DOB' in df.columns:
    dob_col = 'DOB'
    for_canon = ['Last Name', 'First Name', 'DOB']
else:
    raise Exception('DOB column not found')
for_canon[2] = dob_col

# Apply deduplication filter
rows_to_delete = set()
grouped = df.groupby(for_canon, dropna=False, sort=False)
for keys, group in grouped:
    dup_yes = group[group['Duplicate?'].astype(str).str.strip().str.lower() == 'yes']
    if len(dup_yes) > 1:
        if not (dup_yes['Relationship'] == 'Employee').any():
            idx_to_delete = dup_yes.index[0]
            rows_to_delete.add(idx_to_delete)

df_clean = df.drop(list(rows_to_delete)).reset_index(drop=True)

# Write to B2:O42 with openpyxl
wb = load_workbook(input_path)
ws = wb[sheet_name]

# Write header
for i, header in enumerate(headers, start_col):
    ws.cell(row=start_row, column=i, value=header)

# Write data
max_data_rows = end_row - start_row
for r, row in enumerate(df_clean.values.tolist()[:max_data_rows], start=start_row+1):
    for c, value in enumerate(row, start=start_col):
        ws.cell(row=r, column=c, value=value)

wb.save(output_path)
