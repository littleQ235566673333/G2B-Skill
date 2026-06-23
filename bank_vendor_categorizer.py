import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_524-31_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_524-31_tc1/output.xlsx'

# Example mapping (customize as needed)
vendor_category_map = {
    'amazon.com': 'Shopping',
    'amzn mktp': 'Shopping',
    'starbucks': 'Coffee',
    'walmart': 'Groceries',
    'wholefds': 'Groceries',
    'shell oil': 'Gas',
    'uber': 'Transport',
    'lyft': 'Transport',
}

def extract_vendor(description):
    if '*' in description:
        return description.split('*')[0].strip().lower()
    return description.split()[0].strip().lower()

wb = load_workbook(input_path)
ws = wb['Exp-DB']

# Assuming transaction descriptions are in column B (B1:B53)
desc_col = 'B'
output_col = 'E'
start_row = 1
end_row = 53

for row in range(start_row, end_row + 1):
    desc = ws[f'{desc_col}{row}'].value
    if desc:
        vendor = extract_vendor(desc)
        # Try longest prefix match in mapping, fallback to vendor for category
        found = False
        for key in vendor_category_map.keys():
            if key in vendor:
                ws[f'{output_col}{row}'].value = vendor_category_map[key]
                found = True
                break
        if not found:
            ws[f'{output_col}{row}'].value = 'Other'
    else:
        ws[f'{output_col}{row}'].value = ''

wb.save(output_path)
