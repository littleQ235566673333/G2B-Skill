import openpyxl
import pandas as pd

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/before_fix/core_4714/input.xlsx')
ws = wb.active
rows = list(ws.iter_rows(min_row=2, max_row=25, values_only=True))
# Correct indices for the task
col_names = ['Emp Code', 'Month (YYYYMM)', 'Month', 'Hours', 'Moving Average Amount', 'Moving Average (Four months exceed 48)']
df = pd.DataFrame(rows, columns=col_names)
output = []
for idx, row in df.iterrows():
    emp = row['Emp Code']
    mon = row['Month']
    hrs = row['Hours']
    try:
        mon_int = int(mon)
    except:
        output.append('n/a')
        continue
    emp_df = df[(df['Emp Code'] == emp) & (df['Month'].apply(lambda x: isinstance(x, (int, float)) and x <= mon_int and x >= mon_int - 3))]
    if len(emp_df) < 4:
        output.append('n/a')
    else:
        avg = emp_df['Hours'].mean()
        if avg == int(avg):
            avg_fmt = str(int(avg))
        else:
            avg_fmt = f"{avg:.2f}".rstrip('0').rstrip('.')
        output.append(avg_fmt)
for i, v in enumerate(output, start=2):
    ws[f'E{i}'] = v
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/before_fix/core_4714/output.xlsx')
