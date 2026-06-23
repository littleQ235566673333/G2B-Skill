import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_262-17_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_262-17_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

headers = [cell.value for cell in ws[1]]
task_idx = headers.index('Task')
resp_idx = headers.index('Responsibility')

rows = list(ws.iter_rows(min_row=2, max_row=14, values_only=True))

rows_sorted = sorted(rows, key=lambda row: (row[task_idx], row[resp_idx]))

for i, row in enumerate(rows_sorted, start=2):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

wb.save(output_path)
