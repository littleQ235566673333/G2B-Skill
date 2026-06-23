import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_2/group_44017/r1/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_2/group_44017/r1/evolve_44017/output.xlsx'

def get_months(ws):
    # columns 30 (AD) through 41 (AO)
    return [ws.cell(row=9, column=col).value for col in range(30, 42)]

def cumulative_rate(base, increases, n_applied):
    rate = base
    for i in range(n_applied):
        rate *= (1 + increases[i])
    return rate

def months_between(start, end):
    return (end.year - start.year) * 12 + (end.month - start.month)

def main():
    wb = load_workbook(input_path)
    ws = wb['Data']
    months = get_months(ws)
    yellow_fills = [PatternFill(start_color='FFFFFF00', fill_type='solid'),   # openpyxl hex for yellow
                    PatternFill(start_color='FFFF00', fill_type='solid'),
                    PatternFill(bgColor='FFFF00', fill_type='solid')]

    for row in range(14, 43):    # inclusive
        base_rate = ws.cell(row=row, column=23).value
        eff_date  = ws.cell(row=row, column=12).value
        freq      = ws.cell(row=row, column=10).value
        increases = [ws.cell(row=row, column=col).value for col in range(13, 17)]  # M-P

        for i, col in enumerate(range(30, 42)):
            cell = ws.cell(row=row, column=col)
            month_date = months[i]
            # Remove yellow fill if present
            if cell.fill in yellow_fills or (cell.fill.patternType == 'solid' and cell.fill.start_color.rgb in ['FFFF00', 'FFFFFF00']):
                cell.fill = PatternFill()  # Reset fill

            # Validation: if anything missing, blank
            if (base_rate is None or eff_date is None or freq is None or any(v is None for v in increases)
                or not isinstance(month_date, (datetime.date, datetime.datetime))):
                cell.value = None
                continue

            # Effective?
            if month_date < eff_date:
                cell.value = None
                continue

            # How many increases have phased in?
            n_incs = 0
            months_since = months_between(eff_date, month_date)
            while n_incs < 4 and months_since >= freq * (n_incs+1):
                n_incs += 1
            # Compute compounded rate
            computed_rate = cumulative_rate(base_rate, increases, n_incs)
            cell.value = round(computed_rate, 6)

    wb.save(output_path)

if __name__ == '__main__':
    main()
