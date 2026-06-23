from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load workbook and sheet
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/group_56274/r2/evolve_56274/input.xlsx')
ws = wb['Sheet2']

# 1. Read Fiscal Month (D7)
fiscal_month = ws['D7'].value

# 2. Find which column (G-J) in row 3 matches fiscal_month
col_index = None
for col in range(7, 11):  # G=7, H=8, I=9, J=10
    cell_value = ws.cell(row=3, column=col).value
    if cell_value == fiscal_month:
        col_index = col
        break
if col_index is None:
    raise Exception(f"Fiscal Month {fiscal_month} not found in header row 3, cols G-J")
col_letter = get_column_letter(col_index)

# 3. Values needed (rows 4-7 of found column)
rows = [4, 5, 6, 7]
values = [ws.cell(row=r, column=col_index).value for r in rows]

# Option 1: Write formulas to D9:D12 referencing the month table
for i, r in enumerate(range(9, 13)):
    formula = f'={col_letter}{rows[i]}'
    ws.cell(row=r, column=4).value = formula

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/group_56274/r2/evolve_56274/output.xlsx')
