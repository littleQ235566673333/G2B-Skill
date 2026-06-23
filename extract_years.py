from openpyxl import load_workbook
import datetime

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun1/eval_41978_tc1/input.xlsx')
ws = wb.active

years = []
for row in ws.iter_rows(min_row=14, max_row=185, min_col=7, max_col=7):
    val = row[0].value
    yr = None
    if isinstance(val, datetime.date):
        yr = val.year
    elif isinstance(val, int):
        yr = val
    elif isinstance(val, str):
        try:
            if len(val) == 4 and val.isdigit():
                yr = int(val)
            else:
                yr = datetime.datetime.strptime(val, '%Y-%m-%d').year
        except:
            pass
    if yr is not None:
        years.append(yr)
print(sorted(set(years)))
