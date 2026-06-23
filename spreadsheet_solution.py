from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun2/eval_51354_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun2/eval_51354_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

MONTH_MAP = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
REVERSE_MONTH_MAP = {v:k for k,v in MONTH_MAP.items()}

for row in range(2, 7):
    text = ws[f'A{row}'].value
    m = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec) ?(\d{2})$', text)
    if m:
        mon_str = m.group(1)
        yy = int(m.group(2))
        ws[f'D{row}'] = yy
        ws[f'D{row}'].fill = PatternFill(fill_type='solid', fgColor='FFC000')
        mm = MONTH_MAP[mon_str]
        mm_new = mm + 1
        yy_new = yy
        if mm_new > 12:
            mm_new = 1
            yy_new += 1
            yy_new = yy_new % 100
        new_mon_str = REVERSE_MONTH_MAP[mm_new]
        ws[f'E{row}'] = f'{new_mon_str} {yy_new:02d}'
    else:
        ws[f'D{row}'] = ''
        ws[f'E{row}'] = ''

wb.save(output_path)
