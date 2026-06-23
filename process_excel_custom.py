import openpyxl
from datetime import datetime

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_46240_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_46240_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 5):  # J2:J4
    h_val = ws[f'H{row}'].value
    j_val = ''
    if h_val == 'N/A':
        j_val = 'N/A'
    elif h_val is None or (isinstance(h_val, str) and h_val.strip() == ''):
        j_val = 'No'
    elif isinstance(h_val, datetime):
        j_val = 'Yes'
    elif isinstance(h_val, str) and h_val.strip().lower() == 'yes':
        j_val = 'Yes'
    else:
        try:
            # Try parsing as a date
            parsed = datetime.strptime(str(h_val), '%Y-%m-%d')
            j_val = 'Yes'
        except Exception:
            j_val = 'No'
    ws[f'J{row}'] = j_val

wb.save(output_path)