from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_535-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_535-20_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Identify columns to delete (with '/description' in the first row)
del_cols = []
for idx, cell in enumerate(ws[1], 1):
    if cell.value is not None and '/description' in str(cell.value):
        del_cols.append(idx)

# Delete columns from right to left to avoid index shift
for col_idx in sorted(del_cols, reverse=True):
    ws.delete_cols(col_idx)

wb.save(output_path)
