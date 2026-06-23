import openpyxl

# Paths
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_55060_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_55060_tc1/output.xlsx'

# Load workbook and select active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get the value from I12
value = ws['I12'].value

# Apply the logic to cells J23:N23
for col in range(10, 15):  # J=10, K=11, L=12, M=13, N=14
    ws.cell(row=23, column=col, value=(value if value not in (None, '') else None))

# Save the workbook
wb.save(output_path)
