import openpyxl
from datetime import time

# Load workbook and sheet
task_input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_53161_tc1/input.xlsx'
task_output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_53161_tc1/output.xlsx'

wb = openpyxl.load_workbook(task_input_path)
sheet = wb['Sheet1']

# Parse agent-category table (row 4-9, col F-N)
agents = []
agent_categories = {}
for row in range(4, 10):
    agent = sheet.cell(row=row, column=5).value
    if agent is None:
        continue
    agents.append(agent)
    # Category columns are 6-14 (inclusive)
    categories = []
    for col in range(6, 15):
        val = sheet.cell(row=row, column=col).value
        categories.append(val if val else 0)
    agent_categories[agent] = categories
num_categories = len(categories)  # Should be 9

# Parse schedule time intervals (row 12, col F-AE)
time_intervals = []
for col in range(6, 32):
    t = sheet.cell(row=12, column=col).value
    time_intervals.append(t)

# Parse agent schedule table (row 13-18, col E-AE)
schedules = {}
for row in range(13, 19):
    agent = sheet.cell(row=row, column=5).value
    if agent is None:
        continue
    schedule = []
    for col in range(6, 32):
        val = sheet.cell(row=row, column=col).value
        schedule.append(val if val else 0)
    schedules[agent] = schedule

# Calculate total count of categories per interval
result = []
for cat_idx in range(num_categories):
    cat_row = [f'Category {cat_idx+1}']
    for interval_idx in range(len(time_intervals)):
        total = 0
        for agent in agents:
            # If agent is scheduled and has this category
            if schedules.get(agent, [0]*len(time_intervals))[interval_idx]:
                if agent_categories.get(agent, [0]*num_categories)[cat_idx]:
                    total += 1
        cat_row.append(total)
    result.append(cat_row)

# Fill in cells F22:AE29
start_row, end_row = 22, 22+num_categories
start_col = 6
for i, cat_row in enumerate(result):
    for j, val in enumerate(cat_row):
        sheet.cell(row=start_row + i, column=start_col + j, value=val)

wb.save(task_output_path)
