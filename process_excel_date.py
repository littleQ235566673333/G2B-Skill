import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from calendar import month_abbr
import re

# Input and output file paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_51354_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_51354_tc1/output.xlsx'

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Month mapping for abbreviations
month_map = {m: i for i, m in enumerate(month_abbr) if m}

# Define the fill color for column D
yellow_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

for row in range(2, 7):  # Assuming rows 2-6 based on E2:E6 indicated
    cell_value = ws[f'A{row}'].value
    # Extract the last 6 or 7 chars which should be like 'Oct 21', 'Aug 21', etc.
    match = re.search(r'([A-Za-z]{3})\s*(\d{2})$', cell_value.strip())
    if match:
        mon_abbr, year_2d = match.groups()
        # Get current date
        if mon_abbr in month_map:
            mon_num = month_map[mon_abbr]
            # Build full date
            # To avoid ambiguity, take years 00-99 as 2000-2099
            year_full = 2000 + int(year_2d)
            try:
                date = datetime(year_full, mon_num, 1)
                # Add one month
                if date.month == 12:
                    new_year = date.year + 1
                    new_month = 1
                else:
                    new_year = date.year
                    new_month = date.month + 1
                # Output format
                new_mon_abbr = month_abbr[new_month]
                new_year_2d = str(new_year)[2:]
                ws[f'E{row}'].value = f'{new_mon_abbr} {new_year_2d}'
                # Also put just '21' (year) in D column, with color
                ws[f'D{row}'].value = year_2d
                ws[f'D{row}'].fill = yellow_fill
            except Exception as e:
                ws[f'E{row}'].value = 'ERROR'
                ws[f'D{row}'].value = ''
        else:
            ws[f'E{row}'].value = 'ERROR'
            ws[f'D{row}'].value = ''
    else:
        ws[f'E{row}'].value = 'ERROR'
        ws[f'D{row}'].value = ''

# Save the workbook
wb.save(output_path)
