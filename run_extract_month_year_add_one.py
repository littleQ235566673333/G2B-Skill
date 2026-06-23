import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load workbook and worksheet
wb = load_workbook('results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r2/eval_51354_tc1/input.xlsx')
ws = wb['Sheet1']

# Month mapping for date handling
month_map = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}
month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

for i in range(2, 7):  # E2:E6 + D cells
    val = ws.cell(row=i, column=1).value
    # Extract last 6 or 7 chars, to cover all variants
    date_match = re.search(r'([A-Za-z]{3}) ?([0-9]{2})$', val.strip() if val else '')
    out_E = ''
    out_D = ''
    if date_match:
        mon = date_match.group(1).title()
        yr  = date_match.group(2)
        out_D = yr
        # Find next month
        idx = (month_map.get(mon.lower(), 0) % 12)  # wrap-around
        new_mon = month_names[idx]
        new_val = f'{new_mon} {yr}'
        out_E = new_val
    ws.cell(row=i, column=4, value=out_D)
    ws.cell(row=i, column=4).fill = fill
    ws.cell(row=i, column=5, value=out_E)

wb.save('results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r2/eval_51354_tc1/output.xlsx')
print('Done')
