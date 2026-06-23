import openpyxl
from collections import defaultdict

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_50971_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_50971_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read columns A and B into lists, skipping header
data_a = [ws[f'A{row}'].value for row in range(2, ws.max_row + 1)]
data_b = [ws[f'B{row}'].value for row in range(2, ws.max_row + 1)]

# Build dictionary mapping values in A to all B values
agroup = defaultdict(list)
for a, b in zip(data_a, data_b):
    agroup[a].append(b)

# Only keep values in B where A is duplicated (count > 1)
duplicate_b = []
for a in agroup:
    if len(agroup[a]) > 1:
        duplicate_b.extend(agroup[a])

# Write result to G3 downward
for idx, val in enumerate(duplicate_b, start=3):
    ws[f'G{idx}'] = val

wb.save(output_path)
