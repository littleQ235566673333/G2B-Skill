import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_4/task_147-48/r1/evolve_147-48/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_4/task_147-48/r1/evolve_147-48/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Collect non-empty values from column A
col_a_data = []
for row in range(1, 1001):
    val = ws.cell(row=row, column=1).value
    if val is None:
        break
    col_a_data.append(val)

# Transpose and place into C1:K6 (row-major order, 6 rows x 9 columns)
rows, cols = 6, 9
for r in range(rows):
    for c in range(cols):
        idx = r * cols + c
        if idx < len(col_a_data):
            ws.cell(row=1+r, column=3+c, value=col_a_data[idx])
        else:
            ws.cell(row=1+r, column=3+c, value=None)

wb.save(output_path)
