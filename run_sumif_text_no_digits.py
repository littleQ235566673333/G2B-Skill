import re
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_4/regression_gate/before_pass/core_42181/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_4/regression_gate/before_pass/core_42181/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# text to search in criteria
text_to_search = str(ws['A13'].value)
sum_result = 0

for row in range(4, 11):
    criteria_cell = ws[f'I{row}'].value
    sum_cell = ws[f'B{row}'].value
    # Only sum if:
    # 1. criteria_cell contains text_to_search (anywhere)
    # 2. criteria_cell contains NO digits (0-9)
    if isinstance(criteria_cell, str) and text_to_search in criteria_cell:
        if not re.search(r'\d', criteria_cell):
            if isinstance(sum_cell, (int, float)):
                sum_result += sum_cell

ws['B13'] = sum_result
wb.save(output_path)
