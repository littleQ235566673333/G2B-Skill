import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_3/group_4714/r1/evolve_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_3/group_4714/r1/evolve_4714/output.xlsx'

# Read input data
df = pd.read_excel(input_path)

# Identify columns explicitly
name_col = df.columns[0]
month_col = df.columns[2]
hours_col = df.columns[3]

# Convert months to datetime if needed
df[month_col] = pd.to_datetime(df[month_col])
results = []

for idx, row in df.iterrows():
    emp = row[name_col]
    cur_month = row[month_col]
    # All records for this employee, within the rolling 4-month window
    emp_rows = df[df[name_col] == emp]
    start_window = cur_month - pd.DateOffset(months=3)
    window_data = emp_rows[(emp_rows[month_col] <= cur_month) & (emp_rows[month_col] >= start_window)][hours_col]
    if len(window_data) < 4:
        results.append('n/a')
    else:
        avg = window_data.mean()
        # Output as int if decimal is zero, else show 2 decimals
        res = int(avg) if avg == int(avg) else round(avg, 2)
        results.append(res)

# Write results to E2:E25
wb = load_workbook(input_path)
ws = wb.active
for i, val in enumerate(results[:24], start=2):
    ws[f'E{i}'] = val
wb.save(output_path)
