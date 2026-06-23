from openpyxl import load_workbook
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_1/regression_gate/before_pass/core_290-27/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_1/regression_gate/before_pass/core_290-27/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

start_row, end_row = 14, 137
col = 2  # Column B

for row in range(start_row, end_row + 1):
    cell = ws.cell(row=row, column=col)
    val = cell.value
    # Process only if value is a string
    if isinstance(val, str):
        # Pattern 1: 2 uppercase then space then digits, e.g., 'GG 1'
        m1 = re.match(r'^([A-Z]{2} )([0-9]+)$', val)
        # Pattern 2: 3 uppercase then digits, e.g., 'PID1'
        m2 = re.match(r'^([A-Z]{3})([0-9]+)$', val)
        if m1:
            cell.value = m1.group(2)
        elif m2:
            cell.value = m2.group(2)
        else:
            # Remove all leading alpha and spaces
            modified = re.sub(r'^[A-Z ]+', '', val)
            # Only set if the rest is digits
            if modified.isdigit():
                cell.value = modified

wb.save(output_path)
