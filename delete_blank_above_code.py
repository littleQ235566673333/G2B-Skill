import openpyxl
input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_9/task_374-31/r3/evolve_374-31/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_9/task_374-31/r3/evolve_374-31/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']
# Find rows with 'Code' in column A
code_rows = []
for idx, row in enumerate(ws.iter_rows(min_row=1, max_col=1, values_only=True), start=1):
    if row[0] is not None and str(row[0]).strip() == 'Code':
        code_rows.append(idx)
# Collect blank rows above each 'Code' cell
blank_rows_to_delete = set()
for code_row in code_rows:
    r = code_row - 1
    while r > 0:
        cell = ws.cell(row=r, column=1).value
        if cell is None or (isinstance(cell, str) and cell.strip() == ""):
            blank_rows_to_delete.add(r)
            r -= 1
        else:
            break
# Delete blank rows from bottom up
for r in sorted(blank_rows_to_delete, reverse=True):
    ws.delete_rows(r)
wb.save(output_path)
print('SUCCESS')
