import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_7/regression_gate/before_fix/core_263-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_7/regression_gate/before_fix/core_263-1/output.xlsx'

# Explicitly use the detected columns
material_col = 'Mtrl'
width_col = 'Width'
height_col = 'Height'

df = pd.read_excel(input_path)

# Clean-up and calculate area
usable_df = df[[material_col, width_col, height_col]].dropna()
usable_df[width_col] = pd.to_numeric(usable_df[width_col], errors='coerce')
usable_df[height_col] = pd.to_numeric(usable_df[height_col], errors='coerce')
usable_df = usable_df.dropna(subset=[material_col, width_col, height_col])
usable_df['Area'] = usable_df[width_col] * usable_df[height_col]

# Group by material, sum area
result = usable_df.groupby(material_col)['Area'].sum().reset_index()
result = result.sort_values(material_col).reset_index(drop=True)

# Open workbook and write results
wb = load_workbook(input_path)
ws = wb.active

for idx in range(3):
    val = None
    if idx < len(result):
        val = f"{result.iloc[idx, 0]}: {result.iloc[idx, 1]:.2f}"
    ws[f'H{2+idx}'] = val

wb.save(output_path)
