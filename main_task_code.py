import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_3/group_118-50/r0/evolve_118-50/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_3/group_118-50/r0/evolve_118-50/output.xlsx'

# Load input.xlsx Sheet1 and read up to 55555 rows
words = pd.read_excel(input_path, sheet_name='Sheet1', usecols=[0], engine='openpyxl', header=None, names=['word'], nrows=55555)
words = words.dropna(subset=['word'])
words['word'] = words['word'].astype(str)

# Sort the words alphabetically
words.sort_values(by='word', inplace=True)
words.reset_index(drop=True, inplace=True)

# Define endings order
endings_order = ['ING', 'ERS', 'ATE', 'EST', 'ONE', 'IER', 'ILY']
def get_group(word):
    for ending in endings_order:
        if word.endswith(ending):
            return ending
    return word[-3:] if len(word) >= 3 else ''

def transform(word):
    if len(word) >= 5:
        return word[4] + word[:4] + word[5:]
    return None

# Find pairs
results = []
words_set = set(words['word'])
for word in words['word']:
    if len(word) < 5:
        continue
    t = transform(word)
    # Transformation must match another word, different from itself and same ending
    if t and t != word and t in words_set and word[-3:] == t[-3:]:
        group = get_group(word)
        results.append((group, t, word))

# Group and sort results as requested
def ending_key(group):
    if group in endings_order:
        return endings_order.index(group)
    else:
        return len(endings_order)
results.sort(key=lambda x: (ending_key(x[0]), x[1], x[2]))

# Write results to columns C and D in the output
wb = load_workbook(input_path)
ws = wb['Sheet1']
# Clear columns C and D
for row in range(2, 5001):
    ws.cell(row=row, column=3, value=None)
    ws.cell(row=row, column=4, value=None)
# Paste up to 4999 results
for i, (group, transformed, original) in enumerate(results[:4999]):
    ws.cell(row=i+2, column=3, value=transformed)
    ws.cell(row=i+2, column=4, value=original)
wb.save(output_path)
