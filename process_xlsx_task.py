import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_2/regression_gate/after_fix/core_38985/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_2/regression_gate/after_fix/core_38985/output.xlsx'

# Read in Sheet1, no header, to process manually
sheet = pd.read_excel(input_path, sheet_name='Sheet1', header=None)

# Find Table 1 and Table 2 positions
# Table 1 header is 'Column 1', Table 2 header is 'Column 2'
tbl1_col, tbl1_row = None, None
tbl2_col, tbl2_row = None, None
for i, row in sheet.iterrows():
    for j, val in enumerate(row):
        if val == 'Column 1':
            tbl1_col, tbl1_row = j, i
        if val == 'Column 2':
            tbl2_col, tbl2_row = j, i

# Get Table 1 as DataFrame
# Table 1 has: names in one column, starts after header
# Find out how many data rows (stop when hit next NaN)
row = tbl1_row + 1
names_tbl1 = []
while row < len(sheet) and pd.notnull(sheet.iloc[row, tbl1_col]):
    names_tbl1.append(sheet.iloc[row, tbl1_col])
    row += 1

# Get Table 2 as DataFrame
# Table 2 has: name + one or more columns with values (Value 1, Value 2, ...)
tbl2_headers = sheet.iloc[tbl2_row, tbl2_col:tbl2_col+3] # at least 3 cols
col2_idx = tbl2_col
val1_idx = tbl2_col + 1
val2_idx = tbl2_col + 2
row = tbl2_row + 1
table2_list = []
while row < len(sheet):
    name = sheet.iloc[row, col2_idx]
    v1 = sheet.iloc[row, val1_idx]
    v2 = sheet.iloc[row, val2_idx]
    if pd.isnull(name):
        break
    table2_list.append([name, v1, v2])
    row += 1

'table2_list' # [['AAA','Red','Black'], ...]
# Table 2 is long form; transform to: name -> [v1, v2]
lookup = {}
for row in table2_list:
    name = row[0]
    v1, v2 = row[1], row[2]
    lookup.setdefault(name, []).extend([v1, v2])

# For each duplicate name in input table, retrieve data transposed from Table 2.
name_counts = {}
output_vals = []
for name in names_tbl1:
    seen = name_counts.get(name, 0)
    name_counts[name] = seen + 1
    vals = lookup.get(name, [])
    val = vals[seen] if seen < len(vals) else None
    output_vals.append(val)

# Write result to D8:D11
wb = load_workbook(input_path)
ws = wb['Sheet1']
for i in range(4):
    ws[f'D{8+i}'] = output_vals[i] if i < len(output_vals) else None
wb.save(output_path)
