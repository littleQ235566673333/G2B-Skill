import openpyxl
import pandas as pd

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun1/eval_262-17_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun1/eval_262-17_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_file)
ws = wb['Sheet1']

# Read header
header = [ws.cell(row=1, column=i+1).value for i in range(ws.max_column)]

df = pd.read_excel(input_file, sheet_name='Sheet1', header=0)

# Find column indexes (0-based for pandas)
task_col = header.index('Task')
resp_col = header.index('Responsibility')

df_sorted = df.sort_values(by=[header[task_col], header[resp_col]], ascending=[True, True], kind='mergesort')

# Write back sorted data into Sheet1 (A2...)
for i, row in enumerate(df_sorted.itertuples(index=False), start=2):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j).value = val

wb.save(output_file)
