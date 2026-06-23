from openpyxl import load_workbook
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_1/regression_gate/after_pass/core_290-27/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_1/regression_gate/after_pass/core_290-27/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

def process_value(val):
    if isinstance(val, str):
        m = re.match(r'^([A-Z]{2,3}) ?(\d+.*)$', val)
        if m:
            return m.group(2)
        # fallback: remove leading uppercase and spaces
        return re.sub(r'^[A-Z ]+', '', val)
    return val

for row in ws.iter_rows(min_row=15, max_row=138, min_col=2, max_col=2):
    cell = row[0]
    old_val = cell.value
    # Only process if value is a string and follows pattern
    if type(old_val) is str and re.match(r'^[A-Z ]+[0-9]', old_val):
        new_val = process_value(old_val)
        cell.value = new_val

wb.save(output_path)
