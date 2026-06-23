import openpyxl
from collections import defaultdict
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_41978_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_41978_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Cumulative']

years = [ws[f'G{row}'].value for row in range(14, 186)]
statuses = [ws[f'J{row}'].value for row in range(14, 186)]

# Count number of open cases per year
year_open_count = defaultdict(int)
for yr, st in zip(years, statuses):
    if yr is not None and str(st).strip().lower() == 'open':
        year_open_count[yr] += 1

# Get unique years in sorted order
unique_years = sorted(set(y for y in years if y is not None))

# Write counts for each unique year into cells I2:I11, preserving original formatting
for idx, yr in enumerate(unique_years[:10]):  # Only fill up to I11 (10 slots)
    cell = ws[f'I{idx+2}']
    cell.value = year_open_count[yr]

wb.save(output_path)
