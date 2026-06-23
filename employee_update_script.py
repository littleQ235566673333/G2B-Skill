import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_32093_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_32093_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 16):  # Rows 2 to 15 inclusive
    b = ws.cell(row, 2).value
    c = ws.cell(row, 3).value
    d = ws.cell(row, 4).value
    e = ws.cell(row, 5).value
    # Strip spaces, use empty string if None
    b = b.strip() if b else ''
    c = c.strip() if c else ''
    d = d.strip() if d else ''
    e = e.strip() if e else ''

    # Find first non-empty new employee name, else use current employee
    new_emp = c or d or e or b
    ws.cell(row, 6, new_emp)

wb.save(output_path)
