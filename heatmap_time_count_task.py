import openpyxl
from datetime import time

INPUT_FILE = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/group_52305/r2/evolve_52305/input.xlsx'
OUTPUT_FILE = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/group_52305/r2/evolve_52305/output.xlsx'
SHEET_NAME = 'Sheet1'

# Time criteria
START_TIME = time(21, 30)
END_TIME = time(22, 0)

# Load workbook and sheet
wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb[SHEET_NAME]

# Gather account names from the heat map area J6:N24
name_rows = []
for row in ws.iter_rows(min_row=6, max_row=24, min_col=10, max_col=14):
    name_rows.append([cell.value for cell in row])

# Build list of unique names from column C (Name)
names = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    val = row[0].value
    if val and isinstance(val, str):
        names.add(val)
names = sorted(list(names))

# For reference, print names found
def extract_time(dt):
    if hasattr(dt, 'time'):
        return dt.time()
    return None

# Fill J6:N24 based on criteria
row_offset = 6
col_offset = 10
for i, heat_row in enumerate(ws.iter_rows(min_row=row_offset, max_row=24, min_col=col_offset, max_col=col_offset+4)):
    # Determine which Name this row represents
    heat_name = None
    if i < len(names):
        heat_name = names[i]
    else:
        continue  # skip if no matching name
    # For each cell in this row
    for j, cell in enumerate(heat_row):
        # Count matching entries
        count = 0
        for data_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            name_val = data_row[2].value  # Column C
            time_val = data_row[1].value  # Column B
            if name_val == heat_name:
                t = extract_time(time_val)
                if t and START_TIME <= t < END_TIME:
                    count += 1
        cell.value = count

# Save
wb.save(OUTPUT_FILE)
