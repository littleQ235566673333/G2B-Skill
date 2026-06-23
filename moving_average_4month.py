import openpyxl
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_6/regression_gate/after_fix/core_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_6/regression_gate/after_fix/core_4714/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Assume header row is row 1.
# Employees: A2:A25, Months: C2:C25, Hours: D2:D25
names = []
months = []
hours = []
for row in range(2, 26):
    names.append(ws[f'A{row}'].value)
    months.append(ws[f'C{row}'].value)
    hours.append(ws[f'D{row}'].value)

results = []

for idx in range(0, len(names)):
    emp_name = names[idx]
    emp_month = months[idx]
    # For the current row, find all previous rows (including current) for the same employee
    prior_entries = []
    for j in range(0, idx+1):
        if names[j] == emp_name:
            prior_entries.append((months[j], hours[j]))
    # Sort by month ascending (assuming C is sortable as date or integer)
    prior_entries = sorted(prior_entries, key=lambda x: x[0])
    # Take last 4 entries
    if len(prior_entries) < 4:
        results.append("n/a")
    else:
        last_4 = prior_entries[-4:]
        avg = sum([x[1] for x in last_4])/4
        # Formatting: show as int if decimal is zero
        if avg == int(avg):
            avg_display = str(int(avg))
        else:
            avg_display = str(round(avg, 2))
        results.append(avg_display)

# Write E2:E25
for i, val in enumerate(results):
    ws[f'E{i+2}'] = val

wb.save(output_path)
