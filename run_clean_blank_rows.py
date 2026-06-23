from openpyxl import load_workbook

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_2/regression_gate/after_pass/core_160-6/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_2/regression_gate/after_pass/core_160-6/output.xlsx'

wb = load_workbook(in_path)
ws = wb['SH']

# Read rows 2-12, columns 1-12 (A-L)
rows = list(ws.iter_rows(min_row=2, max_row=12, max_col=12, values_only=True))
# Filter out rows where all values are None
filtered = [row for row in rows if any(cell is not None for cell in row)]

# Write filtered rows to target output range A6:L11 (6 rows max)
for rel_i in range(6):
    row = filtered[rel_i] if rel_i < len(filtered) else [None]*12
    for j, value in enumerate(row, 1):
        ws.cell(row=6+rel_i, column=j, value=value)

wb.save(out_path)
