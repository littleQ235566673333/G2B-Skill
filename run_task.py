import openpyxl
from openpyxl.styles import Alignment

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_56378_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_56378_tc1/output.xlsx'

# Load workbook & sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Assume frame 1 is on columns A:J (expand if necessary).
# And assumes row 1 contains headers, with 'QUANTITY UNITS' header somewhere in columns.
headers = [cell.value for cell in ws[1]]
q_col_idx = headers.index('QUANTITY UNITS') + 1  # 1-based index
product_col_idx = headers.index('Product') + 1  # This may appear as 'Product' or similar

# Collect relevant (non-empty QUANTITY UNITS) rows from frame 1 (excluding header)
selected_rows = []
for row in ws.iter_rows(min_row=2, max_row=20):  # Assume first 20 rows cover frame 1
    quantity = row[q_col_idx-1].value
    if quantity not in (None, ''):
        # Collect all columns between L:R or as many as match original frame
        selected_rows.append([cell.value for cell in row])

# Map selected data to L5:R8 for up to 4 products
output_start_row = 5
output_end_row = output_start_row + 4 - 1
output_start_col = openpyxl.utils.column_index_from_string('L')
output_end_col = openpyxl.utils.column_index_from_string('R')

# Headers: Copy from frame 1 to frame 2 (mapped to L:R)
output_headers = headers[product_col_idx-1:product_col_idx+(output_end_col-output_start_col)]
for c, val in enumerate(output_headers, start=output_start_col):
    ws.cell(row=output_start_row, column=c, value=val)

# Data: Write to L6:R8
for i, row_data in enumerate(selected_rows[:output_end_row-output_start_row]):
    for j, val in enumerate(row_data[product_col_idx-1:product_col_idx+(output_end_col-output_start_col)]):
        ws.cell(row=output_start_row + 1 + i, column=output_start_col + j, value=val)

# Formatting: Left align product names in col L, right align O-R
for r in range(output_start_row+1, output_end_row+1):
    ws.cell(row=r, column=output_start_col).alignment = Alignment(horizontal='left')
    for c in range(output_start_col+3, output_end_col+1):
        ws.cell(row=r, column=c).alignment = Alignment(horizontal='right')

wb.save(output_path)
print('Done')
