from openpyxl import load_workbook

# Open the given workbook and select the correct worksheet
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_1/group_44017/r3/evolve_44017/input.xlsx')
ws = wb['Data']

# Column indices
col_base = 23 # W (P07 Old rate)
col_freq = 10 # J
col_eff = 12 # L
col_inc1 = 13 # M
col_inc2 = 14 # N
col_inc3 = 15 # O
col_inc4 = 16 # P
col_out_start = 30 # AD
col_out_end = 41 # AO
row_month_header = 9
row_data_start = 14
row_data_end = 42

for c in range(col_out_start, col_out_end+1):
    for r in range(row_data_start, row_data_end+1):
        month_cell = ws.cell(row=row_month_header, column=c).coordinate
        formula = f'=IF({month_cell}<$L{r},"",$W{r}*'
        months_since_start = f'(DATEDIF($L{r},{month_cell},"m")+1)'
        waves = f'FLOOR(MAX({months_since_start}/$J{r},0),1)+1'
        formula += f'(1+IF({waves}>=1,$M{r},0))'
        formula += f'*(1+IF({waves}>=2,$N{r},0))'
        formula += f'*(1+IF({waves}>=3,$O{r},0))'
        formula += f'*(1+IF({waves}>=4,$P{r},0))'
        formula += ')'
        ws.cell(row=r, column=c, value=formula)

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_1/group_44017/r3/evolve_44017/output.xlsx')
print('Done')
