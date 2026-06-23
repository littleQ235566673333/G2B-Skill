import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import re

def is_formula(cell):
    if cell.data_type == 'f':
        return True
    if isinstance(cell.value, str) and cell.value.strip().startswith('='):
        return True
    return False

def convert_text_formula(cell):
    if isinstance(cell.value, str) and cell.value.strip().startswith('='):
        cell.value = cell.value.strip()
        cell.data_type = 'f'

def is_row_blank(sheet, row_idx, start_col=1, end_col=8):
    for col in range(start_col, end_col+1):
        v = sheet.cell(row=row_idx, column=col).value
        if v not in (None, ''):
            return False
    return True

def get_blocks(sheet, start_row, col_count):
    # Returns [(block_row_start, block_row_end), ...]
    in_block = False
    block_start = None
    blocks = []
    max_row = sheet.max_row
    for row in range(start_row, max_row+2):
        if row > max_row or is_row_blank(sheet, row, 1, col_count):
            if in_block:
                blocks.append((block_start, row-1))
                in_block = False
        else:
            if not in_block:
                in_block = True
                block_start = row
    return blocks

def evaluate_formula_result(cell):
    # openpyxl cannot evaluate formulas, so just return the cached value or None
    return cell.value if cell.data_type != 'f' else None

def get_formula_ref_adjusted(cell, new_row):
    # Adjusts a simple cell ref formula to new row for this task's common =Brow+Crow type formulas
    if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value and cell.value[0]=='='):
        raw_formula = cell.value if not cell.data_type == 'f' else cell.value
        # Substitute all cell refs in the formula string to the appropriate row
        # Only works for simple ref pattern (e.g., =B3+C3)
        def repl(m):
            col, _ = m.group(1), m.group(2)
            return f"{col}{new_row}"
        formula = re.sub(r"([A-Z]+)([0-9]+)", repl, raw_formula)
        return formula
    return cell.value

def apply_number_format(cell, value):
    # number formatting for D,E,F: one decimal if not int, else int with no .0
    if isinstance(value, (int, float)):
        if float(value) == int(value):
            cell.number_format = '#,##0'
        else:
            cell.number_format = '#,##0.0'
    else:
        cell.number_format = 'General'

def format_bold_center(cell):
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)

def process_sheet(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb['Sheet1']
    max_col = 8 # Up to H

    # Ensure any text-stored formulas in D,E,F are formulas
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=4, max_col=6):
        for cell in row:
            convert_text_formula(cell)

    blocks = get_blocks(ws, 1, max_col)

    # Read all rows into memory, process
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col))
    new_rows = []

    for block_start, block_end in blocks:
        block = all_rows[block_start-1:block_end]
        # Convert all formulas in D–F to text so we can sort by evaluated value
        sortable_block = []
        for row in block:
            row_data = []
            for idx, cell in enumerate(row):
                v = cell.value
                # If D,E,F formula, keep the formula string to restore, else copy
                if 3 <= idx <= 5 and is_formula(cell):
                    formula = cell.value if isinstance(cell.value, str) else cell.value
                    row_data.append((v, formula))
                else:
                    row_data.append((v, None))
            sortable_block.append(row_data)
        # Sort the block by column D, descending
        try:
            sortable_block.sort(key=lambda r: float(r[3][0]) if r[3][0] not in (None, '') else float('-inf'), reverse=True)
        except:
            # Sometimes D is not all floattable: skip or use fallback
            pass
        new_rows.extend(sortable_block)
        # Append a blank row as separator (except after last block)
        if block_end != blocks[-1][1]:
            new_rows.append([(None, None)]*max_col)

    # Clear and rebuild ws so we can keep formatting
    for row in ws['A1:H85']:
        for cell in row:
            cell.value = None
            cell.font = Font(color='000000', bold=False)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.number_format = 'General'
    # Fill new values row by row
    for dest_idx, rdata in enumerate(new_rows):
        if dest_idx >= 85:
            break
        for col_idx, (cell_val, cell_formula) in enumerate(rdata):
            cell = ws.cell(row=dest_idx+1, column=col_idx+1)
            if 3 <= col_idx <= 5 and cell_formula and (isinstance(cell_formula, str) or cell_formula):
                # Re-insert formula: =Brow+Crow ref to correct row (adjust row number in formula string)
                new_formula = get_formula_ref_adjusted(ws.cell(row=dest_idx+1, column=col_idx+1), dest_idx+1)
                cell.value = cell_formula if not new_formula else new_formula
                cell.data_type = 'f'
            else:
                cell.value = cell_val
            # Apply formats for D–F
            if 3 <= col_idx <= 5:
                format_bold_center(cell)
                v = cell_val
                if is_formula(cell):
                    # Don't format formulas (display as calculated)
                    cell.number_format = '#,##0.0'
                else:
                    apply_number_format(cell, v)
    # Set column widths to original
    for col in range(1, max_col+1):
        ws.column_dimensions[get_column_letter(col)].width = ws.column_dimensions[get_column_letter(col)].width
    # Save output
    wb.save(output_path)

process_sheet(
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_191-40_tc1/input.xlsx',
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_191-40_tc1/output.xlsx'
)
