from openpyxl import load_workbook

infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_486-17_tc1/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_486-17_tc1/output.xlsx'

wb = load_workbook(infile)
ws = wb['Blad1']
max_row = ws.max_row

for row in range(2, max_row+1):
    val = ws[f'A{row}'].value
    if val and isinstance(val, str) and len(val) == 9 and val[0] == '0':
        year = val[1:5]
        month = val[5:7]
        day = val[7:9]
        ws[f'B{row}'] = f'{year} {month} {day}'
    else:
        ws[f'B{row}'] = None

wb.save(outfile)
