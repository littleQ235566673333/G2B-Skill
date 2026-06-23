import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_3/group_57033/r0/evolve_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_3/group_57033/r0/evolve_57033/output.xlsx'

# Read all sheets
sheets = pd.read_excel(input_path, sheet_name=None)

df4 = sheets['Sheet4']
mdf = sheets['CBtrans']  # Use CBtrans as reference for match

# Columns: find company/account/xchar (case-insensitive)
def find_col(df, colname):
    for c in df.columns:
        if c.strip().lower() == colname.lower():
            return c
    raise ValueError(f'Column {colname} not found')

# Get the right columns
company_col_4 = find_col(df4, 'company')
account_col_4 = find_col(df4, 'account')
xchar_col_4 = find_col(df4, 'xchar')
company_col_md = find_col(mdf, 'company')
account_col_md = find_col(mdf, 'account')
xchar_col_md = find_col(mdf, 'xchar')

# For each row in Sheet4 K2:K7, does (company, account, xchar) appear in CBtrans?
results = []
for i in range(1, 7):  # 2..7 (0-based in pandas)
    item = (df4.at[i, company_col_4], df4.at[i, account_col_4], df4.at[i, xchar_col_4])
    match = ((mdf[company_col_md]==item[0]) & (mdf[account_col_md]==item[1]) & (mdf[xchar_col_md]==item[2])).any()
    results.append('Match' if match else '-')

# Write results using openpyxl
wb = load_workbook(input_path)
ws = wb['Sheet4']
fill = PatternFill(start_color='FF66CC', end_color='FF66CC', fill_type='solid')
for idx, val in enumerate(results, start=2):
    cell = ws[f'K{idx}']
    cell.value = val.capitalize()  # Regular case
    cell.fill = fill
wb.save(output_path)
