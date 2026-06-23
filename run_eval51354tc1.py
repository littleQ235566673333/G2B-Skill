from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import calendar
import re

def add_one_month(month_str, year_str):
    month_map = {v.lower(): k for k,v in enumerate(calendar.month_abbr) if v}
    month_map_full = {v.lower(): k for k,v in enumerate(calendar.month_name) if v}
    month_name = month_str.lower()
    mm = month_map.get(month_name, None) or month_map_full.get(month_name, None)
    if mm is not None:
        year = int(year_str)
        if mm == 12:
            return (1, year+1)
        else:
            return (mm+1, year)
    return (None, None)

wbin = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun2/eval_51354_tc1/input.xlsx'
wout = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun2/eval_51354_tc1/output.xlsx'
wb = load_workbook(wbin)
ws = wb.active

fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

# For rows 2 to 6
for row in range(2, 7):
    text = ws[f'A{row}'].value or ''
    match = re.search(r'([A-Za-z]{3,9})[ .-]*([0-9]{2,4})$', text)
    if match:
        mmm = match.group(1)
        yy = match.group(2)[-2:] # always keep as 2-digit year
        ws[f'D{row}'] = yy
        ws[f'D{row}'].fill = fill
        mm_plus, yyyy_plus = add_one_month(mmm, '20'+yy if len(yy)==2 else yy)
        if mm_plus:
            new_date = f'{calendar.month_abbr[mm_plus]} {str(yyyy_plus)[-2:]}'
            ws[f'E{row}'] = new_date
        else:
            ws[f'E{row}'] = ''
    else:
        ws[f'D{row}'] = ''
        ws[f'E{row}'] = ''
        ws[f'D{row}'].fill = fill

wb.save(wout)
