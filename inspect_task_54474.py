from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_54474_s2/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_54474_s2/output.xlsx'

wb = load_workbook(input_path)
print(wb.sheetnames)
for ws in wb.worksheets:
    print('\nSHEET:', ws.title)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
        print(row)
