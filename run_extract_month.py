import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r3/eval_15380_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r3/eval_15380_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Find the 'codes2' and 'month' columns
codes2, month = [], []
header_row = None
for row in ws.iter_rows(min_row=1, max_row=20):
    if row[0].value == 'codes2' and row[1].value == 'month':
        header_row = row[0].row
        break
if header_row:
    i = 1
    while True:
        code_cell = ws.cell(row=header_row + i, column=1)
        month_cell = ws.cell(row=header_row + i, column=2)
        if not code_cell.value or not month_cell.value:
            break
        codes2.append(str(code_cell.value))
        month.append(str(month_cell.value))
        i += 1

# For A3:A14, search for codes2 element in A and place matching month in B
for i in range(3, 15):
    s = str(ws[f'A{i}'].value) if ws[f'A{i}'].value else ''
    found = ''
    for idx, code in enumerate(codes2):
        if code in s:
            found = month[idx] if idx < len(month) else ''
            break
    ws[f'B{i}'].value = found

wb.save(output_path)
