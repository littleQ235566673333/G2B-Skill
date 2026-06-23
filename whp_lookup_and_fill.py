import openpyxl

# Paths
i_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_7/group_54474/r2/evolve_54474/input.xlsx'
o_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_7/group_54474/r2/evolve_54474/output.xlsx'

wb = openpyxl.load_workbook(i_path)
whp_ws = wb['WHP']
whp_data_ws = wb['WHP DATA']

# Read site names from 'WHP' C7 and C8
site_cells = [(7, 3), (8, 3)]  # row, col for C7, C8
site_names = [whp_ws.cell(row=r, column=c).value for r, c in site_cells]

# Header row in WHP DATA is row 4 (Excel, 1-based)
header_rownum = 4
header_row = next(whp_data_ws.iter_rows(min_row=header_rownum, max_row=header_rownum, values_only=True))
# Map indices
site_col = 1        # 'Local London Partnership' (B)
data_cols = [2, 3, 4]  # EEG (C), Dis (D), LTU (E), 0-based

results = []
# Data starts at row 5
for site in site_names:
    found = False
    for row in whp_data_ws.iter_rows(min_row=header_rownum+1, values_only=True):
        if site and row[site_col] == site:
            extracted = [(row[i] if i < len(row) else None) for i in data_cols]
            results.append(extracted)
            found = True
            break
    if not found:
        results.append([None, None, None])

# Write to 'WHP'!E7:G8
for i, res in enumerate(results):
    for j, val in enumerate(res):
        whp_ws.cell(row=7+i, column=5+j).value = val

wb.save(o_path)
