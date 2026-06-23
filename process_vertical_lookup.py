import openpyxl

# Load the workbook and select the active sheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_8/group_10452/r2/evolve_10452/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_8/group_10452/r2/evolve_10452/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Process rows 4 to 12 (Excel rows)
for row in range(4, 13):
    cell_value = ws[f'B{row}'].value
    # If cell starts with 'PK', copy to E column, else blank
    if isinstance(cell_value, str) and cell_value.startswith('PK'):
        ws[f'E{row}'].value = cell_value
    else:
        ws[f'E{row}'].value = None

wb.save(output_path)
