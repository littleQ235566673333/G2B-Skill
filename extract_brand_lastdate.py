import openpyxl
import pandas as pd

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/after_fix/core_516-46/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/after_fix/core_516-46/output.xlsx'

wb = openpyxl.load_workbook(in_path)
ws = wb['ورقة1']

data = []
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    if row[0] is None:
        break
    data.append(row)

df = pd.DataFrame(data, columns=['date', 'brand', 'batch', 'origin', 'quantity'])

# Find last date per brand
df['date'] = pd.to_datetime(df['date'])
last_dates = df.groupby('brand')['date'].max()

res_rows = []
for brand, last_date in last_dates.items():
    rows_last = df[(df['brand'] == brand) & (df['date'] == last_date)]
    qty_sum = rows_last['quantity'].sum()
    batch = ','.join(rows_last['batch'].astype(str).unique())
    origin = ','.join(rows_last['origin'].astype(str).unique())
    modification = 'merged' if len(rows_last) > 1 else ''
    res_rows.append([last_date, brand, batch, origin, qty_sum, modification])

# Write results to H2:L...
header = ['date', 'brand', 'batch', 'origin', 'quantity', 'modification']
for j, h in enumerate(header):
    ws.cell(row=1, column=8+j, value=h)
for i, row in enumerate(res_rows):
    for j, val in enumerate(row):
        ws.cell(row=2+i, column=8+j, value=val)
wb.save(out_path)
