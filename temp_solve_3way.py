import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_4/regression_gate/before_fix/core_57033/input.xlsx')
ws4 = wb['Sheet4']
wsmd = wb['CBtrans']
# Map CBtrans headers
md_headers = [cell.value for cell in next(wsmd.iter_rows(min_row=1, max_row=1))]
company_md_ix = md_headers.index('company')
account_md_ix = md_headers.index('account')
xchar_md_ix = md_headers.index('xchar')
# Map Sheet4 headers
s4_headers = [cell.value for cell in next(ws4.iter_rows(min_row=1, max_row=1))]
company4_ix = s4_headers.index('Company')
account4_ix = s4_headers.index('account')
xchar4_ix = s4_headers.index('xchar')
# Col K is index 11 (1-based)
col_K = 11
fill = PatternFill(fill_type='solid', fgColor='FF66CC')
for row_idx in range(2, 8):
    s4_row = list(ws4.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    company = s4_row[company4_ix]
    account = s4_row[account4_ix]
    xchar = s4_row[xchar4_ix]
    match = False
    for md_row in wsmd.iter_rows(min_row=2, max_row=wsmd.max_row, values_only=True):
        if md_row[company_md_ix] == company and md_row[account_md_ix] == account and md_row[xchar_md_ix] == xchar:
            match = True
            break
    val = 'Match' if match else '-'
    cell = ws4.cell(row=row_idx, column=col_K, value=val if val == '-' else val.capitalize())
    cell.fill = fill
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_4/regression_gate/before_fix/core_57033/output.xlsx')
