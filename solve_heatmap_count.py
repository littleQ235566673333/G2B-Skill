import pandas as pd
from openpyxl import load_workbook
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_1/group_52305/r2/evolve_52305/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_1/group_52305/r2/evolve_52305/output.xlsx'

# Read all cell data with pandas, assuming main sheet
df = pd.read_excel(input_path)

# Try to infer the structure
possible_time_cols = [c for c in df.columns if 'time' in c.lower()]
possible_name_cols = [c for c in df.columns if 'name' in c.lower()]

# If J5:N5 has headers for time bins, and J6:J24 has names, extract them
wb = load_workbook(input_path)
ws = wb.active

# Get the header for heatmap columns (time bins)
time_bins = [ws.cell(row=5, column=col).value for col in range(10, 15)] # J:N columns (10-14)
# Get the Name list down J6:J24
name_list = [ws.cell(row=row, column=10).value for row in range(6, 25)]

def parse_time(t):
    if pd.isnull(t): return None
    if isinstance(t, datetime.time):
        return t
    if isinstance(t, datetime.datetime):
        return t.time()
    if isinstance(t, str):
        try:
            # Try HH:MM or HH:MM:SS
            parts = t.split(":")
            if len(parts)==2:
                return datetime.time(int(parts[0]), int(parts[1]))
            if len(parts)==3:
                return datetime.time(int(parts[0]), int(parts[1]), int(parts[2]))
        except:
            return None
    return None

# Get correct columns
time_col = possible_time_cols[0] if possible_time_cols else None
name_col = possible_name_cols[0] if possible_name_cols else None

def in_time_range(target, start, end):
    if not target: return False
    # Assume times never wrap midnight
    return start <= target <= end

def parse_bin_range(text):
    # Accept e.g. '21:30-22:00'
    if not text: return None, None
    try:
        span = text.split('-')
        t0 = parse_time(span[0].strip())
        t1 = parse_time(span[1].strip())
        return t0, t1
    except:
        return None, None

# Precompute bin ranges
time_bin_ranges = [parse_bin_range(tb) for tb in time_bins]

# Now count for each cell in heatmap
data_time_col = df[time_col].apply(parse_time) if time_col else None
results = []
for row_idx, name in enumerate(name_list):
    row_counts = []
    for bin_range in time_bin_ranges:
        if not all(bin_range):
            row_counts.append(None)
            continue
        # Entries where name and time match
        mask = (df[name_col] == name) if name_col else pd.Series([False]*len(df))
        t0, t1 = bin_range
        matched = mask & data_time_col.apply(lambda x: in_time_range(x, t0, t1))
        row_counts.append(int(matched.sum()))
    results.append(row_counts)

# Write the results to output file in J6:N24
for i, row_counts in enumerate(results):
    for j, val in enumerate(row_counts):
        ws.cell(row=6 + i, column=10 + j).value = val
wb.save(output_path)
