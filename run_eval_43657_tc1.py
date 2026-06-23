import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_43657_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_43657_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Jul-22 (2)']

def is_green(cell):
    fill: PatternFill = cell.fill
    if fill is not None and fill.fill_type is not None:
        fgColor = fill.fgColor
        if fgColor.type == 'rgb' or (hasattr(fgColor, 'rgb') and fgColor.rgb):
            # Common Excel greens and possible Excel color representations
            return fgColor.rgb in ['00B050', '92D050', 'C6EFCE', 'FF00B050', 'FF92D050', 'FFC6EFCE']
    return False

# Get the names (L2:L8)
name_cells = ['L'+str(r) for r in range(2,9)]
names = [ws[cell].value for cell in name_cells]

counts = []
# For each name, count occurrences in C3:G3 that are highlighted green
for idx, person in enumerate(names):
    count = 0
    for row in ws.iter_rows(min_row=3, max_row=3, min_col=3, max_col=7):
        for cell in row:
            if cell.value is not None and person is not None and str(person) in str(cell.value):
                if is_green(cell):
                    count += 1
    counts.append(count)
    ws[f'K{2+idx}'].value = count

wb.save(output_path)
print(counts)
