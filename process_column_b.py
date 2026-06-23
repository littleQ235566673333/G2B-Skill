from openpyxl import load_workbook
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v3/train/iter_2/regression_gate/after_pass/core_290-27/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v3/train/iter_2/regression_gate/after_pass/core_290-27/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def is_date(value):
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return True
    return False

def process_cell(val):
    if val is None:
        return val
    if is_date(val):
        return val
    m = re.match(r'^[A-Z ]*([0-9]+.*)', str(val))
    if m:
        return m.group(1)
    return val

for row in ws.iter_rows(min_row=14, max_row=137, min_col=2, max_col=2):
    cell = row[0]
    new_val = process_cell(cell.value)
    cell.value = new_val

wb.save(output_path)
