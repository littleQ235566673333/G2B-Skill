import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from calendar import month_abbr

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_51354_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_51354_tc1/output.xlsx'

# RGB #FFC000 => Color for column D
ffc000_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

# Month abbreviation lookup
month_mapping = {abbr: idx for idx, abbr in enumerate(month_abbr) if abbr}

def add_one_month(month, year):
    month_num = month_mapping[month]
    if month_num == 12:
        return (1, year + 1)
    else:
        return (month_num + 1, year)

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 7):  # E2:E6
    cell_a = ws[f'A{row}'].value
    if cell_a is None:
        ws[f'D{row}'] = ''
        ws[f'E{row}'] = ''
        continue
    # Extract last 6 chars, expected format 'MMM YY'
    last_six = cell_a[-6:]  # E.g., 'Oct 21'
    month_str, year_str = last_six[:3], last_six[-2:]
    # Column D: just year part
    ws[f'D{row}'] = year_str
    ws[f'D{row}'].fill = ffc000_fill
    # Add one month
    try:
        month_num = month_mapping[month_str]
        year_num = int(year_str)
        next_month, next_year = add_one_month(month_str, year_num)
        next_month_str = month_abbr[next_month]
        result_str = f'{next_month_str} {str(next_year).zfill(2)}'
    except Exception:
        result_str = ''
    ws[f'E{row}'] = result_str

wb.save(output_path)
