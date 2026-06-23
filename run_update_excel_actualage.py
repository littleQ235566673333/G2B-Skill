import openpyxl
from openpyxl.styles import Alignment, Font
from copy import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_7/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_7/regression_gate/after_pass/core_32337/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
header_row_idx = 2
headers = [str(cell.value).strip() if cell.value else '' for cell in ws[header_row_idx]]
header_map = {h: i + 1 for i, h in enumerate(headers) if h}

dob_col = header_map['DATE OF BIRTH']  # col C (3)
age_year_col = header_map['age (year)']  # e.g. col O (15)
year_age_col = header_map.get('YEAR (age)', 8)  # col H (8), in case used for lookup
cat_col = header_map['CATEGORY']  # col I (9)
actual_age_col = age_year_col + 1
# Insert new column for 'Actual Age' after 'age (year)'
ws.insert_cols(actual_age_col)
ws.cell(row=header_row_idx, column=actual_age_col, value='Actual Age')
age_fill = ws.cell(row=header_row_idx + 1, column=age_year_col).fill
max_row = ws.max_row
# Format and populate Actual Age
for r in range(header_row_idx + 1, max_row + 1):
    ws.cell(row=r, column=actual_age_col).fill = copy(age_fill)
    ws.cell(row=r, column=actual_age_col).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=actual_age_col).font = Font(bold=True)
    ws.cell(row=r, column=actual_age_col).number_format = '0'
ws.cell(row=header_row_idx, column=actual_age_col).alignment = Alignment(horizontal='center', vertical='top')
ws.cell(row=header_row_idx, column=actual_age_col).fill = copy(ws.cell(row=header_row_idx, column=age_year_col).fill)
ws.cell(row=header_row_idx, column=actual_age_col).font = Font(bold=True)

# Actual Age formula using $B$1 as REPORT Date, e.g.: =INT(($B$1-C3)/365.25)
dob_letter = openpyxl.utils.get_column_letter(dob_col)
for r in range(header_row_idx + 1, max_row + 1):
    ws.cell(row=r, column=actual_age_col).value = (
        f'=INT(($B$1-{dob_letter}{r})/365.25)')

# CATEGORY lookup in Expected Result (col E, index 5) for rows 3-15
cat_letter = openpyxl.utils.get_column_letter(cat_col)
age_letter = openpyxl.utils.get_column_letter(age_year_col)
actual_age_letter = openpyxl.utils.get_column_letter(actual_age_col)
for r in range(3, 16):
    # Use Actual Age to match age (year) and get corresponding CATEGORY
    formula = (
        f'=INDEX(${cat_letter}$3:${cat_letter}$15, '
        f'MATCH({actual_age_letter}{r}, ${age_letter}$3:${age_letter}$15, 0))'
    )
    ws.cell(row=r, column=5).value = formula

wb.save(output_path)
