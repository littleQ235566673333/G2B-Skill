import openpyxl
from collections import defaultdict
import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/group_177-6/r2/evolve_177-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/group_177-6/r2/evolve_177-6/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['DATA']
out_ws = wb['combined']
# Map columns
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
field_to_col = {name: idx+1 for idx, name in enumerate(headers)}
rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=18))
# Group rows by ComboKey (column H/8)
groups = defaultdict(list)
for r in rows:
    v = r[7].value
    if v is not None and str(v).strip() != '' and v != '\xa0':
        groups[v].append(r)
merged_rows = []
for groupkey, grouprows in groups.items():
    # Columns A-H: take from the first row in group
    out_row = [copy.copy(cell) for cell in grouprows[0][:8]]
    # Columns I-R: sum per column among group, treat blanks as 0
    for col in range(8, 18):
        total = 0
        for gr in grouprows:
            val = gr[col].value
            try:
                v = float(str(val).replace(',', '').strip()) if val not in [None, '', '\xa0'] else 0
            except Exception:
                v = 0
            total += v
        out_cell = copy.copy(grouprows[0][col])
        # Format: blank if 0, else 2 decimals, no comma
        if total == 0:
            out_cell.value = ''
        else:
            out_cell.value = f'{total:.2f}'
        out_row.append(out_cell)
    merged_rows.append(out_row)
# Write headers to out_ws
for c, cell in enumerate(ws[1][:18], 1):
    out_ws.cell(row=1, column=c).value = cell.value
    out_ws.cell(row=1, column=c)._style = copy.copy(cell._style)
# Write merged data
for r_idx, out_row in enumerate(merged_rows, 2):
    for c_idx, cell in enumerate(out_row, 1):
        out_ws.cell(row=r_idx, column=c_idx).value = cell.value
        out_ws.cell(row=r_idx, column=c_idx)._style = copy.copy(cell._style)
wb.save(output_path)
print('Merged rows:', len(merged_rows))
