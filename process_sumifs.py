import openpyxl
from datetime import datetime, timedelta

# File paths
input_fp = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_51262_tc1/input.xlsx'
output_fp = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_51262_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_fp)
ws = wb.active

# Assume data is in B6:D13 (dates, criteria, values), criteria list in F10:F14, date filter in F5
# Read date filter in F5
filter_date = ws['F5'].value
if isinstance(filter_date, datetime):
    year = filter_date.year
    month = filter_date.month
else:
    filter_date = datetime.strptime(str(filter_date), '%Y-%m-%d')
    year = filter_date.year
    month = filter_date.month

start_date = datetime(year, month, 1)
# Find last day of month
if month == 12:
    end_date = datetime(year+1, 1, 1) - timedelta(days=1)
else:
    end_date = datetime(year, month+1, 1) - timedelta(days=1)

# Read Criteria List in F10:F14
criteria_list = [ws[f'F{i}'].value for i in range(10,15) if ws[f'F{i}'].value is not None]

# Read table data
rows = []
for i in range(6, 14):
    row_date = ws[f'B{i}'].value
    row_criteria = ws[f'C{i}'].value
    row_value = ws[f'D{i}'].value
    rows.append((row_date, row_criteria, row_value))

# Sum values that match date range and criteria list
result_sum = 0
for row_date, row_criteria, row_value in rows:
    if row_date is None or row_criteria is None or row_value is None:
        continue
    # convert Excel date if needed
    if isinstance(row_date, datetime):
        rdate = row_date
    elif isinstance(row_date, (int, float)):
        rdate = openpyxl.utils.datetime.from_excel(row_date)
    else:
        try:
            rdate = datetime.strptime(str(row_date), '%Y-%m-%d')
        except:
            continue
    if start_date <= rdate <= end_date and row_criteria in criteria_list:
        result_sum += row_value

# Write result to F6:H6 as sum
for col, val in zip(['F','G','H'], [result_sum]*3):
    ws[f'{col}6'] = val

wb.save(output_fp)
print('Calculation done and result saved.')