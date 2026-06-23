import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_142-12_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_142-12_tc1/output.xlsx'
sheet_name = 'Sheet1'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb[sheet_name]

# Rows 1–15 (assuming header or content), Columns F (6) and J (10)
for row in range(1, 16):
    cell_f = ws.cell(row=row, column=6)
    cell_j = ws.cell(row=row, column=10)
    
    # Check if Column F is 'Marble Slab Creamery' and Column J is a whole number
    if cell_f.value == 'Marble Slab Creamery':
        amount = cell_j.value
        # Accept numeric or string representations
        if isinstance(amount, (int, float)):
            if float(amount) == int(amount):
                cell_f.value = 'Georgia State WH'
        elif isinstance(amount, str):
            try:
                amt_float = float(amount)
                if amt_float == int(amt_float):
                    cell_f.value = 'Georgia State WH'
            except ValueError:
                pass

# Save the result
wb.save(output_path)
