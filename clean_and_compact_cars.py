import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_5/regression_gate/after_pass/core_493-18/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_5/regression_gate/after_pass/core_493-18/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get all non-empty values in column F (F2:F7)
f_values = set()
for row in ws.iter_rows(min_row=2, max_row=7, min_col=6, max_col=6):
    val = row[0].value
    if val is not None:
        f_values.add(val)

# Gather matching car data
car_rows = []
for row_idx in range(2, 8):
    a_val = ws.cell(row=row_idx, column=1).value
    if a_val in f_values:
        car_rows.append([ws.cell(row=row_idx, column=col).value for col in range(1, 4)])
    # Clear A, B, C for non-matches
    else:
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).value = None

# Compact A, B, C by moving matched rows up, clear below
for idx, car_data in enumerate(car_rows):
    for col in range(1, 4):
        ws.cell(row=idx+2, column=col).value = car_data[col-1]
for row_idx in range(len(car_rows)+2, 8):
    for col in range(1, 4):
        ws.cell(row=row_idx, column=col).value = None

# Retain autofilter in range A1:C7
ws.auto_filter.ref = 'A1:C7'

wb.save(output_path)
