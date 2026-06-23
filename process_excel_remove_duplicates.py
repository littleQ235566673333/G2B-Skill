import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_448-11_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_448-11_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read all rows from A2:J (assume max_col=10)
rows = list(ws.iter_rows(min_row=2, max_col=10, values_only=True))

# Collect unique rows based on Column A
seen = set()
unique_rows = []
for row in rows:
    key = row[0]  # Column A value
    if key not in seen:
        seen.add(key)
        unique_rows.append(row)

# Delete all rows from A2 onwards
for _ in range(len(rows)):
    ws.delete_rows(2)

# Write back unique rows starting at A2:J
for idx, row in enumerate(unique_rows, start=2):
    for j, value in enumerate(row, start=1):
        ws.cell(row=idx, column=j, value=value)

wb.save(output_path)
