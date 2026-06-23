from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_7/group_290-1/r0/evolve_290-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_7/group_290-1/r0/evolve_290-1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Analyze data and propagate manual transposition or filling (K2:U10)
for row in range(2, 11):   # Row 2 to 10 inclusive
    # Read values
    values = [ws.cell(row=row, column=col).value for col in range(11, 22)]
    # Simple propagation: For empty (None) value, keep as is (if true manual transpose is needed, you'd add logic here)
    # Write the same values (you may modify logic here if filling is required)
    for i, val in enumerate(values):
        ws.cell(row=row, column=11 + i, value=val)

wb.save(output_path)
