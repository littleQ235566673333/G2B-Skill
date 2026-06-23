from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
import re

input_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r2/eval_39903_tc1/input.xlsx"
output_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r2/eval_39903_tc1/output.xlsx"

wb = load_workbook(input_path)
sheetnames = wb.sheetnames
if not sheetnames:
    raise Exception("No sheets found in workbook!")
ws = wb[sheetnames[0]]

start_row = 2
end_row = 6
col = 3  # Column C

# Regex pattern matches position names not starting with X/Z, up to colon
def count_individual_bin_locations(cell_value):
    if not cell_value or not isinstance(cell_value, str):
        return 0
    # Split by comma, semicolon, or leave as one: treat all colons as delim
    # Each location is: SomeCode: digits
    # Only count if before colon does not start with X or Z
    # Allow for white-spaces
    parts = re.findall(r'([A-Za-z0-9\-.]+):\s*\d+', cell_value)
    bins = set()
    for part in parts:
        p = part.strip()
        if p and p[0] not in ("X", "Z"):
            bins.add(p)
    return len(bins)

# Count for each cell and apply formatting
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
courier_font = Font(name="Courier New", size=9)
for r in range(start_row, end_row+1):
    val = ws.cell(row=r, column=col).value
    count = count_individual_bin_locations(val)
    cell = ws.cell(row=r, column=col)
    cell.value = count
    cell.border = border
    cell.font = courier_font

wb.save(output_path)
