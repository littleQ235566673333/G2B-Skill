from openpyxl import load_workbook
import datetime

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/group_56274/r2/evolve_56274/input.xlsx')
ws = wb['Sheet2']
for row in ws.iter_rows(min_row=1, max_row=20, max_col=10):
    print([str(cell.value) for cell in row])
