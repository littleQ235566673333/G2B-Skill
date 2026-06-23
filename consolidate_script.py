import openpyxl
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_2/regression_gate/after_pass/core_80-42/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_2/regression_gate/after_pass/core_80-42/output.xlsx'

wb = openpyxl.load_workbook(input_path)
target_ws = wb['Consolidate_ALL']

# Gather target headers and their column index in source
consolidate_headers = [cell.value for cell in next(target_ws.iter_rows(min_row=1, max_row=1))]
sheets_to_consolidate = ['Jack', 'Henry', 'Richard']

consolidated_rows = []
for sheet_name in sheets_to_consolidate:
    ws = wb[sheet_name]
    source_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Map from consolidate header to index in source
    col_indexes = [source_headers.index(header) if header in source_headers else None for header in consolidate_headers[:-1]]
    # Gather rows, skipping header (start from row=2)
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip completely empty rows
        if all(cell is None for cell in row):
            continue
        # Extract only the matched columns in target order, then add sheet name
        values = [row[idx] if idx is not None and idx < len(row) else None for idx in col_indexes]
        values.append(sheet_name)  # For Sheet Name (col L)
        consolidated_rows.append(values)

# Find where to start writing in Consolidate_ALL (below last non-empty in col A)
data_start = 2
max_row = target_ws.max_row
for i in range(max_row, 1, -1):
    if target_ws.cell(row=i, column=1).value:
        data_start = i + 1
        break

for i, values in enumerate(consolidated_rows, start=data_start):
    for j, val in enumerate(values, start=1):
        target_ws.cell(row=i, column=j, value=val)

wb.save(output_path)
