from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/group_44017/r2/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/group_44017/r2/evolve_44017/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Data']

# AD:AO = cols 30:43
for row in range(14, 43):
    for idx, col in enumerate(range(30, 44)):
        col_letter = get_column_letter(col)
        mon_cell = f'{col_letter}$9'
        freq_cell = f'$J{row}'
        eff_cell = f'$L{row}'
        base = f'$W{row}'
        # Build product of increases for up to 4 possible waves
        mult_formula = (
            f'(1+IF($M{row}<>"",IF(wave>=1,$M{row},0),0))'
            f'* (1+IF($N{row}<>"",IF(wave>=2,$N{row},0),0))'
            f'* (1+IF($O{row}<>"",IF(wave>=3,$O{row},0),0))'
            f'* (1+IF($P{row}<>"",IF(wave>=4,$P{row},0),0))'
        )
        # LET allows us to define the wave count once
        formula = (
            f'=IF(OR({mon_cell}<{eff_cell},{eff_cell}="",{base}="",{freq_cell}=""),"",'
            f'LET(wave,MIN(4,1+INT(DATEDIF({eff_cell},{mon_cell},"m")/{freq_cell})), '
            f'{base}*{mult_formula}'
            f'))'
        )
        ws[f'{col_letter}{row}'].value = formula
        # No explicit fill modification to avoid style bugs

wb.save(output_path)
print('Formulas written and file saved!')
