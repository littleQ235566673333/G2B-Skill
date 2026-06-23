from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r3/eval_54638_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r3/eval_54638_tc1/output.xlsx'

wb = load_workbook(input_path)
ws1 = wb['Sheet1']

# Ensure all text is Calibri font
for row in ws1.iter_rows():
    for cell in row:
        if cell.value is not None:
            cell.font = Font(name='Calibri', size=11)

# Recopy all original formulas in A2:A150, fill down formulas to A150, preserve only border in A2:A13
thin = Side(border_style='thin', color='000000')
for r in range(2, 151):
    c = ws1.cell(row=r, column=1)
    # Restore formula if present, otherwise leave as value
    if hasattr(c, 'value') and isinstance(c.value, str) and c.value.startswith('='):
        formula = c.value
    else:
        formula = None
    if formula:
        ws1.cell(row=r, column=1, value=formula)
    if 2 <= r <= 13:
        ws1.cell(row=r, column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    else:
        ws1.cell(row=r, column=1).border = Border()
    ws1.cell(row=r, column=1).font = Font(name='Calibri', size=11)

# Hide Sheet2
if 'Sheet2' in wb.sheetnames:
    wb['Sheet2'].sheet_state = 'hidden'

# Prepare B2:B150 - non-dynamic array formula for listing uniques from A2:A150, with gridline (border) and fill
fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
for r in range(2, 151):
    # The COUNTIF-INDEX-MATCH approach for uniques
    formula = f"=IFERROR(INDEX($A$2:$A$150, MATCH(0, COUNTIF($B$1:B{r-1}, $A$2:$A$150), 0)), \"\")"
    cell = ws1.cell(row=r, column=2, value=formula)
    cell.font = Font(name='Calibri', size=11)
    cell.fill = fill
    # Apply gridline borders
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

wb.save(output_path)
