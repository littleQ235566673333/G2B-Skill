import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun1/eval_55708_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun1/eval_55708_tc1/output.xlsx'

df = pd.read_excel(input_path)
dep_col = 'Department'
status_col = 'Status'
tat_col = 'turnaroud time'

q = df[(df[status_col].isin(['In Progress', 'In Review'])) & (df[tat_col] >= 6)]
departments = df[dep_col].dropna().unique()
avgs = []
for dep in departments:
    subset = q[q[dep_col] == dep]
    if not subset.empty:
        avgs.append(round(subset[tat_col].mean(), 2))
    else:
        avgs.append('')
wb = load_workbook(input_path)
ws = wb.active
for i, val in enumerate(avgs):
    ws.cell(row=11 + i, column=2, value=val if val != '' else None)
wb.save(output_path)
