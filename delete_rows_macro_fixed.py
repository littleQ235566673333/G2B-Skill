import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_374-18_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_374-18_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Imported Data']

header_row = 1
col_e = 5  # Column E

# Loop bottom-up, delete rows with E < 1
for row in range(ws.max_row, header_row, -1):
    cell_val = ws.cell(row=row, column=col_e).value
    try:
        num = float(cell_val)
    except (TypeError, ValueError):
        num = None
    if num is not None and num < 1:
        ws.delete_rows(row)

wb.save(output_path)
