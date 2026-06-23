import openpyxl

def reformat_date(num):
    s = str(num).zfill(9)  # Ensure string is '0yyyymmdd'
    yyyy = s[1:5]
    mm = s[5:7]
    dd = s[7:9]
    return f"{yyyy} {mm} {dd}"

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_486-17_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_486-17_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)
ws = wb["Blad1"]

for row in range(2, 131):  # B2:B130
    val = ws.cell(row=row, column=2).value
    if val is not None:
        ws.cell(row=row, column=2, value=reformat_date(val))

wb.save(output_path)
