import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_3/group_58701/r3/evolve_58701/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_3/group_58701/r3/evolve_58701/output.xlsx'

wb = openpyxl.load_workbook(input_path)
entry_ws = wb['Entry Tab']
table_ws = wb['Table Tab']

# Build location -> office code mapping from Table Tab
location_to_code = {}
# Assume first row is header
for row in table_ws.iter_rows(min_row=2, values_only=True):
    loc, code = row[:2]
    if loc is not None:
        location_to_code[str(loc).strip()] = code

# For E2:E3 (row 2 and 3, col 5) on Entry Tab, read corresponding location from D2:D3
for row in range(2, 4):
    location_cell = entry_ws.cell(row=row, column=4)  # D column = 4
    location = location_cell.value
    code = location_to_code.get(str(location).strip(), '') if location else ''
    entry_ws.cell(row=row, column=5).value = code

wb.save(output_path)
