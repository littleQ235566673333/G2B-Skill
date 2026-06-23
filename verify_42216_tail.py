from openpyxl import load_workbook
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_42216_s1/output.xlsx')
ws = wb[wb.sheetnames[0]]
for r in range(262, 340):
    print(r, ws.cell(r,1).value, ws.cell(r,2).value)
