import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_516-46/r1/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_516-46/r1/evolve_516-46/output.xlsx'
sheet_name = 'ورقة1'

# Step 1: Read original data (pandas for easier grouping, then write back by openpyxl)
df = pd.read_excel(input_path, sheet_name=sheet_name)
# Only columns A-E
df = df.iloc[:, :5]

# Inference: which columns are date, brand, qty?
# Let's auto-detect or assume:
# A: Date, B: Brand, ... Any headers?
colnames = df.columns.tolist()

def guess_col(cols):
    date_col = [c for c in cols if 'date' in str(c).lower() or '\u062a\u0627\u0631\u064a\u062e' in str(c)]
    brand_col = [c for c in cols if 'brand' in str(c).lower() or '\u0627\u0644\u0645\u0627\u0631\u0643\u0647' in str(c)]
    qty_col = [c for c in cols if 'qty' in str(c).lower() or 'quantity' in str(c).lower() or '\u0643\u0645\u064a\u0629' in str(c)]
    # fallback: by order
    return date_col[0] if date_col else cols[0], brand_col[0] if brand_col else cols[1], qty_col[0] if qty_col else cols[2]

# Column order
date_col, brand_col, qty_col = guess_col(colnames)

df[date_col] = pd.to_datetime(df[date_col])
last_dates = df.groupby(brand_col)[date_col].transform('max')
last_date_df = df[df[date_col] == last_dates]

result = last_date_df.groupby([brand_col, date_col], as_index=False).agg({qty_col:'sum'})

modifications = last_date_df.groupby([brand_col, date_col]).size().reset_index(name='Count')
result = result.merge(modifications, on=[brand_col, date_col])
result['Modification'] = result['Count'].apply(lambda x: 'Aggregated' if x > 1 else '')
result = result.drop(columns=['Count'])

wb = load_workbook(input_path)
ws = wb[sheet_name]

# Write header
header = [brand_col, date_col, qty_col, 'Modification']
for j, val in enumerate(header, 8):
    ws.cell(row=2, column=j, value=val)
# Write data (up to 3 rows, as per L4, but could be more or less)
for i, row in enumerate(result.itertuples(index=False), 3):
    ws.cell(row=i, column=8, value=getattr(row, brand_col))
    ws.cell(row=i, column=9, value=getattr(row, date_col))
    ws.cell(row=i, column=10, value=getattr(row, qty_col))
    ws.cell(row=i, column=11, value=getattr(row, 'Modification'))

wb.save(output_path)
