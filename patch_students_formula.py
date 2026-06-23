from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_5/regression_gate/after_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_5/regression_gate/after_pass/core_41601/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Students']
for row in range(2, 8):
    ws[f'E{row}'] = f"=INDIRECT('\''&A{row}&'\'!C2')"
wb.save(output_path)
