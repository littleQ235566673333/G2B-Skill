import openpyxl
import re

# File paths
input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_4/group_39515/r3/evolve_39515/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_4/group_39515/r3/evolve_39515/output.xlsx'

wb = openpyxl.load_workbook(input_fp)
ws = wb.active

def normalize_month(s):
    s = str(s)
    abbr = s[:3].capitalize()
    return abbr

# Read headers from C1:N1
col_headers = list(ws.iter_rows(min_row=1, max_row=1, min_col=3, max_col=14, values_only=True))[0]

# Parse headers as (month, year) tuples
parsed_headers = []
for h in col_headers:
    # Accepts: 'Jan 2022', '2022 Jan', 'Jan-22', etc.
    month_match = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)', str(h), re.I)
    year_match = re.search(r"(20\\d{2}|'\\d{2})", str(h))
    month = normalize_month(month_match.group(1)) if month_match else str(h)
    if year_match:
        raw_year = year_match.group(1)
        if raw_year.startswith("'"):
            year = int("20" + raw_year[1:])
        elif len(raw_year) == 2:
            year = int("20" + raw_year)
        else:
            year = int(raw_year)
    else:
        year = None
    parsed_headers.append((month, year))

for row in range(2, 14):
    # A: target month, B: year
    month = normalize_month(ws[f'A{row}'].value)
    year_val = ws[f'B{row}'].value
    # Try int conversion
    try:
        year = int(year_val)
    except Exception:
        # Parse from str if needed
        year_match = re.search(r'(20\\d{2}|\\d{2})', str(year_val))
        if year_match:
            y = year_match.group(1)
            if len(y) == 2:
                year = int('20' + y)
            else:
                year = int(y)
        else:
            year = None
    # Find header col that matches both
    try:
        idx = [i for i, (m, y) in enumerate(parsed_headers) if m == month and y == year][0]
        val = ws.cell(row=row, column=3+idx).value
    except Exception:
        val = None
    ws[f'O{row}'] = val

wb.save(output_fp)
print('Done')
