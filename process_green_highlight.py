import openpyxl

def is_green(cell):
    if cell.fill is None:
        return False
    fg = cell.fill.fgColor
    if fg.type == 'rgb' and fg.rgb is not None:
        # Standard green RGB hex: FF00FF00 or 00FF00
        return fg.rgb in ['FF00FF00', '00FF00']
    if fg.type == 'indexed':
        return fg.indexed == 3  # Excel indexedColor 3 is green
    return False

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_43657_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_43657_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
results = []
for row in range(2, 9):  # Rows 2 to 8, per your instructions
    name = ws[f'L{row}'].value
    count = 0
    for col in ['C','D','E','F','G']:
        cell = ws[f'{col}{row}']
        if cell.value and name and name in str(cell.value):
            if is_green(cell):
                count += 1
    results.append(count)
for idx, v in enumerate(results, start=2):
    ws[f'K{idx}'].value = v
wb.save(output_path)
