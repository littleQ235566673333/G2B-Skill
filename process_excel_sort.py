import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_262-17_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_262-17_tc1/output.xlsx'

# Load workbook and sheet
wb = load_workbook(input_path)
ws = wb.worksheets[0]

# Read header row to find indices for 'Task' and 'Responsibility'
header = [cell.value for cell in ws[1]]
try:
    task_idx = header.index('Task')
    resp_idx = header.index('Responsibility')
except ValueError:
    raise Exception('Required headers not found')

# Prepare data (A1:F14)
data = list(ws.iter_rows(min_row=1, max_row=14, min_col=1, max_col=6, values_only=True))
header_row = data[0]
content_rows = data[1:]

# Perform sorting: 1st by Task, then by Responsibility (both ascending)
sorted_content = sorted(content_rows, key = lambda r: (r[task_idx], r[resp_idx]))
output = [header_row] + sorted_content

# Write sorted data to worksheet (overwrite original data in A1:F14)
for row_idx, row in enumerate(output, start=1):
    for col_idx, val in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=val)

wb.save(output_path)
