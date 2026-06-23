import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_2/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_2/regression_gate/after_pass/core_50526/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

# Read headers
headers = [cell.value for cell in ws[1]]
# Read lookup value from B6
lookup_value = ws['B6'].value

# Find the row index for the lookup value in column 1 (A)
row_lookup = None
for row in range(2, ws.max_row+1):
    if ws.cell(row=row, column=1).value == lookup_value:
        row_lookup = row
        break

values_to_display = []
if row_lookup:
    for col in range(2, ws.max_column+1):
        val = ws.cell(row=row_lookup, column=col).value
        if isinstance(val, (int, float)) and val > 0:
            # Use column header
            values_to_display.append(ws.cell(row=1, column=col).value)

# Write output values to B9, B10
for i in range(2):
    cell = ws.cell(row=9+i, column=2)
    if i < len(values_to_display):
        cell.value = values_to_display[i]
    else:
        cell.value = None

wb.save(output_path)
