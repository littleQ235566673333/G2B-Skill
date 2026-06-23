from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/before_pass/core_41969/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/before_pass/core_41969/output.xlsx"

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# For A6:C6, fill formulas: =COUNTBLANK(A3:C3), =COUNTBLANK(D3:F3), =COUNTBLANK(G3:I3)
for idx, col in enumerate([1, 4, 7]):    # A=1, D=4, G=7
    start_col = get_column_letter(col)
    end_col = get_column_letter(col+2)
    formula = f"=COUNTBLANK({start_col}3:{end_col}3)"
    cell = ws.cell(row=6, column=col)
    cell.value = formula

wb.save(output_path)
