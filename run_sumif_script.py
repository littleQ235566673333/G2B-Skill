import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_51262_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_51262_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active
# Read ranges
data = ws['B6':'D13']
criteria_list = [ws[f'F{row}'].value for row in range(10, 15) if ws[f'F{row}'].value is not None]
f5_date = ws['F5'].value
# Extract data for DataFrame
data_rows = []
for row in data:
    data_rows.append([cell.value for cell in row])
df = pd.DataFrame(data_rows, columns=['Date', 'Criteria', 'Value'])
# Date range
df['Date'] = pd.to_datetime(df['Date'])
if hasattr(f5_date, 'year') and hasattr(f5_date, 'month'):
    date_from = pd.Timestamp(f5_date.year, f5_date.month, 1)
    date_to = (date_from + pd.offsets.MonthEnd(0))
else:
    date_from = pd.to_datetime(f5_date).replace(day=1)
    date_to = date_from + pd.offsets.MonthEnd(0)
# Filtering
mask = (df['Date'] >= date_from) & (df['Date'] <= date_to) & (df['Criteria'].isin(criteria_list))
sum_value = df.loc[mask, 'Value'].sum()
# Write result to F6
ws['F6'].value = sum_value
ws['G6'].value = ''
ws['H6'].value = ''
wb.save(output_path)
