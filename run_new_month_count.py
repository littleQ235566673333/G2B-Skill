from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_4/group_45707/r1/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_4/group_45707/r1/evolve_45707/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Read all relevant data upfront
dates = [ws[f'A{i}'].value for i in range(2,70)]
col_c = [ws[f'C{i}'].value for i in range(2,70)]

# Prepare a mapping: (year, month) -> list of indices
year_month_indices = {}
for idx, d in enumerate(dates):
    if isinstance(d, datetime):
        ym = (d.year, d.month)
        year_month_indices.setdefault(ym, []).append(idx)

for i in range(2, 70):  # D2:D69
    next_idx = i - 1  # zero-based
    if next_idx + 1 < len(dates) and isinstance(dates[next_idx + 1], datetime):
        next_day = dates[next_idx + 1]
        if next_day.day == 1:
            ym = (next_day.year, next_day.month)
            indices = year_month_indices.get(ym, [])
            count_ones = sum(1 for j in indices if col_c[j] == 1)
            ws[f'D{i}'] = count_ones
        else:
            ws[f'D{i}'].value = None
    else:
        ws[f'D{i}'].value = None

wb.save(output_path)
