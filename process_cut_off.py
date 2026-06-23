import openpyxl
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_35739_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_35739_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

for row in range(2, 101):  # Rows 2 to 100
    val_a = ws[f'A{row}'].value
    val_b = ws[f'B{row}'].value
    
    # If columns A or B are empty, set C blank
    if val_a is None or val_b is None:
        ws[f'C{row}'].value = None
        continue
    
    # Handle time and subtract 30 minutes
    try:
        # If B is a datetime.time
        if isinstance(val_b, datetime) or isinstance(val_b, datetime.time):
            # Extract the time if it's datetime
            if isinstance(val_b, datetime):
                time_val = val_b.time()
            else:
                time_val = val_b
            # Convert to datetime for calculation
            dt = datetime.combine(datetime.today(), time_val)
            cut_off_dt = dt - timedelta(minutes=30)
            cut_off_time = cut_off_dt.time()
            ws[f'C{row}'].value = cut_off_time
        # If B is a string (e.g., '23:15'), try parsing
        elif isinstance(val_b, str):
            try:
                time_val = datetime.strptime(val_b, '%H:%M').time()
            except ValueError:
                try:
                    time_val = datetime.strptime(val_b, '%H:%M:%S').time()
                except ValueError:
                    ws[f'C{row}'].value = None
                    continue
            dt = datetime.combine(datetime.today(), time_val)
            cut_off_dt = dt - timedelta(minutes=30)
            cut_off_time = cut_off_dt.time()
            ws[f'C{row}'].value = cut_off_time
        else:
            ws[f'C{row}'].value = None
    except Exception:
        ws[f'C{row}'].value = None

wb.save(output_path)
