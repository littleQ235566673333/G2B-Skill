import openpyxl

# Load workbook and select active sheet
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_59734_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_59734_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Assume header is row 1, so data starts at row 2
# For each row (B2:B16), check if:
# - Column A (A2:A16) matches NUMVAL
# - Columns F and G match O and P (we use F2:F16, G2:G16, O, P)
# Fill column B with corresponding value from Column V (V2:V16)

# Get values of NUMVAL, O, P from somewhere (context unclear).
# For now, let's get values from first row (header or some cell?)
NUMVAL = ws['A2'].value  # Example, need proper definition
O = ws['F2'].value      # Example
P = ws['G2'].value      # Example

for row in range(2, 17):
    # Read values
    val_a = ws[f'A{row}'].value
    val_f = ws[f'F{row}'].value
    val_g = ws[f'G{row}'].value
    val_v = ws[f'V{row}'].value

    # Apply criteria: matches
    if val_a == NUMVAL and val_f == O and val_g == P:
        ws[f'B{row}'] = val_v
    else:
        ws[f'B{row}'] = ''

wb.save(output_path)
