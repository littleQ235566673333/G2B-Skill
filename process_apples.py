from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_5/group_58687/r0/evolve_58687/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_5/group_58687/r0/evolve_58687/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# 1. Read horizontal apples sold by month (B4:V6)
# Months (B4:V4), Count in (B5:V5)
months = [ws.cell(row=4, column=col).value for col in range(2, 23)]
apples_each_month = []
for col in range(2, 23):
    val = ws.cell(row=5, column=col).value
    try:
        apples_each_month.append(int(val))
    except (ValueError, TypeError):
        apples_each_month.append(0)

# 2. Build an expanding list mapping index 1...92 to month numbers
month_map = []
for month, cnt in enumerate(apples_each_month, 1):
    month_map.extend([month]*cnt)

# 3. Read the "Apples Accum" values (col C9:C100 --> index numbers)
indices = [ws.cell(row=row, column=3).value for row in range(9, 101)]

# 4. Fill B9:B100 with month month for that apple, or add 1 if index missing
for i, idx in enumerate(indices):
    if isinstance(idx, int) and 1 <= idx <= len(month_map):
        ws.cell(row=9+i, column=2).value = month_map[idx-1]
    else:
        # Apple index not found, add 1 to last assigned, or just 1
        prev = ws.cell(row=8+i, column=2).value if (i > 0) else 0
        try:
            ws.cell(row=9+i, column=2).value = int(prev) + 1
        except:
            ws.cell(row=9+i, column=2).value = 1

wb.save(output_path)
