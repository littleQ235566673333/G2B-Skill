import openpyxl

input_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_1563_tc1/input.xlsx"
output_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_1563_tc1/output.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

last_value = None
# Rows 2 to 30 (inclusive)
for row in range(2, 31):
    cell_a = ws[f'A{row}'].value
    if cell_a is not None and cell_a != "":
        last_value = cell_a
    ws[f'B{row}'].value = last_value

wb.save(output_path)
