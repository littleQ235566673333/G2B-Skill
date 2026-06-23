from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_3/regression_gate/before_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_3/regression_gate/before_pass/core_50526/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

lookup_value = ws['B6'].value
lookup_row = None
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=2).value == lookup_value:
        lookup_row = row
        break
if lookup_row:
    result = []
    for col in range(3, ws.max_column + 1):  # skip first two columns
        val = ws.cell(row=lookup_row, column=col).value
        if val is not None and val > 0:
            color = ws.cell(row=1, column=col).value
            result.append(color)
    for i in range(2):
        ws.cell(row=9 + i, column=2).value = result[i] if i < len(result) else None
else:
    for i in range(2):
        ws.cell(row=9 + i, column=2).value = None

wb.save(output_path)
