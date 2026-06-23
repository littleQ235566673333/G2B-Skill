import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

# Load paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_1/group_516-46/r2/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_1/group_516-46/r2/evolve_516-46/output.xlsx'
sheet_name = 'ورقة1'

# Load the Excel data
# Data starts from A2, columns A-E
all_data = pd.read_excel(input_path, sheet_name=sheet_name, header=0)  # Assumes headers in row 1
# Exclude header row. Fetch columns A-E
all_data = all_data.iloc[1:, 0:5].copy()  # Only data, columns A-E
all_data.columns = ['Date', 'Brand', 'Qty', 'C', 'D']

# Standardize types/clean
df = all_data
# Datetime and numeric
try:
    df['Date'] = pd.to_datetime(df['Date'])
except Exception:
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
df['Qty'] = pd.to_numeric(df['Qty'], errors='coerce').fillna(0)
df['Brand'] = df['Brand'].astype(str).str.strip()

# Get latest date
last_date = df['Date'].max()
last_rows = df[df['Date'] == last_date]

# Group by brand and sum the qty for last date
out = last_rows.groupby('Brand', as_index=False).agg({'Date':'first','Qty':'sum'})
out = out[['Date','Brand','Qty']]

# Find duplicated brands (multiple rows in last date per brand)
dupl = last_rows.groupby('Brand').filter(lambda x: len(x) > 1)
modifications = {}
if not dupl.empty:
    for brand, grp in dupl.groupby('Brand'):
        mods = grp[['Brand','Qty']].values.tolist()
        modifications[brand] = mods

# Open workbook, write to H2:L4
wb = load_workbook(input_path)
ws = wb[sheet_name]
# Headers
ws['H2'] = 'Date'
ws['I2'] = 'Brand'
ws['J2'] = 'Qty'
ws['K2'] = 'Mods for Dups'
ws['L2'] = ''
# Content
i_start = 3
for idx, row in enumerate(out.itertuples(index=False), start=i_start):
    ws[f'H{idx}'] = row.Date.strftime('%Y-%m-%d') if pd.notnull(row.Date) else ''
    ws[f'I{idx}'] = row.Brand
    ws[f'J{idx}'] = int(row.Qty)
    modtext = ''
    if row.Brand in modifications:
        entry = [str(int(qty)) for brand, qty in modifications[row.Brand]]
        modtext = ' + '.join(entry) + f' = {int(row.Qty)}'
    ws[f'K{idx}'] = modtext
    ws[f'L{idx}'] = ''

wb.save(output_path)
