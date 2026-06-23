import openpyxl

# Load workbook and worksheet
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_524-31_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_524-31_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Exp-DB']

# Build lookup dictionary from first 2 columns: Short description to category
lookup = {}
for row in ws.iter_rows(min_row=1, max_row=53, max_col=2, values_only=True):
    short_desc, category = row
    if short_desc and category:
        lookup[short_desc.strip().lower()] = category.strip()

# Assign category for each transaction description found in column D (col 4)
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=53, min_col=4, max_col=4, values_only=True), 1):
    desc = row[0]
    category = None
    if desc:
        # Extract prefix before '*' and normalize
        prefix = desc.split('*')[0].strip().lower()
        # Try to match prefix to one of the known short descriptions
        for short in lookup:
            if short in prefix:
                category = lookup[short]
                break
    ws.cell(row=idx, column=5, value=category)

wb.save(output_path)
print('Done: Categories assigned to transactions.')
