from openpyxl import load_workbook

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_46897_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_46897_tc1/output.xlsx"

wb = load_workbook(filename=input_path)
ws = wb.active

# The following values are for demonstration.
# Header name to match is assumed to be in H1
# Value to count is assumed to be in H2
header_to_find = ws["H1"].value  # Which header title to use
value_to_count = ws["H2"].value  # Which value to count

# Find column letter matching the header
from openpyxl.utils import get_column_letter

target_col = None
for cell in ws[1]:
    if cell.value == header_to_find:
        target_col = get_column_letter(cell.column)
        break

if target_col:
    # Write formula for counting occurrences in the detected column
    formula = f'=COUNTIF({target_col}2:{target_col}1048576, "{value_to_count}")'
else:
    formula = "Header not found"

ws["I3"] = formula
wb.save(output_path)
