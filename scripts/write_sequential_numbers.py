import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_38537_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_38537_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

b2_value = ws['B2'].value
counter = 10  # Start at 010 per user request

for row in range(3, 37):  # Covers B3:B36
    b_val = ws[f'B{row}'].value
    if b_val not in (None, ''):
        ws[f'A{row}'].value = f'{b2_value}-{counter:03d}'
        counter += 1
    else:
        ws[f'A{row}'].value = None  # Leave empty for gaps

wb.save(output_path)
