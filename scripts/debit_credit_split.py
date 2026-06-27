from openpyxl import load_workbook

fn_in = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_SAPR-A5-N2-seed1/eval_r3/eval_469-9_tc1/input.xlsx'
fn_out = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_SAPR-A5-N2-seed1/eval_r3/eval_469-9_tc1/output.xlsx'

wb = load_workbook(fn_in)
ws = wb.active

# Set headers
ws['H1'] = 'Debits'
ws['I1'] = 'Credits'

# Process rows 2 to 10 (inclusive)
for row in range(2, 11):
    val = ws.cell(row=row, column=3).value  # Column C
    if val is None or val == 0:
        continue
    if val > 0:
        ws.cell(row=row, column=8).value = abs(val)
    elif val < 0:
        ws.cell(row=row, column=9).value = abs(val)

wb.save(fn_out)
