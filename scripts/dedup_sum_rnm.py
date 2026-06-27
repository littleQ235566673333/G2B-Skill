import openpyxl
from collections import defaultdict
from copy import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_SAPR-A0-N2-seed1/eval_r1/eval_250-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_SAPR-A0-N2-seed1/eval_r1/eval_250-20_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['RNM']

rows = list(ws.iter_rows(min_row=2, max_col=11, values_only=False))
header = [cell.value for cell in ws[1]]

sums = defaultdict(int)
row_cache = {}
for row in rows:
    num = row[1].value
    line = row[2].value
    qty = row[9].value if row[9].value is not None else 0
    key = (num, line)
    sums[key] += qty
    if key not in row_cache:
        row_cache[key] = row
# Prepare deduplicated rows for output
unique_rows = []
for key, row in row_cache.items():
    # Copy cell objects to avoid worksheet mutation
    cells = [copy(cell) for cell in row[:10]]
    # Set the summed MATCHED_QTY in J (index 9)
    cells[9].value = sums[key]
    unique_rows.append(cells)
# Sort output to match first occurrence order
unique_rows.sort(key=lambda r: (rows.index(row_cache[(r[1].value, r[2].value)])))
# Clear output range A2:J20
for r in range(2, 21):
    for c in range(1, 11):
        ws.cell(row=r, column=c, value=None)
# Rewrite header (A1:J1)
for j, val in enumerate(header[:10], start=1):
    ws.cell(row=1, column=j, value=val)
# Write deduplicated data rows
for i, row in enumerate(unique_rows, start=2):
    if i > 20:
        break
    for j, cell in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=cell.value)
wb.save(output_path)
print('Done')
