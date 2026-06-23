from openpyxl import load_workbook
from datetime import datetime

# Paths
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_534-26_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_534-26_tc1/output.xlsx'

# Load workbook and select sheet
wb = load_workbook(input_path)
ws = wb['Sheet1']

# Get header
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

# Current month/first day
now = datetime.today()
current_month_start = datetime(now.year, now.month, 1)

# Find columns to delete
cols_to_delete = []
for idx, val in enumerate(header):
    if isinstance(val, datetime) and val < current_month_start:
        cols_to_delete.append(idx+1)  # 1-based indexing for openpyxl

# Delete columns from right to left (no shifting)
for col in reversed(cols_to_delete):
    ws.delete_cols(col)

# Save the result
wb.save(output_path)
