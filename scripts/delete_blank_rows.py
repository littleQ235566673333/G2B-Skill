from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_6/group_157-4/r1/evolve_157-4/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_6/group_157-4/r1/evolve_157-4/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

filtered_rows = []
for row in ws.iter_rows(min_row=14, max_row=83, min_col=6, max_col=13, values_only=True):
    # row is (F, G, H, I, J, K, L, M)
    # Check if all F to L (indices 0:6) are all None/empty
    if all(cell is None or str(cell).strip() == '' for cell in row[:7]):
        continue
    filtered_rows.append(row)
# Write to E9:M... in output
start_row = 9
start_col = 5  # Col E (openpyxl is 1-based)
for i, row in enumerate(filtered_rows):
    for j, val in enumerate(row):
        ws.cell(row=start_row + i, column=start_col + j, value=val)
wb.save(output_path)
