from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r2/eval_45372_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r2/eval_45372_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

for row in range(2, 16):
    time_val = ws[f'A{row}'].value
    b_val = ws[f'B{row}'].value
    c_val = ws[f'C{row}'].value
    val = None
    cmp_time = None
    if isinstance(time_val, datetime.time):
        cmp_time = time_val
    elif isinstance(time_val, datetime.datetime):
        cmp_time = time_val.time()
    if cmp_time:
        threshold = datetime.time(9, 45)
        if cmp_time < threshold:
            val = b_val
        else:
            val = c_val
    # If the value is 0 or None, clear the E cell
    if val == 0 or val is None:
        ws[f'E{row}'].value = None
    else:
        ws[f'E{row}'].value = val
    # Apply the fill
    ws[f'E{row}'].fill = fill

wb.save(output_path)
