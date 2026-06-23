import openpyxl
import pandas as pd

def remove_row_duplicates(row):
    seen = set()
    result = []
    for elem in row:
        if pd.isnull(elem) or (isinstance(elem, str) and elem.strip() == ""):
            result.append(elem)
        elif elem not in seen:
            seen.add(elem)
            result.append(elem)
        else:
            result.append("")
    return result

inp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/group_227-40/r2/evolve_227-40/input.xlsx'
outp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/group_227-40/r2/evolve_227-40/output.xlsx'
wb = openpyxl.load_workbook(inp)
ws = wb['Sheet1']

rows_to_dedup = []
for r in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=5, values_only=True):
    rows_to_dedup.append(list(r))

processed = [remove_row_duplicates(row) for row in rows_to_dedup]

for r_idx, row in enumerate(processed, start=14):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx).value = value

wb.save(outp)
