import openpyxl

# Input and output paths
i_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_SAPR-A0-N2-seed1/eval_r1/eval_493-5_tc1/input.xlsx'
o_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_SAPR-A0-N2-seed1/eval_r1/eval_493-5_tc1/output.xlsx'

wb = openpyxl.load_workbook(i_path)
ws = wb['Imported Data']

# Read header and data rows (limit to first 9 data rows)
header = [cell.value for cell in ws[1][:6]]
data = []
for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=6, values_only=True):
    # Skip completely empty rows
    if any(x is not None for x in row):
        data.append(row)

n = len(data)
matched_idx = set()

# Compare all pairs of rows
def as_float(val):
    try:
        return float(val)
    except Exception:
        return None
for i in range(n):
    for j in range(i+1, n):
        rowi, rowj = data[i], data[j]
        # Reference and Narrative must match
        if rowi[0] == rowj[0] and rowi[5] == rowj[5]:
            debit_i = as_float(rowi[2])
            credit_i = as_float(rowi[3])
            debit_j = as_float(rowj[2])
            credit_j = as_float(rowj[3])
            # Debit of one == Credit of the other, both ways
            if debit_i is not None and credit_j is not None and debit_i == credit_j \
                and credit_i is not None and debit_j is not None and credit_i == debit_j:
                matched_idx.update([i, j])

# Rows NOT part of any deleted pair
filtered = [data[k] for k in range(n) if k not in matched_idx]

# Write to new workbook (header + up to 9 rows, for A1:F10)
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Imported Data'
ws2.append(header)
for row in filtered[:9]:
    ws2.append(row)

wb2.save(o_path)
