import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_38537_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_38537_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read prefix from B2
prefix = ws['B2'].value

seq_number = 10  # starting as 010

for row in range(3, 37):  # A3 to A36
    b_val = ws[f'B{row}'].value
    if b_val not in (None, ''):  # Only number non-empty rows
        ws[f'A{row}'] = f'{prefix}-{seq_number:03d}'
        seq_number += 1
    else:
        ws[f'A{row}'] = None

wb.save(output_path)
