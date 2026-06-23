import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_33157_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_33157_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Mapping columns to headers for activities
col_to_header = {2: ws['B1'].value, 4: ws['D1'].value, 6: ws['F1'].value, 8: ws['H1'].value}

for row in range(2, 7):  # K2:K6
    ref_date = ws[f'J{row}'].value
    found_header = None
    for col in [2, 4, 6, 8]:
        activity_date = ws.cell(row=row, column=col).value
        if activity_date == ref_date:
            found_header = col_to_header[col]
            break
    ws[f'K{row}'] = found_header

wb.save(output_path)
