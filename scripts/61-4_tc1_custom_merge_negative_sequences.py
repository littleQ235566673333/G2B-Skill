import pandas as pd
from openpyxl import load_workbook

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/eval_seed42_REORDER_TEST/eval_61-4_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/eval_seed42_REORDER_TEST/eval_61-4_tc1/output.xlsx"

# Read the input
df = pd.read_excel(input_path, sheet_name="input")

# Function to summarize consecutive negative blocks
def summarize_neg_groups(group):
    results = []
    i = 0
    n = len(group)
    while i < n:
        # skip non-negative or NaN 'Change'
        row = group.iloc[i]
        if pd.isna(row['Change']) or row['Change'] >= 0:
            i += 1
            continue
        start = i
        # expand through consecutive negative 'Change'
        while i < n and not pd.isna(group.iloc[i]['Change']) and group.iloc[i]['Change'] < 0:
            i += 1
        end = i - 1
        substring = group.iloc[start:end+1]
        results.append({
            'DATE ': substring.iloc[0]['DATE '],
            'Stock Name': substring.iloc[0]['Stock Name'],
            'OPENP* ': substring.iloc[0]['OPENP* '],
            'HIGH ': substring['HIGH '].max(),
            'LOW ': substring['LOW '].min(),
            'CLOSEP* ': substring.iloc[-1]['CLOSEP* '],
            'VOLUME': substring['VOLUME'].sum()
        })
    return pd.DataFrame(results)

# Run per-stock grouping and collect results
full_summary = pd.concat([
    summarize_neg_groups(grp) for name, grp in df.groupby('Stock Name')
], ignore_index=True)

# Write result to output sheet, starting at A2
wb = load_workbook(input_path)
if 'output' in wb.sheetnames:
    ws = wb['output']
else:
    ws = wb.create_sheet('output')
# Clear area A2:G15
for row in ws.iter_rows(min_row=2, max_row=15, min_col=1, max_col=7):
    for cell in row:
        cell.value = None

# Write header if needed
headers = ['DATE ', 'Stock Name', 'OPENP* ', 'HIGH ', 'LOW ', 'CLOSEP* ', 'VOLUME']
for j, h in enumerate(headers, 1):
    ws.cell(row=1, column=j).value = h

for i, record in enumerate(full_summary.itertuples(index=False), 2):
    ws.cell(row=i, column=1).value = record[0]
    ws.cell(row=i, column=2).value = record[1]
    ws.cell(row=i, column=3).value = record[2]
    ws.cell(row=i, column=4).value = record[3]
    ws.cell(row=i, column=5).value = record[4]
    ws.cell(row=i, column=6).value = record[5]
    ws.cell(row=i, column=7).value = record[6]

wb.save(output_path)
