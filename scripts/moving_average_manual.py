import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_3/group_4714/r3/evolve_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_3/group_4714/r3/evolve_4714/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet2']
rows = list(ws.iter_rows(min_row=2, max_row=25, max_col=4, values_only=True))
df = pd.DataFrame(rows, columns=['Emp Code', 'Month (YYYYMM)', 'Month', 'Hours'])
df['Month'] = pd.to_datetime(df['Month'])

def moving_avg(row, df):
    emp = row['Emp Code']
    curr_month = row['Month']
    emp_rows = df[df['Emp Code'] == emp]
    window_start = curr_month - pd.DateOffset(months=3)
    window = emp_rows[(emp_rows['Month'] >= window_start) & (emp_rows['Month'] <= curr_month)]
    if len(window) < 4:
        return 'n/a'
    avg = window['Hours'].mean()
    # Format output: whole number if decimal zero
    if avg == int(avg):
        avg = int(avg)
    else:
        avg = round(avg, 2)
    return avg

df['4mo_average'] = df.apply(lambda row: moving_avg(row, df), axis=1)
# Write to Excel, col E2:E25
for i, val in enumerate(df['4mo_average'], start=2):
    ws[f'E{i}'] = val
wb.save(output_path)
