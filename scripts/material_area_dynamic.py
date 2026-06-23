import pandas as pd
from openpyxl import load_workbook

def calc_material_areas(inp, outp):
    df = pd.read_excel(inp)
    # Only keep rows with all three required columns non-NaN
    df1 = df[['Mtrl','Width','Height']].copy().dropna()
    df1['Area'] = df1['Width'] * df1['Height']
    comp = df1.groupby('Mtrl')['Area'].sum().reset_index()
    # Write in format 'Material: Total Area'
    wb = load_workbook(inp)
    ws = wb.active
    nrows = len(comp)
    for i in range(nrows):
        ws[f'H{i+2}'].value = f'{comp.iloc[i,0]}: {comp.iloc[i,1]}'
    for i in range(nrows,3): # If less than 3 types, blank the rest
        ws[f'H{i+2}'].value = ''
    wb.save(outp)

if __name__ == '__main__':
    calc_material_areas(
        'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_1/task_263-1/r1/evolve_263-1/input.xlsx',
        'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_1/task_263-1/r1/evolve_263-1/output.xlsx'
    )
