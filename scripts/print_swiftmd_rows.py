from openpyxl import load_workbook
import datetime

def safe(x):
    return x.strftime('%Y-%m-%d') if hasattr(x, 'strftime') else x

wb = load_workbook('results/runs/skillgrad_gpt-4.1_ss-gpt41/train/iter_9/evolve_91-34/input.xlsx')
ws = wb['SwiftMD']
header = [cell.value for cell in ws[1]][1:15]
print('HEADER:', header)
rows = list(ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=15))
for idx, row in enumerate(rows, start=2):
    print(f'ROW {idx}:', [safe(cell.value) for cell in row])
