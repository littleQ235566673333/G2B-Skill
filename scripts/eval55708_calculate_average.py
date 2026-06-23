import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_55708_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_55708_tc1/output.xlsx'

# Load sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Extract data to DataFrame
rows = list(ws.values)
header = rows[0]
data_rows = [row for row in rows[1:] if row[0] not in (None, 'Calculations', 'Department') and row[0] and len(str(row[0])) == 1]
df = pd.DataFrame(data_rows, columns=header)

# Filter relevant Status & turnaround >= 6
df = df[df['Status'].isin(['In Progress', 'In Review'])]
df = df[df['turnaroud time'] >= 6]

# Calculate averages by department
groups = df.groupby('Department')['turnaroud time'].mean()

# Departments to fill
out_cells = ['B11', 'B12', 'B13']
departments = [ws[f'A{row}'].value for row in [11, 12, 13]]

for cell, dept in zip(out_cells, departments):
    value = groups.get(dept, None)
    ws[cell].value = None if pd.isna(value) else float(value) if value is not None else None

# Set general format (openpyxl defaults to general)
wb.save(output_path)
