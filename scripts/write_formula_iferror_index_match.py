from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = 'results/runs/ITSM_B_v0_mask_R5_FIX-V-seed2/eval_r3/eval_7902_tc1/input.xlsx'
output_path = 'results/runs/ITSM_B_v0_mask_R5_FIX-V-seed2/eval_r3/eval_7902_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Formula Required']

# D3:J6 = columns 4 to 10, rows 3 to 6
for row in range(3, 7):  # 7 is exclusive, so covers 3,4,5,6
    for col in range(4, 11):  # 11 is exclusive, so 4..10 (D..J)
        cell = ws.cell(row=row, column=col)
        # COLUMN(D$2)-2 part: D=4-2=2 (i.e. 2nd data col in Grouping)
        col_offset = col - 2  # Because 'Grouping' data starts from C (3), but our formula wants to index data columns 3..9 in Grouping
        # Prepare dynamic range addresses for MATCH/INDEX part
        formula = (
            f"=IFERROR(INDEX(Grouping!$C$2:$I$7, MATCH(1, (Grouping!$A$2:$A$7=$C{row})*(Grouping!$B$2:$B$7=$B{row}), 0), {col_offset}), "
            f"VLOOKUP($C{row},Grouping!$A$2:$I$7,{col_offset+1},FALSE))"
        )
        cell.value = formula

wb.save(output_path)
