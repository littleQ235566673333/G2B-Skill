import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_2/regression_gate/after_fix/core_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_2/regression_gate/after_fix/core_57033/output.xlsx'

# Read sheets
df_sheet4 = pd.read_excel(input_path, sheet_name='Sheet4')
df_cbtrans = pd.read_excel(input_path, sheet_name='CBtrans')

# Standardize keys for matching: strip/clean up whitespace if any
def norm(x):
    if pd.isnull(x):
        return ''
    if isinstance(x, float):
        # Numbers shouldn't get lower() call, just string conversion
        return str(int(x)) if x == int(x) else str(x)
    return str(x).strip()

# We assume matching on Company/company, account/account, and xchar/xchar
# For Sheet4, company may be float or NaN, handle carefully
sheet4_company = df_sheet4['Company'].apply(norm)
sheet4_account = df_sheet4['account'].apply(norm)
sheet4_xchar = df_sheet4['xchar'].apply(norm)

cbtrans_company = df_cbtrans['company'].apply(norm)
cbtrans_account = df_cbtrans['account'].apply(norm)
cbtrans_xchar = df_cbtrans['xchar'].apply(norm)

cbtrans_keys = set(zip(cbtrans_company, cbtrans_account, cbtrans_xchar))

result = []
for i in range(len(df_sheet4)):
    key = (sheet4_company[i], sheet4_account[i], sheet4_xchar[i])
    if key in cbtrans_keys:
        result.append('Match')
    else:
        result.append('-')

# Write results to output (Sheet4!K2:K7) and apply fill
wb = load_workbook(input_path)
ws = wb['Sheet4']

pink_fill = PatternFill(start_color='FF66CC', end_color='FF66CC', fill_type='solid')

for row, value in enumerate(result[:6], start=2): # K2:K7 (rows 2 to 7)
    cell = ws.cell(row=row, column=11)  # K is col 11
    cell.value = value
    cell.fill = pink_fill
    if isinstance(cell.value, str):
        cell.value = cell.value.capitalize()

wb.save(output_path)