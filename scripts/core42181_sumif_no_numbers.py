from openpyxl import load_workbook
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_4/regression_gate/before_pass/core_42181/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_4/regression_gate/before_pass/core_42181/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def contains_number(s):
    return bool(re.search(r'[0-9]', str(s)))

# Get filter text from A13
filter_text = ws['A13'].value

# Read values from columns I and B (rows 4 to 10)
col_I = [ws[f'I{row}'].value for row in range(4, 11)]
col_B = [ws[f'B{row}'].value for row in range(4, 11)]

# Calculate the sum
result = 0
for i, val in enumerate(col_I):
    if val is not None:
        val_str = str(val)
        if filter_text in val_str and not contains_number(val_str):
            result += col_B[i] if col_B[i] is not None else 0

# Write the result to B13
ws['B13'] = result
wb.save(output_path)
