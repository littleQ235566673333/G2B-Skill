import openpyxl

# Input and output paths
data_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_55965_tc1/input.xlsx'
out_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_55965_tc1/output.xlsx'

# Load workbook and sheets
wb = openpyxl.load_workbook(data_path)
sheet1 = wb['Foglio1']
# Assume the second sheet is the data sheet
sheets = wb.sheetnames
data_sheet = wb[sheets[1]]

# Read IDs from Sheet1 column E (E2:E18)
ids_sheet1 = [sheet1[f'E{i}'].value for i in range(2, 19)]

# Read all rows in the data sheet and collect rows where column K matches
# Get all columns G to P for each matched row
matches = {id: [] for id in ids_sheet1}
for row in range(2, data_sheet.max_row + 1):
    data_id = data_sheet[f'K{row}'].value
    if data_id in matches:
        # Get columns G to P
        row_values = [data_sheet[f'{col}{row}'].value for col in 'GHIJKLMNOP']
        matches[data_id].append(row_values)

# For each ID, get the last 10 matches (or fewer if not enough matches)
output_rows = []
for id in ids_sheet1:
    matched_rows = matches[id][-10:] if matches[id] else [[None]*8 for _ in range(10)]
    # Always 10 rows per ID
    num_matched = len(matches[id])
    if num_matched < 10:
        # Pad missing with None rows
        matched_rows = [[None]*8]*(10-num_matched) + matched_rows
    output_rows.extend(matched_rows)

# Write result to 'Foglio1'!G2:P18
row_idx = 2
for out_row in output_rows:
    for col_offset, val in enumerate(out_row):
        sheet1.cell(row=row_idx, column=7+col_offset, value=val)  # Column G=7, ... P=16
    row_idx += 1

wb.save(out_path)
