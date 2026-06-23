import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/train/iter_9/evolve_91-34/input.xlsx'
OUTPUT = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/train/iter_9/evolve_91-34/output.xlsx'
SHEET = 'SwiftMD'
FIRST_DATA_ROW = 2
LAST_DATA_ROW = 42
FIRST_DATA_COL = 2  # B
LAST_DATA_COL = 15  # O

def safe_date(x):
    if isinstance(x, datetime.datetime):
        return x.date()
    elif isinstance(x, datetime.date):
        return x
    return x

def main():
    wb = load_workbook(INPUT)
    ws = wb[SHEET]
    # Read all rows B2:O42 (inclusive)
    data = []
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW, max_row=LAST_DATA_ROW, min_col=FIRST_DATA_COL, max_col=LAST_DATA_COL):
        values = [cell.value for cell in row]
        data.append(values)

    # The first row in this block is the header
    header = data[0]
    rows = data[1:]
    # Column indexes
    COL_LAST = header.index('Last Name')
    COL_FIRST = header.index('First Name')
    COL_DOB = header.index('Date Of Birth')
    COL_DUP = header.index('Duplicate?')
    COL_REL = header.index('Relationship')

    # Group rows by (Last Name, First Name, DOB)
    from collections import defaultdict
    groups = defaultdict(list)
    for idx, row in enumerate(rows):
        key = (row[COL_LAST], row[COL_FIRST], safe_date(row[COL_DOB]))
        groups[key].append(idx)

    # Row indexes to delete (0-based within rows, NOT including header)
    to_delete = set()
    for key, idx_list in groups.items():
        if len(idx_list) > 1:
            # Check all Duplicate? == 'Yes' and none Relationship == 'Employee'
            dups = [rows[i][COL_DUP] for i in idx_list]
            rels = [str(rows[i][COL_REL]).strip().lower() if rows[i][COL_REL] else '' for i in idx_list]
            if all(str(d).strip().lower() == 'yes' for d in dups) and not any(r == 'employee' for r in rels):
                # Mark only one (e.g., the first in list) for deletion
                to_delete.add(idx_list[0])

    # Write output: header + all rows except those in to_delete
    new_data = [header] + [rows[i] for i in range(len(rows)) if i not in to_delete]

    # Clear output range first
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW+1):
        for c in range(FIRST_DATA_COL, LAST_DATA_COL+1):
            ws.cell(row=r, column=c, value=None)
    # Write back
    for r_idx, row in enumerate(new_data):
        for c_idx, val in enumerate(row):
            ws.cell(row=FIRST_DATA_ROW + r_idx, column=FIRST_DATA_COL + c_idx, value=val)

    wb.save(OUTPUT)

if __name__ == '__main__':
    main()
