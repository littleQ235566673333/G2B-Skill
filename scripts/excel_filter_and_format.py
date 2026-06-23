import openpyxl
from openpyxl.styles import Alignment

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_56378_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_56378_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Folha1']

# Extract products from Frame 1 (columns C:J, row 5-8) with non-empty QUANTITY UNITS
products = []
for row in ws.iter_rows(min_row=5, max_row=8, min_col=3, max_col=10, values_only=True):
    if row[-1] is not None and str(row[-1]).strip() != '' and row[1]:
        products.append(row)

# Output to Frame 2 (L5:R8)
start_row = 5
start_col = 12  # Excel column L
for i, product in enumerate(products):
    for j, value in enumerate(product):
        ws.cell(row=start_row+i, column=start_col+j, value=value)

# Align product names left (column M), values in columns O:R right (columns 15-18)
for i in range(len(products)):
    ws.cell(row=start_row+i, column=start_col+1).alignment = Alignment(horizontal='left')  # M
    for c in range(start_col+3, start_col+7):  # O:R
        ws.cell(row=start_row+i, column=c).alignment = Alignment(horizontal='right')

wb.save(output_path)
