from openpyxl import load_workbook

def is_blank_row(row):
    return all(cell is None for cell in row)

def main():
    input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_SAPR-T3-A5-seed1/train/iter_7/regression_gate/after_pass/core_160-6/input.xlsx'
    output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_SAPR-T3-A5-seed1/train/iter_7/regression_gate/after_pass/core_160-6/output.xlsx'
    ws_name = 'SH'
    start_output_row = 6
    end_output_row = 11
    start_col = 1  # A
    end_col = 12   # L

    wb = load_workbook(input_path)
    ws = wb[ws_name]

    # Collect all relevant rows (excluding header)
    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=end_col, values_only=True):
        if not is_blank_row(row):
            data_rows.append(row[:end_col])

    # Pad or truncate to fit exactly 6 rows in output
    result_rows = data_rows[:(end_output_row - start_output_row + 1)]
    while len(result_rows) < (end_output_row - start_output_row + 1):
        result_rows.append((None,) * end_col)

    # Write to output region
    for i, row_values in enumerate(result_rows):
        for j, value in enumerate(row_values):
            ws.cell(row=start_output_row + i, column=start_col + j, value=value)

    wb.save(output_path)

if __name__ == '__main__':
    main()
