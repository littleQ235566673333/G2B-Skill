from openpyxl import load_workbook

# Paths
input_path = 'results/runs/ITSM_B_v0_mask_R5_4.1_seed0/eval_r3/eval_395-36_tc1/input.xlsx'
output_path = 'results/runs/ITSM_B_v0_mask_R5_4.1_seed0/eval_r3/eval_395-36_tc1/output.xlsx'

wb = load_workbook(input_path)
main_ws = wb['Main unique ID']
get_ws = wb['Result what i am getting']

# 1. Read IDs from 'Main unique ID'
main_ids = [row[0].value for row in main_ws.iter_rows(min_row=1, max_col=1) if row[0].value is not None]

# 2. Read all rows from 'Result what i am getting'
rows = list(get_ws.iter_rows(values_only=True))
header = rows[0]
data_rows = rows[1:]
get_id_to_row = {row[0]: row for row in data_rows if row[0] is not None}

# 3. To produce output like in expectation: recompose from main_ids order, inserting available rows or blank rows for missing, and also preserving any extra rows in result
existing_ids = set(get_id_to_row.keys())
result_rows = []
for uid in main_ids:
    if uid in existing_ids:
        result_rows.append(get_id_to_row[uid])
    else:
        result_rows.append((uid,) + tuple([None]*(len(header)-1)))
# Add extra rows from result that are not in main_ids, in their original order
for row in data_rows:
    if row[0] not in main_ids:
        result_rows.append(row)
# 5. Write to MyResult sheet
if 'MyResult' in wb.sheetnames:
    del wb['MyResult']
my_ws = wb.create_sheet('MyResult')
for col, val in enumerate(header, 1):
    my_ws.cell(row=1, column=col, value=val)
for r_idx, row in enumerate(result_rows, 2):
    for c_idx, val in enumerate(row, 1):
        my_ws.cell(row=r_idx, column=c_idx, value=val)
wb.save(output_path)
print('Done')
