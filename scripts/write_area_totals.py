import pandas as pd
from openpyxl import load_workbook

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_5/group_263-1/r0/evolve_263-1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_5/group_263-1/r0/evolve_263-1/output.xlsx"

df = pd.read_excel(input_path)
# Drop rows where essential values are missing
cols_needed = ['Mtrl', 'Width', 'Height']
df_valid = df[cols_needed].dropna()
df_valid['Area'] = df_valid['Width'] * df_valid['Height']
summary = df_valid.groupby('Mtrl')['Area'].sum().reset_index()

wb = load_workbook(input_path)
ws = wb.active
# Write grouped total areas to H2:H4 as requested
for i, (_, row) in enumerate(summary.iterrows()):
    ws[f'H{2 + i}'] = row['Area']
wb.save(output_path)
