import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun2/eval_57989_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun2/eval_57989_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Identify header row (containing all weekdays)
weekdays = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
header_row = None
header_map = {}
for row in range(1, ws.max_row+1):
    vals = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column+1)]
    if any(v == 'Monday' for v in vals):
        for col in range(len(vals)):
            if vals[col] in weekdays:
                header_map[vals[col]] = col+1 # 1-based
        header_row = row
        break

# Find driver names and data rows (all in col 1, after the header row)
drivers = []
data_rows = []
row = header_row + 1
while row <= ws.max_row:
    val = ws.cell(row=row, column=1).value
    if val is not None and str(val).strip() != '':
        drivers.append(val)
        data_rows.append(row)
    row += 1

# For each driver, count non-blank values in their row for each weekday column
output_start_row = 25
output_start_col = 2
for i, (driver, data_row) in enumerate(zip(drivers, data_rows)):
    for j, day in enumerate(weekdays):
        col = header_map.get(day)
        count = 0
        if col is not None:
            val = ws.cell(row=data_row, column=col).value
            if val not in (None, '', ' '):
                count = 1
        ws.cell(row=output_start_row+i, column=output_start_col+j, value=count)

wb.save(output_path)
