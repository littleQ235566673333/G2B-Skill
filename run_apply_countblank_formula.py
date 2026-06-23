from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_5/regression_gate/after_pass/core_41969/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_5/regression_gate/after_pass/core_41969/output.xlsx"

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# Place formulas in A6, B6, C6
for i in range(3):
    start_col = 1 + i * 3  # 1-based column index (A=1, D=4, G=7)
    end_col = start_col + 2
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    formula = f"=COUNTBLANK({start_letter}3:{end_letter}3)"
    # Target is A6, B6, C6 (6th row, 1st to 3rd columns)
    ws.cell(row=6, column=1+i).value = formula

wb.save(output_path)
