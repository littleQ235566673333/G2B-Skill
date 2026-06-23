from openpyxl import load_workbook
from datetime import datetime

input_path = r"results/runs/g2b-v8_gpt-5.4_v82/eval_100slice_singleseed/task_8942/input.xlsx"
output_path = r"results/runs/g2b-v8_gpt-5.4_v82/eval_100slice_singleseed/task_8942/output.xlsx"

wb = load_workbook(input_path)

print('Sheets:', wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    print('\nSHEET', s)
    for r in range(1, min(ws.max_row, 15) + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 8) + 1)]
        print(r, vals)
