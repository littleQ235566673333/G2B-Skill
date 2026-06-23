from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_2/group_51-12/r3/evolve_51-12/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_2/group_51-12/r3/evolve_51-12/output.xlsx'

# Load workbook & active sheet
wb = load_workbook(input_path)
ws = wb.active

# Data range
row = 26
start_col = column_index_from_string('L')
end_col = column_index_from_string('SHA')
data = [ws.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]

count = 0
n = len(data)
i = 0
while i < n:
    ref = data[i]
    found = False
    for j in range(i + 1, n):
        # Skip if data or ref is None
        if ref is not None and data[j] is not None:
            try:
                if float(data[j]) >= float(ref) * 1.1:
                    count += 1
                    i = j
                    found = True
                    break
            except Exception:
                pass
    if not found:
        break

ws['B6'] = count
wb.save(output_path)
