import openpyxl

def find_and_paste_date(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    sheet_a = wb['A']
    sheet_b = wb['B']
    # 1. In col P (16th/column 16) of A, find first row with 'Yes' or 'NA'
    p_col = [cell.value for cell in sheet_a['P']]
    try:
        first_idx = next(i for i, val in enumerate(p_col[1:], start=2) if str(val).strip() in ['Yes', 'NA'])
        found = True
    except StopIteration:
        found = False
        first_idx = None

    if found:
        # 2. Search in column L (col 12), check in rows above this
        l_col = [sheet_a.cell(row=x, column=12).value for x in range(2, first_idx+1)] # rows 2 to first_idx-1
        hundred_row = None
        for rel_idx, val in enumerate(l_col[::-1]):  # search from closest above
            if val == 100:
                hundred_row = first_idx - rel_idx - 1
                break
        if hundred_row is not None and hundred_row >= 2:
            result_date = sheet_a.cell(row=hundred_row, column=6).value  # col F
        else:
            # 100 not found in col L above, default to F3
            result_date = sheet_a['F3'].value
    else:
        # If no Yes/NA found
        result_date = sheet_a['F3'].value
    # 3. Compare text in A!F1 to what is in B!C:C
    f1_text = sheet_a['F1'].value
    match_found = False
    for row in sheet_b.iter_rows(min_row=2, min_col=3, max_col=3):
        if row[0].value == f1_text:
            # paste to column F of same row
            sheet_b.cell(row=row[0].row, column=6).value = result_date
            match_found = True
            break
    # 4. Output answer also in B!F4
    sheet_b['F4'].value = result_date
    # save and make A active
    wb.active = wb.sheetnames.index('A')
    wb.save(output_path)

find_and_paste_date('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_333-29_tc1/input.xlsx',
                    'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_333-29_tc1/output.xlsx')
