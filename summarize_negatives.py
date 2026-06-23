import pandas as pd
import openpyxl

def summarize_consecutive_negatives(input_path, output_path):
    df = pd.read_excel(input_path, sheet_name='input', engine='openpyxl')
    df = df[['DATE ', 'Stock Name', 'OPENP* ', 'HIGH ', 'LOW ', 'CLOSEP* ', 'VOLUME', 'Change']]
    stock_col = 'Stock Name'
    date_col = 'DATE '
    open_col = 'OPENP* '
    close_col = 'CLOSEP* '
    high_col = 'HIGH '
    low_col = 'LOW '
    vol_col = 'VOLUME'
    
    for col in [open_col, close_col, high_col, low_col, vol_col, 'Change']:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df = df.dropna(subset=[date_col, close_col])

    summary_rows = []
    for stock, g in df.groupby(stock_col):
        g = g.sort_values(date_col).reset_index(drop=True)
        in_neg = False
        start_idx = None
        for i, row in g.iterrows():
            if row['Change'] < 0:
                if not in_neg:
                    in_neg = True
                    start_idx = i
            else:
                if in_neg:
                    end_idx = i - 1
                    if start_idx <= end_idx:
                        streak = g.loc[start_idx:end_idx]
                        summary_rows.append([
                            stock,
                            streak.iloc[0][date_col],
                            streak.iloc[-1][date_col],
                            streak.iloc[0][open_col],
                            streak.iloc[-1][close_col],
                            streak[high_col].max(),
                            streak[low_col].min(),
                            streak[vol_col].sum(),
                        ])
                    in_neg = False
        if in_neg:
            end_idx = len(g) - 1
            if start_idx <= end_idx:
                streak = g.loc[start_idx:end_idx]
                summary_rows.append([
                    stock,
                    streak.iloc[0][date_col],
                    streak.iloc[-1][date_col],
                    streak.iloc[0][open_col],
                    streak.iloc[-1][close_col],
                    streak[high_col].max(),
                    streak[low_col].min(),
                    streak[vol_col].sum(),
                ])
    # Write output
    book = openpyxl.load_workbook(input_path)
    ws = book['output']
    headers = [stock_col, 'Start Date', 'End Date', 'Open', 'Close', 'High', 'Low', 'Volume']
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    # Clear region
    for r in range(2, 16):
        for c in range(1, 8+1):
            ws.cell(row=r, column=c, value=None)
    # Fill output (A2:G15)
    for i, rec in enumerate(summary_rows[:14], 2):
        ws.cell(row=i, column=1, value=rec[0])
        ws.cell(row=i, column=2, value=rec[1].date() if pd.notnull(rec[1]) else None)
        ws.cell(row=i, column=3, value=rec[2].date() if pd.notnull(rec[2]) else None)
        ws.cell(row=i, column=4, value=rec[3])
        ws.cell(row=i, column=5, value=rec[4])
        ws.cell(row=i, column=6, value=rec[5])
        ws.cell(row=i, column=7, value=rec[6])
        ws.cell(row=i, column=8, value=rec[7])
    book.save(output_path)

summarize_consecutive_negatives(
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_61-4_tc1/input.xlsx',
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_61-4_tc1/output.xlsx'
)
