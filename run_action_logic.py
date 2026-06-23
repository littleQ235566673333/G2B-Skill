from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_3/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_3/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']
rownum = 4

# Read 'Last Contact' and 'YES/NO' from respective columns
last_contact_cell = ws['H' + str(rownum)].value
yn_response = ws['I' + str(rownum)].value

# Default value
result = 'NO ACTION'

today = datetime.today().date()

# Ensure last_contact_cell is a date, else set to None
last_contact = None
if hasattr(last_contact_cell, 'date'):
    last_contact = last_contact_cell.date()  # Already datetime
elif isinstance(last_contact_cell, datetime):
    last_contact = last_contact_cell.date()

if yn_response and yn_response.strip().upper() == 'YES' and last_contact:
    delta_days = (today - last_contact).days
    if delta_days <= 30:
        result = 'HOLD'
    else:
        result = 'TOUCH BASE'

ws['J' + str(rownum)] = result
wb.save(output_path)
