from openpyxl import load_workbook
import numpy as np

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_55468_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_55468_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Read the lookup criteria from the formula driver cells
AC4 = ws['AC4'].value  # 1st horizontal match (X)
AB4 = ws['AB4'].value  # 2nd horizontal match (X-sub)
AE4 = ws['AE4'].value  # 1st vertical match (Y)
AD4 = ws['AD4'].value  # 2nd vertical match (Y-sub)

# Y-axis main labels and sublabels
rows = list(range(5, 11))
y_labels = [ws[f'A{row}'].value for row in rows]
y_sub_labels = [ws[f'B{row}'].value for row in rows]

# X-axis main headers and subheaders
cols = list(range(3, 27))  # C to Z is 3 to 26
x_headers = [ws.cell(row=4, column=col).value for col in cols]
x_sub_headers = [ws.cell(row=3, column=col).value for col in cols]

# Data block
# Data block is C5:Z10 (6 rows, 24 columns)
data = np.array([[ws.cell(row=row, column=col).value for col in cols] for row in rows])

# Find matching row index where both y_labels and y_sub_labels match
row_idx = None
for i in range(len(rows)):
    if y_labels[i] == AE4 and y_sub_labels[i] == AD4:
        row_idx = i
        break

# Find matching column index where both x_headers and x_sub_headers match
col_idx = None
for j in range(len(cols)):
    if x_headers[j] == AC4 and x_sub_headers[j] == AB4:
        col_idx = j
        break

# Extract result
result = None
if row_idx is not None and col_idx is not None:
    result = data[row_idx, col_idx]

ws['AE5'] = result
wb.save(output_path)
