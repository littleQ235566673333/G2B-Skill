import openpyxl

# Paths
input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_57743_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_57743_tc1/output.xlsx'

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read model numbers and price lookup
model_numbers_col_A = [ws[f'A{row}'].value for row in range(2, 20)]  # A2:A19
model_numbers_col_C = [ws[f'C{row}'].value for row in range(2, ws.max_row+1)]
prices_col_D = [ws[f'D{row}'].value for row in range(2, ws.max_row+1)]

# Build lookup dict from column C (model #) to D (price)
lookup = {model: price for model, price in zip(model_numbers_col_C, prices_col_D)}

# Fill B2:B19 with matched prices, leave blank if not found or error
for idx, row in enumerate(range(2, 20)):
    model = model_numbers_col_A[idx]
    price = lookup.get(model, None)
    ws[f'B{row}'].value = price if price is not None else ''

# Save output
wb.save(output_path)
