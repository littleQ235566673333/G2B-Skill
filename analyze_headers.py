from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_120-24_tc1/input.xlsx')
ws = wb['Sheet1']
ncols = ws.max_column
header_map = {}
for col in range(1, ncols+1):
    val = ws.cell(row=1, column=col).value
    if val:
        header_map[val] = {'index': col, 'col_letter': get_column_letter(col)}
        print(f'{col} ({get_column_letter(col)}): {val}')
print('\n-- Sample values under the key headers --')
targets = ['BL', 'BG', 'AY', 'BN']
for t in targets:
    if t in header_map:
        col = header_map[t]['index']
        print(f'Column {t} (#{col}, {get_column_letter(col)}):')
        for r in range(1, 6):
            print(f'  Row {r}:', ws.cell(row=r, column=col).value)