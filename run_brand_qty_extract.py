import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/after_fix/core_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/after_fix/core_516-46/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['ورقة1']

# Extract data A2:E (date, brand, batch, origin, quantity)
data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5, values_only=True):
    if all([cell is None for cell in row]):
        continue
    data.append(row)
if not data:
    raise ValueError('No data found.')

# Build dataframe
cols = ['date','brand','batch','origin','quantity']
df = pd.DataFrame(data, columns=cols)
df = df.dropna(subset=['date', 'brand', 'quantity'])
df['date'] = pd.to_datetime(df['date'])

# Aggregate: For each brand, get the latest date and sum qty for that date
result_rows = []
for brand, group in df.groupby('brand'):
    last_date = group['date'].max()
    last_entries = group[group['date'] == last_date]
    sum_qty = last_entries['quantity'].sum()
    # Combine batch and origin, take the first
    batch = last_entries.iloc[0]['batch']
    origin = last_entries.iloc[0]['origin']
    result_rows.append([last_date, brand, batch, origin, sum_qty])

# Write to output (columns H-L start row 2)
for i, row in enumerate(result_rows, start=2):
    for j, val in enumerate(row, start=8):
        ws.cell(row=i, column=j, value=val)

wb.save(output_path)
