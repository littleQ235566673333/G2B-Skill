import openpyxl

input_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_7665_tc1/input.xlsx"
output_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_7665_tc1/output.xlsx"

# Load workbook and select active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Extract values from column H, skipping header if present
vertical_values = []
for row in ws["H"]:
    if row.row == 1:
        continue  # Skip header
    value = row.value
    if value is not None:
        vertical_values.append(value)

# Get unique sorted values
unique_sorted_values = sorted(set(vertical_values))

# Write values horizontally starting at Q2
start_col = openpyxl.utils.column_index_from_string('Q')
row = 2
for idx, val in enumerate(unique_sorted_values):
    ws.cell(row=row, column=start_col + idx, value=val)

# Save the result
wb.save(output_path)
