from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_4/group_36277/r3/evolve_36277/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_4/group_36277/r3/evolve_36277/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Get headers and worksheet size
data_max_col = ws.max_column
data_max_row = ws.max_row
headers = [ws.cell(row=1, column=col).value for col in range(1, data_max_col+1)]

# Automatic: Use header specified in H1 as lookup, else last column header
lookup_header = ws['H1'].value if ws['H1'].value else headers[-1]

# Calculate Excel column letter
from openpyxl.utils import get_column_letter
max_col_letter = get_column_letter(data_max_col)

# Build references
index_range = f'$A$1:${max_col_letter}${data_max_row}'
header_row_range = f'$A$1:${max_col_letter}$1'

for row in range(2, 6):  # I2:I5
    # Embed the lookup header as a string literal
    formula = f'=INDEX({index_range}, ROW(), MATCH(\"{lookup_header}\", {header_row_range}, 0))'
    ws[f'I{row}'] = formula

wb.save(output_path)
