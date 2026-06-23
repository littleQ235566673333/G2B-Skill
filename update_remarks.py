from openpyxl import load_workbook
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_52541_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_52541_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active
header_row = 5
row_start = 6
row_end = 10
headers = {cell.value: idx for idx, cell in enumerate(ws[header_row], 1)}
col_remarks = headers.get('Remarks', 10)
col_outstanding = headers.get('Amount Outstanding')
col_overdue = headers.get('Over Due Days')
def safe_num(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return None
for row in range(row_start, row_end+1):
    amt = safe_num(ws.cell(row=row, column=col_outstanding).value)
    overdue_raw = ws.cell(row=row, column=col_overdue).value
    overdue = safe_num(overdue_raw)
    if amt is not None and amt < 0:
        ws.cell(row=row, column=col_remarks).value = 'Prepaid'
    elif overdue_raw in (None, ""):
        ws.cell(row=row, column=col_remarks).value = ""
    elif overdue is not None and overdue < 90:
        ws.cell(row=row, column=col_remarks).value = "Call Customer"
    else:
        ws.cell(row=row, column=col_remarks).value = "Bad Debts"
wb.save(output_path)
