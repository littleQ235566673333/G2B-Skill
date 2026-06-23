import openpyxl
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_3/group_45707/r1/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_3/group_45707/r1/evolve_45707/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read all dates (A2:A69) and values in C2:C69
N = 68  # 68 rows

def parse_date(val):
    if isinstance(val, datetime.datetime):
        return val.date()
    elif isinstance(val, datetime.date):
        return val
    elif isinstance(val, str):
        try:
            # Try parsing ISO or common date formats
            return datetime.datetime.strptime(val, "%Y-%m-%d").date()
        except ValueError:
            try:
                return datetime.datetime.strptime(val, "%d/%m/%Y").date()
            except ValueError:
                try:
                    return datetime.datetime.strptime(val, "%m/%d/%Y").date()
                except ValueError:
                    return None
    return None

dates = [parse_date(ws[f'A{i}'].value) for i in range(2, 70)]
vals_c = [ws[f'C{i}'].value for i in range(2, 70)]

for idx in range(N):
    d_next = dates[idx+1] if idx+1 < N else None
    if d_next is not None and hasattr(d_next, 'day') and d_next.day == 1:
        month, year = d_next.month, d_next.year
        count_ones = 0
        for j, dd in enumerate(dates):
            if dd is not None and dd.month == month and dd.year == year and vals_c[j] == 1:
                count_ones += 1
        ws[f'D{idx+2}'].value = count_ones
    else:
        ws[f'D{idx+2}'].value = None

wb.save(output_path)
