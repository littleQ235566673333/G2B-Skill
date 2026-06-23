import openpyxl
from openpyxl.utils import get_column_letter

# Input and output paths
infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_52541_tc1/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_52541_tc1/output.xlsx'

wb = openpyxl.load_workbook(infile)
ws = wb.active

# Header is on row 5. Get indices for quick reference
HEADER_ROW = 5
for idx, cell in enumerate(ws[HEADER_ROW], start=1):
    if cell.value == 'Amount Outstanding':
        amt_col = idx
    if cell.value == 'Over Due Days':
        overdue_col = idx
    if cell.value == 'Remarks':
        remarks_col = idx

# Data is in rows 6 to 10
for row in range(6, 11):
    # Try converting formulas to values (if necessary)
    amt_val = ws.cell(row=row, column=amt_col).value
    overdue_val = ws.cell(row=row, column=overdue_col).value
    
    # Evaluate if possible
    try:
        amt = float(amt_val) if amt_val not in (None, "") else None
    except:
        amt = None
    try:
        overdue = int(overdue_val) if overdue_val not in (None, "") else None
    except:
        overdue = None

    if amt is not None and amt < 0:
        remark = "Prepaid"
    elif overdue is None:
        remark = ""
    elif overdue < 90:
        remark = "Call Customer"
    else:
        remark = "Bad Debts"

    ws.cell(row=row, column=remarks_col).value = remark

wb.save(outfile)
