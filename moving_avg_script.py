import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# File paths
input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_5/regression_gate/after_fix/core_4714/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_5/regression_gate/after_fix/core_4714/output.xlsx'

# Read spreadsheet using pandas for easier manipulation
df = pd.read_excel(input_file)

# Helper: parse month for window
def parse_date(cell):
    if pd.isnull(cell): return None
    if isinstance(cell, datetime): return cell
    try: return pd.to_datetime(cell)
    except: return None

# Month as datetime
month_col = df.columns[2]
df['Month_dt'] = df[month_col].apply(parse_date)

# Calculate moving average for each row 2-25 (excel)
def moving_average_4mo(row_idx, employee):
    row = df.iloc[row_idx]
    curr_month = row['Month_dt']
    if pd.isnull(curr_month):
        return 'n/a'
    start_month = curr_month - pd.DateOffset(months=3)
    # 4-month window (inclusive)
    mask = (df[df.columns[0]] == employee) & (df['Month_dt'] >= start_month) & (df['Month_dt'] <= curr_month)
    emp_rows = df[mask]
    if emp_rows.shape[0] < 4:
        return 'n/a'
    avg_val = emp_rows[df.columns[3]].mean()
    if pd.isnull(avg_val):
        return 'n/a'
    return int(avg_val) if avg_val == int(avg_val) else round(avg_val,1)

output_list = []
for idx in range(1,25): # Excel E2:E25 (indices 1-24 pandas)
    if idx >= len(df):
        output_list.append('n/a')
        continue
    employee = df.iloc[idx][df.columns[0]]
    val = moving_average_4mo(idx, employee)
    output_list.append(val)

# Write E2:E25 in output workbook
wb = load_workbook(input_file)
ws = wb.active
for offset, val in enumerate(output_list):
    ws.cell(row=2+offset, column=5).value = val
wb.save(output_file)
