import openpyxl
from datetime import datetime

in_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_38823_tc1/input.xlsx'
out_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_38823_tc1/output.xlsx'
wb = openpyxl.load_workbook(in_path)
ws = wb.active

# Get start and end dates from F3 and G3
start_date = ws['F3'].value
end_date = ws['G3'].value

# Data rows start from row 4
row = 4
results_data = []
while True:
    date_cell = ws[f'A{row}'].value
    tags_cell = ws[f'B{row}'].value
    units_cell = ws[f'C{row}'].value
    if date_cell is None:
        break
    results_data.append((date_cell, tags_cell, units_cell))
    row += 1

# For each search term in H4:H7
for out_idx, hrow in enumerate(range(4, 8), start=0):
    search_term = ws[f'H{hrow}'].value
    s = 0
    for d, tags, units in results_data:
        if not (d and tags and units):
            continue
        if start_date and end_date and start_date <= d <= end_date and search_term in tags:
            s += units
    ws[f'I{hrow}'] = s

wb.save(out_path)
