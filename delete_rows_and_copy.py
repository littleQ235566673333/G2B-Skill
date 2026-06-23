import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_374-18_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_374-18_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Imported Data']

rows_to_delete = []
for row in range(2, ws.max_row + 1):  # Start from 2 to skip header
    value = ws.cell(row=row, column=5).value
    # Check if the value is numeric and < 1
    try:
        if value is not None and float(value) < 1:
            rows_to_delete.append(row)
    except (ValueError, TypeError):
        pass

# Delete rows in reverse order to avoid shifting
for row in reversed(rows_to_delete):
    ws.delete_rows(row)

wb.save(output_path)

# Copy the resulting A2:G6 from 'Imported Data' to 'Sheet1'
ws_out = wb['Sheet1']
ws_in = wb['Imported Data']

for i, row in enumerate(ws_in.iter_rows(min_row=2, max_row=6, min_col=1, max_col=7), start=2):
    for j, cell in enumerate(row, start=1):
        ws_out.cell(row=i, column=j).value = cell.value

wb.save(output_path)
