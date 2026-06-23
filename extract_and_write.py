import openpyxl

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_178-22_tc1/input.xlsx'
out_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_178-22_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_fp)
ws1 = wb['Sheet1']
data = list(ws1.iter_rows(values_only=True))
headers = data[0]
filtered = [row for row in data[1:] if (row[1] == 'TELIVISION') or (row[2] == 'CLASS III') or (row[2] == 'CLASS IV')]

if 'Sheet2' not in wb.sheetnames:
    wb.create_sheet('Sheet2')
ws2 = wb['Sheet2']

# Clear Sheet2 content first
for row in ws2.iter_rows():
    for cell in row:
        cell.value = None

# Set headers in A2
for j, h in enumerate(headers, 1):
    ws2.cell(row=2, column=j, value=h)

# Put filtered data starting at A3
for i, row in enumerate(filtered, 3):
    for j, val in enumerate(row, 1):
        ws2.cell(row=i, column=j, value=val)

wb.save(out_fp)
