import openpyxl

# Load the workbook and worksheet
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_3/regression_gate/after_pass/core_50526/input.xlsx')
ws = wb.active

# Locate lookup value in B6
lookup_value = ws['B6'].value

# Get headers from row 1, skipping A1 (row headers)
headers = [cell.value for cell in ws[1]][1:]

# Find the row number where first column equals lookup_value
row_lookup = None
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value == lookup_value:
        row_lookup = row
        break

# Gather header names for non-zero (positive) values in that row
result = []
if row_lookup:
    for idx, cell in enumerate(row_lookup[1:]):
        if isinstance(cell.value, (int, float)) and cell.value > 0:
            result.append(headers[idx])

# Write up to 2 results in B9:B10
for i in range(2):
    ws.cell(row=9+i, column=2).value = result[i] if i < len(result) else None

# Save the workbook
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_3/regression_gate/after_pass/core_50526/output.xlsx')
