from openpyxl import load_workbook, styles
from openpyxl.utils import get_column_letter
from copy import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_3/regression_gate/before_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_3/regression_gate/before_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

actual_age_col = 16 # Column P, right after 'age (year)' in O
ws.insert_cols(actual_age_col)
header_cell = ws.cell(row=2, column=actual_age_col)
header_cell.value = 'Actual Age'

# Copy fill color, font, etc. from column O header without reference bugs
age_header = ws.cell(row=2, column=15)
age_fill = copy(age_header.fill)
age_font = copy(age_header.font)
age_alignment = copy(age_header.alignment)
header_cell.fill = age_fill
header_cell.font = styles.Font(bold=True)
header_cell.alignment = styles.Alignment(horizontal='center', vertical='top')

for row in range(3, 16):
    # Excel formula: whole years between report date (B$1) and DOB (C?)
    formula = f'=IF(AND(ISNUMBER($B$1),ISNUMBER(C{row})),DATEDIF(C{row},$B$1,"Y"),"")'
    cell = ws.cell(row=row, column=actual_age_col)
    cell.value = formula
    cell.number_format = '0'
    cell.alignment = styles.Alignment(horizontal='center', vertical='center')
    cell.fill = age_fill
    cell.font = styles.Font(bold=True)

# Set the Expected Result formula in E3:E15 (lookup CATEGORY from 'age (year)')
for row in range(3, 16):
    ws[f'E{row}'] = f'=INDEX($I$3:$I$15,MATCH(O{row},$O$3:$O$15,0))'

wb.save(output_path)
