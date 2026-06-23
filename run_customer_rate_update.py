import openpyxl
input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/group_44017/r1/evolve_44017/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/group_44017/r1/evolve_44017/output.xlsx'
wb = openpyxl.load_workbook(input_file)
ws = wb.active
# Columns: AD (30) to AO (43)
for row in range(14, 43):
    for col in range(30, 43):
        date_cell = ws.cell(row=9, column=col).coordinate  # e.g., AD9
        base = ws.cell(row=row, column=23).coordinate     # W
        eff_date = ws.cell(row=row, column=12).coordinate # L
        freq = ws.cell(row=row, column=10).coordinate     # J
        inc1 = ws.cell(row=row, column=13).coordinate    # M
        inc2 = ws.cell(row=row, column=14).coordinate    # N
        inc3 = ws.cell(row=row, column=15).coordinate    # O
        inc4 = ws.cell(row=row, column=16).coordinate    # P
        # Formula for cumulative increases
        # Only show if date >= eff_date
        # Use Excel LET for clarity
        formula = (
            f'=IF({date_cell}>={eff_date},'
            'LET('
            f'base,{base},freq,{freq},eff,{eff_date},d,{date_cell},'
            f'm1,{inc1},m2,{inc2},m3,{inc3},m4,{inc4},'
            'delta,IF(d>=eff,MAX(0,DATEDIF(eff,d,"m")),0),'
            'wave,IF(freq>0,MIN(4,INT(delta/freq)+1),1),'
            'mult,PRODUCT(1+IF(wave>=1,m1,0),1+IF(wave>=2,m2,0),1+IF(wave>=3,m3,0),1+IF(wave>=4,m4,0)),base*mult),""'
            ')'
        )
        ws.cell(row=row, column=col, value=formula)
wb.save(output_file)
