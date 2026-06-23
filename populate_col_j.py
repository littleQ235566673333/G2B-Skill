from openpyxl import load_workbook
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_5/regression_gate/after_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_5/regression_gate/after_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']
today = datetime.today()

for row in ws.iter_rows(min_row=4, min_col=8, max_col=9):
    h_cell = row[0]
    i_cell = row[1]
    j_cell = ws.cell(row=h_cell.row, column=10)  # J column
    h_val = h_cell.value
    i_val = (i_cell.value or '').strip().upper()
    result = 'NO ACTION'
    if isinstance(h_val, datetime) and i_val == 'YES':
        delta = today - h_val
        if delta.days <= 30:
            result = 'HOLD'
        else:
            result = 'TOUCH BASE'
    j_cell.value = result

wb.save(output_path)
