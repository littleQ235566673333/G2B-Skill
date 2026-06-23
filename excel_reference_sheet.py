from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_5/regression_gate/after_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_5/regression_gate/after_pass/core_41601/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Students']

# Go through rows 2 to 7
for row in range(2, 8):
    # Formula: =INDIRECT("'" & A2 & "'!C2")
    formula = f"=INDIRECT(\"'\"&A{row}&\"'!C2\")"
    ws[f'E{row}'] = formula

wb.save(output_path)
