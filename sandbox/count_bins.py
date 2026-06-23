from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_39903_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_39903_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb['SKUs as Bins']
font = Font(name='Courier New', size=9)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for row in range(2, 7):
    val = ws.cell(row=row, column=2).value
    if not val:
        count = 0
    else:
        bins = [x.strip() for x in val.split(',')]
        bins_ok = set()
        for b in bins:
            loc = b.split(':')[0].strip()
            if not (loc.startswith('X') or loc.startswith('Z')):
                bins_ok.add(loc)
        count = len(bins_ok)
    cell = ws.cell(row=row, column=3)
    cell.value = count
    cell.font = font
    cell.border = border

wb.save(output_path)
