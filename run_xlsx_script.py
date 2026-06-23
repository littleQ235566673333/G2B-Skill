import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_5/regression_gate/after_fix/core_38985/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_5/regression_gate/after_fix/core_38985/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active  # Use the first (main) worksheet

# Step 1: Read the range D8:D11 (names in table 1)
name_cells = ['D8', 'D9', 'D10', 'D11']
table1_names = [ws[cell].value for cell in name_cells]

# Step 2: Try G7 as table2 header by default (usual Excel conventions)
# If you want to auto-detect, adapt further. But here's a semi-robust way:
def find_table2_start(ws, min_row=1, max_row=30, min_col=1, max_col=20):
    for row in range(min_row, max_row + 1):
        values = [ws.cell(row=row, column=col).value for col in range(min_col, max_col + 1)]
        filled = [v for v in values if v not in (None, '\u200b', '')]
        if len(filled) >= 2:
            return row, [i for i,v in enumerate(values) if v][0] + min_col  # Row, col index of first header
    return None, None

t2_row, t2_col = find_table2_start(ws, min_row=5, max_row=20, min_col=7, max_col=10)
table2_values = {}
if t2_row is not None:
    # Columns of table2: Name (t2_col), Value1 (t2_col+1), Value2 ...
    for r in range(t2_row+1, t2_row+100):
        name = ws.cell(row=r, column=t2_col).value
        if name in (None, '', '\u200b'):
            break
        row_values = []
        col = t2_col + 1
        while ws.cell(row=t2_row, column=col).value not in (None, '', '\u200b'):
            row_values.append(ws.cell(row=r, column=col).value)
            col += 1
        table2_values.setdefault(name, []).append(row_values)

# Step 3: For each name in table1_names, find their table2 data and flatten (transpose) as needed.
# We will take all values returned, but if only one set, flatten it.
results = []
for name in table1_names:
    rows = table2_values.get(name, [])
    flat = []
    for row in rows:
        flat.extend(row)
    # If nothing found, leave blank
    results.append(flat[0] if flat else None)  # Per instruction, put the first value

# Step 4: Write result back into D8:D11
for idx, cell in enumerate(name_cells):
    ws[cell] = results[idx]

wb.save(output_path)
