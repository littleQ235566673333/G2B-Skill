import openpyxl
from openpyxl.styles import PatternFill, Font

# File paths
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_48620_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_48620_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Fill and font
fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
font = Font(name='Calibri')

# Range for output: E2:E7
output_range = range(2, 8)

for row in output_range:
    query = ws[f'D{row}'].value
    matches = []
    # Search column A for the query and collect matching column B values
    for i in range(2, ws.max_row+1):
        if ws[f'A{i}'].value == query:
            matches.append(ws[f'B{i}'].value)
    # Join all matches by ',' if multiple, else single value, or blank if none
    if not matches:
        val = ''
    else:
        # Optionally join matches with comma if multiple
        val = ', '.join([str(x) for x in matches if x != 0 and x is not None])
        # Show blank if value is explicitly 0
        if not val:
            val = ''
    ws[f'E{row}'].value = val
    ws[f'E{row}'].fill = fill
    ws[f'E{row}'].font = font

wb.save(output_path)
