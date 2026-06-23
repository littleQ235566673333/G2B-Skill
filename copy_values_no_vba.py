from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/group_48975/r2/evolve_48975/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/group_48975/r2/evolve_48975/output.xlsx'

wb = load_workbook(input_path)
ws_in = wb['Input']
ws_out = wb['Output']

output_row = 11
for row in ws_in.iter_rows(min_row=2, values_only=True):
    if len(row) < 5:
        continue
    a, b, _, _, todo = row[:5]
    if (todo is not None and str(todo).strip().lower() == 'yes' and
        a is not None and str(a).strip() != '' and
        b is not None and str(b).strip() != ''):
        ws_out.cell(row=output_row, column=2, value="{}, {}".format(a, b))
        output_row += 1
        if output_row > 17:
            break

wb.save(output_path)
