import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Load the workbook and the active sheet
input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_51262_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_51262_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Helper for Excel A1 notation
col_B, col_C, col_D = 2, 3, 4

# Fetch values from Excel
values = [ws.cell(row=i, column=col_D).value for i in range(6, 14)]
dates = [ws.cell(row=i, column=col_B).value for i in range(6, 14)]
criteria = [ws.cell(row=i, column=col_C).value for i in range(6, 14)]

# Get the month/year reference (cell F5)
f5_value = ws['F5'].value
if isinstance(f5_value, datetime):
    month = f5_value.month
    year = f5_value.year
else:
    # Try to parse from string if necessary
    try:
        dt = datetime.strptime(str(f5_value), "%Y-%m-%d")
        year = dt.year
        month = dt.month
    except Exception:
        raise ValueError("F5 does not contain a valid date")

# Get criteria list from F10:F14
criteria_list = [ws[f'F{i}'].value for i in range(10, 15)]

# Find sum for each criteria in F10:F14 list (for F6:H6)
result = []
for single_criteria in criteria_list:
    crit_sum = 0
    for dt, val, crit in zip(dates, values, criteria):
        if not isinstance(dt, datetime):
            try:
                dt = datetime.strptime(str(dt), "%Y-%m-%d")
            except:
                continue  # skip invalid dates
        if dt.year == year and dt.month == month and crit == single_criteria:
            crit_sum += val if val is not None else 0
    result.append(crit_sum)

# Write result starting at F6 (spanning across F6, G6, H6, but limited to first three)
for idx, val in enumerate(result[:3]):
    col = 6 + idx  # F=6, G=7, H=8
    ws.cell(row=6, column=col).value = val

wb.save(output_path)
