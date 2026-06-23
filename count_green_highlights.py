import openpyxl
from openpyxl.styles import PatternFill

# Define the green fill color used in Excel (common green hex values, modify if needed)
GREEN_HEXES = {
    'FF00FF00',  # bright green
    'FF92D050',  # Excel default green (light)
    'FF00B050',  # darker green
}

def is_green(cell):
    fill = cell.fill
    if not fill or not isinstance(fill, PatternFill):
        return False
    fg = fill.fgColor.rgb if fill.fgColor.type == 'rgb' else None
    return fg in GREEN_HEXES

# Load workbook
wb = openpyxl.load_workbook('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_43657_tc1/input.xlsx')
ws = wb.active

# Name range L2:L8, output K2:K8, check C3:G3 for each name
for idx, out_row in enumerate(range(2, 9), start=2):
    name = ws[f'L{out_row}'].value
    if not name:
        ws[f'K{out_row}'].value = 0
        continue
    count = 0
    for col in range(3, 8):  # C to G
        cell = ws.cell(row=3, column=col)
        if cell.value:
            cell_text = str(cell.value)
            # Check if name is in cell as substring and cell is green
            if name in cell_text and is_green(cell):
                count += 1
    ws[f'K{out_row}'] = count

# Save results
wb.save('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_43657_tc1/output.xlsx')
