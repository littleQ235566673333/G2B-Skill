import openpyxl
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_370-43_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_370-43_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Before Insert Row']

rows_to_insert = []
# Scan A7:A1000 (1-based indices)
for row in range(7, 1001):
    cell = ws[f'A{row}']
    if cell.value == 'X':
        rows_to_insert.append(row)

# Insert rows (from bottom to top)
for row in reversed(rows_to_insert):
    ws.insert_rows(row)

# Save the modified workbook
wb.save(output_path)
