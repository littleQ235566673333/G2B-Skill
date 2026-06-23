import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r3/eval_341-14_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r3/eval_341-14_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheetnames = wb.sheetnames
if 'problem' not in sheetnames or 'result' not in sheetnames:
    raise Exception(f'Sheet not found. Found: {sheetnames}')
ws_prob = wb['problem']
ws_res = wb['result']

# --- Identify header in problem and read data block
header_row = None
for r in range(1, ws_prob.max_row+1):
    if ws_prob.cell(row=r, column=1).value and ws_prob.cell(row=r, column=2).value:
        header_row = r
        break
if header_row is None:
    raise Exception('Header row not found in problem sheet')

problem_data = []
for r in range(header_row+1, ws_prob.max_row+1):
    val_a = ws_prob.cell(row=r, column=1).value
    val_b = ws_prob.cell(row=r, column=2).value
    if val_a is not None or val_b is not None:
        problem_data.append((val_a, val_b))

problem_dict = {}
for a, b in problem_data:
    if a is not None:
        problem_dict[str(a).strip()] = b

# --- Identify header in result and get the range for output
res_header_row = None
for r in range(1, ws_res.max_row+1):
    va = ws_res.cell(row=r, column=1).value
    vb = ws_res.cell(row=r, column=2).value
    if va is not None or vb is not None:
        res_header_row = r
        break
if res_header_row is None:
    raise Exception('Header row not found in result sheet')

entries = []
for r in range(res_header_row+1, res_header_row+28+1):
    entries.append(ws_res.cell(row=r, column=1).value)

# Write lookup results to col 2 in-place in result sheet
for idx, entry in enumerate(entries):
    result_cell = ws_res.cell(row=res_header_row+1+idx, column=2)
    if entry is not None and str(entry).strip() in problem_dict and problem_dict[str(entry).strip()] is not None:
        result_cell.value = problem_dict[str(entry).strip()]
    else:
        result_cell.value = None

wb.save(output_path)
