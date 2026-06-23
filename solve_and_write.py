import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_1/group_387-16/r0/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_1/group_387-16/r0/evolve_387-16/output.xlsx'

# Read full sheet for header and row details
df = pd.read_excel(input_path, sheet_name='Sheet1', header=1)

value_col = 'Value'
binaries_col = 'Binaries'
result_values_col = 'Target value :'

# The second row contains 'Result values' header, actual data is from row 2 onward
# Find actual result values, skipping the header row with 'Result values'
values_to_remove = df[result_values_col].tolist()[2:]  # skip 'Result values' and header
values_to_remove = [v for v in values_to_remove if pd.notna(v)]

values = df[value_col].tolist()
binaries = df[binaries_col].tolist()

# Remove only one instance from values/binaries for each result value
for key in values_to_remove:
    if key in values:
        idx = values.index(key)
        values.pop(idx)
        binaries.pop(idx)

# Shift-up logic
non_empty_vals = [v for v in values if pd.notna(v) and v != '']
values = non_empty_vals + [''] * (len(df) - len(non_empty_vals))
non_empty_bin = [v for v in binaries if pd.notna(v) and v != '']
binaries = non_empty_bin + [''] * (len(df) - len(non_empty_bin))

df[value_col] = values
df[binaries_col] = binaries

# Solver result is the sum of column A ('Value'), ignoring blanks
solver_result = sum([v for v in df[value_col] if isinstance(v,(int,float))])
# Target value at [0, '65477'] column (from header reading) if present
try:
    target_value = df.columns[df.columns.get_loc(65477)]
    target_num = 65477
except Exception:
    target_num = 0

difference = solver_result - target_num

# Write results
wb = load_workbook(input_path)
ws = wb['Sheet1']

# Write updated columns 'Value', 'Binaries'
for i in range(len(df)):
    ws.cell(row=2+i, column=1, value=df[value_col][i])  # A
    ws.cell(row=2+i, column=2, value=df[binaries_col][i])  # B

# Write solver result and difference as requested (C and D columns: A2:D18)
ws.cell(row=2, column=4, value=solver_result)    # D2
ws.cell(row=3, column=4, value=difference)       # D3

wb.save(output_path)
