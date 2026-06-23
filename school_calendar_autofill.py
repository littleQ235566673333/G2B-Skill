import openpyxl

# File paths
input_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_10/task_50916/r0/evolve_50916/input.xlsx"
output_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_10/task_50916/r0/evolve_50916/output.xlsx"

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Assume cycle day numbers reference table is in A2:A8 and class names in C2:H8
# Schedule entries for days to be filled are in C12:H14, with cycle day number in A12:A14

def get_cycle_map():
    cycle_map = {}
    for i in range(2, 9):  # A2:A8, 7-day
        cycle_num = ws[f"A{i}"].value
        periods = [ws[f"{col}{i}"].value for col in ['C','D','E','F','G','H']]
        cycle_map[cycle_num] = periods
    return cycle_map

# Get cycle number->periods mapping
cycle_map = get_cycle_map()

# Fill C12:H14 using column offsets
for row in range(12, 15):  # rows 12,13,14
    cycle_day = ws[f"A{row}"].value
    if cycle_day in cycle_map:
        periods = cycle_map[cycle_day]
        for j, col in enumerate(['C','D','E','F','G','H']):
            ws[f"{col}{row}"].value = periods[j]

# Save result
wb.save(output_path)
