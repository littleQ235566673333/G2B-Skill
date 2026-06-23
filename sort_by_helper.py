import pandas as pd
from openpyxl import load_workbook

# Load the input spreadsheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_22-47_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_22-47_tc1/output.xlsx'

wb = load_workbook(input_path)
sheet = wb.active

# Read the entire data as list of rows
data = []
for row in sheet.iter_rows(min_row=1, values_only=True):
    data.append(row)

# Find header row (assuming A column contains 'Name' header)
header_idx = 0
for i, row in enumerate(data):
    if row[0] and str(row[0]).lower().strip() in ('name',):
        header_idx = i
        break

columns = list(data[header_idx])
derived_data = data[header_idx:]

# Column indexes
col_A = 0
col_B = 1
col_C = 2
col_F = 5
col_G = 6
col_H = 7
col_J = 9

# Process the data (remove header, empty rows, and duplicate B & C)
seen = set()
filtered_rows = []
for row in derived_data[1:]:
    if not any(row):
        continue
    b, c = row[col_B], row[col_C]
    if (b, c) in seen:
        continue
    seen.add((b, c))
    if (b is None) or (str(b).strip() == ''):
        continue
    filtered_rows.append(row)

# Read helper (J) names in use-order, skipping blanks
j_names = []
for row in derived_data[1:]:
    j_val = row[col_J]
    if j_val and str(j_val).strip():
        j_names.append(str(j_val).strip())
j_names = list(dict.fromkeys(j_names))  # remove duplicates, keep order

# Split filtered into two buckets for grouping
in_j_group = []
rest = []
for row in filtered_rows:
    if str(row[col_B]).strip() in j_names:
        in_j_group.append(row)
    else:
        rest.append(row)

# Compose sorted output
output_rows = []
for name in j_names:
    for row in in_j_group:
        if str(row[col_B]).strip() == name:
            output_rows.append(row)
output_rows.extend(rest)

if not j_names:
    output_rows = sorted(filtered_rows, key=lambda x: str(x[col_B]).strip() if x[col_B] else '')

# Output to F2:H10
ws_out = wb.active
row_max = 9
for i in range(row_max):
    if i < len(output_rows):
        row_data = output_rows[i]
        for j in range(3):
            ws_out.cell(row=2+i, column=6+j, value=row_data[col_F+j])
    else:
        for j in range(3):
            ws_out.cell(row=2+i, column=6+j, value=None)

# For columns G, H: sort only column H (keep G-H rows paired)
gh_rows = [
    [ws_out.cell(row=2+i, column=7).value, ws_out.cell(row=2+i, column=8).value]
    for i in range(row_max)
]
gh_rows_nonempty = [r for r in gh_rows if r[1] is not None]
# Sort by H lowest to highest (try to convert to float, use string fallback)
def tryfloat(x):
    try:
        return float(x)
    except:
        return str(x)
sorted_gh = sorted(gh_rows_nonempty, key=lambda r: tryfloat(r[1]))
# Write sorted G, H back, clear rows below if empty
for i, (g, h) in enumerate(sorted_gh):
    ws_out.cell(row=2+i, column=7, value=g)
    ws_out.cell(row=2+i, column=8, value=h)
for i in range(len(sorted_gh), row_max):
    ws_out.cell(row=2+i, column=7, value=None)
    ws_out.cell(row=2+i, column=8, value=None)

wb.save(output_path)
