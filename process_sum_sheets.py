from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_2/regression_gate/before_pass/core_84-40/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_2/regression_gate/before_pass/core_84-40/output.xlsx'

wb = load_workbook(input_path)
sheets = wb.sheetnames

# Ensure LIST sheet exists
if 'LIST' not in sheets:
    list_ws = wb.create_sheet('LIST')
else:
    list_ws = wb['LIST']

# Clear LIST sheet (except possibly headers)
list_ws.delete_rows(2, list_ws.max_row)

# Entries: [(sheet_name, sum_C, sum_D, sum_C - sum_D)]
entries = []
for sheet in sheets:
    if sheet == 'LIST':
        continue
    ws = wb[sheet]
    # Gather all numeric values from column C and D (3 and 4)
    vals_C = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row+1)
              if isinstance(ws.cell(row=r, column=3).value, (int, float))]
    vals_D = [ws.cell(row=r, column=4).value for r in range(1, ws.max_row+1)
              if isinstance(ws.cell(row=r, column=4).value, (int, float))]
    sum_C = sum(vals_C)
    sum_D = sum(vals_D)
    entries.append((sheet, sum_C, sum_D, sum_C - sum_D))

# Write to LIST sheet
# Headers assumed already exist, otherwise add them
list_ws['B1'] = 'Sheet'
list_ws['C1'] = 'Sum C'
list_ws['D1'] = 'Sum D'
list_ws['E1'] = 'C - D'
for i, entry in enumerate(entries):
    list_ws.cell(row=i+2, column=2).value = entry[0]    # Sheet name in B
    list_ws.cell(row=i+2, column=3).value = entry[1]    # Sum C in C
    list_ws.cell(row=i+2, column=4).value = entry[2]    # Sum D in D
    list_ws.cell(row=i+2, column=5).value = entry[3]    # Diff in E
# Add total row
row_total = len(entries) + 2
list_ws.cell(row=row_total, column=2).value = 'Total'
for c in range(3, 6):
    col_letter = chr(64 + c)
    # Place SUM formula down the column (e.g. C2:C4)
    list_ws.cell(row=row_total, column=c).value = f'=SUM({col_letter}2:{col_letter}{row_total-1})'

wb.save(output_path)
