from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_54474_s2/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_54474_s2/output.xlsx'

wb = load_workbook(input_path)
whp = wb['WHP']
data = wb['WHP DATA']

# Build mapping from site name to (EEG, Dis, LTU) from WHP DATA.
site_map = {}
for row in range(1, data.max_row + 1):
    site = data.cell(row=row, column=2).value
    eeg = data.cell(row=row, column=3).value
    dis = data.cell(row=row, column=4).value
    ltu = data.cell(row=row, column=5).value
    if isinstance(site, str) and site not in {'Go Live 1/3/18', 'Local London Partnership', 'Central London Forward', 'Referral Type', 'Total'}:
        if any(v is not None for v in (eeg, dis, ltu)):
            site_map[site.strip()] = (eeg, dis, ltu)

# Fill E:G for rows 7:8 in WHP based on site names in column C.
for row in range(7, 9):
    site = whp.cell(row=row, column=3).value
    values = site_map.get(site.strip() if isinstance(site, str) else site, (None, None, None))
    for col, val in enumerate(values, start=5):
        whp.cell(row=row, column=col).value = val

wb.save(output_path)

# Verify
check_wb = load_workbook(output_path, data_only=False)
check_ws = check_wb['WHP']
for row in range(7, 9):
    print(row, [check_ws.cell(row=row, column=col).value for col in range(5, 8)])
