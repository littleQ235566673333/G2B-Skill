import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Load workbook and Sheet1
given_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_50631_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_50631_tc1/output.xlsx'
wb = openpyxl.load_workbook(given_path)
ws = wb['Sheet1']

# Calibri font and fill color
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
calibri_font = Font(name='Calibri')

# Helper function to check workdays

def is_workday(date, holidays):
    return date.weekday() < 5 and date not in holidays

# Get holidays from column J (J7:J38 or until blank)
holidays = set()
for row in range(7, 39):
    cell = ws[f'J{row}']
    if cell.value:
        if isinstance(cell.value, datetime):
            holidays.add(cell.value.date())
        elif isinstance(cell.value, str):
            try:
                holidays.add(datetime.strptime(cell.value, '%m/%d/%Y').date())
            except Exception:
                pass

# Start and end dates from B3 and F3
start_date = ws['B3'].value
end_date = ws['F3'].value
if isinstance(start_date, datetime):
    start_date = start_date.date()
if isinstance(end_date, datetime):
    end_date = end_date.date()

# Prepare column H: H7:H38
h_column_start = 7
h_column_end = 38
for i, row in enumerate(range(h_column_start, h_column_end+1)):
    # Compute nth workday after start_date
    n = i  # zero-based index
    current_date = start_date
    workdays_counted = 0
    while workdays_counted < n:
        current_date += timedelta(days=1)
        if is_workday(current_date, holidays):
            workdays_counted += 1
    # Only fill cell if in highlighted range (match yellow fill in existing Sheet1)
    cell = ws[f'H{row}']
    # Check fill color in input
    if cell.fill.start_color.index == 'FFFF00':
        ws[f'H{row}'].value = current_date.strftime('%m/%d/%Y')
        ws[f'H{row}'].font = calibri_font
        ws[f'H{row}'].fill = yellow_fill
    else:
        ws[f'H{row}'].value = None
    ws[f'H{row}'].font = calibri_font

# Apply Calibri font throughout worksheet
for row in ws.iter_rows():
    for cell in row:
        cell.font = calibri_font

# Save workbook
wb.save(output_path)
