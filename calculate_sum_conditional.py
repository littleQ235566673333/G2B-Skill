from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
import calendar

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun1/eval_57590_tc1/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun1/eval_57590_tc1/output.xlsx'

wb = load_workbook(INPUT)
ws = wb.active

a26 = ws['A26'].value

def infer_year_from_colC(month: str, ws):
    for i in range(2, ws.max_row + 1):
        date_val = ws[f'C{i}'].value
        try:
            date_in = pd.to_datetime(date_val)
        except Exception:
            continue
        if isinstance(date_in, pd.Timestamp):
            month_name = date_in.strftime('%B')
            if month_name.lower() == month.lower():
                return date_in.year
    return None

def get_month_bounds(a26, ws):
    if not a26:
        print('A26 is empty or None')
        return None, None
    if isinstance(a26, datetime):
        month_start = a26.replace(day=1)
    else:
        try:
            month_start = pd.to_datetime(a26).replace(day=1)
        except Exception as e:
            try:
                month_num = list(calendar.month_name).index(a26)
            except Exception:
                print('A26 not parseable and not a month name:', a26)
                return None, None
            year = infer_year_from_colC(a26, ws)
            if not year:
                print('Cannot infer year for month', a26)
                return None, None
            month_start = datetime(year, month_num, 1)
    month_end = month_start.replace(day=calendar.monthrange(month_start.year, month_start.month)[1])
    return month_start, month_end

month_start, month_end = get_month_bounds(a26, ws)
if month_start is None or month_end is None:
    ws['B26'] = 'INVALID A26'
    wb.save(OUTPUT)
    raise Exception(f'Cannot determine month/year from A26 value: {a26!r}')

vals = []
for i in range(2, ws.max_row + 1):
    date_val = ws[f'C{i}'].value
    inc_val = ws[f'I{i}'].value
    try:
        date_in = pd.to_datetime(date_val)
        # Only proceed if this is actual datetime
        if not isinstance(date_in, (datetime, pd.Timestamp)):
            continue
    except Exception:
        continue
    if month_start <= date_in <= month_end:
        try:
            vals.append(float(inc_val))
        except:
            pass
result = sum(vals)
ws['B26'] = result
wb.save(OUTPUT)
