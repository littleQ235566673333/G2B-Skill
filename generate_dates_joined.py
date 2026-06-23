import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_7/regression_gate/before_fix/core_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_7/regression_gate/before_fix/core_45896/output.xlsx'

# Read data
wb = load_workbook(input_path)
ws = wb['Volym P5_P6_2023']
df_main = pd.read_excel(input_path, sheet_name='Volym P5_P6_2023')
df_zord = pd.read_excel(input_path, sheet_name='ZORD')

def format_dates(dates):
    return ','.join([pd.to_datetime(d).strftime('%d/%m/%Y') for d in dates if pd.notna(d)])

# For rows 2-10
for ix in range(9):  # Data rows in Pandas are zero-indexed
    key = df_main.iloc[ix, 0]  # 'Material' value
    filtered = df_zord[df_zord['Material'] == key]['Valid from'].dropna().unique()
    res = format_dates(filtered)
    ws.cell(row=ix+2, column=9, value=res)

wb.save(output_path)
