import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_1/group_387-16/r0/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_1/group_387-16/r0/evolve_387-16/output.xlsx'

# Load dataframe
df = pd.read_excel(input_path, sheet_name='Sheet1', header=None)
df_copy = df.copy()

# Identify columns
value_col = None
bin_col = None
result_values_col = None
target_col = None
for col_idx, col in enumerate(df_copy.iloc[0]):
    col_str = str(col).strip().lower()
    if col_str == 'value':
        value_col = col_idx
    if col_str == 'binaries':
        bin_col = col_idx
    if col_str == 'result values':
        result_values_col = col_idx
    if col_str == 'target value':
        target_col = col_idx
if value_col is None or bin_col is None or result_values_col is None:
    raise Exception('Required columns not found.')

# Get all result values (skip header)
result_values = df_copy.iloc[1:, result_values_col].dropna().tolist()
value_indices_done = set()

# Remove one instance where Value == Result values (from Value and Binaries), only one occurrence per value
for match_val in result_values:
    match_indices = df_copy.iloc[1:, value_col][df_copy.iloc[1:, value_col] == match_val].index
    for idx in match_indices:
        if idx not in value_indices_done:
            value_indices_done.add(idx)
            break
rows_to_remove = sorted(value_indices_done)

for idx in rows_to_remove:
    df_copy.iat[idx, value_col] = None
    df_copy.iat[idx, bin_col] = None

def compact_col(col_idx):
    non_blank = df_copy.iloc[1:, col_idx].dropna().tolist()
    for i in range(1, len(df_copy)):
        df_copy.iat[i, col_idx] = non_blank[i-1] if i-1 < len(non_blank) else None

compact_col(value_col)
compact_col(bin_col)

# Calculate solver result and difference
target_value = df_copy.iat[1, target_col] if target_col is not None else None
solver_result = df_copy.iloc[1:, 0].dropna().sum()
diff = solver_result - target_value if target_value is not None else None
# Place results in designated columns (last two)
df_copy.iat[1, df_copy.shape[1]-2] = solver_result
if diff is not None:
    df_copy.iat[1, df_copy.shape[1]-1] = diff

# Write result to Sheet1!A2:D18 (A1 is header)
wb = load_workbook(input_path)
ws = wb['Sheet1']
for r in range(2, 19):
    for c in range(1, 5):
        ws.cell(row=r, column=c).value = (
            df_copy.iat[r-1, c-1] if r-1 < len(df_copy) and c-1 < len(df_copy.columns) else None
        )
wb.save(output_path)
