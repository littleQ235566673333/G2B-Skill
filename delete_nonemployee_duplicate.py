from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_5/regression_gate/after_fix/core_91-34/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_5/regression_gate/after_fix/core_91-34/output.xlsx'

wb = load_workbook(input_path)
ws = wb['SwiftMD']

header = [cell.value for cell in list(ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=15))[0]]
rows = []
for row in ws.iter_rows(min_row=3, max_row=42, min_col=2, max_col=15):
    rows.append([cell.value for cell in row])

idx = {name: i for i, name in enumerate(header)}

# Step 1: group 'Yes' dups by (last, first, dob)
yes_dups = {}
for i, row in enumerate(rows):
    if row[idx['Duplicate?']] == 'Yes':
        k = (row[idx['Last Name']], row[idx['First Name']], row[idx['Date Of Birth']])
        if k not in yes_dups:
            yes_dups[k] = []
        yes_dups[k].append(i)

# Step 2: determine for each group what to delete
to_delete = set()
for ids in yes_dups.values():
    rels = [rows[j][idx['Relationship']] for j in ids]
    if all(r != 'Employee' for r in rels) and len(ids) > 1:
        to_delete.add(ids[0])  # Only delete one

# Step 3: output keeping all not-deleted rows
new_rows = [row for i, row in enumerate(rows) if i not in to_delete]

# Step 4: write back to B2:O, keep header
for i, val in enumerate(header, start=2):
    ws.cell(row=2, column=i, value=val)
for r, row in enumerate(new_rows, start=3):
    for c, val in enumerate(row, start=2):
        ws.cell(row=r, column=c, value=val)
# clear any extra existing rows out to 42
for r in range(3+len(new_rows), 43):
    for c in range(2,16):
        ws.cell(row=r, column=c, value=None)

wb.save(output_path)
