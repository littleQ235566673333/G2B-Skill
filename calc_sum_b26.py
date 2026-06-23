import openpyxl
from openpyxl.utils.datetime import from_excel
from datetime import datetime, timedelta
import calendar

input_path = "results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_57590_tc1/input.xlsx"
output_path = "results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_57590_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get the value in A26, which might be a string like 'January', a date, etc.
cell_a26 = ws["A26"].value

def parse_month(cell_value):
    # Try parsing as datetime first
    if isinstance(cell_value, datetime):
        return cell_value.year, cell_value.month
    # Try parsing as month name
    try:
        month_num = list(calendar.month_name).index(str(cell_value))
        if month_num > 0:
            # Guess year as the year of the latest date in column C
            years = []
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                cell_c = row[0].value
                if isinstance(cell_c, datetime):
                    years.append(cell_c.year)
                elif isinstance(cell_c, str):
                    try:
                        d = datetime.strptime(cell_c, "%Y-%m-%d")
                        years.append(d.year)
                    except:
                        continue
            year = max(years) if years else datetime.now().year
            return year, month_num
    except:
        pass
    # Try as YYYY-MM, YYYY/MM
    for fmt in ("%Y-%m", "%Y/%m", "%B %Y", "%b %Y"):
        try:
            dt = datetime.strptime(str(cell_value), fmt)
            return dt.year, dt.month
        except:
            continue
    raise ValueError("Unrecognized month in A26: " + str(cell_value))

year, month = parse_month(cell_a26)

start_date = datetime(year, month, 1)
# Get last day of the month
last_day = calendar.monthrange(year, month)[1]
end_date = datetime(year, month, last_day, 23, 59, 59)

total_sum = 0.0
for row in ws.iter_rows(min_row=2):
    cell_c = row[2].value  # Column C
    cell_i = row[8].value  # Column I
    if cell_c is None or cell_i is None:
        continue
    # Handle Excel dates (if necessary)
    if isinstance(cell_c, (int, float)):
        cell_c = from_excel(cell_c)
    elif isinstance(cell_c, str):
        try:
            cell_c = datetime.strptime(cell_c, "%Y-%m-%d")
        except:
            continue
    if start_date <= cell_c <= end_date:
        try:
            num_value = float(cell_i)
        except:
            continue
        total_sum += num_value

ws["B26"] = total_sum
wb.save(output_path)
