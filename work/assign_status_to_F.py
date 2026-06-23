from openpyxl import load_workbook
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_6/regression_gate/after_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_6/regression_gate/after_pass/core_55421/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Step 1: Preload all relevant rows
data_rows = []
for row in ws.iter_rows(min_row=2, max_row=20, max_col=6, values_only=True):
    data_rows.append(row)

# Step 2: Group by numbers in col A (index 0)
groups = defaultdict(list)
for idx, row in enumerate(data_rows):
    num = row[0]
    groups[num].append((idx, row))

# Step 3: Logic for each group
results = [''] * len(data_rows)
for num, rowlist in groups.items():
    statuses = set(r[1][3] for r in rowlist)  # Column D (status)
    has_noshow = 'NO SHOW' in statuses
    has_sch = 'SCH' in statuses
    # If only SCH
    if statuses == {'SCH'}:
        for idx, _ in rowlist:
            results[idx] = 'FUTURE'
    # Both SCH and NO SHOW
    elif has_sch and has_noshow:
        for idx, _ in rowlist:
            results[idx] = 'NS/SCHED'
    else:
        # Check remaining rules for NO SHOW
        for idx, row in rowlist:
            status = row[3]
            date_e = row[4]
            if status == 'NO SHOW':
                if date_e is not None and str(date_e).strip() != '':
                    results[idx] = 'NO ACTION NEEDED'
                else:
                    results[idx] = 'CALL PT'

# Step 4: Write results to F2:F20
for i, val in enumerate(results, start=2):
    ws[f'F{i}'] = val

wb.save(output_path)
