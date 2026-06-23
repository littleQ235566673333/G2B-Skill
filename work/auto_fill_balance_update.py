from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_4/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_4/regression_gate/after_pass/core_32337/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

header_row = 2
row_start, row_end = 3, 15

# Helper to map header to column (1-based)
def find_col(header):
    for col in range(1, ws.max_column+1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip().lower() == header.strip().lower():
            return col
    raise Exception(f'Column header not found: {header}')

col_expected_result = find_col('Expected Result')           # E
col_age_year = find_col('age (year)')                      # O
col_category = find_col('CATEGORY')                        # I
col_year_age = find_col('YEAR (age)')                      # H
col_dob = find_col('DATE OF BIRTH')                        # C
col_report_date = 2  # B1 contains REPORT Date
col_actual_age = col_age_year + 1

# Try to match fill color from age(year) or fallback to default
age_year_header_cell = ws.cell(row=header_row, column=col_age_year)
age_year_fill_obj = age_year_header_cell.fill
if hasattr(age_year_fill_obj, 'start_color') and age_year_fill_obj.start_color and age_year_fill_obj.start_color.type == 'rgb':
    match_fill = PatternFill(start_color=age_year_fill_obj.start_color.rgb, end_color=age_year_fill_obj.end_color.rgb, fill_type=age_year_fill_obj.fill_type)
else:
    match_fill = PatternFill(fill_type=None)

# --- Step 1: Fill E3:E15 with a formula for CATEGORY based on age (O), looking up in O & I ---
for r in range(row_start, row_end+1):
    ws.cell(row=r, column=col_expected_result).value = f'=XLOOKUP(O{r},O$3:O$100,I$3:I$100,"")'

# --- Step 2: Insert Actual Age column after age (year) ---
ws.insert_cols(col_actual_age)
ws.cell(row=header_row, column=col_actual_age, value='Actual Age')

# Set header styling
header_cell = ws.cell(row=header_row, column=col_actual_age)
header_cell.font = Font(bold=True)
header_cell.alignment = Alignment(horizontal='center', vertical='top')
header_cell.fill = match_fill

# --- Step 3: Fill Actual Age (whole number) ---
for r in range(row_start, ws.max_row+1):
    dob_cell = ws.cell(row=r, column=col_dob).coordinate
    actual_age_cell = ws.cell(row=r, column=col_actual_age)
    actual_age_cell.value = f'=INT(YEARFRAC({dob_cell},$B$1,1))'
    actual_age_cell.number_format = '0'
    actual_age_cell.font = Font(bold=True)
    actual_age_cell.alignment = Alignment(horizontal='center')
    actual_age_cell.fill = match_fill

wb.save(output_path)
