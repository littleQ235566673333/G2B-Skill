import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/regression_gate/after_pass/core_32337/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/regression_gate/after_pass/core_32337/output.xlsx'

wb = load_workbook(in_path)
ws = wb['Sheet1']
headers = [cell.value for cell in ws[2]]
DOB_col = headers.index('DATE OF BIRTH') + 1
age_year_col = headers.index('age (year)') + 1
cat_col = headers.index('CATEGORY') + 1
exp_result_col = headers.index('Expected Result') + 1
new_col_idx = age_year_col + 1
ws.insert_cols(new_col_idx)
ws.cell(row=2, column=new_col_idx).value = 'Actual Age'
age_fill = copy.copy(ws.cell(row=2, column=age_year_col).fill)
ws.cell(row=2, column=new_col_idx).fill = age_fill
ws.cell(row=2, column=new_col_idx).alignment = Alignment(horizontal='center', vertical='top')
for row in range(3, 16):
    cell = ws.cell(row=row, column=new_col_idx)
    cell.value = f'=DATEDIF(C{row},$B$1,"Y")'
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
for row in range(3, 16):
    ws.cell(row=row, column=exp_result_col).value = f'=INDEX($I$3:$I$15, MATCH(O{row}, $O$3:$O$15, 0))'
wb.save(out_path)
