from openpyxl import load_workbook
from datetime import datetime
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_3/regression_gate/after_fix/core_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_3/regression_gate/after_fix/core_4714/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet2']
data = []
for row in ws.iter_rows(min_row=2, max_row=25, max_col=4, values_only=True):
    emp, ym, month, hours = row
    data.append({'emp': emp, 'month_ym': ym, 'month': month, 'hours': hours})
def is_within_n_months(src, target, n):
    d1 = src.year*12 + src.month
    d2 = target.year*12 + target.month
    return (d1 - d2) >= 0 and (d1 - d2) < n
def format_number(num):
    if num is None:
        return 'n/a'
    if isinstance(num, float) and num.is_integer():
        return str(int(num))
    return f"{num:g}"
results = []
for idx, row in enumerate(data):
    emp = row['emp']
    cur_month = row['month']
    # Find all same-employee rows where month is within cur_month and 3 previous
    window = [d['hours'] for d in data if d['emp'] == emp and is_within_n_months(cur_month, d['month'], 4)]
    if len(window) < 4:
        results.append('n/a')
    else:
        # If more than 4 months, pick latest 4
        if len(window) > 4:
            months_sorted = sorted([d for d in data if d['emp'] == emp and is_within_n_months(cur_month, d['month'], 4)], key=lambda d: d['month'], reverse=True)
            hours_4 = [d['hours'] for d in months_sorted[:4]]
            avg = sum(hours_4)/4
        else:
            avg = sum(window)/4
        results.append(format_number(avg))
# Write to column E2:E25
for i, v in enumerate(results):
    ws.cell(row=2+i, column=5).value = v
wb.save(output_path)
print('Done.')