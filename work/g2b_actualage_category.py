import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import datetime

def get_header_color(cell):
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type == 'rgb':
        return fill.fgColor.rgb
    return None

input_xlsx = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/regression_gate/before_pass/core_32337/input.xlsx'
output_xlsx = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/regression_gate/before_pass/core_32337/output.xlsx'
wb = openpyxl.load_workbook(input_xlsx)
ws = wb['Sheet1']
def find_col(header):
    for i, cell in enumerate(ws[2]):
        if cell.value == header:
            return i
    return -1
# Row 1: ws[0]: Report Date in col 0 (A) value is in B
col_report_date = 0  # 'REPORT Date' is always column 0 (A), value in col 1 (B)
col_report_date_val = 1
col_dob = find_col('DATE OF BIRTH')
col_expected_result = find_col('Expected Result')
col_category = find_col('CATEGORY')
col_year_age = find_col('YEAR (age)')
col_age_year = find_col('age (year)')
insert_at = col_age_year + 2 # 1-based openpyxl

# Get header fill color before insert
header_fill_rgb = get_header_color(ws.cell(row=2, column=col_age_year+1))
ws.insert_cols(insert_at)
ws.cell(row=2, column=insert_at).value = 'Actual Age'
ws.cell(row=2, column=insert_at).alignment = Alignment(horizontal='center', vertical='top')
if header_fill_rgb:
    ws.cell(row=2, column=insert_at).fill = PatternFill(start_color=header_fill_rgb, end_color=header_fill_rgb, fill_type='solid')

# Adjust columns (those >= insert_at-1 shift right)
col_expected_result += 1 if col_expected_result >= insert_at-1 else 0
col_year_age += 1 if col_year_age >= insert_at-1 else 0
col_category += 1 if col_category >= insert_at-1 else 0
col_dob += 1 if col_dob >= insert_at-1 else 0

for r in range(3, 16):
    dob = ws.cell(row=r, column=col_dob+1).value
    report_date = ws.cell(row=1, column=col_report_date_val+1).value
    if isinstance(dob, datetime.datetime) and isinstance(report_date, datetime.datetime):
        age = report_date.year - dob.year - ((report_date.month, report_date.day) < (dob.month, dob.day))
        cell = ws.cell(row=r, column=insert_at)
        cell.value = age
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        if header_fill_rgb:
            cell.fill = PatternFill(start_color=header_fill_rgb, end_color=header_fill_rgb, fill_type='solid')
        cell.number_format = '0'
    year_val = ws.cell(row=r, column=col_year_age+1).value
    match_cat = None
    for catrow in range(3, 16):
        if ws.cell(row=catrow, column=col_year_age+1).value == year_val:
            match_cat = ws.cell(row=catrow, column=col_category+1).value
            break
    ws.cell(row=r, column=col_expected_result+1).value = match_cat

wb.save(output_xlsx)
