import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_50631_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_50631_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read holidays from column J (assuming they're date values, starting from J7)
holidays = []
for row in ws['J']:
    if row.row >= 7 and row.value:
        if isinstance(row.value, datetime):
            holidays.append(row.value.date())
        else:
            try:
                holidays.append(datetime.strptime(str(row.value), '%m/%d/%Y').date())
            except:
                continue

def is_workday(date, holidays):
    if date.weekday() >= 5:
        return False  # Saturday or Sunday
    if date in holidays:
        return False
    return True

def nth_workday_after(start_date, n, holidays):
    i = 0
    current = start_date
    while i < n:
        current += timedelta(days=1)
        if is_workday(current, holidays):
            i += 1
    return current

# Find B3
start_date = ws['B3'].value
if not isinstance(start_date, datetime):
    start_date = datetime.strptime(str(start_date), '%m/%d/%Y')
start_date = start_date.date()

H_start = 7
H_end = 38

for i, row in enumerate(range(H_start, H_end+1), start=0):
    cell = ws[f'H{row}']
    b_candidate = ws[f'B{row}'].value
    if b_candidate is not None:
        # Parse b_candidate to date if needed
        if isinstance(b_candidate, datetime):
            b_date = b_candidate.date()
        else:
            try:
                b_date = datetime.strptime(str(b_candidate), '%m/%d/%Y').date()
            except:
                cell.value = None
                continue
        # Determine workday offset (number of workdays from B3 up to that B value, not counting the start)
        n = 0
        temp = start_date
        while temp < b_date:
            temp += timedelta(days=1)
            if is_workday(temp, holidays):
                n += 1
        # Find the actual workday
        date_result = nth_workday_after(start_date, n, holidays) if n >= 0 else ''
        cell.value = date_result
        cell.number_format = 'mm/dd/yyyy'
    else:
        cell.value = None

calibri = Font(name='Calibri')
for row in ws.iter_rows(min_row=H_start, max_row=H_end, min_col=8, max_col=8):
    for cell in row:
        cell.font = calibri
        if cell.fill and cell.fill.start_color.index == 'FFFFFF00':
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

wb.save(output_path)
