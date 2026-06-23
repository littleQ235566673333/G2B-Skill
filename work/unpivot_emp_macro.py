import pandas as pd
from openpyxl import load_workbook

infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_4/regression_gate/before_fix/core_547-43/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_4/regression_gate/before_fix/core_547-43/output.xlsx'

emp = pd.read_excel(infile, sheet_name='Emp')
lookup = pd.read_excel(infile, sheet_name='Lookup')

metric_vars = [str(v) for v in lookup['Variable'].dropna().tolist()]
result_rows = []
# For each row in Emp, add output for each variable in lookup
for _, row in emp.iterrows():
    if pd.isnull(row['seqno']):
        continue
    for var in metric_vars:
        if var in emp.columns and not pd.isnull(row[var]):
            cat_row = lookup[lookup['Variable']==var].iloc[0]
            result_rows.append([
                row['seqno'], row['Empno'], row['ename'], row['Jobid'],
                var, cat_row['Category'], cat_row['Subcategory'], row[var]
            ])

wb = load_workbook(infile)
ws = wb['Expected']

# Optionally clear old contents under the header
for r in range(2, ws.max_row+1):
    for c in range(1, 9):
        ws.cell(row=r, column=c, value=None)
# Write output rows
for r, vals in enumerate(result_rows, 2):
    for c, v in enumerate(vals, 1):
        ws.cell(row=r, column=c, value=v)
wb.save(outfile)
