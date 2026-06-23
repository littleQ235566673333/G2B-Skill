from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Paths
in_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/32337/input.xlsx'
out_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/32337/output.xlsx'

wb = load_workbook(in_path)
ws = wb['Sheet1']

e_col, i_col, h_col, o_col, p_col = 5, 9, 8, 15, 16
row_start, row_end = 3, 15

# Set Expected Result formula
for r in range(row_start, row_end + 1):
    o_cell = f'{get_column_letter(o_col)}{r}'
    formula = f'=IFERROR(INDEX($I$3:$I$15, MATCH({o_cell}, $H$3:$H$15, 0)), "")'
    ws.cell(row=r, column=e_col).value = formula

# Set header for Actual Age
ws.cell(row=2, column=p_col).value = 'Actual Age'
ws.cell(row=2, column=p_col).alignment = Alignment(horizontal='center', vertical='top')
ws.cell(row=2, column=p_col).font = Font(bold=True)

# Try copying fill from column O
sample_fill = ws.cell(row=2, column=o_col).fill
if isinstance(sample_fill, PatternFill) and sample_fill.fill_type is not None:
    ws.cell(row=2, column=p_col).fill = sample_fill

# Actual Age formula and formatting
for r in range(row_start, row_end + 1):
    formula = f'=DATEDIF(C{r}, $B$1, "Y")'
    cell = ws.cell(row=r, column=p_col)
    cell.value = formula
    cell.number_format = '0'
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    sample_fill = ws.cell(row=r, column=o_col).fill
    if isinstance(sample_fill, PatternFill) and sample_fill.fill_type is not None:
        cell.fill = sample_fill

wb.save(out_path)
print('Done')
