from openpyxl import load_workbook

# Load the workbook and worksheets
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun1/eval_341-14_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun1/eval_341-14_tc1/output.xlsx'
wb = load_workbook(input_path)
ws_prob = wb['problem']
ws_result = wb['result']

# Read the dataset from columns A and B of 'problem'
problem_data = []
for row in ws_prob.iter_rows(min_row=1, max_row=ws_prob.max_row, min_col=1, max_col=2, values_only=True):
    if row[0] is not None or row[1] is not None:
        problem_data.append(row)

# Read what is already in 'result'!A1:B28 as template
result_template = []
for row in ws_result.iter_rows(min_row=1, max_row=28, min_col=1, max_col=2, values_only=True):
    result_template.append(list(row))

# Build mapping from the input
mapping = {}
for k, v in problem_data:
    mapping[k] = v

# Go through results template and fill matching numbers from mapping
for i in range(28):
    key = result_template[i][0]
    if key in mapping and key is not None:
        ws_result.cell(row=i+1, column=2).value = mapping[key]
    elif key is not None:
        ws_result.cell(row=i+1, column=2).value = None

# Save to output
wb.save(output_path)
