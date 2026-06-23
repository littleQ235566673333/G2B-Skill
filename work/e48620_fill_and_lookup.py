import openpyxl
from openpyxl.styles import PatternFill, Font

# Paths
infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_48620_tc1/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_48620_tc1/output.xlsx'

wb = openpyxl.load_workbook(infile)
ws = wb.active

# The source data is in columns A (product), B (serial/desc), search in D2:D7, write output to E2:E7

fill = PatternFill(fill_type='solid', fgColor='FCE4D6')
font = Font(name='Calibri')

# Build a reference list of products and their serials
products = []
serials = []

# Find the last data row in col A (assuming no gaps)
row = 1
while ws.cell(row=row, column=1).value is not None:
    products.append(ws.cell(row=row, column=1).value)
    serials.append(ws.cell(row=row, column=2).value)
    row += 1

# Now, for D2:D7, look up in products, output matching serial in E2:E7
for i in range(2, 8):
    key = ws.cell(row=i, column=4).value
    # Try to find all matches in products
    for j, p in enumerate(products):
        if key == p:
            value = serials[j]
            break
    else:
        value = 0
    # Write result
    ws.cell(row=i, column=5).value = '' if value == 0 else value
    cell = ws.cell(row=i, column=5)
    cell.font = font
    cell.fill = fill

wb.save(outfile)
