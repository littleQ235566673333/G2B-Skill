from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from copy import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_1/regression_gate/before_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_1/regression_gate/before_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active
header = [cell.value for cell in ws[2]]

dob_col = header.index('DATE OF BIRTH') + 1
report_date_coord = 'B1'
age_year_col = header.index('age (year)') + 1
actual_age_col = age_year_col + 1
ws.insert_cols(actual_age_col)
ws.cell(row=2, column=actual_age_col).value = 'Actual Age'
prev_header = ws.cell(row=2, column=actual_age_col - 1)
if prev_header.fill:
    ws.cell(row=2, column=actual_age_col).fill = copy(prev_header.fill)
ws.cell(row=2, column=actual_age_col).font = Font(bold=True)
ws.cell(row=2, column=actual_age_col).alignment = Alignment(horizontal='center', vertical='top')
for r in range(3, 16):
    dob_coord = ws.cell(row=r, column=dob_col).coordinate
    formula = f'=DATEDIF({dob_coord},${report_date_coord},"Y")'
    cell = ws.cell(row=r, column=actual_age_col)
    cell.value = formula
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.number_format = '0'
    # Copy previous column cell fill
    prev_cell = ws.cell(row=r, column=actual_age_col - 1)
    if prev_cell.fill:
        cell.fill = copy(prev_cell.fill)
for r in range(3, 16):
    ws.cell(row=r, column=5).value = f'=XLOOKUP(O{r},H3:H15,I3:I15,"Not found")'
wb.save(output_path)
