from openpyxl import load_workbook
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/group_45707/r0/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/group_45707/r0/evolve_45707/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def str_to_date(x):
    if isinstance(x, datetime):
        return x
    if isinstance(x, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(x, fmt)
            except Exception:
                continue
    if isinstance(x, float) or isinstance(x, int):
        try:
            # Excel number to date
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(x))
        except Exception:
            return None
    return None

rows = list(ws.iter_rows(min_row=2, max_row=69, min_col=1, max_col=3, values_only=True))
all_A = [str_to_date(row[0]) for row in rows]
all_C = [row[2] for row in rows]

for i in range(68):
    cur_date = all_A[i]
    next_date = all_A[i+1] if i+1 < len(all_A) else None
    if next_date and next_date.day == 1:
        # Identify all indices in column A with same month and year as next_date
        target_month = next_date.month
        target_year = next_date.year
        count_1 = 0
        for a, c in zip(all_A, all_C):
            if a and a.month == target_month and a.year == target_year and c == 1:
                count_1 += 1
        ws.cell(row=i+2, column=4, value=count_1)
    else:
        ws.cell(row=i+2, column=4, value=None)

# Last row, i=68
ws.cell(row=69, column=4, value=None)

wb.save(output_path)
