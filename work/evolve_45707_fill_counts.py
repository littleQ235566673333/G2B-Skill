import openpyxl
import datetime

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_5/task_45707/r0/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_5/task_45707/r0/evolve_45707/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# There's always headers in columns 1 (A), 3 (C), and 5 (E), as specified
start_row = 2
end_row = 69
n_rows = end_row - start_row + 1

def parse_date(d):
    if isinstance(d, (datetime.datetime, datetime.date)):
        return d
    if isinstance(d, str):
        try:
            # Try ISO (yyyy-mm-dd)
            return datetime.datetime.strptime(d[:10], '%Y-%m-%d')
        except Exception:
            return None
    return None

# Cache dates (col1) and values in col3
colA_dates = [parse_date(ws.cell(row=r, column=1).value) for r in range(start_row, end_row+1)]
colC_vals = [ws.cell(row=r, column=3).value for r in range(start_row, end_row+1)]

for idx in range(n_rows):
    this_dt = colA_dates[idx]
    if not this_dt:
        ws.cell(row=start_row+idx, column=4).value = None
        continue
    next_day = this_dt + datetime.timedelta(days=1)
    # Only fill if next_day is first of a month
    if next_day.day == 1:
        # Count 1s in colC for all dates in this year/month
        count = 0
        for _dt, _v in zip(colA_dates, colC_vals):
            if _dt and _dt.year == next_day.year and _dt.month == next_day.month and _v == 1:
                count += 1
        ws.cell(row=start_row+idx, column=4).value = count
    else:
        ws.cell(row=start_row+idx, column=4).value = None

# Retain column headers
# D header: leave as-is, add something meaningful if empty
if not ws.cell(row=1, column=4).value:
    ws.cell(row=1, column=4).value = 'Check/Count 1s'
# All other headers left as-is

wb.save(output_path)
