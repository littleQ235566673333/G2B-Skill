from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_333-29_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_333-29_tc1/output.xlsx'

wb = load_workbook(input_path)
ws_a = wb['A']
ws_b = wb['B']

# 1. Find the first 'Yes' or 'NA' in column P of 'A' (starting from row 2)
first_target_row = None
for row in range(2, ws_a.max_row+1):
    val = ws_a.cell(row=row, column=16).value # column P is 16
    if val is not None and str(val).strip() in ('Yes', 'NA'):
        first_target_row = row
        break

fallback_date = ws_a['F3'].value # use F3 if no match
found_date = None
found_any = first_target_row is not None
if found_any:
    # 2. Search above this in col L for value 100
    for row in range(first_target_row-1, 0, -1):
        val_l = ws_a.cell(row=row, column=12).value # column L is 12
        if val_l == 100:
            found_date = ws_a.cell(row=row, column=6).value # column F is 6
            break
    if found_date is None:
        found_date = fallback_date
else:
    found_date = fallback_date

# 3. ws_a['F1'] value
text_a_f1 = ws_a['F1'].value
found_match = False
for row in range(2, ws_b.max_row+1):
    col_c = ws_b.cell(row=row, column=3).value # column C is 3
    if col_c == text_a_f1:
        ws_b.cell(row=row, column=6).value = found_date # column F is 6
        found_match = True

# 4. B!F4 answer
ws_b['F4'] = found_date if found_match else fallback_date

wb.save(output_path)
