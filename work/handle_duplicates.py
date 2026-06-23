import pandas as pd
from openpyxl import load_workbook

input_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/91-34/input.xlsx'
output_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/91-34/output.xlsx'
sheetname = 'SwiftMD'

# Skip the first row of NaNs
header_row = 1  # Second row: actual headers
df = pd.read_excel(input_path, sheet_name=sheetname, engine='openpyxl', header=header_row)

cols = df.columns.tolist()
needed = ['Last Name', 'First Name', 'Date Of Birth', 'Duplicate?', 'Relationship']
for c in needed:
    if c not in cols:
        raise Exception(f'Missing column: {c}')

mask_dupe = (df['Duplicate?'].astype(str).str.lower() == 'yes')
to_delete_idx = set()
grouped = df[mask_dupe].groupby(['Last Name', 'First Name', 'Date Of Birth'])
for _, group in grouped:
    group_df = df.loc[group.index]
    relationships = group_df['Relationship'].astype(str).str.lower().unique()
    if len(group_df) > 1 and all(rel != 'employee' for rel in relationships):
        # more than 1 dupe, no Employee, only remove one
        to_delete_idx.add(group_df.index[0])

df_filtered = df.drop(to_delete_idx, axis=0).reset_index(drop=True)

# Now write updated data back
wb = load_workbook(input_path)
ws = wb[sheetname]

# Write over rows 2:42 (B:O) with new data
start_row = 2
start_col = 2  # 'B'
end_col = 15   # 'O'
rows_to_write = min(41, len(df_filtered))

for i in range(rows_to_write):
    row_values = df_filtered.iloc[i, start_col-2:end_col].tolist()
    for col in range(start_col, end_col+1):
        ws.cell(row=start_row + i, column=col, value=row_values[col - start_col])

wb.save(output_path)
