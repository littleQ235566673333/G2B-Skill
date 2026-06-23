import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_1/group_387-16/r1/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_1/group_387-16/r1/evolve_387-16/output.xlsx'

# Load data with pandas for easier handling
wb = load_workbook(input_path)
ws = wb['Sheet1']

# Find max rows/data range
all_rows = list(ws.values)
headers = all_rows[0]
data_rows = all_rows[1:]
df = pd.DataFrame(data_rows, columns=headers)

# Identify columns
value_col = 'Value' if 'Value' in df.columns else df.columns[1]
binaries_col = 'Binaries'
result_values_col = 'Result values'
target_value_col = 'target value' if 'target value' in df.columns else None

# Prepare result match list
result_values = df[result_values_col].dropna().tolist() if result_values_col in df.columns else []

# Remove only *one* instance in 'Value' (and matching 'Binaries') per 'Result values', prioritize top-down row order
removed_indexes = set()
for val in result_values:
    matches = df.index[df[value_col] == val].difference(removed_indexes)
    if not matches.empty:
        idx = matches[0]
        removed_indexes.add(idx)

# Mask rows to keep
mask = ~df.index.isin(removed_indexes)
new_df = df[mask].copy().reset_index(drop=True)

# After removals, drop empty/shift up for 'Value' and 'Binaries'
for col in [value_col, binaries_col]:
    if col in new_df:
        non_na = new_df[col].dropna()
        non_na = non_na[non_na != ''].reset_index(drop=True)
        filled = pd.Series([non_na[i] if i < len(non_na) else '' for i in range(len(new_df))])
        new_df[col] = filled

# Calculate solver result (sum of valid numbers in value_col)
def try_float(x):
    try:
        return float(x)
    except:
        return 0
val_sum = new_df[value_col].apply(try_float).sum()
solver_result = val_sum

# Target value, if present
if target_value_col is not None and target_value_col in df.columns:
    target_value = df.loc[0, target_value_col]
    try:
        target_value = float(target_value)
    except:
        target_value = None
else:
    target_value = None

difference = (solver_result - target_value) if target_value is not None else None

# Write results to 'Sheet1'!A2:D18
for i in range(17):
    for j in range(4):
        cell = f"{get_column_letter(j+1)}{i+2}"
        if i < len(new_df) and j < len(new_df.columns):
            ws[cell] = new_df.iloc[i, j]
        else:
            ws[cell] = ''
# Output solver result and difference nearby, in E2/F2, E3/F3
ws['E2'] = 'Solver result'
ws['F2'] = solver_result
if difference is not None:
    ws['E3'] = 'Difference'
    ws['F3'] = difference

wb.save(output_path)
