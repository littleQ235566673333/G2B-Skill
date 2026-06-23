from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

def patternfill_from(src_fill):
    # Openpyxl returns colors as either theme, indexed, or rgb; handle only rgb
    if src_fill and src_fill.fill_type and src_fill.start_color and src_fill.start_color.type == 'rgb':
        start_rgb = src_fill.start_color.rgb
    else:
        start_rgb = None
    if src_fill and src_fill.fill_type and src_fill.end_color and src_fill.end_color.type == 'rgb':
        end_rgb = src_fill.end_color.rgb
    else:
        end_rgb = None
    return PatternFill(
        fill_type=src_fill.fill_type if src_fill else None,
        start_color=start_rgb,
        end_color=end_rgb
    ) if (src_fill and src_fill.fill_type) else PatternFill()

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_4/regression_gate/before_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_4/regression_gate/before_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# CATEGORY lookup formula in E3:E15
for i in range(3, 16):
    ws.cell(row=i, column=5).value = '=INDEX($I$3:$I$15, MATCH(O{} , $H$3:$H$15, 0))'.format(i)

# "Actual Age" col P (16), header and data
ws.cell(row=2, column=16).value = "Actual Age"

base_data_fill = ws.cell(row=3, column=15).fill
base_header_fill = ws.cell(row=2, column=15).fill

for i in range(3, 16):
    ws.cell(row=i, column=16).value = '=DATEDIF(C{0},$B$1,"Y")'.format(i)
    ws.cell(row=i, column=16).number_format = '0'
    ws.cell(row=i, column=16).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=i, column=16).font = Font(bold=True)
    ws.cell(row=i, column=16).fill = patternfill_from(base_data_fill)

ws.cell(row=2, column=16).alignment = Alignment(horizontal='center', vertical='top')
ws.cell(row=2, column=16).font = Font(bold=True)
ws.cell(row=2, column=16).fill = patternfill_from(base_header_fill)

wb.save(output_path)
