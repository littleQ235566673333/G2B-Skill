import openpyxl

def assign_animal_rankings(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    labels = [ws.cell(row=1, column=c).value for c in range(1,9)]
    numbers = [ws.cell(row=2, column=c).value for c in range(1,9)]

    output = ['']*8
    for animal in ['cat','dog']:
        pairs = [(i, num) for i, (lab, num) in enumerate(zip(labels, numbers)) if lab == animal]
        if pairs:
            sorted_pairs = sorted(pairs, key=lambda x: x[1], reverse=True)
            if len(sorted_pairs) > 0:
                output[sorted_pairs[0][0]] = f'{animal}1'
            if len(sorted_pairs) > 1:
                output[sorted_pairs[1][0]] = f'{animal}2'
    for i, val in enumerate(output):
        ws.cell(row=4, column=i+1).value = val
    wb.save(output_path)

assign_animal_rankings(
    'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_51289_tc1/input.xlsx',
    'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_51289_tc1/output.xlsx')
