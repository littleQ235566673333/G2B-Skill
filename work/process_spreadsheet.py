import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_6/group_58484/r2/evolve_58484/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_6/group_58484/r2/evolve_58484/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Border style for the task
tb = Side(border_style='thin', color='000000')
border = Border(left=tb, right=tb, top=tb, bottom=tb)

# Header (row 4): bold, bordered, wrap_text
for cell in ws[4]:
    cell.font = Font(bold=True)
    cell.border = border
    cell.alignment = Alignment(wrap_text=True)

# For rows 5-26: apply thin border to columns A-F, clear value+border for G
for r in range(5, 27):
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = border
    ws.cell(row=r, column=7).border = None
    ws.cell(row=r, column=7).value = None

# Clear heading and data in column G (row 4-26)
for r in range(4, 27):
    ws.cell(row=r, column=7).value = None
    ws.cell(row=r, column=7).border = None
    
# Function to process transfers
def count_transfers(ws, start_row, end_row):
    # walk all call rows, collect:
    #   transfer number (C)
    #   operator name (D)
    #   staff transferred to (E)
    calls = []
    for r in range(start_row, end_row+1):
        transfer_num = ws.cell(row=r, column=3).value
        operator = ws.cell(row=r, column=4).value
        to_staff = ws.cell(row=r, column=5).value
        calls.append({'row': r, 'transfer_num': transfer_num, 'operator': operator, 'to_staff': to_staff})
    # Identify valid transfers for 5551234 and staff destination
    transfers = [x for x in calls if x['transfer_num'] == 5551234 and x['to_staff'] not in (None, '', '-')]
    if not transfers:
        # No entries, clear col H
        for r in range(start_row, end_row+1):
            ws.cell(row=r, column=8).value = None
        return
    count = len(transfers)
    last_row = transfers[-1]['row']
    # Only set result in col H for the last transfer
    for t in transfers[:-1]:
        ws.cell(row=t['row'], column=8).value = None
    ws.cell(row=last_row, column=8).value = f'Total transfers: {count}'
count_transfers(ws, 5, 26)

wb.save(output_path)
