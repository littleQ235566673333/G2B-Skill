from openpyxl import load_workbook
from datetime import datetime

# File paths
doc_in = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_5/regression_gate/after_pass/core_41589/input.xlsx'
doc_out = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_5/regression_gate/after_pass/core_41589/output.xlsx'

# Load workbook and worksheet
wb = load_workbook(doc_in)
ws = wb['Contact List']

# Get values from H4 and I4
h4 = ws['H4'].value
try:
    i4 = ws['I4'].value.strip().upper() if ws['I4'].value else ''
except AttributeError:
    i4 = str(ws['I4'].value).upper() if ws['I4'].value else ''

# Today's date
now = datetime.now()

# Determine output for J4
result = 'NO ACTION'
if isinstance(h4, datetime) and i4 == 'YES':
    days_since = (now - h4).days
    if days_since <= 30:
        result = 'HOLD'
    elif days_since > 30:
        result = 'TOUCH BASE'

ws['J4'] = result

# Save changes
wb.save(doc_out)
