import openpyxl
import datetime
from openpyxl.styles import Alignment

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_7/group_59902/r1/evolve_59902/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_7/group_59902/r1/evolve_59902/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

rows = range(5, 29)
name_col = 7 # G
date_col = 6 # F
out_col = 3  # C

# Collect per row: name and date
sales = []
for r in rows:
    name = ws.cell(row=r, column=name_col).value
    date = ws.cell(row=r, column=date_col).value
    sales.append({'row': r, 'name': name, 'date': date})

# Calculate days since previous sale for each person
prev_dates = {}
results = {}
for s in sales:
    staff = s['name']
    date = s['date']
    if not staff or not date:
        results[s['row']] = ''
        continue
    # Ensure date is a date/datetime
    if isinstance(date, str):
        try:
            date = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        except Exception:
            results[s['row']] = ''
            continue
    if staff not in prev_dates:
        results[s['row']] = 0
    else:
        diff = (date - prev_dates[staff]).days
        results[s['row']] = diff
    prev_dates[staff] = date

# Accounting format from Excel (string)
accounting_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

# Write output and apply accounting format
for r in rows:
    cell = ws.cell(row=r, column=out_col)
    val = results[r]
    cell.value = val
    cell.alignment = Alignment(horizontal='right')
    cell.number_format = accounting_format

wb.save(output_path)
