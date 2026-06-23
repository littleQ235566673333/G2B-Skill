import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/before_pass/core_51249/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/before_pass/core_51249/output.xlsx'

# Open the workbook and active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Color fill as required
fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

# Row indices for D1, D5, D9
cell_rows = [1,5,9]
for row in cell_rows:
    b1 = ws[f'B{row}'].value
    b2 = ws[f'B{row+1}'].value
    if b1 == 'Description A' and b2 is None:
        result = 'Single A'
    elif b1 == 'Description B' and b2 is None:
        result = 'Single B'
    elif b1 == 'Description A' and b2 == 'Description B':
        result = 'Multiple'
    else:
        result = ''
    ws[f'D{row}'].value = result
    ws[f'D{row}'].fill = fill

wb.save(output_path)
print('Done.')
