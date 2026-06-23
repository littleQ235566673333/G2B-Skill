import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta, date

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_50631_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_50631_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Calibri font
def set_calibri(ws):
    calibri = Font(name='Calibri')
    for row in ws.iter_rows():
        for cell in row:
            cell.font = calibri

# Get holidays from column J (skip header)
holidays = []
for cell in ws['J'][1:]:  # skip header
    val = cell.value
    if isinstance(val, (datetime, date)):
        holidays.append(val)
    elif isinstance(val, float) or isinstance(val, int):
        try:
            holidays.append(openpyxl.utils.datetime.from_excel(val))
        except Exception:
            pass
    elif isinstance(val, str):
        try:
            holidays.append(datetime.strptime(val, '%m/%d/%Y'))
        except Exception:
            pass

def is_holiday(dt):
    for h in holidays:
        if dt.year == h.year and dt.month == h.month and dt.day == h.day:
            return True
    return False

def is_workday(dt):
    return dt.weekday() < 5 and not is_holiday(dt)

# Get start and end dates from B3 and F3
start_date = ws['B3'].value
end_date = ws['F3'].value
if isinstance(start_date, str):
    start_date = datetime.strptime(start_date, '%m/%d/%Y')
if isinstance(end_date, str):
    end_date = datetime.strptime(end_date, '%m/%d/%Y')
if isinstance(start_date, (datetime, date)) and not isinstance(start_date, datetime):
    start_date = datetime.combine(start_date, datetime.min.time())
if isinstance(end_date, (datetime, date)) and not isinstance(end_date, datetime):
    end_date = datetime.combine(end_date, datetime.min.time())

# Generate all workdays between start_date and end_date
workdays = []
dt = start_date
while dt <= end_date:
    if is_workday(dt):
        workdays.append(dt)
    dt += timedelta(days=1)

# Save existing fills for H7:H38
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
original_fills = {}
for row in range(7, 39):
    cell = ws[f'H{row}']
    fill = cell.fill
    is_yellow = False
    # openpyxl cell.fill may use fgColor or start_color
    if hasattr(fill, 'fgColor') and (
        getattr(fill.fgColor, 'rgb', '').upper() in ['FFFF00', 'FFFFFF00'] or
        getattr(fill.fgColor, 'indexed', None) == 6
    ):
        is_yellow = True
    if hasattr(fill, 'start_color') and (
        getattr(fill.start_color, 'rgb', '').upper() in ['FFFF00', 'FFFFFF00'] or
        getattr(fill.start_color, 'indexed', None) == 6
    ):
        is_yellow = True
    original_fills[row] = is_yellow

for i, row in enumerate(range(7, 39)):
    cell = ws[f'H{row}']
    if i < len(workdays):
        cell.value = workdays[i]
        cell.number_format = 'mm/dd/yyyy'
        if original_fills[row]:
            cell.fill = yellow_fill
        else:
            cell.fill = PatternFill(fill_type=None)
    else:
        cell.value = None
        cell.fill = PatternFill(fill_type=None)

set_calibri(ws)
ws.sheet_view.showGridLines = False
wb.save(output_path)
