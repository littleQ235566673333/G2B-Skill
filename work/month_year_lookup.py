import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_6/regression_gate/before_fix/core_39515/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_6/regression_gate/before_fix/core_39515/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 14):
    month = ws[f'A{row}'].value
    year = ws[f'B{row}'].value
    if not (month and year):
        ws[f'O{row}'] = None
        continue
    found_col = None
    for col in range(3, 15):  # C to N is 3..14
        header = ws.cell(row=1, column=col).value
        if header is None:
            continue
        header_lower = str(header).lower()
        month_lower = str(month).lower()
        if (month_lower in header_lower) and (str(year) in header_lower):
            found_col = col
            break
    if found_col:
        ws.cell(row=row, column=15).value = ws.cell(row=row, column=found_col).value
    else:
        ws.cell(row=row, column=15).value = None

wb.save(output_path)
