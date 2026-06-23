from openpyxl import load_workbook

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_2/regression_gate/after_fix/core_50916/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_2/regression_gate/after_fix/core_50916/output.xlsx'
wb = load_workbook(INPUT)
ws = wb.active

# Detect cycle day (A12:A14)
cycle_rows = [12,13,14]
cycle_col = 'A'
output_cols = ['C','D','E','F','G','H']

# The table for days/classes (cycles 1-7 in A2:A8, class names for periods in C2:H8)
cycle_table_start_row = 2
cycle_table_end_row = 8
period_cols = ['C','D','E','F','G','H']  # 6 periods per day

for dst_row in cycle_rows:
    cycle_val = ws[f'{cycle_col}{dst_row}'].value
    # Find which cycle day row
    match_row = None
    for table_row in range(cycle_table_start_row, cycle_table_end_row+1):
        if ws[f'A{table_row}'].value == cycle_val:
            match_row = table_row
            break
    if match_row is not None:
        for i, col in enumerate(period_cols):
            ws[f'{col}{dst_row}'].value = ws[f'{col}{match_row}'].value

wb.save(OUTPUT)
