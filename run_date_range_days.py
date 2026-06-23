from openpyxl import load_workbook
import re

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun1/eval_43589_tc1/input.xlsx')
ws = wb.active

cell_value = ws['A2'].value
days = None
if isinstance(cell_value, str):
    match = re.match(r'(\d+)\s*to\s*(\d+)', cell_value)
    if match:
        start, end = map(int, match.groups())
        days = end - start + 1

ws['B2'] = days if days is not None else 'ERROR'
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun1/eval_43589_tc1/output.xlsx')
