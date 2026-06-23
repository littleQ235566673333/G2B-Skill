import openpyxl

# File paths
input_fp = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_33157_tc1/input.xlsx'
output_fp = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_33157_tc1/output.xlsx'

# Open the workbook and sheet
wb = openpyxl.load_workbook(input_fp)
ws = wb['Sheet1']

# Set columns and their headers
cols = ['B', 'D', 'F', 'H']
headers = [ws[f'{col}1'].value for col in cols]  # assumes headers in row 1

# Process rows 2 to 6
date_col = 'J'
output_col = 'K'
for row in range(2, 7):
    person_date = ws[f'{date_col}{row}'].value
    found_header = None
    for col, header in zip(cols, headers):
        act_date = ws[f'{col}{row}'].value
        # Compare dates directly (if both are not None)
        if act_date == person_date and person_date is not None:
            found_header = header
            break
    ws[f'{output_col}{row}'] = found_header

# Save output
wb.save(output_fp)
