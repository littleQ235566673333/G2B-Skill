import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_2/task_516-46/r1/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_2/task_516-46/r1/evolve_516-46/output.xlsx'

# Read the input (auto-detecting header/columns) and picking the Arabic sheet name
all_sheets = pd.read_excel(input_path, sheet_name=None)
sheet_name = [s for s in all_sheets if 'ورقة' in s][0]
df = all_sheets[sheet_name]

# Keep only columns A-E (from A2, so skip header probably, but with/without header both work)
df = df.iloc[:, :5]
df.columns = ['A', 'B', 'C', 'D', 'E']

# Try to infer columns: Brand, Qty, Date
def infer_columns(df):
    date_col = None
    qty_col = None
    brand_col = None
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            date_col = col
        elif pd.api.types.is_numeric_dtype(df[col]) and qty_col is None:
            qty_col = col
        elif df[col].dtype == object and brand_col is None:
            brand_col = col
    return brand_col, qty_col, date_col

brand_col, qty_col, date_col = infer_columns(df)
if not (brand_col and qty_col and date_col):
    # fallback to default positions (based on usual patterns, user must confirm)
    brand_col, qty_col, date_col = 'B', 'C', 'E'

df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
df = df.dropna(subset=[brand_col, qty_col, date_col])

# For each brand, get the last date
last_dates = df.groupby(brand_col)[date_col].transform('max')
df_last = df[df[date_col] == last_dates]

# For each brand on its last date, sum duplicate entries by qty
result = df_last.groupby([brand_col, date_col])[qty_col].sum().reset_index()
mod_counts = df_last.groupby([brand_col, date_col]).size().reset_index(name='Modifications')
result_df = pd.merge(result, mod_counts, on=[brand_col, date_col])

# Pad to exactly 3 rows for output range H2:L4
while len(result_df) < 3:
    result_df = pd.concat([result_df, pd.DataFrame([[None]*len(result_df.columns)], columns=result_df.columns)], ignore_index=True)

result_df = result_df[[brand_col, date_col, qty_col, 'Modifications']]

wb = load_workbook(input_path)
ws = wb[sheet_name]

start_row = 2
start_col = 8  # column H
for i, row in enumerate(result_df.values):
    for j, val in enumerate(row):
        ws.cell(row=start_row + i, column=start_col + j).value = val

wb.save(output_path)
