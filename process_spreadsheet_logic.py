import openpyxl
from collections import defaultdict

# File paths
action_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/before_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/before_pass/core_55421/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(action_path)
ws = wb.active

# Read info into groups by number in Col A
records = defaultdict(list)
for row in ws.iter_rows(min_row=2, max_row=20, min_col=1, max_col=6, values_only=True):
    a, _, _, d, e, _ = row
    records[a].append({'row': row, 'status': d, 'date': e})

# Create output map for each unique Col A value based on priorities
action_map = {}
for num, entries in records.items():
    statuses = {entry['status'] for entry in entries}
    ns_entries = [entry for entry in entries if entry['status'] == 'NO SHOW']

    # Priority rules
    # 1. NO SHOW with empty date -> CALL PT
    call_pt = any(entry['status'] == 'NO SHOW' and not entry['date'] for entry in entries)
    if call_pt:
        action_map[num] = 'CALL PT'
        continue
    # 2. NO SHOW with date -> NO ACTION NEEDED
    no_action = any(entry['status'] == 'NO SHOW' and entry['date'] for entry in entries)
    if no_action:
        action_map[num] = 'NO ACTION NEEDED'
        continue
    # 3. Both NO SHOW and SCH -> NS/SCHED
    if 'NO SHOW' in statuses and 'SCH' in statuses:
        action_map[num] = 'NS/SCHED'
        continue
    # 4. Only SCH -> FUTURE
    if statuses == {'SCH'}:
        action_map[num] = 'FUTURE'
        continue
    # fallback
    action_map[num] = ''

# Write output to Col F for rows 2-20
for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=20, min_col=1, max_col=1), start=2):
    a_val = row[0].value
    result = action_map.get(a_val, '')
    ws[f'F{idx}'] = result

wb.save(output_path)
