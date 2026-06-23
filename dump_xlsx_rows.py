from openpyxl import load_workbook

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_8/regression_gate/after_fix/core_263-1/input.xlsx')
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
