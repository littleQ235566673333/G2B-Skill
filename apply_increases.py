from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_1/group_44017/r1/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_1/group_44017/r1/evolve_44017/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Data']
date_row = 9
date_cols = list(range(30, 42))
for row in range(14, 43):
    base_ref = f'$W{row}'
    eff_ref = f'$L{row}'
    freq_ref = f'$J{row}'
    pct_refs = [f'$M{row}', f'$N{row}', f'$O{row}', f'$P{row}']
    for i, col in enumerate(date_cols):
        cell = ws.cell(row=row, column=col)
        month_date_ref = f'{get_column_letter(col)}$9'
        formula = (
            f'=IF({month_date_ref}<{eff_ref},"",'
            f'{base_ref}*PRODUCT('
            f'1+IF(AND(ISNUMBER({pct_refs[0]}),{month_date_ref}>={eff_ref}+0*{freq_ref}),{pct_refs[0]},0),'
            f'1+IF(AND(ISNUMBER({pct_refs[1]}),{month_date_ref}>={eff_ref}+1*{freq_ref}),{pct_refs[1]},0),'
            f'1+IF(AND(ISNUMBER({pct_refs[2]}),{month_date_ref}>={eff_ref}+2*{freq_ref}),{pct_refs[2]},0),'
            f'1+IF(AND(ISNUMBER({pct_refs[3]}),{month_date_ref}>={eff_ref}+3*{freq_ref}),{pct_refs[3]},0)'
            '))'
            ')'
        )
        cell.value = formula
wb.save(output_path)
print('Done')
