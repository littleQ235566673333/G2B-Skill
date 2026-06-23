import openpyxl
import pandas as pd

# Load workbook and sheets
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_55965_tc1/input.xlsx')
sheet1 = wb['Foglio1'] if 'Foglio1' in wb.sheetnames else wb[wb.sheetnames[0]]
# Detect data sheet (usually called 'data', else second sheet)
data_ws = None
for name in wb.sheetnames:
    if name.lower() == 'data' or name.lower().startswith('data'):
        data_ws = wb[name]
if not data_ws:
    data_ws = wb[wb.sheetnames[1]]

# Read data sheet to DataFrame
data = list(data_ws.values)
data_header = data[0]
data_rows = list(data[1:])
data_df = pd.DataFrame(data_rows, columns=data_header)

# Read IDs from Foglio1 E2:E18
id_column_sheet1 = 'E'
data_id_col_data_sheet = 'IG'  # Matching IG with E
output_start_col = 7  # G
output_end_col = 16   # P (inclusive)
output_cols_count = output_end_col - output_start_col + 1
row_start = 2
row_end = 18

for idx, excel_row in enumerate(range(row_start, row_end + 1)):
    search_id = sheet1[f'{id_column_sheet1}{excel_row}'].value
    # Clear output content first
    for col in range(output_start_col, output_end_col + 1):
        sheet1.cell(row=excel_row, column=col).value = None
    if not search_id:
        continue
    matches = data_df[data_df[data_id_col_data_sheet] == search_id]
    if not matches.empty:
        last10 = matches.tail(10)
        # Write LAST ROW of last 10 (most recent)
        fill_row = last10.values.tolist()[-1] if len(last10) else [None] * len(data_header)
        for offset, value in enumerate(fill_row[:output_cols_count]):
            sheet1.cell(row=excel_row, column=output_start_col+offset).value = value

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_55965_tc1/output.xlsx')
