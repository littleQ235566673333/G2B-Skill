import pandas as pd
from openpyxl import load_workbook

input_xlsx = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/regression_gate/after_pass/core_3413/input.xlsx'
output_xlsx = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/regression_gate/after_pass/core_3413/output.xlsx'

# Read with first row as header
actual_df = pd.read_excel(input_xlsx, engine='openpyxl', header=1)
wb = load_workbook(input_xlsx)
ws = wb.active

# For G3:G6, look in rows 2–5 (0-based) in pandas
for df_idx, excel_row in enumerate(range(3, 7)):
    # Reference cells for matching
    dept_val = ws[f'E{excel_row}'].value
    ru_val = ws[f'F{excel_row}'].value
    # Check with both dept+ru
    filt = (actual_df['Dept'] == dept_val) & (actual_df['RU'] == ru_val)
    if actual_df[filt].shape[0] > 0:
        s = actual_df.loc[filt, 'Total'].sum()
    else:
        # Sum all for department regardless of RU
        s = actual_df.loc[actual_df['Dept'] == dept_val, 'Total'].sum()
    ws[f'G{excel_row}'] = s
wb.save(output_xlsx)
