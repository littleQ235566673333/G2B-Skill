import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_36097_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_36097_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(3, 7):
    orig_cost = ws[f'B{row}'].value
    sale_proceeds = ws[f'C{row}'].value
    itv = ws[f'D{row}'].value
    profit = ws[f'E{row}'].value
    loss = ws[f'F{row}'].value

    # Determine value per instructions
    if sale_proceeds < orig_cost:
        # Sold at a loss: recoup ITV adjusted by the loss (add loss to ITV)
        val = itv + abs(loss)
    elif profit < orig_cost:
        # Profit is less than original cost: recoup the full profit
        val = profit
    else:
        # Otherwise: recoup the cost basis reduced by ITV
        val = orig_cost - itv

    # Preserve number format from column I
    ws[f'H{row}'].value = val
    ws[f'H{row}'].number_format = ws[f'I{row}'].number_format

wb.save(output_path)
