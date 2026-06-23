import openpyxl

# Input and output paths
template_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/group_54474/r2/evolve_54474/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/group_54474/r2/evolve_54474/output.xlsx'

wb = openpyxl.load_workbook(template_path)
whp_sheet = wb['WHP']
whp_data_sheet = wb['WHP DATA']

# Get site names from B7:B8 in WHP sheet
sites = [whp_sheet[f'B{row}'].value for row in range(7, 9)]

# Collect all site data blocks
data = {}
collect = False
for row in whp_data_sheet.iter_rows(values_only=True):
    # Detect start of a group
    if row[2] == 'EEG' and row[3] == 'Dis' and row[4] == 'LTU':
        collect = True
        continue
    if collect and row[1] and isinstance(row[1], str) and all(isinstance(v, (int, float, type(None))) for v in row[2:5]):
        site = row[1]
        eeg, dis, ltu = row[2:5]
        data[site] = [eeg, dis, ltu]
    # Detect end/group separator
    if collect and row[1] == 'Total':
        collect = False

# Fill E7:G8 in WHP tab
for i, site in enumerate(sites):
    vals = data.get(site, [None, None, None])
    for j, v in enumerate(vals):
        whp_sheet.cell(row=7+i, column=5+j, value=v)

wb.save(output_path)
