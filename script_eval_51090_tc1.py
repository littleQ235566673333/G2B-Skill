import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_51090_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_51090_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path, data_only=True)
ws_daily = wb['Daily Numbers']
daily_data = list(ws_daily.values)
daily_cols = list(daily_data[1])
daily_df = pd.DataFrame(daily_data[2:], columns=daily_cols)
ws_ir = wb['Inbound Receipts']
ir_data = list(ws_ir.values)
ir_cols = list(next(iter(ir_data)))
ir_df = pd.DataFrame(ir_data[1:], columns=ir_cols)
ws_errors = wb['Errors']
errors_data = list(ws_errors.values)
errors_cols = list(next(iter(errors_data)))
errors_df = pd.DataFrame(errors_data[1:], columns=errors_cols)

warehouse_code = '27'
users = ['CHROGIL1', 'CHDSPOLJ', 'CHSJEFFE', 'CHBTHOMA']
error_codes = ['II', 'IR', 'IT', 'OV', 'PI']
result = []

for idx, row in daily_df.iterrows():
    if str(row['WHS']).strip() == warehouse_code:
        cartons_received = row['Inbound Receipts'] if 'Inbound Receipts' in daily_df.columns else (row['M'] if 'M' in daily_df.columns else None)
        total_errors = 0
        for err_name in error_codes:
            filtered = errors_df[(errors_df['Location'].astype(str).str.strip() == warehouse_code) & (errors_df['Document Type'] == err_name) & (errors_df['User ID'].isin(users))]
            if err_name == 'IR':
                filtered = filtered[pd.to_numeric(filtered['Transaction Quantity'], errors='coerce') > 0]
            err_sum = pd.to_numeric(filtered['Transaction Quantity'], errors='coerce').sum()
            total_errors += err_sum
        if cartons_received is not None:
            try:
                cartons_received = float(cartons_received)
            except:
                cartons_received = 0
            result.append(cartons_received - total_errors)
        else:
            result.append(None)
    else:
        result.append(None)

wb_out = openpyxl.load_workbook(input_path)
ws_out = wb_out['Daily Numbers']
for i, val in enumerate(result):
    ws_out.cell(row=3+i, column=17).value = val  # Q column is 17
wb_out.save(output_path)
print(result[:5])
