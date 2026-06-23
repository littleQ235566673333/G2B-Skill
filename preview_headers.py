import openpyxl
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_7/regression_gate/before_pass/core_9726/input.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active
max_rows = ws.max_row
max_cols = ws.max_column

with open('row_preview.txt', 'w', encoding='utf8') as f:
    for r in range(1, 21):
        vals = [str(ws.cell(row=r, column=c).value).strip() if ws.cell(row=r, column=c).value else '' for c in range(1, max_cols + 1)]
        f.write(f"Row {r}: {vals}\n")
