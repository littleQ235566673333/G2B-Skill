import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_4/group_146-49/r2/evolve_146-49/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_4/group_146-49/r2/evolve_146-49/output.xlsx'
vowels = set('AEIOU')

def is_pair(w1, w2, vowel):
    if len(w1) != len(w2):
        return False
    for c1, c2 in zip(w1, w2):
        if c1 == 'S':
            if c2 != vowel:
                return False
        else:
            if c1 != c2:
                return False
    return True

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Read all words (scan all columns for words in sheet)
data_words = []
for row in ws.iter_rows(values_only=True):
    for value in row:
        if isinstance(value, str) and value.strip():
            data_words.append(value.strip().upper())
# Remove duplicates
data_words = sorted(set(data_words))
results = set()
for w1 in data_words:
    if 'S' not in w1:
        continue
    for v in vowels:
        w2 = w1.replace('S', v)
        if w2 == w1:
            continue
        if w2 in data_words and is_pair(w1, w2, v):
            # Enforce lexicographical order and avoid reverses
            pair = tuple(sorted((w1, w2), key=lambda x: x.upper()))
            results.add(pair)
# Sort pairs (case-insensitive)
sorted_pairs = sorted(results, key=lambda p: (p[0].upper(), p[1].upper()))
# Write to G and H, starting at G1 (so output in Sheet1!G1:H...)
start_row = 1
for idx, (a, b) in enumerate(sorted_pairs):
    ws.cell(row=start_row+idx, column=7, value=a)
    ws.cell(row=start_row+idx, column=8, value=b)
wb.save(output_path)
