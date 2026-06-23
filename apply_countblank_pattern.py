from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/after_pass/core_41969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/after_pass/core_41969/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# We want formulas in A6, B6, C6 that act on A3:C3, D3:F3, G3:I3 respectively
col_starts = [1, 4, 7]
for i, col_start in enumerate(col_starts):
    col_letter_start = chr(64 + col_start)
    col_letter_end = chr(64 + col_start + 2)
    formula = f"=COUNTBLANK({col_letter_start}3:{col_letter_end}3)"
    ws.cell(row=6, column=i+1).value = formula

wb.save(output_path)
print('Formulas written to A6:C6')
