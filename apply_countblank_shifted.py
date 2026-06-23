import openpyxl
from openpyxl.utils import get_column_letter

def col_letter(num):
    return get_column_letter(num)

# Input/output paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_7/regression_gate/after_pass/core_41969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_7/regression_gate/after_pass/core_41969/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# For A6, want =COUNTBLANK(A3:C3)
# For B6, want =COUNTBLANK(D3:F3)
# For C6, want =COUNTBLANK(G3:I3)
row_formula = 6
row_range = 3
start_row = 3

for i, dest_col in enumerate(range(1, 4)):
    start_col = 1 + i * row_range
    end_col = start_col + row_range - 1
    start_letter = col_letter(start_col)
    end_letter = col_letter(end_col)
    formula = f"=COUNTBLANK({start_letter}{start_row}:{end_letter}{start_row})"
    ws.cell(row=row_formula, column=dest_col).value = formula

wb.save(output_path)
