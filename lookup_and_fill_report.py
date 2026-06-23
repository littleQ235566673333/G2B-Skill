import openpyxl
import pandas as pd

# Load workbook and sheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/regression_gate/after_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/regression_gate/after_pass/core_18935/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Helper to convert worksheet table to DataFrame
def ws_to_df(ws, min_row, max_row, min_col, max_col):
    data = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        data.append(row)
    df = pd.DataFrame(data[1:], columns=data[0])
    return df

# Step 1: Identify Data Table range (assume from the top, boundaries by inspection)
# We'll read a large area and let pandas drop all-NaN rows/cols
raw = pd.DataFrame(ws_to_df(ws, 1, 20, 1, 10))
raw = raw.dropna(how="all").dropna(axis=1, how="all")

# Heuristically find the Data Table starting row/header
header_row_idx = None
for idx, row in raw.iterrows():
    if 'Work Criteria' in row.values and 'Category Type' in row.values and 'Material' in row.values:
        header_row_idx = idx
        break
assert header_row_idx is not None, "Header row not found."
data_table = raw.iloc[header_row_idx:]  # keep header and below

# Clean and reset header
real_headers = list(data_table.iloc[0])
data_table = data_table[1:]
data_table.columns = real_headers

# Step 2: Identify report rows (by placement: D17:D22)
report_rows = list(range(17, 23))
results = []
for row in report_rows:
    work_criteria = ws[f'B{row}'].value
    category_type = ws[f'C{row}'].value
    material = ws[f'E{row}'].value
    # Search for match in Data Table (case-insensitive)
    match = data_table[ (data_table['Work Criteria'].astype(str).str.lower() == str(work_criteria).lower()) &\
                      (data_table['Category Type'].astype(str).str.lower() == str(category_type).lower()) &\
                      (data_table['Material'].astype(str).str.lower() == str(material).lower()) ]
    category_value = None
    if not match.empty:
        category_value = match.iloc[0]['Category Value']
    ws[f'D{row}'] = category_value

# Save result
wb.save(output_path)
