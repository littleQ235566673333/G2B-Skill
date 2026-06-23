import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_1/group_91-34/r2/evolve_91-34/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_1/group_91-34/r2/evolve_91-34/output.xlsx'
sheet_name = 'SwiftMD'

# Read B2:O42, row 2 as header
cols = list(range(1, 15))  # B:O (1-based index)
df = pd.read_excel(input_path, sheet_name=sheet_name, usecols=cols, header=1)

key_columns = ['Last Name', 'First Name', 'Date Of Birth']
dup_col = 'Duplicate?'
rel_col = 'Relationship'

# Only consider duplicates marked 'Yes'
filtered = df[df[dup_col] == 'Yes']
grouped = filtered.groupby(key_columns)
to_drop = set()
for keys, group in grouped:
    rels = set(group[rel_col].astype(str))
    if 'Employee' not in rels and len(group) > 1:
        # Mark one index for deletion (not all, just one entry)
        to_drop.add(group.index[0])

df_new = df.drop(list(to_drop)).reset_index(drop=True)

# Write modified table to B2:O42
wb = load_workbook(input_path)
ws = wb[sheet_name]
# Clear B2:O42 first
for row in ws.iter_rows(min_row=2, max_row=42, min_col=2, max_col=15):
    for cell in row:
        cell.value = None

# Write new data (max 41 rows, 14 columns)
values = df_new.values
for i in range(min(len(values), 41)):
    for j in range(min(values.shape[1], 14)):
        ws.cell(row=2+i, column=2+j, value=values[i, j])

wb.save(output_path)
