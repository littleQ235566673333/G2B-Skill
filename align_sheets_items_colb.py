import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_6/regression_gate/after_pass/core_302-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_6/regression_gate/after_pass/core_302-1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_items = wb['ITEMS']
ws1 = wb['SHEET1']
ws2 = wb['SHEET2']

# Read the target column B values from ITEMS into a dict (row idx: value)
items_b = {}
for row in range(2, ws_items.max_row + 1):
    cell_value = ws_items.cell(row=row, column=2).value
    items_b[row] = cell_value

# Helper function to replace column B in sheet based on ITEMS column B
def align_sheet_col_b(ws, max_row):
    for row in range(2, max_row + 1):
        new_value = items_b.get(row, None)
        ws.cell(row=row, column=2).value = new_value

# Apply to both sheets
# Sheet1: rows 2-8
align_sheet_col_b(ws1, 8)
# Sheet2: rows 2-7
align_sheet_col_b(ws2, 7)

wb.save(output_path)
