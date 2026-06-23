import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_8/regression_gate/after_pass/core_56599/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_8/regression_gate/after_pass/core_56599/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
slip_search_cell = 'M5'
search_slip = ws[slip_search_cell].value
found = False

if search_slip is None:
    ws['B2'] = 'not found - please re-enter'
else:
    try:
        slip_int = int(search_slip)
        # Search for slip number
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                if cell.value is None:
                    continue
                # check for 7 digit numbers as string or int
                if (isinstance(cell.value, int) and len(str(cell.value)) == 7) or \
                   (isinstance(cell.value, str) and cell.value.isdigit() and len(cell.value) == 7):
                    value = int(cell.value)
                    if value == slip_int:
                        # Found, calculate book number (1-based)
                        book_number = (value - 1) // 10 + 1
                        ws['B2'] = book_number
                        found = True
                        break
            if found:
                break
        if not found:
            ws['B2'] = 'not found - please re-enter'
    except Exception:
        ws['B2'] = 'not found - please re-enter'

wb.save(output_path)
