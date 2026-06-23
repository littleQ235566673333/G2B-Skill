import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

# Load the workbook and select the active sheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_45372_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_45372_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# The fill color requested
fill = PatternFill(fill_type='solid', fgColor='FFC000')

for row in range(2, 16):  # E2 to E15
    time_cell = ws[f'A{row}']
    value_b = ws[f'B{row}'].value
    value_c = ws[f'C{row}'].value

    # Ensure time is in datetime/time or string hh:mm format
    time_value = time_cell.value
    if isinstance(time_value, (datetime,)):
        current_time = time_value.time()
    elif isinstance(time_value, str):
        try:
            current_time = datetime.strptime(time_value.strip(), '%H:%M').time()
        except:
            ws[f'E{row}'].value = ''
            continue
    else:
        ws[f'E{row}'].value = ''
        continue
    
    # Decision based on the time
    if current_time < datetime.strptime('09:45', '%H:%M').time():
        new_value = value_b
    else:
        new_value = value_c
    # If value is 0 (or zero float or int), blank
    if new_value == 0 or new_value == 0.0:
        ws[f'E{row}'].value = ''
    else:
        ws[f'E{row}'].value = new_value
    # Set fill
    ws[f'E{row}'].fill = fill

# Save the workbook
wb.save(output_path)
