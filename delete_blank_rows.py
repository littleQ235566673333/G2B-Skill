from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_5/regression_gate/before_pass/core_160-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_5/regression_gate/before_pass/core_160-6/output.xlsx'

wb = load_workbook(input_path)
ws = wb['SH']

max_col = ws.max_column
rows_to_delete = []

for row in range(1, ws.max_row+1):
    is_blank = True
    for col in range(1, max_col+1):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value is not None and str(cell_value).strip() != "":
            is_blank = False
            break
    if is_blank:
        rows_to_delete.append(row)

for row in reversed(rows_to_delete):
    ws.delete_rows(row)

wb.save(output_path)
