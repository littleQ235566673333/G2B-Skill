from openpyxl import load_workbook

def colnum_string(n):
    string = ''
    while n > 0:
        n, r = divmod(n-1, 26)
        string = chr(r+65) + string
    return string

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/regression_gate/after_pass/core_41969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/regression_gate/after_pass/core_41969/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

first_cols = [1, 4, 7]  # A, D, G (as 1-based indices)
for i, first_col in enumerate(first_cols):
    col_letter_start = colnum_string(first_col)
    col_letter_end = colnum_string(first_col + 2)
    formula = f'=COUNTBLANK({col_letter_start}3:{col_letter_end}3)'
    ws.cell(row=6, column=i+1).value = formula

wb.save(output_path)
