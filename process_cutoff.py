import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_35739_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_35739_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

for row in range(2, 101):  # C2:C100
    col_a = ws[f'A{row}'].value
    col_b = ws[f'B{row}'].value
    if col_a is None or col_b is None:
        ws[f'C{row}'].value = None
    else:
        # col_b is STD time, typically a time or datetime
        if isinstance(col_b, datetime):
            std_time = col_b
        else:
            # Handle Excel time as float
            try:
                std_time = datetime(1899,12,30) + timedelta(days=col_b)
            except:
                ws[f'C{row}'].value = None
                continue
        cutoff_time = std_time - timedelta(minutes=30)
        # If cutoff_time < 0: wrap to previous day
        if cutoff_time.day != std_time.day:
            cutoff_time_str = cutoff_time.strftime('%H:%M')
        else:
            cutoff_time_str = cutoff_time.strftime('%H:%M')
        ws[f'C{row}'].value = cutoff_time_str

wb.save(output_path)
