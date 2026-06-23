from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_42216_s1/input.xlsx'
wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

for r in range(19, 80):
    a = ws.cell(r,1).value
    b = ws.cell(r,2).value
    print(r, a, b)
