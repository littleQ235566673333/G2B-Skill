import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/after_pass/core_36191/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/after_pass/core_36191/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

def extract_name(cell_value):
    if not cell_value or '@' not in cell_value:
        return ''
    # split before email
    before_email = cell_value.split('@')[0]
    # split into words
    parts = before_email.split()
    # Check for at least first name and surname
    if len(parts) < 2:
        return ''
    # Return everything except email
    return ' '.join(parts)

for row in range(2, 4):
    val = ws[f'B{row}'].value
    ws[f'C{row}'] = extract_name(val)

wb.save(output_path)
