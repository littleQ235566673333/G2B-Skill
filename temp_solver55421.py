from openpyxl import load_workbook
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/regression_gate/before_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/regression_gate/before_pass/core_55421/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

# Read all data for rows 2 to 20 (inclusive, for columns A to F)
rows = list(ws.iter_rows(min_row=2, max_row=20, min_col=1, max_col=6, values_only=True))
# Prepare for per-ID logic
A_to_rows = defaultdict(list)
for idx, row in enumerate(rows, start=2):
    A_val = row[0]
    if A_val is not None:
        A_to_rows[A_val].append((idx, row))

# Result per row
results = {}
for A_val, group in A_to_rows.items():
    statuses = [r[1][3] for r in group]
    dates    = [r[1][4] for r in group]
    # Apply rules per source row
    for (ix, row) in group:
        status = row[3]
        date   = row[4]
        result = None
        if status == 'NO SHOW':
            if date not in (None, ''):        # NO SHOW + date present
                result = 'NO ACTION NEEDED'
            else:                            # NO SHOW + empty date
                result = 'CALL PT'
        elif status == 'SCH':
            if set(statuses) == {'SCH'}:      # SCH only
                result = 'FUTURE'
            elif 'NO SHOW' in statuses:       # SCH + NO SHOW
                result = 'NS/SCHED'
        results[ix] = result

for ix in range(2, 21):
    val = results.get(ix, None)
    if val is not None:
        ws.cell(row=ix, column=6, value=val)
    else:
        ws.cell(row=ix, column=6, value='')

wb.save(output_path)
