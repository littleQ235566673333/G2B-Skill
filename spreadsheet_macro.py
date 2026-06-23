import openpyxl

# User's specified last-three-letter endings (in desired sort order)
GROUP_ORDER = ['ING', 'ERS', 'ATE', 'EST', 'ONE', 'IER', 'ILY']

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_8/group_118-50/r3/evolve_118-50/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_8/group_118-50/r3/evolve_118-50/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
sheet = wb['Sheet1']

names = []

# Read column A (skip header) and store all names
total_rows = sheet.max_row
for row in range(2, total_rows + 1):
    val = sheet.cell(row=row, column=1).value
    if isinstance(val, str):
        names.append(val.strip())

# Sort names alphabetically
names.sort()

# Group matches for each ending
result_groups = {ending: [] for ending in GROUP_ORDER}

for name in names:
    if len(name) >= 8:  # Need at least 8 letters to have 5th letter and 3-letter ending
        ending = name[-3:]
        if ending in GROUP_ORDER:
            if len(name) >= 5:
                fifth_letter = name[4]
                transformed = fifth_letter + name[:4] + name[5:]
                if transformed != name:
                    result_groups[ending].append((transformed, name))

# Arrange results grouped and ordered by ending
pairs = []
for ending in GROUP_ORDER:
    group_pairs = result_groups[ending]
    pairs.extend(group_pairs)

# Write pairs to columns C and D in Sheet1 (starting at row 2)
for idx, (transformed, original) in enumerate(pairs, start=2):
    sheet.cell(row=idx, column=3).value = transformed
    sheet.cell(row=idx, column=4).value = original

wb.save(output_path)
