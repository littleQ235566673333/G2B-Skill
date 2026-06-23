import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/group_341-40/r3/evolve_341-40/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/group_341-40/r3/evolve_341-40/output.xlsx'

# Load workbook and select active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Go through rows 2 to 701 (assuming header is at row 1)
for row in range(2, 702):
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    c = ws.cell(row=row, column=3).value

    if a == 'Government' and b == 'Germany' and c == 'Carretera':
        ws.cell(row=row, column=17, value='Volkswagen')  # Column Q is 17

# Save the updated workbook
wb.save(output_path)
