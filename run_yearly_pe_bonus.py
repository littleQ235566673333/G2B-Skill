import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Configuration
file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_4/group_47766/r0/evolve_47766/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_4/group_47766/r0/evolve_47766/output.xlsx'
df = pd.read_excel(file, sheet_name='Total (2)', header=None)
wb = load_workbook(file)
ws = wb['Total (2)']

block1 = df.iloc[7:37]  # rows 8:38
block2 = df.iloc[40:53] # rows 41:53

def parse_date(val):
    if pd.isna(val):
        return pd.NaT
    # Excel number type
    if isinstance(val, (int, float)):
        return pd.to_datetime(val, unit='d', origin='1899-12-30')
    # Date string
    return pd.to_datetime(val, errors='coerce')

# Get year start dates from N42, N47, N52 (col 14 = N, row=42,47,52)
years = [ws.cell(row=42, column=14).value, ws.cell(row=47, column=14).value, ws.cell(row=52, column=14).value]
# rows to include
data = pd.concat([block1, block2])

results = []
for yr_idx, start in enumerate(years):
    if pd.isnull(start):
        results.append(None)
        continue
    end = years[yr_idx + 1] if yr_idx + 1 < len(years) else None
    # Date column
    date_col = data[5].apply(parse_date)
    # Filter logic
    mask = data[7].astype(str).str.contains('PE', na=False) & (data[6] == 'Y')
    start_dt = pd.to_datetime(start)
    if end is not None and not pd.isnull(end):
        end_dt = pd.to_datetime(end)
        date_mask = (date_col >= start_dt) & (date_col < end_dt)
    else:
        date_mask = (date_col >= start_dt)
    # Commission sum
    filtered = data.loc[mask & date_mask, 2]
    summ = filtered.sum()
    results.append(float(summ))
    # Write output to cells K40, K45, K50
    target_row = 40 + 5 * yr_idx
    ws.cell(row=target_row, column=11, value=float(summ))

wb.save(output_file)
print('Yearly PE commission:', results)
