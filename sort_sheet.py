from openpyxl import load_workbook

def sort_sheet():
    input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_262-17_tc1/input.xlsx'
    output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_262-17_tc1/output.xlsx'
    wb = load_workbook(input_path)
    ws = wb['Sheet1']

    # Read all data
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = list(rows[1:])

    # Find column indexes
    task_idx = header.index('Task')
    resp_idx = header.index('Responsibility')

    # Sort by Task, then Responsibility
    data_sorted = sorted(data, key=lambda row: (str(row[task_idx] or ''), str(row[resp_idx] or '')))

    # Write header
    for c, val in enumerate(header, 1):
        ws.cell(row=1, column=c, value=val)

    # Write sorted data, clear unused rows in target range A1:F14
    for r in range(2, 15):
        for c in range(1, 7):
            ws.cell(row=r, column=c, value=None)
    for r, row in enumerate(data_sorted, start=2):
        if r > 14:
            break
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    wb.save(output_path)

if __name__ == '__main__':
    sort_sheet()
