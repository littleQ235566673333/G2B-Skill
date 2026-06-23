from openpyxl import load_workbook
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_1/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_1/regression_gate/after_pass/core_50526/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

lookup_value = ws['B6'].value
lookup_row = None
for row in range(2, ws.max_row+1):
    cell_value = ws.cell(row=row, column=1).value
    if cell_value == lookup_value:
        lookup_row = row
        break
if lookup_row is None:
    ws.cell(row=9, column=2).value = ""
    ws.cell(row=10, column=2).value = ""
    wb.save(output_path)
    exit()
output = []
for col in range(2, ws.max_column+1):
    val = ws.cell(row=lookup_row, column=col).value
    if val is not None and val > 0:
        output.append(ws.cell(row=1, column=col).value)
ws.cell(row=9, column=2).value = output[0] if len(output) > 0 else ""
ws.cell(row=10, column=2).value = output[1] if len(output) > 1 else ""
wb.save(output_path)
