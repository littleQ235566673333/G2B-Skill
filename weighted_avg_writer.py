from openpyxl import load_workbook
import re

def normalize_header(header):
    return re.sub(r'\s+', '', str(header)).lower() if header else ''

# File paths
in_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed2/eval_183-8_tc1/input.xlsx'
out_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed2/eval_183-8_tc1/output.xlsx'

wb = load_workbook(in_path)
assert 'Sheet1' in wb.sheetnames, 'Sheet1 not found.'
ws = wb['Sheet1']

# Locate headers and rows
header_row_idx = None
headers = []
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    normed = [normalize_header(h) for h in row]
    # Headers should have a block of non-blanks
    if sum(x != '' for x in normed) >= 3:
        header_row_idx = idx
        headers = normed
        break
assert header_row_idx is not None, 'Header row not found.'

# Map columns J, K, L
col_letters = {}
for col in ['j','k','l']:
    for i, h in enumerate(headers):
        if h == col:
            col_letters[col] = ws.cell(row=header_row_idx, column=i+1).column_letter
            break
for c in ['j','k','l']:
    assert c in col_letters, f'Column {c} not found.'

# Find data range by scanning down from header (until first full blank row)
data_start = header_row_idx + 1
data_end = data_start
while True:
    values = [ws[f'{col_letters[c]}{data_end}'].value for c in ['j','k','l']]
    if all(v is None or (isinstance(v, str) and v.strip()=='') for v in values):
        break
    data_end += 1
data_end -= 1  # The last filled row

# Try to find if there's a 'weight' column
weight_col = None
candidate_names = ['weight','weights','w']
for i, h in enumerate(headers):
    if any(n == h for n in candidate_names):
        weight_col = ws.cell(row=header_row_idx, column=i + 1).column_letter
        break
if not weight_col:
    # Try column I as fallback
    weight_col = 'I'

# For each cell in J3:L6, set the weighted average formula for the respective column
for c_idx, c in enumerate(['j','k','l'], start=10):
    col = col_letters[c]
    vals_rng = f'${col}${data_start}:${col}${data_end}'
    weights_rng = f'${weight_col}${data_start}:${weight_col}${data_end}'
    formula = f"=SUMPRODUCT({vals_rng},{weights_rng})/SUM({weights_rng})"
    # Output in row 3 to 6 for the col, mapped to 'Sheet1'!J3, K3, L3, ...
    for r in range(3, 7):
        ws.cell(row=r, column=c_idx+10).value = formula

wb.save(out_path)
