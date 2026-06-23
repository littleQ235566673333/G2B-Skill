from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_52541_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_52541_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

for row in range(6, 11):
    formula = f'=IF(H{row}<0,"Prepaid",IF(I{row}="","",IF(I{row}<90,"Call Customer","Bad Debts")))'
    ws[f'J{row}'] = formula

wb.save(output_path)
