import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_438-18_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_438-18_tc1/output.xlsx'

value_map = {
    'UNATIDE STATES AMERICAN': 'USA',
    'FRANCE': 'FRA',
    'KOREA': 'KOR',
    'GERMANY': 'GR',
    'ITALY': 'IT'
}

wb = openpyxl.load_workbook(input_path)
input_ws = wb['INPUT']
if 'OUTPUT' not in wb:
    wb.create_sheet('OUTPUT')
output_ws = wb['OUTPUT']

# Copy A1:D6 from INPUT to OUTPUT (values before replacement)
for row in range(1, 7):
    for col in range(1, 5):
        cell_value = input_ws.cell(row=row, column=col).value
        output_ws.cell(row=row, column=col, value=cell_value)

# Replace values in column F of INPUT, and put into OUTPUT column D
for row in range(1, 7):
    input_val = input_ws.cell(row=row, column=6).value
    replacement = value_map.get(input_val, input_val)
    output_ws.cell(row=row, column=4, value=replacement)

wb.save(output_path)
