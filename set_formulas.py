from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_5/group_40478/r0/evolve_40478/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_5/group_40478/r0/evolve_40478/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

formula_template = (
    'MID(A{row}, FIND(" ",A{row}, FIND(" ",A{row})+1)+1, ' 
    'FIND(" -",A{row})-FIND(" ",A{row}, FIND(" ",A{row})+1)-1)'
)
for row in range(1, 4):
    ws[f'B{row}'].value = f'={formula_template.format(row=row)}'

wb.save(output_path)
