import openpyxl

def main():
    in_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_38823_tc1/input.xlsx"
    out_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_38823_tc1/output.xlsx"
    wb = openpyxl.load_workbook(in_path)
    ws = wb.active

    # Get date range
    start_date = ws['E4'].value
    end_date = ws['F4'].value

    # Get search terms (assume in H4:H7)
    search_terms = [ws[f'H{row}'].value for row in range(4,8)]

    # Prepare output values
    results = []

    # Data rows are from row 3 to end (skip header at 2, row 1 is title)
    for term in search_terms:
        total = 0
        row = 3
        while True:
            # End of data when Date is None
            date = ws[f'A{row}'].value
            fabric_cell = ws[f'B{row}'].value
            units = ws[f'C{row}'].value
            if date is None:
                break
            # Date filter
            if start_date <= date <= end_date:
                # Fabric contains search term
                if fabric_cell and term in fabric_cell.split(','):
                    if type(units) in [int, float]:
                        total += units
            row += 1
        results.append(total)

    # Write results to I4:I7
    for i, val in enumerate(results):
        ws[f'I{4+i}'] = val

    wb.save(out_path)

if __name__ == "__main__":
    main()
