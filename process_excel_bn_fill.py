import openpyxl
from openpyxl.styles import PatternFill

# Input and output file paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_120-24_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_120-24_tc1/output.xlsx'

# Open the workbook and select the sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Fill for the color #0070C0
blue_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')

# Column indices (1-based: A=1, ..., AY=51, BG=59, BL=64, BN=66)
BL_col = 64  # 'BL'
BG_col = 59  # 'BG'
AY_col = 51  # 'AY'
BN_col = 66  # 'BN'

# BG specific values
BG_str1 = 'OCOGS - Spares - Transfer Price Overhead'
BG_str2 = 'OFSS - ORCL Consulting Prod Cost I/C'

# AY specific values
AY_set = {'BOA_033E', 'BOA_011G'}

# Process rows 2 through 45 inclusive
for row in range(2, 46):
    BL_val = ws.cell(row=row, column=BL_col).value
    BG_val = ws.cell(row=row, column=BG_col).value
    AY_val = ws.cell(row=row, column=AY_col).value

    BN_cell = ws.cell(row=row, column=BN_col)
    apply_fill = False
    # First two criteria: BL == 'LAG' and BG matches
    if BL_val == 'LAG' and BG_val in [BG_str1, BG_str2]:
        # Third: AY in specified set
        if AY_val in AY_set:
            BN_cell.value = BG_str1
        else:
            BN_cell.value = BG_str2
    # Apply fill if BL is 'LAG' and BN is populated
    if BL_val == 'LAG' and BN_cell.value:
        ws.cell(row=row, column=BL_col).fill = blue_fill

# Save result
wb.save(output_path)
