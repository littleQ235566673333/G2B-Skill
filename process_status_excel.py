from openpyxl import load_workbook
import collections

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/regression_gate/before_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/regression_gate/before_pass/core_55421/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Read data rows 2-20, columns A-F
rows = list(ws.iter_rows(min_row=2, max_row=20, min_col=1, max_col=6, values_only=False))
data = [[cell.value for cell in row] for row in rows]

A_vals = [row[0] for row in data]
D_vals = [row[3] for row in data]
E_vals = [row[4] for row in data]

# Build mapping: number in A -> list of (status D, date E, row idx)
num_stat = collections.defaultdict(list)
for idx, (num, status, dt) in enumerate(zip(A_vals, D_vals, E_vals)):
    num_stat[num].append((status, dt, idx))

results = [None]*len(data)
for num, occurrences in num_stat.items():
    statuses = set(x[0] for x in occurrences)
    has_sch = 'SCH' in statuses
    has_ns = 'NO SHOW' in statuses
    only_sch = statuses == {'SCH'}
    both = has_sch and has_ns
    for status, dt, idx in occurrences:
        output = None
        if only_sch:
            output = 'FUTURE'
        elif both:
            output = 'NS/SCHED'
        elif status == 'NO SHOW' and dt:
            output = 'NO ACTION NEEDED'
        elif status == 'NO SHOW' and not dt:
            output = 'CALL PT'
        results[idx] = output

# Write to F2:F20
for i, val in enumerate(results):
    ws.cell(row=2+i, column=6, value=val)

wb.save(output_path)
