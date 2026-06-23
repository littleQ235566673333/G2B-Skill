import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_13284_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_13284_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
base_ws = wb['Base']
streets_ws = wb['Streets']

# Preload Streets data (filter out rows without Start (C) or End (D))
streets_rows = []
for row in streets_ws.iter_rows(min_row=2, values_only=True):
    st_name = row[1]  # 'Streets'!B
    st_min = row[2]   # 'Streets'!C
    st_max = row[3]   # 'Streets'!D
    st_value = row[5] # 'Streets'!F
    # Only consider rows where both Start (C) and End (D) are populated (not None)
    if st_name is not None and st_min is not None and st_max is not None:
        streets_rows.append({'name': st_name, 'min': st_min, 'max': st_max, 'value': st_value})

# Look up for each Base row
for idx, base_row in enumerate(base_ws.iter_rows(min_row=2, max_row=26, min_col=3, max_col=4, values_only=True), start=2):
    base_match_val = base_row[0]  # 'Base'!C
    base_num = base_row[1]        # 'Base'!D
    result = ''
    if base_match_val is not None and base_num is not None:
        for strow in streets_rows:
            if base_match_val == strow['name'] and strow['min'] is not None and strow['max'] is not None:
                if strow['min'] <= base_num <= strow['max']:
                    result = strow['value'] if strow['value'] is not None else ''
                    break
    base_ws[f'E{idx}'] = result

wb.save(output_path)
