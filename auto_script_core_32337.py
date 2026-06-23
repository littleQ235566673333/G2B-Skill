import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime
input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/regression_gate/before_pass/core_32337/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/regression_gate/before_pass/core_32337/output.xlsx'
wb = load_workbook(input_file)
ws = wb.active

def get_age(dob, report_date):
    try:
        return report_date.year - dob.year - ((report_date.month, report_date.day) < (dob.month, dob.day))
    except Exception:
        return None

header_row = 2
headers = {c.value: c.column for c in ws[header_row]}
col_dob = headers.get('DATE OF BIRTH')
col_age = headers.get('age (year)')
col_category = headers.get('CATEGORY')
col_expected = headers.get('Expected Result')

col_actual_age = col_age + 1
ws.insert_cols(col_actual_age)
ws.cell(row=header_row, column=col_actual_age).value = 'Actual Age'
ref_cell = ws.cell(row=header_row, column=col_actual_age-1)
fill_obj = ref_cell.fill
if isinstance(fill_obj, PatternFill):
    ws.cell(row=header_row, column=col_actual_age).fill = PatternFill(fill_obj.patternType, fill_obj.fgColor.rgb, fill_obj.bgColor.rgb)
ws.cell(row=header_row, column=col_actual_age).alignment = Alignment(horizontal='center', vertical='top')

# Read report date from B1
report_val = ws.cell(row=1, column=2).value
if isinstance(report_val, str):
    try:
        report_val = datetime.strptime(report_val, '%Y-%m-%d')
    except:
        try:
            report_val = datetime.strptime(report_val, '%m/%d/%Y')
        except: report_val = None

for row in range(3, 16):
    dob_val = ws.cell(row=row, column=col_dob).value
    if isinstance(dob_val, str):
        try:
            dob_val = datetime.strptime(dob_val, '%Y-%m-%d')
        except:
            try:
                dob_val = datetime.strptime(dob_val, '%m/%d/%Y')
            except: dob_val = None
    actual_age = get_age(dob_val, report_val) if dob_val and report_val else None
    acell = ws.cell(row=row, column=col_actual_age)
    acell.value = actual_age if actual_age is not None else None
    acell.font = Font(bold=True)
    acell.alignment = Alignment(horizontal='center', vertical='center')
    acell.number_format = '0'

for row in range(3, 16):
    formula = (
        f"=INDEX(${get_column_letter(col_category)}$3:${get_column_letter(col_category)}$15, "
        f"MATCH({get_column_letter(col_age)}{row}, ${get_column_letter(col_age)}$3:${get_column_letter(col_age)}$15, 0))"
    )
    ws.cell(row=row, column=col_expected).value = formula

wb.save(output_file)
