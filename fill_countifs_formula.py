from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_4/group_45707/r2/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_4/group_45707/r2/evolve_45707/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Data starts in row 2, ends at row 69
start_row = 2
end_row = 69

for row in range(start_row, end_row+1):
    next_row = row + 1
    if next_row > end_row + 1:
        ws.cell(row=row, column=4, value=None)
        continue
    # Excel formula for the condition
    formula = (
        f'=IF(DAY(A{next_row})=1,'
        f'COUNTIFS(C${start_row}:C${end_row},1,'
        f"MONTH(A${start_row}:A${end_row}),MONTH(A{next_row}),"
        f"YEAR(A${start_row}:A${end_row}),YEAR(A{next_row}))"
        ',"")'
    )
    ws.cell(row=row, column=4, value=formula)

wb.save(output_path)
