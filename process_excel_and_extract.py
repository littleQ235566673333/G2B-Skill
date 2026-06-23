import openpyxl

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/regression_gate/after_pass/core_236-22/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/regression_gate/after_pass/core_236-22/output.xlsx"

wb = openpyxl.load_workbook(input_path)

# Find relevant sheets (from 'BR1' onwards)
start = False
sheets_to_process = []
for name in wb.sheetnames:
    if name == 'BR1':
        start = True
    if start:
        sheets_to_process.append(name)

for sheetname in sheets_to_process:
    ws = wb[sheetname]
    max_row = ws.max_row
    rows_to_delete = set()

    # Collect rows to delete (bottom up to avoid messing up indices)
    i = 1
    while i < max_row:
        cell_val = ws.cell(row=i, column=1).value
        next_cell_val = ws.cell(row=i+1, column=1).value if i+1 <= max_row else None
        # Exclude rows 2 & 9 on any sheet named 'Sheet1'
        if sheetname == 'Sheet1' and (i == 2 or i == 9 or i+1 == 2 or i+1 == 9):
            i += 1
            continue
        if cell_val and "Line No" in str(cell_val) and (next_cell_val is None or str(next_cell_val).strip() == ""):
            rows_to_delete.add(i)
            rows_to_delete.add(i+1)
            i += 2  # skip next row since we're removing both
        else:
            i += 1

    # Convert set to sorted list (descending) so row removal doesn't affect indices above)
    for row in sorted(rows_to_delete, reverse=True):
        if row > 0 and row <= ws.max_row:
            ws.delete_rows(row, 1)

# Save output to specified location
wb.save(output_path)

# Sheet1 A1:E7 to be the answer
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Sheet1'
ws_in = wb['Sheet1']
for r in range(1, 8):
    for c in range(1, 6):
        ws_out.cell(row=r, column=c, value=ws_in.cell(row=r, column=c).value)
wb_out.save(output_path)
