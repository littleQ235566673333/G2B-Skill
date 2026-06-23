import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_54638_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_54638_tc1/output.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_path)

# Get Sheet1
ds = wb['Sheet1']

# Set Calibri font for all cells in Sheet1
calibri = Font(name='Calibri')
for row in ds.iter_rows():
    for cell in row:
        cell.font = calibri

# Recreate formulas in column A up to row 150
for row in range(2, 151):
    formula = ds[f'A{row}'].value
    if isinstance(formula, str) and formula.startswith('='):
        ds[f'A{row}'].value = formula
    # If there's no formula, keep cell as is

# Add borders ONLY up to row 13 for column A
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
for row in range(2, 14):
    ds[f'A{row}'].border = border
# Remove border for A14 and onward
for row in range(14, 151):
    ds[f'A{row}'].border = Border()

# Insert non-dynamic array formula for uniques in B2:B150
for row in range(2, 151):
    ds[f'B{row}'].value = f'=IF(COUNTIF($A$2:A{row},A{row})=1,A{row},"")'
    ds[f'B{row}'].font = calibri
    ds[f'B{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    # Add gridlines (borders) for B2:B150
    ds[f'B{row}'].border = border

# Hide Sheet2
if 'Sheet2' in wb.sheetnames:
    ws2 = wb['Sheet2']
    ws2.sheet_state = 'hidden'

# Save workbook
wb.save(output_path)
