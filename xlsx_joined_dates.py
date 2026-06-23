import openpyxl
import pandas as pd
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/regression_gate/after_fix/core_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/regression_gate/after_fix/core_45896/output.xlsx'
main_sheet = 'Volym P5_P6_2023'
lookup_sheet = 'ZORD'

# Load workbook and relevant sheets
wb = openpyxl.load_workbook(input_path)
ws = wb[main_sheet]

# Use pandas for source lookups (for efficiency)
zord_df = pd.read_excel(input_path, sheet_name=lookup_sheet)

def try_format_date(val):
    if pd.isnull(val):
        return None
    # Excel dates as datetime
    if isinstance(val, pd.Timestamp):
        return val.strftime('%d/%m/%Y')
    if isinstance(val, datetime):
        return val.strftime('%d/%m/%Y')
    # string: try parse
    try:
        return pd.to_datetime(val).strftime('%d/%m/%Y')
    except:
        return str(val)  # fallback - shouldn't hit

# Get the row range (I2:I10)
for row in range(2, 11):
    key = ws[f'A{row}'].value  # Assumes lookup key is in column A
    # Get matching dates from ZORD
    matches = zord_df.loc[zord_df.iloc[:,0] == key, zord_df.columns[2]]  # col 0: A, col 2: C
    unique_dates = matches.drop_duplicates()
    sorted_dates = sorted(unique_dates, key=lambda x: pd.to_datetime(x))
    formatted = [try_format_date(d) for d in sorted_dates if pd.notnull(d)]
    result = ','.join(formatted)
    ws[f'I{row}'] = result

wb.save(output_path)
