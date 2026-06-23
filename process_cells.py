import openpyxl
from openpyxl.utils import get_column_letter

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_55060_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_55060_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)
ws = wb.active

source_value = ws["I12"].value

for col in range(openpyxl.utils.column_index_from_string('J'), openpyxl.utils.column_index_from_string('N')+1):
    cell = ws.cell(row=23, column=col)
    # If cell is a MergedCell, unmerge then write
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        merge_range = None
        for rng in ws.merged_cells.ranges:
            if cell.coordinate in rng:
                merge_range = rng
                break
        if merge_range:
            ws.unmerge_cells(str(merge_range))
        cell = ws.cell(row=23, column=col)  # re-fetch as normal cell
    if source_value in (None, ""):
        cell.value = None
    else:
        cell.value = source_value

wb.save(output_path)
