import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_414-20_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_414-20_tc1/output.xlsx'

# Load the workbook and select Sheet1
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Find the first occurrence of "Invoice No." in column A (case-insensitive)
first_occurrence_row = None
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
    cell_value = str(row[0].value).strip() if row[0].value is not None else ''
    if cell_value.lower() == 'invoice no.':
        first_occurrence_row = row[0].row
        break

# If found, delete all rows above it
if first_occurrence_row and first_occurrence_row > 1:
    ws.delete_rows(1, first_occurrence_row - 1)

# Save the workbook
wb.save(output_path)
