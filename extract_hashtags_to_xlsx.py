import re
from openpyxl import load_workbook

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/regression_gate/after_fix/core_57113/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/regression_gate/after_fix/core_57113/output.xlsx'

wb = load_workbook(input_fp)
ws = wb.active

def extract_hashtags(text):
    if not text:
        return []
    return [tag[1:] for tag in re.findall(r'#\w+', text)]

for row in range(2, 4):  # B2:C3
    tweet = ws.cell(row=row, column=1).value
    tags = extract_hashtags(tweet)
    for i in range(min(2, len(tags))):
        ws.cell(row=row, column=2 + i, value=tags[i])

wb.save(output_fp)
