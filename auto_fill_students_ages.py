from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_5/regression_gate/after_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_5/regression_gate/after_pass/core_41601/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Students']

# Populate E2:E7 with the INDIRECT formula
for row in range(2, 8):  # 2 through 7 inclusive
    student = ws[f'A{row}'].value
    cell = f'E{row}'
    if student and str(student).strip() != '':
        # Place the formula so that it references A{row}
        ws[cell] = f"=INDIRECT(\"'\"&A{row}&\"'!C2\")"
    else:
        ws[cell] = None

wb.save(output_path)
