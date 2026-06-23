from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_535-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_535-20_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

max_col = ws.max_column
# Find columns to delete
cols_to_delete = []
for col in range(1, max_col + 1):
    cell_value = ws.cell(row=1, column=col).value
    if cell_value and '/description' in str(cell_value):
        cols_to_delete.append(col)
# Delete columns in reverse order
for col_idx in sorted(cols_to_delete, reverse=True):
    ws.delete_cols(col_idx)
wb.save(output_path)
