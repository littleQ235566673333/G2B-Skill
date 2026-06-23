import openpyxl
from datetime import time

def time_in_range(t, start, end):
    """Return True if time t is within [start, end)"""
    # Assumes start < end, but start == end handled as closed
    return start <= t < end if start != end else t == start

# Load input workbook
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/regression_gate/after_fix/core_52305/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/regression_gate/after_fix/core_52305/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

data_rows = [r for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4, values_only=True)]

# Names are J5:N5 (columns 10..14, row 5)
name_cols = list(range(10, 15))  # J,N columns
names = [ws.cell(row=5, column=col).value for col in name_cols]
# Time windows H6:I24 (cols 8,9, rows 6..24)
time_windows = [(ws.cell(row=row, column=8).value, ws.cell(row=row, column=9).value) for row in range(6,25)]

# For each cell in J6:N24, count entries matching name & in window
for r_idx, (start, end) in enumerate(time_windows, start=6):
    for c_off, name in enumerate(names):
        count = 0
        for row in data_rows:
            dt = row[1]  # B column: datetime
            nm = row[2]  # C column: Name
            t = dt.time() if hasattr(dt, 'time') else None
            if nm == name and t is not None and time_in_range(t, start, end):
                count += 1
        ws.cell(row=r_idx, column=10 + c_off).value = count

wb.save(output_path)
print(f"Output saved to: {output_path}")
