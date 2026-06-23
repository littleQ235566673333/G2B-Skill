import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r2/eval_567-21_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r2/eval_567-21_tc1/output.xlsx'

# Read the spreadsheet
df = pd.read_excel(input_path, sheet_name='Sheet1')

# Keep only rows with non-empty values in column E (date)
df = df[df['E'].notna()]

# Normalize and parse 'E' to datetime, filter bad formats
def valid_date(val):
    try:
        return pd.to_datetime(str(val), format='%Y/%m', errors='raise')
    except Exception:
        return pd.NaT

df['E_parsed'] = df['E'].apply(valid_date)
df = df[df['E_parsed'].notna()]

# Group by Columns A & B, keep only rows with the latest date per group
grp = df.groupby(['A', 'B'])

def keep_latest_dates(group):
    max_date = group['E_parsed'].max()
    return group[group['E_parsed'] == max_date]

filtered_df = grp.apply(keep_latest_dates).reset_index(drop=True)
filtered_df = filtered_df[df.columns.difference(['E_parsed'])]  # Remove helper

# Write back to new Excel, preserving existing structure and only updating Sheet1!A3:G28
wb = load_workbook(input_path)
ws = wb['Sheet1']

# Clear writing region A3:G28
for row in ws.iter_rows(min_row=3, max_row=28, min_col=1, max_col=7):
    for cell in row:
        cell.value = None

# Write headers from row 2 to row 3 (to match template)
for col_idx, value in enumerate(df.columns, start=1):
    ws.cell(row=3, column=col_idx).value = ws.cell(row=2, column=col_idx).value

# Write the filtered results starting at row 4
for row_idx, row in enumerate(filtered_df.itertuples(index=False), start=4):
    if row_idx > 28:
        break
    for col_idx, value in enumerate(row, start=1):
        if col_idx > 7:
            break
        ws.cell(row=row_idx, column=col_idx).value = value

wb.save(output_path)
