from openpyxl import load_workbook
import datetime

# Input/output paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_5/regression_gate/after_fix/core_56274/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_5/regression_gate/after_fix/core_56274/output.xlsx'

wb = load_workbook(input_path)
sheet = wb[wb.sheetnames[0]]  # Sheet2

# Read D7 for Fiscal Month
fm_value = sheet['D7'].value

# Find the header row with month dates and the rows with Opening, Debits, Credits, and Closing
rows = list(sheet.iter_rows(values_only=True))
header_row = None
opening_row = None
debits_row = None
credits_row = None
closing_row = None
for row in rows:
    if header_row is None and isinstance(row[6], datetime.datetime):
        header_row = row
    if row[5] == 'Opening Bal':
        opening_row = row
    if row[5] == 'Debits':
        debits_row = row
    if row[5] == 'Credits':
        credits_row = row
    if row[5] == 'Closing Bal':
        closing_row = row

# Find the column index for the fiscal month
fm_col = None
if isinstance(fm_value, datetime.datetime):
    for idx, val in enumerate(header_row):
        if val == fm_value:
            fm_col = idx
            break
if fm_col is None:
    raise Exception('Fiscal Month value not found in header')

# Get the values
opening_value = opening_row[fm_col]
debits_value = debits_row[fm_col]
credits_value = credits_row[fm_col]
closing_value = closing_row[fm_col]

# Write values to D9:D12
sheet['D9'] = opening_value
sheet['D10'] = debits_value
sheet['D11'] = credits_value
sheet['D12'] = closing_value

wb.save(output_path)
print('Populated D9:D12 for Fiscal Month', fm_value)
