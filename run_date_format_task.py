from openpyxl import load_workbook
import re

# Parameters
input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed1/eval_486-17_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed1/eval_486-17_tc1/output.xlsx'
start_row = 2
end_row = 130  # inclusive, since B2:B130 is required
sheet_name = 'Blad1'

wb = load_workbook(input_path)
if sheet_name not in wb.sheetnames:
    raise ValueError(f'Sheet {sheet_name} not found. Sheets available: {wb.sheetnames}')
ws = wb[sheet_name]

def format_date_cell(number_val):
    '''Converts a number like 020210305 to yyyy mm dd with spaces.'''
    if number_val is None:
        return ''
    s = str(number_val).strip()
    # Remove leading '0' if present (formats are always 0yyyymmdd)
    if len(s) == 9 and s.startswith('0'):
        s = s[1:]
    # Now s should be 8 digits yyyymmdd
    if not re.match(r'^\d{8}$', s):
        return ''  # blank instead of error
    yyyy = s[:4]
    mm = s[4:6]
    dd = s[6:]
    return f"{yyyy} {mm} {dd}"

for i in range(start_row, end_row + 1):
    orig_val = ws[f'A{i}'].value  # input column is A
    formatted = format_date_cell(orig_val)
    ws[f'B{i}'].value = formatted

wb.save(output_path)
