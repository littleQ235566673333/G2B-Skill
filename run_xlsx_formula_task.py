import openpyxl
from collections import defaultdict
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_3/regression_gate/after_fix/core_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_3/regression_gate/after_fix/core_45707/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

dates = [ws[f'A{row}'].value for row in range(2, 70)]
vals = [ws[f'C{row}'].value for row in range(2, 70)]

# Map (year, month) -> list of indices in that month
date_indices = defaultdict(list)
for idx, d in enumerate(dates):
    if isinstance(d, datetime):
        key = (d.year, d.month)
        date_indices[key].append(idx)

# For each (year, month) count how many 1's in C
ones_count = {key: sum(1 for i in indices if vals[i] == 1) for key, indices in date_indices.items()}

for i in range(68):  # D2:D69 is i from 0 to 67
    next_date = dates[i+1] if i+1 < len(dates) else None
    cell = ws.cell(row=i+2, column=4)
    if isinstance(next_date, datetime) and next_date.day == 1:
        count = ones_count.get((next_date.year, next_date.month), 0)
        cell.value = count
    else:
        cell.value = None
# Save results
wb.save(output_path)
