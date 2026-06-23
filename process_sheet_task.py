import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/group_45707/r3/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/group_45707/r3/evolve_45707/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

dates = []
c_vals = []
for row in ws.iter_rows(min_row=2, max_row=69, min_col=1, max_col=3, values_only=True):
    dates.append(row[0])
    c_vals.append(row[2])

# Build (year, month) -> count of c=1 for month
date_months = defaultdict(int)
for date_val, c_val in zip(dates, c_vals):
    if isinstance(date_val, datetime) and c_val == 1:
        key = (date_val.year, date_val.month)
        date_months[key] += 1

for i in range(2, 70):  # Rows 2 to 69 inclusive
    a_val = ws[f'A{i}'].value
    if not isinstance(a_val, datetime):
        ws[f'D{i}'].value = None
        continue
    next_day = a_val + timedelta(days=1)
    if next_day.day == 1:
        count = date_months.get((next_day.year, next_day.month), 0)
        ws[f'D{i}'].value = count
    else:
        ws[f'D{i}'].value = None

wb.save(output_path)
