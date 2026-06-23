from openpyxl import load_workbook
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_5/group_57113/r2/evolve_57113/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_5/group_57113/r2/evolve_57113/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def extract_hashtags(text):
    if not text or not isinstance(text, str):
        return []
    hashtags = re.findall(r'#(\w+)', text)
    return hashtags

for row in range(2, 4):  # B2:C3 means rows 2 and 3
    tweet = ws.cell(row=row, column=1).value
    tags = extract_hashtags(tweet)
    for idx in range(2):
        ws.cell(row=row, column=2+idx).value = tags[idx] if idx < len(tags) else None

wb.save(output_path)
