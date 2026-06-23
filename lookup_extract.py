import openpyxl
from openpyxl import load_workbook

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_6/regression_gate/before_pass/core_50526/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_6/regression_gate/before_pass/core_50526/output.xlsx'

wb = load_workbook(input_file)
ws = wb.active

# Read the lookup value in B6
lookup_value = ws['B6'].value

# Find the row of the lookup value in column A
lookup_row = None
for row in range(2, ws.max_row + 1):
    if ws[f'A{row}'].value == lookup_value:
        lookup_row = row
        break

# Get headers: color names in row 1, columns B to last
headers = []
for col in range(2, ws.max_column + 1):
    headers.append(ws.cell(row=1, column=col).value)

results = []
if lookup_row:
    # For columns B:x, check if value is > 0 in lookup_row
    for i, col in enumerate(range(2, 2 + len(headers))):
        val = ws.cell(row=lookup_row, column=col).value
        if val is not None and isinstance(val, (int, float)) and val > 0:
            results.append(headers[i])

# Write results to B9, B10
ws['B9'] = results[0] if len(results) > 0 else None
ws['B10'] = results[1] if len(results) > 1 else None

wb.save(output_file)
