import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/regression_gate/before_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/regression_gate/before_pass/core_41601/output.xlsx'

wb = openpyxl.load_workbook(input_path)
students_sheet = wb['Students']

# Read student names from A2:A7
names = []
for row in range(2, 8):
    name = students_sheet[f'A{row}'].value
    names.append(name)

# Lookup C2 from each matching sheet
for idx, name in enumerate(names, start=2):
    age = None
    if name in wb.sheetnames:
        ws = wb[name]
        age = ws['C2'].value
    students_sheet[f'E{idx}'] = age

wb.save(output_path)
