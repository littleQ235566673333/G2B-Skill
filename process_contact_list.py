from openpyxl import load_workbook
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/regression_gate/after_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/regression_gate/after_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']
today = datetime.today()

# Find last row with data in column H
max_row = ws.max_row
for row in range(4, max_row+1):
    h_cell = ws[f'H{row}']
    i_cell = ws[f'I{row}']
    j_cell = ws[f'J{row}']
    h_val = h_cell.value
    i_val = str(i_cell.value).strip().upper() if i_cell.value else ''

    result = 'NO ACTION'
    if isinstance(h_val, datetime) and i_val == 'YES':
        delta = (today - h_val).days
        if 0 <= delta <= 30:
            result = 'HOLD'
        elif delta > 30:
            result = 'TOUCH BASE'
    j_cell.value = result

wb.save(output_path)
