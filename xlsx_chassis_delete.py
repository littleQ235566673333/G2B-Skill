from openpyxl import load_workbook, Workbook

# Define paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_2/group_269-44/r1/evolve_269-44/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_2/group_269-44/r1/evolve_269-44/output.xlsx"

# Load workbook and worksheet (actual sheet name is 'Sheet1')
wb = load_workbook(input_path)
ws = wb['Sheet1']

# Find all the rows where 'Chassis' appears in column A (case sensitive)
chassis_rows = []
for row in range(1, ws.max_row + 1):
    if ws.cell(row=row, column=1).value == 'Chassis':
        chassis_rows.append(row)

# Sort indices in reverse for safe row deletion
# For each chassis row, delete the 4 rows immediately before it
rows_to_delete = []
for chassis_row in chassis_rows:
    if chassis_row > 4:
        rows_to_delete.extend([chassis_row-4, chassis_row-3, chassis_row-2, chassis_row-1])
    else:
        rows_to_delete.extend(list(range(1, chassis_row)))

# Remove duplicates (if overlapping), sort in reverse for safe deletion
rows_to_delete = sorted(set(rows_to_delete), reverse=True)

for del_row in rows_to_delete:
    ws.delete_rows(del_row, 1)

# Save ONLY A1:A15 to conform to instructions
wb_out = Workbook()
ws_out = wb_out.active
for i in range(1, 16):
    ws_out[f"A{i}"] = ws.cell(row=i, column=1).value

wb_out.save(output_path)
