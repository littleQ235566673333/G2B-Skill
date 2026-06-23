import math
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_4/group_44017/r1/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_4/group_44017/r1/evolve_44017/output.xlsx'

# Read dates as values, not formulas
wb = load_workbook(input_path, data_only=True)
ws = wb['Data']
# Frequency mapping from reference table (I6:J8)
freq_map = {ws.cell(row=r, column=9).value: ws.cell(row=r, column=10).value for r in range(6,9)}

row_start, row_end = 14, 42
col_start, col_end = 30, 43  # AD to AO
ref_date_row = 9
# Get the rate period start dates for the 12 output columns
col_dates = [ws.cell(row=ref_date_row, column=c).value for c in range(col_start, col_end+1)]

# Reopen workbook for writing
wb2 = load_workbook(input_path)
ws2 = wb2['Data']

for row in range(row_start, row_end+1):
    base = ws2.cell(row=row, column=23).value  # W
    eff_date = ws2.cell(row=row, column=12).value  # L
    label = ws2.cell(row=row, column=9).value
    val = ws2.cell(row=row, column=10).value
    # Resolve freq value using freq_map if needed
    try:
        freq = int(val)
    except Exception:
        freq = freq_map.get(label, None)
    increases = [ws2.cell(row=row, column=c).value for c in range(13, 17)]  # M-P
    increases_seq = [x for x in increases if isinstance(x, (float, int))]
    for i, c in enumerate(range(col_start, col_end+1)):
        cell = ws2.cell(row=row, column=c)
        cell_date = col_dates[i]
        if not isinstance(eff_date, datetime) or not isinstance(cell_date, datetime) or freq is None:
            cell.value = None
            continue
        # Only fill on/after effective date
        if cell_date < eff_date:
            cell.value = None
            continue
        months = (cell_date.year - eff_date.year) * 12 + (cell_date.month - eff_date.month)
        waves = math.floor(months / freq) if freq > 0 else 0
        waves = min(waves, len(increases_seq))
        rate = base
        for j in range(waves):
            rate *= (1 + increases_seq[j])
        cell.value = round(rate, 2)

wb2.save(output_path)
