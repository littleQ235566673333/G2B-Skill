import openpyxl

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_53161_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_53161_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Step 1: Parse Agent Categories Table (starts at E3)
category_rows = range(4, 10)  # 6 agents
category_cols = range(6, 15)  # Category 1-9 (F-M)
agent_categories = {}  # agent: list of assigned categories
agents = []
categories = []
for i, row in enumerate(category_rows):
    agent = ws.cell(row=row, column=5).value  # Column E
    if agent:
        agents.append(agent)
        agent_categories[agent] = []
        for cat_col in category_cols:
            agent_categories[agent].append(ws.cell(row=row, column=cat_col).value)

for cat_col in category_cols:
    categories.append(ws.cell(row=3, column=cat_col).value)

num_categories = len(categories)

# Step 2: Parse Agent Schedule Table (starts at E13)
schedule_rows = range(14, 20)  # 6 agents
interval_cols = range(6, 36)   # F-AE, total 30 intervals
interval_times = []
for col in interval_cols:
    interval_times.append(ws.cell(row=12, column=col).value)

agent_schedule = {}  # agent: list of scheduled intervals
for row in schedule_rows:
    agent = ws.cell(row=row, column=5).value
    if agent:
        agent_schedule[agent] = []
        for col in interval_cols:
            agent_schedule[agent].append(ws.cell(row=row, column=col).value)

# Step 3: Calculate totals for each category & interval
category_interval_counts = [[0 for _ in interval_cols] for _ in range(num_categories)]

for interval_idx, col in enumerate(interval_cols):
    for agent in agents:
        if agent not in agent_schedule or agent not in agent_categories:
            continue
        if agent_schedule[agent][interval_idx] == 1:
            # Add agent's assigned categories
            for cat_idx in range(num_categories):
                if agent_categories[agent][cat_idx] == 1:
                    category_interval_counts[cat_idx][interval_idx] += 1

# Step 4: Write results to F22:AE29
result_start_row = 22
result_end_row = result_start_row + num_categories  # Row 29
result_start_col = 6  # F
result_end_col = result_start_col + len(interval_cols)  # AE

for cat_idx in range(num_categories):
    ws.cell(row=result_start_row + cat_idx, column=5).value = categories[cat_idx]  # Category name
    for interval_idx in range(len(interval_cols)):
        ws.cell(row=result_start_row + cat_idx, column=result_start_col + interval_idx).value = category_interval_counts[cat_idx][interval_idx]

wb.save(output_path)
