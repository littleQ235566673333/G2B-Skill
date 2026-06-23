import openpyxl

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_23-24_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_23-24_tc1/output.xlsx'

# Load the workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get the set of words to delete from column I (skip blanks)
words_to_delete = set()
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=9, max_col=9):
    val = row[0].value
    if val is not None and str(val).strip() != '':
        words_to_delete.add(str(val))

# We'll check rows 1-1102 in column A
delete_indices = []
for i in range(1, 1103):
    cell_value = ws.cell(row=i, column=1).value
    if cell_value is not None and str(cell_value) in words_to_delete:
        delete_indices.append(i)

# Delete from bottom up to avoid mis-indexing
for idx in reversed(delete_indices):
    ws.delete_rows(idx, 1)

# Save to the output path
wb.save(output_path)
