from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_59224_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_59224_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

date_cell = ws['B2']
project_start = date_cell.value
fill_range = list(range(4, 15))
select_period_row = None

# First, scan for the first row that should be 'Select Period'
for row in fill_range:
    c_val = ws[f'C{row}'].value
    d_val = ws[f'D{row}'].value
    if project_start and c_val and d_val and c_val < project_start < d_val and select_period_row is None:
        select_period_row = row
        break

# Fill all cells up to and including select_period_row with 'Select Period', rest with 'Do Not Select'
for row in fill_range:
    if select_period_row and row <= select_period_row:
        ws[f'E{row}'].value = 'Select Period'
    else:
        ws[f'E{row}'].value = 'Do Not Select'

wb.save(output_path)
