from openpyxl import load_workbook
from datetime import datetime

def is_date(val):
    if isinstance(val, datetime):
        return True
    # Try string coercion of common date formats
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
            try:
                datetime.strptime(val, fmt)
                return True
            except ValueError:
                continue
    return False

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r2/eval_46240_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r2/eval_46240_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

for i in range(2, 5):  # J2:J4 (i = row)
    h = ws[f'H{i}'].value
    # 1. N/A
    if isinstance(h, str) and h.strip().upper() == 'N/A':
        ws[f'J{i}'] = 'N/A'
    # 2. Yes, if H contains a date or 'Yes'
    elif (isinstance(h, str) and h.strip().lower() == 'yes') or is_date(h):
        ws[f'J{i}'] = 'Yes'
    # 3. No, if blank (None or empty string)
    elif (h is None) or (isinstance(h, str) and h.strip()==''):
        ws[f'J{i}'] = 'No'
    else:
        ws[f'J{i}'] = 'No'  # fallback

wb.save(output_path)
