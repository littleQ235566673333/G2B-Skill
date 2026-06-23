from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_42216_s1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_42216_s1/output.xlsx'

wb = load_workbook(input_path)
print(wb.sheetnames)
for ws in wb.worksheets:
    print('SHEET', ws.title)
    for r in range(1, 45):
        vals = [ws.cell(r, c).value for c in range(1, 8)]
        if any(v is not None for v in vals):
            print(r, vals)
    print('max_row', ws.max_row, 'max_col', ws.max_column)
