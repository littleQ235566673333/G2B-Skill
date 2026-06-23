import openpyxl

def insert_row_above_x(input_path, output_path, sheet_name, col_idx, start_row, end_row, max_output_rows):
    wb = openpyxl.load_workbook(input_path)
    ws = wb[sheet_name]

    # Collect positions where we need to insert a row (from start_row to end_row)
    insert_rows = []
    for row in range(start_row, end_row+1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value == "X":
            insert_rows.append(row)

    # Insert rows in reverse order to prevent shifting
    for row in reversed(insert_rows):
        ws.insert_rows(row)

    # Save workbook
    wb.save(output_path)

    # Ensure only the first max_output_rows remain visible in output
    # (optional based on instruction, but let's trim to be safe)
    wb_out = openpyxl.load_workbook(output_path)
    ws_out = wb_out[sheet_name]
    ws_out.delete_rows(max_output_rows+1, ws_out.max_row-max_output_rows)
    wb_out.save(output_path)

insert_row_above_x(
    input_path='results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_370-43_tc1/input.xlsx',
    output_path='results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_370-43_tc1/output.xlsx',
    sheet_name='Before Insert Row',
    col_idx=1,
    start_row=7,
    end_row=1000,
    max_output_rows=70
)
