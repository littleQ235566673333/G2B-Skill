from openpyxl import load_workbook

def norm(val):
    return str(val).strip().lower() if val is not None else ''

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_141-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_141-20_tc1/output.xlsx'

wb = load_workbook(input_path)
ws_pl = wb['PL Recon Items']
ws_st = wb['Statement Recon Items']

# Get header mappings
def get_header_map(ws):
    return {cell.value: idx for idx, cell in enumerate(ws[1])}

pl_map = get_header_map(ws_pl)
st_map = get_header_map(ws_st)

# Gather all data rows (row index, row tuple)
def get_rows(ws, header_len):
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if any(cell is not None for cell in row):
            rows.append((i, tuple(row) + (None,)*(header_len-len(row))))
    return rows

pl_rows = get_rows(ws_pl, len(pl_map))
st_rows = get_rows(ws_st, len(st_map))

# Build sets for cross-matching
pl_keys = { (norm(row[pl_map['Invoice No.']]), norm(row[pl_map['Amount']])): idx for idx, row in pl_rows }
st_keys = { (norm(row[st_map['Reference']]), norm(row[st_map['Total Invoice Value']])): idx for idx, row in st_rows }

# Find common keys, to be deleted
keys_to_delete = set(pl_keys) & set(st_keys)

pl_del_idxs = sorted([pl_keys[k] for k in keys_to_delete], reverse=True)
st_del_idxs = sorted([st_keys[k] for k in keys_to_delete], reverse=True)

# Delete rows from bottom to top (reverse order)
for idx in pl_del_idxs:
    ws_pl.delete_rows(idx)
for idx in st_del_idxs:
    ws_st.delete_rows(idx)

wb.save(output_path)
