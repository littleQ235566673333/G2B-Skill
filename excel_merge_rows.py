import openpyxl
import shutil
from copy import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/regression_gate/before_fix/core_177-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/regression_gate/before_fix/core_177-6/output.xlsx'

shutil.copyfile(input_path, output_path)
wb = openpyxl.load_workbook(output_path)
ws = wb.active
combined = wb.create_sheet('combined') if 'combined' not in wb.sheetnames else wb['combined']

rows = list(ws.iter_rows(min_row=1, max_col=18, values_only=False))
header = [cell.value for cell in rows[0]]
combined.append(header)

ref_col = 7  # H (0-based index: 7)
data_rows = rows[1:]
groups = {}
for r in data_rows:
    key = r[ref_col].value
    if key not in groups:
        groups[key] = []
    groups[key].append(r)

def merge_group(group):
    merged = []
    for i in range(len(group[0])):
        for row in group:
            val = row[i].value
            if val not in (None, ''):
                merged_val = val
                break
        else:
            merged_val = None
        merged.append(merged_val)
    return merged

def copy_format(src_cell, tgt_cell):
    # Only use copy to avoid StyleProxy issues
    if src_cell.font: tgt_cell.font = copy(src_cell.font)
    if src_cell.fill: tgt_cell.fill = copy(src_cell.fill)
    if src_cell.border: tgt_cell.border = copy(src_cell.border)
    if src_cell.alignment: tgt_cell.alignment = copy(src_cell.alignment)
    if src_cell.protection: tgt_cell.protection = copy(src_cell.protection)
    if src_cell.number_format: tgt_cell.number_format = src_cell.number_format

for idx, group in enumerate(groups.values(), start=2):
    ref_row = group[0]
    merged = merge_group(group)
    for c in range(18):
        cell = combined.cell(row=idx, column=c+1, value=merged[c])
        src_cell = ref_row[c]
        copy_format(src_cell, cell)
        if 8 <= c <= 17:  # I-R (0-based)
            v = cell.value
            if isinstance(v, (float, int)) and v == 0:
                cell.value = None
                cell.number_format = '0.00'
            elif isinstance(v, (float, int)):
                cell.value = round(float(v), 2)
                cell.number_format = '0.00'
            elif v is None:
                cell.value = None
                cell.number_format = '0.00'
        
wb.save(output_path)
