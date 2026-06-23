import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun2/eval_22-47_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun2/eval_22-47_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))

A, B, C, J = 0, 1, 2, 9
output_start_row = 2
output_base_col = 5

# Filter data rows (skip header and empty)
data_rows = [r for r in rows[1:] if any(r)]

# Collect all unique helper names (in order of appearance, no blanks)
j_names = []
j_seen = set()
for row in data_rows:
    name = row[J]
    if name and name not in j_seen:
        j_names.append(name)
        j_seen.add(name)

# Remove duplicates (by B and C), skip blanks for B or C
seen_pairs = set()
source_data = []
for row in data_rows:
    if not row[B] or not row[C]:
        continue
    pair = (row[B], row[C])
    if pair in seen_pairs:
        continue
    seen_pairs.add(pair)
    source_data.append(row)

if not any(j_names):  # Helper column J is empty; sort by B A-Z
    sorted_data = sorted(source_data, key=lambda r: str(r[B]).lower())
else:
    # Group by each J name (all matching in source order), then others in original order
    matched = []
    matched_ids = set()
    for name in j_names:
        for row in source_data:
            if row[B] == name and id(row) not in matched_ids:
                matched.append(row)
                matched_ids.add(id(row))
    remainder = [row for row in source_data if id(row) not in matched_ids]
    sorted_data = matched + remainder

out_data = sorted_data[:9]
target_cols = [B, C, J]
for i, row in enumerate(out_data):
    for j, col_idx in enumerate(target_cols):
        ws.cell(row=output_start_row + i, column=output_base_col + j, value=row[col_idx])

# Sort H (col H, output_base_col+2) within output
h_col = output_base_col + 2
rows_to_sort = len(out_data)
h_values = [ws.cell(row=output_start_row + i, column=h_col).value for i in range(rows_to_sort)]
# Use only string or number for sorting, treat blank/None as 'zzz' (sort last)
def h_sortval(val):
    if val is None or (isinstance(val, str) and val.strip() == ''):
        return 'zzz'
    return str(val)
sort_order = sorted(range(rows_to_sort), key=lambda i: h_sortval(h_values[i]))

for j, col_idx in enumerate(target_cols):
    vals = [ws.cell(row=output_start_row + i, column=output_base_col + j).value for i in sort_order]
    for i, v in enumerate(vals):
        ws.cell(row=output_start_row + i, column=output_base_col + j, value=v)

wb.save(output_path)
