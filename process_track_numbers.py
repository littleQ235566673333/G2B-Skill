import openpyxl

# File paths
i_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_16511_tc1/input.xlsx'
o_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_16511_tc1/output.xlsx'

# Load workbook and select sheet
wb = openpyxl.load_workbook(i_path)
ws = wb.active

album_col = 'B'  # Assume album name is in column B (change if needed)
start_row = 2
end_row = 10

album_track_dict = {}

for row in range(start_row, end_row + 1):
    album = ws[f'{album_col}{row}'].value
    if album is None:
        ws[f'F{row}'] = None
        continue
    if album not in album_track_dict:
        album_track_dict[album] = 1
    else:
        album_track_dict[album] += 1
    ws[f'F{row}'] = album_track_dict[album]

wb.save(o_path)
