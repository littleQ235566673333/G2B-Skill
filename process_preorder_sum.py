import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/regression_gate/after_pass/core_50796/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/regression_gate/after_pass/core_50796/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get criteria from B2:B5
criteria = [ws[f'B{i}'].value for i in range(2,6)]
results = []
for crit in criteria:
    total = 0
    for row in range(12, 23):  # Data rows
        if ws[f'B{row}'].value == crit:
            val = ws[f'C{row}'].value
            try:
                total += float(val)
            except (TypeError, ValueError):
                pass
    results.append(total)
# Write results into B2:B5
for idx, result in enumerate(results, start=2):
    ws[f'B{idx}'].value = result
wb.save(output_path)
