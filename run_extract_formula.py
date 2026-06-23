from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_6/regression_gate/after_fix/core_40478/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_6/regression_gate/after_fix/core_40478/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active
# Set Excel formulas in B1:B3 to extract text after 2nd space and before dash in A1:A3
for i in range(1, 4):
    ws[f'B{i}'] = f'=MID(A{i}, FIND(" ",A{i},FIND(" ",A{i})+1)+1, FIND("-",A{i})-FIND(" ",A{i},FIND(" ",A{i})+1)-2)'
wb.save(output_path)
