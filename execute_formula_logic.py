import openpyxl

def get_lookup_value(grouping_sheet, group_name, material, col_offset):
    # Search for exact match first (if material is present)
    for row in grouping_sheet.iter_rows(min_row=2, max_row=7):
        sheet_group = row[0].value
        sheet_material = row[1].value
        if sheet_group == group_name and sheet_material == material:
            return row[2 + col_offset].value
    # If not found, try only by group
    for row in grouping_sheet.iter_rows(min_row=2, max_row=7):
        sheet_group = row[0].value
        sheet_material = row[1].value
        if sheet_group == group_name and sheet_material is None:
            return row[2 + col_offset].value
    return None

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_7902_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_7902_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_file)
ws_form = wb['Formula Required']
ws_group = wb['Grouping']

rows = [3, 4, 5, 6]
cols = list(range(4, 11)) # D:J
input_cols = list(range(11, 18)) # L:R

for row in rows:
    group = ws_form.cell(row=row, column=3).value # C
    material = ws_form.cell(row=row, column=2).value # B
    for i, col in enumerate(cols):
        input_val = ws_form.cell(row=row, column=input_cols[i]).value
        if input_val not in [None, 0]:
            ws_form.cell(row=row, column=col).value = input_val
        else:
            lookup_val = get_lookup_value(ws_group, group, material, i)
            ws_form.cell(row=row, column=col).value = lookup_val

wb.save(output_file)
