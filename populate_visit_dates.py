import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_2/regression_gate/before_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_2/regression_gate/before_pass/core_9726/output.xlsx'

# Load all sheets
sheets = pd.read_excel(input_path, sheet_name=None)
# Identify Table 1 and Table 2; usually first two sheets
sheet_names = list(sheets.keys())
df1 = sheets[sheet_names[0]]
df2 = sheets[sheet_names[1]]

# Normalize date columns, handle errors
if 'Visit date' in df1:
    df1['Visit date'] = pd.to_datetime(df1['Visit date'], errors='coerce')
if 'Submission date' in df2:
    df2['Submission date'] = pd.to_datetime(df2['Submission date'], errors='coerce')

def get_prev_visit(student, sub_date):
    filt = (df1['Student'] == student) & (df1['Visit date'] < sub_date)
    visits = df1.loc[filt, 'Visit date']
    if not visits.empty:
        return visits.max()
    else:
        return pd.NaT

visit_dates = [get_prev_visit(row['Student'], row['Submission date']) for _, row in df2.iterrows()]

# Write visit_dates to Table 2 in J2 downward
wb = load_workbook(input_path)
ws = wb[sheet_names[1]]  # Table 2

for idx, visit_date in enumerate(visit_dates, start=2):
    ws[f'J{idx}'] = visit_date.strftime('%d-%b-%y') if pd.notnull(visit_date) else ''

wb.save(output_path)
