import openpyxl
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_4/group_45707/r1/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_4/group_45707/r1/evolve_45707/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 70):  # D2:D69, corresponding to A2:A69
    a_val = ws[f'A{row}'].value
    if not isinstance(a_val, datetime):
        ws[f'D{row}'].value = None
        continue
    next_day = a_val + timedelta(days=1)
    if next_day.day == 1:
        # Count the number of 1s in column C for this month and year
        month = a_val.month
        year = a_val.year
        count_ones = 0
        for check_row in range(2, 70):
            check_date = ws[f'A{check_row}'].value
            if not isinstance(check_date, datetime):
                continue
            if check_date.month == month and check_date.year == year:
                c_val = ws[f'C{check_row}'].value
                if c_val == 1:
                    count_ones += 1
        ws[f'D{row}'].value = count_ones
    else:
        ws[f'D{row}'].value = None

wb.save(output_path)
