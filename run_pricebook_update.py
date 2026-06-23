import openpyxl

# Load the input Excel file
input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_57743_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_57743_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read model numbers from A2:A19
model_numbers = [ws[f'A{row}'].value for row in range(2, 20)]

# Build dictionary from reference list in C:D
model_price_dict = {}
row = 2
while True:
    model = ws[f'C{row}'].value
    # Break if end of model list
    if model is None:
        break
    price = ws[f'D{row}'].value
    model_price_dict[model] = price
    row += 1

# Lookup prices for model numbers in A2:A19
for i, model_num in enumerate(model_numbers, start=2):
    price = model_price_dict.get(model_num, '')
    ws[f'B{i}'] = price

# Save to output file
wb.save(output_path)
