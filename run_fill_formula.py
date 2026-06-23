import openpyxl

def find_row(ws_grouping, key1, key2):
    for row in ws_grouping.iter_rows(min_row=2, max_row=7, min_col=1, max_col=9):
        values = [cell.value for cell in row]
        if key2 is not None:
            if values[0] == key1 and values[1] == key2:
                return values
        else:
            if values[0] == key1 and (values[1] is None or values[1] == ''):
                return values
    return None

def main():
    path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_7902_tc1/input.xlsx'
    wb = openpyxl.load_workbook(path)
    ws_formula = wb['Formula Required']
    ws_grouping = wb['Grouping']
    # Fill D3:J6
    for r_idx, sheet_row in enumerate(ws_formula.iter_rows(min_row=3, max_row=6, min_col=4, max_col=10), start=3):
        main_group = ws_formula.cell(row=r_idx, column=3).value
        material = ws_formula.cell(row=r_idx, column=2).value
        g_row = find_row(ws_grouping, main_group, material)
        if g_row is None:
            g_row = find_row(ws_grouping, main_group, None)
        for c_offset, cell in enumerate(sheet_row, start=2):
            cell.value = g_row[c_offset] if g_row and len(g_row) > c_offset else None
    wb.save('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_7902_tc1/output.xlsx')
if __name__ == '__main__':
    main()
