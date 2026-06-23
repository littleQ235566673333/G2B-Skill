import pandas as pd
from openpyxl import load_workbook
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_7/regression_gate/before_fix/core_146-49/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_7/regression_gate/before_fix/core_146-49/output.xlsx'

def extract_words(ws):
    # Find first column with words, assume starts in A (col 1, 0-based for DataFrame, 1-based for openpyxl)
    vals = []
    col = 1
    while True:
        header = ws.cell(row=1, column=col).value
        if header and 'word' in str(header).lower():
            break
        col += 1
        if col > 20:
            col = 1
            break
    # Collect all non-empty, unique words (excluding header)
    for r in range(2, ws.max_row+1):
        w = ws.cell(row=r, column=col).value
        if w and isinstance(w, str):
            vals.append(w.strip())
    return vals

wb = load_workbook(input_path)
ws = wb['Sheet1']
words = extract_words(ws)
words = list(dict.fromkeys(words))  # Unique, preserve order
vowels = ['A','E','I','O','U']
pairs = set()
def vowelify(base, v):
    return ''.join(c if c!='S' else v for c in base)
for word in words:
    if 'S' not in word: continue
    for v in vowels:
        alt = vowelify(word, v)
        if alt == word: continue
        if alt in words and len(alt)==len(word):
            # All non-S letters must match
            if all(sc == ac or sc=='S' for sc,ac in zip(word,alt)):
                pair = tuple(sorted((word, alt), key=lambda x: x.lower()))
                pairs.add(pair)
pair_list = sorted(list(pairs), key=lambda x: (x[0].lower(), x[1].lower()))
# Write results to G/H columns, starting at row 2
for i, (w1, w2) in enumerate(pair_list):
    ws.cell(row=2+i, column=7).value = w1
    ws.cell(row=2+i, column=8).value = w2
# Sort pairs case-insensitively as requested; move up to G1/H1
if pair_list:
    temp_list = [(ws.cell(row=2+i, column=7).value, ws.cell(row=2+i, column=8).value) for i in range(len(pair_list))]
    temp_list = sorted(temp_list, key=lambda x: (x[0].lower(), x[1].lower()))
    for i, (w1, w2) in enumerate(temp_list):
        ws.cell(row=1+i, column=7).value = w1
        ws.cell(row=1+i, column=8).value = w2
    # Remove remaining old position, if any (in case pairs < 64 previously written)
    for i in range(len(temp_list), len(pair_list)+1):
        ws.cell(row=1+i, column=7).value = None
        ws.cell(row=1+i, column=8).value = None
wb.save(output_path)
