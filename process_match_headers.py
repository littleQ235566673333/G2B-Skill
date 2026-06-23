from openpyxl import load_workbook
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_3/regression_gate/before_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_3/regression_gate/before_pass/core_50526/output.xlsx'

# Load workbook and relevant worksheet
wb = load_workbook(input_path)
ws = wb.active

# Read data as DataFrame
# We do NOT want to interpret the first column as the index
# It is lookup values (e.g. 'A', 'B')
df = pd.read_excel(input_path)

# Get the lookup value from B6
b6_val = ws['B6'].value

# Function to get headers for positive cells in the row matching lookup
def get_positive_headers(lookup_val):
    row = df[df.iloc[:,0] == lookup_val]
    if row.empty:
        return []
    # Exclude the row label; only check rest of columns
    vals = row.iloc[0,1:]
    headers = df.columns[1:][vals.values > 0]
    return list(headers)

result = get_positive_headers(b6_val)

# Write results into B9, B10
for i in range(2):
    ws.cell(row=9+i, column=2).value = result[i] if i < len(result) else None

wb.save(output_path)
