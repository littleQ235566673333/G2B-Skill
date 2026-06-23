import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/regression_gate/after_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/regression_gate/after_pass/core_9726/output.xlsx'

# Read Sheet1
sheet_name = 'Sheet1'
df = pd.read_excel(input_path, sheet_name=sheet_name)

# Table 1 columns: 'Student ID', 'Visit date'
# Table 2 columns: 'Student ID.1', 'Submission date', to fill 'Visit date.1'
visit_df = df[['Student ID', 'Visit date']].dropna(subset=['Visit date'])
subm_df = df[['Student ID.1', 'Submission date']].copy()

# Convert to datetime
visit_df['Visit date'] = pd.to_datetime(visit_df['Visit date'], errors='coerce', dayfirst=True)
subm_df['Submission date'] = pd.to_datetime(subm_df['Submission date'], errors='coerce', dayfirst=True)

# Build map: Student ID → sorted visit dates
visit_map = {}
for student, group in visit_df.groupby('Student ID'):
    vsorted = sorted(d for d in group['Visit date'] if pd.notna(d))
    visit_map[student] = vsorted

last_visits = []
for idx, row in subm_df.iterrows():
    student = row['Student ID.1']
    subm_date = row['Submission date']
    vdates = visit_map.get(student, [])
    prior_visits = [v for v in vdates if v < subm_date and pd.notna(v)]
    # Get max prior visit or None
    last_visits.append(max(prior_visits) if prior_visits else None)

# Write result in df and also via openpyxl to Visit date.1
wb = load_workbook(input_path)
ws = wb[sheet_name]

# Find the header row and col for Table 2 (where Student ID.1 appears)
header_row = None
visit_date1_col = None
for row in ws.iter_rows(min_row=1, max_row=10):
    headers = [str(c.value).strip() if c.value else '' for c in row]
    if 'Student ID.1' in headers and 'Submission date' in headers:
        header_row = row[0].row
        visit_date1_col = headers.index('Visit date.1') + 1
        break
if header_row is None or visit_date1_col is None:
    raise Exception('Could not find Table 2 header or Visit date.1 col')

# Write the results (assumes Table 2 starts at that header row, just as in pandas)
for i, value in enumerate(last_visits):
    cell = ws.cell(row=header_row + 1 + i, column=visit_date1_col)
    if pd.isna(value) or value is None:
        cell.value = None
    else:
        cell.value = value.strftime('%d-%b-%y')

wb.save(output_path)
