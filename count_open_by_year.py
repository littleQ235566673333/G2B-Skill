from openpyxl import load_workbook
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun1/eval_41978_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun1/eval_41978_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

years = [2013, 2014, 2016, 2017, 2018, 2019, 2020, 2021, 2022]

def extract_year(val):
    if isinstance(val, datetime.date):
        return val.year
    elif isinstance(val, int):
        return val
    elif isinstance(val, str):
        try:
            if len(val) == 4 and val.isdigit():
                return int(val)
            else:
                return datetime.datetime.strptime(val, '%Y-%m-%d').year
        except:
            return None
    return None

for idx, year in enumerate(years):
    count = 0
    for row_idx in range(14, 186):
        gval = ws.cell(row=row_idx, column=7).value  # G
        jval = ws.cell(row=row_idx, column=10).value # J
        yr = extract_year(gval)
        if yr == year and isinstance(jval, str) and jval.strip().lower() == 'open':
            count += 1
    # Write to Column I (9), row = 2 + idx
    ws.cell(row=2+idx, column=9, value=count)

wb.save(output_path)
