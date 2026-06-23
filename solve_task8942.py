from openpyxl import load_workbook
from datetime import datetime
from openpyxl.utils.datetime import from_excel

input_path = r"results/runs/g2b-v8_gpt-5.4_v82/eval_100slice_singleseed/task_8942/input.xlsx"
output_path = r"results/runs/g2b-v8_gpt-5.4_v82/eval_100slice_singleseed/task_8942/output.xlsx"

wb = load_workbook(input_path)
overview = wb['Overview']
pay_ws = wb['Pay Dates']

# Build monthly total from Pay Dates based on the month/year selected in Overview!A2.
selected = overview['A2'].value

if isinstance(selected, (int, float)):
    try:
        selected = from_excel(selected)
    except Exception:
        pass

monthly_total = None
if isinstance(selected, datetime):
    target_year = selected.year
    target_month = selected.month
    monthly_total = 0
    for row in range(2, pay_ws.max_row + 1):
        pay_date = pay_ws.cell(row=row, column=1).value
        if isinstance(pay_date, (int, float)):
            try:
                pay_date = from_excel(pay_date)
            except Exception:
                pass
        if isinstance(pay_date, datetime) and pay_date.year == target_year and pay_date.month == target_month:
            for col in (2, 3):
                val = pay_ws.cell(row=row, column=col).value
                if isinstance(val, (int, float)):
                    monthly_total += val
                elif val is None:
                    monthly_total += 0
                else:
                    try:
                        monthly_total += float(val)
                    except Exception:
                        pass

overview['B5'] = monthly_total

wb.save(output_path)

# Verify saved result
check_wb = load_workbook(output_path, data_only=True)
check_ws = check_wb['Overview']
print(check_ws['B5'].value)
