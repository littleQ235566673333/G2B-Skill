from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_5/regression_gate/before_fix/core_56274/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_5/regression_gate/before_fix/core_56274/output.xlsx'

wb = load_workbook(input_path)
ws = wb["Sheet2"]

# Read fiscal month from D7
fiscal_month = ws['D7'].value
if isinstance(fiscal_month, str):
    # Try to parse string to date if present that way
    try:
        fiscal_month = datetime.strptime(fiscal_month, '%d-%b-%Y')
    except Exception:
        # Handle other formats if necessary
        pass

# Get month headers (row 3, columns G to R (col 7 to 18, 0-based))
month_row = 3
min_col = 7
max_col = 18
month_cells = list(ws.iter_cols(min_row=month_row, max_row=month_row, min_col=min_col, max_col=max_col))

month_idx = None
for idx, col in enumerate(month_cells):
    cell_val = col[0].value
    if cell_val == fiscal_month:
        month_idx = idx
        break
if month_idx is None:
    # Try loose matching for string inputs
    for idx, col in enumerate(month_cells):
        cell_val = col[0].value
        if cell_val and fiscal_month and str(cell_val)[:10] == str(fiscal_month)[:10]:
            month_idx = idx
            break
if month_idx is None:
    raise Exception('Fiscal month value not found in month headers.')

# Get values from rows below, same column:
# Row 4: Opening Bal, Row 5: Debits, Row 6: Credits, Row 7: Closing Bal
open_val = ws.cell(row=4, column=min_col+month_idx).value  # G4+idx
Debit_val = ws.cell(row=5, column=min_col+month_idx).value
Credit_val = ws.cell(row=6, column=min_col+month_idx).value
Close_val = ws.cell(row=7, column=min_col+month_idx).value

# Write values to D9:D12
ws['D9'].value = open_val
ws['D10'].value = Debit_val
ws['D11'].value = Credit_val
ws['D12'].value = Close_val

wb.save(output_path)
print('Done')
