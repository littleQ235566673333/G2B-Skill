import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_183-8_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_183-8_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
sheet = wb['Sheet1']

# For J3:J6: Weighted average of %PLF DAY per FY
for row in range(3, 7):  # 3 to 6
    fy_cell = sheet.cell(row=row, column=8).coordinate  # FY in column H
    # Weighted avg for J: col C (Gen, kwh Day), col D (%PLF DAY), col B (FY)
    formula_j = f"=SUMPRODUCT(($B$3:$B$43={fy_cell}),$C$3:$C$43,$D$3:$D$43)/SUMIFS($C$3:$C$43,$B$3:$B$43,{fy_cell})"
    sheet.cell(row=row, column=10).value = formula_j
    # Simple avg for K (%MC Avail): col E
    formula_k = f"=AVERAGEIFS($E$3:$E$43,$B$3:$B$43,{fy_cell})"
    sheet.cell(row=row, column=11).value = formula_k
    # Simple avg for L (%Grid Avail): col F
    formula_l = f"=AVERAGEIFS($F$3:$F$43,$B$3:$B$43,{fy_cell})"
    sheet.cell(row=row, column=12).value = formula_l

wb.save(output_path)
