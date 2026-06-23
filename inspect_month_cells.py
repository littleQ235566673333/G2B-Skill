import openpyxl
from openpyxl import load_workbook
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/group_56274/r0/evolve_56274/input.xlsx')
ws = wb.active
print('D7:', ws['D7'].value, type(ws['D7'].value))
for col in range(7, 11):
    print(f'Col {col} (row 7):', ws.cell(row=7, column=col).value, type(ws.cell(row=7, column=col).value))
