from openpyxl import load_workbook

in_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42/eval_38537_tc1/input.xlsx'
out_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42/eval_38537_tc1/output.xlsx'

wb = load_workbook(in_file)
ws = wb.active

prefix = ws['B2'].value
counter = 10

for row in range(3, 37):
    val = ws[f'B{row}'].value
    if val is not None and str(val).strip() != '':
        num_str = f'{counter:03}'
        ws[f'A{row}'] = f'{prefix}-{num_str}'
        counter += 1
    else:
        ws[f'A{row}'] = None

wb.save(out_file)
