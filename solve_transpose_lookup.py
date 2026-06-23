import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_7/regression_gate/before_fix/core_38985/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_7/regression_gate/before_fix/core_38985/output.xlsx'

# Read the Excel
sheets = pd.read_excel(input_path, sheet_name=None, header=None)
df = sheets['Sheet1']

# Table 1 header: at col 2 (index 2), row 4 (index 4); names in rows 6-9 (index 6-9)
# Table 2 header: at col 10 (index 10), row 4 (index 4); data rows 5-9 (index 5-9)

# Build table 2: Name = col 10, Value 1 = col 11, Value 2 = col 12
# Table 1: names col 2, rows 5-9 (index 5-9)
name_rows = range(5, 9)
table1_name_col = 2
table2_name_col = 10
table2_val1_col = 11
table2_val2_col = 12

# Create a lookup dict for table 2, mapping name to [value1, value2]
table2_lookup = {}
for r in range(5, 10):
    name = df.iat[r, table2_name_col]
    val1 = df.iat[r, table2_val1_col]
    val2 = df.iat[r, table2_val2_col]
    table2_lookup[name] = [val1, val2]

# Openpyxl to write output in D8:D11 (col 4, rows 8-11), and D/E if two values
wb = load_workbook(input_path)
ws = wb['Sheet1']

for i, ws_row in enumerate(range(8, 12)):
    name = ws.cell(row=ws_row, column=3).value
    vals = table2_lookup.get(name, [None, None])
    for j in range(len(vals)):
        ws.cell(row=ws_row, column=4 + j, value=vals[j])

wb.save(output_path)
