import openpyxl
from collections import defaultdict
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_3/group_45707/r0/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_3/group_45707/r0/evolve_45707/output.xlsx'

def parse_date(dt):
    if dt is None:
        return None
    if hasattr(dt, 'year'):
        return dt
    try:
        # Try to parse if it's a string
        return datetime.strptime(str(dt), '%Y-%m-%d')
    except Exception:
        pass
    try:  # Try another common format
        return datetime.strptime(str(dt), '%m/%d/%Y')
    except Exception:
        pass
    return None

wb = openpyxl.load_workbook(input_path)
ws = wb.active

dates = []
values = []
for row in ws.iter_rows(min_row=2, max_row=69, min_col=1, max_col=3):
    dates.append(parse_date(row[0].value))
    values.append(row[2].value)

def is_first_of_month(dt):
    if dt is None:
        return False
    return hasattr(dt, 'day') and dt.day == 1

counts = defaultdict(int)
for dt, val in zip(dates, values):
    if dt is None or not hasattr(dt, 'year'):
        continue
    if val == 1:
        counts[(dt.year, dt.month)] += 1

total_rows = 68
for i in range(total_rows):
    curr_rownum = i + 2
    next_date = dates[i + 1] if i + 1 < len(dates) else None
    dest_cell = ws.cell(row=curr_rownum, column=4)
    if is_first_of_month(next_date):
        key = (next_date.year, next_date.month)
        dest_cell.value = counts.get(key, 0)
    else:
        dest_cell.value = None

wb.save(output_path)
