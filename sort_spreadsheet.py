import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed0/eval_22-47_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed0/eval_22-47_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

data = list(ws.values)

def is_header(row):
    return (
        row and isinstance(row[1], str)
        and row[1].strip().upper() in {'NAME', 'NAMES', 'B'}
    )

def is_empty(row):
    return not any(row)

# Assemble unified data from both tables, skip headers and empties
records = []
for row in data:
    if is_header(row) or is_empty(row):
        continue
    # Only accept rows with both NAME and REF
    if len(row) > 2 and row[1] and row[2]:
        records.append(row)

# Deduplicate by (NAME, REF)
seen = set()
unique_records = []
for r in records:
    key = (r[1], r[2])
    if key in seen:
        continue
    seen.add(key)
    unique_records.append(r)

# Create a lookup for J (col 9): The helper list (J is col 9, header is 'LIST')
helper_names = [row[9] for row in data if row[9] not in (None,'LIST','', ' ')]
helper_order = []
seen_helpers = set()
for name in helper_names:
    if name not in seen_helpers:
        helper_order.append(name)
        seen_helpers.add(name)

# Compose records dict for output
# We set H as None (since no values found), and take only NAME (B:1), REF (C:2)
ready = [{'B': r[1], 'C': r[2], 'H': None} for r in unique_records]

# Our sort: First those matching helper J, then the rest; keep group order; preserve input order within group
if helper_order:
    group1 = [x for h in helper_order for x in ready if x['B'] == h]
    group2 = [x for x in ready if x['B'] not in helper_order]
    ordered = group1 + group2
else:
    ordered = sorted(ready, key=lambda x: (str(x['B']).lower() if x['B'] else ''))

# Only take the first 9 outputs for F2:H10 (since 9 rows)
ordered = ordered[:9]

# H column: all are None; sorting H low-to-high has no effect, but sort it anyway for logic
ordered = sorted(ordered, key=lambda x: (x['H'] is None, x['H']))

# Write to F2:H10
for i, d in enumerate(ordered):
    ws.cell(row=2+i, column=6).value = d['B'] # F
    ws.cell(row=2+i, column=7).value = d['C'] # G
    ws.cell(row=2+i, column=8).value = d['H'] # H

wb.save(output_path)
