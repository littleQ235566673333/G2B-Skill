import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import numpy as np

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v3/train/iter_1/group_45707/r2/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v3/train/iter_1/group_45707/r2/evolve_45707/output.xlsx'

df = pd.read_excel(input_path)
result = [None] * (len(df) - 1)

for idx in range(len(df) - 1):
    date_this = df.iloc[idx, 0]
    date_next = df.iloc[idx + 1, 0]
    if pd.isnull(date_next) or not isinstance(date_next, (datetime, np.datetime64)):
        result[idx] = None
        continue
    # Convert np.datetime64 to datetime if needed
    if isinstance(date_next, np.datetime64):
        date_next = pd.to_datetime(date_next).to_pydatetime()
    if getattr(date_next, 'day', None) == 1:
        y, m = date_next.year, date_next.month
        mask = df.iloc[:, 0].apply(lambda d: pd.notnull(d) and getattr(d, 'year', None) == y and getattr(d, 'month', None) == m)
        count_ones = (df.loc[mask, df.columns[2]] == 1).sum()
        result[idx] = count_ones
    else:
        result[idx] = None
# Last entry (row 69): no next day to check
result.append(None)

wb = load_workbook(input_path)
ws = wb.active
for i in range(2, 70):
    value = result[i - 2] if i - 2 < len(result) else None
    ws[f'D{i}'] = value if value is not None else None
wb.save(output_path)
