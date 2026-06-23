from openpyxl import load_workbook

# Load workbook and sheet
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_4/regression_gate/after_fix/core_50916/input.xlsx')
ws = wb['21-22 Schedule']

# Map cycle day number to its schedule row
cycle_map = {}
for row in range(2, 9):
    day = ws.cell(row=row, column=2).value
    if isinstance(day, int):
        cycle_map[day] = row

# For each cell in C12:H14, fill with class name mapped by cycle day (from row 12)
for r in range(12, 15):
    for c in range(3, 9):
        day_num = ws.cell(row=12, column=c).value
        period = ws.cell(row=1, column=c).value
        if isinstance(day_num, int) and day_num in cycle_map and period is not None:
            sched_row = cycle_map[day_num]
            class_val = ws.cell(row=sched_row, column=c).value
            ws.cell(row=r, column=c).value = class_val
        else:
            ws.cell(row=r, column=c).value = None

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_4/regression_gate/after_fix/core_50916/output.xlsx')
print('DONE')
