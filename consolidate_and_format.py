from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_4/regression_gate/after_pass/core_80-42/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_4/regression_gate/after_pass/core_80-42/output.xlsx'

src_sheets = ['Jack', 'Henry', 'Richard']
target_sheet = 'Consolidate_ALL'
# Columns with yellow highlight in Jack (1-based)
highlighted_columns = [1, 3, 4, 5, 6, 7, 14, 15, 16, 17, 18]
output_cols = list(range(1, 12))  # A-K (1-11) for data, L (12) for sheet name

wb = load_workbook(input_path)
ws_target = wb[target_sheet]

# Find the first empty row in Consolidate_ALL (start at A2)
def find_first_empty_row(ws, start_row=2, col=1, end_row=8000):
    for r in range(start_row, end_row + 1):
        if ws.cell(row=r, column=col).value in (None, ''):
            return r
    return end_row + 1

# Get number formatting, font, border from the first data row in Consolidate_ALL
fmt_cells = [ws_target.cell(row=2, column=col) for col in output_cols]

cur_row = find_first_empty_row(ws_target)

for sheet in src_sheets:
    ws_src = wb[sheet]
    for r in range(2, ws_src.max_row + 1):
        # Check if all highlighted columns are empty
        if all(ws_src.cell(row=r, column=col).value in (None, '') for col in highlighted_columns):
            continue  # skip blank row
        # Extract/prepare the consolidated row
        row_data = [ws_src.cell(row=r, column=col).value for col in highlighted_columns]
        # Write to target with formatting matching sample (but skip fill)
        for idx, val in enumerate(row_data):
            cell = ws_target.cell(row=cur_row, column=idx + 1, value=val)
            # Copy number format, font, border only
            cell.number_format = fmt_cells[idx].number_format
            cell.font = copy(fmt_cells[idx].font)
            cell.border = copy(fmt_cells[idx].border)
            cell.alignment = copy(fmt_cells[idx].alignment)
        # Add sheet name to L
        cell = ws_target.cell(row=cur_row, column=12, value=sheet)
        cell.number_format = fmt_cells[10].number_format
        cell.font = copy(fmt_cells[10].font)
        cell.border = copy(fmt_cells[10].border)
        cell.alignment = copy(fmt_cells[10].alignment)
        cur_row += 1

wb.save(output_path)
print(f'Data consolidated to {output_path}')
