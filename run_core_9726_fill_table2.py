import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/regression_gate/after_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/regression_gate/after_pass/core_9726/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Extract Table 1: columns 1, 3 (A, C): Student ID, Visit date, rows 2~
table1 = []
row = 2
while True:
    sid = ws.cell(row=row, column=1).value
    vdt = ws.cell(row=row, column=3).value
    if sid is None and vdt is None:
        break
    if sid is not None and vdt is not None:
        table1.append((str(sid).strip(), vdt if isinstance(vdt, datetime) else None))
    row += 1

table1_by_student = {}
for sid, vdt in table1:
    if vdt is None:
        continue
    table1_by_student.setdefault(sid, []).append(vdt)
# Sort visit dates for each student
for sid in table1_by_student:
    table1_by_student[sid] = sorted(d for d in table1_by_student[sid] if d is not None)

# Now go through Table 2 (columns 8, 9, 10) and fill in J2+ as requested
row = 2
while True:
    sid2 = ws.cell(row=row, column=8).value
    subm_dt = ws.cell(row=row, column=9).value
    if sid2 is None and subm_dt is None:
        break
    if sid2 is not None and subm_dt is not None:
        sid2_std = str(sid2).strip()
        visit_list = table1_by_student.get(sid2_std, [])
        # Find the latest visit date before submission date
        found = [d for d in visit_list if d < subm_dt]
        if found:
            last_visit = max(found)
            ws.cell(row=row,column=10).value = last_visit.strftime('%d-%b-%y')
        else:
            ws.cell(row=row,column=10).value = None
    row += 1

wb.save(output_path)
