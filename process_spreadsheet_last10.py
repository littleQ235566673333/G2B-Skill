import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_55965_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_55965_tc1/output.xlsx'

# Load workbook and sheets
twb = openpyxl.load_workbook(input_path)

sheet1 = twb['Foglio1']
# Let's guess that the data sheet is named something else; try the second sheet
sheetnames = twb.sheetnames
datasheet_name = [name for name in sheetnames if name != 'Foglio1'][0]
datasheet = twb[datasheet_name]

# Columns
sheet1_id_col = 'E'  # ID in Sheet1
sheet1_start_row = 2
sheet1_end_row = 18
sheet1_result_cols = list('GHIJKLMNOP')

# Data sheet: ID in column K (11th, openpyxl is 1-based)
data_id_col_idx = 11
# Data sheet results: G-P (7th to 16th columns)
data_result_cols = list(range(7, 17))

for row in range(sheet1_start_row, sheet1_end_row+1):
    sheet1_id = sheet1[f'{sheet1_id_col}{row}'].value
    if sheet1_id is None:
        continue
    # Find rows in datasheet with matching ID in col K
    matching_rows = []
    for drow in range(2, datasheet.max_row+1):  # skip header
        data_id = datasheet.cell(row=drow, column=data_id_col_idx).value
        if data_id == sheet1_id:
            result_values = [datasheet.cell(row=drow, column=col).value for col in data_result_cols]
            matching_rows.append(result_values)
    # Get last 10
    last_10 = matching_rows[-10:] if len(matching_rows) >= 10 else matching_rows
    # Place in sheet1 G-P, G == first result, P == 10th result
    for i, values in enumerate(last_10):
        for j, v in enumerate(values):
            target_col = sheet1_result_cols[j]
            sheet1[f'{target_col}{row}'] = v if i == 0 else None  # Only show the most recent result, as G2:G18 requested

# Save

# Actually, the instruction says to write answer to G2:G18, so only the latest result per ID
for row in range(sheet1_start_row, sheet1_end_row+1):
    sheet1_id = sheet1[f'{sheet1_id_col}{row}'].value
    if sheet1_id is None:
        continue
    # Find rows in datasheet with matching ID in col K
    matching_rows = []
    for drow in range(2, datasheet.max_row+1):
        data_id = datasheet.cell(row=drow, column=data_id_col_idx).value
        if data_id == sheet1_id:
            result_values = [datasheet.cell(row=drow, column=col).value for col in data_result_cols]
            matching_rows.append(result_values)
    if matching_rows:
        latest = matching_rows[-1][0]  # only first result (G)
        sheet1[f'G{row}'] = latest
    else:
        sheet1[f'G{row}'] = None

twb.save(output_path)
