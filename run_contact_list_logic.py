from openpyxl import load_workbook
from datetime import datetime

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_5/regression_gate/before_pass/core_41589/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_5/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(in_path)
ws = wb['Contact List']
today = datetime.today().date()

for row in ws.iter_rows(min_row=4, min_col=8, max_col=10):
    h, i, j = row
    result = 'NO ACTION'
    if h.value and i.value and str(i.value).strip().upper() == 'YES':
        try:
            last_dt = h.value.date() if isinstance(h.value, datetime) else h.value
            diff = (today - last_dt).days
            if diff <= 30:
                result = 'HOLD'
            elif diff > 30:
                result = 'TOUCH BASE'
        except Exception:
            pass
    ws.cell(row=h.row, column=10).value = result

wb.save(out_path)
