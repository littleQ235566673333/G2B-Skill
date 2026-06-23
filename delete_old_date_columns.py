from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_534-26_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_534-26_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet1']
now = datetime.now()
this_month = datetime(now.year, now.month, 1)
del_cols = []
for idx, cell in enumerate(ws[1], 1):
    if isinstance(cell.value, datetime):
        if cell.value < this_month:
            del_cols.append(idx)
for idx in sorted(del_cols, reverse=True):
    ws.delete_cols(idx)
wb.save(output_path)
