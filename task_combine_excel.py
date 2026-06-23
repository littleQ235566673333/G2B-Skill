import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy

# File paths
data_in = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_6/group_130-9/r3/evolve_130-9/input.xlsx'
data_out = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_6/group_130-9/r3/evolve_130-9/output.xlsx'

# Sheet names
source_sheet = 'cdnr'
dest_sheet = 'b2b, sez, de'

wb = openpyxl.load_workbook(data_in)
ws_src = wb[source_sheet]
ws_dst = wb[dest_sheet]

# Get header row (assume it's row 5 for both sheets)
header_row_src = 5
header_row_dst = 5

header_src = [ws_src.cell(row=header_row_src, column=col).value for col in range(1, ws_src.max_column + 1)]
header_dst = [ws_dst.cell(row=header_row_dst, column=col).value for col in range(1, ws_dst.max_column + 1)]

# Dictionary: header -> column index
src_header_map = {h: idx+1 for idx, h in enumerate(header_src) if h}
dst_header_map = {h: idx+1 for idx, h in enumerate(header_dst) if h}

# Find the first empty row in destination (starting at row 7)
start_row = 7
while any(ws_dst.cell(row=start_row, column=col).value for col in range(1, ws_dst.max_column + 1)):
    start_row += 1

# Identify source data rows (below header row)
# Get data from cdnr rows 6 to ws_src.max_row
src_data_rows = [row for row in ws_src.iter_rows(min_row=header_row_src+1, max_row=ws_src.max_row, values_only=False)]

for row_data in src_data_rows:
    # Prepare new row for destination
    new_row_idx = start_row
    for hdr, dst_col in dst_header_map.items():
        if hdr in src_header_map:
            src_col = src_header_map[hdr]
            src_cell = ws_src.cell(row=row_data[0].row, column=src_col)
            dst_cell = ws_dst.cell(row=new_row_idx, column=dst_col)
            value = src_cell.value
            # Columns L & M: -ve value
            if get_column_letter(dst_col) in ['L', 'M']:
                if isinstance(value, (int, float)):
                    value = -abs(value)
            # Preserve number format
            dst_cell.number_format = ws_dst.cell(row=header_row_dst+1, column=dst_col).number_format
            dst_cell.value = value
        else:
            ws_dst.cell(row=new_row_idx, column=dst_col).value = None
    # Set column H blank and column I "Credit Note"
    col_H = dst_header_map.get('H', 8)
    col_I = dst_header_map.get('I', 9)
    ws_dst.cell(row=new_row_idx, column=col_H).value = None
    ws_dst.cell(row=new_row_idx, column=col_I).value = 'Credit Note'
    # Copy cell styles (from prior row if applicable)
    for dst_col in range(1, ws_dst.max_column + 1):
        prior_row = header_row_dst+1
        prior_cell = ws_dst.cell(row=prior_row, column=dst_col)
        new_cell = ws_dst.cell(row=new_row_idx, column=dst_col)
        new_cell.font = copy(prior_cell.font)
        new_cell.alignment = copy(prior_cell.alignment)
        new_cell.border = copy(prior_cell.border)
        new_cell.fill = copy(prior_cell.fill)
        new_cell.number_format = prior_cell.number_format
    start_row += 1

wb.save(data_out)