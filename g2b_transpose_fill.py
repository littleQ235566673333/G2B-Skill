import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun1/eval_341-14_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun1/eval_341-14_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
problem_ws = wb['problem']
result_ws = wb['result']

# Build a mapping from the problem sheet: name -> number
problem_dict = {}
for row in problem_ws.iter_rows(min_row=1, max_row=problem_ws.max_row, min_col=1, max_col=2, values_only=True):
    name, number = row
    if name:  # skip blank names
        problem_dict[name] = number

# Now fill result!B1:B28 according to result!A1:A28
for i in range(1, 29):  # rows 1 to 28 inclusive
    name = result_ws.cell(row=i, column=1).value
    if name is not None and name in problem_dict:
        result_ws.cell(row=i, column=2, value=problem_dict[name])
    else:
        result_ws.cell(row=i, column=2, value=None)  # blank if no match

wb.save(output_path)
