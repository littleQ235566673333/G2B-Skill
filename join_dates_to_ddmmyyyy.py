import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

zord = pd.read_excel('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/regression_gate/after_fix/core_45896/input.xlsx', sheet_name='ZORD')
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/regression_gate/after_fix/core_45896/input.xlsx')
ws = wb['Volym P5_P6_2023']
header = [cell.value for cell in ws[1]]
mat_idx = header.index('Material')+1
for excel_row in range(2, 11):
    mat_value = ws.cell(row=excel_row, column=mat_idx).value
    dates = zord.loc[zord['Material'] == mat_value, 'Valid from'].unique()
    formatted_dates = []
    for d in dates:
        try:
            dt = pd.to_datetime(d)
            formatted_dates.append(dt.strftime('%d/%m/%Y'))
        except Exception:
            pass
    ws.cell(row=excel_row, column=9).value = ', '.join(formatted_dates)
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/regression_gate/after_fix/core_45896/output.xlsx')
