from openpyxl import load_workbook
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/group_45896/r2/evolve_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/group_45896/r2/evolve_45896/output.xlsx'

wb = load_workbook(input_path)
ws_main = wb['Volym P5_P6_2023']
ws_zord = wb['ZORD']

# Columns
mat_col = 1 # 'Material' column (A)
out_col = 9 # Column I (0-indexed, == 9)

# Build data from ZORD
zord_data = list(ws_zord.iter_rows(min_row=2, values_only=True))
def fmt(d):
    return d.strftime('%d/%m/%Y') if isinstance(d, (datetime.date, datetime.datetime)) else str(d)

for r in range(2, 11):  # I2:I10
    mat = ws_main.cell(r, mat_col).value
    # Get all 'Valid from' for matching material
    dates = [row[1] for row in zord_data if row[0] == mat]
    # Remove duplicates, keep order
    unique_dates = []
    seen = set()
    for d in dates:
        if d not in seen:
            unique_dates.append(d)
            seen.add(d)
    # Format
    dates_str = ','.join([fmt(d) for d in unique_dates if d])
    ws_main.cell(r, out_col + 1).value = dates_str if unique_dates else ''

wb.save(output_path)
