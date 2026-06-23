from openpyxl import load_workbook
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_5/group_57113/r0/evolve_57113/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_5/group_57113/r0/evolve_57113/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]
rows_to_process = [2, 3]

def extract_hashtags(tweet):
    if not tweet: return []
    return [tag[1:] for tag in re.findall(r"#\\w+", tweet)]

for row in rows_to_process:
    tweet = ws[f"A{row}"].value
    hashtags = extract_hashtags(tweet)
    for idx, tag in enumerate(hashtags[:2]):  # Only fill B & C
        ws.cell(row=row, column=2+idx, value=tag)
    # Fill empty if less than 2 hashtags
    for idx in range(len(hashtags), 2):
        ws.cell(row=row, column=2+idx, value=None)

wb.save(output_path)
print('Done')
