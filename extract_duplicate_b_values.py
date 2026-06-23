import pandas as pd
from openpyxl import load_workbook

# Load the spreadsheet
wb = load_workbook('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_50971_tc1/input.xlsx')
ws = wb.active

# Read worksheet into DataFrame
rows = list(ws.values)
header = rows[0]
data = rows[1:]
df = pd.DataFrame(data, columns=header)

# Identify duplicate values in Column A
duplicates = df[df.duplicated(subset=[header[0]], keep=False)]
# Group by Column A and get associated values from Column B
results = duplicates.groupby(header[0])[header[1]].apply(list)

# To match the output layout, flatten values
values = []
for group in results:
    values.extend(group)

# Write result to G3:G13
for i, value in enumerate(values):
    ws.cell(row=3+i, column=7, value=value)

wb.save('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_50971_tc1/output.xlsx')
