import openpyxl
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_51262_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_51262_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

date_ref = ws['F5'].value
criteria = [ws[f'F{i}'].value for i in range(10, 15)]

# Only keep valid criteria
criteria = [c for c in criteria if c is not None]

# Date boundaries for August 2021 if ref is 2021-08-01
start_date = datetime(date_ref.year, date_ref.month, 1)
end_date = datetime(date_ref.year, date_ref.month, 31)  # Always 31: Excel's SUMIFS uses <= end

# Load data rows
values = []
for row in ws.iter_rows(min_row=6, max_row=13, min_col=2, max_col=4):
    date_cell, category_cell, value_cell = row
    if type(date_cell.value) is datetime and type(value_cell.value) in [int, float]:
        values.append((date_cell.value, category_cell.value, value_cell.value))

# Sum for each criterion
def sum_by_criterion(criteria):
    results = []
    for crit in criteria:
        s = sum(
            v for d, c, v in values
            if (start_date <= d <= end_date) and (c == crit)
        )
        results.append(s)
    return results

sums = sum_by_criterion(criteria)
# Place results starting from F6
for idx, s in enumerate(sums):
    ws.cell(row=6, column=6+idx, value=s)

# Fill unused output cells in F6:H6 with empty or zero
for i in range(len(sums), 3):
    ws.cell(row=6, column=6+i, value=None)

wb.save(output_path)
