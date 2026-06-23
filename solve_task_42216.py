from openpyxl import load_workbook
from datetime import datetime
import calendar

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_42216_s1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_42216_s1/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

value_by_date = {}

for row in range(3, 14):
    ym = ws.cell(row=row, column=1).value
    if not ym:
        continue
    year_str, month_str = str(ym).split('-')
    year = int(year_str)
    month = int(month_str)
    last_day = calendar.monthrange(year, month)[1]
    for col in range(2, 33):
        day = ws.cell(row=2, column=col).value
        if day in (None, ''):
            continue
        try:
            day = int(day)
        except Exception:
            continue
        if day > last_day:
            continue
        val = ws.cell(row=row, column=col).value
        value_by_date[datetime(year, month, day).date()] = val

for row in range(19, 340):
    dt = ws.cell(row=row, column=1).value
    if dt is None:
        ws.cell(row=row, column=2).value = None
        continue
    d = dt.date() if hasattr(dt, 'date') else dt
    val = value_by_date.get(d, None)

    if d.month == 4 and d not in value_by_date:
        out = 0
    elif val in (None, ''):
        out = 0
    elif isinstance(val, str) and val.strip().upper() == 'NA':
        out = None
    elif isinstance(val, str) and val.strip().upper() == 'M':
        out = None
    elif val in (-9999, -99999, '-9999', '-99999'):
        out = None
    else:
        out = val

    ws.cell(row=row, column=2).value = out

wb.save(output_path)

wb2 = load_workbook(output_path)
ws2 = wb2[wb2.sheetnames[0]]
for r in [19,20,22,24,48,49,50,77,78,108,109,138,139,169,170,200,230,260,261]:
    print(r, ws2.cell(r,1).value, ws2.cell(r,2).value)
