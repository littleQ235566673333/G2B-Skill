from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import datetime

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_1/group_44017/r0/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_1/group_44017/r0/evolve_44017/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Columns
baseline_col = 23  # W
freq_col = 10      # J
eff_date_col = 12  # L
inc_cols = [13, 14, 15, 16]  # M, N, O, P
start_col = 30    # AD (30) ... AO (41)
end_col = 41
row_start = 14
row_end = 42
months_row = 9

# Pre-read the dates from row 9 (columns AD:AO)
month_dates = []
for col in range(start_col, end_col + 1):
    cell = ws.cell(row=months_row, column=col)
    val = cell.value
    if isinstance(val, datetime.datetime) or isinstance(val, datetime.date):
        month_dates.append(val)
    elif val:
        try:
            month_dates.append(datetime.datetime.strptime(str(val), '%Y-%m-%d').date())
        except Exception:
            month_dates.append(None)
    else:
        month_dates.append(None)

no_fill = PatternFill(fill_type=None)

for row in range(row_start, row_end + 1):
    base = ws.cell(row=row, column=baseline_col).coordinate
    freq = ws.cell(row=row, column=freq_col).coordinate
    eff_date = ws.cell(row=row, column=eff_date_col).coordinate
    inc_cells = [ws.cell(row=row, column=col).coordinate for col in inc_cols]
    for idx, col in enumerate(range(start_col, end_col + 1)):
        target_cell = ws.cell(row=row, column=col)
        col_letter = get_column_letter(col)
        formula = (
            f"=IF(AND(ISNUMBER({eff_date}),{col_letter}{months_row}>={eff_date}),"
            f"{base}*"
            f"IF({col_letter}{months_row}>=EDATE({eff_date},{freq}*0),1+{inc_cells[0]},1)"
            f"*IF({col_letter}{months_row}>=EDATE({eff_date},{freq}*1),1+{inc_cells[1]},1)"
            f"*IF({col_letter}{months_row}>=EDATE({eff_date},{freq}*2),1+{inc_cells[2]},1)"
            f"*IF({col_letter}{months_row}>=EDATE({eff_date},{freq}*3),1+{inc_cells[3]},1)"
            f',"","")'
            f')'
        )
        target_cell.value = formula
        target_cell.fill = no_fill

wb.save(output_path)
