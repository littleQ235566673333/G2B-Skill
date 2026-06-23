import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_262-17_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_262-17_tc1/output.xlsx'

# Load the workbook and select Sheet1
wb = openpyxl.load_workbook(input_path)
sheet = wb['Sheet1']

# Find header row and column indices for 'Task' and 'Responsibility'
header = [cell.value for cell in sheet[1]]
task_idx = header.index('Task')  # zero-based
resp_idx = header.index('Responsibility')  # zero-based

# Extract data rows
rows = []
for row in sheet.iter_rows(min_row=2, max_row=14, values_only=True):
    rows.append(row)

# Sort rows: Task asc, Responsibility asc
sorted_rows = sorted(
    rows,
    key=lambda r: (r[task_idx], r[resp_idx])
)

# Overwrite data in Sheet1
for i, row in enumerate([header] + sorted_rows, start=1):
    for j, value in enumerate(row, start=1):
        sheet.cell(row=i, column=j, value=value)

# Save to output
wb.save(output_path)
