import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_120-24_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_120-24_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Columns indexes: AY=51, BG=59, BL=64, BN=66 (A=1)
col_idx = {
    'AY': 51,
    'BG': 59,
    'BL': 64,
    'BN': 66
}

# Fill color (hex #0070C0)
blue_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')

for row in range(2, 46):
    bl_val = ws.cell(row=row, column=col_idx['BL']).value
    bg_val = ws.cell(row=row, column=col_idx['BG']).value
    ay_val = ws.cell(row=row, column=col_idx['AY']).value

    bn_new_val = None
    if bl_val == 'LAG' and bg_val in [
        'OCOGS - Spares - Transfer Price Overhead',
        'OFSS - ORCL Consulting Prod Cost I/C'
    ]:
        if ay_val in ['BOA_033E', 'BOA_011G']:
            bn_new_val = 'OCOGS - Spares - Transfer Price Overhead'
        else:
            bn_new_val = 'OFSS - ORCL Consulting Prod Cost I/C'

    if bn_new_val is not None:
        ws.cell(row=row, column=col_idx['BN']).value = bn_new_val

    # Fill color in BL if criteria met
    bn_val = ws.cell(row=row, column=col_idx['BN']).value
    if bl_val == 'LAG' and bn_val:
        ws.cell(row=row, column=col_idx['BL']).fill = blue_fill

wb.save(output_path)
