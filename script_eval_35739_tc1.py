from openpyxl import load_workbook
from datetime import datetime, timedelta, time

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42/eval_35739_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42/eval_35739_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet1']

def no_data(cell):
    return cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == '')

def subtract_30_minutes(std_time):
    # std_time: datetime.time or datetime.datetime or string 'HH:MM'
    if isinstance(std_time, datetime):
        base_time = std_time.time()
    elif isinstance(std_time, time):
        base_time = std_time
    elif isinstance(std_time, str):
        try:
            base_time = datetime.strptime(std_time.strip(), '%H:%M').time()
        except Exception:
            return None
    else:
        return None
    # Handle 30 min subtraction and midnight wrap
    dt = datetime.combine(datetime(2000,1,1), base_time)
    dt_minus_30 = dt - timedelta(minutes=30)
    # Return as 'HH:MM'
    return dt_minus_30.time().strftime('%H:%M')

for row in range(2, 101):
    cell_a = ws[f'A{row}']
    cell_b = ws[f'B{row}']
    cell_c = ws[f'C{row}']
    # Blank if no data in A and B
    if no_data(cell_a) and no_data(cell_b):
        cell_c.value = None
        continue
    # If only B missing, also blank
    if no_data(cell_b):
        cell_c.value = None
        continue
    # Main logic
    result = subtract_30_minutes(cell_b.value)
    cell_c.value = result

wb.save(output_path)
