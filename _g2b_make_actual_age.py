import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_2/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_2/regression_gate/after_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

row_start = 3
row_end = 15  # inclusive, based on your data
col_expected_result = 5          # E
col_category = 10                # J
col_dob = 3                      # C
col_new_age = 15                 # O

# Read report date
report_date = ws.cell(row=1, column=2).value

# Fill Expected Result (E3:E15) with CATEGORY (J3:J15)
for r in range(row_start, row_end+1):
    ws.cell(row=r, column=col_expected_result).value = ws.cell(row=r, column=col_category).value

# Actual Age: bold, center, header top align, integer only, leave fill as default
for r in range(row_start, row_end+1):
    dob = ws.cell(row=r, column=col_dob).value
    age = None
    if report_date is not None and dob is not None:
        if isinstance(report_date, datetime) and isinstance(dob, datetime):
            age = (report_date - dob).days // 365
    cell = ws.cell(row=r, column=col_new_age)
    cell.value = age
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

header_cell = ws.cell(row=2, column=col_new_age)
header_cell.font = Font(bold=True)
header_cell.alignment = Alignment(horizontal='center', vertical='top')

wb.save(output_path)
