import openpyxl
from openpyxl import load_workbook

# Input and output paths
ingest_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_22-47_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_22-47_tc1/output.xlsx"

wb = load_workbook(ingest_path)
ws = wb.active

data = []
seen_pairs = set()
for row in ws.iter_rows(min_row=2):
    val_b = row[1].value
    val_c = row[2].value
    val_j = row[9].value
    if val_b and val_c:
        pair = (val_b, val_c)
        if pair not in seen_pairs:
            seen_pairs.add(pair)
            data.append({'row': row, 'B': val_b, 'C': val_c, 'J': val_j})

j_vals = []
j_seen = set()
for d in data:
    if d['J'] and (d['J'] not in j_seen):
        j_vals.append(d['J'])
        j_seen.add(d['J'])

j_group = []
other_group = []
if j_vals:
    for name in j_vals:
        for d in data:
            if d['B'] == name:
                j_group.append(d)
    for d in data:
        if d['B'] not in j_vals:
            other_group.append(d)
    sorted_data = j_group + other_group
else:
    sorted_data = sorted(data, key=lambda x: (str(x['B'])))

output_rows = []
for d in sorted_data:
    output_rows.append([d['row'][5].value, d['B'], d['C']])

output_rows = output_rows[:9]

# Always sort column H by string comparison for consistency
output_rows = sorted(output_rows, key=lambda x: str(x[2]))

for i, vals in enumerate(output_rows):
    for j, val in enumerate(vals):
        ws.cell(row=2+i, column=6+j, value=val)

wb.save(output_path)
