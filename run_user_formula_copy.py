import openpyxl

# Load workbook and worksheet
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_55060_tc1/input.xlsx')
ws = wb.active

# Read value from I12
value = ws['I12'].value

# Only write to non-merged cells in J23:N23
for col in range(10, 15):  # J (10) through N (14)
    cell = ws.cell(row=23, column=col)
    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
        cell.value = value if value not in (None, '') else ''

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_55060_tc1/output.xlsx')
