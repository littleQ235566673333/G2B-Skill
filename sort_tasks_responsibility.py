import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_262-17_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_262-17_tc1/output.xlsx'
sheet_name = 'Sheet1'
output_range = (1, 1, 14, 6)  # (min_row, min_col, max_row, max_col) for Sheet1!A1:F14

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb[sheet_name]

# Find header row and column indices for Task and Responsibility
def find_header_indices(ws, headers):
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        hdr_idx = {h: row.index(h) if h in row else None for h in headers}
        return hdr_idx

header_names = ['Task', 'Responsibility']
header_indices = find_header_indices(ws, header_names)

# Ensure headers exist
if None in header_indices.values():
    raise ValueError(f"Could not find all headers. Found: {header_indices}")

task_col = header_indices['Task']
resp_col = header_indices['Responsibility']

# Read all data rows, keep headers
all_rows = list(ws.iter_rows(min_row=1, max_row=14, max_col=6, values_only=True))
header = all_rows[0]
data_rows = all_rows[1:]

# Sort rows by Task, then Responsibility (both ascending)
data_rows_sorted = sorted(
    data_rows,
    key=lambda r: (
        r[task_col],
        r[resp_col]
    )
)

# Write sorted data back to ws
for i, row in enumerate([header] + data_rows_sorted, start=1):
    for j, value in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=value)

# Save to output_path
wb.save(output_path)
