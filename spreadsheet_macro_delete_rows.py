import openpyxl
from openpyxl.styles import PatternFill

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_141-20_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_141-20_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_file)
pl_sheet = wb['PL Recon Items']
st_sheet = wb['Statement Recon Items']

# Collect yellow highlighted (Invoice, Value) from PL Recon Items
pl_yellow = set()
pl_rows = {}
for row in range(2, pl_sheet.max_row + 1):
    invoice = pl_sheet[f'C{row}'].value
    value = pl_sheet[f'D{row}'].value
    fill = pl_sheet[f'A{row}'].fill
    # openpyxl 3.0+ for xlsx, "solid" pattern fill, color in rgb
    if fill is not None and fill.patternType == 'solid' and (
        fill.start_color.rgb == 'FFFFFF00' or fill.start_color.index == 'FFFF00'):
        pl_yellow.add((invoice, value))
        pl_rows[(invoice, value)] = row

# Collect yellow highlighted (Reference, Value) from Statement Recon Items
st_yellow = set()
st_rows = {}
for row in range(2, st_sheet.max_row + 1):
    reference = st_sheet[f'F{row}'].value
    value = st_sheet[f'I{row}'].value
    fill = st_sheet[f'A{row}'].fill
    if fill is not None and fill.patternType == 'solid' and (
        fill.start_color.rgb == 'FFFFFF00' or fill.start_color.index == 'FFFF00'):
        st_yellow.add((reference, value))
        st_rows[(reference, value)] = row

# Compute intersections
common = pl_yellow & st_yellow

# Rows to delete in both sheets
pl_rows_to_delete = [pl_rows[k] for k in common]
st_rows_to_delete = [st_rows[k] for k in common]

# Sort and delete from bottom up
for row in sorted(pl_rows_to_delete, reverse=True):
    pl_sheet.delete_rows(row, 1)
for row in sorted(st_rows_to_delete, reverse=True):
    st_sheet.delete_rows(row, 1)

wb.save(output_file)
