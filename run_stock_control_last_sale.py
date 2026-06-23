import openpyxl
from openpyxl.styles import numbers
from datetime import datetime

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_9448_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_9448_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_file)
ws = wb['Data']

# I-T => columns 9 to 20 (1-based), i.e., 8 to 19 (0-based)
sale_cols = list(range(9, 21))

for row in range(9, 19):  # rows 9 to 18 inclusive
    last_sale_col = None
    for col in sale_cols:
        val = ws.cell(row=row, column=col).value
        if isinstance(val, (int, float)) and val != 0:
            last_sale_col = col
    if last_sale_col is not None:
        # Get header from row 8
        header_value = ws.cell(row=8, column=last_sale_col).value
        # Try parsing header as a date
        date_obj = None
        if isinstance(header_value, datetime):
            date_obj = header_value
        else:
            for fmt in ['%B %Y', '%b %Y', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d']:
                try:
                    date_obj = datetime.strptime(str(header_value), fmt)
                    break
                except Exception:
                    continue
        if date_obj:
            first_of_month = datetime(date_obj.year, date_obj.month, 1)
            cell = ws.cell(row=row, column=21)  # U column
            cell.value = first_of_month
            cell.number_format = numbers.FORMAT_DATE_XLSX14
        else:
            ws.cell(row=row, column=21).value = header_value
    else:
        ws.cell(row=row, column=21).value = None

wb.save(output_file)
