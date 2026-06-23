import openpyxl
import re

# File paths
target_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_40892_tc1/output.xlsx"
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_40892_tc1/input.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read list of allowed colors from column D (ignore header, so D2 downwards)
color_list = []
for cell in ws['D'][1:]:  # Skipping D1 header
    value = cell.value
    if value:
        color_list.append(str(value).strip().lower())

# Prepare regex patterns for colors (ensure exact word match)
color_patterns = [(color, re.compile(r'\\b' + re.escape(color) + r'\\b', re.IGNORECASE)) for color in color_list]

# Apply formula logic to each relevant cell in column A and write to B
for row in range(2, 18):  # B2:B17
    desc = ws[f"A{row}"].value
    found_color = ""
    if desc:
        desc_lower = desc.lower()
        for color, pat in color_patterns:
            if pat.search(desc_lower):
                found_color = color.title()
                break
    ws[f"B{row}"] = found_color

# Save to output file
wb.save(target_path)
