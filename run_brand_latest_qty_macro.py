import pandas as pd
from openpyxl import load_workbook

# Paths and settings
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/group_516-46/r2/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/group_516-46/r2/evolve_516-46/output.xlsx'
sheet_name = 'ورقة1'

# Read the used range, infer rows by nonempty in A
wb = load_workbook(input_path)
ws = wb[sheet_name]
rows = []
for row in ws.iter_rows(min_row=2, min_col=1, max_col=5, values_only=True):
    if all(cell is None for cell in row):
        break
    rows.append(row)

if not rows:
    wb.save(output_path)
    exit()

df = pd.DataFrame(rows)
columns = ['A', 'B', 'C', 'D', 'E']
df.columns = columns

# Assume brand is in column B, date in column C, quantity in column E
brand_col = 'B'
date_col = 'C'
qty_col = 'E'

# Ensure date parsing
df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

# Find the last date per brand
df_last_dates = df.groupby(brand_col)[date_col].max().reset_index()
# Merge to get all rows with the last date for each brand
last_rows = pd.merge(df, df_last_dates, left_on=[brand_col, date_col], right_on=[brand_col, date_col], how='inner')

# Sum qty for each brand + last date
grouped = last_rows.groupby([brand_col, date_col], as_index=False).agg({qty_col: 'sum'})

# (Optional): Note brands that had duplicates
modifications = last_rows.groupby([brand_col, date_col]).size().reset_index(name='count')
modifications = modifications[modifications['count'] > 1]

# Prepare for writing: columns H:L = [brand, date, qty summed, (mod flag/count)]
out_data = []
for _, row in grouped.iterrows():
    mod_row = modifications[(modifications[brand_col] == row[brand_col]) & (modifications[date_col] == row[date_col])]
    mod_flag = ''
    if not mod_row.empty:
        mod_flag = f"MODIFIED (+{mod_row['count'].values[0]} total entries)"
    out_data.append([row[brand_col], row[date_col].strftime('%Y-%m-%d') if pd.notnull(row[date_col]) else '', row[qty_col], mod_flag])

# Write to output Excel
for i, data_row in enumerate(out_data):
    for j, val in enumerate(data_row):
        ws.cell(row=2+i, column=8+j, value=val)  # H=8

wb.save(output_path)
