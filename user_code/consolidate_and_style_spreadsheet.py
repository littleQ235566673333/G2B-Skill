from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill
import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_2/regression_gate/before_pass/core_80-42/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_2/regression_gate/before_pass/core_80-42/output.xlsx'

# Main sheets to consolidate
sheets_to_consolidate = ['Jack', 'Henry', 'Richard']
consolidate_sheet = 'Consolidate_ALL'

wb = load_workbook(input_path)
ws_dest = wb[consolidate_sheet]

# Get dest cell formatting (from first data row in Consolidate_ALL)
format_row = list(ws_dest.iter_rows(min_row=2, max_row=2, max_col=12))[0]
formatting = []
for cell in format_row:
    formatting.append({
        'number_format': cell.number_format,
        'font': copy.copy(cell.font),
        'alignment': copy.copy(cell.alignment),
        'border': copy.copy(cell.border),
        'fill': copy.copy(cell.fill)
    })

# Find first available blank row in Consolidate_ALL (after header)
first_row = 2
for row in ws_dest.iter_rows(min_row=2, max_col=1):
    if row[0].value is None:
        break
    first_row += 1

dest_row = first_row

for sheet in sheets_to_consolidate:
    ws_src = wb[sheet]
    # Detect which columns are yellow highlighted in header (row 1)
    header_row = list(ws_src.iter_rows(min_row=1, max_row=1, max_col=ws_src.max_column))[0]
    yellow_idxs = [i for i, cell in enumerate(header_row) if cell.fill.patternType and getattr(cell.fill.fgColor, 'rgb', None) == 'FFFFFF00']
    # Only consolidate if at least something is selected
    if not yellow_idxs:
        continue
    # For each data row in this sheet after header
    for row in ws_src.iter_rows(min_row=2, max_col=ws_src.max_column):
        # Don't copy empty rows
        if all(cell.value is None for cell in row):
            continue
        # Extract selected columns
        selected_values = [row[i].value for i in yellow_idxs]
        # Fill up to 11 columns, pad with None if required
        dest_values = selected_values + [None]*(11-len(selected_values)) if len(selected_values) < 11 else selected_values[:11]
        # Set cell values in Consolidate_ALL, columns A:K (1-11)
        for col, value in enumerate(dest_values, 1):
            dest_cell = ws_dest.cell(row=dest_row, column=col, value=value)
            # copy style
            fmt = formatting[col-1] if col-1 < len(formatting) else formatting[-1]
            dest_cell.number_format = fmt['number_format']
            dest_cell.font = copy.copy(fmt['font'])
            dest_cell.alignment = copy.copy(fmt['alignment'])
            dest_cell.border = copy.copy(fmt['border'])
            dest_cell.fill = copy.copy(fmt['fill'])
        # Set sheet name in column L
        colL_cell = ws_dest.cell(row=dest_row, column=12, value=sheet)
        fmt = formatting[11] if len(formatting) > 11 else formatting[-1]
        colL_cell.number_format = fmt['number_format']
        colL_cell.font = copy.copy(fmt['font'])
        colL_cell.alignment = copy.copy(fmt['alignment'])
        colL_cell.border = copy.copy(fmt['border'])
        colL_cell.fill = copy.copy(fmt['fill'])
        dest_row += 1

wb.save(output_path)
