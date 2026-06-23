import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_2/regression_gate/before_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_2/regression_gate/before_pass/core_9726/output.xlsx'

# Read all data
sheet = pd.read_excel(input_path, sheet_name='Sheet1')

# Ensure dates are parsed
sheet['Visit date'] = pd.to_datetime(sheet['Visit date'], errors='coerce')
sheet['Submission date'] = pd.to_datetime(sheet['Submission date'], errors='coerce')

# Go row by row, find for each non-null 'Submission date' the last 'Visit date' for the same student before submission
def get_last_visit(row, df):
    if pd.isnull(row['Submission date']) or pd.isnull(row['Student ID.1']):
        return pd.NaT
    student = row['Student ID.1']
    sub_date = row['Submission date']
    # Find all visits for that student
    mask = (df['Student ID'] == student) & (df['Visit date'] < sub_date)
    visits = df.loc[mask, 'Visit date']
    if not visits.empty:
        return visits.max()
    return pd.NaT

sheet['Visit date.1'] = sheet.apply(lambda row: get_last_visit(row, sheet), axis=1)

# Write result to output.xlsx in same structure, so 'Visit date.1' lands in column J
wb = load_workbook(input_path)
ws = wb['Sheet1']
# Overwrite the values in column J (index 10, 1-based)
for i, val in enumerate(sheet['Visit date.1'], start=2):   # Data begins at row 2
    ws.cell(row=i, column=10, value=val.to_pydatetime() if pd.notnull(val) else None)
wb.save(output_path)
