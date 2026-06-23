import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils.cell import get_column_letter
import re

def is_row_empty(row):
    return all([cell.value is None or str(cell.value).strip() == '' for cell in row])

def is_formula(cell):
    return cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('='))

def get_real_formula(cell):
    if isinstance(cell.value, str) and cell.value.startswith('='):
        return cell.value
    return None

def apply_number_format(cell):
    # Applies: round floats to 1 decimal, whole numbers as integer (no .0), others General
    v = cell.value
    if v is None:
        cell.number_format = 'General'
        return
    if is_formula(cell):
        # number format - let Excel display it as number, 1 decimal
        cell.number_format = '#,##0.0'
    else:
        try:
            val = float(v)
            if val == int(val):
                cell.number_format = '#,##0'
            else:
                cell.number_format = '#,##0.0'
        except:
            cell.number_format = 'General'

def make_bold_center(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

def sort_block(sheet, start, end):
    rows = []
    for r in sheet.iter_rows(min_row=start, max_row=end, max_col=8):
        rows.append([c for c in r])
    # Sort by evaluated numeric D
    data = []
    for r in rows:
        d = r[3]
        v = None
        if is_formula(d):
            v = d.value if isinstance(d.value, str) else d.value
            try:
                v = d.internal_value if hasattr(d, 'internal_value') else d.value
            except:
                v = d.value
        else:
            v = d.value
        try:
            v = float(v)
        except Exception:
            v = -float('inf')
        data.append((v, r))
    data = sorted(data, key=lambda x: x[0], reverse=True)
    # Copy back, block is start..end
    for idx, tup in enumerate(data):
        r = tup[1]
        dest_row = start + idx
        for ci, c in enumerate(r):
            dest_cell = sheet.cell(row=dest_row, column=ci+1)
            if is_formula(c):
                dest_cell.value = get_real_formula(c) or c.value
                dest_cell.data_type = 'f'
            else:
                dest_cell.value = c.value
            # Only preserve number format for non-D/E/F
            if ci in [3,4,5]:
                make_bold_center(dest_cell)
                apply_number_format(dest_cell)
            else:
                dest_cell.number_format = c.number_format
    # Set all other cells in range (beyond sorted data) to empty
    for row in range(start + len(data), end+1):
        for ci in range(8):
            dc = sheet.cell(row=row, column=ci+1)
            dc.value = None
            dc.number_format = 'General'
            dc.font = Font(bold=False)
            dc.alignment = Alignment(horizontal='left')

def main():
    input_file = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed2/eval_191-40_tc1/input.xlsx'
    output_file = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed2/eval_191-40_tc1/output.xlsx'
    wb = openpyxl.load_workbook(input_file)
    ws = wb['Sheet1']
    max_row = 85
    max_col = 8
    # Save original column widths
    col_widths = {i: ws.column_dimensions[get_column_letter(i)].width for i in range(1, max_col+1)}
    # Find blocks
    blocks = []
    in_block = False
    block_start = 1
    for i in range(1, max_row+2):
        row = ws[i]
        if not is_row_empty(row) and not in_block:
            block_start = i
            in_block = True
        elif (is_row_empty(row) or i > max_row) and in_block:
            block_end = i-1
            blocks.append((block_start, block_end))
            in_block = False
    # For each block, fix formulas & sort
    for start, end in blocks:
        for r in range(start, end+1):
            for ci in [4,5,6]: # D-F are 4-6, 1-indexed
                c = ws.cell(row=r, column=ci)
                if isinstance(c.value, str) and c.value.startswith('='):
                    c.data_type = 'f'
        sort_block(ws, start, end)
    # Restore column widths
    for i, w in col_widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(output_file)

if __name__ == '__main__':
    main()
