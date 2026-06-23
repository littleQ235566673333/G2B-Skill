import openpyxl

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_5/regression_gate/after_fix/core_547-43/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_5/regression_gate/after_fix/core_547-43/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_emp = wb['Emp']
ws_lookup = wb['Lookup']
ws_expected = wb['Expected']

# Map headers for Emp
emp_headers = next(row for row in ws_emp.iter_rows(min_row=1, max_row=2, values_only=True) if any(cell is not None for cell in row))
emp_col_map = {h: i+1 for i, h in enumerate(emp_headers) if h}

# Map headers for Lookup
table_lookup = []
for i, row in enumerate(ws_lookup.iter_rows(min_row=2, max_row=ws_lookup.max_row, values_only=True), 2):
    if not row[0]:
        continue
    table_lookup.append({
        'Variable': row[0],
        'Category': row[1],
        'Subcategory': row[2]
    })
lookup_map = {row['Variable']: (row['Category'], row['Subcategory']) for row in table_lookup}

# Data rows (stop at first row where seqno is None)
data_rows = []
for row in ws_emp.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        break
    data_rows.append(row)

results = []
for row in data_rows:
    seqno = row[emp_col_map['seqno']-1]
    empno = row[emp_col_map['Empno']-1]
    ename = row[emp_col_map['ename']-1]
    jobid = row[emp_col_map['Jobid']-1]
    for variable in lookup_map:
        if variable in emp_col_map:
            value = row[emp_col_map[variable]-1]
            cat, subcat = lookup_map[variable]
            results.append([seqno, empno, ename, jobid, variable, cat, subcat, value])

# Write to Expected
start_row = 2
for i, out_row in enumerate(results):
    for j, val in enumerate(out_row):
        ws_expected.cell(row=start_row + i, column=1 + j, value=val)

wb.save(output_path)
