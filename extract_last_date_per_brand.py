import pandas as pd
from openpyxl import load_workbook
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_516-46/r0/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_516-46/r0/evolve_516-46/output.xlsx'
sheetname = 'ورقة1'
# Read the relevant range into a dataframe, default header
df = pd.read_excel(input_path, sheet_name=sheetname, usecols='A:E', header=0)
df.columns = df.columns.str.strip()
# Remove any rows where all are NA (blank rows)
df = df.dropna(how='all')
# Debug print columns
print('Detected columns:', df.columns.tolist())
# Infer column names
brand_col = None
date_col = None
qty_col = None
for col in df.columns:
    if isinstance(col, str):
        if 'brand' in col.lower():
            brand_col = col
        if 'date' in col.lower():
            date_col = col
        if 'qty' in col.lower() or 'quantity' in col.lower():
            qty_col = col
if not (brand_col and date_col and qty_col):
    brand_col = df.columns[1]
    date_col = df.columns[0]
    qty_col = df.columns[-1]
df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
df = df.dropna(subset=[brand_col, date_col, qty_col])
last_dates = df.groupby(brand_col)[date_col].transform('max')
df_last = df[df[date_col]==last_dates]
out = df_last.groupby([brand_col, date_col], as_index=False)[qty_col].sum()
out["Modifications"] = ""
mod_mask = df_last.duplicated([brand_col, date_col], keep=False)
if mod_mask.any():
    for idx in out.index:
        bc = out.loc[idx, brand_col]
        dc = out.loc[idx, date_col]
        filtered = df_last[(df_last[brand_col]==bc) & (df_last[date_col]==dc)]
        if len(filtered) > 1:
            out.loc[idx, "Modifications"] = f"Merged {len(filtered)} entries"
wb = load_workbook(input_path)
ws = wb[sheetname]
columns_to_write = [brand_col, date_col, qty_col, "Modifications"]
for ix, col_name in enumerate(columns_to_write):
    ws.cell(row=1, column=8+ix, value=col_name)  # H=8
for i, row in out.iterrows():
    for j, col_name in enumerate(columns_to_write):
        ws.cell(row=2+i, column=8+j, value=row[col_name])
wb.save(output_path)
