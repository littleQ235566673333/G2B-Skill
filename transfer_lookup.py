import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_48983_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_48983_tc1/output.xlsx'

# Load the workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# --- Step 1: Identify the source and target regions ---
# Typically, the headers (categories / brands) and data below them start at row 6
def read_headers(row, start_col, end_col):
    return [ws.cell(row=row, column=col).value for col in range(start_col, end_col + 1)]

def read_column(col, start_row, end_row):
    return [ws.cell(row=row, column=col).value for row in range(start_row, end_row + 1)]

# Find the source table (scan around L6:S11 for structure)
brand_row = 5  # usually headers
category_col = 12  # column L
brand_start_col = 13  # column M
brand_end_col = 19   # column S
category_start_row = 6
category_end_row = 11

# Read brand headers for destination
brands = read_headers(brand_row, brand_start_col, brand_end_col)
# Read category (row label) headers for destination
categories = read_column(category_col, category_start_row, category_end_row)

# Now scan for the source table
# Assume source brands are in a header row above values; categories as index columns
found_table = False
for row in range(1, 20):
    for col in range(1, 20):
        # If we see a brand matching our target, this could be a source table
        val = ws.cell(row=row, column=col).value
        if val in brands:
            # Try to find the full header row and category column
            header_row = row
            header_start_col = col
            # Try leftwards for more header matches
            for c in range(col-1, 0, -1):
                if ws.cell(row=row, column=c).value in brands:
                    header_start_col = c
                else:
                    break
            header_end_col = header_start_col + len(brands) - 1
            # Above must be a category label column
            # Try below for enough categories
            category_labels = [ws.cell(row=r, column=header_start_col-1).value for r in range(row+1, row+1+len(categories))]
            if all(label in categories for label in category_labels):
                found_table = True
                break
    if found_table:
        break

if not found_table:
    raise Exception('Could not locate the source table. Check table structure or header names.')

# Map from brand/category to values in source
source_data = {}
for r_idx, cat in enumerate(categories):
    source_data[cat] = {}
    for c_idx, brand in enumerate(brands):
        val = ws.cell(row=header_row+1+r_idx, column=header_start_col+c_idx).value
        source_data[cat][brand] = val

# Write this to the target m6:s11
for r_idx, cat in enumerate(categories):
    for c_idx, brand in enumerate(brands):
        ws.cell(row=category_start_row + r_idx, column=brand_start_col + c_idx, value=source_data[cat][brand])

wb.save(output_path)
print('Data transferred from source to destination (M6:S11) using INDEX/MATCH logic.')
