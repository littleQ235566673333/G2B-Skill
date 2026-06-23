import openpyxl
from openpyxl.utils import get_column_letter

in_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_409-45_tc1/input.xlsx'
out_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_409-45_tc1/output.xlsx'

wb = openpyxl.load_workbook(in_path)
ws = wb['DATA']

I2_value = ws['I2'].value

rows = list(ws.iter_rows(min_row=1, max_row=19, min_col=1, max_col=9, values_only=False))

# Helper to check if a row is empty in cols A-E

def row_is_empty(row):
    return all(cell.value is None for cell in row[:5])

# All data
all_data = [row for row in rows]

# Remove all-empty rows in place
new_data = []
for row in all_data:
    if not row_is_empty(row):
        new_data.append([cell.value for cell in row])

# Pad to A1:E19 with empty cells for missing rows, and trim if too long
while len(new_data) < 19:
    new_data.append([None]*9)
new_data = new_data[:19]

# Write back to sheet
for r_idx, row in enumerate(new_data, 1):
    for c_idx in range(1, 10):
        ws.cell(row=r_idx, column=c_idx).value = row[c_idx-1]

wb.save(out_path)
