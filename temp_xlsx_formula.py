from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun2/eval_37900_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun2/eval_37900_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active  # Default to first worksheet, adjust if needed

# Place an Excel formula in E5 to lookup today's date in column A and return value from B, or "Not Found"
ws['E5'] = '=XLOOKUP(TODAY(),A:A,B:B,"Not Found")'

wb.save(output_path)
