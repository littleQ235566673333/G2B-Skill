import openpyxl
import numpy as np

# Open workbook
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_2/group_3911/r2/evolve_3911/input.xlsx')
ws = wb['denomination']

# Row indices for employees (Excel rows 2–8)
row_indices = list(range(2, 9))
# Salaries
salaries = [ws[f'E{row}'].value for row in row_indices]
total_salary = sum(salaries)

# Assign each employee a number of notes (sum to 120)
raw_note_quota = np.array([s / total_salary * 120 for s in salaries])
note_quota = np.floor(raw_note_quota)
remainder = 120 - int(note_quota.sum())
remainders = raw_note_quota - note_quota
for i in np.argsort(remainders)[-remainder:]:
    note_quota[i] += 1
note_quota = note_quota.astype(int)

# Denominations H:P (columns 8:16), integer headers only
col_offset = 8
num_denoms = 9

def try_int(x):
    try:
        return int(x)
    except:
        return None

denoms = [try_int(ws.cell(row=1, column=col).value) for col in range(col_offset, col_offset+num_denoms)]
denoms = [d for d in denoms if d is not None]
num_denoms = len(denoms)

# Greedy distribute denominations per employee based on their salary
# n_notes: number of notes assigned to employee
# Amount: that employee's net salary

def distribute_notes(amount, n_notes, denominations):
    result = [0] * len(denominations)
    for _ in range(n_notes):
        # Find largest denom <= amount left
        for i, denom in enumerate(denominations):
            if denom <= amount:
                result[i] += 1
                amount -= denom
                break
        else:
            # If nothing fits, give smallest
            result[-1] += 1
            amount -= denominations[-1]
    return result

for idx, (row, amt, nnotes) in enumerate(zip(row_indices, salaries, note_quota)):
    if nnotes == 0:
        fill = [0] * num_denoms
    else:
        fill = distribute_notes(amt, nnotes, denoms)
    for i, val in enumerate(fill):
        ws.cell(row=row, column=col_offset + i, value=val)

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_2/group_3911/r2/evolve_3911/output.xlsx')
