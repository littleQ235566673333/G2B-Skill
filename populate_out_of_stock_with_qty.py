import openpyxl

# Load input workbook
input_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_45063_tc1/input.xlsx"
output_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_45063_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)
sheet1 = wb["Sheet1"]
sheet2 = wb["Sheet2"]

# Read Sheet2 data into a dictionary for fast lookup
product_to_qty = {}
for row in sheet2.iter_rows(min_row=2, min_col=1, max_col=2):
    product = row[0].value
    qty = row[1].value
    if product is not None:
        product_to_qty[product] = qty

# Apply lookup for products in Sheet1 (A2:A6)
for row in range(2, 7):
    out_of_stock_product = sheet1.cell(row=row, column=1).value
    qty = product_to_qty.get(out_of_stock_product, "")
    sheet1.cell(row=row, column=2, value=qty if qty is not None else "")

# Save workbook
wb.save(output_path)
