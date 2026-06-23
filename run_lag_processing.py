import openpyxl
from openpyxl.styles import PatternFill

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_120-24_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_120-24_tc1/output.xlsx'

# Load workbook and the first sheet (assuming 'Sheet1')
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Column references: BG=59, BL=64, BN=66, AY=51 (openpyxl is 1-based)
col_BL = 64
col_BG = 59
col_BN = 66
col_AY = 51

# Overhead strings
target_BG1 = 'OCOGS - Spares - Transfer Price Overhead'
target_BG2 = 'OFSS - ORCL Consulting Prod Cost I/C'
target_AY = ['BOA_033E', 'BOA_011G']

# Set fill for hex #0070C0
fill_lag = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')

for row in range(2, 46):  # Row 2 to 45 inclusive (Excel 1-based)
    val_BL = ws.cell(row=row, column=col_BL).value
    val_BG = ws.cell(row=row, column=col_BG).value
    val_AY = ws.cell(row=row, column=col_AY).value
    cell_BN = ws.cell(row=row, column=col_BN)

    # Check if BL is 'LAG' and BG matches
    if val_BL == 'LAG' and val_BG in [target_BG1, target_BG2]:
        if val_AY in target_AY:
            cell_BN.value = target_BG1
        else:
            cell_BN.value = target_BG2
    # Set fill if BL is LAG and BN is populated
    if ws.cell(row=row, column=col_BL).value == 'LAG' and cell_BN.value:
        ws.cell(row=row, column=col_BL).fill = fill_lag

wb.save(output_path)
