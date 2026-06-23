import openpyxl

# Load workbook and worksheet
def process_date_range(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Assume input is in A2
    range_text = ws['A2'].value
    if range_text:
        # Extract numbers from 'x to y' style
        parts = range_text.strip().split('to')
        if len(parts) == 2:
            try:
                start = int(parts[0].strip())
                end = int(parts[1].strip())
                day_count = end - start + 1
            except ValueError:
                day_count = 'Invalid range'
        else:
            day_count = 'Invalid range'
    else:
        day_count = 'No input'

    ws['B2'] = day_count
    wb.save(output_path)

process_date_range('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_43589_tc1/input.xlsx',
                   'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_43589_tc1/output.xlsx')
