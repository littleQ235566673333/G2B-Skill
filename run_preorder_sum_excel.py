from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_6/regression_gate/after_pass/core_50796/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_6/regression_gate/after_pass/core_50796/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Get descriptors from A2:A5
row_start = 2
row_end = 5
col_descriptor = 1  # Column A
col_result = 2      # Column B

descriptors = [ws.cell(row=i, column=col_descriptor).value for i in range(row_start, row_end + 1)]

# Source data in B12:C22 (columns 2:3, rows 12:22)
data_start = 12
data_end = 22

# Prepare a mapping from descriptor to sum of values
for output_row, descriptor in enumerate(descriptors, start=row_start):
    total = 0
    for data_row in range(data_start, data_end + 1):
        cell_desc = ws.cell(row=data_row, column=2).value
        cell_val = ws.cell(row=data_row, column=3).value
        if cell_desc == descriptor and isinstance(cell_val, (int, float)):
            total += cell_val
    ws.cell(row=output_row, column=col_result).value = total

wb.save(output_path)
