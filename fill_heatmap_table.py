import openpyxl
from datetime import time

# File paths
def input_path():
    return 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/before_fix/core_52305/input.xlsx'
def output_path():
    return 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/before_fix/core_52305/output.xlsx'

wb = openpyxl.load_workbook(input_path())
ws = wb['Sheet1']

# Read account names from J5:N5 (columns 10-14)
names = [ws.cell(row=5, column=col).value for col in range(10, 15)]

# Read time ranges from rows 6-24 (columns 8 and 9)
time_ranges = []
for row in range(6, 25):
    start = ws.cell(row=row, column=8).value
    end = ws.cell(row=row, column=9).value
    if isinstance(start, time) and isinstance(end, time):
        time_ranges.append((start, end))
    else:
        # fallback to parse string if needed
        try:
            if isinstance(start, str):
                start = time.fromisoformat(start)
            if isinstance(end, str):
                end = time.fromisoformat(end)
            time_ranges.append((start, end))
        except Exception:
            time_ranges.append((None, None))

# Collect data rows
entries = []
for row in ws.iter_rows(min_row=2, max_row=100, max_col=4, values_only=True):
    type_, dt, name, dest = row
    if name in names and hasattr(dt, 'time'):
        entries.append((name, dt.time()))

# Fill heatmap count table (J6:N24)
for r_index, (start, end) in enumerate(time_ranges):
    for c_index, acct_name in enumerate(names):
        if start is None or end is None:
            ws.cell(row=6 + r_index, column=10 + c_index, value='')
            continue
        # Count matching entries
        count = sum(
            1 for n, t in entries
            if n == acct_name and start <= t < end
        )
        ws.cell(row=6 + r_index, column=10 + c_index, value=count)

wb.save(output_path())
print('Heatmap filled and saved to', output_path())