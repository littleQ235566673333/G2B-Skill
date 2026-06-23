import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_2/group_47766/r2/evolve_47766/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_2/group_47766/r2/evolve_47766/output.xlsx'

# Read all sheets
xls = pd.read_excel(input_file, sheet_name=None)

# We'll use openpyxl for writing
wb = load_workbook(input_file)
ws = wb.active

# Find the data range via pandas
sheetname = ws.title
df = xls[sheetname]

# Define your date range for a particular year
start_date = datetime(2023, 1, 1)
end_date = datetime(2023, 12, 31)

# Map excel columns to data frame columns
# Assume: 'F' = 5 = closing date; 'H' = 7 = type; 'C' = 2 = amount
closing_date_col = 5  # F
amount_col = 2        # C
type_col = 7          # H

# Extract the relevant ranges (Excel is 1-based, pandas 0-based, so adjust)
data_rows = pd.concat([
    df.iloc[7:37],     # H8:H37 (indexes 7 to 36)
    df.iloc[40:58],    # H41:H58 (indexes 40 to 57)
    df.iloc[61:74],    # H62:H74 (indexes 61 to 73)
])

# Filter for '*PE*' in type and date range in closing date
def is_date_in_range(val):
    if pd.isna(val): return False
    try:
        d = pd.to_datetime(val)
    except Exception:
        return False
    return start_date <= d <= end_date

filtered = data_rows[data_rows.iloc[:,type_col-1].astype(str).str.contains('PE', na=False)]
filtered = filtered[filtered.iloc[:,closing_date_col-1].apply(is_date_in_range)]

total = filtered.iloc[:,amount_col-1].sum()

# Write the result to K40 (row=40, col=11)
ws.cell(row=40, column=11).value = total

wb.save(output_file)
print('Done. Total:', total)
