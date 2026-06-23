import openpyxl
from datetime import datetime
from calendar import monthrange

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_57590_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_57590_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

def get_year_from_column_c():
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        date_val = row[2].value  # Column C
        if isinstance(date_val, datetime):
            return date_val.year
        elif isinstance(date_val, str):
            for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d-%b-%Y'):
                try:
                    dt = datetime.strptime(date_val, fmt)
                    return dt.year
                except Exception:
                    continue
    return datetime.now().year  # fallback if no date found

# Read month from A26
month_str = ws['A26'].value
if isinstance(month_str, str) and month_str.isalpha():
    try:
        month_num = datetime.strptime(month_str, '%B').month
    except ValueError:
        raise ValueError(f'Could not parse month name: {month_str}')
    year = get_year_from_column_c()
    target_month = datetime(year, month_num, 1)
else:
    # If A26 is a real date, use it directly
    if isinstance(month_str, datetime):
        target_month = month_str
    else:
        raise ValueError('A26 does not contain a recognizable month or date')

start_date = target_month.replace(day=1)
_, last_day = monthrange(target_month.year, target_month.month)
end_date = target_month.replace(day=last_day)

vals = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    date_cell = row[2]  # Column C (index 2)
    val_cell = row[8]   # Column I (index 8)
    date_val = date_cell.value
    val_val = val_cell.value
    # Convert date if needed
    if isinstance(date_val, str):
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d-%b-%Y'):
            try:
                date_val = datetime.strptime(date_val, fmt)
                break
            except Exception:
                continue
        else:
            continue  # skip if no valid date
    if date_val is None or val_val is None:
        continue
    if start_date <= date_val <= end_date:
        try:
            vals.append(float(val_val))
        except Exception:
            continue
sum_val = sum(vals)
ws['B26'].value = sum_val
wb.save(output_path)
