import openpyxl
import pandas as pd
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_5/regression_gate/after_fix/core_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_5/regression_gate/after_fix/core_45896/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Volym P5_P6_2023']
zord = wb['ZORD'] if 'ZORD' in wb.sheetnames else wb[wb.sheetnames[1]]

# When values_only=True, row is a tuple of values, not cell objects
zord_data = [(row[0], row[2]) for row in zord.iter_rows(min_row=2, values_only=True) if row[0] is not None and row[2] is not None]

for idx in range(2, 11):
    key = ws[f'A{idx}'].value
    dates = [d for a, d in zord_data if a == key]
    dt_objs = []
    for d in dates:
        try:
            dt = pd.to_datetime(d, dayfirst=True)
            dt_objs.append(dt)
        except Exception:
            continue
    dt_objs = sorted(dt_objs)
    formatted_dates = [dt.strftime('%d/%m/%Y') for dt in dt_objs]
    val = ','.join(formatted_dates) if formatted_dates else ''
    ws[f'I{idx}'] = val

wb.save(output_path)
