import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_7/regression_gate/after_fix/core_263-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_7/regression_gate/after_fix/core_263-1/output.xlsx'

mat_col = 'Mtrl'
width_col = 'Width'
height_col = 'Height'

# Load all sheets as DataFrames
df = pd.read_excel(input_path)

# Only keep non-NA and numeric rows for area calculation
df = df[[mat_col, width_col, height_col]].dropna()

# Area for each row
df['Area'] = pd.to_numeric(df[width_col], errors='coerce') * pd.to_numeric(df[height_col], errors='coerce')

# Remove rows where area is not valid (e.g., due to non-numeric entries)
df = df.dropna(subset=['Area'])

# Group by material type
res = df.groupby(mat_col)['Area'].sum().reset_index()
# Sort by material name for output stability
res = res.sort_values(mat_col).reset_index(drop=True)

# Load workbook to write output
wb = load_workbook(input_path)
ws = wb.active

# Write results to H2:H4 (rounded), blank if less than 3 entries
for i in range(3):
    cell = f'H{i+2}'
    if i < len(res):
        ws[cell] = f"{res.loc[i, mat_col]}: {round(res.loc[i, 'Area'], 2)}"
    else:
        ws[cell] = ''

wb.save(output_path)
