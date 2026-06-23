from openpyxl import load_workbook
import json

INPUT = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r2/eval_120-24_tc1/input.xlsx'
wb = load_workbook(INPUT)
ws = wb['Sheet1']

out = []
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    out.append([(i, str(cell) if cell is not None else None) for i, cell in enumerate(row)])
    if idx > 10:
        break
print(json.dumps(out, indent=2))
