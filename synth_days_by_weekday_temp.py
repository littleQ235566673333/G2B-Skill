import openpyxl
from openpyxl.utils import get_column_letter

def main():
    input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_57989_tc1/input.xlsx"
    output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_57989_tc1/output.xlsx"

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Guess the header region to extract drivers and days
    # Assume drivers are in column A, days as header row
    header_row = 1
    start_data_row = 2  # after header
    start_data_col = 2  # days start at column B

    days = []
    col_map = {}
    max_col = ws.max_column
    # days are headers in row 1 from col B onwards
    for c in range(start_data_col, max_col + 1):
        day = ws.cell(row=header_row, column=c).value
        if day:
            days.append(day)
            col_map[day] = c

    # drivers are in col A, from row 2 up to last used row
    max_row = ws.max_row
    drivers = []
    for r in range(start_data_row, max_row + 1):
        driver = ws.cell(row=r, column=1).value
        if driver:
            drivers.append(driver)
    drivers = list(dict.fromkeys(drivers))  # keep order, remove duplicates

    # Build the summary table
    synthesis = {}
    for driver in drivers:
        synthesis[driver] = {}
        for day in days:
            synthesis[driver][day] = 0

    # Count for each driver and day
    for r in range(start_data_row, max_row + 1):
        driver = ws.cell(row=r, column=1).value
        if not driver:
            continue
        for day in days:
            c = col_map[day]
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != '':
                synthesis[driver][day] += 1

    # Write result to B25:H43
    base_row = 25
    base_col = 2 # col B

    # Write day names as header row
    for d, day in enumerate(days):
        ws.cell(row=base_row, column=base_col + d, value=day)
    # Write driver names in first column and synthesis data
    for i, driver in enumerate(drivers):
        ws.cell(row=base_row + 1 + i, column=base_col - 1, value=driver)
        for j, day in enumerate(days):
            ws.cell(row=base_row + 1 + i, column=base_col + j, value=synthesis[driver][day])

    wb.save(output_path)

if __name__ == "__main__":
    main()
