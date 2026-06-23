from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_6/group_11276/r2/evolve_11276/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_6/group_11276/r2/evolve_11276/output.xlsx'

wb = load_workbook(input_path)
ws = wb['ATTENDENCE']

start_col = 6  # F
end_col = 36   # AJ
row_formula = 3
row_date = 4

for c in range(start_col, end_col + 1):
    col_letter = get_column_letter(c)
    ws[f'{col_letter}{row_formula}'] = f'=TEXT({col_letter}{row_date},"DDD")'

wb.save(output_path)
