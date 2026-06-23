import openpyxl

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_48983_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_48983_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Define source table - brands in B6:B19, categories in C5:J5
brand_rows = range(6, 20)  # Persil to Bloo
category_cols = range(3, 10)  # C-J
categories = [ws.cell(row=5, column=col).value for col in category_cols]
brands = [ws.cell(row=row, column=2).value for row in brand_rows]

# Build source lookup dictionary
source_dict = {}
for r in brand_rows:
    brand = ws.cell(row=r, column=2).value
    source_dict[brand] = {}
    for c in category_cols:
        cat = ws.cell(row=5, column=c).value
        source_dict[brand][cat] = ws.cell(row=r, column=c).value

# Target table: brands in L6:L11, categories in M5:S5
output_rows = range(6, 12)
output_brand_col = 12  # L
output_category_cols = range(13, 20)  # M:S
output_categories = [ws.cell(row=5, column=col).value for col in output_category_cols]

for i, row in enumerate(output_rows):
    brand = ws.cell(row=row, column=output_brand_col).value
    for j, col in enumerate(output_category_cols):
        category = ws.cell(row=5, column=col).value
        value = source_dict.get(brand, {}).get(category)
        ws.cell(row=row, column=col).value = value

wb.save(output_path)
