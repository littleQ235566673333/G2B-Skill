from openpyxl import load_workbook

def extract_name(emp_str):
    if not emp_str:
        return ''
    # Split string at first occurrence of an email delimiter (e.g., '@')
    name_part = emp_str.split('@')[0].strip()
    # Remove anything after the last space (could be a domain, etc.)
    name_parts = name_part.split()
    if len(name_parts) >= 2:
        return f'{name_parts[0]} {name_parts[1]}'
    return name_part

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_7/regression_gate/after_pass/core_36191/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_7/regression_gate/after_pass/core_36191/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Process C2 and C3 from Employee column (B2, B3)
ws['C2'] = extract_name(ws['B2'].value)
ws['C3'] = extract_name(ws['B3'].value)

wb.save(output_path)
