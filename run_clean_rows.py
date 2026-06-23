import openpyxl

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_374-18_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_374-18_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Collect all row indices (excluding header) where column E (5th col, index 4) < 1
to_delete = []
for i in range(2, ws.max_row + 1):
    val = ws.cell(row=i, column=5).value
    if val is not None and isinstance(val, (int, float)):
        if val < 1:
            to_delete.append(i)

# Delete rows in reverse order
for idx in reversed(to_delete):
    ws.delete_rows(idx)

# Save to new file
wb.save(output_path)
