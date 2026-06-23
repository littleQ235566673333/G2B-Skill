import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_4/regression_gate/after_fix/core_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_4/regression_gate/after_fix/core_45896/output.xlsx'

# Load workbooks/sheets
wb = openpyxl.load_workbook(input_path)
ws = wb['Volym P5_P6_2023']
zord_df = pd.read_excel(input_path, sheet_name='ZORD')

# Iterate I2:I10
for row in range(2, 11):
    key = ws[f'A{row}'].value
    # Find all matching dates in column C (index 2) of ZORD where column A matches
    matches = zord_df[zord_df[zord_df.columns[0]] == key][zord_df.columns[2]]
    # Remove duplicates, drop NaN, coerce to datetime, keep only valid/deduped
    dates = pd.to_datetime(matches.dropna().unique(), errors='coerce')
    # Format as DD/MM/YYYY and sort
    dates = sorted([d.strftime('%d/%m/%Y') for d in dates if pd.notnull(d)])
    value = ', '.join(dates) if dates else ''
    ws[f'I{row}'] = value

wb.save(output_path)
