import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_22-47_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_22-47_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read all rows so we can process columns by position
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
# Use only columns A(0), B(1), C(2), J(9)
A_idx, B_idx, C_idx, J_idx = 0, 1, 2, 9

data = []
for r in rows[1:]:
    b = r[B_idx]
    c = r[C_idx]
    if (b is None or c is None):
        continue  # skip empty
    if isinstance(b, str) and ('name' in b.lower() or 'emplo' in b.lower()):
        continue  # skip header duplicate rows
    data.append({'ITEM': r[A_idx], 'NAME': b, 'REF': c, 'LIST': r[J_idx] if J_idx < len(r) else None})

# Remove duplicates by NAME + REF
seen = set()
filtered = []
for row in data:
    key = (row['NAME'], row['REF'])
    if key not in seen:
        seen.add(key)
        filtered.append(row)

data = filtered
# Find helpers from LIST/J
jvals = [row['LIST'] for row in data if row['LIST'] is not None and not (isinstance(row['LIST'], str) and ('name' in row['LIST'].lower() or 'emplo' in row['LIST'].lower()))]

result_rows = []
used = set()
if jvals:
    # Add matching names in J
    for helper in jvals:
        for row in data:
            if row['NAME'] == helper:
                key = (row['NAME'], row['REF'])
                if key not in used:
                    result_rows.append(row)
                    used.add(key)
    # Add everything else, original order
    for row in data:
        if row['NAME'] not in jvals:
            key = (row['NAME'], row['REF'])
            if key not in used:
                result_rows.append(row)
                used.add(key)
else:
    # Sort by NAME A-Z if helper is empty
    data_sorted = sorted(data, key=lambda x: x['NAME'])
    for row in data_sorted:
        key = (row['NAME'], row['REF'])
        if key not in used:
            result_rows.append(row)
            used.add(key)

# Write F2:H10 (columns 6,7,8)
outrows = result_rows[:9]
for i, row in enumerate(outrows):
    ws.cell(row=2+i, column=6, value=row['ITEM'])
    ws.cell(row=2+i, column=7, value=row['NAME'])
    ws.cell(row=2+i, column=8, value=row['REF'])

# Now G/H sorted only by H ascending
gh_rows = [(ws.cell(row=2+i, column=7).value, ws.cell(row=2+i, column=8).value) for i in range(len(outrows))]
gh_sorted = sorted(gh_rows, key=lambda x: (x[1] if x[1] is not None else float('inf')))
for i, (g, h) in enumerate(gh_sorted):
    ws.cell(row=2+i, column=7, value=g)
    ws.cell(row=2+i, column=8, value=h)

wb.save(output_path)
