import openpyxl

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_55979_tc1/input.xlsx')
sheetnames = wb.sheetnames
for s in ['Supplier_1','Supplier_2','Supplier_3']:
    ws = wb[s]
    print(f'Sheet: {s}')
    for row in ws.iter_rows(values_only=True):
        print(row)