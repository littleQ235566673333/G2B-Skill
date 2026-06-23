import openpyxl

# Load workbook
wb = openpyxl.load_workbook('results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_3/task_58701/r0/evolve_58701/input.xlsx')

# Read Table Tab for location -> office code mapping
table_ws = wb['Table Tab']
header = [cell.value for cell in next(table_ws.iter_rows(min_row=1, max_row=1))]
loc_col_idx = header.index('Location')
office_col_idx = header.index('Office Code')
loc_to_office = {}
for row in table_ws.iter_rows(min_row=2, values_only=True):
    location = row[loc_col_idx]
    office_code = row[office_col_idx]
    if location is not None and office_code is not None:
        loc_to_office[location] = office_code

# Load Entry Tab
entry_ws = wb['Entry Tab']
entry_header = [cell.value for cell in next(entry_ws.iter_rows(min_row=1, max_row=1))]
loc_entry_idx = entry_header.index('Location') + 1  # 1-based for .cell
office_entry_idx = entry_header.index('Office Code') + 1  # 1-based

# For rows 2 and 3, fill Office Code based on Location value
for rownum in range(2, 4):  # rows 2, 3
    loc_val = entry_ws.cell(row=rownum, column=loc_entry_idx).value
    office_code = loc_to_office.get(loc_val, None)
    entry_ws.cell(row=rownum, column=office_entry_idx, value=office_code)

wb.save('results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_3/task_58701/r0/evolve_58701/output.xlsx')
