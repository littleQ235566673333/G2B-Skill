import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_33157_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_33157_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Headers for activity columns B, D, F, H
activity_cols = ['B', 'D', 'F', 'H']
header_map = {}
for col in activity_cols:
    header_map[col] = ws[f'{col}1'].value

# Process rows 2 to 6 (assuming these are the relevant rows)
for row in range(2, 7):
    target_date = ws[f'J{row}'].value
    matched_activity = None
    for col in activity_cols:
        cell_value = ws[f'{col}{row}'].value
        if cell_value == target_date:
            matched_activity = header_map[col]
            break
    ws[f'K{row}'] = matched_activity

# Save the workbook
wb.save(output_path)
