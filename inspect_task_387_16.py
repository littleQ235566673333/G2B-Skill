from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_387-16_s0/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_387-16_s0/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Inspect current data layout to identify columns/rows
for r in range(1, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    print(r, vals)
