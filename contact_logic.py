from openpyxl import load_workbook
from datetime import datetime, timedelta

inp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_1/regression_gate/after_pass/core_41589/input.xlsx'
outp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_1/regression_gate/after_pass/core_41589/output.xlsx'

wb = load_workbook(inp)
ws = wb['Contact List']
today = datetime.today()
h = ws['H4'].value
i = ws['I4'].value

# Parse date
result = 'NO ACTION'
d = None
if isinstance(h, datetime):
    d = h
elif isinstance(h, str):
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d-%b-%Y', '%d/%m/%Y'):
        try:
            d = datetime.strptime(h.strip(), fmt)
            break
        except Exception:
            pass
if i and i.strip().upper() == 'YES' and d is not None:
    days = (today - d).days
    if days <= 30:
        result = 'HOLD'
    elif days > 30:
        result = 'TOUCH BASE'
ws['J4'] = result
wb.save(outp)
