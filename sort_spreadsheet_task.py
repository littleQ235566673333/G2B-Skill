import pandas as pd
from openpyxl import load_workbook

# File paths
infile = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun1/eval_22-47_tc1/input.xlsx"
outfile = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun1/eval_22-47_tc1/output.xlsx"

# Read sheet into pandas DataFrame
wb = load_workbook(infile)
ws = wb.active
rows = list(ws.values)

# Find the data's actual starting row (after header)
header_row = 0
for idx, row in enumerate(rows):
    if row[0] == 'ITEM' and row[1] == 'NAME':
        header_row = idx
        break

# Gather data (columns A, B, C are 0,1,2; J is 9)
data = []
for r in rows[header_row+1:]:
    # skip if all are None or part-header
    if (r[0], r[1], r[2]) == ('ITEM', 'NAME', 'REF'):
        continue
    if r[1] is None and r[2] is None:
        break
    data.append({
        'A': r[0], 'B': r[1], 'C': r[2], 'J': r[9]
    })

df = pd.DataFrame(data)

# Drop rows where both B and C are missing
if not df.empty:
    df = df[~((df['B'].isna()) & (df['C'].isna()))]

# Deduplication: drop duplicates on (B, C), keep first
seen = set()
unique_rows = []
for _, row in df.iterrows():
    key = (row['B'], row['C'])
    if key in seen:
        continue
    seen.add(key)
    unique_rows.append(row)
df_unique = pd.DataFrame(unique_rows)

# Read helper column J names in the order they appear, skip empties
j_names = df_unique['J'].dropna().unique().tolist()
j_names = [n for n in j_names if pd.notnull(n) and str(n).strip() != '']

# Grouping and sorting
result = []
if j_names:
    # Group 1: For each name in J, all rows in that order
    used_idx = set()
    for n in j_names:
        matches = df_unique[df_unique['B'] == n]
        for idx, row in matches.iterrows():
            result.append(row)
            used_idx.add(idx)
    # Group 2: rows whose B is not in J, keep order
    others = df_unique[~df_unique.index.isin(used_idx)]
    for _, row in others.iterrows():
        result.append(row)
else:
    # No helper J, just sort by B alphabetically
    result = df_unique.sort_values(by='B').to_dict('records')

# Now, restrict output to max 9 rows (for F2:H10) and focus on output only G and H (per user)
# Output columns: G (NAME, B), H (REF, C)
final = []
for i in range(min(9, len(result))):
    row = result[i]
    # Keep only those with valid values in B
    if isinstance(row, pd.Series):
        name, ref = row['B'], row['C']
    else:
        name, ref = row['B'], row['C']
    if name is None and ref is None:
        continue
    final.append({'G': name, 'H': ref})

# For final requirement: sort by H (ref) lowest to highest, only in column H, but output must remain in G2:H10
df_final = pd.DataFrame(final)
if not df_final.empty:
    df_final = df_final.sort_values(by='H', ascending=True, ignore_index=True)

# Write to output file in F2:H10 (only G and H; F blank)
ws = wb.active
out_start = 2
for i in range(9):
    g = df_final.iloc[i]['G'] if i < len(df_final) else None
    h = df_final.iloc[i]['H'] if i < len(df_final) else None
    ws[f'F{out_start+i}'] = None
    ws[f'G{out_start+i}'] = g
    ws[f'H{out_start+i}'] = h

wb.save(outfile)
print('Done')
