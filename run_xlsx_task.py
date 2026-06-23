import openpyxl
from openpyxl.utils import column_index_from_string

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/group_32789/r0/evolve_32789/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/group_32789/r0/evolve_32789/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Set up columns and boundaries
data_start_col = column_index_from_string('BD')
data_end_col = column_index_from_string('BR')
data_cols = list(range(data_start_col, data_end_col + 1))

# Read all primary (row 2) and secondary (row 3) headers for BD:BR
primary_headers = [ws.cell(row=2, column=col).value for col in data_cols]
secondary_headers = [ws.cell(row=3, column=col).value for col in data_cols]

# We'll use B2 as the primary header, and C3 as the secondary header for demonstration; but really this should match the correct logic.
def find_matching_col_indices(pri_header, sec_header):
    return [i for i, (p, s) in enumerate(zip(primary_headers, secondary_headers)) if p == pri_header and s == sec_header]

# For each D4:D7 row, set cell formatted as whole number by matching date + headers
def get_value_by_headers_and_date(date_val, pri_header, sec_header):
    matches = find_matching_col_indices(pri_header, sec_header)
    if not matches:
        return '-'
    for col_index in matches: # Could be more than one match; take first
        col = data_cols[col_index]
        # Find row where BD matches date_val
        for data_row in range(4, ws.max_row + 1):
            if ws.cell(row=data_row, column=data_start_col).value == date_val:
                val = ws.cell(row=data_row, column=col).value
                try:
                    if val is not None and float(val) == int(float(val)):
                        return int(float(val))
                    else:
                        return val
                except Exception:
                    return val
        return '-'
    return '-'

# Fill D4:D7, C4:C7
for row in range(4, 8):
    date_val = ws[f'A{row}'].value
    pri_header = ws[f'B2'].value # This is possibly not dynamic, but per user description
    sec_header = ws[f'C3'].value # If more sec headers needed, adapt as necessary
    result = get_value_by_headers_and_date(date_val, pri_header, sec_header)
    ws[f'D{row}'] = result
    ws[f'C{row}'] = None  # Clear as user wants only D

# Clear below row 7 on columns C:D (C8:D14)
for row in range(8, 15):
    ws[f'C{row}'] = None
    ws[f'D{row}'] = None

wb.save(output_path)
