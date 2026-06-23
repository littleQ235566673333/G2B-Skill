import openpyxl
import pandas as pd
from collections import OrderedDict

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_22-47_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_22-47_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

df = pd.read_excel(input_path, sheet_name=ws.title)

headers = list(df.columns)
colB = headers.index('B') if 'B' in headers else 1
colC = headers.index('C') if 'C' in headers else 2
jcol = headers.index('J') if 'J' in headers else 9

data = []
existing_pairs = set()
for idx, row in df.iterrows():
    val_b = row.iloc[colB] if len(row) > colB else None
    val_c = row.iloc[colC] if len(row) > colC else None
    # Blank or header: skip
    if pd.isna(val_b) or pd.isna(val_c):
        continue
    if idx == 0:
        continue  # skip header
    pair = (val_b, val_c)
    if pair in existing_pairs:
        continue
    existing_pairs.add(pair)
    data.append((idx, val_b, val_c, [row.get(col, None) for col in headers], row.iloc[jcol] if len(row) > jcol else None))
# Helper column
j_vals = df.iloc[:, jcol].dropna().tolist() if df.shape[1] > jcol else []
j_vals_clean = [x for x in j_vals if str(x).strip()]
found_j = len(j_vals_clean) > 0
order_j = []
j_seen = set()
for x in j_vals:
    if pd.isna(x):
        continue
    if x in j_seen or str(x).strip() == '':
        continue
    order_j.append(x)
    j_seen.add(x)
# Map from colB to original data row indices
matches = OrderedDict()
for od in data:
    idx, val_b, val_c, fullrow, val_j = od
    if val_b not in matches:
        matches[val_b] = []
    matches[val_b].append((idx, fullrow))
output = []
if found_j and order_j:
    collected = set()
    for name in order_j:
        if name in matches:
            for mm in matches[name]:
                if (name, mm[1][colC]) not in collected:
                    output.append(mm[1])
                    collected.add((name, mm[1][colC]))
    # Now add all that are not in J in original order
    for od in data:
        val_b = od[1]
        if val_b not in order_j:
            output.append(od[3])
else:
    # J is empty: sort A-Z by B
    data.sort(key=lambda x: ('' if pd.isna(x[1]) else x[1]))
    for od in data:
        output.append(od[3])
# Write out to F2:H10 (max 9 rows)
for r in range(2, 11):
    out_row = output[r-2] if r-2 < len(output) else ['']*len(headers)
    for n, cc in enumerate(['F','G','H']):
        ws[f'{cc}{r}'] = out_row[n+5] if n+5 < len(out_row) else ''
# Sort only column H in F2:H10
h_block = []
for r in range(2, 11):
    hval = ws[f'H{r}'].value
    gval = ws[f'G{r}'].value
    fval = ws[f'F{r}'].value
    h_block.append((hval, gval, fval))
h_block_sorted = sorted(h_block, key=lambda x: (float(x[0]) if x[0] is not None and x[0] != '' else float('inf')))
for idx, (hval, gval, fval) in enumerate(h_block_sorted):
    ws[f'F{idx+2}'] = fval
    ws[f'G{idx+2}'] = gval
    ws[f'H{idx+2}'] = hval
wb.save(output_path)
