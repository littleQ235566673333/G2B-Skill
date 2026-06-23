import pandas as pd
from openpyxl import load_workbook
import numpy as np

def main():
    input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun1/eval_567-21_tc1/input.xlsx'
    output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun1/eval_567-21_tc1/output.xlsx'

    df = pd.read_excel(input_path, sheet_name='Sheet1')

    # Remove rows with empty/NaN in E (col 4, zero-based index)
    df = df[~df.iloc[:,4].isna() & (df.iloc[:,4].astype(str).str.strip() != '')]

    # Convert date in E to a comparable form (assume YYYY/MM always)
    def parse_date(val):
        try:
            return pd.to_datetime(str(val), format='%Y/%m')
        except:
            return pd.NaT
    df['_date'] = df.iloc[:,4].apply(parse_date)
    df = df[~df['_date'].isna()]

    # Group by A and B, keep all with max E per group
    idx = df.groupby([df.columns[0], df.columns[1]])['_date'].transform(max) == df['_date']
    df_keep = df[idx].copy()
    df_keep = df_keep.drop(columns=['_date'])

    # Prep for writing to openpyxl starting at A3
    to_write = df_keep.iloc[:26, :7].values.tolist() # Only rows for A3:G28

    wb = load_workbook(input_path)
    ws = wb['Sheet1']

    # Clear area A3:G28
    for row in ws.iter_rows(min_row=3, max_row=28, min_col=1, max_col=7):
        for cell in row:
            cell.value = None

    # Write data
    for i, row in enumerate(to_write, start=3):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    wb.save(output_path)

if __name__ == '__main__':
    main()
