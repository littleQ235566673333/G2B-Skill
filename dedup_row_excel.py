from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/regression_gate/after_fix/core_227-40/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/regression_gate/after_fix/core_227-40/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

def remove_row_duplicates(row):
    seen = set()
    result = []
    for v in row:
        if v not in seen:
            seen.add(v)
            result.append(v)
        else:
            result.append(None)
    return result

# Deduplicate rows A2:E5 and write output at A14:E17
for out_idx, in_idx in enumerate(range(2, 6), start=14):
    row_vals = [ws.cell(row=in_idx, column=col).value for col in range(1, 6)]
    cleaned = remove_row_duplicates(row_vals)
    for col, val in enumerate(cleaned, start=1):
        ws.cell(row=out_idx, column=col, value=val)

wb.save(output_path)
