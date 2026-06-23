import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_3/regression_gate/after_fix/core_58701/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_3/regression_gate/after_fix/core_58701/output.xlsx'

# Load the workbook and sheets
wb = load_workbook(input_path)
ws_entry = wb['Entry Tab']
ws_table = wb['Table Tab']

# ---- Map Entry Tab Columns ---- #
header_row_entry = next(row for row in ws_entry.iter_rows(min_row=1, max_row=3, values_only=True) if any(cell is not None for cell in row))
entry_mapping = {header: idx+1 for idx, header in enumerate(header_row_entry) if header}
location_col_entry = entry_mapping['Location']
office_code_col_entry = entry_mapping['Office Code']

# ---- Map Table Tab Columns ---- #
header_row_table = next(row for row in ws_table.iter_rows(min_row=1, max_row=3, values_only=True) if any(cell is not None for cell in row))
table_mapping = {header: idx+1 for idx, header in enumerate(header_row_table) if header}
location_name_col_table = table_mapping['Location Name']
office_code_col_table = table_mapping['Office Code']

# ---- Build dict Location -> Office Code ---- #
location_to_code = {}
for row in ws_table.iter_rows(min_row=2, max_row=ws_table.max_row, values_only=True):
    location = row[location_name_col_table-1]
    code = row[office_code_col_table-1]
    if location:
        location_to_code[location] = code

# ---- Write Office Codes to Entry Tab ---- #
for r in [2,3]:  # target Entry Tab!E2:E3
    location = ws_entry.cell(row=r, column=location_col_entry).value
    code = location_to_code.get(location, None)
    ws_entry.cell(row=r, column=office_code_col_entry, value=code)

# Save result
wb.save(output_path)
