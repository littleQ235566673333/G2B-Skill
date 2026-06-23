import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_1/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_1/regression_gate/after_pass/core_50526/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

lookup_value = ws['B6'].value
# Build col name mapping
col_names = []
for cell in ws[1]:
    col_names.append(cell.value)
# Find the row of the lookup_value in col 1
lookup_row = None
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=1).value == lookup_value:
        lookup_row = row
        break
if lookup_row is None:
    values = ['', '']
else:
    # Find all headers whose value in that row > 0 (except the first col)
    values = []
    for col in range(2, ws.max_column + 1):
        v = ws.cell(row=lookup_row, column=col).value
        if (isinstance(v, (int, float)) and v > 0):
            values.append(col_names[col-1])
    while len(values) < 2:   # Fill at least 2 cells
        values.append('')
# Write to B9, B10
ws['B9'].value = values[0]
ws['B10'].value = values[1]
wb.save(output_path)
