import openpyxl

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42/eval_55468_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42/eval_55468_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Read matching values from criteria cells
AC4 = ws['AC4'].value
AB4 = ws['AB4'].value
AE4 = ws['AE4'].value
AD4 = ws['AD4'].value

# Define data and header regions, as in the user's question
# Data in C5:Z10, vertical criteria in A/B 5-10, horizontal criteria in C4:Z4 and C3:Z3
num_data_rows = 6  # 5-10 inclusive
num_data_cols = 24 # C-Z inclusive

data = [[ws.cell(row=5 + i, column=3 + j).value for j in range(num_data_cols)] for i in range(num_data_rows)]
C4_Z4 = [ws.cell(row=4, column=3 + j).value for j in range(num_data_cols)]
C3_Z3 = [ws.cell(row=3, column=3 + j).value for j in range(num_data_cols)]
A5_A10 = [ws.cell(row=5 + i, column=1).value for i in range(num_data_rows)]
B5_B10 = [ws.cell(row=5 + i, column=2).value for i in range(num_data_rows)]

# Find the row index
row_idx = None
for i, (a, b) in enumerate(zip(A5_A10, B5_B10)):
    if a == AE4 and b == AD4:
        row_idx = i
        break

# Find the column index
col_idx = None
for j, (c4, c3) in enumerate(zip(C4_Z4, C3_Z3)):
    if c4 == AC4 and c3 == AB4:
        col_idx = j
        break

# Set the result in AE5
if row_idx is not None and col_idx is not None:
    ws['AE5'] = data[row_idx][col_idx]
else:
    ws['AE5'] = '#N/A'

wb.save(output_file)
