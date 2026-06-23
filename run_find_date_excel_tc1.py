from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun1/eval_333-29_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun1/eval_333-29_tc1/output.xlsx'

wb = load_workbook(input_path)
ws_A = wb['A']
ws_B = wb['B']

# Step 1: Find the first cell in column P (16th col) with 'Yes' or 'NA'
row_found = None
for r in range(2, ws_A.max_row+1):
    val = ws_A.cell(row=r, column=16).value
    if val in ('Yes','NA'):
        row_found = r
        break
date_to_paste = None
if row_found is not None:
    # Step 2: Look UP from (row_found-1 downto 2) in column L for value 100
    val_row = None
    for rr in range(row_found-1, 1, -1):
        if ws_A.cell(row=rr, column=12).value == 100:
            val_row = rr
            break
    # Step 3: If found, take column F (date) from this row
    if val_row is not None:
        date_to_paste = ws_A.cell(row=val_row, column=6).value

# Step 4: If nothing found, fallback to cell F3
if row_found is None or date_to_paste is None:
    date_to_paste = ws_A['F3'].value

# Step 5: Check F1 text against column C of sheet B
f1_val = ws_A['F1'].value
match_row = None
for r in range(2, ws_B.max_row+1):
    if ws_B.cell(row=r, column=3).value == f1_val:
        match_row = r
        break

# Step 6: As per instructions, always output answer in B!F4
ws_B.cell(row=4, column=6, value=date_to_paste)

wb.save(output_path)
