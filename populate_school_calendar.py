import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_3/group_50916/r3/evolve_50916/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_3/group_50916/r3/evolve_50916/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Table layout guesses:
# Cycle days and schedules are in rows 2:8.
# Class periods are in columns C (3) through H (8)
# User inputs cycle days in column A, rows 12:14.
# Output class schedule will be written to columns C:H, rows 12:14.

CYCLE_TABLE_START_ROW = 2
CYCLE_TABLE_END_ROW = 8
CLASS_TABLE_START_COL = 3 # C
CLASS_TABLE_END_COL = 8   # H
DATA_INPUT_START_ROW = 12
DATA_INPUT_END_ROW = 14
NUM_CYCLES = CYCLE_TABLE_END_ROW - CYCLE_TABLE_START_ROW + 1
PERIOD_COUNT = CLASS_TABLE_END_COL - CLASS_TABLE_START_COL + 1

# Build cycle-to-row mapping
day_to_row = {}
for i in range(NUM_CYCLES):
    cycle_day = ws.cell(row=CYCLE_TABLE_START_ROW + i, column=1).value
    day_to_row[cycle_day] = CYCLE_TABLE_START_ROW + i

for data_row in range(DATA_INPUT_START_ROW, DATA_INPUT_END_ROW + 1):
    input_cycle_day = ws.cell(row=data_row, column=1).value
    schedule_row = day_to_row.get(input_cycle_day, None)
    for col in range(CLASS_TABLE_START_COL, CLASS_TABLE_END_COL + 1):
        dest_cell = ws.cell(row=data_row, column=col)
        if schedule_row is not None:
            # Copy value from the appropriate cycle row and period column
            v = ws.cell(row=schedule_row, column=col).value
            dest_cell.value = v
        else:
            dest_cell.value = None  # Blank out if input is invalid

wb.save(output_path)
