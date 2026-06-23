import openpyxl
from openpyxl import load_workbook

def col_idx(col):
    return ord(col.upper()) - ord('A')

def is_header_row(row):
    return all(isinstance(v, str) for v in row)

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2-PRUNED/eval_seed42/eval_22-47_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2-PRUNED/eval_seed42/eval_22-47_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
header = rows[0]
data = rows[1:]

col_B = col_idx('B')
col_C = col_idx('C')
col_J = col_idx('J')
# Output columns
col_F = col_idx('F')
col_G = col_idx('G')
col_H = col_idx('H')

# Gather helper list from J (skip header+empties)
helper_items = []
seen_j = set()
for row in data:
    v = row[col_J]
    if v and v not in seen_j:
        helper_items.append(v)
        seen_j.add(v)

def unique_key(row):
    return (row[col_B], row[col_C])
found = set()
filtered = []
for row in data:
    b, c = row[col_B], row[col_C]
    if not b or not c:
        continue  # skip empty
    if is_header_row(row):
        continue
    k = unique_key(row)
    if k in found:
        continue
    found.add(k)
    filtered.append(row)
# Split by priority (helper) and not
main_group = []
secondary_group = []
if helper_items:
    for helper in helper_items:
        for row in filtered:
            if row[col_B] == helper:
                main_group.append(row)
    # All other rows (B not in helper_items)
    for row in filtered:
        if row[col_B] not in helper_items:
            secondary_group.append(row)
    output_rows = main_group + secondary_group
else:
    # Sort alphabetically by B only if helper column is empty
    output_rows = sorted(filtered, key=lambda r: r[col_B])

# Write to output region F2:H10 (max 9 entries)
for i in range(9):
    out_row = output_rows[i] if i < len(output_rows) else ('',) * ws.max_column
    for j, col in enumerate(range(col_F, col_H+1)):
        ws.cell(row=2+i, column=col+1, value=out_row[col])

# Additionally, sort only column H (G and H are output columns) lowest to highest in G2:H10
# Copy output for G and H only, sort by H
gh_rows = []
for i in range(2, 11): # rows 2 to 10
    g = ws.cell(row=i, column=col_G+1).value
    h = ws.cell(row=i, column=col_H+1).value
    if g or h:
        gh_rows.append((g, h))
# Sort gh_rows by H column (skip None for sort key)
gh_rows_sorted = sorted([row for row in gh_rows if row[1] is not None], key=lambda x: x[1])
# Fill remaining with blanks if less than 9
diff = 9 - len(gh_rows_sorted)
if diff > 0:
    gh_rows_sorted += [('', '')]*diff
for idx, (g, h) in enumerate(gh_rows_sorted):
    ws.cell(row=2+idx, column=col_G+1, value=g)
    ws.cell(row=2+idx, column=col_H+1, value=h)

wb.save(output_path)
