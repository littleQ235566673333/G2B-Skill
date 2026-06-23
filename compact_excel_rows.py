from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_6/regression_gate/before_pass/core_160-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_6/regression_gate/before_pass/core_160-6/output.xlsx'

wb = load_workbook(input_path)
ws = wb['SH']

data = []
# Read data rows from A2 to O15 (columns 1-15)
for row in ws.iter_rows(min_row=2, max_row=15, max_col=15, values_only=True):
    # Only append if row is not all Nones (i.e., not blank)
    if any(cell is not None for cell in row):
        data.append(row)
# Write compacted data back, starting from row 2
for i, row in enumerate(data, start=2):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)
# Clear rows after the last row of compacted data
for i in range(len(data)+2, 16):
    for j in range(1, 16):
        ws.cell(row=i, column=j, value=None)

wb.save(output_path)
