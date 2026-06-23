from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

def copy_font(font):
    return Font(
        name=font.name,
        size=font.size,
        bold=font.bold,
        italic=font.italic,
        vertAlign=font.vertAlign,
        underline=font.underline,
        strike=font.strike,
        color=font.color.rgb if font.color and hasattr(font.color, 'rgb') else None
    )

def copy_fill(fill):
    if isinstance(fill, PatternFill):
        return PatternFill(
            fill_type=fill.fill_type,
            fgColor=fill.fgColor.rgb if fill.fgColor and hasattr(fill.fgColor, 'rgb') else None,
            bgColor=fill.bgColor.rgb if fill.bgColor and hasattr(fill.bgColor, 'rgb') else None
        )
    return PatternFill(fill_type=None)

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/regression_gate/after_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Headers (assume row 2)
headers = {cell.value: i+1 for i, cell in enumerate(ws[2])}
row_start = 3
row_end = 15

expected_result_col = headers.get('Expected Result', 5)
age_year_col = headers.get('age (year)', 15)

# CATEGORY table search
cat_table_start_row = None
cat_table_year_col = None
cat_table_category_col = None
for row in ws.iter_rows(min_row=row_end+2, max_row=row_end+10):
    for cell in row:
        if cell.value == 'YEAR (age)':
            cat_table_year_col = cell.column
            cat_table_start_row = cell.row + 1
        if cell.value == 'CATEGORY':
            cat_table_category_col = cell.column
if cat_table_start_row and cat_table_year_col and cat_table_category_col:
    cat_year_col_letter = ws.cell(row=cat_table_start_row-1, column=cat_table_year_col).column_letter
    cat_category_col_letter = ws.cell(row=cat_table_start_row-1, column=cat_table_category_col).column_letter
    for r in range(row_start, row_end+1):
        lookup_formula = f'=XLOOKUP(O{r},${cat_year_col_letter}${cat_table_start_row}:${cat_year_col_letter}${cat_table_start_row+20},${cat_category_col_letter}${cat_table_start_row}:${cat_category_col_letter}${cat_table_start_row+20},"")'
        ws.cell(row=r, column=expected_result_col).value = lookup_formula
else:
    for r in range(row_start, row_end+1):
        ws.cell(row=r, column=expected_result_col).value = 'CATEGORY TABLE MISSING'

# Actual Age calculation (add after age (year))
actual_age_col = age_year_col + 1
report_date_col = headers.get('REPORT DATE')
birth_date_col = headers.get('DATE OF BIRTH')

age_header = ws.cell(row=2, column=age_year_col)
actual_age_header = ws.cell(row=2, column=actual_age_col)
actual_age_header.value = 'Actual Age'
actual_age_header.alignment = Alignment(horizontal='center', vertical='top')
actual_age_header.fill = copy_fill(age_header.fill)
actual_age_header.font = copy_font(age_header.font)

if report_date_col and birth_date_col:
    age_fill = copy_fill(age_header.fill)
    age_font = copy_font(age_header.font)
    for r in range(row_start, row_end+1):
        rep_date = ws.cell(row=r, column=report_date_col).coordinate
        dob = ws.cell(row=r, column=birth_date_col).coordinate
        age_formula = f'=INT(({rep_date} - {dob})/365.25)'
        cell = ws.cell(row=r, column=actual_age_col)
        cell.value = age_formula
        cell.number_format = '0'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, name=age_font.name, size=age_font.size, color=age_font.color)
        cell.fill = age_fill
wb.save(output_path)
