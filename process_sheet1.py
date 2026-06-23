import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_50631_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_50631_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_file)
sheet = wb['Sheet1']

# Extract holidays from column J (J7 downwards)
holidays = set()
row = 7
while True:
    val = sheet[f'J{row}'].value
    if val is None:
        break
    if isinstance(val, datetime):
        holidays.add(val.date())
    row += 1

start_date = sheet['B3'].value
end_date = sheet['F3'].value
if isinstance(start_date, datetime):
    start_date = start_date.date()
if isinstance(end_date, datetime):
    end_date = end_date.date()

# Style variables
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
calibri_font = Font(name='Calibri')

# Helper to calculate the nth workday after start_date
def nth_workday(start, n, holidays):
    current = start
    count = 0
    while count < n:
        if current.weekday() < 5 and current not in holidays:
            count += 1
            if count == n:
                break
        current += timedelta(days=1)
    return current

# Fill column H (H7:H38)
for idx, row in enumerate(range(7, 39)):
    # n is the (row-7)+1 th workday after start_date
    n = idx + 1
    workday = nth_workday(start_date, n, holidays)
    # Only include dates within and including the range
    if workday <= end_date:
        sheet[f'H{row}'].value = workday
        # Apply fill and font
        sheet[f'H{row}'].fill = yellow_fill
        sheet[f'H{row}'].font = calibri_font
    else:
        sheet[f'H{row}'].value = None

# Apply Calibri font to all relevant cells
for col in ['B', 'D', 'F', 'H']:
    for row in range(1, sheet.max_row+1):
        cell = sheet[f'{col}{row}']
        cell.font = calibri_font

# Hide gridlines (not directly possible, but set showGridLines property if possible)
sheet.sheet_view.showGridLines = False

wb.save(output_file)
print('Filled and saved')
