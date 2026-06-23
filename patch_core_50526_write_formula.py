import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# File paths
in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_1/regression_gate/before_pass/core_50526/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_1/regression_gate/before_pass/core_50526/output.xlsx'

wb = load_workbook(in_path)
ws = wb.active

# Figure out the number of columns/headers in the data
first_data_col = 2
last_data_col = ws.max_column
header_range = f'{get_column_letter(first_data_col)}1:{get_column_letter(last_data_col)}1'

# Figure out data rows (excluding header)
first_data_row = 2
last_data_row = ws.max_row
lookup_col = f'A{first_data_row}:A{last_data_row}'
lookup_cell = 'B6'

data_first = f'{get_column_letter(first_data_col)}{first_data_row}'
data_last = f'{get_column_letter(last_data_col)}{last_data_row}'

data_range = f'{data_first}:{data_last}'

# Compose array formula for the dynamic result list
# Will place in B9 (ROW(1:1)), B10 (ROW(2:2)) etc.
# Excel array formula using IF, SMALL, INDEX, MATCH:
# =IFERROR(INDEX($B$1:$E$1,SMALL(IF(INDEX($B$2:$E$6,MATCH($B$6,$A$2:$A$6,0),)>0,COLUMN($B$1:$E$1)-COLUMN($B$1)+1),ROW(1:1))),"")

header_range_str = f'${get_column_letter(first_data_col)}$1:${get_column_letter(last_data_col)}$1'
data_matrix_str = f'${get_column_letter(first_data_col)}${first_data_row}:${get_column_letter(last_data_col)}${last_data_row}'
lookup_col_str = f'$A${first_data_row}:$A${last_data_row}'

# B9, B10 ... formula
formula = f'=IFERROR(INDEX({header_range_str},SMALL(IF(INDEX({data_matrix_str},MATCH({lookup_cell},{lookup_col_str},0),)>0,COLUMN({header_range_str})-COLUMN({get_column_letter(first_data_col)}$1)+1),ROW(1:1))),"")'
formula2 = formula.replace('ROW(1:1)','ROW(2:2)')
ws['B9'] = formula
ws['B10'] = formula2

wb.save(out_path)
