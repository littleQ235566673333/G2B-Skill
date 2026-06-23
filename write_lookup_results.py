import openpyxl

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_1/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_1/regression_gate/after_pass/core_50526/output.xlsx'

# Load workbook and select Sheet1
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Get the lookup value from B6
lookup_value = ws['B6'].value

# Find the data header range from row 1, columns B to F (2 to 6)
headers = [ws.cell(row=1, column=col).value for col in range(2, 7)]

# Find the row where column A matches the lookup_value
match_row = None
for row in range(2, 4):  # data rows are row 2 and 3 (A2 and A3)
    if ws.cell(row=row, column=1).value == lookup_value:
        match_row = row
        break

results = []
if match_row is not None:
    for col in range(2, 7):  # B to F, 2 to 6
        cell_value = ws.cell(row=match_row, column=col).value
        if isinstance(cell_value, (int, float)) and cell_value > 0:
            results.append(ws.cell(row=1, column=col).value)

# Write results to B9 and B10
for i in range(2):
    ws.cell(row=9 + i, column=2, value=results[i] if i < len(results) else None)

wb.save(output_path)
