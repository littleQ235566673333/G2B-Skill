import pandas as pd
from openpyxl import load_workbook
from datetime import time
import numpy as np

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_1/group_52305/r1/evolve_52305/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_1/group_52305/r1/evolve_52305/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

df = pd.read_excel(input_path)

column_names = df.columns.tolist()
name_col = [c for c in column_names if 'name' in c.lower()][0]
time_col_candidates = [c for c in column_names if 'time' in c.lower() or 'hour' in c.lower()]
time_col = time_col_candidates[0] if time_col_candidates else None

start_row = 6
end_row = 24
start_col = 10  # J
end_col = 14    # N

unique_names = df[name_col].dropna().unique().tolist()
unique_names = unique_names[:(end_row-start_row+1)]

column_headers = df.columns[start_col-1:end_col].tolist()

time_start = time(21,30)
time_end = time(22,0)

def extract_time(val):
    if pd.isnull(val):
        return None
    if isinstance(val, pd.Timestamp):
        return val.time()
    if isinstance(val, str):
        try:
            parts = val.strip().split(':')
            h = int(parts[0])
            m = int(parts[1])
            return time(h, m)
        except Exception:
            return None
    if hasattr(val, 'time'):
        return val.time()
    return None

if not time_col:
    raise ValueError('No time column found in input.')
df['__time_obj'] = df[time_col].apply(extract_time)
df_in_time = df[df['__time_obj'].apply(lambda t: t is not None and time_start <= t <= time_end)]

other_cols = [c for c in column_names if c != name_col and c != time_col and df[c].dtype==object]

if len(other_cols) > 0:
    pivot = pd.pivot_table(
        df_in_time,
        index=name_col,
        columns=other_cols[0],
        values=other_cols[1] if len(other_cols)>1 else None,
        aggfunc='count',
        fill_value=0
    )
    if hasattr(pivot.columns, 'levels'):
        pivot.columns = ['_'.join(str(l) for l in col if str(l) != '') for col in pivot.columns.values]
    pivot = pivot.reindex(index=unique_names)
    matrix = pivot.values
else:
    matrix = df_in_time.groupby(name_col).size().reindex(unique_names, fill_value=0).to_frame('Count').values
    if matrix.shape[1] < (end_col-start_col+1):
        matrix = np.pad(matrix, ((0,0),(0,(end_col-start_col+1)-matrix.shape[1])), 'constant')

matrix = np.nan_to_num(matrix, nan=0).astype(int)

for i, name in enumerate(unique_names):
    for j in range(matrix.shape[1]):
        ws.cell(row=start_row+i, column=start_col+j, value=int(matrix[i][j]))

wb.save(output_path)
