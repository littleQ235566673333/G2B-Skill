from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_7/group_53647/r0/evolve_53647/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_7/group_53647/r0/evolve_53647/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

for row in range(7, 17):
    formula = f'=IF(AND(ISBLANK(C{row}), NOT(ISBLANK(D{row}))), D{row}, IF(AND(ISBLANK(D{row}), NOT(ISBLANK(C{row}))), C{row}, ""))'
    ws[f'E{row}'] = formula

wb.save(output_path)
