from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_8/group_39432/r0/evolve_39432/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_8/group_39432/r0/evolve_39432/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def get_purchase_layers(ws, row):
    # Return list of dict {'qty': qty, 'cost': cost} for each valid purchase layer
    # Purchases start at D/E = col 4/5, then F/G = 6/7, ...
    layers = []
    col = 4
    while True:
        qty_cell = ws.cell(row=row, column=col)
        cost_cell = ws.cell(row=row, column=col+1)
        if qty_cell.value is None and cost_cell.value is None:
            break
        if qty_cell.value is not None and cost_cell.value is not None:
            try:
                qty = float(qty_cell.value)
                cost = float(cost_cell.value)
                if qty > 0:
                    layers.append({'qty': qty, 'cost': cost})
            except (ValueError, TypeError):
                pass
        col += 2
    return layers

for row in range(2, 6):  # Fill M2:M5
    beg_inv = ws.cell(row=row, column=2).value
    beg_cost = ws.cell(row=row, column=3).value
    on_hand = ws.cell(row=row, column=12).value
    # Collect inventory layers in FIFO order
    layers = []
    if beg_inv is not None and beg_cost is not None and float(beg_inv) > 0:
        try:
            layers.append({'qty': float(beg_inv), 'cost': float(beg_cost)})
        except (ValueError, TypeError):
            pass
    layers += get_purchase_layers(ws, row)
    # FIFO selection of lots
    qty_needed = float(on_hand) if on_hand is not None else 0
    cost_sum = 0.0
    lot_pick = 0.0
    idx = 0
    while qty_needed > 0 and idx < len(layers):
        pick = min(qty_needed, layers[idx]['qty'])
        cost_sum += pick * layers[idx]['cost']
        lot_pick += pick
        qty_needed -= pick
        idx += 1
    # Write per unit cost
    if lot_pick > 0:
        per_unit_cost = cost_sum / lot_pick
        ws.cell(row=row, column=13, value=round(per_unit_cost, 2))
    else:
        ws.cell(row=row, column=13, value=None)

wb.save(output_path)
