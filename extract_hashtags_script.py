from openpyxl import load_workbook
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/regression_gate/after_fix/core_57113/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/regression_gate/after_fix/core_57113/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def extract_hashtags(text):
    if not isinstance(text, str):
        return []
    # Grab hashtags, remove the #
    return [tag[1:] for tag in re.findall(r'#\w+', text)]

rows = [2, 3]
start_col = 2  # B
max_tags = 2   # Output to B and C

for row in rows:
    tweet = ws.cell(row=row, column=1).value
    hashtags = extract_hashtags(tweet)
    for i in range(max_tags):
        val = hashtags[i] if i < len(hashtags) else ""
        ws.cell(row=row, column=start_col + i, value=val)

wb.save(output_path)
