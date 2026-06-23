from openpyxl import load_workbook
import pandas as pd

# File paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_5/group_54474/r1/evolve_54474/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_5/group_54474/r1/evolve_54474/output.xlsx"

def main():
    # Load workbook and the sheets
    wb = load_workbook(input_path)
    ws_whp = wb["WHP"]
    # Use pandas to load both sheets for flexible data extraction
    df_whp = pd.read_excel(input_path, sheet_name="WHP", header=None)
    df_data = pd.read_excel(input_path, sheet_name="WHP DATA")

    # Try to extract site list from column D (col 4, index 3) for rows 7 and 8 (Excel is 1-based)
    site_cells = [ws_whp[f"D{row}"].value for row in range(7, 9)]

    # If all Nones, use pandas fallback (sometimes merged or weird formatting can make excel None)
    if not any(site_cells):
        site_cells = list(df_whp.iloc[6:8,3])  # 0-based index (rows 7-8, col D)

    # For each site, extract corresponding info from WHP data
    # We'll extract columns 2,3,4 from WHP DATA for the matching site
    result_rows = []
    for site in site_cells:
        if pd.notna(site) and site in df_data.iloc[:, 0].values:
            row = df_data[df_data.iloc[:, 0] == site].iloc[0]
            result = list(row.iloc[1:4])  # B, C, D
        else:
            result = ["", "", ""]
        result_rows.append(result)

    # Write the result to E7:G8 in WHP
    for idx, row_values in enumerate(result_rows):
        excel_row = 7 + idx
        for col_offset, val in enumerate(row_values):
            ws_whp.cell(row=excel_row, column=5 + col_offset, value=val)

    wb.save(output_path)

if __name__ == "__main__":
    main()
