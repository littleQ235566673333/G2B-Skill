import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_416-27_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_416-27_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Find the first row in col A with value, and last row
first_val_row = None
for row_idx in range(1, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=1).value
    if val not in (None, ""):
        first_val_row = row_idx
        break

if first_val_row is None:
    first_val_row = 1

rows_to_delete = []
# Collect indices to delete from last to first to avoid row shifting
for row_idx in range(ws.max_row, first_val_row - 1, -1):
    val = ws.cell(row=row_idx, column=1).value
    if val in (None, ""):
        ws.delete_rows(row_idx)

# Now collect values and write up to 22 rows in A3:A24
values = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
    if row[0] not in (None, ""):
        values.append(row[0])

# Write these values to A3:A24
for i in range(22):
    cell = ws.cell(row=3+i, column=1)
    cell.value = values[i] if i < len(values) else None

wb.save(output_path)
