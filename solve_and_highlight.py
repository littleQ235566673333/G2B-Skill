import openpyxl
from openpyxl.styles import PatternFill
from itertools import combinations

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_254-34_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_254-34_tc1/output.xlsx'

TARGET_MIN = 994108
TARGET_MAX = 994112
HIGHLIGHT = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

def find_sum_combination(numbers, coords, target_min, target_max):
    # Try all combinations from 2 up to len(numbers) (limited for feasibility)
    for r in range(2, min(len(numbers), 20)):
        for idxs in combinations(range(len(numbers)), r):
            s = sum(numbers[i] for i in idxs)
            if target_min <= s <= target_max:
                return [coords[i] for i in idxs], s
    return None, None

wb = openpyxl.load_workbook(input_path)

# Delete Sheet3 if it exists
if 'Sheet3' in wb.sheetnames:
    del wb['Sheet3']

ws = wb['Before']

# Gather all value cells (numeric and not blank)
numbers = []
coords = []
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, (int, float)):
            numbers.append(cell.value)
            coords.append(cell.coordinate)

# Find a valid combination
combo_coords, found_sum = find_sum_combination(numbers, coords, TARGET_MIN, TARGET_MAX)
if combo_coords is None:
    # fallback: try single-value matches
    for i, val in enumerate(numbers):
        if TARGET_MIN <= val <= TARGET_MAX:
            combo_coords = [coords[i]]
            found_sum = val
            break

# Highlight and mark sum in C2
if combo_coords:
    for coord in combo_coords:
        ws[coord].fill = HIGHLIGHT
    ws['C2'].value = found_sum
else:
    ws['C2'].value = 'No combination found'

wb.save(output_path)
