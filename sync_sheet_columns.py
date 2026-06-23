import openpyxl

def sync_column_b(input_path, output_path):
    # Load the workbook and sheets
    wb = openpyxl.load_workbook(input_path)
    items_sheet = wb['ITEMS']
    sheet1 = wb['SHEET1']
    sheet2 = wb['SHEET2']

    # Read column B from ITEMS
    items_col_b = {}
    for row in range(2, items_sheet.max_row + 1):
        cell_value = items_sheet.cell(row=row, column=2).value
        items_col_b[row] = cell_value

    # Function to update column B for a given sheet based on ITEMS
    def update_sheet_col_b(sheet, max_row):
        for row in range(2, max_row + 1):
            item_value = items_col_b.get(row)
            if item_value is not None:
                sheet.cell(row=row, column=2).value = item_value

    # Update SHEET1 column B
    update_sheet_col_b(sheet1, sheet1.max_row)
    # Update SHEET2 column B
    update_sheet_col_b(sheet2, sheet2.max_row)

    # Save to output
    wb.save(output_path)

# File paths
input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/after_pass/core_302-1/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/after_pass/core_302-1/output.xlsx'

# Run the sync
sync_column_b(input_fp, output_fp)
