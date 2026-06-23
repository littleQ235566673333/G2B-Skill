import openpyxl

# Load the workbook
input_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_42198_tc1/input.xlsx"
output_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_42198_tc1/output.xlsx"
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# We assume data starts at row 2 (headers at row 1), goes through row 7
eval_rows = range(2, 8)  # 2 to 7 inclusive

for end_row in eval_rows:
    found = None
    # Look cumulatively from 2 to end_row (inclusive)
    for r in range(2, end_row + 1):
        a_val = ws[f"A{r}"].value
        b_val = ws[f"B{r}"].value
        if a_val == "Potato" and b_val is False:
            found = "Worst"
            break
    if found is None:
        for r in range(2, end_row + 1):
            a_val = ws[f"A{r}"].value
            b_val = ws[f"B{r}"].value
            if a_val == "Tomato" and b_val is False:
                found = "Ignore"
                break
    if found is None:
        for r in range(2, end_row + 1):
            a_val = ws[f"A{r}"].value
            b_val = ws[f"B{r}"].value
            if a_val == "Pickle" and b_val is False:
                found = "Bad"
                break
    if found is None:
        found = "Good"
    ws[f"C{end_row}"].value = found

wb.save(output_path)
