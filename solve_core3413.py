import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_1/regression_gate/after_pass/core_3413/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_1/regression_gate/after_pass/core_3413/output.xlsx'

wb = load_workbook(input_path)
sheet = wb.active

data = []
# Read until a blank or end for A/B/C
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
    if row[0] is None and row[1] is None and row[2] is None:
        break
    data.append(row)
df = pd.DataFrame(data, columns=['Department', 'RU', 'Value'])

for output_row in range(3, 7):
    dept = sheet[f'E{output_row}'].value
    ru = sheet[f'F{output_row}'].value
    mask = (df['Department'] == dept) & (df['RU'] == ru)
    s = df.loc[mask, 'Value'].sum()
    # If no match for Department + RU, sum for just Department
    if s == 0:
        s = df.loc[df['Department'] == dept, 'Value'].sum()
    sheet[f'G{output_row}'] = s

wb.save(output_path)
