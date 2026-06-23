import openpyxl

def is_row_empty(row):
    """Check if all cells in a given row (excluding None) are empty"""
    return all((cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == '')) for cell in row)

# File paths
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_409-45_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_409-45_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb[wb.sheetnames[0]]  # use the first sheet, not the second

I2_value = ws['I2'].value

# Gather all data in the output area ('DATA'!A1:E19) - assuming this means rows 1-19 and cols A-E
start_row, end_row = 1, 19
start_col, end_col = 1, 5   # Columns A-E are 1-5
rows = list(ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col))

keep_rows = [rows[0]] # always keep header

if I2_value:
    # Only process rows where column B matches I2_value
    for row in rows[1:]:
        id_val = row[1].value # column B is index 1 in slice
        if id_val == I2_value:
            if not is_row_empty(row):
                keep_rows.append(row)
else:
    # Remove all empty rows from all ranges
    for row in rows[1:]:
        if not is_row_empty(row):
            keep_rows.append(row)

# Pad to maintain total 19 rows (if less after deletion), otherwise truncate
while len(keep_rows) < 19:
    # create blank row as list of openpyxl.cell.Cell objects with blank values
    blank_row = [openpyxl.cell.cell.Cell(ws, row=0, column=c) for c in range(1, 6)]
    keep_rows.append(blank_row)
if len(keep_rows) > 19:
    keep_rows = keep_rows[:19]

# Write back to ws['A1':'E19']
for r, row in enumerate(keep_rows, start=1):
    for c, cell in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=cell.value)

wb.save(output_path)
