from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun1/eval_42902_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun1/eval_42902_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Read all values from column A, skipping empty cells
colA = [cell.value for cell in ws['A'] if cell.value is not None]
# How many complete groups of 3?
num_records = len(colA) // 3
records = [colA[i*3:(i+1)*3] for i in range(num_records)]

for i, record in enumerate(records):
    for j in range(3):
        val = record[j] if j < len(record) else ""
        ws.cell(row=1+i, column=4+j, value=val)

wb.save(output_path)
