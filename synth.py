import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_57989_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_57989_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']
rows = [row for row in ws.iter_rows(values_only=True)]
header_row = rows[0]
weekdays = ['Friday', 'Saturday', 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday']
weekday_to_columns = {wd: [] for wd in weekdays}
for i, v in enumerate(header_row):
    if v in weekday_to_columns:
        weekday_to_columns[v].append(i)
drivers = [(idx, row[0]) for idx, row in enumerate(rows) if isinstance(row[0], str) and row[0].startswith('Driver')]
# Build the synthesis table
synthesis = []
for driver_idx, driver_name in drivers:
    row_counts = [driver_name]
    for wd in weekdays:
        count = 0
        for col_idx in weekday_to_columns[wd]:
            cell_value = rows[driver_idx][col_idx]
            if cell_value is not None and str(cell_value).strip() != '':
                count += 1
        row_counts.append(count)
    synthesis.append(row_counts)
# Write answer to B25:H43
for i, synth_row in enumerate(synthesis):
    for j, value in enumerate(synth_row):
        ws.cell(row=25+i, column=2+j, value=value)
wb.save(output_path)
print('Done: written synthesis to output.')
