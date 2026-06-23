import openpyxl
import sys

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/group_130-9/r3/evolve_130-9/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/group_130-9/r3/evolve_130-9/output.xlsx'

wb = openpyxl.load_workbook(input_path)
wsrc = wb['cdnr']
wout = wb['b2b, sez, de']

# Find header row and columns (assume headers in row 5 in both src/dest as per sample output region)
src_headers = [c.value for c in next(wsrc.iter_rows(min_row=5, max_row=5))]
dest_headers = [c.value for c in next(wout.iter_rows(min_row=5, max_row=5))]

# Map dest col index (1-based) to dest header
dest_header_map = {h: i+1 for i, h in enumerate(dest_headers)}
# Map source col names to index (0-based)
src_header_map = {h: i for i, h in enumerate(src_headers)}

# Where does data begin & first empty row in dest?
start_row_dest = 7 # Per user info, row 7 is the first empty

# Collect source data rows (skip header row, start at row 6)
data_rows = list(wsrc.iter_rows(min_row=6, values_only=True))
# Filter out completely empty (None) rows
data_rows = [row for row in data_rows if any(x is not None for x in row)]

# Find first empty row in destination sheet (row 7 or after, row is empty if all columns empty)
first_empty_dest_row = start_row_dest
for r in range(start_row_dest, wout.max_row + 2):
    if all(wout.cell(row=r, column=col).value in (None, '') for col in range(1, wout.max_column + 1)):
        first_empty_dest_row = r
        break

# Prepare columns for special handling
l_idx = dest_header_map.get('L', None)
m_idx = dest_header_map.get('M', None)
h_idx = dest_header_map.get('H', None)
i_idx = dest_header_map.get('I', None)

amount_colnames = []
for col_name in ['L', 'M']:
    if col_name in dest_headers:
        amount_colnames.append(col_name)

# For each source row, create a destination row mapped by header
for nrow, src_row in enumerate(data_rows):
    dest_row = [''] * len(dest_headers)
    # Map each column with matching header
    for h in dest_headers:
        if h and h in src_header_map:
            dest_row[dest_header_map[h]-1] = src_row[src_header_map[h]]
    # Negative for cols L and M (if present)
    for colnm in amount_colnames:
        idx = dest_header_map[colnm]-1
        val = dest_row[idx]
        try:
            if val is not None and val != '':
                dest_row[idx] = -abs(float(val)) if float(val) > 0 else float(val)
        except:
            pass  # keep as is if not a number
    # Set col H to blank, I to 'Credit Note'
    if h_idx is not None:
        dest_row[h_idx-1] = ''
    if i_idx is not None:
        dest_row[i_idx-1] = 'Credit Note'
    # Write mapped row to target
    for j, val in enumerate(dest_row):
        # Copy number format if present in original destination header
        cell = wout.cell(row=first_empty_dest_row+nrow, column=j+1)
        cell.value = val
        # preserve number format for L/M (float) if present in sample row 6
        if (dest_headers[j] in amount_colnames) and (wout.cell(row=6, column=j+1).number_format != 'General'):
            cell.number_format = wout.cell(row=6, column=j+1).number_format

# Save the result
wb.save(output_path)
print(f"Rows from 'cdnr' appended to 'b2b, sez, de' starting at row {first_empty_dest_row} (headers matched by text, L/M values negative, H blank, I='Credit Note').")
