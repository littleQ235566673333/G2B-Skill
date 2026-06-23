import openpyxl

# Define paths
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_55468_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_55468_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Extract the user criteria
crit_vertical_1 = ws['AE4'].value  # main Y
crit_vertical_2 = ws['AD4'].value  # sec Y
crit_horizontal_1 = ws['AC4'].value  # main X
crit_horizontal_2 = ws['AB4'].value  # sec X

# Find data boundaries (assume as per user and formula that data is C5:Z10)
data_min_row, data_max_row = 5, 10
data_min_col, data_max_col = 3, 26  # C:Z

def colnum_to_letter(num):
    # 1 -> A
    string = ''
    while num:
        num, rem = divmod(num-1, 26)
        string = chr(rem + ord('A')) + string
    return string

# --- Find the vertical (Y) matching row ---
# Main Y axis (row headers in A5:A10)
y_header_values = [ws[f'A{r}'].value for r in range(data_min_row, data_max_row+1)]
try:
    v1_idx = y_header_values.index(crit_vertical_1)
except ValueError:
    v1_idx = None

# Sec Y axis (row headers in B5:B10)
y2_header_values = [ws[f'B{r}'].value for r in range(data_min_row, data_max_row+1)]
try:
    v2_idx = y2_header_values.index(crit_vertical_2)
except ValueError:
    v2_idx = None

# Both criteria must match on the same row
row_index = None
for r in range(data_min_row, data_max_row+1):
    if ws[f'A{r}'].value == crit_vertical_1 and ws[f'B{r}'].value == crit_vertical_2:
        row_index = r
        break

# --- Find the horizontal (X) matching column ---
# Main X axis (column headers in C4:Z4, row 4)
x_headers = [ws[f'{colnum_to_letter(col)}4'].value for col in range(data_min_col, data_max_col+1)]
# Sec X axis (column headers in C3:Z3, row 3)
x2_headers = [ws[f'{colnum_to_letter(col)}3'].value for col in range(data_min_col, data_max_col+1)]

col_index = None
for i, col in enumerate(range(data_min_col, data_max_col+1)):
    if x_headers[i] == crit_horizontal_1 and x2_headers[i] == crit_horizontal_2:
        col_index = col
        break

# Default output if not found
value = None
if row_index is not None and col_index is not None:
    cell = ws.cell(row=row_index, column=col_index)
    value = cell.value

# Write the result into AE5
ws['AE5'] = value

# Save
wb.save(output_path)
