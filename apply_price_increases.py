from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_2/group_44017/r2/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_2/group_44017/r2/evolve_44017/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

start_row = 14
end_row = 42
start_col = 30  # AD
num_months = 12
incs = ['M', 'N', 'O', 'P']
for row in range(start_row, end_row + 1):
    for m in range(num_months):
        col = start_col + m
        col_letter = get_column_letter(col)
        month_date_cell = f'{col_letter}$9'
        # Build the cumulative multiplier formula
        multiplier = '1'
        for i, inc_col in enumerate(incs):
            # Effective date for each increase
            if i == 0:
                effective = f'$L{row}'
            else:
                effective = f'EDATE($L{row},{i}*$J{row})'
            multiplier += f'*IF({month_date_cell}>={effective},1+${inc_col}{row},1)'
        formula = f'=IF({month_date_cell}<$L{row},"",$W{row}*{multiplier})'
        ws[f'{col_letter}{row}'] = formula

wb.save(output_path)
print('Done')
