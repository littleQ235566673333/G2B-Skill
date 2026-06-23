from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_57590_tc1/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_57590_tc1/output.xlsx'
wb = load_workbook(infile)
ws = wb['Sheet1']

headers = [cell.value for cell in ws[1]]
col_qty = get_column_letter(headers.index('QTY') + 1)
col_date = get_column_letter(headers.index('Date') + 1)
# Use a smaller range, e.g. 2:100, since real data stops before 100
formula = '=SUMIFS({qty}2:{qty}100,{date}2:{date}100,">="&A26,{date}2:{date}100,"<="&EOMONTH(A26,0))'.format(qty=col_qty, date=col_date)
ws['B26'] = formula
wb.save(outfile)
print('Formula written:', formula)
