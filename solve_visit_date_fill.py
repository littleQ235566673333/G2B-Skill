import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_1/regression_gate/after_pass/core_9726/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_1/regression_gate/after_pass/core_9726/output.xlsx'

# Read in both tables fully as column blocks
# Table 1: columns 0,1,2
# Table 2: columns 7,8,9
full_df = pd.read_excel(input_fp)
visits = full_df.iloc[:, [0, 1, 2]].dropna(subset=[full_df.columns[0], full_df.columns[2]])
visits.columns = ['Student ID', 'Visit number', 'Visit date']
visits['Visit date'] = pd.to_datetime(visits['Visit date'], errors='coerce', dayfirst=True)
subs = full_df.iloc[:, [7, 8]].dropna(subset=[full_df.columns[7], full_df.columns[8]])
subs.columns = ['Student ID', 'Submission date']
subs['Submission date'] = pd.to_datetime(subs['Submission date'], errors='coerce', dayfirst=True)

def get_last_visit(student, sub_date):
    filtered = visits[(visits['Student ID'] == student) & (visits['Visit date'] < sub_date)]
    if filtered.empty:
        return pd.NaT
    return filtered['Visit date'].max()

results = []
for idx, row in subs.iterrows():
    student = row['Student ID']
    sub_date = row['Submission date']
    last_visit = get_last_visit(student, sub_date)
    results.append(last_visit)

# Write the result to the 10th column (J), starting from row 2, aligned with Submission date (column 9)
wb = load_workbook(input_fp)
ws = wb[wb.sheetnames[0]]

for i in range(2, ws.max_row + 1):
    student = ws.cell(row=i, column=8).value
    sub_date = ws.cell(row=i, column=9).value
    if student and sub_date and len(results) > 0:
        v = results.pop(0)
        ws.cell(row=i, column=10, value=v.date() if pd.notnull(v) else None)

wb.save(output_fp)
