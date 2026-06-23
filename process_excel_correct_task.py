import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_7665_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_7665_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

vertical_values = []
for cell in ws['H']:
    if cell.row == 1:
        continue  # skip header
    if cell.value is not None:
        vertical_values.append(cell.value)
unique_sorted = sorted(set(vertical_values))
start_col = openpyxl.utils.column_index_from_string('Q')
for idx, value in enumerate(unique_sorted):
    ws.cell(row=2, column=start_col + idx, value=value)
wb.save(output_path)
