import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import re

def is_blank_row(row):
    return all((cell.value is None or str(cell.value).strip() == '') for cell in row)

def is_formula(cell):
    return cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.strip().startswith('='))

def convert_text_to_formula(cell):
    if isinstance(cell.value, str) and cell.value.strip().startswith('='):
        cell.value = cell.value.strip()
        cell.data_type = 'f'


def get_blocks(ws, start_row=1, end_row=85):
    blocks = []
    in_block = False
    block_start = None
    for i in range(start_row, end_row+1):
        row = ws[i]
        if is_blank_row(row):
            if in_block:
                blocks.append((block_start, i-1))
                in_block = False
        else:
            if not in_block:
                block_start = i
                in_block = True
    if in_block:
        blocks.append((block_start, end_row))
    return blocks


def set_number_format(cell):
    if cell.column in [4,5,6]:  # D,E,F
        val = cell.value
        # Only set number format if it's not a formula
        if not is_formula(cell):
            try:
                num = float(val)
                if num == int(num):
                    cell.number_format = '0'
                else:
                    cell.number_format = '0.0'
            except:
                pass
        else:
            cell.number_format = '0.0'


def style_df_cells(ws, r1, r2):
    bold_font = Font(bold=True)
    center = Alignment(horizontal='center')
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=4, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)
            set_number_format(cell)


def sort_block(ws, r1, r2):
    data = []
    for row_idx in range(r1, r2+1):
        row = ws[row_idx]
        # Convert any text formulas to real formulas
        for col_idx in [3,4,5]:
            convert_text_to_formula(row[col_idx])
        val_d = row[3].value
        if is_formula(row[3]):
            val_d_eval = -float('inf') # openpyxl can't evaluate
        else:
            try:
                val_d_eval = float(val_d)
            except:
                val_d_eval = -float('inf')
        data.append((val_d_eval, [cell.value for cell in row], [cell for cell in row]))
    # Sort
    data_sorted = sorted(data, key=lambda x: x[0], reverse=True)
    for i, (_, values, orig_cells) in enumerate(data_sorted):
        for j, value in enumerate(values):
            cell = ws.cell(row=r1 + i, column=j+1, value=value)
            # Assign formula if present
            if is_formula(orig_cells[j]):
                cell.value = orig_cells[j].value
                cell.data_type = orig_cells[j].data_type
            # Only assign specific styling to D–F
            if j+1 in [4,5,6]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                set_number_format(cell)
    # Style D-F
    style_df_cells(ws, r1, r1+len(data_sorted)-1)


def preserve_column_widths(ws, ws_out):
    for col in range(1,9):
        col_letter = get_column_letter(col)
        ws_out.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width

inp = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_191-40_tc1/input.xlsx"
outp = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_191-40_tc1/output.xlsx"
wb = openpyxl.load_workbook(inp)
ws = wb['Sheet1']
blocks = get_blocks(ws, start_row=1, end_row=85)
for block in blocks:
    r1, r2 = block
    sort_block(ws, r1, r2)
for block in blocks:
    r1, r2 = block
    style_df_cells(ws, r1, r2)
preserve_column_widths(ws, ws)
wb.save(outp)
