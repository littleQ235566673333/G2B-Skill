import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_120-24_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_120-24_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Column indices (1-based)
col_BL = 64  # BL
col_BG = 59  # BG
col_AY = 51  # AY
col_BN = 65  # BN

# The range is from row 2 to 45 inclusive
for row in range(2, 46):
    val_BL = ws.cell(row=row, column=col_BL).value
    val_BG = ws.cell(row=row, column=col_BG).value
    val_AY = ws.cell(row=row, column=col_AY).value
    val_BN = ws.cell(row=row, column=col_BN).value

    # Criteria
    is_lag = (val_BL == 'LAG')
    is_target_bg = (val_BG in [
        'OCOGS - Spares - Transfer Price Overhead',
        'OFSS - ORCL Consulting Prod Cost I/C'])
    is_target_ay = (val_AY in ['BOA_033E', 'BOA_011G'])

    if is_lag and is_target_bg:
        if is_target_ay:
            ws.cell(row=row, column=col_BN).value = 'OCOGS - Spares - Transfer Price Overhead'
        else:
            ws.cell(row=row, column=col_BN).value = 'OFSS - ORCL Consulting Prod Cost I/C'

    val_BN_mod = ws.cell(row=row, column=col_BN).value
    # Fill color #0070C0 for BL if 'LAG' and BN populated
    if val_BL == 'LAG' and val_BN_mod not in [None, '']:
        fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        ws.cell(row=row, column=col_BL).fill = fill

wb.save(output_path)
