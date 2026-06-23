from openpyxl import load_workbook

INPUT_PATH = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_9/task_44389/r3/evolve_44389/input.xlsx'
OUTPUT_PATH = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_9/task_44389/r3/evolve_44389/output.xlsx'

wb = load_workbook(INPUT_PATH)
ws = wb['Sheet1']

# Columns B:N = 2:14 in Excel/1-indexed, 1:13 in Python zero-based
header_row = [cell.value for cell in ws[1]]
col_headers = header_row[1:14]  # B to N

for i in range(2, 8):  # Rows 2 to 7, inclusive
    values = []
    for col in range(2, 15):  # B (2) to N (14)
        val = ws.cell(row=i, column=col).value
        # Convert to float if possible
        try:
            num = float(val)
            if num > 0:
                values.append((col_headers[col-2], num))
        except (TypeError, ValueError):
            pass
    if values:
        min_val = min(v for h, v in values)
        min_headers = [h for h, v in values if v == min_val]
        ws.cell(row=i, column=15).value = ','.join(min_headers)
    else:
        ws.cell(row=i, column=15).value = ''

wb.save(OUTPUT_PATH)
