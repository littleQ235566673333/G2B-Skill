from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_4/group_44017/r2/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_4/group_44017/r2/evolve_44017/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Data']

start_row, end_row = 14, 42
start_col, end_col = 30, 41  # AD=30, AO=41

for row in range(start_row, end_row + 1):
    base = f'$W{row}'
    effdate = f'$L{row}'
    freq = f'$J{row}'
    increases = [f'$M{row}', f'$N{row}', f'$O{row}', f'$P{row}']
    for i, col in enumerate(range(start_col, end_col + 1)):
        col_letter = get_column_letter(col)
        date_cell = f'{col_letter}9'
        # Figure out which wave applies
        wave_idx = (
            f'IF({date_cell}>={effdate},MIN(4,MAX(1,INT((DATEDIF({effdate},{date_cell},"m")/{freq})+1)),0),0)'
        )
        mult = base
        # Cumulative increases: multiply by (1+increase%) for each wave that applies
        for j in range(4):
            mult += f'*(1+IF({wave_idx}>={j+1},IF(ISNUMBER({increases[j]}),{increases[j]},0),0))'
        # Only show value if date >= effective date
        formula = f'=IF({date_cell}<{effdate},"",ROUND({mult},2))'
        ws[f'{col_letter}{row}'].value = formula

wb.save(output_path)
