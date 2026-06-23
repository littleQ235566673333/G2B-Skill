import pandas as pd
from openpyxl import load_workbook
import numpy as np

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_2/group_387-16/r0/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_2/group_387-16/r0/evolve_387-16/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Column map: 'Value'=A=1, 'Binaries'=B=2, 'Result values'=D=4
# Get 'Value' and 'Binaries' data rows (A3:A18, B3:B18)
values, binaries = [], []
for row_idx in range(3, 19):  # Inclusive, for rows 3 through 18 (A2:D18 is target region)
    values.append(ws.cell(row=row_idx, column=1).value)
    binaries.append(ws.cell(row=row_idx, column=2).value)

# Get 'Result values' from D5 downward (find first empty row)
result_values = []
row_idx = 5
while True:
    v = ws.cell(row=row_idx, column=4).value
    if v is None:
        break
    result_values.append(v)
    row_idx += 1

# Remove only one instance in order of each result_value (from values AND binaries)
mask = np.ones(len(values), dtype=bool)
for rv in result_values:
    for i, v in enumerate(values):
        if mask[i] and v == rv:
            mask[i] = False
            break
values_kept = [v for i,v in enumerate(values) if mask[i]]
binaries_kept = [b for i,b in enumerate(binaries) if mask[i]]
# Shift up (fill with empty for length alignment)
while len(values_kept) < len(values):
    values_kept.append('')
    binaries_kept.append('')

# Write back the filtered columns to A3:A18, B3:B18
for idx in range(len(values)):
    ws.cell(row=idx+3, column=1, value=values_kept[idx])
    ws.cell(row=idx+3, column=2, value=binaries_kept[idx])

# Solver result: sum of entries in column A3:A18 (excluding '')
solver_result = sum([v for v in values_kept if isinstance(v, (int, float))])
# Target value is E2
try:
    target_val = ws.cell(row=2, column=5).value
    difference = solver_result - target_val if target_val is not None else ''
except Exception:
    target_val = None
    difference = ''

# Write results: C20 = Solver result, D20 = sum, C21 = Difference, D21 = diff
ws['C20'] = 'Solver result'
ws['D20'] = solver_result
ws['C21'] = 'Difference'
ws['D21'] = difference

wb.save(output_path)
