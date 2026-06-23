import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_191-40_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_191-40_tc1/output.xlsx'
sheet_name = 'Sheet1'

# Load workbook
wb = openpyxl.load_workbook(input_path)
ws = wb[sheet_name]

# Get column widths
col_widths = {col: ws.column_dimensions[get_column_letter(col)].width for col in range(1, 9)}

# Helper to check if a row is blank (all empty in A-H)
def is_blank_row(row):
    return all((cell.value is None or str(cell.value).strip() == '') for cell in row)

rows = list(ws.iter_rows(min_row=1, max_row=85, min_col=1, max_col=8))

blocks = []
current_block = []

for row in rows:
    if is_blank_row(row):
        if current_block:
            blocks.append(current_block)
            current_block = []
        blocks.append([row])  # Blank row is its own block
    else:
        current_block.append(row)
if current_block:
    blocks.append(current_block)

# Clean up blocks: merge consecutive blank-row blocks to just one blank row
clean_blocks = []
for b in blocks:
    if len(b) == 1 and is_blank_row(b[0]):
        if not (clean_blocks and len(clean_blocks[-1]) == 1 and is_blank_row(clean_blocks[-1][0])):
            clean_blocks.append(b)
    else:
        clean_blocks.append(b)

# Sort and reformat blocks
sorted_rows = []

for b in clean_blocks:
    if len(b) == 1 and is_blank_row(b[0]):
        sorted_rows.append(b[0])  # Just append blank row
    else:
        # Sort block by column D (index 3, openpyxl is 0-based index)
        block = b
        def get_D_value(row):
            cell_D = row[3]
            # Prefer value (should be the calculated value unless text formula)
            try:
                v = cell_D.value
                if v is None:
                    return float('-inf')
                if isinstance(v, str):
                    try:
                        return float(v)
                    except:
                        return float('-inf')  # sort text formula last
                return float(v)
            except:
                return float('-inf')
        sorted_block = sorted(block, key=get_D_value, reverse=True)
        sorted_rows.extend(sorted_block)
        # Insert a blank row (will be handled later)

# Rebuild rows in Sheet1
for i in range(85):
    for j in range(8):
        ws.cell(row=i+1, column=j+1).value = None

for idx, row in enumerate(sorted_rows):
    for j, cell in enumerate(row):
        new_cell = ws.cell(row=idx+1, column=j+1)
        # Preserve formula
        if cell.data_type == 'f':
            new_cell.value = cell.value
        else:
            new_cell.value = cell.value
        # D–F: apply number formats
        if 3 <= j <= 5:  # D, E, F
            v = new_cell.value
            if isinstance(v, (int, float)):
                if isinstance(v, int) or ((isinstance(v, float) and v == int(v))):
                    new_cell.number_format = '0'
                else:
                    new_cell.number_format = '0.0'
            else:
                new_cell.number_format = 'General'
            new_cell.font = Font(bold=True)
            new_cell.alignment = Alignment(horizontal='center')
        else:
            # Preserve formatting
            new_cell.font = cell.font
            new_cell.alignment = cell.alignment
        # Copy fill, border, etc
        new_cell.fill = cell.fill
        new_cell.border = cell.border

# Restore column widths
for col in range(1, 9):
    ws.column_dimensions[get_column_letter(col)].width = col_widths[col]

# Save
wb.save(output_path)
