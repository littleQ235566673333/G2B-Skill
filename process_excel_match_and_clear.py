import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/regression_gate/after_pass/core_493-18/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/regression_gate/after_pass/core_493-18/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Step 1: Build set of values in column F (ignore header)
col_f_values = set()
for row in ws.iter_rows(min_row=2, max_col=6, max_row=ws.max_row):
    val = row[5].value
    if val is not None:
        col_f_values.add(val)

# Step 2: Clear A/B/C for rows where A not in F, else collect retained data
data_to_keep = []
header = [ws.cell(row=1, column=col).value for col in range(1,7)]
for i in range(2, ws.max_row+1):
    a_val = ws.cell(row=i, column=1).value
    if a_val in col_f_values:
        row_values = [ws.cell(row=i, column=c).value for c in range(1,7)]
        data_to_keep.append(row_values)
    else:
        for c in range(1,4):
            ws.cell(row=i, column=c).value = None

# Step 3: Remove blank rows from A/B/C and shift up.
for idx, vals in enumerate(data_to_keep):
    for c in range(1,7):
        ws.cell(row=idx+2, column=c).value = vals[c-1]
# clear below shifted
for r in range(len(data_to_keep)+2, ws.max_row+1):
    for c in range(1,4):
        ws.cell(row=r, column=c).value = None
# Restore headers
for c in range(1,7):
    ws.cell(row=1, column=c).value = header[c-1]

# Step 4: Ensure autofilter is set
ws.auto_filter.ref = 'A1:C{}'.format(len(data_to_keep)+1)

wb.save(output_path)
