from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_547-43_s1/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_547-43_s1/output.xlsx'

wb = load_workbook(INPUT)
print(wb.sheetnames)
for ws in wb.worksheets:
    print('SHEET', ws.title, ws.max_row, ws.max_column)
    for r in range(1, min(ws.max_row, 25)+1):
        vals = [ws.cell(r,c).value for c in range(1, min(ws.max_column, 12)+1)]
        print(r, vals)
    print('---')
