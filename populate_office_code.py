import openpyxl

def map_header_to_col(ws):
    # Map headers to column indices (1-based)
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return {header: idx + 1 for idx, header in enumerate(headers)}

def main():
    in_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/after_fix/core_58701/input.xlsx"
    out_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/after_fix/core_58701/output.xlsx"

    wb = openpyxl.load_workbook(in_path)
    ws_entry = wb['Entry Tab']
    ws_table = wb['Table Tab']

    # Map headers, 1-based
    entry_headers = [cell.value for cell in next(ws_entry.iter_rows(min_row=1, max_row=1))]
    entry_field_to_col = {header: idx+1 for idx, header in enumerate(entry_headers)}
    table_headers = [cell.value for cell in next(ws_table.iter_rows(min_row=1, max_row=1))]
    table_field_to_col = {header: idx+1 for idx, header in enumerate(table_headers)}

    # Build location->office_code dict from Table Tab
    loc_col = table_field_to_col.get("Location Name")
    code_col = table_field_to_col.get("Office Code")
    location_to_code = {}
    for row in ws_table.iter_rows(min_row=2):
        loc = row[loc_col-1].value
        code = row[code_col-1].value
        if loc is not None and code is not None:
            location_to_code[loc] = code

    # For 'Entry Tab', fill Office Code (col 'E', header 'Office Code') for rows 2 & 3
    entry_loc_col = entry_field_to_col.get("Location")
    entry_code_col = entry_field_to_col.get("Office Code")
    for r in range(2, 4):  # rows 2 and 3
        location = ws_entry.cell(row=r, column=entry_loc_col).value
        code = location_to_code.get(location)
        ws_entry.cell(row=r, column=entry_code_col, value=code)

    wb.save(out_path)

if __name__ == "__main__":
    main()
