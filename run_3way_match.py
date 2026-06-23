import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_1/group_57033/r3/evolve_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_1/group_57033/r3/evolve_57033/output.xlsx'

df4 = pd.read_excel(input_path, sheet_name='Sheet4')
dfcb = pd.read_excel(input_path, sheet_name='CBtrans')

def norm(val):
    if pd.isna(val): return None
    return str(val).strip().lower()

results = []
for idx, row in df4.iloc[0:6].iterrows():
    company = norm(row['Company'])
    account = norm(row['account'])
    xchar = norm(row['xchar'])
    match = ((dfcb['company'].astype(str).str.strip().str.lower() == company) &
             (dfcb['account'].astype(str).str.strip().str.lower() == account) &
             (dfcb['xchar'].astype(str).str.strip().str.lower() == xchar)).any()
    results.append('Match' if match else '-')

wb = load_workbook(input_path)
ws = wb['Sheet4']
fill = PatternFill('solid', fgColor='FF66CC')
for i, val in enumerate(results, start=2):
    cell = ws[f'K{i}']
    cell.value = val.capitalize()
    cell.fill = fill
wb.save(output_path)
