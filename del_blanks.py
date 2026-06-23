import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_416-27_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_416-27_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Find the first row with a value in column A
col_a_values = [row[0].value for row in ws.iter_rows(min_row=1, max_col=1, max_row=ws.max_row)]
first_val_row = next((i+1 for i, v in enumerate(col_a_values) if v not in (None, '')), 1)
last_row = ws.max_row

# Collect rows to delete (blank in Column A, starting from first_val_row to last_row)
rows_to_delete = [i for i in range(last_row, first_val_row-1, -1) if ws.cell(row=i, column=1).value in (None, '')]

for i in rows_to_delete:
    ws.delete_rows(i)

# For output, restrict to A3:A24. Clear previous values outside this range
for row in range(1, ws.max_row+1):
    if row < 3 or row > 24:
        ws.cell(row=row, column=1).value = None

wb.save(output_path)
