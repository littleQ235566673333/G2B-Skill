from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42/eval_50631_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42/eval_50631_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet1']

# Prepare yellow fill and Calibri font
fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
font_calibri = Font(name='Calibri')

# Find range of holidays in J (skip headers)
holiday_rows = []
for r in range(4, ws.max_row + 1):
    if isinstance(ws[f'J{r}'].value, datetime.datetime):
        holiday_rows.append(r)
holiday_formula = f'$J$4:$J${max(holiday_rows) if holiday_rows else 4}' if holiday_rows else ''

# Fill formulas in H7:H38
for idx, row in enumerate(range(7, 39), 1):
    cell = ws[f'H{row}']
    ref = f'B{row}'
    ref_rows = f'$B$7:$B{row}'
    if holiday_formula:
        formula = f'=IF($B$3+ROWS({ref_rows})-1<=$F$3,WORKDAY($B$3+ROWS({ref_rows})-1,0,{holiday_formula}),"")'
    else:
        formula = f'=IF($B$3+ROWS({ref_rows})-1<=$F$3,WORKDAY($B$3+ROWS({ref_rows})-1,0),"")'
    cell.value = formula
    cell.font = font_calibri
    cell.fill = fill_yellow # Per instructions: "Every colored cell in column H should have a date value"

wb.save(output_path)
