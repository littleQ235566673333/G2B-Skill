from openpyxl import load_workbook

input_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_1/task_37554/r0/evolve_37554/input.xlsx"
output_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_1/task_37554/r0/evolve_37554/output.xlsx"

wb = load_workbook(input_path)
ws = wb.active

# The terms to count in column A, in rows 2-17
terms = [ws.cell(row=i, column=1).value for i in range(2, 18)]
# All values in column A to compare for exact match
all_types = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row+1)]

for idx, term in enumerate(terms):
    # Exact match only (not substring)
    count = sum(1 for t in all_types if t == term)
    ws.cell(row=idx + 2, column=6, value=count)

wb.save(output_path)
