import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_8/regression_gate/before_pass/core_493-18/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_8/regression_gate/before_pass/core_493-18/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# 1. Gather F values (assuming A1:F7 = data range, including headers)
f_vals = set()
for r in range(2, 8):
    val = ws.cell(row=r, column=6).value
    if val not in (None, ''):
        f_vals.add(val)

# 2. Clear A/B/C cells (2-7) if A not in F
for r in range(2, 8):
    v = ws.cell(row=r, column=1).value
    if v not in f_vals:
        for c in range(1, 4):
            ws.cell(row=r, column=c).value = None

# 3. Collect nonblank car data rows from A, B, C
car_rows = []
for r in range(2, 8):
    if ws.cell(row=r, column=1).value not in (None, ''):
        car_row = [ws.cell(row=r, column=c).value for c in range(1, 4)]
        car_rows.append(car_row)

# 4. Shift upwards in A,B,C
cur = 2
for row in car_rows:
    for j, val in enumerate(row):
        ws.cell(row=cur, column=j+1).value = val
    cur += 1
# Clear any remaining (if #car_rows < 6)
for r in range(cur, 8):
    for c in range(1, 4):
        ws.cell(row=r, column=c).value = None

# 5. Restore autofilter to A1:C7
ws.auto_filter.ref = 'A1:C7'

wb.save(output_path)
