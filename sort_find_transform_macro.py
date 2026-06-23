import openpyxl
from openpyxl import load_workbook

def sort_names(ws):
    # Determine real data range
    values = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is not None and str(v).strip():
            values.append(str(v).strip())
    values.sort(key=lambda x: x.upper())
    # Write sorted values back
    for idx, val in enumerate(values, 2):
        ws.cell(row=idx, column=1).value = val
    # Blank any leftover
    for idx in range(len(values)+2, ws.max_row+1):
        ws.cell(row=idx, column=1).value = None

def group_order_key(s, sort_groups):
    for i, ending in enumerate(sort_groups):
        if s.endswith(ending):
            return (i, s)
    return (len(sort_groups), s)

def find_pattern_transform(words, group_endings):
    results = []
    for word in words:
        if len(word) >= 8:
            ending = word[-3:]
            # Find all pairs
            for w2 in words:
                if w2 != word and w2[-3:] == ending and len(w2) >= 8:
                    transformed = w2[4] + w2[:4] + w2[5:]
                    if transformed == word:
                        results.append((transformed, w2))
    # Deduplicate pairs
    return list(dict.fromkeys(results))

def main():
    input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_3/group_118-50/r1/evolve_118-50/input.xlsx'
    output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_3/group_118-50/r1/evolve_118-50/output.xlsx'
    wb = load_workbook(input_file)
    ws = wb['Sheet1']
    # Sort names in column A
    sort_names(ws)
    # Read sorted words
    words = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is not None and str(v).strip():
            words.append(str(v).strip())
    group_endings = ['ING','ERS','ATE','EST','ONE','IER','ILY']
    pairs = find_pattern_transform(words, group_endings)
    # Group and order
    pairs.sort(key=lambda x: group_order_key(x[0][-3:], group_endings) + (x[0], x[1]))
    # Write results
    for i, (trans, orig) in enumerate(pairs[:4999]):
        ws.cell(row=2+i, column=3).value = trans
        ws.cell(row=2+i, column=4).value = orig
    for i in range(len(pairs)+2, 5001):
        ws.cell(row=i, column=3).value = None
        ws.cell(row=i, column=4).value = None
    wb.save(output_file)

if __name__ == '__main__':
    main()
