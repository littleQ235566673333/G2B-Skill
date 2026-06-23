from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_7/group_290-1/r1/evolve_290-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_7/group_290-1/r1/evolve_290-1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Transpose data from A2:J10 into K2:U10
for r in range(2, 11):  # Rows 2 to 10
    for j in range(10):  # Columns A-J to K-U
        val = ws.cell(row=r, column=1+j).value
        ws.cell(row=r, column=11+j).value = val

wb.save(output_path)
