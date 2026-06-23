import openpyxl
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_41978_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_41978_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

years = [ws[f'G{row}'].value for row in range(14, 186)]
statuses = [ws[f'J{row}'].value for row in range(14, 186)]
year_open_counts = defaultdict(int)
for yr, st in zip(years, statuses):
    if yr and st and str(st).strip().lower() == 'open':
        year_open_counts[yr] += 1

unique_years = sorted({y for y in years if y is not None and isinstance(y, (int, float))})

for i, yr in enumerate(unique_years[:10]):
    cell = ws[f'I{i+2}']
    cell.value = year_open_counts.get(yr, 0)

wb.save(output_path)
