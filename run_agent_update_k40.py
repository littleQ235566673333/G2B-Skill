import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/group_47766/r3/evolve_47766/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/group_47766/r3/evolve_47766/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def extract_data(start, end):
    data = []
    for row in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=15, values_only=True):
        data.append(row)
    return data

rows = []
rows += extract_data(8, 37)   # block 1
rows += extract_data(41, 58)  # block 2
rows += extract_data(62, 74)  # block 3

columns = list('ABCDEFGHIJKLMNO')
df = pd.DataFrame(rows, columns=columns)

def extract_year(d):
    if isinstance(d, datetime):
        return d.year
    try:
        return pd.to_datetime(d).year
    except Exception:
        return np.nan

df['Year'] = df['F'].apply(extract_year)

pes = df[df['H'].astype(str).str.contains('PE', na=False)]
pes = pes.dropna(subset=['Year'])

# Read the years from J40:J53 (Excel is 1-based)
year_col = []
for i in range(40, 54):  # J40:J53
    val = ws[f'J{i}'].value
    year_col.append(val)

out_vals = []
for y in year_col:
    try:
        val = pes.loc[pes['Year'] == int(y), 'C'].sum()
        if pd.isnull(val):
            out_vals.append('')
        else:
            out_vals.append(val)
    except:
        out_vals.append('')

def get_assignable_cell(ws, col, row):
    # For merged regions, return the top-left cell. Otherwise, return the cell itself.
    for mr in ws.merged_cells.ranges:
        if (row, col) >= (mr.min_row, mr.min_col) and (row, col) <= (mr.max_row, mr.max_col):
            return ws.cell(row=mr.min_row, column=mr.min_col)
    return ws.cell(row=row, column=col)

for idx, v in enumerate(out_vals):
    cell = get_assignable_cell(ws, 11, 40+idx)  # Column 11 = K
    cell.value = v

wb.save(output_path)
