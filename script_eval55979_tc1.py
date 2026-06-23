import openpyxl

# File paths
data_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed0/eval_55979_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed0/eval_55979_tc1/output.xlsx'

# Load workbook and sheets
wb = openpyxl.load_workbook(data_path)
calc_sheet = wb['Calc']
item_number = calc_sheet['A10'].value
supplier = None

# Search other sheets for item number
for sheet_name in wb.sheetnames:
    if sheet_name == 'Calc':
        continue
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        if item_number in row:
            # Get supplier name from sheet name (e.g., 'Supplier_1' -> 'Supplier 1')
            if sheet_name.startswith('Supplier_'):
                supplier = sheet_name.replace('_', ' ')
            else:
                # Handle other possible supplier name formats
                supplier = sheet_name
            break
    if supplier:
        break

if supplier:
    calc_sheet['B7'] = supplier
else:
    calc_sheet['B7'] = 'Not Found'

wb.save(output_path)
print('Supplier written to Calc!B7:', supplier)