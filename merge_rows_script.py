import pandas as pd
from openpyxl import load_workbook
from copy import copy

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_1/group_177-6/r2/evolve_177-6/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_1/group_177-6/r2/evolve_177-6/output.xlsx'

ref_col = 'ComboKey'  # Based on the mapping, this is column H in Excel

# Get the first sheet name
sheetname = pd.ExcelFile(input_file).sheet_names[0]
df_raw = pd.read_excel(input_file, sheet_name=sheetname)
df = df_raw.iloc[:, :18]  # A-R
num_cols = df.columns[8:18]  # I-R

for c in num_cols:
    df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
def agg_nums(x):
    return x.astype(float).sum()
agg_dict = {c: agg_nums if c in num_cols else 'first' for c in df.columns}
merged = df.groupby(ref_col, as_index=False).agg(agg_dict)
for c in num_cols:
    merged[c] = merged[c].apply(lambda v: '' if pd.isna(v) or float(v) == 0 else f'{float(v):.2f}')

wb = load_workbook(input_file)
ws = wb[sheetname]
ws_combined = wb['combined'] if 'combined' in wb.sheetnames else wb.create_sheet('combined')
ws_combined.delete_rows(1, ws_combined.max_row)

def copy_cell_style(source, target):
    if not source or not target:
        return
    if source.has_style:
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)

# Header
for j, col in enumerate(merged.columns, 1):
    cell = ws_combined.cell(row=1, column=j)
    cell.value = str(col)
    copy_cell_style(ws.cell(row=1, column=j), cell)
# Data
for i, row in merged.iterrows():
    ref_val = row[ref_col]
    style_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=8).value == ref_val:
            style_row = r
            break
    for j, col in enumerate(merged.columns, 1):
        cell = ws_combined.cell(row=i + 2, column=j)
        cell.value = row[col]
        # Copy formatting from first matched row
        if style_row:
            copy_cell_style(ws.cell(row=style_row, column=j), cell)
wb.save(output_file)
