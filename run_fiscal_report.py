import openpyxl
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/after_fix/core_56274/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/after_fix/core_56274/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet2']

# Step 1: Get the fiscal month from D7
fiscal_month = ws['D7'].value

# Step 2: Get headers row (row 3, columns G to R)
header = [ws.cell(row=3, column=col).value for col in range(7, 19)]

# Step 3: Locate column for fiscal_month
try:
    idx = header.index(fiscal_month)  # 0-index, maps to column (7+idx)
except ValueError:
    raise Exception(f"Fiscal month {fiscal_month} not found in header {header}")

# Step 4: Get Opening Bal, Debits, Credits, Closing Bal for that column
row_bases = {'Opening Bal': 4, 'Debits': 5, 'Credits': 6, 'Closing Bal': 7}
labels = ['Opening Bal', 'Debits', 'Credits', 'Closing Bal']
results = []
for i, label in enumerate(labels):
    val = ws.cell(row=row_bases[label], column=7+idx).value
    ws[f'D{9+i}'] = val

wb.save(output_path)
print('Done')
