from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']
today = datetime.today()

for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=11):
    last_contact = row[7].value  # H: Last Contact
    yes_no = row[8].value        # I: Regular Contact
    action_cell = row[9]        # J: Action
    if last_contact is not None and yes_no is not None:
        if isinstance(last_contact, datetime) and str(yes_no).strip().upper() == 'YES':
            days_since_contact = (today - last_contact).days
            if days_since_contact <= 30:
                action_cell.value = 'HOLD'
            else:
                action_cell.value = 'TOUCH BASE'
        else:
            action_cell.value = 'NO ACTION'
    else:
        action_cell.value = 'NO ACTION'

wb.save(output_path)
