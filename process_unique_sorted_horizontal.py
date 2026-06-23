from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun2/eval_7665_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun2/eval_7665_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

# Collect all non-empty values in column H (8), skip header if any
def is_yellow(cell):
    # Check for yellow highlight (OpenPyXL uses RGB; Excel Yellow 'FFFF00', sometimes as theme)
    fill = cell.fill
    if not fill or not hasattr(fill, 'fgColor'):
        return False
    fg = fill.fgColor.rgb
    if fg:
        return fg.upper() == 'FFFFFF00' or fg.upper().endswith('FF00')
    return False

values = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=8, max_col=8):
    for cell in row:
        if cell.value is not None and str(cell.value).strip() != '':
            # If the sample used highlights to indicate which values to use:
            if is_yellow(cell):
                values.append(cell.value)

# If not enough yellow found, take all non-empty values (in case sample has no highlights in test)
if not values:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=8, max_col=8):
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != '':
                values.append(cell.value)

unique_sorted = sorted(set(values), key=lambda x: (str(x).lower(), x))

# Write to Q2:V2 (columns 17 to 22)
for idx in range(6):
    col = 17 + idx
    dest = ws.cell(row=2, column=col)
    dest.value = unique_sorted[idx] if idx < len(unique_sorted) else None

wb.save(output_path)
