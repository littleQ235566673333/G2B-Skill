import openpyxl
import datetime

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_5/task_45707/r1/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_5/task_45707/r1/evolve_45707/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Explicitly mapping headers to columns
header_row = [cell.value for cell in ws[1]]
date_col = header_row.index('Date') + 1  # openpyxl is 1-based
occurence_col = header_row.index('Occurence') + 1
# Column D is the 4th column (index 3), per spec
result_col = 4

# Read all rows into memory
rows = list(ws.iter_rows(min_row=2, max_row=69, values_only=True))
dates = [row[date_col-1] for row in rows]
col_c_vals = [row[occurence_col-1] for row in rows]

# Parse dates to datetime.date
def parse_date(d):
    if isinstance(d, (datetime.date, datetime.datetime)):
        return d if isinstance(d, datetime.date) else d.date()
    elif isinstance(d, str):
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
            try:
                return datetime.datetime.strptime(d, fmt).date()
            except Exception:
                continue
    return None

parsed_dates = [parse_date(d) for d in dates]

for i in range(len(parsed_dates)):
    # Check the next day
    if i+1 < len(parsed_dates) and parsed_dates[i+1] is not None:
        next_day = parsed_dates[i+1]
        if next_day.day == 1:
            # Count 1s in Occurence for the same month/year as next_day
            count = 0
            for pd, cv in zip(parsed_dates, col_c_vals):
                if pd is not None and pd.year == next_day.year and pd.month == next_day.month and cv == 1:
                    count += 1
            ws.cell(row=2+i, column=result_col, value=count)
        else:
            ws.cell(row=2+i, column=result_col, value=None)
    else:
        ws.cell(row=2+i, column=result_col, value=None)

wb.save(output_path)
