import openpyxl
from openpyxl.styles import Alignment

# Load workbook and worksheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun2/eval_56378_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun2/eval_56378_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Folha1']

# Find rows in frame 1 (columns C:J), with non-empty 'QUANTITY UNITS'
# Header row: row 4 for frame 1 (C:J), data from row 5 onward
header = [ws.cell(row=4, column=c).value for c in range(3, 10)]
data_rows = []
for i in range(5, 12):  # rows 5 to 11 in input sample
    row_vals = [ws.cell(row=i, column=c).value for c in range(3, 10)]
    if row_vals[-1] is not None:
        data_rows.append(row_vals)

# Write header at L5:R5
for j, val in enumerate(header):
    ws.cell(row=5, column=12+j, value=val)
# Write filtered products to L6:R8 (as many as found)
for i, row in enumerate(data_rows):
    for j, val in enumerate(row):
        ws.cell(row=6+i, column=12+j, value=val)

# Alignment: Product name left (M), O:R right align
for i in range(6, 6+len(data_rows)):
    ws.cell(row=i, column=13).alignment = Alignment(horizontal='left')  # PRODUCT
    for col in range(15, 19):  # O(15),P(16),Q(17),R(18)
        ws.cell(row=i, column=col).alignment = Alignment(horizontal='right')

wb.save(output_path)
