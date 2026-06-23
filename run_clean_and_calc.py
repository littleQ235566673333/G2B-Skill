import openpyxl

# Load workbook and worksheet
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_2/group_387-16/r3/evolve_387-16/input.xlsx')
ws = wb['Sheet1']

# Read Values and Binaries (rows 3-18 = 16 rows)
values = []
binaries = []
for row in ws.iter_rows(min_row=3, max_row=18, min_col=1, max_col=2, values_only=True):
    values.append(row[0])
    binaries.append(row[1])

# Read Result values (column D, from D5 downward)
result_values = []
row = 5
while True:
    cell_value = ws.cell(row=row, column=4).value
    if cell_value is None:
        break
    result_values.append(cell_value)
    row += 1

# Remove only ONE instance of each matching result value from 'values' and same index in binaries
for rv in result_values:
    if rv in values:
        idx = values.index(rv)
        values.pop(idx)
        binaries.pop(idx)

# Fill output range (A3:B18) - move all values up, blanks at end if < 16
for i in range(16):
    v = values[i] if i < len(values) else None
    b = binaries[i] if i < len(binaries) else None
    ws.cell(row=3+i, column=1, value=v)  # Value col
    ws.cell(row=3+i, column=2, value=b)  # Binaries col
    ws.cell(row=3+i, column=3, value=None)
    ws.cell(row=3+i, column=4, value=None)

# Restore headers for columns A/B if needed
ws.cell(row=2, column=1, value='Value')
ws.cell(row=2, column=2, value='Binaries')

# Compute solver result using formula (sum of column A)
ws['H2'] = '=SUM(A3:A18)'
# Difference: H2-E2
ws['K2'] = '=H2-E2'

# Save workbook
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_2/group_387-16/r3/evolve_387-16/output.xlsx')
