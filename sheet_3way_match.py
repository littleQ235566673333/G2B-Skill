import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- Load data with pandas for fast matching ---
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_2/regression_gate/before_fix/core_57033/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_2/regression_gate/before_fix/core_57033/output.xlsx'

# Load all sheets
sheets = pd.read_excel(input_path, sheet_name=None)
df4 = sheets['Sheet4']
# Try common 3-way match table names
md_candidates = [k for k in sheets.keys() if k.lower() == 'md']
md_sheet = md_candidates[0] if md_candidates else list(sheets.keys())[0]
df_md = sheets[md_sheet]

# Lowercase matching columns for robustness
for col in ['company','account','xchar']:
    if col in df4.columns: df4[col+'_l'] = df4[col].astype(str).str.lower()
    if col in df_md.columns: df_md[col+'_l'] = df_md[col].astype(str).str.lower()

# Do the 3-way match for rows 2-7 (index 0-based: 1-6)
results = []
for idx in range(1,7):
    row = df4.iloc[idx]
    cond = (
        (df_md['company_l'] == row.get('company_l', '')) &
        (df_md['account_l'] == row.get('account_l', '')) &
        (df_md['xchar_l'] == row.get('xchar_l', ''))
    )
    match = df_md[cond]
    results.append('Match' if not match.empty else '-')

# Open with openpyxl for formatting
wb = load_workbook(input_path)
ws = wb['Sheet4']

# Hex #FF66CC as Excel fill
fill = PatternFill(start_color='FF66CC', end_color='FF66CC', fill_type='solid')

for i,res in enumerate(results):
    cell = ws.cell(row=2+i, column=11) # K column is col 11
    # Regular casing (capitalize first letter)
    cell.value = res.capitalize()
    cell.fill = fill

wb.save(out_path)
