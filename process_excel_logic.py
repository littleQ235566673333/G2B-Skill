from openpyxl import load_workbook
from collections import defaultdict

def process_excel(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb.active

    # Step 1: collect all row data for grouping by numbers in Column A
    groups = defaultdict(list)  # number -> list of (row_idx, status, date)
    for row in range(2, 21):  # Only rows 2 to 20
        number = ws[f"A{row}"].value
        status = ws[f"D{row}"].value  # status in Column D
        date = ws[f"E{row}"].value    # date in Column E (can be None)
        groups[number].append((row, status, date))

    # Step 2: Evaluate each group and map output for each row
    results = {}
    for number, info_list in groups.items():
        statuses = set([item[1] for item in info_list])
        # For group that is exclusively SCH
        if statuses == {'SCH'}:
            for row, _, _ in info_list:
                results[row] = 'FUTURE'
        # For group with both NO SHOW and SCH
        elif statuses == {'SCH', 'NO SHOW'}:
            for row, status, _ in info_list:
                results[row] = 'NS/SCHED'
        # Group contains only NO SHOW
        elif statuses == {'NO SHOW'}:
            for row, status, date in info_list:
                if date is not None and str(date).strip() != '':
                    results[row] = 'NO ACTION NEEDED'
                else:
                    results[row] = 'CALL PT'
        else:
            # Mixed status or not matching above
            for row, status, date in info_list:
                if status == 'NO SHOW':
                    if date is not None and str(date).strip() != '':
                        results[row] = 'NO ACTION NEEDED'
                    else:
                        results[row] = 'CALL PT'
                elif status == 'SCH':
                    results[row] = 'FUTURE'
                else:
                    results[row] = ''

    # Step 3: Write result to column F
    for row in range(2, 21):
        if row in results:
            ws[f"F{row}"] = results[row]
    wb.save(output_path)

if __name__ == "__main__":
    process_excel(
        "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_2/regression_gate/before_pass/core_55421/input.xlsx",
        "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_2/regression_gate/before_pass/core_55421/output.xlsx"
    )
