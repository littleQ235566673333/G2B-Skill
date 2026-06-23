import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_1/group_45707/r1/evolve_45707/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_1/group_45707/r1/evolve_45707/output.xlsx'

# Read the input file with pandas for easier date and value handling
df = pd.read_excel(input_file)

# We need to check the next row date for each row
count_results = [''] * len(df)

for idx in range(1, len(df)):
    next_date = df.iloc[idx, 0]
    if not isinstance(next_date, pd.Timestamp):
        try:
            next_date = pd.to_datetime(next_date)
        except Exception:
            continue
    # If next day is the first of a new month
    if next_date.day == 1:
        filter_month = next_date.month
        filter_year = next_date.year
        month_mask = pd.to_datetime(df.iloc[:, 0], errors='coerce').dt.month == filter_month
        year_mask = pd.to_datetime(df.iloc[:, 0], errors='coerce').dt.year == filter_year
        value_mask = df.iloc[:, 2] == 1
        cnt = df[month_mask & year_mask & value_mask].shape[0]
        count_results[idx-1] = cnt

# Write this to D2:D69 (D column, skipping header)
wb = load_workbook(input_file)
ws = wb.active
for i, val in enumerate(count_results[:68]):
    ws.cell(row=i+2, column=4, value=val if val != '' else None)

wb.save(output_file)
