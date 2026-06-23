import pandas as pd
from openpyxl import load_workbook

# Define file paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_7/regression_gate/after_fix/core_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_7/regression_gate/after_fix/core_45896/output.xlsx'

# Load workbook and worksheet
wb = load_workbook(input_path)
ws = wb['Volym P5_P6_2023']

# Read ZORD sheet with pandas for convenience
zord_df = pd.read_excel(input_path, sheet_name='ZORD', dtype=object)

# Build mapping: key (A) -> list of unique dates (C)
zord_map = {}
for idx, row in zord_df.iterrows():
    key = row.iloc[0]
    val = row.iloc[2] if len(row) > 2 else None
    if pd.notnull(key) and pd.notnull(val):
        zord_map.setdefault(key, []).append(val)

# For each row in Volym P5_P6_2023!A2:A10
for r in range(2, 11):
    a_val = ws[f'A{r}'].value
    dates = zord_map.get(a_val, [])
    # Remove duplicates, keep order
    seen = set()
    unique_dates = []
    for d in dates:
        if d not in seen:
            seen.add(d)
            unique_dates.append(d)
    # Format as DD/MM/YYYY
    formatted_dates = []
    for d in unique_dates:
        try:
            parsed = pd.to_datetime(d)
            formatted = parsed.strftime('%d/%m/%Y')
            formatted_dates.append(formatted)
        except Exception:
            continue # skip unparseable
    # Join with comma and write result
    ws[f'I{r}'].value = ','.join(formatted_dates)

# Save to new file
wb.save(output_path)
