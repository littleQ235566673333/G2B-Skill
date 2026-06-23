import openpyxl
from openpyxl.utils import column_index_from_string

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/before_fix/core_49667/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/before_fix/core_49667/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Edit']

block_start_col = column_index_from_string('F')
block_end_col = column_index_from_string('AT')
output_cols = ['B', 'C', 'D', 'E']

block_headers = []
for col in range(block_start_col, block_end_col + 1):
    block_headers.append(ws.cell(row=1, column=col).value)

for row in range(2, 17):
    meetings = []
    in_meeting = False
    start_idx = None
    for idx, col in enumerate(range(block_start_col, block_end_col + 1)):
        val = ws.cell(row=row, column=col).value
        if val == 'm':
            if not in_meeting:
                start_idx = idx
                in_meeting = True
        else:
            if in_meeting:
                meetings.append((start_idx, idx - 1))
                in_meeting = False
    if in_meeting:
        meetings.append((start_idx, block_end_col - block_start_col))
    # Up to 2 meetings
    meetings = meetings[:2]
    output_values = []
    for block in meetings:
        start_col, end_col = block
        start_header = block_headers[start_col]
        end_header = block_headers[end_col]
        # Numeric representation
        start_time_num = openpyxl.utils.datetime.from_excel(start_header) if isinstance(start_header, (int, float)) else start_header
        end_time_num = openpyxl.utils.datetime.from_excel(end_header) if isinstance(end_header, (int, float)) else end_header
        output_values.extend([start_header, end_header])
    # Fill blanks if less than 2 meetings
    while len(output_values) < 4:
        output_values.append('')
    for j, val in enumerate(output_values):
        ws[f'{output_cols[j]}{row}'] = val
wb.save(output_path)
