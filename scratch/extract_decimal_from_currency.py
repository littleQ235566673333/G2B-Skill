import openpyxl
import re

def extract_number_and_decimal(text):
    # Remove all currency symbols
    text = re.sub(r'[€$£¥₹₽]', '', text)
    # Identify if "," or "." is decimal point
    # If both exist, the last one is the decimal separator
    comma = text.rfind(',')
    dot = text.rfind('.')
    if comma > dot:
        decimal_sign = ','
    else:
        decimal_sign = '.'
    # Remove thousands separators
    if decimal_sign == ',':
        num = re.sub(r'[.]', '', text)
        num = num.replace(',', '.')
    else:
        num = re.sub(r'[,]', '', text)
    # Extract float value
    matches = re.findall(r'-?\d+\.?\d*', num)
    if matches:
        value = matches[0]
    else:
        value = ''
    return value, decimal_sign

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/group_48257/r0/evolve_48257/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/group_48257/r0/evolve_48257/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 6):
    cell_value = str(ws[f'B{row}'].value)
    num, dec_sign = extract_number_and_decimal(cell_value)
    ws[f'C{row}'].value = f'{num} (decimal: {dec_sign})'

wb.save(output_path)
