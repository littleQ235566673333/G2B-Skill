import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/group_387-16/r2/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/group_387-16/r2/evolve_387-16/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']
rows = list(ws.iter_rows(values_only=True))
# Parse columns
values = [rows[i][0] for i in range(2, 18)]
binaries = [rows[i][1] for i in range(2, 18)]
result_values = [rows[i][3] for i in range(4, 18) if rows[i][3] is not None]
# Remove one instance of each result_value in values, and corresponding binary
values_out, binaries_out = list(values), list(binaries)
for rv in result_values:
    if rv in values_out:
        idx = values_out.index(rv)
        values_out.pop(idx)
        binaries_out.pop(idx)
# Remove resulting blanks and shift up
while len(values_out) < 16:
    values_out.append(None)
    binaries_out.append(None)
# Write output
for i in range(16):
    ws.cell(row=3+i, column=1, value=values_out[i])
    ws.cell(row=3+i, column=2, value=binaries_out[i])
# Calculate Solver result and difference
target_value = rows[1][4]
solver_result = sum([v*b if v is not None and b is not None else 0 for v, b in zip(values_out, binaries_out)])
difference = solver_result - target_value
ws['H2'] = solver_result
ws['J2'] = difference
wb.save(output_path)
