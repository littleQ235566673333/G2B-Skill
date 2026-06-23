import openpyxl

# Load the input workbook
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/before_fix/core_50916/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/before_fix/core_50916/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheet = wb.active

# Map cycle days to class schedules
# Assume cycle days are in A2:A8, classes in C2:H8
class_schedule = {}
for i in range(2, 9):  # Rows 2-8
    cycle_day = sheet[f'A{i}'].value
    # Get classes for the cycle day from columns C-H
    classes = [sheet[f'{openpyxl.utils.get_column_letter(col)}{i}'].value for col in range(3, 9)]
    class_schedule[cycle_day] = classes

# Fill calendar in C12:H14 based on cycle day reference in A12:A14
for row in range(12, 15):  # Rows 12, 13, 14
    cycle_day = sheet[f'A{row}'].value
    classes = class_schedule.get(cycle_day, [None] * 6)
    for col in range(3, 9):  # C=3, H=8
        sheet.cell(row=row, column=col, value=classes[col-3])

# Save to output
wb.save(output_path)
