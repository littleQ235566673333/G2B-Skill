import openpyxl

data_table_start = 5 # First row with data (after header, which is row 4)
data_table_end = 10  # Last non-empty data row from previous preview
report_start = 17
report_end = 22

data_table_header_row = 4
report_header_row = 16

data_table_cols = {
    'Work': 1,
    'Material': 2,
    'Category 1': 3,
    'Category 2': 4,
    'Category 3': 5
}
report_cols = {
    'Work': 1,
    'Category Type': 2,
    'Material': 3,
    'Category Value': 4
}

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_3/regression_gate/before_pass/core_18935/input.xlsx')
ws = wb.active

# Read Data Table into a list of dicts
data_table = []
for row in ws.iter_rows(min_row=data_table_start, max_row=data_table_end, min_col=1, max_col=5, values_only=True):
    data_table.append({
        'Work': row[0],
        'Material': row[1],
        'Category 1': row[2],
        'Category 2': row[3],
        'Category 3': row[4]
    })

# For each report row (D17:D22), perform lookup and write result.
for i, row in enumerate(ws.iter_rows(min_row=report_start, max_row=report_end, min_col=1, max_col=4)):
    work = row[0].value
    category_type = row[1].value
    material = row[2].value
    # Find entry in data table
    value = None
    for entry in data_table:
        if entry['Work'] == work and entry['Material'] == material:
            value = entry.get(category_type)
            break
    # Write result to column D
    ws.cell(row=report_start + i, column=4, value=value)

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_3/regression_gate/before_pass/core_18935/output.xlsx')
