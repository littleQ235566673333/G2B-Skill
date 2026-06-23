from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/regression_gate/before_fix/core_49667/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/regression_gate/before_fix/core_49667/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Edit']

# Map columns
first_data_row = 2
last_data_row = 16
meeting_start_col = 6   # F
meeting_end_col = 46   # AT (1-based, so actual index is AT->46)

for row in range(first_data_row, last_data_row+1):
    meeting_cells = [ws.cell(row=row, column=col).value for col in range(meeting_start_col, meeting_end_col+1)]
    blocks = []
    in_block = False
    start_col = None
    for idx, val in enumerate(meeting_cells):
        if val == 'm' and not in_block:
            in_block = True
            start_col = idx
        elif val != 'm' and in_block:
            in_block = False
            blocks.append((start_col, idx - 1))
    if in_block: # block runs to last column
        blocks.append((start_col, len(meeting_cells) - 1))
    
    # Write up to 2 blocks: B=start_date, C=start_time, D=finish_date, E=finish_time
    for i in range(2):
        base = 2*i
        if i < len(blocks):
            s, e = blocks[i]
            date_val = ws.cell(row=row, column=1).value  # Column A: the date
            # Columns F(6)+s and F(6)+e are the actual start/end columns
            start_time = ws.cell(row=row, column=meeting_start_col + s).value
            finish_time = ws.cell(row=row, column=meeting_start_col + e).value
            ws.cell(row=row, column=2 + base, value=date_val)
            ws.cell(row=row, column=3 + base, value=start_time)
        else:
            ws.cell(row=row, column=2 + base, value=None)
            ws.cell(row=row, column=3 + base, value=None)

wb.save(output_path)
