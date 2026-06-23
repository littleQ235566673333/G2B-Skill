from openpyxl import load_workbook

input_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_6/regression_gate/before/core_54474/input.xlsx"
output_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_6/regression_gate/before/core_54474/output.xlsx"

wb = load_workbook(input_path)
ws_whp = wb["WHP"]
ws_data = wb["WHP DATA"]

# Build lookup from WHP DATA: site -> {header: value}
headers = {}
for c in range(1, ws_data.max_column + 1):
    headers[c] = ws_data.cell(3, c).value

site_lookup = {}
for r in range(4, ws_data.max_row + 1):
    site = ws_data.cell(r, 2).value
    if site not in (None, ""):
        site_lookup[site] = {
            headers[c]: ws_data.cell(r, c).value
            for c in range(1, ws_data.max_column + 1)
        }

# Fill E:G on WHP for listed sites using columns EEG, Dis, LTU
for r in range(7, 9):
    site = ws_whp.cell(r, 3).value
    site_data = site_lookup.get(site, {})
    ws_whp.cell(r, 5).value = site_data.get("EEG")
    ws_whp.cell(r, 6).value = site_data.get("Dis")
    ws_whp.cell(r, 7).value = site_data.get("LTU")

wb.save(output_path)
print("saved", output_path)
