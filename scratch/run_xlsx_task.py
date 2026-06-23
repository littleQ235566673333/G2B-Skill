from openpyxl import load_workbook

input_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/51680/input.xlsx'
output_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/51680/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

headers = [ws.cell(row=1, column=col).value for col in range(1, 7)]
header_colors = []
for col in range(1, 7):
    fill = ws.cell(row=1, column=col).fill
    color = getattr(fill.fgColor, 'rgb', None)
    header_colors.append((color, col, headers[col-1]))
header_colors.sort(key=lambda x: (x[0] if x[0] else '', x[1]))
ordered_columns = [col for _, col, _ in header_colors]

for row in range(2, 15):
    used_areas = []
    for col in ordered_columns:
        val = ws.cell(row=row, column=col).value
        if isinstance(val, str) and val.strip().upper() == 'Y':
            used_areas.append(ws.cell(row=1, column=col).value)
    ws.cell(row=row, column=7).value = ", ".join(used_areas)

wb.save(output_path)
