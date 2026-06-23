import openpyxl
import re
from openpyxl.utils import get_column_letter

# Load workbook and select active sheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/regression_gate/after_pass/core_290-27/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/regression_gate/after_pass/core_290-27/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# The area to process: A14:I137
start_row = 14
end_row = 137
start_col = 1  # Column A
end_col = 9    # Column I

for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
    cell = row[1]  # Column B is index 1
    val = cell.value
    # Only operate on strings (alphanumeric patterns), leave dates and non-strs untouched
    if isinstance(val, str):
        # Look for the pattern with optional space after 2 or 3 caps at start
        m = re.match(r'^([A-Z]{2,3}) ?(\d+.*)$', val)
        if m:
            cell.value = m.group(2)
    # No change if not matching the pattern

wb.save(output_path)
