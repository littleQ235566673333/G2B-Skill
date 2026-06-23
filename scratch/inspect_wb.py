from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_8/regression_gate/before_fix/core_57033/input.xlsx"
output_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_8/regression_gate/before_fix/core_57033/output.xlsx"


def norm(v):
    return " ".join(str(v or "").split()).casefold()

wb = load_workbook(input_path)

# Inspect sheets and headers
for ws in wb.worksheets:
    print(f"SHEET::{ws.title}")
    headers = [ws.cell(row=1, column=c).value for c in range(1, min(ws.max_column, 15) + 1)]
    print(headers)
    for r in range(2, min(ws.max_row, 8) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 12) + 1)]
        print(r, vals)
    print('---')
