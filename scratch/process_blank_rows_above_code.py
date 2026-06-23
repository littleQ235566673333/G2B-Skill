import openpyxl

# Load workbook and sheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_374-31/r2/evolve_374-31/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_374-31/r2/evolve_374-31/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

rows_to_delete = set()

for i in range(2, ws.max_row+1):  # start at 2 to avoid deleting header
    if str(ws[f'A{i}'].value).strip() == 'Code':
        j = i - 1
        while j > 0 and all(ws[f'{col}{j}'].value in (None, '', ' ') for col in ['A','B','C','D']):
            rows_to_delete.add(j)
            j -= 1

for r in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(r)

wb.save(output_path)
