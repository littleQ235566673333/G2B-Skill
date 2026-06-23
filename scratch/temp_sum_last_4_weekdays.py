import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/58109/input.xlsx'
output_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/58109/output.xlsx'

# Read relevant columns from the file (auto detect sheet)
df = pd.read_excel(input_path)

# Find the openpyxl sheet and load workbook
wb = load_workbook(input_path)
ws = wb.active

# Find I4 (date to sum up through)
date_in_I4 = ws['I4'].value
if isinstance(date_in_I4, str):
    date_in_I4 = pd.to_datetime(date_in_I4).date()
elif isinstance(date_in_I4, datetime):
    date_in_I4 = date_in_I4.date()

# Ensure A is dates, B is values
date_col = df.columns[0]
value_col = df.columns[1]

# Keep only dates up through the date in I4, and only on weekdays (by dataset design)
mask = (df[date_col] <= pd.Timestamp(date_in_I4))
weekday_dates = df.loc[mask, date_col].sort_values(ascending=False)
last_4_dates = weekday_dates.unique()[:4]

# Filter only those rows (this handles accounting for no weekends data!!)
sum_rows = df[df[date_col].isin(last_4_dates)][value_col].sum()

# Place sum in I6
ws['I6'] = sum_rows
wb.save(output_path)
print('Sum placed:', sum_rows)
