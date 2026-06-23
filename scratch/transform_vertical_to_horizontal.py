from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_7665_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_7665_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Collect values in column H (8th column), skipping blanks
col_h_values = []
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=8).value
    if val is not None and val != '':
        col_h_values.append(val)

# Convert all values to string for consistent sorting
col_h_values = [str(v) for v in col_h_values]
unique_sorted = sorted(set(col_h_values))

# Write these horizontally starting at Q2 (column 17, row 2)
for i, val in enumerate(unique_sorted):
    ws.cell(row=2, column=17 + i, value=val)
    # Only fill Q2:V2 (6 columns max)
    if i >= 5:
        break

wb.save(output_path)
