from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/before_fix/core_58701/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/before_fix/core_58701/output.xlsx'

wb = load_workbook(input_path)
ws_table = wb['Table Tab']
ws_entry = wb['Entry Tab']

# Build location->office code mapping from 'Table Tab'
location_to_code = {}
for row in ws_table.iter_rows(min_row=2, values_only=True):
    location, code = row
    location_to_code[location] = code

# Process rows in 'Entry Tab', populate E2:E3
for row_idx in range(2, 4):  # Rows 2 and 3
    location_cell = ws_entry.cell(row=row_idx, column=2)  # 'Location' name
    office_code_cell = ws_entry.cell(row=row_idx, column=5)  # 'Office Code'
    location = location_cell.value
    office_code = location_to_code.get(location)
    office_code_cell.value = office_code

wb.save(output_path)
