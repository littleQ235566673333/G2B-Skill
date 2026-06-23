import openpyxl

def count_blanks_between_numeric_entries(row_values):
    # Search for first and second numeric entries
    numeric_indices = [i for i, val in enumerate(row_values) if isinstance(val, (int, float)) and val is not None]
    
    if len(numeric_indices) < 2:
        return 0  # Not enough numeric entries
    
    first_idx = numeric_indices[0]
    second_idx = numeric_indices[1]
    
    first_val = row_values[first_idx]
    # If the FIRST value in the row is greater than 1, output should be 1
    if first_val > 1:
        return 1
    
    # Count blank cells between first and second (exclusive bounds)
    blanks = 0
    for i in range(first_idx+1, second_idx):
        if row_values[i] is None:
            blanks += 1
    return blanks

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/before_fix/core_50521/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/before_fix/core_50521/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Process rows 4, 5, 6 (N4:N6)
for row_idx in range(4, 7):
    # Row data from columns B to M (2 to 13)
    row_values = [ws.cell(row=row_idx, column=col).value for col in range(2, 14)]
    result = count_blanks_between_numeric_entries(row_values)
    ws.cell(row=row_idx, column=14).value = result  # N is column 14

wb.save(output_path)
