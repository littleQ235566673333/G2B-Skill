import openpyxl
from datetime import datetime, time

def excel_time_from_str(s):
    # Convert '08:00' -> Excel float representing time
    t = datetime.strptime(s, '%H:%M').time()
    return (t.hour * 60 + t.minute) / (24 * 60)

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/group_49667/r0/evolve_49667/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/group_49667/r0/evolve_49667/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Timeslot headers (columns F:AT)
time_columns = [cell.value for cell in ws[1][5:46]]
time_numbers = [excel_time_from_str(t) if t else None for t in time_columns]

for row_idx in range(2, 17):  # B2:E16
    row = [cell.value for cell in ws[row_idx][5:46]]

    meetings = []
    block = None
    for i, val in enumerate(row):
        if val == 'm':
            if block is None:
                block = [i, i]  # start_idx, end_idx
            else:
                block[1] = i  # keep updating end
        else:
            if block is not None:
                meetings.append(tuple(block))
                block = None
    if block is not None:
        meetings.append(tuple(block))

    # Only up to 2 blocks
    results = []
    for meet in meetings[:2]:
        start_idx, end_idx = meet
        start_time = time_numbers[start_idx]
        finish_time = time_numbers[end_idx]
        # Use the current day (if needed): Excel stores only time as fraction
        results.append((start_time, finish_time))
    # Fill to two meetings
    while len(results) < 2:
        results.append((None, None))

    # Write results back to columns B-E
    ws.cell(row=row_idx, column=2).value = results[0][0]  # Meeting Start1
    ws.cell(row=row_idx, column=3).value = results[0][1]  # Meeting Finish1
    ws.cell(row=row_idx, column=4).value = results[1][0]  # Meeting Start2
    ws.cell(row=row_idx, column=5).value = results[1][1]  # Meeting Finish2
    # Set number format for time as time
    for col in [2, 3, 4, 5]:
        ws.cell(row=row_idx, column=col).number_format = 'hh:mm'

wb.save(output_path)
print('Done.')
