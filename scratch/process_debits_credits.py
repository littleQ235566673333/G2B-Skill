import openpyxl
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_469-9_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_469-9_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active
# Write headers
ws['H1'] = 'Debits'
ws['I1'] = 'Credits'
for row in range(2, 11):
    cell = ws[f'C{row}'].value
    if cell is None or cell == '':
        continue
    try:
        value = float(cell)
    except Exception:
        continue
    if value < 0:
        ws[f'H{row}'] = abs(value)
    elif value > 0:
        ws[f'I{row}'] = abs(value)
wb.save(output_path)
