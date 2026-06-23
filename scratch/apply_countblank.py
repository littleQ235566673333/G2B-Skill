import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/before_pass/core_41969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/before_pass/core_41969/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Apply COUNTBLANK formula to shifting ranges for A6:C6
for idx, col in enumerate(range(1, 10, 3)):
    start_col = openpyxl.utils.get_column_letter(col)
    end_col = openpyxl.utils.get_column_letter(col+2)
    formula = f'=COUNTBLANK({start_col}3:{end_col}3)'
    ws.cell(row=6, column=idx+1, value=formula)

wb.save(output_path)
