import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/after_pass/core_50526/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

lookup_value = ws['B6'].value
header = [cell.value for cell in ws[1]][1:]  # skip A1
found_row = None
for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    if row[0].value == lookup_value:
        found_row = row[0].row
        break
result = []
if found_row is not None:
    for col, head in enumerate(header, start=2):
        cell_value = ws.cell(row=found_row, column=col).value
        try:
            if cell_value is not None and float(cell_value) > 0:
                result.append(ws.cell(row=1, column=col).value)
        except (ValueError, TypeError):
            continue
for i, v in enumerate(result[:2]):
    ws[f'B{9+i}'] = v
wb.save(output_path)
