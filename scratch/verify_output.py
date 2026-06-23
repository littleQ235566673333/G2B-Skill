from openpyxl import load_workbook

output_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4_v6/train/iter_2/regression_gate/after_fix/core_57033/output.xlsx"
wb = load_workbook(output_path)
ws = wb["Sheet4"]
for row in range(2,8):
    c = ws[f"K{row}"]
    print(row, c.value, c.fill.fill_type, c.fill.start_color.rgb)
