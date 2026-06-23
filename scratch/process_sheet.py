import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/after_pass/core_32337/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Column indices (1-based)
col_report_date = 2  # B
col_dob = 3          # C
col_expected_result = 5  # E
col_age_year = 15    # O
col_category = 9     # I

# Find last row with data in NAME column (A)
last_row = ws.max_row
for row in range(3, ws.max_row + 1):
    if ws[f'A{row}'].value is None:
        last_row = row - 1
        break

# 1. Set 'Expected Result' (E3:E15): formula to look up CATEGORY from age(year) in O
for row in range(3, last_row + 1):
    ws.cell(row=row, column=col_expected_result).value = f"=INDEX($I$3:$I${last_row},MATCH(O{row},$O$3:$O${last_row},0))"

# 2. Insert 'Actual Age' column after 'age (year)' (after col 15)
actual_age_col_idx = col_age_year + 1
actual_age_col_letter = get_column_letter(actual_age_col_idx)
ws.insert_cols(actual_age_col_idx)

# Set header
ws[f'{actual_age_col_letter}2'] = 'Actual Age'

# Copy fill from preceding column header (age (year)) in a safe way
age_year_header_cell = ws[f'{get_column_letter(col_age_year)}2']
age_year_fill = age_year_header_cell.fill
if isinstance(age_year_fill, PatternFill):
    ws[f'{actual_age_col_letter}2'].fill = PatternFill(fill_type=age_year_fill.fill_type,
                                                       fgColor=age_year_fill.fgColor,
                                                       bgColor=age_year_fill.bgColor)
# Top align header
ws[f'{actual_age_col_letter}2'].alignment = Alignment(horizontal='center', vertical='top')

# Copy fill for data cells from age (year) column in a safe way
for row in range(3, last_row + 1):
    # Formula for Actual Age
    ws[f'{actual_age_col_letter}{row}'].value = f'=INT((INDEX($B$1:$B$1,1)-C{row})/365)'
    # Make bold, center
    ws[f'{actual_age_col_letter}{row}'].font = Font(bold=True)
    ws[f'{actual_age_col_letter}{row}'].alignment = Alignment(horizontal='center', vertical='center')
    # Copy fill safely
    age_year_cell = ws[f'{get_column_letter(col_age_year)}{row}']
    age_year_fill = age_year_cell.fill
    if isinstance(age_year_fill, PatternFill):
        ws[f'{actual_age_col_letter}{row}'].fill = PatternFill(fill_type=age_year_fill.fill_type,
                                                             fgColor=age_year_fill.fgColor,
                                                             bgColor=age_year_fill.bgColor)

# Save as new output
wb.save(output_path)
