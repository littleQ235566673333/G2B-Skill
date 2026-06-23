import openpyxl
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/regression_gate/before_pass/core_41589/output.xlsx'

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Contact List']

# Read values
h_cell = ws['H4'].value  # Date of last contact
i_cell = ws['I4'].value  # Yes/No response

result = 'NO ACTION'  # Default

if i_cell and str(i_cell).strip().upper() == 'YES':
    h_date = None
    # Check if the cell is already a datetime or string
    if isinstance(h_cell, datetime):
        h_date = h_cell
    elif isinstance(h_cell, str):
        try:
            # Try to parse ISO format
            h_date = datetime.strptime(h_cell, '%Y-%m-%d')
        except ValueError:
            try:
                # Try to parse other common format
                h_date = datetime.strptime(h_cell, '%m/%d/%Y')
            except ValueError:
                h_date = None
    
    if h_date:
        days_diff = (datetime.now() - h_date).days
        if days_diff <= 30:
            result = 'HOLD'
        else:
            result = 'TOUCH BASE'

ws['J4'] = result
wb.save(output_path)
