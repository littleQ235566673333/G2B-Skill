import pandas as pd
from openpyxl import load_workbook

def process_excel(input_path, output_path, out_col, start_row, end_row):
    # Load data as DataFrame
    df = pd.read_excel(input_path)
    headers = df.columns.tolist()
    # Load workbook
    wb = load_workbook(input_path)
    ws = wb.active

    for idx, row_idx in enumerate(range(start_row, end_row + 1)):
        row = df.iloc[idx].tolist()  # Get row values as list
        filtered = [(headers[i], v) for i, v in enumerate(row) if (isinstance(v, (int, float)) and v > 0)]
        if not filtered:
            result = ''
        else:
            min_val = min([v for _, v in filtered])
            min_headers = [h for h, v in filtered if v == min_val]
            result = ','.join(min_headers)
        ws[f'{out_col}{row_idx}'] = result

    wb.save(output_path)

process_excel(
    input_path='results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/44389/input.xlsx',
    output_path='results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/44389/output.xlsx',
    out_col='P', start_row=2, end_row=7
)
