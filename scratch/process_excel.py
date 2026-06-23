import openpyxl
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/before_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/before_pass/core_9726/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Parse Table 1: Student ID (A), Visit number (B), Visit date (C)
table1 = []
for row in ws.iter_rows(min_row=2, max_col=3):
    student_id = row[0].value
    visit_date = row[2].value
    if student_id is not None and visit_date is not None and isinstance(visit_date, datetime):
        table1.append((student_id, visit_date))

# Parse Table 2: Student ID (H, col 8), Submission date (I, col 9), Visit date output (J, col 10)
table2 = []
for row in ws.iter_rows(min_row=2, max_col=10):
    student_id = row[7].value
    submission_date = row[8].value
    visit_date_cell = row[9]
    if student_id is not None and submission_date is not None and isinstance(submission_date, datetime):
        table2.append({'row': visit_date_cell.row, 'student_id': student_id, 'submission_date': submission_date})

# Populate Table 2 Visit date (col J)
for entry in table2:
    student = entry['student_id']
    submission_date = entry['submission_date']
    prior_visits = [vd for s, vd in table1 if s == student and vd < submission_date]
    if prior_visits:
        most_recent = max(prior_visits)
        ws.cell(row=entry['row'], column=10).value = most_recent.strftime('%d-%b-%y')
    else:
        ws.cell(row=entry['row'], column=10).value = None

wb.save(output_path)
