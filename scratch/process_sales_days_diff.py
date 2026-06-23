import pandas as pd
from openpyxl import load_workbook

input_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/59902/input.xlsx'
output_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/59902/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

dates = []
names = []
for row in range(5, 29):
    dates.append(ws[f'F{row}'].value)
    names.append(ws[f'G{row}'].value)

data = pd.DataFrame({'Date': dates, 'Name': names})
results = []
for i, (name, date) in enumerate(zip(data['Name'], data['Date'])):
    if pd.isnull(name) or pd.isnull(date):
        results.append('')
        continue
    # Find the previous row (if any) for the same name
    prev_idx = data.iloc[:i][data.iloc[:i]['Name']==name]['Date'].last_valid_index() if i > 0 else None
    if prev_idx is not None:
        prev_date = data.loc[prev_idx, 'Date']
        diff = (date - prev_date).days
        results.append(diff)
    else:
        results.append(0)
# Write results and format cells
for j, val in enumerate(results):
    cell = ws[f'C{j+5}']
    if val == '':
        cell.value = ''
    else:
        cell.value = val
    cell.number_format = u'#,##0_);[Red](#,##0)'
wb.save(output_path)
