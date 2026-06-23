import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/regression_gate/before_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/regression_gate/before_pass/core_50526/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Lookup value in B6
lookup_value = ws['B6'].value

# Get header names from row 1
headers = [ws.cell(row=1, column=col).value for col in range(2, ws.max_column + 1)]

# Find the row in column A matching the lookup value
lookup_row = None
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=1).value == lookup_value:
        lookup_row = row
        break

matching_headers = []
if lookup_row is not None:
    for col in range(2, ws.max_column + 1):
        cell_value = ws.cell(row=lookup_row, column=col).value
        if cell_value is not None and cell_value > 0:
            matching_headers.append(ws.cell(row=1, column=col).value)

# Write result to B9:B10 (or fewer if there are less than 2 matches)
for i in range(2):
    cell = ws.cell(row=9 + i, column=2)  # B9, B10
    if i < len(matching_headers):
        cell.value = matching_headers[i]
    else:
        cell.value = None

wb.save(output_path)
