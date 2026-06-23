import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/after_pass/core_41969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/after_pass/core_41969/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# For outputs in A6, B6, C6 -> counting blanks in A3:C3, D3:F3, G3:I3
for i in range(3):
    start_col = 1 + i*3
    end_col = start_col + 2
    values = [ws.cell(row=3, column=col).value for col in range(start_col, end_col+1)]
    blank_count = sum(v is None or v == '' for v in values)
    ws.cell(row=6, column=1+i, value=blank_count)

wb.save(output_path)
