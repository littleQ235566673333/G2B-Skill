from openpyxl import load_workbook
wb=load_workbook('results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_4/regression_gate/before_pass/core_54474/input.xlsx')
ws=wb['WHP DATA']
for r in range(1, ws.max_row+1):
    vals=[ws.cell(r,c).value for c in range(1,6)]
    if any(v is not None for v in vals):
        print(r, vals)
