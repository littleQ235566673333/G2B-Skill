import openpyxl

# Define input and output paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/group_59358/r0/evolve_59358/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/group_59358/r0/evolve_59358/output.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read values from L7:L9 for lookup
lookup_cells = ['L7', 'L8', 'L9']
results = []
for lc in lookup_cells:
    results.append(ws[lc].value)

# Read ranges B2:B23 and D2:D23, and F2:F23
range_b = [ws[f'B{i}'].value for i in range(2, 24)]
range_d = [ws[f'D{i}'].value for i in range(2, 24)]
range_f = [ws[f'F{i}'].value for i in range(2, 24)]

# Find the first match for each value in results, and get the corresponding F value
match_outputs = []
for search_value in results:
    output_text = None
    for idx, b_val in enumerate(range_b):
        if b_val == search_value:
            # look for a match in D col (same value somewhere in D)
            if search_value in range_d:
                # output F at this B's index + 2 (as data starts at row 2)
                output_text = range_f[idx]
                break
    match_outputs.append(output_text)

# Write output to L7:L9
for idx, lc in enumerate(lookup_cells):
    ws[lc] = match_outputs[idx]

wb.save(output_path)
