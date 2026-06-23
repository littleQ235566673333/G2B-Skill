import openpyxl
from openpyxl.styles import Alignment, Font
from math import floor
import copy
from datetime import datetime

FILE = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/before_pass/core_32337/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/before_pass/core_32337/output.xlsx'

def to_datetime(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        if val == '?' or not val.strip():
            return None
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
        return None  # just skip if invalid
    return None

wb = openpyxl.load_workbook(FILE)
ws = wb['Sheet1']
col_map = { 'REPORT Date': 1, 'NAME': 2, 'GENDER': 3, 'DATE OF BIRTH': 4, 'Expected Result': 5, 'YEAR (age)': 8, 'CATEGORY': 9, 'age (year)': 15 }
report_date = to_datetime(ws.cell(row=1, column=2).value)

# Insert 'Actual Age' column after 'age (year)'
actual_age_col = col_map['age (year)'] + 1
ws.insert_cols(actual_age_col)
ws.cell(row=2, column=actual_age_col).value = 'Actual Age'
header_fill = copy.copy(ws.cell(row=2, column=actual_age_col-1).fill)
ws.cell(row=2, column=actual_age_col).fill = header_fill
ws.cell(row=2, column=actual_age_col).alignment = Alignment(horizontal='center', vertical='top')
age_col_letter = openpyxl.utils.get_column_letter(col_map['age (year)'])
category_col_letter = openpyxl.utils.get_column_letter(col_map['CATEGORY'])
for row in range(3,16):
    dob = ws.cell(row=row, column=col_map['DATE OF BIRTH']).value
    dob = to_datetime(dob)
    age_val = None
    if dob and report_date:
        age_val = floor((report_date - dob).days / 365.25)
    if age_val is not None:
        ws.cell(row=row, column=actual_age_col).value = age_val
    cell = ws.cell(row=row, column=actual_age_col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = copy.copy(header_fill)
    formula = (
        f"=INDEX(${category_col_letter}$3:${category_col_letter}$15, "
        f"MATCH(${age_col_letter}{row}, ${age_col_letter}$3:${age_col_letter}$15, 0))"
    ).format(row=row)
    ws.cell(row=row, column=col_map['Expected Result']).value = formula
wb.save(OUTPUT)
print('done')
