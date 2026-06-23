import pandas as pd
from openpyxl import load_workbook

vowels = 'AEIOU'
input_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/146-49/input.xlsx'
output_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/146-49/output.xlsx'

# Read words from all cells in the first sheet
# No assumption about the column or layout; flatten all string cells

df = pd.read_excel(input_path, sheet_name='Sheet1', header=None)
words = pd.Series([str(w) for w in df.values.flatten() if isinstance(w, str)])
word_set = set(words)

pairs = set()

def normalized(word, repl):
    return ''.join(repl if c == 'S' else c for c in word)

def is_pair(w1, w2):
    if len(w1) != len(w2):
        return False
    # Find the vowel (could be A/E/I/O/U) substituted for all S in w1
    sub_vowel = None
    for c1, c2 in zip(w1, w2):
        if c1 == c2:
            continue
        elif c1 == 'S' and c2 in vowels:
            if sub_vowel is None:
                sub_vowel = c2
            elif sub_vowel != c2:
                return False
        else:
            return False
    return sub_vowel is not None

for w1 in words:
    for v in vowels:
        if 'S' not in w1:
            continue
        w2 = normalized(w1, v)
        if w2 != w1 and w2 in word_set:
            if is_pair(w1, w2):
                # lexicographic lower goes first
                pair = tuple(sorted([w1, w2], key=str.lower))
                pairs.add(pair)

# Sort the result pairs as per instructions
sorted_pairs = sorted(pairs, key=lambda x: (x[0].lower(), x[1].lower()))

# Write into columns G, H of Sheet1 starting at G1 (row=1, col=7)
wb = load_workbook(input_path)
ws = wb['Sheet1']
for i, (a, b) in enumerate(sorted_pairs):
    ws.cell(row=i+1, column=7, value=a)
    ws.cell(row=i+1, column=8, value=b)
wb.save(output_path)
