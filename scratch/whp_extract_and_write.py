import openpyxl

def extract_whp_data(input_path, output_path):
    # Load workbook
    wb = openpyxl.load_workbook(input_path)
    whp_sheet = wb['WHP']
    whp_data_sheet = wb['WHP DATA']

    # Collect site info from WHP sheet rows 7 and 8
    site_rows = [7, 8]
    site_names = [whp_sheet[f'C{row}'].value for row in site_rows]

    # Read WHP DATA headers
    types_row = 3
    types = [whp_data_sheet[f'C{types_row}'].value, whp_data_sheet[f'D{types_row}'].value, whp_data_sheet[f'E{types_row}'].value]
    # Build site data map from WHP DATA
    site_data_map = {}
    for i in range(4, whp_data_sheet.max_row+1):
        site = whp_data_sheet[f'B{i}'].value
        if site:
            values = [whp_data_sheet[f'C{i}'].value, whp_data_sheet[f'D{i}'].value, whp_data_sheet[f'E{i}'].value]
            site_data_map[site.strip().lower()] = values

    # Write to WHP sheet in E7:G8
    for idx, row in enumerate(site_rows):
        site_key = site_names[idx].strip().lower() if site_names[idx] else None
        vals = site_data_map.get(site_key, [None]*3)
        for col_off, val in enumerate(vals):
            whp_sheet.cell(row=row, column=5+col_off).value = val

    wb.save(output_path)

extract_whp_data(
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/group_54474/r1/evolve_54474/input.xlsx',
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/group_54474/r1/evolve_54474/output.xlsx'
)
