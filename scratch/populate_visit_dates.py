import openpyxl
from datetime import datetime
import pandas as pd

in_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/after_pass/core_9726/input.xlsx'
out_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/after_pass/core_9726/output.xlsx'

wb = openpyxl.load_workbook(in_file)
ws = wb['Sheet1']

visit_rows = []
submission_rows = []

for row in ws.iter_rows(min_row=2, values_only=True):
    # Table 1: Student ID, Visit number, Visit date are cols 0, 1, 2
    sid1, vnum, vdate = row[0], row[1], row[2]
    if sid1 and vdate:
        # Try parsing date, skip if fails
        try:
            vdate_parsed = pd.to_datetime(vdate)
            visit_rows.append({'Student ID': sid1, 'Visit number': vnum, 'Visit date': vdate_parsed})
        except Exception:
            pass
    # Table 2: Student ID, Submission date are cols 7, 8
    sid2, sdate = None, None
    if len(row) > 8:
        sid2, sdate = row[7], row[8]
    if sid2 and sdate:
        try:
            sdate_parsed = pd.to_datetime(sdate)
            submission_rows.append({'Student ID': sid2, 'Submission date': sdate_parsed})
        except Exception:
            pass

visit_df = pd.DataFrame(visit_rows)
sub_df = pd.DataFrame(submission_rows)

result_visit_dates = []
for idx, subrow in sub_df.iterrows():
    student = subrow['Student ID']
    sub_date = subrow['Submission date']
    visits = visit_df[visit_df['Student ID'] == student]
    visits_before = visits[visits['Visit date'] < sub_date]
    if not visits_before.empty:
        last_visit = visits_before['Visit date'].max()
        result_visit_dates.append(last_visit.strftime('%Y-%m-%d'))
    else:
        result_visit_dates.append('')
# Write to column J (index 10, 1-based)
for i, val in enumerate(result_visit_dates, start=2):
    ws.cell(row=i, column=10, value=val)
wb.save(out_file)
