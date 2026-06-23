import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_54638_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_54638_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws1 = wb['Sheet1']

# Use Calibri font for the entire sheet
calibri_font = Font(name='Calibri')
for row in ws1.iter_rows():
    for cell in row:
        cell.font = calibri_font

# --- COLUMN A: Formulas and Borders ---
A_formula_template = '=IF(Sheet2!A{row}="", "", Sheet2!A{row})'
ws1['A1'].value = 'Time off '
for r in range(2, 151):
    ws1[f'A{r}'].value = A_formula_template.format(row=r)
    ws1[f'A{r}'].font = calibri_font
# Borders for A1:A13 only
thin = Side(border_style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for r in range(1, 14):
    ws1[f'A{r}'].border = border
for r in range(14, 151):
    ws1[f'A{r}'].border = Border()

# --- COLUMN B: Unique Non-Array Formulas, Highlight, Borders ---
yellow_fill = PatternFill('solid', fgColor='FFFF00')
unique_formula_template = '=IFERROR(INDEX($A$2:$A$150, MATCH(0, COUNTIF($B$1:B{row_minus_1}, $A$2:$A$150), 0)), "")'
for r in range(2, 151):
    fmla = unique_formula_template.replace('{row_minus_1}', str(r-1))
    cell = ws1[f'B{r}']
    cell.value = fmla
    cell.fill = yellow_fill
    cell.font = calibri_font
    cell.border = border

# Sheet settings: Remove gridlines for Sheet1
ws1.sheet_view.showGridLines = False

# Hide Sheet2
wb['Sheet2'].sheet_state = 'hidden'

wb.save(output_path)
