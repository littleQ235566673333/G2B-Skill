import openpyxl
from datetime import datetime

INPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/after_fix/core_4714/input.xlsx'
OUTPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/after_fix/core_4714/output.xlsx'
SHEET = 'Sheet2'
START_ROW = 2
END_ROW = 25  # inclusive

wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb[SHEET]

rows = list(ws.iter_rows(min_row=START_ROW, max_row=END_ROW, max_col=4, values_only=True))
# Read headers from first row for context
headers = [cell.value for cell in ws[1][:4]]

# Create a list of dicts for each row for easier manipulation
records = []
for idx, row in enumerate(rows, START_ROW):
    emp_code, _, month_str, hours = row
    # Convert month to datetime object
    if isinstance(month_str, str):
        try:
            month = datetime.strptime(month_str, "%Y-%m")
        except Exception:
            try:
                month = datetime.strptime(month_str, "%Y/%m")
            except Exception:
                month = None
    else:
        month = month_str
    records.append({'rownum': idx, 'emp_code': emp_code, 'month': month, 'hours': hours})

# Calculate 4-month rolling averages and format for column E
output_results = []
for i, rec in enumerate(records):
    emp_code = rec['emp_code']
    month = rec['month']
    hours = rec['hours']
    if emp_code is None or month is None or hours is None:
        output_results.append('n/a')
        continue
    # Find all previous months (including current) for this employee up to 4
    emp_records = [r for r in records if r['emp_code']==emp_code and r['month'] is not None and r['month']<=month]
    emp_records = sorted(emp_records, key=lambda r: r['month'], reverse=True)
    window_records = emp_records[:4]
    if len(window_records) < 4:
        output_results.append('n/a')
    else:
        avg = sum(r['hours'] for r in window_records)/4
        # Output whole number if integer
        if avg == int(avg):
            avg_str = f"{int(avg)}"
        else:
            avg_str = f"{round(avg, 1)}"
        output_results.append(avg_str)

# Write to column E (col 5)
for offset, value in enumerate(output_results):
    ws.cell(row=START_ROW+offset, column=5, value=value)

wb.save(OUTPUT_PATH)
print(f"Saved output to {OUTPUT_PATH}")
