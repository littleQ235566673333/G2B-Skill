from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_8/regression_gate/before_fix/core_57033/input.xlsx"
output_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_8/regression_gate/before_fix/core_57033/output.xlsx"


def norm_text(value):
    return " ".join(str(value or "").split()).casefold()


def title_case_if_text(value):
    if isinstance(value, str):
        return value.strip().title()
    return value

wb = load_workbook(input_path)
ws = wb["Sheet4"]
md = wb["CBtrans"]

lookup = {}
for row in md.iter_rows(min_row=2, values_only=True):
    company, _, _, _, _, _, _, account, _, _, xchar = row[:11]
    key = (norm_text(company), norm_text(account), norm_text(xchar))
    lookup[key] = "Match"

fill = PatternFill(fill_type="solid", fgColor="FF66CC")

affected_cells = []
for r in range(2, 8):
    company = ws[f"B{r}"].value
    xchar = ws[f"E{r}"].value
    account = ws[f"I{r}"].value
    key = (norm_text(company), norm_text(account), norm_text(xchar))
    result = lookup.get(key, "-")
    ws[f"K{r}"] = result
    affected_cells.append(ws[f"K{r}"])

    for col in ["B", "I", "K"]:
        cell = ws[f"{col}{r}"]
        cell.value = title_case_if_text(cell.value)
        affected_cells.append(cell)

for cell in affected_cells:
    cell.fill = copy(fill)

wb.save(output_path)

# Re-open to verify saved values
wb2 = load_workbook(output_path)
ws2 = wb2["Sheet4"]
for r in range(2, 8):
    print(r, ws2[f"K{r}"].value)
