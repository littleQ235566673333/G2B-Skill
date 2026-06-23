import openpyxl

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_39432/r2/evolve_39432/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_39432/r2/evolve_39432/output.xlsx'

wb = openpyxl.load_workbook(input_fp)
ws = wb.active

# Column indices per Excel (1-based):
# A = 1 (Item), B = 2 (Begin Qty), C = 3 (Begin Cost), D...K = purchase orders
# L = 12 (Inventory On Hand), M = 13 (Per Unit Cost, output)

for row in range(2, 6):  # Rows 2-5 for data
    fifo_lots = []

    # Beginning inventory
    begin_qty = ws.cell(row=row, column=2).value or 0
    begin_cost = ws.cell(row=row, column=3).value or 0.0
    if begin_qty > 0:
        fifo_lots.append({'qty': begin_qty, 'cost': begin_cost})
    
    # Subsequent purchase order columns (D-K, 4-11)
    for col in range(4, 12):
        po_qty = ws.cell(row=row, column=col).value or 0
        po_cost = ws.cell(row=row, column=col+8).value if ws.cell(row=1, column=col+8).value and col+8 <= ws.max_column else None
        # Automate cost parsing: each unit column in D-K should have a paired cost column
        if po_qty > 0 and po_cost:
            fifo_lots.append({'qty': po_qty, 'cost': po_cost})
    
    inventory_on_hand = ws.cell(row=row, column=12).value or 0

    # FIFO calculation: grab enough units from earliest lots
    remaining = inventory_on_hand
    taken = []
    for lot in fifo_lots:
        take_qty = min(lot['qty'], remaining)
        if take_qty > 0:
            taken.append({'qty': take_qty, 'cost': lot['cost']})
            remaining -= take_qty
        if remaining == 0:
            break
    
    if taken:
        # Weighted average cost per unit
        total_cost = sum(t['qty'] * t['cost'] for t in taken)
        total_qty = sum(t['qty'] for t in taken)
        per_unit_cost = round(total_cost / total_qty, 2) if total_qty else 0
    else:
        per_unit_cost = 0
    ws.cell(row=row, column=13).value = per_unit_cost

wb.save(output_fp)
