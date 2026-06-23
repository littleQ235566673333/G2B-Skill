import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_3/regression_gate/before_pass/core_41969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_3/regression_gate/before_pass/core_41969/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# For cells A6, B6, C6: ranges A3:C3, D3:F3, G3:I3
for i in range(3):
    start_col = i * 3 + 1  # Columns 1, 4, 7
    end_col = start_col + 3  # up to 3, 6, 9 (exclusive)
    blank_count = 0
    for col in range(start_col, end_col):
        value = ws.cell(row=3, column=col).value
        if value is None or str(value).strip() == '':
            blank_count += 1
    ws.cell(row=6, column=1 + i).value = blank_count

wb.save(output_path)
