import openpyxl
import math

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_44389/r3/evolve_44389/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_44389/r3/evolve_44389/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

header = [cell.value for cell in ws[1]][1:14]  # B1:N1

for row_idx in range(2, 8):  # P2:P7 for results
    # Read row values B:N (columns 2 to 14)
    values = []
    for col_idx in range(2, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        v = cell.value
        try:
            v = float(v)
        except (TypeError, ValueError):
            v = None
        values.append(v)

    nonzero_values = [(idx, v) for idx, v in enumerate(values) if v is not None and v > 0]

    if not nonzero_values:
        min_headers = ''
    else:
        min_value = min(v for idx, v in nonzero_values)
        min_indices = [idx for idx, v in nonzero_values if math.isclose(v, min_value)]
        min_headers = ','.join([header[idx] for idx in min_indices])

    ws.cell(row=row_idx, column=16, value=min_headers)  # Column P = 16

wb.save(output_path)
