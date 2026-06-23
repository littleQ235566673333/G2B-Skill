import openpyxl
from datetime import datetime, timedelta

INPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/group_44017/r3/evolve_44017/input.xlsx'
OUTPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/group_44017/r3/evolve_44017/output.xlsx'

wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb.active

# Output columns (AD:AO) are 30..41
OUTPUT_COLS = list(range(30, 42 + 1))   # AD to AO
OUTPUT_ROWS = list(range(14, 43))  # 14 to 42

# Find a starting date anchor for formulas
# Col 17 is date (row 9)
anchor_date = ws.cell(row=9, column=17).value
assert isinstance(anchor_date, datetime)

# Calculate column dates
col_dates = []
prev_date = anchor_date
for col in range(17, 43):
    cell = ws.cell(row=9, column=col)
    v = cell.value
    if isinstance(v, datetime):
        col_dates.append(v)
        prev_date = v
    elif isinstance(v, str) and v.startswith('=DATE('):
        # move to first day of next month
        next_month = (prev_date.month % 12) + 1
        next_year = prev_date.year + (1 if prev_date.month == 12 else 0)
        new_date = datetime(next_year, next_month, 1)
        col_dates.append(new_date)
        prev_date = new_date
    else:
        col_dates.append(None)

# Only take AD:AO columns
col_dates = col_dates[13:25]

for row in OUTPUT_ROWS:
    base_rate = ws.cell(row=row, column=23).value  # W
    frequency = ws.cell(row=row, column=10).value  # J
    effective_date = ws.cell(row=row, column=12).value  # L
    increases = [ws.cell(row=row, column=c).value for c in range(13, 17)]  # M-P

    try:
        freq_months = int(frequency)
    except:
        freq_months = None
    if freq_months is None or base_rate is None or effective_date is None:
        for i, col in enumerate(OUTPUT_COLS):
            ws.cell(row=row, column=col).value = None
        continue

    for i, (col, col_date) in enumerate(zip(OUTPUT_COLS, col_dates)):
        if col_date is None or col_date < effective_date:
            ws.cell(row=row, column=col).value = None
            continue

        months_since_effective = (col_date.year - effective_date.year) * 12 + (col_date.month - effective_date.month)
        waves = months_since_effective // freq_months if freq_months > 0 else 0
        waves = min(waves, 4)
        new_rate = base_rate
        for wave in range(waves):
            inc = increases[wave]
            if inc is not None:
                new_rate *= (1 + inc)
        ws.cell(row=row, column=col).value = round(new_rate, 2)

wb.save(OUTPUT_PATH)
print('Done.')
