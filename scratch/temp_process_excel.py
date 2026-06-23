import openpyxl
import numpy as np

input_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/49300/input.xlsx'
output_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/49300/output.xlsx'

wb = openpyxl.load_workbook(input_path)
data_ws = wb['Data']
sheet1_ws = wb['Sheet1']

# Read lookup values from Sheet1 (A2, B2), (A3, B3)
criteria = []
for r in range(2, 4):
    a_val = sheet1_ws[f'A{r}'].value
    b_val = sheet1_ws[f'B{r}'].value
    criteria.append((a_val, b_val))

# Read the Data table structure
# - Row 2: column headers (C2:AY2)
# - Column B (B3:B4...): row headers
col_headers = []
for col in range(3, 52):  # C=3 to AY=51
    val = data_ws.cell(row=2, column=col).value
    col_headers.append(val)
row_headers = []
for row in range(3, 5):   # 3 to 4
    val = data_ws.cell(row=row, column=2).value
    row_headers.append(val)

table = np.array([[data_ws.cell(row=r, column=c).value for c in range(3, 52)] for r in range(3,5)])

for idx, (row_label, col_label) in enumerate(criteria):
    # Find row
    try:
        r_idx = row_headers.index(row_label)
    except ValueError:
        sheet1_ws[f'C{2+idx}'] = None
        continue
    # Find column
    try:
        c_idx = col_headers.index(col_label)
    except ValueError:
        sheet1_ws[f'C{2+idx}'] = None
        continue
    val = table[r_idx, c_idx]
    # If not a number, leave blank
    if val is None:
        sheet1_ws[f'C{2+idx}'] = None
    else:
        try:
            sheet1_ws[f'C{2+idx}'] = float(val)
        except Exception:
            sheet1_ws[f'C{2+idx}'] = val

wb.save(output_path)
