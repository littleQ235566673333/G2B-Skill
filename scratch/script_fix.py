import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/regression_gate/before_fix/core_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/regression_gate/before_fix/core_57033/output.xlsx'
wb = openpyxl.load_workbook(input_path)
sheet4 = wb['Sheet4']
cb = wb['CBtrans']
cb_rows = list(cb.iter_rows(min_row=2, max_row=11, values_only=True))

def threeway_match(sheet4_row):
    company = sheet4_row[1]
    account = sheet4_row[8]
    xchar = sheet4_row[2]
    for cb_row in cb_rows:
        if cb_row[0] == company and cb_row[7] == account and cb_row[10] == xchar:
            return 'Match'
    return '-'

fill = PatternFill(start_color='FF66CC', end_color='FF66CC', fill_type='solid')

# Fill K2:K7
for idx, row in enumerate(sheet4.iter_rows(min_row=2, max_row=7, values_only=True), start=2):
    result = threeway_match(row)
    cell = sheet4[f'K{idx}']
    cell.value = result
    cell.fill = fill

wb.save(output_path)
