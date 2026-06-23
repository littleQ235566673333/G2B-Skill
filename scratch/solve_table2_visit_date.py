import openpyxl
from datetime import datetime

def make_date(dt):
    if isinstance(dt, datetime):
        return dt
    if isinstance(dt, str):
        for fmt in ('%d-%b-%y','%Y-%m-%d','%m/%d/%Y'):
            try:
                return datetime.strptime(dt, fmt)
            except:
                pass
    return None

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/after_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/after_pass/core_9726/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheet = wb.active

# Read Table 1: rows 2–7, columns 1–3
visits = {}
for i in range(2, 8):
    student = sheet.cell(row=i, column=1).value
    visit_date = sheet.cell(row=i, column=3).value
    if student is not None and visit_date is not None:
        dt = make_date(visit_date)
        if dt:
            visits.setdefault(student, []).append(dt)

# Sort visit dates descending for each student
for stu in visits:
    visits[stu].sort(reverse=True)

# Table 2: Only row 2, columns 8 (H, Student ID) and 9 (I, Submission date)
student2 = sheet.cell(row=2, column=8).value
submission_date2 = sheet.cell(row=2, column=9).value
submission_dt2 = make_date(submission_date2)
vdates = visits.get(student2, [])
prior = None
for vdt in vdates:
    if vdt < submission_dt2:
        prior = vdt
        break
sheet.cell(row=2, column=10).value = prior.strftime('%d-%b-%y') if prior else None

wb.save(output_path)
