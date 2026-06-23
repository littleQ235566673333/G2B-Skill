from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4_v6/train/iter_3/regression_gate/after_fix/core_57033/input.xlsx"
output_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4_v6/train/iter_3/regression_gate/after_fix/core_57033/output.xlsx"

wb = load_workbook(input_path)
print('Sheets:', wb.sheetnames)
for ws in wb.worksheets:
    print(f'--- {ws.title} ---')
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), min_col=1, max_col=min(ws.max_column, 12), values_only=True):
        print(row)
