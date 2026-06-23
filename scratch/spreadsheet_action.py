import openpyxl
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/after_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/after_pass/core_41589/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Contact List']

# Read values from H4 (date) and I4 (YES/NO)
date_cell = ws['H4'].value
response_cell = ws['I4'].value

action = 'NO ACTION'

if date_cell is not None and response_cell is not None:
    # Ensure date is handled as datetime
    if isinstance(date_cell, datetime.datetime):
        last_contact_date = date_cell
    else:
        try:
            last_contact_date = datetime.datetime.strptime(str(date_cell), '%Y-%m-%d')
        except Exception:
            action = 'NO ACTION'
            last_contact_date = None
    
    if last_contact_date:
        today = datetime.datetime.today()
        days_diff = (today - last_contact_date).days
        if response_cell == 'YES':
            if days_diff <= 30:
                action = 'HOLD'
            else:
                action = 'TOUCH BASE'

ws['J4'] = action
wb.save(output_path)
