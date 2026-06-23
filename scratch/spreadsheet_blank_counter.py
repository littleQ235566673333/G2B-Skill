import openpyxl

def count_blank_cells(row_values):
    # Only consider columns B to M (1-12 month data)
    numeric_indices = [i for i, v in enumerate(row_values[1:13], 1) if isinstance(v, (int, float)) and v is not None]
    
    if len(numeric_indices) < 2:
        return 0
    
    first_idx = numeric_indices[0]
    second_idx = numeric_indices[1]
    first_value = row_values[first_idx]
    
    if first_value > 1:
        return 1
    
    # Count blanks from after the first numeric entry up to just before the second
    blank_count = 0
    for i in range(first_idx + 1, second_idx):
        if row_values[i] is None:
            blank_count += 1
    return blank_count

# Load workbook and worksheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/after_fix/core_50521/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/after_fix/core_50521/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Rows 4, 5, 6
for row_idx, output_idx in zip(range(4, 7), range(4, 7)):
    row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 14)]
    result = count_blank_cells(row_values)
    ws.cell(row=output_idx, column=14, value=result)  # Column N: 14

wb.save(output_path)
