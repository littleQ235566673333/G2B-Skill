import pandas as pd
from openpyxl import load_workbook

# Paths
input_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/6698/input.xlsx'
output_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/6698/output.xlsx'

# Load workbook
wb = load_workbook(input_path)
ws = wb.active

# Read raw data
data = pd.read_excel(input_path, sheet_name=wb.sheetnames[0])
raw = pd.read_excel(input_path, sheet_name='raw 1')

# Get F4:F8 values from the main sheet
F_values = [ws[f'F{i}'].value for i in range(4, 9)]
results = []

# Assumption: 'raw 1' columns: A->0, B->1, C->2
for f in F_values:
    mask_A = raw.iloc[:, 0] == f
    mask_B = raw.iloc[:, 1] != 0
    mask_C = raw.iloc[:, 2] != 0
    count = ((mask_B | mask_C) & mask_A).sum()
    results.append(count)

# Put results in G4:G8
for idx, val in enumerate(results, start=4):
    ws[f'G{idx}'] = val

# Add filter to A1:C1 of the main sheet
ws.auto_filter.ref = 'A1:C1'

# Save
wb.save(output_path)
