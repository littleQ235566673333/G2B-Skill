import openpyxl, re
from itertools import product

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/before_fix/core_325-44/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/before_fix/core_325-44/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_in = wb['Input']
ws_out = wb['Output']
header = [cell.value for cell in ws_out[1]]
result_rows = [header]

def parse_filter(filter_str):
    vals = {}
    if not filter_str:
        return vals
    # Matches key = value
    m = re.findall(r'\("(\w+)"\s*=\s*"([\w]+)"\)', filter_str)
    for k, v in m:
        vals[k] = [v]
    # Matches key IN LIST values
    m = re.findall(r'\("(\w+)"\s+IN\s+LIST\s+([\d,]+)\)', filter_str)
    for k, v in m:
        vals[k] = [str(int(i)) for i in v.split(',')]
    return vals

for row in ws_in.iter_rows(min_row=2, max_row=10, min_col=1, max_col=5):
    wtype_id, name, status, filter_str, _ = [cell.value for cell in row]
    if not wtype_id:
        continue
    vals = parse_filter(filter_str)
    orgIds = vals.get('orgId', [''])
    careTypes = vals.get('careType', [''])
    specialtyIds = vals.get('specialtyId', [''])
    contextIds = vals.get('contextId', [''])
    # For each combination
    if not specialtyIds:
        specialtyIds = ['']
    if not contextIds:
        contextIds = ['']
    combinations = list(product(orgIds, careTypes, specialtyIds, contextIds))
    for orgId, careType, specialtyId, contextId in combinations:
        out_row = [wtype_id, name, status, orgId or None, careType or None, specialtyId or None, contextId or None]
        if not any(out_row[3:]):  # skip rows with all blank split cols
            continue
        result_rows.append(out_row)

# Rewrite Output tab A1:G10
for idx, row in enumerate(ws_out.iter_rows(min_row=1, max_row=10, min_col=1, max_col=7)):
    values = result_rows[idx] if idx < len(result_rows) else [None]*7
    for jdx, cell in enumerate(row):
        cell.value = values[jdx]

wb.save(output_path)
