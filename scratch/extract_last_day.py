from openpyxl import load_workbook
import pandas as pd

input_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/31628/input.xlsx'
output_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/31628/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Read dates from A1:A11
dates = []
for row in range(1, 12):
    val = ws[f'A{row}'].value
    if val is not None:
        dates.append(val)

# Get the last non-empty date
if not dates:
    last_day_number = ''
else:
    last_date = dates[-1]
    # If it's a datetime object, get the day
    try:
        last_day_number = last_date.day
    except AttributeError:
        # If it's a string, try to parse using pandas
        try:
            last_day_number = pd.to_datetime(last_date).day
        except Exception:
            last_day_number = ''

ws['B1'] = last_day_number
wb.save(output_path)
