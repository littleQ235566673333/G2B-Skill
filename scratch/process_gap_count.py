import openpyxl

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/group_50521/r1/evolve_50521/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/group_50521/r1/evolve_50521/output.xlsx'

wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Helper function for the calculation
def process_row(row):
    nums = []
    idxs = []
    # Identify numeric entries and their column indices
    for idx, cell in enumerate(row):
        if isinstance(cell.value, (int, float)) and cell.value != '' and cell.value is not None:
            nums.append(cell.value)
            idxs.append(idx)
    if len(nums) < 2:
        return None
    if nums[0] > 1:
        return 1
    # Count blanks between first and second numeric cells (exclusive)
    start, end = idxs[0], idxs[1]
    if end - start <= 1:
        return 0
    blank_count = sum(1 for cell in row[start+1:end] if cell.value in (None, ''))
    return blank_count

# Rows 4, 5, 6 => customers; columns from B onwards are monthly order numbers
for excel_row, out_cell in zip(range(4, 7), ['N4', 'N5', 'N6']):
    row = list(ws[excel_row])
    # Skip first column (likely customer name/id), so process from column B
    values = row[1:]
    result = process_row(values)
    ws[out_cell] = result

wb.save(output_file)
