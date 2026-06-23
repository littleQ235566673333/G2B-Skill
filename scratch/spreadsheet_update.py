import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import calendar

def increment_month(dt, n=1):
    y = dt.year
    m = dt.month + n
    while m > 12:
        y += 1
        m -= 12
    return datetime(y, m, 1)

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/group_44017/r1/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/group_44017/r1/evolve_44017/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

rows = range(14, 43)
output_cols = ['AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO']
output_col_nums = [openpyxl.utils.column_index_from_string(col) for col in output_cols]

# Build month_dates for columns AD:AO
month_dates = [None]*len(output_col_nums)
for i, col_num in enumerate(output_col_nums):
    val = ws.cell(row=9, column=col_num).value
    if isinstance(val, datetime):
        month_dates[i] = val
    else:
        if i == 0:
            prior = ws.cell(row=9, column=col_num-2).value
            if isinstance(prior, datetime):
                month_dates[i] = increment_month(prior, 1)
            else:
                month_dates[i] = datetime(2000,1,1)
        else:
            month_dates[i] = increment_month(month_dates[i-1], 1)

for row in rows:
    base_rate = ws.cell(row=row, column=23).value  # column W=23
    freq = ws.cell(row=row, column=10).value       # column J=10
    eff_date = ws.cell(row=row, column=12).value   # column L=12
    increases = [ws.cell(row=row, column=col).value or 0 for col in range(13, 17)]  # M:P = 13-16
    if eff_date is None:
        for col_num in output_col_nums:
            cell = ws.cell(row=row, column=col_num)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
        continue
    if not isinstance(eff_date, datetime):
        try:
            eff_date = datetime.strptime(str(eff_date), '%Y-%m-%d')
        except:
            try:
                eff_date = datetime.strptime(str(eff_date), '%m/%d/%Y')
            except:
                for col_num in output_col_nums:
                    cell = ws.cell(row=row, column=col_num)
                    cell.value = None
                    cell.fill = PatternFill(fill_type=None)
                continue
    for i, col_num in enumerate(output_col_nums):
        month = month_dates[i]
        cell = ws.cell(row=row, column=col_num)
        if month < eff_date:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
        else:
            periods = freq or 1
            months_diff = (month.year - eff_date.year)*12 + (month.month - eff_date.month)
            n_increases = min(4, max(0, months_diff//periods + 1))
            rate = base_rate
            for inc_idx in range(n_increases):
                pct = increases[inc_idx] or 0
                rate = rate * (1 + pct)
            cell.value = round(rate,2)
            cell.fill = PatternFill(fill_type=None)

wb.save(output_path)
