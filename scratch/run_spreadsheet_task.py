import openpyxl
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/group_45707/r2/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/group_45707/r2/evolve_45707/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
max_row = 69  # inclusive, as specified

for i in range(2, max_row + 1):
    # Skip the header row (A1, C1, E1 stay unchanged)
    if i == 1:
        continue
    cell_next = ws.cell(row=i + 1, column=1)
    if isinstance(cell_next.value, datetime.datetime):
        next_day = cell_next.value
        if next_day.day == 1:
            search_month = next_day.month
            search_year = next_day.year
            # Count how many '1's in col C for same month/year
            count = 0
            for j in range(2, max_row + 1):
                date_cell = ws.cell(row=j, column=1)
                col_c_cell = ws.cell(row=j, column=3)
                if isinstance(date_cell.value, datetime.datetime):
                    if (date_cell.value.month == search_month and date_cell.value.year == search_year
                        and col_c_cell.value == 1):
                        count += 1
            ws.cell(row=i, column=4).value = count
        else:
            ws.cell(row=i, column=4).value = None
    else:
        ws.cell(row=i, column=4).value = None

wb.save(output_path)
