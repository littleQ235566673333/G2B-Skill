import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/after_pass/core_50526/output.xlsx'

# Load workbook and select the active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get the lookup value from B6
lookup_value = ws['B6'].value

# Get headers from row 1 (starting from column 2, since column 1 is lookup)
header = [cell.value for cell in ws[1]][1:]

# Find row matching the lookup value in column 1
lookup_row = None
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].value == lookup_value:
        lookup_row = row
        break

# Collect headers where cell value in that row is > 0
res = []
if lookup_row:
    for idx, cell in enumerate(lookup_row[1:]):
        if isinstance(cell.value, (int, float)) and cell.value > 0:
            res.append(header[idx])

# Write up to 2 results to B9:B10
for i in range(2):
    cell_addr = f'B{9 + i}'
    ws[cell_addr].value = res[i] if i < len(res) else None

# Save workbook
wb.save(output_path)
