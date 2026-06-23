import openpyxl

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/before_pass/core_41969/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/before_pass/core_41969/output.xlsx"

# Load workbook and select sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Define formula patterns for each cell in A6:C6
for i, col in enumerate(['A', 'B', 'C']):
    start_col = openpyxl.utils.cell.column_index_from_string('A') + i * 3
    end_col = start_col + 2
    start_col_letter = openpyxl.utils.get_column_letter(start_col)
    end_col_letter = openpyxl.utils.get_column_letter(end_col)
    formula = f"=COUNTBLANK({start_col_letter}3:{end_col_letter}3)"
    ws[f"{col}6"] = formula

# Save result
wb.save(output_path)
