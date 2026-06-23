import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/after_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/after_pass/core_18935/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheet = wb['Sheet1']

# Data Table rows: 5-10 (index 4-9)
# Columns: A: Work, B: Material, C: Category 1, D: Category 2, E: Category 3

data_table = []
for row in sheet.iter_rows(min_row=5, max_row=10, min_col=1, max_col=5, values_only=True):
    data_table.append(row)

category_column = {
    'Category 1': 2, # C
    'Category 2': 3, # D
    'Category 3': 4  # E
}

# Report rows: 17-22 (index 16-21)
for report_row in range(17, 23):
    work = sheet.cell(row=report_row, column=1).value
    category_type = sheet.cell(row=report_row, column=2).value
    material = sheet.cell(row=report_row, column=3).value
    
    cat_col = category_column.get(category_type)
    
    found_value = None
    for dt_row in data_table:
        if dt_row[0] == work and dt_row[1] == material:
            found_value = dt_row[cat_col]
            break
    # Write output to column D
    sheet.cell(row=report_row, column=4).value = found_value

wb.save(output_path)