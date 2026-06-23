import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/group_50521/r0/evolve_50521/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/group_50521/r0/evolve_50521/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheet = wb.active

def count_blanks_between_entries(row):
    numeric_indices = []
    values = [sheet.cell(row=row, column=col).value for col in range(2, sheet.max_column+1)]
    for idx, val in enumerate(values):
        if isinstance(val, (int, float)) and val != '':
            numeric_indices.append(idx)
            if len(numeric_indices) == 2:
                break
    if len(numeric_indices) < 2:
        return 0
    first_val = values[numeric_indices[0]]
    if first_val > 1:
        return 1
    blanks = values[numeric_indices[0]+1:numeric_indices[1]]
    return sum(1 for b in blanks if b in [None, '', ' '])

for row in range(4, 7):
    count = count_blanks_between_entries(row)
    sheet.cell(row=row, column=14).value = count

wb.save(output_path)
