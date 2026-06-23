import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_57989_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_57989_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Days in the header row (row 1, col B-U)
day_headers = [ws.cell(row=1, column=col).value for col in range(2, 22)]
# Data rows are A3:A21 (drivers), cols B:U
driver_rows = {ws.cell(row=r, column=1).value: r for r in range(3, 22) if ws.cell(row=r, column=1).value} # Map: driver -> row
weekdays = [ws.cell(row=24, column=col).value for col in range(2, 9)]
# B25:H43
for outrow in range(25, 44):
    driver = ws.cell(row=outrow, column=1).value
    if not driver: continue
    src_row = driver_rows.get(driver)
    if not src_row: continue
    for outcol, weekday in enumerate(weekdays, start=2):
        # Find all columns in B1:U1 that match this weekday
        day_cols = [i+2 for i, day in enumerate(day_headers) if day==weekday]
        # For the driver's row, count non-empty cells in those columns
        count = sum(1 for col in day_cols if ws.cell(row=src_row, column=col).value not in (None, ''))
        ws.cell(row=outrow, column=outcol, value=count)
wb.save(output_path)
