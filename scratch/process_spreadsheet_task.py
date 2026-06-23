import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# File paths
template = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/regression_gate/before_pass/core_32337/input.xlsx'
output   = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_5/regression_gate/before_pass/core_32337/output.xlsx'

wb = openpyxl.load_workbook(template)
ws = wb.active

col_expected_result = 5  # E
col_year_age = 8         # H
col_category = 9         # I
col_age_year = 15        # O
row_start, row_end = 3, 15

# Write formula for CATEGORY lookup in column E (Expected Result)
formula_lookup = '=INDEX($I$3:$I$15, MATCH(O{row}, $H$3:$H$15, 0))'
for row in range(row_start, row_end+1):
    ws.cell(row=row, column=col_expected_result).value = formula_lookup.format(row=row)

# Insert Actual Age column after Age (year) column O
actual_age_col = col_age_year + 1
ws.insert_cols(actual_age_col)
ws.cell(row=2, column=actual_age_col).value = 'Actual Age'

col_dob = 3 # DATE OF BIRTH
# Find REPORT DATE (look for row value using $B$1)
col_report_date = None
for formula_row in range(row_start, row_end+1):
    for cell in ws[formula_row]:
        if isinstance(cell.value, str) and '$B$1' in cell.value:
            col_report_date = cell.column
            break
    if col_report_date:
        break

for row in range(row_start, row_end+1):
    dob = ws.cell(row=row, column=col_dob).value
    report_date = ws.cell(row=row, column=col_report_date).value
    if isinstance(report_date, str) and '$B$1' in report_date:
        report_date = ws.cell(row=1, column=2).value
    if isinstance(dob, datetime) and isinstance(report_date, datetime):
        age = int((report_date - dob).days // 365)
        ws.cell(row=row, column=actual_age_col).value = age
    else:
        ws.cell(row=row, column=actual_age_col).value = None

# Try to copy only relevant simple attributes for fill
preceding_cell = ws.cell(row=3, column=col_age_year)
prev_fill = preceding_cell.fill
if prev_fill.patternType:
    cell_fill = PatternFill(
        fill_type=prev_fill.patternType,
        fgColor=prev_fill.fgColor.rgb if prev_fill.fgColor.type == 'rgb' else 'FFFFFF',
        bgColor=prev_fill.bgColor.rgb if prev_fill.bgColor.type == 'rgb' else 'FFFFFF'
    )
else:
    # fallback: plain white no fill
    cell_fill = PatternFill(fill_type=None)

for row in range(row_start, row_end+1):
    cell = ws.cell(row=row, column=actual_age_col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.number_format = '0'
    cell.fill = cell_fill

header_cell = ws.cell(row=2, column=actual_age_col)
header_cell.alignment = Alignment(horizontal='center', vertical='top')
header_cell.fill = cell_fill

wb.save(output)
