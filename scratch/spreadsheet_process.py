import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/regression_gate/before_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/regression_gate/before_pass/core_32337/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# 1. Fill Expected Result column E with formula referencing CATEGORY column I
for row in range(3, 16):  # E3:E15
    ws[f'E{row}'].value = f'=$I{row}'

# 2. Insert Actual Age column after age (year) (O)
actual_age_col_idx = 16  # after col O (15)
ws.insert_cols(actual_age_col_idx)
ws.cell(row=2, column=actual_age_col_idx).value = 'Actual Age'

# Copy fill color from preceding column (O)
preceding_fill = ws.cell(row=2, column=actual_age_col_idx - 1).fill
if isinstance(preceding_fill, PatternFill):
    fill = PatternFill(fill_type=preceding_fill.fill_type, fgColor=preceding_fill.fgColor, bgColor=preceding_fill.bgColor)
else:
    fill = PatternFill(fill_type=None)

for row in range(3, 16):
    ws.cell(row=row, column=actual_age_col_idx).value = f'=ROUNDDOWN(($B$1-C{row})/365,0)'
    ws.cell(row=row, column=actual_age_col_idx).font = Font(bold=True)
    ws.cell(row=row, column=actual_age_col_idx).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=actual_age_col_idx).fill = PatternFill(fill_type=fill.fill_type, fgColor=fill.fgColor, bgColor=fill.bgColor)

# Header formatting
ws.cell(row=2, column=actual_age_col_idx).font = Font(bold=True)
ws.cell(row=2, column=actual_age_col_idx).alignment = Alignment(horizontal='center', vertical='top')
ws.cell(row=2, column=actual_age_col_idx).fill = PatternFill(fill_type=fill.fill_type, fgColor=fill.fgColor, bgColor=fill.bgColor)

wb.save(output_path)
