import openpyxl
from datetime import datetime

def parse_date(cell):
    if isinstance(cell, datetime):
        return cell
    if isinstance(cell, str):
        for fmt in ('%d-%b-%y', '%d/%m/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(cell, fmt)
            except Exception:
                continue
    return None

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/before_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/before_pass/core_9726/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Table 1: columns A (Student), C (Visit date), rows 2-7
# Table 2: columns H (Student), I (Submission date)
# Output: column J

table1 = []
for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=3):
    student = row[0].value
    visit_date = parse_date(row[2].value)
    table1.append({'Student': student, 'Visit date': visit_date})

table2_rows = []
row_idx = 2
while ws[f'H{row_idx}'].value:
    student = ws[f'H{row_idx}'].value
    submission_date = parse_date(ws[f'I{row_idx}'].value)
    table2_rows.append({'row': row_idx, 'Student': student, 'Submission date': submission_date})
    row_idx += 1

for entry in table2_rows:
    visits = [r['Visit date'] for r in table1 if r['Student'] == entry['Student'] and r['Visit date'] is not None and r['Visit date'] < entry['Submission date']]
    if visits:
        last_visit = max(visits)
        ws.cell(row=entry['row'], column=10).value = last_visit
    else:
        ws.cell(row=entry['row'], column=10).value = None

wb.save(output_path)
