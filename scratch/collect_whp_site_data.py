import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/before_fix/core_54474/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/before_fix/core_54474/output.xlsx'

wb = openpyxl.load_workbook(input_path)
whp_sheet = wb['WHP']
data_sheet = wb['WHP DATA']

# Get mapping of sites to EEG/Dis/LTU from WHP DATA
site_data = {}
for row in data_sheet.iter_rows(min_row=4, values_only=True):
    site = row[1]
    if site:
        eeg, dis, ltu = row[2], row[3], row[4]
        site_data[site.strip()] = (eeg, dis, ltu)

# Find the sites in WHP sheet at C7 and C8
target_rows = [7, 8]
for row_idx in target_rows:
    site_cell = whp_sheet.cell(row=row_idx, column=3)
    site_name = site_cell.value
    if site_name and site_name.strip() in site_data:
        eeg, dis, ltu = site_data[site_name.strip()]
        whp_sheet.cell(row=row_idx, column=5, value=eeg)
        whp_sheet.cell(row=row_idx, column=6, value=dis)
        whp_sheet.cell(row=row_idx, column=7, value=ltu)

wb.save(output_path)
print('Done')
