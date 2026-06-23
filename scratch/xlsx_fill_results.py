import openpyxl
from openpyxl.styles import PatternFill

# File paths
i_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_1/regression_gate/after_pass/core_51249/input.xlsx'
o_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_1/regression_gate/after_pass/core_51249/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(i_path)
ws = wb.active

# Custom fill color (RGB: 226, 239, 218)
fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

# Cells D1, D5, D9
d_cells = ['D1', 'D5', 'D9']

for d_cell in d_cells:
    row = int(d_cell[1:])
    b1 = ws[f'B{row}'].value
    b2 = ws[f'B{row+1}'].value  # B2 is the cell below B1
    # Logic
    if b1 == 'Description A' and (b2 is None or b2 == ''):
        result = 'Single A'
    elif b1 == 'Description B' and (b2 is None or b2 == ''):
        result = 'Single B'
    elif b1 == 'Description A' and b2 == 'Description B':
        result = 'Multiple'
    else:
        result = ''  # leave blank if not matched
    ws[d_cell].value = result
    ws[d_cell].fill = fill

wb.save(o_path)
