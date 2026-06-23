import openpyxl

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/after_pass/core_18935/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/after_pass/core_18935/output.xlsx'

wb = openpyxl.load_workbook(input_file)
sheet = wb['Sheet1']

# Extract the Data Table
# Data Table starts at row 4 and ends at row 10
# Header at row 4: ('Work', 'Material', 'Category 1', 'Category 2', 'Category 3', None, None)
data_table = []
for row in sheet.iter_rows(min_row=5, max_row=10, max_col=5, values_only=True):
    data_table.append(row)

# Map category type to its column index
category_col_map = {'Category 1': 2, 'Category 2': 3, 'Category 3': 4}

# Process Report Format rows (row 17-22)
for idx, report_row in enumerate(sheet.iter_rows(min_row=17, max_row=22, max_col=4, values_only=True)):
    work_criteria = report_row[0]
    category_type = report_row[1]
    material = report_row[2]
    # Lookup
    value = None
    col_idx = category_col_map.get(category_type)
    if col_idx:
        for data_row in data_table:
            if data_row[0] == work_criteria and data_row[1] == material:
                value = data_row[col_idx - 1]  # openpyxl uses 0-based index
                break
    # Write result to cell D17..D22
    sheet[f'D{17+idx}'] = value

wb.save(output_file)
