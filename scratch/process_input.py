import openpyxl

# Load workbook and sheets
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_7902_tc1/input.xlsx')
group_sheet = wb['Grouping']
form_sheet = wb['Formula Required']

# Read Grouping data into a list of rows
# Grouping!A2:I7
reference_data = []
for row in group_sheet.iter_rows(min_row=2, max_row=7, min_col=1, max_col=9, values_only=True):
    reference_data.append(row)

# D3:J6 (rows 3 to 6, columns 4 to 10)
for row_idx in range(3, 7):      # rows 3..6
    lookup_value = form_sheet.cell(row=row_idx, column=3).value  # C column for this row
    for col_idx in range(4, 11):  # cols D=4 .. J=10
        cell = form_sheet.cell(row=row_idx, column=col_idx)
        # If the value matches lookup_value, return it
        if cell.value == lookup_value:
            result = cell.value
        else:
            # Emulate VLOOKUP(C#, Grouping!A2:I7, X, FALSE)
            found = None
            for ref_row in reference_data:
                if ref_row[0] == lookup_value:
                    # The vlookup col index: leftmost is 1, so col_idx=4 (D) -> col 3 in VLOOKUP (C in Grouping)
                    found = ref_row[col_idx-2]
                    break
            result = found
        cell.value = result
# Save
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_7902_tc1/output.xlsx')
