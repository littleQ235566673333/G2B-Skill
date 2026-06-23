import openpyxl
import re

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_40892_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_40892_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Read the list of colors from column D, skipping header
color_list = []
for row in ws.iter_rows(min_row=3, min_col=4, max_col=4):
    val = row[0].value
    if val:
        color_list.append(str(val).strip().lower())

def extract_color(text, color_list):
    if not isinstance(text, str):
        return ''
    text_l = text.lower()
    for color in color_list:
        # Look for whole color words (e.g. 'red', not 'rodeo')
        if re.search(r'\b' + re.escape(color) + r'\b', text_l):
            # Return with original capitalization
            return color.title()
    return ''

# Fill B2:B17 based on matching color in A2:A17
for row in range(2, 18):
    cell_text = ws.cell(row=row, column=1).value
    color_found = extract_color(cell_text, color_list)
    ws.cell(row=row, column=2).value = color_found

wb.save(output_file)
print('Done')
