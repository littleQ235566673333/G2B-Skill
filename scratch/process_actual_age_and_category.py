import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/before_pass/core_32337/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/before_pass/core_32337/output.xlsx'

wb = openpyxl.load_workbook(in_path)
sheet = wb.active

col_dob = 3  # 'DATE OF BIRTH' is column C
col_expected_result = 5  # 'Expected Result' is column E
col_age_year = 15  # 'age (year)' is O
col_actual_age = col_age_year + 1  # P
col_category = 9  # 'CATEGORY' is I
col_year_age = 8  # 'YEAR (age)' is H

header_row = 2
start_row = 3
end_row = 15

actual_age_header = 'Actual Age'
sheet.cell(row=header_row, column=col_actual_age, value=actual_age_header)

# Safely copy fill details
src_fill = sheet.cell(row=header_row, column=col_age_year).fill
if isinstance(src_fill, PatternFill):
    new_fill = PatternFill(
        fill_type=src_fill.fill_type, fgColor=src_fill.fgColor.rgb, bgColor=src_fill.bgColor.rgb
    )
else:
    new_fill = PatternFill(fill_type=None)

for row in range(start_row, end_row+1):
    formula_age = f'=INT(YEARFRAC(C{row},$B$1))'
    cell = sheet.cell(row=row, column=col_actual_age, value=formula_age)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = new_fill

header_cell = sheet.cell(row=header_row, column=col_actual_age)
header_cell.alignment = Alignment(horizontal='center', vertical='top')
header_cell.fill = new_fill
header_cell.font = Font(bold=True)

for row in range(start_row, end_row+1):
    formula_cat = f'=VLOOKUP(P{row},$H$3:$I$15,2,TRUE)'
    sheet.cell(row=row, column=col_expected_result, value=formula_cat)

wb.save(out_path)
