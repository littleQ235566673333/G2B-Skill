from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/before_fix/core_50521/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/before_fix/core_50521/output.xlsx'

wb = load_workbook(input_path)
sheet = wb.active
output_col = 14  # Column N

for row_idx in range(4, 7):  # Rows 4, 5, 6
    row_vals = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(2, sheet.max_column+1)]
    first_idx = None
    second_idx = None
    for idx, val in enumerate(row_vals):
        if isinstance(val, (int, float)):
            if first_idx is None:
                first_idx = idx
            elif second_idx is None:
                second_idx = idx
                break
    result = None
    if first_idx is not None and second_idx is not None:
        first_val = row_vals[first_idx]
        if first_val > 1:
            result = 1
        else:
            result = second_idx - first_idx
    sheet.cell(row=row_idx, column=output_col).value = result

wb.save(output_path)
