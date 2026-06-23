import openpyxl
from openpyxl.styles import PatternFill

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/after_pass/core_51249/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/after_pass/core_51249/output.xlsx"

# Define fill color
fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

# Load workbook
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Target rows (from instruction)
target_rows = [1, 5, 9]

for row in target_rows:
    b1 = ws[f"B{row}"].value
    b2 = ws[f"B{row+1}"].value
    c = ws[f"C{row}"].value
    if c == "Result:":
        # Determine result
        if b1 == "Description A" and (b2 is None or b2 == ""):
            result = "Single A"
        elif b1 == "Description B" and (b2 is None or b2 == ""):
            result = "Single B"
        elif b1 == "Description A" and b2 == "Description B":
            result = "Multiple"
        else:
            result = ""
        ws[f"D{row}"].value = result
        ws[f"D{row}"].fill = fill

wb.save(output_path)
