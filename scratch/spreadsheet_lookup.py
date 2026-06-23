import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/after_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/after_pass/core_18935/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

data_table = []
data_header_row = 4  # Row index in Excel (1-based)
data_first_row = 5  # First data row (1-based)

# Identify header and load data table rows
header = [ws.cell(row=data_header_row, column=col).value for col in range(1, 6)]
for row in range(data_first_row, 11):  # rows 5-10
    work = ws.cell(row=row, column=1).value
    material = ws.cell(row=row, column=2).value
    categories = [ws.cell(row=row, column=c).value for c in range(3, 6)]
    data_table.append({'Work': work, 'Material': material,
                      'Category 1': categories[0],
                      'Category 2': categories[1],
                      'Category 3': categories[2]})

# Helper: category type to column mapping
category_col_map = {'Category 1': 'Category 1',
                   'Category 2': 'Category 2',
                   'Category 3': 'Category 3'}

for i, row in enumerate(range(17, 23)):  # rows 17-22
    work = ws.cell(row=row, column=1).value
    category_type = ws.cell(row=row, column=2).value
    material = ws.cell(row=row, column=3).value

    cat_col = category_col_map.get(category_type)

    # Find matching row in data_table
    match = next((r for r in data_table if r['Work'] == work and r['Material'] == material), None)
    value = match[cat_col] if match and (cat_col) in match else None
    ws.cell(row=row, column=4).value = value

wb.save(output_path)
