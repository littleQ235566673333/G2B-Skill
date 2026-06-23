import openpyxl
from datetime import datetime, timedelta

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/before_pass/core_41589/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Contact List']

# Read values from H4 and I4
last_contact_raw = ws['H4'].value  # Expected as date or string
response_raw = ws['I4'].value

# Parse date
if isinstance(last_contact_raw, datetime):
    last_contact_date = last_contact_raw
elif isinstance(last_contact_raw, str):
    try:
        last_contact_date = datetime.strptime(last_contact_raw, '%Y-%m-%d')
    except ValueError:
        # Try alternative format
        try:
            last_contact_date = datetime.strptime(last_contact_raw, '%m/%d/%Y')
        except ValueError:
            last_contact_date = None
else:
    last_contact_date = None

# Check yes/no
response = response_raw.strip().upper() if isinstance(response_raw, str) else ''

# Logic
result = 'NO ACTION'  # Default
if last_contact_date and response == 'YES':
    days_since = (datetime.now() - last_contact_date).days
    if days_since <= 30:
        result = 'HOLD'
    else:
        result = 'TOUCH BASE'

# Write result to J4
ws['J4'].value = result

# Save workbook
wb.save(output_path)
