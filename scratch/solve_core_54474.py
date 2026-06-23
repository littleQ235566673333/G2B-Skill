from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_4/regression_gate/before_pass/core_54474/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_4/regression_gate/before_pass/core_54474/output.xlsx'

wb = load_workbook(input_path)
ws_whp = wb['WHP']
ws_data = wb['WHP DATA']

lookup = {}
current_headers = None
for row in range(1, ws_data.max_row + 1):
    site_or_label = ws_data.cell(row=row, column=2).value
    c_val = ws_data.cell(row=row, column=3).value
    d_val = ws_data.cell(row=row, column=4).value
    e_val = ws_data.cell(row=row, column=5).value

    if c_val == 'EEG' and d_val == 'Dis' and e_val == 'LTU':
        current_headers = {'EEG': 3, 'Dis': 4, 'LTU': 5}
        continue

    if current_headers and site_or_label not in (None, 'Total'):
        lookup[str(site_or_label).strip()] = {
            'EEG': c_val,
            'Dis': d_val,
            'LTU': e_val,
        }

for row in (7, 8):
    site = ws_whp.cell(row=row, column=3).value
    values = lookup.get(str(site).strip(), {}) if site is not None else {}
    ws_whp.cell(row=row, column=5).value = values.get('EEG')
    ws_whp.cell(row=row, column=6).value = values.get('Dis')
    ws_whp.cell(row=row, column=7).value = values.get('LTU')

wb.save(output_path)

check_wb = load_workbook(output_path)
check_ws = check_wb['WHP']
for r in (7, 8):
    print(r, [check_ws.cell(row=r, column=c).value for c in range(5, 8)])
