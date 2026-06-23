import openpyxl
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/group_387-16/r0/evolve_387-16/input.xlsx')
ws = wb['Sheet1']
data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=18, min_col=1, max_col=4)]

values = [row[0] for row in data[2:]]
binaries = [row[1] for row in data[2:]]
result_values = [row[3] for row in data[4:] if row[3] is not None]

removed_indices = set()
for r_value in result_values:
    for i, v in enumerate(values):
        if i in removed_indices:
            continue
        if v == r_value:
            removed_indices.add(i)
            break
new_values = [v for idx, v in enumerate(values) if idx not in removed_indices]
new_binaries = [b for idx, b in enumerate(binaries) if idx not in removed_indices]

while len(new_values) < 16:
    new_values.append("")
while len(new_binaries) < 16:
    new_binaries.append("")

solver_result = sum([v for v in new_values if isinstance(v, int)])
target_value = data[1][3] if isinstance(data[1][3], int) else 0
ws['C2'] = 'Solver result:'
ws['D2'] = solver_result
ws['C3'] = 'Difference:'
ws['D3'] = solver_result - target_value
for i in range(16):
    ws.cell(row=3+i, column=1, value=new_values[i])
    ws.cell(row=3+i, column=2, value=new_binaries[i])
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/group_387-16/r0/evolve_387-16/output.xlsx')
