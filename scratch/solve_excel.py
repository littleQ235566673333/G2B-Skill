from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4_v6/train/iter_3/regression_gate/after_fix/core_57033/input.xlsx"
output_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4_v6/train/iter_3/regression_gate/after_fix/core_57033/output.xlsx"

wb = load_workbook(input_path)
ws = wb["Sheet4"]
source = wb["CBtrans"]

# Build a set of 3-way match keys from source sheet: company, account, xchar
source_keys = set()
for row in source.iter_rows(min_row=2, values_only=True):
    company = row[0]
    account = row[7]
    xchar = row[10]
    if company is None and account is None and xchar is None:
        continue
    source_keys.add((company, account, xchar))

pink_fill = PatternFill(fill_type="solid", fgColor="FF66CC")

for r in range(2, 8):
    company = ws[f"B{r}"].value
    account = ws[f"I{r}"].value
    xchar = ws[f"E{r}"].value
    result = "Match" if (company, account, xchar) in source_keys else "-"
    cell = ws[f"K{r}"]
    cell.value = result
    cell.fill = pink_fill

wb.save(output_path)
print('Saved', output_path)
for r in range(2, 8):
    print(r, ws[f'K{r}'].value)
