import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_39432/r0/evolve_39432/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_39432/r0/evolve_39432/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

def get_fifo_unit_cost(row):
    beginning_qty = row[1]
    beginning_cost = row[2]
    inventory_on_hand = row[11]

    fifo_sources = []
    if beginning_qty and beginning_cost:
        fifo_sources.append([beginning_qty, beginning_cost])

    # Orders in D,F,H,J and costs in E,G,I,K (1-based: 4/5, 6/7, 8/9, 10/11)
    for unit_idx, cost_idx in zip([3,5,7,9], [4,6,8,10]):
        qty = row[unit_idx]
        cost = row[cost_idx]
        if qty and cost:
            fifo_sources.append([qty, cost])

    if inventory_on_hand is None or inventory_on_hand == 0 or not fifo_sources:
        return 0

    needed = inventory_on_hand
    total_cost = 0
    total_gathered = 0
    for qty, cost in fifo_sources:
        use = min(qty, needed)
        total_cost += use * cost
        total_gathered += use
        needed -= use
        if needed == 0:
            break
    if total_gathered == 0:
        return 0
    return round(total_cost / total_gathered, 2)

for row_idx in range(2,6):
    row = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1,14)]
    unit_cost = get_fifo_unit_cost(row)
    ws.cell(row=row_idx, column=13).value = unit_cost

wb.save(output_path)
