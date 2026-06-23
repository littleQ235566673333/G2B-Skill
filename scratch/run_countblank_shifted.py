import openpyxl
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/before_pass/core_41969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/before_pass/core_41969/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Place the formulas in A6, B6, C6 for the shifting 3-column window
for i in range(3):
    col_start = 1 + 3 * i  # A=1, D=4, G=7
    col_end = col_start + 2
    cell_range = f'{get_column_letter(col_start)}3:{get_column_letter(col_end)}3'
    formula = f'=COUNTBLANK({cell_range})'
    ws.cell(row=6, column=1+i).value = formula  # A6, B6, C6

wb.save(output_path)
