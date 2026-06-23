from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_4/regression_gate/before_pass/core_54474/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4/train/iter_4/regression_gate/before_pass/core_54474/output.xlsx'

wb = load_workbook(input_path)
print(wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('SHEET', name)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=True):
        print(row)
    print('max_row', ws.max_row, 'max_col', ws.max_column)
