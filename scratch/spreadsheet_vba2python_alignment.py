import openpyxl
from openpyxl.utils import get_column_letter

# File paths
i_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/group_130-9/r1/evolve_130-9/input.xlsx"
o_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/group_130-9/r1/evolve_130-9/output.xlsx"

# Load workbook and sheets
wb = openpyxl.load_workbook(i_path)
ws_src = wb["cdnr"]
ws_dst = wb["b2b, sez, de"]

# Get headers for src and dst
src_headers = [cell.value for cell in ws_src[1]]
dst_headers = [cell.value for cell in ws_dst[1]]

# Map destination header to column index
header_to_dst_col = {header: idx+1 for idx, header in enumerate(dst_headers)}
header_to_src_col = {header: idx+1 for idx, header in enumerate(src_headers)}

# Find first empty row in dst, starting from row 7
row = 7
while any(ws_dst.cell(row=row, column=col).value for col in range(1, ws_dst.max_column+1)):
    row += 1

first_empty_row = row

# Columns to set special values
col_H = header_to_dst_col.get('H')    # Set to blank
col_I = header_to_dst_col.get('I')    # Set to 'Credit Note'
col_L = header_to_dst_col.get('L')    # Amount negative
col_M = header_to_dst_col.get('M')    # Amount negative

# For number formats
numfmt_L = ws_dst.cell(row=2, column=col_L).number_format if col_L else None
numfmt_M = ws_dst.cell(row=2, column=col_M).number_format if col_M else None

# Append rows from src to dst, aligning columns
for src_row in ws_src.iter_rows(min_row=2, max_row=ws_src.max_row, values_only=False):
    dst_row_idx = first_empty_row
    for header in dst_headers:
        col_idx = header_to_dst_col[header]
        src_col_idx = header_to_src_col.get(header)
        cell_dst = ws_dst.cell(row=dst_row_idx, column=col_idx)
        # Special columns
        if col_idx == col_H:
            cell_dst.value = ""
        elif col_idx == col_I:
            cell_dst.value = "Credit Note"
        elif col_idx == col_L:
            if src_col_idx:
                val = src_row[src_col_idx-1].value
                cell_dst.value = -val if isinstance(val, (int, float)) and val is not None else val
                if numfmt_L:
                    cell_dst.number_format = numfmt_L
            else:
                cell_dst.value = None
        elif col_idx == col_M:
            if src_col_idx:
                val = src_row[src_col_idx-1].value
                cell_dst.value = -val if isinstance(val, (int, float)) and val is not None else val
                if numfmt_M:
                    cell_dst.number_format = numfmt_M
            else:
                cell_dst.value = None
        else:
            if src_col_idx:
                cell_dst.value = src_row[src_col_idx-1].value
    first_empty_row += 1

# Save to output
wb.save(o_path)
