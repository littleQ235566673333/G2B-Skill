import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_1/regression_gate/before_pass/core_32337/input.xlsx')
ws = wb.active

header_row = 2
headers = []
for i, cell in enumerate(ws[header_row], 1):
    headers.append((i, cell.value))
print(headers)
