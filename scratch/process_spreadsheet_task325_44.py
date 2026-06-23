import openpyxl

input_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/325-44/input.xlsx'
output_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/325-44/output.xlsx'

wb = openpyxl.load_workbook(input_path)
input_ws = wb['Input'] if 'Input' in wb.sheetnames else wb.worksheets[0]
output_ws = wb['Output']

def read_data(ws, max_row, max_col):
    data = []
    for i in range(1, max_row+1):
        row = []
        empty_row = True
        for j in range(1, max_col+1):
            val = ws.cell(row=i, column=j).value
            row.append(val)
            if val is not None and val != '':
                empty_row = False
        if empty_row:
            break
        data.append(row)
    return data

# Read both tabs for format
input_data = read_data(input_ws, 20, 7)
output_format = read_data(output_ws, 10, 7)

# Locate filter column in input header
header = input_data[0]
filter_col_idx = None
for idx, col in enumerate(header):
    if 'filter' in str(col).lower():
        filter_col_idx = idx
        break

# Map output headers
output_header = output_format[0]
output_col_indices = {col: i for i, col in enumerate(output_header) if col is not None}

def parse_filter_string(filter_str):
    filt = {}
    if not filter_str:
        return filt
    for part in str(filter_str).split(','):
        if '=' in part:
            k, v = part.split('=', 1)
            k, v = k.strip(), v.strip()
            try:
                if v.isdigit():
                    v = str(int(v))  # drop leading zeros
            except:
                pass
            filt[k] = v
    return filt

for inrow_idx, row in enumerate(input_data[1:]):
    base = [None]*len(output_header)
    for colname, idx in output_col_indices.items():
        if colname in header:
            base[idx] = row[header.index(colname)]
    filter_val = row[filter_col_idx] if filter_col_idx is not None else None
    parsed = parse_filter_string(filter_val)
    for k, v in parsed.items():
        if k in output_col_indices:
            base[output_col_indices[k]] = v
    for j, cell_val in enumerate(base):
        output_ws.cell(row=inrow_idx+2, column=j+1, value=cell_val)

wb.save(output_path)
