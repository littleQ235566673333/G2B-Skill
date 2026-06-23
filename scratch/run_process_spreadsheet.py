import openpyxl

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_44389/r2/evolve_44389/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_44389/r2/evolve_44389/output.xlsx"

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# We'll process rows 2-7 (Excel indices), columns A-O (1-15), put output to P2:P7 (16)
for row in range(2, 8):  # 8 is exclusive
    values = []
    for col in range(1, 16):  # columns 1 (A) - 15 (O)
        val = ws.cell(row=row, column=col).value
        values.append(val)

    # Filter values > 0
    positive = [(i, v) for i, v in enumerate(values) if (isinstance(v, (int, float)) and v > 0)]
    if not positive:
        label = ''
    else:
        min_val = min(v for i, v in positive)
        # Get all columns where value matches min_val
        cols = [i for i, v in positive if v == min_val]
        # Get headers (assume header is at row 1)
        headers = [ws.cell(row=1, column=col+1).value for col in cols]
        label = ','.join(str(h) for h in headers)
    # Put result in column 16 (P)
    ws.cell(row=row, column=16).value = label

wb.save(output_path)
