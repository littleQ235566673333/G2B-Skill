import openpyxl

# Load the workbook and necessary sheets
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_267-21_tc1/input.xlsx')
ws_merging = wb['merging']
ws_SH1 = wb['SH1']
ws_RP1 = wb['RP1']

# Build ID -> QTY mapping for SH1
id_to_qty_SH1 = {}
for row in ws_SH1.iter_rows(min_row=2, values_only=True):
    id_, qty = row[0], row[1]
    id_to_qty_SH1[id_] = qty

# Build ID -> QTY mapping for RP1
id_to_qty_RP1 = {}
for row in ws_RP1.iter_rows(min_row=2, values_only=True):
    id_, qty = row[0], row[1]
    id_to_qty_RP1[id_] = qty

# For merging sheet, fill C2:C11 from SH1 and D2:D11 from RP1 using mapping, else '-'
for row_idx in range(2, 12):
    id_cell = ws_merging.cell(row=row_idx, column=2)
    target_C = ws_merging.cell(row=row_idx, column=3)
    target_D = ws_merging.cell(row=row_idx, column=4)
    id_ = id_cell.value
    val_c = id_to_qty_SH1.get(id_, '-')
    val_d = id_to_qty_RP1.get(id_, '-')
    if val_c is None or val_c == '':
        val_c = '-'
    if val_d is None or val_d == '':
        val_d = '-'
    target_C.value = val_c
    target_D.value = val_d

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/eval/eval_267-21_tc1/output.xlsx')
