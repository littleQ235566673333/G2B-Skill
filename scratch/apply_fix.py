from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4_v6/train/iter_2/regression_gate/after_fix/core_57033/input.xlsx"
output_path = r"results/runs/g2b-skill-spreadsheet_gpt-5.4_v6/train/iter_2/regression_gate/after_fix/core_57033/output.xlsx"

wb = load_workbook(input_path)
ws = wb["Sheet4"]
md = wb["CBtrans"]

# Build a 3-way match set using company, account, and xchar from CBtrans.
lookup = set()
for row in range(2, md.max_row + 1):
    company = md[f"A{row}"].value
    account = md[f"H{row}"].value
    xchar = md[f"K{row}"].value
    if company is None and account is None and xchar is None:
        continue
    lookup.add((company, account, xchar))

fill = PatternFill(fill_type="solid", start_color="FF66CC", end_color="FF66CC")

def proper_case(value):
    if isinstance(value, str):
        return value.title()
    return value

for row in range(2, 8):
    company = ws[f"B{row}"].value
    account = ws[f"I{row}"].value
    xchar = ws[f"E{row}"].value
    result = "Match" if (company, account, xchar) in lookup else "-"
    cell = ws[f"K{row}"]
    cell.value = proper_case(result)
    cell.fill = fill

wb.save(output_path)
