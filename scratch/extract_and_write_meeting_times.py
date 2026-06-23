import openpyxl
from openpyxl.utils import get_column_letter

def time_from_header(header_str):
    h, m = map(int, header_str.split(':'))
    return h + m / 60

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/group_49667/r2/evolve_49667/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/group_49667/r2/evolve_49667/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get time headers and mapping from col index to time
headers = [cell.value for cell in ws[1]]
time_col_indices = list(range(6, len(headers)+1))  # Excel 1-based, col 6 is F
col_idx_to_time = {idx: time_from_header(headers[idx-1]) for idx in time_col_indices}

# For each row 2–16
for row in range(2, 17):
    values = [ws.cell(row=row, column=col).value for col in time_col_indices]
    blocks = []
    in_block = False
    start_idx = None

    for offset, val in enumerate(values):
        idx = time_col_indices[offset]
        if val == 'm':
            if not in_block:
                in_block = True
                start_idx = idx
        else:
            if in_block:
                in_block = False
                blocks.append((start_idx, time_col_indices[offset-1]))
    if in_block:
        blocks.append((start_idx, time_col_indices[-1]))

    # Output up to 2 blocks
    for i in range(2):
        if i < len(blocks):
            start_col, finish_col = blocks[i]
            ws.cell(row=row, column=2+i*2).value = col_idx_to_time[start_col]
            ws.cell(row=row, column=3+i*2).value = col_idx_to_time[finish_col]+0.25  # 15 min = 0.25h; end is exclusive
        else:
            ws.cell(row=row, column=2+i*2).value = None
            ws.cell(row=row, column=3+i*2).value = None

wb.save(output_path)
