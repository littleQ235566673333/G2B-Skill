import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/regression_gate/before_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/regression_gate/before_pass/core_18935/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Data Table structure (rows 5-10, columns A-E/1-5; headers in row 4)
data_headers = [ws.cell(4, c).value for c in range(1, 6)]
data = []
for r in range(5, 11):  # rows 5-10 (inclusive)
    row = [ws.cell(r, c).value for c in range(1, 6)]
    data.append(dict(zip(data_headers, row)))

# For each report row (17-22), look up
for rr in range(17, 23):
    work = ws.cell(rr, 1).value
    category_type = ws.cell(rr, 2).value  # 'Category 1', etc
    material = ws.cell(rr, 3).value
    val = None
    for d in data:
        if d['Work'] == work and d['Material'] == material:
            val = d.get(category_type)
            break
    ws.cell(rr, 4).value = val  # write value in col D

wb.save(output_path)
