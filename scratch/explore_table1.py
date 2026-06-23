from openpyxl import load_workbook

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_1/regression_gate/after_pass/core_9726/input.xlsx')
ws = wb.active
for i in range(2, 10):
    print(f'Row {i}: A={ws[f"A{i}"].value} B={ws[f"B{i}"].value} C={ws[f"C{i}"].value}')
