from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_3/regression_gate/after_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_3/regression_gate/after_pass/core_9726/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Gather Table 1: Columns A (0), B (1), C (2)
visit_rows = []
for row in ws.iter_rows(min_row=2, max_row=7, values_only=True):
    sid, _, visit_date = row[0], row[1], row[2]
    if sid and visit_date:
        visit_rows.append({'student_id': sid, 'visit_date': visit_date})

# Gather Table 2: Start at row 2, H=8 I=9 J=10
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=9, values_only=True), start=2):
    student_id, submission_date = row[0], row[1]
    if student_id and submission_date:
        # Find the latest visit date before submission date
        candidate_dates = [v['visit_date'] for v in visit_rows if v['student_id'] == student_id and v['visit_date'] < submission_date]
        if candidate_dates:
            latest_visit = max(candidate_dates)
            ws.cell(row=i, column=10, value=latest_visit)

wb.save(output_path)
