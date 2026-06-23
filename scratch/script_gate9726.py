import openpyxl
import pandas as pd
from datetime import datetime

# Load workbook and sheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/after_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/after_pass/core_9726/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read Table 1
visit_data = []
row = 2
while ws[f'A{row}'].value is not None:
    student = ws[f'A{row}'].value
    visit_date = ws[f'C{row}'].value
    visit_data.append({'student': student, 'visit_date': visit_date})
    row += 1
visit_df = pd.DataFrame(visit_data)
visit_df['visit_date'] = pd.to_datetime(visit_df['visit_date'])

# Read Table 2
subm_data = []
row = 2
while ws[f'H{row}'].value is not None:
    student = ws[f'H{row}'].value
    subm_date = ws[f'I{row}'].value
    subm_data.append({'student': student, 'submission_date': subm_date, 'row': row})
    row += 1
subm_df = pd.DataFrame(subm_data)
subm_df['submission_date'] = pd.to_datetime(subm_df['submission_date'])

# For each submission, find the last visit date before the submission
results = []
for s in subm_df.itertuples():
    vdates = visit_df[(visit_df['student'] == s.student) & (visit_df['visit_date'] < s.submission_date)]
    if not vdates.empty:
        last_visit = vdates['visit_date'].max()
        results.append(last_visit)
    else:
        results.append(None)

# Write results in column J
for i, val in enumerate(results):
    ws[f'J{subm_df.iloc[i]["row"]}'] = val.strftime('%d-%b-%y') if pd.notnull(val) else ''

wb.save(output_path)
