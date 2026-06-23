import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/before_pass/core_51249/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/before_pass/core_51249/output.xlsx'

# Load workbook and select the first sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Define the cells to fill
cells_to_fill = [(1, 4), (5, 4), (9, 4)]  # D1, D5, D9
result_cells = [(1, 3), (5, 3), (9, 3)]   # C1, C5, C9

# Fill color: RGB 226,239,218
fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

for (row, col), (res_row, res_col) in zip(cells_to_fill, result_cells):
    # Check if cell in column C says 'Result:'
    if ws.cell(row=res_row, column=res_col).value == 'Result:':
        # B1/B2, B5/B6, B9/B10 for D1/D5/D9
        b1 = ws.cell(row=row, column=2).value
        b2 = ws.cell(row=row+1, column=2).value
        result = None
        if b1 == 'Description A' and (b2 is None or b2 == ''):
            result = 'Single A'
        elif b1 == 'Description B' and (b2 is None or b2 == ''):
            result = 'Single B'
        elif b1 == 'Description A' and b2 == 'Description B':
            result = 'Multiple'
        ws.cell(row=row, column=col).value = result
        if result:
            ws.cell(row=row, column=col).fill = fill

wb.save(output_path)
