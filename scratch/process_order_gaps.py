import openpyxl
from openpyxl.utils import get_column_letter

# File paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/before_fix/core_50521/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/before_fix/core_50521/output.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# The rows to process (N4:N6 => data in rows 4 to 6)
start_row = 4
end_row = 6
# The columns containing order values (assuming from B to M, since N is output)
first_data_col = 2  # Column B
last_data_col = 13  # Column M
output_col = 14     # Column N

def count_blanks_between_first_two_numbers(row):
    numbers = []
    for idx, cell in enumerate(row):
        if isinstance(cell, (int, float)) and cell != '':
            numbers.append(idx)
    if len(numbers) < 2:
        return None  # Not enough entries to compute
    # Check if first value is > 1
    first_value = row[numbers[0]]
    if isinstance(first_value, (int, float)) and first_value > 1:
        return 1
    # Count blanks between first and second numeric entries (inclusive)
    start = numbers[0]
    end = numbers[1]
    # The count should be (end - start)
    return end - start

for row_num in range(start_row, end_row+1):
    # Extract the data row (B:M)
    row_values = [ws.cell(row=row_num, column=col).value for col in range(first_data_col, last_data_col+1)]
    count = count_blanks_between_first_two_numbers(row_values)
    # Write the result in column N
    ws.cell(row=row_num, column=output_col).value = count

# Save to output path
wb.save(output_path)
