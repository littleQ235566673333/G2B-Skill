import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_1/group_57033/r0/evolve_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_1/group_57033/r0/evolve_57033/output.xlsx'

wb = openpyxl.load_workbook(input_path)
s4 = wb['Sheet4']
cb = wb['CBtrans']

def get_column_map(ws):
    return {str(cell.value).strip().lower(): idx+1 for idx, cell in enumerate(ws[1]) if cell.value}

cb_map = get_column_map(cb)
s4_map = get_column_map(s4)

# Extract all CBtrans comparison rows
cb_rows = [
    [
        cb.cell(row=r, column=cb_map['company']).value,
        cb.cell(row=r, column=cb_map['account']).value.strip() if isinstance(cb.cell(row=r, column=cb_map['account']).value, str) else cb.cell(row=r, column=cb_map['account']).value,
        cb.cell(row=r, column=cb_map['xchar']).value
    ]
    for r in range(2, cb.max_row+1)
    if cb.cell(row=r, column=cb_map['company']).value is not None
]

fill = PatternFill(start_color='FF66CC', end_color='FF66CC', fill_type='solid')

for row in range(2, 8):
    company = s4.cell(row=row, column=s4_map['company']).value
    account = s4.cell(row=row, column=s4_map['account']).value
    if isinstance(account, str): account = account.strip()
    xchar = s4.cell(row=row, column=s4_map['xchar']).value

    match = any([
        (company == cb_row[0] and account == cb_row[1] and xchar == cb_row[2])
        for cb_row in cb_rows
    ])

    val = 'Match' if match else '-'
    out_cell = s4.cell(row=row, column=11)
    out_cell.value = val
    out_cell.fill = fill

wb.save(output_path)
