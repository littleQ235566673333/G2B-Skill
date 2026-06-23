import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Load workbook
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_1/regression_gate/before_pass/core_32337/input.xlsx')
ws = wb.active

header_row = 2
# From inspection
DOB_col = 3
REPORT_DATE_CELL = 'B1'
AGE_YEAR_COL = 15
CATEGORY_COL = 9
EXPECTED_RESULT_COL = 5
ACTUAL_AGE_COL = AGE_YEAR_COL + 1

# Insert column for Actual Age after age (year)
ws.insert_cols(ACTUAL_AGE_COL)
ws.cell(row=header_row, column=ACTUAL_AGE_COL).value = 'Actual Age'

# Safely copy fill from previous column (age (year))
age_fill_src = ws.cell(row=header_row, column=AGE_YEAR_COL).fill
if isinstance(age_fill_src, PatternFill):
    age_fill = PatternFill(fill_type=age_fill_src.fill_type, fgColor=age_fill_src.fgColor.rgb, bgColor=age_fill_src.bgColor.rgb)
else:
    age_fill = PatternFill(fill_type=None)
ws.cell(row=header_row, column=ACTUAL_AGE_COL).fill = age_fill
# Top align header and bold
ws.cell(row=header_row, column=ACTUAL_AGE_COL).alignment = Alignment(horizontal='center', vertical='top')
ws.cell(row=header_row, column=ACTUAL_AGE_COL).font = Font(bold=True)

for row in range(3, 16):
    dob_cell = f'{get_column_letter(DOB_col)}{row}'
    age_formula_cell = ws.cell(row=row, column=ACTUAL_AGE_COL)
    age_formula_cell.value = f'=IF(${REPORT_DATE_CELL}>0,INT(YEARFRAC({dob_cell},${REPORT_DATE_CELL})), "")'
    age_formula_cell.alignment = Alignment(horizontal='center', vertical='center')
    age_formula_cell.font = Font(bold=True)
    age_formula_cell.number_format = '0'
    age_formula_cell.fill = age_fill

# Adjust CATEGORY_COL and EXPECTED_RESULT_COL because we inserted one column before them
if CATEGORY_COL >= ACTUAL_AGE_COL:
    CATEGORY_COL += 1
if EXPECTED_RESULT_COL >= ACTUAL_AGE_COL:
    EXPECTED_RESULT_COL += 1
if AGE_YEAR_COL >= ACTUAL_AGE_COL:
    AGE_YEAR_COL += 1

category_range = f'${get_column_letter(CATEGORY_COL)}$3:${get_column_letter(CATEGORY_COL)}$15'
age_range = f'${get_column_letter(AGE_YEAR_COL)}$3:${get_column_letter(AGE_YEAR_COL)}$15'

for row in range(3, 16):
    ws.cell(row=row, column=EXPECTED_RESULT_COL).value = f'=IFERROR(INDEX({category_range},MATCH({get_column_letter(AGE_YEAR_COL)}{row},{age_range},0)),"")'

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_1/regression_gate/before_pass/core_32337/output.xlsx')
print('Done')
