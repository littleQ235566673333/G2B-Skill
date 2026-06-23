import pandas as pd
from openpyxl import load_workbook

input_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/31915/input.xlsx'
output_path = 'results/base_trajectories/master_0_heldout_42/spreadsheet/gpt-4.1/31915/output.xlsx'

# Read input data from Sheet1
df = pd.read_excel(input_path, sheet_name='Sheet1')

# Use 'serial' as the unique key and sum over 'extent' (as there is no 'B' column like 'price')
serial_sums = df.groupby('serial')['extent'].sum()
seen = {}
out_col = []
for idx, row in df.iterrows():
    serial = row['serial']
    if serial not in seen:
        out_col.append(serial_sums[serial])
        seen[serial] = True
    else:
        out_col.append('')
# Ensure length is 10 for C2:C11
out_col = (out_col + ['']*10)[:10]

# Write to Sheet2 C2:C11
wb = load_workbook(input_path)
ws = wb['Sheet2']
for i, val in enumerate(out_col, start=2):
    ws[f'C{i}'] = val
wb.save(output_path)
