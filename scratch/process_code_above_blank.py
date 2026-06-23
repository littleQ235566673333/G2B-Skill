import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/after_fix/core_374-31/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/regression_gate/after_fix/core_374-31/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

def is_blank_row(row):
    return all((cell.value is None or str(cell.value).strip() == '') for cell in row)

rows_to_delete = set()
max_row = ws.max_row
for row_idx in range(max_row, 0, -1):  # bottom to top
    cell_value = ws.cell(row=row_idx, column=1).value
    if cell_value == 'Code':
        check_row = row_idx - 1
        while check_row > 0:
            row_cells = ws[check_row]
            if is_blank_row(row_cells):
                rows_to_delete.add(check_row)
                check_row -= 1
            else:
                break
# Delete rows from bottom to top
for row_num in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row_num, 1)
wb.save(output_path)
