import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/after_fix/core_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_7/regression_gate/after_fix/core_57033/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheet4 = wb['Sheet4']
cbtrans = wb['CBtrans']
cb_data = list(cbtrans.iter_rows(min_row=2, values_only=True))
# Headers
cb_headers = [c.value for c in cbtrans[1]]
s4_headers = [c.value for c in sheet4[1]]

# Get field indexes
s4_company_idx = s4_headers.index('Company')
s4_account_idx = s4_headers.index('account')
s4_xchar_idx = s4_headers.index('xchar')
cb_company_idx = cb_headers.index('company')
cb_account_idx = cb_headers.index('account')
cb_xchar_idx = cb_headers.index('xchar')

values = []
for row in sheet4.iter_rows(min_row=2, max_row=7, values_only=True):
    company = str(row[s4_company_idx]).strip().lower() if row[s4_company_idx] is not None else ''
    account = str(row[s4_account_idx]).strip().lower() if row[s4_account_idx] is not None else ''
    xchar = str(row[s4_xchar_idx]).strip().lower() if row[s4_xchar_idx] is not None else ''
    found = False
    for cb_row in cb_data:
        cb_company = str(cb_row[cb_company_idx]).strip().lower() if cb_row[cb_company_idx] is not None else ''
        cb_account = str(cb_row[cb_account_idx]).strip().lower() if cb_row[cb_account_idx] is not None else ''
        cb_xchar = str(cb_row[cb_xchar_idx]).strip().lower() if cb_row[cb_xchar_idx] is not None else ''
        if (company == cb_company) and (account == cb_account) and (xchar == cb_xchar):
            found = True
            break
    values.append('Match' if found else '-')

# Set casing, fill and write to Sheet4 K2:K7
fill = PatternFill(start_color='FF66CC', end_color='FF66CC', fill_type='solid')
for i, val in enumerate(values, start=2):
    cell = sheet4[f'K{i}']
    cell.value = val.capitalize()
    cell.fill = fill
wb.save(output_path)
