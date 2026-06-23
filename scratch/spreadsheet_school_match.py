import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/group_55427/r0/evolve_55427/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/group_55427/r0/evolve_55427/output.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_path)

# Access sheets
urn_sheet = wb['URN lookup']
compiled_sheet = wb['Compiled and located schools da']

# Build lookup dictionary: URN Sheet K (col 11) -> D (col 4)
urn_lookup = {}
for row in urn_sheet.iter_rows(min_row=2, max_row=1461, min_col=4, max_col=11):
    d_cell = row[0]   # Column D
    k_cell = row[7]   # Column K (col 11, so index 7 in 4:11 slice)
    urn_lookup[k_cell.value] = d_cell.value

# Iterate over Compiled sheet L (col 12)
for i, row in enumerate(compiled_sheet.iter_rows(min_row=2, max_row=1461)):
    l_val = row[11].value  # Column L (12th col)
    dfes_num = urn_lookup.get(l_val, None)
    compiled_sheet.cell(row=i+2, column=2, value=dfes_num)  # B2:B1461

# Save the result
wb.save(output_path)
