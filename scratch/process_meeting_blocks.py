import openpyxl
from datetime import datetime

def time_to_excel_number(timestr):
    dt = datetime.strptime(timestr, '%H:%M')
    return dt.hour + dt.minute/60

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/after_fix/core_49667/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_6/regression_gate/after_fix/core_49667/output.xlsx'

wb = openpyxl.load_workbook(input_fp)
ws = wb.active

# Read time header for columns F1:AT1
header_times = [ws.cell(row=1, column=col).value for col in range(6, 47)]
time_numbers = [time_to_excel_number(str(t)) if t else None for t in header_times]

for row in range(2, 17):
    line = [ws.cell(row=row, column=col).value for col in range(6,47)]
    meeting_blocks = []
    block_start = None
    for idx, v in enumerate(line):
        if v == 'm':
            if block_start is None:
                block_start = idx
        else:
            if block_start is not None:
                block_end = idx - 1
                meeting_blocks.append((block_start, block_end))
                block_start = None
    if block_start is not None:
        meeting_blocks.append((block_start, len(line)-1))

    # Up to two blocks
    for b in range(2):
        if b < len(meeting_blocks):
            start_idx, end_idx = meeting_blocks[b]
            ws.cell(row=row, column=2).value = header_times[start_idx] # Start time as in header
            ws.cell(row=row, column=3).value = time_numbers[start_idx] # Start time as number
            ws.cell(row=row, column=4).value = header_times[end_idx]   # Finish time as in header
            ws.cell(row=row, column=5).value = time_numbers[end_idx]   # Finish time as number
        else:
            ws.cell(row=row, column=2).value = None
            ws.cell(row=row, column=3).value = None
            ws.cell(row=row, column=4).value = None
            ws.cell(row=row, column=5).value = None

wb.save(output_fp)
