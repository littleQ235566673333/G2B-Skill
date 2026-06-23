import openpyxl
from datetime import datetime, timedelta

# File paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/after_pass/core_41589/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_10/regression_gate/after_pass/core_41589/output.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb["Contact List"]

# Read needed values
last_contact_cell = ws["H4"].value  # last contact date
response_cell = ws["I4"].value     # YES/NO response
result = "NO ACTION"                # Default

# Parse contact date
if last_contact_cell is not None and response_cell is not None:
    try:
        # Try parsing as a date
        if isinstance(last_contact_cell, datetime):
            last_contact_date = last_contact_cell
        else:
            last_contact_date = datetime.strptime(str(last_contact_cell), "%Y-%m-%d")

        days_diff = (datetime.today() - last_contact_date).days

        if (days_diff <= 30) and (str(response_cell).strip().upper() == "YES"):
            result = "HOLD"
        elif (days_diff > 30) and (str(response_cell).strip().upper() == "YES"):
            result = "TOUCH BASE"
        else:
            result = "NO ACTION"
    except Exception as e:
        # If date parsing fails, leave as 'NO ACTION'
        pass
else:
    result = "NO ACTION"

# Write result to J4
ws["J4"] = result

# Save the output
wb.save(output_path)
