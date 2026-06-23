import openpyxl
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_91-34/r1/evolve_91-34/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_9/group_91-34/r1/evolve_91-34/output.xlsx'
sheet_name = 'SwiftMD'
start_row, end_row = 3, 42       # Data is row 3 downward (headers in row 2)
header_row = 2
start_col, end_col = 2, 15  # Columns B:O (1-based)

wb = openpyxl.load_workbook(input_path)
ws = wb[sheet_name]

# Read headers from row 2
headers = [cell.value for cell in ws[header_row][start_col-1:end_col]]

last_idx = headers.index('Last Name')
first_idx = headers.index('First Name')
dob_idx = headers.index('Date Of Birth')
dup_idx = headers.index('Duplicate?')
rel_idx = headers.index('Relationship')

# Collect data rows within the B3:O42 block
data_rows = []
for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
    values = [cell.value for cell in row]
    data_rows.append(values)

# Group by (Last Name, First Name, DOB) where 'Duplicate?' == 'Yes'
groups = defaultdict(list)
for i, row in enumerate(data_rows):
    if (row[dup_idx] == 'Yes') and (row[last_idx] and row[first_idx] and row[dob_idx]):
        key = (row[last_idx], row[first_idx], row[dob_idx])
        groups[key].append(i)

# Mark rows to delete
delete_row_indices = set()
for key, idxs in groups.items():
    rels = [data_rows[i][rel_idx] for i in idxs]
    if all(r != 'Employee' for r in rels):
        # For non-Employee duplicates, delete only one (the first)
        delete_row_indices.add(idxs[0])
    # If any 'Employee', do not delete any

# Create new list with deleted rows removed
filtered_rows = [row for i, row in enumerate(data_rows) if i not in delete_row_indices]

# Ensure output fits B3:O42 (pad with blanks or trim as needed)
pad_len = (end_row - start_row + 1)
while len(filtered_rows) < pad_len:
    filtered_rows.append([''] * (end_col - start_col + 1))
filtered_rows = filtered_rows[:pad_len]

# Write back to ws, B3:O42
for r, row in enumerate(filtered_rows, start=start_row):
    for c, val in enumerate(row, start=start_col):
        ws.cell(row=r, column=c, value=val)

wb.save(output_path)
