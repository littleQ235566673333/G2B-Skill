import openpyxl
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/regression_gate/before_pass/core_41589/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Contact List']

today = datetime.today()
row = 4  # Target row
last_contact_cell = ws[f'H{row}']
yesno_cell = ws[f'I{row}']
output_cell = ws[f'J{row}']

# Get H4 value (last contact date)
last_contact_val = last_contact_cell.value
yesno_val = yesno_cell.value
try:
    last_contact_date = None
    if isinstance(last_contact_val, datetime):
        last_contact_date = last_contact_val
    elif isinstance(last_contact_val, str):
        # Try parsing as date string if not empty/None
        try:
            last_contact_date = datetime.strptime(last_contact_val, '%Y-%m-%d')
        except Exception:
            last_contact_date = None

    if yesno_val and str(yesno_val).strip().upper() == 'YES':
        if last_contact_date:
            if (today - last_contact_date).days <= 30:
                output = 'HOLD'
            else:
                output = 'TOUCH BASE'
        else:
            output = 'NO ACTION'
    else:
        output = 'NO ACTION'
except Exception:
    output = 'NO ACTION'

output_cell.value = output
wb.save(output_path)
