import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/after_fix/core_49667/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_8/regression_gate/after_fix/core_49667/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Time columns start at F (5), header row holds times in string (08:00, etc)
time_headers = [cell.value for cell in ws[1]][5:]
num_slots = len(time_headers)

# Helper: Convert 'HH:MM' string to Excel serial as number
# Assume meetings are all on the same date (Excel date 1 = 1899-12-31, but time is fraction)
def time_str_to_excel_serial(tstr):
    base_date = datetime(1899, 12, 31)
    t = datetime.strptime(tstr, "%H:%M")
    delta = timedelta(hours=t.hour, minutes=t.minute)
    return delta.total_seconds() / 86400  # fraction of day

for row_idx in range(2, 17):  # Excel rows 2-16
    row = ws[row_idx]
    blocks = []
    block_start = None
    block_end = None
    in_block = False
    for idx, col in enumerate(range(6, 6 + num_slots)):
        cell_val = ws.cell(row=row_idx, column=col).value
        if cell_val == 'm':
            if not in_block:
                block_start = idx  # slot index
                in_block = True
        else:
            if in_block:
                block_end = idx - 1  # end before current slot
                blocks.append((block_start, block_end))
                in_block = False
    if in_block:
        # Block ends at last slot
        block_end = num_slots - 1
        blocks.append((block_start, block_end))
    # Write up to two blocks
    for b in range(2):
        if b < len(blocks):
            start_idx, end_idx = blocks[b]
            start_time = time_headers[start_idx]
            end_time = time_headers[end_idx]
            # As Excel serials (numbers):
            start_serial = time_str_to_excel_serial(start_time)
            end_serial = time_str_to_excel_serial(end_time)
            ws.cell(row=row_idx, column=2 + b*2).value = start_serial
            ws.cell(row=row_idx, column=3 + b*2).value = end_serial
        else:
            ws.cell(row=row_idx, column=2 + b*2).value = None
            ws.cell(row=row_idx, column=3 + b*2).value = None

wb.save(output_path)
