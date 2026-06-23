import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_2/regression_gate/after_pass/core_32337/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get REPORT DATE from cell B1
report_date = ws['B1'].value

row_start = 3
row_end = 15
DOB_col = 3
Expected_col = 5
Age_col = 15  # O
CATEGORY_MAP_COL = 9  # I
YEAR_AGE_COL = 8      # H

# Build CATEGORY lookup: map each YEAR (age) to its CATEGORY
age_years_to_category = {}
for r in range(row_start, row_end+1):
    age_val = ws.cell(row=r, column=YEAR_AGE_COL).value
    category = ws.cell(row=r, column=CATEGORY_MAP_COL).value
    if age_val is not None and category is not None and isinstance(age_val, (int, float)):
        age_years_to_category[int(round(age_val))] = category

# Insert Actual Age column after 'age (year)'
actual_age_col = Age_col + 1
ws.insert_cols(actual_age_col)
ws.cell(row=2, column=actual_age_col).value = 'Actual Age'

# Inherit fill from previous column if available
prev_fill_obj = ws.cell(row=2, column=Age_col).fill
fill_props = {}
if isinstance(prev_fill_obj, PatternFill):
    fill_props = dict(patternType=prev_fill_obj.patternType,
                      fgColor=prev_fill_obj.fgColor.rgb if prev_fill_obj.fgColor else None,
                      bgColor=prev_fill_obj.bgColor.rgb if prev_fill_obj.bgColor else None)
    pattern_fill = PatternFill(**{k: v for k, v in fill_props.items() if v})
else:
    pattern_fill = None

ws.cell(row=2, column=actual_age_col).alignment = Alignment(horizontal='center', vertical='top')
if pattern_fill:
    ws.cell(row=2, column=actual_age_col).fill = pattern_fill

# Fill Actual Age and Expected Result
for r in range(row_start, row_end+1):
    dob = ws.cell(row=r, column=DOB_col).value
    if isinstance(dob, datetime):
        actual_years = report_date.year - dob.year - ((report_date.month, report_date.day) < (dob.month, dob.day))
        ws.cell(row=r, column=actual_age_col).value = actual_years
        ws.cell(row=r, column=actual_age_col).font = Font(bold=True)
        ws.cell(row=r, column=actual_age_col).alignment = Alignment(horizontal='center', vertical='center')
        if pattern_fill:
            ws.cell(row=r, column=actual_age_col).fill = pattern_fill
        cat = age_years_to_category.get(actual_years)
        ws.cell(row=r, column=Expected_col).value = cat
    else:
        ws.cell(row=r, column=actual_age_col).value = None
        ws.cell(row=r, column=Expected_col).value = None

wb.save(output_path)
