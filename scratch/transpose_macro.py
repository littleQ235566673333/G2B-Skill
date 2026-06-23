import openpyxl

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/group_147-48/r1/evolve_147-48/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_v6/train/iter_4/group_147-48/r1/evolve_147-48/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read data from column A, rows 1 to 6
values = [ws[f'A{row}'].value for row in range(1, 7)]

# Write values to row 1, columns C to H (3 to 8)
for col, value in enumerate(values, start=3):
    ws.cell(row=1, column=col).value = value

# Save output
wb.save(output_path)
