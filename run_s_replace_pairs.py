import openpyxl
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/group_146-49/r1/evolve_146-49/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/group_146-49/r1/evolve_146-49/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read all words from column A (assuming the dictionary is in column A)
words = []
for row in ws.iter_rows(min_col=1, max_col=1, min_row=1):
    val = row[0].value
    if val and isinstance(val, str):
        words.append(val.strip())

# Set of vowels
vowels = set('AEIOU')

def template_and_s_indices(word):
    t = []
    s_indices = []
    for i, c in enumerate(word):
        if c == 'S':
            t.append('_')
            s_indices.append(i)
        else:
            t.append(c)
    return ''.join(t), tuple(s_indices)

# Map (template, S positions) => [words]
by_template = defaultdict(list)
for w in words:
    tpl, s_idx = template_and_s_indices(w)
    by_template[(tpl, s_idx)].append(w)

def s_replace_matches(template_group):
    found_pairs = set()
    n = len(template_group)
    for i in range(n):
        for j in range(i+1, n):
            w1, w2 = template_group[i], template_group[j]
            # Indices where letters differ
            diff_indices = [k for k, c in enumerate(w1) if c != w2[k]]
            if not diff_indices:
                continue
            if all((w1[k] == 'S' and w2[k] in vowels) or (w2[k] == 'S' and w1[k] in vowels) for k in diff_indices):
                s_to_vowel = None
                consistent = True
                for k in diff_indices:
                    if w1[k] == 'S':
                        v = w2[k]
                    else:
                        v = w1[k]
                    if v not in vowels:
                        consistent = False
                        break
                    if s_to_vowel is None:
                        s_to_vowel = v
                    elif s_to_vowel != v:
                        consistent = False
                        break
                if consistent:
                    sorted_pair = tuple(sorted([w1, w2], key=lambda x: x.lower()))
                    found_pairs.add(sorted_pair)
    return found_pairs

all_pairs = set()
for group in by_template.values():
    matches = s_replace_matches(group)
    all_pairs.update(matches)

# Sort pairs as required
sorted_pairs = sorted(all_pairs, key=lambda pr: (pr[0].lower(), pr[1].lower()))

# Write results: columns G (7) and H (8), start at row 1 (G1, H1)
for r, (w1, w2) in enumerate(sorted_pairs, start=1):
    ws.cell(row=r, column=7, value=w1)
    ws.cell(row=r, column=8, value=w2)

wb.save(output_path)
