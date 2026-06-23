import pandas as pd
from openpyxl import load_workbook

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_4/regression_gate/before_pass/core_3413/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_4/regression_gate/before_pass/core_3413/output.xlsx'

# Load data
wb = load_workbook(input_path)
ws = wb.active

df = pd.read_excel(input_path)

# Set column names (better to auto-detect if headers may be different)
cols = [c.lower() for c in df.columns]
try:
    dep_col = [c for c in df.columns if 'dep' in c.lower()][0]
    ru_col  = [c for c in df.columns if 'ru' in c.lower()][0]
    val_col = [c for c in df.columns if 'val' in c.lower()][0]
except IndexError:
    dep_col, ru_col, val_col = df.columns[:3]

# Read E/F columns for the lookup
E_vals = [ws[f'E{i}'].value for i in range(3,7)]
F_vals = [ws[f'F{i}'].value for i in range(3,7)]

results = []
for dep, ru in zip(E_vals, F_vals):
    mask_both = (df[dep_col] == dep) & (df[ru_col] == ru)
    mask_dep  = (df[dep_col] == dep)
    if any(mask_both):
        val = df.loc[mask_both, val_col].sum()
    else:
        val = df.loc[mask_dep, val_col].sum()
    results.append(val)

# Write results to G3:G6
for idx, val in enumerate(results, start=3):
    ws[f'G{idx}'] = val

wb.save(output_path)
