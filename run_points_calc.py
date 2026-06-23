import openpyxl

# Load workbook and select active worksheet
wb = openpyxl.load_workbook("results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/after_fix/core_54274/input.xlsx")
ws = wb.active

total = ws["C4"].value
target = ws["D4"].value
stretch = ws["E4"].value
overstretch = ws["F4"].value

if None in (total, target, stretch, overstretch):
    raise ValueError("Missing required input value in C4-F4!")

if total <= target:
    points = total / target
elif total > target and total < stretch:
    points = min(total / target, 1.09)
elif total == stretch:
    points = min(total / target, 1.1)
elif total > stretch and total < overstretch:
    points = min(total / target, 1.19)
elif total >= overstretch:
    points = min(total / target, 1.2)
else:
    points = None

ws["G4"] = points

wb.save("results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/after_fix/core_54274/output.xlsx")
