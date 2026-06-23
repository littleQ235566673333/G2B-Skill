from openpyxl import load_workbook
import re
import itertools

def parse_filter(filter_str):
    # Patterns for = and IN LIST
    key_value = re.findall(r'"(\w+)" = "?([\w]+)"?', filter_str)
    key_list = re.findall(r'"(\w+)" IN LIST ((?:\d{1,}|0\d{1,}|\d{2,}|\d+,?)+)', filter_str)
    # allow for numbers with leading zeros, but convert to int later
    key_vals = {k: [v] for k, v in key_value}
    for k, v in key_list:
        vals = [x.lstrip('0') or '0' for x in v.split(',')]
        key_vals[k] = vals
    return key_vals

def product_dict(data):
    # Only for orgId, careType, specialtyId, contextId; rest ignored
    keys = ['orgId', 'careType', 'specialtyId', 'contextId']
    items = [(k, data.get(k, [None])) for k in keys]
    values_list = [v for _, v in items]
    for prod in itertools.product(*values_list):
        d = dict(zip(keys, prod))
        yield d

infile = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_7/task_325-44/r2/evolve_325-44/input.xlsx'
outfile = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_7/task_325-44/r2/evolve_325-44/output.xlsx'
wb = load_workbook(infile)
ws_in = wb['Input']
ws_out = wb['Output']

rows = list(ws_in.iter_rows(min_row=2, values_only=True))
result = []
for r in rows:
    wtype_id, name, status, filter_str = r
    if not wtype_id:
        continue
    filters = parse_filter(filter_str)
    for prod in product_dict(filters):
        specialtyId = prod['specialtyId']
        contextId = prod['contextId']
        # If any IN LIST, explode each, otherwise take single value or None
        orgId = prod['orgId']
        careType = prod['careType']
        # Clean up leading zero for specialtyId/contextId
        if specialtyId is not None:
            specialtyId = int(specialtyId) if specialtyId not in (None,'') else None
        if contextId is not None:
            contextId = int(contextId) if contextId not in (None,'') else None
        result.append((wtype_id, name, status, orgId, careType, specialtyId, contextId))

# Write up to 10 rows to Output
ws_out.delete_rows(2, ws_out.max_row)
for i, row in enumerate(result[:9], 2):
    for j, val in enumerate(row, 1):
        ws_out.cell(row=i, column=j, value=val)
wb.save(outfile)
