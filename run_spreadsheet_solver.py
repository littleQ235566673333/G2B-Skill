import openpyxl
from openpyxl import load_workbook
import copy

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_1/group_387-16/r1/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_1/group_387-16/r1/evolve_387-16/output.xlsx'

# Load workbook and select 'Sheet1'
wb = load_workbook(input_path)
ws = wb['Sheet1']

# Get all data in the range A2:D18
MIN_ROW, MAX_ROW = 2, 18
MIN_COL, MAX_COL = 1, 4

data = []
for row in ws.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, min_col=MIN_COL, max_col=MAX_COL, values_only=True):
    data.append(list(row))

# Identify column indices from header, assumed in row 1
headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
col_A = 0
col_Value = 1
col_Binaries = 2
col_Result_values = 3

# Extract lists
A_list = [r[col_A] for r in data]
Value_list = [r[col_Value] for r in data]
Binaries_list = [r[col_Binaries] for r in data]
Result_values_list = [r[col_Result_values] for r in data]

# For each value in Result_values, remove it once from Value & Binaries
Value_mod = copy.copy(Value_list)
Binaries_mod = copy.copy(Binaries_list)

for result_val in Result_values_list:
    if result_val is not None and result_val in Value_mod:
        idx = Value_mod.index(result_val)
        Value_mod.pop(idx)
        Binaries_mod.pop(idx)

# Remove empty cells (None or blanks) and shift up
compact_Value = [v for v in Value_mod if v not in [None, '']]
compact_Binaries = [b for b in Binaries_mod if b not in [None, '']]

# Pad lists to original length (A2:A18 is length 17)
pad_len = len(Value_list)
compact_Value += [None] * (pad_len - len(compact_Value))
compact_Binaries += [None] * (pad_len - len(compact_Binaries))

# Write results back to Sheet1!B2:B18 and C2:C18
for i in range(pad_len):
    ws.cell(row=MIN_ROW + i, column=2).value = compact_Value[i]
    ws.cell(row=MIN_ROW + i, column=3).value = compact_Binaries[i]

# Compute solver result (sum of non-None, non-blank col A)
solver_result = sum([x for x in A_list if isinstance(x, (int, float))])

# Find target value: try D1, else first numeric in Result Values
header_cell = ws.cell(row=1, column=4).value
if isinstance(header_cell, (int, float)):
    target_value = header_cell
else:
    target_value = None
    for val in Result_values_list:
        if isinstance(val, (int, float)):
            target_value = val
            break

# Write solver result and difference to E2, E3, F2, F3
ws['E2'].value = 'solver result'
ws['F2'].value = solver_result
ws['E3'].value = 'difference'
ws['F3'].value = solver_result - target_value if target_value is not None else None

wb.save(output_path)
