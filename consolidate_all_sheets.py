import openpyxl
from copy import copy
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/regression_gate/after_pass/core_80-42/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/regression_gate/after_pass/core_80-42/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_out = wb['Consolidate_ALL']
sheets = [s for s in wb.sheetnames if s not in ['Consolidate_ALL']]

# Find first available output row in A (skip header)
def first_blank_row(ws, col='A', start=2):
    for i in range(start, ws.max_row+10_000):
        if not ws[f'{col}{i}'].value:
            return i
    return ws.max_row + 1

# Safely copy formatting except fill
def copy_format(src_cell, tgt_cell, skip_fill=False):
    tgt_cell.font = copy(src_cell.font)
    tgt_cell.border = copy(src_cell.border)
    tgt_cell.number_format = src_cell.number_format
    tgt_cell.protection = copy(src_cell.protection)
    tgt_cell.alignment = copy(src_cell.alignment)
    if not skip_fill:
        tgt_cell.fill = copy(src_cell.fill)

def is_yellow(cell):
    fill = cell.fill
    if fill is None or fill.patternType != 'solid':
        return False
    fg = getattr(fill, 'fgColor', None)
    if fg is not None:
        code = None
        if hasattr(fg, 'rgb') and fg.rgb:
            code = fg.rgb
        elif hasattr(fg, 'value'):
            code = fg.value
        if code:
            code = str(code).upper()
            if code in {'FFFFFF00', 'FFFF00', 'FF00FFFF00'} or code.endswith('FF00'):
                return True
    return False

columns = 12
col_L_idx = 12
out_row = first_blank_row(ws_out)

for sheet_name in sheets:
    ws_in = wb[sheet_name]
    # Find header row with any values
    header_row = 1
    while header_row <= ws_in.max_row and not any(ws_in.cell(row=header_row, column=col).value for col in range(1, ws_in.max_column+1)):
        header_row += 1
    # Detect yellow highlight columns in header row
    highlight_cols = []
    for col in range(1, ws_in.max_column+1):
        c = ws_in.cell(row=header_row, column=col)
        if is_yellow(c):
            highlight_cols.append(col)
    # Fallback: take A-K (1-11)
    if not highlight_cols:
        highlight_cols = list(range(1, columns))
    ncols = min(columns-1, len(highlight_cols))
    # Start scanning rows below header
    for r_idx in range(header_row+1, ws_in.max_row+1):
        # Skip rows that are completely blank in highlight cols
        if not any(ws_in.cell(row=r_idx, column=col).value is not None for col in highlight_cols):
            continue
        # Copy cell values and formats
        for cidx, col in enumerate(highlight_cols[:ncols]):
            s_cell = ws_in.cell(row=r_idx, column=col)
            t_cell = ws_out.cell(row=out_row, column=cidx+1)
            t_cell.value = s_cell.value
            copy_format(s_cell, t_cell, skip_fill=True)
            refrow = out_row-1 if out_row-1 >= 2 else 2
            t_cell.fill = copy(ws_out.cell(row=refrow, column=cidx+1).fill)
        # Write sheet name in column L
        l_cell = ws_out.cell(row=out_row, column=col_L_idx)
        l_cell.value = sheet_name
        refrow = out_row-1 if out_row-1 >=2 else 2
        copy_format(ws_out.cell(row=refrow, column=col_L_idx), l_cell, skip_fill=False)
        out_row += 1
wb.save(output_path)
