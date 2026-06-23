import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_5/task_45707/r3/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_5/task_45707/r3/evolve_45707/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Map header names to columns
def find_col(header):
    for cell in ws[1]:
        if cell.value == header:
            return cell.column
    raise ValueError(f'Header {header} not found')
col_A = find_col('Date')           # Date column
col_C = find_col('Occurence')     # Target value to count

# Helper: Parse Excel date and other formats
def parse_excel_date(val):
    if isinstance(val, datetime):
        return val
    try:
        if isinstance(val, (int, float)):
            return datetime(1899, 12, 30) + timedelta(days=val)
        return datetime.strptime(str(val), '%Y-%m-%d')
    except Exception:
        return None

N = 68  # For D2:D69
first_row = 2
last_row = first_row + N - 1

dates = [ws.cell(row=r, column=col_A).value for r in range(first_row, last_row + 1)]
dates = [parse_excel_date(v) for v in dates]
col_C_vals = [ws.cell(row=r, column=col_C).value for r in range(first_row, last_row + 1)]

month_groups = defaultdict(list)
for idx, dt in enumerate(dates):
    if dt:
        month_groups[(dt.year, dt.month)].append(idx)

for i in range(N):
    next_dt = dates[i+1] if i+1 < N else None
    target_cell = ws.cell(row=first_row + i, column=4)  # This is always column D
    if next_dt and next_dt.day == 1:
        idxs = month_groups.get((next_dt.year, next_dt.month), [])
        count_1 = sum(1 for idx in idxs if col_C_vals[idx] == 1)
        target_cell.value = count_1
    else:
        target_cell.value = None

wb.save(output_path)
