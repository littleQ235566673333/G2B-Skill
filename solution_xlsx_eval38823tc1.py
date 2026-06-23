from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun1/eval_38823_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun1/eval_38823_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active
# Data starts at row 3
# 'Date' in A3, 'Fabric' in B3, 'Units Sold' in C3
# Date range: E4(start), F4(end)
# Search terms in H5:H8; Output in I5:I8

# Get date range
start_date = ws['E4'].value
end_date = ws['F4'].value

# For each search term in H5:H8
for row in range(5, 9):
    search_term = ws[f'H{row}'].value
    if not search_term:
        ws[f'I{row}'].value = None
        continue
    total = 0
    # Scan data rows from row 4 and down (until first blank date)
    data_row = 4
    while ws[f'A{data_row}'].value:
        dt = ws[f'A{data_row}'].value
        fabric = ws[f'B{data_row}'].value
        units = ws[f'C{data_row}'].value
        # Date check
        if (start_date and dt < start_date) or (end_date and dt > end_date):
            data_row += 1
            continue
        # Search term presence
        if fabric and search_term in fabric and units:
            try:
                total += float(units)
            except Exception:
                pass
        data_row += 1
    ws[f'I{row}'].value = total
wb.save(output_path)
