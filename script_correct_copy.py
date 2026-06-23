from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_56378_tc1/input.xlsx')
ws = wb['Folha1']

src_rows = []
for r in range(5, 11):
    if ws.cell(row=r, column=9).value not in (None, ''):
        src_rows.append(r)

for idx, r in enumerate(src_rows[:4]):
    for offset in range(7): # columns C:I map to L:R
        cell = ws.cell(row=5+idx, column=12+offset)
        src_col = 3+offset
        col_letter = get_column_letter(src_col)
        cell.value = f"=Folha1!{col_letter}{r}"

for ridx in range(5,9):
    ws.cell(row=ridx, column=13).alignment = Alignment(horizontal='left')
    for c in range(15,19):
        ws.cell(row=ridx, column=c).alignment = Alignment(horizontal='right')

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_56378_tc1/output.xlsx')
print('Wrote rows', src_rows[:4], 'to L5:R8.')
