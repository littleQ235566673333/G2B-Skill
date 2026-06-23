import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_52541_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_52541_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

for i in range(6, 11):
    # Amount Outstanding: column 8, Over Due Days: column 9
    amount_outstanding = ws.cell(row=i, column=8).value
    over_due_days = ws.cell(row=i, column=9).value

    # If Amount Outstanding is negative, remark is 'Prepaid'
    if amount_outstanding is not None and isinstance(amount_outstanding, (int, float)) and amount_outstanding < 0:
        ws.cell(row=i, column=10).value = 'Prepaid'
    else:
        # Over Due Days is empty
        if over_due_days in [None, '', 0]:
            ws.cell(row=i, column=10).value = ''
        else:
            try:
                over_due_days_val = float(over_due_days)
            except (ValueError, TypeError):
                ws.cell(row=i, column=10).value = ''
                continue
            if over_due_days_val < 90:
                ws.cell(row=i, column=10).value = 'Call Customer'
            else:
                ws.cell(row=i, column=10).value = 'Bad Debts'

wb.save(output_path)
