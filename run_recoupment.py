import openpyxl

def to_number(val):
    try:
        return float(val)
    except Exception:
        return 0  # treat invalids as zero for business logic

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_36097_tc1/input.xlsx')
ws = wb.active

def recoupment(cost, itv, profit, proceeds):
    if proceeds < cost:  # Sold at a loss
        loss = cost - proceeds
        return itv + loss
    elif profit < cost:  # Profit but less than original cost
        return profit  # Recoup the full profit
    else:
        return cost - itv

for row in range(3, 7):
    cost = to_number(ws.cell(row=row, column=3).value)
    itv = to_number(ws.cell(row=row, column=5).value)
    profit = to_number(ws.cell(row=row, column=6).value)
    proceeds = to_number(ws.cell(row=row, column=7).value)
    value = recoupment(cost, itv, profit, proceeds)
    # Preserve column I formatting
    source_cell = ws.cell(row=row, column=9)
    target_cell = ws.cell(row=row, column=8)
    target_cell.value = value
    if source_cell.has_style:
        target_cell._style = source_cell._style

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_36097_tc1/output.xlsx')
