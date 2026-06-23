from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_4/regression_gate/before_fix/core_58701/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_4/regression_gate/before_fix/core_58701/output.xlsx'

wb = load_workbook(input_path)
table_ws = wb['Table Tab']
entry_ws = wb['Entry Tab']

# Build location→office code mapping from Table Tab (A2:B3)
mapping = {}
for row in table_ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=2, values_only=True):
    name, code = row
    if name:
        mapping[str(name).strip()] = code

# For each entry in Entry Tab B2:B3, look up code, write to E2:E3
for entry_row in range(2, 4):
    loc = entry_ws.cell(row=entry_row, column=2).value
    code = mapping.get(str(loc).strip()) if loc else None
    entry_ws.cell(row=entry_row, column=5, value=code)

wb.save(output_path)
