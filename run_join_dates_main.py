import openpyxl
import datetime
import sys

INPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/group_45896/r1/evolve_45896/input.xlsx'
OUTPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/group_45896/r1/evolve_45896/output.xlsx'
wb = openpyxl.load_workbook(INPUT_PATH)
ws_main = wb['Volym P5_P6_2023']
ws_zord = wb['ZORD']

# Identify column A in main (Material) and column A (Material), C (Dates) in ZORD
material_col_main = 1  # 1-based indexing
material_col_zord = 1
zord_date_col = 3

# Helper to format Excel float date, Excel base date is 1899-12-30
def format_excel_date(val):
    if isinstance(val, datetime.datetime) or isinstance(val, datetime.date):
        return val.strftime('%d/%m/%Y')
    try:
        dt = float(val)
        base = datetime.datetime(1899, 12, 30)
        date_val = base + datetime.timedelta(days=dt)
        return date_val.strftime('%d/%m/%Y')
    except Exception:
        return None

# Load all ZORD data
zord_rows = list(ws_zord.iter_rows(values_only=True))
material_to_dates = {}
for row in zord_rows[1:]:
    mat = row[material_col_zord-1]
    datecell = row[zord_date_col-1]
    if mat and datecell:
        material_to_dates.setdefault(mat, []).append(datecell)

# For each main row, join all matching dates as string in DD/MM/YYYY
for i in range(2, 11):  # I2:I10
    mat = ws_main.cell(row=i, column=material_col_main).value
    dates = material_to_dates.get(mat, [])
    formatted_dates = [format_excel_date(d) for d in dates if format_excel_date(d)]
    joined = ','.join(formatted_dates)
    ws_main.cell(row=i, column=9, value=joined)

wb.save(OUTPUT_PATH)
