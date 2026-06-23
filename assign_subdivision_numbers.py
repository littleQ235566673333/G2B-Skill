import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/regression_gate/before_fix/core_42930/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/regression_gate/before_fix/core_42930/output.xlsx'

# Load with pandas
df = pd.read_excel(input_path)

# Identify required columns
lower_cols = {c.lower(): c for c in df.columns}
sub_col = next((lower_cols[c] for c in lower_cols if 'subdiv' in c), None)
built_col = next((lower_cols[c] for c in lower_cols if 'built' in c), None)
assert sub_col and built_col, f'Could not infer columns for subdivision: {sub_col}, house built: {built_col}'

# Compute earliest house built per subdivision
earliest_built = df.groupby(sub_col)[built_col].min().reset_index()
earliest_built_sorted = earliest_built.sort_values(built_col).reset_index(drop=True)
sub_to_number = {row[sub_col]: i+1 for i,row in earliest_built_sorted.iterrows()}

# Assign the unique number to each house
df['SubdivOrder'] = df[sub_col].map(sub_to_number)
# Sort by new order
df_sorted = df.sort_values(['SubdivOrder', built_col]).reset_index(drop=True)

# Patch result into original file using openpyxl
wb = load_workbook(input_path)
ws = wb.active
# Write the order numbers to C2:C22
for i in range(2, 23):
    val = int(df_sorted.loc[i-2, 'SubdivOrder']) if i-2 < len(df_sorted) else ''
    ws.cell(row=i, column=3, value=val)
wb.save(output_path)
