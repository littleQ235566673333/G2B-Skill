import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_414-20_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_414-20_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']
first_occurrence = None
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1), start=1):
    cell_value = row[0].value
    if cell_value is not None and str(cell_value).strip().lower() == 'invoice no.':
        first_occurrence = i
        break
if first_occurrence and first_occurrence > 1:
    ws.delete_rows(1, first_occurrence - 1)
wb.save(output_path)
