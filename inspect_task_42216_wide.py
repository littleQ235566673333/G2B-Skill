from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_42216_s1/input.xlsx'
wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

for r in range(1, 25):
    vals = [ws.cell(r, c).value for c in range(1, 35)]
    print('ROW', r)
    for c, v in enumerate(vals, start=1):
        if v is not None:
            print(c, v)
