from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/group_341-40/r3/evolve_341-40/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/group_341-40/r3/evolve_341-40/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

max_row = 701  # per user instruction
for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=17):
    if (
        row[0].value == 'Government' and
        row[1].value == 'Germany' and
        row[2].value == 'Carretera'
    ):
        ws.cell(row=row[0].row, column=17, value='Volkswagen')

wb.save(output_path)
