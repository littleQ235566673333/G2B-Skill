from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2-PRUNED/eval_seed42/eval_48983_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2-PRUNED/eval_seed42/eval_48983_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Build mappings for row headers (categories) and col headers (brands)
category_to_row = {}
for r in range(6, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val is not None:
        category_to_row[val] = r
brand_to_col = {}
for c in range(2, 13):
    val = ws.cell(row=5, column=c).value
    if val is not None:
        brand_to_col[val] = c
# Now for M6:S11, copy the corresponding value from (category/brand)
for dest_row in range(6, 12):
    category = ws.cell(row=dest_row, column=1).value
    for dest_col in range(13, 20):
        brand = ws.cell(row=5, column=dest_col).value
        src_row = category_to_row.get(category, None)
        src_col = brand_to_col.get(brand, None)
        value = ws.cell(row=src_row, column=src_col).value if src_row and src_col else None
        ws.cell(row=dest_row, column=dest_col).value = value

wb.save(output_path)
