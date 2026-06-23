from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_120-24_tc1/input.xlsx')
ws = wb['Sheet1']
for col in range(65, 81):
    val = ws.cell(row=1, column=col).value
    print(f'{col} ({get_column_letter(col)}): {val}')
