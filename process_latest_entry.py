import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Load the workbook and the worksheet
in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_58723_tc1/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_58723_tc1/output.xlsx'
wb = openpyxl.load_workbook(in_path)
ws = wb.active

# Figure out the range (M2:M41)
start_row = 2
end_row = 41
name_col = 'C'
time_col = 'I'
output_col = 'M'

# 1. Gather all (Name, Entry Time) pairs for rows under consideration and keep original row for later
name_time_rows = []
for row in range(start_row, end_row+1):
    name = ws[f'{name_col}{row}'].value
    entry_time = ws[f'{time_col}{row}'].value
    # Try to parse date if it's a string
    if isinstance(entry_time, str):
        try:
            entry_time = datetime.strptime(entry_time, '%Y-%m-%d %H:%M:%S')
        except Exception:
            pass
    name_time_rows.append((name, entry_time, row))

# 2. Find the latest entry time for each name
timemap = {}
for name, entry_time, row in name_time_rows:
    if name is None or entry_time is None:
        continue
    if name not in timemap or entry_time > timemap[name]:
        timemap[name] = entry_time

# 3. Mark latest and not latest in column M
for name, entry_time, row in name_time_rows:
    if name is None or entry_time is None or name not in timemap:
        ws[f'{output_col}{row}'] = None
    elif entry_time == timemap[name]:
        ws[f'{output_col}{row}'] = 'Latest'
    else:
        ws[f'{output_col}{row}'] = 'Not Latest'

wb.save(out_path)
