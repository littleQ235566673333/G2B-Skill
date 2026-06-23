from openpyxl import load_workbook

wb = load_workbook('results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed2/eval_52541_tc1/input.xlsx')
ws = wb['Sheet1']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    print('Row', i, [str(cell).strip() if cell is not None else '' for cell in row])
