from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_3/regression_gate/before_fix/core_382-10/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_3/regression_gate/before_fix/core_382-10/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet1']

for row in range(1, 20):  # A1:A19 (1-based)
    cell = ws.cell(row=row, column=1)
    if cell.value != 'Comments':
        cell.value = None

wb.save(output_path)
