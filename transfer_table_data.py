import openpyxl
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_48983_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_48983_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Source headers: Brands in row 5 (C onwards), categories in B6 downward
src_brand_row = 5
src_cat_col = 2
src_brand_start_col = 3
src_brands = []
for col in range(src_brand_start_col, ws.max_column+1):
    val = ws.cell(row=src_brand_row, column=col).value
    if val:
        src_brands.append(val)
    else:
        break
src_cats = []
row = 6
while True:
    val = ws.cell(row=row, column=src_cat_col).value
    if val:
        src_cats.append(val)
        row += 1
    else:
        break

# Destination headers: Brands in M5:S5, categories in L6:L11
# Column M = 13, S = 19
# Row 6 to 11

dest_brand_row = 5
M_col = 13
S_col = 19
dest_brand_cols = range(M_col, S_col+1)
dest_cat_col = 12  # L
output_start_row = 6  # M6
output_end_row = 11  # M11

dest_brands = [ws.cell(row=dest_brand_row, column=col).value for col in dest_brand_cols]
dest_cats = [ws.cell(row=row, column=dest_cat_col).value for row in range(output_start_row, output_end_row+1)]

# Build lookup from (category, brand) to value
src_data = {}
for i, cat in enumerate(src_cats):
    for j, brand in enumerate(src_brands):
        val = ws.cell(row=6+i, column=src_brand_start_col + j).value
        src_data[(cat, brand)] = val

# Copy to output
for r, cat in enumerate(dest_cats):
    for c, brand in enumerate(dest_brands):
        output_row = output_start_row + r
        output_col = M_col + c
        val = src_data.get((cat, brand), None)
        ws.cell(row=output_row, column=output_col, value=val)

wb.save(output_path)
