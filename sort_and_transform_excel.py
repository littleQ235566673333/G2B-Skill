import openpyxl

# Input and output paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/group_118-50/r3/evolve_118-50/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/group_118-50/r3/evolve_118-50/output.xlsx'

ending_order = ['ING', 'ERS', 'ATE', 'EST', 'ONE', 'IER', 'ILY']

def get_last3(word):
    return word[-3:] if len(word) >= 3 else ''

def fifth_to_first(word):
    if len(word) < 5:
        return word
    else:
        return word[4] + word[:4] + word[5:]

# Read words from Sheet1 column A
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']
words = []
for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
    if row[0] is not None and str(row[0]).strip() != '':
        words.append(str(row[0]).strip())
# Sort the names alphabetically and update in place
words_sorted = sorted(words)
for i, word in enumerate(words_sorted, start=2):
    ws.cell(row=i, column=1, value=word)

# Collect word pairs
from collections import defaultdict
pairs_by_ending = defaultdict(list)
word_set = set(words_sorted)
for word in words_sorted:
    if len(word) < 8:
        continue  # Minimum for 5th char change and last 3
    transformed = fifth_to_first(word)
    if transformed != word and transformed in word_set:
        ending = get_last3(word)
        pairs_by_ending[ending].append((transformed, word))

# Write results to columns C (transformed) and D (original), grouped/sorted
row_out = 2
ws = wb['Sheet1']
for ending in ending_order:
    if ending in pairs_by_ending:
        group = sorted(pairs_by_ending[ending], key=lambda x: (x[1], x[0]))
        for t, o in group:
            ws.cell(row=row_out, column=3, value=t)
            ws.cell(row=row_out, column=4, value=o)
            row_out += 1
# Append the rest, alphabetically by group ending, if any
rest_endings = sorted(e for e in pairs_by_ending.keys() if e not in ending_order)
for ending in rest_endings:
    group = sorted(pairs_by_ending[ending], key=lambda x: (x[1], x[0]))
    for t, o in group:
        ws.cell(row=row_out, column=3, value=t)
        ws.cell(row=row_out, column=4, value=o)
        row_out += 1

wb.save(output_path)
