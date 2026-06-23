import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_58032_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_58032_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Style setup
header_font = Font(name='Arial', bold=True, italic=True)
cell_font = Font(name='Arial')
gray_fill = PatternFill(fill_type='solid', fgColor='CCCCCC')
orange_fill = PatternFill(fill_type='solid', fgColor='FCD5B4')
border_style = Side(border_style='thin', color='000000')
border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

max_row = ws.max_row
max_col = ws.max_column

# Set header styling (bold, italic, border for row 1)
for col in range(1, max_col + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.border = border
# Set Arial font for all other cells
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = cell_font
# Fill column D with #CCCCCC from D2 down
for row in range(2, max_row + 1):
    ws[f'D{row}'].fill = gray_fill
# Formula for A2:A35; fill them as well
for row in range(2, 36):
    ws[f'A{row}'].value = f'=INDEX($J$2:$J${max_row},MATCH(B{row}&C{row},$I$2:$I${max_row},0))'
    ws[f'A{row}'].fill = orange_fill
# Hide gridlines
ws.sheet_view.showGridLines = False
wb.save(output_path)
