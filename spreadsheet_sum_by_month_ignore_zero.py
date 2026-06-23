from openpyxl import load_workbook
import calendar
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r2/eval_40757_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r2/eval_40757_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

data = list(ws.iter_rows(min_row=2, values_only=True))
month_values = defaultdict(list)

for row in data:
    if not row or len(row) < 2:
        continue
    month, val = row[0], row[1]
    # Accept various date/month representations
    if isinstance(month, str):
        month_str = month.strip().capitalize()
    elif hasattr(month, 'month') and hasattr(month, 'year'):
        month_str = calendar.month_name[month.month]
    else:
        continue
    try:
        val_flt = float(val)
    except (TypeError, ValueError):
        continue
    if val_flt != 0:
        month_values[month_str].append(val_flt)

# Pick two months from the available ones if at least two
selected_months = list(month_values.keys())[:2]
sums = []
for month in selected_months:
    s = sum(month_values[month])
    sums.append((month, s))
    # Write sum in B10, B11
for i, (_, result) in enumerate(sums):
    ws.cell(row=10+i, column=2).value = result

wb.save(output_path)
