import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_8/regression_gate/before_fix/core_56274/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_8/regression_gate/before_fix/core_56274/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheet = wb.active

# Get fiscal month from D7
fiscal_month = sheet['D7'].value

# Extract values from appropriate rows for Opening/Closing/Debits/Credits
# They are in:
# Row 9: Opening Balance (col 4)
# Row 10: Debits (col 4)
# Row 11: Credits (col 4)
# Row 12: Closing Balance (col 4 or maybe elsewhere, see test below)
rows = {
    'opening_balance': 9,
    'debits': 10,
    'credits': 11,
    'closing_balance': 12,
}

# We trust values are in col D/4 for month in D7, unless otherwise detected
col = 4  # D
sheet['D9'].value = sheet.cell(row=rows['opening_balance'], column=col).value
sheet['D10'].value = sheet.cell(row=rows['debits'], column=col).value
sheet['D11'].value = sheet.cell(row=rows['credits'], column=col).value
sheet['D12'].value = sheet.cell(row=rows['closing_balance'], column=col).value

wb.save(output_path)
