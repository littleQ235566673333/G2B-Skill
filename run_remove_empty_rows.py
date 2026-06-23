import openpyxl

INPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun2/eval_409-45_tc1/input.xlsx'
OUTPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun2/eval_409-45_tc1/output.xlsx'

wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb['DATA']

def is_row_empty(row, columns):
    return all((row[col].value is None or str(row[col].value).strip() == '') for col in columns)

I2_value = ws['I2'].value
start_row, end_row = 1, 19
min_col, max_col = 1, 5
columns_indices = [i for i in range(max_col)]  # 0-based for row cells
rows = list(ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col))

header = [cell.value for cell in rows[0]]
filtered_rows = [rows[0]]  # Always keep the header

if I2_value:
    target_id = I2_value
    # For each row, if column B (cell[1]) matches the target_id and the row is NOT empty, keep it
    # Only keep header and matching, non-empty rows
    for r in rows[1:]:
        if r[1].value == target_id and not is_row_empty(r, columns_indices):
            filtered_rows.append(r)
else:
    # No I2 value: keep all non-empty rows (except header)
    for r in rows[1:]:
        if not is_row_empty(r, columns_indices):
            filtered_rows.append(r)

# Write filtered rows back to A1:E19
for idx, row in enumerate(filtered_rows):
    for j, cell in enumerate(row):
        ws.cell(row=start_row+idx, column=j+1, value=cell.value)
# Blank out remains
for idx in range(len(filtered_rows), end_row):
    for col in range(min_col, max_col+1):
        ws.cell(row=start_row+idx, column=col, value=None)

wb.save(OUTPUT_PATH)
