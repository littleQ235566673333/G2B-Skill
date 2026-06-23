import openpyxl
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r1/eval_250-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r1/eval_250-20_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['RNM']

# Get data rows (A1:J20)
data = list(ws.iter_rows(min_row=1, max_col=10, max_row=20, values_only=True))
header = data[0]
data_rows = data[1:]

# Build a mapping from (B, C) pair to row and sum column J (index 9)
sum_dict = {}
for row in data_rows:
    key = (row[1], row[2])  # (colB, colC)
    if key not in sum_dict:
        sum_dict[key] = list(row)
        # If J is None, treat as 0
        sum_dict[key][9] = row[9] if row[9] is not None else 0
    else:
        sum_dict[key][9] += row[9] if row[9] is not None else 0

# Output data in the same order as first occurrence in input
distinct_keys = []
seen = set()
for row in data_rows:
    key = (row[1], row[2])
    if key not in seen:
        distinct_keys.append(key)
        seen.add(key)

out_rows = [header] + [sum_dict[k] for k in distinct_keys]

# Clear existing data in A1:J20
for r in range(1, 21):
    for c in range(1, 11):
        ws.cell(row=r, column=c, value=None)

# Write new rows back
for ir, row in enumerate(out_rows):
    for ic, val in enumerate(row):
        ws.cell(row=ir+1, column=ic+1, value=val)

wb.save(output_path)
