import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_51289_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_51289_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
labels = [ws.cell(row=1, column=i).value for i in range(1,9)]
numbers = [ws.cell(row=2, column=i).value for i in range(1,9)]
results = ['']*8

# For each animal, find largest and second largest value, note indices
for animal in ['cat','dog']:
    filtered = [(i, num) for i,(lab,num) in enumerate(zip(labels, numbers)) if lab == animal]
    if len(filtered) > 0:
        (max_idx, max_val) = max(filtered, key=lambda x: x[1])
        results[max_idx] = f'{animal}1'
    if len(filtered) > 1:
        filtered_sorted = sorted(filtered, key=lambda x: x[1], reverse=True)
        (second_idx, second_val) = filtered_sorted[1]
        results[second_idx] = f'{animal}2'

for i, val in enumerate(results):
    ws.cell(row=4, column=i+1).value = val

wb.save(output_path)
