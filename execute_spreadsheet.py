import openpyxl
from openpyxl.styles import Alignment

# Load workbook and worksheet
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_56378_tc1/input.xlsx')
ws = wb['Folha1']

# Get input data rows for frame 1 (columns C:I)
input_rows = list(ws.iter_rows(min_row=5, max_row=15, min_col=3, max_col=9, values_only=True))

# Headers for frame 2 (columns L:R), which is columns 12-18 (0-based 11:18)
headers = [cell.value for cell in ws[4][11:18]]

# Filter rows: only keep ones where 'QUANTITY UNITS' (col I, which is index 6) is not None and not 0
filtered = []
for row in input_rows:
    if row[6] is not None and row[6] != 0:
        filtered.append(row)

# Write headers in L5:R5
for c, h in enumerate(headers, start=12):
    ws.cell(row=5, column=c, value=h)

# Write up to 3 matching rows in L6:R8
for i, row in enumerate(filtered[:3]):
    for j, val in enumerate(row):
        ws.cell(row=6+i, column=12+j, value=val)

# Align product (col M = 13) left, values (col O = 15 to R = 18) right
for r in range(6, 9):
    prod_cell = ws.cell(row=r, column=13)
    prod_cell.alignment = Alignment(horizontal='left')
    for c in range(15, 19):
        ws.cell(row=r, column=c).alignment = Alignment(horizontal='right')

# Save output
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_56378_tc1/output.xlsx')
