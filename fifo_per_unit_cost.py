import openpyxl

def fifo_unit_cost(row, lots, needed):
    """
    row: dict of values from Excel
    lots: list of (qty, cost) tuples (in purchase order)
    needed: inventory units to price
    Returns FIFO per unit cost.
    """
    if needed is None or needed <= 0:
        return None
    cost_sum = 0.0
    used = 0
    for units, cost in lots:
        if needed <= 0:
            break
        take = min(needed, units)
        cost_sum += take * cost
        needed -= take
        used += take
    if used == 0:
        return None
    return round(cost_sum / used, 4)


def main():
    infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_8/group_39432/r0/evolve_39432/input.xlsx'
    outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_8/group_39432/r0/evolve_39432/output.xlsx'

    wb = openpyxl.load_workbook(infile)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    # Index mapping
    idx_item = 0
    idx_begin_qty = 1
    idx_begin_cost = 2
    purchase_cols = [(3,4), (5,6), (7,8), (9,10)]  # Each PO and Cost
    idx_cur_inv = 11
    idx_cur_cost = 12  # Where result goes

    # Rows 2-5
    for r in range(2,6):
        row = [ws.cell(row=r, column=i+1).value for i in range(len(header))]
        lots = []
        # Beginning inv
        if row[idx_begin_qty] and row[idx_begin_cost]:
            lots.append((row[idx_begin_qty], row[idx_begin_cost]))
        # Purchases
        for qty_col, cost_col in purchase_cols:
            qty = row[qty_col]
            cost = row[cost_col]
            if qty is not None and cost is not None:
                lots.append((qty, cost))
        needed = row[idx_cur_inv]
        result = fifo_unit_cost(row, lots, needed)
        ws.cell(row=r, column=idx_cur_cost+1, value=result)
    wb.save(outfile)

if __name__ == '__main__':
    main()
