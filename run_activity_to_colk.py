import openpyxl

# File paths
target_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun2/eval_33157_tc1/output.xlsx'
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun2/eval_33157_tc1/input.xlsx'

# Open the workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

activity_cols = ['B', 'D', 'F', 'H']
header_row = 1

def get_col_index(col):
    """Given 'A', 'B', returns 1-based Excel index for openpyxl."""
    return openpyxl.utils.column_index_from_string(col)

# Get headers for activity columns
activity_headers = [ws[f'{col}{header_row}'].value for col in activity_cols]

# Determine max row (be robust: stop at first empty in A or J after header)
row = 2
while ws[f'A{row}'].value is not None or ws[f'J{row}'].value is not None:
    row += 1
max_row = row - 1

for row in range(2, max_row+1):
    target_date = ws[f'J{row}'].value
    found = None
    for col, header in zip(activity_cols, activity_headers):
        v = ws[f'{col}{row}'].value
        if v == target_date and target_date is not None:
            found = header
            break
    ws[f'K{row}'] = found if found else ''

wb.save(target_path)
