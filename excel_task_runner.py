from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INPUT = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r2/eval_120-24_tc1/input.xlsx'
OUTPUT = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r2/eval_120-24_tc1/output.xlsx'
SHEET = 'Sheet1'

# Static column indices from preview: (Python 0-based)
COL_AY = 50
COL_BG = 57
COL_BN = 58
COL_BL = 63

BN_FILL_HEX = '0070C0'
fill = PatternFill(start_color=BN_FILL_HEX, end_color=BN_FILL_HEX, fill_type='solid')

def get_str(val):
    return str(val).strip() if val is not None else ''

target_rows = range(2, 46)  # Excel Sheet1!BN2:BN45 = rows 2-45 inclusive

wb = load_workbook(INPUT)
ws = wb[SHEET]

for row in target_rows:
    val_BL = get_str(ws.cell(row=row, column=COL_BL+1).value)
    val_BG = get_str(ws.cell(row=row, column=COL_BG+1).value)
    val_AY = get_str(ws.cell(row=row, column=COL_AY+1).value)
    val_BN = get_str(ws.cell(row=row, column=COL_BN+1).value)

    # Only overwrite BN & set fill if BL == 'LAG' and BG matches
    intended_BN = ''
    if val_BL == 'LAG' and val_BG in ['OCOGS - Spares - Transfer Price Overhead', 'OFSS - ORCL Consulting Prod Cost I/C']:
        if val_AY in {'BOA_033E', 'BOA_011G'}:
            intended_BN = 'OCOGS - Spares - Transfer Price Overhead'
        else:
            intended_BN = 'OFSS - ORCL Consulting Prod Cost I/C'
        ws.cell(row=row, column=COL_BN+1, value=intended_BN)
        # Fill color for BL if BN is now nonempty
        if intended_BN:
            ws.cell(row=row, column=COL_BL+1).fill = fill

wb.save(OUTPUT)
