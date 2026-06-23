import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/regression_gate/before_pass/core_51249/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/regression_gate/before_pass/core_51249/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# GREEN FILL RGB(226,239,218) as hex
fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

# Cells of interest
row_indices = [1, 5, 9]
for idx in row_indices:
    result_cell = f'D{idx}'
    b1 = ws[f'B{idx}'].value or ''
    b2 = ws[f'B{idx+1}'].value or ''
    output = ''
    if b1 == 'Description A' and b2 == '':
        output = 'Single A'
    elif b1 == 'Description B' and b2 == '':
        output = 'Single B'
    elif b1 == 'Description A' and b2 == 'Description B':
        output = 'Multiple'
    if output:
        ws[result_cell].value = output
        ws[result_cell].fill = fill

wb.save(output_path)
