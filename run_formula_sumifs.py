import openpyxl
from datetime import datetime
import calendar

def month_number_from_name(month_name):
    try:
        return list(calendar.month_name).index(month_name.capitalize())
    except ValueError:
        pass
    try:
        return list(calendar.month_abbr).index(month_name[:3].capitalize())
    except ValueError:
        pass
    return None

def main():
    input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_57590_tc1/input.xlsx'
    output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_57590_tc1/output.xlsx'
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Read month from A26
    ref_month = ws['A26'].value
    month_num = None
    year_num = None

    if isinstance(ref_month, str):
        # If A26 is a month name only, extract month number
        month_num = month_number_from_name(ref_month.strip())
        if month_num is None:
            # Try if the user entered 'January 2024', etc.
            try:
                dt = datetime.strptime(ref_month.strip(), '%B %Y')
                month_num, year_num = dt.month, dt.year
            except Exception:
                pass
    elif isinstance(ref_month, datetime):
        month_num = ref_month.month
        year_num = ref_month.year
    
    if month_num is None:
        raise ValueError('A26 is not a recognizable month')

    # Find last row
    max_row = ws.max_row
    sum_val = 0

    for row in range(2, max_row+1):
        date_val = ws.cell(row=row, column=3).value
        sum_candidate = ws.cell(row=row, column=9).value
        if not date_val:
            continue
        # Parse date string
        if isinstance(date_val, str):
            try:
                date_val = datetime.strptime(date_val, '%Y-%m-%d')
            except ValueError:
                try:
                    date_val = datetime.strptime(date_val, '%Y/%m/%d')
                except ValueError:
                    continue
        elif not isinstance(date_val, datetime):
            continue
        if date_val.month == month_num:
            if year_num is None or date_val.year == year_num:
                sum_val += sum_candidate if sum_candidate else 0
    ws['B26'] = sum_val
    wb.save(output_path)

if __name__ == '__main__':
    main()
