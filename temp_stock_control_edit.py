import openpyxl
from datetime import datetime
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_9448_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_9448_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Data']

col_range = [chr(c) for c in range(ord('I'), ord('T')+1)]
header_row = 7
output_col = 'U'

def parse_month_header(header_value):
    """
    Attempts to parse header_value which may be like 'April 2023', 'Apr 2023', etc.
    Returns first day of that month as datetime, else None.
    """
    # Try month name parsing
    try:
        # If header is already a date
        if isinstance(header_value, datetime):
            return header_value.replace(day=1)
        header_str = str(header_value).strip()
        # Try known formats
        for fmt in ['%B %Y', '%b %Y', '%m/%Y', '%Y-%m', '%d/%m/%Y']:
            try:
                dt = datetime.strptime(header_str, fmt)
                return dt.replace(day=1)
            except Exception:
                pass
        # Fallback: extract month/year
        import re
        mo = re.match(r'(\w+)\s+(\d{4})', header_str)
        if mo:
            monthname, year = mo.groups()
            month = datetime.strptime(monthname, '%B').month if len(monthname)>3 else datetime.strptime(monthname, '%b').month
            return datetime(int(year), month, 1)
        # Try numeric month/year
        mo = re.match(r'(\d{1,2})[/-](\d{4})', header_str)
        if mo:
            month, year = mo.groups()
            return datetime(int(year), int(month), 1)
    except Exception:
        pass
    return None

for row in range(9, 19):  # U9-U18
    last_col = None
    for col in col_range[::-1]:
        val = ws[f'{col}{row}'].value
        if isinstance(val, (int, float)) and val > 0:
            last_col = col
            break
    if last_col:
        header_value = ws[f'{last_col}{header_row}'].value
        date_obj = parse_month_header(header_value)
        ws[f'{output_col}{row}'].value = date_obj
        if date_obj:
            ws[f'{output_col}{row}'].number_format = 'DD/MM/YYYY'
    else:
        ws[f'{output_col}{row}'].value = None

wb.save(output_path)
