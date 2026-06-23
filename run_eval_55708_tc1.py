import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/eval_seed42/eval_55708_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/eval_seed42/eval_55708_tc1/output.xlsx'

# Read the spreadsheet
wb = load_workbook(input_path)
ws = wb.active

# Read dataframe for easier filtering
df = pd.read_excel(input_path)

# Columns
# ['Department', 'Status', 'turnaroud time']
department_col = 'Department'
status_col = 'Status'
tat_col = 'turnaroud time'

# Departments are in A11, A12, A13 for rows 11,12,13 (1-based)
department_cells = ['A11', 'A12', 'A13']
output_cells = ['B11', 'B12', 'B13']
departments = [ws[cell].value for cell in department_cells]

for idx, dept in enumerate(departments):
    mask = (
        (df[department_col] == dept) &
        (df[status_col].isin(['In Progress', 'In Review'])) &
        (df[tat_col] >= 6)
    )
    filtered = df[mask]
    cell = output_cells[idx]
    if not filtered.empty:
        ws[cell] = filtered[tat_col].mean()
    else:
        ws[cell] = None  # blank cell

wb.save(output_path)
