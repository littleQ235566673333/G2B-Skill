import openpyxl

# Input and output file paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_40757_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_40757_tc1/output.xlsx'

# Load the workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Identify month to sum for (for demonstration, pick the first month in the data)
def try_parse_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0

months = [cell.value for cell in ws['A'][1:]] # Skipping header
months = [m for m in months if m is not None]
if months:
    target_month = months[0]
else:
    target_month = None

# Sum values in column B for the target month, coerce to float if possible, ignoring None and non-numeric
values = [try_parse_float(row[1].value) for row in ws.iter_rows(min_row=2) if row[0].value == target_month and row[1].value is not None]
sum_val = sum(values)

ws['B10'] = sum_val if sum_val != 0 else None

# For another demonstration: show for another month in B11, but leave blank if sum is zero
other_month = None
for m in set(months):
    if m != target_month:
        other_month = m
        break
if other_month:
    other_values = [try_parse_float(row[1].value) for row in ws.iter_rows(min_row=2) if row[0].value == other_month and row[1].value is not None]
    other_sum = sum(other_values)
    ws['B11'] = other_sum if other_sum != 0 else ''
else:
    ws['B11'] = ''

wb.save(output_path)
