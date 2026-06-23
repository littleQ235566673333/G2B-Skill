import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42/eval_267-21_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42/eval_267-21_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
merge_ws = wb['merging']
sh1_ws = wb['SH1']
rp1_ws = wb['RP1']

def make_lookup_dict(ws, id_col, val_col):
    lookup = {}
    for row in ws.iter_rows(min_row=2):
        id_val = row[id_col-1].value
        qty_val = row[val_col-1].value
        if id_val is not None and str(id_val) != '':
            lookup[id_val] = qty_val
    return lookup

sh1_lookup = make_lookup_dict(sh1_ws, 2, 3)  # ID in col 2, QTY in col 3
rp1_lookup = make_lookup_dict(rp1_ws, 2, 3)

for row in range(2, 12):
    id_val = merge_ws.cell(row=row, column=2).value
    val_sh1 = sh1_lookup.get(id_val, '-')
    if val_sh1 is None or str(val_sh1) == '':
        val_sh1 = '-'
    val_rp1 = rp1_lookup.get(id_val, '-')
    if val_rp1 is None or str(val_rp1) == '':
        val_rp1 = '-'
    merge_ws.cell(row=row, column=3).value = val_sh1
    merge_ws.cell(row=row, column=4).value = val_rp1

wb.save(output_path)
