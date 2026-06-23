import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_267-21_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_267-21_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_merge = wb['merging']
ws_sh1 = wb['SH1']
ws_rp1 = wb['RP1']

def get_qty_map(sheet):
    id_col = None
    qty_col = None
    # Scan header to find ID and QTY columns
    for col in range(1, sheet.max_column + 1):
        hdr = str(sheet.cell(row=1, column=col).value).strip().upper()
        if hdr == 'ID':
            id_col = col
        elif hdr == 'QTY':
            qty_col = col
    result = {}
    if id_col and qty_col:
        for row in range(2, sheet.max_row + 1):
            id_val = sheet.cell(row=row, column=id_col).value
            qty_val = sheet.cell(row=row, column=qty_col).value
            k = str(id_val).strip() if id_val is not None else None
            v = str(qty_val).strip() if qty_val is not None else '-' 
            if k:
                # If qty blank, store '-'
                result[k] = v if v else '-'
    return result

qty_map_sh1 = get_qty_map(ws_sh1)
qty_map_rp1 = get_qty_map(ws_rp1)

# Fill merging sheet C2:D11
for row in range(2, 12):
    id_cell = ws_merge.cell(row=row, column=2) # Column B
    id_val = str(id_cell.value).strip() if id_cell.value is not None else None

    # Sheet SH1 for col C
    c_val = qty_map_sh1.get(id_val, '-') if id_val else '-'
    ws_merge.cell(row=row, column=3, value=c_val)

    # Sheet RP1 for col D
    d_val = qty_map_rp1.get(id_val, '-') if id_val else '-'
    ws_merge.cell(row=row, column=4, value=d_val)

wb.save(output_path)