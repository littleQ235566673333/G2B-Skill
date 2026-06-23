import openpyxl

# Input/output files
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_38823_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_38823_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Define columns and rows
DATE_COL = 'C'
VALUE_COL = 'E'
TERM_COL = 'G'
OUTPUT_START_ROW, OUTPUT_END_ROW = 4, 7
SEARCH_TERM_COL = 'H'
OUTPUT_COL = 'I'

# Get date range from B2:B3
start_date = ws['B2'].value
end_date = ws['B3'].value

# Read the search terms from H4:H7
search_terms = [ws[f'H{row}'].value for row in range(OUTPUT_START_ROW, OUTPUT_END_ROW + 1)]

# Find how many data rows exist (stop at first empty cell in C)
data_rows = []
row = 4
while ws[f'{DATE_COL}{row}'].value is not None:
    data_rows.append(row)
    row += 1

dates  = [ws[f'{DATE_COL}{r}'].value for r in data_rows]
values = [ws[f'{VALUE_COL}{r}'].value for r in data_rows]
terms  = [ws[f'{TERM_COL}{r}'].value for r in data_rows]

# For each search term, calculate sum for values whose date is in range and whose term includes the search term
for idx, search_term in enumerate(search_terms):
    total = 0
    for d, v, t in zip(dates, values, terms):
        # Only consider valid records
        if t and search_term and (search_term in str(t)) and (d is not None) and (start_date is not None) and (end_date is not None):
            if start_date <= d <= end_date:
                try:
                    val = float(v)
                except (TypeError, ValueError):
                    continue
                total += val
    ws[f'{OUTPUT_COL}{OUTPUT_START_ROW + idx}'] = total

wb.save(output_path)
