import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/group_56274/r0/evolve_56274/input.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

print('D7 (Fiscal Month):', ws['D7'].value)
print('D9 -> D12:')
for row in range(9, 13):
    print(f'D{row}:', ws[f"D{row}"].value)

print('\nA1:H20 snapshot:')
for r in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=8):
    print([cell.value for cell in r])
