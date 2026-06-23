import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_12864_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_12864_tc1/output.xlsx'

# Load data
sheet1 = pd.read_excel(input_path, sheet_name='Sheet1')
sheet2 = pd.read_excel(input_path, sheet_name='Sheet2')

# Look up values from Sheet1's 'Existing Data Fields' in Sheet2's 'Existing Data Fields'
lookup_values = sheet1['Existing Data Fields'].dropna().tolist()
results = []
for key in lookup_values:
    matches = sheet2[sheet2['Existing Data Fields'] == key]
    if not matches.empty:
        # If multiple matches, take the first one
        valb = matches.iloc[0]['Existing Data Fields']
        valc = matches.iloc[0]['deal name']
        results.append((valb, valc))
    else:
        results.append((None, None))
# Pad output for B2:B12, C2:C12
while len(results) < 11:
    results.append((None, None))

# Write back to the output file
wb = load_workbook(input_path)
ws = wb['Sheet2']
for idx, (valb, valc) in enumerate(results[:11], start=2):
    ws[f'B{idx}'] = valb
    ws[f'C{idx}'] = valc
wb.save(output_path)
