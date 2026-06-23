import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_12864_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_12864_tc1/output.xlsx'

# Load the workbook and relevant sheets
wb = openpyxl.load_workbook(input_path)
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']

# Build a lookup dictionary from Sheet2's Column D (target to match) to (B, C)
sheet2_data = {}
for row in sheet2.iter_rows(min_row=2, max_col=3, values_only=True):
    key = row[1]  # Column C of Sheet2 (index 1 for B, 2 for C). We need to match Sheet1 D with Sheet2 C data? Let's confirm.
    # But, user says match Column D Sheet1 to Column B Sheet2, so we index correctly:
    lookup_key = row[0] # Column B value which is at index 0
    if lookup_key is not None:
        sheet2_data[lookup_key] = (row[0], row[1])

# Match values from Sheet1 D and find corresponding data in Sheet2
results = []
for row in sheet1.iter_rows(min_row=2, max_col=4, min_col=4, max_row=12, values_only=True):
    match_value = row[0] # Value from Sheet1 Column D
    result = sheet2_data.get(match_value, (None, None))
    results.append(result)

# Write results into Sheet2 B2:B12 (as per instruction)
for idx, (date_val, associated_val) in enumerate(results, start=2):
    # Writing a tuple as a string (date, associated data), one per cell
    sheet2[f'B{idx}'] = f"{date_val}, {associated_val}" if date_val is not None else ''

wb.save(output_path)
