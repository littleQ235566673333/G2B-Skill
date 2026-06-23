from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import datetime

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_2/regression_gate/before_pass/core_32337/input.xlsx')
ws = wb.active
header_row = 2
headers = {ws.cell(row=header_row, column=col).value: col for col in range(1, ws.max_column+1)}
dob_col = headers.get('DATE OF BIRTH')
report_col = headers.get('REPORT DATE')
age_year_col = headers.get('age (year)')
category_col = headers.get('CATEGORY')
start_row, end_row = 3, 15
actual_age_col = age_year_col + 2
ws.cell(row=header_row, column=actual_age_col, value='Actual Age')
prev_fill = ws.cell(row=header_row, column=actual_age_col-1).fill
for row in range(start_row, end_row + 1):
    dob = ws.cell(row=row, column=dob_col).value
    report_date = ws.cell(row=row, column=report_col).value
    if isinstance(dob, datetime.date) and isinstance(report_date, datetime.date):
        age = report_date.year - dob.year - ((report_date.month, report_date.day) < (dob.month, dob.day))
    else:
        age = None
    cell = ws.cell(row=row, column=actual_age_col, value=age)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = prev_fill
    if age is not None:
        cell.number_format = '0'
header_cell = ws.cell(row=header_row, column=actual_age_col)
header_cell.alignment = Alignment(horizontal='center', vertical='top')
header_cell.fill = prev_fill
header_cell.font = Font(header_cell.font.name, header_cell.font.size, bold=True)
year_age_header_col = None
category_map_header_col = None
for col in range(1, ws.max_column+1):
    header = ws.cell(row=header_row, column=col).value
    if header == 'YEAR (age)':
        year_age_header_col = col
    if header == 'CATEGORY':
        category_map_header_col = col
max_mapping_row = header_row + 1
while ws.cell(row=max_mapping_row, column=year_age_header_col).value not in (None, ''):
    max_mapping_row += 1
max_mapping_row -= 1
def col_letter(colidx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(colidx)
year_age_range = f"${col_letter(year_age_header_col)}${header_row+1}:${col_letter(year_age_header_col)}${max_mapping_row}"
category_range = f"${col_letter(category_map_header_col)}${header_row+1}:${col_letter(category_map_header_col)}${max_mapping_row}"
for row in range(start_row, end_row + 1):
    age_val_cell = ws.cell(row=row, column=age_year_col).coordinate
    formula = f'=IFERROR(INDEX({category_range}, MATCH({age_val_cell}, {year_age_range}, 0)), "")'
    ws.cell(row=row, column=5, value=formula)
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_2/regression_gate/before_pass/core_32337/output.xlsx')
