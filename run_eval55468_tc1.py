import openpyxl

# Load workbook and sheet
infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_55468_tc1/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_55468_tc1/output.xlsx'
wb = openpyxl.load_workbook(infile)
ws = wb['Sheet1']

# Criteria
h1_criteria = ws['AC4'].value # horizontal 1 (row 4)
h2_criteria = ws['AB4'].value # horizontal 2 (row 3)
v1_criteria = ws['AE4'].value # vertical 1 (col A)
v2_criteria = ws['AD4'].value # vertical 2 (col B)

# Find row (A5:A10, B5:B10)
row_found = None
for row in range(5, 11):
    if ws[f'A{row}'].value == v1_criteria and ws[f'B{row}'].value == v2_criteria:
        row_found = row
        break

# Find column (C4:Z4 and C3:Z3)
col_found = None
for col in range(3, 27):  # C is 3, Z is 26
    col_letter = openpyxl.utils.get_column_letter(col)
    if ws[f'{col_letter}4'].value == h1_criteria and ws[f'{col_letter}3'].value == h2_criteria:
        col_found = col
        break

# Value at intersection or blank
result = None
if row_found and col_found:
    cell_val = ws.cell(row=row_found, column=col_found).value
    result = cell_val

# Write to AE5
ws['AE5'] = result
wb.save(outfile)
