import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_8/regression_gate/before_fix/core_54474/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_8/regression_gate/before_fix/core_54474/output.xlsx'

wb = openpyxl.load_workbook(input_path)
whp_ws = wb['WHP']
data_ws = wb['WHP DATA']

data_map = {}
header = None
# Identify header row
for i, row in enumerate(data_ws.iter_rows(values_only=True), start=1):
    if 'EEG' in row:
        header = [str(cell) if cell is not None else '' for cell in row]
        continue
    if header and i > 2 and any(row):
        site = row[1]
        if site:
            data_map[site] = {header[j]: row[j] for j in range(len(header))}

# Sites for E7 and E8
sites = [data_ws['B4'].value, data_ws['B5'].value]
columns = ['EEG', 'Dis', 'LTU']

for idx, site in enumerate(sites):
    if site in data_map:
        for col_idx, col_name in enumerate(columns, start=5):
            whp_ws.cell(row=7+idx, column=col_idx, value=data_map[site][col_name])

wb.save(output_path)
print('Done')
