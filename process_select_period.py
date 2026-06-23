import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# Load the workbook and sheet
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_59224_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_59224_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get the project start date from B2
def excel_date_to_datetime(value):
    if isinstance(value, datetime):
        return value
    return datetime.strptime(str(value), '%Y-%m-%d')

project_start = ws['B2'].value
if not isinstance(project_start, datetime):
    project_start = excel_date_to_datetime(project_start)

# E4:E14
first_select_period_row = None
for row in range(4, 15):
    date_c = ws[f'C{row}'].value
    date_d = ws[f'D{row}'].value
    if not isinstance(date_c, datetime):
        date_c = excel_date_to_datetime(date_c)
    if not isinstance(date_d, datetime):
        date_d = excel_date_to_datetime(date_d)

    if project_start > date_c and project_start < date_d:
        ws[f'E{row}'] = 'Select Period'
        if first_select_period_row is None:
            first_select_period_row = row
    else:
        ws[f'E{row}'] = 'Do Not Select'

# Fill cells above with "Select Period"
if first_select_period_row is not None:
    for row in range(4, first_select_period_row):
        ws[f'E{row}'] = 'Select Period'

wb.save(output_path)
