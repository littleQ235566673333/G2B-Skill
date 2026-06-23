import openpyxl
from datetime import datetime

def parse_month(cell_value):
    """Parse month from string like 'January 29' or similar, or datetime object."""
    if isinstance(cell_value, datetime):
        return cell_value.month
    try:
        return datetime.strptime(str(cell_value), '%B %d').month
    except:
        pass
    try:
        return datetime.strptime(str(cell_value), '%B %d, %Y').month
    except:
        pass
    try:
        return datetime.strptime(str(cell_value), '%Y-%m-%d').month
    except:
        pass
    try:
        return datetime.strptime(str(cell_value), '%m/%d/%Y').month
    except:
        pass
    return None

def main():
    input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_8942_tc1/input.xlsx'
    output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_8942_tc1/output.xlsx'
    
    wb = openpyxl.load_workbook(input_path)
    overview = wb['Overview']
    paydates = wb['Pay Dates']
    
    # Get selected month from Overview!A2
    overview_month_val = overview['A2'].value
    selected_month = parse_month(overview_month_val)
    
    # Identify date and amount/pay columns in Pay Dates
    header_row = next(paydates.iter_rows(min_row=1, max_row=1))
    date_col = None
    pay_col = None
    for idx, cell in enumerate(header_row):
        h = cell.value
        if not h:
            continue
        hc = str(h).lower()
        if 'date' in hc:
            date_col = idx
        if 'pay' in hc or 'amount' in hc:
            pay_col = idx
    
    if date_col is None or pay_col is None or selected_month is None:
        overview['B5'].value = None
        wb.save(output_path)
        return
    
    total = 0.0
    for row in paydates.iter_rows(min_row=2):
        date_cell = row[date_col].value
        pay_cell = row[pay_col].value
        # Parse date
        date_obj = None
        if isinstance(date_cell, datetime):
            date_obj = date_cell
        else:
            try:
                date_obj = datetime.strptime(str(date_cell), '%Y-%m-%d')
            except:
                try:
                    date_obj = datetime.strptime(str(date_cell), '%m/%d/%Y')
                except:
                    try:
                        date_obj = datetime.strptime(str(date_cell), '%B %d, %Y')
                    except:
                        pass
        if date_obj is not None and date_obj.month == selected_month:
            try:
                total += float(pay_cell)
            except:
                continue
    overview['B5'].value = total
    wb.save(output_path)

if __name__ == '__main__':
    main()
