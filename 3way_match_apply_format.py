import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_7/group_57033/r0/evolve_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_7/group_57033/r0/evolve_57033/output.xlsx'

# Read data for matching
cb_df = pd.read_excel(input_path, sheet_name='CBtrans')
s4_df = pd.read_excel(input_path, sheet_name='Sheet4')

# Load sheet for openpyxl edits
wb = load_workbook(input_path)
ws = wb['Sheet4']

# Prepare the fill style
fill = PatternFill(fill_type='solid', fgColor='FF66CC')

for i in range(2, 8):  # K2:K7
    row_idx = i - 2  # DataFrame zero index for first 6 rows
    row = s4_df.iloc[row_idx]
    # 3-way match (case-insensitive for company fields)
    match = ((cb_df['company'].astype(str).str.lower() == str(row['Company']).lower()) &
             (cb_df['account'] == row['account']) &
             (cb_df['xchar'] == row['xchar'])).any()
    value = 'Match' if match else '-'
    cell = ws[f'K{i}']
    cell.value = value
    cell.fill = fill
    cell.value = cell.value.capitalize() if cell.value != '-' else '-'

wb.save(output_path)
