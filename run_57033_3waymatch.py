import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_3/group_57033/r2/evolve_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_3/group_57033/r2/evolve_57033/output.xlsx'

# Load data
sheet4 = pd.read_excel(input_path, sheet_name='Sheet4')
cbtrans = pd.read_excel(input_path, sheet_name='CBtrans')

# Helper: regular casing
def regular_case(x):
    if pd.isnull(x):
        return x
    x = str(x)
    return x[:1].upper() + x[1:].lower() if x.isalpha() else x

# Prepare columns for matching - case insensitive, numeric safe, xchar as string
def safe_str(x):
    if pd.isnull(x):
        return ''
    return str(x).strip().lower()

def safe_num(x):
    try:
        return float(x)
    except Exception:
        return None

# Matching logic: for each row in Sheet4, does (company, account, xchar) triple match in CBtrans?
result = []
for i, row in sheet4.iloc[:6].iterrows():
    company4 = safe_num(row['Company'])
    account4 = safe_str(row['account'])
    xchar4 = safe_str(row['xchar'])

    mask = (
        cbtrans['company'].apply(safe_num) == company4
    ) & (
        cbtrans['account'].apply(safe_str) == account4
    ) & (
        cbtrans['xchar'].apply(safe_str) == xchar4
    )

    match = 'Match' if mask.any() else '-'
    result.append(match)

# Write results with correct casing
def case_result(val):
    return 'Match' if val == 'Match' else '-'

result_reg = [case_result(v) for v in result]

# Write results to Excel
wb = load_workbook(input_path)
ws = wb['Sheet4']

fill = PatternFill(start_color='FF66CC', end_color='FF66CC', fill_type='solid')

for idx, value in enumerate(result_reg, start=2):
    cell = ws[f'K{idx}']
    cell.value = value
    cell.fill = fill

wb.save(output_path)
