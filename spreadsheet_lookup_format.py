import openpyxl
from openpyxl.styles import PatternFill, Font

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_48620_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_48620_tc1/output.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Find range for lookup (A:B), queries in D2:D7
output_col = 'E'
query_cells = ['D2', 'D3', 'D4', 'D5', 'D6', 'D7']
lookup_range = list(ws.iter_rows(min_row=2, max_col=2, values_only=True))

results = []
for cell in query_cells:
    query = ws[cell].value
    match = None
    for prod, color in lookup_range:
        if prod == query:
            match = color
            break
    results.append(match if match is not None else 0)

# Write results to E2:E7, turning 0s into blanks
for i, value in enumerate(results, start=2):
    ws[f'E{i}'] = '' if value == 0 else value

# Styling
fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
font = Font(name='Calibri')
for i in range(2, 8):
    cell = ws[f'E{i}']
    cell.fill = fill
    cell.font = font

wb.save(output_path)
