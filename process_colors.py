import openpyxl
import re

# Load the workbook and relevant sheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_40892_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_40892_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read colors list from column D (D2:Dwhatever, assuming contiguous)
color_cells = ws['D']
color_names = [cell.value for cell in color_cells if cell.row >= 2 and cell.value]

# Pre-compile regex for each color for efficiency, enforcing word boundaries (case-insensitive)
color_patterns = [(color, re.compile(r'\\b' + re.escape(str(color)) + r'\\b', re.IGNORECASE)) for color in color_names]

# Process each entry in column A (A2:A17)
for row in range(2, 18):
    cell_value = ws[f'A{row}'].value
    output_color = ''
    if cell_value:
        for color, pattern in color_patterns:
            if pattern.search(cell_value):
                output_color = color
                break
    ws[f'B{row}'] = output_color

# Save result
wb.save(output_path)
