from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_387-16_s0/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_387-16_s0/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Read Value/Binaries pairs from A3:B down to last non-empty in A or B
pairs = []
row = 3
while True:
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    if a is None and b is None:
        break
    pairs.append([a, b])
    row += 1

# Read result values from D5 downward until blank
result_values = []
row = 5
while True:
    v = ws.cell(row=row, column=4).value
    if v is None:
        break
    result_values.append(v)
    row += 1

# Remove only one matching instance for each result value
for rv in result_values:
    for i, (val, binv) in enumerate(pairs):
        if val == rv:
            pairs.pop(i)
            break

# Write compacted remaining pairs back starting at row 3
start_row = 3
for idx, (val, binv) in enumerate(pairs, start=start_row):
    ws.cell(row=idx, column=1).value = val
    ws.cell(row=idx, column=2).value = binv

# Clear leftover cells through original range
for r in range(start_row + len(pairs), start_row + 19):
    ws.cell(row=r, column=1).value = None
    ws.cell(row=r, column=2).value = None

# Compute solver result as sum of column A remaining values
solver_result = sum(val for val, _ in pairs if isinstance(val, (int, float)))
target_value = ws['E2'].value
ws['H2'] = solver_result
ws['K2'] = solver_result - target_value

wb.save(output_path)

# Verification print for checked range
wb2 = load_workbook(output_path, data_only=False)
ws2 = wb2['Sheet1']
for r in range(2, 19):
    print(r, [ws2.cell(r, c).value for c in range(1, 5)])
print('H2', ws2['H2'].value, 'K2', ws2['K2'].value)
