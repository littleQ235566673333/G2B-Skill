import pandas as pd
from openpyxl import load_workbook
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_9391_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_9391_tc1/output.xlsx'
df_data = pd.read_excel(input_path, sheet_name='Data')
wb = load_workbook(input_path)
ws_front = wb['Front']
target_date = ws_front['B1'].value
results = []
for row in range(2, 13):
    agent = ws_front.cell(row=row, column=3).value
    # ensure correct match, handle possible dt/string
    data_rows = df_data[df_data.iloc[:,0] == target_date]
    if not data_rows.empty:
        match_rows = data_rows[data_rows.iloc[:,1] == agent]
        if not match_rows.empty:
            val = match_rows.iloc[0,5]
        else:
            val = ''
    else:
        val = ''
    results.append(val)
for i, val in enumerate(results):
    ws_front.cell(row=2 + i, column=2).value = val
wb.save(output_path)
