import openpyxl

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_6/group_247-24/r1/evolve_247-24/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_6/group_247-24/r1/evolve_247-24/output.xlsx'

wb = openpyxl.load_workbook(INPUT)
main_ws = wb['Main']
lookup_ws = wb['Lookup']

# Map header to col index
main_header = [c.value for c in next(main_ws.iter_rows(min_row=1,max_row=1))]
header_to_col = {name: idx+1 for idx, name in enumerate(main_header)}

# Build main data as list of dict
main_data = []
for row in main_ws.iter_rows(min_row=2, max_row=main_ws.max_row, max_col=len(main_header), values_only=True):
    if all([cell is None for cell in row]):
        continue
    entry = dict(zip(main_header, row))
    main_data.append(entry)

# Build lookup dict
lookup_header = [c.value for c in next(lookup_ws.iter_rows(min_row=1,max_row=1))]
empid_col, weeks_col = lookup_header.index('Employee ID'), lookup_header.index('Working Weeks')
lookup_dict = {}
for row in lookup_ws.iter_rows(min_row=2, max_row=lookup_ws.max_row, max_col=2, values_only=True):
    emp_id = row[0]
    weeks = row[1]
    lookup_dict[str(emp_id)] = weeks

# 1 & 2: Row deletions
new_rows = []
for entry in main_data:
    if entry['Company'] == 'Motorcycle':
        continue # delete
    if entry['Company'] == 'Ahmed Sons' and entry['Location'] == 'Canada':
        continue # delete
    new_rows.append(entry)
# 3: Change bill rate
for entry in new_rows:
    if entry['Company'] == 'National TV' and entry['Location'] == 'India':
        entry['Bill Rate of Resource'] = 180
# 4: Insert 2 new rows after every Ahmed Sons
main_entries = []
for entry in new_rows:
    main_entries.append(entry)
    if entry['Company'] == 'Ahmed Sons':
        # Add two blank rows, same structure
        main_entries.append({h: None for h in main_header})
        main_entries.append({h: None for h in main_header})
# 5: VLOOKUP
for entry in main_entries:
    empid_col_header = 'Employee ID'
    if entry.get(empid_col_header):
        key = str(entry[empid_col_header])
        entry['Lookup For Working Weeks'] = lookup_dict.get(key, None)
    else:
        entry['Lookup For Working Weeks'] = None

# Write results to Main A2:M70
for i in range(69):
    row_data = main_entries[i] if i < len(main_entries) else {h: None for h in main_header}
    for j, h in enumerate(main_header):
        main_ws.cell(row=2 + i, column=j + 1, value=row_data[h])
# Clear left-over rows if any beyond what we've written
for i in range(2 + len(main_entries), 2 + 69):
    for j in range(1, len(main_header) + 1):
        main_ws.cell(row=i, column=j, value=None)

wb.save(OUTPUT)
