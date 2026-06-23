from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_5/regression_gate/before_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_5/regression_gate/before_pass/core_41601/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Students']

for row in range(2, 8):
    student = ws.cell(row=row, column=1).value
    if student:
        formula = f"=INDIRECT('{student}!C2')"
        ws.cell(row=row, column=5, value=formula)

wb.save(output_path)
print('Wrote formulas to E2:E7.')
