from openpyxl import load_workbook
# Load workbook
inp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_5/regression_gate/before_fix/core_56274/input.xlsx'
outp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_5/regression_gate/before_fix/core_56274/output.xlsx'
wb = load_workbook(inp)
ws = wb.active
#----
# Find where fiscal months and balances exist
months_row = None
for r in ws.iter_rows(min_row=1, max_row=30, max_col=10):
    for c in r:
        if c.value and str(c.value).lower() == 'fiscal month':
            months_row = c.row
            break
    if months_row: break
# Map labels correctly
label_map = {
    'Opening Balance': 'Opening Balance',
    'Total Debits': 'Debits',
    'Total Credits': 'Credits',
    'Closing Balance': 'Closing Balance',
}
if months_row:
    n_cols = 12
    # Find rows for balance labels
    bal_rows = {}
    for rr in range(months_row+2, months_row+10):
        label = ws.cell(row=rr, column=1).value
        if label and str(label).strip() in label_map:
            bal_rows[label_map[str(label).strip()]] = rr
    # Output locations
    out_cells = ['D9', 'D10', 'D11', 'D12']
    col_start = ws.cell(row=months_row+1, column=2).coordinate
    col_end = ws.cell(row=months_row+1, column=2+n_cols-1).coordinate
    months_range = f'${col_start}:${col_end}'
    bal_labels = ['Opening Balance', 'Debits', 'Credits', 'Closing Balance']
    for idx, label in enumerate(bal_labels):
        row = bal_rows[label]
        data_range = f'$B${row}:$M${row}'
        formula = f'=INDEX({data_range}, MATCH($D$7,{months_range},0))'
        ws[out_cells[idx]] = formula
wb.save(outp)
