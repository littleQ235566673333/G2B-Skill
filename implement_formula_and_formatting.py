import openpyxl
from openpyxl.styles import Alignment

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_56378_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_56378_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Header in row 4, data from row 5 down, Frame 1 is columns C:I (index 2:9)
header_row = 4
data_start_row = 5
data_end_row = 20
frame1_cols = list(range(2, 9)) # Columns C to I (0-based)
frame2_start_col = 12  # L

# Prepare output header (to L5:R5)
for idx, col_idx in enumerate(frame1_cols):
    ws.cell(row=5, column=frame2_start_col + idx, value=ws.cell(row=header_row, column=col_idx + 1).value)

# Gather valid products (non-empty QUANTITY UNITS, I column = index 8)
filtered_rows = []
for r in range(data_start_row, data_end_row+1):
    quantity_val = ws.cell(row=r, column=9).value  # I
    if quantity_val is not None and quantity_val != '':
        filtered_rows.append([ws.cell(row=r, column=col+1).value for col in frame1_cols])

# Only 3 data rows to fill 5:8 (L6:R8)
for row_offset, values in enumerate(filtered_rows[:3]):
    for col_offset, val in enumerate(values):
        ws.cell(row=6 + row_offset, column=frame2_start_col + col_offset, value=val)
        # Alignment: L is left, M:R are right
        if col_offset == 1:  # PRODUCT name (column L+1, i.e., M)
            ws.cell(row=6 + row_offset, column=frame2_start_col + col_offset).alignment = Alignment(horizontal='left')
        else:
            ws.cell(row=6 + row_offset, column=frame2_start_col + col_offset).alignment = Alignment(horizontal='right')
# Also left-align the header for PRODUCT, right for rest
for col_offset in range(len(frame1_cols)):
    align = Alignment(horizontal='left') if col_offset == 1 else Alignment(horizontal='right')
    ws.cell(row=5, column=frame2_start_col + col_offset).alignment = align

wb.save(output_path)
print('Done')
