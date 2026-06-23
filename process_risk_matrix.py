import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_50768_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_50768_tc1/output.xlsx'

# Load workbook and active sheet
wb = openpyxl.load_workbook(input_path)
sheet = wb.active

# Locate the header row and columns (Risk ID, Impact, Likelihood)
header_row = None
col_risk = col_impact = col_likelihood = None
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    for cell in row:
        if cell.value == 'Risk ID':
            header_row = cell.row
            col_risk = cell.column
        elif cell.value == 'Impact':
            header_row = cell.row
            col_impact = cell.column
        elif cell.value == 'Likelihood':
            header_row = cell.row
            col_likelihood = cell.column
    if header_row:
        break

if not (header_row and col_risk and col_impact and col_likelihood):
    raise Exception('Header row or required columns not found.')

# Build the matrix: Impact 1-3 (rows, 1=bottom), Likelihood 1-3 (cols, 1=left)
risk_matrix = {(impact, likelihood): [] for impact in range(1,4) for likelihood in range(1,4)}
for row in sheet.iter_rows(min_row=header_row+1, max_row=sheet.max_row):
    risk_id = row[col_risk-1].value
    impact = row[col_impact-1].value
    likelihood = row[col_likelihood-1].value
    try:
        impact = int(impact)
        likelihood = int(likelihood)
        if impact in (1,2,3) and likelihood in (1,2,3):
            risk_matrix[(impact, likelihood)].append(str(risk_id))
    except:
        continue

# Matrix output area: F12:H14 (Impact: 3,2,1 = rows 12,13,14; Likelihood: 1-3 = F-H)
row_base = 14   # Impact=1 is row 14, Impact=2 is row 13, Impact=3 is row 12
col_base = 6    # Likelihood=1 is column 6 (F), 2 is 7 (G), 3 is 8 (H)
for impact in range(1,4):
    for likelihood in range(1,4):
        risks = ','.join(risk_matrix[(impact, likelihood)])
        sheet.cell(row=row_base-(impact-1), column=col_base+(likelihood-1)).value = risks

wb.save(output_path)
print('Matrix written to F12:H14 in output.xlsx')
