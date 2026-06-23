import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_51354_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_51354_tc1/output.xlsx'

month_map = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
}
rev_month_map = {v: k for k, v in month_map.items()}

wb = load_workbook(input_path)
ws = wb['Sheet1']

fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

def add_one_month(month, year):
    if month == 12:
        return 1, year + 1
    else:
        return month + 1, year

for row in range(2, 7):
    cell_val = ws[f'A{row}'].value
    m = re.search(r'([A-Za-z]{3})\s?(\d{2})$', str(cell_val))
    if m:
        mon_abbr = m.group(1)
        yr = int(m.group(2))
        ws[f'D{row}'] = yr
        ws[f'D{row}'].fill = fill
        mon_num = month_map.get(mon_abbr[:3], None)
        if mon_num is not None:
            new_mon, new_yr = add_one_month(mon_num, 2000+yr if yr < 100 else yr)
            # Format back to short month and last two digits of year
            result = f'{rev_month_map[new_mon]} {str(new_yr)[-2:]}'
            ws[f'E{row}'] = result
        else:
            ws[f'E{row}'] = ''
    else:
        ws[f'D{row}'] = ''
        ws[f'E{row}'] = ''

wb.save(output_path)
