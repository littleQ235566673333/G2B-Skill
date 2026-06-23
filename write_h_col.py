import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun2/eval_56419_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun2/eval_56419_tc1/output.xlsx'

# Load both pandas df and openpyxl workbook
try:
    df = pd.read_excel(input_path)
    wb = load_workbook(input_path)
    ws = wb.active
except Exception as e:
    raise RuntimeError(f'Error loading input file: {e}')

# Find required columns
quantity_col = None
type_col = None
for col in df.columns:
    if col.strip().lower() == 'quantity':
        quantity_col = col
    if col.strip().lower() == 'type':
        type_col = col
if quantity_col is None or type_col is None:
    raise ValueError('Missing required Quantity or Type column')

# Collect non-zero values and associated type
try:
    nonzero_types = df.loc[df[quantity_col].fillna(0)!=0, type_col].fillna('').tolist()
except Exception as e:
    nonzero_types = []

# Pad or trim to 26 entries (H2:H27)
h_list = list(nonzero_types)[:26] + ['']*(26-len(nonzero_types))

# Set green background
fill = PatternFill('solid', fgColor='92D050')
for i, val in enumerate(h_list):
    cell = ws.cell(row=2+i, column=8)  # H = col 8
    cell.value = val
    cell.fill = fill

wb.save(output_path)
