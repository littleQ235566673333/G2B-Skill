from openpyxl import load_workbook
from datetime import datetime, timedelta

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_3/regression_gate/before_pass/core_41589/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_3/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(in_path)
ws = wb['Contact List']
today = datetime.today()

for row in ws.iter_rows(min_row=4, min_col=8, max_col=9):
    h_cell, i_cell = row
    result = 'NO ACTION'

    contact_date = None
    if h_cell.value and isinstance(h_cell.value, datetime):
        contact_date = h_cell.value
    elif h_cell.value and isinstance(h_cell.value, str):
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d-%b-%y'):
            try:
                contact_date = datetime.strptime(h_cell.value, fmt)
                break
            except ValueError:
                continue

    i_value = str(i_cell.value).strip().upper() if i_cell.value else ''

    if contact_date and i_value == 'YES':
        days_diff = (today - contact_date).days
        if days_diff <= 30:
            result = 'HOLD'
        else:
            result = 'TOUCH BASE'
    ws.cell(row=h_cell.row, column=10).value = result

wb.save(out_path)
