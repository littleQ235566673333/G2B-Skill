import openpyxl
from openpyxl.utils import get_column_letter

VOWELS = set('AEIOU')

def is_match(word1, word2):
    if len(word1) != len(word2):
        return False
    s_indices = [i for i, c in enumerate(word1) if c == 'S']
    if not s_indices:
        return False
    vowel = None
    for i, c in enumerate(word1):
        if c == 'S':
            if word2[i] not in VOWELS:
                return False
            if vowel is None:
                vowel = word2[i]
            elif vowel != word2[i]:
                return False
        else:
            if word2[i] != c:
                return False
    return vowel is not None

file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/group_146-49/r3/evolve_146-49/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/group_146-49/r3/evolve_146-49/output.xlsx'
wb = openpyxl.load_workbook(file)
ws = wb['Sheet1']

# Read all words in col A
words = set()
for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
    if row[0]:
        words.add(row[0].strip().upper())
words = sorted(words)

used = set()
pairs = []
for w1 in words:
    for w2 in words:
        if w1 == w2 or (w2,w1) in used:
            continue
        if is_match(w1, w2):
            pairs.append((w1, w2))
            used.add((w1, w2))

max_len = 65
if len(pairs) > max_len:
    pairs = pairs[:max_len]

# Write pairs to G2:H2 onwards
for idx in range(1, 1+max_len):
    ws[f'G{idx}'] = None
    ws[f'H{idx}'] = None
for idx, (a, b) in enumerate(pairs):
    ws[f'G{idx+2}'] = a
    ws[f'H{idx+2}'] = b

# Sort G2:H* as described (must move up to G1:H1)
# Only sort the actual found pairs
pairs_sorted = sorted(pairs, key=lambda p: (p[0].lower(), p[1].lower()))
for idx in range(1, 1+max_len):
    ws[f'G{idx}'] = None
    ws[f'H{idx}'] = None
for idx, (a, b) in enumerate(pairs_sorted):
    ws[f'G{idx+1}'] = a
    ws[f'H{idx+1}'] = b

wb.save(output_path)
