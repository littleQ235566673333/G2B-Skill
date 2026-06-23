from openpyxl import load_workbook
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_8/regression_gate/after_fix/core_263-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_8/regression_gate/after_fix/core_263-1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

# Find the indices for the relevant columns
try:
    idx_mtrl = header.index('Mtrl')
    idx_width = header.index('Width')
    idx_height = header.index('Height')
except ValueError:
    raise Exception('Cannot find required columns in header')

totals = defaultdict(float)

for row in ws.iter_rows(min_row=2, values_only=True):
    mtrl, width, height = row[idx_mtrl], row[idx_width], row[idx_height]
    if mtrl and width and height:
        try:
            area = float(width) * float(height)
        except Exception:
            continue
        totals[mtrl] += area

# Sort materials for output - to match what user expects if a template is present
materials_sorted = list(totals.keys())
# If existing values in H2:H4 specify the names/ordering, use that
output_cells = [(2, 'H2'), (3, 'H3'), (4, 'H4')]
for i, (rownum, cell) in enumerate(output_cells):
    label_cell = ws[f'G{rownum}']
    if label_cell.value in totals:
        ws[f'H{rownum}'] = round(totals[label_cell.value], 2)
    else:
        ws[f'H{rownum}'] = ''  # wipe if not present

wb.save(output_path)
