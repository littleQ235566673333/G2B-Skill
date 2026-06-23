from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/regression_gate/before_pass/core_160-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/regression_gate/before_pass/core_160-6/output.xlsx'

wb = load_workbook(input_path)
ws = wb['SH']

all_rows = list(ws.iter_rows(min_row=2, max_row=12, max_col=12, values_only=True))
compact_rows = [row for row in all_rows if any(cell is not None for cell in row)]

for r in range(6, 12):
    # Clear out region 'A6:L11'
    for c in range(1, 13):
        ws.cell(row=r, column=c, value=None)

for i, row in enumerate(compact_rows):
    if i >= 6:  # Only fill up to row 11
        break
    for c, v in enumerate(row, 1):
        ws.cell(row=6+i, column=c, value=v)

wb.save(output_path)
