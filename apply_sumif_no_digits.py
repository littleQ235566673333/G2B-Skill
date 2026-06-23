from openpyxl import load_workbook
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_2/regression_gate/after_pass/core_42181/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_2/regression_gate/after_pass/core_42181/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# A13 has the reference text
ref_text = ws['A13'].value

sum_b = 0
for row in range(4, 11):  # Only rows 4 to 10 inclusive
    col_i_val = ws[f'I{row}'].value
    col_b_val = ws[f'B{row}'].value
    # Check if col_i_val is a string, contains ref_text, and contains NO digits
    if isinstance(col_i_val, str) and ref_text in col_i_val:
        if not re.search(r'\d', col_i_val):  # No digits present
            if isinstance(col_b_val, (int, float)):
                sum_b += col_b_val
ws['B13'] = sum_b
wb.save(output_path)
