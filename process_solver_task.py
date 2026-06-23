import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/group_387-16/r1/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/group_387-16/r1/evolve_387-16/output.xlsx'

# Read using pandas.
df = pd.read_excel(input_path, sheet_name='Sheet1', header=None)

# Column indices
i_value = 0  # 'Value'
i_binaries = 1  # 'Binaries'
i_result_values = 3  # 'Result values', starting row 4

target_value = df.iloc[1, i_result_values]  # Row 2 (zero-based 1), col 3

# Read Values & Binaries
value_col = df.iloc[2:18, i_value].tolist()  # Rows 3-18
binaries_col = df.iloc[2:18, i_binaries].tolist()
# Read Result values (appearing in same (row,col) as sample)
result_values = df.iloc[3:19, i_result_values].tolist()  # Rows 4-19
result_values = [x for x in result_values if pd.notnull(x)]

def remove_one_instance(values, binaries, value_to_remove):
    for idx, v in enumerate(values):
        if v == value_to_remove:
            values.pop(idx)
            binaries.pop(idx)
            return
# Make copies
import copy
values = copy.deepcopy(value_col)
binaries = copy.deepcopy(binaries_col)
for v in result_values:
    remove_one_instance(values, binaries, v)
# Remove blanks
filtered_values = [x for x in values if pd.notnull(x)]
filtered_binaries = [x for x in binaries if pd.notnull(x)]

# Compose output DataFrame, pad to 16 rows
pad_len = 16 - len(filtered_values)
output_values = filtered_values + [None]*pad_len
output_binaries = filtered_binaries + [None]*pad_len
output_result_values = df.iloc[3:19, i_result_values].tolist()
output_target_value = [target_value] + [None]*15

# Compute Solver result and difference
try:
    solver_result = sum([float(x) for x in filtered_values if pd.notnull(x)])
except Exception as e:
    solver_result = None
try:
    target_val = float(target_value)
except Exception as e:
    target_val = None
diff = solver_result - target_val if solver_result is not None and target_val is not None else None

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Write headers at A2
headers = ['Value', 'Binaries', 'Result values', 'Target value', 'Solver result', 'Difference']
for cidx, head in enumerate(headers):
    ws.cell(row=2, column=cidx+1, value=head)
# Write data rows (A3:D18) + two computed cols
for ridx in range(16):
    ws.cell(row=3+ridx, column=1, value=output_values[ridx])
    ws.cell(row=3+ridx, column=2, value=output_binaries[ridx])
    ws.cell(row=3+ridx, column=3, value=output_result_values[ridx] if ridx < len(output_result_values) else None)
    ws.cell(row=3+ridx, column=4, value=output_target_value[ridx])
# Write Solver result and Difference to E3, F3
ws.cell(row=3, column=5, value=solver_result)
ws.cell(row=3, column=6, value=diff)

wb.save(output_path)
