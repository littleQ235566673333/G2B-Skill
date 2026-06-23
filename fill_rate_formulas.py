from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def main():
    input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_1/group_44017/r0/evolve_44017/input.xlsx'
    output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_1/group_44017/r0/evolve_44017/output.xlsx'
    wb = load_workbook(input_path)
    ws = wb['Data']

    BASE_RATE_COL = 23  # W
    FREQ_COL = 10      # J
    EFFECTIVE_COL = 12 # L
    INCREASES = [13, 14, 15, 16]  # M:N:O:P
    START_ROW = 14
    END_ROW = 42
    START_COL = 30  # AD
    END_COL = 41   # AO

    for row in range(START_ROW, END_ROW+1):
        for col in range(START_COL, END_COL+1):
            base = f'${get_column_letter(BASE_RATE_COL)}{row}'
            freq = f'${get_column_letter(FREQ_COL)}{row}'
            eff = f'${get_column_letter(EFFECTIVE_COL)}{row}'
            incs = [f'${get_column_letter(c)}{row}' for c in INCREASES]
            date_cell = f'${get_column_letter(col)}$9'
            # Build per-wave factors
            factors = []
            for i in range(4):
                # Each wave triggers at EDATE(eff, i*freq)
                inc = incs[i]
                # If increase cell is blank, use 0%
                cond = f'IF({inc}<>"",IF({date_cell}>=EDATE({eff},{i}*{freq}),1+{inc},1),1)'
                factors.append(cond)
            increase_formula = '*'.join(factors)
            formula = f'=IF({date_cell}<{eff},"",{base}*{increase_formula})'
            ws.cell(row=row, column=col).value = formula
            # Remove yellow fill (if any)
            ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)
    wb.save(output_path)

if __name__ == '__main__':
    main()
