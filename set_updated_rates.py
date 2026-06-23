from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_1/group_44017/r2/evolve_44017/input.xlsx')
ws = wb['Data']

first_data_row = 14
last_data_row = 42
base_col = 'W'
eff_date_col = 'L'
freq_col = 'J'
inc_cols = ['M', 'N', 'O', 'P']
date_row = 9
output_start_col = 30  # AD
output_end_col = 41   # AO

for row in range(first_data_row, last_data_row+1):
    for output_col in range(output_start_col, output_end_col+1):
        out_cell = f'{get_column_letter(output_col)}{row}'
        date_cell = f'{get_column_letter(output_col)}{date_row}'
        months_expr = f'(YEAR({date_cell})-YEAR({eff_date_col}{row}))*12+MONTH({date_cell})-MONTH({eff_date_col}{row})'
        formula = (
            f'=IF(AND(ISNUMBER({freq_col}{row}), {date_cell}>={eff_date_col}{row}),'
            f'{base_col}{row}'
            f'*PRODUCT('
            f'1+IF({months_expr}>=0, {inc_cols[0]}{row}, 0),'
            f'1+IF({months_expr}>={freq_col}{row}, {inc_cols[1]}{row}, 0),'
            f'1+IF({months_expr}>={freq_col}{row}*2, {inc_cols[2]}{row}, 0),'
            f'1+IF({months_expr}>={freq_col}{row}*3, {inc_cols[3]}{row}, 0)'
            f'),""'
            f')'
        )
        ws[out_cell] = formula

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_1/group_44017/r2/evolve_44017/output.xlsx')
