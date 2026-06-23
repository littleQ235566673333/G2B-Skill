from openpyxl import load_workbook
from datetime import datetime, timedelta

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_6/regression_gate/before_pass/core_41589/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_6/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(in_path)
ws = wb['Contact List']
today = datetime.today()

for row in ws.iter_rows(min_row=4, min_col=8, max_col=10):
    h_cell, i_cell, j_cell = row
    h_date = h_cell.value
    i_val = (i_cell.value or '').strip().upper()  # Defensive for blank
    j_val = 'NO ACTION'
    if isinstance(h_date, datetime) and i_val == 'YES':
        delta_days = (today - h_date).days
        if delta_days <= 30:
            j_val = 'HOLD'
        elif delta_days > 30:
            j_val = 'TOUCH BASE'
    j_cell.value = j_val

wb.save(out_path)
