from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_54474_s2/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-5.4_v4/multiseed_full_eval/task_54474_s2/output.xlsx'

wb = load_workbook(input_path)
ws = wb['WHP DATA']
for r in range(1, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, 6)]
    if any(v is not None for v in vals):
        print(r, vals)
