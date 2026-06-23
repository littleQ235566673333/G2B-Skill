import openpyxl
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/group_10452/r2/evolve_10452/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/group_10452/r2/evolve_10452/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active
# Scan B4:B12, collect those starting with 'PK'.
b_range = range(4, 13)  # inclusive of 12
output_vals = []
for r in b_range:
    val = ws.cell(row=r, column=2).value
    if isinstance(val, str) and val.startswith('PK'):
        output_vals.append(val)
# Write results contiguously in E4 downward, rest blank
for idx, r in enumerate(b_range):
    ws.cell(row=r, column=5, value=output_vals[idx] if idx < len(output_vals) else "")
wb.save(output_path)
