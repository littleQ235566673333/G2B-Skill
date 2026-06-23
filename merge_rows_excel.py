import openpyxl

def is_number(val):
    try:
        float(val)
        return True
    except (TypeError, ValueError):
        return False

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_4/regression_gate/before_fix/core_177-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_4/regression_gate/before_fix/core_177-6/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['DATA']

header_row = [c.value for c in ws[1]]
header_to_col = {header_row[i]: i for i in range(len(header_row))}

merge_col = 'ComboKey'
merge_idx = header_to_col[merge_col]

rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
groups = {}
for row in rows:
    key = row[merge_idx]
    if key not in groups:
        groups[key] = []
    groups[key].append(row)

out_rows = []
out_rows.append(header_row[:18])
for group_rows in groups.values():
    base = list(group_rows[0][:8])
    combined_nums = [0] * 10
    for row in group_rows:
        for j in range(8, 18):
            val = row[j]
            if is_number(val):
                combined_nums[j-8] += float(val)
    formatted_nums = []
    for x in combined_nums:
        if round(x, 2) == 0:
            formatted_nums.append('')
        else:
            formatted_nums.append(f"{x:.2f}")
    merged = base + formatted_nums
    out_rows.append(merged)

ws_out = wb['combined']
for i, row in enumerate(out_rows[:8]):
    for j, val in enumerate(row):
        ws_out.cell(row=i+1, column=j+1, value=val)

wb.save(output_path)
