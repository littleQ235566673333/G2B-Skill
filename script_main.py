import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def get_ids(sheet, col=1):
    # Read all values in the first column (skipping the header row)
    return [row[0] for row in sheet.iter_rows(min_col=col, max_col=col, min_row=2, values_only=True) if row[0] is not None]

def append_missing_ids(main_ids, getting_ids, getting_sheet):
    # Find which IDs from main_ids are missing in getting_ids
    missing_ids = [id for id in main_ids if id not in getting_ids]
    # Append each missing ID as a new row in getting_sheet; rest columns blank
    max_col = getting_sheet.max_column
    for id in missing_ids:
        new_row = [id] + [None]*(max_col-1)
        getting_sheet.append(new_row)

# Main processing
wb = openpyxl.load_workbook('results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_395-36_tc1/input.xlsx')
main_ids_sheet = wb['Main unique ID']
getting_sheet = wb['Result what i am getting']

main_id_list = get_ids(main_ids_sheet)
getting_id_list = get_ids(getting_sheet)

append_missing_ids(main_id_list, getting_id_list, getting_sheet)

# Write output to new sheet 'MyResult' with a maximum of A1:O20
if 'MyResult' in wb.sheetnames:
    del wb['MyResult']
result_sheet = wb.create_sheet('MyResult')

# Write table with up to 20 rows and 15 columns (A1:O20)
for i, row in enumerate(getting_sheet.iter_rows(values_only=True)):
    if i >= 20:
        break
    for j, value in enumerate(row):
        if j >= 15:
            break
        result_sheet.cell(row=i+1, column=j+1, value=value)

wb.save('results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_395-36_tc1/output.xlsx')
print('Done')
