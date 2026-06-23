import openpyxl
from openpyxl.styles import Alignment

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_56378_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_56378_tc1/output.xlsx'

# Load workbook and select the first sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Assume Frame 1 has headers in row 4 and products start from row 5
header_row = 4
start_row = 5

# Identify columns
headers = [cell.value for cell in ws[header_row]]
product_col_idx = headers.index('PRODUCT') + 1 if 'PRODUCT' in headers else None
quantity_col_idx = headers.index('QUANTITY UNITS') + 1 if 'QUANTITY UNITS' in headers else None

# Collect products with non-empty quantity units
filtered_rows = []
for row in ws.iter_rows(min_row=start_row, max_row=start_row+20, values_only=False):  # Scan 20 products
    product = row[product_col_idx-1].value if product_col_idx else None
    quantity = row[quantity_col_idx-1].value if quantity_col_idx else None
    # Check for non-empty QUANTITY UNITS
    if quantity is not None and str(quantity).strip() != '':
        filtered_rows.append([cell.value for cell in row])

# Only use up to 4 rows for Frame 2 (L5:R8)
filtered_rows = filtered_rows[:4]

# Output is L5:R8 => cols 12 to 18 (L=12, M=13, ..., R=18)
out_start_row = 5
out_start_col = 12
out_end_col = 18

for i, filtered in enumerate(filtered_rows):
    for j in range(out_start_col, out_end_col + 1):
        ws.cell(row=out_start_row + i, column=j, value=filtered[j - out_start_col])
        # Alignment: Left for Product (L), Right for O:R
        if j == out_start_col:
            ws.cell(row=out_start_row + i, column=j).alignment = Alignment(horizontal='left')
        elif j >= 15:
            ws.cell(row=out_start_row + i, column=j).alignment = Alignment(horizontal='right')

wb.save(output_path)
