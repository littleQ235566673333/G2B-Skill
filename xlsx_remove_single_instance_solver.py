import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/regression_gate/after_fix/core_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/regression_gate/after_fix/core_387-16/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

def get_col_values(ws, col_letter, min_row, max_row):
    return [ws[f'{col_letter}{i}'].value for i in range(min_row, max_row + 1)]

def set_col_values(ws, col_letter, start_row, values):
    for offset, value in enumerate(values):
        ws[f'{col_letter}{start_row + offset}'].value = value

# Data region A2:D18
min_row, max_row = 2, 18
colA = get_col_values(ws, 'A', min_row, max_row)
colB = get_col_values(ws, 'B', min_row, max_row)  # 'Value'
colC = get_col_values(ws, 'C', min_row, max_row)  # 'Binaries'
colD = get_col_values(ws, 'D', min_row, max_row)  # 'Result values'

# Remove only one instance for each result value in B and C (keeping index alignment of B/C)
B_mask = [True] * len(colB)
C_mask = [True] * len(colC)
for target in colD:
    for i, bval in enumerate(colB):
        if bval == target and B_mask[i]:
            B_mask[i] = False
            C_mask[i] = False
            break

colA_kept = [colA[i] for i in range(len(colA)) if B_mask[i]]
colB_kept = [colB[i] for i in range(len(colB)) if B_mask[i]]
colC_kept = [colC[i] for i in range(len(colC)) if C_mask[i]]

# Remove blanks, pad with None to keep length 17
pad_len = max_row - min_row + 1
def pad_list(L, pad_len):
    return L + [None]*(pad_len - len(L))

colA_new = pad_list(colA_kept, pad_len)
colB_new = pad_list(colB_kept, pad_len)
colC_new = pad_list(colC_kept, pad_len)

# Compute solver result (sum of non-None numeric values in colA_new)
solver_result = sum([x for x in colA_new if isinstance(x, (int, float))])
# Find target value (last non-empty numeric in colD)
target_val = None
for x in reversed(colD):
    if isinstance(x, (int, float)):
        target_val = x
        break
diff_val = solver_result - target_val if target_val is not None else None

for i in range(pad_len):
    ws[f'A{min_row + i}'].value = colA_new[i]
    ws[f'B{min_row + i}'].value = colB_new[i]
    ws[f'C{min_row + i}'].value = colC_new[i]

ws['D2'].value = 'solver result'
ws['D3'].value = solver_result
ws['D4'].value = 'difference'
ws['D5'].value = diff_val
ws['D6'].value = 'target value'
ws['D7'].value = target_val

wb.save(output_path)
