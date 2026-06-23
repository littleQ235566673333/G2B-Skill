import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_209-30_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_209-30_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Data to Import']

# Iterate from row 2 to the last non-empty row in column C
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=3)  # column 3 is C
    value = cell.value
    if isinstance(value, str) and len(value) > 3:
        cell.value = value[:-3]
    # Optionally, if it's less than or equal to 3 chars, you may want to make it blank or leave as is.

wb.save(output_path)
