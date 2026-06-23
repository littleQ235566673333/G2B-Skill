import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_1/group_547-43/r2/evolve_547-43/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_1/group_547-43/r2/evolve_547-43/output.xlsx'

# Read data using pandas for convenience
emp = pd.read_excel(input_path, sheet_name='Emp')
lookup = pd.read_excel(input_path, sheet_name='Lookup')

rows_out = []
for idx, row in emp.iterrows():
    # Skip any fully empty rows
    if pd.isna(row['seqno']):
        continue
    for _, lrow in lookup.iterrows():
        var = lrow['Variable']
        cat = lrow['Category']
        subcat = lrow['Subcategory']
        # Make sure the variable/column exists in emp
        if var in row:
            val = row[var]
            if pd.notna(val):
                rows_out.append([
                    row['seqno'], row['Empno'], row['ename'], row['Jobid'],
                    var, cat, subcat, val
                ])

# Load output workbook using openpyxl for writing
wb = openpyxl.load_workbook(input_path)
ws = wb['Expected']

# Clear existing output rows except the header
ws.delete_rows(2, ws.max_row - 1)

# Write output
for r_idx, out_row in enumerate(rows_out, start=2):
    for c_idx, value in enumerate(out_row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(output_path)
