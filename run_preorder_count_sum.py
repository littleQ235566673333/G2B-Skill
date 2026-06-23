import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/regression_gate/after_pass/core_50796/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/regression_gate/after_pass/core_50796/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Output categories to process (A2:A5): get their values
output_categories = [ws[f'A{i}'].value for i in range(2, 6)]

# Data range is B13:C22 (descriptors and quantities)
start_data_row = 13
end_data_row = 22

# Helper: sum C for each category where B matches
for idx, category in enumerate(output_categories):
    total = 0
    for row in range(start_data_row, end_data_row+1):
        descriptor = ws[f'B{row}'].value
        quantity = ws[f'C{row}'].value
        if descriptor == category and isinstance(quantity, (int, float)):
            total += quantity
    ws[f'B{idx+2}'] = total

wb.save(output_path)
