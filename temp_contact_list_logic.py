from openpyxl import load_workbook
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_3/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_3/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']

today = datetime.today()
h_value = ws['H4'].value
i_value = ws['I4'].value
result = 'NO ACTION'

if isinstance(h_value, datetime) and i_value and str(i_value).strip().upper() == 'YES':
    if h_value >= today - timedelta(days=30):
        result = 'HOLD'
    else:
        result = 'TOUCH BASE'

ws['J4'] = result
wb.save(output_path)
