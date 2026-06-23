from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_3/group_50916/r1/evolve_50916/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_3/group_50916/r1/evolve_50916/output.xlsx'

wb = load_workbook(input_path)
ws = wb['21-22 Schedule']

for row in range(12, 15):  # Rows 12, 13, 14
    for col in range(3, 9):  # Columns C=3 to H=8
        col_letter = chr(64 + col)
        # Dynamic Excel formula (INDEX/MATCH lookup)
        formula = f'=INDEX($C$2:$H$8, MATCH($B{row}, $B$2:$B$8, 0), COLUMN()-2)'
        ws[f'{col_letter}{row}'] = formula

wb.save(output_path)
