import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_2/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_2/regression_gate/after_pass/core_50526/output.xlsx'

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active
# Load as DataFrame for easier row/col logic
# (ignoring empty/metadata rows after real data)
df = pd.read_excel(input_path)
# Get lookup value from B6
lookup_value = ws['B6'].value
results = []
if lookup_value is not None:
    # The leftmost column ("Unnamed: 0") contains keys
    row = df[df['Unnamed: 0'].astype(str) == str(lookup_value)]
    if not row.empty:
        # For each color (columns except lookup col), keep where value > 0
        for col in df.columns[1:]:
            val = row.iloc[0][col]
            if pd.notnull(val) and val > 0:
                results.append(col)
# Write to B9 and B10
for i in range(2):
    ws[f'B{9+i}'] = results[i] if i < len(results) else ''
wb.save(output_path)
