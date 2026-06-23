import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_48983_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_48983_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Output region is M6:S11 (columns 13 to 19, rows 6 to 11)
category_row = 5   # The row containing category headers for both tables
brand_col = 11     # Column K (11) where brands for table 2 exist
output_start_col = 13 # M
output_end_col = 19   # S
output_start_row = 6
output_end_row = 11

# Brands for output rows (in K6:K11)
br_names = [ws.cell(row=row, column=brand_col).value for row in range(output_start_row, output_end_row + 1)]
# Categories for output columns (in M5:S5)
cat_names = [ws.cell(row=category_row, column=col).value for col in range(output_start_col, output_end_col + 1)]

# Main table (A/B/C): Brand column is B (2), Data columns are C:J (3:10), brands in 6:15, headers in 5
data_brand_col = 2
main_cat_start = 3  # C
main_cat_end = 10   # J
main_brand_rows = list(range(6,16))  # Data from row 6 to 15
main_cat_names = [ws.cell(row=category_row, column=col).value for col in range(main_cat_start, main_cat_end + 1)]
main_br_names = [ws.cell(row=row, column=data_brand_col).value for row in main_brand_rows]
# Build mapping
cat_to_col = {name: col for col, name in zip(range(main_cat_start, main_cat_end + 1), main_cat_names) if name}
brand_to_row = {name: row for row, name in zip(main_brand_rows, main_br_names) if name}

# Fill the output table (M6:S11)
for ocol, cat in zip(range(output_start_col, output_end_col + 1), cat_names):
    if not cat or cat not in cat_to_col:
        continue
    mcol = cat_to_col[cat]
    for orow, br in zip(range(output_start_row, output_end_row + 1), br_names):
        if not br or br not in brand_to_row:
            continue
        mrow = brand_to_row[br]
        val = ws.cell(row=mrow, column=mcol).value
        ws.cell(row=orow, column=ocol).value = val

wb.save(output_path)
