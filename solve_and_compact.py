import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_2/group_387-16/r2/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_2/group_387-16/r2/evolve_387-16/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Value and Binaries are in cols A (1) and B (2), start at row 3
values = []
binaries = []
for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
    values.append(row[0])
    binaries.append(row[1])

# Target value is in E2 (column 5, row 2)
target_value = ws.cell(row=2, column=5).value

# Result values label is at D4, values from D5 down
result_values = []
row_rv = 5
while True:
    rv = ws.cell(row=row_rv, column=4).value
    if rv is None:
        break
    result_values.append(rv)
    row_rv += 1

# Remove one instance of each result value from values/binaries
values_work = list(values)
binaries_work = list(binaries)
for rv in result_values:
    for idx, v in enumerate(values_work):
        if v == rv:
            values_work[idx] = None
            binaries_work[idx] = None
            break
# Compact: remove blanks and keep up to 17 slots
compacted = [(v, b) for v, b in zip(values_work, binaries_work) if v is not None]
while len(compacted) < 17:
    compacted.append((None, None))
# Solver result is sum of all non-None in col A (0)
solver_result = sum([v for v, b in compacted if isinstance(v, (int, float))])
diff = solver_result - (target_value if target_value else 0)

# Write Values and Binaries to A2:B18
for i in range(17):
    ws.cell(row=2+i, column=1).value = compacted[i][0]
    ws.cell(row=2+i, column=2).value = compacted[i][1]
# Also fill C and D columns with None for format
for i in range(17):
    ws.cell(row=2+i, column=3).value = None
    ws.cell(row=2+i, column=4).value = None

# Write summary
ws['B20'] = 'solver result'
ws['C20'] = solver_result
ws['B21'] = 'Difference'
ws['C21'] = diff
wb.save(output_path)
