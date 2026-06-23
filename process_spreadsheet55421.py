from openpyxl import load_workbook
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_3/regression_gate/before_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_3/regression_gate/before_pass/core_55421/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def get_col_values(col, min_row, max_row):
    return [ws[f'{col}{row}'].value for row in range(min_row, max_row + 1)]

rows = range(2, 21)  # F2:F20 (rows 2 through 20 inclusive)
A_vals = get_col_values('A', 2, 20)
D_vals = get_col_values('D', 2, 20)
E_vals = get_col_values('E', 2, 20)

# Map: value in column A -> list of row indices (0-based, relative to start)
A_to_rows = defaultdict(list)
for idx, a in enumerate(A_vals):
    if a is not None:
        A_to_rows[a].append(idx)

for idx, a in enumerate(A_vals):
    result = ''
    if a is None:
        ws[f'F{idx+2}'] = ''
        continue
    d_statuses = [D_vals[i] for i in A_to_rows[a]]
    e_dates = [E_vals[i] for i in A_to_rows[a]]
    status_set = set(str(s or '') for s in d_statuses)
    # Only 'SCH': all statuses are exactly SCH
    if status_set == {'SCH'}:
        result = 'FUTURE'
    # Both 'SCH' and 'NO SHOW': mixture
    elif 'NO SHOW' in status_set and 'SCH' in status_set:
        result = 'NS/SCHED'
    # Only NO SHOW
    elif status_set == {'NO SHOW'}:
        # E has a date (any non-empty)
        if any(bool(e) for e in e_dates):
            result = 'NO ACTION NEEDED'
        else:
            result = 'CALL PT'
    ws[f'F{idx+2}'] = result

wb.save(output_path)
