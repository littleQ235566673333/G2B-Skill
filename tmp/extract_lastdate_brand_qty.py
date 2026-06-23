import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_3/regression_gate/after_fix/core_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_3/regression_gate/after_fix/core_516-46/output.xlsx'
sheet_name = 'ورقة1'

df = pd.read_excel(input_path, sheet_name=sheet_name)
df = df.iloc[:, 0:5]
df.columns = ['date', 'brand', 'batch', 'origin', 'qty']
df['date'] = pd.to_datetime(df['date'])
df_grouped = df.groupby('brand')
final_rows = []
for brand, group in df_grouped:
    last_date = group['date'].max()
    subset = group[group['date'] == last_date]
    qty_sum = subset['qty'].sum()
    batch = subset.iloc[0]['batch'] if not subset.empty else ''
    origin = subset.iloc[0]['origin'] if not subset.empty else ''
    entry = [last_date, brand, batch, origin, qty_sum]
    if len(subset) > 1:
        entry.append('Combined {} entries'.format(len(subset)))
    else:
        entry.append('')
    final_rows.append(entry)
out_df = pd.DataFrame(final_rows, columns=['date', 'brand', 'batch', 'origin', 'qty', 'modifications'])

wb = load_workbook(input_path)
ws = wb[sheet_name]
for r, row in enumerate(out_df.values.tolist(), start=2):
    for c, value in enumerate(row, start=8):
        ws.cell(row=r, column=c, value=value)
wb.save(output_path)
print(out_df)
