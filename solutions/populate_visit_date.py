import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/regression_gate/after_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/regression_gate/after_pass/core_9726/output.xlsx'

# Read the relevant columns
sheet = 'Sheet1'
df = pd.read_excel(input_path, sheet_name=sheet)

# Table 1: ['Student ID', 'Visit date']
t1 = df[['Student ID', 'Visit date']].dropna(subset=['Student ID', 'Visit date'])
t1['Visit date'] = pd.to_datetime(t1['Visit date'], errors='coerce')

# Table 2: ['Student ID.1', 'Submission date']
t2 = df[['Student ID.1', 'Submission date']].copy()
t2['Submission date'] = pd.to_datetime(t2['Submission date'], errors='coerce')

# Prepare dictionary of visit dates per student
visit_by_student = t1.groupby('Student ID')['Visit date'].apply(list).to_dict()

def find_latest_prior(visit_dates, cut_date):
    dates = pd.Series(visit_dates)
    prior_dates = dates[dates < cut_date]
    if not prior_dates.empty:
        return prior_dates.max()
    return pd.NaT

output_dates = []
for idx, row in t2.iterrows():
    student = row['Student ID.1']
    submission_date = row['Submission date']
    visits = visit_by_student.get(student, [])
    dt = find_latest_prior(visits, submission_date)
    output_dates.append(dt.strftime('%d-%b-%y') if pd.notnull(dt) else '')

# Write to Visit date.1 column (J)
wb = load_workbook(input_path)
ws = wb[sheet]
start_row = 2
output_col = 10 # 'J'
for i, val in enumerate(output_dates, start=start_row):
    ws.cell(row=i, column=output_col).value = val
wb.save(output_path)
