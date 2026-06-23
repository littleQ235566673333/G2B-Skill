import openpyxl
input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_9/task_44389/r2/evolve_44389/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_9/task_44389/r2/evolve_44389/output.xlsx'
from openpyxl import load_workbook
wb = load_workbook(input_path)
# Step 1: Detect and verify the sheet name and header row.
sheetnames = wb.sheetnames
ws = wb[sheetnames[0]]  # Use 1st sheet (user did not specify, brief robust check)
# Find header row
header_row_idx = None
headers = None
for row in ws.iter_rows(min_row=1, max_row=5):
    values = [str(cell.value).strip() if cell.value is not None else '' for cell in row]
    nonblank = [v for v in values if v]
    if len(nonblank) >= 2 and all(v for v in values):
        header_row_idx = row[0].row
        headers = values
        break
if not headers:
    raise Exception('Unable to find header row by first 5 rows')
# Map columns (to indexes)
target_cols = [i for i, h in enumerate(headers) if h not in ('', None)]
# We'll process rows 2-7 for output
def is_number(val):
    try:
        return (val is not None) and (isinstance(val, (int, float)) or (isinstance(val, str) and val not in ('', None) and float(val) == float(val)))
    except Exception:
        return False
answer_cells = [(r, 16) for r in range(header_row_idx+1, header_row_idx+7)]  # P2:P7 (col 16 = P)
for ofs, (row_idx, col_idx) in enumerate(answer_cells):
    values = []
    for c in target_cols:
        val = ws.cell(row=row_idx, column=c+1).value
        if is_number(val):
            try: val = float(val)
            except: val = None
        else:
            val = None
        values.append(val)
    # Exclude non-numeric or missing; only those >0
    with_headers = [(headers[i], v) for i, v in enumerate(values) if is_number(v) and v > 0]
    if not with_headers:
        result = ''
    else:
        min_val = min(v for h, v in with_headers)
        res_headers = [h for h, v in with_headers if v == min_val]
        result = ','.join(str(h) for h in res_headers)
    ws.cell(row=row_idx, column=col_idx, value=result)
wb.save(output_path)
