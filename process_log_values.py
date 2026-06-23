import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_5835_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_5835_tc1/output.xlsx'

# Load workbook and sheets
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Column C and I indices (C=3, I=9)
C_col = 3
I_col = 9

# Range for C: rows 3 to 19
for row in range(3, 20):
    log_value = ws.cell(row=row, column=I_col).value
    ws.cell(row=row, column=C_col).value = log_value

# Save to output
wb.save(output_path)
