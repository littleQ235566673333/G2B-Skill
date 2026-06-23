from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_1/group_44017/r0/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_1/group_44017/r0/evolve_44017/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Data']

# Remove fill style
no_fill = PatternFill(fill_type=None)

for col_idx in range(30, 42):  # AD (30) to AO (41)
    col_letter = get_column_letter(col_idx)
    for row in range(14, 43):  # Rows 14 to 42
        date_cell = f'{col_letter}$9'
        base = f'$W{row}'
        eff_date = f'$L{row}'
        freq = f'$J{row}'
        incs = [f'$M{row}', f'$N{row}', f'$O{row}', f'$P{row}']
        step_dates = [
            f'DATE(YEAR({eff_date}),MONTH({eff_date})+({i}*{freq}),DAY({eff_date}))' for i in range(4)
        ]
        part = []
        for n in range(4):
            part.append(f'IF({date_cell}>={step_dates[n]},1+{incs[n]},1)')
        multiplier = '*'.join(part)
        formula = f'=IF({date_cell}<{eff_date},"",{base}*({multiplier}))'
        cell = ws[f'{col_letter}{row}']
        cell.value = formula
        cell.fill = no_fill

wb.save(output_path)
