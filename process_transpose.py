import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_1/group_38985/r2/evolve_38985/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_1/group_38985/r2/evolve_38985/output.xlsx'

# Load all sheets with pandas to locate the tables
excel_obj = pd.ExcelFile(input_path)
sheet_names = excel_obj.sheet_names
content = {name: pd.read_excel(input_path, sheet_name=name) for name in sheet_names}

def find_table1_table2(content):
    t1 = t2 = None
    for df in content.values():
        cols = df.columns
        for col in cols:
            vc = df[col].value_counts()
            if vc.max() > 1 and len(vc) > 2: # duplicate names
                t1 = df
                break
    for df in content.values():
        cols = df.columns
        for col in cols:
            vc = df[col].value_counts()
            if vc.max() == 1 and len(vc) > 2: # unique names
                t2 = df
                break
    return t1, t2

table1, table2 = find_table1_table2(content)

# If structure isn't matched, default to Sheet1 and Sheet2
if table1 is None or table2 is None:
    table1 = content.get('Sheet1', list(content.values())[0])
    table2 = content.get('Sheet2', list(content.values())[1])

# Assume first column in both is 'Name'
table1_names = table1.iloc[:, 0]
table2_names = table2.iloc[:, 0]
value_cols = table2.columns[1:]

# Prepare mapping: name -> list of values in appearance order (as there may be multiple table2 vals per name)
name_to_values = defaultdict(list)
for _, row in table2.iterrows():
    name = row.iloc[0]
    values = row.iloc[1:].tolist()
    name_to_values[name].append(values)

# For each name in table1, get next available set of associated values, and flatten them
output_vals = []
used_counter = defaultdict(int)
for name in table1_names.iloc[7:11]:  # D8:D11 (rows 8 to 11, 0-based index 7-10)
    vals_list = name_to_values.get(name, [[None]])
    idx = used_counter[name]
    used_counter[name] += 1
    vals = vals_list[min(idx, len(vals_list) - 1)]
    # Concatenate values for transposed effect
    if len(vals) > 1:
        flat = ', '.join([str(v) for v in vals if pd.notnull(v)])
    elif len(vals) == 1:
        flat = str(vals[0])
    else:
        flat = ''
    output_vals.append(flat)

# Write those to D8:D11
wb = load_workbook(input_path)
wss = wb.sheetnames
ws = wb[wss[0]]  # main sheet
for i, val in enumerate(output_vals):
    ws[f'D{8 + i}'] = val
wb.save(output_path)
