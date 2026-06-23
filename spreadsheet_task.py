import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_58032_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_58032_tc1/output.xlsx'

# Load input workbook
wb = openpyxl.load_workbook(input_path)
ws_in = wb['Sheet1']

# Duplicate Sheet1 to new sheet
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Sheet1'

max_row = ws_in.max_row
max_col = ws_in.max_column

# Helper styles
arial_font = Font(name='Arial')
title_font = Font(name='Arial', bold=True, italic=True)
fill_colD = PatternFill(fill_type='solid', fgColor='CCCCCC')
fill_formula = PatternFill(fill_type='solid', fgColor='FCD5B4')
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

# Copy values and formatting
for row in ws_in.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
    for cell in row:
        o_cell = ws_out.cell(row=cell.row, column=cell.col_idx, value=cell.value)
        o_cell.font = arial_font
        if cell.row == 1:
            o_cell.font = title_font
            o_cell.border = border

# Fill Column D (from D2) with #CCCCCC
for r in range(2, max_row+1):
    ws_out[f'D{r}'].fill = fill_colD

# Formula for A2:A35
# Assumes Title is column B, Dept code is column C, search names in columns I (Title) and J (Dept code)
for r in range(2, min(36, max_row+1)):
    formula = ("=IFERROR(INDEX($H$2:$H$100, MATCH(B{row}&C{row}, $I$2:$I$100&$J$2:$J$100, 0)), 'Not found')".format(row=r))
    ws_out[f'A{r}'].value = formula
    ws_out[f'A{r}'].fill = fill_formula
    ws_out[f'A{r}'].font = arial_font

# Hide gridlines
ws_out.sheet_view.showGridLines = False

wb_out.save(output_path)
