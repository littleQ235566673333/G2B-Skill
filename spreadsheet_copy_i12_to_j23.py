import openpyxl

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_55060_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_55060_tc1/output.xlsx"

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Unmerge J23:N23 if merged
from openpyxl.utils import get_column_letter
for merged_range in list(ws.merged_cells.ranges):
    # Check if the merged range intersects with J23:N23
    m_start_col, m_start_row, m_end_col, m_end_row = merged_range.bounds
    for col in range(9, 15):  # 9 = J, 14 = N
        if m_start_row <= 23 <= m_end_row and m_start_col <= col <= m_end_col:
            ws.unmerge_cells(range_string=str(merged_range))
            break

val_i12 = ws['I12'].value
if val_i12 is None or (isinstance(val_i12, str) and val_i12.strip() == ""):
    value_to_set = ""
else:
    value_to_set = val_i12

for col in range(openpyxl.utils.column_index_from_string('J'), openpyxl.utils.column_index_from_string('N')+1):
    ws.cell(row=23, column=col, value=value_to_set)

wb.save(output_path)
