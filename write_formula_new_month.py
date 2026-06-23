from openpyxl import load_workbook

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_4/group_45707/r0/evolve_45707/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_4/group_45707/r0/evolve_45707/output.xlsx'

wb = load_workbook(input_fp)
ws = wb.active

# Map headers
header_row = next(row for row in ws.iter_rows(min_row=1, max_row=10, values_only=True) if any(cell is not None for cell in row))
header_map = {header: idx + 1 for idx, header in enumerate(header_row) if header}
col_date = header_map['Date']
col_occur = header_map['Occurence']
col_output = 4  # D

start_row, end_row = 2, 69
for row in range(start_row, end_row + 1):
    if row + 1 <= end_row:
        cell_next_addr = f"A{row+1}"
        # Formula: Only output if DAY(next-date) = 1; then count Occurence=1 for that month+year
        formula = (
            f'=IF(DAY({cell_next_addr})=1,'
            f' SUMPRODUCT((C$2:C$69=1)*(MONTH(A$2:A$69)=MONTH({cell_next_addr}))*(YEAR(A$2:A$69)=YEAR({cell_next_addr}))),"" )'
        )
        ws.cell(row=row, column=col_output, value=formula)
    else:
        ws.cell(row=row, column=col_output, value=None)

wb.save(output_fp)
