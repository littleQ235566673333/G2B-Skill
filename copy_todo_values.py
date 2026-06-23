from openpyxl import load_workbook

# File paths
task_input = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_6/regression_gate/after_fix/core_48975/input.xlsx'
task_output = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_6/regression_gate/after_fix/core_48975/output.xlsx'

# Load workbook and sheets
wb = load_workbook(task_input)
ws_in = wb['Input']
ws_out = wb['Output']

# Prepare data
output_values = []
for row in ws_in.iter_rows(min_row=9, max_row=100, min_col=1, max_col=5):
    a, b, c, d, todo = [cell.value for cell in row]
    # Must have 'yes' in column E, 'A', 'B' not empty
    if (str(todo).strip().lower() == 'yes'
        and a is not None and b is not None and str(a).strip() != '' and str(b).strip() != ''):
        output_values.append((a, b))

# Write only 'A' values starting at B11 (as per instruction cells: Output'!B11:B17)
for i, (a_val, b_val) in enumerate(output_values):
    ws_out.cell(row=11+i, column=2, value=a_val)  # Column B (2)
    # If you instead wanted (b_val), change a_val to b_val

wb.save(task_output)
