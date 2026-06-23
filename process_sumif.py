import re
import pandas as pd
from openpyxl import load_workbook

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_6/regression_gate/after_pass/core_42181/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_6/regression_gate/after_pass/core_42181/output.xlsx'

# Load workbook/worksheet
wb = load_workbook(input_path)
ws = wb.active

a13 = ws['A13'].value

row_start = 4
row_end = 10

b_values = []
for row in range(row_start, row_end + 1):
    i_val = ws[f'I{row}'].value
    b_val = ws[f'B{row}'].value
    if isinstance(i_val, str) and (a13 in i_val) and not re.search(r'[0-9]', i_val):
        if isinstance(b_val, (int, float)):
            b_values.append(b_val)

result = sum(b_values)
ws['B13'] = result
wb.save(output_path)
