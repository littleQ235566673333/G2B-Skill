import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_1/group_387-16/r2/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_1/group_387-16/r2/evolve_387-16/output.xlsx'

# Load all rows, no headers
xls = pd.read_excel(input_path, sheet_name='Sheet1', header=None)

# Values: column 0, rows 2 onwards (index >= 2)
# Binaries: column 1, same rows
# Result values: column 3, rows from where 'Result values' label exists + 1

value_col_idx = 0
binaries_col_idx = 1
result_label_row = xls[xls[3]=='Result values'].index[0]
first_data_row = 2
result_vals_start_row = result_label_row + 1

# Get Value/Binaries block (aligned rows)
values = xls.loc[first_data_row:, value_col_idx].reset_index(drop=True)
binaries = xls.loc[first_data_row:, binaries_col_idx].reset_index(drop=True)

# Get Result values block, stopping at first nan
result_vals = []
for r in range(result_vals_start_row, xls.shape[0]):
    val = xls.iat[r, 3]
    if pd.isna(val):
        break
    result_vals.append(val)

# Remove one instance when Value matches Result values
used_indices = set()
for res_val in result_vals:
    match_idxs = values[(values == res_val) & (~values.index.isin(used_indices))]
    if not match_idxs.empty:
        idx = match_idxs.index[0]
        values.iloc[idx] = ''
        binaries.iloc[idx] = ''
        used_indices.add(idx)

# Compact columns (remove blanks/shifts up)
def compact(col):
    items = [x for x in col if pd.notna(x) and str(x) != '']
    return items + [''] * (len(col) - len(items))
values = pd.Series(compact(values))
binaries = pd.Series(compact(binaries))

# Solver result = sum of values in col 0 (after filtering)
try:
    solver_result = sum(float(val) for val in values if pd.notna(val) and str(val) != '')
except Exception:
    solver_result = ''
# Target value (from cell [1,4])
target_value = xls.iat[1,4] if xls.shape[1] > 4 else None
try:
    difference = solver_result - float(target_value) if target_value is not None and pd.notna(target_value) else ''
except Exception:
    difference = ''

# Write result to output at A2:D18, as requested
wb = load_workbook(input_path)
ws = wb['Sheet1']
# Clear A2:D18
for r in range(2,19):
    for c in range(1,5):
        ws.cell(row=r, column=c).value = None
# Write columns: A=values, B=binaries (C and D left empty, as in input)
for i in range(17):
    ws.cell(row=2+i, column=1).value = values.iloc[i] if i < len(values) else None
    ws.cell(row=2+i, column=2).value = binaries.iloc[i] if i < len(binaries) else None
# Insert solver result and difference
ws['A19'] = 'Solver Result'
ws['B19'] = solver_result
ws['A20'] = 'Difference'
ws['B20'] = difference
wb.save(output_path)
