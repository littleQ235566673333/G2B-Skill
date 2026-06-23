import openpyxl

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_55468_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_55468_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Helper to get values
cell = lambda coord: ws[coord].value

# Table bounds and criteria cell layout
# - C5:Z10 is data (as per formula)
# - C4:Z4, C3:Z3: column headers, two levels
# - A5:A10, B5:B10: row headers, two levels
# Criterion cells (as per user formula):
ac4 = cell('AC4') # col_header 1
ab4 = cell('AB4') # col_header 2

ae4 = cell('AE4') # row_header 1
ad4 = cell('AD4') # row_header 2

# Read column header blocks
col_header_1 = [ws.cell(row=4, column=col).value for col in range(3, 27)]  # C4:Z4
col_header_2 = [ws.cell(row=3, column=col).value for col in range(3, 27)]  # C3:Z3

# Read row header blocks
row_header_1 = [ws.cell(row=row, column=1).value for row in range(5, 11)]  # A5:A10
row_header_2 = [ws.cell(row=row, column=2).value for row in range(5, 11)]  # B5:B10

# Find row index: both vertical criteria must match
row_idx = None
for i, (v1, v2) in enumerate(zip(row_header_1, row_header_2)):
    if v1 == ae4 and v2 == ad4:
        row_idx = i
        break

# Find column index: both horizontal criteria must match
col_idx = None
for j, (h1, h2) in enumerate(zip(col_header_1, col_header_2)):
    if h1 == ac4 and h2 == ab4:
        col_idx = j
        break

result = None
if row_idx is not None and col_idx is not None:
    # C5:Z10 maps to [row 5 + row_idx, col 3 + col_idx]
    result = ws.cell(row=5 + row_idx, column=3 + col_idx).value

# Write result to AE5
ws['AE5'] = result

# Save workbook
wb.save(output_path)
