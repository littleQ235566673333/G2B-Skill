from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/regression_gate/after_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Identify column indexes
headers = [cell.value for cell in ws[2]]
col_map = {v: i+1 for i, v in enumerate(headers) if v}

header_row = 2
start_row = 3
end_row = ws.max_row

# Insert "Actual Age" column after "age (year)"
age_col = col_map['age (year)']
actual_age_col = age_col + 1
ws.insert_cols(actual_age_col)
ws.cell(row=header_row, column=actual_age_col, value='Actual Age')

# Replicate fill color from "age (year)" header
orig_fill = ws.cell(row=header_row, column=age_col).fill
if orig_fill:
    copied_fill = PatternFill(
        fill_type=orig_fill.fill_type,
        start_color=orig_fill.start_color.rgb if orig_fill.start_color else None,
        end_color=orig_fill.end_color.rgb if orig_fill.end_color else None
    )
else:
    copied_fill = PatternFill()
ws.cell(row=header_row, column=actual_age_col).fill = copied_fill
ws.cell(row=header_row, column=actual_age_col).alignment = Alignment(horizontal='center', vertical='top')

# Formula for Actual Age in each row (whole years)
dob_col = col_map['DATE OF BIRTH']
report_col = 13 # M (REPORT DATE per sample formulas)
for row in range(start_row, end_row+1):
    formula = f"=IF(AND(ISNUMBER(M{row}),ISNUMBER(C{row})),INT((M{row}-C{row})/365),\"\")"
    cell = ws.cell(row=row, column=actual_age_col, value=formula)
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    cell.fill = copied_fill

# Formula for Expected Result (E3:E15)
year_col = col_map['YEAR (age)']
category_col = col_map['CATEGORY']
expected_col = col_map['Expected Result']
for row in range(start_row, min(end_row+1, 16)):
    formula = f"=IFERROR(INDEX($I$3:$I$15, MATCH(O{row}, $H$3:$H$15, 0)), \"\")"
    ws.cell(row=row, column=expected_col, value=formula)

wb.save(output_path)
