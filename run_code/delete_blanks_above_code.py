from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/regression_gate/before_fix/core_374-31/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/regression_gate/before_fix/core_374-31/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Step 1: Find the first row where column A == 'Code'
anchor_row = None
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1), start=1):
    if (row[0].value is not None) and (str(row[0].value).strip() == 'Code'):
        anchor_row = i
        break

# Step 2: Collect blank rows strictly above anchor_row
rows_to_delete = []
if anchor_row is not None:
    for i in range(1, anchor_row):
        # Is entire row blank?
        is_blank = all(ws.cell(row=i, column=j).value in (None, "") for j in range(1, ws.max_column+1))
        if is_blank:
            rows_to_delete.append(i)

    # Step 3: Delete rows from bottom to top to avoid row shifting
    for idx in reversed(rows_to_delete):
        ws.delete_rows(idx)

wb.save(output_path)
