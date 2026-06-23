from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_3/regression_gate/before_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_3/regression_gate/before_pass/core_50526/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# B9 formula (drag down to B10):
# Dynamic lookup headers in B1:G1 (cols 2-7), lookup values in A2:A3
# Lookup value in B6
# Formula for B9, which returns the k-th matching header (k=1 for B9, k=2 for B10...)

for r in range(9, 11):
    k = r - 8  # 1 for B9, 2 for B10
    formula = (
        f'=IFERROR('
        f'INDEX($B$1:$G$1, '
        f'SMALL('
        f'IF(INDEX($B$2:$G$3, MATCH($B$6, $A$2:$A$3, 0), 0) > 0, '
        f'COLUMN($B$1:$G$1)-COLUMN($B$1)+1), '
        f'{k})), "")'
    )
    ws[f'B{r}'] = formula

wb.save(output_path)
