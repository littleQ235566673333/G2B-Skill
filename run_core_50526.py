from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_7/regression_gate/before_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_7/regression_gate/before_pass/core_50526/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

# Find lookup value from B6
lookup_value = ws['B6'].value

# Map all columns for easy reference
header_row = 1
lookup_col = 1
field_names = [ws.cell(row=header_row, column=col).value for col in range(2, ws.max_column+1)]
lookup_dict = {}
for row in range(2, ws.max_row+1):
    key = ws.cell(row=row, column=lookup_col).value
    if key is not None:
        vals = [ws.cell(row=row, column=col).value for col in range(2, ws.max_column+1)]
        lookup_dict[key] = vals
# Now, get index of the lookup value
vals = lookup_dict.get(lookup_value, [])
output_row = 9
fills = 0
for idx, v in enumerate(vals):
    if isinstance(v, (int, float)) and v > 0:
        ws.cell(row=output_row, column=2).value = field_names[idx]
        output_row += 1
        fills += 1
    if fills >= 2:
        break
wb.save(output_path)
