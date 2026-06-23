import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_32093_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_32093_tc1/output.xlsx'

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Process rows 2 to 15
for row in range(2, 16):
    # Get current employee name (B)
    current_employee = ws[f'B{row}'].value
    # Get new employee names (C, D, E)
    new_employees = [ws[f'{col}{row}'].value for col in ['C', 'D', 'E']]
    # Find first non-empty new employee name
    new_name = next((name for name in new_employees if name not in [None, '']), None)
    # If none, use current employee
    result = new_name if new_name else current_employee
    # Set result in column F
    ws[f'F{row}'] = result

# Save to output path
wb.save(output_path)
