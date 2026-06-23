import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_8/regression_gate/after_pass/core_50796/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_8/regression_gate/after_pass/core_50796/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 6):
    search_value = ws[f'A{row}'].value
    total = 0
    for data_row in range(12, 23):
        if ws[f'B{data_row}'].value == search_value:
            c_val = ws[f'C{data_row}'].value
            if c_val is not None and isinstance(c_val, (int, float)):
                total += c_val
    ws[f'B{row}'].value = total

wb.save(output_path)
