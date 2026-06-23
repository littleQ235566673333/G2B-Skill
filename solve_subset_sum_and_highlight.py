import openpyxl
from openpyxl.styles import PatternFill

# Path constants
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_254-34_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_254-34_tc1/output.xlsx'

# The valid target sums
valid_targets = list(range(994108, 994113))
found_comb = None
found_target = None

# Open workbook
wb = openpyxl.load_workbook(input_path)

# Delete sheet3 if present
if 'Sheet3' in wb.sheetnames:
    del wb['Sheet3']

ws = wb['Before']

# Collect all (cell, value) pairs that are numeric (skip headers and blanks)  
cells_vals = []
for row in ws.iter_rows():
    for cell in row:
        # Try to get only numeric (integer) values, skip header row/cols
        if isinstance(cell.value, (int, float)):
            cells_vals.append((cell, cell.value))

# Subset sum search
from itertools import combinations

def find_subset_sum(cells_vals, targets):
    values = [v for (c, v) in cells_vals]
    # Try all subset sizes from 2 to len(values) (single cells unlikely unless target is same as value)
    for r in range(2, min(len(values), 18)):
        for combo in combinations(enumerate(values), r):
            idxs, vals = zip(*combo)
            s = sum(vals)
            if s in targets:
                return [i for i in idxs], s
    # Try single values too
    for i, v in enumerate(values):
        if v in targets:
            return [i], v
    return None, None

idxs, found_target = find_subset_sum(cells_vals, valid_targets)
if idxs is not None and found_target is not None:
    # Set all highlights to none first
    for cell, _ in cells_vals:
        cell.fill = PatternFill(fill_type=None)
    # Apply highlight to the found combination
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    for i in idxs:
        cells_vals[i][0].fill = green_fill
    # Place the found target value in C2
    ws['C2'] = found_target
else:
    ws['C2'] = 'No combination found'

# Save the workbook
wb.save(output_path)
