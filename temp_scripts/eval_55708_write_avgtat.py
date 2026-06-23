import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun1/eval_55708_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun1/eval_55708_tc1/output.xlsx'
start_row = 10  # B11 is row 11 (0-based index is 10)
start_col = 2   # Column B (1-based index)

df = pd.read_excel(input_path)
department_col = 'Department'
status_col = 'Status'
tat_col = 'turnaroud time'  # Correct typo from inspection

results = []
groups = df.groupby(department_col)
for dept, group in groups:
    filtered = group[(group[status_col].isin(['In Progress', 'In Review'])) & (group[tat_col] >= 6)]
    if filtered.empty:
        results.append('')
    else:
        results.append(filtered[tat_col].mean())

wb = load_workbook(input_path)
ws = wb.active
for i, val in enumerate(results):
    row_num = start_row + i + 1
    ws.cell(row=row_num, column=start_col, value=val if val != '' else None)
wb.save(output_path)
