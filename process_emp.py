import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/group_547-43/r0/evolve_547-43/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/group_547-43/r0/evolve_547-43/output.xlsx'

# Read all sheets
ds = pd.read_excel(input_path, sheet_name=None)

emp_sheet, lookup_sheet = None, None
for sname in ds:
    if 'emp' in sname.lower():
        emp_sheet = sname
    if 'lookup' in sname.lower():
        lookup_sheet = sname
if not emp_sheet or not lookup_sheet:
    raise ValueError('Missing Emp or lookup sheet!')
df_emp = ds[emp_sheet]
df_lookup = ds[lookup_sheet]

comp_col = None
for col in df_emp.columns:
    if 'comp1' in str(col).lower():
        comp_col = col
        break
if comp_col is None:
    raise ValueError('Comp1-performance column not found!')

fixed_vars = [c for c in df_emp.columns if c != comp_col]
long_df = df_emp.melt(id_vars=fixed_vars, value_vars=[comp_col], var_name='measure', value_name='value')

# Merge by matching measure with Variable in lookup
merge_df = long_df.merge(df_lookup, left_on='measure', right_on='Variable', how='left')
if 'Category' in merge_df:
    merge_df.rename(columns={'Category':'category', 'Subcategory':'subcategory'}, inplace=True)
if 'category' not in merge_df:
    merge_df['category'] = ''
if 'subcategory' not in merge_df:
    merge_df['subcategory'] = ''

# Write to output Excel ('Expected'!A2:H19): output 18 rows, 8 columns
wb = load_workbook(input_path)
ws = wb['Expected'] if 'Expected' in wb.sheetnames else wb.create_sheet('Expected')

# Clear A2:H19
def clear(ws):
    for row in ws.iter_rows(min_row=2, max_row=19, min_col=1, max_col=8):
        for cell in row:
            cell.value = None
clear(ws)

out_df = merge_df.head(18)

for r, row in enumerate(out_df.itertuples(index=False), 2):
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val)

wb.save(output_path)
