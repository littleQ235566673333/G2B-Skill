from openpyxl import load_workbook
from decimal import Decimal

input_path = r"analysis/precheck_a/eval_spreadsheet_modeselector/task_142-12/input.xlsx"
output_path = r"analysis/precheck_a/eval_spreadsheet_modeselector/task_142-12/output.xlsx"

wb = load_workbook(input_path)
ws = wb["Sheet1"]

for row in range(1, ws.max_row + 1):
    f_val = ws.cell(row=row, column=6).value
    j_val = ws.cell(row=row, column=10).value

    if f_val == "Marble Slab Creamery" and j_val is not None:
        is_whole = False

        if isinstance(j_val, int):
            is_whole = True
        elif isinstance(j_val, float):
            is_whole = j_val.is_integer()
        elif isinstance(j_val, Decimal):
            is_whole = j_val == j_val.to_integral_value()
        else:
            try:
                num = float(j_val)
                is_whole = num.is_integer()
            except (TypeError, ValueError):
                is_whole = False

        if is_whole:
            ws.cell(row=row, column=6).value = "Georgia State WH"

wb.save(output_path)

# Verification pass
wb_check = load_workbook(output_path, data_only=False)
ws_check = wb_check["Sheet1"]
for cell in ["F10", "F13"]:
    pass
