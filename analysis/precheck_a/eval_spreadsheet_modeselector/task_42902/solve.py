from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_42902/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_42902/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# Collect values from column A, stopping at the last non-empty row.
values = []
for row in range(1, ws.max_row + 1):
    v = ws.cell(row=row, column=1).value
    if v is not None and v != '':
        values.append(v)

# Transpose each group of 3 rows into columns D:F, one record per row.
for i in range(0, len(values), 3):
    group = values[i:i+3]
    out_row = i // 3 + 1
    for j, val in enumerate(group):
        ws.cell(row=out_row, column=4 + j).value = val

wb.save(output_path)

# Verification
wb2 = load_workbook(output_path)
ws2 = wb2[wb2.sheetnames[0]]
for r in range(1, 8):
    print(r, [ws2.cell(row=r, column=c).value for c in range(4, 7)])
