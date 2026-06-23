from openpyxl import load_workbook
from pathlib import Path
import re

input_path = Path('analysis/precheck_a/eval_spreadsheet_modeselector/task_15380/input.xlsx')
output_path = Path('analysis/precheck_a/eval_spreadsheet_modeselector/task_15380/output.xlsx')

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

codes_range = wb.defined_names['codes2']
month_range = wb.defined_names['month']

code_values = []
month_values = []

for sheet_name, coord in codes_range.destinations:
    ws_codes = wb[sheet_name]
    for row in ws_codes[coord]:
        for cell in row:
            if cell.value is not None:
                code_values.append(str(cell.value).strip())

for sheet_name, coord in month_range.destinations:
    ws_month = wb[sheet_name]
    for row in ws_month[coord]:
        for cell in row:
            if cell.value is not None:
                month_values.append(str(cell.value).strip())

lookup = dict(zip(code_values, month_values))

for row in range(3, 15):
    text = ws[f'A{row}'].value
    result = None
    if text is not None:
        text = str(text)
        match = re.search(r'_(\d{2})_', text)
        if match:
            code = match.group(1)
            result = lookup.get(code)
        else:
            for code, month in lookup.items():
                if re.search(rf'(?<!\d){re.escape(code)}(?!\d)', text):
                    result = month
                    break
    ws[f'B{row}'] = result

wb.save(output_path)

wb2 = load_workbook(output_path)
ws2 = wb2[wb2.sheetnames[0]]
for row in range(3, 15):
    print(f'A{row}={ws2[f"A{row}"].value!r} -> B{row}={ws2[f"B{row}"].value!r}')
