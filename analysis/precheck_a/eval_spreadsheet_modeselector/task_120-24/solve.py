from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_120-24/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_120-24/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

lag_fill = PatternFill(fill_type='solid', fgColor='0070C0')

valid_bg = {
    'OCOGS - Spares - Transfer Price Overhead',
    'OFSS - ORCL Consulting Prod Cost I/C',
}
valid_ay = {'BOA_033E', 'BOA_011G'}

for row in range(2, 46):
    bl = ws[f'BL{row}'].value
    bg = ws[f'BG{row}'].value
    ay = ws[f'AY{row}'].value

    bl_text = str(bl).strip() if bl is not None else ''
    bg_text = str(bg).strip() if bg is not None else ''
    ay_text = str(ay).strip() if ay is not None else ''

    if bl_text == 'LAG' and bg_text in valid_bg:
        if ay_text in valid_ay:
            ws[f'BN{row}'] = 'OCOGS - Spares - Transfer Price Overhead'
        else:
            ws[f'BN{row}'] = 'OFSS - ORCL Consulting Prod Cost I/C'

    bn_val = ws[f'BN{row}'].value
    if bl_text == 'LAG' and bn_val not in (None, ''):
        ws[f'BL{row}'].fill = lag_fill

wb.save(output_path)
print(output_path)
