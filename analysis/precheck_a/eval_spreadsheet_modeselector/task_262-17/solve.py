from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_262-17/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_262-17/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# Detect headers dynamically from the first row
header_row = 1
headers = {}
for col in range(1, ws.max_column + 1):
    value = ws.cell(row=header_row, column=col).value
    if value is not None:
        headers[str(value).strip()] = col

required = ['Task', 'Responsibility']
for name in required:
    if name not in headers:
        raise ValueError(f"Required header '{name}' not found")

task_col = headers['Task']
responsibility_col = headers['Responsibility']

# Determine the real last data row
last_row = 1
for row in range(ws.max_row, 0, -1):
    if any(ws.cell(row=row, column=col).value is not None for col in range(1, ws.max_column + 1)):
        last_row = row
        break

# Read data rows, sort by Task ascending then Responsibility ascending
rows = []
for row in range(2, last_row + 1):
    rows.append([ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)])

rows.sort(key=lambda r: (
    '' if r[task_col - 1] is None else str(r[task_col - 1]),
    '' if r[responsibility_col - 1] is None else str(r[responsibility_col - 1])
))

# Write sorted rows back to Sheet1
for idx, row_values in enumerate(rows, start=2):
    for col, value in enumerate(row_values, start=1):
        ws.cell(row=idx, column=col).value = value

wb.save(output_path)
