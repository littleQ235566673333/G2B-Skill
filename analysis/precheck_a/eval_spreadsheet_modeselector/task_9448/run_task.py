from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

input_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_9448/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_9448/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Data']

# Inspect header dates in I:T and write formula in U9:U18 that returns the header
# corresponding to the rightmost numeric positive sale in the row.
for row in range(9, 19):
    start_col = 9   # I
    end_col = 20    # T

    # Build LOOKUP formula compatible with Excel 2010.
    # It returns the header from row 7 for the last cell in I:T on the same row with a numeric value > 0.
    sales_range = f"I{row}:T{row}"
    header_range = "$I$7:$T$7"
    formula = f'=IFERROR(LOOKUP(2,1/({sales_range}>0),{header_range}),"")'
    cell = ws[f'U{row}']
    cell.value = formula
    cell.number_format = 'mm/dd/yyyy'

wb.save(output_path)
