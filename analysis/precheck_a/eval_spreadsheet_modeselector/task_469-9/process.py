from openpyxl import load_workbook
from math import fabs

input_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_469-9/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_469-9/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

ws['H1'] = 'Debits'
ws['I1'] = 'Credits'

for row in range(2, ws.max_row + 1):
    value = ws[f'C{row}'].value
    ws[f'H{row}'] = None
    ws[f'I{row}'] = None
    if isinstance(value, (int, float)):
        if value < 0:
            ws[f'H{row}'] = abs(value)
        elif value > 0:
            ws[f'I{row}'] = abs(value)
    elif value is not None:
        try:
            num = float(value)
            if num < 0:
                ws[f'H{row}'] = abs(num)
            elif num > 0:
                ws[f'I{row}'] = abs(num)
        except (TypeError, ValueError):
            pass

wb.save(output_path)

# Verify saved output in target range
wb2 = load_workbook(output_path)
ws2 = wb2[wb2.sheetnames[0]]
for cell in ['H1','I1','H2','I2','H3','I3','H4','I4','H5','I5','H6','I6','H7','I7','H8','I8','H9','I9','H10','I10']:
    print(cell, ws2[cell].value)
