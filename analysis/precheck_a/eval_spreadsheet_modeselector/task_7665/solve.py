from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_7665/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_7665/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

values = []
row = 2
while True:
    val = ws[f'H{row}'].value
    if val is None:
        break
    values.append(val)
    row += 1

unique_sorted = sorted(set(values))

for idx in range(6):  # Q:V
    cell = ws.cell(row=2, column=17 + idx)
    cell.value = unique_sorted[idx] if idx < len(unique_sorted) else None

wb.save(output_path)
