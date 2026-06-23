from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_5835/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_modeselector/task_5835/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# Build lookup from the Basic Table (columns G:I)
lookup = {}
for row in range(3, ws.max_row + 1):
    year = ws.cell(row=row, column=7).value  # G
    stock = ws.cell(row=row, column=8).value  # H
    log_value = ws.cell(row=row, column=9).value  # I
    if year is None and stock is None:
        continue
    if log_value == '':
        log_value = None
    lookup[(stock, year)] = log_value

# Fill requested table Log Value in column C based on Stock+Year in columns A:B
for row in range(3, 20):
    stock = ws.cell(row=row, column=1).value  # A
    year = ws.cell(row=row, column=2).value   # B
    ws.cell(row=row, column=3).value = lookup.get((stock, year), None)

wb.save(output_path)

# Verification
wb2 = load_workbook(output_path)
ws2 = wb2[wb2.sheetnames[0]]
for row in range(3, 20):
    print(row, ws2.cell(row=row, column=3).value)
