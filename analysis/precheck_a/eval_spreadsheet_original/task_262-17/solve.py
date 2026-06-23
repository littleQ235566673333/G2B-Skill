from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_262-17/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_262-17/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# Read header row and dynamically locate the requested sort columns.
headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
header_to_col = {header: idx + 1 for idx, header in enumerate(headers) if header is not None}

task_col = header_to_col['Task']
resp_col = header_to_col['Responsibility']

# Capture all data rows from the first worksheet.
data_rows = []
for row in range(2, ws.max_row + 1):
    values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
    if any(value is not None for value in values):
        data_rows.append(values)

# Sort by Task ascending, then by Responsibility ascending.
data_rows.sort(key=lambda r: (r[task_col - 1], r[resp_col - 1]))

# Write the sorted result back to Sheet1, preserving the header row.
for row_idx, values in enumerate(data_rows, start=2):
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx).value = value

wb.save(output_path)
