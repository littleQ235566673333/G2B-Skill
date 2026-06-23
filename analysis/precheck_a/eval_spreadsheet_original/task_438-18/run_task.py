from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_438-18/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_438-18/output.xlsx'

mapping = {
    'UNATIDE STATES AMERICAN': 'USA',
    'FRANCE': 'FRA',
    'KOREA': 'KOR',
    'GERMANY': 'GR',
    'ITALY': 'IT',
}

wb = load_workbook(input_path)
ws_in = wb['INPUT']
ws_out = wb['OUPUT']

# Write headers for the output area
ws_out['A1'] = 'ITEM'
ws_out['B1'] = 'BRANDS'
ws_out['C1'] = 'TYPE'
ws_out['D1'] = 'ORIGIN'

# Copy data from INPUT to OUPUT with origin replacement
for in_row, out_row in zip(range(2, ws_in.max_row + 1), range(2, 7)):
    ws_out.cell(row=out_row, column=1).value = ws_in.cell(row=in_row, column=1).value  # ITEM
    ws_out.cell(row=out_row, column=2).value = ws_in.cell(row=in_row, column=3).value  # BRANDS
    ws_out.cell(row=out_row, column=3).value = ws_in.cell(row=in_row, column=5).value  # TYPE
    origin = ws_in.cell(row=in_row, column=6).value
    ws_out.cell(row=out_row, column=4).value = mapping.get(origin, origin)

wb.save(output_path)

# Verify output cells A1:D6
check_wb = load_workbook(output_path)
check_ws = check_wb['OUPUT']
for row in check_ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=4, values_only=True):
    print(row)
