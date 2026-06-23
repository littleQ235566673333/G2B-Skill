from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_141-20/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_141-20/output.xlsx'

wb = load_workbook(input_path)
pl_ws = wb['PL Recon Items']
st_ws = wb['Statement Recon Items']

def norm(v):
    if v is None:
        return ''
    if isinstance(v, str):
        return v.strip()
    return v

# Gather data rows excluding header row
pl_matches = {}
for r in range(2, pl_ws.max_row + 1):
    inv = norm(pl_ws.cell(r, 3).value)
    val = norm(pl_ws.cell(r, 4).value)
    if inv != '' and val != '':
        pl_matches.setdefault((inv, val), []).append(r)

st_matches = {}
for r in range(2, st_ws.max_row + 1):
    ref = norm(st_ws.cell(r, 6).value)
    val = norm(st_ws.cell(r, 9).value)
    if ref != '' and val != '':
        st_matches.setdefault((ref, val), []).append(r)

common_keys = set(pl_matches) & set(st_matches)
pl_delete = []
st_delete = []
for key in common_keys:
    count = min(len(pl_matches[key]), len(st_matches[key]))
    pl_delete.extend(pl_matches[key][:count])
    st_delete.extend(st_matches[key][:count])

for r in sorted(set(pl_delete), reverse=True):
    pl_ws.delete_rows(r, 1)
for r in sorted(set(st_delete), reverse=True):
    st_ws.delete_rows(r, 1)

wb.save(output_path)

# verify workbook can be reopened and target sheets exist
wb2 = load_workbook(output_path)
assert 'PL Recon Items' in wb2.sheetnames
assert 'Statement Recon Items' in wb2.sheetnames
print('Saved', output_path)
print('PL deleted rows:', sorted(pl_delete))
print('Statement deleted rows:', sorted(st_delete))
for rng in [('PL Recon Items',1,2,1,4), ('Statement Recon Items',1,3,1,10)]:
    ws = wb2[rng[0]]
    vals = []
    for rr in range(rng[1], rng[2]+1):
        vals.append([ws.cell(rr, cc).value for cc in range(rng[3], rng[4]+1)])
    print(rng[0], vals)
