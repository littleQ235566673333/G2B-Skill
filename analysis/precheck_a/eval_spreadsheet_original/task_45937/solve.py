from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_45937/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_45937/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# Column D returns category A-E from the B-value legend in H6:I11.
# Complete column E with a formula that picks the percentage from the table H17:N22,
# using the kms range in H20:I22 and the category headers J18:N18 with percentages in J20:N22.
for row in range(7, 10):
    ws[f'E{row}'] = (
        f'=INDEX($J$20:$N$22,'
        f'MATCH(B{row},$H$20:$H$22,1),'
        f'MATCH(D{row},$J$18:$N$18,0))'
    )

wb.save(output_path)
