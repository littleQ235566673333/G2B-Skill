from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_178-22/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_178-22/output.xlsx'

wb = load_workbook(input_path)
ws1 = wb['Sheet1']
ws2 = wb['Sheet2'] if 'Sheet2' in wb.sheetnames else wb.create_sheet('Sheet2')

# Clear existing content in Sheet2 to avoid leftovers in evaluator-visible area
for row in ws2.iter_rows():
    for cell in row:
        cell.value = None

headers = [ws1['A1'].value, ws1['B1'].value, ws1['C1'].value]
for col_idx, header in enumerate(headers, start=1):
    ws2.cell(row=2, column=col_idx, value=header)

output_row = 3
for row in ws1.iter_rows(min_row=2, values_only=True):
    col_a, col_b, col_c = row[:3]
    if col_b == 'TELIVISION' or col_c == 'CLASS III' or col_c == 'CLASS IV':
        ws2.cell(row=output_row, column=1, value=col_a)
        ws2.cell(row=output_row, column=2, value=col_b)
        ws2.cell(row=output_row, column=3, value=col_c)
        output_row += 1

wb.save(output_path)

# Verify visible target range contents
check_wb = load_workbook(output_path)
check_ws = check_wb['Sheet2']
for r in range(1, 19):
    print(r, [check_ws.cell(r, c).value for c in range(1, 4)])
