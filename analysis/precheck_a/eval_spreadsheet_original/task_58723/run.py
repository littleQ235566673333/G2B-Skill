from openpyxl import load_workbook
from collections import defaultdict

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_58723/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_58723/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

latest_by_name = {}
for row in range(2, 42):
    name = ws.cell(row=row, column=3).value  # Column C
    entry_time = ws.cell(row=row, column=9).value  # Column I
    if name is None or entry_time is None:
        continue
    if name not in latest_by_name or entry_time > latest_by_name[name]:
        latest_by_name[name] = entry_time

for row in range(2, 42):
    name = ws.cell(row=row, column=3).value
    entry_time = ws.cell(row=row, column=9).value
    target = ws.cell(row=row, column=13)  # Column M
    if name is None or entry_time is None:
        target.value = None
    elif entry_time == latest_by_name.get(name):
        target.value = 'Latest'
    else:
        target.value = 'Not Latest'

wb.save(output_path)

# verify
wb2 = load_workbook(output_path)
ws2 = wb2[wb2.sheetnames[0]]
for row in range(2, 42):
    assert ws2.cell(row=row, column=13).value in ('Latest', 'Not Latest')
print('saved', output_path)
