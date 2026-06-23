from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_38537/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_38537/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

prefix = ws['B2'].value or ''
start_num = 10
count = 0

for row in range(3, 37):
    b_val = ws[f'B{row}'].value
    if b_val not in (None, ''):
        seq = f"{start_num + count:03d}"
        ws[f'A{row}'] = f"{prefix}-{b_val}-{seq}"
        count += 1
    else:
        ws[f'A{row}'] = None

wb.save(output_path)

# verify
wb2 = load_workbook(output_path)
ws2 = wb2[wb2.sheetnames[0]]
for row in range(3, 37):
    _ = ws2[f'A{row}'].value
