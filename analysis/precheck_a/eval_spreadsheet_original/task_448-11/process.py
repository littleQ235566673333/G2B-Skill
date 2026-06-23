from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_448-11/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_448-11/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Current region starts at A2 and includes the header row plus contiguous data block.
# Delete duplicate data rows based only on Column A, keeping the first occurrence.
start_row = 2
end_row = start_row
while end_row <= ws.max_row and any(ws.cell(end_row, c).value is not None for c in range(1, ws.max_column + 1)):
    end_row += 1
end_row -= 1

seen = set()
rows_to_delete = []
for row in range(start_row + 1, end_row + 1):  # skip header at row 2
    key = ws.cell(row=row, column=1).value
    if key in seen:
        rows_to_delete.append(row)
    else:
        seen.add(key)

for row in reversed(rows_to_delete):
    ws.delete_rows(row, 1)

wb.save(output_path)
