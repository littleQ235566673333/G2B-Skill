from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_32093/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_32093/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

for row in range(2, 16):
    formula = f'=IF(C{row}<>"",C{row},IF(D{row}<>"",D{row},IF(E{row}<>"",E{row},B{row})))'
    ws[f'F{row}'] = formula

wb.save(output_path)
