from openpyxl import load_workbook

INPUT_PATH = 'analysis/precheck_a/eval_spreadsheet_original/task_524-31/input.xlsx'
OUTPUT_PATH = 'analysis/precheck_a/eval_spreadsheet_original/task_524-31/output.xlsx'
SHEET_NAME = 'Exp-DB'

wb = load_workbook(INPUT_PATH)
ws = wb[SHEET_NAME]

# Build a vendor-to-category map from columns A:B.
vendor_map = []
for row in range(1, 54):
    vendor = ws.cell(row=row, column=1).value
    category = ws.cell(row=row, column=2).value
    if vendor and category:
        vendor_map.append((str(vendor), category))

# Match each transaction description in column D against the shortened vendor text.
for row in range(1, 54):
    description = ws.cell(row=row, column=4).value
    matched_category = None

    if isinstance(description, str):
        for vendor, category in vendor_map:
            if description.startswith(vendor):
                matched_category = category
                break

    ws.cell(row=row, column=5).value = matched_category

wb.save(OUTPUT_PATH)
