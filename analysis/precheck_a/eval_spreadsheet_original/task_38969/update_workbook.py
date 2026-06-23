from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_38969/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_38969/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

for row in range(2, 21):
    formula = (
        f'=IF(ISERROR(S{row}),"Upload",'
        f'IF(AND(U{row}>=-1,U{row}<=1),"Do not Upload",'
        f'IF(U{row}>1,"to Check","")))'
    )
    ws[f'R{row}'] = formula

wb.save(output_path)
