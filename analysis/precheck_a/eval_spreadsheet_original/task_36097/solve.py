from openpyxl import load_workbook

input_path = 'analysis/precheck_a/eval_spreadsheet_original/task_36097/input.xlsx'
output_path = 'analysis/precheck_a/eval_spreadsheet_original/task_36097/output.xlsx'

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

for row in range(3, 7):
    cost = ws[f'C{row}'].value or 0
    itv = ws[f'E{row}'].value or 0
    profit = ws[f'F{row}'].value or 0

    if profit < 0:
        recoupment = itv + profit
    elif profit < cost:
        recoupment = profit
    else:
        recoupment = cost - itv

    ws[f'H{row}'] = recoupment

wb.save(output_path)
