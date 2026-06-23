import pandas as pd
from openpyxl import load_workbook
import numpy as np

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/group_3911/r3/evolve_3911/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/group_3911/r3/evolve_3911/output.xlsx'

# Load the data
df = pd.read_excel(input_path)
salary_col = 'NetSalary'

# Remove non-numeric salary entries (likely headers, totals, etc.)
df = df[pd.to_numeric(df[salary_col], errors='coerce').notna()]
salaries = df[salary_col].astype(float).values
n_notes = 120
total_salary = salaries.sum()

# Initial allocation by proportion
note_alloc = np.floor(n_notes * salaries / total_salary).astype(int)
allocated = note_alloc.sum()
remainder = n_notes - allocated
if remainder > 0:
    fractions = n_notes * salaries / total_salary - note_alloc
    order = np.argsort(-fractions)
    for i in range(remainder):
        note_alloc[order[i]] += 1

# Write result to H2:P8
wb = load_workbook(input_path)
ws = wb.active
row0, col0 = 2, 8
for idx, val in enumerate(note_alloc):
    ws.cell(row=row0+idx, column=col0).value = val
wb.save(output_path)
