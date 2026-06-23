from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_341-14_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_341-14_tc1/output.xlsx'

wb = load_workbook(input_path)
ws_problem = wb['problem']
ws_result = wb['result']

# Read data from 'problem'
A_vals = []
B_vals = []
for row in ws_problem.iter_rows(min_row=1, max_row=ws_problem.max_row, min_col=1, max_col=2):
    a, b = row
    if a.value is None and b.value is None:
        continue
    A_vals.append(a.value)
    B_vals.append(b.value)

# Build the layout of the result sheet
# Read what's already in the 'result' sheet
result_vals = []
for row in ws_result.iter_rows(min_row=1, max_row=28, min_col=1, max_col=2):
    a, b = row
    result_vals.append([a.value, b.value])

# Build a lookup from A -> B, include only if A is not None
lookup = {a: b for a, b in zip(A_vals, B_vals)}

# For each entry in column A of the 'result' sheet (A1:A28),
# If it is a key in lookup, fill B; else, leave as None
for i, (label, num) in enumerate(result_vals):
    out_label = label
    out_val = lookup.get(label, None)
    ws_result.cell(row=1 + i, column=1).value = out_label
    ws_result.cell(row=1 + i, column=2).value = out_val

wb.save(output_path)
