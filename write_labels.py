from openpyxl import load_workbook

# Load workbook and worksheet
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_2/regression_gate/before_pass/core_50526/input.xlsx')
ws = wb['Sheet1']

# Get lookup value from B6
lookup_cell = ws['B6'].value

# Build header and lookup map (first row as header, next rows as data)
header = [cell.value for cell in ws[1]][1:]  # Skip first cell (none)
lookup_map = {}
for row in ws.iter_rows(min_row=2, max_row=3, values_only=True):
    key = row[0]
    values = row[1:]
    lookup_map[key] = values

# Get the labels corresponding to the values > 0 for the selected lookup value
labels = header
selected_values = lookup_map.get(lookup_cell, [])
result = [labels[i] for i, v in enumerate(selected_values) if v and v > 0]

# Write found results to B9 and B10
for idx, val in enumerate(result):
    ws.cell(row=9+idx, column=2).value = val
# Clear out extra cells if fewer than 2 values found
for clear_idx in range(len(result), 2):
    ws.cell(row=9+clear_idx, column=2).value = None

# Save workbook
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_2/regression_gate/before_pass/core_50526/output.xlsx')
