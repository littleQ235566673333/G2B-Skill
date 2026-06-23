from openpyxl import load_workbook

def process_sheet(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]
    min_row, max_row = 3, 36  # inclusive, per instructions
    # Step 1: Clear F where E is empty in the block
    for row in range(min_row, max_row + 1):
        e_val = ws.cell(row=row, column=5).value  # E
        if e_val is None or (str(e_val).strip() == ''):
            ws.cell(row=row, column=6).value = None  # F

    # Step 2: Copy A downward, capped at an empty F, but skip if A matches known index ('column1')
    # First, get all the index anchor values in Col A in range
    index_values = set()
    for row in range(min_row, max_row + 1):
        av = ws.cell(row=row, column=1).value
        if av is not None and av not in (None, '') and str(av).startswith('column'):
            index_values.add(av)
    # Now iterate and fill
    current_anchor = None
    for row in range(min_row, max_row + 1):
        a = ws.cell(row=row, column=1).value
        f = ws.cell(row=row, column=6).value
        # Only set current anchor if it's NOT an index
        if a not in index_values and a is not None:
            current_anchor = a
        # If A is empty and F is not blank, copy anchor (unless next in index row)
        if (a is None or a == '') and f not in (None, ''):
            # Peek ahead to next A for index check
            is_index_row = False
            if row + 1 <= max_row:
                a_next = ws.cell(row=row + 1, column=1).value
                # If the next A value is a known index, don't fill this one
                if a_next in index_values:
                    is_index_row = True
            if (not is_index_row) and current_anchor is not None and str(current_anchor) not in index_values:
                ws.cell(row=row, column=1).value = current_anchor
        # Stop copy process if F is blank
        if f in (None, ''):
            current_anchor = None

    wb.save(output_path)

process_sheet('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun2/eval_304-35_tc1/input.xlsx',
              'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun2/eval_304-35_tc1/output.xlsx')
