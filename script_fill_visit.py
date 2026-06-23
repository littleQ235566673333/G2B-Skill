import pandas as pd
from openpyxl import load_workbook

df = pd.read_excel('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_2/regression_gate/before_pass/core_9726/input.xlsx')

# Table 1
table1 = df[['Student ID', 'Visit number', 'Visit date']].dropna(subset=['Student ID', 'Visit number', 'Visit date'])
# Table 2
# rows with Student ID.1 and Submission date (Submission date non-null)
table2 = df[['Student ID.1', 'Submission date']].dropna(subset=['Student ID.1', 'Submission date']).reset_index()

# Ensure dates are datetime
table1['Visit date'] = pd.to_datetime(table1['Visit date'])
table2['Submission date'] = pd.to_datetime(table2['Submission date'])

# For each row in table2, find the latest visit date for the same student id < submission date
results = []
for idx, row in table2.iterrows():
    sid = row['Student ID.1']
    sdate = row['Submission date']
    visits = table1[(table1['Student ID'] == sid) & (table1['Visit date'] < sdate)]
    if not visits.empty:
        val = visits['Visit date'].max().date()
    else:
        val = None
    results.append(val)

table2['Visit date.1'] = results

# Open Excel and write results to the correct cells (J2..)
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_2/regression_gate/before_pass/core_9726/input.xlsx')
ws = wb['Sheet1']

for i, date_val in zip(table2['index']+2, table2['Visit date.1']):
    ws.cell(row=i, column=10).value = date_val

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_2/regression_gate/before_pass/core_9726/output.xlsx')
