import openpyxl

# Load workbook
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_48983_tc1/input.xlsx')
ws = wb.active

# Parse source table headers (categories): C5:I5 (col 3 to 9, row 5)
source_cat_row = 5
source_cat_cols = list(range(3, 10))
source_categories = [ws.cell(row=source_cat_row, column=col).value for col in source_cat_cols]

# Parse source table brands: B6:B11 (col 2, rows 6-11)
source_brand_col = 2
source_brand_rows = list(range(6, 12))
source_brands = [ws.cell(row=row, column=source_brand_col).value for row in source_brand_rows]

# Build source value lookup: {(brand, category): value}
lookup = {}
for i, row in enumerate(source_brand_rows):
    brand = ws.cell(row=row, column=source_brand_col).value
    for j, col in enumerate(source_cat_cols):
        category = ws.cell(row=source_cat_row, column=col).value
        value = ws.cell(row=row, column=col).value
        if value is not None:
            lookup[(brand, category)] = value

# Target table: M6:S11 (cols 13:19, rows 6:11)
output_start_col = 13
output_end_col = 19
output_start_row = 6
output_end_row = 11

# Target column headers (categories) in row 6: M6:S6
output_categories = [ws.cell(row=6, column=col).value for col in range(output_start_col, output_end_col+1)]
# Target row headers (brands) in col M: M7:M11
output_brands = [ws.cell(row=row, column=13).value for row in range(7, 12)]

# Fill in the output table
for i, row in enumerate(range(7, 12)):
    brand = ws.cell(row=row, column=13).value
    if brand is None:
        continue
    for j, col in enumerate(range(output_start_col+1, output_end_col+1)):
        category = ws.cell(row=6, column=col).value
        value = lookup.get((brand, category), None)
        ws.cell(row=row, column=col, value=value)

# Save the result
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_48983_tc1/output.xlsx')
print('Done')
