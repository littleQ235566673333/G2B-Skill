import openpyxl
import pandas as pd

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r3/eval_61-4_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r3/eval_61-4_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheetnames = wb.sheetnames
if 'output' not in sheetnames:
    ws_out = wb.create_sheet('output')
else:
    ws_out = wb['output']

input_sheet = [s for s in sheetnames if s.lower() != 'output'][0]
ws = wb[input_sheet]

# Only columns A:H are canonical for the first data table
header_row_idx = 1
row_gen = ws.iter_rows(min_row=header_row_idx+1, max_col=8, values_only=True)
header_cells = [str(c).strip().lower().replace('*','').replace(' ','') for c in ws[header_row_idx][:8]]
# Map for normalized header columns
header_map = {h:i for i,h in enumerate(header_cells)}

# Canonical map: date=A, stock=B, open=C, high=D, low=E, close=F, volume=G, change=H
col_date = header_map.get('date',0)
col_stock = header_map.get('stockname',1)
col_open = header_map.get('openp',2)
col_high = header_map.get('high',3)
col_low = header_map.get('low',4)
col_close = header_map.get('closep',5)
col_vol = header_map.get('volume',6)
col_change = header_map.get('change',7)

records = []
for row in ws.iter_rows(min_row=header_row_idx+1, max_col=8, values_only=True):
    if all(cell is None for cell in row):
        continue
    records.append([row[col_date], row[col_stock], row[col_open], row[col_high], row[col_low], row[col_close], row[col_vol], row[col_change]])
df = pd.DataFrame(records, columns=['date','stock','open','high','low','close','volume','change'])

# Parse dates and sort
try:
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
except:
    pass
df = df.dropna(subset=['date']).sort_values(['stock','date']).reset_index(drop=True)

result = []
for stock, sdf in df.groupby('stock'):
    sdf = sdf.reset_index(drop=True)
    is_neg = sdf['change'].apply(pd.to_numeric, errors='coerce') < 0
    i = 0
    while i < len(sdf):
        if is_neg.iloc[i]:
            start = i
            while i+1 < len(sdf) and is_neg.iloc[i+1]:
                i += 1
            end = i
            block = sdf.iloc[start:end+1]
            open_ = block.iloc[0]['open']
            close_ = block.iloc[-1]['close']
            high_ = pd.to_numeric(block['high'], errors='coerce').max()
            low_ = pd.to_numeric(block['low'], errors='coerce').min()
            vol_ = pd.to_numeric(block['volume'], errors='coerce').sum()
            result.append([
                stock,
                block.iloc[0]['date'],
                block.iloc[-1]['date'],
                open_,
                close_,
                high_,
                low_,
                vol_
            ])
        i += 1

output_headers = ['Stock Name','From Date','To Date','Opening','Closing','High','Low','Volume']
for j, h in enumerate(output_headers, 1):
    ws_out.cell(row=1, column=j, value=h)
for i, row in enumerate(result, 2):
    for j, v in enumerate(row, 1):
        ws_out.cell(row=i, column=j, value=v)
wb.save(output_path)
print('Wrote output to', output_path)