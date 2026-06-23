import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_209-30_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_209-30_tc1/output.xlsx'

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Data to Import']

# Process column C starting from row 2 to the last used row in C
for row in range(2, ws.max_row + 1):
    cell = ws[f'C{row}']
    val = str(cell.value) if cell.value is not None else ''
    if len(val) > 3:
        cell.value = val[:-3]
    else:
        cell.value = ''

# Save the updated workbook
wb.save(output_path)
