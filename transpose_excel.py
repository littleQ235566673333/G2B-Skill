from openpyxl import load_workbook

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/train/iter_4/evolve_147-48/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/train/iter_4/evolve_147-48/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Collect non-empty values from column A
col_a = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1) if ws.cell(row=i, column=1).value is not None]

start_row, start_col = 1, 3  # C1 is (1, 3)
rows, cols = 6, 9  # Up to K6 (row 6, column 11)

for idx, val in enumerate(col_a):
    if idx >= rows * cols:
        break
    row = start_row + idx // cols
    col = start_col + idx % cols
    ws.cell(row=row, column=col).value = val

wb.save(output_path)
