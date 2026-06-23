from openpyxl import load_workbook

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_7/regression_gate/before_pass/core_56599/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_7/regression_gate/before_pass/core_56599/output.xlsx"

wb = load_workbook(input_path)
sheet = wb.active

# Get slip number to look for from M5
slip_cell = sheet['M5'].value

# Handle blank input
if slip_cell is None or str(slip_cell).strip() == "":
    sheet['B2'] = 'not found - please re-enter'
else:
    # Gather all slip numbers into a flat list (excluding blanks, must be 7 digit numbers)
    slip_numbers = []
    slip_map = {} # mapping: slip_number (int) -> position (index)
    for row in sheet.iter_rows():
        for cell in row:
            val = cell.value
            if val is not None and str(val).isdigit() and len(str(val)) == 7:
                num = int(val)
                slip_map[num] = len(slip_numbers) # allows lookup
                slip_numbers.append(num)
    # Find slip position
    try:
        slip_cell_int = int(slip_cell)
        if slip_cell_int in slip_map:
            idx = slip_map[slip_cell_int]
            book_number = (idx // 10) + 1 # Book number 1-based
            sheet['B2'] = book_number
        else:
            sheet['B2'] = 'not found - please re-enter'
    except Exception:
        sheet['B2'] = 'not found - please re-enter'

wb.save(output_path)
