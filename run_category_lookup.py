import openpyxl
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_6/regression_gate/after_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_6/regression_gate/after_pass/core_18935/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Data Table: header row 4, data rows 5–10, columns A-F
headers = [ws.cell(row=4, column=c).value for c in range(1, 6)]
data = []
for r in range(5, 11):
    row = [ws.cell(row=r, column=c).value for c in range(1, 6)]
    if row[0] is None:
        break
    data.append(row)

# Build search structure
data_rows = []
for row in data:
    d = dict(zip(headers, row))
    data_rows.append(d)

# Category columns in the header (excluding Work and Material)
category_cols = [h for h in headers if h not in ("Work", "Material")]

# For each report row (17-22), fill in D with the looked-up value
for r in range(17, 23):
    work = ws[f'A{r}'].value
    category_type = ws[f'B{r}'].value  # e.g., Category 1
    material = ws[f'C{r}'].value

    # Find row in data_rows matching Work and Material
    val = ''
    for d in data_rows:
        if d['Work'] == work and d['Material'] == material:
            val = d.get(category_type, '')
            break
    ws[f'D{r}'].value = val

wb.save(output_path)
