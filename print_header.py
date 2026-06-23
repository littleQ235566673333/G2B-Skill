from openpyxl import load_workbook
ws = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_4/group_177-6/r2/evolve_177-6/input.xlsx').active
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(header)