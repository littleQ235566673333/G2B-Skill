import openpyxl
import pandas as pd

# Load workbook and worksheet
your_input = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_493-5_tc1/input.xlsx'
your_output = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_493-5_tc1/output.xlsx'
wb = openpyxl.load_workbook(your_input)
ws = wb['Imported Data']
data = list(ws.values)

# Parse headers and DataFrame
cols = data[0]
df = pd.DataFrame(data[1:], columns=cols)
df['__rownum'] = range(len(df))

# Find pairs to delete based on criteria
to_delete = set()
for i, row_i in df.iterrows():
    for j, row_j in df.iterrows():
        if (
            i != j and
            row_i[cols[0]] == row_j[cols[0]] and  # Reference Number
            row_i[cols[5]] == row_j[cols[5]] and  # Narrative
            row_i[cols[2]] == row_j[cols[3]]      # Col C of one == Col D of another
        ):
            to_delete.add(i)
            to_delete.add(j)

# Keep rows not marked for deletion
df_keep = df.drop(list(to_delete)).reset_index(drop=True)
out_data = [cols] + df_keep.drop(columns='__rownum').values.tolist()

# Write output to new workbook
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Imported Data'
for row_idx, row in enumerate(out_data[:10], 1):
    for col_idx, val in enumerate(row, 1):
        ws_out.cell(row=row_idx, column=col_idx, value=val)
wb_out.save(your_output)
