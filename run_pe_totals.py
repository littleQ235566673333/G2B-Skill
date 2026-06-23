import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_2/group_47766/r3/evolve_47766/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_2/group_47766/r3/evolve_47766/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Data blocks
prod_blocks = [(8, 37), (41, 58), (62, 74)]
pe_totals_by_year = {}
for block_start, block_end in prod_blocks:
    for row in range(block_start, block_end + 1):
        tag = ws[f'H{row}'].value
        amt = ws[f'C{row}'].value
        close = ws[f'F{row}'].value
        # Check for PE in tag and valid close date
        if tag and 'PE' in str(tag) and amt:
            try:
                close_dt = close
                if isinstance(close_dt, str):
                    close_dt = pd.to_datetime(close_dt, errors='coerce')
                if pd.isnull(close_dt):
                    continue
                year = close_dt.year
                pe_totals_by_year.setdefault(year, 0)
                pe_totals_by_year[year] += float(amt)
            except Exception:
                continue

# Write results in K40 (col 11, row 40) and to the right
# Read column headers from row 39, cols 11-15 (K-O)
years = []
for col in range(11, 16):
    yr = ws.cell(row=39, column=col).value
    years.append(yr)
for i, year in enumerate(years):
    val = round(pe_totals_by_year.get(year, 0), 2) if year else ''
    ws.cell(row=40, column=11 + i).value = val if val != 0 else ''

wb.save(output_path)
