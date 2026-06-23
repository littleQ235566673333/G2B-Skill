from openpyxl import load_workbook
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']

today = datetime.today()
cell_h = ws['H4'].value
cell_i = ws['I4'].value

result = 'NO ACTION'
if cell_h and cell_i and str(cell_i).strip().upper() == 'YES':
    h_date = None
    if isinstance(cell_h, datetime):
        h_date = cell_h
    else:
        try:
            h_date = datetime.strptime(str(cell_h), '%Y-%m-%d')
        except:
            try:
                h_date = datetime.strptime(str(cell_h), '%m/%d/%Y')
            except:
                h_date = None
    if h_date:
        if (today - h_date).days <= 30:
            result = 'HOLD'
        else:
            result = 'TOUCH BASE'
ws['J4'] = result
wb.save(output_path)
