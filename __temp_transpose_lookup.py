import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_7/regression_gate/after_fix/core_38985/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_7/regression_gate/after_fix/core_38985/output.xlsx'

# Load the sheet
sheets = pd.read_excel(input_path, sheet_name=None, header=None)
df = sheets['Sheet1']

# Table 1: names in col 4 (Unnamed: 2), rows 6–9 (zero indexed: rows 5–8)
names_col1 = df.iloc[5:9, 2].tolist()

# Table 2: keys in col 12 (Unnamed: 11), values in col 13 (Unnamed: 12) and/or col 14 (Unnamed: 13)
t2_start_row = 5
keys = df.iloc[t2_start_row:t2_start_row+4, 11].tolist()
vals1 = df.iloc[t2_start_row:t2_start_row+4, 12].tolist()
vals2 = df.iloc[t2_start_row:t2_start_row+4, 13].tolist()

# For this task, assume we want vals1 (Value 1) unless further instruction—can adapt for both
lookup_dict = defaultdict(list)
for k, v in zip(keys, vals1):
    lookup_dict[k].append(v)

# Perform the lookup (main Table 1 col values across Table 2)
output_values = []
for name in names_col1:
    vals = lookup_dict.get(name, [None])
    if len(vals) > 1:
        output = ', '.join(str(v) for v in vals)
    else:
        output = vals[0]
    output_values.append(output)

# Write results to D8:D11 (rows 7,8,9,10 index, 1-based)
wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]
for i, val in enumerate(output_values):
    ws[f'D{8+i}'] = val
wb.save(output_path)
