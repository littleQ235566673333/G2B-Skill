from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_1/regression_gate/before_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_1/regression_gate/before_pass/core_41601/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Students']

student_names = [ws[f'A{r}'].value for r in range(2, 8)]

for i, name in enumerate(student_names):
    cell = ws[f'E{i+2}']
    if name:
        # Surround with single quotes if not just alphanumeric
        sheet_ref = f"'{name}'" if not str(name).isalnum() else str(name)
        cell.value = f"={sheet_ref}!C2"

wb.save(output_path)
