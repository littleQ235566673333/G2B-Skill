import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_267-21_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_267-21_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sh_merge = wb['merging']
sh_sh1 = wb['SH1']
sh_rp1 = wb['RP1']

# Build ID to QTY mapping for SH1 and RP1
sh1_dict = {}
rp1_dict = {}

for row in sh_sh1.iter_rows(min_row=2, values_only=True):
    id_val = row[0]  # Assuming ID is in A
    qty = row[1]     # QTY is in B
    if id_val is not None:
        sh1_dict[id_val] = qty

for row in sh_rp1.iter_rows(min_row=2, values_only=True):
    id_val = row[0]
    qty = row[1]
    if id_val is not None:
        rp1_dict[id_val] = qty

for i in range(2, 12):  # Rows 2-11
    row_id = sh_merge[f'B{i}'].value
    # SH1 QTY
    qty_sh1 = sh1_dict.get(row_id, '-')
    if qty_sh1 in [None, '']:
        qty_sh1 = '-'
    sh_merge[f'C{i}'] = qty_sh1
    # RP1 QTY
    qty_rp1 = rp1_dict.get(row_id, '-')
    if qty_rp1 in [None, '']:
        qty_rp1 = '-'
    sh_merge[f'D{i}'] = qty_rp1

wb.save(output_path)
