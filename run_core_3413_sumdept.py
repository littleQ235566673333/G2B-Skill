import openpyxl
import pandas as pd

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/regression_gate/before_pass/core_3413/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/regression_gate/before_pass/core_3413/output.xlsx'
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Build DataFrame from A2:C (Dept, RU, Total)
data_rows = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3, values_only=True):
    if all(cell is None for cell in row):
        break
    data_rows.append(row)
df = pd.DataFrame(data_rows, columns=['Dept', 'RU', 'Total'])

# For each G3:G6 (rows 3-6)
results = []
for i in range(3, 7):  # Excel rows 3,4,5,6
    dept = ws[f'E{i}'].value
    ru = ws[f'F{i}'].value
    # Try to match both Dept and RU
    mask = (df['Dept'] == dept) & (df['RU'] == ru)
    sum_match = df.loc[mask, 'Total'].sum()
    if sum_match == 0:
        sum_dept = df.loc[df['Dept'] == dept, 'Total'].sum()
        val = sum_dept
    else:
        val = sum_match
    ws[f'G{i}'] = val
wb.save(output_file)
