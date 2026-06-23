import openpyxl

# Load the workbook and worksheet
input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_472-15_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r3/eval_472-15_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read the value from cell A1
value = ws['A1'].value

# Map according to specified rules
lookup = {
    '4Ozark': 1,
    '3Tall': 2,
    '1Jasper': 3,
    '2GWood': 4,
    '5Dawson': 5,
    '8CPark': 6,
}

result = None
for key, val in lookup.items():
    if value and key in value:
        result = val
        break

# Set the value in cell B2
ws['B2'].value = result

# Save the workbook
wb.save(output_path)
