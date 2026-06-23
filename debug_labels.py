from openpyxl import load_workbook
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_5/regression_gate/before_fix/core_56274/input.xlsx')
ws = wb.active
# Find 'fiscal month' row
months_row = None
for r in ws.iter_rows(min_row=1, max_row=30, max_col=10):
    for c in r:
        if c.value and str(c.value).lower() == 'fiscal month':
            months_row = c.row
            break
    if months_row: break
print('months_row', months_row)
if months_row:
    # Print labels below
    for rr in range(months_row+2, months_row+10):
        print(rr, repr(ws.cell(row=rr, column=1).value))
