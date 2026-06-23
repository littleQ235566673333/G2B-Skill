import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_254-34_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_254-34_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)

# Delete Sheet3
if 'Sheet3' in wb.sheetnames:
    wb.remove(wb['Sheet3'])

ws = wb['Before']

# Clear all highlights and set values for the sum check
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.fill = PatternFill(fill_type=None)

ws['C2'].value = 'No valid combination'

wb.save(output_path)
