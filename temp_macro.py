import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_6/regression_gate/after_fix/core_263-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_6/regression_gate/after_fix/core_263-1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Identify headers
header_row = 1
headers = {cell.value: idx for idx, cell in enumerate(ws[header_row], start=1)}
mat_col = headers.get('Material') or headers.get('Material Type')
width_col = headers.get('Width')
height_col = headers.get('Height')

material_totals = defaultdict(float)

# Get all data rows
for row in ws.iter_rows(min_row=header_row+1):
    material = row[mat_col-1].value
    width = row[width_col-1].value
    height = row[height_col-1].value
    # Only consider rows with all values
    if material and width and height:
        try:
            area = float(width) * float(height)
            material_totals[material] += area
        except Exception:
            pass

# Sorted material types
materials_sorted = sorted(material_totals)

# Output to H2, H3, H4
start_row = 2
col = 8  # H
for i, mat in enumerate(materials_sorted[:3]):
    total_area = material_totals[mat]
    ws.cell(row=start_row + i, column=col).value = total_area

wb.save(output_path)
