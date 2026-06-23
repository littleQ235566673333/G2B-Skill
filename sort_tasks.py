import openpyxl

# Load the input workbook and worksheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_262-17_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_262-17_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Get header and identify target columns
def get_header_row(ws):
    return [cell.value for cell in ws[1]]

header = get_header_row(ws)

try:
    task_col_idx = header.index('Task')
    resp_col_idx = header.index('Responsibility')
except ValueError:
    raise Exception('Required columns not found in header!')

# Read data rows (assuming data is from row 2 to row 14, columns A-F)
data_rows = []
for row in ws.iter_rows(min_row=2, max_row=14, max_col=6):
    data_rows.append([cell.value for cell in row])

# Sort data: first by Task, then by Responsibility (both ascending)
sorted_rows = sorted(data_rows, key=lambda x: (x[task_col_idx], x[resp_col_idx]))

# Write sorted data back to worksheet
for i, row in enumerate(sorted_rows, start=2):
    for j, val in enumerate(row):
        ws.cell(row=i, column=j+1, value=val)

# Save to output file
wb.save(output_path)
