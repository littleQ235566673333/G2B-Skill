import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/after_pass/core_32337/input.xlsx')
ws = wb.active

# Map headers and identify relevant columns
header_row_idx = 2
header_row = [c.value for c in ws[header_row_idx]]
header_map = {name: idx+1 for idx, name in enumerate(header_row) if name}
# Columns
dob_col = header_map['DATE OF BIRTH']
report_date_cell = 'B1'
year_age_col = header_map['YEAR (age)']
category_col = header_map['CATEGORY']
age_actual_col = header_map['age (year)'] + 1  # next column for new 'Actual Age'
expected_result_col = header_map['Expected Result']

# Determine range
data_start, data_end = 3, 15

# Insert header for Actual Age
ws.cell(row=header_row_idx, column=age_actual_col).value = 'Actual Age'
preceding_header_cell = ws.cell(row=header_row_idx, column=age_actual_col-1)
# Copy only fill properties using PatternFill
if isinstance(preceding_header_cell.fill, PatternFill) and preceding_header_cell.fill.fgColor.type != 'indexed':
    copied_fill = PatternFill(
        fill_type=preceding_header_cell.fill.fill_type,
        fgColor=preceding_header_cell.fill.fgColor.rgb,
        bgColor=preceding_header_cell.fill.bgColor.rgb
    )
else:
    copied_fill = PatternFill()
ws.cell(row=header_row_idx, column=age_actual_col).fill = copied_fill
# Top align header
ws.cell(row=header_row_idx, column=age_actual_col).alignment = Alignment(horizontal='center', vertical='top')
ws.cell(row=header_row_idx, column=age_actual_col).font = Font(bold=True)

# Write Actual Age formula + formatting for each row
for row in range(data_start, data_end+1):
    dob_cell = f'{get_column_letter(dob_col)}{row}'
    actual_age_cell = ws.cell(row=row, column=age_actual_col)
    actual_age_cell.value = f'=DATEDIF({dob_cell},${report_date_cell},"y")'
    actual_age_cell.number_format = '0'
    actual_age_cell.alignment = Alignment(horizontal='center', vertical='center')
    actual_age_cell.font = Font(bold=True)
    actual_age_cell.fill = copied_fill

# Write formula for Expected Result in E3:E15
cat_col = get_column_letter(category_col)
year_col = get_column_letter(year_age_col)
for row in range(data_start, data_end+1):
    formula = f'=INDEX(${cat_col}${data_start}:${cat_col}${data_end},MATCH(O{row},${year_col}${data_start}:${year_col}${data_end},0))'
    ws.cell(row=row, column=expected_result_col).value = formula

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/after_pass/core_32337/output.xlsx')
