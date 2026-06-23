import openpyxl
import numpy as np

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_55468_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_55468_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read relevant headers and criteria
AC4 = ws['AC4'].value   # horizontal header criterion 1
AB4 = ws['AB4'].value   # horizontal header criterion 2
AE4 = ws['AE4'].value   # vertical header criterion 1
AD4 = ws['AD4'].value   # vertical header criterion 2

# Get header ranges for 2D table location
# Assumes:    C4:Z4 and C3:Z3 are dual X-axis headings
#             A5:A10 and B5:B10 are dual Y-axis headings
# Table Data: C5:Z10

# Get column headers (top two rows)
x1_headers = [ws.cell(row=4, column=col).value for col in range(3, 27)]  # C4:Z4 (24 columns)
x2_headers = [ws.cell(row=3, column=col).value for col in range(3, 27)]  # C3:Z3

# Get row headers (left two columns)
y1_headers = [ws.cell(row=row, column=1).value for row in range(5, 11)]  # A5:A10 (6 rows)
y2_headers = [ws.cell(row=row, column=2).value for row in range(5, 11)]  # B5:B10

# Table values
values = np.array([[ws.cell(row=row, column=col).value for col in range(3, 27)] for row in range(5, 11)])

# Find column indices matching both X criteria
col_matches = [i for i, (h1, h2) in enumerate(zip(x1_headers, x2_headers)) if h1 == AC4 and h2 == AB4]
# Find row indices matching both Y criteria
y_matches = [i for i, (h1, h2) in enumerate(zip(y1_headers, y2_headers)) if h1 == AE4 and h2 == AD4]

# Get the value, else None
result = None
if col_matches and y_matches:
    result = values[y_matches[0], col_matches[0]]

ws['AE5'].value = result
wb.save(output_path)
