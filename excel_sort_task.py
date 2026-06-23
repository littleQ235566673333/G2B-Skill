from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_262-17_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_262-17_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet1']

# 1. Find header indices for 'Task' and 'Responsibility'
header_row = 1
headers = {cell.value: col_idx for col_idx, cell in enumerate(ws[header_row], 1)}
task_col = headers.get('Task')
resp_col = headers.get('Responsibility')

if not (task_col and resp_col):
    raise Exception('Could not find required columns')

# 2. Extract data rows for sorting (A-F columns)
data_rows = []
for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, max_col=6, values_only=True):
    if any(row):
        data_rows.append(row)

# Remove trailing fully-empty rows
def strip_empty(rows):
    while rows and not any(rows[-1]):
        rows.pop()
    return rows

data_rows = strip_empty(data_rows)

# 3. Sort: first by Task (primary), then by Responsibility (secondary), both ascending
def sort_key(r):
    key0 = r[task_col-1]
    key1 = r[resp_col-1]
    # Handles None values by pushing them to the end
    return ('' if key0 is None else key0, '' if key1 is None else key1)
sorted_data = sorted(data_rows, key=sort_key)

# 4. Write output: fill Sheet1 A1:F14
# Copy headers A1-F1
for cidx in range(1, 7):
    ws.cell(row=1, column=cidx).value = ws.cell(row=1, column=cidx).value
for ridx, rd in enumerate(sorted_data[:13], start=2):
    for cidx in range(1, 7):
        ws.cell(row=ridx, column=cidx).value = rd[cidx-1] if cidx <= len(rd) else None
# Clear any below row 14, columns A-F
for ridx in range(len(sorted_data)+2, 15):
    for cidx in range(1, 7):
        ws.cell(row=ridx, column=cidx).value = None
wb.save(output_path)
