import openpyxl

# Load the input file
input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed0/eval_267-21_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed0/eval_267-21_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)

# Get relevant sheets
sh_merging = wb['merging']
sh_sh1 = wb['SH1']
sh_rp1 = wb['RP1']

# Helper function to create ID->QTY mapping
def sheet_id_qty(sheet):
    id_col, qty_col = None, None
    for col in range(1, sheet.max_column+1):
        header = sheet.cell(row=1, column=col).value
        if header == 'ID': id_col = col
        if header == 'QTY': qty_col = col
    mp = {}
    if id_col is not None and qty_col is not None:
        for row in range(2, sheet.max_row+1):
            _id = sheet.cell(row=row, column=id_col).value
            _qty = sheet.cell(row=row, column=qty_col).value
            mp[_id] = _qty
    return mp

sh1_mp = sheet_id_qty(sh_sh1)
rp1_mp = sheet_id_qty(sh_rp1)

# Process merging sheet (IDs in B2:B11, fill output in C and D)
for r in range(2, 12):
    id_val = sh_merging.cell(row=r, column=2).value
    val_c = sh1_mp.get(id_val, '-') if id_val is not None and id_val != '' else '-'
    val_d = rp1_mp.get(id_val, '-') if id_val is not None and id_val != '' else '-'
    sh_merging.cell(row=r, column=3).value = val_c if val_c not in [None, ''] else '-'
    sh_merging.cell(row=r, column=4).value = val_d if val_d not in [None, ''] else '-'

wb.save(output_path)
