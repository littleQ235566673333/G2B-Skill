from openpyxl import load_workbook
from datetime import datetime

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r3/eval_33157_tc1/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r3/eval_33157_tc1/output.xlsx'

wb = load_workbook(in_path)
ws = wb['Sheet1']

ACTIVITY_COLS = ['B', 'D', 'F', 'H']
ACTIVITY_HEADERS = [ws[f'{col}1'].value for col in ACTIVITY_COLS]

def safe_date(cell_val):
    if isinstance(cell_val, datetime):
        return cell_val
    try:
        return datetime.strptime(str(cell_val), '%Y-%m-%d')
    except:
        pass
    try:
        return datetime.strptime(str(cell_val), '%m/%d/%Y')
    except:
        pass
    return None

for row in range(2, 7):  # For K2:K6
    ref_date = ws[f'J{row}'].value
    ref_date_parsed = safe_date(ref_date)
    found = False
    if ref_date_parsed is not None:
        for i, col in enumerate(ACTIVITY_COLS):
            activity_date = ws[f'{col}{row}'].value
            activity_date_parsed = safe_date(activity_date)
            if activity_date_parsed == ref_date_parsed:
                ws[f'K{row}'] = ACTIVITY_HEADERS[i]
                found = True
                break
    if not found:
        ws[f'K{row}'] = ''

wb.save(out_path)
