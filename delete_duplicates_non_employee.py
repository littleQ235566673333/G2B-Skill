import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_1/group_91-34/r2/evolve_91-34/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_1/group_91-34/r2/evolve_91-34/output.xlsx'
sheet_name = 'SwiftMD'

# Read with correct header row
header_row = 1  # 2nd row in Excel, 0-based index

df = pd.read_excel(input_path, sheet_name=sheet_name, header=header_row, engine='openpyxl')
orig_cols = df.columns.tolist()

df['__row__'] = df.index
# Step 1: Filter only 'Yes' in Duplicate?
to_consider = df[df['Duplicate?'] == 'Yes']

# DOB column is called "Date Of Birth"
def check_group(g):
    non_emps = g[g['Relationship'].str.lower() != 'employee']
    if len(non_emps) > 1 and len(non_emps) == len(g):
        return non_emps.iloc[1:]['__row__'].tolist() # keep one
    return []

to_drop = []
groups = to_consider.groupby(['Last Name', 'First Name', 'Date Of Birth'])
for _, g in groups:
    marks = check_group(g)
    to_drop.extend(marks)

df_cleaned = df[~df['__row__'].isin(to_drop)].copy()
if '__row__' in df_cleaned.columns:
    df_cleaned = df_cleaned.drop(columns='__row__')

# Write results to B2:O42
wb = load_workbook(input_path)
ws = wb[sheet_name]
# Start at B2
for i in range(2, 43):
    for j, col in enumerate(orig_cols, start=2):
        value = df_cleaned.iloc[i-2, j-2] if (i-2) < len(df_cleaned) and (j-2) < len(df_cleaned.columns) else None
        ws.cell(row=i, column=j, value=value)
wb.save(output_path)
