from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_4/group_39515/r2/evolve_39515/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_4/group_39515/r2/evolve_39515/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
header_to_col = {header: idx+1 for idx, header in enumerate(header_row)}
col_month = header_to_col['Month Available']
col_year = header_to_col['Year Available']
col_jan = header_to_col['JAN']
col_dec = header_to_col['DEC']
col_form = header_to_col['Increase Based on Availability']

for row in range(2, 14):  # O2:O13 => rows 2 to 13
    month_cell = ws.cell(row=row, column=col_month).coordinate
    year_cell = ws.cell(row=row, column=col_year).coordinate
    jan_cell = ws.cell(row=row, column=col_jan).coordinate
    # data range for index: C2:N2, etc.
    data_range = f"{ws.cell(row=row, column=col_jan).coordinate}:{ws.cell(row=row, column=col_dec).coordinate}"
    # Build the formula string:
    formula = f"=IF({year_cell}=2022,{jan_cell},INDEX({data_range},MATCH({month_cell},$C$1:$N$1,0)))"
    ws.cell(row=row, column=col_form).value = formula

wb.save(output_path)
