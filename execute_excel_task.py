import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_333-29_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_333-29_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)

a_ws = wb['A']
b_ws = wb['B']

# Step 1: Find first 'Yes' or 'NA' in column P of sheet 'A'
first_condition_row = None
for row in range(2, a_ws.max_row + 1):  # Assuming header at row 1
    value = a_ws[f'P{row}'].value
    if value is not None and str(value).strip() in ['Yes', 'NA']:
        first_condition_row = row
        break

if first_condition_row:
    # Step 2: Find value 100 in column L above the found row
    date_row = None
    for row in range(first_condition_row-1, 1, -1):  # Go upwards
        l_value = a_ws[f'L{row}'].value
        if l_value == 100:
            date_row = row
            break
    if date_row:
        date_value = a_ws[f'F{date_row}'].value
    else:
        date_value = a_ws[f'F{first_condition_row}'].value  # fallback if 100 not found
else:
    # Step 5: Fallback to date in F3
    date_value = a_ws['F3'].value

# Step 4: Check F1 in sheet A against column C in sheet B
f1_text = a_ws['F1'].value
match_row_B = None
for row in range(2, b_ws.max_row + 1):  # B sheet header in row 1
    if b_ws[f'C{row}'].value == f1_text:
        match_row_B = row
        break

if match_row_B:
    b_ws[f'F{match_row_B}'] = date_value
else:
    # Insert result into F4
    b_ws['F4'] = date_value

# Ensure answer in F4 as per instructions
b_ws['F4'] = date_value

# Finish with 'A' as active sheet (not applicable in openpyxl, but workbook may be saved)
wb.save(output_path)
