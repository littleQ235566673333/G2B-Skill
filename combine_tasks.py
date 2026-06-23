import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_7/regression_gate/after_fix/core_60-7/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_7/regression_gate/after_fix/core_60-7/output.xlsx'

existing = pd.read_excel(input_path, sheet_name='Existing Task')
additions = pd.read_excel(input_path, sheet_name='Additions')
retired = pd.read_excel(input_path, sheet_name='Retired')

# Concatenate Existing + Additions
combined = pd.concat([existing, additions], ignore_index=True)
# Remove rows in Retired from combined
final = combined.merge(retired.drop_duplicates(), on=list(retired.columns), how='left', indicator=True)
net_activities = final[final['_merge'] == 'left_only'].drop('_merge', axis=1)

# Load the workbook and select Consolidated Tracker sheet
wb = load_workbook(input_path)
ws = wb['Consolidated Tracker']

# Write result to A3 according to instructions
start_row = 3
start_col = 1
max_rows = 9  # E11 - A3 + 1
max_cols = 5  # A-E
data_to_write = net_activities.iloc[:max_rows, :max_cols].values  # just in case
# Write header
for c in range(max_cols):
    ws.cell(row=start_row, column=start_col + c, value=net_activities.columns[c])
# Write data
for r in range(data_to_write.shape[0]):
    for c in range(data_to_write.shape[1]):
        ws.cell(row=start_row + 1 + r, column=start_col + c, value=data_to_write[r, c])

wb.save(output_path)
