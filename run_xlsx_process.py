from openpyxl import load_workbook

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_3/regression_gate/before_pass/core_493-18/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_3/regression_gate/before_pass/core_493-18/output.xlsx'
wb = load_workbook(INPUT)
ws = wb.active

# Gather F values to match
match_values = set()
for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
    v = row[0].value
    if v is not None and str(v).strip() != '':
        match_values.add(v)

# Identify and process rows to clear in A, B, C
maxrow = ws.max_row
for r in range(2, maxrow + 1):
    a_val = ws.cell(row=r, column=1).value
    if a_val not in match_values:
        # Clear (A,B,C)
        for c in range(1, 4):
            ws.cell(row=r, column=c).value = None

# Now, compress any gaps in A-C (shift up non-blanks, no row deletion)
read_rows = []
for r in range(2, maxrow + 1):
    if ws.cell(row=r, column=1).value is not None and str(ws.cell(row=r, column=1).value).strip() != '':
        row_data = [ws.cell(row=r, column=col).value for col in range(1, 4)]
        read_rows.append(row_data)

# Blank all A, B, C
for r in range(2, maxrow + 1):
    for c in range(1, 4):
        ws.cell(row=r, column=c).value = None

# Write compressed rows to top
for idx, row in enumerate(read_rows):
    for c, value in enumerate(row, 1):
        ws.cell(row=idx + 2, column=c).value = value

# Reapply autofilter for A-C
ws.auto_filter.ref = f'A1:C{maxrow}'

wb.save(OUTPUT)
