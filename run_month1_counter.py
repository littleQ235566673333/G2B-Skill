import openpyxl
from datetime import datetime

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_5/task_45707/r1/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_5/task_45707/r1/evolve_45707/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Detect header row and actual column indexes
header_row = 1
headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column+1)]

colA_idx = 1  # Date
colC_idx = 3  # Values to check for 1

# Read dates and values from columns

def get_date(cell):
    v = cell.value
    if isinstance(v, datetime):
        return v
    elif isinstance(v, str):
        try:
            return datetime.strptime(v, '%Y-%m-%d')
        except Exception:
            try:
                return datetime.strptime(v, '%d/%m/%Y')
            except Exception:
                return None
    return None

dates = [get_date(ws.cell(row=r, column=colA_idx)) for r in range(2, 70)]
colC = [ws.cell(row=r, column=colC_idx).value for r in range(2, 70)]
colC = [int(v) if v is not None and str(v).strip() != '' else None for v in colC]

# Process D2:D69
for i in range(68):
    this_row = i + 2
    nextdate = dates[i+1] if (i+1) < len(dates) else None
    val = ''
    if nextdate and nextdate.day == 1:
        target_month, target_year = nextdate.month, nextdate.year
        count_ones = sum((d.month == target_month and d.year == target_year and v == 1)
                         for d, v in zip(dates, colC) if d is not None and v is not None)
        val = count_ones
    ws.cell(row=this_row, column=4).value = val

wb.save(output_path)
