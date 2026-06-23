import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_45937_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_45937_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Gather the percentage ranges and their category/percent from N20:N22 and H20:I22
ranges = []
for row in range(20, 23):
    min_km = ws.cell(row=row, column=8).value
    max_km = ws.cell(row=row, column=9).value
    percent = ws.cell(row=row, column=14).value
    ranges.append((min_km, max_km, percent))

# Fill E7:E9 based on B7:B9
for i, row in enumerate(range(7, 10)):
    kms = ws.cell(row=row, column=2).value  # B7, B8, B9
    value_to_write = 'Unknown'
    for min_km, max_km, percent in ranges:
        if min_km is not None and max_km is not None:
            if min_km <= kms <= max_km:
                value_to_write = percent
                break
    ws.cell(row=row, column=5).value = value_to_write  # write to E7, E8, E9

wb.save(output_path)
print('Completed filling column E for rows 7-9 based on kms ranges.')
