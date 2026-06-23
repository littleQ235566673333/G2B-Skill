from openpyxl import load_workbook
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/regression_gate/before_pass/core_41589/output.xlsx'
sheet_name = 'Contact List'

wb = load_workbook(input_path)
ws = wb[sheet_name]

today = datetime.today()
# Start at row 4 and continue until blank in Column H
row = 4
while True:
    h_val = ws[f'H{row}'].value
    i_val = ws[f'I{row}'].value
    if h_val is None and (i_val is None or i_val == ''):
        break
    # Handle blank H (contact date) as not within last 30 days
    in_last_30 = False
    if isinstance(h_val, datetime):
        delta = today - h_val
        in_last_30 = 0 <= delta.days <= 30
    action = 'NO ACTION'
    if i_val is not None and str(i_val).strip().upper() == 'YES':
        if in_last_30:
            action = 'HOLD'
        else:
            action = 'TOUCH BASE'
    ws[f'J{row}'] = action
    row += 1
wb.save(output_path)
