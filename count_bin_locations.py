from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font

# Load the input file
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42/eval_39903_tc1/input.xlsx')
ws = wb.active

output_cells = ['C2', 'C3', 'C4', 'C5', 'C6']
results = []

def count_non_pallet_locations(cell_val):
    if not isinstance(cell_val, str):
        return 0
    # Each location is "<loc>: <number>", 1 per cell
    # Only count if doesn't start with 'X' or 'Z'
    loc = cell_val.split(':')[0].strip() if ':' in cell_val else cell_val.strip()
    if len(loc) == 0:
        return 0
    if loc[0].upper() not in ['X', 'Z']:
        return 1
    return 0

for cell in output_cells:
    val = ws[cell].value
    results.append(count_non_pallet_locations(val))

# Write results
for cell, res in zip(output_cells, results):
    ws[cell].value = res

# Formatting: Courier New, 9pt, all borders
font = Font(name='Courier New', size=9)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
for cell in output_cells:
    ws[cell].font = font
    ws[cell].border = border

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42/eval_39903_tc1/output.xlsx')
