import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import NumberFormatDescriptor
import re

INPUT = "results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_191-40_tc1/input.xlsx"
OUTPUT = "results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_191-40_tc1/output.xlsx"
SHEET = "Sheet1"
OUT_RANGE = (1, 1, 85, 8)  # (row1, col1, row2, col2) for A1:H85

# Utility: determine if a cell is blank (for block breaks)
def is_blank_row(row):
    return all(cell.value is None or(str(cell.value).strip() == '') for cell in row)

def convert_text_to_formula(cell):
    if cell.data_type == 's' and isinstance(cell.value, str):
        txt = cell.value.strip()
        if txt.startswith('='):
            cell.value = txt
            cell.data_type = 'f'  # set as formula

def get_formula_and_vals(row):
    # Store formulas (or None if not formula), and evaluated values for columns D, E, F
    out = []
    for cell in row[3:6]:
        if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
            out.append(('formula', cell.value))
        else:
            out.append(('value', cell.value))
    return out

def format_number(val):
    # Convert to float, round to 1 decimal, display as ##0.0 but whole numbers as integer
    try:
        num = float(val)
        if num == int(num):
            return int(num)
        else:
            return round(num, 1)
    except Exception:
        return val

def main():
    wb = openpyxl.load_workbook(INPUT)
    ws = wb[SHEET]
    # Store original column widths:
    col_widths = {i: ws.column_dimensions[get_column_letter(i)].width for i in range(1, ws.max_column+1)}

    # Step 1: Ensure calculation mode is automatic (openpyxl has no direct setting, this is for Excel UI only)
    # Step 2: Convert text formulas in D–F to real formulas
    for row in ws.iter_rows(min_row=OUT_RANGE[0], max_row=OUT_RANGE[2], min_col=4, max_col=6):
        for cell in row:
            convert_text_to_formula(cell)

    # Step 3: Find blocks (runs of non-blank rows separated by one blank row)
    blocks = []  # list of (start_row, end_row) (both inclusive, 1-based)
    in_block = False
    start = None
    for r in range(OUT_RANGE[0], OUT_RANGE[2]+1):
        row = list(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=OUT_RANGE[3]))[0]
        if not is_blank_row(row):
            if not in_block:
                start = r
                in_block = True
        else:
            if in_block:
                blocks.append((start, r-1))
                in_block = False
    if in_block:  # handle block ending at last row
        blocks.append((start, OUT_RANGE[2]))

    # Step 4: Process each block
    for (b_start, b_end) in blocks:
        # snapshot formulas (for columns D-F) and rows (A-H)
        orig_rows = []
        formula_infos = []
        d_values = []
        for r in range(b_start, b_end+1):
            row = list(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=OUT_RANGE[3]))[0]
            orig_rows.append([cell.value for cell in row])
            formula_infos.append([cell.value if (cell.data_type=='f' or (isinstance(cell.value,str) and cell.value.startswith('='))) else None for cell in row[3:6]])
            # To get sorting value, try to read cell's calculated value or fallback to value
            dval = row[3].value # col D (4th base-0)
            # Try to coerce to float
            try:
                fval = float(dval)
            except Exception:
                fval = float('-inf')
            d_values.append(fval)
        # Sort indices by D descending, stable sort for ties
        sorted_idx = sorted(range(len(d_values)), key=lambda i: (d_values[i], -i), reverse=True)
        # Rearrange the rows, re-apply formulas in D-F
        for dest_i, src_i in enumerate(sorted_idx):
            row_num = b_start + dest_i
            row = list(ws.iter_rows(min_row=row_num,max_row=row_num,min_col=1,max_col=OUT_RANGE[3]))[0]
            # Copy A-C and G-H as pure values
            for c in [0,1,2,6,7]:
                row[c].value = orig_rows[src_i][c]
            # D-F: restore formulas or values
            for c in range(3,6):
                val = formula_infos[src_i][c-3]
                if val is not None and isinstance(val,str) and val.startswith('='):
                    row[c].value = val
                    row[c].data_type = 'f'
                else:
                    row[c].value = orig_rows[src_i][c]
            # G-H
            for c in [6,7]:
                row[c].value = orig_rows[src_i][c]

    # Step 5: Formatting (bold, center D/E/F; number format; widths)
    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    numberfmt_decimal = '0.0'  # will edit per cell
    numberfmt_integer = '0'
    for r in range(OUT_RANGE[0], OUT_RANGE[2]+1):
        row = list(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=OUT_RANGE[3]))[0]
        for idx in range(3,6): # D/E/F
            cell = row[idx]
            # Bold/center
            cell.font = bold
            cell.alignment = center
            # Number formatting
            # If present value is whole number, apply integer format; else, decimal
            v = cell.value
            # Only apply format if formula or value is number
            # Can't evaluate formula here; will set both formats (Excel will render)
            if v is not None and (isinstance(v, (int, float)) or (isinstance(v, str) and v.startswith('='))):
                # If formula OR value is float/int
                # See if the *value* should be integer or .0 shown
                try:
                    num = float(v) if not (isinstance(v,str) and v.startswith('=')) else None
                except:
                    num = None
                if num is not None and num == int(num):
                    cell.number_format = numberfmt_integer
                else:
                    cell.number_format = numberfmt_decimal

    # Step 6: Restore column widths
    for i, w in col_widths.items():
        if w is not None:
            ws.column_dimensions[get_column_letter(i)].width = w

    # Step 7: Keep one blank row between blocks, compress any extras
    block_rows = set()
    for b in blocks:
        block_rows.update(range(b[0], b[1]+1))
    # Remove any extra blank rows between data
    next_write = OUT_RANGE[0]
    for (b_start, b_end) in blocks:
        if next_write < b_start:
            # Move block up, preserving one blank in between
            move_by = b_start - next_write
            for r in range(b_end, b_start-1, -1):
                for c in range(1, OUT_RANGE[3]+1):
                    ws.cell(row=r-move_by+1, column=c).value = ws.cell(row=r, column=c).value
                    ws.cell(row=r, column=c).value = None
            next_write = b_start - move_by + (b_end-b_start+1) + 1
        else:
            next_write = b_end + 2  # one blank row after block
    # Cut off any trailing extra rows
    for r in range(next_write, OUT_RANGE[2]+1):
        for c in range(1, OUT_RANGE[3]+1):
            ws.cell(row=r, column=c).value = None
    wb.save(OUTPUT)

main()
