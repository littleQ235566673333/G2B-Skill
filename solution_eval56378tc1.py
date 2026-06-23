import openpyxl
from openpyxl.styles import Alignment
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/eval_seed42_rerun2/eval_56378_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/eval_seed42_rerun2/eval_56378_tc1/output.xlsx'

# These blocks are offset: frames start at col 2 and col 11, header at row 4
header_row = 4
frame1_start_col = 2
frame2_start_col = 11
cols_in_frame = 7
rows_in_frame = 4

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.worksheets[0]

# Read all values in the range needed as reference
# Get the header for frame 1 (left block)
header = [ws.cell(row=header_row, column=frame1_start_col + i).value for i in range(cols_in_frame)]

# Build a list of frame 1 product rows below the header
source_rows = []
for r in range(header_row + 1, header_row + 1 + 8):  # Up to 8, enough for sample
    values = [ws.cell(row=r, column=frame1_start_col + i).value for i in range(cols_in_frame)]
    # Only add rows with non-empty first cell (product code)
    if any(v is not None for v in values):
        source_rows.append(values)

# Find the index of 'QUANTITY UNITS' in the header
try:
    qty_idx = [h.lower() for h in header].index('quantity units')
except Exception as e:
    raise RuntimeError('Could not find QUANTITY UNITS in header: '+repr(header))

# Filter products where QUANTITY UNITS is not None/blank
filtered = [row for row in source_rows if row[qty_idx] is not None and str(row[qty_idx]).strip() != '']

# Write filtered results to frame 2 (starting at L5)
for i, row in enumerate(filtered[:rows_in_frame]):
    for j in range(cols_in_frame):
        cell = ws.cell(row=header_row + 1 + i, column=frame2_start_col + j)
        cell.value = row[j]
        # Left align first col (product name), right align the rest
        if j == 1:
            cell.alignment = Alignment(horizontal='left')
        elif j > 1:
            cell.alignment = Alignment(horizontal='right')

wb.save(output_path)
