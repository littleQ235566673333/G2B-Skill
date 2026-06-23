from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/regression_gate/before_fix/core_40478/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/regression_gate/before_fix/core_40478/output.xlsx'

def extract(text):
    if text is None:
        return None
    # Split by space
    parts = text.split(' ')
    if len(parts) < 3:
        return ''
    # Join the parts after the second space
    after_second_space = ' '.join(parts[2:])
    # Find the first dash
    dash_idx = after_second_space.find('-')
    if dash_idx == -1:
        return after_second_space.strip()
    return after_second_space[:dash_idx].strip()

wb = load_workbook(input_path)
ws = wb.active

for row in range(1, 4):
    val = ws[f'A{row}'].value
    ws[f'B{row}'] = extract(val)

wb.save(output_path)
