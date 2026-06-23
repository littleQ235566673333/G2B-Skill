import openpyxl

def eval_priority(rng_a, rng_b):
    pairs = list(zip(rng_a, rng_b))
    for veg, result in [('Potato', 'Worst'), ('Tomato', 'Ignore'), ('Pickle', 'Bad')]:
        if any((a == veg and b is False) for a, b in pairs):
            return result
    return 'Good'

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_42198_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_42198_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 8):  # Rows 2 to 7 inclusive
    rng_a = [ws[f'A{r}'].value for r in range(2, row + 1)]
    rng_b = [ws[f'B{r}'].value for r in range(2, row + 1)]
    ws[f'C{row}'] = eval_priority(rng_a, rng_b)

wb.save(output_path)
