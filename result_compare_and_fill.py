from openpyxl import load_workbook

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_395-36_tc1/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_395-36_tc1/output.xlsx'

wb = load_workbook(INPUT)
ws_main = wb['Main unique ID']
ws_get = wb['Result what i am getting']

unique_ids = [row[0] for row in ws_main.iter_rows(min_row=1, max_col=1, values_only=True) if row[0] is not None]
get_rows = list(ws_get.iter_rows(min_row=2, values_only=True))
header = tuple(ws_get.iter_rows(min_row=1, max_row=1, values_only=True))[0]
get_id_map = { row[0]: row for row in get_rows }

if 'MyResult' in wb.sheetnames:
    ws_res = wb['MyResult']
    wb.remove(ws_res)
wres = wb.create_sheet('MyResult')
wres.append(header)
for uid in unique_ids:
    row = list(get_id_map.get(uid, (uid,) + (None,)*(len(header)-1)))
    wres.append(row)
for _ in range(len(unique_ids)+1, 21):
    wres.append([None]*15)
wb.save(OUTPUT)
