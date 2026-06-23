import openpyxl

# Load input workbook and sheet
target_input = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_493-5_tc1/input.xlsx"
target_output = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_493-5_tc1/output.xlsx"
wb = openpyxl.load_workbook(target_input)
ws = wb["Imported Data"]

# Read the first 10 rows (A1:F10)
data = list(ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=6, values_only=False))
header = [cell.value for cell in data[0]]
data_rows = data[1:]

# Store each row as a dictionary for easier comparison
rows = []
for idx, row in enumerate(data_rows):
    rows.append({
        'sheet_idx': idx+2,  # Sheet index (starts at 2; header is 1)
        'A': row[0].value,
        'C': row[2].value,
        'D': row[3].value,
        'F': row[5].value,
        'values': [cell.value for cell in row],
    })

# Find all row pairs to delete
rows_to_delete = set()
for i in range(len(rows)):
    if rows[i]['sheet_idx'] in rows_to_delete:
        continue
    for j in range(len(rows)):
        if i == j or rows[j]['sheet_idx'] in rows_to_delete:
            continue
        # Only delete if all 3 criteria are met
        if (
            rows[i]['C'] == rows[j]['D'] and
            rows[i]['A'] == rows[j]['A'] and
            rows[i]['F'] == rows[j]['F']
        ):
            rows_to_delete.add(rows[i]['sheet_idx'])
            rows_to_delete.add(rows[j]['sheet_idx'])
            break

# Build filtered rows; always include the header
filtered_rows = [header]
for row in rows:
    if row['sheet_idx'] not in rows_to_delete:
        filtered_rows.append(row['values'])
# Pad with empty rows if needed or trim to 9 data rows after header
while len(filtered_rows) < 10:
    filtered_rows.append(['']*6)
filtered_rows = filtered_rows[:10]

# Write to new output workbook
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Imported Data'
for r_idx, row in enumerate(filtered_rows, 1):
    for c_idx, val in enumerate(row, 1):
        ws_out.cell(row=r_idx, column=c_idx, value=val)
wb_out.save(target_output)
