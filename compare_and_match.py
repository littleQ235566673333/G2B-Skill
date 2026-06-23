import openpyxl
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_1/group_48969/r3/evolve_48969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_1/group_48969/r3/evolve_48969/output.xlsx'

def excel_date(dt):
    if isinstance(dt, datetime):
        return dt.date()
    return dt

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Source 1 headings are row 2: dates B2-G2; cities B3-B5, values C3-G5
# Source 2 headings are row 7: dates C7-G7; cities B8-B10, values C8-G10

# Read dates
s1_dates = [excel_date(ws.cell(row=2, column=col).value) for col in range(3, 8)]
s2_dates = [excel_date(ws.cell(row=7, column=col).value) for col in range(3, 8)]

# Map date to column
s1_date_col = {d: col for d, col in zip(s1_dates, range(3, 8))}
s2_date_col = {d: col for d, col in zip(s2_dates, range(3, 8))}

# Cities
s1_cities = [ws.cell(row=row, column=2).value for row in range(3, 6)]
s2_cities = [ws.cell(row=row, column=2).value for row in range(8, 11)]

# Map city to row
s2_city_row = {city: row for city, row in zip(s2_cities, range(8, 11))}

# Find common dates
common_dates = [d for d in s1_dates if d in s2_dates]

# Output: start at J3 (col=10, row=3)
for i, city in enumerate(s1_cities):
    s1_row = 3 + i
    s2_row = s2_city_row.get(city)
    for j, d in enumerate(common_dates):
        col = 10 + j
        if s2_row is None or d not in s2_date_col:
            ws.cell(row=s1_row, column=col).value = None
        else:
            s1_val = ws.cell(row=s1_row, column=s1_date_col[d]).value
            s2_val = ws.cell(row=s2_row, column=s2_date_col[d]).value
            ws.cell(row=s1_row, column=col).value = (s1_val == s2_val)
wb.save(output_path)
print('Done')
