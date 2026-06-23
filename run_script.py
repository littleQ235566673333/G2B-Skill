from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import copy

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_1/regression_gate/before_pass/core_32337/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_1/regression_gate/before_pass/core_32337/output.xlsx'
wb = load_workbook(INPUT)
ws = wb.active

header_row = 2
headers = {cell.value: cell.column for cell in ws[header_row]}
def getcol(name):
    # Allow for partial/case-insensitive matches
    v = headers.get(name)
    if v is None:
        for k in headers:
            if k and name.lower() in str(k).lower():
                return headers[k]
    return v

age_col = getcol('age (year)')
if not age_col:
    raise Exception('age (year) column not found')
category_col = getcol('CATEGORY') or 3

start, end = 3, 15
# Write XLOOKUP in E3:E15
for row in range(start, end+1):
    ws.cell(row=row, column=5).value = f'=XLOOKUP(O{row},$O$3:$O$15,$C$3:$C$15)'

# Insert Actual Age column after age (year)
ws.insert_cols(age_col+1)
ws.cell(row=header_row, column=age_col+1, value='Actual Age')
# Copy fill from prev column using deep copy
fill = ws.cell(row=header_row, column=age_col).fill
for row in range(start, end+1):
    dob_col = getcol('birth')
    rdt_col = getcol('report')
    col_dob = get_column_letter(dob_col) if dob_col else 'L'
    col_rdt = get_column_letter(rdt_col) if rdt_col else 'D'
    formula = f'=DATEDIF(${col_dob}{row},${col_rdt}{row},"Y")'
    c = ws.cell(row=row, column=age_col+1)
    c.value = formula
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')
    c.number_format = '0'
    # Use deep copy for PatternFill
    if isinstance(fill, PatternFill) and fill.patternType is not None:
        c.fill = copy.copy(fill)
head_cell = ws.cell(row=header_row, column=age_col+1)
head_cell.font = Font(bold=True)
head_cell.alignment = Alignment(horizontal='center', vertical='top')
if isinstance(fill, PatternFill) and fill.patternType is not None:
    head_cell.fill = copy.copy(fill)

wb.save(OUTPUT)
