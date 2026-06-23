from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/group_11276/r2/evolve_11276/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/group_11276/r2/evolve_11276/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# F=6, AJ=36, so cols 6 to 36 inclusive
for col in range(6, 37):
    col_letter = get_column_letter(col)
    cell = f'{col_letter}3'
    ref_cell = f'{col_letter}4'
    # Set to weekday name, e.g. "Mon", "Tue" etc.
    ws[cell] = f'=TEXT({ref_cell},"DDD")'

wb.save(output_path)
