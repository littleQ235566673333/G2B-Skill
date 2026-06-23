import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_32093_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_32093_tc1/output.xlsx'

# Load workbook and select active worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 16):  # Rows 2 to 15 (inclusive)
    current_employee = ws[f'B{row}'].value
    new_employees = [ws[f'C{row}'].value, ws[f'D{row}'].value, ws[f'E{row}'].value]
    # Find first non-empty new employee name
    new_name = None
    for name in new_employees:
        if name is not None and str(name).strip() != '':
            new_name = name
            break
    # If no new employee name, use current employee
    if new_name is None:
        ws[f'F{row}'] = current_employee
    else:
        ws[f'F{row}'] = new_name

wb.save(output_path)
