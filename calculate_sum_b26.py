import openpyxl
from datetime import datetime
from calendar import monthrange
import re

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_57590_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun1/eval_57590_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Helper: extract year from first non-null date in Column C
first_year = None
for r in range(2, ws.max_row + 1):
    cval = ws.cell(row=r, column=3).value
    if cval:
        if hasattr(cval, 'year') and hasattr(cval, 'month'):
            first_year = cval.year
            break
        else:
            try:
                dt = datetime.strptime(str(cval), '%Y-%m-%d')
                first_year = dt.year
                break
            except:
                continue

base = ws['A26'].value
month = None
year = None

if hasattr(base, 'year') and hasattr(base, 'month'):
    year = base.year
    month = base.month
elif isinstance(base, str):
    # Try to match month name and year, e.g. 'January 2023'
    match = re.match(r'(\w+)\s*(\d{4})?', base)
    if match:
        month_str = match.group(1)
        month = datetime.strptime(month_str, '%B').month
        year = match.group(2)
        if not year and first_year:
            year = first_year
        elif not year:
            raise ValueError('No year found for month in A26 and no year in Column C!')
        year = int(year)
    else:
        raise ValueError('Could not parse month/year from A26')
else:
    raise ValueError('Unrecognized A26 format')

first_day = datetime(year, month, 1)
last_day = first_day.replace(day=monthrange(year, month)[1])

total = 0
for r in range(2, ws.max_row + 1):
    cval = ws.cell(row=r, column=3).value
    if not cval:
        continue
    dt = None
    if hasattr(cval, 'year') and hasattr(cval, 'month'):
        dt = cval
    else:
        try:
            dt = datetime.strptime(str(cval), '%Y-%m-%d')
        except:
            continue
    if first_day <= dt <= last_day:
        ival = ws.cell(row=r, column=9).value
        try:
            val = float(ival)
            total += val
        except (TypeError, ValueError):
            continue
ws['B26'] = total
wb.save(output_path)
