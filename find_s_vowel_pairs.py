import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/regression_gate/after_fix/core_146-49/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/regression_gate/after_fix/core_146-49/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Collect all words from column A (assuming first column; skip empty)
words = []
for row in ws.iter_rows(min_row=1, max_col=1):
    val = row[0].value
    if val and isinstance(val, str):
        words.append(val.upper())

# Unique while preserving order
seen = set()
words = [x for x in words if not (x in seen or seen.add(x))]

vowels = 'AEIOU'
results = set()

for w1 in words:
    s_indices = [i for i, c in enumerate(w1) if c == 'S']
    if not s_indices:
        continue
    for vowel in vowels:
        w2_chars = list(w1)
        for i in s_indices:
            w2_chars[i] = vowel
        w2 = ''.join(w2_chars)
        if w2 != w1 and w2 in words:
            # Sort to deduplicate ('S...','A...') and ('A...','S...')
            pair = tuple(sorted([w1, w2]))
            results.add(pair)

# Write pairs to G2:H... (1-based index)
results = sorted(list(results), key=lambda x: (x[0].lower(), x[1].lower()))
for idx, (wA, wB) in enumerate(results, 2):
    ws.cell(row=idx, column=7).value = wA
    ws.cell(row=idx, column=8).value = wB

# Remove anything after row len(results)+2 up to row 65 (as requested)
for r in range(len(results)+2, 66):
    ws.cell(row=r, column=7).value = None
    ws.cell(row=r, column=8).value = None

# After writing, sort the result range G2:H65 case-insensitively
pairs = []
for row in ws.iter_rows(min_row=2, max_row=65, min_col=7, max_col=8):
    v1 = row[0].value
    v2 = row[1].value
    if v1 and v2:
        pairs.append((v1, v2))
# Sort the pairs case-insensitively
pairs.sort(key=lambda x: (x[0].lower(), x[1].lower()))
# Write back, starting at G1
for i, (v1, v2) in enumerate(pairs, 1):
    ws.cell(row=i, column=7).value = v1
    ws.cell(row=i, column=8).value = v2
# Clear the rest
for r in range(len(pairs)+1, 66):
    ws.cell(row=r, column=7).value = None
    ws.cell(row=r, column=8).value = None

wb.save(output_path)
