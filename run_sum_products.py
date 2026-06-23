import openpyxl

# Load workbook
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_2/regression_gate/after_pass/core_50796/input.xlsx')
ws = wb.active

# Read criteria (product types) from B2:B5
criteria = []
for row in range(2, 6):
    val = ws[f'B{row}'].value
    if val is not None:
        criteria.append((row, val))

# Scan B12:C22 for counting
results = {}
for target_row, crit in criteria:
    s = 0
    for row in range(12, 23):
        if ws[f'B{row}'].value == crit:
            c_val = ws[f'C{row}'].value
            try:
                s += float(c_val)
            except (TypeError, ValueError):
                pass
    results[target_row] = s

# Write results back to B2:B5
for r in results:
    ws[f'B{r}'] = results[r]

# Save to output file
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_2/regression_gate/after_pass/core_50796/output.xlsx')
