import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_1/task_263-1/r0/evolve_263-1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_1/task_263-1/r0/evolve_263-1/output.xlsx'

# Correct columns based on actual headers:
df = pd.read_excel(input_path, sheet_name=0)
matcol = 'Mtrl'
widcol = 'Width'
heicol = 'Height'

# Calculate area for each row
df['Area'] = df[widcol] * df[heicol]
area_total = df.groupby(matcol)['Area'].sum().reset_index()
area_total['display'] = area_total.apply(lambda row: f'Total area for {row[matcol]}: {row["Area"]}', axis=1)

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]
for i in range(len(area_total)):
    ws[f'H{i+2}'] = area_total.loc[i,'display']
wb.save(output_path)
