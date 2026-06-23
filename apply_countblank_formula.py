from openpyxl import load_workbook

# Paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/after_pass/core_41969/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/after_pass/core_41969/output.xlsx"

wb = load_workbook(input_path)
ws = wb.active

# Write formulas A6, B6, C6
for i in range(3):
    start_col = 1 + i*3  # 1=A, 4=D, 7=G
    end_col = start_col + 2
    start_letter = chr(64 + start_col)
    end_letter = chr(64 + end_col)
    formula = f'=COUNTBLANK({start_letter}3:{end_letter}3)'
    ws.cell(row=6, column=1 + i, value=formula)

wb.save(output_path)
