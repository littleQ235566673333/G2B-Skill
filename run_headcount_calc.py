import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/regression_gate/after_fix/core_59129/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/regression_gate/after_fix/core_59129/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read hire and termination dates
hire_dates = [ws[f'A{row}'].value for row in range(2, 23)]
term_dates = [ws[f'B{row}'].value for row in range(2, 23)]

# Read months from E1:P1 (assuming these are date/datetime columns or string month names)
months = [ws.cell(row=1, column=col).value for col in range(5, 17)]

# Handle string -> datetime conversion for months if necessary
for i, m in enumerate(months):
    if isinstance(m, str):
        try:
            months[i] = datetime.strptime(m, '%b %Y')
        except ValueError:
            try:
                months[i] = datetime.strptime(m, '%B %Y')
            except ValueError:
                pass

def is_active(hire, term, month_end):
    if not hire:
        return False
    if hire > month_end:
        return False
    if not term:
        return True
    return term > month_end

results = []
for month in months:
    # Assume the end of the month for testing headcount (i.e., last day of month)
    month_end = month
    if isinstance(month, datetime):
        # get last possible day of month
        next_month = month.replace(day=28) + timedelta(days=4)
        month_end = next_month - timedelta(days=next_month.day)
    count = 0
    for hire, term in zip(hire_dates, term_dates):
        # Convert Excel dates to python datetime if they are not already
        if type(hire).__name__ == 'float' or type(hire).__name__ == 'int':
            hire = openpyxl.utils.datetime.from_excel(hire)
        if type(term).__name__ == 'float' or type(term).__name__ == 'int':
            term = openpyxl.utils.datetime.from_excel(term)
        if is_active(hire, term, month_end):
            count += 1
    results.append(count)

# Write to E2:P2
for idx, val in enumerate(results):
    ws.cell(row=2, column=5 + idx).value = val

wb.save(output_path)
