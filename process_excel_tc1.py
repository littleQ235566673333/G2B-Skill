import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_304-35_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_304-35_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)

# The available sheet is named 'start'
ws = wb['start']

start_row = 3
end_row = 36

# Step 1: Delete values from column F where column E is empty
for row in range(start_row, end_row+1):
    col_e = ws[f'E{row}'].value
    if col_e is None or (isinstance(col_e, str) and col_e.strip() == ''):
        ws[f'F{row}'].value = None

# Step 2: Copy values in column A downward, stopping at empty cell in column F,
# and never copying identical index cells (e.g., A2->A3 if A3 is identical)
prev_val = None
for row in range(start_row, end_row+1):
    val_a = ws[f'A{row}'].value
    val_f = ws[f'F{row}'].value
    # Only copy if cell A is empty, cell F is not empty, and prev_val exists and is not identical index
    if (val_a is None or (isinstance(val_a, str) and val_a.strip() == '')) and val_f is not None:
        # Check if prev_val is not an index (shouldn't copy when index repeats)
        if prev_val is not None and ws[f'A{row-1}'].value != ws[f'A{row}'].value:
            ws[f'A{row}'].value = prev_val
    # Do not continue copying if F is empty
    if val_f is None:
        prev_val = None
    else:
        prev_val = ws[f'A{row}'].value

wb.save(output_path)
