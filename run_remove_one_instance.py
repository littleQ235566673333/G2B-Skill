import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_5/regression_gate/after_fix/core_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_5/regression_gate/after_fix/core_387-16/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Columns:
col_value = 1  # A
col_binaries = 2  # B
col_result_values = 4  # D
val_start_row = 3  # Data starts at row 3 for 'Value'/'Binaries'
# Find last non-empty row in column A or B; conservative up to 20 rows
value_list = []
binaries_list = []
for row in range(val_start_row, 20):
    v = ws.cell(row=row, column=col_value).value
    b = ws.cell(row=row, column=col_binaries).value
    if v is None:
        break
    value_list.append(v)
    binaries_list.append(b)
# Result values start at D5 (row 5)
result_values = []
for row in range(5, 30):  # 5 ..
    v = ws.cell(row=row, column=col_result_values).value
    if v is None:
        break
    result_values.append(v)
# Mark rows to remove only one instance per match
to_remove = [False]*len(value_list)
for target in result_values:
    for idx, val in enumerate(value_list):
        if val == target and not to_remove[idx]:
            to_remove[idx] = True
            break
new_values = [v for v, rem in zip(value_list, to_remove) if not rem]
new_binaries = [v for v, rem in zip(binaries_list, to_remove) if not rem]
# Clear results in output area A2:D18
for r in range(2, 19):
    for c in range(1, 5):
        ws.cell(row=r, column=c).value = None
# Write new values into output (A2, B2, C2 ... down)
for idx, (v, b) in enumerate(zip(new_values, new_binaries)):
    ws.cell(row=2+idx, column=1).value = v
    ws.cell(row=2+idx, column=2).value = b
# Solver result: sum of col A2..A18 (just computed values)
solver_sum = sum([x for x in new_values if isinstance(x, (int, float))])
ws.cell(row=2, column=4).value = solver_sum
# Target value (cell E2, col 5 row 2)
target_value = ws.cell(row=2, column=5).value
if isinstance(target_value, (int, float)):
    ws.cell(row=3, column=4).value = solver_sum - target_value
wb.save(output_path)
