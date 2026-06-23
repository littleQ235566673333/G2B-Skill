import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_6/group_42930/r3/evolve_42930/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_6/group_42930/r3/evolve_42930/output.xlsx'

# Load data with pandas for easier handling
# Use the first sheet

df = pd.read_excel(input_path)

# Guess column names for subdivision and build year
sub_col = [c for c in df.columns if 'subdiv' in c.lower()][0]
built_col = [c for c in df.columns if 'year' in c.lower() or 'built' in c.lower()][0]

# Assign unique number to each subdivision by earliest house built (lowest build year)
earliest = df.groupby(sub_col)[built_col].min().reset_index()
earliest = earliest.sort_values(built_col).reset_index(drop=True)
earliest['subdiv_num'] = earliest.index + 1

# Map this number back to each house
merged = df.merge(earliest[[sub_col, 'subdiv_num']], on=sub_col, how='left')

# Sort all houses so that subdivisions are clustered by their earliest built (subdiv_num)
sorted_houses = merged.sort_values(['subdiv_num', built_col]).reset_index(drop=True)

# Write subdiv_num for first 21 houses to C2:C22
wb = load_workbook(input_path)
ws = wb.active
for i in range(21):
    ws[f'C{i+2}'] = sorted_houses['subdiv_num'].iloc[i] if i < len(sorted_houses) else None

wb.save(output_path)
