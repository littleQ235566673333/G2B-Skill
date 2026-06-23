import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_2/regression_gate/before_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_2/regression_gate/before_pass/core_9726/output.xlsx'

# Read all sheets
all_sheets = pd.read_excel(input_path, sheet_name=None)

# Assuming Table 1 and Table 2 are on the first sheet
sheet_name = list(all_sheets.keys())[0]
df = all_sheets[sheet_name]

# Find likely Table 1 and Table 2 by columns
cols = df.columns
# Table 1 columns should have Student and Visit date
# Table 2 columns should have Student and Submission date and a Visit date column to fill (likely empty)

# Typical Table 1: Student, ..., Visit date
# Typical Table 2: Student, ..., Submission date, Visit date [to fill]

# Infer positions
tab1_student_col = [c for c in cols if "student" in c.lower()][0]
tab1_visit_col = [c for c in cols if "visit" in c.lower() and not "submission" in c.lower()][0]

tab2_student_col = tab1_student_col
tab2_submission_col = [c for c in cols if "submission" in c.lower()][0]
tab2_visit_col = [c for c in cols if "visit" in c.lower() and not "submission" in c.lower()][-1]

# Split the tables: scan for a blank row (all NaN)
tab1_rows = []
tab2_start = None
for idx, row in df.iterrows():
    if row.isnull().all():
        tab2_start = idx + 1
        break
    tab1_rows.append(idx)

if tab2_start is None:
    raise Exception('Could not find Table 2 start')

tab1 = df.loc[tab1_rows].dropna(how='all')
tab2 = df.loc[tab2_start:].reset_index(drop=True)

# Clean Table 1/2 (strip extra header na row)
tab2 = tab2.dropna(how='all').reset_index(drop=True)

# Parse visit and submission dates
tab1[tab1_visit_col] = pd.to_datetime(tab1[tab1_visit_col], errors='coerce', dayfirst=True)
tab2[tab2_submission_col] = pd.to_datetime(tab2[tab2_submission_col], errors='coerce', dayfirst=True)

# Prepare result column for Table 2
result_dates = []
for _, row in tab2.iterrows():
    student = row[tab2_student_col]
    submission_date = row[tab2_submission_col]
    # All visit dates for this student, strictly less than submission date
    visits = tab1[(tab1[tab1_student_col] == student) & (tab1[tab1_visit_col] < submission_date)]
    if visits.empty:
        result = None
    else:
        result = visits[tab1_visit_col].max()
    result_dates.append(result)

# Write back to the original sheet in the correct cell
wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]
# Write results to J2 onwards (as specified)
for i, dt in enumerate(result_dates):
    cell = f'J{i+2}'
    ws[cell] = dt.strftime('%d-%b-%y') if pd.notnull(dt) else ''

wb.save(output_path)
