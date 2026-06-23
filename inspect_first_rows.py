import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/regression_gate/before_fix/core_387-16/input.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet1']

for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=10):
    print([cell.value for cell in row])
