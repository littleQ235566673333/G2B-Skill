from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/after_pass/core_51249/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/after_pass/core_51249/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# The RGB color E2EFDA is specified as 'E2EFDA' (no #)
fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

targets = [1, 5, 9]

for r in targets:
    b1 = ws.cell(row=r, column=2).value
    b2 = ws.cell(row=r+1, column=2).value
    out = None
    if b1 == 'Description A' and (b2 is None or b2 == ''):
        out = 'Single A'
    elif b1 == 'Description B' and (b2 is None or b2 == ''):
        out = 'Single B'
    elif b1 == 'Description A' and (b2 == 'Description B'):
        out = 'Multiple'
    if out is not None:
        dcell = ws.cell(row=r, column=4)
        dcell.value = out
        dcell.fill = fill

wb.save(output_path)
