import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_6/group_49667/r1/evolve_49667/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_6/group_49667/r1/evolve_49667/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Assume meeting blocks are in columns F (6) through AT (46)
block_cols = list(range(6, 47))
output_area = (2, 16, 2, 5)  # B2:E16

def col_to_time(col_idx):
    base_hour = 8  # F=col 6 is 8.0, G=6.5, ...
    return base_hour + (col_idx - 6) * 0.5

def col_to_date(col_idx):
    # Get the header value for the meeting column (row 1)
    return ws.cell(row=1, column=col_idx).value

for row in range(2, 17):  # rows 2-16 (inclusive)
    meeting_blocks = []
    start_col = None
    end_col = None
    for col in block_cols:
        val = ws.cell(row=row, column=col).value
        if val == 'm':
            if start_col is None:
                start_col = col
            end_col = col  # always update end_col for consecutive 'm'
        else:
            if start_col is not None:
                meeting_blocks.append((start_col, end_col))
                start_col = None
                end_col = None
    if start_col is not None:   # handle block segment running to the end
        meeting_blocks.append((start_col, end_col))

    # Only up to two blocks per requirements
    for i in range(2):
        if i < len(meeting_blocks):
            s_col, e_col = meeting_blocks[i]
            date1 = col_to_date(s_col)
            time1 = col_to_time(s_col)
            date2 = col_to_date(e_col)
            time2 = col_to_time(e_col + 1)  # End time is half-hour after last 'm' col
            # Place into B,C (first) or D,E (second) blocks
            ws.cell(row=row, column=2 + i*2, value=date1)
            ws.cell(row=row, column=3 + i*2, value=time1)
            ws.cell(row=row, column=4 + i*2, value=date2)
            ws.cell(row=row, column=5 + i*2, value=time2)
        else:
            # Clear if not present
            ws.cell(row=row, column=2 + i*2, value=None)
            ws.cell(row=row, column=3 + i*2, value=None)
            ws.cell(row=row, column=4 + i*2, value=None)
            ws.cell(row=row, column=5 + i*2, value=None)

wb.save(output_path)
