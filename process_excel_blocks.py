import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import re

infile = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42/eval_191-40_tc1/input.xlsx"
outfile = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42/eval_191-40_tc1/output.xlsx"
wb = openpyxl.load_workbook(infile, data_only=False)
wb_data = openpyxl.load_workbook(infile, data_only=True)
ws = wb["Sheet1"]
ws_data = wb_data["Sheet1"]

start_row = 1
end_row = 85
start_col = 1
end_col = 8 # A-H

blocks = []
block_start = None

# Identify block starts/ends
for r in range(start_row, end_row + 1):
    is_blank = all((ws.cell(row=r, column=c).value is None or str(ws.cell(row=r, column=c).value).strip() == "") for c in range(start_col, end_col+1))
    if not is_blank and block_start is None:
        block_start = r
    elif is_blank and block_start is not None:
        blocks.append((block_start, r-1))
        block_start = None
if block_start is not None:
    blocks.append((block_start, end_row))

def get_num_fmt(val):
    fmt = "#,##0.0;[Red]-#,##0.0"
    fmt_int = "#,##0;[Red]-#,##0"
    if isinstance(val, (int, float)) and float(val) == int(val):
        return fmt_int
    return fmt

def shift_formula_refs(formula, src_row, tgt_row):
    delta = tgt_row - src_row
    # Only handle single-letter columns (A-Z)
    def repl(m):
        col = m.group(1)
        row = int(m.group(2))
        return f"{col}{row + delta}"
    return re.sub(r'([A-Z])([0-9]+)', repl, formula)

# Style for D–F
def style_df(cell, value):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = get_num_fmt(value)

for bstart, bend in blocks:
    # Gather block data for sorting
    rows = []
    for r in range(bstart, bend+1):
        row_data = []
        for c in range(start_col, end_col+1):
            cell = ws.cell(row=r, column=c)
            row_data.append(cell.value)
        rows.append((r, row_data))
    # Evaluate D for sorting
    evals = []
    for idx, (row_num, row_data) in enumerate(rows):
        d_val = ws_data.cell(row=row_num, column=4).value
        try:
            num = float(d_val) if d_val is not None else float('-inf')
        except Exception:
            num = float('-inf')
        evals.append((num, idx, row_num, row_data))
    evals.sort(reverse=True)
    # Write block back sorted by D
    for dest_off, (_, orig_idx, orig_row, orig_row_data) in enumerate(evals):
        target_r = bstart + dest_off
        for c in range(start_col, end_col+1):
            src_cell = ws.cell(row=orig_row, column=c)
            tgt_cell = ws.cell(row=target_r, column=c)
            value = src_cell.value
            if c in (4,5,6):
                # Restore or move true formulas, adjusting references
                if src_cell.data_type == "f" or (isinstance(value, str) and str(value).startswith("=")):
                    # Adjust formula relative to new row
                    formula = str(value)
                    formula_new = shift_formula_refs(formula, orig_row, target_r)
                    tgt_cell.value = formula_new
                else:
                    tgt_cell.value = value
                # Style
                v_eval = ws_data.cell(row=orig_row, column=c).value
                style_df(tgt_cell, v_eval)
            else:
                tgt_cell.value = value
    # After writing block, format D-F
    for dest_off in range(bend-bstart+1):
        r = bstart + dest_off
        for c in (4,5,6):
            v_eval = ws_data.cell(row=rows[evals[dest_off][1]][0], column=c).value
            cell = ws.cell(row=r, column=c)
            style_df(cell, v_eval)

# Ensure exactly one blank row between blocks
prev_end = None
for i, (bstart, bend) in enumerate(blocks):
    if prev_end is not None and bstart - prev_end > 1:
        for r in range(prev_end+2, bstart):
            for c in range(start_col, end_col+1):
                ws.cell(row=r, column=c).value = None
    prev_end = bend

# Redo style for D-F everywhere (just in case)
for r in range(start_row, end_row+1):
    for c in (4,5,6):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        if val is not None:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

wb.active = wb.sheetnames.index("Sheet1")
wb.save(outfile)
