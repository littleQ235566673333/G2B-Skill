import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_6/group_42930/r0/evolve_42930/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_6/group_42930/r0/evolve_42930/output.xlsx'

# Read the data to pandas
df = pd.read_excel(input_path)

# Guess column names
cols = df.columns.tolist()
# Try to identify subdivision and build year columns
subdiv_col = [c for c in cols if 'sub' in c.lower()][0]
year_col = [c for c in cols if 'year' in c.lower() or 'built' in c.lower()][0]

# Find the earliest build year per subdivision
early = df.groupby(subdiv_col)[year_col].min().reset_index()
early = early.sort_values(year_col).reset_index(drop=True)
# Assign unique numbers to subdivisions based on earliest build year
early['subdiv_num'] = early.index + 1
# Map each house to its subdivision's number
subdiv2num = dict(zip(early[subdiv_col], early['subdiv_num']))
df['subdiv_num'] = df[subdiv_col].map(subdiv2num)
# Sort main df by subdiv_num, then by build year
df_sorted = df.sort_values(['subdiv_num', year_col]).reset_index(drop=True)

# Load with openpyxl to preserve structure
wb = load_workbook(input_path)
ws = wb.active

# Write result C2:C22
values = df_sorted['subdiv_num'].iloc[:21].tolist()
for i, val in enumerate(values):
    ws.cell(row=2+i, column=3, value=val)
wb.save(output_path)
