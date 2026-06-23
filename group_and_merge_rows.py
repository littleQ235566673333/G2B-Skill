import openpyxl
import pandas as pd
import numpy as np

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_2/regression_gate/after_fix/core_177-6/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_2/regression_gate/after_fix/core_177-6/output.xlsx'

# Load the workbook and sheets
wb = openpyxl.load_workbook(in_path)
ws_data = wb['DATA']
ws_combined = wb['combined']

# Get headers and data
headers = [cell.value for cell in ws_data[1][:18]]  # A:R => 18 columns
rows = []
for row in ws_data.iter_rows(min_row=2, max_col=18, values_only=True):
    rows.append(list(row))

# Create DataFrame
cols = headers
df = pd.DataFrame(rows, columns=cols)

def is_blank(x):
    return pd.isna(x) or x == '' or (isinstance(x, str) and x.isspace())

df = df.dropna(how='all')  # Remove fully blank rows
key_col = headers[7]  # 'ComboKey', H

# Filter out blank key values
df = df[~df[key_col].apply(is_blank)]

def sum_merge(g):
    take = g.iloc[0, :].copy()
    for col in headers[8:18]:  # I:R (0-based idx 8 to 17)
        vals = g[col].replace([chr(160), '', np.nan], 0)
        vals = pd.to_numeric(vals, errors='coerce').fillna(0)
        val_sum = vals.sum()
        if np.isclose(val_sum, 0):
            take[col] = ''
        else:
            take[col] = f"{val_sum:.2f}"
    return take

grouped = df.groupby(key_col, dropna=False)
merged_rows = [sum_merge(group) for _, group in grouped]
df_out = pd.DataFrame(merged_rows, columns=headers)

# Clear existing contents in 'combined'
for row in ws_combined['A1':'R8']:
    for cell in row:
        cell.value = None

# Write headers
for cidx, val in enumerate(headers, 1):
    ws_combined.cell(row=1, column=cidx, value=val)

# Write merged data rows
for ridx, row in enumerate(df_out.values, 2):  # Start at row 2
    if ridx > 8:
        break
    for cidx, val in enumerate(row, 1):
        ws_combined.cell(row=ridx, column=cidx, value=val)

wb.save(out_path)
