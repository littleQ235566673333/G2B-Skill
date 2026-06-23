import openpyxl
from datetime import datetime, timedelta
import math

# Paths
infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/group_44017/r1/evolve_44017/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/group_44017/r1/evolve_44017/output.xlsx'

wb = openpyxl.load_workbook(infile)
ws = wb['Data']

def get_excel_dt(cell):
    v = cell.value
    if v is None:
        return None
    # Excel serial (float)
    if isinstance(v, (float, int)):
        return openpyxl.utils.datetime.from_excel(v)
    # Excel formula: unsupported
    if isinstance(v, str) and v.startswith('='):
        return None
    # Datetime obj
    if isinstance(v, datetime):
        return v
    # Fallback: try parse
    try: return datetime.fromisoformat(v)
    except: return None

# Gather month header dates (row 9, AD:AO i.e. cols 30-41)
month_dates = []
for col in range(30, 42):
    dt = get_excel_dt(ws.cell(row=9, column=col))
    month_dates.append(dt)

for row in range(14, 43):
    base = ws.cell(row=row, column=23).value
    eff_date = ws.cell(row=row, column=12).value
    freq = ws.cell(row=row, column=10).value
    # Make sure freq is int (handle possible formula result)
    if isinstance(freq, str):
        try: freq = int(freq)
        except: freq = None
    try: 
        # Excel formula evaluating freq is not directly possible
        if freq is not None: freq = int(float(freq))
    except: freq = None
    increases = []
    for col in range(13, 17):
        p = ws.cell(row=row, column=col).value
        if p is None or p == '':
            increases.append(0.0)
        else:
            increases.append(float(p))
    # Effective date as datetime
    if isinstance(eff_date, str):
        try: eff_date = datetime.fromisoformat(eff_date)
        except: eff_date = None
    # If eff_date is float (Excel serial)
    if isinstance(eff_date, (float, int)):
        eff_date = openpyxl.utils.datetime.from_excel(eff_date)
    elif not isinstance(eff_date, datetime):
        eff_date = None
    for idx, month_col in enumerate(range(30, 42)):
        col_dt = month_dates[idx]
        cell = ws.cell(row=row, column=month_col)
        if (eff_date is None or col_dt is None or freq is None or base is None):
            cell.value = None
            continue
        if col_dt < eff_date:
            cell.value = None
            continue
        # Determine which increases are live
        phases = [eff_date + timedelta(days=30*freq*n) for n in range(4)]
        cumulative = float(base)
        for i in range(4):
            # Only if increase started
            if col_dt >= phases[i]:
                cumulative *= (1 + increases[i])
        cell.value = round(cumulative,2)
        cell.fill = openpyxl.styles.PatternFill(fill_type=None) # Remove fill

wb.save(outfile)
print('done')
