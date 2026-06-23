import openpyxl
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_41978_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_41978_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

year_col = 'G'
status_col = 'J'
min_row, max_row = 14, 185

# Count 'open' cases per year
year_status = defaultdict(int)
for row in range(min_row, max_row+1):
    year = ws[f'{year_col}{row}'].value
    status = ws[f'{status_col}{row}'].value
    if year is not None and str(status).strip().lower() == 'open':
        year_status[year] += 1

sorted_years = sorted(year_status)
for idx, year in enumerate(sorted_years[:10], 2):
    ws[f'I{idx}'].value = year_status[year]

wb.save(output_path)
