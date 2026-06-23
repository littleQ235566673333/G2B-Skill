from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

INPUT_FN = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_2/regression_gate/after_pass/core_32337/input.xlsx'
OUTPUT_FN = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_2/regression_gate/after_pass/core_32337/output.xlsx'

wb = load_workbook(INPUT_FN)
ws = wb['Sheet1']

# Step 1: Set formula for the 'Expected Result' column E3:E15
for row in range(3, 16):
    ws.cell(row=row, column=5).value = f"=INDEX($I$3:$I$15, MATCH(O{row}, $O$3:$O$15, 0))"

# Step 2: Insert new column after 'age (year)' (O=15), so at col 16 (P)
ws.insert_cols(16)
ws.cell(row=2, column=16).value = 'Actual Age'
for row in range(3, 16):
    ws.cell(row=row, column=16).value = f'=DATEDIF($C{row},$B$1,"Y")'
    ws.cell(row=row, column=16).number_format = '0'

# Step 3: Format new 'Actual Age': duplicate only PatternFill components
age_year_fill = ws.cell(row=2, column=15).fill
if isinstance(age_year_fill, PatternFill):
    # Construct a new PatternFill from the copied one (handles openpyxl copy bug)
    new_fill = PatternFill(
        fill_type=age_year_fill.fill_type,
        fgColor=age_year_fill.fgColor.rgb,
        bgColor=age_year_fill.bgColor.rgb
    )
else:
    new_fill = None

for row in range(3, 16):
    cell = ws.cell(row=row, column=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    if new_fill is not None:
        cell.fill = new_fill
header_cell = ws.cell(row=2, column=16)
header_cell.alignment = Alignment(horizontal='center', vertical='top')
header_cell.font = Font(bold=True)
if new_fill is not None:
    header_cell.fill = new_fill

wb.save(OUTPUT_FN)
print('Done')
