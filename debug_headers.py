from openpyxl import load_workbook
input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_1/regression_gate/after_pass/core_9726/input.xlsx'
wb = load_workbook(input_fp)
sheetnames = wb.sheetnames
ws = wb[sheetnames[0]]
header_row = 1
colnames = [ws.cell(row=header_row, column=j+1).value for j in range(ws.max_column)]
print(repr(colnames))
