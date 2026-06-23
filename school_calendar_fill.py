import openpyxl

input_file = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_10/task_50916/r1/evolve_50916/input.xlsx'
output_file = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_10/task_50916/r1/evolve_50916/output.xlsx'

wb = openpyxl.load_workbook(input_file)
ws = wb[wb.sheetnames[0]]  # Use the first sheet (always check)

# Detect cycle day numbers and class schedule (A2:A8 and B2:H8)
cycle_days = []
class_schedule = []
for i in range(2, 9):  # Excel is 1-based, Python is 0-based
    day_num = ws[f'A{i}'].value
    row_classes = [ws.cell(row=i, column=j).value for j in range(2, 9)]  # Columns B-H
    cycle_days.append(day_num)
    class_schedule.append(row_classes)

# Fill out C12:H14 based on cycle day in A12:A14
for target_row in range(12, 15):  # Rows 12-14
    cycle_day = ws.cell(row=target_row, column=1).value  # A12, A13, A14
    try:
        idx = cycle_days.index(cycle_day)
        classes = class_schedule[idx]
    except (ValueError, IndexError):
        classes = [None] * 7  # Can't find
    for target_col in range(3, 9):  # C:H = 3~8 (fixed range)
        ws.cell(row=target_row, column=target_col, value=classes[target_col-2])

wb.save(output_file)
