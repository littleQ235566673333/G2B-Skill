import openpyxl

# Load workbook and worksheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_49945_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_49945_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read models to check from G4:G6
models = [ws[f'G{r}'].value for r in range(4, 7)]

# Read C4:C11 and D4:D11 for existing model-quantity pairs
model_col = [ws[f'C{r}'].value for r in range(4, 12)]
qty_col = [ws[f'D{r}'].value for r in range(4, 12)]

# Calculate sum for each model in G4:G6
for idx, m in enumerate(models):
    total = sum(q for model, q in zip(model_col, qty_col) if model == m and isinstance(q, (int, float)))
    ws[f'G{4+idx}'] = total

wb.save(output_path)
