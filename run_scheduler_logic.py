import openpyxl
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_2/regression_gate/before_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_2/regression_gate/before_pass/core_55421/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read all relevant records
rows = list(ws.iter_rows(min_row=2, max_row=20, min_col=1, max_col=5, values_only=True))

# Group by Column A (number)
by_number = defaultdict(list)
for row in rows:
    number = row[0]
    by_number[number].append(row)

results = {}
for number, group in by_number.items():
    statuses = {r[3] for r in group}
    has_ns = 'NO SHOW' in statuses
    has_sch = 'SCH' in statuses
    only_sch = statuses == {'SCH'}
    both_ns_sch = statuses == {'NO SHOW', 'SCH'}
    if only_sch:
        for r in group:
            idx = rows.index(r) + 2
            results[idx] = 'FUTURE'
    elif both_ns_sch:
        for r in group:
            idx = rows.index(r) + 2
            results[idx] = 'NS/SCHED'
    elif has_ns and not has_sch:
        for r in group:
            idx = rows.index(r) + 2
            if r[4]:  # Date in Column E (index 4)
                results[idx] = 'NO ACTION NEEDED'
            else:
                results[idx] = 'CALL PT'
    else:
        for r in group:
            idx = rows.index(r) + 2
            results.setdefault(idx, '')

# Write results in F2:F20
for i in range(2, 21):
    ws[f'F{i}'] = results.get(i, '')

wb.save(output_path)
