import openpyxl

def sort_spreadsheet(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Read data range
    data = []
    header = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not header:
            header = row
            continue
        # Columns: B, C, J (indices 1, 2, 9)
        b, c, j = row[1], row[2], row[9]
        # Skip rows where B and C are empty
        if not b and not c:
            continue
        data.append({'B': b, 'C': c, 'J': j, 'row': row})

    # Remove duplicates: identical B & C
    unique = []
    seen = set()
    for item in data:
        key = (item['B'], item['C'])
        if key in seen:
            continue
        seen.add(key)
        unique.append(item)
    data = unique

    # Identify header and skip it if present (row 1)
    # Process helper column J
    j_values = [item['J'] for item in data if item['J'] not in [None, '', ' ']]
    j_values_ordered = []
    seen_j = set()
    for j in j_values:
        if j not in seen_j:
            j_values_ordered.append(j)
            seen_j.add(j)
    # If J is empty for all, sort alphabetically by B
    output_rows = []
    if j_values_ordered:
        # Group by J, then remainder
        listed = []
        others = []
        for j in j_values_ordered:
            group = [item for item in data if item['B'] == j]
            listed.extend(group)
        for item in data:
            if item['B'] not in j_values_ordered:
                others.append(item)
        output_rows = listed + others
    else:
        # Sort alphabetically A-Z by B
        output_rows = sorted(data, key=lambda x: (str(x['B']).lower() if x['B'] else ''))

    # Remove empty cells and header from output_rows
    final_rows = [item for item in output_rows if item['B'] not in [None, '', header[1]]]

    # Output in F2:H10
    out_start_row = 2
    out_end_row = 10
    output_range = final_rows[:out_end_row-out_start_row+1]
    # Columns F,G,H are 5,6,7; Output G and H, sorted by H lowest-highest
    to_output = [(row['row'][6], row['row'][7]) for row in output_range]  # G,H
    # Sort only column H for output
    to_output_sorted = sorted(to_output, key=lambda x: (x[1] if x[1] is not None else float('inf')))

    for idx, (g, h) in enumerate(to_output_sorted, start=out_start_row):
        ws.cell(row=idx, column=7, value=g)  # G
        ws.cell(row=idx, column=8, value=h)  # H

    wb.save(output_path)

sort_spreadsheet(
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_22-47_tc1/input.xlsx',
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_22-47_tc1/output.xlsx'
)
