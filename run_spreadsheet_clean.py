import numpy as np
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/group_387-16/r1/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/group_387-16/r1/evolve_387-16/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

datarows = []
for r in range(2, 19):  # A2:D18
    datarows.append([ws.cell(r, c).value for c in range(1, 5)])

data = np.array(datarows, dtype=object)
col_a, col_value, col_binaries, col_result_values = 0, 1, 2, 3
values = list(data[:, col_value])
binaries = list(data[:, col_binaries])
colas = list(data[:, col_a])
result_values = [v for v in data[:, col_result_values] if v is not None]

removed_indices = set()
for test_val in result_values:
    for idx, val in enumerate(values):
        if idx in removed_indices:
            continue
        if val == test_val:
            removed_indices.add(idx)
            break
# Keep only non-removed indices
def keep_indices(orig_list):
    return [x for i, x in enumerate(orig_list) if i not in removed_indices]
values_cleaned = keep_indices(values)
binaries_cleaned = keep_indices(binaries)
col_a_cleaned = keep_indices(colas)

# Write cleaned columns back, shift up, clear below
final_row_ct = len(values_cleaned)
for i in range(17):  # rows A2:D18
    ws.cell(i+2, col_value+1).value = values_cleaned[i] if i < final_row_ct else None
    ws.cell(i+2, col_binaries+1).value = binaries_cleaned[i] if i < final_row_ct else None
    ws.cell(i+2, col_a+1).value = col_a_cleaned[i] if i < final_row_ct else None

# Solver result is sum of col A (cleaned)
solver_result = sum([v for v in col_a_cleaned if isinstance(v, (int, float))])
ws['C1'] = 'Solver result'
ws['C2'] = solver_result
# Target value from D1 or D2

target_value = None
if ws['D1'].value is not None and isinstance(ws['D1'].value, (int, float)):
    target_value = ws['D1'].value
elif ws['D2'].value is not None and isinstance(ws['D2'].value, (int, float)):
    target_value = ws['D2'].value
diff = solver_result - (target_value if target_value is not None else 0)
ws['D1'] = 'Difference'
ws['D2'] = diff

wb.save(output_path)
