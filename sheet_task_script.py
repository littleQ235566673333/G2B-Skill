import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_50631_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_50631_tc1/output.xlsx'

# Open the workbook
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Set Calibri font throughout
for row in ws.iter_rows():
    for cell in row:
        cell.font = Font(name='Calibri')

# Hide gridlines
ws.sheet_view.showGridLines = False

# Setup fill color for column H (yellow)
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Get holidays from column J (assuming all are date-formatted, starting from J3)
holidays = []
for row in range(3, ws.max_row + 1):
    holiday_cell = ws[f'J{row}']
    if holiday_cell.value:
        holidays.append(holiday_cell.coordinate)

# Find the holidays range (for formula)
holiday_range = f'J3:J{ws.max_row}'

# Insert WORKDAY formula in H7:H38
for i, row in enumerate(range(7, 39)):
    cell = ws[f'H{row}']
    # Formula: =IF($B$3+ROWS($B$7:$B7)-1<=$F$3,WORKDAY($B$3+ROWS($B$7:$B7)-1,0,$J$3:$J$38),"")
    formula = f'=IF($B$3+ROWS($B$7:$B{row})-1<=$F$3,WORKDAY($B$3+ROWS($B$7:$B{row})-1,0,{holiday_range}),"")'
    cell.value = formula
    cell.fill = yellow_fill
    cell.font = Font(name="Calibri")

wb.save(output_path)
