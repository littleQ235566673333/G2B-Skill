from openpyxl import load_workbook
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2-PRUNED/eval_seed42/eval_46240_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2-PRUNED/eval_seed42/eval_46240_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def is_date(val):
    return isinstance(val, (datetime.datetime, datetime.date))

for row in range(2, 5):  # J2:J4 (rows 2-4)
    h_val = ws[f'H{row}'].value
    if h_val == 'N/A':
        result = 'N/A'
    elif is_date(h_val):
        result = 'Yes'
    elif str(h_val).strip().lower() == 'yes':
        result = 'Yes'
    elif h_val in (None, ''):
        result = 'No'
    else:
        result = 'No'
    ws[f'J{row}'] = result

wb.save(output_path)
