from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_1/regression_gate/before_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_1/regression_gate/before_pass/core_18935/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Map category to its column index in the data table
# Data table area identified by inspection: headers at (row 4), data from row 5-10, columns 1-5 (A-E)
data_table_start_row = 5
data_table_end_row = 10
category_map = {'Category 1': 3, 'Category 2': 4, 'Category 3': 5} # Columns C,D,E

for r in range(17, 23):
    work = ws[f'A{r}'].value
    category_type = ws[f'B{r}'].value
    material = ws[f'C{r}'].value
    value_found = None
    if work is not None and category_type in category_map and material is not None:
        # Search the data table for matching row
        for data_r in range(data_table_start_row, data_table_end_row + 1):
            dt_work = ws.cell(row=data_r, column=1).value
            dt_material = ws.cell(row=data_r, column=2).value
            if dt_work == work and dt_material == material:
                lookup_col = category_map[category_type]
                value_found = ws.cell(row=data_r, column=lookup_col).value
                break
    ws[f'D{r}'] = value_found

wb.save(output_path)
