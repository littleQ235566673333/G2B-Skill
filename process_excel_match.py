import openpyxl

def match_and_fill(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    base_ws = wb['Base']
    streets_ws = wb['Streets']

    # Collect Streets match data, skip rows with missing C or D
    streets_data = []
    for row in streets_ws.iter_rows(min_row=2, max_col=6, values_only=True):
        street_b, street_c, street_d, street_f = row[1], row[2], row[3], row[5]
        if street_c is None or street_d is None:
            continue
        streets_data.append({
            'B': street_b,
            'C': street_c,
            'D': street_d,
            'F': street_f
        })

    # Fill Base E2:E26
    for base_row in range(2, 27):
        base_c = base_ws.cell(row=base_row, column=3).value
        base_d = base_ws.cell(row=base_row, column=4).value
        result = ""
        if base_c is not None and base_d is not None:
            for s in streets_data:
                if (
                    s['B'] == base_c and
                    isinstance(base_d, (int, float)) and
                    isinstance(s['C'], (int, float)) and
                    isinstance(s['D'], (int, float)) and
                    s['C'] <= base_d <= s['D']
                ):
                    result = s['F'] if s['F'] is not None else ""
                    break
        base_ws.cell(row=base_row, column=5).value = result

    wb.save(output_path)

match_and_fill(
    "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_13284_tc1/input.xlsx",
    "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_13284_tc1/output.xlsx"
)
