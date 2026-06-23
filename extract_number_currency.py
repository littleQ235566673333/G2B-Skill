import openpyxl
import re

def extract_number_and_decimal_sign(text):
    # Remove currency symbols (common ones), spaces, and thousands separators for detection
    clean = re.sub(r'[\$€£¥₹złfrs₽元₩₺₫₦₴₲₵₡₢₣₤₥₧₱₲₳₴₵₶]+', '', str(text))
    clean = clean.replace(' ', '')
    # Try to detect correct decimal and thousand separator
    possibles = re.findall(r'((\d{1,3}(?:[.,\s]\d{3})+)([.,]\d+))', clean)
    if possibles:
        # Example: 1,234.56 or 1.234,56
        number_part = possibles[0][0]
        # Detect which separator is used for decimals (the last one)
        if ',' in possibles[0][2]:
            decimal_sign = ','
        elif '.' in possibles[0][2]:
            decimal_sign = '.'
        else:
            decimal_sign = None
        # Remove all except digits and the decimal separator
        result = re.sub(r'[^0-9'+re.escape(decimal_sign)+']', '', number_part)
        try:
            if decimal_sign == ',':
                value = float(result.replace('.', '').replace(',', '.'))
            else:
                value = float(result.replace(',', ''))
        except:
            value = None
        return (value, decimal_sign)
    # Single decimal separator
    matches = re.findall(r'\d+[.,]\d+', clean)
    if matches:
        number_part = matches[0]
        decimal_sign = '.' if '.' in number_part else ','
        try:
            value = float(number_part.replace(',', '.') if decimal_sign == ',' else number_part)
        except:
            value = None
        return (value, decimal_sign)
    # Only integer
    matches = re.findall(r'\d+', clean)
    if matches:
        value = int(matches[0])
        return (value, None)
    return (None, None)

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_4/task_48257/r2/evolve_48257/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_4/task_48257/r2/evolve_48257/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active
for row in range(2, 6):
    val = ws[f'A{row}'].value
    num, dec = extract_number_and_decimal_sign(val)
    if num is None:
        ws[f'C{row}'] = ''
    elif dec is not None:
        ws[f'C{row}'] = f'{num} (decimal: {dec})'
    else:
        ws[f'C{row}'] = f'{num}'
wb.save(output_path)
