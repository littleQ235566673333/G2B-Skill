import openpyxl
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/group_59902/r3/evolve_59902/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/group_59902/r3/evolve_59902/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Read names and dates from G5:G28 and F5:F28
names = [ws[f'G{row}'].value for row in range(5, 29)]
dates = [ws[f'F{row}'].value for row in range(5, 29)]

# Compute days since last sale per staff (ascending dates)
past_date_per_person = {}
diff_days = []
for idx, (name, date) in enumerate(zip(names, dates)):
    if name not in past_date_per_person or past_date_per_person[name] is None:
        diff_days.append(0)
    else:
        prev_date = past_date_per_person[name]
        if isinstance(date, datetime) and isinstance(prev_date, datetime):
            diff_days.append((date - prev_date).days)
        else:
            diff_days.append('')
    past_date_per_person[name] = date

# Write results to C5:C28 with Accounting format
for idx, val in enumerate(diff_days):
    cell = ws[f'C{5 + idx}']
    cell.value = val
    # Accounting format without currency symbol (shows dashes for 0 or blanks)
    cell.number_format = '_(* #,##0_);_(* \-#,##0);_(* "-"??_);_(@_)'

wb.save(output_path)
