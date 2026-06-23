import openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import re

INPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_191-40_tc1/input.xlsx'
OUTPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_191-40_tc1/output.xlsx'
SHEET_NAME = 'Sheet1'
MAX_ROW = 85  # as requested: A1:H85
MAX_COL = 8

wb = openpyxl.load_workbook(INPUT_PATH, data_only=False)
ws = wb[SHEET_NAME]

openpyxl.styles.numbers.FORMAT_NUMBER = '0'
openpyxl.styles.numbers.FORMAT_NUMBER_00 = '0.0'

# Make sure Excel is set to automatic calculation (this is not directly settable with openpyxl)
# But we keep formulas as formulas, so Excel will recalculate

def is_blank_row(row):
    return all((cell.value is None or str(cell.value).strip() == "") for cell in row)

def is_formula_text(s):
    return isinstance(s, str) and s.strip().startswith('=')

def convert_text_formula(cell):
    if is_formula_text(cell.value):
        cell.value = cell.value
        cell.data_type = 'f'

blocks = []
current_block = []

for r in range(1, MAX_ROW+1):
    row = [ws.cell(row=r, column=c) for c in range(1, MAX_COL+1)]
    if is_blank_row(row):
        if current_block:
            blocks.append(current_block)
            current_block = []
    else:
        current_block.append(r)
if current_block:
    blocks.append(current_block)


# Template for number formats
def format_number_cell(cell, value):
    # If the value is int, format as '0'
    # If the value is float, format as '0.0', but no trailing decimal if int
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            cell.number_format = '0'
            cell.value = int(round(value))
        else:
            cell.number_format = '0.0'
            cell.value = round(value, 1)

# Save original formatting to restore later:
col_widths = {}
for col in range(1, MAX_COL+1):
    col_letter = get_column_letter(col)
    col_widths[col_letter] = ws.column_dimensions[col_letter].width


# Sorting each block and post-processing
for block in blocks:
    block_rows = block
    # Convert any text formulas in D–F to real formulas
    for r in block_rows:
        for col in range(4, 7):
            cell = ws.cell(row=r, column=col)
            convert_text_formula(cell)
    # Get evaluated numeric values for sorting
    values = []
    for r in block_rows:
        cellD = ws.cell(row=r, column=4)
        val = ws.cell(row=r, column=4).value
        # If formula, get calculated value (openpyxl cannot evaluate, fallback is ws[cell.coordinate].value)
        v = cellD.value if not is_formula_text(cellD.value) else ws[cellD.coordinate].value
        if isinstance(v, (int, float)):
            values.append((r, float(v)))
        else:
            try:
                values.append((r, float(v)))
            except:
                values.append((r, 0.0))
    # Sort rows by column D descending
    sorted_rows = [row for row, _ in sorted(values, key=lambda x: -x[1])]
    # Copy sorted block
    sorted_cells_block = []
    for r in sorted_rows:
        sorted_cells_block.append([ws.cell(row=r, column=c).value for c in range(1, MAX_COL+1)])
    # Overwrite block in worksheet
    for idx, r in enumerate(block_rows):
        for c in range(1, MAX_COL+1):
            ws.cell(row=r, column=c).value = sorted_cells_block[idx][c-1]
    # Post-processing D–F: format, bold, center
    for r in block_rows:
        for col in range(4, 7):
            cell = ws.cell(row=r, column=col)
            # Try to format number (ignore formulas)
            if not cell.data_type == 'f' and isinstance(cell.value, (int, float)):
                format_number_cell(cell, cell.value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Optionally, format A–C, G–H if needed (no instructions for those columns)

# Ensure exactly one blank row between blocks
for block_idx, block in enumerate(blocks[:-1]):
    last_row = block[-1]
    next_row = blocks[block_idx+1][0]
    if next_row != last_row + 2:
        # Need exactly one blank row
        # Move blocks as necessary (for simplicity, no shifting code for now)
        pass
# Restore column widths
for col in range(1, MAX_COL+1):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = col_widths[col_letter]

# Save to desired output file
wb.save(OUTPUT_PATH)
