import openpyxl
from copy import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/group_208-20/r1/evolve_208-20/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/group_208-20/r1/evolve_208-20/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Q']

# Get Actual headers A2:F2
headers = [ws.cell(row=2, column=col).value for col in range(1, 7)]
# Read all data rows (A3:F37)
data_rows = []
raw_rows = []
for row in ws.iter_rows(min_row=3, max_row=37, min_col=1, max_col=6):
    if any(cell.value for cell in row):
        data_rows.append([copy(cell) for cell in row])
        raw_rows.append(row)

# Sort rows: by E (col index 4, 0-based), None/blank last, descending
def rdays_key(row):
    v = row[4].value
    return (v is not None, v if v is not None else float('-inf'))

data_sorted = sorted(data_rows, key=rdays_key, reverse=True)
# Write sorted rows back to ws (A3:F37)
for row_idx, row in enumerate(data_sorted, start=3):
    for col_idx, cell in enumerate(row, start=1):
        dest = ws.cell(row=row_idx, column=col_idx)
        dest.value = cell.value
        dest.data_type = cell.data_type
        dest.number_format = cell.number_format
        dest.font = copy(cell.font)
        dest.fill = copy(cell.fill)
        dest.border = copy(cell.border)
        dest.alignment = copy(cell.alignment)
        dest.protection = copy(cell.protection)
# Save result
wb.save(output_path)
