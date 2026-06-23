from openpyxl import load_workbook

def get_cell_display_value(cell):
    if cell.is_date:
        return cell.value.strftime('%Y-%m-%d')
    if cell.data_type == 'f':
        return str(cell.value) if cell.value is not None else ''
    return str(cell.value) if cell.value is not None else ''

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_5/regression_gate/before_fix/core_48975/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_5/regression_gate/before_fix/core_48975/output.xlsx'
wb = load_workbook(input_path)
ws_in = wb['Input']
ws_out = wb['Output']
results = []
for row in ws_in.iter_rows(min_row=2):
    a, b, _, _, todo = row[:5]
    # Skip if 'To Do' column isn't 'yes'
    if (str(todo.value or '').strip().lower() != 'yes'):
        continue
    # Skip if either A or B is empty
    if not (a.value and b.value):
        continue
    val_a = get_cell_display_value(a)
    val_b = get_cell_display_value(b)
    if val_a and val_b:
        results.append((val_a, val_b))
# Write results to Output!B11:B17
start_row = 11
for i, (v1, v2) in enumerate(results):
    ws_out.cell(row=start_row + i, column=2, value=v1)
    # Optionally, copy B values as well
    # ws_out.cell(row=start_row + i, column=3, value=v2)
wb.save(output_path)
