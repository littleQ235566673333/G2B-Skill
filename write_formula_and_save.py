import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_2/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_2/regression_gate/after_pass/core_50526/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# The array formula to list color headers where corresponding lookup value is >0
formula = '=IFERROR(INDEX($B$1:$G$1, SMALL(IF(INDEX($B$2:$G$3, MATCH($B$6, $A$2:$A$3, 0), 0)>0, COLUMN($B$1:$G$1)-COLUMN($B$1)+1), ROW(1:1))), "")'

# Enter formula in B9 and B10
ws['B9'].value = formula
ws['B10'].value = formula.replace('ROW(1:1)', 'ROW(2:2)')

wb.save(output_path)
