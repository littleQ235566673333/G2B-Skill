import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wbin = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_5/regression_gate/before_pass/core_32337/input.xlsx'
wout = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_5/regression_gate/before_pass/core_32337/output.xlsx'
wb = openpyxl.load_workbook(wbin)
ws = wb[wb.sheetnames[0]]

# Map headers from row 2
header_row_idx = 2
headers = [cell.value for cell in ws[header_row_idx]]
header_map = {h: i + 1 for i, h in enumerate(headers) if h}

dob_col = header_map['DATE OF BIRTH']
report_date_col = 2 # $B$1
age_col = header_map['age (year)']
category_col = header_map['CATEGORY']
expected_col = header_map['Expected Result']
year_age_col = header_map['YEAR (age)']

# Place new column after 'age (year)'
actual_age_col = age_col + 1
actual_age_col_letter = get_column_letter(actual_age_col)
data_start, data_end = 3, 15

# Try to grab the fill pattern from age (year), but fall back to no fill if unsupported
p_header_cell = ws.cell(row=header_row_idx, column=age_col)
age_fill = p_header_cell.fill
if not isinstance(age_fill, PatternFill) or (age_fill.patternType is None):
    age_fill = PatternFill(fill_type=None)

# 1. Fill in Expected Result (E) by looking up value in CATEGORY by year
for r in range(data_start, data_end + 1):
    ws.cell(row=r, column=expected_col).value = (
        f'=INDEX(${get_column_letter(category_col)}$3:${get_column_letter(category_col)}$15, '
        f'MATCH(${get_column_letter(year_age_col)}{r}, ${get_column_letter(year_age_col)}$3:${get_column_letter(year_age_col)}$15, 0))'
    )

# 2. Insert 'Actual Age' after ‘age (year)’ column
header_cell = ws.cell(row=header_row_idx, column=actual_age_col)
header_cell.value = 'Actual Age'
header_cell.fill = age_fill
header_cell.alignment = Alignment(horizontal='center', vertical='top')
header_cell.font = Font(bold=True)
for r in range(data_start, data_end + 1):
    c = ws.cell(row=r, column=actual_age_col)
    c.value = f'=ROUNDDOWN((($B$1-C{r})/365),0)'
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center')
    c.fill = age_fill

wb.save(wout)
