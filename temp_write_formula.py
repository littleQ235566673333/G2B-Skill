from openpyxl import load_workbook

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_3/regression_gate/before_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_3/regression_gate/before_pass/core_50526/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Formulas to be written
formula_b9 = '=IFERROR(INDEX($B$1:$Z$1,SMALL(IF(INDEX($B$2:$Z$100,MATCH($B$6,$A$2:$A$100,0),)>0,COLUMN($B$1:$Z$1)-COLUMN($B$1)+1),ROW(1:1))),"")'
formula_b10 = '=IFERROR(INDEX($B$1:$Z$1,SMALL(IF(INDEX($B$2:$Z$100,MATCH($B$6,$A$2:$A$100,0),)>0,COLUMN($B$1:$Z$1)-COLUMN($B$1)+1),ROW(2:2))),"")'

ws['B9'].value = formula_b9
ws['B10'].value = formula_b10

wb.save(output_path)
