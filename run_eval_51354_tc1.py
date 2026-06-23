from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def parse_month_year(text):
    s = text.strip()[-6:]
    # Handle if there is an extra space or longer month (e.g., 'Sept 21')
    if len(s) < 6 or s[3] != ' ':
        s = text.strip()[-7:]
    parts = s.strip().split(' ')
    if len(parts) == 2:
        month_str, year_str = parts
    else:
        month_str, year_str = '', ''
    try:
        dt = datetime.strptime(month_str + ' ' + year_str, '%b %y')
    except Exception:
        return '', year_str, ''
    # Add one month
    m, y = dt.month, dt.year
    if m == 12:
        nm, ny = 1, y + 1
    else:
        nm, ny = m + 1, y
    next_month = datetime(ny, nm, 1).strftime('%b %y')
    return month_str, year_str, next_month

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun2/eval_51354_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun2/eval_51354_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

for row in range(2, 7):
    val = ws[f'A{row}'].value
    if val is None or str(val).strip() == '':
        ws[f'D{row}'] = ''
        ws[f'E{row}'] = ''
        continue
    month_str, year_str, next_month = parse_month_year(str(val))
    ws[f'D{row}'] = year_str
    ws[f'E{row}'] = next_month
    ws[f'D{row}'].fill = PatternFill(fill_type='solid', fgColor='FFC000')

wb.save(output_path)
