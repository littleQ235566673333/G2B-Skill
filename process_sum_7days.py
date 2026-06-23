import openpyxl
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_2/group_59595/r3/evolve_59595/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_2/group_59595/r3/evolve_59595/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Extract dates and points for target rows (C4:C19 <- rows 4-19)
dates = []
points = []
for row in range(4, 20):
    a = ws[f'A{row}'].value
    b = ws[f'B{row}'].value
    dates.append(a)
    points.append(b)

for idx, date in enumerate(dates):
    if not isinstance(date, (datetime.datetime, datetime.date)):
        ws[f'C{4 + idx}'] = None
        continue
    # Support both datetime and date
    if isinstance(date, datetime.datetime):
        date_as_date = date.date()
    else:
        date_as_date = date
    window_start = date_as_date - datetime.timedelta(days=6)
    total = 0
    for j, adate in enumerate(dates):
        if not isinstance(adate, (datetime.datetime, datetime.date)):
            continue
        if isinstance(adate, datetime.datetime):
            adate_as_date = adate.date()
        else:
            adate_as_date = adate
        if window_start <= adate_as_date <= date_as_date:
            val = points[j]
            if isinstance(val, (int, float)):
                total += val
    ws[f'C{4 + idx}'] = total

wb.save(output_path)
