import pandas as pd
from datetime import timedelta
from openpyxl import load_workbook

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_2/group_59595/r2/evolve_59595/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_2/group_59595/r2/evolve_59595/output.xlsx'

# Read with pandas, the actual data starts from row 3, header=1, and skip row 0
raw = pd.read_excel(input_path, header=1)
# Remove the redundant first row (which is likely column names repeated)
df = raw[1:].copy()

# Fix column names
col_date = 'Reporting'
col_points = 'Points'
df[col_date] = pd.to_datetime(df[col_date])
df[col_points] = pd.to_numeric(df[col_points], errors='coerce')

# Calculate rolling 7-day sum for each row
def rolling_7day_sum(idx, df, date_col, points_col):
    current_date = df.iloc[idx][date_col]
    window_start = current_date - timedelta(days=6)
    mask = (df[date_col] >= window_start) & (df[date_col] <= current_date)
    return df.loc[mask, points_col].sum()

sums = [rolling_7day_sum(i, df, col_date, col_points) for i in range(len(df))]

# Write results to C4:C19 in the original workbook
wb = load_workbook(input_path)
ws = wb.active
for target_row, value in zip(range(4, 20), sums):
    ws.cell(row=target_row, column=3, value=value)
wb.save(output_path)
print('Done.')
