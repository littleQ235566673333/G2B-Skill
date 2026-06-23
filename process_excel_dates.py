import openpyxl
import pandas as pd

input_xlsx = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/group_45896/r0/evolve_45896/input.xlsx'
output_xlsx = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/group_45896/r0/evolve_45896/output.xlsx'

# Load workbooks and sheets
wb = openpyxl.load_workbook(input_xlsx)
ws = wb['Volym P5_P6_2023']
zord_df = pd.read_excel(input_xlsx, sheet_name='ZORD')

for row in range(2, 11):  # Rows 2 to 10
    key = ws.cell(row=row, column=1).value
    if key is not None:
        matches = zord_df[zord_df.iloc[:, 0] == key].iloc[:, 2]
        # Remove nulls, unique, format as date string
        dates = pd.unique(matches.dropna())
        date_strs = []
        for d in dates:
            try:
                # Try parsing each value into a date
                dt = pd.to_datetime(d)
                date_strs.append(dt.strftime('%d/%m/%Y'))
            except Exception:
                pass  # Ignore failures to parse as date (e.g. if cell is not a date)
        ws.cell(row=row, column=9).value = ','.join(date_strs) if date_strs else None

wb.save(output_xlsx)
