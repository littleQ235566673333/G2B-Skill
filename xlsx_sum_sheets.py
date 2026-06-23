from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# File paths
target_output = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/regression_gate/after_pass/core_84-40/output.xlsx'
input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/regression_gate/after_pass/core_84-40/input.xlsx'

wb = load_workbook(input_fp)
list_ws = wb['LIST']

# Clear LIST except headers
for row in list_ws.iter_rows(min_row=2, max_row=list_ws.max_row, min_col=2, max_col=5):
    for cell in row:
        cell.value = None

# Sheets to process (exclude LIST)
sheets = [s for s in wb.sheetnames if s != 'LIST']

for idx, sheet_name in enumerate(sheets, start=2):
    ws = wb[sheet_name]
    # Sum IMPORT (C) and EXPORT (D) columns (skip header)
    sum_import, sum_export = 0, 0
    for r in ws.iter_rows(min_row=2, min_col=3, max_col=4, values_only=True):
        imp, exp = r
        if imp is not None:
            sum_import += imp
        if exp is not None:
            sum_export += exp
    list_ws[f'B{idx}'] = sheet_name
    list_ws[f'C{idx}'] = sum_import
    list_ws[f'D{idx}'] = sum_export
    list_ws[f'E{idx}'] = f"=C{idx}-D{idx}"

# Write total row
data_end = 2+len(sheets)
list_ws[f'B{data_end}'] = 'TOTAL'
list_ws[f'C{data_end}'] = f"=SUM(C2:C{data_end-1})"
list_ws[f'D{data_end}'] = f"=SUM(D2:D{data_end-1})"
list_ws[f'E{data_end}'] = f"=C{data_end}-D{data_end}"

wb.save(target_output)
