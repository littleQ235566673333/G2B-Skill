from openpyxl import load_workbook
from datetime import datetime

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_1/regression_gate/before_pass/core_41589/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_1/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(in_path)
ws = wb['Contact List']
last_contact = ws['H4'].value
reg_contact = ws['I4'].value if ws['I4'].value is not None else ''
today = datetime.today()
ans = 'NO ACTION'

if isinstance(last_contact, datetime):
    days_ago = (today - last_contact).days
    if 0 <= days_ago <= 30 and str(reg_contact).strip().upper() == 'YES':
        ans = 'HOLD'
    elif days_ago > 30 and str(reg_contact).strip().upper() == 'YES':
        ans = 'TOUCH BASE'

ws['J4'] = ans
wb.save(out_path)
