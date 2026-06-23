from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/regression_gate/before_fix/core_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/regression_gate/before_fix/core_45896/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Volym P5_P6_2023']
wz = wb['ZORD']

# Gather all valid from dates per Material as a dict
mat_to_dates = {}
for row in wz.iter_rows(min_row=2, values_only=True):
    mat, valid_from = row[0], row[1]
    if mat is not None:
        if mat not in mat_to_dates:
            mat_to_dates[mat] = []
        if valid_from is not None:
            mat_to_dates[mat].append(valid_from)

# For rows 2-10, column I (9), set joined, formatted dates string
for i in range(2, 11):
    val = ws.cell(row=i, column=1).value
    if val in mat_to_dates:
        # Unique dates, sorted
        dates = sorted(set(mat_to_dates[val]))
        formatted = ','.join([d.strftime('%d/%m/%Y') for d in dates if isinstance(d, datetime)])
        ws.cell(row=i, column=9).value = formatted
    else:
        ws.cell(row=i, column=9).value = ''

wb.save(output_path)
