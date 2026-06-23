import openpyxl
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_4/regression_gate/after_pass/core_41601/input.xlsx')
ws = wb['Students']
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
header_to_col = {h: i+1 for i, h in enumerate(header_row)}
name_col = header_to_col['Name']
age_col = header_to_col['Age']
for row in range(2, 8):
    name_cell = ws.cell(row=row, column=name_col)
    age_cell = ws.cell(row=row, column=age_col)
    if name_cell.value and str(name_cell.value).strip():
        age_cell.value = f"=INDIRECT(\"'\"&A{row}&\"'!C2\")"
    else:
        age_cell.value = None
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_4/regression_gate/after_pass/core_41601/output.xlsx')
