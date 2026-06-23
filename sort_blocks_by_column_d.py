import openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell
import re

# User input/output
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_191-40_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_191-40_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path, data_only=False)
sheet = wb['Sheet1']

# For formatting
bold_center = NamedStyle(name="bold_center")
bold_center.font = Font(bold=True)
bold_center.alignment = Alignment(horizontal="center", vertical="center")

num_format_1d = '#,##0.0'
num_format_int = '#,##0'

# Helper: Check if a row is blank (all A–H empty or whitespace)
def is_blank_row(row):
    return all((cell.value is None or str(cell.value).strip() == '') for cell in row)

# Helper: Check if cell contains a formula as text (starts with '=' and is string)
def is_formula_as_text(cell):
    return isinstance(cell.value, str) and cell.value.strip().startswith('=')

# Gather rows into blocks
max_row = 85
max_col = 8
rows = [list(sheet.iter_rows(min_row=i, max_row=i, max_col=max_col)) for i in range(1, max_row+1)]
blocks = []
block = []
for row in rows:
    cells = [cell for cell in row[0]]
    if is_blank_row(cells):
        if block:
            blocks.append(block)
            block = []
    else:
        block.append(cells)
if block:
    blocks.append(block)

# For each block, process:
processed_blocks = []
for block in blocks:
    block_rows = []
    for cells in block:
        # Convert formula-as-text in D–F back to formulas
        for col in range(3, 6):
            cell = cells[col]
            if is_formula_as_text(cell):
                cell.value = cell.value.strip()
                cell.data_type = 'f'
        block_rows.append(cells)
    
    # Evaluate column D (numeric sort), data_only mode won't update while writing
    def get_numeric_d(row):
        cell = row[3]
        try:
            val = cell.value
            if cell.data_type == 'f' or (is_formula_as_text(cell)):
                # If formula, cannot evaluate here; just use cached value if present
                return cell.internal_value if hasattr(cell, 'internal_value') else 0
            return float(val) if val is not None else 0
        except Exception:
            return 0
    block_rows_sorted = sorted(block_rows, key=get_numeric_d, reverse=True)
    processed_blocks.append(block_rows_sorted)

# Write back blocks to sheet, preserving one blank row between blocks
row_idx = 1
for block in processed_blocks:
    for cells in block:
        for col_idx, cell in enumerate(cells):
            out_cell = sheet.cell(row=row_idx, column=col_idx+1)
            out_cell.value = cell.value
            # Copy formulas
            if is_formula_as_text(cell):
                out_cell.value = cell.value.strip()
                out_cell.data_type = 'f'
            elif cell.data_type == 'f':
                out_cell.value = cell.value
                out_cell.data_type = 'f'
            # Copy formatting
            if col_idx in [3,4,5]:
                out_cell.style = bold_center
            # Format numbers (numeric or formula)
            if col_idx in [3,4,5]:
                if out_cell.data_type == 'f':
                    out_cell.number_format = num_format_1d
                else:
                    v = out_cell.value
                    if isinstance(v, float):
                        out_cell.number_format = num_format_1d
                    elif isinstance(v, int):
                        out_cell.number_format = num_format_int
            # (Preserve column width at end)
        row_idx += 1
    # Insert single blank row only if not last block
    if processed_blocks.index(block) != len(processed_blocks) - 1:
        for col_idx in range(1, max_col+1):
            sheet.cell(row=row_idx, column=col_idx).value = None
        row_idx += 1

# Recheck D–F bold center, and number formatting
for r in range(1, max_row+1):
    for c in [4,5,6]:
        cell = sheet.cell(row=r, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        v = cell.value
        if isinstance(v, float):
            cell.number_format = num_format_1d
        elif isinstance(v, int):
            cell.number_format = num_format_int
        elif cell.data_type == 'f':
            cell.number_format = num_format_1d

# Preserve column widths
def preserve_column_widths(orig_sheet, mod_sheet):
    for col in range(1, max_col+1):
        letter = get_column_letter(col)
        mod_sheet.column_dimensions[letter].width = orig_sheet.column_dimensions[letter].width
preserve_column_widths(sheet, sheet)

wb.save(output_path)
