import openpyxl

def remove_duplicates_by_column_a(input_path, output_path, sheet_name, start_cell, end_cell):
    # Load input workbook and sheet
    wb = openpyxl.load_workbook(input_path)
    ws = wb[sheet_name]
    
    # Calculate target range
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    start_col = ord(start_cell[0]) - ord('A') + 1
    end_row = int(''.join(filter(str.isdigit, end_cell)))
    end_col = ord(end_cell[0]) - ord('A') + 1

    # Read all rows in the region
    rows = list(ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True))

    # Remove duplicates based on col A (index 0)
    seen = set()
    filtered = []
    for row in rows:
        col_a_val = row[0]
        if col_a_val not in seen:
            seen.add(col_a_val)
            filtered.append(row)
    
    # Pad/truncate to fit exactly the region
    while len(filtered) < (end_row - start_row + 1):
        filtered.append(tuple([''] * (end_col - start_col + 1)))
    filtered = filtered[:end_row - start_row + 1]

    # Create new workbook for output
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = sheet_name

    # Paste filtered rows to correct region
    for idx, row in enumerate(filtered, start=start_row):
        for jdx, value in enumerate(row, start=start_col):
            out_ws.cell(row=idx, column=jdx, value=value)

    # Save output
    out_wb.save(output_path)

remove_duplicates_by_column_a(
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_448-11_tc1/input.xlsx',
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_448-11_tc1/output.xlsx',
    sheet_name='Sheet1',
    start_cell='A2',
    end_cell='J16'
)
