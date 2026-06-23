from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/group_47766/r1/evolve_47766/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/group_47766/r1/evolve_47766/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Total (2)']

# Get year boundaries from header row (row 39, columns K:O i.e., 11-15)
year_starts = [ws.cell(row=39, column=col).value for col in range(11, 16)]  # K39:O39
year_ends = year_starts[1:] + [datetime(2100, 1, 1)]

# Store results here
results = [0, 0, 0, 0, 0]  # for K40:O40

# Go through data rows (row 41 to 58 per previous dump) and sum commissions for '*PE*' and within date range
for row in range(41, 59):
    date_val = ws.cell(row=row, column=6).value  # column F (6)
    agent_val = ws.cell(row=row, column=8).value  # column H (8)
    comm_val = ws.cell(row=row, column=3).value  # column C (3)
    # Only process real rows
    if isinstance(date_val, datetime) and agent_val and 'PE' in str(agent_val):
        # Handle if commission is a number
        try:
            if isinstance(comm_val, str) and comm_val.startswith('='):
                continue  # Skip formula cells
            comm = float(comm_val)
        except:
            continue
        for i, (start, end) in enumerate(zip(year_starts, year_ends)):
            if isinstance(start, datetime) and date_val >= start and date_val < end:
                results[i] += comm
                break
# Write results to K40:O40
for i in range(5):
    ws.cell(row=40, column=11+i, value=results[i])

wb.save(output_path)
print('Done!')
