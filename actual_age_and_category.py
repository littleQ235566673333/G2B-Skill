from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/regression_gate/before_pass/core_32337/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/regression_gate/before_pass/core_32337/output.xlsx'
wb = load_workbook(infile)
ws = wb.active

header_row = 2
# Map column label to column index (1-based)
col_map = {cell.value: cell.column for cell in ws[header_row] if cell.value}
col_age = col_map['age (year)']
col_dob = col_map['DATE OF BIRTH']
col_e_idx = col_map['Expected Result']
col_year_age = col_map['YEAR (age)']
col_category = col_map['CATEGORY']

# Add 'Actual Age' col after age (year)
ws.insert_cols(col_age + 1)
col_actual_age_idx = col_age + 1
col_actual_age = get_column_letter(col_actual_age_idx)

# Style header
ws[f'{col_actual_age}{header_row}'] = 'Actual Age'
pre_fill_src = ws[f'{get_column_letter(col_age)}{header_row}'].fill
# Create a fresh fill if the source has a color
if pre_fill_src.patternType is not None:
    pre_fill = PatternFill(fill_type=pre_fill_src.patternType,
                          fgColor=pre_fill_src.fgColor,
                          bgColor=pre_fill_src.bgColor)
else:
    pre_fill = PatternFill()
ws[f'{col_actual_age}{header_row}'].fill = pre_fill
ws[f'{col_actual_age}{header_row}'].alignment = Alignment(horizontal='center', vertical='top')
ws[f'{col_actual_age}{header_row}'].font = Font(bold=True)

# Compute lookup range for XLOOKUP for categories
lookup_row1 = header_row + 1
lookup_row2 = lookup_row1
while ws[f'{get_column_letter(col_year_age)}{lookup_row2}'].value is not None:
    lookup_row2 += 1
lookup_row2 -= 1

for r in range(3, 16):
    # Formula for Actual Age
    dob_cell = f'{get_column_letter(col_dob)}{r}'
    formula = f'=INT(($B$1 - {dob_cell})/365.25)'
    cell = ws[f'{col_actual_age}{r}']
    cell.value = formula
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = pre_fill

    # Formula for CATEGORY (Expected Result column)
    age_cell = f'{get_column_letter(col_age)}{r}'
    xlookup = (
        f'=XLOOKUP({age_cell},'
        f'${get_column_letter(col_year_age)}${lookup_row1}:${get_column_letter(col_year_age)}${lookup_row2},'
        f'${get_column_letter(col_category)}${lookup_row1}:${get_column_letter(col_category)}${lookup_row2},"")'
    )
    ws[f'{get_column_letter(col_e_idx)}{r}'] = xlookup

wb.save(outfile)
