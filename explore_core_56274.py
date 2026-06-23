from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_5/regression_gate/before_fix/core_56274/input.xlsx'
wb = load_workbook(input_path)

# Print sheet names
print('Sheetnames:', wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f'--- Sheet: {sheetname} ---')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20)):
        values = [repr(cell.value) for cell in row]
        print(f'Row {i+1}:', values)
    print()