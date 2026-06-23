import openpyxl
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_262-17_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_262-17_tc1/output.xlsx'

# Load workbook and active sheet (Sheet1)
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read all data inside Sheet1 into a list
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
data = list(rows[1:])

# Dynamically get columns for 'Task' and 'Responsibility'
task_col_idx = header.index('Task')
resp_col_idx = header.index('Responsibility')

# Sort: first by Task, then by Responsibility (both ascending)
sorted_data = sorted(data, key=lambda row: (row[task_col_idx], row[resp_col_idx]))

# Clear Sheet1
for row in ws['A1:F14']:
    for cell in row:
        cell.value = None

# Write header
for c, value in enumerate(header[:6], 1):
    ws.cell(row=1, column=c, value=value)
# Write sorted data (up to 13 rows)
for r, row in enumerate(sorted_data[:13], 2):
    for c, value in enumerate(row[:6], 1):
        ws.cell(row=r, column=c, value=value)

wb.save(output_path)
