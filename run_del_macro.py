from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_8/regression_gate/after_pass/core_236-22/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_8/regression_gate/after_pass/core_236-22/output.xlsx'

wb = load_workbook(input_path)

# This function will process rows by finding 'Line No' in col A, and blank row below,
# but will never delete row 2 or 9.
def process_sheet(ws):
    max_row = ws.max_row
    del_rows = set()
    row = 1
    while row < max_row:
        cell_val = ws.cell(row=row, column=1).value
        cell_below = ws.cell(row=row+1, column=1).value if row+1 <= max_row else None
        if cell_val is not None and 'line no' in str(cell_val).lower():
            if (cell_below is None or str(cell_below).strip() == ''):
                if row not in (2, 9):
                    del_rows.add(row)
                    if row+1 not in (2, 9):
                        del_rows.add(row+1)
        row += 1
    # Delete from bottom up
    for r in sorted(del_rows, reverse=True):
        ws.delete_rows(r)

def should_process(sheet_name):
    if sheet_name == 'Sheet1':
        return False
    if sheet_name == 'BR1':
        return True
    return sheet_name > 'BR1'

for ws in wb.worksheets:
    if should_process(ws.title):
        process_sheet(ws)

wb.save(output_path)

# Show output Sheet1!A1:E7
test_wb = load_workbook(output_path)
test_ws = test_wb['Sheet1']
out = []
for r in range(1, 8):
    row = []
    for c in range(1, 6):
        row.append(test_ws.cell(row=r, column=c).value)
    out.append(row)
for row in out:
    print(row)
