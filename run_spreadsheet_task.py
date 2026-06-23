from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_1/group_38985/r1/evolve_38985/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_1/group_38985/r1/evolve_38985/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Gather lookup names from Table 1
lookup_names = [ws[f'C{row}'].value for row in range(7, 12)]

# Gather table 2 as list of dicts
table2 = []
for row in range(7, 12):
    name = ws[f'L{row}'].value
    v1 = ws[f'M{row}'].value
    v2 = ws[f'N{row}'].value
    table2.append({'name': name, 'val1': v1, 'val2': v2})

# For each name in Table 1, collect all matching pairs from Table 2
for idx, name in enumerate(lookup_names):
    matches = [f"{r['val1']},{r['val2']}" for r in table2 if r['name'] == name]
    cell_value = '; '.join(matches) if matches else ''
    ws[f'D{8+idx}'] = cell_value

wb.save(output_path)
