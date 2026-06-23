from openpyxl import load_workbook

def extract_text_after_second_space_before_dash(s):
    """Extract the substring after the second space and before the dash ('-')."""
    if not isinstance(s, str):
        return ''
    # Find second space
    parts = s.split(' ')
    if len(parts) < 3:
        return ''
    after_second = ' '.join(parts[2:])
    dash = after_second.find('-')
    if dash == -1:
        return after_second.strip()
    return after_second[:dash].strip()

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_5/group_40478/r2/evolve_40478/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_5/group_40478/r2/evolve_40478/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Apply the extraction to rows 1-3 (B1:B3)
for row in range(1, 4):
    val = ws[f'A{row}'].value
    result = extract_text_after_second_space_before_dash(val)
    ws[f'B{row}'] = result

wb.save(output_path)
