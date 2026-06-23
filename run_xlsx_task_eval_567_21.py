import openpyxl
import datetime
import collections
from openpyxl import load_workbook

input_file = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY/eval_567-21_tc1/input.xlsx'
output_file = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY/eval_567-21_tc1/output.xlsx'

wb = load_workbook(input_file)
ws = wb['Sheet1']

# The 'header row' is not present, data starts at row 3 (per previous output)
start_data_row = 3
max_data_row = ws.max_row
rows = list(ws.iter_rows(min_row=start_data_row, max_row=max_data_row, max_col=7, values_only=True))

def parse_date(val):
    try:
        if isinstance(val, str) and '/' in val and len(val.strip())>=7:
            return datetime.datetime.strptime(val[:7], '%Y/%m').date()
        elif isinstance(val, datetime.datetime):
            return val.date()
        elif isinstance(val, str) and '-' in val:
            # Support yyyy-mm-dd
            return datetime.datetime.strptime(val[:10], '%Y-%m-%d').date()
    except Exception:
        return None
    return None

data = []
for row in rows:
    if not any(row): continue
    a = row[0]  # Col A
    b = row[1]  # Col B
    e = row[4]  # Col E
    parsed_e = parse_date(e)
    if e in (None, '', '-') or parsed_e is None:
        continue  # skip rows without valid date
    data.append((a, b, parsed_e, row))

groups = collections.defaultdict(list)
for a, b, e_parsed, r in data:
    groups[(a, b)].append((e_parsed, r))

final_rows = []
for k, vals in groups.items():
    max_date = max(e for e, _ in vals)
    max_rows = [r for e, r in vals if e == max_date]
    final_rows.extend(max_rows)

output_block = final_rows[:25]  # Max to fit in G28 (including start row)

wb_out = load_workbook(input_file)
ws_out = wb_out['Sheet1']
start_row, start_col = 3, 1
for r_idx, row in enumerate(output_block):
    for c_idx, val in enumerate(row[:7]):
        ws_out.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
wb_out.save(output_file)
