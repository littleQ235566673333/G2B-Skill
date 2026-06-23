import openpyxl

# Load the workbook and worksheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_16511_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_16511_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read album names from B2:B10
albums = [ws[f'B{i}'].value for i in range(2, 11)]

track_numbers = []
album_counts = {}
for album in albums:
    if album not in album_counts:
        album_counts[album] = 1
    else:
        album_counts[album] += 1
    track_numbers.append(album_counts[album])

# Write track numbers to F2:F10
for idx, track_num in enumerate(track_numbers, start=2):
    ws[f'F{idx}'] = track_num

# Save result
wb.save(output_path)
