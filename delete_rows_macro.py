from openpyxl import load_workbook
infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_7/regression_gate/after_pass/core_236-22/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_7/regression_gate/after_pass/core_236-22/output.xlsx'
wb = load_workbook(infile)
sheets = wb.sheetnames
for sheet in sheets[wb.sheetnames.index('BR1'):] if 'BR1' in wb.sheetnames else wb.sheetnames:
    ws = wb[sheet]
    rows_to_delete = set()
    max_row = ws.max_row
    col_a = [ws.cell(row=i, column=1).value for i in range(1, max_row+1)]
    n = len(col_a)
    i = 0
    while i < n - 1:
        # Check the "Line No" pattern and below blank
        if col_a[i] == 'Line No' and (col_a[i+1] is None or str(col_a[i+1]).strip() == ''):
            # Check not to delete protected rows (i+2 == 2 or 9, i.e., 2-based rows 2 and 9)
            if (i+1) != 1 and (i+2) != 9:
                rows_to_delete.add(i)
                rows_to_delete.add(i+1)
            i += 2
        else:
            i += 1
    # Delete rows from bottom to top to avoid shifting
    for idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(idx+1)
wb.save(outfile)
