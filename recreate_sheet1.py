from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

inp_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY/eval_58032_tc1/input.xlsx"
out_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY/eval_58032_tc1/output.xlsx"

wb = load_workbook(inp_path)
assert "Sheet1" in wb.sheetnames, "Sheet1 not found"
ws = wb["Sheet1"]

max_row = 35  # Output for A2:A35

# Map headers to index
headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
header_map = {h.lower(): i+1 for i, h in enumerate(headers)}

# Style setup
arial_font = Font(name="Arial")
bold_italic_font = Font(name="Arial", bold=True, italic=True)
orange_fill = PatternFill(fill_type="solid", fgColor="FCD5B4")
gray_fill = PatternFill(fill_type="solid", fgColor="CCCCCC")
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header row styling
for col in range(1, ws.max_column+1):
    cell = ws.cell(row=1, column=col)
    cell.font = bold_italic_font
    cell.border = border

# All cells: Arial font
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            cell.font = arial_font

# Fill D2:D35 with gray
for i in range(2, max_row+1):
    ws[f"D{i}"].fill = gray_fill

# Formula for A2:A35, highlighted with orange
for i in range(2, max_row+1):
    cell = ws[f"A{i}"]
    # This formula searches for a match in I:J with CONCAT of Title and Department code
    cell.value = "=IFERROR(INDEX($I$2:$I$35, MATCH(CONCAT(B{0},C{0}), $J$2:$J$35 & $K$2:$K$35, 0)), \"\")".format(i)
    cell.fill = orange_fill
    cell.font = arial_font

# Hide gridlines where possible
if hasattr(ws, "sheet_view"):
    ws.sheet_view.showGridLines = False

wb.save(out_path)
