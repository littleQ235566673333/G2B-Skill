import openpyxl
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_41978_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_41978_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Extract years and open cases from specified ranges
open_cases = defaultdict(int)
years = []
for row in ws.iter_rows(min_row=14, max_row=185, min_col=7, max_col=7):
    for cell in row:
        year_val = cell.value
        years.append(year_val)

status_list = []
for row in ws.iter_rows(min_row=14, max_row=185, min_col=10, max_col=10):
    for cell in row:
        status_list.append(str(cell.value).strip().lower() if cell.value else '')

# Compose a year->open cases mapping
for y, s in zip(years, status_list):
    if y and s == 'open':
        open_cases[y] += 1

# Aggregate years (sorted, unique, not None)
years_sorted = sorted(set(y for y in years if y))

# Write results to I2:I11 and retain formatting
for i, y in enumerate(years_sorted[:10]):
    cell = ws[f'I{i+2}']
    cell.value = open_cases.get(y, 0)

wb.save(output_path)
