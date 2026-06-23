import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_36277/r1/evolve_36277/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_36277/r1/evolve_36277/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

max_col = ws.max_column
max_row = ws.max_row

from openpyxl.utils import get_column_letter
start_col_letter = 'A'
end_col_letter = get_column_letter(max_col)

for row in range(2, 6):
    formula = f'=INDEX($A$2:${end_col_letter}${max_row},ROW()-1,MATCH($H{row},$A$1:${end_col_letter}$1,0))'
    ws[f'I{row}'] = formula

wb.save(output_path)
