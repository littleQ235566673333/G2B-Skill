import openpyxl

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_472-15_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_472-15_tc1/output.xlsx"

# Load the input workbook and select the active worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read value from A1
value = ws['A1'].value

# Map values according to the user's rules
mapping = {
    '4Ozark': 1,
    '3Tall': 2,
    '1Jasper': 3,
    '2GWood': 4,
    '5Dawson': 5,
    '8CPark': 6
}

result = mapping.get(value, None)
if result is not None:
    ws['B2'] = result
else:
    ws['B2'] = ''  # Or leave it blank if there's no match

# Save to the output file
wb.save(output_path)
