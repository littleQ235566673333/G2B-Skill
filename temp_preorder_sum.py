from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_2/regression_gate/before_pass/core_50796/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_2/regression_gate/before_pass/core_50796/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# For rows 2-5 in column A, lookup label; Output sum in B2-B5
for output_row in range(2, 6):
    criteria = ws[f'A{output_row}'].value
    total = 0
    for row in range(12, 23):
        if ws[f'B{row}'].value == criteria:
            val = ws[f'C{row}'].value
            try:
                total += float(val) if val is not None else 0
            except Exception:
                pass
    ws[f'B{output_row}'] = total

wb.save(output_path)
