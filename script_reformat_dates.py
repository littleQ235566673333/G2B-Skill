import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_486-17_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_486-17_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Blad1']

for i in range(2, 131):
    cell = ws.cell(row=i, column=1).value
    if cell is None:
        ws.cell(row=i, column=2, value='')
        continue
    s = str(cell)
    if len(s) == 9 and s.startswith('0'):
        yyyy = s[1:5]
        mm = s[5:7]
        dd = s[7:9]
        result = f'{yyyy} {mm} {dd}'
    else:
        result = ''
    ws.cell(row=i, column=2, value=result)

wb.save(output_path)
