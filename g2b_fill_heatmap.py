import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_5/regression_gate/before_fix/core_52305/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_5/regression_gate/before_fix/core_52305/output.xlsx'

# Read main sheet
wb = load_workbook(input_path)
ws = wb.active

# Read with pandas for easy filtering
# Clean headers
cols = [str(ws.cell(row=1, column=i+1).value).strip().replace(' ', '') for i in range(ws.max_column)]
data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4, values_only=True):
    if all(x is None for x in row): continue
    data.append(row)
df = pd.DataFrame(data, columns=cols[:4])

def extract_time_only(dt):
    if pd.isnull(dt): return None
    if isinstance(dt, datetime): return dt.time()
    if isinstance(dt, str):
        try:
            return datetime.strptime(dt, '%H:%M:%S').time()
        except:
            pass
        try:
            return datetime.strptime(dt, '%H:%M').time()
        except:
            return None
    return None

df['TimeOnly'] = df['Time'].apply(extract_time_only)

# Read heatmap axes from J5:N5 (columns), I6:I24 (rows)
dest_headers = [ws.cell(row=5, column=j).value for j in range(10, 15)]
name_headers = [ws.cell(row=i, column=9).value for i in range(6, 25)]

start_time = datetime.strptime('21:30', '%H:%M').time()
end_time = datetime.strptime('22:00', '%H:%M').time()

for row_idx, name in enumerate(name_headers, start=6):
    if not name: continue
    for col_off, dest in enumerate(dest_headers):
        if not dest: continue
        count = df[(df['Name'] == name) & (df['Destination'] == dest) & \
                 (df['TimeOnly'] >= start_time) & (df['TimeOnly'] <= end_time)].shape[0]
        ws.cell(row=row_idx, column=10 + col_off).value = count

wb.save(output_path)
print('Heatmap filled.')
