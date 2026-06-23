import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_8/task_44017/r3/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_8/task_44017/r3/evolve_44017/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Columns for output (AD=30, AO=41)
col_start = 30  # AD
col_end = 41   # AO
row_date = 9   # Row containing dates
row_start = 14 # First customer row
row_end = 42   # Last customer row

# Columns for inputs
col_base_rate = 23 # W
col_effective_date = 12 # L
col_frequency = 10     # J
col_increase_start = 13 # M
col_increase_end = 16   # P

# Read the dates for each output column
output_dates = []
for col in range(col_start, col_end+1):
    cell_val = ws.cell(row=row_date, column=col).value
    if isinstance(cell_val, datetime):
        output_dates.append(cell_val)
    else:
        # Try to parse if it's a string
        if isinstance(cell_val, str):
            try:
                output_dates.append(datetime.strptime(cell_val, '%Y-%m-%d'))
            except:
                try:
                    output_dates.append(datetime.strptime(cell_val, '%m/%d/%Y'))
                except:
                    output_dates.append(None)
        else:
            output_dates.append(None)

for row in range(row_start, row_end+1):
    base_rate = ws.cell(row=row, column=col_base_rate).value
    effective_date_cell = ws.cell(row=row, column=col_effective_date).value
    freq = ws.cell(row=row, column=col_frequency).value
    # Try to parse effective date
    if isinstance(effective_date_cell, datetime):
        effective_date = effective_date_cell
    elif isinstance(effective_date_cell, str):
        try:
            effective_date = datetime.strptime(effective_date_cell, '%Y-%m-%d')
        except:
            try:
                effective_date = datetime.strptime(effective_date_cell, '%m/%d/%Y')
            except:
                effective_date = None
    else:
        effective_date = None
    
    # Get increases
    increases = []
    for inc_col in range(col_increase_start, col_increase_end+1):
        inc = ws.cell(row=row, column=inc_col).value
        if inc is not None:
            try:
                inc_float = float(inc)
            except:
                inc_float = 0.0
        else:
            inc_float = 0.0
        increases.append(inc_float)
    
    # Compute for each column
    for idx, col in enumerate(range(col_start, col_end+1)):
        col_date = output_dates[idx]
        cell = ws.cell(row=row, column=col)
        if col_date is None or effective_date is None or base_rate is None:
            cell.value = None
            continue
        if col_date < effective_date:
            cell.value = None
            continue
        # Calculate how many waves of increases have happened by this date
        months_since_effective = (col_date.year - effective_date.year)*12 + (col_date.month - effective_date.month)
        waves = 0
        if freq is not None:
            try:
                freq_int = int(freq)
                if freq_int > 0:
                    waves = months_since_effective // freq_int
            except:
                waves = 0
        # Apply up to four increases (but only those that have taken effect)
        applied_increases = increases[:min(waves, len(increases))]
        new_rate = base_rate
        for pct in applied_increases:
            new_rate *= (1 + pct/100.0)
        cell.value = round(new_rate, 2)
        # No fill applied (no yellow fill requested)

wb.save(output_path)
