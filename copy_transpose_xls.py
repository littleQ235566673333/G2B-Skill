import openpyxl

# Define paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/group_290-1/r1/evolve_290-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/group_290-1/r1/evolve_290-1/output.xlsx'

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Manually define the column correspondence for K:U
columns_src = list('ABCDEFGHIJ')  # Columns A-J
columns_dst = list('KLMNOPQRSTU')  # Columns K-U

for row in range(2, 11):  # Rows 2 to 10 inclusive
    for i, (src, dst) in enumerate(zip(columns_src, columns_dst)):
        src_cell = f'{src}{row}'
        dst_cell = f'{dst}{row}'
        ws[dst_cell].value = ws[src_cell].value  # Copy value
        ws[dst_cell]._style = ws[src_cell]._style  # Copy formatting (limited but retains basic style)

# Save as output.xlsx
wb.save(output_path)
