import openpyxl
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/group_59129/r1/evolve_59129/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/group_59129/r1/evolve_59129/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Get hire and term dates (A2:A22 and B2:B22)
hire_dates = [ws[f'A{row}'].value for row in range(2, 23)]
term_dates = [ws[f'B{row}'].value for row in range(2, 23)]

# Get the month values in E1:P1
months = [ws.cell(row=1, column=col).value for col in range(5, 17)]

# Helper to convert header to first of the month (if string)
def parse_month(m):
    if isinstance(m, datetime):
        return datetime(m.year, m.month, 1)
    elif isinstance(m, str):
        for fmt in ('%b-%Y', '%b %Y', '%Y-%m', '%m/%Y', '%Y/%m'):
            try:
                dt = datetime.strptime(m, fmt)
                return datetime(dt.year, dt.month, 1)
            except Exception:
                continue
        # Try just year if present
        try:
            dt = datetime.strptime(m, '%Y')
            return datetime(dt.year, 1, 1)
        except:
            return None
    return None

parsed_months = [parse_month(m) for m in months]

# Count active employees for each month (E2:P2)
headcounts = []
for ref_month in parsed_months:
    if ref_month is None:
        headcounts.append('')
        continue
    count = 0
    for h, t in zip(hire_dates, term_dates):
        # Must be hired on or before this month
        if isinstance(h, datetime) and h <= ref_month:
            # If terminated, termination must be after the month
            if t is None or (isinstance(t, datetime) and t > ref_month):
                count += 1
    headcounts.append(count)

# Write headcounts to E2:P2
for i, val in enumerate(headcounts):
    ws.cell(row=2, column=5+i, value=val)

wb.save(output_path)
