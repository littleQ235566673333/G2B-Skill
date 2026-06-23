import openpyxl
from openpyxl import load_workbook
import pandas as pd
import numpy as np

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_2/group_387-16/r0/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_2/group_387-16/r0/evolve_387-16/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']
df = pd.read_excel(input_path, sheet_name='Sheet1', header=None)
colnames = df.iloc[0].tolist()
data = df.iloc[1:].reset_index(drop=True)
data.columns = colnames

value_col = 'Value' if 'Value' in data.columns else colnames[1]
bin_col = 'Binaries' if 'Binaries' in data.columns else colnames[2]
res_col = 'Result values' if 'Result values' in data.columns else colnames[3]
colA = colnames[0]

def remove_one_instance(values, binaries, remove_values):
    values = list(values)
    binaries = list(binaries)
    for rem in remove_values:
        try:
            idx = values.index(rem)
            values.pop(idx)
            binaries.pop(idx)
        except ValueError:
            continue
    return values, binaries

values = list(data[value_col].dropna())
binaries = list(data[bin_col].dropna())
remove_values = list(data[res_col].dropna())

values_new, binaries_new = remove_one_instance(values, binaries, remove_values)

data[value_col] = pd.Series(values_new + ['']*(len(data)-len(values_new)))
data[bin_col] = pd.Series(binaries_new + ['']*(len(data)-len(binaries_new)))

# Solver result as proper scalar
solver_sum = data[colA].dropna()
try:
    solver_sum = float(solver_sum.astype(float).sum())
except Exception:
    try:
        solver_sum = float(solver_sum.sum())
    except Exception:
        solver_sum = ''
if isinstance(solver_sum, float) and (np.isnan(solver_sum) or pd.isnull(solver_sum)):
    solver_sum = ''

# Target value as proper scalar
target_val = None
for cname in colnames:
    if 'target' in str(cname).lower():
        tq = data[cname].dropna()
        if not tq.empty:
            target_val = tq.iloc[0]
            break
if target_val is None:
    possible = ws['D1'].value
    if isinstance(possible, (int, float)):
        target_val = possible
if pd.isnull(target_val) or (isinstance(target_val, float) and np.isnan(target_val)):
    target_val = ''

# Difference
diff = ''
try:
    if target_val != '' and solver_sum != '':
        diff = float(solver_sum) - float(target_val)
except Exception:
    diff = ''

rows_to_write = min(17, len(data))
for r in range(rows_to_write):
    for c in range(4):
        ws.cell(row=2+r, column=1+c).value = data.iloc[r, c]
ws.cell(row=2, column=4).value = target_val
ws.cell(row=3, column=4).value = solver_sum
ws.cell(row=4, column=4).value = diff

wb.save(output_path)
