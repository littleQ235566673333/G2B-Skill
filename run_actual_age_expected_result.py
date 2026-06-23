from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from copy import copy

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_2/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_2/regression_gate/after_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

header = [cell.value for cell in ws[2]]
col_names = {name: idx for idx, name in enumerate(header, 1)}
DOB_col = col_names['DATE OF BIRTH']
YEARAGE_col = col_names['YEAR (age)']
CATEGORY_col = col_names['CATEGORY']
AGEYEAR_col = 15 # 'age (year)'

first_row, last_row = 3, 15

ws.insert_cols(AGEYEAR_col + 1)
actual_age_col = AGEYEAR_col + 1
actual_age_col_letter = get_column_letter(actual_age_col)

# Set header for new column
ws.cell(row=2, column=actual_age_col, value='Actual Age')
# Copy fill from the preceding column
age_header_cell = ws.cell(row=2, column=AGEYEAR_col)
if isinstance(age_header_cell.fill, PatternFill):
    ws.cell(row=2, column=actual_age_col).fill = copy(age_header_cell.fill)
else:
    ws.cell(row=2, column=actual_age_col).fill = PatternFill(fill_type=None)
ws.cell(row=2, column=actual_age_col).alignment = Alignment(vertical='top', horizontal='center')
ws.cell(row=2, column=actual_age_col).font = Font(bold=True)

for row in range(first_row, last_row+1):
    # Write Actual Age formula
    formula = f"=INT((($B$1 - {get_column_letter(DOB_col)}{row})/365.25))"
    cell = ws.cell(row=row, column=actual_age_col, value=formula)
    # Copy fill from 'age (year)'
    source_cell = ws.cell(row=row, column=AGEYEAR_col)
    if isinstance(source_cell.fill, PatternFill):
        cell.fill = copy(source_cell.fill)
    else:
        cell.fill = PatternFill(fill_type=None)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)

for row in range(first_row, last_row+1):
    formula = f'=INDEX(${get_column_letter(CATEGORY_col)}$3:${get_column_letter(CATEGORY_col)}$15, MATCH({actual_age_col_letter}{row}, ${get_column_letter(YEARAGE_col)}$3:${get_column_letter(YEARAGE_col)}$15, 0))'
    ws.cell(row=row, column=5, value=formula)

wb.save(output_path)
