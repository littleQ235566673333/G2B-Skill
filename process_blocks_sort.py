import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import re

def is_blank_row(row):
    return all((cell.value is None or (not isinstance(cell.value, str) or cell.value.strip() == '')) for cell in row)

def is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")

def restore_formula(cell):
    # If formula is stored as text with a ' (e.g. "'=B2+C2"), restore to real formula
    if isinstance(cell.value, str) and cell.value.startswith("'="):
        cell.value = cell.value[1:] # Remove leading '
    return cell

def format_number_cell(cell):
    number_format_int = '#,##0'
    number_format_dec = '#,##0.0'
    if cell.value is not None and not is_formula(cell):
        try:
            val = float(cell.value)
            if val == int(val):
                cell.number_format = number_format_int
                cell.value = int(val)
            else:
                cell.number_format = number_format_dec
                cell.value = round(val, 1)
        except Exception:
            pass
    else:
        cell.number_format = number_format_dec
    return cell

def set_bold_centered(cell):
    cell.font = Font(bold=True, name=cell.font.name if cell.font else None)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    return cell

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_191-40_tc1/input.xlsx')
wsa = wb['Sheet1']

rows = list(wsa.iter_rows(min_row=1, max_row=wsa.max_row, max_col=8))
blocks = []
block_start = 0
for idx, row in enumerate(rows):
    if is_blank_row(row):
        if block_start < idx:
            blocks.append((block_start, idx-1))
        block_start = idx+1
if block_start < len(rows):
    blocks.append((block_start, len(rows)-1))

col_widths = [wsa.column_dimensions[get_column_letter(col+1)].width for col in range(8)]

# We'll write all output to a new set of rows
output_rows = []
for block in blocks:
    start, end = block
    if end < start:
        continue
    block_rows = rows[start:end+1]
    # Clean any text formulas, reset styles
    for row in block_rows:
        for col in (3,4,5):
            cell = row[col]
            restore_formula(cell)
            format_number_cell(cell)
            set_bold_centered(cell)
    # Sort block rows by evaluated column D
    def block_sortkey(r):
        val = r[3].value
        if val is None:
            return float('-inf')
        try:
            return float(val)
        except:
            return float('-inf')
    block_rows_sorted = sorted(block_rows, key=block_sortkey, reverse=True)
    output_rows.extend(block_rows_sorted)
    output_rows.append([None]*8) # blank row

# Limit to 85 rows
output_rows = output_rows[:85]

# Apply output to worksheet
for i in range(85):
    for j in range(8):
        cell = wsa.cell(row=i+1, column=j+1)
        val = None
        font = None
        align = None
        numfmt = None
        if i < len(output_rows):
            source = output_rows[i][j] if isinstance(output_rows[i], (list, tuple)) else None
            if source:
                val = source.value
                font = source.font
                align = source.alignment
                numfmt = source.number_format
        cell.value = val
        if font is not None:
            cell.font = Font(bold=font.bold, name=font.name)
        if align is not None:
            cell.alignment = Alignment(horizontal=align.horizontal, vertical=align.vertical)
        if numfmt is not None:
            cell.number_format = numfmt
# Restore column widths
for col in range(8):
    if col_widths[col]:
        wsa.column_dimensions[get_column_letter(col+1)].width = col_widths[col]
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_191-40_tc1/output.xlsx')
