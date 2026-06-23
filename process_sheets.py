import openpyxl
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/before_pass/core_84-40/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/before_pass/core_84-40/output.xlsx'

wb = openpyxl.load_workbook(input_path)

# Find the LIST sheet (case insensitive)
list_sheet = None
for sheet in wb.sheetnames:
    if sheet.lower() == 'list':
        list_sheet = wb[sheet]
        break

if list_sheet is None:
    raise Exception('No sheet named list found')

# Clear existing data (everything except header) in columns B-E
for row in range(2, list_sheet.max_row + 1):
    for col in range(2, 6):  # B (2) to E (5)
        list_sheet.cell(row=row, column=col).value = None

sheet_summaries = []
for sheetname in wb.sheetnames:
    if sheetname.lower() == 'list':
        continue
    ws = wb[sheetname]
    # Sum C and D
    col_c_sum = 0
    col_d_sum = 0
    # Assume first row is header, start from row 2
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4, values_only=True):
        c, d = row
        if c is not None and isinstance(c, (int, float)):
            col_c_sum += c
        if d is not None and isinstance(d, (int, float)):
            col_d_sum += d
    sheet_summaries.append((sheetname, col_c_sum, col_d_sum))

# Write to LIST sheet starting at row 2
for i, (sheetname, c_sum, d_sum) in enumerate(sheet_summaries, start=2):
    list_sheet.cell(row=i, column=2, value=sheetname)  # B
    list_sheet.cell(row=i, column=3, value=c_sum)      # C
    list_sheet.cell(row=i, column=4, value=d_sum)      # D
    list_sheet.cell(row=i, column=5, value=c_sum - d_sum)  # E

# Write totals in the row after the last
total_row = len(sheet_summaries) + 2
list_sheet.cell(row=total_row, column=2, value='TOTAL')
list_sheet.cell(row=total_row, column=3, value=sum(s[1] for s in sheet_summaries))
list_sheet.cell(row=total_row, column=4, value=sum(s[2] for s in sheet_summaries))
list_sheet.cell(row=total_row, column=5, value=(sum(s[1] for s in sheet_summaries) - sum(s[2] for s in sheet_summaries)))

wb.save(output_path)
