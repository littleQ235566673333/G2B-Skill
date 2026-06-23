import openpyxl
from openpyxl.styles import PatternFill

# Load workbook
in_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_254-34_tc1/input.xlsx'
out_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_254-34_tc1/output.xlsx'
wb = openpyxl.load_workbook(in_path)

# Delete Sheet3 if it exists
if 'Sheet3' in wb.sheetnames:
    del wb['Sheet3']

# Find values and rows in Before sheet
ws = wb['Before']
values = []
cells = []
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, (int, float)):
            values.append(cell.value)
            cells.append(cell)

# We need all subsets that sum to [994108, 994109, 994110, 994111, 994112]
from itertools import combinations

target_range = range(994108, 994113)
found = False
for r in range(2, min(25, len(values))+1):  # Limit to reasonable size subsets
    for comb in combinations(enumerate(values), r):
        idxs, nums = zip(*comb)
        s = sum(nums)
        if s in target_range:
            # Highlight these cells
            fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            for i in idxs:
                cells[i].fill = fill
            ws['C2'] = s
            found = True
            break
    if found:
        break

wb.save(out_path)
