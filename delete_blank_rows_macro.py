from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/group_157-4/r0/evolve_157-4/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/group_157-4/r0/evolve_157-4/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Check F14:M83 (columns 6-13, rows 14-83)
def is_blank_row(cells):
    return all(cell.value in (None, '') for cell in cells)

rows_to_keep = []
for row_num in range(14, 84):
    f_to_m_cells = [ws.cell(row=row_num, column=col) for col in range(6, 14)]
    if not is_blank_row(f_to_m_cells):
        # Store E:M (for output)
        em_values = [ws.cell(row=row_num, column=col).value for col in range(5, 14)]
        rows_to_keep.append(em_values)

# Write kept rows into E9:M83
write_row = 9
for vals in rows_to_keep:
    for col_offset, val in enumerate(vals):
        ws.cell(row=write_row, column=5 + col_offset, value=val)
    write_row += 1

# Clear anything left in E9:M83
for r in range(write_row, 84):
    for c in range(5, 14):
        ws.cell(row=r, column=c, value=None)

wb.save(output_path)
