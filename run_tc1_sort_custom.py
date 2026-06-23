import pandas as pd
from openpyxl import load_workbook

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun2/eval_22-47_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun2/eval_22-47_tc1/output.xlsx'

wb = load_workbook(input_file)
sheetname = wb.sheetnames[0]
ws = wb[sheetname]

# Read all data into DataFrame
data = list(ws.values)
df = pd.DataFrame(data)

# Clean headers and rows
df = df.dropna(how='all').reset_index(drop=True)
headers = df.iloc[0]
df = df[1:]
df.columns = headers
df = df.reset_index(drop=True)

# Define relevant columns
col_B = 'B' if 'B' in df.columns else df.columns[1]
col_C = 'C' if 'C' in df.columns else df.columns[2]
col_J = 'J' if 'J' in df.columns else df.columns[9] if len(df.columns)>9 else None
col_F = 'F' if 'F' in df.columns else df.columns[5] if len(df.columns)>5 else None
col_G = 'G' if 'G' in df.columns else df.columns[6] if len(df.columns)>6 else None
col_H = 'H' if 'H' in df.columns else df.columns[7] if len(df.columns)>7 else None

# Remove duplicates, empty, and header-likes in B,C
used = set()
def row_key(row):
    return (row[col_B], row[col_C])
def headerish(val):
    if val is None:
        return True
    if isinstance(val, str) and (val.strip().lower() in ['name', 'header', '', None]):
        return True
    return False
filtered = df[~df[[col_B, col_C]].isnull().any(axis=1)]
rows = []
for i, row in filtered.iterrows():
    k = row_key(row)
    if k in used:
        continue
    if headerish(row[col_B]) or headerish(row[col_C]):
        continue
    rows.append(row)
    used.add(k)

# Pull J names in written order without blanks/duplicates
j_names = []
if col_J:
    for val in df[col_J]:
        if pd.isna(val) or headerish(val) or val in j_names:
            continue
        j_names.append(val)

# Each row: sort order. For each name in J, group rows.
sorted_rows = []
if j_names:
    # Grouped by J order then rest
    for name in j_names:
        for row in rows:
            if row[col_B] == name:
                sorted_rows.append(row)
    # Now rest of rows (not in j_names)
    for row in rows:
        if row[col_B] not in j_names:
            sorted_rows.append(row)
else:
    # Sort alphabetically on col_B
    sorted_rows = sorted(rows, key=lambda r: str(r[col_B]))

# Output range: F2:H10 (so up to 9 rows, skip if fewer)
wb2 = load_workbook(input_file)
ws2 = wb2[sheetname]
# Write to F2:H10. Fill F, G, H cols from source cols (see if named or by order)
for i in range(9):
    excel_row = 2 + i
    if i >= len(sorted_rows):
        ws2[f'F{excel_row}'].value = None
        ws2[f'G{excel_row}'].value = None
        ws2[f'H{excel_row}'].value = None
    else:
        row = sorted_rows[i]
        ws2[f'F{excel_row}'].value = row.get(col_F, None) if col_F else None
        ws2[f'G{excel_row}'].value = row.get(col_G, None) if col_G else None
        ws2[f'H{excel_row}'].value = row.get(col_H, None) if col_H else None

# Now sort only column H lowest to highest within written F2:H10
output_rows = []
for i in range(9):
    excel_row = 2 + i
    vals = (
        ws2[f'F{excel_row}'].value,
        ws2[f'G{excel_row}'].value,
        ws2[f'H{excel_row}'].value
    )
    if any(v is not None for v in vals):
        output_rows.append(vals)
output_rows_sorted = sorted(output_rows, key=lambda x: (x[2] if x[2] is not None else float('inf')))
for i, vals in enumerate(output_rows_sorted):
    excel_row = 2 + i
    ws2[f'F{excel_row}'].value, ws2[f'G{excel_row}'].value, ws2[f'H{excel_row}'].value = vals

wb2.save(output_file)
