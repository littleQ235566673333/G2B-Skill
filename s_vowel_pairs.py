from openpyxl import load_workbook

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_7/regression_gate/after_fix/core_146-49/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_7/regression_gate/after_fix/core_146-49/output.xlsx'

vowels = 'AEIOU'
def get_words(sheet):
    words = set()
    for row in sheet.iter_rows(values_only=True):
        for val in row:
            if isinstance(val, str) and val.strip():
                words.add(val.strip())
    return words

def generate_candidates(word):
    if 'S' not in word:
        return []
    results = []
    for v in vowels:
        candidate = ''.join(v if c == 'S' else c for c in word)
        if candidate != word:
            results.append(candidate)
    return results

def s_vowel_pairs(words):
    pairs = set()
    wordlist = list(words)
    for w in wordlist:
        candidates = generate_candidates(w)
        for cand in candidates:
            if cand in words:
                # To avoid duplicate directional pairs
                pair = tuple(sorted((w, cand), key=lambda x: x.lower()))
                pairs.add(pair)
    # sort as per spec (case-insensitive)
    return sorted(pairs, key=lambda x: (x[0].lower(), x[1].lower()))

wb = load_workbook(input_file)
ws = wb['Sheet1']
words = get_words(ws)
pairs = s_vowel_pairs(words)

# Write to G1:H ...
for rowi, (w1, w2) in enumerate(pairs):
    ws.cell(row=rowi+1, column=7).value = w1
    ws.cell(row=rowi+1, column=8).value = w2

# No headers. Just pairs.
wb.save(output_file)
