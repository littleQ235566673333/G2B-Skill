from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_1/group_44017/r3/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke4/train/iter_1/group_44017/r3/evolve_44017/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

col_base_rate = 23  # W
col_freq = 10       # J
col_eff_date = 12   # L
col_inc_1 = 13      # M
col_inc_2 = 14      # N
col_inc_3 = 15      # O
col_inc_4 = 16      # P

min_row = 14
max_row = 42
col_start = 30  # AD
col_end = 41    # AO
header_row = 9

# Get month dates cell references AD9:AO9 (row 9)
month_date_coords = [get_column_letter(col) + str(header_row) for col in range(col_start, col_end+1)]

for row in range(min_row, max_row+1):
    base_cell = get_column_letter(col_base_rate) + str(row)
    freq_cell = get_column_letter(col_freq) + str(row)
    eff_cell = get_column_letter(col_eff_date) + str(row)
    inc_cells = [get_column_letter(c) + str(row) for c in range(col_inc_1, col_inc_4+1)]
    for idx, col in enumerate(range(col_start, col_end+1)):
        target_cell = ws.cell(row=row, column=col)
        month_cell = get_column_letter(col) + str(header_row)
        # The formula uses EDATE for rolling increases and only applies when date >= effective date
        factors = []
        # There can be up to 4 increases, from columns M through P.
        for i in range(4):
            factors.append(f"IF({month_cell}>=EDATE(${eff_cell},${freq_cell}*{i}),1+${inc_cells[i]}/100,1)")
        # Show blank if before effective date; otherwise calculate new rate
        formula = f"=IF({month_cell}<${eff_cell},\"\",${base_cell}*" + '*'.join(factors) + ")"
        target_cell.value = formula
        # Remove yellow fill (if present; just set pattern to none)
        target_cell.fill = PatternFill(fill_type=None)

wb.save(output_path)
