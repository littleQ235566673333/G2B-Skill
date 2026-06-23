from openpyxl import load_workbook
from datetime import datetime
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun2/eval_46240_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun2/eval_46240_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def check_output(val):
    if val is None or (isinstance(val, str) and val.strip() == ''):
        return 'No'
    if isinstance(val, str):
        if val.strip().upper() == 'N/A':
            return 'N/A'
        if val.strip().lower() == 'yes':
            return 'Yes'
        # Try regex date check (basic)
        if re.match(r"\d{1,2}/\d{1,2}/\d{2,4}", val.strip()) or re.match(r"\d{4}-\d{1,2}-\d{1,2}", val.strip()):
            return 'Yes'
    if isinstance(val, datetime):
        return 'Yes'
    return 'No'

for row in range(2, 5):  # J2:J4, H2:H4
    h_val = ws[f'H{row}'].value
    ws[f'J{row}'] = check_output(h_val)

wb.save(output_path)
