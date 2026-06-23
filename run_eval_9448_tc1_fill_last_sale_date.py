from openpyxl import load_workbook
import datetime
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r3/eval_9448_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r3/eval_9448_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Data']
header_row = 8
month_cols = list(range(9, 21))  # I to T columns (1-based)

def parse_month_header(header_val):
    if isinstance(header_val, datetime.date):
        return datetime.date(header_val.year, header_val.month, 1)
    if not header_val or not isinstance(header_val, str):
        return None
    # Try 'Apr 2022', 'April 2022', etc.
    match = re.search(r'(\w+)\s*(\d{4})', header_val)
    if match:
        month_name, year = match.groups()
        try:
            month = datetime.datetime.strptime(month_name[:3], '%b').month
        except:
            try:
                month = datetime.datetime.strptime(month_name, '%B').month
            except:
                month = 1
        return datetime.date(int(year), month, 1)
    # Try '2022-04', '2022/04'
    match = re.search(r'(\d{4})[/-](\d{1,2})', header_val)
    if match:
        year, month = match.groups()
        return datetime.date(int(year), int(month), 1)
    return None

for row in range(9, 19):  # U9:U18 is for rows 9 to 18
    last_sold_col = None
    for col in reversed(month_cols):
        val = ws.cell(row=row, column=col).value
        if isinstance(val, (int, float)) and val != 0:
            last_sold_col = col
            break
    if last_sold_col:
        header_val = ws.cell(row=header_row, column=last_sold_col).value
        date_out = parse_month_header(header_val)
        if date_out:
            ws.cell(row=row, column=21).value = date_out  # U column
        else:
            ws.cell(row=row, column=21).value = header_val
    else:
        ws.cell(row=row, column=21).value = None

wb.save(output_path)
