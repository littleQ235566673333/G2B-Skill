import openpyxl

def process_top_animals(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    # Read types from row 1 and numbers from row 2 (A1:H1 and A2:H2)
    types = [ws.cell(row=1, column=i+1).value for i in range(8)]
    nums = [ws.cell(row=2, column=i+1).value for i in range(8)]
    
    # Collect indices and values for 'cat' and 'dog'
    cats = [(i, n) for i, (t, n) in enumerate(zip(types, nums)) if t == 'cat']
    dogs = [(i, n) for i, (t, n) in enumerate(zip(types, nums)) if t == 'dog']
    
    # Sort by value descending for ranking
    cats_sorted = sorted(cats, key=lambda x: x[1], reverse=True)
    dogs_sorted = sorted(dogs, key=lambda x: x[1], reverse=True)
    
    # Assign values to row 4 (A4:H4)
    row4 = ['']*8
    if len(dogs_sorted) > 0:
        row4[dogs_sorted[0][0]] = 'dog1'
    if len(dogs_sorted) > 1:
        row4[dogs_sorted[1][0]] = 'dog2'
    if len(cats_sorted) > 0:
        row4[cats_sorted[0][0]] = 'cat1'
    if len(cats_sorted) > 1:
        row4[cats_sorted[1][0]] = 'cat2'
    
    # Write to row 4
    for i, value in enumerate(row4):
        ws.cell(row=4, column=i+1).value = value
    
    wb.save(output_path)

process_top_animals(
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_51289_tc1/input.xlsx',
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_51289_tc1/output.xlsx'
)
