import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/before_fix/core_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/before_fix/core_4714/output.xlsx'

# Read
df = pd.read_excel(input_path, sheet_name='Sheet2')
window_size = 4
# column names observed in structure printing
key_col = 'Emp Code'
sort_col = 'Month'
metric_col = 'Hours'

# Rolling mean logic
results = []
for key, group in df.groupby(key_col):
    group = group.sort_values(by=sort_col).reset_index(drop=True)
    rm = group[metric_col].rolling(window=window_size, min_periods=window_size)
    group['window_mean'] = rm.mean()
    results.append(group)
df_out = pd.concat(results).sort_index()

def fmt(x):
    if pd.isna(x): return 'n/a'
    if isinstance(x, float) and x.is_integer(): return int(x)
    return round(x, 1)

df_out['window_mean_fmt'] = df_out['window_mean'].apply(fmt)

# Write results only to E2:E25 (rows 2–25, col E)
wb = load_workbook(input_path)
ws = wb['Sheet2']
for i in range(2, 26):
    if i-2 < len(df_out):
        val = df_out.iloc[i-2]['window_mean_fmt']
    else:
        val = ''
    ws.cell(row=i, column=5, value=val)

wb.save(output_path)
