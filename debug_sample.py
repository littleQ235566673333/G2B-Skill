import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_8/task_44017/r2/evolve_44017/input.xlsx')
ws = wb.active

target_rows = range(14, 43)
date_row = 9
col_start = 30  # AD
col_end = 42    # AO (exclusive)
dates = [ws.cell(row=date_row, column=c).value for c in range(col_start, col_end)]
output = []
for col, dt in zip(range(col_start, col_end), dates):
    output.append((get_column_letter(col), dt))
print('Output columns and date formulas:', output)
row = 14
freq = ws.cell(row=row, column=10).value
eff_date = ws.cell(row=row, column=12).value
incrs = [ws.cell(row=row, column=col).value for col in range(13,17)]
baserate = ws.cell(row=row, column=23).value
print('Sample row:', {'freq':freq, 'eff_date':eff_date, 'incrs':incrs, 'baserate':baserate})
