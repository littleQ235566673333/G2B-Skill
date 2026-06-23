import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_1/group_269-44/r0/evolve_269-44/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_1/group_269-44/r0/evolve_269-44/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Step 1: Identify all row indices where 'Chassis' occurs in column A
rows_to_delete_before = set()
for row in range(1, ws.max_row + 1):
    if ws.cell(row=row, column=1).value == 'Chassis':
        # For each occurrence, mark the 4 rows BEFORE (if in bounds)
        for del_row in range(row - 4, row):
            if del_row > 0:
                rows_to_delete_before.add(del_row)

# Step 2: Delete those rows in reverse order (high to low)
for del_row in sorted(rows_to_delete_before, reverse=True):
    ws.delete_rows(del_row)

# Save the updated sheet
wb.save(output_path)
