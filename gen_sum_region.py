import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r2/eval_48365_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r2/eval_48365_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Dashboard']
df = pd.read_excel(input_path, sheet_name='Data')
df_dash = pd.read_excel(input_path, sheet_name='Dashboard', header=None)
selected_regions = []
# Read region choices, up to 3 region slots assumed in rows 9 and further in column C (index 2)
for i in range(9, 12):
    try:
        region = df_dash.iloc[i, 2]
        if pd.notna(region) and str(region).strip() != '' and str(region).strip().upper() != 'CHOOSE':
            selected_regions.append(str(region).strip())
    except Exception:
        pass
# Check also if the word 'All' appears in the available All options, which may start in column Q (index 16) row 0, 1, 2, ...
for i in range(0, 5):
    try:
        all_opt = df_dash.iloc[i, 16] # col Q
        if pd.notna(all_opt) and str(all_opt).strip().upper() == 'ALL':
            selected_regions = ['All']
            break
    except Exception:
        pass
if 'All' in selected_regions:
    _sum = df['M1'].sum()
else:
    # Get product from cell C1 (row 0, col 2)
    product = df_dash.iloc[0, 2] if df_dash.shape[1] > 2 else None
    # Apply boolean criteria
    crit = (df['Product'] == product)
    if selected_regions:
        crit = crit & df['Region'].isin(selected_regions)
    _sum = df.loc[crit, 'M1'].sum()
ws['C4'] = _sum
wb.save(output_path)
