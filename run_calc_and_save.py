import openpyxl

# Load workbook and select sheet
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/regression_gate/after_pass/core_3413/input.xlsx')
ws = wb['Sheet1']

def get_cell(row, col):
    return ws.cell(row=row, column=col).value

# Read all department and RU data from A3:C7 (rows 3,4,5,6,7)
source_records = []
for r in range(3, 8):
    dept = get_cell(r, 1)
    ru   = get_cell(r, 2)
    val  = get_cell(r, 3)
    source_records.append({'row': r, 'dept': dept, 'ru': ru, 'val': val})

# Fill results in G3:G6 based on E and F
for rr in range(3, 7):
    target_dept = get_cell(rr, 5)
    target_ru = get_cell(rr, 6)

    # Try to find all rows in A/B matching BOTH dept and RU
    matching = [rec for rec in source_records if rec['dept']==target_dept and rec['ru']==target_ru]
    if matching and target_ru != 'ALL' and target_ru is not None:
        s = sum(rec['val'] for rec in matching if rec['val'] is not None)
    else:
        # If none matching both, or RU is ALL, sum all for dept only regardless of RU
        s = sum(rec['val'] for rec in source_records if rec['dept']==target_dept and rec['val'] is not None)
    ws.cell(row=rr, column=7, value=s)  # col G

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/regression_gate/after_pass/core_3413/output.xlsx')
