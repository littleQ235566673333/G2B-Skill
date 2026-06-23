import openpyxl

# Load the workbook and select first sheet
wb = openpyxl.load_workbook('results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_43589_tc1/input.xlsx')
ws = wb.active

# Read the range string from A2
range_value = ws['A2'].value

# Function to calculate days from 'n to m' (both inclusive)
def days_from_range(text):
    if text is None or 'to' not in text:
        return ''
    try:
        parts = text.split('to')
        start = int(parts[0].strip())
        end = int(parts[1].strip())
        return end - start + 1 if end >= start else ''
    except Exception:
        return ''

# Calculate the answer
result = days_from_range(range_value)

ws['B2'] = result
wb.save('results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_43589_tc1/output.xlsx')
print('done')
