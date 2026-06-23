from openpyxl import load_workbook
from datetime import datetime, timedelta

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/regression_gate/after_pass/core_41589/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/regression_gate/after_pass/core_41589/output.xlsx'

wb = load_workbook(input_file)
ws = wb['Contact List']

# Read values
cell_h = ws['H4'].value
cell_i = ws['I4'].value

today = datetime.today()

def write_action(cell_h, cell_i):
    # Case 1: H4 is a datetime, I4 == YES
    if isinstance(cell_h, datetime):
        days_diff = (today - cell_h).days
        if cell_i == 'YES':
            if days_diff <= 30:
                return 'HOLD'
            else:
                return 'TOUCH BASE'
    # Case 2: H4 blank/not date, I4 == YES
    if cell_i == 'YES':
        return 'NO ACTION'
    # All other cases
    return 'NO ACTION'

ws['J4'] = write_action(cell_h, cell_i)
wb.save(output_file)
