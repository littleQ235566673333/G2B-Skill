import openpyxl
from openpyxl.styles import Border, Side, Font

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_39903_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_39903_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

borders = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
font = Font(name='Courier New', size=9)

for row in range(2, 7):
    cell = ws[f'C{row}']
    text = str(cell.value).strip() if cell.value is not None else ''
    count = 0
    for item in text.split(','):
        location = item.split(':')[0].strip()
        if not location or location[0] in ('X', 'Z'):
            continue
        count += 1
    out_cell = ws[f'C{row}']
    out_cell.value = count
    out_cell.font = font
    out_cell.border = borders

wb.save(output_path)
