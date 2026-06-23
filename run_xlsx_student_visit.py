from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_2/regression_gate/after_pass/core_9726/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_2/regression_gate/after_pass/core_9726/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Table 1: A (Student ID), B (Visit number), C (Visit date) - Rows 2-7
table1 = []
for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=3, values_only=True):
    table1.append(row)

# Table 2: H (Student ID), I (Submission date) - e.g., Row 2 (H2, I2)
table2 = []
for row in ws.iter_rows(min_row=2, max_row=2, min_col=8, max_col=9, values_only=True):
    table2.append(row)

results = []
for student_id2, submission_date in table2:
    # Find all visit dates for this student strictly before submission_date
    valid_dates = [visit_date for student_id1, _, visit_date in table1 if student_id1 == student_id2 and visit_date is not None and visit_date < submission_date]
    if valid_dates:
        result_date = max(valid_dates)
    else:
        result_date = None
    results.append(result_date)
    ws['J2'] = result_date

wb.save(output_path)
