# FIFO per-unit cost calculation for Excel sheet
# Fills M2:M5 with per-unit cost for current inventory

from openpyxl import load_workbook

inp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/group_39432/r1/evolve_39432/input.xlsx'
outp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/group_39432/r1/evolve_39432/output.xlsx'

# Columns info
#  A: Item
#  B: Beginning Inventory
#  C: Beginning Cost
#  D: PO-1 units
#  E: PO-1 cost
#  F: PO-2 units
#  G: PO-2 cost
#  H: PO-3 units
#  I: PO-3 cost
#  J: PO-4 units
#  K: PO-4 cost
#  L: Current Inv (on hand)
#  M: Per Unit Cost (to fill)

wb = load_workbook(inp)
ws = wb['Sheet1']

for row in range(2,6):
    inventory = []
    # Add beginning inventory
    beg_qty = ws.cell(row=row, column=2).value or 0
    beg_cost = ws.cell(row=row, column=3).value or 0
    if beg_qty > 0:
        inventory.append((beg_qty, beg_cost))
    # Add POs as (qty, cost) pairs
    for off in [(4,5),(6,7),(8,9),(10,11)]:
        qty = ws.cell(row=row, column=off[0]).value
        cost = ws.cell(row=row, column=off[1]).value
        if qty:
            inventory.append((qty, cost))
    # Each entry in inventory is (qty, cost per unit) in order acquired
    total_on_hand = ws.cell(row=row, column=12).value or 0
    if total_on_hand == 0:
        ws.cell(row=row, column=13).value = None
        continue
    # FIFO: inventory is used oldest first, so remaining on hand is from latest lots
    # Reverse inventory to newest first, take as much as needed to sum to current inv
    need = total_on_hand
    running_units = []
    for qty, cost in reversed(inventory):
        if need <= 0:
            break
        use = min(qty, need)
        running_units.append((use, cost))
        need -= use
    # Compute weighted average of these units
    numerator = sum(n*c for n,c in running_units)
    denominator = sum(n for n,_ in running_units)
    avg_cost = numerator/denominator if denominator else None
    ws.cell(row=row, column=13).value = round(avg_cost, 4) if avg_cost is not None else None

wb.save(outp)
