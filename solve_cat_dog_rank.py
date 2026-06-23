import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_51289_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_51289_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read tags and numbers
tags = [ws.cell(row=1, column=i).value for i in range(1, 9)]
numbers = [ws.cell(row=2, column=i).value for i in range(1, 9)]

# Find indices for each species
dog_indices = [i for i, tag in enumerate(tags) if tag == 'dog']
cat_indices = [i for i, tag in enumerate(tags) if tag == 'cat']

output = ['']*8

# For 'dog', get numbers and indices for sorting
dogs_with_numbers = sorted([(numbers[i], i) for i in dog_indices], reverse=True)
if len(dogs_with_numbers) > 0:
    output[dogs_with_numbers[0][1]] = 'dog1'
if len(dogs_with_numbers) > 1:
    output[dogs_with_numbers[1][1]] = 'dog2'

# For 'cat', get numbers and indices for sorting
cats_with_numbers = sorted([(numbers[i], i) for i in cat_indices], reverse=True)
if len(cats_with_numbers) > 0:
    output[cats_with_numbers[0][1]] = 'cat1'
if len(cats_with_numbers) > 1:
    output[cats_with_numbers[1][1]] = 'cat2'

# Write to row 4, columns 1-8
for i in range(8):
    ws.cell(row=4, column=i+1, value=output[i])

wb.save(output_path)
