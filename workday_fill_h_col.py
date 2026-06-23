import openpyxl
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_50631_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_50631_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

holidays = [ws[f'J{r}'].value for r in range(4, 20) if isinstance(ws[f'J{r}'].value, datetime.datetime)]

def is_workday(date, holidays):
    return date.weekday() < 5 and date not in holidays

start_date = ws['B3'].value
end_date = ws['F3'].value
rows = range(7, 39)  # H7:H38

for idx, row in enumerate(rows):
    workdays = idx
    date = start_date
    found = 0
    while found < workdays:
        date += datetime.timedelta(days=1)
        if is_workday(date, holidays):
            found += 1
    if date <= end_date:
        ws[f'H{row}'].value = date
    else:
        ws[f'H{row}'].value = None

wb.save(output_path)
