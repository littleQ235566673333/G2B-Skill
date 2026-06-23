import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_4/group_177-6/r3/evolve_177-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_4/group_177-6/r3/evolve_177-6/output.xlsx'

# Load workbook and sheet
wb = load_workbook(input_path)
ws = wb.active

# Use openpyxl to get column headers as Excel sees them
headers = [ws.cell(row=1, column=i+1).value for i in range(ws.max_column)]
df = pd.read_excel(input_path, sheet_name=0, dtype=str)
reference_col_name = headers[7]  # Column H, 'ComboKey'
num_cols = 18  # Columns A:R (0-17)

# Make sure to use correct reference_col
try:
    groups = df.groupby(reference_col_name, sort=False)
except KeyError:
    raise Exception(f"Reference column '{reference_col_name}' not found. Headers: {df.columns}")

rows_to_write = []
formats = []
for value, group in groups:
    rep_row = group.iloc[0].copy()
    # For columns I:R, merge as sum (or keep as-is if not)
    if len(group) > 1:
        for i in range(8, num_cols):
            vals = pd.to_numeric(group.iloc[:, i], errors='coerce')
            rep_row.iloc[i] = vals.sum() if not vals.isna().all() else ''
    rows_to_write.append(rep_row)
    row_idx = group.index[0]+2
    formats.append([ws.cell(row=row_idx, column=i+1)._style for i in range(num_cols)])

# Create/replace combined sheet
if 'combined' in wb.sheetnames:
    wb.remove(wb['combined'])
combined = wb.create_sheet('combined')

# Write header
for col_idx in range(1, num_cols+1):
    cell = combined.cell(row=1, column=col_idx, value=ws.cell(row=1, column=col_idx).value)
    cell._style = ws.cell(row=1, column=col_idx)._style
# Write merged rows
for out_idx, (row, style_row) in enumerate(zip(rows_to_write, formats), start=2):
    for col_idx, val in enumerate(row, start=1):
        cell = combined.cell(row=out_idx, column=col_idx, value=val)
        cell._style = style_row[col_idx-1]
        # For I:R, float with 2 decimals, blank for 0
        if 9 <= col_idx <= 18:
            try:
                v = float(cell.value)
                if v == 0:
                    cell.value = ''
                else:
                    cell.value = '{:.2f}'.format(v).replace(',', '')
            except (ValueError, TypeError):
                pass
wb.save(output_path)
