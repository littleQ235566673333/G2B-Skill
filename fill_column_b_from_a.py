from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_1563_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_1563_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

last_value = None
for row in range(2, 31):
    val = ws.cell(row=row, column=1).value
    if val is not None:
        last_value = val
    ws.cell(row=row, column=2).value = last_value

wb.save(output_path)
print('Done')
