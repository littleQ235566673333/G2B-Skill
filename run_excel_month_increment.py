import openpyxl
from openpyxl.styles import PatternFill

def increment_month(month_str, year_str):
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    month_index = months.index(month_str)
    year = int("20" + year_str) if len(year_str) == 2 else int(year_str)
    if month_index == 11:
        new_month = 'Jan'
        year += 1
    else:
        new_month = months[month_index + 1]
    return f'{new_month} {str(year)[-2:]}'

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_51354_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_51354_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active
fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
for row in range(2, 7):
    cell_value = ws[f'A{row}'].value
    date_str = cell_value[-6:] if cell_value else ''
    if len(date_str.split()) == 2:
        month_str, year_str = date_str.split()
        if month_str in ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'] and year_str.isdigit():
            ws[f'D{row}'] = year_str
            ws[f'D{row}'].fill = fill
            ws[f'E{row}'] = increment_month(month_str, year_str)
        else:
            ws[f'D{row}'] = ''
            ws[f'E{row}'] = ''
    else:
        ws[f'D{row}'] = ''
        ws[f'E{row}'] = ''
wb.save(output_path)
