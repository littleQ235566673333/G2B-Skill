import openpyxl
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_43589_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun1/eval_43589_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

text = ws['A2'].value

days = None
if text:
    m = re.match(r'(\d+)\s*to\s*(\d+)', text)
    if m:
        start, end = map(int, m.groups())
        days = end - start + 1

if days is not None:
    ws['B2'] = days
else:
    ws['B2'] = 'Invalid range'

wb.save(output_path)
