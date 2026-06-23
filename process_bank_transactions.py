import openpyxl
from openpyxl.utils import get_column_letter

# Define a sample mapping from shortened vendor descriptions to categories
CATEGORY_MAP = {
    'amazon.com': 'Shopping',
    'AMZN MKTP': 'Shopping',
    'STARBUCKS': 'Coffee',
    'UBER': 'Transport',
    'WALMART': 'Groceries',
    'SHELL': 'Gas',
    'MCDONALD': 'Food',
    'APPLE.COM': 'Tech',
    # Add/expand as needed
}

def assign_category(desc):
    """
    Match known shortened descriptions to categories.
    Default to 'Other' if not matched.
    """
    desc_upper = str(desc).upper()
    for key in CATEGORY_MAP:
        if key in desc_upper:
            return CATEGORY_MAP[key]
    return 'Other'

# File paths
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_524-31_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_524-31_tc1/output.xlsx'

# Load the workbook and required sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Exp-DB']

# Assume descriptions are in column B (update if different)
desc_col = 2  # column B
cat_col = 5   # column E
for row in range(1, 54):  # E1:E53 (1-based indexing)
    desc = ws.cell(row=row, column=desc_col).value
    cat = assign_category(desc)
    ws.cell(row=row, column=cat_col, value=cat)

wb.save(output_path)
