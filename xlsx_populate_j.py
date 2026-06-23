from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_2/regression_gate/after_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_2/regression_gate/after_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']

row_idx = 4
h = ws[f'H{row_idx}'].value  # Last Contact (date)
i = ws[f'I{row_idx}'].value  # Regular Contact (YES/NO/None)
today = datetime.today()

result = 'NO ACTION'
if isinstance(h, datetime) and i and str(i).strip().upper() == 'YES':
    delta_days = (today - h).days
    if delta_days <= 30:
        result = 'HOLD'
    else:
        result = 'TOUCH BASE'
ws[f'J{row_idx}'] = result
wb.save(output_path)
