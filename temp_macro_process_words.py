import openpyxl
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/group_118-50/r2/evolve_118-50/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/group_118-50/r2/evolve_118-50/output.xlsx'
endings_priority = ['ING','ERS','ATE','EST','ONE','IER','ILY']
wanted_patterns = set(e.upper() for e in endings_priority)
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']
# Read and sort names in column A (skip header)
words = []
for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
    val = row[0]
    if val and isinstance(val, str):
        words.append(val)
words.sort()
def transform_word(word):
    if len(word) < 5:
        return None
    return word[4] + word[:4] + word[5:]
groups = {e: [] for e in endings_priority}
ending_regex = re.compile(r'(.{3})$')
for word in words:
    m = ending_regex.search(word.upper())
    if not m:
        continue
    ending = m.group(1)
    if ending in wanted_patterns and len(word) >= 5:
        transformed = transform_word(word)
        if transformed != word:
            groups[ending].append((transformed, word))
final = []
for end in endings_priority:
    group_sorted = sorted(groups[end], key=lambda t: t[1])
    final.extend(group_sorted)
for i, (trans, orig) in enumerate(final[:4999]):
    ws.cell(row=i+2, column=3, value=trans)
    ws.cell(row=i+2, column=4, value=orig)
wb.save(output_path)
