import openpyxl

def main():
    infile = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_524-31_tc1/input.xlsx'
    outfile = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_524-31_tc1/output.xlsx'
    wb = openpyxl.load_workbook(infile)
    ws = wb['Exp-DB']

    # First, create a list of (vendor_short, category) pairs from columns A and B
    vendor_map = []
    for row in range(1, 54):
        vendor = ws.cell(row=row, column=1).value
        cat = ws.cell(row=row, column=2).value
        if vendor and cat:
            vendor_map.append((str(vendor).lower(), cat))

    # Now, for each transaction description, assign category
    for row in range(1, 54):
        desc = ws.cell(row=row, column=4).value
        category_assigned = None
        if desc:
            desc_lower = str(desc).lower()
            for vendor_short, cat in vendor_map:
                if vendor_short in desc_lower:
                    category_assigned = cat
                    break
        ws.cell(row=row, column=5).value = category_assigned

    wb.save(outfile)

if __name__ == '__main__':
    main()
