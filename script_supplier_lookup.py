import openpyxl

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_55979_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_55979_tc1/output.xlsx"

# Load workbook
wb = openpyxl.load_workbook(input_path)

calc_sheet = wb['Calc']
item_number = calc_sheet['A10'].value

supplier_found = None

for sheet_name in wb.sheetnames:
    if sheet_name == 'Calc':
        continue
    sheet = wb[sheet_name]
    # Search for item_number in this sheet
    for row in sheet.iter_rows(values_only=True):
        if item_number in row:
            # Supplier name is part of the sheet name (e.g., Supplier_1, Supplier_2, etc.)
            # Extract 'Supplier_1' from the sheet name
            if sheet_name.startswith('Supplier_'):
                supplier = sheet_name.split('_')[0] + '_' + sheet_name.split('_')[1]
            else:
                supplier = sheet_name
            supplier_found = supplier
            break
    if supplier_found:
        break

# Write supplier to 'Calc'!B7
calc_sheet['B7'] = supplier_found if supplier_found else 'Not Found'

# Save to output file
wb.save(output_path)
