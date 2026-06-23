from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_4/group_56274/r3/evolve_56274/input.xlsx'
wb = load_workbook(input_path)
ws = wb.active

for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
    if row and any(col is not None and str(col).strip() != '' for col in row):
        print(f"Row {i}: {row}")
