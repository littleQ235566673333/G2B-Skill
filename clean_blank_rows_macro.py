from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_7/regression_gate/before_pass/core_160-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_7/regression_gate/before_pass/core_160-6/output.xlsx'

wb = load_workbook(input_path)
ws = wb['SH']

# Read all data from the relevant range (A6:L11)
data = []
for row in ws.iter_rows(min_row=6, max_row=11, min_col=1, max_col=12, values_only=True):
    data.append(list(row))

# Remove rows that are entirely blank (all None or empty string)
nonblank_rows = [r for r in data if any(cell not in (None, '') for cell in r)]

# Now clear the range in the sheet (A6:L11)
for row_idx in range(6, 12):
    for col_idx in range(1, 13):
        ws.cell(row=row_idx, column=col_idx, value=None)

# Write compacted non-blank rows back to A6
for i, row in enumerate(nonblank_rows):
    for j, val in enumerate(row):
        ws.cell(row=6 + i, column=1 + j, value=val)

wb.save(output_path)
