import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_1/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_1/regression_gate/after_pass/core_32337/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active
header_row_idx = 2
headers = {cell.value: idx+1 for idx, cell in enumerate(ws[header_row_idx]) if cell.value}
age_col = headers.get('age (year)')
expected_result_col = headers.get('Expected Result')
year_col = headers.get('YEAR (age)')
category_col = headers.get('CATEGORY')
dob_col = headers.get('DATE OF BIRTH')
start_row = 3
end_row = 15
actual_age_col = age_col + 1
ws.insert_cols(actual_age_col)
ws.cell(row=header_row_idx, column=actual_age_col).value = 'Actual Age'
# Try to copy fill color from previous column
base_fill = ws.cell(row=start_row, column=actual_age_col-1).fill
fill_args = {'fill_type': base_fill.fill_type}
if hasattr(base_fill.start_color, 'rgb') and isinstance(base_fill.start_color.rgb, str):
    fill_args['start_color'] = base_fill.start_color.rgb
copy_fill = PatternFill(**fill_args)
# Get report date (B1)
report_date_ref = f'$B$1'
for r in range(start_row, end_row + 1):
    dob_letter = get_column_letter(dob_col)
    age_formula = f'=IF(AND(ISNUMBER({report_date_ref}),ISNUMBER({dob_letter}{r})),DAYS({report_date_ref},{dob_letter}{r})/365.25, "")'
    c = ws.cell(row=r, column=actual_age_col)
    c.value = age_formula
    c.number_format = '0'
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center', vertical='bottom')
    c.fill = copy_fill
h = ws.cell(row=header_row_idx, column=actual_age_col)
h.alignment = Alignment(horizontal='center', vertical='top')
h.fill = copy_fill
if year_col and category_col:
    from_col = get_column_letter(category_col)
    lookup_col = get_column_letter(year_col)
    category_range = f'${from_col}${start_row}:${from_col}${end_row}'
    year_range = f'${lookup_col}${start_row}:${lookup_col}${end_row}'
    for r in range(start_row, end_row+1):
        formula = f'=IFERROR(INDEX({category_range},MATCH({lookup_col}{r},{year_range},0)),"")'
        ws.cell(row=r, column=expected_result_col).value = formula
wb.save(output_path)
