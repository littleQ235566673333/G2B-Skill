from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r3/eval_58032_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY_r3/eval_58032_tc1/output.xlsx'

wb = load_workbook(input_path)
assert 'Sheet1' in wb.sheetnames, 'Sheet1 not found'
ws = wb['Sheet1']

# Set Arial font for all cells
for row in ws.iter_rows():
    for cell in row:
        cell.font = Font(name='Arial', size=cell.font.size if cell.font else 11)

# Format header row (bold, italic, bordered)
header = next(ws.iter_rows(min_row=1, max_row=1))
sides = Side(style='thin')
border = Border(left=sides, right=sides, top=sides, bottom=sides)
for cell in header:
    cell.font = Font(name='Arial', bold=True, italic=True)
    cell.border = border

# Fill D2:D35 with #CCCCCC
gray_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
for row in range(2, 36):
    ws[f'D{row}'].fill = gray_fill

# Mapping main and lookup columns
col_title_left = 2  # B
col_dept_left = 3   # C
col_name_right = 8  # H
col_title_right = 9 # I
col_dept_right = 10 # J
B = get_column_letter(col_title_left)
C = get_column_letter(col_dept_left)
H = get_column_letter(col_name_right)
I = get_column_letter(col_title_right)
J = get_column_letter(col_dept_right)

start_row = 2
end_row = ws.max_row
# Formula in A2:A35, matching Title+DeptCode on right table (I+J) and returning Name from H
fill_ans = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')
for row in range(2,36):
    formula = (
        f"=IFERROR(INDEX(${H}${start_row}:${H}${end_row}, "
        f"MATCH(1, ((${I}${start_row}:${I}${end_row}={B}{row})*"
        f"(${J}${start_row}:${J}${end_row}={C}{row})), 0)), \"\")"
    )
    ws[f'A{row}'].value = formula
    ws[f'A{row}'].fill = fill_ans

# Hide gridlines
ws.sheet_view.showGridLines = False

wb.save(output_path)
