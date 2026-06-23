import openpyxl
import re
from openpyxl.utils import get_column_letter

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/after_pass/core_290-27/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/after_pass/core_290-27/output.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Process only cells in B14:B137
for row in range(14, 138):
    cell = ws[f'B{row}']
    value = cell.value
    if value and isinstance(value, str):
        # Match patterns like 'PID1', 'PID2', 'GG 1' (two/three uppercase letters + optional space + number)
        m = re.match(r"^([A-Z]{2,3}) ?([0-9]+)$", value)
        if m:
            cell.value = m.group(2)

# Save the result
wb.save(output_path)
