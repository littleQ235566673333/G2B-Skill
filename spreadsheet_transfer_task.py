import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_6/group_58484/r2/evolve_58484/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_6/group_58484/r2/evolve_58484/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

# Determine data range (5 to 26 for H, but we want to replicate rows 1-26)
header_row = 4
first_data_row = 5
last_data_row = 26

# 1. Style HEADER (bold, borders)
header_font = Font(bold=True)
side = Side(border_style='thin', color='000000')
border = Border(left=side, right=side, top=side, bottom=side)

for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=header_row, column=col)
    cell.font = header_font
    cell.border = border
    # Wrap text for entire row 4
    cell.alignment = Alignment(wrapText=True)

# 2. Clear column G (7) (no borders, no values)
for row in range(1, last_data_row + 1):
    g_cell = ws.cell(row=row, column=7)
    g_cell.value = None
    g_cell.border = Border()

# 3. For rows 5-26, process transfer counting for operator 5551234
operator_number = '5551234'
entity_col = 4 # D
staff_col = 5 # E
output_col = 8 # H

# Collect indices of transfer rows for this operator
transfer_indices = []
for row in range(first_data_row, last_data_row + 1):
    if str(ws.cell(row=row, column=entity_col).value) == operator_number \
        and ws.cell(row=row, column=staff_col).value not in (None, ''):
        transfer_indices.append(row)

# Write count ONLY next to last transfer for this operator in this range
for row in range(first_data_row, last_data_row + 1):
    cell = ws.cell(row=row, column=output_col)
    if transfer_indices and row == transfer_indices[-1]:
        cell.value = len(transfer_indices)
    else:
        cell.value = None

# Ensure headings beyond column G (e.g. H) are bordered/bolded too if present
for col in range(8, ws.max_column + 1):
    cell = ws.cell(row=header_row, column=col)
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(wrapText=True)

wb.save(output_path)
print('Done')
