import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun1/eval_48983_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42_rerun1/eval_48983_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Source categories (C4:I4) are in cols 3 to 9, row 4
source_cat_row = 4
source_cat_cols = list(range(3, 10))
source_cats = [ws.cell(row=source_cat_row, column=col).value for col in source_cat_cols]

# Target/output categories (M4:S4 -> cols 13 to 19)
target_cat_row = 4
target_cat_cols = list(range(13, 20))
target_cats = [ws.cell(row=target_cat_row, column=col).value for col in target_cat_cols]

# Source brands (A6:A20 -> rows 6+)
source_brand_col = 1
source_brand_rows = list(range(6, 21))
source_brands = [ws.cell(row=row, column=source_brand_col).value for row in source_brand_rows]

# Target/output brands (K6:K11 -> rows 6 to 11, col 11)
target_brand_col = 11
target_brand_rows = list(range(6, 12))
target_brands = [ws.cell(row=row, column=target_brand_col).value for row in target_brand_rows]

# Build lookup dictionaries for brands and categories
brand_index = {b: i for i, b in enumerate(source_brands) if b is not None}
cat_index = {c: i for i, c in enumerate(source_cats) if c is not None}

# Now write to output M6:S11
for t_row_i, t_brand in enumerate(target_brands):
    for t_col_j, t_cat in enumerate(target_cats):
        # Find source indices
        if t_brand in brand_index and t_cat in cat_index:
            s_row = source_brand_rows[brand_index[t_brand]]
            s_col = source_cat_cols[cat_index[t_cat]]
            val = ws.cell(row=s_row, column=s_col).value
        else:
            val = None
        # Write to output
        ws.cell(row=6 + t_row_i, column=13 + t_col_j).value = val

wb.save(output_path)
