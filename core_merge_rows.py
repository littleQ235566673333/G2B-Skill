import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy

file_in = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_2/regression_gate/after_fix/core_177-6/input.xlsx'
file_out = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_2/regression_gate/after_fix/core_177-6/output.xlsx'

wb = openpyxl.load_workbook(file_in)
data_ws = wb['DATA']
out_ws = wb['combined']

rows = list(data_ws.iter_rows(min_row=1, max_row=8, max_col=18))
header = rows[0]
data_rows = rows[1:]

groups = {}
for row in data_rows:
    key = row[7].value  # Column H, index 7
    if key not in groups:
        groups[key] = []
    groups[key].append(row)

results = []
results.append(header)
for key, g_rows in groups.items():
    if not g_rows:
        continue
    merged = []
    # columns A-H: take from the first row of group
    for col in range(8):
        merged.append(g_rows[0][col])
    # columns I-R: sum numeric, otherwise non-null, 0s to blanks
    for col in range(8, 18):
        vals = [r[col].value for r in g_rows]
        is_numeric = all((v in [None, '', '\xa0'] or isinstance(v,(int,float))) for v in vals)
        if is_numeric:
            total = sum(float(v) for v in vals if isinstance(v,(int,float)) and v != 0)
            val = total if total != 0 else ''
        else:
            non_blank = [v for v in vals if (isinstance(v,str) and v.strip() and v != '\xa0') or (v not in [None, '', '\xa0'])]
            val = non_blank[0] if non_blank else ''
        merged.append(val)
    results.append(merged)

# Write to output
for r_idx, row in enumerate(results, 1):
    for c_idx, cell in enumerate(row, 1):
        # Value
        val = cell.value if hasattr(cell,'value') else cell
        out_cell = out_ws.cell(row=r_idx, column=c_idx, value=val)
        # Copy style for columns A-H from first in group
        if r_idx>1 and c_idx<=8 and hasattr(cell,'_style'):
            out_cell._style = copy(cell._style)
        # Format for I:R
        if c_idx>=9 and r_idx>1:
            v = out_cell.value
            # Blank if 0 or blank
            if v in [None, '', 0, '0', '0.0', '\xa0']:
                out_cell.value = ''
            else:
                try:
                    v = float(v)
                    out_cell.value = f"{v:.2f}" if v != 0 else ''
                except:
                    pass
            out_cell.number_format = '0.00' if out_cell.value != '' else '@'

wb.save(file_out)
print('Done')
