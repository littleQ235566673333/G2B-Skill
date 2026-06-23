import openpyxl

# Load workbook and worksheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_4/group_56274/r3/evolve_56274/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_4/group_56274/r3/evolve_56274/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read the fiscal month value
fiscal_month = ws['D7'].value

# Try to locate table headers for month, opening, debits, credits, closing
found_table = False
header_row = None
headers = None
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    values = [cell.value for cell in row]
    lower_values = [str(v).lower() if v is not None else '' for v in values]
    if any('fiscal month' in v for v in lower_values):
        headers = values
        header_row = row[0].row
        found_table = True
        break
if not found_table:
    print("Headers detected in sheet (first 10 rows):")
    for i in range(10):
        row = [cell.value for cell in ws[i+1]]
        print(f"Row {i+1}: {row}")
    raise ValueError('Header row containing "Fiscal Month" not found.')

# Print detected headers for diagnosis
print(f"Detected headers: {headers}")

def find_partial(header_list, keywords):
    for idx, h in enumerate(header_list):
        if h is not None:
            val = str(h).lower().replace(' ','')
            for keyword in keywords:
                if keyword in val:
                    return idx
    return None

month_idx = find_partial(headers, ['fiscalmonth','month'])
opening_idx = find_partial(headers, ['opening'])
debits_idx = find_partial(headers, ['debit'])
credits_idx = find_partial(headers, ['credit'])
closing_idx = find_partial(headers, ['closing','bal'])

print('Header indices:', {
    'month_idx': month_idx,
    'opening_idx': opening_idx,
    'debits_idx': debits_idx,
    'credits_idx': credits_idx,
    'closing_idx': closing_idx
})

def get_row_value_safe(row, idx):
    return row[idx].value if idx is not None else None

# Search for the row for the selected fiscal month
result = [None] * 4  # [opening, debits, credits, closing]
for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
    this_month = get_row_value_safe(row, month_idx)
    if str(this_month) == str(fiscal_month):
        result[0] = get_row_value_safe(row, opening_idx)
        result[1] = get_row_value_safe(row, debits_idx)
        result[2] = get_row_value_safe(row, credits_idx)
        result[3] = get_row_value_safe(row, closing_idx)
        break

ws['D9'] = result[0]
ws['D10'] = result[1]
ws['D11'] = result[2]
ws['D12'] = result[3]

wb.save(output_path)
