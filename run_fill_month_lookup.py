import openpyxl

def extract_code_and_lookup_month(text, codes, months):
    """
    Given a text string and a list of codes/months, find the first code present in the text,
    and return the corresponding month. If none found, return None.
    """
    for code, month in zip(codes, months):
        if code in text:
            return month
    return None

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_15380_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_15380_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet3']

# Extract codes and months from the input file (cols D and E)
codes = []
months = []
for row in ws.iter_rows(min_row=3, max_row=14, min_col=4, max_col=5, values_only=True):
    code, month = row
    codes.append(str(code))
    months.append(str(month))

# Process rows A3:A14 and populate B3:B14
for idx, row in enumerate(ws.iter_rows(min_row=3, max_row=14, min_col=1, max_col=1, values_only=True), start=3):
    cell_text = row[0]
    res = extract_code_and_lookup_month(cell_text if cell_text else '', codes, months)
    ws[f'B{idx}'] = res if res else ''

wb.save(output_path)
