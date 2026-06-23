import pandas as pd
from openpyxl import load_workbook

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/regression_gate/after_pass/core_3413/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/regression_gate/after_pass/core_3413/output.xlsx'

# Read with no header so we can use Excel-like indices
src_df = pd.read_excel(input_fp, header=None)
wb = load_workbook(input_fp)
ws = wb.active

def safe_float(val):
    try:
        return float(val)
    except Exception:
        return 0

# G3:G6 means row indices 3,4,5,6 (Excel 1-based), i=2,3,4,5 (0-based)
for excel_row in range(3, 7):
    dept_lookup = ws.cell(row=excel_row, column=5).value  # col E
    ru_lookup = ws.cell(row=excel_row, column=6).value    # col F

    # src table - columns are Dept(col0) RU(col1) Total(col2)
    department = src_df[0].astype(str).fillna("")
    ru = src_df[1].astype(str).fillna("")
    value_col = src_df[2]

    # Get only valid numeric for value_col for matching rows
    both_match = (department == str(dept_lookup)) & (ru == str(ru_lookup))
    partial_match = (department == str(dept_lookup))
    # main sum logic
    if both_match.any():
        total = sum([safe_float(v) for v, ok in zip(value_col, both_match) if ok])
    else:
        total = sum([safe_float(v) for v, ok in zip(value_col, partial_match) if ok])
    ws.cell(row=excel_row, column=7, value=total)  # col G

wb.save(output_fp)
