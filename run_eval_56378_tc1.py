import openpyxl
from openpyxl.styles import Alignment

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_56378_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_56378_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Header is on row 4 (from inspection)
header_row = 4
headers = {cell.value: idx for idx, cell in enumerate(ws[header_row], 1)}
product_col = headers.get('PRODUCT')
qty_col = headers.get('QUANTITY UNITS')

if not product_col or not qty_col:
    raise Exception('Cannot find required columns: PRODUCT, QUANTITY UNITS')

# Gather rows with non-empty QUANTITY UNITS, max 4 (for output)
data_rows = []
for row in ws.iter_rows(min_row=header_row + 1, min_col=1, max_col=ws.max_column, values_only=False):
    if not row[product_col-1].value:
        break
    if row[qty_col-1].value not in (None, ''):
        data_rows.append(row)

# Copy up to 4 rows, columns D:J (4:10, 7 cols) to L5:R8 (cols 12–18)
start_row, start_col = 5, 12
num_rows = min(4, len(data_rows))
num_cols = 7
src_start_col = 4 - 1 # D (openpyxl is 0-based)
for i in range(num_rows):
    for j in range(num_cols):
        val = data_rows[i][src_start_col + j].value
        cell = ws.cell(row=start_row + i, column=start_col + j)
        cell.value = val
        # Left align Product (first col), right align O:R (cols 15:18; output cols 3-6)
        if j == 0:
            cell.alignment = Alignment(horizontal='left')
        elif 3 <= j <= 6:
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='general')

wb.save(output_path)
