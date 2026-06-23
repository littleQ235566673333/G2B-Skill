from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_36277/r0/evolve_36277/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_36277/r0/evolve_36277/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Write the lookup formula in I2:I5
table_range = '$B$2:$E$5'
header_range = '$B$1:$E$1'
for row in range(2, 6):
    header_cell = f'H{row}'
    dest_cell = f'I{row}'
    # The formula works for each row, relying on the header in H2, H3, etc. and the data in table_range
    formula = f'=INDEX({table_range},ROW()-1,MATCH({header_cell},{header_range},0))'
    ws[dest_cell] = formula

wb.save(output_path)
