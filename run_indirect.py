from openpyxl import load_workbook

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/before_pass/core_41601/input.xlsx')
ws = wb['Students']
for row in range(2, 8):
    ws[f'E{row}'] = f'=INDIRECT("\'" & A{row} & "\'!C2")'
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/before_pass/core_41601/output.xlsx')
