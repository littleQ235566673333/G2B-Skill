import openpyxl
from openpyxl.styles import Alignment

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed0/eval_56378_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed0/eval_56378_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Folha1']

# Frame 1 columns and header
header_row = 4
out_start_col = 12  # L -> 12
out_end_col = 18    # R -> 18
out_start_row = 5
out_end_row = 8

frame1_start, frame1_end = 3, 9
headers = []
for idx in range(frame1_start, frame1_end+1):
    headers.append(str(ws.cell(row=header_row, column=idx).value or '').strip().lower())
product_col = None
quantity_col = None
for idx, name in enumerate(headers, start=frame1_start):
    if name == 'product':
        product_col = idx
    elif name == 'quantity units':
        quantity_col = idx

# Collect all frame 1 data (not just Product and Quantity, but all columns from frame 1)
data_rows = []
for row in range(header_row+1, ws.max_row+1):
    row_values = [ws.cell(row=row, column=col).value for col in range(frame1_start, frame1_end+1)]
    prod = row_values[product_col-frame1_start] if product_col else None
    qty = row_values[quantity_col-frame1_start] if quantity_col else None
    if prod is not None and qty is not None and str(qty).strip() != '' and str(qty).strip() != 'None':
        data_rows.append(row_values)

# Write to frame 2 at L5:R8
for idx, values in enumerate(data_rows[:out_end_row-out_start_row+1]):
    excel_row = out_start_row + idx
    for j, val in enumerate(values):
        col = out_start_col + j
        ws.cell(row=excel_row, column=col, value=val)
        # Set alignment: Product (col L) left, O:P:Q:R (cols 15:18) right
        if j == product_col - frame1_start:
            ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal='left')
        elif col >= 15 and col <= 18:
            ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal='right')
        else:
            ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal='general')

wb.save(output_path)
