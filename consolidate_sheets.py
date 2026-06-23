import openpyxl
from openpyxl import load_workbook
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_3/regression_gate/before_pass/core_80-42/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_3/regression_gate/before_pass/core_80-42/output.xlsx'

wb = load_workbook(input_path)
ws_result = wb['Consolidate_ALL']

# Headers for reference
headers_result = [cell.value for cell in ws_result[1][:12]]
def find_first_blank(ws, start_row, col=1):
    """Find first completely blank row in the given worksheet (in col) starting from start_row."""
    row = start_row
    while any(ws.cell(row=row, column=c).value for c in range(1, 13)):
        row += 1
    return row

start_row = find_first_blank(ws_result, 2)

# Column selection mapping for 'Jack', 'Henry', 'Richard' for 'Consolidate_ALL' columns A-K
# Using printed headers:
# Consolidate_ALL and Jack:     1 2 3 4 5 6  7 8 9 10 11 12
# Henry/Richard:               1 2 3 4 5 6  7 8 9 10 11 12 13 14 15 16 17 18
# 'Consolidate_ALL' fields:    Date, Name, Description, Parent Node, Parent Description, CC External Code, Available in CORE, OAL CC Type, CC Default for LOB, CC Fiscal Recharge Owner, Ancestor List, Sheet Name
#
# For Henry/Richard, mapping to Consolidate_ALL (col: header):
# 1:Date
# 3:Name
# 4:Description
# 5:Parent Node
# 6:Parent Description
# 7:CC External Code
# - [Field 14] Available in CORE
# 15:OAL CC Type
# 16:CC Default for LOB
# 17:CC Fiscal Recharge Owner
# 18:Ancestor List
# 12:Sheet Name <- we set this manually.
henry_map = [0, 2, 3, 4, 5, 6, 13, 14, 15, 16, 17]  # zero-based indices for columns to copy

def extract_rows(ws, col_map, skip_header=True):
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if i == 1 and skip_header:
            continue
        if all(cell is None for cell in row):
            continue
        # Defensive: only extract if enough columns
        if max(col_map) < len(row):
            rows.append([row[j] for j in col_map])
    return rows

def append_to_sheet(ws_dest, first_row, rows, source_sheet_name):
    for data in rows:
        # data: 11 columns from mapping
        # set the 12th col to source_sheet_name
        for col_idx, val in enumerate(data, 1):
            ws_dest.cell(row=first_row, column=col_idx, value=val)
        ws_dest.cell(row=first_row, column=12, value=source_sheet_name)
        first_row += 1
    return first_row

# Append Jack rows (1:1 map as in result, from col 1--12, but use only 11 and set name manually)
ws_jack = wb['Jack']
jack_rows = extract_rows(ws_jack, list(range(0,11)))
start_row = append_to_sheet(ws_result, start_row, jack_rows, 'Jack')

ws_henry = wb['Henry']
henry_rows = extract_rows(ws_henry, henry_map)
start_row = append_to_sheet(ws_result, start_row, henry_rows, 'Henry')

ws_richard = wb['Richard']
richard_rows = extract_rows(ws_richard, henry_map)
start_row = append_to_sheet(ws_result, start_row, richard_rows, 'Richard')

# Save
wb.save(output_path)
print(f"Consolidation completed and output saved to {output_path}")
