import openpyxl

# File paths
input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_6/regression_gate/before_fix/core_50916/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_6/regression_gate/before_fix/core_50916/output.xlsx'

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Extract cycle day to class mappings from rows 2-8
cycle_day_map = {}
for i in range(2, 9):
    day = ws[f'A{i}'].value
    classes = [ws[f'{col}{i}'].value for col in ['C','D','E','F','G','H']]
    cycle_day_map[day] = classes

# Target rows and columns for output (C12:H14)
for row in range(12, 15):
    cycle_day = ws[f'A{row}'].value
    classes = cycle_day_map.get(cycle_day, [None]*6)
    for j, col in enumerate(['C','D','E','F','G','H']):
        ws[f'{col}{row}'] = classes[j]

# Save the modified workbook
wb.save(output_file)
