from openpyxl import load_workbook
from datetime import datetime

in_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/regression_gate/before_pass/core_41589/input.xlsx"
out_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/regression_gate/before_pass/core_41589/output.xlsx"

wb = load_workbook(in_path)
ws = wb['Contact List']

today = datetime.now()

for row in ws.iter_rows(min_row=4, min_col=8, max_col=9):
    h_cell, i_cell = row
    j_val = 'NO ACTION'
    last_contact = None
    # Parse the date in H column
    try:
        if isinstance(h_cell.value, datetime):
            last_contact = h_cell.value
        elif isinstance(h_cell.value, str):
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
                try:
                    last_contact = datetime.strptime(h_cell.value, fmt)
                    break
                except ValueError:
                    continue
    except Exception:
        last_contact = None
    if last_contact and isinstance(i_cell.value, str) and i_cell.value.strip().upper() == 'YES':
        days_diff = (today - last_contact).days
        if days_diff <= 30:
            j_val = 'HOLD'
        else:
            j_val = 'TOUCH BASE'
    ws.cell(row=h_cell.row, column=10, value=j_val)

wb.save(out_path)
