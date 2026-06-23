import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numpy as np
from datetime import datetime, timedelta

# Parameters
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_2/group_44017/r0/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_2/group_44017/r0/evolve_44017/output.xlsx'
sheet = 'Data'

# Read Excel
wb = load_workbook(input_path)
ws = wb[sheet]

# Helper: get all month datetime objects from AD9:AO9
month_cols = list(range(30, 42+1))  # AD:AO (Excel col 30:42)
month_dates = [ws.cell(row=9, column=c).value for c in month_cols]

# For each customer row (14 to 42)
for row in range(14, 43):
    base_rate = ws.cell(row=row, column=23).value    # W
    eff_date  = ws.cell(row=row, column=12).value    # L
    freq      = ws.cell(row=row, column=10).value    # J
    pcts = [ws.cell(row=row, column=col).value for col in range(13, 17)] # M:P
    
    # Try to parse frequency (handle str/formula fallback)
    try:
        freq_val = int(freq) if freq is not None else None
    except:
        try:
            freq_val = int(float(freq))
        except:
            freq_val = None
    if freq_val is None or (not isinstance(eff_date, datetime)) or base_rate is None or np.isnan(base_rate):
        # Clear any prior fill
        for col, m in zip(month_cols, month_dates):
            ws.cell(row=row, column=col).value = None
        continue
    # Clean up increase percentages
    incs = [x if isinstance(x, (int,float)) and pd.notnull(x) else 0.0 for x in pcts]
    for idx, (col, col_dt) in enumerate(zip(month_cols, month_dates)):
        # Only fill if date cell is a datetime
        if not isinstance(col_dt, datetime):
            ws.cell(row=row, column=col).value = None
            continue
        # Only apply when date >= effective date
        if col_dt < eff_date:
            ws.cell(row=row, column=col).value = None
            continue
        # Determine how many increases have occurred for this date
        months_since = (col_dt.year - eff_date.year) * 12 + (col_dt.month - eff_date.month)
        nth_wave = int(months_since // freq_val)
        nth_wave = min(nth_wave, 4)
        applicable = incs[:nth_wave]
        rate = base_rate
        for pct in applicable:
            rate *= (1 + pct)
        ws.cell(row=row, column=col).value = round(rate, 2)
        # Remove fill if any
        ws.cell(row=row, column=col).fill = ws.cell(row=row, column=col).fill.copy(patternType=None)

wb.save(output_path)
print('Done')
