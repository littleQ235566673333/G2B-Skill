import pandas as pd
from openpyxl import load_workbook

# Input and output paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_6/regression_gate/before_fix/core_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_6/regression_gate/before_fix/core_4714/output.xlsx'

# Read in the spreadsheet data, assuming relevant columns are A (Employee), B (?), C (Month), D (Hours)
df = pd.read_excel(input_path, engine='openpyxl')

# We need columns A, C, D; output in E2:E25
# Normalizing column names in case
columns_map = {col: col for col in df.columns}
df.columns = [str(c) for c in df.columns]

# Helper: convert column C (Month) to sortable value
# If Month is YYYY-MM or similar, parse as date, else fallback to order of appearance
if df.columns[2].lower().startswith('month'):
    df['Month'] = pd.to_datetime(df[df.columns[2]], errors='coerce', format='%Y-%m')
else:
    df['Month'] = pd.to_datetime(df[df.columns[2]], errors='coerce')

df['Employee'] = df[df.columns[0]] # Just for clarity
df['Hours'] = df[df.columns[3]]

# Sort by Employee, Month
sdf = df.sort_values(['Employee','Month']).reset_index(drop=True)

# Prepare result list for E2-E25
results = [''] * sdf.shape[0]

for idx, row in sdf.iterrows():
    emp = row['Employee']
    dt = row['Month']
    mask = (sdf['Employee'] == emp) & (sdf['Month'] <= dt)
    # Last 4 months including current
    last4 = sdf[mask].sort_values('Month').tail(4)
    if last4.shape[0] < 4:
        results[idx] = 'n/a'
    else:
        avg = last4['Hours'].mean()
        if avg == int(avg):
            results[idx] = str(int(avg))
        else:
            results[idx] = f"{avg:.2f}".rstrip('0').rstrip('.')

# Write results to E2:E25 in the original workbook
wb = load_workbook(input_path)
ws = wb.active
for i, val in enumerate(results[:24], start=2): # E2-E25
    ws[f'E{i}'] = val
wb.save(output_path)
