import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_57743_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_57743_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Build a search dictionary from model numbers (C) to prices (D)
model_to_price = {}
for row in range(2, ws.max_row + 1):
    model = ws[f'C{row}'].value
    price = ws[f'D{row}'].value
    if model is not None:
        model_to_price[model] = price

# For each model in column A, lookup its price from the dictionary
# Write results to column B (B2:B19)
for row in range(2, 20):
    model = ws[f'A{row}'].value
    price = model_to_price.get(model, '') # Leave blank if not found
    ws[f'B{row}'] = price

wb.save(output_path)
print('Done')
