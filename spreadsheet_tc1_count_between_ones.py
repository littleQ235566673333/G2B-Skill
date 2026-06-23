import openpyxl

# Input and output file paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_30930_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_30930_tc1/output.xlsx'

# Load workbook and select Sheet1
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Define the index range for rows (2 to 66 inclusive)
start_row, end_row = 2, 66

# Store indices of rows where Column B == 1
indices_b1 = []
for row in range(start_row, end_row + 1):
    val_b = ws[f'B{row}'].value
    if val_b == 1:
        indices_b1.append(row)

# Calculate counts between B==1 occurrences
prev_b1_row = start_row - 1
for idx in indices_b1:
    # Range for counting is (prev_b1_row+1) to (idx-1) or just to idx if first
    count = 0
    for r in range(prev_b1_row + 1, idx):
        val_a = ws[f'A{r}'].value
        if val_a is not None and val_a > 0:
            count += 1
    # Write count to Column C of idx
    ws[f'C{idx}'].value = count
    prev_b1_row = idx

# Leave blank in other rows (optional, to ensure C2:C66 is clear except these)
for row in range(start_row, end_row + 1):
    if row not in indices_b1:
        ws[f'C{row}'].value = None

# Save workbook
wb.save(output_path)
