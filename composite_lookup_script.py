from openpyxl import load_workbook

infile = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/regression_gate/after_fix/core_39515/input.xlsx"
outfile = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_7/regression_gate/after_fix/core_39515/output.xlsx"

wb = load_workbook(infile)
ws = wb.active

# Gather composite keys from headers (C1:N1)
header_keys = []
for col in range(3, 15):
    val = ws.cell(row=1, column=col).value
    if val is None:
        header_keys.append("")
    else:
        header_keys.append(str(val).strip())

# For each row (2:13), compute the composite key and write value to O
for row in range(2, 14):
    month = ws.cell(row=row, column=1).value
    year = ws.cell(row=row, column=2).value
    if month and year:
        composite = f"{month}-{year}"
    else:
        composite = ""
    found = False
    for idx, key in enumerate(header_keys):
        if key == composite:
            val = ws.cell(row=row, column=3 + idx).value
            ws.cell(row=row, column=15, value=val)
            found = True
            break
    if not found:
        ws.cell(row=row, column=15, value=None)

wb.save(outfile)
