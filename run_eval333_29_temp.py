import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun2/eval_333-29_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun2/eval_333-29_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_A = wb['A']
ws_B = wb['B']

# Step 1: Find first cell in column P (16) that is 'Yes' or 'NA'
found = False
for idx, cell in enumerate(ws_A['P'], start=1):
    val = str(cell.value).strip() if cell.value is not None else ''
    if val in ['Yes', 'NA']:
        found = True
        idx_first = idx
        break
# Step 2: If found, from above row in column L (12), find value 100
found_100 = False
date_value = None
if found:
    for j in range(idx_first - 1, 0, -1):
        if ws_A.cell(row=j, column=12).value == 100:
            found_100 = True
            date_value = ws_A.cell(row=j, column=6).value
            break
    if not found_100:
        date_value = None
else:
    # Use F3 if not found
    date_value = ws_A['F3'].value

# Step 3: Match cell F1 text to column C of B and write date to column F in B for those matches
ws_B_F4_result = ''
text_F1 = str(ws_A['F1'].value).strip() if ws_A['F1'].value is not None else ''
matched = False
for row in range(2, ws_B.max_row + 1): # assuming headers in row 1
    if str(ws_B.cell(row=row, column=3).value).strip() == text_F1:
        ws_B.cell(row=row, column=6, value=date_value)
        matched = True
        ws_B_F4_result = date_value
        break # Only the first match according to description
if not matched:
    ws_B_F4_result = date_value

# Step 4: Write answer to B!F4
ws_B['F4'] = ws_B_F4_result

wb.save(output_path)
