import openpyxl

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_41978_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun1/eval_41978_tc1/output.xlsx'

# Load the workbook and select the sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Cumulative']

# Read years and status
years = [ws[f'G{row}'].value for row in range(14, 186)]
statuses = [ws[f'J{row}'].value for row in range(14, 186)]

# Get sorted unique years, ignoring empty
unique_years = sorted(set(y for y in years if y is not None))

# Count 'open' cases for each year
year_open_counts = []
for year in unique_years:
    count = sum(1 for y, status in zip(years, statuses) if y == year and str(status).strip().lower() == 'open')
    year_open_counts.append((year, count))

# Write counts to I2:I11 (first 10 unique years)
for i, (year, count) in enumerate(year_open_counts[:10]):
    ws[f'I{i+2}'].value = count

# Save the file
wb.save(output_path)
