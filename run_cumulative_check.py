from openpyxl import load_workbook
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun1/eval_42198_tc1/input.xlsx')
ws = wb.active
for row in range(2, 8):
    worst = False
    ignore = False
    bad = False
    for r in range(2, row+1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a == "Potato" and b is False:
            worst = True
        elif a == "Tomato" and b is False:
            ignore = True
        elif a == "Pickle" and b is False:
            bad = True
    if worst:
        ws.cell(row=row, column=3).value = "Worst"
    elif ignore:
        ws.cell(row=row, column=3).value = "Ignore"
    elif bad:
        ws.cell(row=row, column=3).value = "Bad"
    else:
        ws.cell(row=row, column=3).value = "Good"
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun1/eval_42198_tc1/output.xlsx')
