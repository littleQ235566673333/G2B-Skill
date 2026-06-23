from openpyxl import load_workbook
import re
from datetime import datetime

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_4/regression_gate/after_pass/core_290-27/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_4/regression_gate/after_pass/core_290-27/output.xlsx'

wb = load_workbook(input_fp)
ws = wb[wb.sheetnames[0]]

for row in ws.iter_rows(min_row=14, max_row=137, min_col=2, max_col=2):
    cell = row[0]
    val = cell.value
    # Only operate on strings, skip dates and numerics
    if isinstance(val, str):
        # Only operate if it starts with 2-3 all uppercase letters possible space, directly followed by a number
        new_val = re.sub(r'^[A-Z]{2,3} ?', '', val)
        cell.value = new_val if (new_val != val) else val

wb.save(output_fp)
print('Done')
