from openpyxl import load_workbook
from datetime import datetime, timedelta, time

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun2/eval_35739_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun2/eval_35739_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

for row in range(2, 101):
    a = ws[f'A{row}'].value
    b = ws[f'B{row}'].value
    if (a is None or a == '' or b is None or b == ''):
        ws[f'C{row}'].value = None
        continue
    val = b
    try:
        # Excel may store time as float (fraction of a day), datetime, or python time type
        if isinstance(val, float):
            # Excel serial time; 0.5 means 12:00 PM
            orig = datetime(1899, 12, 30) + timedelta(days=val)
            result = orig - timedelta(minutes=30)
            ws[f'C{row}'].value = result.time()
        elif isinstance(val, datetime):
            result = (val - timedelta(minutes=30)).time()
            ws[f'C{row}'].value = result
        elif isinstance(val, time):
            t = datetime.combine(datetime(2000, 1, 1), val)  # arbitrary date
            result = (t - timedelta(minutes=30)).time()
            ws[f'C{row}'].value = result
        else:
            ws[f'C{row}'].value = None
    except Exception:
        ws[f'C{row}'].value = None

wb.save(output_path)
