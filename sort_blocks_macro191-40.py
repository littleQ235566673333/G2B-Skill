import openpyxl
from openpyxl.styles import Alignment, Font
import re

# Paths
infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_191-40_tc1/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_191-40_tc1/output.xlsx'

wb = openpyxl.load_workbook(infile)
ws = wb['Sheet1']

def is_blank_row(row):
    return all(cell.value is None for cell in row)

def is_formula(cell):
    v = cell.value
    return isinstance(v, str) and v.strip().startswith('=')

def round_n_fmt(val):
    if val is None:
        return None, None
    try:
        f = float(val)
        if f.is_integer():
            return int(f), '0'  # Same as ##0
        else:
            return round(f, 1), '0.0'  # Same as ##0.0
    except:
        return val, None

rows = list(ws.iter_rows(min_row=1, max_col=8, max_row=85))
block_ranges = []
start = None
for i, row in enumerate(rows):
    if not is_blank_row(row):
        if start is None:
            start = i
    else:
        if start is not None:
            block_ranges.append((start, i-1))
            start = None
if start is not None:  # Add last block if sheet doesn't end with blank
    block_ranges.append((start, len(rows)-1))

for bstart, bend in block_ranges:
    block = rows[bstart:bend+1]
    # Ensure formulas restored in D, F
    for ridx, row in enumerate(block):
        r = bstart + ridx + 1
        d = ws.cell(row=r, column=4)
        if not is_formula(d):
            d.value = f'=B{r}+C{r}'
        f = ws.cell(row=r, column=6)
        if not is_formula(f):
            f.value = f'=D{r}+E{r}'
        # (if E is also formula, handle if needed)
    # Evaluate D values using a read-only workbook
    wb2 = openpyxl.load_workbook(infile, data_only=True)
    ws2 = wb2['Sheet1']
    vals_rows = []
    for ridx, row in enumerate(block):
        r = bstart + ridx + 1
        dval = ws2.cell(row=r, column=4).value
        vals_rows.append((dval if dval is not None else float('-inf'), row))
    # Sort by D, descending numeric
    vals_rows.sort(key=lambda x: (x[0] if isinstance(x[0], (int, float)) else float('-inf')), reverse=True)
    # Paste sorted rows back, update formulas to reference correct rows
    for new_idx, (dval, original_row) in enumerate(vals_rows):
        target_row = ws[ws.min_row + bstart + new_idx]
        for col in range(1, 9):
            cell = ws.cell(row=bstart + 1 + new_idx, column=col)
            src = original_row[col - 1]
            # For D & F, ensure formulas reference the correct target row
            if col == 4:
                cell.value = f'=B{bstart + 1 + new_idx}+C{bstart + 1 + new_idx}'
            elif col == 6:
                cell.value = f'=D{bstart + 1 + new_idx}+E{bstart + 1 + new_idx}'
            else:
                cell.value = src.value
            # Apply number formatting
            if col in [2, 3, 4, 5, 6]:
                vcell = cell
                v = ws2.cell(row=(bstart + 1 + new_idx), column=col).value
                disp, nf = round_n_fmt(v)
                if nf == '0':
                    vcell.number_format = '#,##0'
                elif nf == '0.0':
                    vcell.number_format = '#,##0.0'
        # Bold/center (D,E,F)
        for col in [4, 5, 6]:
            c = ws.cell(row=bstart + 1 + new_idx, column=col)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')
    # Blank row after block
    if bend + 1 < len(rows):
        for col in range(1, 9):
            ws.cell(row=bend + 2, column=col).value = None
# Save
wb.save(outfile)
