import openpyxl

# Load workbook and sheet
target_file = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_seed0/eval_56378_tc1/input.xlsx'
wb = openpyxl.load_workbook(target_file)
ws = wb['Folha1']

# Identify frame 1 columns for headers on row 4
header_row = 4
frame1_start, frame1_end = 3, 9  # C:I, 1-based
headers = []
for idx in range(frame1_start, frame1_end+1):
    headers.append(str(ws.cell(row=header_row, column=idx).value or '').strip().lower())

product_col = None
quantity_col = None
for idx, name in enumerate(headers, start=frame1_start):
    if name == 'product':
        product_col = idx
    elif name == 'quantity units':
        quantity_col = idx

# Collect eligible products from frame 1 (rows 5-20)
data = []
for row in range(header_row+1, ws.max_row+1):
    prod = ws.cell(row=row, column=product_col).value if product_col else None
    qty = ws.cell(row=row, column=quantity_col).value if quantity_col else None
    if prod is not None and qty is not None and str(qty).strip() != '' and str(qty).strip() != 'None':
        data.append((prod, qty, row))

print('CANDIDATE FRAME1:', data)
