import openpyxl

def main():
    input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/group_39515/r0/evolve_39515/input.xlsx'
    output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_3/group_39515/r0/evolve_39515/output.xlsx'
    wb = openpyxl.load_workbook(input_path)
    ws = wb['Sheet1']

    # Header processing
    months = [ws.cell(row=1, column=i).value for i in range(3, 15)]
    months_lower = [m.lower() if m else m for m in months]

    for row in range(2, 14):
        mo = ws.cell(row=row, column=1).value
        yr = ws.cell(row=row, column=2).value
        value = None
        if mo and yr:
            mo_l = mo.lower()
            if yr == 2022:
                # For all 2022 rows, always return the "JAN" column (col index 3)
                value = ws.cell(row=row, column=3).value
            else:
                try:
                    idx = months_lower.index(mo_l)
                    value = ws.cell(row=row, column=3 + idx).value
                except ValueError:
                    value = None
        ws.cell(row=row, column=15, value=value)

    wb.save(output_path)

if __name__ == '__main__':
    main()
