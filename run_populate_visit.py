import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def parse_date(val):
    if isinstance(val, (datetime, pd.Timestamp)):
        return val
    try:
        return pd.to_datetime(val, dayfirst=True, errors='coerce')
    except Exception:
        return pd.NaT

wb_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_2/regression_gate/before_pass/core_9726/input.xlsx'
wb = load_workbook(wb_path)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))

visit_header_idx = submission_header_idx = None
for idx, row in enumerate(rows):
    if row and 'Visit date' in row:
        visit_header_idx = idx
    if row and 'Submission date' in row:
        submission_header_idx = idx
    if visit_header_idx is not None and submission_header_idx is not None and submission_header_idx > visit_header_idx:
        break

if visit_header_idx is None or submission_header_idx is None:
    raise Exception('Could not find expected headers for both tables.')

header1 = [c for c in rows[visit_header_idx] if c is not None]
data1 = rows[visit_header_idx+1:submission_header_idx]
while data1 and all(x is None for x in data1[-1]):
    data1 = data1[:-1]
table1 = pd.DataFrame(data1, columns=header1)
# Fix duplicate columns
counts = {}
newcols = []
for c in table1.columns:
    if c not in counts:
        counts[c] = 1
        newcols.append(c)
    else:
        counts[c] += 1
        newcols.append(f"{c}_{counts[c]}")
table1.columns = newcols
table1 = table1.dropna(how='all').reset_index(drop=True)
if 'Student ID' in table1.columns:
    table1 = table1[table1['Student ID'] != 'Student ID']

header2 = [c for c in rows[submission_header_idx] if c is not None]
data2 = []
for r in rows[submission_header_idx+1:]:
    if not any((x is not None and x != "") for x in r):
        break
    data2.append(r[:len(header2)])
table2 = pd.DataFrame(data2, columns=header2)
counts2 = {}
newcols2 = []
for c in table2.columns:
    if c not in counts2:
        counts2[c] = 1
        newcols2.append(c)
    else:
        counts2[c] += 1
        newcols2.append(f"{c}_{counts2[c]}")
table2.columns = newcols2
table2 = table2.dropna(how='all').reset_index(drop=True)
if 'Student ID' in table2.columns:
    table2 = table2[table2['Student ID'] != 'Student ID']

if 'Student ID' in table1.columns:
    table1 = table1[table1['Student ID'].notna()]
    table1['Student ID'] = table1['Student ID'].astype(str)
    table1['Visit date'] = table1['Visit date'].apply(parse_date)
    table1 = table1[table1['Visit date'].notna()]
if 'Student ID' in table2.columns:
    table2 = table2[table2['Student ID'].notna()]
    table2['Student ID'] = table2['Student ID'].astype(str)
    table2['Submission date'] = table2['Submission date'].apply(parse_date)

visit_map = {}
for sid, group in table1.groupby('Student ID'):
    visit_map[sid] = sorted([d for d in group['Visit date'] if not pd.isna(d)])

visit_on_submission = []
for _, row in table2.iterrows():
    sid = row['Student ID']
    sub_date = row['Submission date']
    visits = visit_map.get(sid, [])
    last_visit = None
    for v in visits:
        if v < sub_date:
            last_visit = v
        else:
            break
    visit_on_submission.append(last_visit)
table2['Visit date'] = visit_on_submission

wb2 = load_workbook(wb_path)
ws2 = wb2.active
ws2['J2'] = visit_on_submission[0].strftime('%d-%b-%y') if visit_on_submission[0] else ''
wb2.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot/train/iter_2/regression_gate/before_pass/core_9726/output.xlsx')
print('Wrote to J2:', ws2['J2'].value)
