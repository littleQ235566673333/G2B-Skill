import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_3/regression_gate/after_fix/core_4714/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_3/regression_gate/after_fix/core_4714/output.xlsx'

# Read input with pandas

df = pd.read_excel(input_fp)

# We need to work on rows 2 to 25 (E2:E25) - zero indexed 1 to 24
num_rows = 24

def parse_month(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m')
    try:
        return pd.to_datetime(str(val)).strftime('%Y-%m')
    except Exception:
        return val

# Ensure we have a comparable month column
col_employee = df.columns[0]
col_month = df.columns[2]
col_hours = df.columns[3]
df['Month_fmt'] = df[col_month].apply(parse_month)

results = []
for idx in range(num_rows):
    row = df.iloc[idx]
    emp = row[col_employee]
    cur_month = row['Month_fmt']
    emp_df = df[df[col_employee] == emp].sort_values('Month_fmt').reset_index()
    # Find the position of current row in sorted emp_df
    emp_months = list(emp_df['Month_fmt'])
    try:
        ix = emp_months.index(cur_month)
    except ValueError:
        results.append('n/a')
        continue
    if ix < 3:
        results.append('n/a')
        continue
    last4 = emp_df.iloc[ix-3:ix+1]
    if len(last4) < 4:
        results.append('n/a')
        continue
    hours = last4[col_hours]
    avg = hours.mean()
    # Show as whole number if possible
    if float(int(avg)) == avg:
        avg = int(avg)
    results.append(avg)

# Write to E2:E25
wb = load_workbook(input_fp)
ws = wb.active
for i, val in enumerate(results):
    ws[f'E{i+2}'] = val
wb.save(output_fp)
