import openpyxl

# Input and output paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/after_pass/core_41601/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/after_pass/core_41601/output.xlsx"

# Load workbook
wb = openpyxl.load_workbook(input_path)
ws_students = wb['Students']

# Collect student names from A2:A7
student_cells = list(ws_students['A'][1:7])  # A2 to A7
ages = []
for cell in student_cells:
    student_name = cell.value
    age = None
    if student_name in wb.sheetnames:
        ws = wb[student_name]
        age = ws['C2'].value
    ages.append(age)
    ws_students[f'E{cell.row}'] = age

# Save to output
wb.save(output_path)
