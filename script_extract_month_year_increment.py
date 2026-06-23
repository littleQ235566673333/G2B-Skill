import openpyxl
from openpyxl.styles import PatternFill
import calendar

def add_one_month(mon, year):
    try:
        mon_abbr = mon[:3].capitalize()
        mon_num = list(calendar.month_abbr).index(mon_abbr)
        if mon_num == 12:
            return ('Jan', str(int(year)+1)[-2:])
        elif mon_num == 0:  # Not a valid month
            return ('???', year)
        else:
            return (calendar.month_abbr[mon_num+1], year)
    except Exception:
        return ('???', year)

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_51354_tc1/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_51354_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_file)
ws = wb.active
fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
for row in range(2, 7):
    text = str(ws[f'A{row}'].value)
    date_part = text[-6:]
    mon = date_part[:3]
    yr = date_part[-2:]
    ws[f'D{row}'] = yr
    ws[f'D{row}'].fill = fill
    next_mon, next_yr = add_one_month(mon, yr)
    ws[f'E{row}'] = f'{next_mon} {next_yr}'
wb.save(output_file)
