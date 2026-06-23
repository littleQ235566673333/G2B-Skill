import openpyxl
import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_3/group_130-9/r3/evolve_130-9/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_3/group_130-9/r3/evolve_130-9/output.xlsx'

wb = openpyxl.load_workbook(input_path)
src = wb['cdnr']
dst = wb['b2b, sez, de']

# --- Helper functions ---
def get_headers(ws):
    for r in range(1, ws.max_row+1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
        if any(row_vals):
            return r, {v: c for c, v in enumerate(row_vals, 1) if v}
    return None, {}

src_header_row, src_headers = get_headers(src)
dst_header_row, dst_headers = get_headers(dst)

# Find first empty row in destination (start at 7 as per instructions)
row = 7
max_col = dst.max_column
while any([dst.cell(row=row, column=col).value for col in range(1, max_col+1)]):
    row += 1
insert_row = row

# Prepare mapping from destination headers to source columns, None if not present in source
header_map = {d_head: src_headers.get(d_head) for d_head in dst_headers}

# For each data row in source, map and append to destination
for src_row in range(src_header_row+1, src.max_row+1):
    # skip if source row is empty
    if all([src.cell(row=src_row, column=c).value is None for c in src_headers.values()]):
        continue
    for dst_col in range(1, max_col+1):
        d_head = None
        for head, col in dst_headers.items():
            if col == dst_col:
                d_head = head
                break
        cell = dst.cell(row=insert_row, column=dst_col)
        # Special handling for columns H (8) and I (9)
        if dst_col == 8:  # H
            cell.value = None
        elif dst_col == 9:  # I
            cell.value = 'Credit Note'
        # Handle columns L and M (12 and 13): write negative values
        elif dst_col in (12, 13):
            src_col = header_map.get(d_head)
            val = src.cell(row=src_row, column=src_col).value if src_col else None
            # Only integers/floats converted to negative
            if isinstance(val, (int, float)):
                cell.value = -abs(val)
            else:
                cell.value = val
            # Copy number format from prior row if present
            fmt = dst.cell(row=insert_row-1, column=dst_col).number_format
            if fmt:
                cell.number_format = fmt
        else:
            src_col = header_map.get(d_head)
            cell.value = src.cell(row=src_row, column=src_col).value if src_col else None
            # Copy number format from prior row if present
            fmt = dst.cell(row=insert_row-1, column=dst_col).number_format
            if fmt:
                cell.number_format = fmt
    insert_row += 1

wb.save(output_path)
