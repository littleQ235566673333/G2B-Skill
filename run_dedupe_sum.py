import openpyxl
import collections

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r2/eval_250-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r2/eval_250-20_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['RNM']

# Read A1:J20
rows = []
for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10, values_only=True):
    rows.append(list(row))
header = rows[0]
data = rows[1:]

# Group by columns B & C (indices 1,2) and sum J (index 9)
output_dict = collections.OrderedDict()
for r in data:
    key = (r[1], r[2])
    if key not in output_dict:
        output_dict[key] = list(r)
    else:
        existing_sum = output_dict[key][9] if output_dict[key][9] else 0
        new_val = r[9] if r[9] else 0
        output_dict[key][9] = existing_sum + new_val

output_data = [header] + list(output_dict.values())

# Pad to reach 20 rows if needed, clean up Nones
while len(output_data) < 20:
    output_data.append(['']*10)
for r in output_data:
    for i in range(len(r)):
        if r[i] is None:
            r[i] = ''

# Write back to ws ('RNM'!A1:J20)
for row_idx, row in enumerate(output_data, 1):
    for col_idx, val in enumerate(row, 1):
        ws.cell(row=row_idx, column=col_idx, value=val)

wb.save(output_path)
