from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2-PRUNED/eval_seed42/eval_50631_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2-PRUNED/eval_seed42/eval_50631_tc1/output.xlsx"

wb = load_workbook(input_path)
ws = wb["Sheet1"]

# Styles
yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
calibri = Font(name="Calibri")

formula = "=IF(ROW()-6<=NETWORKDAYS($B$3,$F$3,$J$7:$J$38),WORKDAY($B$3,ROW()-7,$J$7:$J$38),\"\")"

for row in range(7, 39):
    cell = ws[f"H{row}"]
    cell.font = calibri
    cell.value = formula
    if cell.fill.fgColor.rgb == "FFFFFF00" or (cell.fill.start_color and cell.fill.start_color.rgb == "FFFFFF00"):
        cell.fill = yellow

wb.save(output_path)
