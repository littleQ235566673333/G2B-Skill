import openpyxl
import re

def group_lookup(group_table, group_key, col_idx, material=None):
    # group_key: value to lookup, col_idx: 1-based index (excel conventions)
    for row in group_table:
        if material:
            if row[0] == group_key and row[1] == material:
                return row[col_idx]
        else:
            if row[0] == group_key:
                return row[col_idx]
    return None

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_7902_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed1/eval_7902_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
group_ws = wb['Grouping']
group_table = [tuple(cell.value for cell in row) for row in group_ws.iter_rows(min_row=2, max_row=7, max_col=9)]
ws = wb['Formula Required']

for i, row_idx in enumerate(range(3,7)):
    row_vals = [ws.cell(row=row_idx, column=col).value for col in range(3,10)]
    group_key = row_vals[0]
    material = row_vals[1] if row_vals[1] not in [None, '', '=VLOOKUP(B4,Grouping!$B$2:$I$7,2,FALSE)'] else None

    for j, col_idx in enumerate(range(4,11)):
        cell = ws.cell(row=row_idx, column=col_idx)
        formula = cell.value
        if formula and isinstance(formula, str) and formula.startswith('=VLOOKUP'):
            m = re.search(r'VLOOKUP\(([^,]+),Grouping!\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+),(\d+),FALSE\)', formula)
            if m:
                lookup_col_str = m.group(1).strip()
                lookup_col = lookup_col_str
                # Excel: C3, B4 etc -> get their value from worksheet
                if lookup_col_str[0] in ['A','B','C','D','E','F','G','H','I']:
                    lookup_cell = ws[lookup_col_str]
                    lookup_val = lookup_cell.value
                else:
                    lookup_val = group_key
                return_col = int(m.group(4))
                # If value present in same row at col-j (excluding formula cols), use value
                # Otherwise lookup
                if j < len(row_vals) and row_vals[j+1] not in [None, '', formula]:
                    cell.value = row_vals[j+1]
                else:
                    result = group_lookup(group_table, lookup_val, return_col-1, material)
                    cell.value = result
wb.save(output_path)
print('Output written.')
