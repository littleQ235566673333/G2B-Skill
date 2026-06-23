import openpyxl
from openpyxl.utils import get_column_letter

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_409-45_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_409-45_tc1/output.xlsx"
sheet_name = "DATA"

wb = openpyxl.load_workbook(input_path)
ws = wb[sheet_name]

I2 = ws['I2'].value

def is_row_empty(row_cells):
    return all((cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == "")) for cell in row_cells)

# Find all ranges by the ID in column B (col 2)
rows = list(ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row))

def delete_rows_by_ids(ws, target_id):
    to_delete = []
    in_range = False
    for i, row in enumerate(rows, 2):  # start at row 2
        cell_b = row[1].value
        if cell_b == target_id:
            in_range = True
        elif cell_b is not None and cell_b != target_id:
            in_range = False
        if in_range and is_row_empty(row):
            to_delete.append(i)
    # Remove from bottom to top
    for row_idx in reversed(to_delete):
        ws.delete_rows(row_idx, 1)

def delete_all_empty_rows(ws):
    to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row), 2):
        if is_row_empty(row):
            to_delete.append(i)
    for row_idx in reversed(to_delete):
        ws.delete_rows(row_idx, 1)

if I2:
    # Delete based on ID in column B
    target_id = I2
    delete_rows_by_ids(ws, target_id)
else:
    # Delete all empty rows in all ranges
    delete_all_empty_rows(ws)

wb.save(output_path)
