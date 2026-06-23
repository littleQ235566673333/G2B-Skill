import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_8/regression_gate/after_fix/core_40478/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_8/regression_gate/after_fix/core_40478/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

def extract_text(s):
    if not isinstance(s, str):
        return ''
    first = s.find(' ')
    if first == -1:
        return ''
    second = s.find(' ', first + 1)
    if second == -1:
        return ''
    after_second = s[second + 1:]
    dash_pos = after_second.find('-')
    if dash_pos == -1:
        return after_second.strip()
    return after_second[:dash_pos].strip()

for row in range(1, 4):
    cell_val = ws[f'A{row}'].value
    ws[f'B{row}'] = extract_text(cell_val)

wb.save(output_path)
