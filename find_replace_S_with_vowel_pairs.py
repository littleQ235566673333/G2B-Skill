import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_4/group_146-49/r2/evolve_146-49/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_4/group_146-49/r2/evolve_146-49/output.xlsx'

# Read all unique words in the file
words_df = pd.read_excel(input_path, sheet_name='Sheet1', header=None)
words = words_df.stack().dropna().astype(str).unique()
vowels = set('AEIOU')
pair_set = set()
words_set = set(words)

for w1 in words:
    s_positions = [i for i, c in enumerate(w1) if c == 'S']
    if not s_positions:
        continue
    for v in vowels:
        w2 = list(w1)
        for pos in s_positions:
            w2[pos] = v
        w2 = ''.join(w2)
        if w2 != w1 and w2 in words_set:
            # Ensure pair is unique in both orders
            pair_set.add(tuple(sorted((w1, w2))))

# Sort the pairs (case-insensitive)
pairs_sorted = sorted(pair_set, key=lambda x: (x[0].lower(), x[1].lower()))

# Write results to columns G and H, starting at row 1 (G1:H1)
wb = load_workbook(input_path)
ws = wb['Sheet1']

for idx, (w1, w2) in enumerate(pairs_sorted):
    ws.cell(row=idx+1, column=7).value = w1
    ws.cell(row=idx+1, column=8).value = w2

wb.save(output_path)
