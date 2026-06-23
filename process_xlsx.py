from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_1/regression_gate/after_pass/core_493-18/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_1/regression_gate/after_pass/core_493-18/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Step 1: Collect all values in column F (distinct, non-empty strings)
f_values = set()
for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True):
    val = row[0]
    if val is not None:
        f_values.add(str(val).strip())

# Step 2: Identify rows to keep (where A value matches some F)
rows_to_keep = []
for row in ws.iter_rows(min_row=2, max_col=6, values_only=False):
    a_val = row[0].value
    if a_val is not None and str(a_val).strip() in f_values:
        rows_to_keep.append([cell.value for cell in row[:3]])
# Step 3: Copy up kept A/B/C rows (fill from row 2)
for i, values in enumerate(rows_to_keep, start=2):
    for j, val in enumerate(values, start=1):
        ws.cell(row=i, column=j).value = val
# Step 4: Clear the rest of A/B/C rows (no duplicate shifting)
for i in range(2 + len(rows_to_keep), ws.max_row + 1):
    for j in range(1, 4):
        ws.cell(row=i, column=j).value = None
# Step 5: Ensure autofilter is on A1:C1
ws.auto_filter.ref = f'A1:C{ws.max_row}'
wb.save(output_path)
