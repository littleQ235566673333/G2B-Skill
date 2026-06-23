import openpyxl

def normalize(val):
    if val is None:
        return ""
    return str(val).strip().lower()

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_4/regression_gate/after_fix/core_547-43/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_4/regression_gate/after_fix/core_547-43/output.xlsx"

# Load input workbook and relevant sheets
wb = openpyxl.load_workbook(input_path)
ws_emp = wb["Emp"]
ws_lookup = wb["Lookup"]

# Map Emp headers -> col index
emp_header_map = {c.value: i for i, c in enumerate(ws_emp[1], 1)}
lookup_header_map = {c.value: i for i, c in enumerate(ws_lookup[1], 1)}

# Build lookup dictionary: variable -> (category, subcategory)
lookup_dict = {}
for row in ws_lookup.iter_rows(min_row=2, values_only=True):
    var = row[0]
    if not var:
        continue
    cat = row[1]
    subcat = row[2]
    lookup_dict[normalize(var)] = (cat, subcat)

# List of variables to transpose per input example (all in Lookup sheet)
variables = [row[0] for row in ws_lookup.iter_rows(min_row=2, values_only=True) if row[0]]

rows_out = [("seqno", "Empno", "ename", "Jobid", "Variable", "Category", "Subcategory", "Value")]

for emp_row in ws_emp.iter_rows(min_row=2, values_only=True):
    if not emp_row[0]:
        continue
    seqno = emp_row[emp_header_map['seqno']-1]
    empno = emp_row[emp_header_map['Empno']-1]
    ename = emp_row[emp_header_map['ename']-1]
    jobid = emp_row[emp_header_map['Jobid']-1]
    for v in variables:
        v_idx = emp_header_map.get(v)
        if v_idx is None:
            continue
        value = emp_row[v_idx-1]
        cat, subcat = lookup_dict[normalize(v)]
        rows_out.append((seqno, empno, ename, jobid, v, cat, subcat, value))

# Write to Expected sheet
ws_expected = wb["Expected"]

for row_idx, row_data in enumerate(rows_out, start=1):
    for col_idx, v in enumerate(row_data, start=1):
        ws_expected.cell(row=row_idx, column=col_idx, value=v)

wb.save(output_path)
