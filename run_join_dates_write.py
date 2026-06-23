import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_2/group_45896/r3/evolve_45896/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_2/group_45896/r3/evolve_45896/output.xlsx'

# Load dataframes
main_df = pd.read_excel(input_path, sheet_name='Volym P5_P6_2023')
zord_df = pd.read_excel(input_path, sheet_name='ZORD')

# Openpyxl workbook
wb = load_workbook(input_path)
ws = wb['Volym P5_P6_2023']

start_row = 2
end_row = 10
for excel_row in range(start_row, end_row + 1):
    material = ws[f'A{excel_row}'].value
    if material:
        # Find unique 'Valid from' dates for this material
        matches = zord_df[zord_df['Material'] == material]['Valid from']
        unique_dates = pd.to_datetime(matches.dropna().unique(), errors='coerce')
        formatted_dates = [d.strftime('%d/%m/%Y') for d in unique_dates if pd.notnull(d)]
        joined_dates = ','.join(formatted_dates) if formatted_dates else None
        ws.cell(row=excel_row, column=9).value = joined_dates
    else:
        ws.cell(row=excel_row, column=9).value = None

wb.save(output_path)
