import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/group_387-16/r3/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/group_387-16/r3/evolve_387-16/output.xlsx'
sheet_name = 'Sheet1'

# Read raw, no header
raw = pd.read_excel(input_path, sheet_name=sheet_name, header=None)

# Extract main values and binaries (rows 2–14, columns 0 & 1)
main_values = raw.iloc[2:15, 0].tolist()
main_binaries = raw.iloc[2:15, 1].tolist()
# Extract result values (rows 4–14, column 3)
result_values = raw.iloc[4:15, 3].dropna().tolist()
# Remove only one occurrence per result value (from top)
removed_indices = set()
for v in result_values:
    for idx, candidate in enumerate(main_values):
        if candidate == v and idx not in removed_indices:
            removed_indices.add(idx)
            break
filtered_values = [v for i, v in enumerate(main_values) if i not in removed_indices]
filtered_binaries = [b for i, b in enumerate(main_binaries) if i not in removed_indices]
# Pack upward (no blanks)
filtered_values = [v for v in filtered_values if pd.notna(v) and v != '']
filtered_binaries = [b for b in filtered_binaries if pd.notna(b) and b != '']
# Fill up to 13 rows with results, pad blanks
final_values = filtered_values + [''] * (13 - len(filtered_values))
final_binaries = filtered_binaries + [''] * (13 - len(filtered_binaries))
# Solver result = sum of packed Value column
try:
    solver_result = sum([float(v) for v in filtered_values if pd.notna(v) and v!=''])
except Exception as e:
    solver_result = ''
# Target value in cell (1,4)
target_value = raw.iloc[1,4]
difference = solver_result - target_value if solver_result != '' and pd.notna(target_value) else ''
# Load workbook and update A2:D18 as required
wb = load_workbook(input_path)
ws = wb[sheet_name]
for i in range(13):
    ws.cell(row=2+i, column=1).value = final_values[i]
    ws.cell(row=2+i, column=2).value = final_binaries[i]
    ws.cell(row=2+i, column=3).value = ''
    ws.cell(row=2+i, column=4).value = ''
ws['C15'].value = solver_result
ws['D15'].value = difference
wb.save(output_path)
