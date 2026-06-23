import openpyxl

# Load input workbook and worksheet
input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_42198_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_42198_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Logic for C2:C7
for i in range(2, 8):  # rows 2 to 7 inclusive
    found = None
    # Collect cumulative data from rows 2 to current
    for row in ws.iter_rows(min_row=2, max_row=i, min_col=1, max_col=2, values_only=True):
        a_val, b_val = row
        if a_val == 'Potato' and (b_val is False or str(b_val).upper() == 'FALSE'):
            found = 'Worst'
            break  # Priority 1
    else: # Only check these if Potato not found
        for row in ws.iter_rows(min_row=2, max_row=i, min_col=1, max_col=2, values_only=True):
            a_val, b_val = row
            if a_val == 'Tomato' and (b_val is False or str(b_val).upper() == 'FALSE'):
                found = 'Ignore'
                break  # Priority 2
        else:
            for row in ws.iter_rows(min_row=2, max_row=i, min_col=1, max_col=2, values_only=True):
                a_val, b_val = row
                if a_val == 'Pickle' and (b_val is False or str(b_val).upper() == 'FALSE'):
                    found = 'Bad'
                    break  # Priority 3
    
    if not found:
        found = 'Good'
    ws[f'C{i}'] = found

# Save to the output file
wb.save(output_path)
