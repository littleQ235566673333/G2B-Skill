from openpyxl import load_workbook
from datetime import datetime

inp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_4/regression_gate/before_pass/core_41589/input.xlsx'
outp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_4/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(inp)
ws = wb['Contact List']
h = ws['H4'].value
i = str(ws['I4'].value).strip().upper() if ws['I4'].value else ''

result = 'NO ACTION'
# Parse date if present
if h:
    if isinstance(h, datetime):
        date_val = h
    elif isinstance(h, str):
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d'):
            try:
                date_val = datetime.strptime(h, fmt)
                break
            except Exception:
                continue
        else:
            date_val = None
    else:
        date_val = None
    if date_val:
        days_ago = (datetime.now() - date_val).days
        if i == 'YES':
            if days_ago <= 30:
                result = 'HOLD'
            else:
                result = 'TOUCH BASE'

ws['J4'] = result
wb.save(outp)
