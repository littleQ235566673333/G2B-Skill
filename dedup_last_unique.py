import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_4/regression_gate/after_fix/core_280-17/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_4/regression_gate/after_fix/core_280-17/output.xlsx'

# 1. Read the input file and deduplicate
# (Assume the key column is the first column)
df = pd.read_excel(input_path)
keycol = df.columns[0]
deduped = df.drop_duplicates(subset=keycol, keep='last')
deduped.to_excel(output_path, index=False)

# 2. Clear all cells except A1:B12
wb = load_workbook(output_path)
ws = wb.active
for row in ws.iter_rows():
    for cell in row:
        if cell.column > 2 or cell.row > 12:
            cell.value = None
wb.save(output_path)
