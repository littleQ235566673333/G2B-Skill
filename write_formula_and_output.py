import openpyxl

# Load workbook and worksheet
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/regression_gate/before_pass/core_3413/input.xlsx')
ws = wb.active

def get_data_rows():
    # from row 3 onward: Dept (A), RU (B), Total (C)
    rows = []
    for i in range(3, ws.max_row+1):
        dept = ws[f'A{i}'].value
        ru = ws[f'B{i}'].value
        value = ws[f'C{i}'].value
        if dept and ru and (value is not None):
            rows.append({'dept': dept, 'ru': ru, 'value': value})
    return rows

data_rows = get_data_rows()

def compute_result(target_dept, target_ru):
    if not target_dept:
        return None
    if target_ru == 'ALL':
        return sum(row['value'] for row in data_rows if row['dept'] == target_dept)
    # Find exact matches
    filtered_vals = [row['value'] for row in data_rows if row['dept'] == target_dept and row['ru'] == target_ru]
    if len(filtered_vals) > 0:
        return sum(filtered_vals)
    # If not, sum all for dept
    return sum(row['value'] for row in data_rows if row['dept'] == target_dept)

# Fill G3:G6
for i in range(3, 7):
    dept = ws[f'E{i}'].value
    ru = ws[f'F{i}'].value
    val = compute_result(dept, ru)
    ws[f'G{i}'].value = val

wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/regression_gate/before_pass/core_3413/output.xlsx')
