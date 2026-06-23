from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r3/eval_230-16_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r3/eval_230-16_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Before']

for row in range(1, 13):
    val = ws.cell(row=row, column=1).value
    ws.cell(row=row, column=2).value = val
    ws.cell(row=row, column=1).value = None

wb.save(output_path)
