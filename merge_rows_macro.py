import openpyxl
import re
from openpyxl.utils import get_column_letter
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/regression_gate/before_fix/core_177-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_8/regression_gate/before_fix/core_177-6/output.xlsx'

wb = openpyxl.load_workbook(input_path)
src_ws = wb['DATA']
dst_ws = wb['combined']

# Build header mapping
header = next(src_ws.iter_rows(min_row=1, max_row=1, values_only=True))
header_to_idx = {h: i for i, h in enumerate(header)}

# The group key column is H (index 7), header 'ComboKey'
key_idx = header_to_idx['ComboKey']

# Collect all rows and group by ComboKey
rows = list(src_ws.iter_rows(min_row=2, max_row=src_ws.max_row, values_only=False))
groups = defaultdict(list)
for row in rows:
    key = row[key_idx].value
    if key is not None and re.sub(r'\s+', '', str(key)) != '':
        groups[key].append(row)

# Helper to merge group
def merge_group(rows):
    merged = [None]*18  # Columns A:R
    # Copy first non-empty cell for columns A:H
    for i in range(0,8):
        for row in rows:
            val = row[i].value
            if val is not None and re.sub(r'\s+', '', str(val)) != '':
                merged[i] = val
                break
    # For columns I:R (8:17), sum numerics, else first non-empty
    for i in range(8,18):
        total = 0.0
        found = False
        all_zero = True
        for row in rows:
            val = row[i].value
            if isinstance(val, (int, float)):
                if val != 0: all_zero = False
                total += float(val)
                found = True
            elif val is not None and re.sub(r'\s+', '', str(val)) != '':
                try:
                    fval = float(str(val).replace(',',''))
                    if fval != 0: all_zero = False
                    total += fval
                    found = True
                except:
                    if not found: merged[i] = val
                    found = True
        if found:
            if not all_zero:
                merged[i] = round(total, 2)
            else:
                merged[i] = ''
        else:
            merged[i] = ''
    return merged

# Compose result matrix: header + one merged row per key
result_matrix = [header[:18]]
for group_rows in groups.values():
    merged_row = merge_group(group_rows)
    result_matrix.append(merged_row)

# Write to combined sheet, region A1:R8
start_row, start_col = 1, 1
for r_i, row in enumerate(result_matrix[:8]):
    for c_i, val in enumerate(row):
        cell = dst_ws.cell(row=start_row + r_i, column=start_col + c_i, value=val)
        # Format columns I:R if it's a data row
        if r_i > 0 and 8 <= c_i < 18:
            v = val
            if isinstance(v, (int, float)):
                if v == 0:
                    cell.value = ''
                else:
                    cell.value = f'{v:.2f}'
            elif isinstance(v, str):
                try:
                    fval = float(v.replace(',',''))
                    if fval == 0:
                        cell.value = ''
                    else:
                        cell.value = f'{fval:.2f}'
                except:
                    pass

wb.save(output_path)
