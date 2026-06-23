from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_8/regression_gate/after_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_8/regression_gate/after_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']

now = datetime.now()
row = 4
while True:
    last_contact = ws[f'H{row}'].value
    yes_no = ws[f'I{row}'].value

    # Stop if both columns are empty (assume end of data)
    if last_contact is None and yes_no is None:
        break

    action = 'NO ACTION'
    if isinstance(last_contact, datetime) and isinstance(yes_no, str):
        yes_no = yes_no.strip().upper()
        if yes_no == 'YES':
            days_diff = (now - last_contact).days
            if days_diff <= 30:
                action = 'HOLD'
            else:
                action = 'TOUCH BASE'
    ws[f'J{row}'] = action
    row += 1

wb.save(output_path)
