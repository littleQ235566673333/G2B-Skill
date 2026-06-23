from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_2/regression_gate/after_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

REPORT_DATE_CELL = 'B1'
DOB_COL = 3   # 'C'
CATEGORY_COL = 9  # 'I'
EXPECTED_RESULT_COL = 5 # 'E'
AGE_YEAR_COL = 15 # 'O'

start_row = 3
end_row = 15

# Insert new column after 'age (year)' (col O => after col 15)
actual_age_col = AGE_YEAR_COL + 1
ws.insert_cols(actual_age_col)

# Update header for new Actual Age column
ws.cell(row=2, column=actual_age_col).value = 'Actual Age'

# Copy fill from previous column (O), clone it for openpyxl
source_fill = ws.cell(row=2, column=AGE_YEAR_COL).fill
if isinstance(source_fill, PatternFill):
    new_fill = PatternFill(fill_type=source_fill.fill_type, fgColor=source_fill.fgColor, bgColor=source_fill.bgColor)
else:
    new_fill = PatternFill()

# Apply fill and formatting to new column header
ws.cell(row=2, column=actual_age_col).fill = new_fill
ws.cell(row=2, column=actual_age_col).alignment = Alignment(horizontal='center', vertical='top')
ws.cell(row=2, column=actual_age_col).font = Font(bold=True)

# Write formulas and style for rows 3-15
for i in range(start_row, end_row+1):
    # 1. Expected Result: get CATEGORY for this row
    ws.cell(row=i, column=EXPECTED_RESULT_COL).value = f'={get_column_letter(CATEGORY_COL)}{i}'
    
    # 2. Actual Age formula: integer years between REPORT DATE and DOB
    actual_age = ws.cell(row=i, column=actual_age_col)
    formula = f'=INT(($B$1-C{i})/365.25)'
    actual_age.value = formula
    actual_age.font = Font(bold=True)
    actual_age.alignment = Alignment(horizontal='center', vertical='center')
    actual_age.fill = new_fill

wb.save(output_path)
print('Done')
