import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_9/task_44389/r0/evolve_44389/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_9/task_44389/r0/evolve_44389/output.xlsx'

# Columns to check
cols_to_check = ['Amazon', 'HN', 'Catch', 'Costco', 'DJ', 'THX', 'H', 'KWH', 'MBX', 'MMX', 'MY', 'POK', 'VB']
df = pd.read_excel(input_path)

result = []
for idx, row in df.iterrows():
    values = [row[col] for col in cols_to_check]
    # filter for non-null and numbers only, greater than zero
    positives = [(cols_to_check[i], v) for i, v in enumerate(values)
                 if pd.notnull(v) and isinstance(v, (int, float)) and v > 0]
    if positives:
        min_val = min(v for h, v in positives)
        min_headers = [h for h, v in positives if v == min_val]
        result.append(','.join(min_headers))
    else:
        result.append('')

wb = load_workbook(input_path)
ws = wb.active
for i, val in enumerate(result[:6]):  # P2-P7 == first 6 rows
    ws[f'P{i+2}'] = val
wb.save(output_path)
