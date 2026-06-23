from openpyxl import load_workbook

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_7/group_290-1/r0/evolve_290-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_7/group_290-1/r0/evolve_290-1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Infer mapping from K2:U10 examples.
def infer_column_mapping():
    mapping = {}
    # K=11, U=21, examine rows 2-10
    for r in range(2, 11):
        for out_col in range(11, 22):
            val = ws.cell(row=r, column=out_col).value
            if val is not None:
                for i in range(1, 11):
                    if ws.cell(row=r, column=i).value == val:
                        mapping[out_col] = i
    # Guarantee order by out_col
    sorted_map = [mapping.get(col, None) for col in range(11,22)]
    return sorted_map

mapping = infer_column_mapping()  # list of source cols for 11-21 (K-U)

def fill_transposed_data():
    for r in range(2, 2002):  # up to 2000 rows
        for j, src in enumerate(mapping):
            if src is not None:
                ws.cell(row=r, column=11+j).value = ws.cell(row=r, column=src).value

fill_transposed_data()

wb.save(output_path)
