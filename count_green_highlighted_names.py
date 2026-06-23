from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main():
    wb = load_workbook('results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY/eval_43657_tc1/input.xlsx')
    ws = wb['Jul-22 (2)']

    # List of names in L2:L8 (col 12, rows 2-8)
    names = [ws.cell(row=r, column=12).value for r in range(2, 9)]

    # Prepare to count green highlights for each name
    green_rgb = 'FF92D050'
    counts = []

    # Count in C3:H9 (col 3-8, row 3-9, inclusive)
    for name in names:
        count = 0
        for row in range(3, 10):
            for col in range(3, 9):
                cell = ws.cell(row=row, column=col)
                fill = cell.fill
                # Check green highlight (pattern type solid and correct RGB)
                if fill.patternType == 'solid' and hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb == green_rgb:
                    value = str(cell.value) if cell.value else ''
                    # Split cells with multiple names (e.g. 'Alicia, Ally')
                    for part in value.split(','):
                        if part.strip() == name:
                            count += 1
        counts.append(count)

    # Write results to K2:K8 (col 11, rows 2-8)
    for idx, val in enumerate(counts):
        ws.cell(row=2+idx, column=11, value=val)

    wb.save('results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/eval_NOW_TIME_VERIFY/eval_43657_tc1/output.xlsx')

if __name__ == '__main__':
    main()
