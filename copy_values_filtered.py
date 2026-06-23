from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/regression_gate/before_fix/core_48975/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_4/regression_gate/before_fix/core_48975/output.xlsx'

wb = load_workbook(input_path)
ws_in = wb['Input']
ws_out = wb['Output']

target_rows = list(range(11, 18))  # 11-17
copy_rows = []
for row in ws_in.iter_rows(min_row=2):  # Assumes header
    a = row[0].value
    b = row[1].value
    todo = row[4].value
    if todo and str(todo).strip().lower() == 'yes':
        if a is not None and b is not None and str(a).strip() != '' and str(b).strip() != '':
            copy_rows.append((a, b))

for i, (a, b) in enumerate(copy_rows):
    if i >= len(target_rows):
        break
    cell = 'B' + str(target_rows[i])
    ws_out[cell] = f"{a}, {b}"

wb.save(output_path)
