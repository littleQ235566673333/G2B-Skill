from openpyxl import load_workbook
import shutil

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_2/regression_gate/after_pass/core_160-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_2/regression_gate/after_pass/core_160-6/output.xlsx'

shutil.copyfile(input_path, output_path)
wb = load_workbook(output_path)
ws = wb['SH']

start_row = 6
end_row = 11
start_col = 1
end_col = 12  # A=1, L=12

# Read rows in A6:L11
rows = []
for r in range(start_row, end_row + 1):
    values = [ws.cell(row=r, column=c).value for c in range(start_col, end_col + 1)]
    rows.append(values)

# Remove blank rows (where all values are None or empty string)
non_blank_rows = [row for row in rows if any(cell not in (None, '') for cell in row)]

# Overwrite A6:L11 with non-blank rows, blank the rest
for i, r in enumerate(range(start_row, end_row + 1)):
    if i < len(non_blank_rows):
        for c, val in enumerate(non_blank_rows[i], start=1):
            ws.cell(row=r, column=c).value = val
    else:
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None

wb.save(output_path)
print('Done.')
