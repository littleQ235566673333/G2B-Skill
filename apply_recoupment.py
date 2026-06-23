import openpyxl
from openpyxl.styles import Font, Fill, Border, Alignment

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_36097_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_36097_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

for row in range(3, 7):
    cost = ws[f'C{row}'].value
    itv = ws[f'E{row}'].value
    profit = ws[f'F{row}'].value

    # Some values might be formulas (strings), ensure numeric
    cost = float(cost) if cost is not None and not isinstance(cost, str) else None
    itv = float(itv) if itv is not None and not isinstance(itv, str) else None
    profit = float(profit) if profit is not None and not isinstance(profit, str) else None

    if cost is None or itv is None or profit is None:
        recoup = None
    elif profit < 0:
        recoup = itv + profit
    elif 0 <= profit < cost:
        recoup = profit
    elif profit >= cost:
        recoup = cost - itv
    else:
        recoup = None

    # Write result to H and copy formatting from I
    hcell = ws[f'H{row}']
    icell = ws[f'I{row}']
    hcell.value = recoup
    hcell.number_format = icell.number_format
    hcell.font = icell.font.copy()
    hcell.fill = icell.fill.copy()
    hcell.border = icell.border.copy()
    hcell.alignment = icell.alignment.copy()

wb.save(output_path)
