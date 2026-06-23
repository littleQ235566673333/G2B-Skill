import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_7/task_130-9/r1/evolve_130-9/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_7/task_130-9/r1/evolve_130-9/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_src = wb['cdnr']
ws_dst = wb['b2b, sez, de']

def get_headers(ws, search_rows=6):
    for row in ws.iter_rows(min_row=1, max_row=search_rows):
        headers = [cell.value for cell in row]
        nonempty = [h for h in headers if h not in (None, '')]
        if len(nonempty) >= max(3, int(len(headers) * 0.7)):
            return row[0].row, headers
    raise ValueError('No header row found')

src_hdr_row, src_headers = get_headers(ws_src)
dst_hdr_row, dst_headers = get_headers(ws_dst)

src_header_to_col = {str(h): i for i, h in enumerate(src_headers) if h is not None and h != ''}
dst_header_to_col = {str(h): i for i, h in enumerate(dst_headers) if h is not None and h != ''}

start_append_row = 7

src_data_rows = list(ws_src.iter_rows(min_row=src_hdr_row+1, max_row=ws_src.max_row, values_only=True))
src_data_rows = [r for r in src_data_rows if any(x is not None and x != '' for x in r)]

inserted_rows = []
for src_row in src_data_rows:
    new_row = [None] * len(dst_headers)
    for dst_head, dst_col in dst_header_to_col.items():
        if dst_head in src_header_to_col:
            src_val = src_row[src_header_to_col[dst_head]]
            new_row[dst_col] = src_val
    # Amount fields to negative (L and M)
    for amount_col in ['L', 'M']:
        idx = dst_header_to_col.get(amount_col)
        if idx is not None:
            val = new_row[idx]
            if val is not None:
                try:
                    fval = float(val)
                    new_row[idx] = -abs(fval)  # always negative
                except:
                    pass
    # Column H blank, I = 'Credit Note'
    for force_h in ['H', 'h']:
        if force_h in dst_header_to_col:
            new_row[dst_header_to_col[force_h]] = ''
    for force_i in ['I', 'i']:
        if force_i in dst_header_to_col:
            new_row[dst_header_to_col[force_i]] = 'Credit Note'
    inserted_rows.append(new_row)

for row_offset, row_values in enumerate(inserted_rows):
    xl_row = start_append_row + row_offset
    for col_idx, value in enumerate(row_values, 1):
        cell = ws_dst.cell(row=xl_row, column=col_idx)
        fmt = cell.number_format
        cell.value = value
        cell.number_format = fmt

wb.save(output_path)
