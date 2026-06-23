from openpyxl import load_workbook

def find_category(bankdesc, shortnames, categories):
    if not isinstance(bankdesc, str):
        return None
    for shortname, cat in zip(shortnames, categories):
        if shortname and shortname in bankdesc:
            return cat
    return None

def main():
    in_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_524-31_tc1/input.xlsx"
    out_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_524-31_tc1/output.xlsx"
    wb = load_workbook(in_path)
    ws = wb["Exp-DB"]
    # Read all the shortnames and categories first (assume first 36 rows are the reference)
    shortnames = [ws.cell(row=i, column=1).value for i in range(1, 37)]
    categories = [ws.cell(row=i, column=2).value for i in range(1, 37)]
    # Fill E1:E53 with the matched category for the description in D
    for i in range(1, 54):
        desc = ws.cell(row=i, column=4).value
        cat = find_category(desc, shortnames, categories)
        ws.cell(row=i, column=5, value=cat)
    wb.save(out_path)

if __name__ == "__main__":
    main()
