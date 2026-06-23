import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_58032_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_58032_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

arial_font = Font(name='Arial')
bold_italic = Font(name='Arial', bold=True, italic=True)
thin = Side(border_style='thin', color='000000')
title_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Set all cells to Arial
for row in ws.iter_rows():
    for cell in row:
        cell.font = arial_font

# Row 1 (titles): bold, italic, bordered
for cell in ws[1]:
    cell.font = bold_italic
    cell.border = title_border

# Fill column D from D2:Dmax with #CCCCCC
gray_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, max_row=ws.max_row):
    for cell in row:
        cell.fill = gray_fill

# Formula for column A (A2:A35) and fill color #FCD5B4
peach_fill = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')
for r in range(2, 36):
    ws[f'A{r}'].fill = peach_fill
    ws[f'A{r}'].font = arial_font
    # Array formula for searching columns I:J based on Title & Department code (assuming B=Title, C=Dept, I=Combo, J=Name, H=Dept Code)
    ws[f'A{r}'].value = '=INDEX($J$2:$J$100, MATCH(B{r}&C{r}, $I$2:$I$100&$H$2:$H$100, 0))'

# Hide gridlines
ws.sheet_view.showGridLines = False

wb.save(output_path)
