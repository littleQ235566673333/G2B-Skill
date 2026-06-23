import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun1/eval_55468_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/eval_seed42_rerun1/eval_55468_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Step 1: Read user criteria from UI
row_crit_1 = ws['AE4'].value
row_crit_2 = ws['AD4'].value
col_crit_1 = ws['AC4'].value
col_crit_2 = ws['AB4'].value

# Step 2: Layout discovery
# Let's assume per original formula: Rows 5-10 (A5:A10/B5:B10 for left labels), Cols C:Z
lookup_top_row1 = 3
lookup_top_row2 = 4
lookup_first_data_row = 5
# Use all rows until empty in A (as likely real data rows)
data_rows = []
for r in range(lookup_first_data_row, ws.max_row+1):
    if ws[f'A{r}'].value is None:
        break
    data_rows.append(r)
# Left side headers are in A and B
side_header1 = [ws[f'A{r}'].value for r in data_rows]
side_header2 = [ws[f'B{r}'].value for r in data_rows]
# Top headers (C onwards)
col_start = 3
col_end = ws.max_column
col_letters = [openpyxl.utils.get_column_letter(c) for c in range(col_start, col_end+1)]
top_header1 = [ws[f'{col}{lookup_top_row1}'].value for col in col_letters]
top_header2 = [ws[f'{col}{lookup_top_row2}'].value for col in col_letters]
# Find matching data row
row_idx = None
for idx, (s1, s2) in enumerate(zip(side_header1, side_header2)):
    if s1 == row_crit_1 and s2 == row_crit_2:
        row_idx = idx
        break
# Find matching data column
col_idx = None
for idx, (h1, h2) in enumerate(zip(top_header1, top_header2)):
    if h1 == col_crit_1 and h2 == col_crit_2:
        col_idx = idx
        break
result = ''
if row_idx is not None and col_idx is not None:
    dest_row = data_rows[row_idx]
    dest_col = openpyxl.utils.column_index_from_string(col_letters[col_idx])
    result = ws.cell(row=dest_row, column=dest_col).value
ws['AE5'] = result
wb.save(output_path)
