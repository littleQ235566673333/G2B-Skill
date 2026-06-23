from openpyxl import load_workbook

def next_row(ws):
    for r in range(2, ws.max_row+2):
        if all((ws.cell(r, c).value is None for c in range(1, 13))):
            return r
    return ws.max_row+1

# Input/output paths
i_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_6/regression_gate/after_pass/core_80-42/input.xlsx'
o_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_6/regression_gate/after_pass/core_80-42/output.xlsx'
wb = load_workbook(i_path)
ws_out = wb['Consolidate_ALL']
output_headers = [c.value for c in ws_out[1][:12]]
src_sheets = ['Jack', 'Henry', 'Richard']

# Explicit mapping: output header -> source header
header_map = {
    'Date': 'Date',
    'Name': 'Name',
    'Description': 'Description',
    'Parent Node': 'Parent Node',
    'Parent Description': 'Parent Description',
    'CC External Code': 'CC External Code',
    'Available in CORE': 'Available in CORE',
    'OAL CC Type': 'OAL CC Type',
    'CC Default for LOB': 'CC Default for LOB',
    'CC Fiscal Recharge Owner': 'CC Fiscal Recharge Owner',
    'Ancestor List': 'Ancestor List'
    # 'Sheet Name': not in source, set afterwards
}
rows_added = 0

for sheet in src_sheets:
    ws_src = wb[sheet]
    src_headers = [c.value for c in ws_src[1]]
    col_map = [src_headers.index(header_map[h]) for h in output_headers[:-1]]  # Exclude 'Sheet Name'
    for row in ws_src.iter_rows(min_row=2, values_only=False):
        # skip blank rows
        if all((row[idx].value is None for idx in col_map)):
            continue
        r_out = next_row(ws_out)
        for i, dst_col in enumerate(range(1, 12)):
            idx = col_map[i]
            src_cell = row[idx]
            dst_cell = ws_out.cell(r_out, dst_col)
            dst_cell.value = src_cell.value
            dst_cell.data_type = src_cell.data_type
            if hasattr(dst_cell, '_style') and hasattr(src_cell, '_style'):
                dst_cell._style = src_cell._style
        # Fill in sheet name last
        ws_out.cell(r_out, 12).value = sheet
        ws_out.cell(r_out, 12).data_type = 's'
        rows_added += 1

wb.save(o_path)
print('Consolidated', rows_added, 'rows')
