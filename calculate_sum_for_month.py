import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import calendar

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_57590_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_57590_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Get A26 value (could be month name, month+year, or a date)
A26_val = ws['A26'].value
if isinstance(A26_val, str):
    # Try extracting month name
    try:
        # e.g. 'January' or 'January 2023'
        tokens = A26_val.split()
        month_str = tokens[0]
        month_num = list(calendar.month_name).index(month_str)
    except Exception:
        raise ValueError(f'Unexpected value for A26: {A26_val}')
else:
    # If it's a datetime, extract the month
    month_num = A26_val.month

values = []
for row in range(2, ws.max_row + 1):  # Assuming header in row 1
    date_cell = ws.cell(row=row, column=3).value  # Column C
    val_cell = ws.cell(row=row, column=9).value   # Column I
    if date_cell is None or val_cell is None:
        continue
    try:
        the_date = pd.to_datetime(date_cell)
    except Exception:
        continue
    if the_date.month == month_num:
        try:
            values.append(float(val_cell))
        except Exception:
            pass
result = sum(values)
ws['B26'] = result
wb.save(output_path)
print(f'Sum for month {month_num} as in A26 is {result} written to B26.')
