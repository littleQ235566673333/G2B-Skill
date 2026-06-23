import openpyxl

# Load the workbook and select the relevant sheet
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_230-16_tc1/input.xlsx')
ws = wb['Before']

# For rows 1 through 12 in column A
for row in ws.iter_rows(min_row=1, max_row=12, min_col=1, max_col=1):
    for cell in row:
        val = cell.value
        substr = ''
        if isinstance(val, str):
            # Find substring starting from the first letter
            for idx, char in enumerate(val):
                if char.isalpha():
                    substr = val[idx:]
                    break
            ws.cell(row=cell.row, column=2, value=substr)  # Paste substring to column B
            ws.cell(row=cell.row, column=1, value=val.replace(substr, '', 1)) # Delete substring from A
        else:
            ws.cell(row=cell.row, column=2, value=None)

# Save the modified workbook
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_230-16_tc1/output.xlsx')
