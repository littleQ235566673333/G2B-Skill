import openpyxl
from collections import Counter

# Input and output file paths
data_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_50971_tc1/input.xlsx'
out_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_50971_tc1/output.xlsx'

# Load workbook and active sheet
wb = openpyxl.load_workbook(data_file)
ws = wb.active

# Read column A and B values (assuming header in row 1)
col_a = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1)]
col_b = [row[1].value for row in ws.iter_rows(min_row=2, max_col=2)]

# Find duplicates in column A
counter = Counter(col_a)
duplicates = {val for val, cnt in counter.items() if cnt > 1}

# Extract values from column B where column A is a duplicate
result = [b for a, b in zip(col_a, col_b) if a in duplicates]

# Output results to column G, rows 3 to 13 (one per row)
for i, value in enumerate(result[:11], start=3):  # up to G13
    ws.cell(row=i, column=7, value=value)  # column G is 7

# Save to output Excel file
wb.save(out_file)