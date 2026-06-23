import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_1/group_516-46/r1/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_1/group_516-46/r1/evolve_516-46/output.xlsx'
sheet_name = 'ورقة1'

# Read in the relevant columns (A to E)
df = pd.read_excel(input_path, sheet_name=sheet_name, usecols='A:E')
df = df.dropna(subset=['brand', 'date'])
df['date'] = pd.to_datetime(df['date'])

# For brands' latest date, extract and combine duplicate quantities
last_date = df['date'].max()
df_last = df[df['date'] == last_date]

# Sum quantity for each brand on the last date, keeping first batch and origin
result = df_last.groupby(['brand', 'date'], as_index=False).agg({'quantity':'sum', 'batch':'first', 'origin':'first'})
result = result[['brand','batch','origin','date','quantity']]

wb = load_workbook(input_path)
ws = wb[sheet_name]

# Write result starting from H2 (Row 2, Col 8)
for i, row in enumerate(result.values):
    for j, val in enumerate(row):
        ws.cell(row=2+i, column=8+j, value=val)

wb.save(output_path)
