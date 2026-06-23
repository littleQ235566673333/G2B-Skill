import openpyxl

def find_matching_row(ws, grade, flute):
    # Rows 5-10 correspond to A5:A10 (Grade) and B5:B10 (Flute)
    for row in range(5, 11):
        grade_cell = ws.cell(row=row, column=1).value  # Col A
        flute_cell = ws.cell(row=row, column=2).value  # Col B
        if grade_cell == grade and flute_cell == flute:
            return row
    return None

def find_matching_col(ws, qty_code, group):
    # Cols C to Z correspond to headers: C4:Z4 (qty code), C3:Z3 (group)
    for col in range(3, 27):  # C=3 to Z=26
        qty_code_header = ws.cell(row=4, column=col).value
        group_header = ws.cell(row=3, column=col).value
        if qty_code_header == qty_code and group_header == group:
            return col
    return None

# Load the workbook and sheet
input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_55468_tc1/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_55468_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_fp)
ws = wb.active

# Get criteria from the worksheet
qty_code = ws['AC4'].value
qty_group = ws['AB4'].value
grade = ws['AE4'].value
flute = ws['AD4'].value

row = find_matching_row(ws, grade, flute)
col = find_matching_col(ws, qty_code, qty_group)

result = None
if row and col:
    result = ws.cell(row=row, column=col).value

ws['AE5'] = result
wb.save(output_fp)