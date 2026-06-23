import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_42902_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_42902_tc1/output.xlsx'

# Load workbook and active sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Read all values from the first column
values = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]

# Process in groups of 3 rows, writing them to D:F columns
row_out = 1
for i in range(0, 21, 3):
    group = values[i:i+3]
    if len(group) == 3:
        ws.cell(row=row_out, column=4).value = group[0]  # Column D
        ws.cell(row=row_out, column=5).value = group[1]  # Column E
        ws.cell(row=row_out, column=6).value = group[2]  # Column F
        row_out += 1

wb.save(output_path)
