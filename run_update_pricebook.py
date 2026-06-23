import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_57743_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42_rerun2/eval_57743_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Create model-to-price lookup from columns C and D
model_to_price = {}
for row in range(2, ws.max_row + 1):
    model = ws[f'C{row}'].value
    price = ws[f'D{row}'].value
    if model is not None:
        model_to_price[model] = price

# For each model in A2:A19, write matched price to B2:B19, blank if not found
for row in range(2, 20):
    model = ws[f'A{row}'].value
    price = model_to_price.get(model, '')
    ws[f'B{row}'] = price

wb.save(output_path)
