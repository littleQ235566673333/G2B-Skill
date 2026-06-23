import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_3/regression_gate/after_pass/core_3413/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_3/regression_gate/after_pass/core_3413/output.xlsx'

# Read real data from row 2 forward
df = pd.read_excel(input_path, header=1)

wb = load_workbook(input_path)
ws = wb.active

# For G3-G6, columns E and F in Excel (zero based = 4,5), get values
for row in range(3, 7):
    dept = ws.cell(row=row, column=5).value  # E
    ru   = ws.cell(row=row, column=6).value  # F

    # sum all for department+ru if it exists, else sum for department
    mask = (df['Dept'] == dept) & (df['RU'] == ru)
    if df.loc[mask, 'Total'].notnull().any():
        val = df.loc[mask, 'Total'].fillna(0).sum()
    else:
        val = df.loc[df['Dept'] == dept, 'Total'].fillna(0).sum()
    ws.cell(row=row, column=7, value=val)  # write to G

wb.save(output_path)
