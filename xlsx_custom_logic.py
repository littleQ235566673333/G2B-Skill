import openpyxl
from collections import defaultdict

wb_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_1/regression_gate/before_pass/core_55421/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_1/regression_gate/before_pass/core_55421/output.xlsx'
wb = openpyxl.load_workbook(wb_path)
ws = wb.active

# Gather all relevant rows and their values
rows = []
for row in ws.iter_rows(min_row=2, max_row=20, min_col=1, max_col=5, values_only=True):
    rows.append(list(row))

# Group row indices by value in Column A
number_to_indices = defaultdict(list)
for idx, row in enumerate(rows):
    number = row[0]
    number_to_indices[number].append(idx)

# Store results for each row
results = [''] * len(rows)
for number, indices in number_to_indices.items():
    statuses = [rows[i][3] for i in indices]
    dates = [rows[i][4] for i in indices]
    # Case detection
    only_sch = all(s == 'SCH' for s in statuses)
    has_no_show = any(s == 'NO SHOW' for s in statuses)
    has_sch = any(s == 'SCH' for s in statuses)
    for i in indices:
        status = rows[i][3]
        date_val = rows[i][4]
        # (1) NO SHOW + blank date
        if status == 'NO SHOW' and (date_val is None or str(date_val).strip() == ''):
            results[i] = 'CALL PT'
        # (2) NO SHOW + date
        elif status == 'NO SHOW' and date_val not in [None, '']:
            results[i] = 'NO ACTION NEEDED'
        # (3) Exclusively SCH for this number
        elif only_sch:
            results[i] = 'FUTURE'
        # (4) Both SCH and NO SHOW present
        elif has_sch and has_no_show:
            results[i] = 'NS/SCHED'
        else:
            results[i] = ''
# Write to F2:F20
for idx, val in enumerate(results):
    ws.cell(row=idx + 2, column=6).value = val
wb.save(out_path)
