import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/regression_gate/after_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_7/regression_gate/after_pass/core_18935/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Data Table locations
DATA_HEADER_ROW = 4
DATA_START_ROW = 5
DATA_END_ROW = 10  # Up to before empty rows
DATA_WORK_COL = 1  # 'Work'
DATA_MATERIAL_COL = 2  # 'Material'
DATA_CATEGORY1_COL = 3  # 'Category 1'
DATA_CATEGORY2_COL = 4  # 'Category 2'
DATA_CATEGORY3_COL = 5  # 'Category 3'
CATEGORY_COLUMNS = {
    'Category 1': DATA_CATEGORY1_COL,
    'Category 2': DATA_CATEGORY2_COL,
    'Category 3': DATA_CATEGORY3_COL
}

# Report starts from row 17, columns: Work (A), Category Type (B), Material (C)
for i in range(17, 23):  # D17:D22
    work = ws.cell(row=i, column=1).value
    cat_type = ws.cell(row=i, column=2).value
    material = ws.cell(row=i, column=3).value

    # Identify the correct Category column
    category_col = CATEGORY_COLUMNS.get(cat_type)
    category_value = None
    # Search data table
    for j in range(DATA_START_ROW, DATA_END_ROW + 1):
        data_work = ws.cell(row=j, column=DATA_WORK_COL).value
        data_material = ws.cell(row=j, column=DATA_MATERIAL_COL).value
        if data_work == work and data_material == material:
            category_value = ws.cell(row=j, column=category_col).value
            break
    # Write value to D column
    ws.cell(row=i, column=4, value=category_value)

wb.save(output_path)
