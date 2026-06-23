import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed2/eval_472-15_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed2/eval_472-15_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
val = ws['A1'].value

mapping = {
    '4Ozark': 1,
    '3Tall': 2,
    '1Jasper': 3,
    '2GWood': 4,
    '5Dawson': 5,
    '8CPark': 6
}

ws['B2'] = mapping.get(val, '')
wb.save(output_path)
