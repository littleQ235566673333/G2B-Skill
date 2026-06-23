from openpyxl import load_workbook
ws = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_1/regression_gate/after_pass/core_9726/input.xlsx')['Sheet1']
for row in ws.iter_rows(min_row=1, max_row=10):
    print([repr(c.value) for c in row])
