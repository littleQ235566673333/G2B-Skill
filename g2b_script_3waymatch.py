import pandas as pd
from openpyxl import load_workbook, styles

# Input and output paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_2/regression_gate/before_fix/core_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_2/regression_gate/before_fix/core_57033/output.xlsx'

# Load data
wb = load_workbook(input_path)
sheet4 = wb['Sheet4']
df4 = pd.read_excel(input_path, sheet_name='Sheet4')
cbtrans = pd.read_excel(input_path, sheet_name='CBtrans')

# Prepare cell style for fill and font
target_fill = styles.PatternFill(fill_type='solid', fgColor='FF66CC')
target_font = styles.Font(bold=False)

# For the first 6 relevant rows (2:7, 0-based index 0:6)
results = []
for i, row in df4.iloc[0:6].iterrows():
    matches = cbtrans[(cbtrans['company'] == row['Company']) & (cbtrans['account'] == row['account']) & (cbtrans['xchar'] == row['xchar'])]
    results.append('Match' if not matches.empty else '-')
    
# Write results and styling
for idx, value in enumerate(results, start=2):  # K2:K7
    cell = sheet4[f'K{idx}']
    cell.value = value
    cell.fill = target_fill
    cell.font = target_font

# Save
wb.save(output_path)
