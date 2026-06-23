import pandas as pd
from openpyxl import load_workbook

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_5/group_263-1/r1/evolve_263-1/input.xlsx'
df = pd.read_excel(input_fp)

# Use correct column names found: Mtrl, Width, Height
df = df.dropna(subset=["Mtrl", "Width", "Height"])  # Only keep valid rows
df['Area'] = df['Width'] * df['Height']
agg = df.groupby('Mtrl', as_index=False)['Area'].sum()

wb = load_workbook(input_fp)
ws = wb.active
for idx, row in agg.iterrows():
    if idx >= 3:
        break
    cell = f'H{2 + idx}'
    ws[cell] = f"{row['Mtrl']}: {row['Area']}"

output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_5/group_263-1/r1/evolve_263-1/output.xlsx'
wb.save(output_fp)
