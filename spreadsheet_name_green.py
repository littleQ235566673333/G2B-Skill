from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_43657_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_43657_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def is_green(cell):
    fill = cell.fill
    if fill and fill.patternType == 'solid':
        fg_color = fill.fgColor.rgb
        if fg_color in (None, '00000000'):
            return False
        green_hexes = ["FF00FF00", "FF92D050", "FF00B050"] # common green fills
        return fg_color.upper() in green_hexes
    return False

# List of names in L2:L8
names = [ws[f'L{row}'].value for row in range(2, 9)]
# Range to check: C2:G8
result_rows = range(2, 9)
check_cols = range(3, 8) # C(3)-G(7)
for idx, (name, row) in enumerate(zip(names, result_rows), start=2):
    count = 0
    if name is not None:
        for col in check_cols:
            cell = ws.cell(row=row, column=col)
            if cell.value and name in str(cell.value) and is_green(cell):
                count += 1
    ws[f'K{row}'].value = count
wb.save(output_path)
