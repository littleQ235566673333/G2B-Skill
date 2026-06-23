from openpyxl import load_workbook

def get_category_column(category_type):
    mapping = {'Category 1': 2, 'Category 2': 3, 'Category 3': 4}  # zero-based
    return mapping.get(category_type, None)

# Load workboook
wb_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_2/regression_gate/before_pass/core_18935/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_2/regression_gate/before_pass/core_18935/output.xlsx'
wb = load_workbook(wb_path)
ws = wb['Sheet1']

# Step 1: Read Data Table into a list of dicts (from row 5)
data_table = []
for row in ws.iter_rows(min_row=5, max_row=10, min_col=1, max_col=5, values_only=True):
    work, material, *category_values = row
    data_table.append({'Work': work, 'Material': material, 'Category 1': category_values[0], \
                      'Category 2': category_values[1], 'Category 3': category_values[2]})

# Step 2: For each report row (rows 17-22), perform lookup
for idx in range(17, 23):
    work = ws[f'A{idx}'].value
    category_type = ws[f'B{idx}'].value
    material = ws[f'C{idx}'].value
    # Lookup
    match = next((row for row in data_table if row['Work'] == work and row['Material'] == material), None)
    cat_val = None
    if match:
        cat_val = match.get(category_type, None)
    ws[f'D{idx}'] = cat_val

# Save output
wb.save(out_path)
