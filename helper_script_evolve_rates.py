import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_2/group_44017/r0/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_2/group_44017/r0/evolve_44017/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active
# Get all month dates from AD9:AO9 (columns 30 to 41, openpyxl is 1-based)
col_months = [ws.cell(row=9, column=col).value for col in range(30, 42)]
for row in range(14, 43):
    base = ws[f'W{row}'].value
    freq = ws[f'J{row}'].value
    eff_date = ws[f'L{row}'].value
    incs = [ws[f'{col}{row}'].value for col in ['M','N','O','P']]
    # Check for critical values missing or improper
    if base is None or eff_date is None or freq is None:
        for colidx in range(30, 42):
            ws.cell(row=row, column=colidx).value = None
        continue
    try:
        freq_int = int(freq)
    except Exception:
        for colidx in range(30, 42):
            ws.cell(row=row, column=colidx).value = None
        continue
    for ci, col_date in enumerate(col_months):
        out_col = 30 + ci
        cell = ws.cell(row=row, column=out_col)
        # If the period is before effective, blank
        if not isinstance(col_date, (datetime.date, datetime.datetime)) or col_date < eff_date:
            cell.value = None
            continue
        # Find how many increases have phased in
        month_delta = (col_date.year - eff_date.year)*12 + (col_date.month - eff_date.month)
        inc_waves = 1 + (month_delta // freq_int)
        # Clamp to number of available increases
        cum_inc = 1.0
        for inc_i in range(min(inc_waves, 4)):
            try:
                inc_pct = float(incs[inc_i])
                if inc_pct:
                    cum_inc *= (1 + inc_pct)
            except (TypeError, ValueError):
                pass
        cell.value = base * cum_inc if base != 0 else 0
wb.save(output_path)
