import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun1/eval_30930_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun1/eval_30930_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read values from A2:A66 and B2:B66
values_a = [ws[f'A{row}'].value for row in range(2, 67)]
values_b = [ws[f'B{row}'].value for row in range(2, 67)]

result_c = [''] * 65  # For 65 rows (C2:C66)

start = 0
for i in range(65):
    if values_b[i] == 1:
        # Count number of values > 0 in values_a[start:i] (not including i)
        count = sum(1 for x in values_a[start:i] if x is not None and x > 0)
        result_c[i] = count
        start = i+1

# Write result in C2:C66
for idx, val in enumerate(result_c, 2):
    ws[f'C{idx}'] = val if val != '' else None

wb.save(output_path)
