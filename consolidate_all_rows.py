import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_8/regression_gate/after_pass/core_80-42/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_8/regression_gate/after_pass/core_80-42/output.xlsx'
wb = openpyxl.load_workbook(input_path)
main_ws = wb['Consolidate_ALL']
main_headers = [cell.value for cell in main_ws[1][:12]]
src_sheets = ['Jack', 'Henry', 'Richard']
header_map_by_sheet = {}
for s_name in src_sheets:
    ws = wb[s_name]
    header_map = {cell.value: idx for idx, cell in enumerate(ws[1])}
    header_map_by_sheet[s_name] = header_map
# Find first blank row in Consolidate_ALL from row 2
row_idx = 2
while main_ws.cell(row=row_idx, column=1).value is not None and row_idx <= 8000:
    row_idx += 1
# Reference 2nd row for formatting
ref_row_idx = 2 if main_ws.max_row>=2 else 1
for s_name in src_sheets:
    ws = wb[s_name]
    header_map = header_map_by_sheet[s_name]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row_idx > 8000:
            break
        if all(cell.value is None for cell in row):
            continue
        for col_idx, main_header in enumerate(main_headers, 1):
            if main_header in header_map and header_map[main_header] < len(row):
                src_cell = row[header_map[main_header]]
                dest_cell = main_ws.cell(row=row_idx, column=col_idx)
                dest_cell.value = src_cell.value
                dest_cell.data_type = src_cell.data_type
                # Match number_format, but match fill to Consolidate_ALL's 2nd row (to avoid source fills)
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(main_ws.cell(row=ref_row_idx, column=col_idx).fill) # Only fill from target
                    dest_cell.number_format = src_cell.number_format
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)
        # Set source sheet name in L (12)
        main_ws.cell(row=row_idx, column=12).value = s_name
        # Match L fill to rest
        main_ws.cell(row=row_idx, column=12).fill = copy(main_ws.cell(row=ref_row_idx, column=12).fill)
        row_idx += 1
wb.save(output_path)
print('[DONE]')
