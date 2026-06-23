from openpyxl import load_workbook
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42/eval_42198_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42/eval_42198_tc1/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active  # assume active sheet
for r in range(2, 8):
    found = None
    for priority, label in [("Potato", "Worst"), ("Tomato", "Ignore"), ("Pickle", "Bad")]:
        for i in range(2, r+1):
            a = ws['A{}'.format(i)].value
            b = ws['B{}'.format(i)].value
            is_false = (b is False) or (isinstance(b, str) and b.strip().upper() == 'FALSE')
            if a == priority and is_false:
                found = label
                break
        if found:
            break
    ws['C{}'.format(r)] = found if found else 'Good'
wb.save(output_path)
