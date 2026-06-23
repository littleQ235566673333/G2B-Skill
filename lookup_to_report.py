from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/regression_gate/before_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/regression_gate/before_pass/core_18935/output.xlsx'

wb = load_workbook(input_path)
sheet = wb.active

# Data Table header is at row 4
category_types = ['Category 1', 'Category 2', 'Category 3']
category_col_lookup = {'Category 1': 2, 'Category 2': 3, 'Category 3': 4}  # relative to data_cols
# Build lookup dict: key = (Work, Material, Category Type) => value
lookup_dict = {}
data_table_rows = range(5, 11)  # inclusive of row 10, as per dump
for row in data_table_rows:
    work = sheet[f'A{row}'].value
    material = sheet[f'B{row}'].value
    for idx, cat in enumerate(category_types):
        value = sheet.cell(row=row, column=3+idx).value  # columns C,D,E => 3,4,5
        key = (work, material, cat)
        lookup_dict[key] = value

# Report section starts row 17
for row in range(17, 23):
    work = sheet[f'A{row}'].value
    category_type = sheet[f'B{row}'].value
    material = sheet[f'C{row}'].value
    value = lookup_dict.get((work, material, category_type), None)
    sheet[f'D{row}'] = value

wb.save(output_path)
