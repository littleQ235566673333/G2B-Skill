import openpyxl
from collections import defaultdict

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/after_fix/core_263-1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/regression_gate/after_fix/core_263-1/output.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Updated headers based on actual sheet
headers = ['Mtrl', 'Width', 'Height']

def find_col_indices(ws, headers):
    col_map = {}
    for col in ws.iter_cols(1, ws.max_column):
        header = col[0].value
        if header in headers:
            col_map[header] = col[0].column
    return col_map

col_indices = find_col_indices(ws, headers)

mat_type_col = col_indices['Mtrl']
width_col = col_indices['Width']
height_col = col_indices['Height']

# Calculate total area for each material type
totals = defaultdict(float)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    material = row[mat_type_col-1].value
    width = row[width_col-1].value
    height = row[height_col-1].value
    if material and width and height:
        try:
            area = float(width) * float(height)
            totals[material] += area
        except Exception:
            continue

# Display dynamic results in H2, H3, H4 as example requires
result_cells = ['H2', 'H3', 'H4']
materials_sorted = sorted(totals.keys())
for idx, mat in enumerate(materials_sorted):
    if idx < len(result_cells):
        ws[result_cells[idx]] = f'{mat}: {totals[mat]:.2f}'

# Save the workbook
wb.save(output_path)
