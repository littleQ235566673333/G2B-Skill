import pandas as pd
from openpyxl import load_workbook

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_1/regression_gate/before_pass/core_9726/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_1/regression_gate/before_pass/core_9726/output.xlsx'

# Read the only sheet
sheet = pd.read_excel(input_fp, sheet_name='Sheet1')

# Table 1
visit_ids = sheet['Student ID']
visit_dates = pd.to_datetime(sheet['Visit date'], errors='coerce')

# Table 2
sub_ids = sheet['Student ID.1']
sub_dates = pd.to_datetime(sheet['Submission date'], errors='coerce')

results = []
for sid, sdt in zip(sub_ids, sub_dates):
    # Find matching visits with date < submission
    mask = (visit_ids == sid) & (visit_dates < sdt)
    prior_visits = visit_dates[mask]
    val = prior_visits.max() if not prior_visits.empty else None
    results.append(val)

# Write result to 'Visit date.1' column (which is column J, index 9), data starts from row 2
wb = load_workbook(input_fp)
ws = wb['Sheet1']
for idx, val in enumerate(results, start=2):
    ws.cell(row=idx, column=10, value=val if pd.isna(val) == False else None)
wb.save(output_fp)
