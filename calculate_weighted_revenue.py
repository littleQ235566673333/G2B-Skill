import openpyxl

# File paths
dir_prefix = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_18645_tc1/'
input_path = dir_prefix + 'input.xlsx'
output_path = dir_prefix + 'output.xlsx'

wb = openpyxl.load_workbook(input_path)
opps_sheet = wb['Opportunities']
data_sheet = wb['Data']

# Read stage-to-percentage mapping
stage_percent = {}
for row in data_sheet.iter_rows(min_row=2):
    stage = row[0].value
    percent = row[1].value
    if stage and percent:
        stage_percent[str(stage).strip()] = float(percent)

# Use correct column indices from printed headers
# Pipeline Stage = 4, Est. Revenue (£) = 5, Weighted Value (£) = 7 (1-based for openpyxl)
stage_col = 4
revenue_col = 5
weighted_col = 7

# Process rows 2-3 (G2:G3)
for row_num in range(2, 4):
    stage = opps_sheet.cell(row=row_num, column=stage_col).value
    revenue = opps_sheet.cell(row=row_num, column=revenue_col).value
    percent = stage_percent.get(str(stage).strip(), 0)
    weighted = revenue * percent if revenue is not None else None
    opps_sheet.cell(row=row_num, column=weighted_col).value = weighted

wb.save(output_path)
