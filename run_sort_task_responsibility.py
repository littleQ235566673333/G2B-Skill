import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun1/eval_262-17_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun1/eval_262-17_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheet = wb['Sheet1']
rows = list(sheet.iter_rows(values_only=True))
header = rows[0]
task_idx = header.index('Task')
resp_idx = header.index('Responsibility')
data = rows[1:]
sorted_data = sorted(data, key=lambda x: (x[task_idx], x[resp_idx]))

# Write header and sorted data back, preserving range A1:F14 (header + 13 rows)
for r, row in enumerate([header] + sorted_data[:13], start=1):
    for c, val in enumerate(row[:6], start=1):
        sheet.cell(row=r, column=c, value=val)
wb.save(output_path)
