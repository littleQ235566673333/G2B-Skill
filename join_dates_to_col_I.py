from openpyxl import load_workbook
from datetime import datetime

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_2/group_45896/r3/evolve_45896/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_2/group_45896/r3/evolve_45896/output.xlsx'

wb = load_workbook(INPUT)
ws = wb['Volym P5_P6_2023']
ws_z = wb['ZORD']

materials = [ws[f'A{r}'].value for r in range(2, 11)]
zord_rows = list(ws_z.iter_rows(min_row=2, values_only=True))

for idx, mat in enumerate(materials):
    dates = [row[1] for row in zord_rows if row[0] == mat and isinstance(row[1], datetime)]
    dates_fmt = [d.strftime('%d/%m/%Y') for d in dates]
    joined = ','.join(dates_fmt)
    ws.cell(row=2 + idx, column=9, value=joined)

wb.save(OUTPUT)
