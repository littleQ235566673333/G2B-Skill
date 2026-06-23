from openpyxl import load_workbook, utils

wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42/eval_120-24_tc1/input.xlsx')
ws = wb['Sheet1']

header_data = []
for i in range(1, 4):
    row_values = [ws.cell(row=i, column=j).value for j in range(utils.column_index_from_string('AX'), utils.column_index_from_string('BN')+1)]
    header_data.append(row_values)
for idx, row in enumerate(header_data, 1):
    print(f'Row {idx}:', row)
