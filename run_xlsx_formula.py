from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_2/group_45707/r0/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-v2-smoke/train/iter_2/group_45707/r0/evolve_45707/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

# Read dates and C values
all_dates = []
all_c_vals = []
for row in ws.iter_rows(min_row=2, max_row=69, min_col=1, max_col=3, values_only=True):
    dt, _, c_val = row
    all_dates.append(dt)
    all_c_vals.append(c_val)

def count_ones_in_month(year, month):
    total = 0
    for dt, c_val in zip(all_dates, all_c_vals):
        if isinstance(dt, datetime) and dt.year == year and dt.month == month and c_val == 1:
            total += 1
    return total

for i in range(2, 70):  # D2:D69
    this_row_next_date = ws.cell(row=i+1, column=1).value
    d_cell = ws.cell(row=i, column=4)
    if isinstance(this_row_next_date, datetime) and this_row_next_date.day == 1:
        # The next row is 1st of month; count 1's in that month/year
        d_cell.value = count_ones_in_month(this_row_next_date.year, this_row_next_date.month)
    else:
        d_cell.value = None  # Leave cell empty

wb.save(output_path)
