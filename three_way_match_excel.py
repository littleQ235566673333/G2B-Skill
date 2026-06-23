import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_2/group_57033/r2/evolve_57033/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-S-smoke16/train/iter_2/group_57033/r2/evolve_57033/output.xlsx'

# Read relevant sheets
sheet4 = pd.read_excel(input_path, sheet_name='Sheet4')
md = pd.read_excel(input_path, sheet_name='CBtrans')

# Harmonize columns for matching
sheet4 = sheet4.rename(columns={'Company': 'company'})
key_fields = ['company', 'account', 'xchar']

sheet4_lower = sheet4.copy()
md_lower = md.copy()
for col in key_fields:
    sheet4_lower[col] = sheet4_lower[col].astype(str).str.lower()
    md_lower[col] = md_lower[col].astype(str).str.lower()

def three_way_match(row):
    mask = (
        (md_lower['company'] == row['company']) &
        (md_lower['account'] == row['account']) &
        (md_lower['xchar'] == row['xchar'])
    )
    return 'Match' if mask.any() else '-'

# Only checking/placing results in K2:K7 (rows index 1 to 6)
match_rows = sheet4_lower.iloc[1:7]
results = match_rows.apply(three_way_match, axis=1).tolist()

# Write and format with openpyxl
wb = load_workbook(input_path)
ws = wb['Sheet4']
fill = PatternFill(fill_type='solid', fgColor='FF66CC')
for i, val in enumerate(results, start=2):  # K2:K7
    cell = ws[f'K{i}']
    cell.value = val.title() if val == 'Match' else val
    cell.fill = fill

wb.save(output_path)
