import openpyxl
import os
from openpyxl.utils import get_column_letter

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/regression_gate/before_fix/core_146-49/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/regression_gate/before_fix/core_146-49/output.xlsx'
vowels = set('AEIOU')

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read all non-empty string words from the sheet (case-insensitive, store uppercase) (all columns)
words = []
for row in ws.iter_rows(min_row=1):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            word = cell.value.strip().upper()
            if word:
                words.append(word)

# Remove duplicates
words = sorted(set(words))
word_set = set(words)
pairs = set()

def get_vowel_replacements(base, s_indices):
    results = []
    for v in vowels:
        res = list(base)
        for idx in s_indices:
            res[idx] = v
        new_word = ''.join(res)
        if new_word != base:
            results.append(new_word)
    return results

for word in words:
    s_indices = [i for i, c in enumerate(word) if c == 'S']
    if s_indices:
        vowel_versions = get_vowel_replacements(word, s_indices)
        for vword in vowel_versions:
            if vword in word_set:
                pair = tuple(sorted([word, vword]))
                pairs.add(pair)

# Write pairs to G2:H...
pairs = sorted(list(pairs), key=lambda x: (x[0].lower(), x[1].lower()))
start_row = 2
col_g = 7
col_h = 8
for i, (w1, w2) in enumerate(pairs):
    ws.cell(row=start_row + i, column=col_g, value=w1)
    ws.cell(row=start_row + i, column=col_h, value=w2)

# Sort G2:H... by both cols, move to G1:H(1+N)
sort_rows = []
for i in range(start_row, start_row + len(pairs)):
    g = ws.cell(row=i, column=col_g).value
    h = ws.cell(row=i, column=col_h).value
    sort_rows.append((g, h))
sort_rows.sort(key=lambda x: (x[0].lower(), x[1].lower()))
for idx, (g, h) in enumerate(sort_rows):
    ws.cell(row=1 + idx, column=col_g).value = g
    ws.cell(row=1 + idx, column=col_h).value = h

wb.save(output_path)
print(f'Wrote to {output_path}')
