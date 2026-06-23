import openpyxl
from collections import defaultdict

# Suffix sort order list (priority order)
SUFFIX_ORDER = ['ING', 'ERS', 'ATE', 'EST', 'ONE', 'IER', 'ILY']

def custom_suffix_sort(w):
    end = w[-3:]
    try:
        num = SUFFIX_ORDER.index(end)
    except ValueError:
        num = 99  # Put others at end
    return (num, end, w)

def transform_word(word):
    # Move the 5th letter to the front, keep other letters in order
    if len(word) < 5:
        return None  # Not enough letters
    return word[4] + word[:4] + word[5:]

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_8/group_118-50/r2/evolve_118-50/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_8/group_118-50/r2/evolve_118-50/output.xlsx"

wb = openpyxl.load_workbook(input_path)
ws = wb["Sheet1"]

# Read column A into a list and filter out None/empty strings
words = [row[0].value for row in ws.iter_rows(min_row=2, min_col=1, max_col=1) if row[0].value and isinstance(row[0].value,str)]

# Sort names in column A and write back
words_sorted = sorted(words, key=lambda w: w.upper())
for i, w in enumerate(words_sorted, start=2):
    ws.cell(row=i, column=1, value=w)

# Find candidate words by suffix
groups = defaultdict(list)
for w in words_sorted:
    if len(w) >= 8:  # Must be long enough for transformation
        suffix = w[-3:]
        for test_suffix in SUFFIX_ORDER:
            if suffix == test_suffix:
                transformed = transform_word(w)
                if transformed and transformed != w:
                    groups[suffix].append((transformed, w))
                break

def sort_grouped_pairs(groups):
    pairs = []
    for suf in SUFFIX_ORDER:
        group = sorted(groups[suf], key=lambda p: (p[1].upper(), p[0].upper()))
        pairs.extend(group)
    return pairs

result_pairs = sort_grouped_pairs(groups)

# Output to C2:D5000 (no header)
for idx, (trans, orig) in enumerate(result_pairs[:4999], start=2):
    ws.cell(row=idx, column=3, value=trans)  # C
    ws.cell(row=idx, column=4, value=orig)   # D
# Clear any remaining cells in C and D past results up to row 5000
for idx in range(len(result_pairs)+2, 5001):
    ws.cell(row=idx, column=3, value=None)
    ws.cell(row=idx, column=4, value=None)

wb.save(output_path)
