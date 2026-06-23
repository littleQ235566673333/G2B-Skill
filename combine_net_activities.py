import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_6/group_60-7/r3/evolve_60-7/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_6/group_60-7/r3/evolve_60-7/output.xlsx'

# Read data
df_exist = pd.read_excel(input_path, sheet_name='Existing Task')
df_add = pd.read_excel(input_path, sheet_name='Additions')
df_retired = pd.read_excel(input_path, sheet_name='Retired')

# Use only columns in 'Existing Task' as the reference (to avoid mismatch issues)
common_cols = df_exist.columns[:5]  # Only need A-E output
# Ensure dataframes align
for df in [df_add, df_retired]:
    if set(df.columns) != set(df_exist.columns):
        df = df[df_exist.columns]

df_add = df_add[common_cols]
df_retired = df_retired[common_cols]

# Step 1. Concatenate Existing Task + Additions
all_rows = pd.concat([df_exist[common_cols], df_add[common_cols]], ignore_index=True)

# Step 2. Remove rows that are RETIRED (match all columns)
# Left join indicator and keep only 'left_only'
if not df_retired.empty:
    left = all_rows.merge(df_retired.drop_duplicates(), how='left', indicator=True)
    result = left[left['_merge'] == 'left_only'][common_cols]
else:
    result = all_rows

# Step 3. Load with openpyxl for output
wb = load_workbook(input_path)
ws = wb['Consolidated Tracker']

start_row = 3
start_col = 1  # Column A
rows_available = 8  # A4 to A11
cols_available = 5  # A-E

# Step 4. Write header in A3:E3
for j, colname in enumerate(common_cols):
    ws.cell(row=start_row, column=start_col + j, value=colname)

# Step 5. Write data starting from A4
for i in range(rows_available):
    if i < len(result):
        row_vals = result.iloc[i].tolist()
        for j, val in enumerate(row_vals):
            ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
    else:
        # Clear cell if no data
        for j in range(cols_available):
            ws.cell(row=start_row + 1 + i, column=start_col + j, value=None)

wb.save(output_path)
