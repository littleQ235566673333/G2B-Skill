import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_1/group_516-46/r3/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_1/group_516-46/r3/evolve_516-46/output.xlsx'
sheet_name = 'ورقة1'

# Read the input data
xl = pd.ExcelFile(input_path)
df = xl.parse(sheet_name)

# Data is assumed in A2:E?, so skip first row header read by pandas, then take rows starting from index 1
# We only consider first 5 columns
columns = df.columns[:5]
data = df.loc[:, columns].iloc[1:].copy()

# Guess columns: will try to identify brand, date, qty
names_l = [str(c).lower() for c in data.columns]
def find_col(sub):
    for i, name in enumerate(names_l):
        if sub in name:
            return data.columns[i]
    return None

date_col = find_col('date') or data.columns[0]
brand_col = find_col('brand') or data.columns[1]
qty_col = find_col('qty') or data.columns[-1]

# Parse date and drop rows missing brand or date

data[date_col] = pd.to_datetime(data[date_col], errors='coerce')
data = data.dropna(subset=[brand_col, date_col])

# Find most recent date per brand
data['last_date'] = data.groupby(brand_col)[date_col].transform('max')

# Only keep those matching brand's latest date
last_rows = data[data[date_col] == data['last_date']]

# Sum quantities per brand for latest date (combine if brand appears more than once on that date)
res = (last_rows.groupby([brand_col, date_col])
    .agg({qty_col: 'sum',
          **{col:'first' for col in data.columns if col not in [brand_col, date_col, qty_col, 'last_date']}})
    .reset_index())

# Optional: show columns in A-E order
show_cols = [col for col in columns if col in res.columns]
if qty_col not in show_cols:
    show_cols.append(qty_col)
res = res[show_cols]

# Write result to worksheet
wb = load_workbook(input_path)
ws = wb[sheet_name]

# Clear H2:L4
for row in range(2,5):
    for col in range(8,13):
        ws.cell(row=row, column=col, value=None)

# Write df to H2 (row=2,col=8)
for r_idx, (_, row) in enumerate(res.iterrows()):
    for c_idx, val in enumerate(row.values):
        ws.cell(row=2+r_idx, column=8+c_idx, value=val)

wb.save(output_path)
