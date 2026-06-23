import openpyxl
from collections import defaultdict
import math

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_3/group_4714/r2/evolve_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_3/group_4714/r2/evolve_4714/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active
# Get all rows as values
rows = list(ws.iter_rows(min_row=2, max_row=25, min_col=1, max_col=4, values_only=True))
# Each row: (Employee, SomeCol, Month, Hours)
employees = defaultdict(list)
for idx, row in enumerate(rows, start=2):
    emp, _, month, hours = row
    if emp is not None and month is not None and hours is not None:
        employees[emp].append({'row_num': idx, 'month': month, 'hours': hours})
# Build a sorted list per employee by month (assume months are sortable)
for emp in employees:
    employees[emp].sort(key=lambda x: x['month'])

# Map: (row num) => moving avg or n/a
results = {}
for emp, records in employees.items():
    for i, rec in enumerate(records):
        window = records[max(0, i-3):i+1]
        if len(window) < 4:
            results[rec['row_num']] = 'n/a'
        else:
            avg = sum(w['hours'] for w in window) / 4
            # Format as int if possible
            avg_disp = str(int(avg)) if math.isclose(avg, int(avg)) else f'{avg:.2f}'
            results[rec['row_num']] = avg_disp

for row in range(2,26):
    val = results.get(row, '')
    ws.cell(row=row, column=5, value=val)

wb.save(output_path)
