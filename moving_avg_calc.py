import openpyxl
import math

def parse_month(cstr):
    cstr = str(cstr)
    # handles YYYY-MM, YYYY/MM, or YYYY.MM, or YYYY M
    for sep in ('-', '/', '.', ' '):
        if sep in cstr:
            y, m = cstr.split(sep)
            return int(y), int(m)
    # fallback for YYYYMM
    y, m = int(cstr[:4]), int(cstr[4:])
    return y, m

def month_to_num(year, month):
    return year * 12 + month

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/after_fix/core_4714/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/after_fix/core_4714/output.xlsx'
wb = openpyxl.load_workbook(input_fp)
ws = wb.active

rows = list(ws.iter_rows(min_row=2, max_col=4, max_row=25, values_only=True))
data = []
for idx, row in enumerate(rows):
    emp, _, month, hours = row
    if emp is None or month is None or hours is None:
        continue
    try:
        year, mon = parse_month(month)
        data.append({
            'row_idx': idx + 2,
            'employee': emp,
            'month': month,
            'year': year,
            'mon': mon,
            'hours': hours,
        })
    except Exception:
        continue

output = [''] * 24
for i, row in enumerate(data):
    emp = row['employee']
    curr_y, curr_m = row['year'], row['mon']
    curr_num = month_to_num(curr_y, curr_m)
    window_months = []
    for j in range(4):
        m = curr_m - j
        y = curr_y
        if m <= 0:
            m += 12
            y -= 1
        window_months.append((y, m))
    window_nums = set(month_to_num(y, m) for y, m in window_months)
    window_rows = [d for d in data if d['employee'] == emp and month_to_num(d['year'], d['mon']) in window_nums and month_to_num(d['year'], d['mon']) <= curr_num]
    if len(window_rows) < 4:
        val = 'n/a'
    else:
        avg = sum(d['hours'] for d in window_rows) / 4
        if math.isclose(avg, round(avg)):
            val = str(int(round(avg)))
        else:
            val = f'{avg:.2f}'.rstrip('0').rstrip('.')
    output[i] = val

for i, val in enumerate(output):
    ws[f'E{i+2}'] = val

wb.save(output_fp)
