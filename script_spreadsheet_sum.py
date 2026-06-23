import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_38823_tc1/input.xlsx')
ws = wb.active

search_terms = [ws[f'H{i}'].value for i in range(4, 8)]
criteria_date_start = ws['F4'].value
criteria_date_end = ws['G4'].value

# If end date is missing, default to latest date in input data (col A, from row 3)
dates = [row[0] for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1, values_only=True) if isinstance(row[0], datetime)]
if not criteria_date_end:
    criteria_date_end = max(dates) if dates else criteria_date_start

sums = []
for term in search_terms:
    s = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3, values_only=True):
        date, text, units = row
        if not (isinstance(date, datetime) and isinstance(units, (int, float)) and isinstance(text, str)):
            continue
        if criteria_date_start <= date <= criteria_date_end and term.lower() in text.lower():
            s += units
    sums.append(s)
for idx, s in enumerate(sums, start=4):
    ws[f'I{idx}'].value = s
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_38823_tc1/output.xlsx')
