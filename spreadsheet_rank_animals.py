from openpyxl import load_workbook

# File paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_51289_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_51289_tc1/output.xlsx"

wb = load_workbook(input_path)
ws = wb.active

# Get the labels and numbers
labels = [ws.cell(row=1, column=i).value for i in range(1, 9)]
numbers = [ws.cell(row=2, column=i).value for i in range(1, 9)]

results = ["" for _ in range(8)]

def find_ranking(animal):
    filtered = [(idx, num) for idx, (lbl, num) in enumerate(zip(labels, numbers)) if lbl == animal]
    # Sort descending, take up to two
    sorted_by_value = sorted(filtered, key=lambda x: x[1], reverse=True)
    return [idx for idx, _ in sorted_by_value[:2]]

dog_idxs = find_ranking("dog")
cat_idxs = find_ranking("cat")

for i in range(8):
    tags = []
    if i == (dog_idxs[0] if len(dog_idxs) > 0 else -1):
        tags.append("dog1")
    elif i == (dog_idxs[1] if len(dog_idxs) > 1 else -1):
        tags.append("dog2")
    if i == (cat_idxs[0] if len(cat_idxs) > 0 else -1):
        tags.append("cat1")
    elif i == (cat_idxs[1] if len(cat_idxs) > 1 else -1):
        tags.append("cat2")
    results[i] = ",".join(tags)

# Write results to row 4, columns A-H
for i in range(8):
    ws.cell(row=4, column=i+1, value=results[i])

wb.save(output_path)
