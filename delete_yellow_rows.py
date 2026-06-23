import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_141-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_141-20_tc1/output.xlsx'

# Load workbook and sheets
wb = openpyxl.load_workbook(input_path)
ws_pl = wb['PL Recon Items']
ws_stmt = wb['Statement Recon Items']

# Helper: Check if row is yellow (fill RGB is 'FFFF00' or 'FFF00')
def is_yellow_row(row):
    for cell in row:
        if cell.fill and cell.fill.patternType and cell.fill.fgColor and cell.fill.fgColor.rgb:
            rgb = cell.fill.fgColor.rgb
            # Accept both full and short (openpyxl may report RGB as 'FFFFFF00')
            if rgb.endswith('FFFF00') or rgb.endswith('FFF00'):
                return True
    return False

# Gather matching (invoice, value) pairs in PL
pl_rows_to_delete = []
pl_pairs = []
for row_idx, row in enumerate(ws_pl.iter_rows(min_row=2), start=2):
    invoice = row[2].value  # Column C
    value = row[3].value    # Column D
    if invoice is not None and value is not None and is_yellow_row(row):
        pl_pairs.append((invoice, value))
        pl_rows_to_delete.append(row_idx)

# Gather matching (reference, value) pairs in Statement
stmt_rows_to_delete = []
stmt_pairs = []
for row_idx, row in enumerate(ws_stmt.iter_rows(min_row=2), start=2):
    reference = row[5].value  # Column F
    value = row[8].value      # Column I
    if reference is not None and value is not None and is_yellow_row(row):
        stmt_pairs.append((reference, value))
        stmt_rows_to_delete.append(row_idx)

# Determine common pairs
common_pairs = set(pl_pairs) & set(stmt_pairs)

# Delete matching rows from PL Recon Items
pl_rows_final = [row_idx for pair, row_idx in zip(pl_pairs, pl_rows_to_delete) if pair in common_pairs]
for row_idx in sorted(pl_rows_final, reverse=True):
    ws_pl.delete_rows(row_idx)

# Delete matching rows from Statement Recon Items
stmt_rows_final = [row_idx for pair, row_idx in zip(stmt_pairs, stmt_rows_to_delete) if pair in common_pairs]
for row_idx in sorted(stmt_rows_final, reverse=True):
    ws_stmt.delete_rows(row_idx)

# Save output
wb.save(output_path)
