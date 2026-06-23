from openpyxl import load_workbook, styles
from copy import copy

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_3/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/train/iter_3/regression_gate/after_pass/core_32337/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']
header_row = 2
start_row, end_row = 3, 15
# Fill formulas in E3:E15 (Expected Result) as =I[row]
for r in range(start_row, end_row+1):
    ws.cell(row=r, column=5).value = f'=I{r}'

# Insert 'Actual Age' col (P) after 'age (year)' (O, col 15 -> new col 16)
ws.insert_cols(16)
ws.cell(row=header_row, column=16).value = 'Actual Age'
# Set formulas for Actual Age using REPORT Date (B1) and DOB (C)
for r in range(start_row, end_row+1):
    ws.cell(row=r, column=16).value = f'=INT(($B$1-C{r})/365.25)'

# Copy fill from O to P for header and cells, using copy to avoid StyleProxy error
for r in range(header_row, end_row+1):
    ws.cell(row=r, column=16).fill = copy(ws.cell(row=r, column=15).fill)

# Format header (bold, center, top-aligned)
hdr = ws.cell(row=header_row, column=16)
hdr.font = styles.Font(bold=True)
hdr.alignment = styles.Alignment(horizontal='center', vertical='top')
# Data cells: center-aligned, bold, integer format
for r in range(start_row, end_row+1):
    cell = ws.cell(row=r, column=16)
    cell.alignment = styles.Alignment(horizontal='center')
    cell.font = styles.Font(bold=True)
    cell.number_format = '0'

wb.save(output_path)
