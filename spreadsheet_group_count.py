import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_30930_tc1/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed0/eval_30930_tc1/output.xlsx'

# Load workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

start_row = 2
end_row = 66
count = 0

for row in range(start_row, end_row + 1):
    a_val = ws[f'A{row}'].value
    b_val = ws[f'B{row}'].value
    # Only count values > 0 in Column A
    if b_val == 1:
        # Write count in Column C if this row in B is 1
        ws[f'C{row}'].value = count
        count = 0  # Reset for next group
    else:
        ws[f'C{row}'].value = None  # Blank cell
        if a_val is not None:
            try:
                if float(a_val) > 0:
                    count += 1
            except Exception:
                pass  # Ignore non-numeric values

wb.save(output_path)
