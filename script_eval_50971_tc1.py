import openpyxl
import collections

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_50971_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_50971_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))

# Extract columns A and B
colA = [r[0] for r in rows]
colB = [r[1] for r in rows]

# Count occurrences in A
ctr = collections.Counter([a for a in colA if a is not None])
dups = set(key for key, count in ctr.items() if count > 1)

# Collect B values where A is duplicated
results = [b for a, b in zip(colA, colB) if a in dups and b is not None]

# Write results to G3:G13
for i, val in enumerate(results[:11]):
    ws.cell(row=3+i, column=7, value=val)

wb.save(output_path)
