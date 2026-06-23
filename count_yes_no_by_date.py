from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_5/group_59160/r3/evolve_59160/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_5/group_59160/r3/evolve_59160/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Find all unique dates in column A
unique_dates = set()
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if isinstance(val, (str, type(None))):
        continue
    unique_dates.add(val)
unique_dates = sorted(unique_dates)
if not unique_dates:
    raise Exception('No dates found in column A')
focus_date = unique_dates[0]  # take the first date as the sample

# Count Yes/No for focus_date in all columns except first
count_yes = 0
count_no = 0
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val == focus_date:
        for col in range(2, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if v == 'Yes':
                count_yes += 1
            elif v == 'No':
                count_no += 1
ws['D3'] = 'Date'
ws['E3'] = 'Yes'
ws['F3'] = 'No'
ws['G3'] = 'Total'
ws['D4'] = focus_date
ws['E4'] = count_yes
ws['F4'] = count_no
ws['G4'] = count_yes + count_no

wb.save(output_path)
