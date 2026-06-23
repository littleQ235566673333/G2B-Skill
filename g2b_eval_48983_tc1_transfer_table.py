import openpyxl

INPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r3/eval_48983_tc1/input.xlsx'
OUTPUT = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r3/eval_48983_tc1/output.xlsx'

wb = openpyxl.load_workbook(INPUT)
ws = wb['Hárok1']

# Target headers (categories) and brands in the L:S table
cat_headers = [ws.cell(row=5, column=col).value for col in range(13, 20)]  # M:S
brand_rows = list(range(6, 12))
brand_names = [ws.cell(row=r, column=12).value for r in brand_rows]  # brands in L
# Source mapping: Brands in B6:B18 and categories in C5:J5
source_brand_row = {}
for i in range(6, 19):
    brand = ws.cell(row=i, column=2).value
    if brand:
        source_brand_row[brand] = i
source_cat_col = {cat: idx+3 for idx, cat in enumerate([ws.cell(row=5, column=c).value for c in range(3, 10)])}
# Fill M6:S11
target_start_row = 6
target_start_col = 13
for target_r, brand in enumerate(brand_names):
    src_row = source_brand_row.get(brand)
    for target_c, cat in enumerate(cat_headers):
        src_col = source_cat_col.get(cat)
        value = None
        if src_row and src_col:
            value = ws.cell(row=src_row, column=src_col).value
        ws.cell(row=target_start_row+target_r, column=target_start_col+target_c).value = value
wb.save(OUTPUT)
print('Done')
