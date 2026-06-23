import openpyxl

def is_row_empty(row):
    return all((cell.value is None or str(cell.value).strip() == '') for cell in row)

def delete_empty_rows(sheet, id_value=None):
    rows = list(sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row))
    ranges = []
    current_id = None
    start_idx = None
    # Identify ranges by ID in column B
    for idx, row in enumerate(rows):
        row_id = row[1].value
        if row_id is not None:
            if current_id != row_id:
                if current_id is not None and start_idx is not None:
                    ranges.append((current_id, start_idx + 2, idx + 1))
                current_id = row_id
                start_idx = idx
    if current_id is not None and start_idx is not None:
        ranges.append((current_id, start_idx + 2, len(rows) + 1))

    rows_to_delete = set()
    # If I2 is filled, only delete within matching ID ranges
    if id_value:
        for r in ranges:
            if r[0] == id_value:
                # r[1]: start_row, r[2]: end_row (1-indexed)
                for i in range(r[1], r[2] + 1):
                    excel_idx = i
                    row_cells = list(sheet.iter_rows(min_row=excel_idx, max_row=excel_idx, min_col=1, max_col=sheet.max_column))[0]
                    if is_row_empty(row_cells):
                        rows_to_delete.add(excel_idx)
    else: # Delete empty in all ranges
        for i in range(2, sheet.max_row + 1):
            row_cells = list(sheet.iter_rows(min_row=i, max_row=i, min_col=1, max_col=sheet.max_column))[0]
            if is_row_empty(row_cells):
                rows_to_delete.add(i)
    # Delete from bottom up to prevent shifting issues
    for row_num in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_num)

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_409-45_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_409-45_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
sheet = wb['DATA']
I2_val = sheet['I2'].value

# Apply the deletion logic
delete_empty_rows(sheet, I2_val)

wb.save(output_path)
