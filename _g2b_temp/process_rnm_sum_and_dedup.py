import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r1/eval_250-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r1/eval_250-20_tc1/output.xlsx'

# Read RNM sheet
df = pd.read_excel(input_path, sheet_name='RNM', header=None)

# Header is first row, A1:J1 (row 0), data begins at row 1
header = df.iloc[0]
data = df.iloc[1:]
data.columns = header

# Drop completely blank rows
cleaned = data.dropna(how='all')

col_B = header[1]
col_C = header[2]
col_J = header[9]

# Only group by rows where both B and C are not blank
groups = cleaned.dropna(subset=[col_B, col_C])

# For columns other than col_J, take first value. For J, sum.
def _agg(series):
    if series.name == col_J:
        return series.sum()
    return series.iloc[0]

sum_df = groups.groupby([col_B, col_C], as_index=False).agg(_agg)

# Restore the original order of columns
sum_df = sum_df[header]

# Compose full table (header, plus results) using concat
output_df = pd.concat([pd.DataFrame([header]), sum_df], ignore_index=True)

# Pad/cut to 20 rows, 10 columns
output_df = output_df.reindex(range(20))

wb = load_workbook(input_path)
ws = wb['RNM']
for r in range(20):
    for c in range(10):
        val = output_df.iloc[r, c] if r < len(output_df) and c < len(output_df.columns) else None
        ws.cell(row=1+r, column=1+c).value = val

wb.save(output_path)
