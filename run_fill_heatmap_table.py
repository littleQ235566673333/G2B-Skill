import openpyxl
import pandas as pd
from datetime import datetime, time

# Load workbook and sheet
infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/regression_gate/before_fix/core_52305/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/regression_gate/before_fix/core_52305/output.xlsx'
wb = openpyxl.load_workbook(infile)
ws = wb.active

# Get Names (column headers)
name_headers = [ws.cell(row=5, column=col).value for col in range(10, 15)]  # J=10 to N=14

# Get time ranges from H and I columns (cols 8 & 9), rows 6-24
ranges = [(ws.cell(row=row, column=8).value, ws.cell(row=row, column=9).value) for row in range(6, 25)]

# Extract all data (Type, Time, Name, Dest)
data = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
    vals = [cell.value for cell in row]
    if vals[1] is not None and vals[2] is not None:  # Must have Time and Name
        # Convert Time value to datetime.time if necessary
        if isinstance(vals[1], datetime):
            tval = vals[1].time()
        elif isinstance(vals[1], time):
            tval = vals[1]
        else:
            continue  # Skip if Time col is not a recognizable time/datetime
        data.append({'Name': vals[2], 'Time': tval})

df = pd.DataFrame(data)

# Fill table J6:N24
for i, (tstart, tend) in enumerate(ranges):
    if not (isinstance(tstart, time) and isinstance(tend, time)):
        continue  # skip invalid rows
    for j, name in enumerate(name_headers):
        if not name:
            continue
        cnt = df[(df['Name'] == name) & (df['Time'] >= tstart) & (df['Time'] < tend)].shape[0]
        ws.cell(row=6+i, column=10+j, value=cnt)

wb.save(outfile)
