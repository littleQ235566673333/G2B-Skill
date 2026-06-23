import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_438-18_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_438-18_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
input_ws = wb['INPUT']

# Remove 'OUPUT' sheet if it exists
overwrite_sheet = 'OUPUT'
if overwrite_sheet in wb.sheetnames:
    std = wb[overwrite_sheet]
    wb.remove(std)

# Create new 'OUPUT' sheet
output_ws = wb.create_sheet('OUPUT')

# Copy all values from 'INPUT' to 'OUPUT'
for i, row in enumerate(input_ws.iter_rows(values_only=True), start=1):
    for j, value in enumerate(row, start=1):
        output_ws.cell(row=i, column=j, value=value)

# Map for replacement
replace_map = {
    'UNATIDE STATES AMERICAN': 'USA',
    'FRANCE': 'FRA',
    'KOREA': 'KOR',
    'GERMANY': 'GR',
    'ITALY': 'IT'
}

# Replace values in column F (6th col), rows 2 to 6
for row in output_ws.iter_rows(min_row=2, max_row=6, min_col=6, max_col=6):
    for cell in row:
        if cell.value in replace_map:
            cell.value = replace_map[cell.value]

wb.save(output_path)
