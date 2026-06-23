import openpyxl
from collections import defaultdict
from openpyxl.utils import get_column_letter
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_1/group_516-46/r1/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_1/group_516-46/r1/evolve_516-46/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['ورقة1']

data = []
for row in ws.iter_rows(min_row=2, max_row=8, min_col=1, max_col=5, values_only=True):
    if row[0] is None or row[1] is None:
        continue
    # (date, brand, batch, origin, quantity)
    data.append(row)

brands = set([r[1] for r in data])
results = []
for brand in brands:
    brand_entries = [r for r in data if r[1] == brand]
    max_date = max(r[0] for r in brand_entries)
    last_date_entries = [r for r in brand_entries if r[0] == max_date]
    if len(last_date_entries) > 1:
        total_qty = sum(r[4] for r in last_date_entries)
        batch = last_date_entries[0][2]
        origin = last_date_entries[0][3]
        results.append((max_date, brand, batch, origin, total_qty))
    else:
        (date, brand, batch, origin, qty) = last_date_entries[0]
        results.append((date, brand, batch, origin, qty))

output_start_row = 2
output_columns = ['H','I','J','K','L']
headers = ['date', 'brand', 'batch', 'origin', 'quantity']
for col, header in zip(output_columns, headers):
    ws[f'{col}{output_start_row}'] = header

for i, result in enumerate(results, start=output_start_row+1):
    for j, value in enumerate(result):
        ws[f'{output_columns[j]}{i}'] = value

wb.save(output_path)
print('Done!')
