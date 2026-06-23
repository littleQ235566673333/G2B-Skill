import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_50768_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_50768_tc1/output.xlsx'

# Load workbook and sheet
template_wb = openpyxl.load_workbook(input_path)
template_ws = template_wb.active

# Find the correct header row and columns
header_row = None
for row in template_ws.iter_rows(min_row=1, max_row=10):
    cell_values = [str(cell.value).strip().lower() if cell.value else '' for cell in row]
    if 'id' in cell_values and 'impact' in cell_values and ('likelihood' in cell_values or 'liklihood' in cell_values):
        header_row = row[0].row
        break

# Get column indices for ID, Impact, Likelihoood (with typo)
for cell in template_ws[header_row]:
    val = str(cell.value).strip().lower() if cell.value else ''
    if val == 'id':
        risk_id_col = cell.column
    elif val == 'impact':
        impact_col = cell.column
    elif val in ('likelihood', 'liklihood'):
        likelihood_col = cell.column

# Collect risks by Impact-Likelihood
matrix_dict = {}
for row in template_ws.iter_rows(min_row=header_row+1):
    rid = row[risk_id_col-1].value
    impact = row[impact_col-1].value
    likelihood = row[likelihood_col-1].value
    if all([rid, impact, likelihood]):
        try:
            impact_num = int(impact)
            likelihood_num = int(likelihood)
        except Exception:
            continue
        key = (impact_num, likelihood_num)
        matrix_dict.setdefault(key, []).append(str(rid))

# Place values in matrix cells F12:H14 (Impact 1=bottom, Likelihood 1=left)
for i in range(3):
    for j in range(3):
        impact = 3 - i  # top row: impact=3, bottom: impact=1
        likelihood = j + 1  # left col: likelihood=1
        risk_ids = matrix_dict.get((impact, likelihood), [])
        outval = ', '.join(risk_ids) if risk_ids else ''
        row_num = 12 + i
        col_num = 6 + j  # F=6, G=7, H=8
        template_ws.cell(row=row_num, column=col_num, value=outval)

# Save the output
template_wb.save(output_path)
print('Matrix populated and saved to', output_path)
