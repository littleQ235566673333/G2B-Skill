import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_32093_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42_rerun2/eval_32093_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 16):
    current = ws[f'B{row}'].value
    c = ws[f'C{row}'].value
    d = ws[f'D{row}'].value
    e = ws[f'E{row}'].value
    # Find the first non-empty value from C, D, E, else use current (B)
    chosen = next((name for name in (c, d, e) if name not in (None, '', ' ')), current)
    ws[f'F{row}'] = chosen

wb.save(output_path)
