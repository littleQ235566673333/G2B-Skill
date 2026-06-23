from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/regression_gate/before_pass/core_41969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/regression_gate/before_pass/core_41969/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Fill A6 with =COUNTBLANK(A3:C3), B6 with =COUNTBLANK(D3:F3), C6 with =COUNTBLANK(G3:I3)
col_starts = [0, 3, 6] # A, D, G
for idx, offset in enumerate(col_starts):
    col1 = chr(ord('A') + offset)
    col2 = chr(ord('A') + offset + 2)
    formula = f'=COUNTBLANK({col1}3:{col2}3)'
    ws.cell(row=6, column=1+idx).value = formula

wb.save(output_path)
