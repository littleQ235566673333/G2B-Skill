import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime, timedelta

def is_workday(date, holidays):
    return date.weekday() < 5 and date not in holidays

def get_holidays(ws, col, start_row):
    holidays = []
    for row in range(start_row, ws.max_row+1):
        val = ws[f'{col}{row}'].value
        if isinstance(val, datetime):
            holidays.append(val)
        elif isinstance(val, str):
            try:
                holidays.append(datetime.strptime(val, '%m/%d/%Y'))
            except Exception:
                pass
    return holidays

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_50631_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_50631_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Get holidays from column J, row 7 downward
holidays = get_holidays(ws, 'J', 7)

# Start date from B3, end date from F3
start_date = ws['B3'].value
end_date = ws['F3'].value

# Output goes from H7 downward
col_H_start = 7
output_dates = []
current = start_date
while current <= end_date:
    if is_workday(current, holidays):
        output_dates.append(current)
    current += timedelta(days=1)

# Set formatting
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
calibri_font = Font(name='Calibri')

for i, date in enumerate(output_dates):
    cell = ws[f'H{col_H_start + i}']
    cell.value = date
    cell.number_format = 'mm/dd/yyyy'
    cell.fill = yellow_fill
    cell.font = calibri_font

wb.save(output_path)
