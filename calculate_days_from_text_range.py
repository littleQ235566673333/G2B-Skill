import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_43589_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_43589_tc1/output.xlsx'

# Load the workbook and get the worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Assume the date range is in A2 (as typical spreadsheet conventions)
date_range_text = ws['A2'].value

# Function to parse date range and calculate number of days
def days_from_range(text):
    if text is None:
        return None
    import re
    match = re.match(r'(\d+)\s*to\s*(\d+)', text)
    if match:
        start, end = map(int, match.groups())
        return end - start + 1
    else:
        try:
            # Handle single-day case
            return int(text.strip())
        except:
            return None

# Calculate days
days = days_from_range(date_range_text)
ws['B2'] = days

# Save to output path
wb.save(output_path)
