import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r2/eval_254-34_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r2/eval_254-34_tc1/output.xlsx'

found_indices = (0, 1, 3, 6, 8, 13, 15)
found_sum = 994110.0
# But there are only 15 cells—index 15 is out of range, so let's fix this to a valid match
# The actual valid indices from previous comb would have to be <= 14

# From the previous nums printed:
nums = [22704, 345510, 1238917, 30424.000000000004, 564077.969, 406152.99999999994, 60073, 436880, 158419, 277936.04750000004, 293976, 827362, 342830, 146354, 1146945]
positions = ['A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'A11', 'A12', 'A13', 'A14', 'A15', 'A16']

# Let's rerun the combination logic with the proper range and apply the first found solution:
import itertools
TARGET_RANGE = range(994108, 994113)
found = False
for k in range(2, len(nums)+1):
    for combo in itertools.combinations(enumerate(nums), k):
        idxs, vals = zip(*combo)
        s = sum(vals)
        if s in TARGET_RANGE:
            found_indices = idxs
            found_sum = s
            found = True
            break
    if found:
        break

wb = openpyxl.load_workbook(input_path)
if 'Sheet3' in wb.sheetnames:
    del wb['Sheet3']
ws = wb['Before']

# Reset highlight (no color)
for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.fill = PatternFill()
GREEN = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
for pos_idx in found_indices:
    ws[positions[pos_idx]].fill = GREEN
ws['C2'] = found_sum
wb.save(output_path)
print(f"Done. Highlighted, sum in C2: {found_sum}")
