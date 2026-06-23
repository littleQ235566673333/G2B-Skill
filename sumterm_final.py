import pandas as pd
from openpyxl import load_workbook

def main():
    input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_38823_tc1/input.xlsx'
    output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun2/eval_38823_tc1/output.xlsx'

    # Data starts at row 3 (Excel, 0-based pandas row 2)
    df = pd.read_excel(input_path, sheet_name=0, header=1, usecols=[0, 1, 2])
    df = df.dropna(subset=['Date']) # Only data rows

    # Get search terms from H4:H7
    wb = load_workbook(input_path)
    ws = wb.active
    search_terms = [ws[f'H{row}'].value for row in range(4, 8)]

    # Get date range from F4 (start), G4 (end)
    start_date = ws['F4'].value
    end_date = ws['G4'].value
    # If no end date in G4, set to max date in data
    if end_date is None:
        end_date = df['Date'].max()

    # Ensure pandas date type
    df['Date'] = pd.to_datetime(df['Date'])

    # For each search term (row 4-7)
    for idx, row in enumerate(range(4, 8), 0):
        term = search_terms[idx]
        if not term:
            ws[f'I{row}'] = ''
            continue
        # Case-insensitive substring match in 'Fabric'
        cond = (
            (df['Date'] >= pd.to_datetime(start_date)) &
            (df['Date'] <= pd.to_datetime(end_date)) &
            (df['Fabric'].str.contains(term, case=False, na=False))
        )
        val = df.loc[cond, 'Units Sold'].sum()
        ws[f'I{row}'] = float(val) if pd.notnull(val) else 0
    wb.save(output_path)

if __name__ == '__main__':
    main()
