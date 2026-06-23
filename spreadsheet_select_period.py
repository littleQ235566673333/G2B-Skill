from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/eval_seed42/eval_59224_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/eval_seed42/eval_59224_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active
project_start_date = ws['B2'].value

def as_date(x):
    if isinstance(x, datetime): return x
    if hasattr(x, 'date'):
        return x.date()
    return x

select_matched_row = None
for i, row in enumerate(range(4, 15)):
    c_val = ws[f'C{row}'].value
    d_val = ws[f'D{row}'].value
    if c_val is None or d_val is None:
        continue
    psd, c_date, d_date = as_date(project_start_date), as_date(c_val), as_date(d_val)
    try:
        if psd > c_date and psd < d_date:
            select_matched_row = row
            break
    except Exception:
        continue

for i, row in enumerate(range(4, 15)):
    c_val = ws[f'C{row}'].value
    d_val = ws[f'D{row}'].value
    e_cell = f'E{row}'
    if c_val is None or d_val is None:
        ws[e_cell] = ''
        continue
    if select_matched_row is not None and row <= select_matched_row:
        ws[e_cell] = 'Select Period'
    else:
        ws[e_cell] = f'=IF(AND($B$2>C{row},$B$2<D{row}),"Select Period","Do Not Select")'

wb.save(output_path)
