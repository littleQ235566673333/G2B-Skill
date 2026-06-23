import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

src_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_567-21_tc1/input.xlsx'
os_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_567-21_tc1/output.xlsx'

wb = openpyxl.load_workbook(src_path)
ws = wb['Sheet1']

data = []
for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
    if all(cell is None for cell in row):
        continue
    data.append(row)

columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
df = pd.DataFrame(data, columns=columns)

# Drop rows where column E is missing (None)
df = df.dropna(subset=['E'])

# Convert 'E' to year/month string for processing
if not pd.api.types.is_datetime64_any_dtype(df['E']):
    df['E'] = pd.to_datetime(df['E'], errors='coerce')
df = df.dropna(subset=['E'])

# For each (A, B), keep all rows with the max date in E
df['max_E'] = df.groupby(['A', 'B'])['E'].transform('max')
df_filtered = df[df['E'] == df['max_E']].copy()
df_filtered = df_filtered.drop('max_E', axis=1)
df_filtered = df_filtered.sort_values(['A', 'B', 'E'])

# Save output
df_filtered.reset_index(drop=True, inplace=True)
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Sheet1'

# Write output starting at row 3 as requested, pad empty first two rows
ws_out.append([None for _ in columns])
ws_out.append([None for _ in columns])
for r in dataframe_to_rows(df_filtered, index=False, header=True):
    ws_out.append(r)
wb_out.save(os_path)
