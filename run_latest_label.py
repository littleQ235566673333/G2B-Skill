import openpyxl
from collections import defaultdict

# File paths
INPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_58723_tc1/input.xlsx'
OUTPUT_PATH = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_58723_tc1/output.xlsx'
SHEET = 'Лист1'

wb = openpyxl.load_workbook(INPUT_PATH)
ws = wb[SHEET]

# Read Name (C) and Entry Time (I), rows 2 to 41
name_to_times = defaultdict(list)
rows = list(ws.iter_rows(min_row=2, max_row=41, min_col=3, max_col=9, values_only=True))
for idx, row in enumerate(rows):
    name = row[0]
    entry_time = row[6]
    name_to_times[name].append((idx, entry_time))

# For each Name, find the latest Entry Time
tags = [''] * len(rows)
for name, items in name_to_times.items():
    latest_time = max(t[1] for t in items)
    for idx, t in items:
        if t == latest_time:
            tags[idx] = 'Latest'
        else:
            tags[idx] = 'Not Latest'

# Write results to M2:M41 (col 13)
for i, tag in enumerate(tags, start=2):
    ws.cell(row=i, column=13, value=tag)

wb.save(OUTPUT_PATH)
print('Done')
