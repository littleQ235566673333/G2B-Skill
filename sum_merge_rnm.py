import pandas as pd
from openpyxl import load_workbook

# File paths
i_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r3/eval_250-20_tc1/input.xlsx"
o_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r3/eval_250-20_tc1/output.xlsx"

# Read data from RNM sheet into DataFrame
wb = load_workbook(i_path)
ws = wb["RNM"]
data = list(ws.values)
headers = data[0]
rows = data[1:]
df = pd.DataFrame(rows, columns=headers)

# If column types are unclear (numbers/NaN as str/float), clean column J for safe summing
df[headers[9]] = pd.to_numeric(df[headers[9]], errors='coerce').fillna(0)

# Group by columns B and C (index 1 and 2), sum J, take first for the rest
agg = {col: 'first' for col in headers}
agg[headers[9]] = 'sum' # Sum for J
result_df = df.groupby([headers[1], headers[2]], as_index=False).agg(agg)

# Prepare the output range (up to 20 rows, 10 columns)
output_rows = [list(headers)] # header
for _, row in result_df.iterrows():
    output_rows.append([row[h] for h in headers])
    if len(output_rows) >= 20:
        break
while len(output_rows) < 20:
    output_rows.append(['']*len(headers)) # pad with empty rows if needed

# Write back to new RNM sheet, overwriting A1:J20
wb_out = load_workbook(i_path)
if 'RNM' in wb_out.sheetnames:
    idx = wb_out.sheetnames.index('RNM')
    wb_out.remove(wb_out['RNM'])
ws_out = wb_out.create_sheet('RNM', 0)
for i, r in enumerate(output_rows):
    for j, val in enumerate(r):
        ws_out.cell(row=i+1, column=j+1, value=val)
wb_out.save(o_path)
