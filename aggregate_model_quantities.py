import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_49945_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42/eval_49945_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Build list of all unique models in C4:C11
models = [ws[f'C{r}'].value for r in range(4, 12) if ws[f'C{r}'].value]
unique_models = []
for m in models:
    if m not in unique_models:
        unique_models.append(m)

# Build a model->sum mapping across C4:C11 and corresponding D4:D11
model_sums = {}
for r in range(4, 12):
    model = ws[f'C{r}'].value
    qty = ws[f'D{r}'].value
    if model is not None and qty is not None:
        model_sums[model] = model_sums.get(model, 0) + (qty if isinstance(qty, (int, float)) else 0)

# Fill G4:G6 with the aggregated sums for the top 3 unique models
for idx, model in enumerate(unique_models[:3]):
    ws[f'G{4+idx}'] = model_sums.get(model, 0)

wb.save(output_path)
