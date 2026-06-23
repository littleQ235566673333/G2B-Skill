import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/group_59160/r1/evolve_59160/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_5/group_59160/r1/evolve_59160/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# The relevant dates are in D1, F1 (columns 4 and 6)
date_cols = [4, 6]  # D and F
result_cells = [4, 5, 6, 7]  # D3, E3, F3, G3

counts = []
for col in date_cols:
    date_ref = ws.cell(row=1, column=col).value
    yes_count = 0
    no_count = 0
    for row in range(2, ws.max_row+1):
        if ws.cell(row=row, column=1).value == date_ref:
            val = ws.cell(row=row, column=col).value
            if val == 'Yes':
                yes_count += 1
            elif val == 'No':
                no_count += 1
    counts.append((yes_count, no_count))

# Write results to D3:G3: D3=Yes col4, E3=No col4, F3=Yes col6, G3=No col6
results = [counts[0][0], counts[0][1], counts[1][0], counts[1][1]]
for i, value in enumerate(results):
    ws.cell(row=3, column=result_cells[i]).value = value

wb.save(output_path)
