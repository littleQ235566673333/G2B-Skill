import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_5/regression_gate/before_fix/core_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_5/regression_gate/before_fix/core_387-16/output.xlsx'

# Load the workbook and worksheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Utility to read a column (excluding header), up to row 18 (for A2:D18)
def read_col(col_idx, max_row=18):
    return [ws.cell(row=row, column=col_idx).value for row in range(2, max_row+1)]

A = read_col(1)
Value = read_col(2)
Binaries = read_col(3)
ResultValues = read_col(4)

# Remove only one instance per matching reference value, compact blanks
Value_mod = Value.copy()
Binaries_mod = Binaries.copy()
for res_val in ResultValues:
    for idx, v in enumerate(Value_mod):
        if v == res_val:
            Value_mod[idx] = None
            Binaries_mod[idx] = None
            break

def compact_column(col):
    vals = [x for x in col if x is not None]
    return vals + [None]*(len(col)-len(vals))

Value_new = compact_column(Value_mod)
Binaries_new = compact_column(Binaries_mod)

# Write columns back to worksheet (A2:D18)
for i in range(len(Value_new)):
    ws.cell(row=i+2, column=2).value = Value_new[i]
    ws.cell(row=i+2, column=3).value = Binaries_new[i]

# Solver result: sum of column A (ignore None/non-numeric)
solver_result = sum(x for x in A if isinstance(x, (int, float)))
ws.cell(row=2, column=4).value = solver_result

# Target value (usually the original D2 if numeric, or fallback to first of ResultValues)
target_value = None
if isinstance(ResultValues[0], (int, float)):
    target_value = ResultValues[0]
if target_value is not None:
    ws.cell(row=3, column=4).value = solver_result - target_value

wb.save(output_path)
