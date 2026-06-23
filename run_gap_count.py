from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/regression_gate/after_fix/core_50521/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/regression_gate/after_fix/core_50521/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active  # Use first sheet, or set the name explicitly if needed

def count_blanks_between_numbers(row):
    indices = [i for i, val in enumerate(row) if isinstance(val, (int, float))]
    if len(indices) < 2:
        return 0
    first, second = indices[0], indices[1]
    if isinstance(row[first], (int, float)) and row[first] > 1:
        return 1
    return sum(1 for v in row[first+1:second] if v in ('', None))

# Fill N4:N6 with calculated values
target_rows = range(4, 7)  # 4, 5, 6
for i in target_rows:
    # A-M columns = 1-13
    values = [ws.cell(row=i, column=j).value for j in range(1, 14)]
    result = count_blanks_between_numbers(values)
    ws.cell(row=i, column=14, value=result)  # N = 14

wb.save(output_path)
