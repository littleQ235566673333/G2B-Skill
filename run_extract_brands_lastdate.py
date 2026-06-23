import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_516-46/r2/evolve_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_1/group_516-46/r2/evolve_516-46/output.xlsx'
sheet_name = 'ورقة1'

# Read data with pandas, picking up the right sheet
df = pd.read_excel(input_path, sheet_name=sheet_name)

# Trim to only columns A-E (first 5 columns) and from row 2 onwards
data = df.iloc[1:, :5]  # skip header row 0, use rows 1+ (Excel's A2:Ex)
data.columns = df.iloc[0, :5]   # set header from the first row

# Ensure the first column is datetime
if not pd.api.types.is_datetime64_any_dtype(data.iloc[:,0]):
    try:
        data.iloc[:,0] = pd.to_datetime(data.iloc[:,0])
    except Exception:
        pass

col_date = data.columns[0]
col_brand = data.columns[1]
col_qty = data.columns[-1]

# Find the last date
max_date = data[col_date].max()
last_date_df = data[data[col_date] == max_date]

# Group by brand and sum quantities
result = last_date_df.groupby(col_brand).agg({col_qty: 'sum'}).reset_index()
result[col_date] = max_date

# Mark modifications for duplicated brand+last_date
mods = last_date_df.groupby(col_brand).size().reset_index(name='count')
result = result.merge(mods, on=col_brand, how='left')
result['modification'] = result['count'].apply(lambda x: 'modified' if x > 1 else '')

cols = [col_date, col_brand, col_qty, 'modification']
output = result[cols].copy()

# Open workbook and write output
wb = load_workbook(input_path)
ws = wb[sheet_name]

start_row, start_col = 2, 8  # H2 = col 8, row 2
for i, row in output.iterrows():
    for j, val in enumerate(row):
        ws.cell(row=start_row + i, column=start_col + j, value=val)

wb.save(output_path)
