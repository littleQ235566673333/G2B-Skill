import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun1/eval_22-47_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42_rerun1/eval_22-47_tc1/output.xlsx'

# Read all columns
# Assume first row is header

df = pd.read_excel(input_path)
orig_cols = list(df.columns)

# Remove fully empty rows
df = df.dropna(how='all')

def get_col(df, col):
    c = df[col] if col in df.columns else df.iloc[:, ord(col)-65]
    return c

# Remove duplicate (B,C) pairs (keep first)
seen = set()
rows = []
for _, row in df.iterrows():
    key = (row[orig_cols[1]], row[orig_cols[2]])
    if pd.isna(key[0]) and pd.isna(key[1]):
        continue
    if key not in seen:
        seen.add(key)
        rows.append(row)
df_nodup = pd.DataFrame(rows, columns=orig_cols)

# Remove rows where both B and C are empty (for exclusion of empty data rows)
df_nodup = df_nodup.dropna(subset=[orig_cols[1], orig_cols[2]], how='all')

# Helper column J: if empty, sort A-Z by B
col_J = get_col(df_nodup, 'J')
helper_list = [str(x).strip() for x in col_J if pd.notna(x) and str(x).strip() != '']
if len(helper_list) > 0:
    try:
        helper_list = pd.unique(pd.Series(helper_list))
    except Exception:
        helper_list = list(dict.fromkeys(helper_list))

# Prepare selection from original data
output_records = []
used_idx = set()

if len(helper_list) > 0:
    # 1. For names in J (in their first-used order), output all rows with B == name (keep source order)
    for name in helper_list:
        for idx, row in df_nodup.iterrows():
            if str(row[orig_cols[1]]) == name and idx not in used_idx:
                output_records.append(row)
                used_idx.add(idx)
    # 2. Then add remaining rows not listed in J (preserve order)
    for idx, row in df_nodup.iterrows():
        if idx not in used_idx:
            output_records.append(row)
else:
    # No helper, sort alphabetically by B
    output_records = sorted(df_nodup.itertuples(index=False), key=lambda r: ('' if pd.isna(r[1]) else r[1]))

# Only first 9 rows
output_records = output_records[:9]

# Coerce to normal row objects
if not isinstance(output_records[0], pd.Series):
    output_df = pd.DataFrame([list(r) for r in output_records], columns=orig_cols)
else:
    output_df = pd.DataFrame(output_records, columns=orig_cols)

# Write to F2:H10 (output columns: A,B,C into F,G,H)
wb = load_workbook(input_path)
ws = wb.active

for i, row in enumerate(output_df.values):
    for j in range(3):
        ws.cell(row=2 + i, column=6 + j, value=row[j])

# Helper for sorting: unify key type
def sortable_val(x):
    if x is None or pd.isna(x) or str(x).strip() == '':
        return (2, None)  # empties last
    try:
        fval = float(x)
        return (0, fval) # numbers first
    except Exception:
        return (1, str(x).lower()) # strings next

# Now sort output G2:H10 (columns 7:8). Only sort G and H by H asc, leave F fixed.
output_sub = [(ws.cell(row=2 + i, column=6).value,    # F
               ws.cell(row=2 + i, column=7).value,    # G
               ws.cell(row=2 + i, column=8).value)    # H
              for i in range(len(output_df))]

output_sub_sorted = sorted(output_sub, key=lambda t: sortable_val(t[2]))
for i, row in enumerate(output_sub_sorted):
    ws.cell(row=2 + i, column=7, value=row[1])
    ws.cell(row=2 + i, column=8, value=row[2])

wb.save(output_path)
