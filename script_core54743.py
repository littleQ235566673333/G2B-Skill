import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/regression_gate/before_fix/core_547-43/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/regression_gate/before_fix/core_547-43/output.xlsx'

# Load all sheets
sheets = pd.read_excel(input_path, sheet_name=None)

# Identify main and lookup tables
emp_table = None
lookup_table = None
expected_output = None
for sheetname, df in sheets.items():
    if 'Emp' in sheetname or 'emp' in sheetname:
        emp_table = df
    elif 'lookup' in sheetname.lower() or ('category' in df.columns and 'subcategory' in df.columns):
        lookup_table = df
    elif sheetname == 'Expected':
        expected_output = df

if emp_table is None:
    raise Exception('Emp table not found.')
if lookup_table is None:
    raise Exception('Lookup table not found.')

# Try to find Comp1-performance column and the ID variable
comp_perf_cols = [c for c in emp_table.columns if 'Comp1-performance' in str(c)]
if not comp_perf_cols:
    comp_perf_cols = [c for c in emp_table.columns if 'Comp1' in str(c) or 'performance' in str(c)]

# Context columns are those in Emp not in comp_perf_cols
context_cols = [c for c in emp_table.columns if c not in comp_perf_cols]

output_rows = []
for idx, row in emp_table.iterrows():
    for col in comp_perf_cols:
        record = {context: row[context] for context in context_cols}
        record['variable'] = col
        record['value'] = row[col] if pd.notnull(row[col]) else ''
        # Lookup match by variable
        cat = subcat = ''
        if 'variable' in lookup_table.columns and col in lookup_table['variable'].values:
            match = lookup_table.loc[lookup_table['variable']==col]
            if not match.empty:
                if 'category' in match.columns:
                    cat = match.iloc[0]['category']
                if 'subcategory' in match.columns:
                    subcat = match.iloc[0]['subcategory']
        record['category'] = cat
        record['subcategory'] = subcat
        output_rows.append(record)

output_df = pd.DataFrame(output_rows)

# Reorder to match expected output structure, if present
if expected_output is not None:
    output_df = output_df.reindex(columns=expected_output.columns)

# Write to the correct location in the output sheet
wb = load_workbook(input_path)
if 'Expected' not in wb.sheetnames:
    ws = wb.create_sheet('Expected')
else:
    ws = wb['Expected']

# Write header to A1 (openpyxl is 1-based)
for i, col in enumerate(output_df.columns):
    ws.cell(row=1, column=i+1, value=col)
# Data to start at A2
for idx, data_row in enumerate(output_df.values, 2):
    for j, val in enumerate(data_row, 1):
        ws.cell(row=idx, column=j, value=val)

wb.save(output_path)
