from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun2/eval_414-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42_rerun2/eval_414-20_tc1/output.xlsx'
ws_name = 'Sheet1'

wb = load_workbook(input_path)
ws = wb[ws_name]

# Find the first row where column A contains (case-insensitive) 'Invoice No.'
first_row = None
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and isinstance(val, str) and val.strip().lower() == 'invoice no.':
        first_row = row
        break

if first_row is not None and first_row > 1:
    ws.delete_rows(1, first_row - 1)

wb.save(output_path)
