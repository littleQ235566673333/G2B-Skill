import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_6/regression_gate/after_fix/core_4714/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/train/iter_6/regression_gate/after_fix/core_4714/output.xlsx'

# Read the file
# Use engine openpyxl to ensure compatibility

df = pd.read_excel(input_path, engine='openpyxl')

# Convert date column (C) to pandas Period (monthly), assume header row is present
month_col = df.columns[2]
hours_col = df.columns[3]

def to_month_period(val):
    try:
        if pd.isnull(val):
            return None
        return pd.to_datetime(val).to_period('M')
    except Exception:
        return None

df['MonthPeriod'] = df[month_col].apply(to_month_period)

results = []

for idx, row in df.iterrows():
    emp = row[0]
    month = row['MonthPeriod']
    if pd.isnull(month) or pd.isnull(row[hours_col]):
        results.append('n/a')
        continue

    emp_window = df[(df[df.columns[0]] == emp) & (df['MonthPeriod'].notnull())]
    # Filter window: current month and previous 3
    window = emp_window[(emp_window['MonthPeriod'] <= month) & 
                        (emp_window['MonthPeriod'] >= (month - 3))]

    # Only consider rows with valid hours values
    hours_vals = window[hours_col].dropna().values
    if len(hours_vals) < 4:
        results.append('n/a')
    else:
        avg = sum(hours_vals[-4:]) / 4  # Only average the LAST 4 (rolling)
        # Format result
        if avg == int(avg):
            s = str(int(avg))
        else:
            s = str(round(avg, 2))
        if avg > 48:
            s += '*'
        results.append(s)

# Write results into E2:E25
wb = load_workbook(input_path)
ws = wb.active
for i in range(24):  # For E2:E25
    cell = f'E{i+2}'
    if i < len(results):
        ws[cell] = results[i]
    else:
        ws[cell] = ''

wb.save(output_path)
