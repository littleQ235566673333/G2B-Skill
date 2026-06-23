import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/regression_gate/after_pass/core_56599/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_7/regression_gate/after_pass/core_56599/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Get slip number from M5
slip_number_cell = ws['M5'].value

if slip_number_cell is None or not str(slip_number_cell).isdigit() or len(str(slip_number_cell)) != 7:
    ws['B2'].value = 'not found - please re-enter'
else:
    slip_number = int(slip_number_cell)
    # Find all slip numbers (7 digit numbers) in the sheet
    all_slips = []
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None and str(val).isdigit() and len(str(val)) == 7:
                all_slips.append(int(val))
    if slip_number not in all_slips:
        ws['B2'].value = 'not found - please re-enter'
    else:
        min_slip = min(all_slips)
        book_number = ((slip_number - min_slip) // 10) + 1
        ws['B2'].value = book_number
wb.save(output_path)
