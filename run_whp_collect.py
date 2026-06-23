import pandas as pd
from openpyxl import load_workbook

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/group_54474/r2/evolve_54474/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_3/group_54474/r2/evolve_54474/output.xlsx'

# 1. Load workbook and worksheets
wb = load_workbook(input_file)
whp_ws = wb['WHP']

# 2. Read site names from WHP!B7:B8
site_names = []
for row in whp_ws.iter_rows(min_row=7, max_row=8, min_col=2, max_col=2):
    for cell in row:
        site_names.append(cell.value)

# 3. Read WHP DATA sheet, actual header is in third row (row index 2)
whp_data = pd.read_excel(input_file, sheet_name='WHP DATA', header=2)

# 4. For each site, find the corresponding row in WHP DATA
results = []
site_col = whp_data.columns[0]  # Should be 'Local London Partnership' or whatever is in first col
for site in site_names:
    row = whp_data[whp_data[site_col] == site]
    if not row.empty:
        values = row.iloc[0, :3].tolist()  # Take the first three columns (site/group, EEG, Dis)
        results.append(values)
    else:
        results.append([None, None, None])

# 5. Write the results to WHP!E7:G8
for i, result_row in enumerate(results):
    for j, value in enumerate(result_row):
        whp_ws.cell(row=7+i, column=5+j, value=value)

wb.save(output_file)
