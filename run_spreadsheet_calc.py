import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_51090_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_51090_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)

# Identify source and dest sheets
er_table_sheet = 'Errors Table' if 'Errors Table' in wb.sheetnames else wb.sheetnames[0]
ws = wb[er_table_sheet]
data = list(ws.values)
columns = data[0]
df = pd.DataFrame(data[1:], columns=columns)

users_for_code = {
    'II': ['CHROGIL1'],
    'IR': ['CHDSPOLJ'],
    'IT': ['CHSJEFFE'],
    'OV': ['CHBTHOMA'],
    'PI': ['CHROGIL1', 'CHDSPOLJ', 'CHSJEFFE', 'CHBTHOMA']
}
code_cols = {'II': 'N', 'IR': 'O', 'IT': 'P', 'OV': 'Q', 'PI': 'R'}
results = []
for idx in range(2, 24):  # Q3:Q24 = idx 2..23
    row = df.iloc[idx-2] if idx-2 < len(df) else None
    value = None
    if row is not None:
        warehouse_match = str(row.get('Warehouse', '')).strip() == '27'
        if warehouse_match:
            # Total Cartons (col M)
            total_cartons = row['M'] if 'M' in df.columns else None
            try:
                total_cartons = int(total_cartons)
            except:
                total_cartons = 0
            error_sum = 0
            for code, col in code_cols.items():
                err_val = row.get(col, 0)
                try:
                    num = int(err_val)
                except:
                    num = 0
                # For IR, only positive numbers
                if code == 'IR' and num <= 0:
                    continue
                # Must match relevant users for each code
                if 'User' in row and row['User'] in users_for_code[code]:
                    error_sum += num
            value = total_cartons - error_sum
        results.append(value)
    else:
        results.append(None)
ws_out = wb['Daily Numbers'] if 'Daily Numbers' in wb.sheetnames else wb.worksheets[0]
for i, v in enumerate(results):
    ws_out[f'Q{i+3}'] = v
wb.save(output_path)
print('done')
