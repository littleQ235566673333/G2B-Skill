import openpyxl
from datetime import timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_2/group_59595/r1/evolve_59595/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_2/group_59595/r1/evolve_59595/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active
# Read dates and points for C4:C19 (rows 4-19)
dates = []
points = []
for row in range(4, 20):  # Rows A4:B19
    dates.append(ws[f'A{row}'].value)
    points.append(ws[f'B{row}'].value)
# For each row, calculate the 7-day window sum
for i, (dt, pt) in enumerate(zip(dates, points)):
    row_num = i + 4
    if not dt:
        ws[f'C{row_num}'] = None
        continue
    window_start = dt - timedelta(days=6)
    total = 0
    for cmp_date, cmp_pt in zip(dates, points):
        if cmp_date is not None and window_start <= cmp_date <= dt:
            total += cmp_pt if cmp_pt is not None else 0
    ws[f'C{row_num}'] = total
wb.save(output_path)
