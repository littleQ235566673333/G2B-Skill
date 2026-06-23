import openpyxl
import pandas as pd

# Load workbook and sheet
wb = openpyxl.load_workbook('results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_55708_tc1/input.xlsx')
ws = wb['Sheet1']

# Read relevant data rows
# (from visible earlier output: rows 2-7 cover the data; columns 1-3)
data = list(ws.iter_rows(min_row=2, max_row=7, max_col=3, values_only=True))
df = pd.DataFrame(data, columns=['Department','Status','turnaroud time'])

# Apply filters and compute averages
filtered = df[df['Status'].isin(['In Progress','In Review']) & (df['turnaroud time'] >= 6)]
avg_by_dept = filtered.groupby('Department')['turnaroud time'].mean()

# Departments list in order found in the table for output rows
departments = ['A','B','C']
avgs = []
for dept in departments:
    val = avg_by_dept.get(dept, None)
    avgs.append(val if pd.notnull(val) else None)

# Write results to B11:B13
for idx, val in enumerate(avgs):
    cell = f'B{11+idx}'
    ws[cell].value = val

wb.save('results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed1/eval_55708_tc1/output.xlsx')
