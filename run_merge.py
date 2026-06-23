import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun1/eval_267-21_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed0/eval_seed42_rerun1/eval_267-21_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_merge = wb['merging']
ws_sh1 = wb['SH1']
ws_rp1 = wb['RP1']

def get_id_qty(sheet):
    data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            id_val = row[0]
            qty_val = row[1]
            data[id_val] = qty_val
    return data

sh1_map = get_id_qty(ws_sh1)
rp1_map = get_id_qty(ws_rp1)

for row in range(2, 12):  # C2:D11
    id_cell = ws_merge.cell(row=row, column=2)
    id_val = id_cell.value
    # Sheet SH1, if error or blank, use "-"
    val_c = sh1_map.get(id_val, '-')
    val_c = val_c if val_c not in [None, '', ' '] else '-'
    # Sheet RP1, if error or blank, use "-"
    val_d = rp1_map.get(id_val, '-')
    val_d = val_d if val_d not in [None, '', ' '] else '-'
    ws_merge.cell(row=row, column=3, value=val_c)
    ws_merge.cell(row=row, column=4, value=val_d)

wb.save(output_path)
