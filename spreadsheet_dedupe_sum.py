import pandas as pd
from openpyxl import load_workbook

inp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_250-20_tc1/input.xlsx'
outp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/eval_seed42_rerun1/eval_250-20_tc1/output.xlsx'

# Read the data
df = pd.read_excel(inp, sheet_name='RNM')

# Group by columns B and C (index 1, 2), sum column J (index 9), keep first values otherwise
group_cols = [df.columns[1], df.columns[2]]
def aggfunc(x):
    return x.iloc[0]
agg_dict = {col: 'first' for col in df.columns}
agg_dict[df.columns[9]] = 'sum'   # Sum J

df_grouped = df.groupby(group_cols, as_index=False).agg(agg_dict)

# Load the workbook and worksheet
wb = load_workbook(inp)
ws = wb['RNM']

# Write back grouped data to worksheet (A1:J20)
header = list(df.columns)
for c, value in enumerate(header, 1):
    ws.cell(row=1, column=c, value=value)
for row_idx in range(2, 21):  # rows 2..20
    if row_idx-2 < len(df_grouped):
        row = df_grouped.iloc[row_idx-2]
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    else:
        # Clear any previous content beyond result
        for col_idx in range(1, 11):
            ws.cell(row=row_idx, column=col_idx, value=None)

# Save to output
wb.save(outp)
