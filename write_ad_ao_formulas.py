from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_4/group_44017/r3/evolve_44017/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_4/group_44017/r3/evolve_44017/output.xlsx'

wb = load_workbook(input_file)
ws = wb.active

# Columns: AD=30 to AO=43, rows 14:42
for row in range(14, 43):
    for idx, col in enumerate(range(30, 43)):
        col_letter = get_column_letter(col)
        date_cell = f'{col_letter}$9'
        eff_date_cell = f'$L{row}'
        freq_cell = f'$J{row}'
        base_rate_cell = f'$W{row}'
        # Increases
        increases = [f'$M{row}', f'$N{row}', f'$O{row}', f'$P{row}']
        # nth wave: 0, 1, 2, 3
        waves = []
        for n in range(4):
            # This cond will be 1 if the nth wave is in effect at this month
            cond = f'MIN(1,MAX(0,INT((DATEDIF({eff_date_cell},{date_cell},\"m\")/{freq_cell})>={n})))'
            waves.append(f'(1+IF({increases[n]}<>\"\",{increases[n]},0))^{cond}')
        formula = f'=IF(OR({eff_date_cell}="",{date_cell}<{eff_date_cell}),"",{base_rate_cell}*' + '*'.join(waves) + ')'
        ws.cell(row=row, column=col).value = formula

wb.save(output_file)
print('Formulas written and saved.')
