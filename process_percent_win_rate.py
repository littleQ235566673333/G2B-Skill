import openpyxl
import re

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_49036_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_49036_tc1/output.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_path)
ws = wb['Dashboard']

# Read cells B5 and B6 values (might be formula, or resulting text if openpyxl read as values)
def extract_number(val):
    if val is None:
        return None
    # If string like '24 TRADES', extract leading digits
    m = re.search(r'([0-9]+)', str(val))
    return int(m.group(1)) if m else None

# Attempt to read value for B5 and B6
val_b5 = ws['B5'].value
val_b6 = ws['B6'].value

# If the value is a formula, we need to look for cached value (openpyxl does not evaluate formulas)
# By default, .value gives the formula string, not the formula result unless data_only=True.
# Let's reload the workbook with data_only=True
wb_data = openpyxl.load_workbook(input_path, data_only=True)
ws_data = wb_data['Dashboard']
val_b5_data = ws_data['B5'].value
val_b6_data = ws_data['B6'].value

# Extract numbers
num_b5 = extract_number(val_b5_data)
num_b6 = extract_number(val_b6_data)

if (num_b5 is None) or (num_b6 is None) or (num_b5 == 0):
    ws['B8'] = 'N/A WIN RATE'
else:
    percent = (num_b6 / num_b5) * 100
    percent_str = f'{percent:.2f}% WIN RATE'
    ws['B8'] = percent_str

wb.save(output_path)
print('Dashboard!B8 value:', ws['B8'].value, '| num_b5:', num_b5, '| num_b6:', num_b6)
