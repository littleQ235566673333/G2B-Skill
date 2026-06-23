from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_3/regression_gate/after_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_3/regression_gate/after_pass/core_55421/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

def get_statuses(num):
    statuses = set()
    no_show_dates = []
    for row in range(2, 21):
        if ws[f'A{row}'].value == num:
            status = str(ws[f'D{row}'].value).strip() if ws[f'D{row}'].value else ''
            date_val = ws[f'E{row}'].value
            statuses.add(status)
            if status == 'NO SHOW':
                no_show_dates.append((date_val, row))
    return statuses, no_show_dates

for row in range(2, 21):
    num = ws[f'A{row}'].value
    if num is None:
        ws[f'F{row}'] = ''
        continue
    statuses, no_show_dates = get_statuses(num)
    if statuses == {'SCH'}:
        ws[f'F{row}'] = 'FUTURE'
    elif {'SCH', 'NO SHOW'} == statuses:
        ws[f'F{row}'] = 'NS/SCHED'
    elif 'NO SHOW' in statuses:
        # Check if in this row it's NO SHOW, and date field
        status_here = str(ws[f'D{row}'].value).strip() if ws[f'D{row}'].value else ''
        if status_here == 'NO SHOW':
            date_val = ws[f'E{row}'].value
            if date_val:
                ws[f'F{row}'] = 'NO ACTION NEEDED'
            else:
                ws[f'F{row}'] = 'CALL PT'
        else:
            ws[f'F{row}'] = ''
    else:
        ws[f'F{row}'] = ''

wb.save(output_path)
