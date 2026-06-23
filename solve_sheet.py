import openpyxl

input_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_10/task_387-16/r3/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41-fix/train/iter_10/task_387-16/r3/evolve_387-16/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']
data = list(ws.values)

# Find header row
def norm(t):
    return str(t).strip().lower() if t is not None else ''
header_row = None
for i, row in enumerate(data):
    normed = [norm(x) for x in row]
    if 'value' in normed and 'binaries' in normed:
        header_row = i
        break
headers = [norm(x) for x in data[header_row]]
value_col = headers.index('value')
binary_col = headers.index('binaries')

# Locate 'Result values' block by looking for 'Result values' in col D (index 3)
result_values = []
found_block = False
for idx, row in enumerate(data):
    if isinstance(row[3], str) and row[3].strip().lower() == 'result values':
        found_block = True
        # Start collecting below this
        for sub_row in data[idx+1:]:
            if sub_row[3] is not None and isinstance(sub_row[3], (int, float)):
                result_values.append(sub_row[3])
            else:
                break
        break
assert found_block, "Result values block not found"

# Extract Values & Binaries (rows right after header)
values_rows = []
for r in data[header_row+1:]:
    v, b = r[value_col], r[binary_col]
    if isinstance(v, (int, float)) and b is not None:
        values_rows.append([v, b])

# Remove one instance per result value
used = set()
filtered_rows = []
for row in values_rows:
    v = row[0]
    if v in result_values and v not in used:
        used.add(v)
        continue
    filtered_rows.append(row)

# Shift output rows so there are no blanks
row_base = header_row + 2 # data starts after the header
for i in range(len(values_rows)):
    v = filtered_rows[i][0] if i < len(filtered_rows) else None
    b = filtered_rows[i][1] if i < len(filtered_rows) else None
    ws.cell(row=row_base + i, column=value_col + 1, value=v)
    ws.cell(row=row_base + i, column=binary_col + 1, value=b)

# Write formulas (columns E=target, H=solver result, J=difference in the template)
ws['H2'] = '=SUMPRODUCT(A3:A21, B3:B21)'
ws['J2'] = '=SUM(H2-E2)'

# Save
wb.save(output_path)
print('Done')
