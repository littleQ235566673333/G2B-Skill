import openpyxl
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_46240_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_46240_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

for row in range(2, 5):
    h_val = ws[f'H{row}'].value
    result = ''
    if h_val is None or (isinstance(h_val, str) and h_val.strip() == ''):
        result = 'No'
    elif isinstance(h_val, str) and h_val.strip().upper() == 'N/A':
        result = 'N/A'
    elif isinstance(h_val, datetime):
        result = 'Yes'
    elif isinstance(h_val, str) and h_val.strip().upper() == 'YES':
        result = 'Yes'
    else:
        # Check if it is a string that represents a date (rare, for completeness)
        try:
            datetime.strptime(h_val, '%Y-%m-%d')
            result = 'Yes'
        except Exception:
            try:
                datetime.strptime(h_val, '%m/%d/%Y')
                result = 'Yes'
            except Exception:
                result = 'No'
    ws[f'J{row}'].value = result

wb.save(output_path)
