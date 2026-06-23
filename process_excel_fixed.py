import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_120-24_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-B/eval_120-24_tc1/output.xlsx'

bg_options = [
    'OCOGS - Spares - Transfer Price Overhead', 'OFSS - ORCL Consulting Prod Cost I/C'
]
ay_options = ['BOA_033E', 'BOA_011G']

# Column mappings (Excel: A=1)
col_BG = 59  # BG
col_BL = 50  # BL
col_AY = 41  # AY
col_BN = 52  # BN

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

fill_lag = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')

for row in range(2, 46):
    bl = ws.cell(row=row, column=col_BL).value
    bg = ws.cell(row=row, column=col_BG).value
    ay = ws.cell(row=row, column=col_AY).value
    bn_cell = ws.cell(row=row, column=col_BN)

    if bl == 'LAG' and bg in bg_options:
        if ay in ay_options:
            bn_cell.value = 'OCOGS - Spares - Transfer Price Overhead'
        else:
            bn_cell.value = 'OFSS - ORCL Consulting Prod Cost I/C'
        if bn_cell.value:
            ws.cell(row=row, column=col_BL).fill = fill_lag
    else:
        ws.cell(row=row, column=col_BL).fill = PatternFill()

wb.save(output_path)
