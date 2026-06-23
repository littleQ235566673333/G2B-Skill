import pandas as pd
from openpyxl import load_workbook

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42/eval_250-20_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/eval_seed42/eval_250-20_tc1/output.xlsx'

# Read only the RNM sheet for processing with pandas
df = pd.read_excel(input_path, sheet_name='RNM', header=0)

# If the file contains extra trailing NaN rows, drop them
df = df.dropna(how='all')

header = df.columns.tolist()

# Group by columns B & C (the second and third column by label, not index)
grouped = df.groupby([header[1], header[2]], as_index=False)

def aggfunc(x):
    out = x.iloc[0].copy()
    out[header[9]] = x[header[9]].sum()   # sum column J
    return out

result_df = grouped.apply(aggfunc)
result_df.reset_index(drop=True, inplace=True)

# Pad/truncate to ensure A1:J20 output as requested (including header)
final_df = result_df.reindex(range(20))
final_df.columns = header

# Save with openpyxl and preserve other sheets
wb = load_workbook(input_path)
if 'RNM' in wb.sheetnames:
    ws = wb['RNM']
    # Clear rows 2 to 20 (A2:J20)
    for row in ws.iter_rows(min_row=2, max_row=20, min_col=1, max_col=10):
        for cell in row:
            cell.value = None
    # Write back the result DataFrame (A2:J20)
    for r_idx, row in enumerate(final_df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
wb.save(output_path)
