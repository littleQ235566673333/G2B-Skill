from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_7/regression_gate/after_fix/core_374-31/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_7/regression_gate/after_fix/core_374-31/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']

rows = list(ws.iter_rows(values_only=True))

# Find all 'Code' rows and store their indices
code_indices = []
for i, row in enumerate(rows):
    # Defensive: skip if row is None
    if row[0] is not None and str(row[0]).strip() == 'Code':
        code_indices.append(i)

# Prepare a mask of rows to keep
keep = [True] * len(rows)
prev_code = -1
for code_idx in code_indices:
    # Check for blanks above this 'Code' (but below previous 'Code')
    for j in range(prev_code + 1, code_idx):
        if all((c is None or str(c).strip() == '') for c in rows[j]):
            keep[j] = False
    prev_code = code_idx

# Filter out blank rows above 'Code' and assemble new row set
new_rows = [rows[i] for i in range(len(rows)) if keep[i]]

# Write back the rows to the worksheet, filling up to A1:D5
ws.delete_rows(1, ws.max_row)
for row in new_rows[:5]:
    row_filled = list(row) + [None] * (4 - len(row)) if len(row) < 4 else list(row[:4])
    ws.append(row_filled)

# If less than 5 rows after deletion, pad with empty rows
for _ in range(len(new_rows), 5):
    ws.append([None]*4)

wb.save(output_path)
