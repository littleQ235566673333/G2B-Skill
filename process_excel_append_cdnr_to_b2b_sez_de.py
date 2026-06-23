import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/group_130-9/r1/evolve_130-9/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_6/group_130-9/r1/evolve_130-9/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws_src = wb['cdnr']
ws_dest = wb['b2b, sez, de']

# Get headers and build header-to-column maps (by header text, not position)
def get_header_map(ws, header_row=5):
    return {str(cell.value).strip(): cell.col_idx for cell in ws[header_row] if cell.value is not None}

src_headers = get_header_map(ws_src)
dest_headers = get_header_map(ws_dest)

# Determine row appending start
def find_first_blank_row(ws, col_idx, start_row):
    r = start_row
    while ws.cell(row=r, column=col_idx).value not in [None, '']:
        r += 1
    return r

append_row = find_first_blank_row(ws_dest, 1, 7) # start at row 7 as user says
first_data_row = 6  # Source data starts after header
src_data_rows = list(ws_src.iter_rows(min_row=first_data_row, max_row=ws_src.max_row, values_only=False))
for row_cells in src_data_rows:
    # Check if row is empty (all cells empty)
    if all([c.value in [None, ''] for c in row_cells]):
        continue
    # Write each matching header
    for header, dcol in dest_headers.items():
        if header not in src_headers:
            continue
        scol = src_headers[header]
        svalue = row_cells[scol - 1].value
        # Special handling for Amount columns (L, M == col 12,13): write as negative if present
        if dcol in [12, 13] and isinstance(svalue, (int, float)) and svalue is not None:
            ws_dest.cell(row=append_row, column=dcol, value=-abs(svalue))
        else:
            ws_dest.cell(row=append_row, column=dcol, value=svalue)
    # Set H blank and I to 'Credit Note' per user
    ws_dest.cell(row=append_row, column=8, value=None)
    ws_dest.cell(row=append_row, column=9, value='Credit Note')
    append_row += 1

wb.save(output_path)
