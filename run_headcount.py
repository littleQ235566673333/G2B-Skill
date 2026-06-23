import openpyxl
import pandas as pd

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/regression_gate/before_fix/core_59129/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_7/regression_gate/before_fix/core_59129/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Load employee data
rows = list(ws.iter_rows(min_row=2, max_row=22, min_col=1, max_col=2, values_only=True))
df = pd.DataFrame(rows, columns=['Hire Date', 'Term Date'])

# Determine the month columns (E1:P1 -> columns 5 to 16)
months = [ws.cell(row=1, column=i).value for i in range(5, 17)]

results = []
for month in months:
    # Employees whose hire date is on or before the month and (no term date or term date is after or equal to the month)
    active = df[(df['Hire Date'] <= month) & ((df['Term Date'].isnull()) | (df['Term Date'] >= month))]
    results.append(len(active))

# Write results to E2:P2
for i, val in enumerate(results):
    ws.cell(row=2, column=5 + i, value=val)

wb.save(output_path)
