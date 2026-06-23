from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_1/group_44017/r3/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_1/group_44017/r3/evolve_44017/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Data']
# Mapping
first_data_row, last_data_row = 14, 42
start_col, end_col = 30, 41 # AD:AO (excel cols)
date_row = 9
for row in range(first_data_row, last_data_row+1):
    base_rate = f'$W{row}'
    eff_date = f'$L{row}'
    freq = f'$J{row}'
    increases = [f'$M{row}', f'$N{row}', f'$O{row}', f'$P{row}']
    for c in range(start_col, end_col+1):
        col_letter = get_column_letter(c)
        month_date = f'{col_letter}{date_row}'
        # Formula to check if this column date is before effective date
        formula = f'=IF({month_date}<$L{row}, "", '
        # Factors for up to 4 waves
        factors = [f'(1+IF({cell}<>"",{cell},0))' for cell in increases]
        # Boundaries: when each wave starts
        boundaries = [f'EDATE($L{row},$J{row}*{i})' for i in range(1,4)]
        # Build multiplier: multiply by factor if past each wave boundary
        # Start with base multiplier
        factor_formula = f'(1'
        for idx, boundary in enumerate(boundaries):
            # Use factors idx+1 because first factor is always applied once eff_date is hit
            factor_formula += f'*IF({month_date}>={boundary},{factors[idx+1]},1)'
        factor_formula += f'*IF({month_date}>=$L{row},{factors[0]},1)'
        factor_formula += ')'
        rate_formula = formula + f'$W{row}*{factor_formula})'
        ws.cell(row=row, column=c).value = rate_formula
        ws.cell(row=row, column=c).fill = PatternFill(fill_type=None)
wb.save(output_path)
