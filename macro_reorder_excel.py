import openpyxl
from collections import defaultdict

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_3/group_118-50/r1/evolve_118-50/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_3/group_118-50/r1/evolve_118-50/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Collect all non-empty words in column A
words = []
for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
    val = row[0]
    if val is not None and str(val).strip() != '':
        words.append(str(val).strip())

# Step 1: Sort names in column A
sorted_words = sorted(words)
for idx, word in enumerate(sorted_words):
    ws.cell(row=idx+1, column=1, value=word)
# Optional: clear rest of column A if original was longer
for idx in range(len(sorted_words)+1, ws.max_row+1):
    ws.cell(row=idx, column=1).value = None

# Step 2: Macro pattern match (per user specification)
def transform_word(word):
    if len(word) < 5:
        return None
    return word[4] + word[:4] + word[5:]

# Group words by last three letters
grouped = defaultdict(list)
for w in sorted_words:
    if len(w) >= 3:
        grouped[w[-3:]].append(w)

# Find matching pairs
def find_pairs():
    found_pairs = []
    seen = set()
    for end3, group in grouped.items():
        group_set = set(group)
        for orig in group:
            if len(orig) < 5:
                continue
            trans = transform_word(orig)
            if trans != orig and trans in group_set:
                key = (end3, trans, orig)
                if key not in seen:
                    found_pairs.append((end3, trans, orig))
                    seen.add(key)
    return found_pairs

found_pairs = find_pairs()

SPECIAL = ['ING', 'ERS', 'ATE', 'EST', 'ONE', 'IER', 'ILY']
def ending_sort_key(x):
    end3 = x[0]
    try:
        return (SPECIAL.index(end3), end3)
    except ValueError:
        return (len(SPECIAL), end3)

found_pairs.sort(key=ending_sort_key)

# Write results to C2:D5000
for i, (_, trans, orig) in enumerate(found_pairs[:4999]):
    ws.cell(row=i+2, column=3, value=trans)
    ws.cell(row=i+2, column=4, value=orig)
# Clear trailing cells if less than 4999 results
for i in range(len(found_pairs)+2, 5001):
    ws.cell(row=i, column=3, value=None)
    ws.cell(row=i, column=4, value=None)

wb.save(output_path)
print(f'Sorted column A and wrote {len(found_pairs[:4999])} pairs to C2:D5000.')
