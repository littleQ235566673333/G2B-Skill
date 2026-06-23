from openpyxl import load_workbook
wb = load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_5/regression_gate/before_pass/core_41601/input.xlsx')
ws = wb['Students']
for r in range(2, 8):
    ws[f'E{r}'] = '=INDIRECT("\'"&A{}&"\'!C2")'.format(r)
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed2/train/iter_5/regression_gate/before_pass/core_41601/output.xlsx')
print('Formulas written to Students!E2:E7')
