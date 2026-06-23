import openpyxl
from openpyxl import load_workbook
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_4/regression_gate/after_fix/core_516-46/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_4/regression_gate/after_fix/core_516-46/output.xlsx'

wb = load_workbook(input_path)
ws = wb['ورقة1']

# Read data from A2:E (first block)
data = []
for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
    if row[0] is None:
        break
    data.append(row)

# Columns: 0:date, 1:brand, 2:batch, 3:origin, 4:quantity
df = pd.DataFrame(data, columns=['date','brand','batch','origin','quantity'])

# Find last date for each brand
last_dates = df.groupby('brand')['date'].max().reset_index()

# For each brand, get all entries with last date and combine quantities by brand for that date, sum qty, merge batch/origin if multiple
results = []
for _, row in last_dates.iterrows():
    brand = row['brand']
    last_date = row['date']
    mask = (df['brand'] == brand) & (df['date'] == last_date)
    sub = df[mask]
    sum_qty = sub['quantity'].sum()
    batches = ', '.join([str(x) for x in sub['batch']])
    origins = ', '.join([str(x) for x in sub['origin']])
    results.append([last_date, brand, batches, origins, sum_qty])

# Write header to H2
header = ['date','brand','batch','origin','quantity']
for j, val in enumerate(header, start=8):
    ws.cell(row=2, column=j, value=val)

for i, res in enumerate(results, start=3):
    for j, val in enumerate(res, start=8):
        ws.cell(row=i, column=j, value=val)

wb.save(output_path)
print('Saved result to', output_path)
