from openpyxl import load_workbook

input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/regression_gate/before_pass/core_41969/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_1/regression_gate/before_pass/core_41969/output.xlsx'

wb = load_workbook(input_fp)
ws = wb.active

# For cells A6, B6, C6, apply shifting COUNTBLANK formulas
for i, col in enumerate(['A', 'B', 'C']):
    start_col_idx = i * 3 + 1  # A=1, D=4, G=7
    end_col_idx = start_col_idx + 2
    # Convert numeric index to Excel column letter (1=A, 2=B, ...)
    def num_to_col(n):
        res = ''
        while n:
            n, r = divmod(n-1, 26)
            res = chr(65 + r) + res
        return res
    start_col_letter = num_to_col(start_col_idx)
    end_col_letter = num_to_col(end_col_idx)
    formula = f"=COUNTBLANK({start_col_letter}3:{end_col_letter}3)"
    ws[f'{col}6'] = formula

wb.save(output_fp)
