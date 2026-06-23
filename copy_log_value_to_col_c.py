import openpyxl

# Paths
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_5835_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_5835_tc1/output.xlsx'

# Load workbook and select the first sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# The range is C3:C19 (rows 3 to 19)
for row in range(3, 20):
    log_value = ws[f'I{row}'].value  # Get value from column I
    ws[f'C{row}'].value = log_value  # Write/log value to column C

# Save to output path
wb.save(output_path)
