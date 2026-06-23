import openpyxl
from datetime import datetime

# File paths
input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_58723_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_58723_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Prepare to group entry times by name
from collections import defaultdict
time_dict = defaultdict(list)

# Read all data in range (rows 2..41, columns C and I)
for row in range(2, 42):
    name = ws[f'C{row}'].value
    entry_time_val = ws[f'I{row}'].value
    # handle both datetime and string (parse if string)
    if isinstance(entry_time_val, str):
        try:
            entry_time = datetime.strptime(entry_time_val, '%Y-%m-%d %H:%M:%S')
        except Exception:
            entry_time = None
    else:
        entry_time = entry_time_val
    if name and entry_time:
        time_dict[name].append(entry_time)

# Get the latest entry time for each name
def get_latest_entry(times):
    if not times:
        return None
    return max(times)

latest_entry = {name: get_latest_entry(times) for name, times in time_dict.items()}

# Mark 'Latest' or 'Not Latest' in column M
for row in range(2, 42):
    name = ws[f'C{row}'].value
    entry_time_val = ws[f'I{row}'].value
    # handle both datetime and string
    if isinstance(entry_time_val, str):
        try:
            entry_time = datetime.strptime(entry_time_val, '%Y-%m-%d %H:%M:%S')
        except Exception:
            entry_time = None
    else:
        entry_time = entry_time_val
    label = ''
    if name and entry_time and name in latest_entry:
        if entry_time == latest_entry[name]:
            label = 'Latest'
        else:
            label = 'Not Latest'
    ws[f'M{row}'] = label

# Save the workbook
wb.save(output_path)
