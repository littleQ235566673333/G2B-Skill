import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/after_pass/core_50526/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/after_pass/core_50526/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

lookup_value = ws['B6'].value
lookup_row = None
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=1).value == lookup_value:
        lookup_row = row
        break

output = []
if lookup_row:
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=lookup_row, column=col).value
        if isinstance(val, (int, float)) and val > 0:
            output.append(ws.cell(row=1, column=col).value)

for i in range(2):
    ws.cell(row=9 + i, column=2, value=output[i] if i < len(output) else '')

wb.save(output_path)
