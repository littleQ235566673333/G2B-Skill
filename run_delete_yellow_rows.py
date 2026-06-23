import openpyxl
from openpyxl.styles import PatternFill

yellow_fills = ['FFFF00', 'FFFFFF00']

def is_yellow(cell):
    fill = cell.fill
    if isinstance(fill, PatternFill) and fill.fgColor is not None:
        color = fill.fgColor.rgb or fill.fgColor.indexed
        if color:
            color = str(color).replace('0x', '').upper()
            for y in yellow_fills:
                if color.endswith(y):
                    return True
    return False

def main():
    wb = openpyxl.load_workbook('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_141-20_tc1/input.xlsx')
    ws_pl = wb['PL Recon Items']
    ws_stmt = wb['Statement Recon Items']

    # Find all highlighted pairs in PL sheet
    matching_pairs = set()
    rows_to_delete_pl = []
    for row in ws_pl.iter_rows(min_row=2, max_row=ws_pl.max_row):
        # Check if both C and D are yellow (columns 2, 3)
        if is_yellow(row[2]) and is_yellow(row[3]):
            matching_pairs.add((str(row[2].value), str(row[3].value)))
            rows_to_delete_pl.append(row[0].row)
    
    # Find matching pairs in the Statement sheet
    rows_to_delete_stmt = []
    for row in ws_stmt.iter_rows(min_row=2, max_row=ws_stmt.max_row):
        if is_yellow(row[5]) and is_yellow(row[8]): # F and I
            ref = (str(row[5].value), str(row[8].value))
            if ref in matching_pairs:
                rows_to_delete_stmt.append(row[0].row)

    # Delete rows in reverse order
    for idx in sorted(rows_to_delete_pl, reverse=True):
        ws_pl.delete_rows(idx, 1)
    for idx in sorted(rows_to_delete_stmt, reverse=True):
        ws_stmt.delete_rows(idx, 1)
    
    wb.save('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_141-20_tc1/output.xlsx')

if __name__ == '__main__':
    main()
