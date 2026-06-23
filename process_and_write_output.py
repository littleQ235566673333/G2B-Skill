import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_1/group_387-16/r2/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_1/group_387-16/r2/evolve_387-16/output.xlsx'

# Load data with no header, manually define column names based on file structure
df = pd.read_excel(input_path, sheet_name='Sheet1', header=None)
subdf = df.iloc[1:18, [0,1,3,9]].copy() # Only relevant columns for user output
subdf.columns = ['Value', 'Binaries', 'Result values', 'Difference']

# Work only on first 17 data rows (A2:D18 as required)
values = subdf['Value'].tolist()
binaries = subdf['Binaries'].tolist()
result_values = subdf['Result values'].tolist()

# Find and remove only one instance of any value in result_values, if present, from Value+Binaries
cleaned_values = []
cleaned_binaries = []
removed_idxs = set()
for rv in result_values:
    # Remove only first instance
    for idx, v in enumerate(values):
        if idx not in removed_idxs and v == rv:
            removed_idxs.add(idx)
            break
# Output lists without removed values
for idx, (v, b) in enumerate(zip(values, binaries)):
    if idx not in removed_idxs:
        cleaned_values.append(v)
        cleaned_binaries.append(b)

# Pad or shrink to 17 rows
while len(cleaned_values) < 17:
    cleaned_values.append('')
    cleaned_binaries.append('')
if len(cleaned_values) > 17:
    cleaned_values = cleaned_values[:17]
    cleaned_binaries = cleaned_binaries[:17]

subdf['Value'] = cleaned_values
subdf['Binaries'] = cleaned_binaries

# Solver result: sum of non-blank values in cleaned_values
def safe_num(v):
    try:
        return float(v)
    except: return 0
solver_result = sum(safe_num(v) for v in cleaned_values if v != '')

# Try to determine target (from top-right cell [1,3] or column 'Result values' of first non-blank)
target_value = None
try:
    tv = df.iloc[1,3]
    target_value = float(tv)
except:
    # fallback: use first numeric in result_values
    for val in result_values:
        try:
            target_value = float(val)
            break
        except:
            continue

# Compute and assign
subdf['Solver result'] = [''] * 17
subdf['Difference'] = [''] * 17
if target_value is not None:
    subdf.at[0,'Solver result'] = solver_result
    subdf.at[0,'Difference'] = solver_result - target_value

# Write back to output as A2:D18
wb = load_workbook(input_path)
ws = wb['Sheet1']
for r, row in enumerate(subdf[['Value','Binaries','Solver result','Difference']].values, start=2):
    for c, v in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=v)
wb.save(output_path)
