import openpyxl
from datetime import datetime

# Load the workbook and active worksheet
wb = openpyxl.load_workbook('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_59224_tc1/input.xlsx')
ws = wb.active

# Read Project Start Date (B2)
B2 = ws['B2'].value
if isinstance(B2, str):
    try:
        B2 = datetime.strptime(B2, '%Y-%m-%d')
    except Exception:
        B2 = None

found = False
# Loop over E4:E14
for row in range(4, 15):
    c = ws[f'C{row}'].value
    d = ws[f'D{row}'].value
    # Handle date string conversion
    for label, v in [('c', c), ('d', d)]:
        if isinstance(v, str) and v.strip() == '':
            if label == 'c': c = None
            else: d = None
        elif isinstance(v, str):
            try:
                dt = datetime.strptime(v, '%Y-%m-%d')
            except Exception:
                dt = None
            if label == 'c': c = dt
            else: d = dt

    if B2 and c and d:
        if not found and c < B2 < d:
            found = True
            ws[f'E{row}'] = 'Select Period'
        elif not found:
            ws[f'E{row}'] = 'Select Period'
        else:
            ws[f'E{row}'] = 'Do Not Select'
    else:
        ws[f'E{row}'] = 'Do Not Select'

wb.save('results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_59224_tc1/output.xlsx')
