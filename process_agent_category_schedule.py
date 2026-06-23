import openpyxl
import datetime

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_53161_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_53161_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# --- Extract agent-category matrix ---
category_header_row = 3  # 1-based
category_start_col = 6
category_end_col = 14
agent_start_row = 4
agent_end_row = 8
categories = [ws.cell(row=category_header_row, column=col).value for col in range(category_start_col, category_end_col+1)]
agents = [ws.cell(row=row, column=5).value for row in range(agent_start_row, agent_end_row+1)]
# Dict: agent -> [1/None per category]
agent_categories = {
    ws.cell(row=row, column=5).value: [ws.cell(row=row, column=col).value for col in range(category_start_col, category_end_col+1)]
    for row in range(agent_start_row, agent_end_row+1)
}

# --- Extract agent schedule matrix ---
sched_header_row = 12
sched_start_col = 6
sched_end_col = 31
sched_time_cols = [ws.cell(row=sched_header_row, column=col).value for col in range(sched_start_col, sched_end_col+1)]
# Dict: agent -> [1/None per interval]
agent_schedules = {
    ws.cell(row=row, column=5).value: [ws.cell(row=row, column=col).value for col in range(sched_start_col, sched_end_col+1)]
    for row in range(13, 18)
}

# --- Calculate for each category (row) and time interval (col) ---
output_start_row = 22
output_start_col = 6

def get_agent_idx(agent):
    try:
        return agents.index(agent)
    except:
        return None

def get_cat_idx(cat):
    try:
        return categories.index(cat)
    except:
        return None

for cat_i, cat in enumerate(categories):
    out_row = output_start_row + cat_i
    for t_i, t in enumerate(sched_time_cols):
        out_col = output_start_col + t_i
        # For this interval, count eligible scheduled agents
        count = 0
        for agent in agents:
            ac_idx = get_agent_idx(agent)
            if ac_idx is None:
                continue
            category_allowed = agent_categories[agent][cat_i]
            scheduled = agent_schedules[agent][t_i]
            if scheduled == 1 and category_allowed == 1:
                count += 1
        ws.cell(row=out_row, column=out_col).value = count

wb.save(output_path)
print('Output written to:', output_path)
