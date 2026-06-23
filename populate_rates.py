from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/group_44017/r0/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v2/train/iter_1/group_44017/r0/evolve_44017/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active

# Excel uses 1-based columns
month_col_start = 30  # AD
month_col_end = 41    # AO
row_start = 14
row_end = 42
base_col = 23  # W
freq_col = 10  # J
start_col = 12 # L
eff_date_col = 'L'
freq_col_letter = get_column_letter(freq_col)
base_col_letter = get_column_letter(base_col)
incr_cols = [13, 14, 15, 16]  # M,N,O,P
incr_col_letters = [get_column_letter(c) for c in incr_cols]
col_letters = [get_column_letter(c) for c in range(month_col_start, month_col_end+1)]

for row in range(row_start, row_end+1):
    for idx, col_letter in enumerate(col_letters):
        month_cell = f'{col_letter}{row}'
        # Excel: only calculate if month >= effective date
        # We'll use (column date - effective date)/frequency to index which waves have occurred
        formula = f'=IF({col_letter}$9>=$L{row}, $W{row}*'
        # Build the cumulative product
        wave_factors = []
        for wave, incr_col_letter in enumerate(incr_col_letters):
            # The increase is included if floor((month_date - eff_date)/freq) >= wave
            # This is: IF((({col}$9-$L{row})/$J{row})>={wave}, 1+$M{row}, 1)
            factor = f'IF((({col_letter}$9-$L{row})/${freq_col_letter}{row})>={wave},1+${incr_col_letter}{row},1)'
            wave_factors.append(factor)
        formula += '*'.join(wave_factors)
        formula += ',"")'  # Else blank
        ws[month_cell] = formula
        # Remove yellow fill if present
        ws[month_cell].fill = PatternFill(fill_type=None)

wb.save(output_path)
