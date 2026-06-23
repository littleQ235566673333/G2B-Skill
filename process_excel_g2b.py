from openpyxl import load_workbook
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_1/regression_gate/after_pass/core_55421/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_1/regression_gate/after_pass/core_55421/output.xlsx'

# Read with pandas to ease filtering
df = pd.read_excel(input_path, sheet_name=0, usecols='A:F', header=None)
df.columns = ['A', 'B', 'C', 'D', 'E', 'F']

wb = load_workbook(input_path)
ws = wb.active

results = [''] * 19  # F2:F20

# Check for each row, grouping by Column A (number)
for idx in range(1, 20):  # Excel rows 2 to 20; pandas rows 1 to 19
    number = df.at[idx, 'A']
    # Find all rows with this number (including all appearances in A2:A20)
    group = df[df['A'] == number]
    statuses = set(str(x).strip().upper() for x in group['D'] if pd.notnull(x))
    # 1. Only SCH
    if statuses == {'SCH'}:
        val = 'FUTURE'
    # 2. SCH and NO SHOW
    elif statuses == {'NO SHOW', 'SCH'}:
        val = 'NS/SCHED'
    # 3. Only NO SHOW(s), check E for any date
    elif statuses == {'NO SHOW'}:
        ns_with_date = any(pd.notnull(row['E']) and str(row['E']).strip() != '' for i, row in group.iterrows() if str(row['D']).upper() == 'NO SHOW')
        if ns_with_date:
            val = 'NO ACTION NEEDED'
        else:
            val = 'CALL PT'
    else:
        val = ''
    ws.cell(row=idx+1, column=6).value = val
wb.save(output_path)
