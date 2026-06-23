import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/regression_gate/before_pass/core_3413/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/regression_gate/before_pass/core_3413/output.xlsx'

# Read the relevant table section only (skip headers/extra rows)
df_raw = pd.read_excel(input_path, header=None)
# The left (source) table starts at row 2, columns A-C (0-based: row 1, cols 0,1,2)
df = df_raw.iloc[1:6, 0:3].copy()
df.columns = ['Department', 'RU', 'Value']
df['Value'] = pd.to_numeric(df['Value'], errors='coerce').fillna(0)

wb = load_workbook(input_path)
ws = wb.active

def sum_for_condition(dept, ru):
    mask_dept = df['Department'] == dept
    mask_ru = df['RU'] == ru
    both_mask = mask_dept & mask_ru
    if both_mask.any() and dept is not None and ru is not None:
        return df.loc[both_mask, 'Value'].sum()
    else:
        return df.loc[mask_dept, 'Value'].sum()

for excel_row in range(3, 7):
    dept = ws[f'E{excel_row}'].value
    ru = ws[f'F{excel_row}'].value
    ws[f'G{excel_row}'] = sum_for_condition(dept, ru)

wb.save(output_path)
