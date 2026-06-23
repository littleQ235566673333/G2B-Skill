from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/before_fix/core_54274/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/train/iter_6/regression_gate/before_fix/core_54274/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active  # Assume single sheet

# Place the nested IF formula as per requirements in G4
correct_formula = (
    "=IF(B4<=C4,B4/C4,"
    "IF(B4<D4,MIN(B4/C4,1.09),"
    "IF(B4=D4,MIN(B4/C4,1.1),"
    "IF(B4<E4,MIN(B4/C4,1.19),MIN(B4/C4,1.2)))))"
)

ws['G4'] = correct_formula

wb.save(output_path)
