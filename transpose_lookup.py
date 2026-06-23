from openpyxl import load_workbook
import pandas as pd
import collections

# File paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_8/regression_gate/before_fix/core_38985/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_8/regression_gate/before_fix/core_38985/output.xlsx"

# Load workbook and worksheet
wb = load_workbook(input_path)
ws = wb.active  # Assuming first sheet

# Table 1 (names in B8:B11; output D8:D11 as per instructions)
table1_start, table1_end = 8, 11  # rows, inclusive
table1_names = [ws[f'B{row}'].value for row in range(table1_start, table1_end+1)]

# Table 2 (Assume names in C, data in D; heuristically search rows 1-30 for filled rows)
table2 = []
for row in ws.iter_rows(min_row=1, max_row=30, min_col=3, max_col=4):    # C and D
    name = row[0].value
    val = row[1].value
    if name and str(name).strip().lower() not in ("name", "names"):
        table2.append((name, val))

# Map names to all corresponding values
name_to_vals = collections.defaultdict(list)
for name, val in table2:
    name_to_vals[str(name).strip()].append(val)

# For each name in table1, output their table2 values transposed in D/E/F... columns
for idx, name in enumerate(table1_names):
    vals = name_to_vals.get(str(name).strip(), [])
    for j, v in enumerate(vals):
        ws.cell(row=table1_start+idx, column=4+j).value = v

wb.save(output_path)
