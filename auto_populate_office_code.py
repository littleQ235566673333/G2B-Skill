from openpyxl import load_workbook

# Paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/before_fix/core_58701/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/regression_gate/before_fix/core_58701/output.xlsx'

wb = load_workbook(input_path)

# Table Tab: build dict from Location Name => Office Code
tab_ws = wb['Table Tab']
tab_headers = [cell.value for cell in next(tab_ws.iter_rows(min_row=1, max_row=1))]
tab_field_to_col = {name: i for i, name in enumerate(tab_headers)}

location_to_code = {}
for row in tab_ws.iter_rows(min_row=2, max_col=2, values_only=True):
    loc, code = row
    if loc and code is not None:
        location_to_code[loc] = code

# Entry Tab: find which col is 'Location' and which is 'Office Code'
entry_ws = wb['Entry Tab']
entry_headers = [cell.value for cell in next(entry_ws.iter_rows(min_row=1, max_row=1))]
entry_field_to_col = {name: i for i, name in enumerate(entry_headers)}

# Defensive fallback: Locate columns by scanning header row
loc_col = entry_field_to_col.get('Location', 1)  # default B (index 1)
office_code_col = entry_field_to_col.get('Office Code', 4)  # default E (index 4)

# Fill in cells E2:E3 (rows 2 and 3)
for row_idx in range(2, 4):
    loc_cell = entry_ws.cell(row=row_idx, column=loc_col+1)
    code_cell = entry_ws.cell(row=row_idx, column=office_code_col+1)
    location = loc_cell.value
    code = location_to_code.get(location, None)
    code_cell.value = code

wb.save(output_path)
