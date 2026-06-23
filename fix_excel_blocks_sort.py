import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_191-40_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_191-40_tc1/output.xlsx'
sheetname = 'Sheet1'

wb = openpyxl.load_workbook(input_path, data_only=False)
ws = wb[sheetname]

rows = list(ws.iter_rows(min_row=1, max_row=85, max_col=8, values_only=False))
block_starts = []
block_ends = []

in_block = False
for idx, row in enumerate(rows):
    if not in_block and any(cell.value is not None for cell in row):
        block_starts.append(idx)
        in_block = True
    elif in_block and all(cell.value is None for cell in row):
        block_ends.append(idx-1)
        in_block = False
if in_block:
    block_ends.append(len(rows)-1)

wb.calculation.active = True

for bstart, bend in zip(block_starts, block_ends):
    block_rows = rows[bstart:bend+1]
    sortable = []
    for row in block_rows:
        dcell = row[3]  # D
        # Convert text formulas to real formulas if needed
        if isinstance(dcell.value, str) and dcell.value.startswith('='):
            ws.cell(row=dcell.row, column=4).value = dcell.value
            ws.cell(row=dcell.row, column=4).data_type = 'f'
        # Get evaluated value for sorting (using data_only workbook)
        try:
            value_for_sort = wb[sheetname].cell(row=dcell.row, column=4).value
            if isinstance(value_for_sort, str) and value_for_sort.startswith('='):
                value_for_sort = 0
            sortable.append((float(value_for_sort) if value_for_sort is not None else 0, row))
        except Exception:
            sortable.append((0, row))
    sortable = sorted(sortable, key=lambda x: x[0], reverse=True)

    for idx, (_, sorted_row) in enumerate(sortable):
        for col in range(8):
            cell = sorted_row[col]
            target_cell = ws.cell(row=bstart+idx+1, column=col+1)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                target_cell.value = cell.value
                target_cell.data_type = 'f'
            else:
                target_cell.value = cell.value
    # Set exactly one blank row after the block
    if bend+1 < len(rows):
        for col in range(8):
            ws.cell(row=bend+2, column=col+1).value = None

# Formatting in D–F: bold, center, numeric format
bold_center = Font(bold=True)
alignment_center = Alignment(horizontal='center')

for row in ws.iter_rows(min_row=1, max_row=85, min_col=4, max_col=6):
    for cell in row:
        cell.font = bold_center
        cell.alignment = alignment_center
        val = cell.value
        try:
            if cell.data_type == 'f':
                cell.number_format = '0.0'
            elif val is not None:
                fval = float(val)
                cell.number_format = '0' if fval == int(fval) else '0.0'
        except Exception:
            cell.number_format = 'General'

for col in range(1, 9):
    width = ws.column_dimensions[get_column_letter(col)].width
    ws.column_dimensions[get_column_letter(col)].width = width

wb.save(output_path)
