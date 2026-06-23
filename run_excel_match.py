from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_6/group_59358/r2/evolve_59358/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_6/group_59358/r2/evolve_59358/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# For L7, L8, L9, do the matching as described
for l_row in [7, 8, 9]:
    # The value to search for (from L cell)
    search_value = ws[f'L{l_row}'].value
    result = None
    # Search each pair of (B2:B23, D2:D23)
    for i in range(2, 24):
        b_val = ws[f'B{i}'].value
        d_val = ws[f'D{i}'].value
        # Check if the value exists in both B and D in the same row
        if b_val == d_val and b_val == search_value and d_val is not None:
            # Found the first match
            result = ws[f'F{i}'].value
            break
    # Otherwise, as per user, find first occurrence B in D, return F at B (if more matches expected)
    if result is None:
        for i in range(2, 24):
            b_val = ws[f'B{i}'].value
            # Only trigger if current B matches search_value and this value appears in column D
            if b_val == search_value and b_val is not None:
                for j in range(2, 24):
                    if ws[f'D{j}'].value == b_val:
                        result = ws[f'F{i}'].value
                        break
                if result:
                    break
    ws[f'L{l_row}'] = result

wb.save(output_path)
