import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/group_227-40/r2/evolve_227-40/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_7/group_227-40/r2/evolve_227-40/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Get the data from rows 2-4, cols A-E (original data with potential row duplicates)
data_rows = list(ws.iter_rows(min_row=2, max_row=4, min_col=1, max_col=5, values_only=True))

def dedup_row(row):
    seen = set()
    result = []
    for v in row:
        if v is not None and v not in seen:
            seen.add(v)
            result.append(v)
        elif v is not None:
            result.append(None)
    # If less than 5 values, pad with None
    if len(result) < 5:
        result += [None] * (5 - len(result))
    return result

new_rows = [dedup_row(row) for row in data_rows]

# Write result to A14:E17
start_row = 14
for r, row in enumerate(new_rows):
    for c, val in enumerate(row):
        ws.cell(row=start_row + r, column=1 + c, value=val)

wb.save(output_path)
