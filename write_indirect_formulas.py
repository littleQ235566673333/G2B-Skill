from openpyxl import load_workbook

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_6/regression_gate/before_pass/core_41601/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_6/regression_gate/before_pass/core_41601/output.xlsx"
wb = load_workbook(input_path)
ws = wb["Students"]
for row in range(2, 8):
    student_name = ws.cell(row=row, column=1).value
    cell = ws.cell(row=row, column=5)
    if student_name:
        formula = "=INDIRECT(\"'\"&A{}&\"'!C2\")".format(row)
        cell.value = formula
    else:
        cell.value = None
wb.save(output_path)
