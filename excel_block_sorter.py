import openpyxl
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
import re
import os

# File paths
i_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_191-40_tc1/input.xlsx"
o_path = "results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42/eval_191-40_tc1/output.xlsx"
start_cell = 'A1'
end_cell = 'H85'
sheet_name = 'Sheet1'

# Number formats
def num_format(val):
    if val is None:
        return 'General'
    try:
        fval = float(val)
        if fval == int(fval):
            return '0'  # No decimal point for whole numbers
        else:
            return '0.0'  # One decimal place for decimals
    except:
        return 'General'

# Helper to check if a row is completely blank
def is_blank_row(row):
    return all([cell.value in (None, '') for cell in row])

# Helper to identify text-formulas and convert back to formulas
def fix_formula_text(cell):
    # Looks for '=SOMETHING' as string, replaces with actual formula
    if cell.data_type == 's' and str(cell.value).strip().startswith('='):
        cell.value = str(cell.value).strip()
        cell.data_type = 'f'

# Open workbook and sheet
wb = openpyxl.load_workbook(i_path)
ws = wb[sheet_name]

# Preserve column widths
col_widths = {col: ws.column_dimensions[get_column_letter(col)].width for col in range(1, 9)}

# Read all rows in the region
rows = list(ws[start_cell:end_cell])

# Identify block boundaries (row indexes)
blocks = []
cur_block = []
for idx, row in enumerate(rows):
    if is_blank_row(row):
        if cur_block:
            blocks.append(cur_block)
            cur_block = []
    else:
        cur_block.append(idx)
# Add last block if present
if cur_block:
    blocks.append(cur_block)

# Process each block
output_data = [[] for _ in range(len(rows))]  # Pre-allocate output
row_cursor = 0
for block in blocks:
    block_cells = [rows[i] for i in block]
    # Fix formulas (D,E,F == 4,5,6)
    for r, row in enumerate(block_cells):
        for col in [3,4,5]:  # 0-based index (D,E,F)
            cell = row[col]
            fix_formula_text(cell)
    # Evaluate column D for sorting; if formula, leave as is, but get value
    # As openpyxl cannot evaluate formulas, sort by current calculated values
    sortable = []
    for r, row in enumerate(block_cells):
        d = row[3].value
        d_disp = None
        try:
            d_disp = float(d) if d is not None else float('-inf')
        except:
            # If formula or cannot convert, use -inf to send to bottom
            d_disp = float('-inf')
        sortable.append((d_disp, r, row))
    sortable_sorted = sorted(sortable, key=lambda t: t[0], reverse=True)
    # After sorting, adjust formulas in D-F to correct row
    # Copy block rows to output, re-number formulas
    rel_start = row_cursor
    for new_r, (_, old_r_idx, old_row) in enumerate(sortable_sorted):
        for c in range(8):
            src_cell = old_row[c]
            tgt_row_idx = row_cursor
            tgt_col_idx = c
            tgt_cell = ws.cell(row=1+tgt_row_idx, column=1+tgt_col_idx)
            # Copy value or formula
            if src_cell.data_type == 'f' or (isinstance(src_cell.value, str) and src_cell.value.startswith('=')):
                # Translate the formula
                if isinstance(src_cell.value, str):
                    f = src_cell.value
                else:
                    f = src_cell.value
                old_row_num = src_cell.row
                new_row_num = tgt_cell.row
                # Translate formula to new row
                new_fmla = Translator(f, origin=f'{get_column_letter(c+1)}{old_row_num}').translate_formula(f'{get_column_letter(c+1)}{new_row_num}')
                tgt_cell.value = new_fmla
            else:
                tgt_cell.value = src_cell.value
            # Formatting
            tgt_cell.font = src_cell.font.copy()
            tgt_cell.fill = src_cell.fill.copy()
            tgt_cell.border = src_cell.border.copy()
            tgt_cell.alignment = src_cell.alignment.copy()
            tgt_cell.number_format = src_cell.number_format
        row_cursor += 1
    # Add one blank row after block (unless last block)
    if block != blocks[-1]:
        row_cursor += 1

# Set all columns to original widths
for col in range(1, 9):
    letter = get_column_letter(col)
    ws.column_dimensions[letter].width = col_widths[col]

# Apply number + style formatting to D,E,F columns over whole area
for row in ws[start_cell:end_cell]:
    for col in [4,5,6]: # D,E,F
        cell = row[col-1]
        val = None
        if cell.data_type == 'f':
            # Can't evaluate formula, so just use number format
            cell.number_format = '0.0'
        else:
            try:
                val = float(cell.value)
                if val == int(val):
                    cell.number_format = '0'
                else:
                    cell.number_format = '0.0'
            except:
                cell.number_format = 'General'
        # Bold and center
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

# Save output
wb.save(o_path)
