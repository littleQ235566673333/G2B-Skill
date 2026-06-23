import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_55708_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-A/eval_55708_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

def normalize_header(h):
    return h.strip().lower().replace(' ', '') if h else ''

header_row = 1
headers = {normalize_header(cell.value): idx for idx, cell in enumerate(ws[header_row], start=1)}

# Find the right columns
status_col = headers.get('status')
tat_col = headers.get('turnaroudtime')  # matches actual header spelling
# Try fallback for 'turnaroundtime' in case of typo fix
if not tat_col:
    tat_col = headers.get('turnaroundtime')
department_col = headers.get('department')

# Department names and output cells
department_cells = ['A11', 'A12', 'A13']
output_cells = ['B11', 'B12', 'B13']
departments = [ws[cell].value for cell in department_cells]

rows = list(ws.iter_rows(min_row=2, values_only=True))

for dept, cell in zip(departments, output_cells):
    if not dept:
        ws[cell] = ''
        continue
    tat_values = []
    for row in rows:
        row_status = row[status_col - 1]
        row_tat = row[tat_col - 1]
        row_dept = row[department_col - 1]
        if (
            row_dept == dept and
            row_status in ('In Progress', 'In Review') and
            isinstance(row_tat, (int, float)) and
            row_tat >= 6
        ):
            tat_values.append(row_tat)
    if tat_values:
        avg_tat = sum(tat_values) / len(tat_values)
        ws[cell] = avg_tat
    else:
        ws[cell] = ''

wb.save(output_path)
