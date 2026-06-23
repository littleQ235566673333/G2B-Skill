import pandas as pd
from openpyxl import load_workbook
import numpy as np
import os

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_3/regression_gate/after_fix/core_177-6/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_3/regression_gate/after_fix/core_177-6/output.xlsx'
ref_col_letter = 'H'
ref_col_idx = ord(ref_col_letter) - ord('A')
start_col = 0  # A
end_col = 18   # R (python exclusive)

# 1. Read with pandas
pd_all = pd.read_excel(input_path, None)
src_sheet = list(pd_all.keys())[0]
df = pd_all[src_sheet]
df_cut = df.iloc[:, start_col:end_col+1].copy()
columns = df_cut.columns.tolist()
ref_col = columns[ref_col_idx]

def numeric_merge(vals):
    vals = [float(x) if pd.notnull(x) and str(x).strip() not in ('', '-') else 0 for x in vals]
    if all(np.isclose(v, 0) for v in vals):
        return 0
    return sum(vals)

agg_dict = {}
for i, c in enumerate(columns):
    if i < 8:  # A-H
        agg_dict[c] = 'first'
    else:      # I-R
        agg_dict[c] = numeric_merge

merged = df_cut.groupby(ref_col, dropna=False).agg(agg_dict).reset_index(drop=True)

# 2. Write output sheet, keep formatting
wb = load_workbook(input_path)
if 'combined' in wb.sheetnames:
    wb.remove(wb['combined'])
ws_c = wb.create_sheet('combined')
ws_src = wb[src_sheet]

# Write header
for j, col in enumerate(columns, 1):
    ws_c.cell(row=1, column=j).value = col

# Write merged rows (max 7, per output A1:R8, header + 7 rows)
for r, row in merged.iloc[:7].iterrows():  # 7 data rows
    excel_row = r + 2
    ref_value = row[ref_col]
    src_row = df_cut[df_cut[ref_col] == ref_value].index[0] + 2
    for j, col in enumerate(columns, 1):
        val = row[col]
        cell = ws_c.cell(row=excel_row, column=j, value=val)
        src_cell = ws_src.cell(row=src_row, column=j)
        # Copy style
        if src_cell.has_style:
            cell._style = src_cell._style
        if j >= 9:  # I-R
            try:
                fval = float(val)
                if np.isclose(fval, 0):
                    cell.value = None
                else:
                    cell.value = f'{fval:.2f}'
            except Exception:
                pass
wb.save(output_path)
assert os.path.exists(output_path)
from openpyxl import load_workbook as lbout
wb_out = lbout(output_path)
ws_out = wb_out['combined']
vals = [[ws_out.cell(row=x, column=y).value for y in range(1, 19)] for x in range(1, 9)]
assert len(vals) == 8
