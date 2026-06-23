import openpyxl
from datetime import datetime

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_58723_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_58723_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Collect Entry Times for each Name
name_to_times = {}
for row in range(2, 42):  # Rows 2 to 41
    name = ws[f'C{row}'].value
    entry_time_raw = ws[f'I{row}'].value
    if not name or not entry_time_raw:
        continue
    # Convert entry_time to datetime if not already
    if isinstance(entry_time_raw, datetime):
        entry_time = entry_time_raw
    else:
        try:
            entry_time = datetime.strptime(str(entry_time_raw), '%Y-%m-%d %H:%M:%S')
        except Exception:
            entry_time = entry_time_raw
    if name not in name_to_times:
        name_to_times[name] = []
    name_to_times[name].append((row, entry_time))

# Find latest time for each Name
name_to_latest_time = {}
for name, entries in name_to_times.items():
    # Ignore entries with invalid datetime
    valid_entries = [(row, et) for row, et in entries if isinstance(et, datetime)]
    if valid_entries:
        latest_time = max(et for _, et in valid_entries)
        name_to_latest_time[name] = latest_time
    else:
        name_to_latest_time[name] = None

# Write into Column M
for name, entries in name_to_times.items():
    latest_time = name_to_latest_time[name]
    for row, entry_time in entries:
        cell = ws[f'M{row}']
        # Mark 'Latest' if entry_time matches latest_time, else 'Not Latest'
        if latest_time and isinstance(entry_time, datetime) and entry_time == latest_time:
            cell.value = 'Latest'
        else:
            cell.value = 'Not Latest'

wb.save(output_path)
