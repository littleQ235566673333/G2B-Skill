import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_1/group_387-16/r3/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/train/iter_1/group_387-16/r3/evolve_387-16/output.xlsx'

# Read the full sheet with pandas (gets all columns for easy manipulation)
df = pd.read_excel(input_path, sheet_name='Sheet1', header=None)

# Identify columns by inspecting row 0
colnames = df.iloc[0].tolist()
df.columns = colnames
df = df[1:]
df.reset_index(drop=True, inplace=True)

# Locate target columns
val_col = 'Value'
bin_col = 'Binaries'
result_col = 'Result values'
target_value_col = 'Target value'
A_col = colnames[0]

# Make sure 'Value', 'Binaries', and 'Result values' exist
if not all(c in df.columns for c in [val_col, bin_col, result_col]):
    raise Exception('Expected columns not found in input.')

# Remove only ONE instance for each matching value
def remove_one_instance_per_match(df, val_col, bin_col, result_col):
    vals = df[val_col].tolist()
    bins = df[bin_col].tolist()
    matched = df[result_col].dropna().tolist()
    remove_indices = []
    used_idxs = set()
    for m in matched:
        for i, v in enumerate(vals):
            if i in used_idxs:
                continue
            if v == m:
                remove_indices.append(i)
                used_idxs.add(i)
                break
    mask = [i not in remove_indices for i in range(len(vals))]
    # rebuild the columns
    vals_new = [v for i, v in enumerate(vals) if mask[i]]
    bins_new = [b for i, b in enumerate(bins) if mask[i]]
    # pad with empty to original length
    pad_len = len(vals) - len(vals_new)
    vals_new.extend([''] * pad_len)
    bins_new.extend([''] * pad_len)
    df[val_col] = vals_new
    df[bin_col] = bins_new
    return df

df = remove_one_instance_per_match(df, val_col, bin_col, result_col)

# Remove blank cells by shifting up nonempty in 'Value' and 'Binaries'
def compress_column(df, col):
    non_blanks = [v for v in df[col] if pd.notnull(v) and v != '']
    pad_len = len(df) - len(non_blanks)
    non_blanks.extend([''] * pad_len)
    df[col] = non_blanks
    return df

df = compress_column(df, val_col)
df = compress_column(df, bin_col)

# Compute solver result as sum of column A
try:
    sum_A = df[A_col].apply(pd.to_numeric, errors='coerce').sum()
except Exception:
    sum_A = ''

# Difference as Solver Result - Target Value
try:
    target = df[target_value_col].dropna().iloc[0]
    diff = sum_A - float(target)
except Exception:
    diff = ''

# Write columns: A-D. Range A2:D18 (so, rows 1-17 in 0-index)
# Compose export
export = df.iloc[:17, :4].copy()
export.columns = colnames[:4]
export.reset_index(drop=True, inplace=True)
export[A_col] = export[A_col].apply(lambda x: '' if pd.isnull(x) else x)

# Load workbook, set A2:D18
wb = load_workbook(input_path)
ws = wb['Sheet1']
# Write A2:D18
for rowi in range(17):
    for coli in range(4):
        val = export.iloc[rowi, coli] if rowi < len(export) else ''
        ws.cell(row=2+rowi, column=1+coli).value = val
# Write solver result & difference if there is space: e.g. D2, D3
if ws['A2'].value != '':
    ws['D2'] = sum_A
    ws['D3'] = diff
wb.save(output_path)
