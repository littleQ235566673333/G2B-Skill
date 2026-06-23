import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/after_pass/core_32337/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/after_pass/core_32337/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# The headers are actually on the second row (row 2)
headers = {cell.value: idx+1 for idx, cell in enumerate(ws[2]) if cell.value}
dob_col = headers.get('DATE OF BIRTH')
# 'REPORT Date' is always in cell A1
report_date_cell = ws.cell(row=1, column=1)
report = report_date_cell.value if isinstance(report_date_cell.value, datetime) else None
age_col = headers.get('age (year)')
category_col = headers.get('CATEGORY')
expected_col = headers.get('Expected Result')

# Insert "Actual Age" column after 'age (year)'
actual_age_col = age_col + 1 if age_col else ws.max_column + 1
ws.insert_cols(actual_age_col)
ws.cell(row=2, column=actual_age_col, value='Actual Age')

# Get just the PatternFill for formatting
preceding_fill = ws.cell(row=3, column=actual_age_col-1).fill
if not isinstance(preceding_fill, PatternFill):
    preceding_fill = PatternFill()

# Format for header
top_header = ws.cell(row=2, column=actual_age_col)
top_header.font = Font(bold=True)
top_header.alignment = Alignment(horizontal='center', vertical='top')
top_header.fill = preceding_fill

for row in range(3, 16):
    dob = ws.cell(row=row, column=dob_col).value
    # Calculate Actual Age
    if dob is not None and report is not None:
        if isinstance(dob, str):
            dob = datetime.strptime(dob[:10], '%Y-%m-%d')
        actual_age = report.year - dob.year - ((report.month, report.day) < (dob.month, dob.day))
    else:
        actual_age = ''
    acell = ws.cell(row=row, column=actual_age_col, value=actual_age)
    acell.font = Font(bold=True)
    acell.alignment = Alignment(horizontal='center')
    acell.fill = preceding_fill
    acell.number_format = '0'

# Formula for E3:E15 (Expected Result) -- get CATEGORY value from adjacent column
if category_col and expected_col:
    cat_letter = openpyxl.utils.get_column_letter(category_col)
    for row in range(3, 16):
        ws.cell(row=row, column=expected_col, value=f'={cat_letter}{row}')

wb.save(output_path)
