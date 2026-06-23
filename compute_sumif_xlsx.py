from openpyxl import load_workbook

def compute_sumif_xlsx(input_path, output_path):
    wb = load_workbook(input_path)
    dashboard = wb['Dashboard']
    data = wb['Data']

    # Read product and month from dashboard
    product = dashboard['C2'].value
    month = dashboard['C3'].value  # only M1 is present in Data

    # Read region choices
    regions = [dashboard['C11'].value, dashboard['C12'].value, dashboard['C13'].value]
    regions = [r for r in regions if r and r != 'All']
    all_selected = not regions  # If no regions are chosen, treat as All

    # Find columns in Data
    headers = [cell.value for cell in next(data.iter_rows(min_row=1, max_row=1))]
    product_col = headers.index('Product')
    region_col = headers.index('Region')
    m1_col = headers.index('M1')

    total = 0
    for row in data.iter_rows(min_row=2, values_only=True):
        row_product = row[product_col]
        row_region = row[region_col]
        row_m1 = row[m1_col]
        if row_product == product:
            if all_selected or row_region in regions:
                total += row_m1

    # Write result to Dashboard!C4
    dashboard['C4'].value = total
    wb.save(output_path)

compute_sumif_xlsx(
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_48365_tc1/input.xlsx',
    'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0-PRUNED/eval_seed42/eval_48365_tc1/output.xlsx'
)
