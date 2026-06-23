import openpyxl

# Input and output path
input_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_8/task_32789/r1/evolve_32789/input.xlsx"
output_path = "results/runs/g2b-v8_gpt-4.1_ss-gpt41/train/iter_8/task_32789/r1/evolve_32789/output.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Columns and range info
# Headers in BD:BR (56:70), secondary in row 3, primary in row 2
first_col, last_col = 56, 70
row_prim = 2
row_sec = 3

# Output target rows and columns
out_rows = range(4, 8)
out_cols = [3, 4]  # C and D

for idx, out_row in enumerate(out_rows):
    date_value = ws[f"A{out_row}"].value
    # Find matching column by date in row 2 (matching column in BD:BR)
    match_col = None
    for col in range(first_col, last_col+1):
        if ws.cell(row=row_prim, column=col).value == date_value:
            match_col = col
            break
    if match_col:
        # Get the secondary header (row 3)
        sec_value = ws.cell(row=row_sec, column=match_col).value
        # Find column where both primary and secondary header match
        found_col = None
        for col in range(first_col, last_col+1):
            if (ws.cell(row=row_prim, column=col).value == date_value and
                ws.cell(row=row_sec, column=col).value == sec_value):
                found_col = col
                break
        if found_col:
            val = ws.cell(row=out_row, column=found_col).value
            # Whole number formatting
            try:
                val = int(round(float(val)))
            except (ValueError, TypeError):
                val = "-"
        else:
            val = "-"
    else:
        val = "-"
    # Write result to column D (4)
    ws.cell(row=out_row, column=4, value=val)

# Clear C:D below row 7
top_clear = 8
max_row = ws.max_row
for col in out_cols:
    for row in range(top_clear, max_row+1):
        ws.cell(row=row, column=col, value=None)

wb.save(output_path)
print("Done")
