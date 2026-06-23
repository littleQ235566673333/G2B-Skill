import openpyxl
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42/eval_56419_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/eval_seed42/eval_56419_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Correct ARGB color format for openpyxl
fill = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid')

row_offset = 2    # H2 is where results start
max_rows = 26      # H2:H27
out_col = 8        # H

# Collect 'Type' and 'Quantity'
types = []
qtys = []
for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
    types.append(row[0])
    qtys.append(row[1])

# Filter non-zero indices
results = [types[i] for i in range(len(types)) if qtys[i] is not None and qtys[i] != 0]

# Write filtered type list to H2:H27, with color, handle exhaustion
for idx in range(max_rows):
    cell = ws.cell(row=row_offset+idx, column=out_col)
    if idx < len(results):
        val = results[idx]
        cell.value = val
        cell.fill = fill
    else:
        cell.value = None  # Do not set .fill to None

wb.save(output_path)
print('done')
