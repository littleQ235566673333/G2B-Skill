import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_45372_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_seed42/eval_45372_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

f_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
threshold_time = datetime.strptime('09:45', '%H:%M').time()

for row in range(2, 16):
    time_cell = ws[f'A{row}'].value
    b_val = ws[f'B{row}'].value
    c_val = ws[f'C{row}'].value
    output = ''
    current_time = None
    if isinstance(time_cell, datetime):
        current_time = time_cell.time()
    elif isinstance(time_cell, str):
        try:
            current_time = datetime.strptime(time_cell, '%H:%M').time()
        except Exception:
            current_time = None
    if current_time:
        if current_time < threshold_time:
            value = b_val
        else:
            value = c_val
        if value == 0 or value == '0':
            output = ''
        else:
            output = value
    ws[f'E{row}'].value = output
    ws[f'E{row}'].fill = f_fill

wb.save(output_path)
