import pandas as pd
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_2/group_387-16/r3/evolve_387-16/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_2/group_387-16/r3/evolve_387-16/output.xlsx'

# Read without headers
raw = pd.read_excel(input_path, sheet_name='Sheet1', header=None)

# Target value from row 1, col 4
try:
    target_value = float(raw.iloc[1, 4])
except:
    target_value = 0
# Data block from row 2 (skip headers), until 'Result values' in col 3
value_rows = []
for idx in range(2, len(raw)):
    if str(raw.iloc[idx,3]).strip() == 'Result values':
        result_values_start = idx + 1
        break
    value_rows.append(idx)
value_list = [raw.iloc[i,0] for i in value_rows]
binary_list = [raw.iloc[i,1] for i in value_rows]
original_A = value_list.copy()  # For sum

# Result values are in column 3, starting at result_values_start
result_values = []
for idx in range(result_values_start, len(raw)):
    val = raw.iloc[idx,3]
    if pd.isna(val):
        break
    result_values.append(val)

# Remove ONE instance of each result_value (preserve order, both lists shift)
removed_indices = set()
for val in result_values:
    try:
        idx = next(i for i, x in enumerate(value_list) if x == val and i not in removed_indices)
        removed_indices.add(idx)
    except StopIteration:
        continue
new_value_list = [x for i,x in enumerate(value_list) if i not in removed_indices]
new_binary_list = [x for i,x in enumerate(binary_list) if i not in removed_indices]
new_A_list = [x for i,x in enumerate(original_A) if i not in removed_indices]

# Shift and fill up to length 17 (for A2:A18)
N = 17
pad = N - len(new_A_list)
A_out = new_A_list + [None]*pad
Value_out = new_value_list + [None]*pad
Binaries_out = new_binary_list + [None]*pad
# Solver result: sum of non-None A_out
solver_result = sum([x for x in A_out if pd.notna(x)])
difference = solver_result - target_value

# Write with openpyxl to A2:D18
wb = load_workbook(input_path)
ws = wb['Sheet1']
for row in range(2, 19):  # rows 2-18 (17 rows)
    idx = row-2
    ws.cell(row=row, column=1).value = A_out[idx] if idx < len(A_out) else None
    ws.cell(row=row, column=2).value = Value_out[idx] if idx < len(Value_out) else None
    ws.cell(row=row, column=3).value = Binaries_out[idx] if idx < len(Binaries_out) else None
    if idx==0:
        ws.cell(row=row, column=4).value = solver_result
        ws.cell(row=row, column=5).value = difference
    else:
        ws.cell(row=row, column=4).value = None
        ws.cell(row=row, column=5).value = None
wb.save(output_path)
