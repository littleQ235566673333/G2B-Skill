import openpyxl
import numpy as np

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_55708_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_seed42/eval_55708_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Find header row
headers = {cell.value: i for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))}

dep_col = headers.get('Department')
status_col = headers.get('Status')
tat_col = headers.get('Turnaround Time')

# Read departments from A11:A13
departments = [ws[f'A{i}'].value for i in range(11, 14)]
results = []

for dep in departments:
    values = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[dep_col].value == dep and row[status_col].value in ('In Progress', 'In Review'):
            try:
                tat = float(row[tat_col].value)
            except (TypeError, ValueError):
                continue
            if tat >= 6:
                values.append(tat)
    if values:
        avg = np.mean(values)
        results.append(round(avg, 2))
    else:
        results.append(None)

for idx, val in enumerate(results):
    ws[f'B{11+idx}'].value = val

wb.save(output_path)
