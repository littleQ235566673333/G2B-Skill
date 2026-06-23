from openpyxl import load_workbook

def colnum_to_letter(n):
    result = ''
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_2/regression_gate/before_pass/core_41969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed0/train/iter_2/regression_gate/before_pass/core_41969/output.xlsx'
wb = load_workbook(input_path)
ws = wb.active
for i, col in enumerate([1, 4, 7]):
    start_col_letter = colnum_to_letter(col)
    end_col_letter = colnum_to_letter(col + 2)
    formula = f'=COUNTBLANK({start_col_letter}3:{end_col_letter}3)'
    ws.cell(row=6, column=i+1).value = formula
wb.save(output_path)
