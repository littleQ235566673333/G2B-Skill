import openpyxl

# File paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_57743_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_57743_tc1/output.xlsx"

# Load workbook and sheets
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Build lookup dict from column C (model numbers) and column D (prices)
lookup = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=4):
    model_cell, price_cell = row
    model = model_cell.value
    price = price_cell.value
    if model is not None:
        lookup[model] = price

# For A2:A19, lookup in dict and write result to B2:B19
for i in range(2, 20):  # Excel rows 2 to 19
    model = ws.cell(row=i, column=1).value  # Column A
    price = lookup.get(model, None)
    ws.cell(row=i, column=2).value = price if price is not None else ""

# Save result
wb.save(output_path)
