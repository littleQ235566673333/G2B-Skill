import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/train/iter_7/evolve_48975/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/train/iter_7/evolve_48975/output.xlsx'

# Load workbook
wb = openpyxl.load_workbook(input_path)
ws_input = wb['Input']
ws_output = wb['Output']

output_values = []

for row in ws_input.iter_rows(min_row=2):
    a = row[0].value
    b = row[1].value
    todo = row[4].value
    if todo and str(todo).strip().lower() == 'yes':
        if a is not None and str(a).strip() != '':
            output_values.append(a)
        if b is not None and str(b).strip() != '':
            output_values.append(b)

# Write up to 7 values to Output!B11:B17
for idx, val in enumerate(output_values[:7]):
    ws_output.cell(row=11+idx, column=2, value=val)

wb.save(output_path)
