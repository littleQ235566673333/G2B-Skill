import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_374-18_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42/eval_374-18_tc1/output.xlsx'

# Load workbook and target sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['Imported Data']

# Collect rows to delete where Column E (5) < 1 (except header)
rows_to_delete = []
for row in range(2, ws.max_row + 1):  # start from 2 to skip header
    val = ws.cell(row=row, column=5).value
    try:
        if val is not None and float(val) < 1:
            rows_to_delete.append(row)
    except (TypeError, ValueError):
        continue

# Delete rows in reverse order
for row in reversed(rows_to_delete):
    ws.delete_rows(row)

# Save after deletion
wb.save(output_path)

# Load output and create Sheet1, copy A2:G6 from Imported Data
wb_out = openpyxl.load_workbook(output_path)
ws_imp = wb_out['Imported Data']
if 'Sheet1' in wb_out.sheetnames:
    sheet1 = wb_out['Sheet1']
else:
    sheet1 = wb_out.create_sheet('Sheet1')

for i, row in enumerate(ws_imp.iter_rows(min_row=2, max_row=6, max_col=7, values_only=True), start=2):
    for j, val in enumerate(row, start=1):
        sheet1.cell(row=i, column=j, value=val)

wb_out.save(output_path)
