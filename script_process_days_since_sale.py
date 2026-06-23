from openpyxl import load_workbook
import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_7/regression_gate/before_fix/core_59902/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_7/regression_gate/before_fix/core_59902/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Read names and dates from F5:F28 (dates) and G5:G28 (names)
rows = list(ws.iter_rows(min_row=5, max_row=28, min_col=6, max_col=7, values_only=True))
output = []
last_sale = {}  # key: name, value: last date (datetime)

for idx, (date_val, name) in enumerate(rows):
    if date_val is None or name is None:
        output.append("")
        continue
    prev_date = last_sale.get(name)
    if prev_date is None:
        days = 0
    else:
        days = (date_val - prev_date).days
    last_sale[name] = date_val
    output.append(days)

# Write results to C5:C28
accounting_fmt = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
for i, val in enumerate(output):
    cell = ws.cell(row=5 + i, column=3)
    cell.value = val if val != "" else None
    cell.number_format = accounting_fmt

wb.save(output_path)
