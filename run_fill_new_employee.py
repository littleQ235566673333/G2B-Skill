import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_32093_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_32093_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

for row in range(2, 16):
    current = ws[f'B{row}'].value
    c = ws[f'C{row}'].value
    d = ws[f'D{row}'].value
    e = ws[f'E{row}'].value

    # Pick the first non-empty new employee, or current if all are empty
    if c:
        new_emp = c
    elif d:
        new_emp = d
    elif e:
        new_emp = e
    else:
        new_emp = current
    ws[f'F{row}'] = new_emp

wb.save(output_path)
