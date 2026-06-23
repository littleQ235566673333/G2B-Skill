import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/regression_gate/after_fix/core_59129/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_6/regression_gate/after_fix/core_59129/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Get header row for months (assume in E1:P1)
months = [ws.cell(row=1, column=col).value for col in range(5, 17)]
# Read employee data
start_dates = [ws[f'A{row}'].value for row in range(2, 23)]
end_dates = [ws[f'B{row}'].value for row in range(2, 23)]

def to_month(dt):
    if isinstance(dt, datetime): 
        return dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if isinstance(dt, str):
        try:
            # try parsing as YYYY-MM or YYYY-MM-DD
            if len(dt) >= 10:
                return datetime.strptime(dt[:10], '%Y-%m-%d').replace(day=1)
            else:
                return datetime.strptime(dt[:7], '%Y-%m').replace(day=1)
        except:
            return None
    return None

months = [to_month(m) for m in months]

headcount = []
for m in months:
    count = 0
    for start, end in zip(start_dates, end_dates):
        if not isinstance(start, datetime):
            continue
        still_employed = (start <= m) and (end is None or (isinstance(end, datetime) and end > m))
        count += int(still_employed)
    headcount.append(count)

# Write results to E2:P2
for j, val in enumerate(headcount):
    ws.cell(row=2, column=5 + j).value = val

wb.save(output_path)
