from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def is_green(cell):
    # Excel green (standard fill) is sometimes 'FF00B050', '00B050', or other forms
    if cell.fill and cell.fill.patternType == 'solid':
        rgb = getattr(cell.fill.fgColor, 'rgb', None) or getattr(cell.fill.fgColor, 'value', None)
        if rgb:
            return rgb.upper() in ['FF00B050', '00B050']
    return False

in_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun1/eval_43657_tc1/input.xlsx'
out_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/eval_seed42_rerun1/eval_43657_tc1/output.xlsx'

wb = load_workbook(in_path)
ws = wb.active

for idx, row in enumerate(range(2,9), start=2):
    name = ws[f'L{row}'].value
    count = 0
    for cell in ws.iter_rows(min_row=row, max_row=row, min_col=3, max_col=7):  # generator yields tuple for one row
        for cell_item in cell:
            if name and cell_item.value and str(name).lower() in str(cell_item.value).lower() and is_green(cell_item):
                count += 1
    ws[f'K{row}'] = count

wb.save(out_path)
