import pandas as pd
from openpyxl import load_workbook

# File paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/regression_gate/after_pass/core_3413/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_8/regression_gate/after_pass/core_3413/output.xlsx'

# Load data with pandas
# By default pandas reads the first sheet, which matches our scenario
# Columns are assumed as:
# A: Department (0), B: RU (1), C: Value (2), D, E: Dept criteria (4), F: RU criteria (5), G: Output (6)
df = pd.read_excel(input_path, dtype=str)
# Make sure numeric columns are converted
if df.columns.size >= 3:
    df.iloc[:,2] = pd.to_numeric(df.iloc[:,2], errors='coerce')

# Open with openpyxl as well for output
wb = load_workbook(input_path)
ws = wb.active

# For G3:G6 (Excel 1-based rows 3 to 6)
for row in range(3, 7):
    dept_criteria = ws[f'E{row}'].value
    ru_criteria = ws[f'F{row}'].value

    # If either criteria is missing, just leave blank
    if (dept_criteria is None) or (ru_criteria is None):
        ws[f'G{row}'] = None
        continue

    # Sum where both Department and RU match
    bothmask = (df.iloc[:,0] == str(dept_criteria)) & (df.iloc[:,1] == str(ru_criteria))
    if bothmask.any():
        val = df.loc[bothmask, df.columns[2]].sum()
    else:
        onlydept = (df.iloc[:,0] == str(dept_criteria))
        val = df.loc[onlydept, df.columns[2]].sum()

    ws[f'G{row}'] = float(val) if pd.notna(val) else ''

wb.save(output_path)
