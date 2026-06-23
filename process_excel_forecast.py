import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_45372_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_45372_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

color = 'FFC000'  # hex without # for openpyxl
fill = PatternFill(fill_type='solid', fgColor=color)

switch_time = datetime.strptime('09:45', '%H:%M')

for row in range(2, 16):  # rows 2 to 15 inclusive
    time_cell = ws[f'A{row}']
    b_value = ws[f'B{row}'].value
    c_value = ws[f'C{row}'].value
    val = None
    # Try to interpret time
    if isinstance(time_cell.value, str):
        try:
            cell_time = datetime.strptime(time_cell.value.strip(), '%H:%M')
        except Exception:
            cell_time = None
    elif isinstance(time_cell.value, datetime):
        cell_time = time_cell.value
    else:
        cell_time = None
    # Determine value
    if cell_time is not None:
        if cell_time < switch_time:
            val = b_value
        else:
            val = c_value
    else:
        val = None  # If time can't be interpreted, leave blank
    # Handle blanks and zeros
    if val == 0 or val is None:
        ws[f'E{row}'].value = ''
    else:
        ws[f'E{row}'].value = val
    ws[f'E{row}'].fill = fill

wb.save(output_path)
