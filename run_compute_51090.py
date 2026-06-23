import pandas as pd
import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_51090_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun1/eval_51090_tc1/output.xlsx'

# Relevant users for filtering errors
users = ['CHROGIL1','CHDSPOLJ','CHSJEFFE','CHBTHOMA']

# Load Daily Numbers
# skiprows=1 to drop the header row (row 2 is the real column headers)
df_daily = pd.read_excel(input_path, sheet_name='Daily Numbers', skiprows=1)
df_daily = df_daily[df_daily['WHS'].astype(str) == '27']

# Load Errors
err_df = pd.read_excel(input_path, sheet_name='Errors')
err_codes = ['II','IR','IT','OV','PI']

# Only consider errors for warehouse 27 & relevant users
err_27 = err_df[(err_df['Business Unit']==27) & (err_df['User ID'].isin(users)) & (err_df['Document Type'].isin(err_codes))]

# OV errors only (since only OV exists for warehouse 27)
ov_err_sum = err_27[err_27['Document Type']=='OV']['Transaction Quantity'].sum()

# Only one row in Daily Numbers for warehouse 27
result = []
for idx, row in df_daily.iterrows():
    cartons_received = row['Inbound Receipts']
    # For IR, only positive errors, but none exist for warehouse 27
    total_errors = ov_err_sum
    value = cartons_received - total_errors
    result.append(value)

# Load workbook for output
wb = openpyxl.load_workbook(input_path)
ws = wb['Daily Numbers']

# Write result into Q3:Q24 (only one date present, so others will be empty)
for i in range(3, 25):
    if i == 3 and result:
        ws['Q%d' % i] = result[0]
    else:
        ws['Q%d' % i] = None

wb.save(output_path)
