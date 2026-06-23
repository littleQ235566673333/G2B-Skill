import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/regression_gate/before_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_4/regression_gate/before_pass/core_41601/output.xlsx'

wb = openpyxl.load_workbook(input_path)
students_ws = wb['Students']
# Get student names from A2:A7
student_names = [students_ws[f'A{i}'].value for i in range(2, 8)]
ages = []
for name in student_names:
    age = None
    if name in wb.sheetnames:
        ws = wb[name]
        age = ws['C2'].value
    ages.append(age)
# Write ages to E2:E7
for idx, age in enumerate(ages, start=2):
    students_ws[f'E{idx}'] = age
wb.save(output_path)
