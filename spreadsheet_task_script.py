from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import calendar
import re

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun1/eval_51354_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed1/eval_seed42_rerun1/eval_51354_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active  # Assume first sheet

# Map MMM to month number
mon_map = {m: i for i, m in enumerate(calendar.month_abbr) if m}

# Color for column D
fill_yellow = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

def extract_and_increment(text):
    """Extract 'MMM YY' (last 6 chars), output ('next_MMM YY', 'YY') for output."""
    term = text.strip()[-6:]
    match = re.match(r'(\w{3})\s(\d{2})', term)
    if not match:
        return ('', '')
    mon, yr = match.groups()
    mon_num = mon_map.get(mon, 0)
    if not mon_num:
        return ('', yr)
    yr_num = int(yr)
    mon_num += 1
    if mon_num > 12:
        mon_num = 1
        yr_num += 1
    new_mon = calendar.month_abbr[mon_num]
    return (f'{new_mon} {str(yr_num).zfill(2)}', match.group(2))

for row in range(2, 7):  # Rows 2 to 6
    text = ws[f'A{row}'].value or ''
    res, yy = extract_and_increment(text)
    ws[f'D{row}'] = yy
    ws[f'D{row}'].fill = fill_yellow
    ws[f'E{row}'] = res

wb.save(output_path)
