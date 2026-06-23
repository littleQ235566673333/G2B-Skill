import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_55708_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_55708_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Extract data rows
rows = list(ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=3, values_only=True))

departments = ['A', 'B', 'C']
output_cells = ['B11', 'B12', 'B13']

for i, dept in enumerate(departments):
    filtered = [r[2] for r in rows if r[0] == dept and r[1] in ('In Progress', 'In Review') and isinstance(r[2], (int, float)) and r[2] >= 6]
    avg = sum(filtered) / len(filtered) if filtered else None
    ws[output_cells[i]].value = avg if avg is not None else None

wb.save(output_path)
