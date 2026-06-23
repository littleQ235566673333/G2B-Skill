import openpyxl
from openpyxl.styles import PatternFill
import itertools
import numpy as np

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_254-34_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun2/eval_254-34_tc1/output.xlsx'

# Target sum range
target_min, target_max = 994108, 994112

def get_numeric_cells(sheet):
    values = []
    cells = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                values.append(cell.value)
                cells.append(cell)
    return values, cells

wb = openpyxl.load_workbook(input_path)
if 'Sheet3' in wb.sheetnames:
    wb.remove(wb['Sheet3'])
sheet = wb['Before']

# Find numeric values and their cell references
values, cells = get_numeric_cells(sheet)
# Try combinations to match the desired sum - brute force up to 20 elements for practicality
found = False
for r in range(2, min(len(values), 20)):
    for combo in itertools.combinations(enumerate(values), r):
        idxs, nums = zip(*combo)
        sum_nums = sum(nums)
        if target_min <= sum_nums <= target_max:
            highlight_indices = idxs
            final_sum = sum_nums
            found = True
            break
    if found:
        break
# If a combo is found, highlight it:
if found:
    for i in highlight_indices:
        cells[i].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    sheet['C2'].value = final_sum
else:
    sheet['C2'].value = 'No combination found'
wb.save(output_path)
