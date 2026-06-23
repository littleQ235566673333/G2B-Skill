from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_4/regression_gate/after_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_4/regression_gate/after_pass/core_41601/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Students']

# Write formulas to E2:E7 based on student names in A2:A7
for row in range(2, 8):
    student_name = ws[f"A{row}"].value
    if student_name:
        ws[f"E{row}"] = f"='{student_name}'!C2"

wb.save(output_path)
