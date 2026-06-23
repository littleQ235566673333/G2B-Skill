import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Input/output file paths per user task
doc_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_1/regression_gate/after_fix/core_177-6/input.xlsx"
out_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v2/train/iter_1/regression_gate/after_fix/core_177-6/output.xlsx"
sheet_name = None # auto-detect
ref_col = 'H'
columns_needed = [chr(i) for i in range(ord('A'), ord('R')+1)]  # A:R

# Step 1: First, read entire data as a DataFrame
df = pd.read_excel(doc_path, sheet_name=None)
if isinstance(df, dict):
    # Take the first sheet which user example probably has
    sheet_name = list(df.keys())[0]
    df = df[sheet_name]
else:
    sheet_name = None

colnames = df.columns[:18]
df = df.iloc[:, :18]

# Step 2: Merge rows by reference column (H)
merge_col = colnames[7]  # H is 8th (0-based index 7)
grouped = df.groupby(merge_col, sort=False)
merged_rows = []
for key, group in grouped:
    agg = {}
    for i, col in enumerate(colnames[:8]):
        agg[col] = group[col].dropna().iloc[0] if not group[col].dropna().empty else ''
    for i, col in enumerate(colnames[8:]):
        # Coerce to numeric for safe sum
        val = pd.to_numeric(group[col], errors='coerce').sum()
        agg[col] = val
    merged_rows.append(agg)

merged_df = pd.DataFrame(merged_rows, columns=colnames)

# Step 3: Write to 'combined' sheet, preserving formats
wb = load_workbook(doc_path)
if 'combined' in wb.sheetnames:
    ws = wb['combined']
    ws.delete_rows(1, ws.max_row)
else:
    ws = wb.create_sheet('combined')

# Write headers
for c, v in enumerate(colnames, 1):
    ws.cell(row=1, column=c, value=v)

row_idx = 2
for i, (_, group) in enumerate(grouped):
    src_row_idx = group.index[0] + 2  # +2 for header
    for c, col in enumerate(colnames, 1):
        cell = ws.cell(row=row_idx, column=c)
        val = merged_df.iloc[i, c-1]
        if c > 8:
            cell.number_format = '0.00'
            cell.value = "" if pd.isna(val) or val == 0 else f"{val:.2f}"
        else:
            cell.value = val
        ws_in = wb[sheet_name]
        try:
            style_src = ws_in.cell(row=src_row_idx, column=c)
            cell.font = style_src.font
            cell.border = style_src.border
            cell.fill = style_src.fill
            cell.number_format = style_src.number_format if c <= 8 else '0.00'
            cell.protection = style_src.protection
            cell.alignment = style_src.alignment
        except Exception:
            pass
    row_idx += 1

# Keep only 8 rows after header (A1:R8 per user)
for _ in range(ws.max_row - 8):
    ws.delete_rows(ws.max_row)

wb.save(out_path)
