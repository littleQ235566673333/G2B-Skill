import openpyxl

# Load the workbook and select the first sheet
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/after_fix/core_50521/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/train/iter_5/regression_gate/after_fix/core_50521/output.xlsx"
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# For each row from 4 to 6
for row in range(4, 7):
    values = [ws.cell(row=row, column=col).value for col in range(2, 14)]  # B:M (2 to 13)
    
    # Find positions and values of numeric (non-blank, non-string) entries
    numeric_positions = [(idx, val) for idx, val in enumerate(values) if isinstance(val, (int, float))]
    
    if len(numeric_positions) < 2:
        ws.cell(row=row, column=14).value = None  # N
        continue

    first_idx, first_val = numeric_positions[0]
    second_idx, _ = numeric_positions[1]

    if first_val is not None and isinstance(first_val, (int, float)) and first_val > 1:
        ws.cell(row=row, column=14).value = 1
    else:
        # Count blanks between first and second numeric entries (exclusive of both bounds)
        gap = 0
        for i in range(first_idx + 1, second_idx):
            if values[i] is None or (isinstance(values[i], str) and values[i].strip() == ""):
                gap += 1
        ws.cell(row=row, column=14).value = gap

# Save the workbook to output
wb.save(output_path)
