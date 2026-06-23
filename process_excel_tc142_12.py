import openpyxl
from openpyxl.utils import get_column_letter

def is_whole_number(value):
    try:
        num = float(value)
        return num.is_integer()
    except (ValueError, TypeError):
        return False

# Load workbook and select sheet
wb = openpyxl.load_workbook('results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_142-12_tc1/input.xlsx')
ws = wb['Sheet1']

# Iterate through F (6th col) and J (10th col)
for row in ws.iter_rows(min_row=2, max_row=15, min_col=6, max_col=10):
    cell_f = row[0]  # F
    cell_j = row[4]  # J
    if cell_f.value == 'Marble Slab Creamery' and is_whole_number(cell_j.value):
        cell_f.value = 'Georgia State WH'

# Save results
out_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r2/eval_142-12_tc1/output.xlsx'
wb.save(out_path)
