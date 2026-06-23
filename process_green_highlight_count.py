import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r1/eval_43657_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_FIXED_r1/eval_43657_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

def is_green_fill(cell):
    fill = cell.fill
    if fill.patternType != 'solid':
        return False
    fgColor = ''
    if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
        fgColor = fill.fgColor.rgb.lower()
    elif hasattr(fill.fgColor, 'theme') and fill.fgColor.theme is not None:
        # theme color, can't reliably check (Excel stores green highlight as RGB)
        return False
    # main Excel greens
    return fgColor.startswith('ff00b0') or fgColor == 'ff00b050'

for out_row in range(2, 9):  # K2:K8, L2:L8, C:G
    name = ws[f'L{out_row}'].value
    count = 0
    for col in range(3, 8): # columns C=3 to G=7
        cell = ws.cell(row=out_row, column=col)
        if (cell.value == name) and is_green_fill(cell):
            count += 1
    ws[f'K{out_row}'] = count

wb.save(output_path)
