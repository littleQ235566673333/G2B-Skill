import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Paths
input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_50631_tc1/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_TIME-C/eval_50631_tc1/output.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb["Sheet1"]

# Font and fill for yellow
calibri_font = Font(name='Calibri')
yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00')

# Find holidays in column J (assuming header at J3 and values start from J4)
holidays = []
for row in range(4, ws.max_row + 1):
    cell = ws[f'J{row}']
    if cell.value is not None:
        holidays.append(cell.coordinate)

# Build reference string for the holidays, e.g. $J$4:$J$15
# Find the last non-empty cell in J for holiday range
holiday_start = 4
holiday_end = holiday_start
for row in range(holiday_start, ws.max_row + 1):
    if ws[f'J{row}'].value:
        holiday_end = row
    else:
        break
holiday_range = f"$J${holiday_start}:$J${holiday_end}" if holiday_end >= holiday_start else ""

# Insert WORKDAY formula in H7:H38
dest_col = 'H'
start_row = 7
end_row = 38
start_date_cell = '$B$3'
end_date_cell = '$F$3'

for i, row in enumerate(range(start_row, end_row + 1), 0):
    # Build the formula: only if B3+ROWS($B$7:$B7)-1 <= F3
    formula = f"=IF({start_date_cell}+ROWS($B$7:$B{row})-1<={end_date_cell},WORKDAY({start_date_cell}+ROWS($B$7:$B{row})-1,0,{holiday_range}),\"\")"
    ws[f'{dest_col}{row}'].value = formula
    # Retain previous formatting if the cell was yellow
    fill = ws[f'{dest_col}{row}'].fill
    if fill.start_color.rgb == 'FFFFFF00' or fill.fgColor.rgb == 'FFFFFF00': # Excel may use FFFFFF00 or FFFF00
        ws[f'{dest_col}{row}'].fill = yellow_fill
    ws[f'{dest_col}{row}'].font = calibri_font

wb.save(output_path)
