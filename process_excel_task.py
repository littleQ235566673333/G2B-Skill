import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_1/group_45707/r3/evolve_45707/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke/train/iter_1/group_45707/r3/evolve_45707/output.xlsx'

# Load into pandas for processing
df = pd.read_excel(input_path)
D = [None] * len(df)

for i in range(1, len(df)):
    try:
        next_date = pd.to_datetime(df.iloc[i, 0])
    except Exception:
        continue
    if next_date.day == 1:
        year, month = next_date.year, next_date.month
        dates_col = pd.to_datetime(df.iloc[:, 0], errors='coerce')
        mask = (dates_col.dt.year == year) & (dates_col.dt.month == month)
        D[i-1] = (df.loc[mask, df.columns[2]] == 1).sum()

# Write results back with openpyxl
gb = load_workbook(input_path)
ws = gb.active

# Copy original headers from A, C, E to A, C, E, and retain any existing for D (for compatibility as user wants headers untouched)
ws.cell(row=1, column=1, value=ws.cell(row=1, column=1).value)
ws.cell(row=1, column=3, value=ws.cell(row=1, column=3).value)
ws.cell(row=1, column=5, value=ws.cell(row=1, column=5).value)

# Write results to D2:D69
for idx in range(2, 70):  # D2 to D69 => df index 0..67
    value = D[idx-2]
    if value is not None:
        ws.cell(row=idx, column=4, value=value)
    else:
        ws.cell(row=idx, column=4, value=None)

gb.save(output_path)
