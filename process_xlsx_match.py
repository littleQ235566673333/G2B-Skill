from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_5/regression_gate/after_fix/core_48969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_5/regression_gate/after_fix/core_48969/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# Read the matching area
# Collect tables for comparison
# Assumption: There are two blocks of data to compare in fixed locations (adjust if needed)

# Read left table: E3:G5
left_table = [[ws.cell(row=i, column=j).value for j in range(5, 8)] for i in range(3, 6)]
# Read right table: H3:J5
right_table = [[ws.cell(row=i, column=j).value for j in range(8, 11)] for i in range(3, 6)]

# Compare and write Boolean results to J3:N5
for i in range(3):
    for j in range(5):
        left_val = left_table[i][j] if j < len(left_table[0]) else None
        right_val = right_table[i][j] if j < len(right_table[0]) else None
        if j < 3:  # Since only 3 columns per block in input, only fill to L
            ws.cell(row=3 + i, column=10 + j).value = (left_table[i][j] == right_table[i][j])

wb.save(output_path)
