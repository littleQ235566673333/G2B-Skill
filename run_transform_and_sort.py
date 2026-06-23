import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_3/group_118-50/r2/evolve_118-50/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_3/group_118-50/r2/evolve_118-50/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Step 1: Get all values in col A (including empty): up to 55555
words = [ws.cell(row=i, column=1).value for i in range(1, 55556)]
words = [w for w in words if isinstance(w, str) and w.strip() != '']  # Filter out blanks

# Step 2: Sort col A alphabetically (write back sorted names)
sorted_words = sorted(words)
for idx, word in enumerate(sorted_words):
    ws.cell(row=idx+1, column=1, value=word)

# Step 3: Prepare for finding matches and transformations
end_priority = ['ING', 'ERS', 'ATE', 'EST', 'ONE', 'IER', 'ILY']
grouped_results = {e: [] for e in end_priority}
other_results = {}

for word in words:
    w = word.strip()
    if len(w) < 5:
        continue
    ending = w[-3:]
    # Only process if at least 8 letters for transformation
    if len(w) >= 8:
        # Move 5th letter to start
        transformed = w[4] + w[:4] + w[5:]
        data = (transformed, w)
        if ending in grouped_results:
            grouped_results[ending].append(data)
        else:
            if ending not in other_results:
                other_results[ending] = []
            other_results[ending].append(data)

# Step 4: Paste output in specified order
results = []
for ending in end_priority:
    results.extend(grouped_results[ending])
for ending in sorted(other_results.keys()):
    results.extend(other_results[ending])

# Max 4999 results to fit C2:D5000
for i, (trans, orig) in enumerate(results[:4999]):
    ws.cell(row=i+2, column=3, value=trans)
    ws.cell(row=i+2, column=4, value=orig)

wb.save(output_path)
