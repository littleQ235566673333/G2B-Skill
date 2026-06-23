from openpyxl import load_workbook

def transform_emp_to_expected(in_path, out_path):
    wb = load_workbook(in_path)
    ws = wb['Emp']
    ws_lookup = wb['Lookup']
    
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    emp_rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    lookup_map = {}
    for row in ws_lookup.iter_rows(min_row=2, values_only=True):
        if row[0]:
            lookup_map[row[0]] = (row[1], row[2])
    
    output_data = []
    for emp in emp_rows:
        # Only process rows with all of seqno, Empno, ename, Jobid present
        if None in emp[:4] or all([v is None for v in emp[:4]]):
            continue
        base = list(emp[:4])
        for ci, col in enumerate(headers[4:], 4):
            var = col
            cat, subcat = lookup_map.get(var, ('',''))
            val = emp[ci]
            output_data.append(base + [var, cat, subcat, val])

    # Debug: print output_data
    print('Output Data:', output_data)
    
    # Write to output sheet at correct location
    if 'Expected' in wb.sheetnames:
        ws_out = wb['Expected']
        # clear area first (A2:H19)
        for row in ws_out.iter_rows(min_row=2, max_row=19, min_col=1, max_col=8):
            for cell in row:
                cell.value = None
    else:
        ws_out = wb.create_sheet('Expected')

    # Write header
    ws_out.cell(row=1, column=1, value=headers[0])
    ws_out.cell(row=1, column=2, value=headers[1])
    ws_out.cell(row=1, column=3, value=headers[2])
    ws_out.cell(row=1, column=4, value=headers[3])
    ws_out.cell(row=1, column=5, value='Variable')
    ws_out.cell(row=1, column=6, value='Category')
    ws_out.cell(row=1, column=7, value='Subcategory')
    ws_out.cell(row=1, column=8, value='Value')
    for i, row in enumerate(output_data, start=2):
        for j, v in enumerate(row, start=1):
            ws_out.cell(row=i, column=j, value=v)
    wb.save(out_path)

if __name__ == '__main__':
    transform_emp_to_expected(
        'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_1/regression_gate/before_fix/core_547-43/input.xlsx',
        'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_1/regression_gate/before_fix/core_547-43/output.xlsx')
