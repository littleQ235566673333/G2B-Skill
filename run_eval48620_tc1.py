import openpyxl
from openpyxl.styles import PatternFill, Font

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_48620_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r2/eval_48620_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Collect all product names and features from A2:A7, B2:B7
products = [ws[f'A{i}'].value for i in range(2, 8)]
features = [ws[f'B{i}'].value for i in range(2, 8)]
# Lookup values in D2:D7
lookups = [ws[f'D{i}'].value for i in range(2, 8)]

# For each lookup, return the B value from the row where A == lookup and index i
output = []
for idx, lookup in enumerate(lookups):
    # output the B value for the same row if D matches A in the same row
    if lookup == products[idx]:
        val = features[idx]
    else:
        val = None
    output.append(val)

# Write output to E2:E7, showing 0 as blank
for idx, val in enumerate(output):
    cell = ws[f'E{idx+2}']
    cell.value = '' if val == 0 or val == '0' or val is None else val

# Shade E2:E9, Calibri font
fill = PatternFill(fill_type='solid', fgColor='FCE4D6')
font = Font(name='Calibri')
for r in range(2, 10):
    cell = ws[f'E{r}']
    cell.fill = fill
    cell.font = font

wb.save(output_path)
