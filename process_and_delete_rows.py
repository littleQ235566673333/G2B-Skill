import openpyxl

# File paths
i_input = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_23-24_tc1/input.xlsx'
i_output = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42/eval_23-24_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(i_input)
ws = wb.active

# Collect words in column I (excluding header, assuming starts at row 1)
words_to_delete = set()
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=9, max_col=9):
    cell_val = row[0].value
    if cell_val is not None:
        words_to_delete.add(str(cell_val).strip())

# Iterate and mark rows to delete based on column A
rows_to_delete = []
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1), start=1):
    cell_val = row[0].value
    if cell_val is not None and str(cell_val).strip() in words_to_delete:
        rows_to_delete.append(idx)

# Delete marked rows (from bottom up to avoid index shift)
for row_idx in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(row_idx)

# Save result: only first 1102 rows, columns A-E
wb.save(i_output)
