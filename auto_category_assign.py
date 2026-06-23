from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_524-31_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_524-31_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Exp-DB']

category_map = {
    'amazon': 'Shopping',
    'auto': 'Transportation',
    'coffee': 'Dining',
    'ins': 'Insurance',
    'groceries': 'Groceries',
    'dining': 'Dining',
    'gas': 'Transportation',
}

for i in range(1, 54):
    short = ws.cell(row=i, column=2).value
    cat = ''
    if short:
        key = str(short).strip().lower()
        cat = category_map.get(key, 'Uncategorized')
    ws.cell(row=i, column=5, value=cat)

wb.save(output_path)
