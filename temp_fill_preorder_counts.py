from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/regression_gate/after_pass/core_50796/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_4/regression_gate/after_pass/core_50796/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# A1:B5 shows template inputs/outputs for values described in A12:C22
# A2:A5 = category
# B2:B5 = result cells (to fill)
# Data: A12:C22 (assume this is the actual sales/preorder table)

# Read which labels we need from A2:A5
criteria = [ws[f'A{i}'].value for i in range(2,6)]
results = []
# Loop over each label, sum C if B matches the criterion in rows 12:22
for crit in criteria:
    total = 0
    for row in range(12, 23):
        b = ws[f'B{row}'].value
        c = ws[f'C{row}'].value
        if b == crit and isinstance(c, (int, float)):
            total += c
    results.append(total)
# Write results to B2:B5
for idx, val in enumerate(results, start=2):
    ws[f'B{idx}'] = val
wb.save(output_path)
