import openpyxl
import pandas as pd
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/group_91-34/r0/evolve_91-34/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_8/group_91-34/r0/evolve_91-34/output.xlsx'

# Load workbook and select the 'SwiftMD' sheet
wb = openpyxl.load_workbook(input_path)
ws = wb['SwiftMD']

# Read header and data
header = [cell.value for cell in ws[2]][1:]
rows = []
for row in ws.iter_rows(min_row=3, min_col=2, max_col=15, values_only=True):
    if not any(row):
        continue
    rows.append(list(row))

df = pd.DataFrame(rows, columns=header)

# Detect duplicate candidates: marked 'Yes' and not Employee
mask = (df['Duplicate?'] == 'Yes') & (df['Relationship'] != 'Employee')
g = df[mask].groupby(['Last Name', 'First Name', 'Date Of Birth'])

rows_to_delete = set()
for k, group in g:
    # If any 'Employee' for same person+dob, do not delete any
    mask_employee = (
        (df['Last Name'] == k[0]) &
        (df['First Name'] == k[1]) &
        (df['Date Of Birth'] == k[2]) &
        (df['Relationship'] == 'Employee')
    )
    if df[mask_employee].shape[0]:
        continue
    # Otherwise, if more than one, delete exactly one (the first)
    if len(group) > 1:
        rows_to_delete.add(group.index[0])

# Drop selected duplicates
new_df = df.drop(rows_to_delete).reset_index(drop=True)

# Clear output area (B2:O42)
for row in ws.iter_rows(min_row=2, max_row=42, min_col=2, max_col=15):
    for cell in row:
        cell.value = None

# Write header
for j, h in enumerate(header, start=2):
    ws.cell(row=2, column=j, value=h)

# Write new data
for i, row in enumerate(new_df.itertuples(index=False), start=3):
    for j, val in enumerate(row, start=2):
        ws.cell(row=i, column=j, value=val)

wb.save(output_path)
