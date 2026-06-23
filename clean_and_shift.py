from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_4/regression_gate/after_pass/core_493-18/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed1/train/iter_4/regression_gate/after_pass/core_493-18/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# 1. Read all values from F into a set
f_values = set()
for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
    v = row[0].value
    if v is not None:
        f_values.add(v)

# 2. For each row in A, if A not in F, clear A, B, C (not row delete)
ac_rows = []
for row in ws.iter_rows(min_row=2, max_col=6, values_only=False):
    a_val = row[0].value
    if a_val is not None and a_val in f_values:
        ac_rows.append([cell.value for cell in row[:3]])
    else:
        for cell in row[:3]:
            cell.value = None

# 3. Shift data in A, B, C upwards to fill empty slots
for idx, ac_row in enumerate(ac_rows, start=2):
    for j, val in enumerate(ac_row):
        ws.cell(row=idx, column=j+1, value=val)
# Clear down any remainder rows in A-C
for idx in range(2+len(ac_rows), ws.max_row+1):
    for col in range(1,4):
        ws.cell(row=idx, column=col, value=None)

# 4. Add autofilter to columns A, B, C
ws.auto_filter.ref = 'A1:C7'

wb.save(output_path)
