import openpyxl
from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_22-47_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_22-47_tc1/output.xlsx'

wb = load_workbook(input_path)
sheet = wb.active

def is_header(row):
    return any(str(cell).strip().lower() in ['name', 'header', 'names', 'id'] for cell in row)

def is_empty(row):
    return all(cell is None or str(cell).strip() == '' for cell in row)

def is_duplicate(row, seen):
    key = (str(row[0]).strip(), str(row[1]).strip())
    if key in seen:
        return True
    seen.add(key)
    return False

# Extract data rows
rows = []
seen = set()

for i in range(1, sheet.max_row+1):
    b_val = sheet[f'B{i}'].value
    c_val = sheet[f'C{i}'].value
    j_val = sheet[f'J{i}'].value
    original_row = [b_val, c_val, j_val, i]
    if is_empty([b_val, c_val]) or is_header([b_val, c_val]):
        continue
    if is_duplicate([b_val, c_val], seen):
        continue
    rows.append(original_row)

# Helper column J
helper_list = []
for i in range(1, sheet.max_row+1):
    j_val = sheet[f'J{i}'].value
    if j_val and str(j_val).strip() != '' and not is_header([j_val]):
        helper_list.append(str(j_val).strip())

output_rows = []

if helper_list:
    # Grouped sort: take all names in J, their matching rows (in source order), then the rest
    matched = set()
    for name in helper_list:
        for row in rows:
            if str(row[0]).strip() == name:
                output_rows.append(row)
                matched.add(row[3])
    # Add non-listed names, preserving original order
    for row in rows:
        if row[3] not in matched:
            output_rows.append(row)
else:
    # If helper is empty, sort alphabetically by column B
    output_rows = sorted(rows, key=lambda x: str(x[0]).strip())

# Only write up to 9 rows, in F2:H10
max_out = 9
out_slice = output_rows[:max_out]

# Output F: Name, G: Col C, H: sorted ascending
h_vals = sorted([row[1] for row in out_slice if row[1] is not None])
for idx, row in enumerate(out_slice):
    sheet[f'F{idx+2}'] = row[0]
    sheet[f'G{idx+2}'] = row[1]
    sheet[f'H{idx+2}'] = h_vals[idx] if idx < len(h_vals) else ''

wb.save(output_path)
