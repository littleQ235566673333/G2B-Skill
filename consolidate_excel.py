import openpyxl
from openpyxl.styles import PatternFill

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_2/regression_gate/after_pass/core_80-42/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/train/iter_2/regression_gate/after_pass/core_80-42/output.xlsx"

def first_blank_row(ws, min_row=2, max_row=8000, col=1):
    for row in range(min_row, max_row+1):
        if ws.cell(row=row, column=col).value is None:
            return row
    return max_row+1

def get_yellow_columns(ws):
    """Returns zero-based column indices for columns with yellow fill in the first row"""
    yellow_cols = []
    for col in range(1, ws.max_column+1):
        fill = ws.cell(row=1, column=col).fill
        if isinstance(fill, PatternFill):
            rgb = fill.fgColor.rgb
            # Check for Excel's default yellow colors
            if rgb in ('FFFFFF00', 'FFFF00', 'FFFFFF99', 'FFFF99'):
                yellow_cols.append(col)
    return yellow_cols

wb = openpyxl.load_workbook(input_path)
consolidated_ws = wb["Consolidate_ALL"]

sheets_to_consolidate = ['Jack', 'Henry', 'Richard']
for sheetname in sheets_to_consolidate:
    ws = wb[sheetname]
    yellow_cols = get_yellow_columns(ws)
    if not yellow_cols:
        continue  # nothing to consolidate
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # skip those with no primary data (check col 1)
        if row[0].value is None:
            continue
        values = [row[col-1].value for col in yellow_cols]
        # Only include up to 11 cols (A:K) if too many yellow found
        values = values[:11]
        # Pad to 11 cols if less
        while len(values) < 11:
            values.append(None)
        values.append(sheetname)  # Originating sheet name in Col L
        dest_row = first_blank_row(consolidated_ws)
        for c, v in enumerate(values, start=1):
            cell = consolidated_ws.cell(row=dest_row, column=c, value=v)
            # Optionally: match formatting with previous row, if present
            if dest_row > 2:
                src_cell = consolidated_ws.cell(row=dest_row-1, column=c)
                cell.font = src_cell.font
                cell.alignment = src_cell.alignment
                cell.border = src_cell.border
                cell.number_format = src_cell.number_format
                # Do not copy fill

def remove_fills(ws, min_row=2, max_col=12):
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            cell.fill = PatternFill()  # No fill

remove_fills(consolidated_ws)

wb.save(output_path)
