from openpyxl import load_workbook
from datetime import datetime, timedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_2/regression_gate/before_pass/core_41589/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-smoke16/train/iter_2/regression_gate/before_pass/core_41589/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Contact List']

def evaluate_action(last_contact_date, response):
    # Handle string or date value for last_contact_date
    if not isinstance(last_contact_date, (datetime, type(None))):
        try:
            last_contact_date = datetime.strptime(str(last_contact_date), '%Y-%m-%d')
        except Exception:
            return 'NO ACTION'
    if response is not None:
        response = str(response).strip().upper()
    else:
        response = ''
    today = datetime.today()
    if response == 'YES':
        if last_contact_date is not None and isinstance(last_contact_date, datetime):
            if today - last_contact_date <= timedelta(days=30):
                return 'HOLD'
            else:
                return 'TOUCH BASE'
    return 'NO ACTION'

last_contact = ws['H4'].value
yes_no = ws['I4'].value
result = evaluate_action(last_contact, yes_no)
ws['J4'] = result

wb.save(output_path)
