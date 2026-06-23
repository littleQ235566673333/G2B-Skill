from openpyxl import load_workbook

def build_lookup(sheet, brand_col, cat_row, val_start_col, val_end_col, val_start_row, val_end_row):
    # Reads categories and brand list from the sheet to build a (brand,category) to value dict
    categories = [sheet.cell(row=cat_row, column=col).value for col in range(val_start_col, val_end_col+1)]
    brand_lookup = {}
    for r in range(val_start_row, val_end_row+1):
        brand = sheet.cell(row=r, column=brand_col).value
        if not brand:
            continue
        for idx, cat in enumerate(categories):
            val = sheet.cell(row=r, column=val_start_col+idx).value
            if cat and val is not None:
                brand_lookup[(brand, cat)] = val
    return brand_lookup

def transfer_values(input_path, output_path):
    wb = load_workbook(input_path)
    ws = wb.active
    # Source table layout inferred from preview
    source_brand_col = 2  # B
    source_cat_row = 5
    source_first_col = 3  # C
    source_last_col = 10  # J
    source_first_row = 6
    source_last_row = 13
    lookup = build_lookup(ws, source_brand_col, source_cat_row, source_first_col, source_last_col, source_first_row, source_last_row)

    # Destination area M6:S11 (col 13 to 19, row 6 to 11)
    for dest_row in range(6, 12):
        brand = ws.cell(row=dest_row, column=13).value  # L (col 12+1=13)
        if not brand:
            continue
        for dest_col in range(14, 20):  # M:S (cols 14 to 19)
            cat = ws.cell(row=5, column=dest_col).value
            if (brand, cat) in lookup:
                ws.cell(row=dest_row, column=dest_col, value=lookup[(brand, cat)])
            else:
                ws.cell(row=dest_row, column=dest_col, value=None)
    wb.save(output_path)

transfer_values(
    'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_48983_tc1/input.xlsx',
    'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun1/eval_48983_tc1/output.xlsx'
)
