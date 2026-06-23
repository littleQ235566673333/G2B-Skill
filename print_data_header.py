import openpyxl
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-v2-ss41-seed2/eval_FIXED_r1/eval_55965_tc1/input.xlsx')
data_ws = None
for name in wb.sheetnames:
    if name.lower() == 'data' or name.lower().startswith('data'):
        data_ws = wb[name]
if not data_ws:
    data_ws = wb[wb.sheetnames[1]]
header = [str(cell.value) for cell in next(data_ws.iter_rows(min_row=1, max_row=1))]
print(header)
