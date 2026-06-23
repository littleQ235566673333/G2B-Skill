import openpyxl
import pandas as pd

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_250-20_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_NOW_TIME_VERIFY_r1/eval_250-20_tc1/output.xlsx'
sheetname = 'RNM'

# Load workbook and sheet
data_wb = openpyxl.load_workbook(input_path)
data_ws = data_wb[sheetname]

# Read rows A1:J20
rows = list(data_ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10, values_only=True))

# First row is header
header = rows[0]
data = rows[1:]

# Create DataFrame
cols = list(header)
df = pd.DataFrame(data, columns=cols)

# Drop rows that are completely empty
df = df.dropna(how='all')

# Group by cols B and C
col_b, col_c, col_j = cols[1], cols[2], cols[9]
# For summed J:
sum_j = df.groupby([col_b, col_c], as_index=False)[col_j].sum()
# For first appearance:
firsts = df.groupby([col_b, col_c], as_index=False).first()
# Set summed J back to firsts:
firsts = pd.merge(firsts, sum_j, on=[col_b, col_c], suffixes=('', '_SUM'))
firsts[col_j] = firsts[col_j + '_SUM']
firsts = firsts[cols]  # Arrange columns as original

# Combine header and result, max 20 rows in total
out_rows = [header] + firsts.values.tolist()
out_rows = out_rows[:20]  # For 'RNM'!A1:J20

# Clear cells
for row in data_ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
    for cell in row:
        cell.value = None

# Write output
for r_idx, row_data in enumerate(out_rows, start=1):
    for c_idx, val in enumerate(row_data, start=1):
        data_ws.cell(row=r_idx, column=c_idx, value=val)

data_wb.save(output_path)
