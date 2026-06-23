import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_1/group_44017/r2/evolve_44017/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke-v3/train/iter_1/group_44017/r2/evolve_44017/output.xlsx"

wb = load_workbook(input_path)
ws = wb['Data']

# Index settings
first_row, last_row = 14, 42
first_col, last_col = 30, 41  # AD:AO
# Where the month dates are
month_dates = [ws.cell(row=9, column=col).value for col in range(first_col, last_col+1)]

for row in range(first_row, last_row + 1):
    base_rate = f"$W{row}"
    freq_cell = f"$J{row}"
    eff_cell = f"$L{row}"
    inc_cells = [f"$M{row}", f"$N{row}", f"$O{row}", f"$P{row}"]
    for idx, col in enumerate(range(first_col, last_col+1)):
        cell = ws.cell(row=row, column=col)
        # The target month header cell reference, e.g., AD9
        month_cell_ref = ws.cell(row=9, column=col).coordinate
        # Build the formula
        # For each wave, DATEDIF($L14,AD$9,"m") >= 0, >=freq, >=2*freq, >=3*freq
        wave_triggers = [
            f"(DATEDIF({eff_cell},{month_cell_ref},\"m\")>={freq_mult}*{freq_cell})"
            for freq_mult in range(4)
        ]
        main_cond = f"({month_cell_ref}>={eff_cell})"
        factors = [
            f"(1+{inc_cells[i]})^({wave_triggers[i]})" for i in range(4)
        ]
        formula = (
            f"=IF({main_cond}",
            f", {base_rate}*" + "*".join(factors) + ",\"\")"
        )
        cell.value = "".join(formula)
        # Remove any yellow fill by setting to pattern fill none
        cell.fill = PatternFill(fill_type=None)

wb.save(output_path)
print(f"Wrote updated Excel to {output_path}")
