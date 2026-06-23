import openpyxl

# Load original and data_only version for formulas
wb = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_3/regression_gate/before_pass/core_56599/input.xlsx')
wb_do = openpyxl.load_workbook('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_3/regression_gate/before_pass/core_56599/input.xlsx', data_only=True)
ws = wb['Sheet1']
ws_do = wb_do['Sheet1']

# Get slip number to find from M5 (column M, row 5)
slip_search = ws['M5'].value
result = 'not found - please re-enter'

if slip_search is not None:
    # Iterate from row 6 downward until a blank in G (Slip fm)
    row = 6
    while True:
        slip_fm = ws_do.cell(row=row, column=7).value  # G
        slip_to = ws_do.cell(row=row, column=8).value  # H
        book_no = ws_do.cell(row=row, column=10).value  # J
        if slip_fm is None:
            break
        if slip_to is None or book_no is None:
            row +=1
            continue
        # Check if the slip_search falls into this book's range
        try:
            if int(slip_fm) <= int(slip_search) <= int(slip_to):
                result = book_no
                break
        except Exception:
            pass
        row += 1

# Write result in B2
ws['B2'] = result
wb.save('results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed1/train/iter_3/regression_gate/before_pass/core_56599/output.xlsx')
print('Done.')
