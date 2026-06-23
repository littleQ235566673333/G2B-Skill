import openpyxl
import pandas as pd

INPUT_PATH = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_61-4_tc1/input.xlsx"
OUTPUT_PATH = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun3/eval_61-4_tc1/output.xlsx"

def main():
    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb['input']
    # Read up to 100 rows for safety of demo, adjust as needed
    data = list(ws.iter_rows(min_row=2, max_row=100, max_col=8, values_only=True))
    df = pd.DataFrame(data, columns=['DATE', 'Stock Name', 'OPENP', 'HIGH', 'LOW', 'CLOSEP', 'VOLUME', 'Change'])
    # Drop all rows with Stock Name nan or date nan - blank rows
    df = df[df['Stock Name'].notna() & df['DATE'].notna()]
    result = []
    for stock, g in df.groupby('Stock Name'):
        g = g.sort_values('DATE').reset_index(drop=True)
        neg_mask = g['Change'] < 0
        grp = (neg_mask != neg_mask.shift()).cumsum()
        for _, group in g.groupby(grp):
            if group['Change'].iloc[0] < 0:
                # Consecutive negative block
                result.append({
                    'START_DATE': group['DATE'].iloc[0],
                    'END_DATE': group['DATE'].iloc[-1],
                    'Stock Name': stock,
                    'OPENP': group['OPENP'].iloc[0],
                    'HIGH': group['HIGH'].max(),
                    'LOW': group['LOW'].min(),
                    'CLOSEP': group['CLOSEP'].iloc[-1],
                    'VOLUME': group['VOLUME'].sum(),
                })
    outdf = pd.DataFrame(result)
    # Fill output sheet: A2:G
    wsout = wb['output']
    # Clear area
    for i in range(2, 16):
        for j in range(1, 8):
            wsout.cell(row=i, column=j).value = None
    # Write
    for i, row in outdf.iterrows():
        if i >= 14:
            break  # To stay inside A2:G15
        wsout.cell(row=i+2, column=1).value = row['START_DATE']
        wsout.cell(row=i+2, column=2).value = row['END_DATE']
        wsout.cell(row=i+2, column=3).value = row['Stock Name']
        wsout.cell(row=i+2, column=4).value = row['OPENP']
        wsout.cell(row=i+2, column=5).value = row['HIGH']
        wsout.cell(row=i+2, column=6).value = row['LOW']
        wsout.cell(row=i+2, column=7).value = row['CLOSEP']
    wb.save(OUTPUT_PATH)

if __name__ == '__main__':
    main()
