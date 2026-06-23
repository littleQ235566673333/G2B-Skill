import openpyxl
from datetime import datetime

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_58723_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed1/eval_seed42_rerun1/eval_58723_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

name_col = 3  # column C
entry_time_col = 9  # column I
output_col = 13  # column M
start_row, end_row = 2, 41

rows = []
for row in range(start_row, end_row+1):
    name = ws.cell(row=row, column=name_col).value
    entry = ws.cell(row=row, column=entry_time_col).value
    if isinstance(entry, str):
        try:
            entry = datetime.strptime(entry, '%Y-%m-%d %H:%M:%S')
        except Exception:
            try:
                entry = datetime.strptime(entry, '%m/%d/%Y %H:%M')
            except Exception:
                entry = None
    rows.append((row, name, entry))

from collections import defaultdict
latest_times = defaultdict(lambda: None)
for row, name, entry in rows:
    if entry is not None:
        if latest_times[name] is None or entry > latest_times[name]:
            latest_times[name] = entry

for row, name, entry in rows:
    if entry is not None and entry == latest_times[name]:
        ws.cell(row=row, column=output_col, value='Latest')
    elif entry is not None:
        ws.cell(row=row, column=output_col, value='Not Latest')
    else:
        ws.cell(row=row, column=output_col, value=None)

wb.save(output_path)
