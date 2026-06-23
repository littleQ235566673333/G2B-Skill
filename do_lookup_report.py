import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/regression_gate/before_pass/core_18935/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-seed0/train/iter_6/regression_gate/before_pass/core_18935/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Locate the data table
# Headers start (row 4): 'Work', 'Material', 'Category 1', ...
data_table_start = 5
data_table_end = 11  # exclusive, since one past last data row

# Build lookup from data table
data = []
for row in ws.iter_rows(min_row=data_table_start, max_row=data_table_end, min_col=1, max_col=5, values_only=True):
    work, material, cat1, cat2, cat3 = row
    data.append({'Work': work, 'Material': material, 'Category 1': cat1, 'Category 2': cat2, 'Category 3': cat3})

# Fill report format (header at row 16, data rows start at 17)
for r in range(17, 23):
    work = ws[f'A{r}'].value
    cat_type = ws[f'B{r}'].value
    material = ws[f'C{r}'].value
    # Lookup
    match = next((row for row in data if row['Work'] == work and row['Material'] == material), None)
    val = match[cat_type] if match else None
    ws[f'D{r}'] = val

wb.save(output_path)