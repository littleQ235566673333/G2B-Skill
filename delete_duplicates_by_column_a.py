import openpyxl

def delete_duplicates_by_column_a(input_path, output_path, sheet_name):
    wb = openpyxl.load_workbook(input_path)
    ws = wb[sheet_name]

    # Read all rows starting from row 2 (exclude header)
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    unique_keys = set()
    unique_rows = []

    for row in rows:
        key = row[0]  # Column A
        if key not in unique_keys:
            unique_keys.add(key)
            unique_rows.append(row)

    # Clear existing rows (except header)
    for i in range(2, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            ws.cell(row=i, column=j).value = None

    # Write the deduplicated rows back
    for idx, row in enumerate(unique_rows, start=2):
        for jdx, val in enumerate(row, start=1):
            ws.cell(row=idx, column=jdx, value=val)

    wb.save(output_path)

# File paths and sheet info
input_file = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_448-11_tc1/input.xlsx'
output_file = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed0/eval_seed42_rerun2/eval_448-11_tc1/output.xlsx'
sheet = 'Sheet1'

delete_duplicates_by_column_a(input_file, output_file, sheet)
