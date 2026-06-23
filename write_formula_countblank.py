from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/before_pass/core_41969/input.xlsx"
output_path = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/before_pass/core_41969/output.xlsx"

wb = load_workbook(input_path)
ws = wb[wb.sheetnames[0]]

# We want:
# A6 = '=COUNTBLANK(A3:C3)'
# B6 = '=COUNTBLANK(D3:F3)'
# C6 = '=COUNTBLANK(G3:I3)'

for i in range(3):
    start_col = i * 3 + 1  # 1-based
    end_col = start_col + 2
    col1 = get_column_letter(start_col)
    col2 = get_column_letter(end_col)
    dest_cell = f"{get_column_letter(i+1)}6"  # A6, B6, C6
    ws[dest_cell] = f"=COUNTBLANK({col1}3:{col2}3)"

wb.save(output_path)
