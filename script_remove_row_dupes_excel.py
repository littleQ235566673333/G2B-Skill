from openpyxl import load_workbook
import pandas as pd

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_7/regression_gate/after_fix/core_227-40/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_7/regression_gate/after_fix/core_227-40/output.xlsx'

# Load the workbook and worksheet
wb = load_workbook(input_path)
ws = wb['Sheet1']

# Read with pandas to locate demonstration block exact rows
df = pd.read_excel(input_path, sheet_name='Sheet1', header=None)
# Find the 'The Result' string to locate demo output area
res_row = df[df.apply(lambda row: row.astype(str).str.contains('The Result').any(), axis=1)].index[0]
# Copy header (next row under the result block label)
header = df.iloc[res_row+1,0:5].tolist()
data = df.iloc[res_row+2:res_row+6,0:5].values.tolist()

# Remove duplicates in each row, keeping first occurrence

def remove_dupes_row(row):
    seen = set()
    result = []
    for cell in row:
        if pd.isna(cell):
            result.append(cell)
            continue
        if cell not in seen:
            result.append(cell)
            seen.add(cell)
        else:
            result.append('')  # Replace duplicate with blank per typical pattern
    return result

processed = [remove_dupes_row(row) for row in data]

# Write the header and processed result to A14:E17
start_row = 14
for col,label in enumerate(header,1):
    ws.cell(row=start_row, column=col).value = label
for i, row in enumerate(processed):
    for j, value in enumerate(row, 1):
        ws.cell(row=start_row+1+i, column=j).value = value

wb.save(output_path)
