import openpyxl

# Load workbook and worksheet
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_49945_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed0/eval_seed42_rerun2/eval_49945_tc1/output.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb['Sheet1']

# Read model names from G4:G6
models = [ws[f'G{i}'].value for i in range(4, 7)]
results = []
for model in models:
    total = 0
    for row in range(4, 12):
        if ws[f'C{row}'].value == model:
            val = ws[f'D{row}'].value
            if isinstance(val, (int, float)):
                total += val
    results.append(total)
# Write results to G4:G6
for i, result in enumerate(results, 4):
    ws[f'G{i}'] = result
wb.save(output_path)
