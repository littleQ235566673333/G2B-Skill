import openpyxl
from openpyxl.utils import get_column_letter

infile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/group_44017/r0/evolve_44017/input.xlsx'
outfile = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed0/train/iter_4/group_44017/r0/evolve_44017/output.xlsx'

wb = openpyxl.load_workbook(infile)
ws = wb.active

output_start_col = 30  # AD
output_end_col = 41    # AO
output_start_row = 14
output_end_row = 42
date_row = 9

for r in range(output_start_row, output_end_row + 1):
    base_rate_cell = f'W{r}'
    eff_date_cell = f'L{r}'
    freq_cell = f'J{r}'
    inc1_cell = f'M{r}'
    inc2_cell = f'N{r}'
    inc3_cell = f'O{r}'
    inc4_cell = f'P{r}'
    for c in range(output_start_col, output_end_col + 1):
        col_letter = get_column_letter(c)
        month_date_cell = f'{col_letter}{date_row}'
        formula = (
            f'=IF({month_date_cell}<$${eff_date_cell},"",'
            f'$${base_rate_cell}'
            f'* (1 + IF((YEAR({month_date_cell})-YEAR($${eff_date_cell}))*12+'
            f'(MONTH({month_date_cell})-MONTH($${eff_date_cell})) >= 0, $${inc1_cell}, 0))'
            f'* (1 + IF((YEAR({month_date_cell})-YEAR($${eff_date_cell}))*12+'
            f'(MONTH({month_date_cell})-MONTH($${eff_date_cell})) >= $${freq_cell}, $${inc2_cell}, 0))'
            f'* (1 + IF((YEAR({month_date_cell})-YEAR($${eff_date_cell}))*12+'
            f'(MONTH({month_date_cell})-MONTH($${eff_date_cell})) >= $${freq_cell}*2, $${inc3_cell}, 0))'
            f'* (1 + IF((YEAR({month_date_cell})-YEAR($${eff_date_cell}))*12+'
            f'(MONTH({month_date_cell})-MONTH($${eff_date_cell})) >= $${freq_cell}*3, $${inc4_cell}, 0))'
            f')'
        )
        # Replace $$ with $ for Excel absolute refs
        formula = formula.replace('$$', '$')
        ws[f'{col_letter}{r}'] = formula

wb.save(outfile)
