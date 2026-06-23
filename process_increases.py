import openpyxl
from datetime import datetime
from dateutil.relativedelta import relativedelta

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_2/group_44017/r3/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-smoke/train/iter_2/group_44017/r3/evolve_44017/output.xlsx'
wb = openpyxl.load_workbook(input_path, data_only=True)
ws = wb.active

col_ad_ao = list(range(30, 43))  # AD:AO
dates = [ws.cell(row=9, column=c).value for c in col_ad_ao]
for row in range(14, 43):
    base_rate = ws.cell(row=row, column=23).value  # column W
    eff_date = ws.cell(row=row, column=12).value  # column L
    freq = ws.cell(row=row, column=10).value      # column J
    increases = [ws.cell(row=row, column=col).value or 0 for col in range(13, 17)]  # M-P
    if not freq:
        freq = 0
    for idx, col in enumerate(col_ad_ao):
        date = dates[idx]
        cell = ws.cell(row=row, column=col)
        if not date or not eff_date or base_rate is None:
            cell.value = None
            continue
        if date < eff_date or not isinstance(freq, (int, float)) or freq <= 0:
            cell.value = None
            continue
        # Calculate waves
        waves = [eff_date + relativedelta(months=i*freq) for i in range(4)]
        increases_applied = [i for i, w in enumerate(waves) if date >= w and increases[i]]
        rate = base_rate
        for i in increases_applied:
            rate *= (1 + increases[i])
        cell.value = round(rate, 6)
wb.save(output_path)
