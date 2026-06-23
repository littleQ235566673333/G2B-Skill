from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42/eval_37900_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42/eval_37900_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Sheet1']
# Write the formula to E5
ws['E5'] = '=XLOOKUP(TODAY(),A5:A10,B5:B10,"Not found")'
wb.save(output_path)
