import openpyxl
from openpyxl.styles import PatternFill

# File paths
input_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_51354_tc1/input.xlsx'
output_fp = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1-PRUNED/eval_seed42/eval_51354_tc1/output.xlsx'

# Map month abbreviations to their numbers, and vice versa
month_map = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
num_to_month = {v:k for k,v in month_map.items()}

wb = openpyxl.load_workbook(input_fp)
ws = wb.active

# Fill for #FFC000 (Excel gold)
gold_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

for row in range(2, 7):  # E2:E6 (corresponding rows)
    val = ws[f'A{row}'].value
    if val:
        last_six = val[-6:]
        # Try to parse month and year
        month_abbr = last_six[:3]
        year_part = last_six[-2:]
        # Only fill D if year_part is digits
        ws[f'D{row}'].value = year_part
        ws[f'D{row}'].fill = gold_fill
        # Increment month
        month_num = month_map.get(month_abbr, 0)
        if month_num == 0 or not year_part.isdigit():
            ws[f'E{row}'].value = ''
            continue
        inc_month = month_num + 1
        inc_year = int(year_part)
        if inc_month > 12:
            inc_month = 1
            inc_year += 1
        ws[f'E{row}'].value = f'{num_to_month[inc_month]} {inc_year:02d}'
    else:
        ws[f'D{row}'].value = ''
        ws[f'E{row}'].value = ''

wb.save(output_fp)