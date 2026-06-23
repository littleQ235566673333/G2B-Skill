import pandas as pd
from openpyxl import load_workbook

def main():
    inp = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_5/regression_gate/before_fix/core_91-34/input.xlsx"
    outp = "results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_5/regression_gate/before_fix/core_91-34/output.xlsx"
    sheet = "SwiftMD"

    wb = load_workbook(inp)
    ws = wb[sheet]

    # Header is in row 2
    header = [ws.cell(row=2, column=col).value for col in range(2,16)]
    # Data is rows 3-42
    data = [[ws.cell(row=row, column=col).value for col in range(2,16)] for row in range(3,43)]
    df = pd.DataFrame(data, columns=header)
    df['_orig_idx'] = range(3, 43)  # for reference

    to_delete = set()
    # Group by (Last Name, First Name, Date Of Birth) for 'Duplicate? == Yes'
    df_dup = df[df['Duplicate?'] == 'Yes']
    grouped = df_dup.groupby(["Last Name", "First Name", "Date Of Birth"])
    for key, group in grouped:
        relationships = set(group['Relationship'])
        group_idxs = group.index.tolist()
        if group_idxs and all(r != 'Employee' for r in relationships):
            # Mark only one for deletion in this group (delete first)
            to_delete.add(group_idxs[0])
    # Remove marked rows
    df_cleaned = df.drop(index=to_delete)
    # Pad/truncate to always fit 40 rows (B3:O42)
    n_rows = 40
    rows = df_cleaned.shape[0]
    if rows < n_rows:
        add_empty = n_rows - rows
        df_cleaned = pd.concat([
            df_cleaned,
            pd.DataFrame([[None]*len(header+['_orig_idx'])]*add_empty, columns=header+['_orig_idx'])
        ], ignore_index=True)
    else:
        df_cleaned = df_cleaned.iloc[:n_rows]
    # Write header back to row 2
    for j, h in enumerate(header):
        ws.cell(row=2, column=2+j, value=h)
    # Write data
    for i, row in enumerate(df_cleaned[header].itertuples(index=False, name=None)):
        for j, val in enumerate(row):
            ws.cell(row=3+i, column=2+j, value=val)
    wb.save(outp)

if __name__=='__main__':
    main()
