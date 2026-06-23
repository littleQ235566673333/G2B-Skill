import openpyxl
from openpyxl import load_workbook

input_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/group_56274/r0/evolve_56274/input.xlsx'
output_file = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/train/iter_8/group_56274/r0/evolve_56274/output.xlsx'

wb = load_workbook(input_file)
ws = wb.active

fiscal_month = ws['D7'].value
if fiscal_month is None:
    raise Exception('D7 (fiscal month) is empty!')

# Find the fiscal month column in row 3, columns 7-10 (G3:J3)
month_col = None
for col in range(7, 11):
    cell_value = ws.cell(row=3, column=col).value
    if cell_value == fiscal_month:
        month_col = col
        break
if month_col is None:
    raise Exception(f'Could not find Fiscal Month {fiscal_month} in row 3 (columns G-J).')

# Now extract the required items for that column:
# Row 4: Opening Bal
# Row 5: Debits
# Row 6: Credits
# Row 7: Closing Bal
ws['D9'] = ws.cell(row=4, column=month_col).value  # Opening Balance
ws['D10'] = ws.cell(row=5, column=month_col).value # Debits
ws['D11'] = ws.cell(row=6, column=month_col).value # Credits
ws['D12'] = ws.cell(row=7, column=month_col).value # Closing Balance

wb.save(output_file)
print('Done.')
