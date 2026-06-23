import openpyxl

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed2/train/iter_8/group_56274/r0/evolve_56274/input.xlsx'
wb = openpyxl.load_workbook(input_path)
ws = wb.active

fiscal_month = ws['D7'].value
print('D7:', fiscal_month, type(fiscal_month))

print('Row 7 values and types (columns 1-15):')
for col in range(1, 16):
    cell = ws.cell(row=7, column=col)
    print(f'col {col}:', cell.value, type(cell.value))
