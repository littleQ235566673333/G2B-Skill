from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# File paths
input_xlsx = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_1/group_44017/r0/evolve_44017/input.xlsx'
output_xlsx = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_1/group_44017/r0/evolve_44017/output.xlsx'
wb = load_workbook(input_xlsx)
ws = wb.active

start_row, end_row = 14, 42
start_col, end_col = 30, 43  # AD (30) to AO (42)

col_letters = [get_column_letter(col) for col in range(start_col, end_col)]

for row in range(start_row, end_row+1):
    for i, col in enumerate(range(start_col, end_col)):
        base_rate = f'$W{row}'
        eff_date = f'$L{row}'
        freq = f'$J{row}'
        incs = [f'$M{row}', f'$N{row}', f'$O{row}', f'$P{row}']
        month_cell = f'{col_letters[i]}$9'
        # Build cumulative Excel formula, applying up to 4 increases in sequence
        formula = f'=IF({month_cell} < {eff_date}, "", {base_rate}'
        for w in range(4):
            wave_idx = w + 1
            # Step triggers: DATEDIF(effective, month, "m") >= wave_idx * freq
            condition = f'(DATEDIF({eff_date}, {month_cell}, "m") >= {wave_idx}*{freq})'
            increase = f'*(1+IF({incs[w]}="",0,{incs[w]})*IF({condition},1,0))'
            formula += increase
        formula += ')'
        ws.cell(row=row, column=col).value = formula

wb.save(output_xlsx)
print('Done')
