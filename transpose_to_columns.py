import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_42902_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42_rerun2/eval_42902_tc1/output.xlsx'

# Load workbook and sheet
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Extract values from the first column (A)
values = [row[0] for row in ws.iter_rows(min_row=1, max_col=1, values_only=True)]
values = [v for v in values if v is not None]

# Split into groups of 3
records = [values[i:i+3] for i in range(0, len(values), 3)]

# Only up to 7 records for D1:F7
for idx, rec in enumerate(records[:7]):
    # Write to D, E, F
    ws.cell(row=1+idx, column=4).value = rec[0] if len(rec)>0 else None
    ws.cell(row=1+idx, column=5).value = rec[1] if len(rec)>1 else None
    ws.cell(row=1+idx, column=6).value = rec[2] if len(rec)>2 else None

wb.save(output_path)
print('Done!')
