import openpyxl
from openpyxl.styles import Alignment

# Input and output paths
input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_56378_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-smoke16/eval_seed42_rerun2/eval_56378_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Folha1']

# Gather Frame 1 data rows (C5:J10) where QUANTITY UNITS (I, col=9) is not empty
indices = []
for r in range(5, 11):  # C5:J10 => rows 5 to 10
    value = ws.cell(row=r, column=9).value
    if value not in (None, ''):
        indices.append(r)

# Up to 4 non-blank rows (to fill L5:R8)
for i, source_row in enumerate(indices[:4]):
    for j in range(0, 7):  # 7 columns: C:D:E:F:G:H:I
        dest_cell = ws.cell(row=5 + i, column=12 + j) # L is 12
        src_col_letter = openpyxl.utils.get_column_letter(3 + j)  # C=3
        src_cell_ref = f"{src_col_letter}{source_row}"
        dest_cell.value = f"={src_cell_ref}"
        # Alignment: PRODUCT (M, j==1) left, O-R (j>=2) right
        if j == 1:
            dest_cell.alignment = Alignment(horizontal='left')
        elif j >= 2:
            dest_cell.alignment = Alignment(horizontal='right')
        else:
            dest_cell.alignment = Alignment(horizontal='general')

wb.save(output_path)
