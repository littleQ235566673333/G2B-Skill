import openpyxl
from datetime import datetime

input_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_9448_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_SG-FIXU-SEED-N3-seed2/eval_seed42/eval_9448_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['Data']

col_letters = list('IJKLMNOPQRST')

for row in range(9, 19):
    last_col = None
    for idx, col_letter in enumerate(reversed(col_letters)):
        cell = ws[f'{col_letter}{row}'].value
        if cell is not None and str(cell).strip() != '':
            last_col = len(col_letters) - idx - 1
            break
    if last_col is not None:
        header_value = ws[f'{col_letters[last_col]}7'].value
        try:
            # Try parsing common formats
            if isinstance(header_value, datetime):
                month = header_value.month
                year = header_value.year
            elif isinstance(header_value, str):
                try:
                    dt = datetime.strptime(header_value, '%b %Y')
                except ValueError:
                    try:
                        dt = datetime.strptime(header_value, '%B %Y')
                    except ValueError:
                        try:
                            dt = datetime.strptime(header_value, '%d/%m/%Y')
                        except ValueError:
                            parts = header_value.replace(',', '').split()
                            if len(parts) == 2:
                                mn_str, yr_str = parts
                                for fmt in ('%b', '%B'):
                                    try:
                                        mn = datetime.strptime(mn_str, fmt).month
                                        yr = int(yr_str)
                                        dt = datetime(yr, mn, 1)
                                        break
                                    except Exception:
                                        pass
                                else:
                                    raise ValueError
                            else:
                                raise ValueError(f'Unknown header: {header_value}')
                month = dt.month
                year = dt.year
            else:
                raise ValueError(f'Unknown header: {header_value}')
            dt = datetime(year, month, 1)
            ws[f'U{row}'].value = dt
            ws[f'U{row}'].number_format = 'DD/MM/YYYY'
        except Exception:
            ws[f'U{row}'].value = 'ERR'
    else:
        ws[f'U{row}'].value = None

wb.save(output_path)
