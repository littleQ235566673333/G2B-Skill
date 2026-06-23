from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_6/regression_gate/after_pass/core_41601/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_6/regression_gate/after_pass/core_41601/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Students']

for row in range(2, 8):  # Rows 2 through 7
    student_cell = f'A{row}'
    col = 'E'
    name = ws[student_cell].value
    if name:
        # Build the Excel formula: =INDIRECT("'" & A2 & "'!C2")
        formula = f"=INDIRECT(\"'\" & {student_cell} & \"'!C2\")"
        ws[f'{col}{row}'] = formula
    else:
        ws[f'{col}{row}'] = ""

wb.save(output_path)
