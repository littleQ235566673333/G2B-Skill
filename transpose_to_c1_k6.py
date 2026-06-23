from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_2/group_147-48/r2/evolve_147-48/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-N3-seed2/train/iter_2/group_147-48/r2/evolve_147-48/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet1']

col_a = []
for row in ws.iter_rows(min_col=1, max_col=1):
    val = row[0].value
    if val is not None:
        col_a.append(val)

rows, cols = 6, 9
for i in range(rows):
    for j in range(cols):
        idx = i * cols + j
        value = col_a[idx] if idx < len(col_a) else ""
        ws.cell(row=1+i, column=3+j, value=value)

wb.save(output_path)
