from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_4/group_56274/r3/evolve_56274/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed1/train/iter_4/group_56274/r3/evolve_56274/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# D7 holds the fiscal month (datetime)
fiscal_month = ws['D7'].value

# Row 3 is the header row for fiscal months (col index starts from 1)
month_col = None
for col in range(7, ws.max_column+1):  # commonly starts at column G (7)
    if ws.cell(row=3, column=col).value == fiscal_month:
        month_col = col
        break

# Default to clear if not found
open_bal = debits = credits = close_bal = None
if month_col:
    open_bal = ws.cell(row=4, column=month_col).value
    debits = ws.cell(row=5, column=month_col).value
    credits = ws.cell(row=6, column=month_col).value
    close_bal = ws.cell(row=7, column=month_col).value

ws['D9'].value = open_bal
ws['D10'].value = debits
ws['D11'].value = credits
ws['D12'].value = close_bal

wb.save(output_path)
