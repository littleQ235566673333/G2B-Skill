from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_1/regression_gate/after_pass/core_41969/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-ST-smoke16/train/iter_1/regression_gate/after_pass/core_41969/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active  # Use the first sheet (index 0)

# Output cells: A6, B6, C6
# Apply formulas: =COUNTBLANK(A3:C3), =COUNTBLANK(D3:F3), =COUNTBLANK(G3:I3)
row = 6
num_outputs = 3
for i in range(num_outputs):
    start_col_idx = i * 3
    start_col_letter = chr(ord('A') + start_col_idx)
    end_col_letter = chr(ord('A') + start_col_idx + 2)
    formula = f'=COUNTBLANK({start_col_letter}3:{end_col_letter}3)'
    output_cell = f'{chr(ord("A") + i)}{row}'
    ws[output_cell] = formula

wb.save(output_path)
