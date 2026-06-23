from openpyxl import load_workbook

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r3/eval_40892_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-ss41-seed1/eval_FIXED_r3/eval_40892_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

# 1. Retrieve list of colors (Column D, skipping header)
color_set = set()
for cell in ws['D'][1:]:
    val = cell.value
    if val:
        color_set.add(val.strip().lower())

# 2. For each cell in A2:A17, check if it contains any color from the list
for row in range(2, 18):
    desc = ws[f'A{row}'].value
    found_color = ''
    if desc:
        desc_lower = desc.lower()
        for color in color_set:
            # match whole word or substring
            if color in desc_lower:
                found_color = color.title()
                break
    ws[f'B{row}'] = found_color

wb.save(output_path)
