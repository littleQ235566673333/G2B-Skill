from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v3/train/iter_1/group_44017/r0/evolve_44017/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-pilot-v3/train/iter_1/group_44017/r0/evolve_44017/output.xlsx'

wb = load_workbook(input_path)
ws = wb['Data']

# Constants
start_row = 14
end_row = 42
first_formula_col = 30 # AD
last_formula_col = 43  # AO + 1

no_fill = PatternFill(fill_type=None)

for row in range(start_row, end_row + 1):
    for col in range(first_formula_col, last_formula_col):
        cell = ws.cell(row=row, column=col)
        base_rate = f'$W{row}'
        freq = f'$J{row}'
        eff_date = f'$L{row}'
        inc1 = f'$M{row}'
        inc2 = f'$N{row}'
        inc3 = f'$O{row}'
        inc4 = f'$P{row}'
        month_date = f'{get_column_letter(col)}$9'
        # Formula logic
        formula = (
            f'=IF({month_date}<$L{row}, "", '
            f'$W{row}*'
            f'(1+IF($M{row}<>0, IF(DATEDIF($L{row},{month_date},"m")>=$J{row}*0,$M{row},0),0))'
            f'*(1+IF($N{row}<>0, IF(DATEDIF($L{row},{month_date},"m")>=$J{row}*1,$N{row},0),0))'
            f'*(1+IF($O{row}<>0, IF(DATEDIF($L{row},{month_date},"m")>=$J{row}*2,$O{row},0),0))'
            f'*(1+IF($P{row}<>0, IF(DATEDIF($L{row},{month_date},"m")>=$J{row}*3,$P{row},0),0))'
            f')'
        )
        cell.value = formula
        cell.fill = no_fill  # Remove fill, ensure no yellow

wb.save(output_path)
