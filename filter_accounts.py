import openpyxl
from collections import defaultdict
from datetime import datetime

INPUT_PATH = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_567-21_tc1/input.xlsx'
OUTPUT_PATH = 'results/runs/g2b-v8_gpt-4.1_ss-gpt41/eval_seed2/eval_567-21_tc1/output.xlsx'

def parse_date(val):
    # Accept 'YYYY/MM', datetime, and skip others
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.replace(day=1)
    if isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), '%Y/%m')
        except:
            return None
    return None

def main():
    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    ws = wb['Sheet1']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]

    groups = defaultdict(list)
    for row in data:
        # Only process if both col A and B are present
        key = (row[0], row[1])
        date_val = row[4]
        parsed_date = parse_date(date_val)
        if parsed_date is None:
            continue  # skip if date E is empty or invalid
        groups[key].append((parsed_date, row))

    # For each (A, B), find most recent date, keep all rows with that date
    result_rows = []
    for key, entries in groups.items():
        if not entries:
            continue
        max_date = max(d for d, r in entries)
        for d, r in entries:
            if d == max_date:
                result_rows.append(r)

    # Sort result for consistent placement: by account, then by assoc, then by date descending
    result_rows.sort(key=lambda r: (r[0], r[1], parse_date(r[4]) if parse_date(r[4]) else datetime.min), reverse=False)

    # Output to new Sheet1, placing results on A3:G28
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Sheet1'

    # Pad up to row 3 and write header to row 3
    out_ws.append(['']*7)
    out_ws.append(['']*7)
    out_ws.append(list(header[:7]))
    for row in result_rows[:25]:
        out_ws.append(list(row)[:7])
    # Pad to G28 if needed
    while out_ws.max_row < 28:
        out_ws.append(['']*7)

    out_wb.save(OUTPUT_PATH)

if __name__ == '__main__':
    main()
