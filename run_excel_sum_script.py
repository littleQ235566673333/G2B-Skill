import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import calendar

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_57590_tc1/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-Q-ss41-N3-seed2/eval_seed42/eval_57590_tc1/output.xlsx'

wb = load_workbook(input_path)
ws = wb.active

data = []
row = 2
while True:
    i_val = ws[f'I{row}'].value
    c_val = ws[f'C{row}'].value
    if i_val is None and c_val is None:
        break
    data.append((i_val, c_val))
    row += 1

# Read the month to match from cell A26
month_val = ws['A26'].value

# Infer year from the first date in column C
first_date = None
for _, c_val in data:
    if c_val is not None:
        try:
            first_date = pd.to_datetime(c_val)
            break
        except Exception:
            continue
if first_date is None:
    raise ValueError('No valid date in column C to infer year.')

if isinstance(month_val, str):
    try:
        # Try to parse e.g. 'January' as a month in the first_date.year
        month_num = list(calendar.month_name).index(month_val)
        if month_num == 0:
            raise ValueError(f'Invalid month name: {month_val}')
        year = first_date.year
        start_month = pd.Timestamp(year=year, month=month_num, day=1)
        # end of month
        end_month = start_month + pd.offsets.MonthEnd(0)
    except Exception:
        # Fallback: try parsing it as a date string
        start_month = pd.to_datetime(month_val).replace(day=1)
        end_month = (start_month + pd.DateOffset(months=1)) - pd.DateOffset(days=1)
else:
    start_month = pd.to_datetime(month_val).replace(day=1)
    end_month = (start_month + pd.DateOffset(months=1)) - pd.DateOffset(days=1)

filtered_sum = 0
for i_val, c_val in data:
    if pd.isna(i_val) or pd.isna(c_val):
        continue
    try:
        c_date = pd.to_datetime(c_val)
    except Exception:
        continue
    if start_month <= c_date <= end_month:
        filtered_sum += i_val

ws['B26'] = filtered_sum
wb.save(output_path)
