import openpyxl

input_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed2/eval_55060_tc1/input.xlsx'
output_path = 'results/runs/skillgrad_gpt-4.1_ss-gpt41/eval_seed2/eval_55060_tc1/output.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb.active

# Unmerge J23:N23 if necessary
from openpyxl.utils import get_column_letter
for merged_range in list(ws.merged_cells.ranges):
    if merged_range.coord == 'J23:N23':
        ws.unmerge_cells('J23:N23')

# Read value from I12
val = ws['I12'].value

# Set J23:N23 to val or blank
for col in range(10, 15):  # J=10, K=11, L=12, M=13, N=14
    ws.cell(row=23, column=col, value=(val if val not in (None, '') else None))

wb.save(output_path)
