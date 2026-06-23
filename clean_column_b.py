import re
from openpyxl import load_workbook
from datetime import datetime

input_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/before_pass/core_290-27/input.xlsx'
output_path = 'results/runs/g2b-skill-spreadsheet_gpt-4.1_c-topo-FIX-U-N3-seed0/train/iter_6/regression_gate/before_pass/core_290-27/output.xlsx'
wb = load_workbook(input_path)
ws = wb['Sheet1']

def clean_cell(val):
    if isinstance(val, str):
        # Remove all leading uppercase alpha codes (2 or 3) and optional space
        # Match 'PID1' => '1', 'PID 1' => '1', 'GG 1' => '1', 'GG 45' => '45'
        mt = re.match(r'^([A-Z]{2,3}) ?([0-9]+)$', val)
        if mt:
            return mt.group(2)
        return val
    return val

for row in range(14, 138):
    v = ws[f'B{row}'].value
    if isinstance(v, datetime):
        continue
    ws[f'B{row}'].value = clean_cell(v)

wb.save(output_path)
print('Completed cleaning and output saved.')
